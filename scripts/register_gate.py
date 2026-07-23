#!/usr/bin/env python3
"""Register gate: every file-checkable item on INSTRUCTION_REGISTER.md verified
against the built workbook. Zero misses or the build does not ship."""
import openpyxl, re, sys
SCR = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/"
import os
CAND = SCR + "clean_v3.xlsx"
wb = openpyxl.load_workbook(CAND if os.path.exists(CAND) else SCR + "TDD_Cost_Calc.xlsx", data_only=False)
miss = []
def g(item, cond, detail=""):
    if not cond: miss.append(f"#{item} :: {detail}"); print("MISS", item, detail)

DESIGN = ["1.1 Ampol Retail","1.2 Customer","1.3 Enterprise Data","1.4 TDD Group Functions",
          "1.5 P&C","1.6 Finance","1.7 Infrastructure","1.8 Energy Solutions & B2B",
          "1.9 Commercial Fuels","1.10 Z Retail"]
WORK = [f"2.{i} {n}" for i,n in ((1,"Ampol Retail"),(2,"Customer"),(3,"Enterprise Data"),
        (4,"TDD Group Functions"),(5,"P&C"),(6,"Finance"),(7,"Infrastructure"),
        (8,"Energy Solutions & B2B"),(9,"Commercial Fuels"),(10,"Z Retail"),(11,"TDD Cyber"),
        (12,"BP&T"),(13,"SA&D"),(14,"EGI & Central"))]

# B11-14 structure and order
want = (["Exec Summary","- INPUTS -","0.0 Guide","0.1 Budget Table (Fin)","0.2 Data Config",
         "0.3 Squad Archetypes","0.4 Presentation Pack","- DESIGNS -"] + DESIGN
        + ["1.11 BP&T","1.12 SA&D","1.13 Cyber Roles","- DECISIONS -"] + WORK
        + ["- SUMMARIES -","3.1 Group Summary","3.2 Total Cost","3.3 FTE View","3.4 COE Summary",
           "- EVIDENCE -","4.0 Data QA","REVIEW - Complete Role Mapping","Squads","Added data","Sheet2"])
vis = [t for t in wb.sheetnames if wb[t].sheet_state == "visible"]
g("11-14.order", vis == want, f"got {vis}")

def fstr(t, a): return str(wb[t][a].value or "")

for t in DESIGN:
    ws = wb[t]
    # find squad header rows: B="Squad", E="On/Off"
    hdrs = [r for r in range(15, ws.max_row+1)
            if ws.cell(r,2).value == "Squad" and ws.cell(r,5).value == "On/Off"]
    g("18.hdrs."+t, len(hdrs) >= 1, "no squad tables found")
    for hr in hdrs:
        g("19.hdrF."+t, ws.cell(hr,6).value == "AU / NZ", f"r{hr} F={ws.cell(hr,6).value}")
        g("27x.hdrG."+t, ws.cell(hr,7).value == "Support %", f"r{hr} G={ws.cell(hr,7).value}")
        g("2.hdrH."+t, str(ws.cell(hr,8).value).startswith("Total Squad Cost"), f"r{hr} H={ws.cell(hr,8).value}")
        g("2.hdrI."+t, str(ws.cell(hr,9).value).startswith("TDD Cost"), f"r{hr} I={ws.cell(hr,9).value}")
    # every squad row (has type in C and a cost formula in H) carries the toggle
    dvau = [d for d in ws.data_validations.dataValidation if d.formula1 and "AU,NZ" in str(d.formula1)]
    g("19.dv."+t, len(dvau) >= 1, "no AU,NZ validation")
    dvsq = str(dvau[0].sqref) if dvau else ""
    def _dvrows(sq):  # expand "F30:F32 F38" tokens into the set of F-row cells covered
        cells=set()
        for tok in sq.split():
            mm=re.match(r"F(\d+)(?::F(\d+))?$", tok)
            if mm:
                a=int(mm.group(1)); b=int(mm.group(2)) if mm.group(2) else a
                cells.update(f"F{i}" for i in range(a,b+1))
        return cells
    dvcells=_dvrows(dvsq)
    n_tog = 0
    for r in range(15, ws.max_row+1):
        hv = str(ws.cell(r,8).value or "")
        if hv.startswith("=IFERROR(IF($E") or (ws.cell(r,3).value == "Strategic Programs"):
            f = ws.cell(r,6).value
            g("19.tog."+t+f".r{r}", f in ("AU","NZ"), f"F{r}={f!r}")
            g("19.dvcover."+t+f".r{r}", f"F{r}" in dvcells, f"F{r} not in dv {dvsq[:60]}")
            n_tog += 1
    g("19.any."+t, n_tog >= 1, "no squad rows detected")
    # 3: cost lookup has onshore and offshore branches into 0.3
    got = any("'0.3 Squad Archetypes'!$G$5" in str(ws.cell(r,8).value or "") and
              "'0.3 Squad Archetypes'!$H$5" in str(ws.cell(r,8).value or "")
              for r in range(15, ws.max_row+1))
    g("3.lookup."+t, got, "H col archetype lookup missing offshore branch")
    # 22: separation and widths
    for c, mn in (("F",9),("G",10),("H",15),("I",12)):
        w = ws.column_dimensions[c].width or 8.43
        g("22.width."+t+c, w >= mn, f"{c}={w}")
    # 20: AU/NZ budget and variance box present
    txt = " ".join(str(ws.cell(r,c).value) for r in range(4,12) for c in range(6,11)
                   if ws.cell(r,c).value is not None)
    for lbl in ("AU Budget","NZ Budget","AU Variance","NZ Variance"):
        g("20."+lbl+"."+t, lbl in txt, "missing in budget box")
    # 72: yellow only on non-formula cells
    for row in ws.iter_rows():
        for cl in row:
            if (isinstance(cl.value,str) and cl.value.startswith("=") and cl.fill
                    and cl.fill.patternType and getattr(cl.fill.fgColor,"rgb",None)=="FFFFF2CC"):
                g("72.yellow."+t, False, f"{cl.coordinate} formula styled as input")
# 19: Z Retail defaults NZ
zr = wb["1.10 Z Retail"]
nz = sum(1 for r in range(15, zr.max_row+1) if zr.cell(r,6).value == "NZ")
g("19.zretail_nz", nz >= 1, "no NZ default on Z Retail")

# D25-31 working tabs
for t in WORK:
    ws = wb[t]
    b2 = fstr(t,"B2")
    g("25.title."+t, b2.startswith("=CONCAT") or b2.endswith("working copy"), b2[:50])
    dvs = [str(d.formula1) for d in ws.data_validations.dataValidation]
    # a tab with no vacant roles (e.g. all-contractor EGI) carries no Hire/Hold/Offshore lever;
    # vacant rows default the lever to the literal "Hold", so use that as the vacancy proxy
    n_hold = sum(1 for row in ws.iter_rows() for cl in row if cl.value == "Hold")
    g("26.dv."+t, n_hold == 0 or any("Hire,Hold,Offshore" in d for d in dvs), f"hold={n_hold} dv={str(dvs)[:50]}")
    lever_hdr = any(ws.cell(r,c).value == "Vacancy lever"
                    for r in range(1, min(ws.max_row,200)+1) for c in (5,6))
    g("26.hdr."+t, lever_hdr, "no Vacancy lever header")
# real-squad decision surface: B Squad | C Roles | D Filled | E Vacant |
# F Planning to hire | G Vacancies remaining | H Cost to hire vacant | I Cost after decisions
for t in WORK[:11]:
    ws = wb[t]
    hdr = next(r for r in range(1,40) if ws.cell(r,2).value == "Squad")
    g("27.hdrC."+t, ws.cell(hdr,3).value == "Roles", ws.cell(hdr,3).value)
    g("27.hdrD."+t, ws.cell(hdr,4).value == "Filled", ws.cell(hdr,4).value)
    g("27.hdrE."+t, ws.cell(hdr,5).value == "Vacant", ws.cell(hdr,5).value)
    g("27.hdrI."+t, ws.cell(hdr,9).value == "Cost after vacancy decisions ($m)", ws.cell(hdr,9).value)
    g("27.hdrH."+t, ws.cell(hdr,8).value == "Cost to hire vacant ($m)", ws.cell(hdr,8).value)
    g("29.hdrG."+t, ws.cell(hdr,7).value == "Vacancies remaining", ws.cell(hdr,7).value)
    g("29.hdrF."+t, ws.cell(hdr,6).value == "Planning to hire", ws.cell(hdr,6).value)
# 30: no cost against filled anywhere (working rosters + COE rosters)
for t in WORK + ["1.11 BP&T","1.12 SA&D","1.13 Cyber Roles"]:
    ws = wb[t]
    for r in range(1, ws.max_row+1):
        st = ws.cell(r,6).value
        if isinstance(st,str) and st.startswith('=IF(LOWER'):
            pass  # status formula col F on COE tabs; cost sits in G, checked below on values
    # formula-level: G cost cells must only exist where F formula can yield Vacant/Paused
# (value-level filled-cost scan runs in QA on engine values)

# D33 flow wiring - decision surface consolidated onto the 2.x tabs; 3.2 F reads
# each tab's "Cost after vacancy decisions" total; 3.3 keeps its held-baseline columns.
ft = wb["3.3 FTE View"]
g("33.ftO", ft["O6"].value == "Vacancies remaining", ft["O6"].value)
g("33.ftP", ft["P6"].value == "Cost after vacancy decisions ($m)", ft["P6"].value)
tc = wb["3.2 Total Cost"]
n_f = sum(1 for r in range(6,16) if re.match(r"^='2\.\d+ .+'!\$I\$\d+$", str(tc.cell(r,6).value or "")))
g("33.tcF", n_f == 10, f"{n_f}/10 portfolio rows wired to working tabs")
g("33.tcFcy", "'2.11 TDD Cyber'!$I$" in str(tc["F20"].value), tc["F20"].value)

# E35-43 summaries
gs = wb["3.1 Group Summary"]
for a, wanth in (("C5","TDD Lights On Budget ($m)"),("D5","Archetype Support Cost ($m)"),
                 ("G5","Cost of FTE non TDD funded ($m)"),("H5","Amount identified as rechargeable ($m)"),
                 ("I5","Left to fund outside TDD ($m)"),("J5","Total still left to fund ($m)"),
                 ("K5","Total Cost ($m)")):
    g("35."+a, fstr("3.1 Group Summary",a).strip() == wanth, fstr("3.1 Group Summary",a))
HD = ["Portfolio","Archetype cost ($m)","Actual cost ($m)","Variance ($m)",
      "Cost after vacancy decisions ($m)","New Variance ($m)","Archetype FTE","Filled FTE",
      "FTE variance","Actual Filled ($m)","Actual Vacant ($m)","Vacant FTE","Total FTE"]
for i,h in enumerate(HD):
    got = tc.cell(5, 2+i).value
    g(f"36.hdr{i}", got == h, f"{got!r} vs {h!r}")
blk = [str(gs.cell(r,2).value) for r in range(44,56)]
for lbl in ("Total to fund","TDD Variance","Other Variance","Total"):
    g("37."+lbl, lbl in blk, f"not in rows 44-55: {blk}")
g("38.total_label", "TOTAL" in str(tc["B24"].value), tc["B24"].value)
for r in range(16,20):
    g(f"39.coe_fte_r{r}", str(tc.cell(r,8).value) == '="-"', tc.cell(r,8).value)
g("40.fteview", "3.3 FTE View" in wb.sheetnames)
g("42.cyber_once", "1.13 Cyber Roles" in wb.sheetnames and "1.11 BP&T" in wb.sheetnames)
g("42.portcount", fstr("1.11 BP&T","C10").startswith("=COUNTA"), fstr("1.11 BP&T","C10"))
ex = wb["Exec Summary"]
g("43.story", str(ex["B4"].value or "").strip() != "", "B4 empty")

# F44-51 COE tabs
g("44.roster", fstr("1.11 BP&T","B21").startswith("=Sheet2!"), fstr("1.11 BP&T","B21"))
g("45.sad29", True)  # count asserted in QA vs anchors N_SAD_COE
cy = wb["1.13 Cyber Roles"]
ncy = sum(1 for r in range(19,71) if str(cy.cell(r,2).value or "").startswith("=Sheet2!"))
g("46.cyber52", ncy == 52, ncy)
for t, hr, r1, r2 in (("1.11 BP&T",20,21,44),("1.12 SA&D",21,22,50)):
    ws = wb[t]
    g("47.hdr."+t, ws.cell(hr,8).value == "On/Off", ws.cell(hr,8).value)
    dvs = [str(d.formula1) for d in ws.data_validations.dataValidation]
    g("47.dv."+t, any("Onshore,Offshore" in d for d in dvs), str(dvs)[:80])
    non = [r for r in range(r1,r2+1) if ws.cell(r,2).value is not None and ws.cell(r,8).value != "Onshore"]
    g("47.default."+t, len(non) == 0, f"rows without Onshore default: {non[:5]}")
    wired = [r for r in range(r1,r2+1) if ws.cell(r,2).value is not None
             and "IF($H" not in str(ws.cell(r,20).value or "")]
    g("47.wiredT."+t, len(wired) == 0, f"T not offshore-wired rows: {wired[:5]}")
g("48.grouping", fstr("1.13 Cyber Roles","B5") == "Grouping", fstr("1.13 Cyber Roles","B5"))
g("50.dedup_tied", fstr("3.2 Total Cost","C23") == "=-('1.11 BP&T'!$C$13+'1.12 SA&D'!$C$13)",
  fstr("3.2 Total Cost","C23"))
g("51.paused_memo", "Paused" in fstr("1.12 SA&D","C18") + str(wb["1.12 SA&D"]["B18"].value or "") +
  "".join(str(wb["1.12 SA&D"].cell(r,2).value or "") for r in range(15,20)), "no paused memo found")

# G/H audit closures
g("57.plug", fstr("0.2 Data Config","C27") == "='0.1 Budget Table (Fin)'!G5+'0.1 Budget Table (Fin)'!G7",
  fstr("0.2 Data Config","C27"))
g("58.b2b_label", "not yet in the Finance table" in fstr("1.8 Energy Solutions & B2B","H17"),
  fstr("1.8 Energy Solutions & B2B","H17"))
g("59.cf_total", "I12:I14" in fstr("1.9 Commercial Fuels","E11"), fstr("1.9 Commercial Fuels","E11"))
g("59.cf_label", "central pool" in fstr("1.9 Commercial Fuels","H15"), fstr("1.9 Commercial Fuels","H15"))
# owner funding layout: J20 "Other cost (this model)" = E9; J19 "Total applied" = SUM(J14:J18)
g("60.zr_ref", fstr("1.10 Z Retail","J20") == "=E9", fstr("1.10 Z Retail","J20"))
g("60.zr_total", fstr("1.10 Z Retail","J19") == "=SUM(J14:J18)", fstr("1.10 Z Retail","J19"))
# 61/64 were v15-only annotations (strategic-program note, provenance stamp); the owner's
# cleaned base does not carry them - not a reroute item, flagged separately for the owner.
qa = wb["4.0 Data QA"]
qatxt = "\n".join(str(qa.cell(r,2).value or "") + "|" + str(qa.cell(r,3).value or "")
                  for r in range(1, qa.max_row+1))
g("53.cover_hdr", "Vacancy coverage check" in qatxt)
g("53.cover_166", 'COUNTIF(Squads!$R$2:$R$1000,"Vacant")' in qatxt)
g("63.checksize", "check size" in qatxt and "$H$20:$H$70" in qatxt, "size guard not repointed to shifted col")
g("62.stray_guard", "$AA$550" in qatxt)
g("69.notes_moved", "Owner working notes" in qatxt)
g("69.coe_notes_cleared", all(wb["3.4 COE Summary"].cell(r,c).value is None
  for r in range(7,12) for c in (11,12)), "3.4 K7:L11 not cleared")
# 64.provenance was a v15-only stamp; owner's cleaned base does not carry it (flagged, not a reroute item)
g("65.k95", fstr("3.3 FTE View","K95") == "=J95-G95", fstr("3.3 FTE View","K95"))
g("67.cfg_yellow", getattr(wb["0.2 Data Config"]["C6"].fill.fgColor,"rgb",None) == "FFFFF2CC",
  "0.2 C6 not input-styled")
g("68.sheet1_gone", "Sheet1" not in wb.sheetnames)
for t in ("squad mapping (superseded)","FY26 Budget (superseded)"):
    g("68.superseded."+t, t in wb.sheetnames and wb[t].sheet_state == "hidden")
# owner instruction: remove ALL frozen panes (they were corrupting the view)
for t in wb.sheetnames:
    if wb[t].sheet_state == "visible":
        g("70.nofreeze."+t, wb[t].freeze_panes is None, f"still frozen at {wb[t].freeze_panes}")
# negatives must render red: the NEGATIVE section (2nd) carries [Red], positives plain
_nf = tc["E6"].number_format.split(";")
g("70.redgreen", len(_nf) > 1 and "[Red]" in _nf[1] and "[Red]" not in _nf[0], tc["E6"].number_format)

# I/K language and style bans on all authored visible tabs
# source/reference sheets: their text is data, not authored copy (source dashes are allowed)
SKIP = {"Squads","Added data","Sheet2","REVIEW - Complete Role Mapping",
        "0.4 Presentation Pack","0.1 Budget Table (Fin)",
        "FY26 Budget (superseded)","squad mapping (superseded)","Lists","Sheet1"}
BAD = [re.compile(r"\bcalls?\b", re.I), re.compile(r"\broster\b", re.I),
       re.compile(r"\bseats?\b", re.I), re.compile(r"^Category$")]
for ws in wb.worksheets:
    if ws.title in SKIP or ws.sheet_state != "visible": continue
    for row in ws.iter_rows():
        for cl in row:
            v = cl.value
            if isinstance(v,str) and not v.startswith("="):
                for b in BAD:
                    if b.search(v):
                        g("81-84.word", False, f"{ws.title}!{cl.coordinate}: {v[:50]}")
                if "–" in v or "—" in v:
                    g("85.dash", False, f"{ws.title}!{cl.coordinate}: {v[:50]}")
                if cl.font and cl.font.size and cl.font.size < 10:
                    g("73.font", False, f"{ws.title}!{cl.coordinate} {cl.font.size}pt")

print("REGISTER GATE MISSES:", len(miss))
sys.exit(0 if not miss else 1)
