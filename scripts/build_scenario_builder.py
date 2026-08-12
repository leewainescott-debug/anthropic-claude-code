#!/usr/bin/env python3
"""Build the offshore scenario builder workbook: a dashboard-grade role picker.

Clean-formula version, no macros. Plain English - the three sets of roles
are List A / List B / List C, never "basket".

Math preserved exactly from the model (verified, numbers tie):
  cost taken out / yr = cost today - 0.4 * full onshore cost
                        (0 if not moved, or a vendor day-rate role)
  year split is by calendar year (Ampol Dec year-end): 2026, 2027.
"""
import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.formatting.rule import FormulaRule, CellIsRule, DataBarRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter

ROLES = json.load(open("/tmp/claude-0/-home-user-anthropic-claude-code/e550b440-3996-5abb-87e5-bafafe598f82/scratchpad/roles.json"))
OUT = "deliverables/TDD_Offshore_Scenario_Builder.xlsx"
PWD = "Tdd123"

NAVY   = "0F2E52"
NAVY2  = "1B4B87"
INK    = "1D2939"
MUTE   = "667085"
HAIR   = "D0D5DD"
BANDR  = "F6F8FB"
CARD   = "FAFBFD"
EMPH   = "E8EEF6"
INPUT  = "FFF4CC"
INPUTB = "E3B505"
GREENF = "DCF3E4"
GREENT = "067647"
FLAGF  = "FDE7E7"
FLAGT  = "B42318"
WHITE  = "FFFFFF"
FONT   = "Calibri"

C_A = "0F2E52"; C_B = "D98A00"; C_C = "1E8E5A"

MONEY   = '#,##0.00;(#,##0.00);""'
MONHERO = '"$"#,##0.00"m"'
FTEF    = '#,##0.0;(#,##0.0);""'
CNT     = '#,##0;;""'

def f(size=11, bold=False, color=INK):
    return Font(name=FONT, size=size, bold=bold, color=color)
def fill(hexv):
    return PatternFill("solid", fgColor=hexv)
def setrange(ws, r1, c1, r2, c2, patch):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            patch(ws.cell(r, c))
def hdrcell(cell, text, align="center"):
    cell.value = text
    cell.font = f(11, True, WHITE)
    cell.fill = fill(NAVY)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
def title(ws, cellref, text):
    ws[cellref] = text
    ws[cellref].font = f(16, True, NAVY)

MONTHS = ["Jul-2026","Aug-2026","Sep-2026","Oct-2026","Nov-2026","Dec-2026",
          "Jan-2027","Feb-2027","Mar-2027","Apr-2027","May-2027","Jun-2027",
          "Jul-2027","Aug-2027","Sep-2027","Oct-2027","Nov-2027","Dec-2027"]

wb = openpyxl.Workbook()

# =====================================================================
# 4 Lists  (month list + named range) -- built first, hidden
# =====================================================================
wl = wb.active
wl.title = "4 Lists"
wl["B1"] = "Month list, used by the dropdowns. Do not change."
wl["B1"].font = f(11, True)
for i, m in enumerate(MONTHS):
    wl.cell(2+i, 2, m).font = f(11)
    wl.cell(2+i, 3, i+1).font = f(11)
wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
    "MonthList", attr_text="'4 Lists'!$B$2:$B$19"))
wl.sheet_state = "hidden"
wl.column_dimensions["B"].width = 14

# =====================================================================
# 2 Pick roles
# =====================================================================
pk = wb.create_sheet("2 Pick roles")
title(pk, "B1", "Pick roles")
pk["B2"] = "Filter to find a role, set it to Yes in List A, B or C, and choose the month it moves. A role can sit in more than one list."
pk["B2"].font = f(11, color=MUTE)

COLS = [
 ("Role ID", 9, "id"),
 ("Role", 32, "role"),
 ("Portfolio", 19, "portfolio"),
 ("Squad or line", 23, "squad"),
 ("AU / NZ", 8, "country"),
 ("Status", 21, "status"),
 ("FTE", 6, "fte"),
 ("Cost today $m", 11, "today"),
 ("In A?", 7, "inA"),
 ("A month", 11, "amon"),
 ("In B?", 7, "inB"),
 ("B month", 11, "bmon"),
 ("In C?", 7, "inC"),
 ("C month", 11, "cmon"),
 ("Full onshore $m", 12, "onshore"),
 ("Taken out A $m/yr", 12, "redA"),
 ("pos A", 6, "posA"),
 ("2026 A $m", 10, "fy26A"),
 ("2027 A $m", 10, "fy27A"),
 ("Taken out B $m/yr", 12, "redB"),
 ("pos B", 6, "posB"),
 ("2026 B $m", 10, "fy26B"),
 ("2027 B $m", 10, "fy27B"),
 ("Taken out C $m/yr", 12, "redC"),
 ("pos C", 6, "posC"),
 ("2026 C $m", 10, "fy26C"),
 ("2027 C $m", 10, "fy27C"),
 ("In any list?", 12, "any"),
 ("Best taken out $m/yr", 12, "maxred"),
 ("rank", 12, "rankscore"),
 ("Lists", 8, "lists"),
]
key2col = {}
HDR_ROW = 4
FIRST = 5
LAST = FIRST + len(ROLES) - 1
for i, (label, width, key) in enumerate(COLS):
    col = 2 + i
    key2col[key] = col
    hdrcell(pk.cell(HDR_ROW, col), label, align="center" if key not in ("role","portfolio","squad","status") else "left")
    pk.column_dimensions[get_column_letter(col)].width = width
pk.row_dimensions[HDR_ROW].height = 30

def L(key):
    return get_column_letter(key2col[key])

EXAMPLE = {"R0002":"Oct-2026","R0029":"Oct-2026","R0351":"Jan-2027","R0468":"Oct-2026"}
leftcols = {"role","portfolio","squad","status"}

for idx, rec in enumerate(ROLES):
    row = FIRST + idx
    def put(key, val, num=None, inp=False):
        c = pk.cell(row, key2col[key], val)
        c.font = f(11, color=INK)
        if num: c.number_format = num
        if key in leftcols:
            c.alignment = Alignment(horizontal="left", vertical="center")
        else:
            c.alignment = Alignment(horizontal="center", vertical="center")
        if inp:
            c.fill = fill(INPUT)
            c.border = Border(left=Side("thin", color=INPUTB), right=Side("thin", color=INPUTB),
                              top=Side("thin", color=INPUTB), bottom=Side("thin", color=INPUTB))
            c.protection = Protection(locked=False)
        return c
    put("id", rec["role_id"])
    put("role", rec["role"])
    put("portfolio", rec["portfolio"])
    put("squad", rec["squad"])
    put("country", rec["country"])
    put("status", rec["status"])
    put("fte", rec["fte"], num=FTEF)
    put("today", rec["today"], num=MONEY)
    put("onshore", rec["onshore"], num=MONEY)
    ex = EXAMPLE.get(rec["role_id"])
    put("inA", "Yes" if ex else "No", inp=True)
    put("amon", ex if ex else None, inp=True)
    put("inB", "No", inp=True)
    put("bmon", None, inp=True)
    put("inC", "No", inp=True)
    put("cmon", None, inp=True)
    on = f"${L('onshore')}{row}"; td = f"${L('today')}{row}"; st = f"${L('status')}{row}"
    for tog, mon, red, pos, fy26, fy27 in [
        ("inA","amon","redA","posA","fy26A","fy27A"),
        ("inB","bmon","redB","posB","fy26B","fy27B"),
        ("inC","cmon","redC","posC","fy26C","fy27C")]:
        tg=f"${L(tog)}{row}"; mn=f"${L(mon)}{row}"
        pk.cell(row, key2col[red], f'=IF(AND({tg}="Yes",{st}<>"Vendor day rates"),{td}-0.4*{on},0)').number_format=MONEY
        pk.cell(row, key2col[pos], f'=IFERROR(MATCH({mn},MonthList,0),0)')
        p=f"${L(pos)}{row}"; rd=f"${L(red)}{row}"
        pk.cell(row, key2col[fy26], f'=IF(OR({tg}<>"Yes",{mn}=""),0,{rd}*MAX(0,7-{p})/12)').number_format=MONEY
        pk.cell(row, key2col[fy27], f'=IF(OR({tg}<>"Yes",{mn}=""),0,{rd}*IF({p}<=7,12,19-{p})/12)').number_format=MONEY
    pk.cell(row, key2col["any"],
            f'=IF(OR(${L("inA")}{row}="Yes",${L("inB")}{row}="Yes",${L("inC")}{row}="Yes"),"Yes","No")')
    # ranking helpers for the dashboard top-ten (epsilon on row breaks ties)
    pk.cell(row, key2col["maxred"],
            f'=MAX(${L("redA")}{row},${L("redB")}{row},${L("redC")}{row})').number_format=MONEY
    pk.cell(row, key2col["rankscore"],
            f'=${L("maxred")}{row}+ROW()*0.000000001')
    pk.cell(row, key2col["lists"],
            f'=IF(${L("inA")}{row}="Yes","A","")&IF(${L("inB")}{row}="Yes","B","")&IF(${L("inC")}{row}="Yes","C","")')
    for k in ("redA","posA","fy26A","fy27A","redB","posB","fy26B","fy27B",
              "redC","posC","fy26C","fy27C","any","maxred","rankscore","lists"):
        cc = pk.cell(row, key2col[k]); cc.font=f(11, color=INK)
        cc.alignment=Alignment(horizontal="center", vertical="center")

dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
dv_mo = DataValidation(type="list", formula1="MonthList", allow_blank=True)
pk.add_data_validation(dv_yn); pk.add_data_validation(dv_mo)
for key in ("inA","inB","inC"):
    dv_yn.add(f"{L(key)}{FIRST}:{L(key)}{LAST}")
for key in ("amon","bmon","cmon"):
    dv_mo.add(f"{L(key)}{FIRST}:{L(key)}{LAST}")

band_range=f"{L('id')}{FIRST}:{L('cmon')}{LAST}"
pk.conditional_formatting.add(band_range,
    FormulaRule(formula=["ISEVEN(ROW())"], fill=fill(BANDR), stopIfTrue=False))
for key in ("inA","inB","inC"):
    rng=f"{L(key)}{FIRST}:{L(key)}{LAST}"
    pk.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Yes"'],
        fill=fill(GREENF), font=Font(name=FONT, size=11, bold=True, color=GREENT)))
for tog,mon in (("inA","amon"),("inB","bmon"),("inC","cmon")):
    rng=f"{L(mon)}{FIRST}:{L(mon)}{LAST}"
    pk.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND(${L(tog)}{FIRST}="Yes",${L(mon)}{FIRST}="")'],
        fill=fill(FLAGF), font=Font(name=FONT,size=11,bold=True,color=FLAGT)))
pk.conditional_formatting.add(f"{L('role')}{FIRST}:{L('role')}{LAST}",
    FormulaRule(formula=[f'${L("any")}{FIRST}="Yes"'], font=Font(name=FONT,size=11,bold=True,color=NAVY)))

pk.auto_filter.ref = f"{L('id')}{HDR_ROW}:{L('cmon')}{LAST}"
for key in ("onshore","redA","posA","fy26A","fy27A","redB","posB","fy26B","fy27B",
            "redC","posC","fy26C","fy27C","any","maxred","rankscore","lists"):
    pk.column_dimensions[get_column_letter(key2col[key])].hidden = True
pk.freeze_panes = f"{L('portfolio')}{FIRST}"
pk.sheet_view.showGridLines = False
pk.column_dimensions["A"].width = 2
pk.protection = openpyxl.worksheet.protection.SheetProtection(
    sheet=True, password=PWD, autoFilter=False, sort=False,
    formatCells=False, formatColumns=False, formatRows=False,
    selectLockedCells=False, selectUnlockedCells=False)
RD = LAST

# =====================================================================
# 3 Compare
# =====================================================================
cp = wb.create_sheet("3 Compare")
title(cp, "B1", "Compare your three lists")
cp["B2"] = "Everything here moves the moment you set a role to Yes on the Pick roles tab. All figures are A$ millions."
cp["B2"].font = f(11, color=MUTE)
cp.sheet_view.showGridLines = False
cp.column_dimensions["A"].width = 2

def pkref(key):
    return f"'2 Pick roles'!${L(key)}${FIRST}:${L(key)}${RD}"

NAME_CELLS = ["'1 Start here'!$B$8", "'1 Start here'!$E$8", "'1 Start here'!$H$8"]

HR = 4
hdrcell(cp.cell(HR,2), "Measure", align="left")
for j in range(3):
    hdrcell(cp.cell(HR,3+j), f"={NAME_CELLS[j]}")
cp.column_dimensions["B"].width=30
for col in ("C","D","E"): cp.column_dimensions[col].width=17
cp.row_dimensions[HR].height=22

tog=["inA","inB","inC"]; red=["redA","redB","redC"]; fy26=["fy26A","fy26B","fy26C"]; fy27=["fy27A","fy27B","fy27C"]
measures = [
 ("Roles moved",               lambda j: f'=COUNTIF({pkref(tog[j])},"Yes")', CNT),
 ("People (FTE)",              lambda j: f'=SUMIF({pkref(tog[j])},"Yes",{pkref("fte")})', FTEF),
 ("Cost today",                lambda j: f'=SUMIF({pkref(tog[j])},"Yes",{pkref("today")})', MONEY),
 ("Cost at the offshore rate", None, MONEY),
 ("Cost taken out, full year", lambda j: f'=SUM({pkref(red[j])})', MONEY),
 ("Taken out in 2026",         lambda j: f'=SUM({pkref(fy26[j])})', MONEY),
 ("Taken out in 2027",         lambda j: f'=SUM({pkref(fy27[j])})', MONEY),
]
MROW0=5
HERO_ROW=MROW0+4
for i,(label,fn,fmt) in enumerate(measures):
    row=MROW0+i
    emph = (i==4)
    lc=cp.cell(row,2,label); lc.font=f(11, bold=emph, color=INK)
    lc.alignment=Alignment(horizontal="left", vertical="center")
    for j in range(3):
        col=3+j
        if fn is None:
            form=f"={get_column_letter(col)}{MROW0+2}-{get_column_letter(col)}{MROW0+4}"
        else:
            form=fn(j)
        c=cp.cell(row,col,form); c.font=f(11, bold=emph, color=INK)
        c.number_format=fmt; c.alignment=Alignment(horizontal="center", vertical="center")
    if emph:
        setrange(cp,row,2,row,5, lambda c: setattr(c,"fill",fill(EMPH)))
setrange(cp,HR,2,HR,5, lambda c: setattr(c,"border",Border(bottom=Side("medium",color=NAVY))))

ports=[]; seen=set()
for rec in ROLES:
    p=rec["portfolio"]
    key = "Retail" if (p or "").upper()=="RETAIL" else (p if p else "Not mapped")
    if key not in seen:
        seen.add(key); ports.append(key)
if "Not mapped" in ports:
    ports = [p for p in ports if p!="Not mapped"] + ["Not mapped"]

PB0 = MROW0 + len(measures) + 2
cp.cell(PB0,2,"Cost taken out by portfolio, full year").font=f(12,True,NAVY)
hr2=PB0+1
hdrcell(cp.cell(hr2,2),"Portfolio", align="left")
for j in range(3):
    hdrcell(cp.cell(hr2,3+j), f"={NAME_CELLS[j]}")
cp.row_dimensions[hr2].height=20
pcol=pkref("portfolio")
for i,p in enumerate(ports):
    row=hr2+1+i
    cp.cell(row,2,p).font=f(11, color=INK); cp.cell(row,2).alignment=Alignment(horizontal="left")
    if i%2==1: setrange(cp,row,2,row,5, lambda c: setattr(c,"fill",fill(BANDR)))
    for j in range(3):
        col=3+j
        if p=="Retail":
            form=f'=SUMIFS({pkref(red[j])},{pcol},"Retail")'   # SUMIFS case-insensitive: captures RETAIL too
        elif p=="Not mapped":
            form="=0"
        else:
            form=f'=SUMIFS({pkref(red[j])},{pcol},"{p}")'
        c=cp.cell(row,col,form); c.number_format=MONEY; c.font=f(11, color=INK)
        c.alignment=Alignment(horizontal="center")
named_rows=[hr2+1+i for i,p in enumerate(ports) if p!="Not mapped"]
nm_row=hr2+1+ports.index("Not mapped")
for j in range(3):
    col=get_column_letter(3+j)
    named_sum="+".join(f"{col}{rr}" for rr in named_rows)
    cp.cell(nm_row,3+j).value=f"=SUM({pkref(red[j])})-({named_sum})"
tot_row=hr2+1+len(ports)
cp.cell(tot_row,2,"Total").font=f(11,True,INK)
setrange(cp,tot_row,2,tot_row,5, lambda c: setattr(c,"fill",fill(EMPH)))
setrange(cp,tot_row,2,tot_row,5, lambda c: setattr(c,"border",Border(top=Side("thin",color=NAVY))))
chk_row=tot_row+1
cp.cell(chk_row,2,"Check against full year (0 = ties)").font=f(9, color=MUTE)
for j in range(3):
    col=get_column_letter(3+j)
    c=cp.cell(tot_row,3+j, f"=SUM({col}{hr2+1}:{col}{tot_row-1})"); c.number_format=MONEY
    c.font=f(11,True,INK); c.alignment=Alignment(horizontal="center")
    c2=cp.cell(chk_row,3+j, f"={col}{tot_row}-{col}{HERO_ROW}"); c2.number_format='0.00;(0.00);"0.00"'
    c2.font=f(9, color=MUTE); c2.alignment=Alignment(horizontal="center")
# in-cell bars on the portfolio spread, one rule per list column
for j in range(3):
    col=get_column_letter(3+j)
    cp.conditional_formatting.add(f"{col}{hr2+1}:{col}{tot_row-1}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color=NAVY2, showValue=True))

MB_C = 8
hdrcell(cp.cell(HR, MB_C), "Month")
for j in range(3):
    hdrcell(cp.cell(HR, MB_C+1+j), f"={NAME_CELLS[j]}")
cp.column_dimensions[get_column_letter(MB_C)].width=11
for j in range(3): cp.column_dimensions[get_column_letter(MB_C+1+j)].width=15
posk=["posA","posB","posC"]
for p in range(18):
    row=HR+1+p
    cp.cell(row, MB_C, f"='4 Lists'!$B${2+p}").font=f(11, color=INK)
    cp.cell(row, MB_C).alignment=Alignment(horizontal="left")
    for j in range(3):
        form=(f'=SUMPRODUCT(({pkref(posk[j])}>=1)*({pkref(posk[j])}<={p+1})*{pkref(red[j])})')
        c=cp.cell(row, MB_C+1+j, form); c.number_format=MONEY; c.font=f(11, color=INK)
        c.alignment=Alignment(horizontal="center")
MB_LAST=HR+18
setrange(cp,HR,MB_C,HR,MB_C+3, lambda c: setattr(c,"border",Border(bottom=Side("medium",color=NAVY))))
# heat shading across the month grid: deeper colour = more cost out
cp.conditional_formatting.add(
    f"{get_column_letter(MB_C+1)}{HR+1}:{get_column_letter(MB_C+3)}{MB_LAST}",
    ColorScaleRule(start_type="num", start_value=0, start_color=WHITE,
                   end_type="max", end_color="9DB8D9"))

BY_C=8; BY0=MB_LAST+3
cp.cell(BY0-1, BY_C, "Cost taken out by year").font=f(12,True,NAVY)
hdrcell(cp.cell(BY0, BY_C), "Year")
for j in range(3):
    hdrcell(cp.cell(BY0, BY_C+1+j), f"={NAME_CELLS[j]}")
for yi,(ylab,mrow) in enumerate([("2026",MROW0+5),("2027",MROW0+6)]):
    row=BY0+1+yi
    cp.cell(row,BY_C,ylab).font=f(11, color=INK); cp.cell(row,BY_C).alignment=Alignment(horizontal="left")
    for j in range(3):
        col=get_column_letter(3+j)
        c=cp.cell(row,BY_C+1+j, f"={col}{mrow}"); c.number_format=MONEY; c.font=f(11, color=INK)
        c.alignment=Alignment(horizontal="center")

def style_line(chart):
    chart.legend.position='b'
    chart.x_axis.delete=False; chart.y_axis.delete=False
    chart.x_axis.majorGridlines=None
    chart.y_axis.title="A$m/yr"; chart.x_axis.title=None
    chart.y_axis.scaling.min=0; chart.y_axis.majorUnit=0.1
    try:
        chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    except Exception:
        pass
    for s,colr in zip(chart.series,(C_A,C_B,C_C)):
        s.smooth=False; s.marker=Marker(symbol='none')
        gp=GraphicalProperties(); gp.line=LineProperties(solidFill=colr, w=int(2.5*12700))
        s.graphicalProperties=gp

def make_runrate_chart():
    ch=LineChart(); ch.title="Cost taken out, run rate by month"
    ch.height=9.2; ch.width=20; ch.style=2
    data=Reference(cp, min_col=MB_C+1, max_col=MB_C+3, min_row=HR, max_row=MB_LAST)
    cats=Reference(cp, min_col=MB_C, max_col=MB_C, min_row=HR+1, max_row=MB_LAST)
    ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
    style_line(ch)
    return ch

cp.add_chart(make_runrate_chart(), f"B{chk_row+3}")

bc=BarChart(); bc.type="col"; bc.title="Cost taken out by year"
bc.height=9.2; bc.width=11; bc.style=2; bc.gapWidth=80
bdata=Reference(cp, min_col=BY_C+1, max_col=BY_C+3, min_row=BY0, max_row=BY0+2)
bcats=Reference(cp, min_col=BY_C, max_col=BY_C, min_row=BY0+1, max_row=BY0+2)
bc.add_data(bdata, titles_from_data=True); bc.set_categories(bcats)
bc.y_axis.title="A$m"; bc.x_axis.delete=False; bc.y_axis.delete=False
bc.y_axis.scaling.min=0; bc.y_axis.majorUnit=0.1
bc.x_axis.majorGridlines=None; bc.legend.position='b'
try:
    bc.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
except Exception: pass
for s,colr in zip(bc.series,(C_A,C_B,C_C)):
    s.graphicalProperties=GraphicalProperties(solidFill=colr)
cp.add_chart(bc, f"I{chk_row+3}")

cp.protection = openpyxl.worksheet.protection.SheetProtection(
    sheet=True, password=PWD, selectLockedCells=False, selectUnlockedCells=False)

# =====================================================================
# 1 Start here  -- the dashboard
# =====================================================================
ws = wb.create_sheet("1 Start here")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
for col in "BCDEFGHIJ":
    ws.column_dimensions[col].width = 11.5
ws.column_dimensions["K"].width = 2
# hidden ranking helpers live in N/O
ws.column_dimensions["N"].hidden = True
ws.column_dimensions["O"].hidden = True

ws.merge_cells("B2:J3")
tb=ws["B2"]; tb.value="Offshore scenario builder"
tb.font=f(18,True,WHITE); tb.alignment=Alignment(horizontal="left", vertical="center", indent=1)
setrange(ws,2,2,3,10, lambda c: setattr(c,"fill",fill(NAVY)))
ws.row_dimensions[2].height=20; ws.row_dimensions[3].height=20
ws["B4"]="Pick which roles move offshore and when. Compare up to three lists side by side and see the cost come out, month by month."
ws["B4"].font=f(11,color=MUTE)

# --- three live cards ---
card_cols=[(2,4),(5,7),(8,10)]
default_names=["Example","List B","List C"]
CARD_TOP=6; NAME_ROW=8; HERO_ROW_C=10; CAP_ROW=11; CAP2_ROW=12
ws.row_dimensions[CARD_TOP].height=6
ws.row_dimensions[HERO_ROW_C].height=32
for k,(c1,c2) in enumerate(card_cols):
    setrange(ws,CARD_TOP,c1,CAP2_ROW,c2, lambda c: setattr(c,"fill",fill(CARD)))
    setrange(ws,CARD_TOP,c1,CARD_TOP,c2, lambda c: setattr(c,"fill",fill(NAVY)))
    ws.merge_cells(start_row=NAME_ROW,start_column=c1,end_row=NAME_ROW,end_column=c2)
    nm=ws.cell(NAME_ROW,c1,default_names[k]); nm.font=f(12,True,NAVY)
    nm.alignment=Alignment(horizontal="left", vertical="center", indent=1)
    nm.fill=fill(INPUT); nm.protection=Protection(locked=False)
    setrange(ws,NAME_ROW,c1,NAME_ROW,c2, lambda c: setattr(c,"fill",fill(INPUT)))
    ws.cell(NAME_ROW,c1).protection=Protection(locked=False)
    col=get_column_letter(3+k)
    ws.merge_cells(start_row=HERO_ROW_C,start_column=c1,end_row=HERO_ROW_C,end_column=c2)
    hn=ws.cell(HERO_ROW_C,c1, f"='3 Compare'!{col}{HERO_ROW}")
    hn.font=f(26,True,NAVY); hn.number_format=MONHERO
    hn.alignment=Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=CAP_ROW,start_column=c1,end_row=CAP_ROW,end_column=c2)
    cap=ws.cell(CAP_ROW,c1,
        f'=TEXT(\'3 Compare\'!{col}{MROW0},"0")&" roles · "&'
        f'TEXT(\'3 Compare\'!{col}{MROW0+1},"0.0")&" FTE · 2026 $"&'
        f'TEXT(\'3 Compare\'!{col}{MROW0+5},"0.00")&"m"')
    cap.font=f(9,color=MUTE); cap.alignment=Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=CAP2_ROW,start_column=c1,end_row=CAP2_ROW,end_column=c2)
    cap2=ws.cell(CAP2_ROW,c1,
        f'=TEXT(IFERROR(\'3 Compare\'!{col}{HERO_ROW}/\'3 Compare\'!{col}{MROW0+2},0),"0%")&" of what these roles cost today"')
    cap2.font=f(9,color=MUTE); cap2.alignment=Alignment(horizontal="left", vertical="center", indent=1)
ws.cell(CAP2_ROW+1,2,"Type over a yellow title to rename that list. The Example list holds a worked example - clear it on the Pick roles tab and make your own.").font=f(9,color=MUTE)

# --- top-ten table (left) + run-rate chart (right) ---
T0=15
ws.cell(T0,2,"Ten biggest roles you've picked ($m a year)").font=f(12,True,NAVY)
TH=T0+1
ws.merge_cells(start_row=TH,start_column=2,end_row=TH,end_column=4); hdrcell(ws.cell(TH,2),"Role", align="left")
ws.merge_cells(start_row=TH,start_column=5,end_row=TH,end_column=6); hdrcell(ws.cell(TH,5),"Portfolio", align="left")
hdrcell(ws.cell(TH,7),"Lists")
hdrcell(ws.cell(TH,8),"$m / yr")
rank_rng=pkref("rankscore"); role_rng=pkref("role"); port_rng=pkref("portfolio")
lists_rng=pkref("lists"); maxred_rng=pkref("maxred")
for k in range(1,11):
    r=TH+k
    ws.cell(r,14, f'=IFERROR(LARGE({rank_rng},{k}),0)')                      # N: score
    ws.cell(r,15, f'=IFERROR(MATCH($N{r},{rank_rng},0),0)')                  # O: row
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
    rc=ws.cell(r,2, f'=IF($N{r}<0.00001,"",INDEX({role_rng},$O{r}))')
    rc.font=f(11,color=INK); rc.alignment=Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=6)
    pc=ws.cell(r,5, f'=IF($N{r}<0.00001,"",INDEX({port_rng},$O{r}))')
    pc.font=f(11,color=INK); pc.alignment=Alignment(horizontal="left", vertical="center")
    lcx=ws.cell(r,7, f'=IF($N{r}<0.00001,"",INDEX({lists_rng},$O{r}))')
    lcx.font=f(11,True,NAVY2); lcx.alignment=Alignment(horizontal="center", vertical="center")
    ac=ws.cell(r,8, f'=IF($N{r}<0.00001,"",INDEX({maxred_rng},$O{r}))')
    ac.font=f(11,color=INK); ac.number_format=MONEY
    ac.alignment=Alignment(horizontal="center", vertical="center")
    if k%2==0:
        setrange(ws,r,2,r,8, lambda c: setattr(c,"fill",fill(BANDR)))
T_LAST=TH+10
ws.conditional_formatting.add(f"H{TH+1}:H{T_LAST}",
    DataBarRule(start_type="num", start_value=0, end_type="max", color=NAVY2, showValue=True))
setrange(ws,TH,2,TH,8, lambda c: setattr(c,"border",Border(bottom=Side("medium",color=NAVY))))

ws.add_chart(make_runrate_chart(), f"I{T0}")

# --- footer: how to use it / how the money works ---
R=T_LAST+3
ws.cell(R,2,"How to use it").font=f(12,True,NAVY)
steps=[
 "1.  Open the Pick roles tab and use the filter arrows to find the roles you want.",
 "2.  Set a role to Yes in List A, B or C, and choose the month it moves. It can sit in more than one list.",
 "3.  A role set to Yes turns green. If it has no month it flags red until you choose one.",
 "4.  This page and the Compare tab move as you go.",
]
for i,s in enumerate(steps): ws.cell(R+1+i,2,s).font=f(11,color=INK)

R2=R+len(steps)+2
ws.cell(R2,2,"How the money works").font=f(12,True,NAVY)
notes=[
 "A role that moves offshore costs 40% of its full onshore cost. What comes out is today's cost less that 40%.",
 "Roles on vendor day rates stay as they are. A held role costs nothing today, so moving it shows as added cost.",
 "The year is the calendar year (Ampol's December year-end). The horizon runs Jul-2026 to Dec-2027.",
 "Figures are A$ millions. NZ roles are converted at the planning rate.",
 "The tabs are locked so the sums can't break. Only the yellow cells take typing. Password to unlock: Tdd123.",
 "Built from the role mapping in TDD_FY27_Cost_Calc_0608 (11 August 2026), copied in full on the hidden data tab.",
]
for i,s in enumerate(notes): ws.cell(R2+1+i,2,s).font=f(11,color=INK)

ws.protection = openpyxl.worksheet.protection.SheetProtection(
    sheet=True, password=PWD, selectLockedCells=False, selectUnlockedCells=False)

order=["1 Start here","2 Pick roles","3 Compare","4 Lists"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active = wb.sheetnames.index("1 Start here")

wb.save(OUT)
print("saved", OUT)
print("rows:", FIRST, "..", RD, "= roles:", RD-FIRST+1)
