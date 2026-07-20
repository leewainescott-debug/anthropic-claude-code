#!/usr/bin/env python3
"""Builder for Ampol Retail — TDD Cost Calculator."""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/Ampol_Retail_TDD_Cost_Calculator.xlsx"

# ---------- palette ----------
BRAND_DARK  = "0B2E3C"   # dark teal (header band)
BRAND_MED   = "1C5A66"   # teal (section headers)
BRAND_LT    = "D9E7EA"   # pale teal (table headers / bands)
BRAND_LT2   = "EDF4F5"   # very pale teal (zebra)
ACCENT      = "E8A87C"   # peach accent
INPUT_FILL  = "FFF6D5"   # pale yellow (input cells)
TOTAL_FILL  = "CFE0E4"   # total band
GRAND_FILL  = "0B2E3C"   # grand total band (dark)
POS_FILL    = "E2EFDA"   # green (surplus)
NEG_FILL    = "FBE4D5"   # red-ish (shortfall)
WHITE       = "FFFFFF"
GREY_TX     = "5A6B6F"

# fonts
def F(sz=10, b=False, color="1A1A1A", italic=False, name="Arial"):
    return Font(name=name, size=sz, bold=b, color=color, italic=italic)
INPUT_FONT   = F(10, False, "0000CC")     # blue = user input
LINK_FONT    = F(10, False, "007A33")     # green = cross-sheet link
FORMULA_FONT = F(10, False, "1A1A1A")     # black = formula

def fill(c): return PatternFill("solid", fgColor=c)
thin = Side(style="thin", color="B8C9CC")
med  = Side(style="medium", color=BRAND_MED)
def box(l=thin,r=thin,t=thin,b=thin): return Border(left=l,right=r,top=t,bottom=b)
ALL = box()
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENnw= Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LFTt= Alignment(horizontal="left", vertical="top", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")

MONEY = '$#,##0.00;($#,##0.00);"–"'
MONEY0= '$#,##0.0;($#,##0.0);"–"'
PCT   = '0%'
NUM   = '#,##0.0'

wb = Workbook()

# =====================================================================
# SHEET 2: Squads (reference lookup)
# =====================================================================
sq = wb.active
sq.title = "Squads"
sq.sheet_view.showGridLines = False

SQ_ROWS = [
    # type, size, archetype, roles, onshore, defsupport
    ("Engineering","S","Lean build squad",4.5,0.8,0.2),
    ("Engineering","M","Standard Engineering Squad",9,2.0,0.2),
    ("Engineering","L","High-scale squad",12,2.7,0.2),
    ("Configuration / Integration","XS","MS Squad",1,3.2,0.2),
    ("Configuration / Integration","S","Simple Configuration Squad",6,0.4,0.2),
    ("Configuration / Integration","M","Standard Configuration Squad",9,1.4,0.2),
    ("Configuration / Integration","L","Complex Integration Squad",14,2.1,0.2),
    ("Product","S","Lean Product Squad",5.5,1.3,0.2),
    ("Product","M","Balanced Product Squad",8.5,1.9,0.2),
    ("Product","L","Scaled Product Squad",12,2.8,0.2),
    ("Operations","S","Lean operations squad",4,0.8,0.2),
    ("Operations","M","Balanced operations squad",6,1.4,0.2),
    ("Operations","L","High demand operations squad",10,2.3,0.2),
    ("Enterprise Data and Insights","S","Lean Data squad",4,0.9,0.2),
    ("Enterprise Data and Insights","M","Standard Data squad",7.5,1.7,0.2),
    ("Enterprise Data and Insights","L","Advanced Data / AI squad",9,2.0,0.2),
    ("Build and Run","S","Lean build and run squad",4,0.9,0.2),
    ("Build and Run","M","Standard build and run squad",9,2.0,0.2),
    ("Build and Run","L","Scaled build and run squad",13,2.8,0.2),
]
# title
sq.merge_cells("B2:I2")
sq["B2"]="Squad Reference — Cost Library"; sq["B2"].font=F(15,True,BRAND_DARK); sq["B2"].alignment=LFT
sq.merge_cells("B3:I3")
sq["B3"]="Standard squad archetypes and onshore/offshore cost per squad, per annum ($m). Offshore cost = onshore × offshore efficiency (0.4). The calculator looks up cost by Squad Type + Size."
sq["B3"].font=F(9,italic=True,color=GREY_TX); sq["B3"].alignment=LFT
sq.row_dimensions[3].height=28

hdr = ["Squad Type","Size","Archetype","# of roles","Cost onshore ($m)","Offshore efficiency","Cost offshore ($m)","Default support %","Key"]
hr=5
for j,h in enumerate(hdr):
    c=sq.cell(hr,2+j,h); c.fill=fill(BRAND_MED); c.font=F(9,True,WHITE); c.alignment=CEN; c.border=ALL
r=hr+1
sq_key_first=r
for (typ,size,arch,roles,onsh,dsup) in SQ_ROWS:
    sq.cell(r,2,typ).font=F(10); sq.cell(r,2).alignment=LFT
    sq.cell(r,3,size).alignment=CENnw
    sq.cell(r,4,arch).alignment=LFT
    sq.cell(r,5,roles).alignment=CENnw; sq.cell(r,5).number_format=NUM
    oc=sq.cell(r,6,onsh); oc.number_format=MONEY; oc.font=F(10)
    ec=sq.cell(r,7,0.4); ec.number_format=PCT; ec.alignment=CENnw
    fc=sq.cell(r,8,f"=F{r}*G{r}"); fc.number_format=MONEY
    dc=sq.cell(r,9,dsup); dc.number_format=PCT; dc.alignment=CENnw
    kc=sq.cell(r,10,f'=B{r}&"|"&C{r}'); kc.font=F(9,color="9AA9AC")
    for col in range(2,11):
        cc=sq.cell(r,col); cc.border=ALL
        if (r-sq_key_first)%2==1:
            if cc.fill.fgColor.rgb in (None,"00000000"): cc.fill=fill(BRAND_LT2)
    r+=1
sq_key_last=r-1
widths={"A":2,"B":30,"C":7,"D":34,"E":11,"F":16,"G":15,"H":16,"I":15,"J":22}
for col,w in widths.items(): sq.column_dimensions[col].width=w
sq.freeze_panes="B6"

# named-ish ranges via explicit refs (used by calculator)
SQ_KEY = f"Squads!$J${sq_key_first}:$J${sq_key_last}"
SQ_ON  = f"Squads!$F${sq_key_first}:$F${sq_key_last}"
SQ_OFF = f"Squads!$H${sq_key_first}:$H${sq_key_last}"

# =====================================================================
# SHEET 3: Data Config
# =====================================================================
dc = wb.create_sheet("Data Config")
dc.sheet_view.showGridLines=False
dc.merge_cells("B2:H2")
dc["B2"]="Data Config — Budget Allocation & Overheads"; dc["B2"].font=F(15,True,BRAND_DARK); dc["B2"].alignment=LFT
dc.merge_cells("B3:H3")
dc["B3"]="TDD budget allocation by area ($m) and the role-based build-up of portfolio & platform overheads. The calculator pulls Ampol Retail budget and the overhead subtotals from here."
dc["B3"].font=F(9,italic=True,color=GREY_TX); dc["B3"].alignment=LFT
dc.row_dimensions[3].height=28

# --- TDD Budget Allocation table (B5..E..) ---
dc.cell(5,2,"TDD Budget Allocation ($m)").font=F(11,True,BRAND_MED)
for j,h in enumerate(["Area","AU","NZ","Total"]):
    c=dc.cell(6,2+j,h); c.fill=fill(BRAND_MED); c.font=F(9,True,WHITE); c.alignment=CEN; c.border=ALL
ALLOC=[("COE - Strat Arch",1.5,0.5),("COE - Cyber",1.5,0.5),("COE - Trans",1.5,0.5),
       ("COE - BP",1.5,0.5),("COE - Data",1.5,0.5),
       ("Ampol Retail",2.5,None),("Z Retail",None,2.5),("Ampol Customer",2.5,None),
       ("Z Customer",0,2.5),("CF",2.5,None),("ES & B2B",2.5,None),("Infra",2.5,None),
       ("P&C",1,1),("Finance",1,1),("Legal",0,None),("TDD",4.5,1),("TDD Data",2.5,1),
       ("TDD Cyber",1,0.5),("EG",0,None),("EGI",0,None)]
r=7
dc_alloc_first=r
retail_budget_row=None
for (area,au,nz) in ALLOC:
    ac=dc.cell(r,2,area); ac.font=F(10,b=(area=="Ampol Retail")); ac.alignment=LFT
    bc=dc.cell(r,3,au); bc.number_format=MONEY; bc.font=F(10)
    cc=dc.cell(r,4,nz); cc.number_format=MONEY; cc.font=F(10)
    tc=dc.cell(r,5,f"=SUM(C{r}:D{r})"); tc.number_format=MONEY; tc.font=F(10,True)
    if area=="Ampol Retail":
        retail_budget_row=r
        for col in range(2,6): dc.cell(r,col).fill=fill(INPUT_FILL)
    for col in range(2,6): dc.cell(r,col).border=ALL
    r+=1
dc_alloc_last=r-1
dc.cell(r,2,"Total Cost").font=F(10,True); dc.cell(r,2).alignment=LFT
dc.cell(r,3,f"=SUM(C{dc_alloc_first}:C{dc_alloc_last})").number_format=MONEY
dc.cell(r,4,f"=SUM(D{dc_alloc_first}:D{dc_alloc_last})").number_format=MONEY
dc.cell(r,5,f"=SUM(E{dc_alloc_first}:E{dc_alloc_last})").number_format=MONEY
for col in range(2,6):
    dc.cell(r,col).border=ALL; dc.cell(r,col).fill=fill(TOTAL_FILL); dc.cell(r,col).font=F(10,True)
    if col>2: dc.cell(r,col).font=F(10,True)
DC_RETAIL_BUDGET = f"'Data Config'!$E${retail_budget_row}"   # 2.5

# --- Overheads build-up (right side, H..L) ---
oc0=8  # column H
dc.cell(5,oc0,"Overhead Build-up — TDD funded ($m)").font=F(11,True,BRAND_MED)
# Portfolio overhead
for j,h in enumerate(["Portfolio role","Cost ($m)","Alloc %","Portfolio cost ($m)"]):
    c=dc.cell(6,oc0+j,h); c.fill=fill(BRAND_MED); c.font=F(9,True,WHITE); c.alignment=CEN; c.border=ALL
PORT=[("Head of Tech",0.275,0.3),("Business Partner",0.43,0.4),("Domain Architect",0.28,0.5),("Leadership Overhead",1.0,0.3)]
r=7; pf_first=r
for (role,cost,pct) in PORT:
    dc.cell(r,oc0,role).alignment=LFT; dc.cell(r,oc0).font=F(10)
    dc.cell(r,oc0+1,cost).number_format=MONEY; dc.cell(r,oc0+1).font=F(10)
    dc.cell(r,oc0+2,pct).number_format=PCT; dc.cell(r,oc0+2).alignment=CENnw
    dc.cell(r,oc0+3,f"={get_column_letter(oc0+1)}{r}*{get_column_letter(oc0+2)}{r}").number_format=MONEY
    for k in range(4): dc.cell(r,oc0+k).border=ALL
    r+=1
pf_last=r-1
dc.cell(r,oc0,"Portfolio overhead subtotal").font=F(10,True); dc.cell(r,oc0).alignment=LFT
dc.cell(r,oc0+3,f"=SUM({get_column_letter(oc0+3)}{pf_first}:{get_column_letter(oc0+3)}{pf_last})").number_format=MONEY
dc.cell(r,oc0+3).font=F(10,True)
for k in range(4): dc.cell(r,oc0+k).border=ALL; dc.cell(r,oc0+k).fill=fill(TOTAL_FILL)
PORT_OH_ROW=r
DC_PORT_OH=f"'Data Config'!${get_column_letter(oc0+3)}${r}"   # 0.6945
r+=2
# Platform overhead
for j,h in enumerate(["Platform role","Cost ($m)","Alloc %","Platform cost ($m)"]):
    c=dc.cell(r,oc0+j,h); c.fill=fill(BRAND_MED); c.font=F(9,True,WHITE); c.alignment=CEN; c.border=ALL
r+=1
PLAT=[("Delivery Manager",0.28,0.3),("Tech Manager",0.27,0.3)]
plf_first=r
for (role,cost,pct) in PLAT:
    dc.cell(r,oc0,role).alignment=LFT; dc.cell(r,oc0).font=F(10)
    dc.cell(r,oc0+1,cost).number_format=MONEY; dc.cell(r,oc0+1).font=F(10)
    dc.cell(r,oc0+2,pct).number_format=PCT; dc.cell(r,oc0+2).alignment=CENnw
    dc.cell(r,oc0+3,f"={get_column_letter(oc0+1)}{r}*{get_column_letter(oc0+2)}{r}").number_format=MONEY
    for k in range(4): dc.cell(r,oc0+k).border=ALL
    r+=1
plf_last=r-1
dc.cell(r,oc0,"Platform overhead subtotal").font=F(10,True); dc.cell(r,oc0).alignment=LFT
dc.cell(r,oc0+3,f"=SUM({get_column_letter(oc0+3)}{plf_first}:{get_column_letter(oc0+3)}{plf_last})").number_format=MONEY
dc.cell(r,oc0+3).font=F(10,True)
for k in range(4): dc.cell(r,oc0+k).border=ALL; dc.cell(r,oc0+k).fill=fill(TOTAL_FILL)
PLAT_OH_ROW=r
DC_PLAT_OH=f"'Data Config'!${get_column_letter(oc0+3)}${r}"   # 0.165
r+=1
dc.cell(r,oc0,"Portfolio + platform overhead, TDD").font=F(10,True,BRAND_DARK); dc.cell(r,oc0).alignment=LFT
dc.cell(r,oc0+3,f"={get_column_letter(oc0+3)}{PORT_OH_ROW}+{get_column_letter(oc0+3)}{PLAT_OH_ROW}").number_format=MONEY
dc.cell(r,oc0+3).font=F(10,True)
for k in range(4): dc.cell(r,oc0+k).border=ALL; dc.cell(r,oc0+k).fill=fill(BRAND_LT)

for col,w in {"A":2,"B":20,"C":9,"D":9,"E":10,"F":3,"G":2,"H":26,"I":11,"J":9,"K":17}.items():
    dc.column_dimensions[col].width=w

# =====================================================================
# SHEET 4: FY26 Budget (from presentation pack)
# =====================================================================
fb = wb.create_sheet("FY26 Budget")
fb.sheet_view.showGridLines=False
fb.merge_cells("B2:H2")
fb["B2"]="FY2026 Technology Budget — by Business Segment"; fb["B2"].font=F(15,True,BRAND_DARK); fb["B2"].alignment=LFT
fb.merge_cells("B3:H3")
fb["B3"]="FY2026 budget ($m) pulled from the TDD presentation pack. Total = Lights On + Initiatives (Project OpEx) + Depreciation + Significant Items + CapEx."
fb["B3"].font=F(9,italic=True,color=GREY_TX); fb["B3"].alignment=LFT
fb.row_dimensions[3].height=28
# segment, lightson, initiatives, depreciation, sigitems, capex
SEG=[("Retail",13.925855,1.31,0.0,6.845,11.8),
     ("Z-Energy",52.22526,1.44624,11.306521,0.0,11.725494),
     ("Commercial Fuels",13.655581,5.861721,0.0,14.219208,6.797713),
     ("Energy Solutions",4.041939,0.97,0.0,0.0,2.997585),
     ("Infrastructure",3.594596,0.763362,0.0,0.0,4.891418),
     ("Finance & Other Corp",2.945287,0.935487,0.0,0.0,0.0),
     ("People & Culture",2.148038,1.975,0.0,1.36,0.33),
     ("Customer",1.691533,0.0,0.0,5.0,5.0)]
hdr=["Business Segment","Lights On","Initiatives\n(Project OpEx)","Depreciation","Significant\nItems","CapEx","Total FY26"]
hr=5
for j,h in enumerate(hdr):
    c=fb.cell(hr,2+j,h); c.fill=fill(BRAND_MED); c.font=F(9,True,WHITE); c.alignment=CEN; c.border=ALL
fb.row_dimensions[hr].height=30
r=hr+1; seg_first=r
for i,(seg,lo,ini,dep,si,cx) in enumerate(SEG):
    sc=fb.cell(r,2,seg); sc.alignment=LFT; sc.font=F(10,b=(seg=="Retail"))
    for k,val in enumerate([lo,ini,dep,si,cx]):
        cc=fb.cell(r,3+k,val); cc.number_format=MONEY; cc.font=F(10)
    fb.cell(r,8,f"=SUM(C{r}:G{r})").number_format=MONEY; fb.cell(r,8).font=F(10,True)
    for col in range(2,9):
        cc=fb.cell(r,col); cc.border=ALL
        if seg=="Retail": cc.fill=fill(BRAND_LT)
        elif i%2==1: cc.fill=fill(BRAND_LT2)
    r+=1
seg_last=r-1
# total row
fb.cell(r,2,"Total Technology").font=F(10,True); fb.cell(r,2).alignment=LFT
for k in range(5):
    col=3+k
    fb.cell(r,col,f"=SUM({get_column_letter(col)}{seg_first}:{get_column_letter(col)}{seg_last})").number_format=MONEY
    fb.cell(r,col).font=F(10,True)
fb.cell(r,8,f"=SUM(C{r}:G{r})").number_format=MONEY; fb.cell(r,8).font=F(10,True)
for col in range(2,9): fb.cell(r,col).border=ALL; fb.cell(r,col).fill=fill(TOTAL_FILL)
fb_total_row=r
FB_RETAIL_ROW=seg_first  # Retail is first
for col,w in {"A":2,"B":22,"C":12,"D":13,"E":13,"F":12,"G":12,"H":13}.items():
    fb.column_dimensions[col].width=w
fb.freeze_panes="B6"
# note
fb.cell(r+2,2,"Source: TDD presentation pack, 'For Presentation Pack' — FY2026 Budget by segment. Depreciation shown where separately reported in the pack.").font=F(8,italic=True,color=GREY_TX)
fb.merge_cells(start_row=r+2,start_column=2,end_row=r+2,end_column=8)

# =====================================================================
# SHEET 1: Ampol Retail (main calculator)  -- placed first
# =====================================================================
ca = wb.create_sheet("Ampol Retail", 0)
ca.sheet_view.showGridLines=False

# ---- data validations ----
dv_type = DataValidation(type="list",
    formula1='"Engineering,Configuration / Integration,Product,Operations,Enterprise Data and Insights,Build and Run"',
    allow_blank=True); dv_type.error="Pick a squad type from the list"; dv_type.errorTitle="Squad Type"
dv_size = DataValidation(type="list", formula1='"XS,S,M,L"', allow_blank=True)
dv_shore= DataValidation(type="list", formula1='"Onshore,Offshore"', allow_blank=True)
dv_sup  = DataValidation(type="list", formula1='"0,0.2,1"', allow_blank=True)
for dv in (dv_type,dv_size,dv_shore,dv_sup): ca.add_data_validation(dv)

def money(c): c.number_format=MONEY; return c

# ---------- Title band ----------
ca.merge_cells("B2:I2")
t=ca["B2"]; t.value="Ampol Retail  —  TDD Cost Calculator"; t.font=F(18,True,WHITE); t.alignment=Alignment(horizontal="left",vertical="center")
for col in range(2,10): ca.cell(2,col).fill=fill(BRAND_DARK)
ca.row_dimensions[2].height=32
ca.merge_cells("B3:I3")
ca["B3"]="Cost build-up: Squad → Platform → Portfolio, split into TDD-funded and Other. Compared against the TDD lights-on budget to reveal the funding gap."
ca["B3"].font=F(10,italic=True,color=WHITE); ca["B3"].alignment=Alignment(horizontal="left",vertical="center")
for col in range(2,10): ca.cell(3,col).fill=fill(BRAND_MED)
ca.row_dimensions[3].height=22
ca.merge_cells("B4:I4")
ca["B4"]="All figures $m per annum • FY2026 • Cells shaded yellow with blue text are your selections"
ca["B4"].font=F(9,color=GREY_TX); ca["B4"].alignment=LFT

# ---------- How to use ----------
ca.merge_cells("B6:I6")
h=ca["B6"]; h.value=("HOW TO USE  —  For each squad choose: ①Squad Type  ②Size (XS/S/M/L)  ③On/Off (Onshore or Offshore)  ④Support %.  "
 "Total Squad Cost is looked up from the Squads sheet; Cost TDD = Total × Support %; Cost Other = Total × (1 − Support %). "
 "Platforms add a $0.165m overhead; the Portfolio adds a $0.6945m overhead. Support-% options are 0 / 20% / 100%.")
h.font=F(9,color="30414A"); h.alignment=LFTt; h.fill=fill(BRAND_LT2); h.border=ALL
for col in range(2,10): ca.cell(6,col).fill=fill(BRAND_LT2); ca.cell(6,col).border=box(thin,thin,thin,thin)
ca.row_dimensions[6].height=46

def section(r,text):
    ca.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9)
    c=ca.cell(r,2,text); c.font=F(12,True,WHITE); c.alignment=Alignment(horizontal="left",vertical="center")
    for col in range(2,10): ca.cell(r,col).fill=fill(BRAND_MED)
    ca.row_dimensions[r].height=22

def tbl_header(r, labels):
    # labels: dict col->text ; cols 2..9
    for col,txt in labels.items():
        c=ca.cell(r,col,txt); c.fill=fill(BRAND_LT); c.font=F(9,True,BRAND_DARK); c.alignment=CEN; c.border=ALL
    for col in range(2,10):
        if col not in labels:
            ca.cell(r,col).fill=fill(BRAND_LT); ca.cell(r,col).border=ALL

# =====================================================================
# BUILD PLATFORM DETAIL FIRST (row 42+) to collect refs
# =====================================================================
DEFAULTS={
 "POS":("Configuration / Integration","M","Onshore",0.2),
 "Payments":("Configuration / Integration","S","Onshore",0.2),
 "Retail Operations":("Operations","M","Onshore",0.2),
 "Deployment":("Build and Run","S","Onshore",0.2),
 "Merchandising & Supply Chain":("Engineering","M","Onshore",0.2),
 "Pricing & WFM":("Product","M","Onshore",0.2),
 "AmPOS":("Engineering","L","Onshore",0.2),
 "Network & QSR":("Engineering","L","Onshore",0.2),
}
PLATFORMS=[
 ("Store Operations",["POS","Payments","Retail Operations","Deployment"]),
 ("Merchandising / Supply Chain",["Merchandising & Supply Chain"]),
 ("Pricing & WFM",["Pricing & WFM"]),
 ("AmPOS",["AmPOS"]),
 ("Network / QSR",["Network & QSR"]),
]
DETAIL_HDR={2:"Squad",3:"Squad Type",4:"Size",5:"On/Off",6:"Support %",7:"Total Squad Cost ($m)",8:"Cost TDD ($m)",9:"Cost Other ($m)"}

pr=29
section(27,"④  PLATFORM DETAIL  —  select squad type, size, shore & support for every squad")
plat_refs=[]
for pidx,(pname,squads) in enumerate(PLATFORMS,1):
    # platform header
    ca.merge_cells(start_row=pr,start_column=2,end_row=pr,end_column=9)
    ph=ca.cell(pr,2,f"Platform {pidx}  —  {pname}"); ph.font=F(11,True,WHITE)
    ph.alignment=Alignment(horizontal="left",vertical="center")
    for col in range(2,10): ca.cell(pr,col).fill=fill(BRAND_DARK)
    ca.row_dimensions[pr].height=20
    pr+=1
    tbl_header(pr,DETAIL_HDR); ca.row_dimensions[pr].height=26
    pr+=1
    squad_H=[]; squad_I=[]
    sq_first=pr
    for s in squads:
        typ,size,shore,supp=DEFAULTS[s]
        ca.cell(pr,2,s).font=F(10); ca.cell(pr,2).alignment=LFT; ca.cell(pr,2).border=ALL
        # inputs
        ic_t=ca.cell(pr,3,typ); ic_s=ca.cell(pr,4,size); ic_o=ca.cell(pr,5,shore); ic_p=ca.cell(pr,6,supp)
        for ic,al in ((ic_t,LFT),(ic_s,CENnw),(ic_o,CENnw),(ic_p,CENnw)):
            ic.fill=fill(INPUT_FILL); ic.font=INPUT_FONT; ic.alignment=al; ic.border=ALL
        ic_p.number_format=PCT
        dv_type.add(ic_t); dv_size.add(ic_s); dv_shore.add(ic_o); dv_sup.add(ic_p)
        # formulas
        g=ca.cell(pr,7,f'=IFERROR(IF($E{pr}="Offshore",INDEX({SQ_OFF},MATCH($C{pr}&"|"&$D{pr},{SQ_KEY},0)),INDEX({SQ_ON},MATCH($C{pr}&"|"&$D{pr},{SQ_KEY},0))),0)')
        money(g); g.font=FORMULA_FONT; g.border=ALL; g.alignment=RGT
        hh=ca.cell(pr,8,f"=G{pr}*F{pr}"); money(hh); hh.font=FORMULA_FONT; hh.border=ALL; hh.alignment=RGT
        ii=ca.cell(pr,9,f"=G{pr}*(1-F{pr})"); money(ii); ii.font=FORMULA_FONT; ii.border=ALL; ii.alignment=RGT
        squad_H.append(f"H{pr}"); squad_I.append(f"I{pr}")
        pr+=1
    sq_last=pr-1
    # ---- roll-up rows ----
    # Platform Overhead
    ca.cell(pr,2,"Platform Overhead (TDD-funded)").font=F(10,italic=True); ca.cell(pr,2).alignment=LFT
    ca.merge_cells(start_row=pr,start_column=2,end_row=pr,end_column=6)
    oh_tdd=ca.cell(pr,8,f"={DC_PLAT_OH}"); money(oh_tdd); oh_tdd.font=LINK_FONT; oh_tdd.alignment=RGT
    oh_oth=ca.cell(pr,9,0); money(oh_oth); oh_oth.alignment=RGT
    oh_tot=ca.cell(pr,7,f"=H{pr}+I{pr}"); money(oh_tot); oh_tot.alignment=RGT
    for col in range(2,10): ca.cell(pr,col).border=ALL; ca.cell(pr,col).fill=fill(BRAND_LT2)
    oh_row=pr; pr+=1
    # Support Cost
    ca.cell(pr,2,"Support Cost  (Σ squads)").font=F(10,italic=True); ca.cell(pr,2).alignment=LFT
    ca.merge_cells(start_row=pr,start_column=2,end_row=pr,end_column=6)
    sup_tdd=ca.cell(pr,8,f"=SUM(H{sq_first}:H{sq_last})"); money(sup_tdd); sup_tdd.alignment=RGT
    sup_oth=ca.cell(pr,9,f"=SUM(I{sq_first}:I{sq_last})"); money(sup_oth); sup_oth.alignment=RGT
    sup_tot=ca.cell(pr,7,f"=H{pr}+I{pr}"); money(sup_tot); sup_tot.alignment=RGT
    for col in range(2,10): ca.cell(pr,col).border=ALL; ca.cell(pr,col).fill=fill(BRAND_LT2)
    sup_row=pr; pr+=1
    # Platform Total
    ca.cell(pr,2,f"PLATFORM TOTAL — {pname}").font=F(10,True,WHITE); ca.cell(pr,2).alignment=LFT
    ca.merge_cells(start_row=pr,start_column=2,end_row=pr,end_column=6)
    tt_tdd=ca.cell(pr,8,f"=H{oh_row}+H{sup_row}"); money(tt_tdd); tt_tdd.font=F(10,True,WHITE); tt_tdd.alignment=RGT
    tt_oth=ca.cell(pr,9,f"=I{oh_row}+I{sup_row}"); money(tt_oth); tt_oth.font=F(10,True,WHITE); tt_oth.alignment=RGT
    tt_tot=ca.cell(pr,7,f"=H{pr}+I{pr}"); money(tt_tot); tt_tot.font=F(10,True,WHITE); tt_tot.alignment=RGT
    for col in range(2,10): ca.cell(pr,col).border=ALL; ca.cell(pr,col).fill=fill(BRAND_MED)
    tot_row=pr; pr+=2
    plat_refs.append(dict(name=pname,oh=f"H{oh_row}",sup_t=f"H{sup_row}",sup_o=f"I{sup_row}",
                          tot_t=f"H{tot_row}",tot_o=f"I{tot_row}",tot=f"G{tot_row}"))

# =====================================================================
# SUMMARY (rows 8-13), BUDGET (15-19), FY26 context (21-24)
# =====================================================================
section(8,"①  PORTFOLIO SUMMARY  —  Ampol Retail")
tbl_header(9,{2:"Cost component",7:"Total ($m)",8:"TDD ($m)",9:"Other ($m)"})
ca.merge_cells("B9:F9")
sum_oh = "+".join(p["oh"] for p in plat_refs)
sum_st = "+".join(p["sup_t"] for p in plat_refs)
sum_so = "+".join(p["sup_o"] for p in plat_refs)
rows_sum=[
 ("Total Portfolio Overhead", f"={DC_PORT_OH}", "0", True),
 ("Total Platform Overhead", f"={sum_oh}", "0", False),
 ("Total Support Costs", f"={sum_st}", f"={sum_so}", False),
]
r=10
first_sum=r
for label,tdd,oth,islink in rows_sum:
    ca.cell(r,2,label).font=F(10); ca.cell(r,2).alignment=LFT
    ca.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    hc=ca.cell(r,8,tdd); money(hc); hc.alignment=RGT; hc.font=(LINK_FONT if islink else FORMULA_FONT)
    ic=ca.cell(r,9,oth); money(ic); ic.alignment=RGT; ic.font=(LINK_FONT if oth.startswith("=") and False else FORMULA_FONT)
    gc=ca.cell(r,7,f"=H{r}+I{r}"); money(gc); gc.alignment=RGT
    for col in range(2,10): ca.cell(r,col).border=ALL
    r+=1
# TOTAL COST grand row
ca.cell(r,2,"TOTAL COST").font=F(11,True,WHITE); ca.cell(r,2).alignment=LFT
ca.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
gt_t=ca.cell(r,8,f"=SUM(H{first_sum}:H{r-1})"); money(gt_t); gt_t.font=F(11,True,WHITE); gt_t.alignment=RGT
gt_o=ca.cell(r,9,f"=SUM(I{first_sum}:I{r-1})"); money(gt_o); gt_o.font=F(11,True,WHITE); gt_o.alignment=RGT
gt_g=ca.cell(r,7,f"=H{r}+I{r}"); money(gt_g); gt_g.font=F(11,True,WHITE); gt_g.alignment=RGT
for col in range(2,10): ca.cell(r,col).fill=fill(GRAND_FILL); ca.cell(r,col).border=ALL
TOTAL_TDD_CELL=f"H{r}"; TOTAL_OTHER_CELL=f"I{r}"; TOTAL_ALL_CELL=f"G{r}"
grand_row=r

# ---- BUDGET & VARIANCE ----
section(15,"②  BUDGET & VARIANCE  —  TDD Lights-On")
budget_rows=[
 ("TDD Lights-On Budget  (Ampol Retail — Data Config)", f"={DC_RETAIL_BUDGET}", LINK_FONT),
 ("TDD Cost implied by squad selections  (Total TDD above)", f"={TOTAL_TDD_CELL}", FORMULA_FONT),
]
r=16
for label,val,fnt in budget_rows:
    ca.cell(r,2,label).font=F(10); ca.cell(r,2).alignment=LFT
    ca.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
    vc=ca.cell(r,8,val); money(vc); vc.alignment=RGT; vc.font=fnt
    ca.merge_cells(start_row=r,start_column=8,end_row=r,end_column=9)
    for col in range(2,10): ca.cell(r,col).border=ALL
    r+=1
budget_cell=f"H16"
# variance row
ca.cell(r,2,"VARIANCE  (Budget − TDD Cost)   →   surplus / (shortfall to fund)").font=F(10,True,BRAND_DARK); ca.cell(r,2).alignment=LFT
ca.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
var=ca.cell(r,8,f"=H16-H17"); money(var); var.alignment=RGT; var.font=F(11,True)
ca.merge_cells(start_row=r,start_column=8,end_row=r,end_column=9)
for col in range(2,10): ca.cell(r,col).border=ALL; ca.cell(r,col).fill=fill(ACCENT)
var_row=r; r+=1
# verdict text
ca.cell(r,2,f'=IF(H{var_row}>=0,"✔ Surplus of $"&TEXT(H{var_row},"0.00")&"m within the TDD lights-on allocation.","✘ Shortfall of $"&TEXT(-H{var_row},"0.00")&"m — additional funding required beyond the $"&TEXT(H16,"0.0")&"m TDD lights-on budget.")')
ca.cell(r,2).font=F(9,italic=True,color="30414A"); ca.cell(r,2).alignment=LFT
ca.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9)

# ---- FY26 RETAIL BUDGET context ----
section(21,"③  FY2026 RETAIL BUDGET  (envelope from the presentation pack)")
tbl_header(22,{2:"Lights On",4:"Initiatives (Project OpEx)",5:"Depreciation",6:"Significant Items",7:"CapEx",9:"Total FY26"})
ca.merge_cells("B22:C22"); ca.merge_cells("D22:D22")
# values pulled from FY26 Budget sheet Retail row
map_cols={2:("C",False),4:("D",False),5:("E",False),6:("F",False),7:("G",False),9:("H",True)}
for tcol,(scol,istot) in map_cols.items():
    v=ca.cell(23,tcol,f"='FY26 Budget'!{scol}{FB_RETAIL_ROW}"); money(v); v.font=(F(10,True) if istot else LINK_FONT); v.alignment=RGT; v.border=ALL
ca.merge_cells("B23:C23"); ca.cell(23,2).alignment=RGT
for col in range(2,10): ca.cell(23,col).border=ALL
ca.cell(24,2,"Note: The TDD Lights-On budget used for the variance above ($2.5m) is the TDD-funded lights-on allocation for Ampol Retail from Data Config; the table above is the full FY26 Retail technology envelope for context.").font=F(8,italic=True,color=GREY_TX)
ca.merge_cells("B24:I24"); ca.cell(24,2).alignment=LFT; ca.row_dimensions[24].height=26

# ---------- column widths ----------
for col,w in {"A":2.5,"B":30,"C":26,"D":8,"E":11,"F":11,"G":16,"H":14,"I":14}.items():
    ca.column_dimensions[col].width=w
ca.sheet_view.zoomScale=100

# ---------- reorder so calculator is first (already index 0) ----------
wb.save(OUT)
print("Built. grand_row(TDD/Other/Total):",grand_row, TOTAL_TDD_CELL,TOTAL_OTHER_CELL,TOTAL_ALL_CELL)
print("Budget cell H16, Var row",var_row)
print("Platform refs:")
for p in plat_refs: print("  ",p)
print("SAVED",OUT)
