#!/usr/bin/env python3
"""Add '3.5 Source Reconciliation' - a live, by-portfolio comparison of the new
REVIEW sheet against the old Squads sheet, so the owner can see where role counts
and vacancies moved between the two datasets. Everything is COUNTIFS (live)."""
import openpyxl
from openpyxl.utils import get_column_letter as gl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC="clean_v5.xlsx"; OUT="clean_v6.xlsx"
wb=openpyxl.load_workbook(SRC)
REV="'REVIEW - Complete Role Mapping'"; SQ="Squads"
name="3.5 Source Reconciliation"
if name in wb.sheetnames: del wb[name]
ws=wb.create_sheet(name, wb.sheetnames.index("3.4 COE Summary")+1)

# each row: common label, Squads Model-Portfolio criteria (col N), REVIEW MTab criteria list (col AJ)
ROWS=[
 ("Ampol Retail",            ['Ampol Retail'],            ['Ampol Retail']),
 ("Customer",                ['Customer'],                ['Customer']),
 ("Enterprise Data",         ['Enterprise Data'],         ['Enterprise Data']),
 ("TDD Group Functions",     ['TDD Group Functions'],     ['TDD Group Functions']),
 ("P&C",                     ['P&C'],                     ['P&C']),
 ("Finance",                 ['Finance'],                 ['Finance']),
 ("Infrastructure",          ['Infrastructure'],          ['Infrastructure']),
 ("Energy Solutions & B2B",  ['Energy Solutions & B2B'],  ['Energy Solutions & B2B']),
 ("Commercial Fuels",        ['Commercial Fuels'],        ['Commercial Fuels']),
 ("Z Retail",                ['Z Retail'],                ['Z Retail']),
 ("COE - Cyber",             ['TDD Cyber'],               ['COE Cyber']),
 ("COE - Partnering, Strategy, Arch & Data", ['COE'],     ['COE BP&T','COE SA&D']),
 ("EGI & Central",           ['EGI'],                     ['EGI & Central']),
 ("Unmapped (Squads only)",  ['Unmapped'],                []),
]

NAVY=PatternFill("solid",fgColor="1F4E79"); GREY=PatternFill("solid",fgColor="F2F2F2")
TOT=PatternFill("solid",fgColor="D9D9D9")
WF=Font(color="FFFFFF",bold=True); BF=Font(bold=True); NF=Font(); IT=Font(italic=True,size=10)
thin=Side(style="thin",color="BFBFBF"); BOX=Border(thin,thin,thin,thin)
CEN=Alignment(horizontal="center",vertical="center")
D0='#,##0;[Red]-#,##0;"-"'   # deltas show red when negative (roles/vacancies dropped)
def sc(ref,v,font=NF,fill=None,al=None,fmt=None,box=True):
    x=ws[ref]; x.value=v; x.font=font
    if fill:x.fill=fill
    if al:x.alignment=al
    if fmt:x.number_format=fmt
    if box:x.border=BOX
    return x

def sq_c(crit, status=None):
    n=",".join(f'{SQ}!$N:$N,"{c}"' for c in crit)  # not used directly; single crit per row
    c=crit[0]
    if status: return f'=COUNTIFS({SQ}!$N:$N,"{c}",{SQ}!$R:$R,"{status}")'
    return f'=COUNTIF({SQ}!$N:$N,"{c}")'
def rev_c(crit, status=None):
    if not crit: return '=0'
    parts=[]
    for c in crit:
        if status: parts.append(f'COUNTIFS({REV}!$AJ:$AJ,"{c}",{REV}!$AK:$AK,"{status}")')
        else: parts.append(f'COUNTIF({REV}!$AJ:$AJ,"{c}")')
    return "="+"+".join(parts)

sc("B2","Source Reconciliation - new REVIEW mapping vs old Squads sheet, by portfolio",Font(bold=True,size=14),box=False)
sc("B3","Positive delta = REVIEW has more than Squads; negative (red) = dropped versus the old sheet.",IT,box=False)
# Occupied = roles - vacant (someone in the seat), so both sheets compare on the
# same basis - REVIEW's "Filled" status excludes contractors while Squads' does not.
hdr=["Portfolio","Squads roles","Squads occupied","Squads vacant","REVIEW roles","REVIEW occupied","REVIEW vacant","Δ roles","Δ occupied","Δ vacant"]
for i,h in enumerate(hdr): sc(f"{gl(2+i)}5",h,WF,NAVY,CEN)
r=6; first=r
for label,sqc,revc in ROWS:
    sc(f"B{r}",label,NF,GREY if (r-first)%2 else None)
    sc(f"C{r}",sq_c(sqc),NF,al=CEN)
    sc(f"D{r}",f"=C{r}-E{r}",NF,al=CEN)                 # Squads occupied = roles - vacant
    sc(f"E{r}",sq_c(sqc,"Vacant"),NF,al=CEN)
    sc(f"F{r}",rev_c(revc),NF,al=CEN)
    sc(f"G{r}",f"=F{r}-H{r}",NF,al=CEN)                 # REVIEW occupied = roles - vacant
    sc(f"H{r}",rev_c(revc,"Vacant"),NF,al=CEN)
    sc(f"I{r}",f"=F{r}-C{r}",BF,al=CEN,fmt=D0)
    sc(f"J{r}",f"=G{r}-D{r}",BF,al=CEN,fmt=D0)
    sc(f"K{r}",f"=H{r}-E{r}",BF,al=CEN,fmt=D0)
    r+=1
last=r-1
sc(f"B{r}","Total",BF,TOT)
for col in "CDEFGH": sc(f"{col}{r}",f"=SUM({col}{first}:{col}{last})",BF,TOT,CEN)
for col in "IJK": sc(f"{col}{r}",f"=SUM({col}{first}:{col}{last})",BF,TOT,CEN,D0)
sc(f"B{r+2}",'Mapping: Squads "TDD Cyber" = COE - Cyber; Squads "COE" = COE Partnering + Strategy/Arch/Data (REVIEW splits these); Squads "EGI" = EGI & Central; "Unmapped" has no REVIEW equivalent.',IT,box=False)
sc(f"B{r+3}","Squads status uses its Status column; REVIEW status uses the MStatus helper (Vacant where the source marks the role vacant).",IT,box=False)
for col,wd in {"B":40,"C":13,"D":13,"E":13,"F":13,"G":13,"H":13,"I":10,"J":10,"K":10}.items():
    ws.column_dimensions[col].width=wd
ws.sheet_view.selection=[]
wb.save(OUT)
print("saved",OUT,"| recon rows:",len(ROWS),"| total row:",r)
