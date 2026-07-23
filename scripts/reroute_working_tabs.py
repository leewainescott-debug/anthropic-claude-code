#!/usr/bin/env python3
"""Real-squad reroute of the 2.x working tabs that KEEPS the owner's full
archetype scorecard columns. Squad rows are the REAL squads from REVIEW; the FTE
list references REVIEW live. Archetype type/roles/cost/size are looked up from
3.3 FTE View by portfolio+squad (a matched design squad shows its target, an
unmatched real squad shows "-" and the Flag says it is outside the archetype
model). Squads / Added data / Sheet2 stay in place."""
import openpyxl
from openpyxl.utils import column_index_from_string as ci, get_column_letter as gl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.views import Selection
from collections import defaultdict, OrderedDict

SRC="clean_reroute.xlsx"; OUT="clean_v4.xlsx"
wb=openpyxl.load_workbook(SRC)
REV="REVIEW - Complete Role Mapping"; Q="'"+REV+"'"; F3="'3.3 FTE View'"
ws=wb[REV]
def c(r,cc): return ws.cell(r,ci(cc)).value

PMAP={"ampol retail":"Ampol Retail","retail":"Ampol Retail","z":"Z Retail","customer":"Customer",
"ampol customer":"Customer","enterprise data":"Enterprise Data","tdd":"TDD Group Functions",
"infrastructure":"Infrastructure","b2b & energy solutions":"Energy Solutions & B2B",
"commercial fuels":"Commercial Fuels","finance":"Finance","p&c":"P&C","p&c, finance & legal":"P&C",
"egi":"EGI & Central","coe - cyber, risk & operations":"COE Cyber",
"coe - partnering & transformation":"COE BP&T","coe - strategy, architecture, data":"COE SA&D"}
SALIAS={"ampos":"AmPOS","customer, ai":"Customer AI","manuacturing group projects":"Manufacturing Group Projects",
"integration & process automation":"Integration & Process Automation","technology suport":"Technology Support",
"data - au":"Data AU","data - nz":"Data NZ","na":"(all roles)","":"(unassigned)"}
def nsq(s):
    s=" ".join(str(s or "").split()); return SALIAS.get(s.lower(), s)
def is_vac(r): return str(c(r,'Q') or "").strip().lower()=="v"
def mtab(r): return PMAP.get(str(c(r,'I') or "").strip().lower())

# helper MSquad + collect rows per tab/squad
AO=ci("AO"); ws.cell(1,AO).value="MSquad"
data=defaultdict(lambda:OrderedDict())
for r in range(2,ws.max_row+1):
    if not ws.cell(r,2).value:
        ws.cell(r,AO).value=None; continue
    sq=nsq(c(r,'K')); ws.cell(r,AO).value=sq
    t=mtab(r)
    if not t: continue
    data[t].setdefault(sq,[]).append(dict(dr=r, vac=is_vac(r)))
if str(ws.cell(445,ci('R')).value or "").strip(): ws.cell(445,ci('R')).value=None

# 3.3 FTE View: portfolio|squad key in a hidden helper column (for archetype lookup)
f3=wb["3.3 FTE View"]; KEYCOL=ci("AD")
f3.cell(6,KEYCOL).value="_key"
for r in range(7,95):
    b=f3.cell(r,2).value; d=f3.cell(r,4).value
    if b is not None and d is not None:
        f3.cell(r,KEYCOL).value=f'=$B{r}&"|"&$D{r}'
f3.column_dimensions[gl(KEYCOL)].hidden=True
KC=gl(KEYCOL)

# styling
NAVY=PatternFill("solid",fgColor="1F4E79"); GREY=PatternFill("solid",fgColor="F2F2F2")
TOT=PatternFill("solid",fgColor="D9D9D9"); YEL=PatternFill("solid",fgColor="FFF2CC")
SQH=PatternFill("solid",fgColor="DDEBF7"); ARCH=PatternFill("solid",fgColor="EDEDED")
WF=Font(color="FFFFFF",bold=True); BF=Font(bold=True); NF=Font(); BL=Font(color="0000FF"); IT=Font(italic=True,size=10)
thin=Side(style="thin",color="BFBFBF"); BOX=Border(thin,thin,thin,thin)
M2='#,##0.00;[Red](#,##0.00);"-"'; D0='#,##0;[Red](#,##0);"-"'; DOL='#,##0'
CEN=Alignment(horizontal="center",vertical="center"); RT=Alignment(horizontal="right")
def sc(w,ref,v,font=NF,fill=None,al=None,fmt=None,box=True):
    x=w[ref]; x.value=v; x.font=font
    if fill:x.fill=fill
    if al:x.alignment=al
    if fmt:x.number_format=fmt
    if box:x.border=BOX
    return x

WT=OrderedDict([("2.1 Ampol Retail","Ampol Retail"),("2.2 Customer","Customer"),
("2.3 Enterprise Data","Enterprise Data"),("2.4 TDD Group Functions","TDD Group Functions"),
("2.5 P&C","P&C"),("2.6 Finance","Finance"),("2.7 Infrastructure","Infrastructure"),
("2.8 Energy Solutions & B2B","Energy Solutions & B2B"),("2.9 Commercial Fuels","Commercial Fuels"),
("2.10 Z Retail","Z Retail"),("2.11 TDD Cyber","COE Cyber"),("2.12 BP&T","COE BP&T"),
("2.13 SA&D","COE SA&D"),("2.14 EGI & Central","EGI & Central")])
# B Squad C ArchType D ArchRoles E Filled F Vacant G Plan-hire H Vac-remain
# I vs-arch J Flag K ArchCost L ActualCost M CostAfter N NewVar O ArchSize
HDR=["Squad","Archetype type","Archetype roles","Filled","Vacant","Planning to hire",
     "Vacancies remaining","vs archetype","Flag","Archetype cost ($m)","Actual cost ($m)",
     "Cost after vacancy decisions ($m)","New Variance ($m)","Archetype size"]  # B..O
ARCHCOLS={"C","D","K","O"}  # grey archetype cells
mtot={}; hire_memo={}
for tabname,mt in WT.items():
    w=wb[tabname]
    for row in list(w.iter_rows(min_row=3)):
        for cl in row:
            cl.value=None; cl.fill=PatternFill(); cl.border=Border()
            cl.alignment=Alignment(); cl.number_format="General"
    for dv in list(w.data_validations.dataValidation): w.data_validations.dataValidation.remove(dv)
    squads=data.get(mt,{})
    sc(w,"B2",f"{mt} - working copy",Font(bold=True,size=15),box=False)
    sc(w,"B4","Position by squad - archetype vs actual, with the vacancy lever",WF,NAVY,box=False)
    for i,h in enumerate(HDR): sc(w,f"{gl(2+i)}5",h,WF,NAVY,CEN)
    first=6; n=len(squads); total_row=first+n
    rost_hdr=total_row+8
    roster_rows={}; rr=rost_hdr+1
    for sq,roles in squads.items():
        roster_rows[sq]=(rr+1, rr+len(roles)); rr+=1+len(roles)
    dv=DataValidation(type="list",formula1='"Hire,Hold,Offshore"',allow_blank=False); w.add_data_validation(dv)
    r=first
    for sq,roles in squads.items():
        a,b=roster_rows[sq]; sqq=sq.replace('"','""')
        key=f'"{mt}|{sqq}"'
        look=lambda col: f'IFERROR(INDEX({F3}!${col}:${col},MATCH({key},{F3}!${KC}:${KC},0)),"-")'
        sc(w,f"B{r}",sq,NF,GREY if (r-first)%2 else None)
        sc(w,f"C{r}",f"={look('E')}",NF,ARCH,CEN)                       # Archetype type
        sc(w,f"D{r}",f"={look('G')}",NF,ARCH,CEN,fmt=D0)                # Archetype roles (FTE)
        sc(w,f"E{r}",f'=COUNTIF($D${a}:$D${b},"Filled")',NF,al=CEN,fmt=D0)
        sc(w,f"F{r}",f'=COUNTIF($D${a}:$D${b},"Vacant")',NF,al=CEN,fmt=D0)
        sc(w,f"G{r}",f'=COUNTIF($E${a}:$E${b},"Hire")',NF,al=CEN,fmt=D0)
        sc(w,f"H{r}",f'=F{r}-G{r}-COUNTIF($E${a}:$E${b},"Offshore")',NF,al=CEN,fmt=D0)
        sc(w,f"I{r}",f'=IFERROR(E{r}+G{r}-D{r},"-")',NF,al=CEN,fmt=D0)  # vs archetype
        sc(w,f"J{r}",f'=IF(ISNUMBER(D{r}),IF(E{r}>D{r},"Filled already over archetype",IF(E{r}+G{r}>D{r},"Over archetype after hire","")),"Outside the archetype model - no target set")',NF)
        sc(w,f"K{r}",f"={look('M')}",NF,ARCH,RT,fmt=M2)                 # Archetype cost
        base=f'SUMIFS({Q}!$AA:$AA,{Q}!$AJ:$AJ,"{mt}",{Q}!$AO:$AO,"{sqq}",{Q}!$AK:$AK,"Filled")'
        sc(w,f"L{r}",f'=({base})/1000000',NF,al=RT,fmt=M2)             # Actual cost (filled)
        dec=f'SUMIFS($F${a}:$F${b},$E${a}:$E${b},"Hire")+0.4*SUMIFS($F${a}:$F${b},$E${a}:$E${b},"Offshore")'
        sc(w,f"M{r}",f'=L{r}+({dec})/1000000',NF,al=RT,fmt=M2)         # Cost after decisions
        sc(w,f"N{r}",f'=IFERROR(M{r}-K{r},"-")',NF,al=RT,fmt=M2)       # New Variance
        sc(w,f"O{r}",f"={look('F')}",NF,ARCH,CEN)                      # Archetype size
        r+=1
    last=r-1
    sc(w,f"B{total_row}","Total",BF,TOT)
    for col in "DEFGHI": sc(w,f"{col}{total_row}",f"=SUM({col}{first}:{col}{last})",BF,TOT,CEN,D0)
    for col in "KLMN": sc(w,f"{col}{total_row}",f"=SUM({col}{first}:{col}{last})",BF,TOT,RT,M2)
    for col in "CJO": sc(w,f"{col}{total_row}","",BF,TOT)
    mtot[tabname]=f"M{total_row}"
    ra,rb=rost_hdr+1, rr-1
    sc(w,f"B{total_row+1}","Cost to hire all vacancies ($m)",NF,box=False)
    sc(w,f"E{total_row+1}",f'=SUMIFS($F${ra}:$F${rb},$D${ra}:$D${rb},"Vacant")/1000000',NF,al=RT,fmt=M2,box=False)
    hire_memo[tabname]=f"E{total_row+1}"
    sc(w,f"B{total_row+2}","Cost of roles marked Hire ($m)",NF,box=False)
    sc(w,f"E{total_row+2}",f'=SUMIF($E${ra}:$E${rb},"Hire",$F${ra}:$F${rb})/1000000',NF,al=RT,fmt=M2,box=False)
    sc(w,f"B{total_row+3}","Vacant roles are priced at standard title rates - indicative until an offer is made.",IT,box=False)
    sc(w,f"B{total_row+4}",'"-" in the archetype columns means this real squad has no design archetype target; the comparison sits at portfolio level on 3.2.',IT,box=False)
    # roster
    sc(w,f"B{rost_hdr-1}",f"{mt} FTE",Font(bold=True,size=12),box=False)
    for i,h in enumerate(["Name","Role","Status","Vacancy lever","Cost if hired ($)"]):
        sc(w,f"{gl(2+i)}{rost_hdr}",h,WF,NAVY,CEN)
    rr=rost_hdr+1
    for sq,roles in squads.items():
        sc(w,f"B{rr}",sq,BF,SQH)
        for col in "CDEF": sc(w,f"{col}{rr}","",BF,SQH)
        rr+=1
        for x in roles:
            dr=x["dr"]
            sc(w,f"B{rr}",f"={Q}!$B${dr}",NF)
            sc(w,f"C{rr}",f"={Q}!$C${dr}",NF)
            sc(w,f"D{rr}",f"={Q}!$AK${dr}",NF,al=CEN)
            if x["vac"]:
                cell=sc(w,f"E{rr}","Hold",BL,YEL,CEN); dv.add(cell)
                sc(w,f"F{rr}",f"={Q}!$AA${dr}",NF,al=RT,fmt=DOL)
            else:
                sc(w,f"E{rr}","",NF); sc(w,f"F{rr}","",NF)
            rr+=1
    widths={"B":30,"C":16,"D":14,"E":9,"F":9,"G":15,"H":18,"I":12,"J":30,"K":18,"L":16,"M":26,"N":16,"O":14}
    for col,wd in widths.items(): w.column_dimensions[col].width=wd

# 3.2 F (cost after vacancy decisions) -> each tab's Cost-after-decisions total (M)
tc=wb["3.2 Total Cost"]
F32={6:"2.1 Ampol Retail",7:"2.2 Customer",8:"2.3 Enterprise Data",9:"2.4 TDD Group Functions",
10:"2.5 P&C",11:"2.6 Finance",12:"2.7 Infrastructure",13:"2.8 Energy Solutions & B2B",
14:"2.9 Commercial Fuels",15:"2.10 Z Retail",20:"2.11 TDD Cyber"}
for r,t in F32.items():
    tc[f"F{r}"].value=f"='{t}'!${mtot[t][0]}${mtot[t][1:]}"
# Exec C52 -> sum of each tab's "cost to hire all vacancies" memo
picks=list(F32.values())
wb["Exec Summary"]["C52"].value="="+"+".join(f"'{t}'!${hire_memo[t][0]}${hire_memo[t][1:]}" for t in picks)
# 3.3 O/P held baseline (decisions live on 2.x)
for r in range(7, 95):
    if f3.cell(r,9).value is not None:
        f3.cell(r,15).value=f"=I{r}"; f3.cell(r,16).value=f"=N{r}"
# 4.0 QA -> new status column D
qa=wb["4.0 Data QA"]; qa["C258"].value=qa["C258"].value.replace("$F$1:$F$500","$D$1:$D$500")
# remove ALL frozen panes cleanly
for w in wb.worksheets:
    w.freeze_panes=None; w.sheet_view.pane=None
    w.sheet_view.selection=[Selection(activeCell="A1",sqref="A1")]
wb.save(OUT)
print("saved",OUT,"| M-totals sample:",{k:mtot[k] for k in list(mtot)[:3]})
