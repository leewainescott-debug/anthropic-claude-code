#!/usr/bin/env python3
"""v9: patch Lee's edited workbook (Cost_Calc_Lee_edits22.xlsx) in place - never
regenerate. Sheet2 is the roster source for Cyber / EGI / P&T / SA&D. Squads
(raw data) is never written. Every roster cell is a live reference, every count
is a formula - no hard-coded names, words or numbers in model cells."""
import shutil, json, copy, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

SCR = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/"
SRC = "/root/.claude/uploads/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/4beb5516-Cost_Calc_Lee_edits22.xlsx"
OUT = SCR + "TDD_Cost_Calc_v9.xlsx"
shutil.copy(SRC, OUT)
wb = openpyxl.load_workbook(OUT, data_only=False)

S2 = "Sheet2"
NPORT = "COUNTA('2.0 Group Summary'!$B$6:$B$15)"   # live portfolio count (10 - cyber is a COE)
OFF = "'0.1 Squads'!$K$5"                           # offshore factor cell

# ---------- shared style kit (matches the workbook's existing look) ----------
NAVY   = PatternFill("solid", fgColor="FF002F6C")
MIDBLU = PatternFill("solid", fgColor="FF1F4E79")
LGREY  = PatternFill("solid", fgColor="FFF2F2F2")
YELL   = PatternFill("solid", fgColor="FFFFF2CC")
WHITEF = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
BOLD   = Font(name="Calibri", size=11, bold=True)
NORM   = Font(name="Calibri", size=11)
THIN   = Border(*[Side(style="thin", color="FFBFBFBF")]*4)
M0  = "0.0"      # fte / counts
M2  = "0.00"     # $m
D0  = "#,##0"    # $
def sc(ws, addr, val, font=None, fill=None, fmt=None, align=None, wrap=False, border=True):
    for mr in list(ws.merged_cells.ranges):
        c1, r1, c2, r2 = mr.bounds
        cell0 = ws[addr]
        if r1 <= cell0.row <= r2 and c1 <= cell0.column <= c2:
            ws.unmerge_cells(str(mr))
    c = ws[addr]
    c.value = val
    if font: c.font = font
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    if border: c.border = THIN
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c
def clear_region(ws, r1, r2, c1=1, c2=15):
    for mr in list(ws.merged_cells.ranges):
        b = mr.bounds
        if not (b[3] < r1 or b[1] > r2):
            ws.unmerge_cells(str(mr))
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
            cell.border = Border()
            cell.font = NORM

# ---------- read Sheet2 (structure only - the workbook keeps live refs) ------
s2 = wb[S2]
def s2v(r, c): return s2.cell(r, c).value
def low(v): return str(v).strip().lower() if v is not None else ""
rows2 = []
for r in range(2, s2.max_row+1):
    if s2v(r, 6) is None: continue
    rows2.append(dict(r=r, name=s2v(r,2), title=s2v(r,3), div=str(s2v(r,6)).strip(),
                      dept=str(s2v(r,7) or "").strip(), port=low(s2v(r,9)),
                      plat=low(s2v(r,10)), squad=low(s2v(r,11)), ctry=s2v(r,13),
                      typ=low(s2v(r,17))))
CYB = [x for x in rows2 if x["div"] == "Cyber, Risk & Operations"]
EGI = [x for x in rows2 if x["div"] == "EGI"]
PT  = [x for x in rows2 if x["div"] == "Partnering & Transformation"]
SAD = [x for x in rows2 if x["div"] == "Strategy, Architecture & Data"]
assert (len(CYB), len(EGI), len(PT)) == (52, 16, 24), (len(CYB), len(EGI), len(PT))

# Squads index for the SA&D exclusion rule + reconciliation (READ ONLY)
sq = wb["Squads"]
sq_rows = []
for r in range(2, sq.max_row+1):
    if sq.cell(r,2).value is None: continue
    sq_rows.append(dict(r=r, name=low(sq.cell(r,2).value), title=low(sq.cell(r,3).value),
                        div=low(sq.cell(r,6).value), N=str(sq.cell(r,14).value or "").strip(),
                        cls=str(sq.cell(r,17).value or "").strip(),
                        st=str(sq.cell(r,18).value or "").strip()))
def sq_match(x):
    """best-effort match of a Sheet2 row to a Squads row (same division first)"""
    cands = [q for q in sq_rows if q["name"] == low(x["name"]) and q["title"] == low(x["title"])]
    div_c = [q for q in cands if q["div"] == low(x["div"])]
    return (div_c or cands or [None])[0]

def blankish(v): return v in ("", "na")
# 2.4 SA&D roster - THE OWNER'S RUTHLESS RULE (their pasted list, 30 roles):
# all of Architecture, Technology Strategy & AI Capability, Delivery SADA and
# the GM office ('na'), plus Group Data roles on the Data Capability team and
# the Head of Technology. Engineering/Science/Operations/Reporting stay squads.
def s2team(r): return str(s2.cell(r, 8).value or "").strip()
def s2fte(r):
    v = s2.cell(r, 15).value
    return float(v) if isinstance(v, (int, float)) else 1.0
def is_sad_coe(x):
    d = x["dept"]
    if d in ("Architecture", "Technology Strategy & AI Capability", "na"):
        return True
    if d == "Delivery, SADA":
        # a Delivery role already seconded into a squad at partial FTE stays
        # with the squad (the owner's 29/31 split - Kina Birkby)
        return blankish(x["squad"]) or s2fte(x["r"]) >= 1.0
    if d == "Group Data" and (s2team(x["r"]) == "Data Capability"
                              or str(x["title"]).startswith("Head of Technology")):
        return True
    return False
sad_coe = [x for x in SAD if is_sad_coe(x)]
assert len(sad_coe) == 29, len(sad_coe)
# update the raw-data MODEL columns (N-P/Q) where they contradict the owner's
# rule - only rows matched inside the SA&D division; every change is logged
CHANGELOG = []
for x in sad_coe:
    m = sq_match(x)
    if not m or m["div"] != "strategy, architecture & data":
        continue
    bucket = "COE - Data" if x["dept"] == "Group Data" else "COE - Strategy Architecture"
    row = m["r"]
    for col, newv in ((14, "COE"), (15, bucket), (16, bucket), (17, "COE")):
        old = sq.cell(row, col).value
        if str(old or "").strip() != newv:
            CHANGELOG.append((row, get_column_letter(col), old, newv, str(x["name"])))
            sq.cell(row, col).value = newv
print("2.4 roster:", len(sad_coe), "| raw mapping cells updated:", len(CHANGELOG))
# mirror the mapping change into the Added data ledger helper columns (ours),
# so the 2.1 actuals move with the roles and nothing is counted twice
ad_ = wb["Added data"]
AD_CHANGES = 0
adx = {}
for r in range(2, ad_.max_row + 1):
    nm, tt = low(ad_.cell(r, 2).value), low(ad_.cell(r, 3).value)
    if nm: adx.setdefault((nm, tt), []).append(r)
for x in sad_coe:
    m = sq_match(x)
    if not m or m["div"] != "strategy, architecture & data": continue
    bucket = "COE - Data" if x["dept"] == "Group Data" else "COE - Strategy Architecture"
    for r in adx.get((low(x["name"]), low(x["title"])), []):
        for c, newv in ((29, "COE"), (30, bucket), (31, bucket), (32, "COE")):
            if str(ad_.cell(r, c).value or "").strip() != newv:
                ad_.cell(r, c).value = newv; AD_CHANGES += 1
print("Added data helper cells mirrored:", AD_CHANGES)

# formula helpers - every cell a live ref into Sheet2
def f_name(r):   return f"={S2}!$B${r}"
def f_title(r):  return f'=SUBSTITUTE(SUBSTITUTE({S2}!$C${r},"–","-"),"—","-")'
def f_dept(r):   return f"={S2}!$G${r}"
def f_ctry(r):   return f"={S2}!$M${r}"
def f_status(r):
    q = f"LOWER({S2}!$Q${r})"
    return (f'=IF({q}="v","Vacant",IF({q}="pause","Paused",'
            f'IF({q}="cxc","Contractor","Filled")))')
def f_cost(r):   return f"={S2}!$AA${r}"
def f_model(row_here, paused_zero=True):
    base = f'IF($H{row_here}="Offshore",$I{row_here}*{OFF},$I{row_here})/1000000'
    if paused_zero:
        return f'=IF($F{row_here}="Paused",0,{base})'
    return "=" + base

def write_roster(ws, first, entries, cat_formula, paused_zero):
    """no salaries for filled people - cost shows only on vacancies; a hidden
    helper column T carries every row's model cost so summaries can foot"""
    for i, x in enumerate(entries):
        rr = first + i; r2r = x["r"]
        sc(ws, f"B{rr}", f_name(r2r), NORM)
        sc(ws, f"C{rr}", f_title(r2r), NORM, wrap=True)
        sc(ws, f"D{rr}", f_dept(r2r), NORM)
        sc(ws, f"E{rr}", f_ctry(r2r), NORM)
        sc(ws, f"F{rr}", f_status(r2r), NORM, align="center")
        if x["typ"] in ("v", "pause"):
            sc(ws, f"G{rr}", f_cost(r2r), NORM, fmt=D0, align="right")
        helper = f"={S2}!$AA${r2r}/1000000"
        if paused_zero:
            helper = f'=IF($F{rr}="Paused",0,{S2}!$AA${r2r}/1000000)'
        c = ws[f"T{rr}"]; c.value = helper; c.number_format = M2
    ws.column_dimensions["T"].hidden = True
    return first + len(entries) - 1

def strip_dv_cf(ws):
    """remove stale data validations / conditional formats before a rebuild -
    overlapping DV ranges make Excel 'repair' the file"""
    ws.data_validations.dataValidation = []
    for rng in list(ws.conditional_formatting):
        del ws.conditional_formatting[rng.sqref]

def roster_headers(ws, r):
    for col, h in zip("BCDEFG", ["Name","Position Title","Department","Country",
                                 "Status","Cost if hired ($)"]):
        sc(ws, f"{col}{r}", h, WHITEF, MIDBLU, align="center", wrap=True)
    for col in "HIJ":  # stale upload headers right of the roster (old On/Off, Full Cost)
        cell = ws[f"{col}{r}"]
        cell.value = None
        cell.fill = PatternFill(); cell.border = Border(); cell.font = NORM

# =====================================================================
# 2.3 BP&T - roster from Sheet2, funding with real formulas, both budgets
# =====================================================================
ws = wb["2.3 BP&T"]
strip_dv_cf(ws)
clear_region(ws, 4, ws.max_row+4, 2, 12)
sc(ws, "B4", "Summary", WHITEF, NAVY)
for col, h in zip("BCDEFGHIJ", ["Grouping","Roles","Filled","Vacant","Planned spend ($m)",
                                "Budget to draw down ($m)","Left to fund ($m)",
                                "Cost - AU ($m)","Cost - NZ ($m)"]):
    sc(ws, f"{col}5", h, WHITEF, MIDBLU, align="center", wrap=True)
n_pt = len(PT)
R1, R2 = 21, 21 + n_pt - 1   # FTE block
BPCRIT = [("TDD Business Partner",), ("Commercial",)]
def dsum(rng_col, crits, extra=""):
    terms = [f'SUMIFS({rng_col},$D${R1}:$D${R2},"{c[0]}"{extra})' for c in crits]
    return "+".join(terms)
for i, (cat, crits) in enumerate([("Business Partnering", [("TDD Business Partner",), ("Commercial",)]),
                                  ("Transformation", [("Transformation",)])]):
    r = 6 + i
    sc(ws, f"B{r}", cat, BOLD)
    cnt_t = "+".join(f'COUNTIFS($D${R1}:$D${R2},"{c[0]}")' for c in crits)
    sc(ws, f"C{r}", f"={cnt_t}", NORM, fmt="0", align="center")
    fil_t = "+".join(f'COUNTIFS($D${R1}:$D${R2},"{c[0]}",$F${R1}:$F${R2},"Filled")' for c in crits)
    vac_t = "+".join(f'COUNTIFS($D${R1}:$D${R2},"{c[0]}",$F${R1}:$F${R2},"Vacant")' for c in crits)
    sc(ws, f"D{r}", f"={fil_t}", NORM, fmt="0", align="center")
    sc(ws, f"E{r}", f"={vac_t}", NORM, fmt="0", align="center")
    sc(ws, f"F{r}", f'={dsum(f"$T${R1}:$T${R2}", crits)}', NORM, fmt=M2, align="right")
    sc(ws, f"H{r}", f"=MAX(0,F{r}-G{r})", NORM, fmt=M2, align="right")
    au_t = "+".join(f'SUMIFS($T${R1}:$T${R2},$D${R1}:$D${R2},"{c[0]}",$E${R1}:$E${R2},"<>NZ")' for c in crits)
    nz_t = "+".join(f'SUMIFS($T${R1}:$T${R2},$D${R1}:$D${R2},"{c[0]}",$E${R1}:$E${R2},"NZ")' for c in crits)
    sc(ws, f"I{r}", f"={au_t}", NORM, fmt=M2, align="right")
    sc(ws, f"J{r}", f"={nz_t}", NORM, fmt=M2, align="right")
    sc(ws, f"G{r}", "=C15" if cat == "Business Partnering" else "=C16", NORM, fmt=M2, align="right")
sc(ws, "B8", "Total", BOLD, LGREY)
for col in "CDEFGHIJ":
    fmt = "0" if col in "CDE" else M2
    sc(ws, f"{col}8", f"=SUM({col}6:{col}7)", BOLD, LGREY, fmt=fmt,
       align="center" if col in "CDE" else "right")
# funding block - every number is a formula
fund = [
    ("Portfolios funded (2.0 Group Summary)",              f"={NPORT}", "0"),
    ("Business Partner allocation per portfolio (FTE - 0.0 Data Config)", "='0.0 Data Config'!$K$7", "0.0"),
    ("Business Partner FTEs funded by portfolio overheads", "=C10*C11", "0.0"),
    ("Business Partner funding from portfolio overheads ($m)", f"={NPORT}*'0.0 Data Config'!$L$7", M2),
    ("COE - Business Partnering allocation ($m) - 0.0 Data Config", "='0.0 Data Config'!$E$9", M2),
    ("Total Business Partnering budget ($m)",              "=C13+C14", M2),
    ("COE - Transformation allocation ($m) - 0.0 Data Config", "='0.0 Data Config'!$E$8", M2),
    ("Total budget to draw down ($m)",                     "=C15+C16", M2),
]
for i, (lab, fx, fmt) in enumerate(fund):
    r = 10 + i
    bold = lab.startswith("Total")
    sc(ws, f"B{r}", lab, BOLD if bold else NORM)
    sc(ws, f"C{r}", fx, BOLD if bold else NORM, fmt=fmt, align="right")
sc(ws, "B19", "Roles", WHITEF, NAVY)
roster_headers(ws, 20)
def cat_pt(rr, r2r):
    return f'=IF({S2}!$G${r2r}="Transformation","Transformation","Business Partnering")'
last = write_roster(ws, R1, PT, None, paused_zero=False)
sc(ws, f"B{last+1}", "Check - roles listed vs counted (must be 0)", NORM)
sc(ws, f"C{last+1}", f"=COUNTA(B{R1}:B{last})-C8", NORM, fmt="0", align="center")
sc(ws, f"B{last+3}",
   "Commercial roles sit in the Business Partnering bucket - the Department column shows where each role reports.",
   NORM, border=False)
clear_region(ws, last+4, max(ws.max_row, last+10), 2, 12)
PT_CHECK = last+1

# =====================================================================
# 2.4 SA&D - full division coverage (architecture, tech strategy & AI, data)
# =====================================================================
ws = wb["2.4 SA&D"]
strip_dv_cf(ws)
clear_region(ws, 4, ws.max_row+6, 2, 12)
sc(ws, "B4", "Summary", WHITEF, NAVY)
for col, h in zip("BCDEFGHIJK", ["Grouping","Roles","Filled","Vacant","Paused",
                                 "Planned spend ($m)","Budget to draw down ($m)","Left to fund ($m)",
                                 "Cost - AU ($m)","Cost - NZ ($m)"]):
    sc(ws, f"{col}5", h, WHITEF, MIDBLU, align="center", wrap=True)
n_sad = len(sad_coe)
S1, S2R = 22, 22 + n_sad - 1
TS, DS, ES, FS = f"$T${S1}:$T${S2R}", f"$D${S1}:$D${S2R}", f"$E${S1}:$E${S2R}", f"$F${S1}:$F${S2R}"
sc(ws, "B6", "Strategy & Architecture", BOLD)
sc(ws, "C6", f'=COUNTA($B${S1}:$B${S2R})-COUNTIFS({DS},"Group Data")', NORM, fmt="0", align="center")
sc(ws, "D6", f'=COUNTIFS({FS},"Filled")-COUNTIFS({DS},"Group Data",{FS},"Filled")', NORM, fmt="0", align="center")
sc(ws, "E6", f'=COUNTIFS({FS},"Vacant")-COUNTIFS({DS},"Group Data",{FS},"Vacant")', NORM, fmt="0", align="center")
sc(ws, "F6", f'=COUNTIFS({FS},"Paused")-COUNTIFS({DS},"Group Data",{FS},"Paused")', NORM, fmt="0", align="center")
sc(ws, "G6", f'=SUM({TS})-SUMIFS({TS},{DS},"Group Data")', NORM, fmt=M2, align="right")
sc(ws, "H6", "=C15", NORM, fmt=M2, align="right")
sc(ws, "I6", "=MAX(0,G6-H6)", NORM, fmt=M2, align="right")
sc(ws, "J6", f'=SUMIFS({TS},{ES},"<>NZ")-SUMIFS({TS},{DS},"Group Data",{ES},"<>NZ")', NORM, fmt=M2, align="right")
sc(ws, "K6", f'=SUMIFS({TS},{ES},"NZ")-SUMIFS({TS},{DS},"Group Data",{ES},"NZ")', NORM, fmt=M2, align="right")
sc(ws, "B7", "Data", BOLD)
sc(ws, "C7", f'=COUNTIFS({DS},"Group Data")', NORM, fmt="0", align="center")
sc(ws, "D7", f'=COUNTIFS({DS},"Group Data",{FS},"Filled")', NORM, fmt="0", align="center")
sc(ws, "E7", f'=COUNTIFS({DS},"Group Data",{FS},"Vacant")', NORM, fmt="0", align="center")
sc(ws, "F7", f'=COUNTIFS({DS},"Group Data",{FS},"Paused")', NORM, fmt="0", align="center")
sc(ws, "G7", f'=SUMIFS({TS},{DS},"Group Data")', NORM, fmt=M2, align="right")
sc(ws, "H7", "=C16", NORM, fmt=M2, align="right")
sc(ws, "I7", "=MAX(0,G7-H7)", NORM, fmt=M2, align="right")
sc(ws, "J7", f'=SUMIFS({TS},{DS},"Group Data",{ES},"<>NZ")', NORM, fmt=M2, align="right")
sc(ws, "K7", f'=SUMIFS({TS},{DS},"Group Data",{ES},"NZ")', NORM, fmt=M2, align="right")
sc(ws, "B8", "Total", BOLD, LGREY)
for col in "CDEFGHIJK":
    fmt = "0" if col in "CDEF" else M2
    sc(ws, f"{col}8", f"=SUM({col}6:{col}7)", BOLD, LGREY, fmt=fmt,
       align="center" if col in "CDEF" else "right")
fund = [
    ("Portfolios funded (2.0 Group Summary)",             f"={NPORT}", "0"),
    ("Domain Architect allocation per portfolio (FTE - 0.0 Data Config)", "='0.0 Data Config'!$K$8", "0.0"),
    ("Domain Architect FTEs funded by portfolio overheads", "=C10*C11", "0.0"),
    ("Domain Architect funding from portfolio overheads ($m)", f"={NPORT}*'0.0 Data Config'!$L$8", M2),
    ("COE - Strategy Architecture allocation ($m) - 0.0 Data Config", "='0.0 Data Config'!$E$6", M2),
    ("Total Strategy & Architecture budget ($m)",         "=C13+C14", M2),
    ("COE - Data allocation ($m) - 0.0 Data Config",      "='0.0 Data Config'!$E$10", M2),
    ("Total budget to draw down ($m)",                    "=C15+C16", M2),
]
for i, (lab, fx, fmt) in enumerate(fund):
    r = 10 + i
    bold = lab.startswith("Total")
    sc(ws, f"B{r}", lab, BOLD if bold else NORM)
    sc(ws, f"C{r}", fx, BOLD if bold else NORM, fmt=fmt, align="right")
sc(ws, "B18", "Paused roles - cost if released ($m)", NORM)
sc(ws, "C18", f'=SUMIFS($G${S1}:$G${S2R},$F${S1}:$F${S2R},"Paused")/1000000', NORM, fmt=M2, align="right")
sc(ws, "B20", "Roles", WHITEF, NAVY)
roster_headers(ws, 21)
def cat_sad(rr, r2r):
    return f'=IF({S2}!$G${r2r}="Group Data","Data","Strategy & Architecture")'
# (Holgate and the Data Capability team fall under Data; Architecture, Tech
# Strategy & AI Capability, Delivery SADA and the GM office under S&A)
last = write_roster(ws, S1, sad_coe, None, paused_zero=True)
sc(ws, f"B{last+1}", "Check - roles listed vs counted (must be 0)", NORM)
sc(ws, f"C{last+1}", f"=COUNTA(B{S1}:B{last})-C8", NORM, fmt="0", align="center")
sc(ws, f"B{last+3}",
   "Squad-based SA&D roles sit in their portfolio squads (1.3 / 4.3). Leadership roles sit on 3.0 FTE View. Planned spend excludes Paused roles - their cost is the memo line above.",
   NORM, border=False, wrap=False)
clear_region(ws, last+4, max(ws.max_row, last+10), 2, 12)
SAD_CHECK = last+1

# =====================================================================
# 2.5 Cyber Roles - re-source all 52 roles from Sheet2 (Lee's costing)
# =====================================================================
ws = wb["2.5 Cyber Roles"]
strip_dv_cf(ws)
n_cy = len(CYB)
C1, C2R = 19, 19 + n_cy - 1
clear_region(ws, 19, ws.max_row+6, 2, 12)
sc(ws, "B5", "Grouping", WHITEF, MIDBLU, align="center")
TC, DC, EC, FC = f"$T${C1}:$T${C2R}", f"$D${C1}:$D${C2R}", f"$E${C1}:$E${C2R}", f"$F${C1}:$F${C2R}"
for r, cat in ((6, "Cyber & Risk"), (7, "Service Operations")):
    sc(ws, f"B{r}", cat, BOLD)
    if cat == "Service Operations":
        sc(ws, f"C{r}", f'=COUNTIFS({DC},"Service Op & Assurance")', NORM, fmt="0", align="center")
        sc(ws, f"D{r}", f'=COUNTIFS({DC},"Service Op & Assurance",{FC},"Filled")', NORM, fmt="0", align="center")
        sc(ws, f"E{r}", f'=COUNTIFS({DC},"Service Op & Assurance",{FC},"Vacant")', NORM, fmt="0", align="center")
        sc(ws, f"F{r}", f'=SUMIFS({TC},{DC},"Service Op & Assurance")', NORM, fmt=M2, align="right")
        sc(ws, f"I{r}", f'=SUMIFS({TC},{DC},"Service Op & Assurance",{EC},"<>NZ")', NORM, fmt=M2, align="right")
        sc(ws, f"J{r}", f'=SUMIFS({TC},{DC},"Service Op & Assurance",{EC},"NZ")', NORM, fmt=M2, align="right")
    else:
        sc(ws, f"C{r}", f'=COUNTA($B${C1}:$B${C2R})-COUNTIFS({DC},"Service Op & Assurance")', NORM, fmt="0", align="center")
        sc(ws, f"D{r}", f'=COUNTIFS({FC},"Filled")-COUNTIFS({DC},"Service Op & Assurance",{FC},"Filled")', NORM, fmt="0", align="center")
        sc(ws, f"E{r}", f'=COUNTIFS({FC},"Vacant")-COUNTIFS({DC},"Service Op & Assurance",{FC},"Vacant")', NORM, fmt="0", align="center")
        sc(ws, f"F{r}", f'=SUM({TC})-SUMIFS({TC},{DC},"Service Op & Assurance")', NORM, fmt=M2, align="right")
        sc(ws, f"I{r}", f'=SUMIFS({TC},{EC},"<>NZ")-SUMIFS({TC},{DC},"Service Op & Assurance",{EC},"<>NZ")', NORM, fmt=M2, align="right")
        sc(ws, f"J{r}", f'=SUMIFS({TC},{EC},"NZ")-SUMIFS({TC},{DC},"Service Op & Assurance",{EC},"NZ")', NORM, fmt=M2, align="right")
def cat_cy(rr, r2r):
    return (f'=IF({S2}!$G${r2r}="Service Op & Assurance","Service Operations","Cyber & Risk")')
roster_headers(ws, 18)
last = write_roster(ws, C1, CYB, None, paused_zero=False)
sc(ws, f"B{last+1}", "Check - roles listed vs counted (must be 0)", NORM)
sc(ws, f"C{last+1}", f"=COUNTA(B{C1}:B{last})-C8", NORM, fmt="0", align="center")
sc(ws, f"B{last+3}",
   "Roles and costs come straight from Sheet2 (updated cyber, risk and operations roster).",
   NORM, border=False)
clear_region(ws, last+4, max(ws.max_row, last+10), 2, 12)
CY_CHECK = last+1

# =====================================================================
# 2.2 COE - draw-down formulas with BOTH allocations + portfolio count
# =====================================================================
ws = wb["2.2 COE"]
ws["E8"].value  = f"='0.0 Data Config'!$E$6+{NPORT}*'0.0 Data Config'!$L$8"
ws["E11"].value = f"='0.0 Data Config'!$E$9+{NPORT}*'0.0 Data Config'!$L$7"
# re-point SA&D refs (2.4 summary moved one column right: F->G, H->I)
ws["D8"].value  = "='2.4 SA&D'!$G$6"
ws["F8"].value  = "='2.4 SA&D'!$I$6"
ws["D12"].value = "='2.4 SA&D'!$G$7"
ws["F12"].value = "='2.4 SA&D'!$I$7"
# cyber COE line: keep the zeros as formulas, not typed constants
ws["D9"].value = "=0"
ws["E9"].value = "=0"
# the old "COE (unspecified)" parking block is gone - Sheet2 maps every role
clear_region(ws, 23, 32, 2, 10)
sc(ws, "B23",
   "Every COE role is now mapped on 2.3, 2.4 or 2.5 from the updated FTE lists on Sheet2 - no unspecified roles remain.",
   NORM, border=False)

# =====================================================================
# 1.11 TDD Cyber - keep the single-source tie to the new 2.5 cells
# =====================================================================
ws = wb["1.11 TDD Cyber"]
ws["G24"].value = "='2.5 Cyber Roles'!$F$6"
ws["G25"].value = "='2.5 Cyber Roles'!$F$7"
ws["B24"].value = "Cyber & Risk"
ws["B25"].value = "Service Operations"

# =====================================================================
# 2.1 Total Cost - ONE consolidated table: leadership + overheads inside
# the portfolio cost, COEs in the same table, $ AND FTE, total at the END
# =====================================================================
ws = wb["2.1 Total Cost"]
strip_dv_cf(ws)
clear_region(ws, 2, ws.max_row+2, 2, 13)
sc(ws, "B2", "Total Cost - archetype model vs actual organisation", Font(name="Calibri", size=14, bold=True, color="FF002F6C"), border=False)
sc(ws, "B4", "Portfolio cost in one number - squads, strategic programs, leadership and overheads", WHITEF, NAVY)
HDR21 = ["Portfolio", "Archetype cost ($m)", "Archetype squad FTE", "Actual Filled ($m)",
         "Filled FTE", "Actual Vacant ($m)", "Vacant FTE", "Actual Total ($m)",
         "Total FTE", "Over/(under) archetype ($m)", "FTE vs archetype"]
for i, h in enumerate(HDR21):
    sc(ws, f"{get_column_letter(2+i)}5", h, WHITEF, MIDBLU, align="center", wrap=True)
PORTS = [("1.1 Ampol Retail", 9, 22), ("1.2 Customer", 9, 36), ("1.3 Enterprise Data", 9, 43),
         ("1.4 TDD Group Functions", 9, 51), ("1.5 P&C", 9, 57), ("1.6 Finance", 9, 63),
         ("1.7 Infrastructure", 10, 71), ("1.8 Energy Solutions & B2B", 9, 77),
         ("1.9 Commercial Fuels", 9, 85), ("1.10 Z Retail", 9, 91)]
AD = "'Added data'"
def sumifs_port(row, status, cls):
    return (f"SUMIFS({AD}!$AA:$AA,{AD}!$AC:$AC,$B{row},{AD}!$AF:$AF,\"{cls}\","
            f"{AD}!$AG:$AG,\"{status}\")")
def countifs_port(row, status, cls):
    return (f"COUNTIFS(Squads!$N:$N,$B{row},Squads!$Q:$Q,\"{cls}\","
            f"Squads!$R:$R,\"{status}\")")
r = 6
for tab, ecell, ftrow in PORTS:
    cyber = tab == "1.11 TDD Cyber"
    sc(ws, f"B{r}", f"='{tab}'!$B$2", BOLD)
    sc(ws, f"C{r}", f"='{tab}'!$F${ecell}", NORM, fmt=M2, align="right")
    # cyber is priced from its actual roles (2.5) - no archetype FTE contract
    sc(ws, f"D{r}", '="-"' if cyber else f"='3.0 FTE View'!$G${ftrow}", NORM, fmt=M0, align="center")
    for col, st in (("E", "Filled"), ("G", "Vacant")):
        parts = "+".join(sumifs_port(r, st, c) for c in ("Squad", "Strategic Program", "Leadership"))
        sc(ws, f"{col}{r}", f"=({parts})/1000000", NORM, fmt=M2, align="right")
    for col, st in (("F", "Filled"), ("H", "Vacant")):
        parts = "+".join(countifs_port(r, st, c) for c in ("Squad", "Strategic Program", "Leadership"))
        sc(ws, f"{col}{r}", f"={parts}", NORM, fmt="0", align="center")
    sc(ws, f"I{r}", f"=E{r}+G{r}", NORM, fmt=M2, align="right")
    sc(ws, f"J{r}", f"=F{r}+H{r}", NORM, fmt="0", align="center")
    sc(ws, f"K{r}", f"=ROUND(I{r}-C{r},6)", NORM, fmt=M2, align="right")
    sc(ws, f"L{r}", '="-"' if cyber else f"=ROUND(J{r}-D{r},1)", NORM, fmt=M0, align="center")
    r += 1
# COE rows - dept-grouped, all live refs into the COE tabs
def coe_ev(sheet, rng1, rng2, crits, status):
    T_, D_, F_ = f"{sheet}!$T${rng1}:$T${rng2}", f"{sheet}!$D${rng1}:$D${rng2}", f"{sheet}!$F${rng1}:$F${rng2}"
    if crits == "ALL":
        return f'=SUMIFS({T_},{F_},"{status}")'
    if crits == "NOT_GD":
        return f'=SUMIFS({T_},{F_},"{status}")-SUMIFS({T_},{D_},"Group Data",{F_},"{status}")'
    return "=" + "+".join(f'SUMIFS({T_},{D_},"{c}",{F_},"{status}")' for c in crits)
coe_rows = [
    ("'2.2 COE'!$B$6", "'2.3 BP&T'!$F$6", "'2.3 BP&T'!$C$12", "'2.3 BP&T'", R1, PT_CHECK-1,
     ["TDD Business Partner", "Commercial"], "'2.3 BP&T'!$D$6", "'2.3 BP&T'!$E$6"),
    ("'2.2 COE'!$B$7", "'2.3 BP&T'!$F$7", None, "'2.3 BP&T'", R1, PT_CHECK-1,
     ["Transformation"], "'2.3 BP&T'!$D$7", "'2.3 BP&T'!$E$7"),
    ("'2.2 COE'!$B$8", "'2.4 SA&D'!$G$6", "'2.4 SA&D'!$C$12", "'2.4 SA&D'", S1, SAD_CHECK-1,
     "NOT_GD", "'2.4 SA&D'!$D$6", "'2.4 SA&D'!$E$6"),
    ("'2.2 COE'!$B$9", "'2.4 SA&D'!$G$7", None, "'2.4 SA&D'", S1, SAD_CHECK-1,
     ["Group Data"], "'2.4 SA&D'!$D$7", "'2.4 SA&D'!$E$7"),
    (None, "'2.5 Cyber Roles'!$F$8", None, "'2.5 Cyber Roles'", C1, CY_CHECK-1,
     "ALL", "'2.5 Cyber Roles'!$D$8", "'2.5 Cyber Roles'!$E$8"),
]
COE_FIRST = r
for lab, arch, archfte, sheet, g1, g2, crits, dcnt, ecnt in coe_rows:
    sc(ws, f"B{r}", f"={lab}" if lab else "COE - Cyber, Risk & Service Operations", BOLD)
    sc(ws, f"C{r}", f"={arch}", NORM, fmt=M2, align="right")
    sc(ws, f"E{r}", coe_ev(sheet, g1, g2, crits, "Filled"), NORM, fmt=M2, align="right")
    sc(ws, f"G{r}", coe_ev(sheet, g1, g2, crits, "Vacant"), NORM, fmt=M2, align="right")
    sc(ws, f"I{r}", f"=E{r}+G{r}", NORM, fmt=M2, align="right")
    sc(ws, f"K{r}", f"=ROUND(I{r}-C{r},6)", NORM, fmt=M2, align="right")
    for col in "DFHJL":
        sc(ws, f"{col}{r}", '="-"', NORM, align="center")
    r += 1
# central leadership + unmapped + de-dup
LEAD_ROW = r
sc(ws, f"B{r}", "Central leadership (COE + unmapped) - funded by overheads", BOLD)
sc(ws, f"C{r}", "=0", NORM, fmt=M2, align="right")
sc(ws, f"D{r}", '="-"', NORM, align="center")
sc(ws, f"E{r}", "=(" + "+".join(
    f'SUMIFS({AD}!$AA:$AA,{AD}!$AC:$AC,"{g}",{AD}!$AF:$AF,"Leadership",{AD}!$AG:$AG,"Filled")'
    for g in ("COE", "Unmapped")) + ")/1000000", NORM, fmt=M2, align="right")
sc(ws, f"F{r}", "=" + "+".join(
    f'COUNTIFS(Squads!$N:$N,"{g}",Squads!$Q:$Q,"Leadership",Squads!$R:$R,"Filled")'
    for g in ("COE", "Unmapped")), NORM, fmt="0", align="center")
sc(ws, f"G{r}", "=(" + "+".join(
    f'SUMIFS({AD}!$AA:$AA,{AD}!$AC:$AC,"{g}",{AD}!$AF:$AF,"Leadership",{AD}!$AG:$AG,"Vacant")'
    for g in ("COE", "Unmapped")) + ")/1000000", NORM, fmt=M2, align="right")
sc(ws, f"H{r}", "=" + "+".join(
    f'COUNTIFS(Squads!$N:$N,"{g}",Squads!$Q:$Q,"Leadership",Squads!$R:$R,"Vacant")'
    for g in ("COE", "Unmapped")), NORM, fmt="0", align="center")
sc(ws, f"I{r}", f"=E{r}+G{r}", NORM, fmt=M2, align="right")
sc(ws, f"J{r}", f"=F{r}+H{r}", NORM, fmt="0", align="center")
sc(ws, f"K{r}", f"=ROUND(I{r}-C{r},6)", NORM, fmt=M2, align="right")
sc(ws, f"L{r}", '="-"', NORM, align="center")
r += 1
UNM_ROW = r
sc(ws, f"B{r}", "Roles not mapped to any squad or COE", BOLD)
sc(ws, f"C{r}", "=0", NORM, fmt=M2, align="right")
sc(ws, f"D{r}", '="-"', NORM, align="center")
sc(ws, f"E{r}", f'=(SUMIFS({AD}!$AA:$AA,{AD}!$AF:$AF,"Unmapped",{AD}!$AG:$AG,"Filled"))/1000000', NORM, fmt=M2, align="right")
sc(ws, f"F{r}", '=COUNTIFS(Squads!$Q:$Q,"Unmapped",Squads!$R:$R,"Filled")', NORM, fmt="0", align="center")
sc(ws, f"G{r}", f'=(SUMIFS({AD}!$AA:$AA,{AD}!$AF:$AF,"Unmapped",{AD}!$AG:$AG,"Vacant"))/1000000', NORM, fmt=M2, align="right")
sc(ws, f"H{r}", '=COUNTIFS(Squads!$Q:$Q,"Unmapped",Squads!$R:$R,"Vacant")', NORM, fmt="0", align="center")
sc(ws, f"I{r}", f"=E{r}+G{r}", NORM, fmt=M2, align="right")
sc(ws, f"J{r}", f"=F{r}+H{r}", NORM, fmt="0", align="center")
sc(ws, f"K{r}", f"=ROUND(I{r}-C{r},6)", NORM, fmt=M2, align="right")
sc(ws, f"L{r}", '="-"', NORM, align="center")
r += 1
DEDUP_ROW = r
sc(ws, f"B{r}", "Less: Business Partner & Domain Architect funded inside portfolio overheads", BOLD)
sc(ws, f"C{r}", f"=-({NPORT}*'0.0 Data Config'!$L$7+{NPORT}*'0.0 Data Config'!$L$8)", NORM, fmt=M2, align="right")
for col in "EGI":
    sc(ws, f"{col}{r}", "=0", NORM, fmt=M2, align="right")
for col in "DFHJ":
    sc(ws, f"{col}{r}", '="-"', NORM, align="center")
sc(ws, f"K{r}", f"=ROUND(I{r}-C{r},6)", NORM, fmt=M2, align="right")
sc(ws, f"L{r}", '="-"', NORM, align="center")
r += 1
TOT_ROW = r
sc(ws, f"B{r}", "TOTAL OPERATING MODEL", WHITEF, NAVY)
for col in "CEGIK":
    sc(ws, f"{col}{r}", f"=SUM({col}6:{col}{r-1})", WHITEF, NAVY, fmt=M2, align="right")
for col in "DFHJL":
    sc(ws, f"{col}{r}", f"=SUM({col}6:{col}{r-1})", WHITEF, NAVY, fmt=M0, align="center")
r += 2
sc(ws, f"B{r}", "Restatement vs the Added data cost ledger ($m) - COE and cyber roles now costed from Sheet2, see 3.1", NORM, border=False)
sc(ws, f"I{r}", f"=ROUND(I{TOT_ROW}-SUM({AD}!$AA$2:$AA$549)/1000000,3)", NORM, fmt=M2, align="right")
RESTATE_ROW = r
r += 1
sc(ws, f"B{r}", "Memo: EGI strategic delivery FTE (Sheet2) - funded from Significant Items, not in the total above ($m)", NORM, border=False)
EGI_MEMO_ROW = r
r += 1
sc(ws, f"B{r}", "Archetype FTE covers squads. Leadership has no archetype FTE - it is funded as a dollar overhead inside the portfolio cost.", NORM, border=False)
ws.column_dimensions["B"].width = 46
for col in "CDEFGHIJKL":
    ws.column_dimensions[col].width = 13
# remap every reference to the old 2.1 layout, workbook-wide
REMAP = {"$C$43": f"$C${DEDUP_ROW}", "$C$52": f"$C${TOT_ROW}", "$D$52": f"$E${TOT_ROW}",
         "$E$52": f"$G${TOT_ROW}", "$F$52": f"$I${TOT_ROW}", "$G$52": f"$K${TOT_ROW}",
         "$F$49": f"$I${UNM_ROW}"}
pat21 = re.compile(r"('2\.1 Total Cost'!)(\$[A-Z]+\$\d+)")
for sheet in wb.worksheets:
    if sheet.title in ("Squads", "Added data", "Sheet2", "2.1 Total Cost"): continue
    for row_ in sheet.iter_rows():
        for cell in row_:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "'2.1 Total Cost'!" in v:
                cell.value = pat21.sub(lambda m: m.group(1) + REMAP.get(m.group(2), m.group(2)), v)

# =====================================================================
# 3.0 FTE View - EGI strategic delivery FTE + role language
# =====================================================================
ws = wb["3.0 FTE View"]
EGI_T = 167
sc(ws, f"B{EGI_T}", "EGI strategic delivery FTE (Sheet2) - funded from Significant Items, outside the archetype model", WHITEF, NAVY)
for col, h in zip("BCDEF", ["Name", "Role", "Engagement", "Country", "Cost ($m)"]):
    sc(ws, f"{col}{EGI_T+1}", h, WHITEF, MIDBLU, align="center")
for i, x in enumerate(EGI):
    rr = EGI_T + 2 + i
    sc(ws, f"B{rr}", f_name(x["r"]), NORM)
    sc(ws, f"C{rr}", f_title(x["r"]), NORM)
    sc(ws, f"D{rr}", f_status(x["r"]), NORM, align="center")
    sc(ws, f"E{rr}", f_ctry(x["r"]), NORM)
    sc(ws, f"F{rr}", f"={S2}!$AA${x['r']}/1000000", NORM, fmt=M2, align="right")
EGI_TOT = EGI_T + 2 + len(EGI)
sc(ws, f"B{EGI_TOT}", "Total EGI strategic delivery", BOLD, LGREY)
sc(ws, f"C{EGI_TOT}", f"=COUNTA(B{EGI_T+2}:B{EGI_TOT-1})", BOLD, LGREY, fmt="0", align="center")
sc(ws, f"F{EGI_TOT}", f"=SUM(F{EGI_T+2}:F{EGI_TOT-1})", BOLD, LGREY, fmt=M2, align="right")
# wire the 2.1 memo row to this block
w21 = wb["2.1 Total Cost"]
w21[f"I{EGI_MEMO_ROW}"].value = f"='3.0 FTE View'!$F${EGI_TOT}"
w21[f"I{EGI_MEMO_ROW}"].number_format = M2
w21[f"J{EGI_MEMO_ROW}"].value = f"='3.0 FTE View'!$C${EGI_TOT}"
w21[f"J{EGI_MEMO_ROW}"].number_format = "0"
# seat -> role language on 3.0
for addr, txt in [("H3", "Roles above archetype"), ("J6", "Roles"), ("K6", "Roles vs archetype"),
                  ("K155", "Roles"), ("B161", "COE roles (detail on 2.3 / 2.4 / 2.5)")]:
    if ws[addr].value is not None: ws[addr].value = txt

wb.save(OUT)
print("stage 2 saved. rows: dedup", DEDUP_ROW, "total", TOT_ROW, "unmapped", UNM_ROW,
      "coe_first", COE_FIRST, "egi_tot", EGI_TOT)
# =====================================================================
# 4.x GM tabs - plain titles, vacancy impact, no seat language, no gaps
# =====================================================================
GM_TABS = [t for t in wb.sheetnames if t.startswith("4.")]
gm_anchor = {}
for tab in GM_TABS:
    ws = wb[tab]
    num, suffix = tab.split(" ", 1)[0][2:], tab.split(" ", 1)[1]
    onex = f"1.{num} {suffix}"
    assert onex in wb.sheetnames, onex
    # title as a live ref to the portfolio tab's own name cell
    title_ref = None
    for c in range(1, 5):
        v = wb[onex].cell(2, c).value
        if isinstance(v, str) and v.strip():
            title_ref = f"'{onex}'!${get_column_letter(c)}$2"
            break
    ws["B2"].value = (f'=CONCATENATE({title_ref}," working copy")' if title_ref
                      else f"{suffix} working copy")
    # locate the summary table and roster
    hdr = tot = rost_hdr = None
    for r in range(3, 45):
        b = ws.cell(r, 2).value
        if b == "Squad" and hdr is None: hdr = r
        elif b == "Total" and hdr and tot is None: tot = r
        elif b == "Name" and tot: rost_hdr = r; break
    assert hdr and tot and rost_hdr, (tab, hdr, tot, rost_hdr)
    # roster blocks: squad label -> (first person row, last person row)
    blocks, cur = {}, None
    last_person = rost_hdr
    for r in range(rost_hdr+1, ws.max_row+1):
        b = ws.cell(r, 2).value
        if isinstance(b, str) and b.startswith("=Squads!"):
            if cur: blocks[cur][1] = r
            last_person = r
        elif isinstance(b, str) and b.strip() and not b.startswith("="):
            if b.startswith(("Check", "Cost", "Vacant", "Leadership", "Cyber")): break
            cur = b.strip(); blocks[cur] = [r+1, r]
    # summary rows
    ws.cell(hdr, 4).value = "Archetype roles"
    ws.cell(hdr, 8).value = "Vacancies after calls"
    for r in range(hdr+1, tot):
        name = str(ws.cell(r, 2).value or "").strip()
        g = ws.cell(r, 7).value
        if not (isinstance(g, str) and g.startswith("=")):
            blk = blocks.get(name)
            ws.cell(r, 7).value = (f'=COUNTIF(E{blk[0]}:E{blk[1]},"Hire")' if blk else "=0")
            ws.cell(r, 7).number_format = "0"
        ws.cell(r, 8).value = f"=F{r}-G{r}"
        ws.cell(r, 9).value = f'=IFERROR(E{r}+G{r}-D{r},"-")'
        ws.cell(r, 10).value = (f'=IF(ISNUMBER(D{r}),IF(E{r}>D{r},"Filled already over archetype",'
                                f'IF(E{r}+G{r}>D{r},"Over archetype after hire calls","")),'
                                f'"Outside the archetype model - no target set")')
    ws.cell(tot, 7).value = f"=SUM(G{hdr+1}:G{tot-1})"
    ws.cell(tot, 8).value = f"=SUM(H{hdr+1}:H{tot-1})"
    # cyber: the archetype column has no target - say so plainly, no text-in-number cells
    dvals = [ws.cell(r, 4).value for r in range(hdr+1, tot)]
    if not any(isinstance(v, str) and v.startswith("=") for v in dvals):
        for r in range(hdr+1, tot):
            sc(ws, f"C{r}", "Priced from actual roles on 2.5", NORM, wrap=True)
            sc(ws, f"D{r}", '="-"', NORM, align="center")
        sc(ws, f"D{tot}", '="-"', BOLD, LGREY, align="center")
    # labels / language + roster band
    for r in range(1, ws.max_row+1):
        v = ws.cell(r, 2).value
        if not isinstance(v, str) or v.startswith("="): continue
        if v == "Your position by squad": ws.cell(r, 2).value = "Position by squad"
        elif "Cost to hire all vacant" in v: ws.cell(r, 2).value = "Cost to hire all vacancies ($m)"
        elif "Cost of the seats you chose" in v: ws.cell(r, 2).value = "Cost of roles marked Hire ($m)"
        elif "Vacant seats are priced" in v:
            ws.cell(r, 2).value = "Vacant roles are priced at standard title rates - indicative until an offer is made."
        elif v.startswith("Your people"):
            ws.cell(r, 2).value = (f'=CONCATENATE({title_ref}," FTE")' if title_ref else "FTE")
            for c in range(2, 7): ws.cell(r, c).fill = NAVY; ws.cell(r, c).font = WHITEF
            for c in range(7, 11):
                ws.cell(r, c).fill = PatternFill(); ws.cell(r, c).border = Border()
            if ws.cell(r-1, 2).value is None:
                sc(ws, f"B{r-1}", "vs archetype: positive = over the allowance, negative = under.",
                   NORM, border=False)
    ws.cell(rost_hdr, 5).value = "Vacancy lever"
    for dvv in ws.data_validations.dataValidation:
        if dvv.formula1 and "Hire,Hold" in str(dvv.formula1) and "Offshore" not in str(dvv.formula1):
            dvv.formula1 = '"Hire,Hold,Offshore"'

    # close the white gap the way the owner hinted on 4.2 / 4.9: column C shows
    # the archetype type and size, live from the same FTE View row D points at.
    # K/L: archetype cost vs real cost per squad - the detail the owner asked
    # for in every working copy.
    sc(ws, f"C{hdr}", "Archetype type and size", WHITEF, MIDBLU, align="center", wrap=True)
    sc(ws, f"K{hdr}", "Archetype cost ($m)", WHITEF, MIDBLU, align="center", wrap=True)
    sc(ws, f"L{hdr}", "Actual cost ($m)", WHITEF, MIDBLU, align="center", wrap=True)
    ws.column_dimensions["K"].width = 13
    ws.column_dimensions["L"].width = 13
    cyber_tab = tab.endswith("TDD Cyber")
    for r in range(hdr+1, tot):
        d = ws.cell(r, 4).value
        m = re.match(r"^='3\.0 FTE View'!\$G\$(\d+)$", str(d or ""))
        if m:
            fr = m.group(1)
            e_, f_ = f"'3.0 FTE View'!$E${fr}", f"'3.0 FTE View'!$F${fr}"
            sc(ws, f"C{r}", f'=IF(OR({f_}=0,{f_}=""),{e_},{e_}&" - "&{f_})',
               NORM, wrap=True)
            sc(ws, f"K{r}", f"='3.0 FTE View'!$M${fr}", NORM, fmt=M2, align="right")
            sc(ws, f"L{r}", f"='3.0 FTE View'!$N${fr}", NORM, fmt=M2, align="right")
        else:
            sc(ws, f"C{r}", '="-"', NORM, align="center")
            sc(ws, f"K{r}", '="-"', NORM, align="center")
            if cyber_tab:
                sc(ws, f"L{r}", "='2.5 Cyber Roles'!$F$8", NORM, fmt=M2, align="right")
            else:
                sc(ws, f"L{r}", '="-"', NORM, align="center")
    sc(ws, f"M{hdr}", "Cost after calls ($m)", WHITEF, MIDBLU, align="center", wrap=True)
    sc(ws, f"N{hdr}", "Change vs actual ($m)", WHITEF, MIDBLU, align="center", wrap=True)
    ws.column_dimensions["M"].width = 13
    ws.column_dimensions["N"].width = 12
    for r in range(hdr+1, tot):
        name = str(ws.cell(r, 2).value or "").strip()
        blk = blocks.get(name)
        lv = str(ws.cell(r, 12).value or "")
        if blk and lv.startswith("=") and "-" not in lv[:4]:
            a, b = blk
            if cyber_tab or not title_ref:
                # cyber's L and its roster block share one Sheet2 source, so held
                # and offshored vacancies subtract from the same total cleanly
                mf = (f'=L{r}-(SUMIFS(F{a}:F{b},E{a}:E{b},"Hold")'
                      f'+0.6*SUMIFS(F{a}:F{b},E{a}:E{b},"Offshore"))/1000000')
            else:
                # squad landing cost: filled people at ledger cost, plus Hire at
                # title rate and Offshore at 0.4x; a held vacancy adds nothing.
                # L nets vacancies from ledger naming that may not match the
                # block, so building up from Filled is the only floor-safe form.
                mf = (f"=(SUMIFS('Added data'!$AA:$AA,'Added data'!$AC:$AC,{title_ref},"
                      f"'Added data'!$AE:$AE,$B{r},'Added data'!$AG:$AG,\"Filled\")"
                      f'+SUMIFS(F{a}:F{b},E{a}:E{b},"Hire")'
                      f'+0.4*SUMIFS(F{a}:F{b},E{a}:E{b},"Offshore"))/1000000')
            sc(ws, f"M{r}", mf, NORM, fmt=M2, align="right")
            sc(ws, f"N{r}", f"=M{r}-L{r}", NORM, fmt=M2, align="right")
        else:
            sc(ws, f"M{r}", '="-"', NORM, align="center")
            sc(ws, f"N{r}", '="-"', NORM, align="center")
    for col in "KL":
        sc(ws, f"{col}{tot}", f"=SUM({col}{hdr+1}:{col}{tot-1})", BOLD, LGREY, fmt=M2, align="right")
    sc(ws, f"M{tot}", f"=SUM(M{hdr+1}:M{tot-1})", BOLD, LGREY, fmt=M2, align="right")
    sc(ws, f"N{tot}", f"=SUM(N{hdr+1}:N{tot-1})", BOLD, LGREY, fmt=M2, align="right")
    gm_anchor[tab] = dict(hdr=hdr, tot=tot, rost_hdr=rost_hdr,
                          blocks={k: v for k, v in blocks.items()})
# sheet order: 4.9 before 4.10
idx9, idx10 = wb.sheetnames.index("4.9 Commercial Fuels"), wb.sheetnames.index("4.10 Z Retail")
if idx9 > idx10:
    wb.move_sheet("4.9 Commercial Fuels", offset=idx10-idx9)

# =====================================================================
# Exec Summary - role language, no baked-in numbers, no possessives
# =====================================================================
ws = wb["Exec Summary"]
EXEC_TXT = {
 "B6": "The test: can each portfolio fund its archetype cost (TDD + business)? If yes, live within it. If not, start with the vacancies (4.x GM tabs), then archetype size (1.x dropdowns).",
 "B11": "Squads are priced from the archetype library on 0.1 Squads (type x size). Offshore is priced at the offshore rate set on 0.1 Squads.",
 "B12": "Each portfolio pays one overhead: Head of Tech, Business Partner and Domain Architect shares, and leadership - rates on 0.0 Data Config.",
 "B7": "Each portfolio tab (1.x) shows the squads, sizes, support %, budget draw-downs and what is left to fund.",
 "B8": "Next step: agree funding for what is left to fund, and decide which vacancies to hire or hold.",
 "B36": "Roles the archetypes allow - squads at their set sizes",
 "B41": "Filled - people in roles today",
 "B55": "The main lever is the vacancies: they are raised but not hired, so holding them impacts nobody. Make the call role by role on the 4.x GM tab.",
 "B37": "Roles actually raised in those squads - filled + vacant",
 "B38": "Roles raised beyond the archetypes",
 "B39": "Roles in squads priced outside archetypes (AmPOS, EGI, cyber)",
 "B43": "of which squad roles - the GM hire or hold lever",
 "B44": "of which leadership, COE and unmapped roles",
 "B46": "Vacant = open roles in the raw data. Sheet2 status updates (Paused, ring-fenced) show on 2.3 to 2.5 and 3.1.",
 "B49": "Today's filled roles cost ($m)",
 "B50": "Filled roles over/(under) the archetype cost ($m)",
 "B51": "Hiring every vacancy would add ($m)",
 "B52": "of which squad roles - the 4.x GM lever ($m)",
 "B53": "Vacant roles are priced at standard title rates - indicative until an offer is made.",
 "B57": "Roles not mapped to any squad or COE ($m)",
 "B71": "Archetype squad roles allowed",
 "B72": "Org roles (excl leadership)",
 "B76": "Org roles include roles outside the archetype model; squad-only counts are on 3.0 FTE View.",
}
for addr, txt in EXEC_TXT.items():
    ws[addr].value = txt

# =====================================================================
# 3.1 Data QA - Sheet2 reconciliation with live refs (no typed data)
# =====================================================================
ws = wb["3.1 Data QA"]
for addr, txt in [("B71", "Roles by model squad - raw data vs Added data (differences only)"),
                  ("C72", "raw data roles"), ("D72", "Added data roles"),
                  ("B11", "In Added data only - joined the org after the raw data cut, or naming mismatch"),
                  ("B51", "In raw data only - left the org, or naming mismatch")]:
    if ws[addr].value is not None: ws[addr].value = txt
qr = ws.max_row + 3
def qhdr(text):
    global qr
    sc(ws, f"B{qr}", text, WHITEF, NAVY); qr += 1
def qcols(*labels):
    global qr
    for i, l in enumerate(labels):
        sc(ws, f"{get_column_letter(2+i)}{qr}", l, WHITEF, MIDBLU, align="center")
    qr += 1
def qnote(text):
    global qr
    sc(ws, f"B{qr}", text, NORM, border=False); qr += 1

# pair Sheet2 rows to Squads rows: exact name+title, then title within division
sq_by_key, sq_by_div = {}, {}
for q in sq_rows:
    sq_by_key.setdefault((q["name"], q["title"]), []).append(q)
    sq_by_div.setdefault(q["div"], []).append(q)
used = set()
pairs, s2_only = [], []
for x in rows2:
    k = (low(x["name"]), low(x["title"]))
    cand = [q for q in sq_by_key.get(k, []) if q["r"] not in used and q["div"] == low(x["div"])] \
        or [q for q in sq_by_key.get(k, []) if q["r"] not in used]
    if cand:
        used.add(cand[0]["r"]); pairs.append((x, cand[0])); continue
    # title-only pairing is only safe for unnamed rows (Vacant / ring fenced);
    # a NAMED person with no exact match is a new joiner and must be listed
    if "vacant" in low(x["name"]) or "ring fenced" in low(x["name"]):
        tcand = [q for q in sq_by_div.get(low(x["div"]), [])
                 if q["r"] not in used and q["title"] == low(x["title"])]
        if tcand:
            used.add(tcand[0]["r"]); pairs.append((x, tcand[0])); continue
    s2_only.append(x)
DIV4 = {"cyber, risk & operations", "egi", "partnering & transformation",
        "strategy, architecture & data"}
sq_only = [q for q in sq_rows if q["div"] in DIV4 and q["r"] not in used]
def s2_open(x): return x["typ"] in ("v", "pause")
def sq_open(q): return q["st"] == "Vacant"
status_diffs = [(x, q) for x, q in pairs
                if s2_open(x) != sq_open(q) or (x["typ"] == "pause" and q["st"] == "Vacant")]

qhdr("Sheet2 reconciliation - the updated FTE lists vs the raw data (Squads). Sheet2 drives 2.3 / 2.4 / 2.5 and the EGI FTE list; Squads drives the portfolio squads.")
qcols("Name (Sheet2)", "Role (Sheet2)", "Division", "Raw data status", "Sheet2 status")
for x, q in status_diffs:
    sc(ws, f"B{qr}", f_name(x["r"]), NORM)
    sc(ws, f"C{qr}", f_title(x["r"]), NORM)
    sc(ws, f"D{qr}", f"={S2}!$F${x['r']}", NORM)
    sc(ws, f"E{qr}", f"=Squads!$R${q['r']}", NORM, align="center")
    sc(ws, f"F{qr}", f_status(x["r"]), NORM, align="center")
    qr += 1
qnote("Where the two disagree, the Sheet2 status is used on 2.3 / 2.4 / 2.5. The raw data still drives the portfolio squad counts on 1.x / 3.0 / 4.x.")
qr += 1
qhdr("Roles in Sheet2 with no matching raw data row")
qcols("Name (Sheet2)", "Role (Sheet2)", "Division", "Where it now shows")
for x in s2_only:
    if x["div"] == "EGI": where = "3.0 FTE View - EGI strategic delivery FTE"
    elif x["port"] == "enterprise data" and not blankish(x["squad"]):
        where = "In an Enterprise Data squad per Sheet2 - needs a raw data row to join the 1.3 / 4.3 counts"
    elif x["port"] == "enterprise data": where = "Nowhere yet - needs a raw data row to join the Enterprise Data portfolio"
    elif x["div"] == "Cyber, Risk & Operations": where = "2.5 Cyber Roles"
    elif x["div"] == "Partnering & Transformation": where = "2.3 BP&T"
    else: where = "2.4 SA&D"
    sc(ws, f"B{qr}", f_name(x["r"]), NORM)
    sc(ws, f"C{qr}", f_title(x["r"]), NORM)
    sc(ws, f"D{qr}", f"={S2}!$F${x['r']}", NORM)
    sc(ws, f"E{qr}", where, NORM)
    qr += 1
qr += 1
if sq_only:
    qhdr("Raw data rows (these four divisions) with no matching Sheet2 row - check if dropped or renamed")
    qcols("Name (raw data)", "Role (raw data)", "Model portfolio", "Status")
    for q in sq_only:
        sc(ws, f"B{qr}", f"=Squads!$B${q['r']}", NORM)
        sc(ws, f"C{qr}", f'=SUBSTITUTE(SUBSTITUTE(Squads!$C${q["r"]},"–","-"),"—","-")', NORM)
        sc(ws, f"D{qr}", f"=Squads!$N${q['r']}", NORM)
        sc(ws, f"E{qr}", f"=Squads!$R${q['r']}", NORM, align="center")
        qr += 1
    qr += 1
qhdr("Raw data integrity audit")
qnote("Zero rows were added to the raw data: the Squads tab in this workbook matches the owner's uploaded copy cell for cell.")
qnote("16 mapping cells on 4 Commercial rows were changed in an earlier build (Unmapped to COE - Business Partnering). The uploaded copy already includes that change, so the two files match today. The names below are live references to those rows.")
for rr in (260, 269, 270, 281):
    sc(ws, f"B{qr}", f"=Squads!$B${rr}", NORM)
    sc(ws, f"C{qr}", f'=SUBSTITUTE(SUBSTITUTE(Squads!$C${rr},"–","-"),"—","-")', NORM)
    sc(ws, f"D{qr}", f"=Squads!$P${rr}", NORM)
    sc(ws, f"E{qr}", f"=Squads!$R${rr}", NORM, align="center")
    qr += 1
qnote("Sheet2 confirms these four roles sit in Partnering & Transformation (Commercial), inside the Business Partnering bucket on 2.3.")
qr += 1
qhdr("Cost bases")
qnote("2.3 / 2.4 / 2.5 and the EGI FTE list are costed from Sheet2 (Full Cost AUD, column AA). Portfolio actuals on 2.1 are costed from the Added data ledger. The difference is the restatement line on 2.1.")

# =====================================================================
# STAGE 4a: 2.0 Group Summary - the owner's column spec + AU/NZ variances
# =====================================================================
g0 = wb["2.0 Group Summary"]
HDR20 = {"C": "TDD Lights On Budget ($m)", "D": "Support Cost ($m)",
         "E": "Variance ($m) = budget - cost", "G": "Cost of non-TDD funding ($m)",
         "H": "Amount that can be recharged ($m)", "I": "Left to fund outside TDD ($m)",
         "J": "Total still left to fund ($m)", "K": "Total Cost ($m)"}
for col, h in HDR20.items():
    sc(g0, f"{col}5", h, WHITEF, MIDBLU, align="center", wrap=True)
for r in list(range(6, 17)) + list(range(18, 23)):
    oldj = g0.cell(r, 10).value
    if oldj is None: continue
    sc(g0, f"K{r}", oldj, NORM, fmt=M2, align="right")
    sc(g0, f"J{r}", f"=MAX(0,-E{r})+I{r}", NORM, fmt=M2, align="right")
for r, kfx in ((17, "=SUM(K6:K16)"), (24, "=SUM(K17,K18,K19,K20,K21,K22)")):
    sc(g0, f"K{r}", kfx, BOLD, LGREY, fmt=M2, align="right")
    sc(g0, f"J{r}", "=SUM(J6:J16)" if r == 17 else "=SUM(J17,J18,J19,J20,J21,J22)",
       BOLD, LGREY, fmt=M2, align="right")
# net-cost lines move to K; note 2.0 is the archetype view
for r in (25, 26):
    v = g0.cell(r, 10).value
    if v is not None:
        sc(g0, f"K{r}", str(v).replace("J24", "K24"), BOLD, fmt=M2, align="right")
        g0.cell(r, 10).value = None
sc(g0, "B3", "Archetype view: what the designed model costs against the TDD budgets. The actual organisation is on 2.1.", NORM, border=False)
# remap references to old 2.0 J cells (total cost) -> K
patJ = re.compile(r"('2\.0 Group Summary'!\$?J\$?)(\d+)")
for sheet in wb.worksheets:
    if sheet.title in ("Squads", "Added data", "Sheet2", "2.0 Group Summary"): continue
    for row_ in sheet.iter_rows():
        for cell in row_:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "'2.0 Group Summary'!" in v:
                nv = v.replace("'2.0 Group Summary'!$J$6:$J$16", "'2.0 Group Summary'!$K$6:$K$16")
                cell.value = nv.replace("'2.0 Group Summary'!$J$", "'2.0 Group Summary'!$K$")

# =====================================================================
# STAGE 4b: Fund toggles (AU/NZ) on every 1.x squad table + tab subtotals
# =====================================================================
AUNZ = {}
dvfund_all = []
for tab in [t for t in wb.sheetnames if t.startswith("1.") and t[2] in ".0123456789"]:
    ws = wb[tab]
    tables = []
    for r in range(15, ws.max_row + 1):
        if ws.cell(r, 2).value == "Squad" and ws.cell(r, 8).value == "TDD Cost ($m)":
            rows = []
            rr = r + 1
            while rr <= ws.max_row:
                b = str(ws.cell(rr, 2).value or "")
                if not b or "Overhead" in b or b.endswith("Total"): break
                rows.append(rr); rr += 1
            if rows: tables.append((r, rows))
    if not tables: continue
    allrows = [rr for _, rows in tables for rr in rows]
    fcol = next(c for c in range(10, 15)
                if all(ws.cell(rr, c).value is None for rr in allrows)
                and all(ws.cell(h, c).value is None for h, _ in tables))
    fL = get_column_letter(fcol)
    dv = DataValidation(type="list", formula1='"AU,NZ"', allow_blank=False, showErrorMessage=True)
    ws.add_data_validation(dv)
    for h, rows in tables:
        sc(ws, f"{fL}{h}", "Fund", WHITEF, MIDBLU, align="center")
        plat = ""
        for up in range(h - 1, max(1, h - 6), -1):
            t_ = str(ws.cell(up, 2).value or "")
            if t_.startswith("Platform"): plat = t_; break
        default = "NZ" if (tab == "1.10 Z Retail" or " Z " in plat + " " or plat.startswith("Platform: Z")) else "AU"
        for rr in rows:
            sc(ws, f"{fL}{rr}", default, NORM, YELL, align="center")
            dv.add(f"{fL}{rr}")
    terms_au = "+".join(f'SUMIF({fL}{rows[0]}:{fL}{rows[-1]},"AU",H{rows[0]}:H{rows[-1]})' for _, rows in tables)
    terms_nz = "+".join(f'SUMIF({fL}{rows[0]}:{fL}{rows[-1]},"NZ",H{rows[0]}:H{rows[-1]})' for _, rows in tables)
    base = ws.max_row + 2
    sc(ws, f"B{base}", "TDD squad cost - AU funded (Fund toggles) ($m)", NORM)
    sc(ws, f"C{base}", f"={terms_au}", NORM, fmt=M2, align="right")
    sc(ws, f"B{base+1}", "TDD squad cost - NZ funded (Fund toggles) ($m)", NORM)
    sc(ws, f"C{base+1}", f"={terms_nz}", NORM, fmt=M2, align="right")
    AUNZ[tab] = (f"C{base}", f"C{base+1}")
au_sum = "+".join(f"'{t}'!${a.replace('C','C$')}" for t, (a, _) in AUNZ.items())
nz_sum = "+".join(f"'{t}'!${b.replace('C','C$')}" for t, (_, b) in AUNZ.items())
r0 = 37
sc(g0, f"B{r0}", "AU / NZ funding split - from the squad Fund toggles on the 1.x tabs", WHITEF, NAVY)
aunz_rows = [
    ("AU budget - 0.0 Data Config ($m)", "='0.0 Data Config'!$C$27"),
    ("AU allocated to COEs - 0.0 Data Config ($m)", "=SUM('0.0 Data Config'!$C$6:$C$10)"),
    ("TDD squad cost - AU funded ($m)", f"={au_sum}"),
    ("Variance vs AU budget ($m)", f"=C{r0+1}-C{r0+2}-C{r0+3}"),
    ("NZ budget - 0.0 Data Config ($m)", "='0.0 Data Config'!$D$27"),
    ("NZ allocated to COEs - 0.0 Data Config ($m)", "=SUM('0.0 Data Config'!$D$6:$D$10)"),
    ("TDD squad cost - NZ funded ($m)", f"={nz_sum}"),
    ("Variance vs NZ budget ($m)", f"=C{r0+5}-C{r0+6}-C{r0+7}"),
]
for i, (lab, fx) in enumerate(aunz_rows):
    rr = r0 + 1 + i
    bold = lab.startswith("Variance")
    sc(g0, f"B{rr}", lab, BOLD if bold else NORM)
    sc(g0, f"C{rr}", fx, BOLD if bold else NORM, fmt=M2, align="right")
sc(g0, f"B{r0+9}", "Portfolio overheads and platform overheads sit inside the COE and portfolio allocations on 0.0 Data Config.", NORM, border=False)
# =====================================================================
# STAGE 4c: working tabs 4.12 / 4.13 / 4.14 - every role gets a 4.x home
# =====================================================================
def mk_working(name, title, after):
    if name in wb.sheetnames: del wb[name]
    ws = wb.create_sheet(name)
    wb.move_sheet(name, offset=wb.sheetnames.index(after) - wb.sheetnames.index(name) + 1)
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 3), ("B", 30), ("C", 40), ("D", 26), ("E", 12), ("F", 10), ("G", 16)):
        ws.column_dimensions[col].width = w
    sc(ws, "B2", title, Font(name="Calibri", size=14, bold=True, color="FF002F6C"), border=False)
    return ws
def w_hdrs(ws, r):
    for col, h in zip("BCDEFG", ["Name", "Role", "Department", "Status", "Vacancy lever", "Cost if hired ($)"]):
        sc(ws, f"{col}{r}", h, WHITEF, MIDBLU, align="center")
def w_rows_sheet2(ws, first, entries):
    dv = DataValidation(type="list", formula1='"Hire,Hold,Offshore"', allow_blank=False, showErrorMessage=True)
    ws.add_data_validation(dv)
    for i, x in enumerate(entries):
        rr = first + i
        sc(ws, f"B{rr}", f_name(x["r"]), NORM)
        sc(ws, f"C{rr}", f_title(x["r"]), NORM, wrap=True)
        sc(ws, f"D{rr}", f_dept(x["r"]), NORM)
        sc(ws, f"E{rr}", f_status(x["r"]), NORM, align="center")
        if x["typ"] in ("v", "pause"):
            sc(ws, f"F{rr}", "Hold", NORM, YELL, align="center"); dv.add(f"F{rr}")
            sc(ws, f"G{rr}", f_cost(x["r"]), NORM, fmt=D0, align="right")
    return first + len(entries) - 1
def w_summary(ws, first, last, top=4):
    sc(ws, f"B{top}", "Position", WHITEF, NAVY)
    items = [("Roles", f"=COUNTA(B{first}:B{last})", "0"),
             ("Filled", f'=COUNTIF(E{first}:E{last},"Filled")+COUNTIF(E{first}:E{last},"Contractor")', "0"),
             ("Vacant", f'=COUNTIF(E{first}:E{last},"Vacant")', "0"),
             ("Paused", f'=COUNTIF(E{first}:E{last},"Paused")', "0"),
             ("Cost to hire all vacancies ($m)", f'=SUMIF(E{first}:E{last},"Vacant",G{first}:G{last})/1000000', M2),
             ("Cost of roles marked Hire ($m)", f'=SUMIF(F{first}:F{last},"Hire",G{first}:G{last})/1000000', M2),
             ("Cost of calls - Hire + 0.4 x Offshore ($m)",
              f'=(SUMIF(F{first}:F{last},"Hire",G{first}:G{last})+0.4*SUMIF(F{first}:F{last},"Offshore",G{first}:G{last}))/1000000', M2)]
    for i, (lab, fx, fmt) in enumerate(items):
        sc(ws, f"B{top+1+i}", lab, NORM)
        sc(ws, f"C{top+1+i}", fx, NORM, fmt=fmt, align="right")
def w_money(ws, spend, budget, left):
    for i, (lab, fx) in enumerate([("Planned spend ($m)", spend),
                                   ("Budget to draw down ($m)", budget),
                                   ("Left to fund ($m)", left)]):
        sc(ws, f"E{5+i}", lab, NORM)
        sc(ws, f"F{5+i}", fx, NORM, fmt=M2, align="right")
ws = mk_working("4.12 BP&T", "Business Partnering & Transformation working copy", "4.11 TDD Cyber")
w_hdrs(ws, 13); last = w_rows_sheet2(ws, 14, PT); w_summary(ws, 14, last)
w_money(ws, "='2.3 BP&T'!$F$8", "='2.3 BP&T'!$G$8", "='2.3 BP&T'!$H$8")
sc(ws, f"B{last+2}", "Funding for these roles is on 2.3 BP&T. Roles and costs come from Sheet2 - costs show on vacancies only.", NORM, border=False)
W412 = (14, last)
ws = mk_working("4.13 SA&D", "Strategy, Architecture & Data working copy", "4.12 BP&T")
w_hdrs(ws, 13); last = w_rows_sheet2(ws, 14, sad_coe); w_summary(ws, 14, last)
w_money(ws, "='2.4 SA&D'!$G$8", "='2.4 SA&D'!$H$8", "='2.4 SA&D'!$I$8")
sc(ws, f"B{last+2}", "COE roles only - squad-based SA&D roles sit on 4.3 (Group Data portfolio). Funding is on 2.4 SA&D.", NORM, border=False)
W413 = (14, last)
ws = mk_working("4.14 EGI & Central", "EGI & Central Roles working copy", "4.13 SA&D")
sc(ws, "B12", "EGI strategic delivery (Sheet2) - contractors, funded from Significant Items", WHITEF, NAVY)
w_hdrs(ws, 13); last = w_rows_sheet2(ws, 14, EGI)
egi_last = last
# every Squads row must have a 4.x home: rows already referenced on 4.1-4.11,
# COE rows carried by the Sheet2 rosters on 4.12/4.13, leadership here, and
# EVERYTHING left over (unmapped, stray squads, raw EGI rows) here too
refset = set()
for t4 in [t for t in wb.sheetnames if t.startswith("4.") and t not in ("4.12 BP&T", "4.13 SA&D", "4.14 EGI & Central")]:
    for row_ in wb[t4].iter_rows(min_col=2, max_col=2):
        m_ = re.match(r"^=Squads!\$B\$(\d+)$", str(row_[0].value or ""))
        if m_: refset.add(int(m_.group(1)))
s2cov = set()
for x in PT + sad_coe:            # rows carried on 4.12 / 4.13 via Sheet2
    m_ = sq_match(x)
    if m_: s2cov.add(m_["r"])
lead_rows = [q for q in sq_rows if q["cls"] == "Leadership"]
leftover_rows = [q for q in sq_rows if q["r"] not in refset and q["r"] not in s2cov
                 and q["cls"] != "Leadership"]
sec = last + 2
sc(ws, f"B{sec}", "Leadership roles - funded via portfolio overheads", WHITEF, NAVY)
w_hdrs(ws, sec + 1)
dv2 = DataValidation(type="list", formula1='"Hire,Hold,Offshore"', allow_blank=False, showErrorMessage=True)
ws.add_data_validation(dv2)
def w_rows_squads(ws, first, entries, dv):
    for i, q in enumerate(entries):
        rr = first + i; n = q["r"]
        sc(ws, f"B{rr}", f'=IF(Squads!$R${n}="Vacant","Vacant",Squads!$B${n})', NORM)
        sc(ws, f"C{rr}", f'=SUBSTITUTE(SUBSTITUTE(Squads!$C${n},"–","-"),"—","-")', NORM, wrap=True)
        sc(ws, f"D{rr}", f"=Squads!$N${n}", NORM)
        sc(ws, f"E{rr}", f"=Squads!$R${n}", NORM, align="center")
        if q["st"] == "Vacant":
            sc(ws, f"F{rr}", "Hold", NORM, YELL, align="center"); dv.add(f"F{rr}")
            sc(ws, f"G{rr}", f"=SUMIF('Added data'!$C$2:$C$549,Squads!$C${n},'Added data'!$AA$2:$AA$549)"
                             f"/COUNTIF('Added data'!$C$2:$C$549,Squads!$C${n})", NORM, fmt=D0, align="right")
    return first + len(entries) - 1
last = w_rows_squads(ws, sec + 2, lead_rows, dv2)
LEAD_BLK = (sec + 2, last)
sec2 = last + 2
sc(ws, f"B{sec2}", "Every other role - unmapped, outside the archetype model, or raw data EGI rows", WHITEF, NAVY)
w_hdrs(ws, sec2 + 1)
last = w_rows_squads(ws, sec2 + 2, leftover_rows, dv2)
UNM_BLK = (sec2 + 2, last)
# new-in-Sheet2 people with no raw data row yet - they still need a 4.x home
newins = [x for x in rows2 if x["div"] != "EGI" and x not in PT and x not in sad_coe
          and sq_match(x) is None]
sec3 = last + 2
NEW_BLK = (0, -1)
if newins:
    sc(ws, f"B{sec3}", "New in Sheet2 - no raw data row yet. Add them to the Squads tab to join the portfolio counts.", WHITEF, NAVY)
    w_hdrs(ws, sec3 + 1)
    dv3 = DataValidation(type="list", formula1='"Hire,Hold,Offshore"', allow_blank=False, showErrorMessage=True)
    ws.add_data_validation(dv3)
    first3 = sec3 + 2
    for i, x in enumerate(newins):
        rr = first3 + i
        sc(ws, f"B{rr}", f_name(x["r"]), NORM)
        sc(ws, f"C{rr}", f_title(x["r"]), NORM, wrap=True)
        sc(ws, f"D{rr}", f"={S2}!$F${x['r']}", NORM)
        sc(ws, f"E{rr}", f_status(x["r"]), NORM, align="center")
        if x["typ"] in ("v", "pause"):
            sc(ws, f"F{rr}", "Hold", NORM, YELL, align="center"); dv3.add(f"F{rr}")
            sc(ws, f"G{rr}", f_cost(x["r"]), NORM, fmt=D0, align="right")
    last = first3 + len(newins) - 1
    NEW_BLK = (first3, last)
w_summary(ws, 14, last)
# roles count must span only the person rows of the blocks
cnt = (f"=COUNTA(B14:B{egi_last})+COUNTA(B{LEAD_BLK[0]}:B{LEAD_BLK[1]})"
       f"+COUNTA(B{UNM_BLK[0]}:B{UNM_BLK[1]})")
if newins: cnt += f"+COUNTA(B{NEW_BLK[0]}:B{NEW_BLK[1]})"
sc(ws, "C5", cnt, NORM, fmt="0", align="right")
sc(ws, f"B{last+2}", "Every role that is not in a portfolio squad (4.1-4.11) or a COE (4.12-4.13) lives here - nothing is missed.", NORM, border=False)

# =====================================================================
# STAGE 4d: 0.5 Guide - how the whole workbook flows, what to edit
# =====================================================================
if "0.5 Guide" in wb.sheetnames: del wb["0.5 Guide"]
gd = wb.create_sheet("0.5 Guide")
wb.move_sheet("0.5 Guide", offset=wb.sheetnames.index("0.4 Budget Table (Fin)") - wb.sheetnames.index("0.5 Guide") + 1)
gd.sheet_view.showGridLines = False
gd.column_dimensions["A"].width = 3
gd.column_dimensions["B"].width = 120
sc(gd, "B2", "How this workbook flows", Font(name="Calibri", size=16, bold=True, color="FF002F6C"), border=False)
GUIDE = [
 ("h", "The spine"),
 ("t", "Inputs (0.x) feed the archetype model (1.x), which rolls into the funding view (2.0). The actual organisation (Squads tab + Sheet2 + Added data) rolls into the affordability view (2.1). GMs make Hire or Hold calls on the working tabs (4.x). Evidence and checks live on 3.1."),
 ("h", "What you can edit (yellow cells are inputs)"),
 ("t", "0.0 Data Config - TDD budgets by portfolio (AU and NZ), overhead rates, COE allocations."),
 ("t", "0.1 Squads - the archetype price library (type x size) and the offshore rate."),
 ("t", "1.x squad tables - squad type, size, On/Off, support %, and the Fund toggle (AU or NZ) that decides which budget the squad draws."),
 ("t", "Squads tab - the raw organisation and its model mapping (columns N to R). This drives every role count."),
 ("t", "Sheet2 - the updated FTE lists for BP&T, SA&D, Cyber and EGI. This drives 2.3, 2.4, 2.5 and the EGI FTE list."),
 ("t", "4.x Call cells - Hire or Hold on every vacancy."),
 ("h", "If you change X, Y updates"),
 ("t", "Change a squad type, size, On/Off, support % or Fund toggle on 1.x - the squad price, 2.0, 2.1 archetype column, 3.0 and the Exec Summary all update."),
 ("t", "Change a status or mapping in the Squads tab - role counts on 3.0 and 4.x, vacancy counts everywhere, and the actual columns of 2.1 update."),
 ("t", "Change a person or cost in Sheet2 - 2.3 / 2.4 / 2.5, the EGI FTE list, 2.2, the COE rows of 2.0 and 2.1, and tabs 4.12 / 4.13 / 4.14 update."),
 ("t", "Change a budget on 0.0 Data Config - budget and variance columns on 2.0, the draw-downs on 2.2 / 2.3 / 2.4, and left to fund everywhere update."),
 ("t", "Make a Hire or Hold call on any 4.x tab - that tab's cost of roles marked Hire updates. Totals only move when a role is actually hired in the raw data."),
 ("h", "What each tab is"),
 ("t", "Exec Summary - the whole story on one page. | 0.0-0.4 - inputs and reference tables. | 1.1-1.11 - one archetype build per portfolio. | 2.0 - archetype cost vs the TDD budgets (the funding view). | 2.1 - archetype vs what the org actually costs (the affordability view), portfolio level; squad detail on 3.0. | 2.2 - COE budget roll-up and the TDD corporate funding pool. | 2.3 / 2.4 / 2.5 - the COE and cyber rosters and their funding. | 3.0 - squad-level detail: archetype FTE and cost vs actual, drills 2.1. | 3.1 - data checks, reconciliations and the audit trail. | 4.1-4.11 - GM working copies per portfolio. | 4.12-4.14 - working copies for BP&T, SA&D and EGI & central roles. | Squads / Sheet2 / Added data - source data."),
 ("h", "How to trust it"),
 ("t", "Every tab with a row labelled Check must show 0. All reconciliations and the raw-data audit live on 3.1. Every role in the organisation appears on exactly one 4.x working tab - if a role were missed, the checks on 3.1 would not foot."),
]
gr_ = 4
for kind, txt in GUIDE:
    if kind == "h":
        sc(gd, f"B{gr_}", txt, WHITEF, NAVY); gr_ += 1
    else:
        c = sc(gd, f"B{gr_}", txt, NORM, border=False, wrap=True)
        gd.row_dimensions[gr_].height = max(15, 15 * (1 + len(txt) // 130))
        gr_ += 1
sc(gd, f"B{gr_+1}", "3.0 FTE View is the squad-level drill of 2.1. 2.0 shows the archetype view only - actuals live on 2.1.", NORM, border=False)

# retitle 3.0 so its place in the flow is explicit
wb["3.0 FTE View"]["B2"].value = "Squad detail - archetype vs actual by portfolio, platform and squad (drills 2.1)"
# 2.2 header: roll-up only
wb["2.2 COE"]["B3"].value = "Roll-up only - edit budgets on 0.0 Data Config, rosters on Sheet2. Detail on 2.3 / 2.4 / 2.5."

# =====================================================================
# STAGE 4e: the mapping change log on 3.1 (full disclosure)
# =====================================================================
if CHANGELOG:
    qr += 1
    qhdr("Model mapping updates made in this build - the owner's SA&D COE list (ruthless mapping)")
    qcols("Raw row", "Cell", "Old value", "New value", "Role")
    for row, colL, old, new, nm in CHANGELOG:
        sc(ws31_ := wb["3.1 Data QA"], f"B{qr}", row, NORM, align="center")
        sc(ws31_, f"C{qr}", f"{colL}{row}", NORM, align="center")
        sc(ws31_, f"D{qr}", str(old), NORM)
        sc(ws31_, f"E{qr}", new, NORM)
        sc(ws31_, f"F{qr}", f"=Squads!$B${row}&\" - \"&Squads!$C${row}", NORM)
        qr += 1
    qnote("These cells were updated at the owner's direction: every role on the owner's SA&D COE list now maps to COE - Data or COE - Strategy Architecture. The Group Data portfolio keeps the engineering, science, operations and reporting squads.")

# =====================================================================
# STAGE 6: cyber is ONE COE (3 funding buckets), COE Summary redesign,
# AU/NZ inside each 1.x Portfolio Summary, exec story, Customer AI
# =====================================================================
# 6a. cyber funding buckets on the roles tab
wc = wb["2.5 Cyber Roles"]
clear_region(wc, 10, 16, 2, 4)
fundc = [
    ("Funding buckets to draw down", None, None),
    ("COE - Cyber allocation ($m) - 0.0 Data Config", "='0.0 Data Config'!$E$7", M2),
    ("TDD Cyber budget ($m) - 0.0 Data Config", "='0.0 Data Config'!$E$23", M2),
    ("Cyber CapEx - Monitoring ($m) - input", 0.5, M2),
    ("Total budget to draw down ($m)", "=SUM(C11:C13)", M2),
    ("Left to fund ($m)", "=MAX(0,F8-C14)", M2),
]
for i6, (lab6, fx6, fmt6) in enumerate(fundc):
    rr6 = 10 + i6
    if fx6 is None:
        sc(wc, f"B{rr6}", lab6, WHITEF, NAVY); continue
    bold6 = lab6.startswith(("Total", "Left"))
    sc(wc, f"B{rr6}", lab6, BOLD if bold6 else NORM)
    sc(wc, f"C{rr6}", fx6, BOLD if bold6 else NORM,
       YELL if isinstance(fx6, float) else None, fmt=fmt6, align="right")
wc["G8"].value = "=C14"
wc["H8"].value = "=MAX(0,F8-G8)"
sc(wc, "I8", "=SUM(I6:I7)", BOLD, LGREY, fmt=M2, align="right")
sc(wc, "J8", "=SUM(J6:J7)", BOLD, LGREY, fmt=M2, align="right")
sc(wc, "I5", "Cost - AU ($m)", WHITEF, MIDBLU, align="center", wrap=True)
sc(wc, "J5", "Cost - NZ ($m)", WHITEF, MIDBLU, align="center", wrap=True)
# 6b. remove the 1.11 portfolio tab - cyber lives once, as a COE
for sheet6 in wb.worksheets:
    for row6 in sheet6.iter_rows():
        for cell6 in row6:
            v6 = cell6.value
            if isinstance(v6, str) and v6.startswith("=") and "'1.11 TDD Cyber'!" in v6:
                if "$C$9" in v6 or "$E$9" in v6:
                    cell6.value = v6.replace("'1.11 TDD Cyber'!$C$9", "'2.5 Cyber Roles'!$F$8").replace("'1.11 TDD Cyber'!$E$9", "'2.5 Cyber Roles'!$F$8")
                elif "$H$5" in v6:
                    cell6.value = v6.replace("'1.11 TDD Cyber'!$H$5", "('0.0 Data Config'!$E$7+'0.0 Data Config'!$E$23)")
                elif "I$13" in v6:
                    cell6.value = v6.replace("+'1.11 TDD Cyber'!$I$13", "").replace("'1.11 TDD Cyber'!$I$13", "0")
                elif "$B$2" in v6:
                    cell6.value = "TDD Cyber working copy"
                else:
                    cell6.value = "='2.5 Cyber Roles'!$F$8"
w411f = wb["4.11 TDD Cyber"]
for r6 in range(3, 20):
    if r6 != 2 and w411f.cell(r6, 2).value == "TDD Cyber working copy":
        w411f.cell(r6, 2).value = "TDD Cyber FTE"
ftc = wb["3.0 FTE View"]
ftc["E92"].value = "Strategic Programs"
ftc["F92"].value = "-"
ftc["M92"].value = "='2.5 Cyber Roles'!$F$8"
ftc["N92"].value = "='2.5 Cyber Roles'!$F$8"
del wb["1.11 TDD Cyber"]
# 6c. Group Summary: cyber out of the portfolio block, real COE cyber row
g6 = wb["2.0 Group Summary"]
clear_region(g6, 16, 16, 2, 12)
for col6 in "CDEFGHIJK":
    v17 = g6[f"{col6}17"].value
    if isinstance(v17, str):
        g6[f"{col6}17"].value = v17.replace(f"{col6}6:{col6}16", f"{col6}6:{col6}15")
for r6 in range(6, 16):
    g6[f"E{r6}"].value = f"=C{r6}-D{r6}"
g6["B19"].value = "COE - Cyber, Risk & Service Operations"
g6["C19"].value = "='0.0 Data Config'!$E$7+'0.0 Data Config'!$E$23"
g6["D19"].value = "='2.5 Cyber Roles'!$F$8"
g6["E19"].value = "=C19-D19"
g6["G19"].value = "=0"
g6["H19"].value = "='2.5 Cyber Roles'!$G$8"
g6["I19"].value = "='2.5 Cyber Roles'!$H$8"
g6["J19"].value = "=MAX(0,-E19)+I19"
g6["K19"].value = "=D19"
g6["C29"].value = "=SUM('0.0 Data Config'!$E$6:$E$10)+'0.0 Data Config'!$E$23"
# 6d. COE Summary redesign - five COEs, one clean grid
w22 = wb["2.2 COE"]
strip_dv_cf(w22)
clear_region(w22, 2, w22.max_row + 2, 2, 10)
sc(w22, "B2", "COE Summary", Font(name="Calibri", size=14, bold=True, color="FF002F6C"), border=False)
sc(w22, "B4", "The five COEs - roles, spend and funding", WHITEF, NAVY)
for col6, h6 in zip("BCDEFGH", ["COE", "Roles", "Filled", "Vacant", "Planned spend ($m)",
                                "Budget to draw down ($m)", "Left to fund ($m)"]):
    sc(w22, f"{col6}5", h6, WHITEF, MIDBLU, align="center", wrap=True)
COEG = [("Business Partnering", "'2.3 BP&T'", ["C6", "D6", "E6", "F6", "G6", "H6"]),
        ("Transformation", "'2.3 BP&T'", ["C7", "D7", "E7", "F7", "G7", "H7"]),
        ("Strategy & Architecture", "'2.4 SA&D'", ["C6", "D6", "E6", "G6", "H6", "I6"]),
        ("Data", "'2.4 SA&D'", ["C7", "D7", "E7", "G7", "H7", "I7"]),
        ("Cyber, Risk & Service Operations", "'2.5 Cyber Roles'", ["C8", "D8", "E8", "F8", "G8", "H8"])]
for i6, (nm6, sh6, cells6) in enumerate(COEG):
    rr6 = 6 + i6
    sc(w22, f"B{rr6}", nm6, BOLD)
    for j6, cl6 in enumerate(cells6):
        col6 = "CDEFGH"[j6]
        fmt6 = "0" if col6 in "CDE" else M2
        sc(w22, f"{col6}{rr6}", f"={sh6}!${cl6[0]}${cl6[1:]}", NORM, fmt=fmt6,
           align="center" if col6 in "CDE" else "right")
sc(w22, "B11", "Total", BOLD, LGREY)
for col6 in "CDEFGH":
    fmt6 = "0" if col6 in "CDE" else M2
    sc(w22, f"{col6}11", f"=SUM({col6}6:{col6}10)", BOLD, LGREY, fmt=fmt6,
       align="center" if col6 in "CDE" else "right")
sc(w22, "I5", "Cost - AU ($m)", WHITEF, MIDBLU, align="center", wrap=True)
sc(w22, "J5", "Cost - NZ ($m)", WHITEF, MIDBLU, align="center", wrap=True)
AUNZ_GRID = [("'2.3 BP&T'", "I6", "J6"), ("'2.3 BP&T'", "I7", "J7"), ("'2.4 SA&D'", "J6", "K6"),
             ("'2.4 SA&D'", "J7", "K7"), ("'2.5 Cyber Roles'", "I8", "J8")]
for i6, (sh6, a6, n6) in enumerate(AUNZ_GRID):
    sc(w22, f"I{6+i6}", f"={sh6}!${a6[0]}${a6[1:]}", NORM, fmt=M2, align="right")
    sc(w22, f"J{6+i6}", f"={sh6}!${n6[0]}${n6[1:]}", NORM, fmt=M2, align="right")
for col6 in "IJ":
    sc(w22, f"{col6}11", f"=SUM({col6}6:{col6}10)", BOLD, LGREY, fmt=M2, align="right")
sc(w22, "B17", "TDD Corporate funding pool ($m)", WHITEF, NAVY)
for col6, h6 in zip("BCDE", ["Funding line", "Pool ($m)", "Drawn ($m)", "Remaining ($m)"]):
    sc(w22, f"{col6}18", h6, WHITEF, MIDBLU, align="center")
POOL = [("OpEx Initiatives", "='0.4 Budget Table (Fin)'!$N$5", "=0"),
        ("Significant Items", "='0.4 Budget Table (Fin)'!$O$5",
         "='1.3 Enterprise Data'!$I$13+'1.4 TDD Group Functions'!$I$13"),
        ("CapEx", "='0.4 Budget Table (Fin)'!$P$5", "='1.9 Commercial Fuels'!$I$15+'2.5 Cyber Roles'!$C$13")]
for i6, (nm6, pf6, df6) in enumerate(POOL):
    rr6 = 19 + i6
    sc(w22, f"B{rr6}", nm6, NORM)
    sc(w22, f"C{rr6}", pf6, NORM, fmt=M2, align="right")
    sc(w22, f"D{rr6}", df6, NORM, fmt=M2, align="right")
    sc(w22, f"E{rr6}", f"=C{rr6}-D{rr6}", NORM, fmt=M2, align="right")
sc(w22, "B22", "Total", BOLD, LGREY)
for col6 in "CDE":
    sc(w22, f"{col6}22", f"=SUM({col6}19:{col6}21)", BOLD, LGREY, fmt=M2, align="right")
sc(w22, "B24", "Roll-up only - budgets on 0.0 Data Config, FTE lists on Sheet2, detail on the three COE tabs.", NORM, border=False)
cfg6 = wb["0.0 Data Config"]
for addr6, ref6 in (("F6", "='2.2 COE'!$F$8"), ("F7", "='2.2 COE'!$F$10"), ("F8", "='2.2 COE'!$F$7"),
                    ("F9", "='2.2 COE'!$F$6"), ("F10", "='2.2 COE'!$F$9")):
    cfg6[addr6].value = ref6
wex6 = wb["Exec Summary"]
wex6["C59"].value = "='2.2 COE'!$H$11"
for r6, refs6 in {18: ("='0.0 Data Config'!$E$6", "='2.2 COE'!$F$8", "='2.2 COE'!$G$8", "='2.2 COE'!$H$8"),
                  20: ("='0.0 Data Config'!$E$8", "='2.2 COE'!$F$7", "='2.2 COE'!$G$7", "='2.2 COE'!$H$7"),
                  21: ("='0.0 Data Config'!$E$9", "='2.2 COE'!$F$6", "='2.2 COE'!$G$6", "='2.2 COE'!$H$6"),
                  22: ("='0.0 Data Config'!$E$10", "='2.2 COE'!$F$9", "='2.2 COE'!$G$9", "='2.2 COE'!$H$9")}.items():
    g6[f"C{r6}"].value = refs6[0]
    g6[f"D{r6}"].value = refs6[1]
    g6[f"E{r6}"].value = f"=C{r6}-D{r6}"
    g6[f"H{r6}"].value = refs6[2]
    g6[f"I{r6}"].value = refs6[3]
    g6[f"J{r6}"].value = f"=MAX(0,-E{r6})+I{r6}"
    g6[f"K{r6}"].value = f"=D{r6}"

wex6["B58"].value = "TDD Cyber - needs more than its buckets, see 2.3 Cyber Roles ($m)"
cfg6["B7"].value = "COE - Cyber, Risk & Service Ops (see 2.3 Cyber Roles)"
wb["Lists"]["K1"].value = "Archetype roles"
# 6e. AU/NZ inside every 1.x Portfolio Summary + budget box variances
CFGROW = {"1.1 Ampol Retail": (11,), "1.2 Customer": (13, 14), "1.3 Enterprise Data": (22,),
          "1.4 TDD Group Functions": (21,), "1.5 P&C": (18,), "1.6 Finance": (19,),
          "1.7 Infrastructure": (17,), "1.8 Energy Solutions & B2B": (16,),
          "1.9 Commercial Fuels": (15,), "1.10 Z Retail": (12,)}
AUNZ2 = {}
for tab6, cfgrows6 in CFGROW.items():
    ws6 = wb[tab6]
    rows6 = {}
    hdr6 = None
    for r6 in range(3, 14):
        b6 = str(ws6.cell(r6, 2).value or "")
        if b6 == "Cost": hdr6 = r6
        for key6, tag6 in (("Portfolio Overhead", "ovh"), ("Platform Overheads", "plat"),
                           ("Squad Support Costs", "squad"), ("Total Cost", "tot")):
            if b6.startswith(key6): rows6[tag6] = r6
    if len(rows6) < 4 or hdr6 is None:
        print("AUNZ skip", tab6, rows6); continue
    aub6 = "+".join(f"'0.0 Data Config'!$C${cr}" for cr in cfgrows6)
    nzb6 = "+".join(f"'0.0 Data Config'!$D${cr}" for cr in cfgrows6)
    for col6, h6 in zip("CDEF", ["TDD AU ($m)", "TDD NZ ($m)", "Other ($m)", "Total ($m)"]):
        sc(ws6, f"{col6}{hdr6}", h6, WHITEF, MIDBLU, align="center", wrap=True)
    au_c, nz_c = AUNZ.get(tab6, (None, None))
    for tag6 in ("ovh", "plat", "squad"):
        r6 = rows6[tag6]
        oldC6 = str(ws6.cell(r6, 3).value or "0")
        oldD6 = str(ws6.cell(r6, 4).value or "")
        body6 = oldC6[1:] if oldC6.startswith("=") else oldC6
        if tag6 == "squad" and au_c:
            au_f6 = str(ws6.cell(int(au_c[1:]), 3).value)
            nz_f6 = str(ws6.cell(int(nz_c[1:]), 3).value)
            sc(ws6, f"C{r6}", au_f6, NORM, fmt=M2, align="right")
            sc(ws6, f"D{r6}", nz_f6, NORM, fmt=M2, align="right")
        else:
            sc(ws6, f"C{r6}", f"=IF(({nzb6})>({aub6}),0,{body6})", NORM, fmt=M2, align="right")
            sc(ws6, f"D{r6}", f"=IF(({nzb6})>({aub6}),{body6},0)", NORM, fmt=M2, align="right")
        sc(ws6, f"E{r6}", oldD6 if oldD6.startswith("=") else "=0", NORM, fmt=M2, align="right")
        sc(ws6, f"F{r6}", f"=C{r6}+D{r6}+E{r6}", NORM, fmt=M2, align="right")
    r6 = rows6["tot"]
    a6, b6_ = rows6["ovh"], rows6["squad"]
    for col6 in "CDEF":
        sc(ws6, f"{col6}{r6}", f"=SUM({col6}{a6}:{col6}{b6_})", BOLD, LGREY, fmt=M2, align="right")
    for i6, (lab6, fx6) in enumerate([("AU Budget ($m)", f"={aub6}"), ("NZ Budget ($m)", f"={nzb6}"),
                                      ("AU Variance ($m)", f"=H6-C{r6}"), ("NZ Variance ($m)", f"=H7-D{r6}")]):
        rr6 = 6 + i6
        sc(ws6, f"G{rr6}", lab6, BOLD if "Variance" in lab6 else NORM)
        sc(ws6, f"H{rr6}", fx6, BOLD if "Variance" in lab6 else NORM, fmt=M2, align="right")
    if au_c:
        clear_region(ws6, int(au_c[1:]), int(nz_c[1:]), 2, 4)
        AUNZ2[tab6] = (f"C{rows6['squad']}", f"D{rows6['squad']}")
if AUNZ2:
    AUNZ = AUNZ2
    au_sum6 = "+".join(f"'{t6}'!${a6[0]}${a6[1:]}" for t6, (a6, b6_) in AUNZ.items())
    nz_sum6 = "+".join(f"'{t6}'!${b6_[0]}${b6_[1:]}" for t6, (a6, b6_) in AUNZ.items())
    g6[f"C{r0+3}"].value = f"={au_sum6}"
    g6[f"C{r0+7}"].value = f"={nz_sum6}"
# 6f. the exec story - five lines a CTO can read cold
for i6, (lab6, fx6) in enumerate([
        ("The story", None),
        ("1. The contract: squad archetypes were set as the design - they allow this many squad roles", "='3.0 FTE View'!$G$4"),
        ("2. What happened: roles actually raised across squads, leadership and COEs - a third vacant", "='3.0 FTE View'!$C$4"),
        ("3. What the raised organisation costs today ($m)", "='2.1 Total Cost'!$I$24"),
        ("4. Over the archetype design by ($m)", "='2.1 Total Cost'!$K$24"),
        ("5. The decision: hire, hold or offshore every vacancy on the 4.x working copies - value on the table ($m)", "='2.1 Total Cost'!$G$24")]):
    rr6 = 4 + i6
    if fx6 is None:
        sc(wex6, f"B{rr6}", lab6, WHITEF, NAVY)
    else:
        sc(wex6, f"B{rr6}", lab6, BOLD, border=False)
        sc(wex6, f"C{rr6}", fx6, Font(name="Calibri", size=12, bold=True, color="FF002F6C"),
           fmt="0.0", align="right", border=False)
# 6g. Customer AI - the 1.2 archetype squad matches the raw data name
import re as re6
for sheet6 in wb.worksheets:
    if sheet6.title in ("Squads", "Added data", "Sheet2"): continue
    for row6 in sheet6.iter_rows():
        for cell6 in row6:
            v6 = cell6.value
            if isinstance(v6, str) and "AI Enablement" in v6:
                if v6.startswith("="):
                    cell6.value = re6.sub(r'"AI Enablement\s*"', '"Customer AI"', v6)
                elif v6.strip() == "AI Enablement":
                    cell6.value = "Customer AI"
clear_region(wb["3.0 FTE View"], 156, 156, 2, 12)
# Customer AI is a real squad now (the owner renamed it) - class follows
for r6 in range(2, sq.max_row + 1):
    if (str(sq.cell(r6, 14).value) == "Customer" and str(sq.cell(r6, 16).value) == "Customer AI"
            and str(sq.cell(r6, 17).value) == "Unmapped"):
        CHANGELOG.append((r6, "Q", "Unmapped", "Squad", "Customer AI"))
        sq.cell(r6, 17).value = "Squad"
for r6 in range(2, ad_.max_row + 1):
    if str(ad_.cell(r6, 31).value) == "Customer AI" and str(ad_.cell(r6, 32).value) == "Unmapped":
        ad_.cell(r6, 32).value = "Squad"
        ad_.cell(r6, 29).value = "Customer"
# 'Other unmapped' residual subtracts only the two true overlap rows
ftx = wb["3.0 FTE View"]
ftx["I159"].value = '=COUNTIFS(Squads!$Q:$Q,"Unmapped",Squads!$R:$R,"Filled")-SUM(I157:I158)'
ftx["J159"].value = '=COUNTIFS(Squads!$Q:$Q,"Unmapped",Squads!$R:$R,"Vacant")-SUM(J157:J158)' 


# 6h. the 1.x summary columns moved (C AU, D NZ, E Other, F Total) - repoint
# every external reference to the old C/D/E total cells
TOTROW = {"1.1 Ampol Retail": 9, "1.2 Customer": 9, "1.3 Enterprise Data": 9,
          "1.4 TDD Group Functions": 9, "1.5 P&C": 9, "1.6 Finance": 9,
          "1.7 Infrastructure": 10, "1.8 Energy Solutions & B2B": 9,
          "1.9 Commercial Fuels": 9, "1.10 Z Retail": 9}
import re as re6
for sheet6 in wb.worksheets:
    if sheet6.title in TOTROW: continue
    for row6 in sheet6.iter_rows():
        for cell6 in row6:
            v6 = cell6.value
            if not (isinstance(v6, str) and v6.startswith("=")): continue
            nv6 = v6
            for tab6, t6 in TOTROW.items():
                q6 = f"'{tab6}'!"
                if q6 not in nv6: continue
                nv6 = nv6.replace(f"{q6}$E${t6}", f"{q6}#F#{t6}")
                nv6 = nv6.replace(f"{q6}$D${t6}", f"{q6}#E#{t6}")
                nv6 = nv6.replace(f"{q6}$C${t6}", f"({q6}#C#{t6}+{q6}#D#{t6})")
            if nv6 != v6:
                nv6 = re6.sub(r"#([A-F])#", r"$\1$", nv6)
                cell6.value = nv6


# no 'roster' anywhere visible - the owner's word is FTE
for sheet7 in wb.worksheets:
    if sheet7.title in ("Squads", "Added data", "Sheet2", "0.4 Presentation Pack",
                        "0.1 Budget Table (Fin)", "FY26 Budget (ref)", "Sheet1"):
        continue
    for row7 in sheet7.iter_rows():
        for cell7 in row7:
            v7 = cell7.value
            if isinstance(v7, str) and not v7.startswith("=") and "roster" in v7.lower():
                cell7.value = v7.replace("rosters", "FTE lists").replace("roster", "FTE list") \
                                .replace("Rosters", "FTE lists").replace("Roster", "FTE list")

# =====================================================================
# STAGE 5: numbering = flow. Renumber tabs, rewrite every reference,
# separators + tab colours, reorder, and stale text mentions fixed.
# =====================================================================
RENAME = {
    "0.4 Budget Table (Fin)": "0.1 Budget Table (Fin)",
    "0.0 Data Config": "0.2 Data Config",
    "0.1 Squads": "0.3 Squad Archetypes",
    "0.3 For Presentation Pack (2)": "0.4 Presentation Pack",
    "0.5 Guide": "0.0 Guide",
    "0.2 FY26 Budget": "FY26 Budget (ref)",
    "2.3 BP&T": "2.1 BP&T",
    "2.4 SA&D": "2.2 SA&D",
    "2.5 Cyber Roles": "2.3 Cyber Roles",
    "2.2 COE": "2.4 COE Summary",
    "2.0 Group Summary": "2.5 Group Summary",
    "2.1 Total Cost": "2.6 Total Cost",
    "3.0 FTE View": "2.7 Squad Detail",
    "3.1 Data QA": "5.0 Data QA",
}
# rewrite every formula reference (full quoted sheet names - unambiguous)
for sheet in wb.worksheets:
    for row_ in sheet.iter_rows():
        for cell in row_:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "'" in v:
                nv = v
                for old, new in RENAME.items():
                    nv = nv.replace(f"'{old}'!", f"'{new}'!")
                if nv != v: cell.value = nv
# stale tab mentions inside plain text labels (ours only; longest first)
TEXTMAP = [("2.3 to 2.5 and 3.1", "2.1 to 2.3 and 5.0"),
           ("on 2.3, 2.4 or 2.5", "on 2.1, 2.2 or 2.3"),
           ("2.3 / 2.4 / 2.5", "2.1 / 2.2 / 2.3"),
           ("2.5 Cyber Roles", "2.3 Cyber Roles"),
           ("2.3 BP&T", "2.1 BP&T"), ("2.4 SA&D", "2.2 SA&D"),
           ("0.0 Data Config", "0.2 Data Config"), ("0.1 Squads", "0.3 Squad Archetypes"),
           ("3.0 FTE View", "2.7 Squad Detail"), ("drills 2.1", "drills 2.6"),
           ("see 3.1", "see 5.0"), ("and 3.1", "and 5.0"), ("on 3.1", "on 5.0"),
           ("on 3.0", "on 2.7"), ("see 2.2 (", "see 2.4 ("), ("is on 2.1.", "is on 2.6."),
           ("sit on 2.4", "sit on 2.2"), ("2.2 COE", "2.4 COE Summary")]
SKIP_TXT = {"Squads", "Added data", "Sheet2", "0.4 Presentation Pack",
            "0.1 Budget Table (Fin)", "FY26 Budget (ref)", "Sheet1"}
for sheet in wb.worksheets:
    if sheet.title in SKIP_TXT or sheet.title in ("0.3 For Presentation Pack (2)", "0.4 Budget Table (Fin)"):
        continue
    for row_ in sheet.iter_rows():
        for cell in row_:
            v = cell.value
            if isinstance(v, str) and not v.startswith("="):
                nv = v
                for old, new in TEXTMAP: nv = nv.replace(old, new)
                if nv != v: cell.value = nv
for old, new in RENAME.items():
    if old in wb.sheetnames: wb[old].title = new
# separators + group colours
GROUPS = [("- INPUTS -", "808080", ["0.0 Guide", "0.1 Budget Table (Fin)", "0.2 Data Config",
                                    "0.3 Squad Archetypes", "0.4 Presentation Pack"]),
          ("- DESIGNS -", "1F4E79", [f"1.{i} " for i in range(1, 12)]),
          ("- ROLL-UP -", "002F6C", ["2.1 BP&T", "2.2 SA&D", "2.3 Cyber Roles", "2.4 COE Summary",
                                     "2.5 Group Summary", "2.6 Total Cost", "2.7 Squad Detail"]),
          ("- DECISIONS -", "BF8F00", [f"4.{i} " for i in range(1, 12)] + ["4.12 BP&T", "4.13 SA&D", "4.14 EGI & Central"]),
          ("- EVIDENCE -", "375623", ["5.0 Data QA", "Squads", "Added data", "Sheet2"])]
def color_of(title):
    for gname, col, members in GROUPS:
        for m in members:
            if title == m or title.startswith(m): return col
    return None
for gname, col, _ in GROUPS:
    if gname in wb.sheetnames: del wb[gname]
    sp = wb.create_sheet(gname)
    sp.sheet_properties.tabColor = col
    sp.sheet_view.showGridLines = False
    sp.column_dimensions["A"].width = 2
    sp.column_dimensions["B"].width = 60
    sc(sp, "B2", gname.strip("- ") + "  >>>", Font(name="Calibri", size=16, bold=True, color="FF" + col), border=False)
for t in wb.sheetnames:
    c = color_of(t)
    if c: wb[t].sheet_properties.tabColor = c
ORDER = (["Exec Summary", "- INPUTS -", "0.0 Guide", "0.1 Budget Table (Fin)", "0.2 Data Config",
          "0.3 Squad Archetypes", "0.4 Presentation Pack", "- DESIGNS -"]
         + [t for t in wb.sheetnames if t.startswith("1.")]
         + ["- ROLL-UP -", "2.1 BP&T", "2.2 SA&D", "2.3 Cyber Roles", "2.4 COE Summary",
            "2.5 Group Summary", "2.6 Total Cost", "2.7 Squad Detail", "- DECISIONS -"]
         + [t for t in wb.sheetnames if t.startswith("4.")]
         + ["- EVIDENCE -", "5.0 Data QA", "Squads", "Added data", "Sheet2"])
rest = [t for t in wb.sheetnames if t not in ORDER]
final = ORDER + rest
wb._sheets = [wb[t] for t in final]
if "Sheet1" in wb.sheetnames: wb["Sheet1"].sheet_state = "hidden"
# guide rewrite: the two-stream flow with the final numbering
gd = wb["0.0 Guide"]
clear_region(gd, 2, gd.max_row + 2, 1, 6)
gd.column_dimensions["B"].width = 118
sc(gd, "B2", "How this workbook flows", Font(name="Calibri", size=16, bold=True, color="FF002F6C"), border=False)
GUIDE2 = [
 ("h", "The flow - two streams that meet in the middle"),
 ("t", "STREAM 1, THE DESIGN: 0.1 Budget Table (the money Finance gave) feeds 0.2 Data Config (how it is allocated per portfolio, AU and NZ, plus overhead rates and the four $2m COE allocations). 0.3 Squad Archetypes prices the contract (type x size). Each GM's design sits on 1.1 to 1.10, priced from 0.3 and funded per 0.2."),
 ("t", "STREAM 2, THE ORG: the Squads tab holds every role actually raised (536, filled and vacant) and its model mapping. Sheet2 holds the updated FTE lists for BP&T, SA&D, Cyber and EGI. Added data holds what each person actually costs."),
 ("t", "THE ROLL-UP: the COE FTE lists (2.1 BP&T, 2.2 SA&D, 2.3 Cyber Roles) and their funding roll into 2.4 COE Summary. 2.5 Group Summary answers: can the budgets fund the design (with AU and NZ variance). 2.6 Total Cost is the test: the design vs what the org actually costs. 2.7 Squad Detail is the same test squad by squad."),
 ("t", "THE DECISIONS: tabs 4.1 to 4.14 - every role in the organisation appears on exactly one of them. GMs mark Hire or Hold on every vacancy. The Exec Summary tells the result. 5.0 Data QA is the evidence behind every number."),
 ("h", "What you can edit (yellow cells are inputs)"),
 ("t", "0.2 Data Config: budgets, rates, allocations. | 0.3 Squad Archetypes: prices and the offshore rate. | 1.x: squad type, size, On/Off, support %, Fund (AU or NZ). | 4.x: the Hire or Hold Call cells. | Squads / Sheet2 / Added data: your source data."),
 ("h", "If you change X, Y updates"),
 ("t", "A squad's type, size, On/Off, support % or Fund toggle on 1.x updates 2.5, 2.6, 2.7 and the Exec Summary."),
 ("t", "A status or mapping in the Squads tab updates every role count: 2.6, 2.7 and the 4.x tabs."),
 ("t", "A person or cost in Sheet2 updates 2.1, 2.2, 2.3, the EGI FTE list, 2.4, the COE lines of 2.5 and 2.6, and tabs 4.12 to 4.14."),
 ("t", "A budget on 0.2 updates the budget and variance columns of 2.5, the draw-downs on 2.1 / 2.2 / 2.4, and left to fund everywhere."),
 ("t", "A Hire or Hold call on 4.x updates that tab's cost of roles marked Hire. Totals move only when a role is actually hired in the raw data."),
 ("h", "What is safe to change in the source data - and what is not"),
 ("t", "SAFE: any number (cost, base, rate), any status, any person's details. All model cells are live references."),
 ("t", "CAUTION: renaming a person in Squads without renaming them in Added data breaks that person's cost lookup (it falls back to the title average); 5.0 shows the drift."),
 ("t", "NOT SAFE ALONE: renaming a squad or portfolio (counts match the name as text), deleting rows (references break loudly and 5.0 stops footing - change the status instead), or adding rows at the bottom (the FTE lists need a refresh). Ask for a refresh after structural edits."),
 ("h", "How to trust it"),
 ("t", "Every row labelled Check must read 0. All reconciliations, the Sheet2 update log, the mapping change log and the raw data audit live on 5.0 Data QA. Every role has exactly one 4.x home - if one were missed, the checks would not foot."),
]
gr_ = 4
for kind, txt in GUIDE2:
    if kind == "h":
        sc(gd, f"B{gr_}", txt, WHITEF, NAVY); gr_ += 1
    else:
        sc(gd, f"B{gr_}", txt, NORM, border=False, wrap=True)
        gd.row_dimensions[gr_].height = max(15, 15 * (1 + len(txt) // 115))
        gr_ += 1

# =====================================================================
# STAGE 7: the owner's structure. 1.x = designs + COE roles, 2.x = the
# working tabs, 3.x = summaries, 4.0 = evidence. Flow wired 1 > 2 > 3.
# Plus the confirmed audit fixes and the banned words purged.
# =====================================================================
import re as re7

# ---------- (A) content fixes on current names, before any renaming ----------
# A1. working tabs: type and size separated, vacancies remaining includes
# Offshore, headers in the owner's words, New Variance vs archetype.
for gt7, an7 in gm_anchor.items():
    w7 = wb[gt7]
    h7, t7 = an7["hdr"], an7["tot"]
    w7.cell(h7, 3).value = "Archetype type"
    sc(w7, f"O{h7}", "Archetype size", WHITEF, MIDBLU, align="center", wrap=True)
    w7.column_dimensions["O"].width = 11
    w7.cell(h7, 8).value = "Vacancies remaining"
    w7.cell(h7, 13).value = "Cost after vacancy decisions ($m)"
    w7.cell(h7, 14).value = "New Variance ($m)"
    for r7 in range(h7 + 1, t7):
        kf7 = str(w7.cell(r7, 11).value or "")
        mfr7 = re7.search(r"'2\.7 Squad Detail'!\$M\$(\d+)", kf7)
        if mfr7:
            fr7 = mfr7.group(1)
            w7.cell(r7, 3).value = f"='2.7 Squad Detail'!$E${fr7}"
            sc(w7, f"O{r7}", f"='2.7 Squad Detail'!$F${fr7}", NORM, align="center")
            w7.cell(r7, 14).value = f"=M{r7}-K{r7}"
        else:
            sc(w7, f"O{r7}", '="-"', NORM, align="center")
        gf7 = str(w7.cell(r7, 7).value or "")
        mg7 = re7.match(r'^=COUNTIF\(E(\d+):E(\d+),"Hire"\)$', gf7)
        if mg7:
            a7, b7 = mg7.group(1), mg7.group(2)
            w7.cell(r7, 8).value = (f'=F{r7}-G{r7}-COUNTIF(E{a7}:E{b7},"Offshore")')
    sc(w7, f"O{t7}", '="-"', BOLD, LGREY, align="center")
    w7.freeze_panes = f"A{h7+1}"

# A2. FTE view: the working tab results land here. Vacancies remaining and
# cost after vacancy decisions per squad, live from the levers.
ft7 = wb["2.7 Squad Detail"]
sc(ft7, "O6", "Vacancies remaining", WHITEF, MIDBLU, align="center", wrap=True)
sc(ft7, "P6", "Cost after vacancy decisions ($m)", WHITEF, MIDBLU, align="center", wrap=True)
ft7.column_dimensions["O"].width = 12
ft7.column_dimensions["P"].width = 14
for gt7, an7 in gm_anchor.items():
    w7 = wb[gt7]
    for r7 in range(an7["hdr"] + 1, an7["tot"]):
        kf7 = str(w7.cell(r7, 11).value or "")
        mfr7 = re7.search(r"'2\.7 Squad Detail'!\$M\$(\d+)", kf7)
        if mfr7:
            fr7 = int(mfr7.group(1))
            sc(ft7, f"O{fr7}", f"='{gt7}'!$H${r7}", NORM, fmt=M0, align="center")
            sc(ft7, f"P{fr7}", f"='{gt7}'!$M${r7}", NORM, fmt=M2, align="right")
sc(ft7, "O95", "=SUM(O7:O94)", BOLD, LGREY, fmt=M0, align="center")
sc(ft7, "P95", "=SUM(P7:P94)", BOLD, LGREY, fmt=M2, align="right")
ft7["K95"].value = "=J95-G95"
ft7.freeze_panes = "A7"

# A3. Total Cost in the owner's pasted structure: Archetype cost, Actual cost,
# Variance, Cost after vacancy decisions, New Variance, then the FTE columns,
# detail kept to the right so nothing downstream loses its source.
tc7 = wb["2.6 Total Cost"]
old7 = {}
for r7 in range(6, TOT_ROW + 1):
    old7[r7] = {c7: tc7.cell(r7, c7).value for c7 in range(3, 13)}
HD7 = ["Portfolio", "Archetype cost ($m)", "Actual cost ($m)", "Variance ($m)",
       "Cost after vacancy decisions ($m)", "New Variance ($m)", "Archetype FTE",
       "Filled FTE", "FTE variance", "Actual Filled ($m)", "Actual Vacant ($m)",
       "Vacant FTE", "Total FTE"]
for i7, hh7 in enumerate(HD7):
    sc(tc7, f"{get_column_letter(2+i7)}5", hh7, WHITEF, MIDBLU, align="center", wrap=True)
for c7, wd7 in zip("CDEFGHIJKLMN", (13, 13, 12, 16, 13, 11, 10, 11, 14, 14, 10, 10)):
    tc7.column_dimensions[c7].width = wd7
for r7 in range(6, TOT_ROW):
    o7 = old7[r7]
    bv7 = str(tc7.cell(r7, 2).value or "")
    tc7.cell(r7, 11).value = o7[5]   # K actual filled $  (old E)
    tc7.cell(r7, 12).value = o7[7]   # L actual vacant $  (old G)
    tc7.cell(r7, 8).value = o7[4]    # H archetype fte    (old D)
    tc7.cell(r7, 9).value = o7[6]    # I filled fte       (old F)
    tc7.cell(r7, 13).value = o7[8]   # M vacant fte       (old H)
    dash7 = isinstance(o7[4], str) and o7[4].startswith('="-"')
    tc7.cell(r7, 14).value = '="-"' if dash7 else f"=H{r7}+M{r7}"    # N total fte
    tc7.cell(r7, 10).value = '="-"' if dash7 else f"=I{r7}-H{r7}"    # J fte variance
    tc7.cell(r7, 4).value = f"=K{r7}+L{r7}"                          # D actual $
    tc7.cell(r7, 5).value = f"=ROUND(D{r7}-C{r7},6)"                 # E variance
    m17 = re7.match(r"^='(1\.\d+ [^']+)'!\$B\$2$", bv7)
    if m17:                       # portfolio row: after decisions from its working tab
        gm7 = "4." + m17.group(1).split(" ", 1)[0].split(".")[1] + " " + m17.group(1).split(" ", 1)[1]
        tc7.cell(r7, 6).value = f"='{gm7}'!$M${gm_anchor[gm7]['tot']}"
    elif r7 == COE_FIRST + 4:     # cyber row
        tc7.cell(r7, 6).value = f"='4.11 TDD Cyber'!$M${gm_anchor['4.11 TDD Cyber']['tot']}"
    elif r7 == DEDUP_ROW:
        tc7.cell(r7, 6).value = f"=C{r7}"
    else:                         # COE, leadership, unmapped: filled people, vacancies held
        tc7.cell(r7, 6).value = f"=K{r7}"
    tc7.cell(r7, 7).value = f"=ROUND(F{r7}-C{r7},6)"                 # G new variance
    for c7, fm7 in ((4, M2), (5, M2), (6, M2), (7, M2), (11, M2), (12, M2)):
        tc7.cell(r7, c7).number_format = fm7
for c7 in "CDEFGHIJKLMN":
    sc(tc7, f"{c7}{TOT_ROW}", f"=SUM({c7}6:{c7}{TOT_ROW-1})", BOLD, LGREY,
       fmt=(M2 if c7 in "CDEFGKL" else M0),
       align=("right" if c7 in "CDEFGKL" else "center"))
tc7.cell(RESTATE_ROW, 4).value = str(old7.get(RESTATE_ROW, {}).get(9) or tc7.cell(RESTATE_ROW, 9).value or "").replace("I24", "D24") or None
rst7 = tc7.cell(RESTATE_ROW, 9).value
if isinstance(rst7, str) and rst7.startswith("="):
    tc7.cell(RESTATE_ROW, 4).value = rst7.replace(f"I{TOT_ROW}", f"D{TOT_ROW}")
    tc7.cell(RESTATE_ROW, 4).number_format = M2
    tc7.cell(RESTATE_ROW, 9).value = None
sc(tc7, f"B{RESTATE_ROW+1}",
   "Role counts come from the Squads tab; dollars come from the Added data ledger. They are different snapshots and the difference is reconciled on the Data QA tab.",
   NORM, border=False, wrap=True)
sc(tc7, f"B{RESTATE_ROW+2}",
   "FTE columns cover the portfolio squads. COE and cyber roles are counted on their own role tabs.",
   NORM, border=False, wrap=True)
tc7.freeze_panes = "A6"
# every external reference into 2.6 follows the new column layout
COLMAP7 = {"D": "H", "E": "K", "F": "I", "G": "L", "H": "M", "I": "D", "J": "N", "K": "E", "L": "J"}
pat7 = re7.compile(r"('2\.6 Total Cost'!\$?)([D-L])(\$?\d+)")
for sh7 in wb.worksheets:
    if sh7.title == "2.6 Total Cost":
        continue
    for row7 in sh7.iter_rows():
        for cl7 in row7:
            v7 = cl7.value
            if isinstance(v7, str) and v7.startswith("=") and "2.6 Total Cost" in v7:
                nv7 = pat7.sub(lambda m7: m7.group(1) + COLMAP7[m7.group(2)] + m7.group(3), v7)
                if nv7 != v7:
                    cl7.value = nv7

# A4. Group Summary in the owner's pasted words, plus the Total to fund block.
gs7 = wb["2.5 Group Summary"]
gs7["C5"].value = "TDD Lights On Budget ($m)"
gs7["D5"].value = "Archetype Support Cost ($m)"
gs7["G5"].value = "Cost of FTE non TDD funded ($m)"
gs7["H5"].value = "Amount identified as rechargeable ($m)"
tf7 = r0 + 11
sc(gs7, f"B{tf7}", "Total to fund", WHITEF, NAVY)
sc(gs7, f"C{tf7}", "", WHITEF, NAVY)
sc(gs7, f"B{tf7+1}", "TDD Variance", NORM)
sc(gs7, f"C{tf7+1}", "=J17-I17", NORM, fmt=M2, align="right")
sc(gs7, f"B{tf7+2}", "Other Variance", NORM)
sc(gs7, f"C{tf7+2}", "=I17", NORM, fmt=M2, align="right")
sc(gs7, f"B{tf7+3}", "Total", BOLD, LGREY)
sc(gs7, f"C{tf7+3}", f"=C{tf7+1}+C{tf7+2}", BOLD, LGREY, fmt=M2, align="right")
gs7.freeze_panes = "A6"

# A5. the audit fixes on model formulas (owner inputs untouched)
wb["0.2 Data Config"]["C27"].value = "='0.1 Budget Table (Fin)'!G5+'0.1 Budget Table (Fin)'!G7"
cf7 = wb["1.9 Commercial Fuels"]
cf7["E11"].value = "=SUM(H12:H14)+SUM(H16:H17)"
if isinstance(cf7["G15"].value, str) and "central" not in cf7["G15"].value:
    cf7["G15"].value = cf7["G15"].value.rstrip() + " (central pool, reference only)"
zr7 = wb["1.10 Z Retail"]
if str(zr7["I18"].value).strip() == "=D9":
    zr7["I18"].value = "=E9"
zr7["I19"].value = "=I18-I17"
es7 = wb["1.8 Energy Solutions & B2B"]
if isinstance(es7["G17"].value, str) and "Finance" not in es7["G17"].value:
    es7["G17"].value = es7["G17"].value.rstrip() + " (not yet in the Finance table)"
# strategic program squads: show the real people cost beside the owner input
for tb7, rr7 in (("1.1 Ampol Retail", 46), ("1.1 Ampol Retail", 64),
                 ("1.2 Customer", 52), ("1.4 TDD Group Functions", 33),
                 ("1.5 P&C", 31), ("1.6 Finance", 31)):
    d17 = wb[tb7]
    nm7 = d17.cell(rr7, 2).value
    if nm7:
        sc(d17, f"M{rr7}",
           f'="People in this program today cost "&TEXT(SUMIFS(\'Added data\'!$AA:$AA,\'Added data\'!$AE:$AE,$B{rr7})/1000000,"0.00")&"m. Set the agreed cost in the yellow cell."',
           NORM, border=False)
# input formatting on 0.2 Data Config: yellow marks every true input
dc7 = wb["0.2 Data Config"]
BLUF7 = Font(name="Calibri", size=11, color="FF0000FF")
for rng7 in ("C6:D25", "J6:K9", "J14:K15"):
    for row7 in dc7[rng7]:
        for cl7 in row7:
            if cl7.value is not None and not (isinstance(cl7.value, str) and cl7.value.startswith("=")):
                cl7.fill = YELL
                cl7.font = BLUF7
# provenance beside the hard typed funding allocations
for tb7, r17, r27 in (("1.2 Customer", 13, 21), ("1.8 Energy Solutions & B2B", 13, 19),
                      ("1.9 Commercial Fuels", 13, 17)):
    d17 = wb[tb7]
    if d17["K12"].value is None:
        sc(d17, "K12", "Agreed by / date", NORM, align="center", wrap=True)
        for rr7 in range(r17, r27 + 1):
            if d17.cell(rr7, 11).value is None:
                sc(d17, f"K{rr7}", "", NORM, fill=YELL)
# owner working notes move to the evidence tab, logged with their source
qa7 = wb["5.0 Data QA"]
nr7 = qa7.max_row + 3
sc(qa7, f"B{nr7}", "Owner working notes (moved off the presentation tabs, preserved verbatim)", WHITEF, NAVY)
nr7 += 1
co7 = wb["2.4 COE Summary"]
for rr7 in range(7, 12):
    for cc7 in (11, 12):
        v7 = co7.cell(rr7, cc7).value
        if v7 is not None:
            sc(qa7, f"B{nr7}", f"2.4 COE Summary {get_column_letter(cc7)}{rr7}", NORM)
            sc(qa7, f"C{nr7}", str(v7), NORM)
            co7.cell(rr7, cc7).value = None
            nr7 += 1
for tb7, cells7 in (("1.1 Ampol Retail", ("J33", "J46", "J47", "J58", "J63")),
                    ("1.2 Customer", ("J37", "J38", "J39")),
                    ("1.4 TDD Group Functions", ("J25",)),
                    ("1.8 Energy Solutions & B2B", ("J6", "J28"))):
    d17 = wb[tb7]
    for ad7 in cells7:
        v7 = d17[ad7].value
        if isinstance(v7, str) and v7.strip():
            sc(qa7, f"B{nr7}", f"{tb7} {ad7}", NORM)
            sc(qa7, f"C{nr7}", v7, NORM)
            d17[ad7].value = None
            nr7 += 1
# vacancy coverage: the working tabs must cover the raw counts exactly
nr7 += 2
sc(qa7, f"B{nr7}", "Vacancy coverage check (must read 0)", WHITEF, NAVY); nr7 += 1
sc(qa7, f"B{nr7}", "Vacant on all working tabs", NORM)
terms7 = "+".join(f"COUNTIF('{t7}'!$F$1:$F$500,\"Vacant\")" for t7 in sorted(gm_anchor))
terms7 += "+COUNTIF('4.12 BP&T'!$F$1:$F$500,\"Vacant\")+COUNTIF('4.13 SA&D'!$F$1:$F$500,\"Vacant\")+COUNTIF('4.14 EGI & Central'!$F$1:$F$500,\"Vacant\")"
sc(qa7, f"C{nr7}", f"={terms7}", NORM, align="center"); cov7 = nr7; nr7 += 1
sc(qa7, f"B{nr7}", "Vacant in Squads (the 166)", NORM)
sc(qa7, f"C{nr7}", '=COUNTIF(Squads!$R$2:$R$1000,"Vacant")', NORM, align="center"); nr7 += 1
sc(qa7, f"B{nr7}", "Vacant in Sheet2 (COE and cyber FTE lists)", NORM)
sc(qa7, f"C{nr7}", '=SUMPRODUCT(--(LOWER(Sheet2!$Q$2:$Q$200)="v"))', NORM, align="center"); nr7 += 1
sc(qa7, f"B{nr7}", "Check - coverage minus raw counts", BOLD)
sc(qa7, f"C{nr7}", f"=C{cov7}-C{cov7+1}-C{cov7+2}", BOLD, LGREY, align="center"); nr7 += 1
sc(qa7, f"B{nr7}", "Squads with a size the archetype library cannot price (must be 0)", NORM)
chksz7 = "+".join(f"COUNTIF('1.{i7} {n7}'!$G$20:$G$70,\"check size\")"
                  for i7, n7 in ((1, "Ampol Retail"), (2, "Customer"), (3, "Enterprise Data"),
                                 (4, "TDD Group Functions"), (5, "P&C"), (6, "Finance"),
                                 (7, "Infrastructure"), (8, "Energy Solutions & B2B"),
                                 (9, "Commercial Fuels"), (10, "Z Retail")))
sc(qa7, f"C{nr7}", f"={chksz7}", NORM, align="center"); nr7 += 1
sc(qa7, f"B{nr7}", "Stray totals parked below the Added data ledger (bounded references keep them out of every sum)", NORM)
sc(qa7, f"C{nr7}", "=COUNT('Added data'!$AA$550:$AA$1000)", NORM, align="center"); nr7 += 1
qa7.freeze_panes = "A6"
# stale hidden sheets clearly out of play
if "Sheet1" in wb.sheetnames:
    del wb["Sheet1"]
for old7s, new7s in (("squad mapping", "squad mapping (superseded)"),
                     ("FY26 Budget (ref)", "FY26 Budget (superseded)")):
    if old7s in wb.sheetnames:
        wb[old7s].title = new7s
# exec: the lever line in the owner's words, plus the snapshot disclosure
ex7 = wb["Exec Summary"]
ex7["B55"].value = ("The main lever is the vacancies: they are raised but not hired, so holding them impacts nobody. "
                    "Decide role by role on the 2.x working tab.")
sc(ex7, "B77",
   "Role counts come from the Squads tab; dollars come from the Added data ledger. Different snapshots, reconciled on the Data QA tab.",
   NORM, border=False, wrap=True)

# ---------- (B) the banned words purged from every authored label ----------
SKIP7 = {"Squads", "Added data", "Sheet2", "0.4 Presentation Pack",
         "0.1 Budget Table (Fin)", "FY26 Budget (superseded)", "squad mapping (superseded)"}
for sh7 in wb.worksheets:
    if sh7.title in SKIP7:
        continue
    for row7 in sh7.iter_rows():
        for cl7 in row7:
            v7 = cl7.value
            if isinstance(v7, str) and not v7.startswith("="):
                nv7 = v7.replace("Cost of calls", "Cost of vacancy decisions")
                nv7 = nv7.replace("after calls", "after vacancy decisions")
                nv7 = nv7.replace("Hire or Hold Call cells", "Vacancy lever cells")
                nv7 = re7.sub(r"\bcalls\b", "vacancy decisions", nv7, flags=re7.I)
                nv7 = re7.sub(r"\bcall\b", "decision", nv7, flags=re7.I)
                nv7 = nv7.replace("–", "-").replace("—", "-")
                if nv7 != v7:
                    cl7.value = nv7

# ---------- (C) renumber to the owner's structure and rewrite every ref ----------
REN7A = {"2.1 BP&T": "1.11 BP&T", "2.2 SA&D": "1.12 SA&D", "2.3 Cyber Roles": "1.13 Cyber Roles",
         "2.4 COE Summary": "3.4 COE Summary", "2.5 Group Summary": "3.1 Group Summary",
         "2.6 Total Cost": "3.2 Total Cost", "2.7 Squad Detail": "3.3 FTE View",
         "5.0 Data QA": "4.0 Data QA"}
REN7B = {f"4.{i7} {n7}": f"2.{i7} {n7}" for i7, n7 in
         ((1, "Ampol Retail"), (2, "Customer"), (3, "Enterprise Data"), (4, "TDD Group Functions"),
          (5, "P&C"), (6, "Finance"), (7, "Infrastructure"), (8, "Energy Solutions & B2B"),
          (9, "Commercial Fuels"), (10, "Z Retail"), (11, "TDD Cyber"), (12, "BP&T"),
          (13, "SA&D"), (14, "EGI & Central"))}
REN7 = {**REN7A, **REN7B}
for sh7 in wb.worksheets:
    for row7 in sh7.iter_rows():
        for cl7 in row7:
            v7 = cl7.value
            if isinstance(v7, str) and v7.startswith("=") and "'" in v7:
                nv7 = v7
                for o7, n7 in REN7.items():
                    nv7 = nv7.replace(f"'{o7}'!", f"'{n7}'!")
                if nv7 != v7:
                    cl7.value = nv7
for o7, n7 in REN7A.items():
    if o7 in wb.sheetnames:
        wb[o7].title = n7
for o7, n7 in REN7B.items():
    if o7 in wb.sheetnames:
        wb[o7].title = n7
gm_anchor = {REN7.get(k7, k7): v7 for k7, v7 in gm_anchor.items()}
# stale numbering inside authored prose
TXT7 = [("2.1 to 2.3 and 5.0", "1.11 to 1.13 and 4.0"),
        ("2.1 / 2.2 / 2.3", "1.11 / 1.12 / 1.13"),
        ("on 2.1, 2.2 or 2.3", "on 1.11, 1.12 or 1.13"),
        ("tabs 4.12 / 4.13 / 4.14", "tabs 2.12 / 2.13 / 2.14"),
        ("4.12 / 4.13", "2.12 / 2.13"), ("4.12/4.13/4.14", "2.12/2.13/2.14"),
        ("2.1 BP&T", "1.11 BP&T"), ("2.2 SA&D", "1.12 SA&D"),
        ("2.3 Cyber Roles", "1.13 Cyber Roles"), ("2.4 COE Summary", "3.4 COE Summary"),
        ("2.5 Group Summary", "3.1 Group Summary"), ("2.6 Total Cost", "3.2 Total Cost"),
        ("2.7 Squad Detail", "3.3 FTE View"), ("5.0 Data QA", "4.0 Data QA"),
        ("drills 2.6", "drills 3.2"), ("is on 2.6.", "is on 3.2."),
        ("on 2.7", "on 3.3"), ("see 5.0", "see 4.0"), ("on 5.0", "on 4.0"),
        ("and 5.0", "and 4.0"), ("live on 5.0", "live on 4.0"),
        ("one 4.x home", "one 2.x home"), ("the 4.x GM tab", "the 2.x working tab"),
        ("the 4.x GM lever", "the 2.x working tab lever"),
        ("the 4.x tabs", "the 2.x working tabs"), ("4.x GM tabs", "2.x working tabs"),
        ("the 4.x working", "the 2.x working"), ("4.x", "2.x"),
        ("2.0, 2.1 archetype", "the summaries' archetype"),
        ("counts on 3.0", "counts on 3.3"), ("2.3 to 2.5 and 3.1", "1.11 to 1.13 and 4.0")]
for sh7 in wb.worksheets:
    if sh7.title in SKIP7:
        continue
    for row7 in sh7.iter_rows():
        for cl7 in row7:
            v7 = cl7.value
            if isinstance(v7, str) and not v7.startswith("="):
                nv7 = v7
                for o7, n7 in TXT7:
                    nv7 = nv7.replace(o7, n7)
                if nv7 != v7:
                    cl7.value = nv7

# ---------- (D) separators, colours, order: numbering = flow ----------
for g7 in ("- ROLL-UP -", "- DESIGNS -", "- DECISIONS -", "- SUMMARIES -"):
    if g7 in wb.sheetnames:
        del wb[g7]
G7 = [("- DESIGNS -", "1F4E79"), ("- DECISIONS -", "BF8F00"), ("- SUMMARIES -", "002F6C")]
for g7, col7 in G7:
    sp7 = wb.create_sheet(g7)
    sp7.sheet_properties.tabColor = col7
    sp7.sheet_view.showGridLines = False
    sp7.column_dimensions["A"].width = 2
    sp7.column_dimensions["B"].width = 60
    sc(sp7, "B2", g7.strip("- ") + "  >>>", Font(name="Calibri", size=16, bold=True, color="FF" + col7), border=False)
def col7_of(t7):
    if t7.startswith("1."): return "1F4E79"
    if t7.startswith("2."): return "BF8F00"
    if t7.startswith("3."): return "002F6C"
    if t7 in ("4.0 Data QA", "Squads", "Added data", "Sheet2"): return "375623"
    return None
for t7 in wb.sheetnames:
    c7 = col7_of(t7)
    if c7:
        wb[t7].sheet_properties.tabColor = c7
def keyn7(t7):
    return float(t7.split(" ")[0].split(".")[1]) if "." in t7.split(" ")[0] else 0
ORD7 = (["Exec Summary", "- INPUTS -", "0.0 Guide", "0.1 Budget Table (Fin)", "0.2 Data Config",
         "0.3 Squad Archetypes", "0.4 Presentation Pack", "- DESIGNS -"]
        + sorted([t7 for t7 in wb.sheetnames if t7.startswith("1.")], key=keyn7)
        + ["- DECISIONS -"]
        + sorted([t7 for t7 in wb.sheetnames if t7.startswith("2.")], key=keyn7)
        + ["- SUMMARIES -", "3.1 Group Summary", "3.2 Total Cost", "3.3 FTE View", "3.4 COE Summary",
           "- EVIDENCE -", "4.0 Data QA", "Squads", "Added data", "Sheet2"])
rest7 = [t7 for t7 in wb.sheetnames if t7 not in ORD7]
wb._sheets = [wb[t7] for t7 in (ORD7 + rest7)]

# =====================================================================
# STAGE 8: the AU/NZ toggle goes INSIDE the squad table where the owner's
# example put it (after On/Off), with proper column separation; and the
# On/Off toggle on the 1.11 / 1.12 COE roles at 0.4x offshore.
# =====================================================================
import re as re8
from copy import copy as cp8
from openpyxl.utils import column_index_from_string as ci8

DESIGN8 = ["1.1 Ampol Retail", "1.2 Customer", "1.3 Enterprise Data", "1.4 TDD Group Functions",
           "1.5 P&C", "1.6 Finance", "1.7 Infrastructure", "1.8 Energy Solutions & B2B",
           "1.9 Commercial Fuels", "1.10 Z Retail"]
INS8 = 6   # new column F

REF8 = re8.compile(r"(?:'([^']+)'!)?(\$?)([A-Z]{1,3})(\$?)(\d{1,5})"
                   r"(?::(\$?)([A-Z]{1,3})(\$?)(\d{1,5}))?")
def shift8(formula, on_tab, target):
    """shift column letters >= F by one for refs belonging to `target` tab,
    covering BOTH endpoints of a range under the same sheet prefix."""
    def bump(col):
        return get_column_letter(ci8(col) + 1) if ci8(col) >= INS8 else col
    def rep(m):
        sheet, d1, c1, d2, r1, e1, c2, e2, r2 = m.groups()
        belongs = (sheet == target) or (sheet is None and on_tab == target)
        if belongs:
            c1 = bump(c1)
            if c2: c2 = bump(c2)
        pre = f"'{sheet}'!" if sheet else ""
        out = f"{pre}{d1}{c1}{d2}{r1}"
        if c2: out += f":{e1}{c2}{e2}{r2}"
        return out
    return REF8.sub(rep, formula)

def shift_sqref8(sq):
    out = []
    for part in str(sq).split():
        cells = part.split(":")
        conv = []
        for cref in cells:
            m = re8.match(r"^(\$?)([A-Z]{1,3})(\$?)(\d+)$", cref)
            if m and ci8(m.group(2)) >= INS8:
                conv.append(m.group(1) + get_column_letter(ci8(m.group(2)) + 1) + m.group(3) + m.group(4))
            else:
                conv.append(cref)
        out.append(":".join(conv))
    return " ".join(out)

for tab8 in DESIGN8:
    ws8 = wb[tab8]
    maxc8 = ws8.max_column + 1
    # merges lifted first - a merged child cell is read-only during the shift
    merges8 = [str(m8) for m8 in ws8.merged_cells.ranges]
    for m8 in merges8:
        ws8.unmerge_cells(m8)
    # move every cell from F rightwards one column, styles included
    for r8 in range(1, ws8.max_row + 1):
        for c8 in range(maxc8, INS8 - 1, -1):
            src8 = ws8.cell(r8, c8)
            dst8 = ws8.cell(r8, c8 + 1)
            dst8.value = src8.value
            dst8.font = cp8(src8.font); dst8.fill = cp8(src8.fill)
            dst8.border = cp8(src8.border); dst8.alignment = cp8(src8.alignment)
            dst8.number_format = src8.number_format
        f8 = ws8.cell(r8, INS8)
        f8.value = None; f8.font = cp8(NORM); f8.fill = PatternFill()
        f8.border = Border(); f8.alignment = Alignment(); f8.number_format = "General"
    # widths follow, then the new column and proper separation
    wd8 = {c: ws8.column_dimensions[c].width for c in list(ws8.column_dimensions)}
    for c8 in range(ws8.max_column + 1, INS8 - 1, -1):
        L8 = get_column_letter(c8); P8 = get_column_letter(c8 + 1)
        if wd8.get(L8):
            ws8.column_dimensions[P8].width = wd8[L8]
    ws8.column_dimensions["F"].width = 10
    ws8.column_dimensions["G"].width = 11
    ws8.column_dimensions["H"].width = 17
    ws8.column_dimensions["I"].width = 14
    # merges re-applied at their shifted positions; validations follow
    for m8 in merges8:
        ws8.merge_cells(shift_sqref8(m8))
    for dv8 in ws8.data_validations.dataValidation:
        dv8.sqref = shift_sqref8(dv8.sqref)
    for rng8 in list(ws8.conditional_formatting):
        pass  # ranges keep working - the shifted columns carry their formats with them
    # same-tab formulas follow the shift
    for row8 in ws8.iter_rows():
        for cl8 in row8:
            if isinstance(cl8.value, str) and cl8.value.startswith("="):
                nv8 = shift8(cl8.value, ws8.title, tab8)
                if nv8 != cl8.value:
                    cl8.value = nv8
    # every other sheet's refs into this tab follow too
    for oth8 in wb.worksheets:
        if oth8.title == tab8:
            continue
        for row8 in oth8.iter_rows():
            for cl8 in row8:
                if isinstance(cl8.value, str) and cl8.value.startswith("=") and tab8 in cl8.value:
                    nv8 = shift8(cl8.value, oth8.title, tab8)
                    if nv8 != cl8.value:
                        cl8.value = nv8
    # the Fund column moves into the table at F, wherever this tab kept it
    fund_dv8 = None
    for dv8 in ws8.data_validations.dataValidation:
        if dv8.formula1 and "AU,NZ" in str(dv8.formula1):
            fund_dv8 = dv8
    newsq8 = []
    if fund_dv8 is not None:
        cells8 = str(fund_dv8.sqref).split()
        cols8 = sorted({re8.match(r"\$?([A-Z]+)", c8).group(1) for c8 in cells8})
        for c8ref in cells8:
            m8 = re8.match(r"\$?([A-Z]+)\$?(\d+)", c8ref)
            colL8, rr8 = m8.group(1), int(m8.group(2))
            lv8 = ws8[f"{colL8}{rr8}"].value
            sc(ws8, f"F{rr8}", lv8 if lv8 in ("AU", "NZ") else "AU", NORM, YELL, align="center")
            ws8[f"{colL8}{rr8}"].value = None
            ws8[f"{colL8}{rr8}"].fill = PatternFill(); ws8[f"{colL8}{rr8}"].border = Border()
            newsq8.append(f"F{rr8}")
        for colL8 in cols8:
            for rr8 in range(1, ws8.max_row + 1):
                if ws8[f"{colL8}{rr8}"].value == "Fund":
                    ws8[f"{colL8}{rr8}"].value = None
                    ws8[f"{colL8}{rr8}"].fill = PatternFill(); ws8[f"{colL8}{rr8}"].border = Border()
        fund_dv8.sqref = " ".join(newsq8)
        for addr8 in ("C8", "C9", "D8", "D9"):
            v8 = ws8[addr8].value
            if isinstance(v8, str) and v8.startswith("="):
                nv8 = v8
                for colL8 in cols8:
                    if colL8 not in ("C", "D", "E", "F", "G", "H", "I"):
                        nv8 = re8.sub(colL8 + r"(\d+):" + colL8 + r"(\d+)", r"F\1:F\2", nv8)
                if nv8 != v8:
                    ws8[addr8].value = nv8
    # header over the toggle on every squad table
    for rr8 in range(1, ws8.max_row + 1):
        if ws8.cell(rr8, 2).value == "Squad" and ws8.cell(rr8, 5).value == "On/Off":
            sc(ws8, f"F{rr8}", "AU / NZ", WHITEF, MIDBLU, align="center")

# ---------- On/Off toggle on the 1.11 / 1.12 COE roles, offshore at 0.4x ----------
dvoo8 = None
for coe8, hr8, r18, r28 in (("1.11 BP&T", 20, 21, 44), ("1.12 SA&D", 21, 22, 50)):
    wc8 = wb[coe8]
    sc(wc8, f"H{hr8}", "On/Off", WHITEF, MIDBLU, align="center", wrap=True)
    wc8.column_dimensions["H"].width = 11
    dvoo8 = DataValidation(type="list", formula1='"Onshore,Offshore"', allow_blank=True)
    wc8.add_data_validation(dvoo8)
    for r8 in range(r18, r28 + 1):
        if wc8.cell(r8, 2).value is None:
            continue
        cell8 = sc(wc8, f"H{r8}", "Onshore", NORM, YELL, align="center")
        dvoo8.add(cell8)
        tv8 = wc8.cell(r8, 20).value   # T model cost helper
        if isinstance(tv8, str) and tv8.startswith("=") and "IF($H" not in tv8:
            wc8.cell(r8, 20).value = f'=({tv8[1:]})*IF($H{r8}="Offshore",0.4,1)'
        gv8 = wc8.cell(r8, 7).value    # G cost if hired (vacant rows only)
        if isinstance(gv8, str) and gv8.startswith("=") and "IF($H" not in gv8:
            wc8.cell(r8, 7).value = f'=({gv8[1:]})*IF($H{r8}="Offshore",0.4,1)'
    sc(wc8, f"B{r28+5}",
       "On/Off: set a role to Offshore and it is priced at 40% of the onshore cost. The totals above and every summary follow.",
       NORM, border=False, wrap=True)

# ---------- register closures: freeze on role tabs, one dedup source,
# red/green variance formats, yellow strictly for inputs ----------
wb["1.11 BP&T"].freeze_panes = "A21"
wb["1.12 SA&D"].freeze_panes = "A22"
wb["1.13 Cyber Roles"].freeze_panes = "A19"
wb["3.2 Total Cost"]["C23"].value = "=-('1.11 BP&T'!$C$13+'1.12 SA&D'!$C$13)"
REDPOS8 = '[Red]#,##0.00;[Green](#,##0.00);"-"'
REDNEG8 = '#,##0.00;[Red](#,##0.00);"-"'
tc8b = wb["3.2 Total Cost"]
for r8 in range(6, 25):
    tc8b.cell(r8, 5).number_format = REDPOS8
    tc8b.cell(r8, 7).number_format = REDPOS8
gs8b = wb["3.1 Group Summary"]
for r8 in range(6, 21):
    gs8b.cell(r8, 5).number_format = REDNEG8
    gs8b.cell(r8, 10).number_format = REDNEG8
for tab8 in DESIGN8:
    ws8b = wb[tab8]
    for row8 in ws8b.iter_rows():
        for cl8 in row8:
            if (isinstance(cl8.value, str) and cl8.value.startswith("=") and cl8.fill
                    and cl8.fill.patternType and getattr(cl8.fill.fgColor, "rgb", None) == "FFFFF2CC"):
                cl8.fill = PatternFill()

wb.save(OUT)
print("stage 5 saved. order:", wb.sheetnames[:12], "...")
print("stage 4 saved. AUNZ:", AUNZ)
print("stage 3 saved. GM anchors:", {t: (a["hdr"], a["tot"], a["rost_hdr"]) for t, a in gm_anchor.items()})
json.dump(dict(DEDUP=DEDUP_ROW, TOT=TOT_ROW, UNM=UNM_ROW, LEAD=LEAD_ROW, COE_FIRST=COE_FIRST,
               RESTATE=RESTATE_ROW, EGI_MEMO=EGI_MEMO_ROW, EGI_TOT=EGI_TOT,
               PT_R1=R1, PT_CHECK=PT_CHECK, SAD_R1=S1, SAD_CHECK=SAD_CHECK,
               CY_R1=C1, CY_CHECK=CY_CHECK, N_SAD_COE=len(sad_coe),
               STATUS_DIFFS=len(status_diffs), S2_ONLY=len(s2_only), SQ_ONLY=len(sq_only),
               GM={t: dict(hdr=a["hdr"], tot=a["tot"], rost_hdr=a["rost_hdr"]) for t, a in gm_anchor.items()},
               AUNZ=AUNZ, W412=W412, W413=W413, LEAD_BLK=LEAD_BLK, UNM_BLK=UNM_BLK,
               CHANGELOG=[[c[0], c[1]] for c in CHANGELOG], AD_CHANGES=AD_CHANGES,
               AUNZ_ROW=r0),
          open(SCR + "anchors_v9.json", "w"), indent=1)

