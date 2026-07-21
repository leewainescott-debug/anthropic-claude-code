#!/usr/bin/env python3
"""v5 on the user's Working Copy:
 A. discrepancy fixes (1.3 sums, 1.11 placeholders, 2.0 COE refs + unwired cols)
 B. Significant Items EGI split (1.1, 1.2, 1.5, 1.6) - AmPOS/CTRM stay on plain Sig Items
 C. TDD Cyber: Data Config roll-in, two squads (TDD COE hardcode + TDD Cyber archetype),
    Cyber CapEx yellow line
 D. 2.1 COE: planned-spend restructure (budget | planned | funded | left | total)
 E. Above Store combine on 1.1 with impact note
 F. raw data: normalisation helper columns N-R
 G. new '3.0 FTE View': archetype FTE vs filled/vacant, leadership roll-up,
    COE planned FTE, squads-not-in-archetype highlight
"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

U = "/root/.claude/uploads/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/"
SRC = U + "a5084be4-TDD_Cost_Calc_Working_Copy.xlsx"
OUT = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/TDD_Cost_Calc_v5.xlsx"
FIN = "'0.4 Budget Table (Fin)'"

wb = openpyxl.load_workbook(SRC, data_only=False)
wbv = openpyxl.load_workbook(SRC, data_only=True)

def frow(ws, col, text, exact=True, start=1):
    ci = openpyxl.utils.column_index_from_string(col)
    for r in range(start, ws.max_row + 2):
        v = ws.cell(r, ci).value
        if v is None: continue
        s = str(v).strip()
        if (exact and s == text) or (not exact and text in s):
            return r
    return None

YELLOW = PatternFill("solid", fgColor="FFFFF2CC")
def make_input(c, fmt=None):
    c.fill = YELLOW
    c.font = Font(name="Calibri", size=10, color="FF0000FF")
    if fmt: c.number_format = fmt

MONEY = '#,##0.00;\\(#,##0.00\\);\\-'

# =====================================================================
# A1. 1.3: include the user's new Operations squad in the sums
# =====================================================================
ws = wb["1.3 Enterprise Data"]
ws["C8"] = "=SUM(H24,H25,H26,H27,H28)"
ws["D8"] = "=SUM(I24,I25,I26,I27,I28)"
ws["I25"] = '=IFERROR($G25*(1-$F25),"")'
ws["I25"]._style = copy(ws["I24"]._style)
ws["C8"].comment = Comment("Fix: your new 'Operations' squad (row 25) was missing from the "
                           "portfolio sums and its Funded-outside formula was gone.", "QA")

# =====================================================================
# B. Significant Items EGI split
# =====================================================================
# ---- 1.1: rewrite funding block rows 12-19 ----
ws = wb["1.1 Ampol Retail"]
def clone(ws, dst, src):
    ws[dst]._style = copy(ws[src]._style); return ws[dst]
# row 16 becomes Sig Items EGI; applied/other/left shift to 17/18/19
for col in "GHIJ":
    clone(ws, f"{col}16", f"{col}15")
ws["G16"] = "Significant Items EGI"
ws["H16"] = None
ws["I16"] = "=I64"
ws["J16"] = None
ws["I15"] = "=I46"
ws["I15"].comment = Comment("Strategic Programs split: AmPOS stays on Significant Items; "
                            "EGI Retail moves to the Significant Items EGI line below.", "QA")
ws["I16"].comment = Comment("EGI Retail strategic-programme cost (row 64).", "QA")
old = {r: {c: ws[f"{c}{r}"].value for c in "GHIJ"} for r in (16, 17, 18)}
for col in "GHIJ":
    clone(ws, f"{col}17", f"{col}16"); clone(ws, f"{col}18", f"{col}17"); clone(ws, f"{col}19", f"{col}18")
ws["G17"] = "Total applied"; ws["H17"] = None; ws["I17"] = "=SUM(I12:I16)"; ws["J17"] = None
ws["G18"] = "Other cost (this model)"; ws["H18"] = None; ws["I18"] = "=D9"; ws["J18"] = None
ws["G19"] = "Left to fund"; ws["H19"] = None; ws["I19"] = "=I18-I17"; ws["J19"] = None
ws["G17"]._style = copy(ws["G16"]._style)
# restyle: G17/I17 like old applied row, G19/I19 like old left row (grey)
# move user's 'Total to fund' from B16/C16 to B19/C19
b_style = copy(ws["B16"]._style); c_style = copy(ws["C16"]._style)
ws["B19"]._style = b_style; ws["C19"]._style = c_style
ws["B19"] = "Total to fund"; ws["C19"] = "=SUM(-H7,I19)"
ws["B16"] = None; ws["C16"] = None
for col in "BC":
    ws[f"{col}16"]._style = copy(ws[f"{col}15"]._style)
ws["E11"] = "=SUM(H12:H16)"

# ---- 1.2: fix the inverted wiring ----
ws = wb["1.2 Customer"]
ws["I15"] = "=I52"
ws["I15"]._style = copy(ws["I16"]._style)
ws["I15"].comment = Comment("EGI Customer strategic-programme cost (row 52) - split out from "
                            "plain Significant Items per the EGI line-item rule.", "QA")
ws["I16"] = 0
# complete the user's two half-built squads and wire them into every total
LKP = ('=IFERROR(IF($E{r}="Onshore",INDEX(\'0.1 Squads\'!$G$5:$G$23,MATCH($C{r}&"|"&$D{r},'
       "'0.1 Squads'!$A$5:$A$23,0)),INDEX('0.1 Squads'!$H$5:$H$23,MATCH($C{r}&\"|\"&$D{r},"
       "'0.1 Squads'!$A$5:$A$23,0))),\"check size\")")
ws["I39"] = '=IFERROR($G39*(1-$F39),"")'          # Digital Support NZ
ws["I39"]._style = copy(ws["I38"]._style)
ws["G41"] = "=SUM(G37:G39)"; ws["I41"] = "=SUM(I37:I39)"
for c in "FGHI":                                    # AI Enablement
    ws[f"{c}46"]._style = copy(ws[f"{c}45"]._style)
ws["F46"] = 0.2
ws["G46"] = LKP.format(r=46)
ws["H46"] = '=IFERROR($G46*$F46,"")'
ws["I46"] = '=IFERROR($G46*(1-$F46),"")'
ws["F46"].comment = Comment("Fix: this new squad had no Support % or cost formulas and was "
                            "excluded from every total. Defaulted 20% - adjust.", "QA")
ws["G48"] = "=SUM(G45:G46)"; ws["I48"] = "=SUM(I45:I46)"
ws["C8"] = "=SUM(H29,H30,H31,H37,H38,H39,H45,H46,H52)"
ws["D8"] = "=SUM(I29,I30,I31,I37,I38,I39,I45,I46,I52)"
for dv in ws.data_validations.dataValidation:
    f = str(dv.formula1)
    if f == "SquadTypes": dv.add("C46")
    elif f == "SquadSizes": dv.add("D46")
    elif f == "OnOff": dv.add("E46")
    elif f == "SupportPct": dv.add("F46"); dv.add("F39")
# ---- 1.5 / 1.6: insert Sig Items EGI line ----
for tab, egirow in [("1.5 P&C", 31), ("1.6 Finance", 31)]:
    w = wb[tab]
    # current: 15 sig (I15='=I31'), 16 dep, 17 applied, 18 other, 19 left
    for col in "GHIJ":
        clone(w, f"{col}20", f"{col}19"); clone(w, f"{col}19", f"{col}18")
        clone(w, f"{col}18", f"{col}17"); clone(w, f"{col}17", f"{col}16")
        clone(w, f"{col}16", f"{col}15")
    w["G16"] = "Significant Items EGI"; w["H16"] = None; w["I16"] = f"=I{egirow}"; w["J16"] = None
    w["I15"] = 0
    w["G17"] = "Depreciation"; w["H17"] = w["H16"].value; w["J17"] = "=IFERROR(H17-I17,0)"
    # re-point dep line to its Fin ref (was on row 16)
    finrow = {"1.5 P&C": 12, "1.6 Finance": 11}[tab]
    w["H17"] = f"={FIN}!M{finrow}"; w["I17"] = 0
    w["G18"] = "Total applied"; w["H18"] = None; w["I18"] = "=SUM(I12:I17)"; w["J18"] = None
    w["G19"] = "Other cost (this model)"; w["H19"] = None; w["I19"] = "=D9"; w["J19"] = None
    w["G20"] = "Left to fund"; w["H20"] = None; w["I20"] = "=I19-I18"; w["J20"] = None
    w["E11"] = "=SUM(H12:H17)"
    w["I16"].comment = Comment(f"EGI strategic-programme cost (row {egirow}).", "QA")

# =====================================================================
# C. TDD Cyber restructure
# =====================================================================
dc = wb["0.0 Data Config"]
dc["C23"] = 2.5; dc["D23"] = 1.0
make_input(dc["C23"]); make_input(dc["D23"])
dc["C23"].comment = Comment(
    "COE - Cyber (1.5 AU / 0.5 NZ) rolled into the TDD Cyber bucket: now 2.5 AU / 1.0 NZ = 3.5 total. "
    "You said 'the four million' - current numbers give 3.5; adjust here if 4.0 is right.", "QA")
dc["B7"] = "COE - Cyber, Risk & Service Ops (rolled into TDD Cyber)"
dc["C7"] = 0; dc["D7"] = 0

cy = wb["1.11 TDD Cyber"]
# row 24: TDD COE (hardcoded); row 25: TDD Cyber (archetype)
cy["B24"] = "TDD COE"
cy["C24"] = "Strategic Programs"; cy["D24"] = None; cy["E24"] = "Onshore"; cy["F24"] = 1.0
cy["G24"] = 2.0
make_input(cy["G24"], MONEY)
cy["G24"].comment = Comment("Hard-coded: the rolled-in COE Cyber allocation (2.0) as a starting "
                            "point - overwrite with the real COE people cost. 100% TDD funded; "
                            "lower Support % to model recharges to the business.", "QA")
for col in "BCDEFGHIJ":
    cy[f"{col}25"]._style = copy(cy[f"{col}24"]._style)
cy["B25"] = "TDD Cyber"
cy["C25"] = "Operations"; cy["D25"] = "M"; cy["E25"] = "Onshore"; cy["F25"] = 1.0
cy["G25"] = ('=IFERROR(IF($E25="Onshore",INDEX(\'0.1 Squads\'!$G$5:$G$23,MATCH($C25&"|"&$D25,'
             "'0.1 Squads'!$A$5:$A$23,0)),INDEX('0.1 Squads'!$H$5:$H$23,MATCH($C25&\"|\"&$D25,"
             "'0.1 Squads'!$A$5:$A$23,0))),\"check size\")")
cy["G25"].font = Font(name="Calibri", size=10, color="FF008000")
cy["G25"].fill = PatternFill()
cy["H25"] = '=IFERROR($G25*$F25,"")'
cy["I25"] = '=IFERROR($G25*(1-$F25),"")'
cy["F25"].comment = Comment("Both squads default to 100% TDD funded ('some is going to be "
                            "recharged back... I don't think a lot will be') - lower to model recharges. "
                            "Your 0.1/'Hardcoded' placeholders were superseded by this build.", "QA")
for c in ("C24", "D24", "E24", "F24", "C25", "D25", "E25", "F25"):
    make_input(cy[c]) if c[0] in "F" else None
for c in ("C24","D24","E24","F24","C25","D25","E25","F25"):
    cy[c].fill = YELLOW
cy["C8"] = "=SUM(H24,H25)"; cy["D8"] = "=SUM(I24,I25)"
cy["G31"] = "=SUM(G24:G25)"; cy["I31"] = "=SUM(I24:I25)"
# Cyber CapEx line (number still to come)
cy["H16"] = 0; make_input(cy["H16"], MONEY)
cy["H16"].comment = Comment("Cyber CapEx (Monitoring) budget - number still to come; hard-code here.", "QA")
cy["I16"] = 0; make_input(cy["I16"], MONEY)
cy["I14"] = "=I12-I13-I16"
cy["I14"].comment = Comment("Left to fund = other cost less TDD Corporate pool funding less Cyber CapEx applied.", "QA")
# dropdowns for the two squad rows
for dv in cy.data_validations.dataValidation:
    f = str(dv.formula1)
    if f == "SquadTypes": dv.add("C24"); dv.add("C25")
    elif f == "SquadSizes": dv.add("D24"); dv.add("D25")
    elif f == "OnOff": dv.add("E24"); dv.add("E25")
    elif f == "SupportPct": dv.add("F24"); dv.add("F25")

# 2.1 Cyber COE row: rolled into 1.11
coe = wb["2.1 COE"]
coe["B9"] = "COE - Cyber, Risk & Service Ops (rolled into 1.11)"

# =====================================================================
# D. 2.1 COE planned-spend restructure
# =====================================================================
coe["C7"] = "TDD Lights On Budget ($m)"
coe["D7"] = "Planned spend ($m)"
coe["E7"] = "Funded ($m)"
coe["F7"] = "Left to fund ($m)"
coe["G7"]._style = copy(coe["F7"]._style); coe["G7"] = "Total cost ($m)"
for r in range(8, 13):
    coe[f"D{r}"] = 0; make_input(coe[f"D{r}"], MONEY)
    coe[f"E{r}"] = 0; make_input(coe[f"E{r}"], MONEY)
    coe[f"F{r}"] = f"=D{r}-E{r}"
    coe[f"G{r}"]._style = copy(coe[f"F{r}"]._style)
    coe[f"G{r}"] = f"=D{r}"
coe["D8"].comment = Comment("Hard-code: how much we plan to spend in this COE. Funded = what has "
                            "been sourced (allocation / envelopes); Left to fund = planned less funded; "
                            "Total cost = planned spend.", "QA")
coe["D9"] = 0; coe["E9"] = 0   # Cyber rolled into 1.11
coe["G13"]._style = copy(coe["F13"]._style)
for col in "CDEFG":
    coe[f"{col}13"] = f"=SUM({col}8:{col}12)"
# 0.0 Spend refs -> planned spend
for i, r in enumerate(range(6, 11)):
    dc[f"F{r}"] = f"='2.1 COE'!$D${8+i}"

# =====================================================================
# E. Above Store combine on 1.1
# =====================================================================
ws = wb["1.1 Ampol Retail"]
ws["B30"] = "Platform: Above Store"
ws["B32"] = "Above Store"; ws["C32"] = "Configuration / Integration"; ws["D32"] = "L"
ws["B35"] = "Above Store Total"
ws["J32"] = None
ws["B32"].comment = Comment(
    "Op-model choice: Merchandising / Supply Chain (C/I M, 9 roles, $1.4m) + Pricing & WFM "
    "(C/I S, 6 roles, $0.4m) combined into one Above Store squad (C/I L, 14 roles, $2.1m). "
    "Archetype impact: roles 15 to 14 (-1; op-model note says -4 vs current staffing), squad cost "
    "$1.8m to $2.1m (+0.3), platform overheads 2x$0.165 to 1x$0.165 (-0.165) - net +$0.135m. "
    "Org data already shows an Above Store platform with 7 filled + 9 vacant = 16 seats.", "QA")
# decommission Pricing & WFM block (rows 37-42)
ws["B37"] = "Platform: Pricing & WFM - combined into Above Store"
for c in "BCDEFGHIJ":
    ws[f"{c}39"].value = None
ws["H41"] = None
ws["B42"] = "(combined into Above Store)"
ws["G42"] = None; ws["H42"] = None; ws["I42"] = None
ws["C7"] = "=SUM(H27,H34,H47,H53,H59)"
ws["C8"] = "=SUM(H23,H24,H25,H26,H32,H33,H46,H52,H58,H64)"
ws["D8"] = "=SUM(I23,I24,I25,I26,I32,I33,I46,I52,I58,I64)"

# =====================================================================
# A2. 2.0 Group Summary: fix COE refs, wire H/I, totals
# =====================================================================
gs = wb["2.0 Group Summary"]
# per-tab anchors for applied/left
TABS = ["1.1 Ampol Retail","1.2 Customer","1.3 Enterprise Data","1.4 TDD Group Functions",
        "1.5 P&C","1.6 Finance","1.7 Infrastructure","1.8 Energy Solutions & B2B",
        "1.9 Commercial Fuels","1.10 Z Retail","1.11 TDD Cyber"]
refs = {}
for t in TABS:
    w = wb[t]
    ar = frow(w, "G", "Total applied")
    if ar is None:
        ar = frow(w, "G", "Funded from TDD Corporate pool", exact=False)
    lr = frow(w, "G", "Left to fund")
    refs[t] = (ar, lr)
for i, t in enumerate(TABS):
    r = 6 + i
    ar, lr = refs[t]
    gs[f"H{r}"] = f"='{t}'!$I${ar}"
    gs[f"I{r}"] = f"='{t}'!$I${lr}"
    gs[f"H{r}"]._style = copy(gs[f"G{r}"]._style)
    gs[f"I{r}"]._style = copy(gs[f"G{r}"]._style)
gs["H17"] = "=SUM(H6:H16)"; gs["I17"] = "=SUM(I6:I16)"
gs["H17"]._style = copy(gs["G17"]._style); gs["I17"]._style = copy(gs["G17"]._style)
# COE rows 18-22: fix scrambled refs, wire all columns
COE_ROWS = {18: ("COE - Strategy Architecture", 6, 8), 19: ("COE - Cyber, Risk & Service Ops (rolled into 1.11)", 7, 9),
            20: ("COE - Transformation", 8, 10), 21: ("COE - Business Partnering", 9, 11),
            22: ("COE - Data", 10, 12)}
for r, (label, dcrow, coerow) in COE_ROWS.items():
    gs[f"B{r}"] = label
    gs[f"C{r}"] = f"='0.0 Data Config'!$E${dcrow}"
    gs[f"D{r}"] = f"='2.1 COE'!$D${coerow}"
    gs[f"E{r}"] = f"=C{r}-D{r}"
    gs[f"G{r}"] = 0
    gs[f"H{r}"] = f"='2.1 COE'!$E${coerow}"
    gs[f"I{r}"] = f"='2.1 COE'!$F${coerow}"
    gs[f"J{r}"] = f"='2.1 COE'!$G${coerow}"
    for col in "HIJ":
        gs[f"{col}{r}"]._style = copy(gs[f"G{r}"]._style)
gs["C18"].comment = Comment("Fix: these COE budget refs were scrambled (pointed at wrong Data Config "
                            "rows/columns - e.g. a variance cell). Now each row pulls its own allocation; "
                            "Cyber shows 0 because it rolled into 1.11.", "QA")
gs["E23"] = None
gs["H24"] = "=SUM(H17,H18,H19,H20,H21,H22)"
gs["I24"] = "=SUM(I17,I18,I19,I20,I21,I22)"
gs["H24"]._style = copy(gs["G24"]._style); gs["I24"]._style = copy(gs["G24"]._style)

# =====================================================================
# F. raw data normalisation helper columns N-R
# =====================================================================
rd = wb["raw data"]; rdv = wbv["raw data"]
CY_DEPTS = {"cyber strat & tech","cyber sec ops","cyber risk","cyber grc","service op & assurance"}
COE_MAP = {"transformation": "COE - Transformation", "architecture": "COE - Strategy Architecture",
           "technology strategy & ai capability": "COE - Strategy Architecture",
           "delivery, sada": "COE - Strategy Architecture",
           "tdd business partner": "COE - Business Partnering",
           "group data": "COE - Data"}
SQUAD_MAP = {  # normalised raw squad -> (portfolio, platform, model squad, class)
 "pos": ("Ampol Retail","Store Operations","POS","Squad"),
 "payments": ("Ampol Retail","Store Operations","Payments","Squad"),
 "store operations": ("Ampol Retail","Store Operations","Store Operations","Squad"),
 "deployment": ("Ampol Retail","Store Operations","Deployment","Squad"),
 "above store": ("Ampol Retail","Above Store","Above Store","Squad"),
 "merchandising & supply chain": ("Ampol Retail","Above Store","Above Store","Squad"),
 "pricing & wfm": ("Ampol Retail","Above Store","Above Store","Squad"),
 "ampos": ("Ampol Retail","AmPOS","AmPOS","Strategic Program"),
 "network & qsr": ("Ampol Retail","Network / QSR","Network / QSR","Squad"),
 "data - au": ("Ampol Retail","Data AU","Data AU","Squad"),
 "egi retail": ("Ampol Retail","EGI Retail","EGI Retail","Strategic Program"),
 "ampol app": ("Customer","Ampol Digital","Ampol App","Squad"),
 "ampol web": ("Customer","Ampol Digital","Ampol Web","Squad"),
 "digital ops support": ("Customer","Ampol Digital","Digital Operations","Squad"),
 "z energy apps": ("Customer","Customer Z","Z Energy Apps","Squad"),
 "z energy martech": ("Customer","Customer Z","Z Energy Martech","Squad"),
 "au crm & martech": ("Customer","Group Customer Platforms","AU CRM & Martech","Squad"),
 "loyalty & martech": ("Customer","Group Customer Platforms","AU CRM & Martech","Squad"),
 "customer, ai": ("Customer","-","Customer AI","Unmapped"),
 "egi customer": ("Customer","EGI Customer","EGI Customer","Strategic Program"),
 "data science": ("Enterprise Data","Group Data","Data Science","Squad"),
 "coe, data science": ("Enterprise Data","Group Data","Data Science","Squad"),
 "coe, data operations": ("Enterprise Data","Group Data","Operations","Squad"),
 "reporting & analytics": ("Enterprise Data","Group Data","Reporting & Analytics","Squad"),
 "data platform": ("Enterprise Data","Group Data","Data Platforms","Squad"),
 "enterprise data delivery": ("Enterprise Data","Group Data","Enterprise Data Delivery","Squad"),
 "coe, data capability": ("COE","COE - Data","COE - Data","COE"),
 "workplace & enterprise tooling": ("TDD Group Functions","TDD Group Functions","Workplace & Enterprise Tooling","Squad"),
 "cloud network & infra ops": ("TDD Group Functions","TDD Group Functions","Network & Infrastructure","Squad"),
 "cloud, network & infra ops": ("TDD Group Functions","TDD Group Functions","Network & Infrastructure","Squad"),
 "devops & qe": ("TDD Group Functions","TDD Group Functions","DevOps & Engineering","Squad"),
 "integration & process automation": ("TDD Group Functions","TDD Group Functions","Integration","Squad"),
 "egi tdd": ("TDD Group Functions","EGI TDD","EGI TDD","Strategic Program"),
 "p&c": ("P&C","P&C","P&C","Squad"),
 "p&c rta": ("P&C","P&C","P&C - RTA","Squad"),
 "egi p&c": ("P&C","EGI P&C","EGI P&C","Strategic Program"),
 "sap erp": ("Finance","AU Finance","AU Finance","Squad"),
 "nz finance": ("Finance","NZ Finance","NZ Finance","Squad"),
 "egi finance": ("Finance","EGI Finance","EGI Finance","Strategic Program"),
 "distribution, sales & services": ("Infrastructure","Distribution","Distribution, Sales & Services","Squad"),
 "manufacturing group projects": ("Infrastructure","Manufacturing","Manufacturing & Group Projects","Squad"),
 "manuacturing group projects": ("Infrastructure","Manufacturing","Manufacturing & Group Projects","Squad"),
 "technology suport": ("Infrastructure","Manufacturing","Technology Support","Squad"),
 "technology support": ("Infrastructure","Manufacturing","Technology Support","Squad"),
 "data & insights": ("Infrastructure","Data & Insights","Data & Insights","Squad"),
 "energy": ("Energy Solutions & B2B","Energy Solutions","Energy","Squad"),
 "evci": ("Energy Solutions & B2B","Energy Solutions","EVCI","Squad"),
 "b2b": ("Energy Solutions & B2B","B2B","B2B","Squad"),
 "trading & shipping": ("Commercial Fuels","Trading & Shipping","Trading & Shipping","Squad"),
 "trading & shipping data": ("Commercial Fuels","Trading & Shipping","Trading & Shipping Data","Squad"),
 "supply": ("Commercial Fuels","Supply","Supply","Squad"),
 "ctrm": ("Commercial Fuels","CTRM","CTRM","Strategic Program"),
 "z supply": ("Z Retail","Z Supply","Z Supply","Squad"),
 "site systems": ("Z Retail","Z Customer","Site Systems","Squad"),
 "z retail backend": ("Z Retail","Z Customer","Z Retail Backend","Squad"),
 "z retail backend & supply": ("Z Retail","Z Customer","Z Retail Backend","Squad"),
 "data - nz": ("Z Retail","-","Data NZ","Unmapped"),
 "egi": ("EGI","EGI","EGI (unassigned)","Unmapped"),
 "coe, architecture": ("COE","COE - Strategy Architecture","COE - Strategy Architecture","COE"),
 "coe": ("COE","COE (unspecified)","COE (unspecified)","COE"),
 "coe, leadership": ("COE","Leadership","Leadership","Leadership"),
}
PORT_NORM = {"z":"Z Retail","retail":"Ampol Retail","customer":"Customer",
             "b2b & energy solutions":"Energy Solutions & B2B","egi integration":"EGI","egi":"EGI",
             "tdd":"TDD Group Functions","enterprise data":"Enterprise Data",
             "infrastructure":"Infrastructure","finance":"Finance","p&c":"P&C",
             "p&c, finance & legal":"P&C","commercial fuels":"Commercial Fuels","na":"","":""}
hdr_style = copy(rd.cell(1, 2)._style)
for j, h in enumerate(["Model Portfolio","Model Platform","Model Squad","Class","Status"]):
    c = rd.cell(1, 14 + j); c.value = h; c._style = hdr_style
mapped = 0
for r in range(2, rd.max_row + 1):
    g = lambda cc: (str(rdv.cell(r, cc).value).strip() if rdv.cell(r, cc).value is not None else "")
    name, dept, port, plat, squad = g(2), (g(7) or g(6)), g(9), g(10), g(11)
    if not (name or squad or port): continue
    status = "Vacant" if name.lower().strip() == "vacant" else "Filled"
    sq = squad.lower().strip()
    pn = PORT_NORM.get(port.lower().strip(), port)
    dp = dept.lower().strip()
    if sq == "leadership" or plat.lower().strip() == "leadership":
        mp, mpl, ms, cls = (pn or "Unmapped"), "Leadership", "Leadership", "Leadership"
    elif sq in SQUAD_MAP:
        mp, mpl, ms, cls = SQUAD_MAP[sq]
    elif dp in CY_DEPTS:
        mp, mpl, ms, cls = "TDD Cyber", "TDD Cyber, Risk & Service Ops", "TDD Cyber", "Squad"
    elif dp in COE_MAP:
        mp, mpl, ms, cls = "COE", COE_MAP[dp], COE_MAP[dp], "COE"
    elif dp == "loyalty & martech":
        mp, mpl, ms, cls = "Customer", "Group Customer Platforms", "AU CRM & Martech", "Squad"
    else:
        mp, mpl, ms, cls = (pn or "Unmapped"), (plat or "-"), (squad or dept or "-"), "Unmapped"
    for j, v in enumerate([mp, mpl, ms, cls, status]):
        rd.cell(r, 14 + j).value = v
    mapped += 1
print("raw data rows mapped:", mapped)

# =====================================================================
# G. 3.0 FTE View
# =====================================================================
tpl = wb["1.1 Ampol Retail"]
fte = wb.create_sheet("3.0 FTE View", wb.sheetnames.index("2.1 COE") + 1)
fte.sheet_view.showGridLines = False
for col, w in {"A":3,"B":22,"C":26,"D":28,"E":26,"F":6,"G":15,"H":8,"I":8,"J":11,"K":16,"L":16}.items():
    fte.column_dimensions[col].width = w
fte.row_dimensions[2].height = 21
fte["B2"]._style = copy(tpl["B2"]._style); fte["B2"] = "FTE View - archetypes vs actual org"

def band(row, text, c1=2, c2=12):
    for c in range(c1, c2 + 1):
        fte.cell(2, 2)  # noop guard
        fte.cell(row, c)._style = copy(tpl["B4"]._style)
    fte.cell(row, 2).value = text
    fte.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)

HDR = ["Portfolio","Platform","Squad","Squad Type","Size","Archetype FTE","Filled","Vacant",
       "Total seats","Seats vs archetype","Archetype cost ($m)"]
def hdr_row(row):
    for j, h in enumerate(HDR):
        c = fte.cell(row, 2 + j); c._style = copy(tpl["B5"]._style); c.value = h

RD = "'raw data'"
def cnt(port, squad, status):
    return (f"=COUNTIFS({RD}!$N:$N,\"{port}\",{RD}!$P:$P,\"{squad}\",{RD}!$R:$R,\"{status}\")")

# harvest modelled squads from the tabs (post-modification)
def blocks(w):
    out = []
    rows = [r for r in range(1, w.max_row + 1)
            if isinstance(w.cell(r, 2).value, str) and w.cell(r, 2).value.startswith("Platform: ")]
    for tr in rows:
        hr = tr + 1; ohr = None
        for r in range(hr + 1, w.max_row + 2):
            bv = w.cell(r, 2).value
            if bv == "Platform Overhead" or (isinstance(bv, str) and bv.endswith(" Total")) or (isinstance(bv,str) and bv.startswith("(combined")):
                ohr = r; break
        out.append((tr, hr, hr + 1, (ohr or hr + 2) - 1))
    return out

PORT_LABEL = {"1.1 Ampol Retail":"Ampol Retail","1.2 Customer":"Customer","1.3 Enterprise Data":"Enterprise Data",
 "1.4 TDD Group Functions":"TDD Group Functions","1.5 P&C":"P&C","1.6 Finance":"Finance",
 "1.7 Infrastructure":"Infrastructure","1.8 Energy Solutions & B2B":"Energy Solutions & B2B",
 "1.9 Commercial Fuels":"Commercial Fuels","1.10 Z Retail":"Z Retail","1.11 TDD Cyber":"TDD Cyber"}

r = 4
band(r, "Modelled squads - archetype vs org data (live: archetype follows the tab dropdowns; counts follow raw data cols N-R)")
r += 1
hdr_row(r); r += 1
first_data = r
for t in TABS:
    w = wb[t]
    for (tr, hrr, s0, s1) in blocks(w):
        pname = str(w.cell(tr, 2).value).replace("Platform: ", "")
        if "combined into" in pname: continue
        for sr in range(s0, s1 + 1):
            sqn = w.cell(sr, 2).value
            if not sqn: continue
            fte.cell(r, 2).value = PORT_LABEL[t]
            fte.cell(r, 3).value = pname
            fte.cell(r, 4).value = sqn
            fte.cell(r, 5).value = f"='{t}'!$C${sr}"
            fte.cell(r, 6).value = f"='{t}'!$D${sr}"
            fte.cell(r, 7).value = (f"=IFERROR(INDEX('0.1 Squads'!$F$5:$F$23,MATCH('{t}'!$C${sr}&\"|\"&'{t}'!$D${sr},"
                                    f"'0.1 Squads'!$A$5:$A$23,0)),\"-\")")
            fte.cell(r, 8).value = cnt(PORT_LABEL[t], sqn, "Filled")
            fte.cell(r, 9).value = cnt(PORT_LABEL[t], sqn, "Vacant")
            fte.cell(r, 10).value = f"=H{r}+I{r}"
            fte.cell(r, 11).value = f'=IFERROR(J{r}-G{r},"-")'
            fte.cell(r, 12).value = f"='{t}'!$G${sr}"
            for c in range(2, 13):
                fte.cell(r, c)._style = copy(tpl["B23"]._style)
            fte.cell(r, 12).number_format = MONEY
            r += 1
last_data = r - 1
for c, colL in [(7, "G"), (8, "H"), (9, "I"), (10, "J")]:
    fte.cell(r, c).value = f"=SUM({colL}{first_data}:{colL}{last_data})"
fte.cell(r, 2).value = "Total modelled squads"
fte.cell(r, 12).value = f"=SUM(L{first_data}:L{last_data})"
fte.cell(r, 12).number_format = MONEY
for c in range(2, 13):
    fte.cell(r, c)._style = copy(tpl["B9"]._style)
totals_row = r
r += 2

# Leadership roll-up
band(r, "Leadership roles - roll up to platform / portfolio overhead (Data Config roles), not to squads")
r += 1
for j, h in enumerate(["Portfolio","","","","","","Filled","Vacant","Total"]):
    c = fte.cell(r, 2 + j); c._style = copy(tpl["B5"]._style); c.value = h
r += 1
lead_first = r
for p in ["Ampol Retail","Customer","Enterprise Data","TDD Group Functions","P&C","Finance",
          "Infrastructure","Energy Solutions & B2B","Commercial Fuels","Z Retail","TDD Cyber","COE","Unmapped"]:
    fte.cell(r, 2).value = p
    fte.cell(r, 8).value = f"=COUNTIFS({RD}!$N:$N,\"{p}\",{RD}!$Q:$Q,\"Leadership\",{RD}!$R:$R,\"Filled\")"
    fte.cell(r, 9).value = f"=COUNTIFS({RD}!$N:$N,\"{p}\",{RD}!$Q:$Q,\"Leadership\",{RD}!$R:$R,\"Vacant\")"
    fte.cell(r, 10).value = f"=H{r}+I{r}"
    for c in range(2, 13): fte.cell(r, c)._style = copy(tpl["B23"]._style)
    r += 1
fte.cell(r, 2).value = "Total leadership"
for c, colL in [(8,"H"),(9,"I"),(10,"J")]:
    fte.cell(r, c).value = f"=SUM({colL}{lead_first}:{colL}{r-1})"
for c in range(2, 13): fte.cell(r, c)._style = copy(tpl["B9"]._style)
lead_tot = r
r += 2

# COE planned FTE
band(r, "Centres of Excellence - planned FTE (hard-code) vs org data")
r += 1
for j, h in enumerate(["COE","","","","","Planned FTE","Filled","Vacant","Total seats","Seats vs planned"]):
    c = fte.cell(r, 2 + j); c._style = copy(tpl["B5"]._style); c.value = h
r += 1
coe_first = r
for coe_name in ["COE - Strategy Architecture","COE - Transformation","COE - Business Partnering",
                 "COE - Data","COE (unspecified)"]:
    fte.cell(r, 2).value = coe_name
    fte.cell(r, 7).value = 0; make_input(fte.cell(r, 7))
    fte.cell(r, 8).value = f"=COUNTIFS({RD}!$P:$P,\"{coe_name}\",{RD}!$R:$R,\"Filled\")"
    fte.cell(r, 9).value = f"=COUNTIFS({RD}!$P:$P,\"{coe_name}\",{RD}!$R:$R,\"Vacant\")"
    fte.cell(r, 10).value = f"=H{r}+I{r}"
    fte.cell(r, 11).value = f"=J{r}-G{r}"
    for c in [2,3,4,5,6,8,9,10,11,12]: fte.cell(r, c)._style = copy(tpl["B23"]._style)
    r += 1
fte.cell(coe_first, 7).comment = Comment("Hard-code the planned FTE for each COE.", "QA")
fte.cell(r, 2).value = "Total COE"
for c, colL in [(7,"G"),(8,"H"),(9,"I"),(10,"J")]:
    fte.cell(r, c).value = f"=SUM({colL}{coe_first}:{colL}{r-1})"
for c in range(2, 13): fte.cell(r, c)._style = copy(tpl["B9"]._style)
coe_tot = r
r += 2

# Squads not in archetype model
band(r, "Squads / groups NOT in the archetype model (in op model or org data) - decide: map, archetype, or retire")
r += 1
for j, h in enumerate(["Portfolio","","Squad (org data)","","","","Filled","Vacant","Total"]):
    c = fte.cell(r, 2 + j); c._style = copy(tpl["B5"]._style); c.value = h
r += 1
un_first = r
UNMAPPED = [("Customer","Customer AI"),("Z Retail","Data NZ"),("EGI","EGI (unassigned)")]
for (p, s) in UNMAPPED:
    fte.cell(r, 2).value = p
    fte.cell(r, 4).value = s
    fte.cell(r, 8).value = cnt(p, s, "Filled")
    fte.cell(r, 9).value = cnt(p, s, "Vacant")
    fte.cell(r, 10).value = f"=H{r}+I{r}"
    for c in range(2, 13):
        fte.cell(r, c)._style = copy(tpl["B23"]._style)
        fte.cell(r, c).fill = PatternFill("solid", fgColor="FFFBE4D5")
    r += 1
fte.cell(r, 2).value = "Other unmapped (Class = Unmapped)"
fte.cell(r, 8).value = f"=COUNTIFS({RD}!$Q:$Q,\"Unmapped\",{RD}!$R:$R,\"Filled\")-SUM(H{un_first}:H{r-1})"
fte.cell(r, 9).value = f"=COUNTIFS({RD}!$Q:$Q,\"Unmapped\",{RD}!$R:$R,\"Vacant\")-SUM(I{un_first}:I{r-1})"
fte.cell(r, 10).value = f"=H{r}+I{r}"
for c in range(2, 13):
    fte.cell(r, c)._style = copy(tpl["B23"]._style)
    fte.cell(r, c).fill = PatternFill("solid", fgColor="FFFBE4D5")
un_last = r
r += 2

# cross-check block
band(r, "Cross-check to raw data")
r += 1
checks = [
 ("Org records mapped (raw data)", f"=COUNTA({RD}!$R$2:$R$10000)"),
 ("Accounted for above (squads + leadership + COE + unmapped)",
  f"=J{totals_row}+J{lead_tot}+J{coe_tot}+SUM(J{un_first}:J{un_last})"),
 ("Difference (should be 0)", None),
]
for (label, f) in checks:
    fte.cell(r, 2).value = label
    fte.cell(r, 7).value = f if f else f"=G{r-2}-G{r-1}"
    fte.cell(r, 2)._style = copy(tpl["B8"]._style)
    fte.cell(r, 7)._style = copy(tpl["C8"]._style)
    r += 1

# =====================================================================
# H. conditional formatting refresh (stale ranges from older layout)
# =====================================================================
GREEN_F = PatternFill("solid", fgColor="FFE2EFDA"); GREEN_T = Font(color="FF006100", bold=True)
RED_F = PatternFill("solid", fgColor="FFFBE4D5"); RED_T = Font(color="FF9C0006", bold=True)
def pos_good(w, rng):
    w.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_F, font=GREEN_T))
    w.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_F, font=RED_T))
def pos_bad(w, rng):
    w.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=RED_F, font=RED_T))
    w.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=GREEN_F, font=GREEN_T))
for t in TABS:
    w = wb[t]
    try: w.conditional_formatting._cf_rules.clear()
    except Exception: pass
    vr = frow(w, "G", "Variance (budget less cost)")
    pos_good(w, f"H{vr}")
    ar, lr = refs[t]
    pos_bad(w, f"I{lr}")
try: gs.conditional_formatting._cf_rules.clear()
except Exception: pass
pos_good(gs, "E6:E24")
pos_bad(gs, "I6:I24")

wb.save(OUT)
print("saved", OUT)
print("2.0 applied/left refs:", refs)
