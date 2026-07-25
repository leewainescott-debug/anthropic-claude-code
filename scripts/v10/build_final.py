"""Stage 7: defects a second reader found in the control tabs.

D3  Lists!AK13 read "COE - Cyber, Risk & Service Ops (see 1.13)" where 0.2!B7 says
    "...(see 1.13 Cyber Roles)". My budget map. The MATCH fails, IFERROR returns "", and
    the COE Cyber budget line is silently dropped from 3.1. Zero impact only because
    0.2!E7 is 0 today - put a budget on that line and it vanishes.
D2  0.1 Budget Table column G is empty for rows 8-13, so six 1.x tabs read a zero people
    allowance. The real figures are in the lower table, column I rows 19-24 ($9.1m).
D10 Lists AG/AH hardcoded the rates and unit counts that 0.2 already computes, so editing
    0.2 would not move the allowance.
D12 0.3 column A hardcoded the type|size join key instead of building it.
D1  1.12's visible role list is still missing Rihan Schalkwyk; its own control reads -1.
"""
import openpyxl
from openpyxl.styles import Font
import model

ITAL = Font(italic=True)
LR = model.LAST_ROW
REV = f"'{model.REVIEW}'"

def run(src, dst):
    wb = openpyxl.load_workbook(src); out = []

    # D3 - match the budget label to 0.2 exactly, by reading it rather than retyping it
    ws = wb["Lists"]; dc = wb["0.2 Data Config"]
    labels = {str(dc.cell(r, 2).value or "").strip(): r for r in range(6, 26)}
    fixed = 0
    for r in range(2, 20):
        v = str(ws.cell(r, 37).value or "").strip()        # column AK = 37
        if not v or v in labels:
            continue
        hit = [k for k in labels if k[:26] == v[:26]]
        if hit:
            ws.cell(r, 37).value = f"='0.2 Data Config'!$B${labels[hit[0]]}"
            fixed += 1
    if fixed:
        out.append(f"Lists AK: {fixed} budget label(s) now read 0.2 directly (the COE Cyber "
                   f"line was silently dropping out of 3.1)")

    # D10 - allowance rates and unit counts come from 0.2 / 3.1, not from typed numbers
    for i, cell in enumerate(("$L$6", "$L$7", "$L$8", "$L$14", "$L$15", "$L$9")):
        ws.cell(2 + i, 33).value = f"='0.2 Data Config'!{cell}"
    ws["AH2"] = "=COUNTA('3.1 Group Summary'!$B$6:$B$15)"
    ws["AH3"] = "=$AH$2"; ws["AH4"] = "=$AH$2"; ws["AH7"] = "=$AH$2"
    out.append("Lists AG/AH now reference 0.2 and 3.1 instead of hardcoded rates and counts")

    # D12 - build the archetype join key rather than typing it
    a3 = wb["0.3 Squad Archetypes"]
    for r in range(5, 24):
        if a3.cell(r, 3).value:
            a3.cell(r, 1).value = f'=$C{r}&"|"&$D{r}'
    out.append("0.3 column A key is now =C&\"|\"&D on all 19 rows")

    # D2 - point the six tabs at the people allowance that actually exists
    src_row = {"1.9 Commercial Fuels": 19, "1.1 Ampol Retail": 20, "1.7 Infrastructure": 21,
               "1.6 Finance": 22, "1.5 P&C": 23, "1.2 Customer": 24}
    cellmap = {"1.9 Commercial Fuels": "J13", "1.1 Ampol Retail": "J14",
               "1.7 Infrastructure": "J14", "1.6 Finance": "J14",
               "1.5 P&C": "J13", "1.2 Customer": "J14"}
    for tab, r in src_row.items():
        wb[tab][cellmap[tab]] = f"='0.1 Budget Table (Fin)'!$I${r}"
    out.append("1.1/1.2/1.5/1.6/1.7/1.9: people allowance read from 0.1 column I "
               "(column G is empty for those six segments, so all six read zero)")

    # D4 - a role that counts in headcount but carries no cost
    rv = wb[model.REVIEW]
    rv["AC491"] = ("No cost in column AA - this role counts in the 525 headcount but "
                   "contributes nothing to the $115.113m")
    rv["AC491"].font = ITAL
    out.append("REVIEW r491 (Nidhi Aggarwal) flagged: counts in headcount, no cost in AA")

    wb.save(dst); return out

if __name__ == "__main__":
    for x in run("v2.xlsx", "v3.xlsx"): print("  ", x)
