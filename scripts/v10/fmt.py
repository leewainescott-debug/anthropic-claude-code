"""One formatting standard for the whole workbook.

Every table in the file uses these and nothing else. The rules:

- one header style, dark navy with white bold text, wrapped
- subtotal rows light grey, block totals darker grey with a line above
- a group heading inside a list is pale blue
- money and counts never colour themselves. The old format painted negatives red, so
  a portfolio UNDER budget showed red and one OVER budget showed black, which is
  backwards. No judgement colouring anywhere.
- yellow means an input you can change, and nothing else is yellow
- notes are plain black at 10pt. The first version of this file set them 9pt grey, which
  breaks two of the owner's standing rules at once: nothing under 10pt, and no grey note
  styling. A formatting standard that breaks the formatting rules is worse than none.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FN = "Calibri"          # openpyxl defaults to Cambria, a serif. The file is Calibri.
BAR = "FF002F6C"        # section bar, the darker navy from 1.11 BP&T
NAVY = "FF1F4E79"       # column header row, the lighter navy
PALE = "FFDDEBF7"
GREY = "FFF2F2F2"
MID = "FFD9D9D9"
YELLOW = "FFFFFF00"

HDR_FILL = PatternFill("solid", start_color=NAVY, end_color=NAVY)
GRP_FILL = PatternFill("solid", start_color=PALE, end_color=PALE)
SUB_FILL = PatternFill("solid", start_color=GREY, end_color=GREY)
TOT_FILL = PatternFill("solid", start_color=MID, end_color=MID)
IN_FILL = PatternFill("solid", start_color=YELLOW, end_color=YELLOW)

BAR_FILL = PatternFill("solid", start_color=BAR, end_color=BAR)
HDR_FONT = Font(name=FN, bold=True, color="FFFFFFFF", size=11)
BOLD = Font(name=FN, bold=True, size=11)
BODY = Font(name=FN, size=11)
TITLE = Font(name=FN, bold=True, size=16)
SECTION = Font(name=FN, bold=True, size=12)

WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
CENTRE = Alignment(horizontal="center", vertical="center")

_thin = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
TOPLINE = Border(top=Side(style="thin", color="FF808080"))

# no red, no brackets-plus-colour, one dash for nothing
MONEY_M = '#,##0.00;(#,##0.00);"-"'
MONEY = '#,##0;(#,##0);"-"'
COUNT = '#,##0;(#,##0);"-"'
COUNT1 = '#,##0.0;(#,##0.0);"-"'
TEXT = "General"


def header(ws, row, col0, labels, widths=None):
    """Write a header row and set the column widths under it."""
    from openpyxl.utils import get_column_letter
    for i, lab in enumerate(labels):
        c = ws.cell(row, col0 + i)
        c.value = lab
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = WRAP
        c.border = BOX
        if widths:
            ws.column_dimensions[get_column_letter(col0 + i)].width = widths[i]
    ws.row_dimensions[row].height = 32


def band(ws, row, col0, ncol, fill, font=None, line=False):
    for i in range(ncol):
        c = ws.cell(row, col0 + i)
        c.fill = fill
        c.font = font or BOLD
        if line:
            c.border = TOPLINE


def money(ws, row, col, formula, fmt=MONEY_M, font=None):
    c = ws.cell(row, col)
    c.value = formula
    c.number_format = fmt
    c.alignment = RIGHT
    c.font = font or BODY
    return c


def clear(ws, r0, r1, c0, c1):
    """Wipe values and styling from a block so nothing stale survives a rebuild."""
    import copy
    import openpyxl
    blank = copy.copy(openpyxl.cell.cell.Cell(ws)._style)
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell._style = copy.copy(blank)
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
