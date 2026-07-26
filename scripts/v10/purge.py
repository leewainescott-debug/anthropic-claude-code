"""Delete the retired sources. REVIEW is the only source of truth, so nothing else may
sit in the file claiming to be one.

Keeping them cost real time. `Sheet2` holds an older cyber list - 52 roles, 10 vacant -
against REVIEW's 46 and 4, and that disagreement was reported as a six-role hole in the
ledger. It is not a hole: it is a stale tab. No formula anywhere read it; the only thing
pointing at it was a note on 1.13 saying "roles and costs come straight from Sheet2", which
had not been true since the roles list was repointed at REVIEW.

`3.5 Source Reconciliation` goes with `Squads` for the same reason. Its entire job was to
reconcile the retired Squads tab to REVIEW, thirty-three formulas of it. Once Squads is gone
the tab is reconciling to nothing, and a tab whose whole purpose is to compare the source of
truth against a source that no longer exists is worse than no tab.

A sheet is only deleted when no formula in the workbook reads it. A note that mentions it by
name is removed, because a note pointing at a tab that is not there is how the last round of
stale references got written.
"""
import re

import openpyxl

DROP = ["Sheet2", "Added data", "FY26 Budget (superseded)", "squad mapping (superseded)",
        "3.5 Source Reconciliation", "Squads"]
# notes whose subject is about to be deleted
NOTES = ("Roles and costs come straight from Sheet2",)


def readers(wb, name):
    """Formulas that read this sheet. Text that merely mentions it does not count."""
    out = []
    for ws in wb.worksheets:
        if ws.title == name:
            continue
        for row in ws.iter_rows():
            v = None
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("=") and name in v:
                    out.append(f"{ws.title}!{c.coordinate}")
    return out


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = []
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip().startswith(NOTES):
                    c.value = None
                    n += 1
    out.append(f"{n} notes removed that pointed at a tab being deleted")
    # deleted in the order given, so a tab is gone before the tab that read it is checked
    for name in DROP:
        if name not in wb.sheetnames:
            out.append(f"{name}: not present")
            continue
        r = readers(wb, name)
        if r:
            out.append(f"{name}: KEPT, {len(r)} formulas read it - {', '.join(r[:4])}")
            continue
        del wb[name]
        out.append(f"{name}: deleted")
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
