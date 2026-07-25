"""Repair REVIEW's derived columns, then rebuild the 2.x tabs so the lever actually works.

Defects replaced (all reproduced in the current file):
  * REVIEW col AP (MSquadC) invented squad groupings that contradict Lee's col K
    (it showed Z Energy Martech = 16 where col K says 2, and dropped Z Loyalty &
    Martech, Ampol Loyalty & Martech and Z App and Web entirely);
  * REVIEW col AQ (MLead) flagged Business Analysts and QA Analysts as leadership
    while col K already carries a literal "Leadership" squad;
  * 2.x squad rows read FIXED row ranges (COUNTIF($E$26:$E$35,"Hire")) while the role
    list was alphabetical by name with no squad column, so the ranges were arbitrary;
  * "new cost" was M + (Hire + 0.4*Offshore), which ADDED 40% of an offshored role on
    top of its full cost - offshoring a filled role made cost go UP, and Hold did nothing;
  * costs and statuses were hardcoded literals, so nothing followed REVIEW;
  * columns E/F/J/P carried headers Lee wrote but no formula behind them.

Lever: Filled/Hire = 1.0x, Hold = 0.0x, Offshore = 0.4x, applied per role, live for
filled and vacant alike, and every squad and portfolio total reads the levered column.
"""
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

import model

REV = f"'{model.REVIEW}'"
LR = model.LAST_ROW
LEVER_LIST = '"Filled,Hire,Hold,Offshore"'

# REVIEW derived columns. AJ (portfolio) and AK (status) already compute correctly, so
# they stay. AP and AQ were fabricated and are rewritten from Lee's raw columns.
C_PF, C_STATUS, C_SQUAD, C_LEAD = "AJ", "AK", "AP", "AQ"

HDRS = {
    "B": "Squad", "C": "Archetype type", "D": "Archetype roles",
    "E": "Total roles (incl. vacant roles)",
    "F": "Variance between archetype and actual squads",
    "G": "Vacant", "H": "Planning to hire or offshore", "I": "Vacancies remaining",
    "J": "New total number of roles in squad", "K": "Flag",
    "L": "Archetype cost ($m)", "M": "Full cost of model ($m)",
    "N": "Variance between archetype & actuals", "O": "New cost of squads ($m)",
    "P": "Total cost after overhead ($m)", "Q": "New Variance ($m)", "R": "Archetype size",
}
ROLE_HDRS = {"B": "Name", "C": "Role", "D": "Status", "E": "Vacancy lever",
             "F": "Squad", "G": "Cost if hired ($)", "H": "Effective cost ($)"}

# 2.x tab -> (design tab, portfolio-overhead row, platform-overhead row).
# COE and EGI tabs get no portfolio overhead: they ARE the shared capacity that the
# portfolio overhead pays for, so charging them again would double count.
DESIGN = {
    "2.1 Ampol Retail": ("1.1 Ampol Retail", 6, 7),
    "2.2 Customer": ("1.2 Customer", 6, 7),
    "2.3 Enterprise Data": ("1.3 Enterprise Data", 6, 7),
    "2.4 TDD Group Functions": ("1.4 TDD Group Functions", 6, 7),
    "2.5 P&C": ("1.5 P&C", 6, 7),
    "2.6 Finance": ("1.6 Finance", 6, 7),
    "2.7 Infrastructure": ("1.7 Infrastructure", 7, 8),
    "2.8 Energy Solutions & B2B": ("1.8 Energy Solutions & B2B", 6, 7),
    "2.9 Commercial Fuels": ("1.9 Commercial Fuels", 6, 7),
    "2.10 Z Retail": ("1.10 Z Retail", 6, 7),
}


def fix_review(wb, roles):
    """Rewrite AP/AQ from Lee's raw columns; label them for what they are."""
    ws = wb[model.REVIEW]
    by_row = {r["row"]: r for r in roles}
    ws[f"{C_SQUAD}1"] = "Squad (from col K)"
    ws[f"{C_LEAD}1"] = "Leadership"
    for i in range(2, LR + 1):
        rec = by_row.get(i)
        ws[f"{C_SQUAD}{i}"] = rec["squad"] if rec else None
        ws[f"{C_LEAD}{i}"] = (1 if rec["leadership"] else 0) if rec else None
    return ws


def archetype_index(wb):
    ws = wb["0.3 Squad Archetypes"]
    return {(str(ws.cell(r, 3).value).strip(), str(ws.cell(r, 4).value).strip()): r
            for r in range(5, 24) if ws.cell(r, 3).value and ws.cell(r, 4).value}


def squad_design(wb):
    """squad (lower) -> (type, size) as Lee has chosen it on his 1.x design tabs."""
    out = {}
    for sn in wb.sheetnames:
        if not sn.startswith("1."):
            continue
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            nm, typ, size = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value
            if not nm or not typ:
                continue
            n = str(nm).strip()
            low = n.lower()
            if low in ("squad", "platform overhead", "cost") or low.endswith("total"):
                continue
            out.setdefault(low, (str(typ).strip(), str(size).strip() if size else ""))
    return out


def build(path_in, path_out, roles):
    wb = openpyxl.load_workbook(path_in)
    fix_review(wb, roles)
    arch = archetype_index(wb)
    design = squad_design(wb)
    report = []

    for tab, pf in model.TAB_PORTFOLIO.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        mine = [r for r in roles if r["portfolio"] == pf]

        def order(s):
            tail = 2 if s == model.BLANK_SQUAD_LABEL else (1 if s == "Leadership" else 0)
            return (tail, -sum(1 for r in mine if r["squad"] == s), s)
        squads = sorted({r["squad"] for r in mine}, key=order)

        title = ws.cell(2, 2).value
        for r in range(4, ws.max_row + 1):
            for c in range(1, max(ws.max_column, 18) + 1):
                if ws.cell(r, c).value is not None:
                    ws.cell(r, c).value = None
        for dv in list(ws.data_validations.dataValidation):
            ws.data_validations.dataValidation.remove(dv)

        hdr, first = 5, 6
        last = first + len(squads) - 1
        tot = last + 1
        m1, m2, m3 = tot + 2, tot + 3, tot + 4
        rl_hdr = m3 + 3
        rl0 = rl_hdr + 1
        rl1 = rl0 + len(mine) - 1
        RL = lambda col: f"${col}${rl0}:${col}${rl1}"

        ws.cell(4, 2).value = ("Position by squad - archetype vs actual "
                               "(roles, status and cost from REVIEW)")
        for col, h in HDRS.items():
            cell = ws[f"{col}{hdr}"]
            cell.value = h
            cell.alignment = Alignment(wrap_text=True, vertical="bottom")

        oh = None
        if tab in DESIGN:
            dt, r_pf, r_pl = DESIGN[tab]
            oh = f"('{dt}'!$F${r_pf}+'{dt}'!$F${r_pl})"

        for i, sq in enumerate(squads):
            r = first + i
            ws[f"B{r}"] = sq
            d = design.get(sq.lower())
            if d and d in arch:
                ar = arch[d]
                ws[f"C{r}"] = f"='0.3 Squad Archetypes'!$C${ar}"
                ws[f"R{r}"] = f"='0.3 Squad Archetypes'!$D${ar}"
                ws[f"D{r}"] = f"='0.3 Squad Archetypes'!$F${ar}"
                ws[f"L{r}"] = f"='0.3 Squad Archetypes'!$G${ar}"
            else:
                # No archetype exists for this squad in 0.3. Say so, rather than showing
                # the silent zero that made these squads look free.
                ws[f"C{r}"] = "Outside archetype model"
                ws[f"R{r}"] = "-"
                ws[f"D{r}"] = "-"
                ws[f"L{r}"] = "-"
                report.append((tab, f"no archetype: {sq}",
                               sum(1 for x in mine if x["squad"] == sq), ""))

            ws[f"E{r}"] = f"=COUNTIFS({RL('F')},$B{r})"
            ws[f"F{r}"] = f'=IFERROR($E{r}-$D{r},"-")'
            ws[f"G{r}"] = f'=COUNTIFS({RL("F")},$B{r},{RL("D")},"Vacant")'
            ws[f"H{r}"] = (f'=COUNTIFS({RL("F")},$B{r},{RL("D")},"Vacant",{RL("E")},"Hire")'
                           f'+COUNTIFS({RL("F")},$B{r},{RL("D")},"Vacant",{RL("E")},"Offshore")')
            ws[f"I{r}"] = f"=$G{r}-$H{r}"
            ws[f"J{r}"] = f'=$E{r}-COUNTIFS({RL("F")},$B{r},{RL("E")},"Hold")'
            ws[f"K{r}"] = (f'=IF(NOT(ISNUMBER($D{r})),"Outside the archetype model",'
                           f'IF($J{r}>$D{r},"Over archetype after decisions",'
                           f'IF($J{r}<$D{r},"Under archetype","On archetype")))')
            ws[f"M{r}"] = f"=SUMIFS({RL('G')},{RL('F')},$B{r})/1000000"
            ws[f"N{r}"] = f'=IFERROR($M{r}-$L{r},"-")'
            ws[f"O{r}"] = f"=SUMIFS({RL('H')},{RL('F')},$B{r})/1000000"
            ws[f"P{r}"] = (f"=$O{r}+{oh}*$E{r}/$E${tot}" if oh else f"=$O{r}")
            ws[f"Q{r}"] = f'=IFERROR($P{r}-$L{r},"-")'

        ws[f"B{tot}"] = "Total"
        for col in "DEFGHIJLNOPQ":
            ws[f"{col}{tot}"] = f"=SUM({col}{first}:{col}{last})"
        # M ties straight to the ledger, so a squad that failed to join is visible as a
        # gap against SUM(M) rather than quietly shrinking the portfolio total.
        ws[f"M{tot}"] = (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!${C_PF}$2:${C_PF}${LR},"{pf}")'
                         f"/1000000")
        ws[f"E{tot}"] = (f'=COUNTIFS({REV}!${C_PF}$2:${C_PF}${LR},"{pf}")')

        ws[f"B{m1}"] = "Cost to fill every vacancy at full onshore cost ($m)"
        ws[f"E{m1}"] = f'=SUMIFS({RL("G")},{RL("D")},"Vacant")/1000000'
        ws[f"B{m2}"] = "Of which currently marked Hire or Offshore ($m)"
        ws[f"E{m2}"] = (f'=(SUMIFS({RL("G")},{RL("D")},"Vacant",{RL("E")},"Hire")'
                        f'+0.4*SUMIFS({RL("G")},{RL("D")},"Vacant",{RL("E")},"Offshore"))'
                        f"/1000000")
        ws[f"B{m3}"] = ("Leadership already inside this portfolio ($m) - the overhead in "
                        "column P is an allowance for the same people")
        ws[f"E{m3}"] = (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!${C_PF}$2:${C_PF}${LR},"{pf}",'
                        f'{REV}!${C_LEAD}$2:${C_LEAD}${LR},1)/1000000')

        ws.cell(rl_hdr - 1, 2).value = f"{pf} roles - from REVIEW, one row per named role"
        for col, h in ROLE_HDRS.items():
            cell = ws[f"{col}{rl_hdr}"]
            cell.value = h
            cell.alignment = Alignment(wrap_text=True, vertical="bottom")

        for i, rec in enumerate(sorted(mine, key=lambda x: (order(x["squad"]),
                                                            x["vacant"], x["name"]))):
            r = rl0 + i
            src = rec["row"]
            ws[f"B{r}"] = f"=INDEX({REV}!$B:$B,{src})"
            ws[f"C{r}"] = f"=INDEX({REV}!$C:$C,{src})"
            ws[f"D{r}"] = f'=IF(ISNUMBER(SEARCH("vacant",$B{r})),"Vacant","Filled")'
            ws[f"E{r}"] = rec["lever"]
            ws[f"F{r}"] = rec["squad"]
            ws[f"G{r}"] = f"=INDEX({REV}!$AA:$AA,{src})"
            ws[f"H{r}"] = (f'=$G{r}*IF($E{r}="Offshore",{model.OFFSHORE_RATE},'
                           f'IF($E{r}="Hold",0,1))')

        dv = DataValidation(type="list", formula1=LEVER_LIST, allow_blank=False)
        dv.errorTitle, dv.error = "Vacancy lever", "Choose Filled, Hire, Hold or Offshore."
        ws.add_data_validation(dv)
        dv.add(f"E{rl0}:E{rl1}")

        ws.cell(2, 2).value = title
        report.append((tab, "BUILT", len(mine),
                       f"{len(squads)} squads, total row {tot}, roles {rl0}-{rl1}"))

    wb.save(path_out)
    return report


if __name__ == "__main__":
    roles = model.load("NEW.xlsx")
    for t, a, b, c in build("NEW.xlsx", "stage1.xlsx", roles):
        print(f"  {t:28s} {str(a)[:36]:38s} {b:4} {c}")
