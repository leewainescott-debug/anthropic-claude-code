"""The sixth pass: every new figure on the 1.x tabs, rebuilt from the ledger.

The five standing passes were written before the 1.x tabs carried an actual cost, so none of
them looks at the new columns. This one does, and it does it the way recompute.py does - from
REVIEW, in Python, knowing nothing about the workbook's formulas.

It finds the squad rows its own way. Reusing the writer's block finder would hide the one
failure that matters most: a squad table the writer never saw, left with no comparison on it.
Here a squad row is any row on a 1.x tab whose name matches a group on the working tab, which
is the definition the reader would use.

And it runs twice - once as shipped, once with a lever pulled - because "equal to the actual
cost after decisions" is a claim about what the cell does when the reader changes something,
not about the number it happens to show today.
"""
import collections
import json
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter as L

import wbio

REVIEW = "REVIEW - Complete Role Mapping"
TOL = 1e-6
B_BAR = "Archetype against actual - every squad on this tab"
COE = ("1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles")


def ledger(wv):
    """Cost per (portfolio, squad-or-overhead-line), straight off REVIEW."""
    cost = collections.Counter()
    R = wv[REVIEW]
    for i in range(2, 529):
        if str(R.cell(i, 2).value or "").strip():
            cost[(str(R.cell(i, 36).value), str(R.cell(i, 46).value))] += \
                R.cell(i, 27).value or 0
    return cost


def levered(wv, tab):
    """Cost after decisions per squad, summed off the role detail block on a working tab.

    The detail block is where a decision is set and the only place it is set, so reading it
    here keeps the recomputation clear of every summary formula in the file. The squad
    heading row carries the block's own total, so it is skipped rather than added twice.
    """
    ws = wv[tab]
    out, squad = collections.Counter(), None
    for r in range(1, ws.max_row + 1):
        c = str(ws.cell(r, 3).value or "").strip()
        if re.match(r"^\d+ roles?$", c):                     # 'Above Store' | '11 roles'
            squad = str(ws.cell(r, 2).value or "").strip()
            continue
        if squad and isinstance(ws.cell(r, 7).value, (int, float)) \
                and str(ws.cell(r, 5).value or "").strip():  # a role row carries a status
            out[squad] += ws.cell(r, 7).value
    return out


def groups(wv, tab, lo, hi):
    """Every group named on a working tab between its first squad row and its total."""
    ws = wv[tab]
    return {str(ws.cell(r, 2).value).strip() for r in range(lo, hi)
            if isinstance(ws.cell(r, 2).value, str) and ws.cell(r, 2).value.strip()
            and not ws.cell(r, 2).value.strip().endswith(" total")}


def variant_of(wb):
    for t in wb.sheetnames:
        if re.match(r"^1\.\d+ ", t):
            for r in range(1, wb[t].max_row + 1):
                if str(wb[t].cell(r, 2).value or "").strip() == B_BAR:
                    return "B"
    return "A"


def untouched(path, base="cand.xlsx"):
    """Nothing that was already in the shipped workbook may have changed.

    Design A appended two columns at a fixed K and L, and on 1.6 the squad table is five
    columns wider than the others: K and L held the owner's own Nbr Archetype Roles and
    Published Roles, typed numbers, and they were silently replaced. No formula broke, no
    total moved and all six other passes stayed green, because a typed number that is
    overwritten by a formula returning a different number is still just a number.
    """
    b, w = (openpyxl.load_workbook(base), openpyxl.load_workbook(path))
    bv, wv = (openpyxl.load_workbook(base, data_only=True),
              openpyxl.load_workbook(path, data_only=True))
    lost, moved = [], []
    for t in b.sheetnames:
        if t not in w.sheetnames:
            lost.append(f"{t}: the tab is gone")
            continue
        for r in range(1, b[t].max_row + 1):
            # the row is searched on the same footing as the cell being judged: a note that
            # is a formula has one text when read as a formula and another when read as a
            # value, and matching a cached string against a row of formulas found nothing
            rows = {id(w[t]): {w[t].cell(r, c).value for c in range(1, w[t].max_column + 1)},
                    id(wv[t]): {wv[t].cell(r, c).value
                                for c in range(1, wv[t].max_column + 1)}}
            for c in range(1, b[t].max_column + 1):
                for f, g, what in ((b[t], w[t], "formula"), (bv[t], wv[t], "value")):
                    row = rows[id(g)]
                    was = f.cell(r, c).value
                    if was is None or was == g.cell(r, c).value:
                        continue
                    # widening a table under an annotation pushes the annotation along its
                    # own row. That is a move, not a loss, but it is reported either way -
                    # folding moves silently into "unchanged" hid six table headers being
                    # carried out of their own tables.
                    if isinstance(was, str) and was in row:
                        if what == "formula":
                            moved.append(f"{t} row {r}: {was[:56]!r} left {L(c)}")
                        continue
                    lost.append(f"{t}!{L(c)}{r} {what} was {was!r}, now "
                                f"{g.cell(r, c).value!r}")
    return lost, moved


KEY, ARCH = "Squad", "Total Squad Cost ($m)"
ACT, VAR = "Actual cost after decisions ($m)", "Variance to archetype ($m)"


def tables(wsv):
    """Every table on the tab that carries an actual cost, as (header row, {label: col}).

    The columns are found by reading the header text, never by knowing where the writer put
    them - design A puts its pair at K on most tabs and at P on 1.6, and a checker that
    assumed a column would have confirmed the wrong cells.
    """
    out = []
    for r in range(1, wsv.max_row + 1):
        m = {}
        for c in range(2, 26):
            v = wsv.cell(r, c).value
            if isinstance(v, str) and v.strip() in (KEY, ARCH, ACT, VAR):
                m.setdefault(v.strip(), c)
        if KEY in m and ACT in m and ARCH in m:
            out.append((r, m))
    return out


def named(wb, wv, one, valid, v):
    """(row, squad, actual, archetype, variance) for every squad row under every table."""
    # only rows under a header row are candidates, which is what keeps the tab title out:
    # 1.5 is titled P&C and also carries a squad called P&C, and a whole-sheet scan read
    # the title as a squad row missing its figure
    ws, out = wv[one], []
    heads = tables(ws)
    for i, (hr, m) in enumerate(heads):
        end = heads[i + 1][0] if i + 1 < len(heads) else ws.max_row + 1
        for r in range(hr + 1, end):
            b = str(ws.cell(r, m[KEY]).value or "").strip()
            if b not in valid:
                continue
            out.append((r, b, ws.cell(r, m[ACT]).value, ws.cell(r, m[ARCH]).value,
                        ws.cell(r, m.get(VAR, m[ACT] + 1)).value))
    return out


def scan(path, anchors="anchors_final.json"):
    wb = openpyxl.load_workbook(path)
    wv = openpyxl.load_workbook(path, data_only=True)
    a = json.load(open(anchors))
    cost = ledger(wv)
    live = {str(wv[t]["C3"].value): t for t in wb.sheetnames if t.startswith("2.")}
    v = variant_of(wb)
    bad, seen, tabs = [], 0, 0
    for one in sorted(t for t in wb.sheetnames if re.match(r"^1\.\d+ ", t)):
        if one in COE:
            note = [r for r in range(1, min(wv[one].max_row, 30) + 1)
                    if str(wv[one].cell(r, 2).value or "").startswith("These are centres")]
            if not note:
                bad.append(f"{one}: a COE tab with no note saying why it has no comparison")
            continue
        anchor = next((x for x in a.values() if x["tab"].split(" ", 1)[-1] ==
                       one.split(" ", 1)[-1]), None)
        if anchor is None:
            bad.append(f"{one}: no working tab pairs with it")
            continue
        tab = live[anchor["pf"]]
        lev = levered(wv, tab)
        valid = groups(wv, tab, anchor["first_squad"], anchor["total_row"])
        rows = named(wb, wv, one, valid, v)
        if not rows:
            bad.append(f"{one}: names a group but carries no actual cost")
            continue
        tabs += 1
        for r, name, actual, arch, var in rows:
            seen += 1
            exp = lev.get(name, cost.get((anchor["pf"], name)))
            if exp is None:
                bad.append(f"{one} row {r} {name}: no such group in the ledger")
                continue
            if not isinstance(actual, (int, float)):
                bad.append(f"{one} row {r} {name}: actual reads {actual!r}")
                continue
            if abs(actual - exp / 1e6) > TOL:
                bad.append(f"{one} row {r} {name}: tab {actual!r}, ledger {exp / 1e6!r}")
            num = isinstance(arch, (int, float))
            if num and isinstance(var, (int, float)):
                if abs(var - (actual - arch)) > TOL:
                    bad.append(f"{one} row {r} {name}: variance {var!r}, actual less "
                               f"archetype {actual - arch!r}")
            elif num != isinstance(var, (int, float)):
                bad.append(f"{one} row {r} {name}: archetype {arch!r}, variance {var!r} - "
                           f"one is a figure and the other is not")
    return bad, seen, tabs, v


def totals(path, anchors="anchors_final.json"):
    """Platform totals add up their own squad rows, and every control reads zero.

    The run of squad rows is bounded by the platform bar above it and the total below it,
    so a figure elsewhere on the tab - a budget block, the footer - cannot drift into the
    sum. Reading every numeric cell in the column instead put 75.78 against Z Supply.
    """
    wb = openpyxl.load_workbook(path)
    wv = openpyxl.load_workbook(path, data_only=True)
    a = json.load(open(anchors))
    live = {str(wv[t]["C3"].value): t for t in wb.sheetnames if t.startswith("2.")}
    v, bad = variant_of(wb), []
    for one in sorted(t for t in wb.sheetnames if re.match(r"^1\.\d+ ", t)):
        anchor = next((x for x in a.values() if x["tab"].split(" ", 1)[-1] ==
                       one.split(" ", 1)[-1]), None)
        if anchor is None:
            continue
        ws = wv[one]
        valid = groups(wv, live[anchor["pf"]], anchor["first_squad"], anchor["total_row"])
        heads = tables(ws)
        for i, (hr, m) in enumerate(heads):
            end = heads[i + 1][0] if i + 1 < len(heads) else ws.max_row + 1
            run = []
            for r in range(hr + 1, end):
                lab = str(ws.cell(r, 2).value or "").strip()
                x = ws.cell(r, m[ACT]).value
                if (lab.endswith(" Total") or lab.endswith(" total")) and run:
                    s = sum(run)
                    if not isinstance(x, (int, float)) or abs(x - s) > TOL:
                        bad.append(f"{one} row {r} {lab}: total {x!r}, its "
                                   f"{len(run)} squad rows add to {s!r}")
                    run = []
                elif str(ws.cell(r, m[KEY]).value or "").strip() in valid \
                        and isinstance(x, (int, float)):
                    run.append(x)
    ctl = []
    for t in wv.sheetnames:
        ws = wv[t]
        cols = [m[ACT] for _, m in tables(ws)]
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 2).value or "").startswith("Control - every line above"):
                got = [ws.cell(r, c).value for c in set(cols)
                       if isinstance(ws.cell(r, c).value, (int, float))]
                ctl.append((t, r, got[0] if got else "no figure in any actual column"))
    return bad, ctl


def lever(path):
    """Pull one lever and confirm the new 1.x figures move by exactly the role's cost."""
    wb = openpyxl.load_workbook(path)
    wv0 = openpyxl.load_workbook(path, data_only=True)
    v = variant_of(wb)
    tab = next(t for t in wb.sheetnames if t.startswith("2.")
               and str(wv0[t]["C3"].value) == "Ampol Retail")
    one = next(t for t in wb.sheetnames if re.match(r"^1\.\d+ ", t)
               and t.split(" ", 1)[-1] == "Ampol Retail")
    ws, wsv = wb[tab], wv0[tab]
    # find a role in the detail block by its heading and its cost, never by row number, and
    # read the headings off the calculated values - on the formula sheet '11 roles' is a
    # formula, so matching there found nothing at all
    squad = pick = None
    for r in range(1, ws.max_row + 1):
        if re.match(r"^\d+ roles?$", str(wsv.cell(r, 3).value or "").strip()):
            squad = str(wsv.cell(r, 2).value or "").strip()
        elif squad and isinstance(wsv.cell(r, 6).value, (int, float)) \
                and wsv.cell(r, 6).value > 150000 \
                and str(wsv.cell(r, 4).value or "").strip() == "Filled":
            pick = (r, squad, wsv.cell(r, 6).value)
            break
    if pick is None:
        return ["lever: no role found to pull"]
    r, squad, amt = pick
    a = json.load(open("anchors_final.json"))
    anchor = next(x for x in a.values() if x["pf"] == "Ampol Retail")
    valid = groups(wv0, tab, anchor["first_squad"], anchor["total_row"])
    before = {n: c for _, n, c, _, _ in named(wb, wv0, one, valid, v)}
    ws.cell(r, 5).value = "Hold"
    tmp = path.replace(".xlsx", "_lever.xlsx")
    wb.save(tmp)
    wv = openpyxl.load_workbook(wbio.recalc(tmp), data_only=True)
    after = {n: c for _, n, c, _, _ in named(openpyxl.load_workbook(tmp), wv, one, valid, v)}
    out = [f"lever test: {tab} row {r}, a ${amt:,.0f} filled role in {squad} set to Hold"]
    moved = 0
    for n in before:
        d = (after.get(n) or 0) - (before.get(n) or 0)
        want = -amt / 1e6 if n == squad else 0.0
        ok = abs(d - want) < TOL
        if not ok or n == squad:
            out.append(f"  {'ok ' if ok else 'FAIL'} {one} {n} {before[n]:.6f} -> "
                       f"{after.get(n):.6f} ({d:+.6f}, expected {want:+.6f})")
        moved += 0 if ok else 1
    out.append(f"  {'ok ' if not moved else 'FAIL'} the other {len(before) - 1} squads on "
               f"{one} unmoved")
    return out


if __name__ == "__main__":
    p = sys.argv[1] if len(sys.argv) > 1 else "cand_A.xlsx"
    bad, seen, tabs, v = scan(p)
    tbad, ctl = totals(p)
    lost, moved = untouched(p)
    print("=" * 74)
    print(f"{p}: design {v}, {seen} squad rows across {tabs} tabs carry an actual cost")
    print(f"FIGURES ON THE 1.x TABS DISAGREEING WITH THE LEDGER: {len(bad) + len(tbad)}")
    for x in (bad + tbad)[:40]:
        print("   ", x)
    print(f"CELLS OF THE SHIPPED WORKBOOK OVERWRITTEN: {len(lost)}")
    for x in lost[:20]:
        print("   ", x)
    print(f"NOTES MOVED ALONG THEIR OWN ROW: {len(moved)}")
    for x in moved[:20]:
        print("   ", x)
    nz = [c for c in ctl if not isinstance(c[2], (int, float)) or abs(c[2]) > TOL]
    print(f"CONTROLS: {len(ctl)} cells, {len(nz)} not zero {nz[:6]}")
    for x in lever(p):
        print("  ", x)
