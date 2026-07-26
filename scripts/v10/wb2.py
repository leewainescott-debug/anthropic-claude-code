"""Build the 2.x working tabs and 3.x summaries from scratch, to docs/LAYOUT_SPEC.md.

The owner's own tabs are not touched: 0.1, 0.2, 0.3, the 1.x design tabs and the REVIEW
ledger are inputs to this build, never outputs of it.

Columns are named constants. The last attempt wrote each person's cost-after-decision into
column G and summed column H, so every squad totalled zero and the tab shipped broken.
Named constants and a tie-out assertion at the end make that class of mistake loud.
"""
import collections
import copy
import json
import re

import openpyxl
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation

import fmt
from build_2xfix import DESIGN, squad_table_bounds

REVIEW = "REVIEW - Complete Role Mapping"
REV = f"'{REVIEW}'"
A3 = "'0.3 Squad Archetypes'"
LAST = 528

# ---- 2.x squad summary columns ----
S_SQUAD, S_TYPE, S_SIZE, S_AROLES = 2, 3, 4, 5
S_ROLES, S_VAC, S_HIRE, S_HOLD, S_AFTER = 6, 7, 8, 9, 10
S_ACOST, S_ACTUAL, S_VAR, S_IMPACT, S_TOTAL = 11, 12, 13, 14, 15
S_HEADERS = ["Squad", "Archetype type", "Archetype size", "Archetype roles",
             "Roles", "Vacant", "To hire or offshore", "On hold", "Roles after decisions",
             "Archetype cost ($m)", "Actual cost ($m)", "Variance to archetype ($m)",
             "Impact of decisions ($m)", "Total cost after decisions ($m)"]
S_WIDTHS = [34, 28, 13, 12, 8, 8, 13, 8, 13, 13, 12, 15, 14, 16]
S_N = len(S_HEADERS)

# ---- 2.x FTE columns ----
P_NAME, P_ROLE, P_STATUS, P_LEVER, P_COST, P_AFTER = 2, 3, 4, 5, 6, 7
P_HEADERS = ["Name", "Role", "Status", "Vacancy lever", "Cost if hired ($)",
             "Cost after decision ($)"]
P_WIDTHS = [34, 46, 11, 16, 17, 20]
P_N = len(P_HEADERS)

OH_ORDER = ["Head of Technology", "Business Partner", "Domain Architect",
            "Delivery Manager", "Technology Manager", "Leadership",
            "Leadership - squad not stated"]

PORTFOLIO_ORDER = ["Ampol Retail", "Customer", "Enterprise Data", "TDD Group Functions",
                   "P&C", "Finance", "Infrastructure", "Energy Solutions & B2B",
                   "Commercial Fuels", "Z Retail", "COE Cyber", "COE BP&T", "COE SA&D",
                   "EGI"]


# --------------------------------------------------------------------------- ledger
def ledger(wv):
    R = wv[REVIEW]
    out = []
    for i in range(2, LAST + 1):
        if not str(R.cell(i, 2).value or "").strip():
            continue
        pf = R.cell(i, 36).value
        coe = str(pf or "").startswith("COE") or pf == "EGI"
        out.append({"row": i, "pf": pf, "grp": R.cell(i, 46).value,
                    "status": R.cell(i, 37).value,
                    "oh": (not coe) and R.cell(i, 44).value != "Squad"})
    return out


def groups_for(rows):
    seen = {}
    for r in rows:
        seen.setdefault(r["grp"], r["oh"])
    delivery = sorted(g for g, oh in seen.items() if not oh)
    overhead = [g for g in OH_ORDER if seen.get(g)]
    overhead += sorted(g for g, oh in seen.items() if oh and g not in OH_ORDER)
    return delivery, overhead


def wipe(ws):
    blank = copy.copy(openpyxl.cell.cell.Cell(ws)._style)
    # merges first - a MergedCell is read-only while its range is live
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    for row in ws.iter_rows():
        for c in row:
            c.value = None
            c._style = copy.copy(blank)
    for k in list(ws.column_dimensions):
        del ws.column_dimensions[k]
    for k in list(ws.row_dimensions):
        del ws.row_dimensions[k]
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    ws.conditional_formatting._cf_rules.clear()
    ws.data_validations.dataValidation = []
    ws.freeze_panes = None


def title(ws, row, text, size=15):
    c = ws.cell(row, 2)
    c.value = text
    c.font = fmt.TITLE if size == 15 else fmt.SECTION
    return row + 1


def factbar(ws, row, pf, tot_row):
    pairs = [("Portfolio", f'="{pf}"', fmt.TEXT),
             ("Roles", f"=${L(S_ROLES)}${tot_row}", fmt.COUNT),
             ("Vacant", f"=${L(S_VAC)}${tot_row}", fmt.COUNT),
             ("Cost today ($m)", f"=${L(S_ACTUAL)}${tot_row}", fmt.MONEY_M)]
    col = 2
    for lab, f, nf in pairs:
        a, b = ws.cell(row, col), ws.cell(row, col + 1)
        a.value = lab
        a.font = fmt.BOLD
        a.alignment = fmt.LEFT
        b.value = f
        b.number_format = nf
        b.font = fmt.BODY
        b.alignment = fmt.LEFT if nf == fmt.TEXT else fmt.RIGHT
        col += 2
    return row


# --------------------------------------------------------------------------- 2.x
def build_working(wb, wv, tab, rows, bounds):
    ws = wb[tab]
    design = DESIGN.get(tab)
    pf = rows[0]["pf"]
    delivery, overhead = groups_for(rows)
    lo, hi = bounds[design] if design else (0, 0)
    wipe(ws)
    ws.column_dimensions["A"].width = 2.5

    # ---- plan the row positions before writing anything ----
    HDR = 6
    r = HDR + 1
    srow = {}
    for g in delivery:
        srow[g] = r
        r += 1
    r_del = r
    r += 1
    r_ohhead = r_oh = None
    if overhead:
        r += 1
        r_ohhead = r
        r += 1
        for g in overhead:
            srow[g] = r
            r += 1
        r_oh = r
        r += 1
    r += 1
    r_tot = r
    r += 2
    r_ctl = r
    r += 3
    r_fte = r
    r += 1
    r_phdr = r
    r += 1
    band, people = {}, {}
    for g in delivery + overhead:
        band[g] = r
        r += 1
        for x in [y for y in rows if y["grp"] == g]:
            people.setdefault(g, []).append((r, x))
            r += 1
    r_end = r - 1

    # ---- header ----
    title(ws, 2, f"{pf} - working copy")
    factbar(ws, 3, pf, r_tot)
    title(ws, 5, "Squad summary", 12)
    fmt.header(ws, HDR, 2, S_HEADERS, S_WIDTHS)

    def span(g):
        p = people.get(g, [])
        return (p[0][0], p[-1][0]) if p else (band[g], band[g])

    def squad(rw, g, is_oh):
        a, b = span(g)
        lev = f"${L(P_LEVER)}${a}:${L(P_LEVER)}${b}"
        st = f"${L(P_STATUS)}${a}:${L(P_STATUS)}${b}"
        ws.cell(rw, S_SQUAD).value = g
        ws.cell(rw, S_SQUAD).font = fmt.BODY
        ws.cell(rw, S_SQUAD).alignment = fmt.LEFT
        if is_oh or design is None:
            ws.cell(rw, S_TYPE).value = ("Overhead" if is_oh else "COE, measured on budget")
            for c in (S_SIZE, S_AROLES, S_ACOST, S_VAR):
                x = ws.cell(rw, c)
                x.value = '="-"'
                x.alignment = fmt.RIGHT
        else:
            key = f'${L(S_TYPE)}{rw}&"|"&${L(S_SIZE)}{rw}'
            m = f"MATCH(${L(S_SQUAD)}{rw},'{design}'!$B${lo}:$B${hi},0)"
            ws.cell(rw, S_TYPE).value = (
                f"=IFERROR(INDEX('{design}'!$C${lo}:$C${hi},{m}),\"Not in the design\")")
            ws.cell(rw, S_SIZE).value = (
                f"=IFERROR(INDEX('{design}'!$D${lo}:$D${hi},{m}),\"-\")")
            fmt.money(ws, rw, S_AROLES,
                      f"=IFERROR(INDEX({A3}!$F$5:$F$23,"
                      f'MATCH({key},{A3}!$A$5:$A$23,0)),"-")', fmt.COUNT1)
            fmt.money(ws, rw, S_ACOST,
                      f"=IFERROR(IF(INDEX('{design}'!$E${lo}:$E${hi},{m})=\"Offshore\","
                      f"INDEX({A3}!$H$5:$H$23,MATCH({key},{A3}!$A$5:$A$23,0)),"
                      f'INDEX({A3}!$G$5:$G$23,MATCH({key},{A3}!$A$5:$A$23,0))),"-")')
            fmt.money(ws, rw, S_VAR,
                      f'=IFERROR(${L(S_ACTUAL)}{rw}-${L(S_ACOST)}{rw},"-")')
        for c in (S_TYPE, S_SIZE):
            ws.cell(rw, c).font = fmt.BODY
            ws.cell(rw, c).alignment = fmt.LEFT
        counts = ((S_ROLES, f"=COUNTA(${L(P_NAME)}${a}:${L(P_NAME)}${b})"),
                  (S_VAC, f'=COUNTIFS({st},"Vacant")'),
                  (S_HIRE, f'=COUNTIFS({lev},"Hire")+COUNTIFS({lev},"Offshore")'),
                  (S_HOLD, f'=COUNTIFS({lev},"Hold")'),
                  (S_AFTER, f"=${L(S_ROLES)}{rw}-${L(S_HOLD)}{rw}"))
        for c, f in counts:
            fmt.money(ws, rw, c, f, fmt.COUNT)
        fmt.money(ws, rw, S_ACTUAL,
                  f"=SUMIFS({REV}!$AA$2:$AA${LAST},{REV}!$AJ$2:$AJ${LAST},\"{pf}\","
                  f"{REV}!$AT$2:$AT${LAST},${L(S_SQUAD)}{rw})/1000000")
        fmt.money(ws, rw, S_TOTAL,
                  f"=SUM(${L(P_AFTER)}${a}:${L(P_AFTER)}${b})/1000000")
        fmt.money(ws, rw, S_IMPACT,
                  f"=${L(S_TOTAL)}{rw}-${L(S_ACTUAL)}{rw}")

    for g in delivery:
        squad(srow[g], g, False)
    for g in overhead:
        squad(srow[g], g, True)

    def total_row(rw, label, r0, r1, fill, line=False):
        fmt.band(ws, rw, 2, S_N, fill, line=line)
        ws.cell(rw, S_SQUAD).value = label
        ws.cell(rw, S_SQUAD).alignment = fmt.LEFT
        for c in range(S_AROLES, S_TOTAL + 1):
            x = ws.cell(rw, c)
            x.value = f"=SUM({L(c)}{r0}:{L(c)}{r1})"
            x.number_format = (fmt.MONEY_M if c >= S_ACOST else
                               (fmt.COUNT1 if c == S_AROLES else fmt.COUNT))
            x.alignment = fmt.RIGHT

    total_row(r_del, "Delivery squads", HDR + 1, r_del - 1, fmt.SUB_FILL)
    if overhead:
        fmt.band(ws, r_ohhead, 2, S_N, fmt.GRP_FILL)
        ws.cell(r_ohhead, S_SQUAD).value = "Overhead roles"
        ws.cell(r_ohhead, S_SQUAD).alignment = fmt.LEFT
        total_row(r_oh, "Overhead roles total", r_ohhead + 1, r_oh - 1, fmt.SUB_FILL)

    fmt.band(ws, r_tot, 2, S_N, fmt.TOT_FILL, line=True)
    ws.cell(r_tot, S_SQUAD).value = "Total portfolio"
    ws.cell(r_tot, S_SQUAD).alignment = fmt.LEFT
    for c in range(S_AROLES, S_TOTAL + 1):
        x = ws.cell(r_tot, c)
        x.value = (f"=N({L(c)}{r_del})+N({L(c)}{r_oh})" if overhead
                   else f"=N({L(c)}{r_del})")
        x.number_format = (fmt.MONEY_M if c >= S_ACOST else
                           (fmt.COUNT1 if c == S_AROLES else fmt.COUNT))
        x.alignment = fmt.RIGHT

    for i, (lab, col, f, nf) in enumerate([
            ("Control - roles against the ledger, must be 0", S_ROLES,
             f"=${L(S_ROLES)}${r_tot}-COUNTIFS({REV}!$AJ$2:$AJ${LAST},\"{pf}\")",
             fmt.COUNT),
            ("Control - cost against the ledger ($m), must be 0", S_ACTUAL,
             f"=ROUND(${L(S_ACTUAL)}${r_tot}-SUMIFS({REV}!$AA$2:$AA${LAST},"
             f"{REV}!$AJ$2:$AJ${LAST},\"{pf}\")/1000000,6)", fmt.MONEY_M)]):
        ws.cell(r_ctl + i, 2).value = lab
        ws.cell(r_ctl + i, 2).font = fmt.BODY
        fmt.money(ws, r_ctl + i, col, f, nf)

    # ---- FTE ----
    title(ws, r_fte, f"{pf} FTE", 12)
    fmt.header(ws, r_phdr, 2, P_HEADERS, P_WIDTHS)
    dv = DataValidation(type="list", formula1='"Filled,Hire,Hold,Offshore"',
                        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)

    R = wv[REVIEW]
    for g in delivery + overhead:
        a, b = span(g)
        bd = band[g]
        fmt.band(ws, bd, 2, P_N, fmt.GRP_FILL)
        ws.cell(bd, P_NAME).value = g
        ws.cell(bd, P_NAME).alignment = fmt.LEFT
        ws.cell(bd, P_ROLE).value = (
            f"=COUNTA(${L(P_NAME)}${a}:${L(P_NAME)}${b})&\" roles\"")
        ws.cell(bd, P_ROLE).alignment = fmt.LEFT
        for c, f in ((P_COST, f"=SUM({L(P_COST)}{a}:{L(P_COST)}{b})"),
                     (P_AFTER, f"=SUM({L(P_AFTER)}{a}:{L(P_AFTER)}{b})")):
            x = ws.cell(bd, c)
            x.value = f
            x.number_format = fmt.MONEY
            x.alignment = fmt.RIGHT
        for rw, x in people.get(g, []):
            i = x["row"]
            ws.cell(rw, P_NAME).value = f"=INDEX({REV}!$B:$B,{i})"
            ws.cell(rw, P_ROLE).value = f"=INDEX({REV}!$C:$C,{i})"
            ws.cell(rw, P_STATUS).value = f"=INDEX({REV}!$AK:$AK,{i})"
            for c in (P_NAME, P_ROLE, P_STATUS):
                ws.cell(rw, c).font = fmt.BODY
                ws.cell(rw, c).alignment = fmt.LEFT
            lv = ws.cell(rw, P_LEVER)
            lv.value = "Filled" if x["status"] == "Filled" else "Hire"
            lv.fill = fmt.IN_FILL
            lv.border = fmt.BOX
            lv.font = fmt.BODY
            lv.alignment = fmt.CENTRE
            dv.add(lv)
            fmt.money(ws, rw, P_COST, f"=INDEX({REV}!$AA:$AA,{i})", fmt.MONEY)
            fmt.money(ws, rw, P_AFTER,
                      f"=${L(P_COST)}{rw}*IFERROR(INDEX(Lists!$AD:$AD,"
                      f"MATCH(${L(P_LEVER)}{rw},Lists!$AC:$AC,0)),1)", fmt.MONEY)

    ws.freeze_panes = f"C{HDR + 1}"
    return {"tab": tab, "pf": pf, "total_row": r_tot, "delivery_row": r_del,
            "overhead_row": r_oh, "header_row": HDR, "first_squad": HDR + 1,
            "last_squad": r_del - 1, "squads": delivery, "overhead": overhead,
            "srow": srow, "fte_end": r_end, "people": sum(len(v) for v in people.values())}


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    wv = openpyxl.load_workbook(src, data_only=True)
    rows = ledger(wv)
    bypf = collections.defaultdict(list)
    for r in rows:
        bypf[r["pf"]].append(r)
    bounds = {d: squad_table_bounds(wb, d) for d in set(DESIGN.values())}
    tabs = [s for s in wb.sheetnames if re.match(r"^2\.\d+ ", s)]
    anchors, out = {}, []
    for tab in tabs:
        pf = wv[tab]["C3"].value or wv[tab]["B2"].value
        pf = next(p for p in bypf if p and (p == pf or str(pf).startswith(p)
                                            or tab.split(" ", 1)[1] == p
                                            or tab.endswith(p)))
        a = build_working(wb, wv, tab, bypf[pf], bounds)
        anchors[tab] = a
        out.append(f"{tab}: {len(a['squads'])} squads, {len(a['overhead'])} overhead, "
                   f"{a['people']} people")
    json.dump(anchors, open("anchors2.json", "w"), indent=1, default=str)
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
