"""Stage 8: the last four live defects, found by a second reader of the 3.x tabs.

Exec C9   headline #5 "value on the table" read 3.2!I21 (cost AFTER decisions) and so
          printed 86.579 - the same figure as "of which filled" three lines below it. The
          value at stake is the vacant cost, M21 = 28.534.
Exec C59  read 3.4!H11, which the EGI insert turned into EGI's own left-to-fund. H12.
3.4 H6/H7 pulled 1.11's "Variance" column (budget - spend) into a "Left to fund" column
          (spend - budget), so two of six rows carried the opposite sign and the total
          failed its own F - G arithmetic check.
3.4 F     the note at B13 says F is net of the amount funded inside portfolio overheads,
          but F equalled K in all six rows - column L ($3.6m) was deducted nowhere.
Exec C27/C33/C50 compared all-525 actuals against archetype-only design figures.
3.3 row 3 an orphaned pre-rebuild header band two rows above the real one, assigning
          different meanings to the same columns - the exact thing Lists!K was written
          against, so leaving it invites the bug back.
4.0       had no check on 3.4 or 3.1, the two tabs that actually broke.
"""
import openpyxl
from openpyxl.styles import Font
import model

ITAL = Font(italic=True); BOLD = Font(bold=True)
REV = f"'{model.REVIEW}'"; LR = model.LAST_ROW

def run(src, dst):
    wb = openpyxl.load_workbook(src); out = []

    # --- 3.4: sign, and make F genuinely net ---
    s4 = wb["3.4 COE Summary"]
    for r in (6, 7):
        s4[f"H{r}"] = f"=$F{r}-$G{r}"
    for r in range(6, 12):
        s4[f"F{r}"] = f"=$K{r}-$L{r}"
        s4[f"H{r}"] = f"=$F{r}-$G{r}"
    s4["B13"] = ("Planned spend (F) is net: gross people cost (K) less the amount funded "
                 "inside portfolio overheads (L). Left to fund (H) = F - G on every row.")
    s4["B13"].font = ITAL
    # I and J came from the 1.11/1.12 role lists while K was rebuilt onto the ledger, so
    # AU + NZ fell $0.473m short of gross cost. Both now read the ledger by country.
    AJ, AA, CT = (f"{REV}!$AJ$2:$AJ${LR}", f"{REV}!$AA$2:$AA${LR}", f"{REV}!$M$2:$M${LR}")
    G = f"{REV}!$G$2:$G${LR}"
    spec = [(6, "COE BP&T", "Transformation", True), (7, "COE BP&T", "Transformation", False),
            (8, "COE SA&D", "Group Data", True), (9, "COE SA&D", "Group Data", False),
            (10, "COE Cyber", None, None), (11, "EGI", None, None)]
    for r, pf, dept, inv in spec:
        for col, ctry in (("I", "Australia"), ("J", "NZ")):
            if dept is None:
                f = f'SUMIFS({AA},{AJ},"{pf}",{CT},"{ctry}")'
            elif inv:
                f = (f'SUMIFS({AA},{AJ},"{pf}",{CT},"{ctry}")'
                     f'-SUMIFS({AA},{AJ},"{pf}",{G},"{dept}",{CT},"{ctry}")')
            else:
                f = f'SUMIFS({AA},{AJ},"{pf}",{G},"{dept}",{CT},"{ctry}")'
            s4[f"{col}{r}"] = f"=({f})/1000000"
    # One COE role sits outside both countries - Neil Reilly, TDD BP - T&S, Singapore -
    # so AU + NZ cannot equal gross cost without naming the residual.
    s4["B16"] = "Cost outside AU and NZ ($m) - TDD BP - T&S is based in Singapore"
    s4["B16"].font = ITAL
    s4["F16"] = (f'=SUMIFS({AA},{AJ},"COE*",{CT},"<>Australia",{CT},"<>NZ")/1000000'
                 f'+SUMIFS({AA},{AJ},"EGI",{CT},"<>Australia",{CT},"<>NZ")/1000000')
    s4["B17"] = "Control - I + J + row 16 must equal K ($m)"
    s4["B17"].font = ITAL
    s4["F17"] = "=ROUND($I$12+$J$12+$F$16-$K$12,6)"
    out.append("3.4: H sign corrected on rows 6-7, F now net of L, I+J vs K control added")

    # --- Exec Summary: the four mispointed lines ---
    ex = wb["Exec Summary"]
    ex["C9"] = "='3.2 Total Cost'!$M$21"
    ex["B9"] = ("5. The decision: hire, hold or offshore every vacancy - value on the "
                "table ($m)")
    ex["C59"] = "='3.4 COE Summary'!$H$12"
    ex["B59"] = "COEs - left to fund after budgets, see 3.4 ($m)"
    # like-for-like: compare the archetype-based portfolios to the archetype design
    ex["C27"] = "='3.2 Total Cost'!$F$16-('3.2 Total Cost'!$G$16+Lists!$AJ$8)"
    ex["C33"] = "='3.2 Total Cost'!$F$16-('3.2 Total Cost'!$G$16+Lists!$AJ$8)"
    ex["C50"] = "='3.2 Total Cost'!$L$16-'3.2 Total Cost'!$G$16"
    ex["B27"] = ("Actual over the designed cost by ($m) - archetype-based portfolios, "
                 "like for like")
    ex["B33"] = "Difference vs the designed cost ($m) - archetype-based portfolios"
    ex["B50"] = ("Filled roles over/(under) the squad archetype cost ($m) - archetype-based "
                 "portfolios")
    out.append("Exec C9/C59 repointed; C27/C33/C50 now compare like for like")

    # --- 3.3: remove the orphaned header band ---
    h = wb["3.3 FTE View"]
    n = 0
    for c in range(2, 14):
        if h.cell(3, c).value is not None:
            h.cell(3, c).value = None; n += 1
    if n:
        out.append(f"3.3 row 3: {n} orphaned pre-rebuild headers cleared")

    # --- 4.0: cover the tabs that actually broke ---
    q = wb["4.0 Data QA"]
    extra = [
        ("3.4 roles vs the COE+EGI ledger", "='3.4 COE Summary'!$C$14"),
        ("3.4 cost vs the COE+EGI ledger ($m)", "='3.4 COE Summary'!$F$15"),
        ("3.4 AU + NZ + other vs gross cost ($m)", "='3.4 COE Summary'!$F$17"),
        ("3.1 total cost vs 3.2 ($m)",
         "=ROUND('3.1 Group Summary'!$D$20-'3.2 Total Cost'!$F$21,6)"),
        ("3.2 filled + vacant vs total cost ($m)",
         "=ROUND('3.2 Total Cost'!$L$21+'3.2 Total Cost'!$M$21-'3.2 Total Cost'!$F$21,6)"),
        ("3.2 overhead + squad - GMs vs ledger ($m)", "='3.2 Total Cost'!$F$38"),
    ]
    base = 15
    for i, (lab, f) in enumerate(extra):
        r = base + i
        q.cell(r, 2).value = lab
        q.cell(r, 3).value = f
        q.cell(r, 4).value = 0
        q.cell(r, 5).value = f"=ROUND($C{r}-$D{r},6)"
    last = base + len(extra)
    q.cell(last, 2).value = "Checks failing"
    q.cell(last, 2).font = BOLD
    q.cell(last, 5).value = f'=COUNTIF(E6:E{last-1},"<>0")'
    q["B4"] = ("Data QA. Column C reads the model; column D is the expected value, which "
               "for the ledger rows is the agreed ground truth and elsewhere is zero.")
    out.append(f"4.0: {len(extra)} checks added covering 3.4, 3.1 and the overhead block")

    wb.save(dst); return out

if __name__ == "__main__":
    for x in run("v3.xlsx", "v4.xlsx"): print("  ", x)
