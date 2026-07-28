"""All fourteen working tabs, layout 2A, out of one function.

Every 2.x tab is written by build() and nothing else touches them, so they cannot drift
apart the way the old ones did (nine different column-width profiles across fourteen tabs,
two title styles, because three different scripts had written them over time).

Layout, top to bottom:

    title
    Squads priced by an archetype     bar + header + rows + subtotal
    Directly funded                   bar + header + rows + subtotal
    Overhead roles                    bar + header + rows + subtotal
    Total portfolio
    two control lines, both must read 0
    <Portfolio> FTE        bar
      header
      squad band carrying that squad's two totals
        one row per person, the lever the only yellow cell

All three blocks appear on all fourteen tabs. Where a portfolio has nothing in a block the
block carries one row reading "None" rather than being dropped, so the fourteen tabs are
structurally identical and the absence is stated rather than left to be inferred - the COEs
have no overhead by decision, and a reader should be able to see that on the tab.
"""
import collections
import copy
import json
import re

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation

import opts
from build_2xfix import DESIGN, squad_table_bounds

# a sentence explaining a line above it, not a figure and not a heading
NOTE = Font(name=opts.FN, size=11, italic=True)

REVIEW = "REVIEW - Complete Role Mapping"
REV = f"'{REVIEW}'"
A3 = "'0.3 Squad Archetypes'"
LAST = None                             # set from the ledger by set_last()

# ---- squad summary columns, B onwards ----
# "Roles after decisions" replaced "Vacancies remaining", which counted vacancies set to
# Hold and so read as though a cancelled vacancy were still outstanding. Roles after
# decisions is roles less anything put On hold, which is what D20 promised: pull Hold and
# the headcount moves, not just the cost.
# FTE sits beside the headcount because the owner reconciles in FTE, not heads: seven roles
# in the ledger are part-time, so Customer is 83 people and 82.4 FTE and his own figures are
# the second one. Both are stated rather than one replacing the other - "525 roles" is used
# throughout the file and in every control.
S = dict(squad=2, type=3, size=4, aroles=5, roles=6, fte=7, filled=8, vacant=9, hire=10,
         offshore=11, hold=12, rafter=13, acost=14, actual=15, var=16, after=17,
         newvar=18)
# The owner's own column names, taken off his markup of 2.8.
S_HDR = ["Squad", "Archetype Type", "Squad Size", "Archetype roles", "Total roles", "FTE",
         "Filled", "Vacant", "To hire", "To offshore", "On hold",
         "Total roles after decisions", "Archetype cost ($m)", "Actual cost ($m)",
         "Variance to archetype ($m)", "Squad cost after decisions ($m)",
         "New variance ($m)"]
# the owner renamed these on his own copy - "Squads" and "Total" - so the build uses his
# words rather than overwriting them on the next run
SECTION = {"arch": "Squads",
           "direct": "Directly funded programmes and platforms",
           "none": "No archetype in 1.x tabs",
           "oh": "Overhead roles"}
# One pattern for every subtotal on every tab: "<section, short> total". The Squads block
# used to print a bare "Total" while its three siblings printed "<section> total", so the
# one row a reader scans for on fourteen tabs was labelled two different ways.
SUBTOTAL = {"arch": "Squads total",
            "direct": "Directly funded total",
            "none": "No archetype total",
            "oh": "Overhead roles total"}
# the label goes in the Squad column, not the Archetype Type column - a sentence in a data
# column is still a sentence in a data column
NONE_TEXT = {"arch": "None - no squad here is priced by an archetype",
             "direct": "None - no directly funded programme here",
             "none": "None - every group here has a figure to compare",
             "oh": "None - the COEs carry no overhead"}
KINDS = ("arch", "direct", "none", "oh")

# The archetype prices overhead. 0.2 Data Config allows a fraction of a Head of Technology,
# a Business Partner, a Domain Architect and the Leadership layer per portfolio - 0.7975 -
# and a fraction of a Delivery Manager and a Technology Manager per platform - 0.165. The
# 1.x tab has been adding both into its Total Cost all along, which is why 1.8 reads 9.03
# where the squads alone are 7.90. The 2.x tabs were showing the 7.90 and none of the rest,
# so every portfolio's archetype was short by its own overhead allowance.
#
# The two sheets name the same lines differently - "Head of Tech" against "Head of
# Technology" - so the map is written out rather than matched on text.
OH_RATE = {"Head of Technology": ("$N$6", "$M$6", False),
           "Business Partner": ("$N$7", "$M$7", False),
           "Domain Architect": ("$N$8", "$M$8", False),
           "Delivery Manager": ("$N$14", "$M$14", True),
           "Technology Manager": ("$N$15", "$M$15", True)}
CFG = "'0.2 Data Config'"
ELSEWHERE = ("Allowed for in the archetype, sitting outside this portfolio")
ELSEWHERE_WHY = "Business Partners and Domain Architects sit in the COEs"
# The Squad column carries the longest text on the tab: a section label, the "allowed for
# elsewhere" line, and the two control labels under the table. The controls now state their
# figure in the column beside them, so nothing spills any more and the column has to be wide
# enough to hold "Control - cost against the ledger ($m), must be 0" on its own. One width
# for all fourteen tabs - the point of the family is that they are the same shape.
S_W = [50, 26, 11, 11, 9, 8, 8, 8, 8, 11, 8, 14, 13, 12, 14, 17, 14]

# ---- FTE columns ----
P = dict(name=2, role=3, status=4, lever=5, cost=6, after=7)
P_HDR = ["Name", "Role", "Status", "Vacancy lever", "Role cost ($)",
         "Cost after decision ($)"]
P_W = [30, 44, 11, 15, 15, 19]

OH_ORDER = ["Head of Technology", "Business Partner", "Domain Architect",
            "Delivery Manager", "Technology Manager", "Leadership",
            "Leadership - squad not stated"]


def set_last(wb):
    global LAST
    LAST = opts.ledger_last(wb)
    return LAST


def ledger(wv):
    R = wv[REVIEW]
    out = []
    for i in range(2, LAST + 1):
        if not str(R.cell(i, 2).value or "").strip():
            continue
        pf = R.cell(i, 36).value
        is_coe = str(pf or "").startswith("COE") or pf == "EGI"
        out.append({"row": i, "pf": pf, "grp": R.cell(i, 46).value,
                    "status": R.cell(i, 37).value,
                    "oh": (not is_coe) and R.cell(i, 44).value != "Squad"})
    return out


# the three overhead lines the archetype draws inside a portfolio. The other three -
# Business Partner, Domain Architect and the Leadership layer - are allowed for per
# portfolio but their people sit in the COEs and above the ledger.
IN_PF = ["Head of Technology", "Delivery Manager", "Technology Manager"]


def groups(rows, design=None):
    """Delivery squads, and the overhead lines this portfolio is measured on.

    The three portfolio-drawn overhead lines are always listed, whether or not anyone is
    in them. The archetype allows a Delivery Manager per platform whether the seat is
    filled or not, and listing only the lines with people in them handed that allowance to
    the "sitting outside this portfolio" line - Finance has no Delivery Manager, so its
    allowance quietly left the portfolio and the group allowance came out 0.31 light.
    """
    seen = {}
    for r in rows:
        seen.setdefault(r["grp"], r["oh"])
    delivery = sorted(g for g, oh in seen.items() if not oh)
    want = set(IN_PF) if design else set()
    overhead = [g for g in OH_ORDER if seen.get(g) or g in want]
    overhead += sorted(g for g, oh in seen.items() if oh and g not in OH_ORDER)
    return delivery, overhead


def has_archetype(wb, wv, design, squad):
    """True when the squad appears in the design tab's squad table AND that row carries a
    type and size the archetype library prices. Strategic programmes deliberately do not
    (register item 21), and a squad missing from the design cannot either.

    This partition matters: the old build put the archetype for SOME squads against the
    actual for ALL of them in one subtotal, so the group read 11.49m over design when the
    squads that actually have an archetype are 0.56m under it.
    """
    if design is None:
        return False
    lo, hi = squad_table_bounds(wb, design)
    ws = wv[design]
    keys = {str(wv["0.3 Squad Archetypes"].cell(r, 1).value or "").strip()
            for r in range(5, 24)}
    for r in range(lo, hi + 1):
        if str(ws.cell(r, 2).value or "").strip() != squad:
            continue
        ty = str(ws.cell(r, 3).value or "").strip()
        sz = str(ws.cell(r, 4).value or "").strip()
        return f"{ty}|{sz}" in keys
    return False


def in_design(wb, wv, design, squad):
    """True when the design tab's squad table has a row for this squad."""
    if design is None:
        return False
    lo, hi = squad_table_bounds(wb, design)
    ws = wv[design]
    return any(str(ws.cell(r, 2).value or "").strip() == squad
               for r in range(lo, hi + 1))


def wipe(ws):
    blank = copy.copy(openpyxl.cell.cell.Cell(ws)._style)
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    for row in ws.iter_rows():
        for c in row:
            c.value = None
            c._style = copy.copy(blank)
    for k in list(ws.column_dimensions):
        del ws.column_dimensions[k]
    for k in list(ws.row_dimensions):
        del ws.row_dimensions[k]
    ws.conditional_formatting._cf_rules.clear()
    ws.data_validations.dataValidation = []
    ws.freeze_panes = None
    ws.sheet_view.showGridLines = False


def oh_rows(wv, design):
    """The two archetype overhead rows on a design tab, found by label.

    Portfolio Overhead is the same 0.7975 on all ten; Platform Overheads is 0.165 for each
    platform the portfolio has, which is how the platform count is read back out.
    """
    if design is None:
        return None, None
    ws = wv[design]
    got = {}
    for r in range(1, 20):
        b = str(ws.cell(r, 2).value or "").strip()
        if b.startswith("Portfolio Overhead"):
            got["pf"] = r
        elif b.startswith("Platform Overheads"):
            got["plat"] = r
    return got.get("pf"), got.get("plat")


def build(wb, wv, tab, rows, bounds):
    ws = wb[tab]
    design = DESIGN.get(tab)
    pf = rows[0]["pf"]
    oh_pf, oh_plat = oh_rows(wv, design)
    # the archetype's own overhead for this portfolio, straight off the design tab, and the
    # number of platforms behind the per-platform half of it
    ALLOW = (f"(N('{design}'!$F${oh_pf})+N('{design}'!$F${oh_plat}))"
             if oh_pf and oh_plat else '0')
    PLAT = f"N('{design}'!$F${oh_plat})/{CFG}!$N$16" if oh_plat else "0"
    delivery, overhead = groups(rows, design)
    lo, hi = bounds[design] if design else (0, 0)
    # split the delivery squads by whether the archetype library actually prices them
    arch = [g for g in delivery if has_archetype(wb, wv, design, g)]
    rest = [g for g in delivery if g not in arch]
    # a directly funded programme has an amount typed against it on the design tab, or is
    # EGI (priced at actual on the owner's ruling), or sits on a COE tab where the planned
    # spend is the comparison. Anything else - the two Leadership groups - has no figure to
    # compare at all, and mixing it in made the subtotal charge 1.30 of cost against nothing.
    direct = [g for g in rest
              if design is None or g.startswith("EGI") or in_design(wb, wv, design, g)]
    nofig = [g for g in rest if g not in direct]
    wipe(ws)
    ws.column_dimensions["A"].width = 2

    # ---- plan every row position first ----
    # Every block gets at least one row. An empty block used to collapse to a subtotal
    # whose SUM range ran backwards from the header - SUM(M7:M6), which Excel reads as
    # SUM(M6:M7) and so includes the subtotal cell itself. That was a circular reference
    # on all four COE tabs.
    HDR = 6
    r = HDR + 1
    # A section with nothing in it is not on the tab at all. It used to print a label, a
    # filler row reading "None - every group here has a figure to compare" and a subtotal of
    # dashes: three rows saying nothing, on most tabs twice over.
    #
    # A section with one row does not get a subtotal either. "Directly funded programmes and
    # platforms total, 10" directly under "EGI Customer, 10" is the same number twice, and
    # the tab carried five total rows where it needed three - which is what buried the one
    # total that matters.
    #
    # One rule, all fourteen tabs: a section gets a subtotal row when it has two or more
    # data rows AND the tab has two or more sections carrying data. The second half is what
    # kills the duplicate on the four COE tabs, where a single section's subtotal was
    # followed immediately by a "Total portfolio" grand total holding the identical figure -
    # the same number twice, one row apart, with two different names.
    data = {"arch": arch, "direct": direct, "none": nofig, "oh": overhead}
    live = [k for k in KINDS if data[k]]
    many = len(live) > 1
    srow, label, sub, anchor, srange = {}, {}, {}, {}, {}
    oh_else = oh_note = None
    for kind in live:
        names = data[kind]
        label[kind] = r
        r += 1
        first = r
        for g in names:
            srow[g] = r
            r += 1
        srange[kind] = (first, r - 1)
        if len(names) > 1 and many:
            sub[kind] = r
            r += 1
        # a one-row section anchors on its own row; a multi-row section anchors on its
        # subtotal, and where the rule gives it none there is no single cell holding that
        # section's total, so there is no anchor to hand out either. Every total that used
        # to be built off these anchors is now built off the section's own data rows.
        anchor[kind] = sub.get(kind) if len(names) > 1 else srow[names[0]]
        # The archetype allows for six overhead lines per portfolio; three of them have
        # nobody in a portfolio at all - the Business Partners and Domain Architects sit in
        # the COEs and the eight GMs sit above the ledger. Their allowance is part of what
        # the archetype prices this portfolio at, so it gets a line, and the line says where
        # those people actually are.
        #
        # It sits BELOW the overhead subtotal, not inside it. Inside, the subtotal compared
        # a 1.13 allowance against 1.12 of actual and read (0.01) - overhead bang on plan -
        # when the portfolio's own three overhead lines are 0.65 over their own allowance
        # and the difference was being covered by allowance for people who are not there.
        if kind == "oh" and design is not None:
            oh_else = r
            r += 1
            # the reason those people are not here is a note, not a data column. It used to
            # sit in C of the same row, where it truncated the label in B beside it - B
            # holds a 59-character sentence and C holds a 54-character one - so a reader saw
            # "Allowed for in the archetype, si" against half a note. The note now has its
            # own row under the line it explains, in B, italic, with nothing beside it.
            oh_note = r
            r += 1
    r_tot = r
    r += 2
    r_ctl = r
    r += 3
    r_ftebar = r
    r_ftehdr = r + 1
    r = r_ftehdr + 1
    band, people = {}, {}
    # a line the archetype allows for but nobody fills has no detail block - a heading with
    # nothing under it reads as a table that failed to load
    staffed = [g for g in delivery + overhead if any(y["grp"] == g for y in rows)]
    for g in staffed:
        band[g] = r
        r += 1
        for x in [y for y in rows if y["grp"] == g]:
            people.setdefault(g, []).append((r, x))
            r += 1

    # ---- header ----
    ws.cell(2, 2).value = f"{pf} - working copy"
    ws.cell(2, 2).font = opts.TITLE
    ws.cell(3, 2).value = "Portfolio"
    ws.cell(3, 2).font = opts.BOLD
    ws.cell(3, 3).value = pf
    ws.cell(3, 3).font = opts.BODY
    opts.bar(ws, 4, 2, len(S_HDR), f"{pf} - archetype cost against actual cost")
    # One table, one header row, the overhead lines inside it. They used to sit in a second
    # table below with a header of their own; the owner's instruction is that the overheads
    # belong up top with everything else.
    W = list(S_W)
    opts.head(ws, HDR, 2, list(S_HDR), W)
    ws.row_dimensions[5].height = 6

    def span(g):
        p = people.get(g, [])
        return (p[0][0], p[-1][0]) if p else (0, 0)

    def squad(rw, g, kind):
        a, b = span(g)
        lev = f"${L(P['lever'])}${a}:${L(P['lever'])}${b}"
        st = f"${L(P['status'])}${a}:${L(P['status'])}${b}"
        ws.cell(rw, S["squad"]).value = g
        ws.cell(rw, S["squad"]).font = opts.BODY
        ws.cell(rw, S["squad"]).alignment = opts.LFT
        if kind == "oh" or design is None:
            # the bar above each block already says what the block is, so the type
            # column stays a dash rather than repeating a sentence on every row
            ws.cell(rw, S["type"]).value = '="-"'
            ws.cell(rw, S["type"]).alignment = opts.RGT
            for k in ("size",):
                x = ws.cell(rw, S[k])
                x.value = '="-"'
                x.alignment = opts.RGT
            if kind == "oh":
                # the allowance this line draws in this portfolio, and the FTE behind it.
                # A per-platform line draws once for every platform on the design tab, and
                # the platform count comes off the tab itself rather than being counted
                # here - Platform Overheads divided by the per-platform rate.
                rate, alloc, per_platform = OH_RATE.get(g, (None, None, False))
                if rate is None or design is None:
                    for k in ("aroles", "acost", "var", "newvar"):
                        ws.cell(rw, S[k]).value = '="-"'
                        ws.cell(rw, S[k]).alignment = opts.RGT
                else:
                    n = f"*({PLAT})" if per_platform else ""
                    _m(ws, rw, S["aroles"], f"={CFG}!{alloc}{n}", opts.C1)
                    _m(ws, rw, S["acost"], f"={CFG}!{rate}{n}")
                    _m(ws, rw, S["var"],
                       f'=ROUND(${L(S["actual"])}{rw}-${L(S["acost"])}{rw},6)')
                    _m(ws, rw, S["newvar"],
                       f'=ROUND(${L(S["after"])}{rw}-${L(S["acost"])}{rw},6)')
            else:
                # A COE squad has no archetype and nothing else prices it either. The
                # column used to carry "=actual", which made every COE row show a variance
                # of exactly nil - not a comparison, the same number written twice. A
                # figure that is the actual restated tells the reader nothing and inflates
                # both sides of every total above it, so it is a dash.
                for k in ("aroles", "acost", "var", "newvar"):
                    x = ws.cell(rw, S[k])
                    x.value = '="-"'
                    x.alignment = opts.RGT
        else:
            key = f'${L(S["type"])}{rw}&"|"&${L(S["size"])}{rw}'
            m = f"MATCH(${L(S['squad'])}{rw},'{design}'!$B${lo}:$B${hi},0)"
            ws.cell(rw, S["type"]).value = (
                f"=IFERROR(INDEX('{design}'!$C${lo}:$C${hi},{m}),\"Not on the 1.x tab\")")
            # INDEX over an empty cell returns 0, not an error, so a strategic programme
            # with no size printed 0 in the Size column
            ws.cell(rw, S["size"]).value = (
                f"=IFERROR(IF(INDEX('{design}'!$D${lo}:$D${hi},{m})=\"\",\"-\","
                f"INDEX('{design}'!$D${lo}:$D${hi},{m})),\"-\")")
            if kind == "arch":
                _m(ws, rw, S["aroles"],
                   f"=IFERROR(INDEX({A3}!$F$5:$F$23,MATCH({key},{A3}!$A$5:$A$23,0)),"
                   f'"-")', opts.C1)
                # the design tab's H is the squad's archetype cost - the same library
                # lookup on nine tabs (now with the owner's Hybrid state priced as the
                # onshore/offshore midpoint), and his own typed bottom-up figures on 1.2.
                # Reading H rather than re-deriving from the library keeps the working
                # tab equal to the tab he edits, whichever of the two H happens to be.
                hh = f"INDEX('{design}'!$H${lo}:$H${hi},{m})"
                _m(ws, rw, S["acost"],
                   f'=IFERROR(IF(ISNUMBER({hh}),{hh},"-"),"-")')
            else:
                # Directly funded. No archetype prices a strategic programme - the owner's
                # instruction - so the figure to compare against is the amount he typed on
                # its own platform block, column H of the 1.x tab.
                #
                # EGI used to be forced to "=actual" here. That did two things wrong. It
                # printed six rows whose funded column was the actual written twice, with a
                # variance of nil, which is what the owner picked up. And on EGI Retail it
                # threw away a real figure: 1.52 is typed on 1.1 and the column showed the
                # actual 1.22 instead, hiding a 0.30 underspend he had set up himself.
                #
                # The rule is now the same for every directly funded row. Where a funded
                # figure is set, compare against it. Where it is blank or zero - five of the
                # six EGI rows - there is nothing to compare against and the column says so,
                # rather than manufacturing a comparison out of the actual.
                _m(ws, rw, S["aroles"], '="-"', opts.C1)
                idx = f"INDEX('{design}'!$H${lo}:$H${hi},{m})"
                idx_i = f"INDEX('{design}'!$I${lo}:$I${hi},{m})"
                # the funded amount lives in H (Total Squad Cost); where the owner typed
                # it into I (TDD Cost) instead - EGI Customer's 2.21 - the I figure is
                # the funded amount and the comparison uses it
                _m(ws, rw, S["acost"],
                   f'=IFERROR(IF(N({idx})>0,{idx},IF(N({idx_i})>0,{idx_i},"-")),"-")')
            # rounded, or a squad priced exactly at its archetype shows (0.00) on a
            # residual of a few billionths
            _m(ws, rw, S["var"],
               f'=IFERROR(ROUND(${L(S["actual"])}{rw}-${L(S["acost"])}{rw},6),"-")')
            _m(ws, rw, S["newvar"],
               f'=IFERROR(ROUND(${L(S["after"])}{rw}-${L(S["acost"])}{rw},6),"-")')
        for k in ("type", "size"):
            ws.cell(rw, S[k]).font = opts.BODY
            ws.cell(rw, S[k]).alignment = opts.LFT
        # To hire only applies to a vacancy. To offshore and On hold can apply to a
        # filled role too, because offshoring a filled resource is a decision the tool
        # exists to price. Vacancies remaining therefore has to look at vacant rows
        # only: counting an offshored filled person against it drove the count negative.
        # An overhead line the archetype allows for but nobody fills has no detail block to
        # count, so the counts read zero rather than counting a range that starts at row 0 -
        # COUNTIFS($D$0:$D$0,"Filled") is #NAME?, and it took Exec down with it.
        #
        # The FTE and the actual cost do not count that block. They join the ledger on
        # portfolio and squad name, so they hold on a line with nobody in it exactly as
        # they do on a line with forty. Those two were being written as a literal =0 on
        # three rows - 2.3's Delivery Manager and 2.6's Head of Technology and Delivery
        # Manager - which is the right figure today and a typed constant tomorrow: put a
        # Delivery Manager into Finance in REVIEW and the line goes on reading nothing, and
        # the tab's controls report the miss as a portfolio-level break with no row to
        # point at. The family SUMIFS goes on every row, filled or not, so the line that is
        # wrong is the line that says so.
        FTE = (f"=ROUND(SUMIFS({REV}!$O$2:$O${LAST},{REV}!$AJ$2:$AJ${LAST},$C$3,"
               f"{REV}!$AT$2:$AT${LAST},${L(S['squad'])}{rw}),2)")
        ACTUAL = (f"=SUMIFS({REV}!$AA$2:$AA${LAST},{REV}!$AJ$2:$AJ${LAST},$C$3,"
                  f"{REV}!$AT$2:$AT${LAST},${L(S['squad'])}{rw})/1000000")
        if not people.get(g):
            for k in ("roles", "filled", "vacant", "hire", "offshore", "hold", "rafter"):
                _m(ws, rw, S[k], "=0", opts.CT)
            _m(ws, rw, S["fte"], FTE, opts.C1)
            _m(ws, rw, S["actual"], ACTUAL)
            # cost after decisions sums the levers in the detail block, and there is none
            _m(ws, rw, S["after"], "=0")
        else:
            for k, f in (("roles", f"=COUNTA(${L(P['name'])}${a}:${L(P['name'])}${b})"),
                         ("filled", f'=COUNTIFS({st},"Filled")'),
                         ("vacant", f'=COUNTIFS({st},"Vacant")'),
                         ("hire", f'=COUNTIFS({st},"Vacant",{lev},"Hire")'),
                         ("offshore", f'=COUNTIFS({lev},"Offshore")'),
                         ("hold", f'=COUNTIFS({lev},"Hold")'),
                         ("rafter", f'=${L(S["roles"])}{rw}-${L(S["hold"])}{rw}')):
                _m(ws, rw, S[k], f, opts.CT)
            _m(ws, rw, S["fte"], FTE, opts.C1)
            _m(ws, rw, S["actual"], ACTUAL)
            _m(ws, rw, S["after"],
               f"=SUM(${L(P['after'])}${a}:${L(P['after'])}${b})/1000000")

    def none_row(rw, why):
        """A block with nothing in it still gets a row, so the tab states the absence."""
        ws.cell(rw, S["squad"]).value = "None"
        ws.cell(rw, S["squad"]).font = opts.BODY
        ws.cell(rw, S["squad"]).alignment = opts.LFT
        ws.cell(rw, S["type"]).value = why
        ws.cell(rw, S["type"]).font = opts.BODY
        ws.cell(rw, S["type"]).alignment = opts.LFT
        for k in ("size", "aroles", "roles", "fte", "filled", "vacant", "hire", "offshore",
                  "hold", "rafter", "acost", "actual", "var", "after", "newvar"):
            x = ws.cell(rw, S[k])
            x.value = '="-"'
            x.font, x.alignment = opts.BODY, opts.RGT

    for g in arch:
        squad(srow[g], g, "arch")
    for g in direct:
        squad(srow[g], g, "direct")
    for g in nofig:
        squad(srow[g], g, "direct")
    for g in overhead:
        squad(srow[g], g, "oh")

    def total(rw, label, r0, r1, bg, line=False):
        opts.row(ws, rw, 2, [label] + [None] * (len(S_HDR) - 1),
                 [None] * len(S_HDR), bg=bg, bold=True, top=line)
        ws.cell(rw, S["squad"]).alignment = opts.LFT
        for k in ("aroles", "roles", "fte", "filled", "vacant", "hire", "offshore", "hold",
                  "rafter", "acost", "actual", "var", "after", "newvar"):
            c = S[k]
            x = ws.cell(rw, c)
            # a variance on a total is actual less archetype. Summing the row variances
            # drops every row whose archetype column holds "-", which understated the
            # directly funded subtotal on six of the fourteen tabs.
            # and where the block has no archetype at all - the overhead lines - the
            # variance is a dash, not the whole cost measured against zero
            full = f"COUNT({L(S['acost'])}{r0}:{L(S['acost'])}{r1})={r1 - r0 + 1}"
            if k == "acost":
                # SUM over a block of dashes is 0, and the overhead subtotal printed 0.00
                # as though an archetype had priced 43 people at nothing. A dash where
                # nothing is priced - but where some rows are, the sum of those rows. The
                # stricter rule dropped 2.92 of priced programmes off Ampol Retail's total
                # because one row beside them had no funded figure.
                x.value = (f'=IF(COUNT({L(c)}{r0}:{L(c)}{r1})=0,"-",'
                           f'SUM({L(c)}{r0}:{L(c)}{r1}))')
            elif k == "var":
                # the variance still needs every row priced, or it measures an archetype
                # covering some of the block against an actual covering all of it
                x.value = (f'=IF({full},'
                           f"ROUND(${L(S['actual'])}{rw}-${L(S['acost'])}{rw},6),\"-\")")
            elif k == "newvar":
                x.value = (f'=IF({full},'
                           f"ROUND(${L(S['after'])}{rw}-${L(S['acost'])}{rw},6),\"-\")")
            else:
                x.value = f"=SUM({L(c)}{r0}:{L(c)}{r1})"
            x.number_format = (opts.M2 if c >= S["acost"] else
                               (opts.C1 if k == "aroles" else opts.CT))
            x.alignment = opts.RGT

    if oh_else:
        # what the archetype allows this portfolio for overhead, less what the lines above
        # have already drawn. Written as the remainder rather than as a list, so it stays
        # right if a portfolio ever does carry a Business Partner of its own.
        drawn = "+".join(f"N(${L(S['acost'])}{srow[g]})" for g in overhead) or "0"
        nfte = "+".join(f"N(${L(S['aroles'])}{srow[g]})" for g in overhead) or "0"
        ws.cell(oh_else, S["squad"]).value = ELSEWHERE
        ws.cell(oh_else, S["squad"]).font = opts.BODY
        ws.cell(oh_else, S["squad"]).alignment = opts.LFT
        # C stays empty on this row, so the label in B is read in full
        ws.cell(oh_note, S["squad"]).value = ELSEWHERE_WHY
        ws.cell(oh_note, S["squad"]).font = NOTE
        ws.cell(oh_note, S["squad"]).alignment = opts.LFT
        _m(ws, oh_else, S["aroles"],
           f"=ROUND(SUM({CFG}!$M$6:$M$9)+SUM({CFG}!$M$14:$M$15)*({PLAT})-({nfte}),6)",
           opts.C1)
        _m(ws, oh_else, S["acost"], f"=ROUND({ALLOW}-({drawn}),6)")
        for k in ("size", "roles", "fte", "filled", "vacant", "hire", "offshore", "hold",
                  "rafter", "actual", "var", "after", "newvar"):
            x = ws.cell(oh_else, S[k])
            x.value = '="-"'
            x.font, x.alignment, x.border = opts.BODY, opts.RGT, opts.BOX

    for kind in live:
        opts.row(ws, label[kind], 2, [SECTION[kind]] + [None] * (len(S_HDR) - 1),
                 [None] * len(S_HDR), bg=opts.PALE, bold=True)
        ws.cell(label[kind], 2).alignment = opts.LFT
        if kind in sub:
            total(sub[kind], SUBTOTAL[kind], srange[kind][0], srange[kind][1], opts.GREY)

    opts.row(ws, r_tot, 2, ["Total portfolio"] + [None] * (len(S_HDR) - 1),
             [None] * len(S_HDR), bg=opts.MID, bold=True, top=True)
    ws.cell(r_tot, S["squad"]).alignment = opts.LFT
    for k in ("aroles", "roles", "fte", "filled", "vacant", "hire", "offshore", "hold",
              "rafter", "acost", "actual", "var", "after", "newvar"):
        c = S[k]
        x = ws.cell(r_tot, c)
        # built off each section's own data rows, never off its subtotal: a section that
        # does not earn a subtotal under the rule above still has to be counted here, and
        # in full. The ranges hold data rows only - no label row and no subtotal is inside
        # one - so nothing can be added twice.
        src = [f"{L(c)}{srange[kk][0]}:{L(c)}{srange[kk][1]}" for kk in live]
        if oh_else:
            src.append(f"{L(c)}{oh_else}")
        src = ",".join(src)
        # The archetype total is what the archetype prices across the whole portfolio, and
        # the variance beside it is actual less that. It is the number the owner asked for
        # and it is the same one the 1.x tab states; what it must not do is pretend the
        # archetype covers everything. It does not - the overhead lines and any programme
        # with no funded figure carry a dash and are not in it - so the label says which
        # rows it prices and the sections above show exactly which those are.
        if k in ("var", "newvar"):
            a, b = L(S["acost"]), L(S["actual"] if k == "var" else S["after"])
            x.value = (f'=IF(ISNUMBER(${a}{r_tot}),'
                       f'ROUND(${b}{r_tot}-${a}{r_tot},6),"-")')
        elif k == "acost":
            # SUM ignores the dashes, so this is the archetype where there is one
            x.value = f'=IF(COUNT({src})=0,"-",SUM({src}))'
        else:
            x.value = f"=SUM({src})"
        x.number_format = (opts.M2 if c >= S["acost"] else
                           (opts.C1 if k == "aroles" else opts.CT))
        x.alignment = opts.RGT

    # Both controls state their figure in C, beside the label in B. They used to sit under
    # the column they were derived from - one in F, one in O - so the two rows that must
    # both read zero were eleven columns apart and neither was next to the words that say
    # what it is. The reader checks two cells in one place now, on all fourteen tabs.
    CTL_COL = S["type"]
    for i, (lab, f, nf) in enumerate([
            ("Control - roles against the ledger, must be 0",
             f"=${L(S['roles'])}${r_tot}-COUNTIFS({REV}!$AJ$2:$AJ${LAST},$C$3)",
             opts.CTL_C),
            ("Control - cost against the ledger ($m), must be 0",
             f"=ROUND(${L(S['actual'])}${r_tot}-SUMIFS({REV}!$AA$2:$AA${LAST},"
             f"{REV}!$AJ$2:$AJ${LAST},$C$3)/1000000,6)", opts.CTL_M)]):
        ws.cell(r_ctl + i, 2).value = lab
        ws.cell(r_ctl + i, 2).font = opts.BODY
        _m(ws, r_ctl + i, CTL_COL, f, nf)

    # ---- FTE ----
    opts.bar(ws, r_ftebar, 2, len(P_HDR), f"{pf} FTE")
    for i, w in enumerate(P_W):
        W[i] = max(W[i], w)
    opts.head(ws, r_ftehdr, 2, list(P_HDR), W)
    dv = DataValidation(type="list", formula1='"Filled,Hire,Hold,Offshore"',
                        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    R = wv[REVIEW]
    for g in staffed:
        a, b = span(g)
        bd = band[g]
        opts.row(ws, bd, 2, [g] + [None] * (len(P_HDR) - 1), [None] * len(P_HDR),
                 bg=opts.PALE, bold=True)
        ws.cell(bd, P["name"]).alignment = opts.LFT
        n = f'COUNTA(${L(P["name"])}${a}:${L(P["name"])}${b})'
        ws.cell(bd, P["role"]).value = f'={n}&IF({n}=1," role"," roles")'
        ws.cell(bd, P["role"]).alignment = opts.LFT
        for k in ("cost", "after"):
            x = ws.cell(bd, P[k])
            x.value = f"=SUM({L(P[k])}{a}:{L(P[k])}{b})"
            x.number_format, x.alignment = opts.M0, opts.RGT
        for rw, x in people.get(g, []):
            i = x["row"]
            for k, col in (("name", "B"), ("role", "C"), ("status", "AK")):
                cell = ws.cell(rw, P[k])
                cell.value = f"=INDEX({REV}!${col}:${col},{i})"
                cell.font, cell.alignment = opts.BODY, opts.LFT
            lv = ws.cell(rw, P["lever"])
            lv.value = "Filled" if x["status"] == "Filled" else "Hire"
            lv.fill, lv.border = opts.fl(opts.YEL), opts.BOX
            lv.font, lv.alignment = opts.BODY, opts.CEN
            dv.add(lv)
            _m(ws, rw, P["cost"], f"=INDEX({REV}!$AA:$AA,{i})", opts.M0)
            _m(ws, rw, P["after"],
               f"=${L(P['cost'])}{rw}*IFERROR(INDEX(Lists!$AD:$AD,"
               f"MATCH(${L(P['lever'])}{rw},Lists!$AC:$AC,0)),1)", opts.M0)

    # no frozen panes anywhere in this workbook - owner's instruction
    # A section that is not on the tab has no anchor, and a one-row section anchors on its
    # own row rather than on a subtotal it no longer has. Every consumer reads these rather
    # than working a row number out for itself.
    return {"tab": tab, "pf": pf, "total_row": r_tot,
            "delivery_row": anchor.get("arch"), "direct_row": anchor.get("direct"),
            "nofig_row": anchor.get("none"), "overhead_row": anchor.get("oh"),
            "elsewhere_row": oh_else, "note_row": oh_note,
            # first and last data row of each section that is on the tab, so a consumer
            # that needs a section total can add the rows even where the section has no
            # subtotal row of its own
            "spans": {k: list(v) for k, v in srange.items()},
            "control_col": CTL_COL, "control_rows": [r_ctl, r_ctl + 1],
            "header_row": HDR,
            "first_squad": min(srow.values()) if srow else HDR + 1,
            "last_squad": max(srow.values()) if srow else HDR + 1,
            "squads": arch, "direct": direct, "nofig": nofig, "overhead": overhead,
            "srow": srow, "people": sum(len(v) for v in people.values()),
            "cols": S}


def _m(ws, r, c, formula, fmt=None):
    x = ws.cell(r, c)
    x.value = formula
    x.number_format = fmt or opts.M2
    x.alignment = opts.RGT
    x.font = opts.BODY
    return x


def run(src, dst):
    _boot_last(src)
    wb = openpyxl.load_workbook(src)
    wv = openpyxl.load_workbook(src, data_only=True)
    rows = ledger(wv)
    bypf = collections.defaultdict(list)
    for x in rows:
        bypf[x["pf"]].append(x)
    bounds = {d: squad_table_bounds(wb, d) for d in set(DESIGN.values())}
    tabs = [s for s in wb.sheetnames if re.match(r"^2\.\d+ ", s)]
    anchors, out = {}, []
    # tab -> portfolio: the tab's own C3 cell on tabs this build has written before, a
    # name-suffix match second, and the tab number last - the review workbook's 2.x tabs
    # are the old generation and carry neither the C3 cell nor the current names
    BY_NUM = {"2.1": "Ampol Retail", "2.2": "Customer", "2.3": "Enterprise Data",
              "2.4": "TDD Group Functions", "2.5": "P&C", "2.6": "Finance",
              "2.7": "Infrastructure", "2.8": "Energy Solutions & B2B",
              "2.9": "Commercial Fuels", "2.10": "Z Retail", "2.11": "COE Cyber",
              "2.12": "COE BP&T", "2.13": "COE SA&D", "2.14": "EGI"}
    for tab in tabs:
        pf = wv[tab]["C3"].value
        if pf not in bypf:
            pf = next((p for p in bypf if p and tab.endswith(str(p))),
                      BY_NUM.get(tab.split(" ")[0]))
        if pf not in bypf:
            raise SystemExit(f"{tab}: no portfolio found (C3, suffix and number all miss)")
        a = build(wb, wv, tab, bypf[pf], bounds)
        anchors[tab] = a
        out.append(f"{tab}: {len(a['squads'])} archetyped, {len(a['direct'])} directly "
                   f"funded, {len(a['nofig'])} with no figure, {len(a['overhead'])} "
                   f"overhead, {a['people']} people")
    json.dump(anchors, open("anchors_final.json", "w"), indent=1)
    wb.save(dst)
    return out


def _boot_last(src):
    set_last(openpyxl.load_workbook(src))


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
