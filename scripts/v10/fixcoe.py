"""What the COE audit found that is still mine to fix.

This file used to rewrite `1.11!C15` and `1.12!C15` from the owner's `=C14` to
`=C14+C13`, on the finding that the portfolio-funded line above was feeding nothing. That
was right two generations ago, when planned spend on both tabs was gross. It is wrong now.
In rev he restructured both tabs so planned spend is net of C13 - his own note at
`1.11!B9` says so in as many words - and the rewrite compared a net spend to a gross
budget, which put both tabs out against 0.2 Data Config by exactly C13: 2.2 on 1.11, 1.4
on 1.12. His `=C14` is correct and it survives.

What remains: the banned Category columns survived as three defined names - BPTCat,
SADCat, CYBCat - pointing at Lists!E2:G4. No formula used them and no COE tab has a
Category column, so the money was never affected, but the columns the owner ruled out
were still in the file and would reappear in any dropdown built off a name.

--- the "formulas that make no sense" round -------------------------------------------

THE LEVER PRICE LIVES IN ONE PLACE.  `Lists!AC2:AD5` is the model's lever table - Filled
1, Hire 1, Hold 0, Offshore 0.4 - and it is what the 2.x working tabs price their levers
off. The three COE design tabs did not read it: all 102 of their T-column cost engines
(24 on 1.11, 26 on 1.12, 52 on 1.13) carried
`IF($H21="Hold",0,IF($H21="Offshore",0.4,1))` inline, so the 0.4 the owner can retype on
Lists moved the working tabs and left the design tabs exactly where they were. They
look it up now, in the house idiom the 2.x tabs already use:

    IFERROR(INDEX(Lists!$AD$2:$AD$5,MATCH($H21,Lists!$AC$2:$AC$5,0)),1)

The IFERROR default of 1 is not decoration - it is what preserves the behaviour exactly.
The inline IF charged full price for anything that was not "Hold" or "Offshore", which on
these tabs means "Onshore" and a blank cell, and neither is in the lever table; both fall
through to 1 as they did before. "Filled" and "Hire" are in the table and are 1 there, so
they are unmoved either way. 1.11's Hold branch, added a round earlier in fix1x, survives
as the Hold row of the table.

THE PLATFORM TOTAL ROWS.  Column H of a platform total stops at the last squad row while
column I runs one further and takes in the Platform Overhead line. It reads like a typo
and it is not: an overhead is not a squad and has no archetype price, so it exists in the
TDD Cost column only, and the H cell on every overhead row is empty. Widening H would
double-count against the portfolio summary, which already reads the overhead out of I.
Checked on all 29 total rows across all fourteen 1.x tabs - every one of them is the same
way round. They carry a comment now so the next reader does not "fix" them. This runs
here rather than in repair_design because fix1x collapses the empty platform blocks in
between, and a comment on a collapsed row would outlive its cell.
"""
import copy
import re

import openpyxl
from openpyxl.styles import Border, PatternFill

import opts

BANNED_NAMES = ("BPTCat", "SADCat", "CYBCat")


def drop_names(wb):
    out = []
    for n in BANNED_NAMES:
        if n in wb.defined_names:
            del wb.defined_names[n]
            out.append(f"defined name {n} removed")
    l = wb["Lists"]
    n = 0
    for r in range(1, 6):
        for c in (5, 6, 7):
            if l.cell(r, c).value is not None:
                l.cell(r, c).value = None
                l.cell(r, c).fill = PatternFill()
                l.cell(r, c).border = Border()
                n += 1
    out.append(f"Lists!E1:G5 cleared, {n} cells - the banned Category columns")
    return out


# ---------------------------------------------------------------- the design tabs
def fix_validation(wb):
    """SupportPct stopped at 90% while six live cells hold 100%.

    Lists!D11 is 1, sitting just outside the named range, so `1.1!G27`, `1.2!G33`, `G39`,
    `G40`, `1.4!G21` and `1.6!G27` all breached their own dropdown.
    """
    out = []
    dn = wb.defined_names.get("SupportPct")
    if dn is not None and dn.value.endswith("$D$10"):
        dn.value = dn.value.replace("$D$10", "$D$11")
        out.append("SupportPct extended to include 100%")
    return out


# tab -> {cell: (what it should say or compute, why)}
SIGNS = {
    # 1.4's variance ran budget less spend and was labelled "(Over)/ Under", the opposite
    # convention to every other design tab, and the funding block below it reads that cell
    # as though it were spend less budget
    "1.4 TDD Group Functions": {"H8": "TDD over/(under) budget ($m)", "I8": "=I7-I6"},
    # 1.7's budget box sits one row higher than the others, so I8+I9 added the NZ variance
    # to a total that already contained it. It read correctly only because NZ is zero.
    "1.7 Infrastructure": {"C17": "=$I$7+$I$8"},
    # 1.5 carried the same variance twice under the same label, and the funding block read
    # the second copy
    "1.5 P&C": {"H10": None, "I10": None, "C16": "=$I$8"},
}
HEADERS = {"1.4 TDD Group Functions": {"F20": "AU / NZ", "F29": "AU / NZ"},
           "1.5 P&C": {"F24": "AU / NZ"}}

def fix_signs(wb):
    out = []
    for tab, cells in SIGNS.items():
        ws = wb[tab]
        for cell, val in cells.items():
            was = ws[cell].value
            ws[cell] = val
            out.append(f"{tab}!{cell} {was!r} -> {val!r}")
    for tab, cells in HEADERS.items():
        ws = wb[tab]
        for cell, val in cells.items():
            if ws[cell].value != val:
                out.append(f"{tab}!{cell} {ws[cell].value!r} -> {val!r}")
                ws[cell] = val
    return out


def drop_orphan_overheads(wb):
    """A "Platform Overhead" label with nothing beside it reads as an unfinished line.

    The strategic-programme platforms have no squads to support, so they carry no platform
    overhead. The label was left behind; the blank cell it points at is summed into the
    portfolio overhead line either way, so removing the label changes no figure.
    """
    out = []
    for tab in [t for t in wb.sheetnames if t.startswith("1.") and " " in t]:
        ws = wb[tab]
        for r in range(1, min(ws.max_row, 90) + 1):
            if str(ws.cell(r, 2).value or "").strip() == "Platform Overhead" \
                    and ws.cell(r, 9).value is None:
                ws.cell(r, 2).value = None
                out.append(f"{tab}!B{r} orphan 'Platform Overhead' label removed")
    return out


LISTS_LEVER = ("IFERROR(INDEX(Lists!$AD$2:$AD$5,"
               "MATCH({h},Lists!$AC$2:$AC$5,0)),1)")
# the inline engine as all three tabs write it, with the lever cell captured
INLINE = re.compile(r'IF\((\$[A-Z]{1,2}\d+)="Hold",0,IF\(\1="Offshore",0\.4,1\)\)')
COE_TABS = ("1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles")


def lever_from_lists(wb):
    """The T-column engines look the lever factor up instead of carrying it."""
    out = []
    if "Lists" not in wb.sheetnames:
        return ["Lists is not in the workbook - lever factors left inline"]
    l = wb["Lists"]
    table = {str(l.cell(r, 29).value): l.cell(r, 30).value for r in range(2, 6)}
    # AD5 is either the literal 0.4 or, since wave K, a live read of the owner's own
    # Offshore rate on 0.3 (one input driving the lever and the archetype side alike).
    # This guard reads the formula view, so the link is accepted by its text - refusing
    # it left every COE engine carrying an inline 0.4 the owner could no longer move.
    off_ok = (table.get("Offshore") == 0.4
              or str(table.get("Offshore")) == "='0.3 Squad Archetypes'!$K$5")
    if table.get("Hold") != 0 or not off_ok:
        return [f"Lists!AC2:AD5 reads {table} - not the factors the engines carry, "
                f"left inline"]
    for tab in COE_TABS:
        if tab not in wb.sheetnames:
            out.append(f"{tab}: not in the workbook")
            continue
        ws = wb[tab]
        n = 0
        # the engines live in T (column 20); the walk is bounded so a tab with a wide
        # used range does not cost a full-sheet scan
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=30):
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                new, k = INLINE.subn(
                    lambda m: LISTS_LEVER.format(h=m.group(1)), v)
                if k:
                    c.value = new
                    n += k
        out.append(f"{tab}: {n} lever factors now read Lists!AC2:AD5"
                   if n else f"{tab}: no inline lever factor left to repoint")
    out.append("lever factors (Hold 0, Offshore 0.4, everything else 1) now priced from "
               "Lists!AC2:AD5 on the design tabs as well as the working tabs")
    return out


TOTAL_NOTE = ("Total Squad Cost stops at the last squad row on purpose. A platform "
              "overhead is not a squad and has no archetype price, so it exists in the "
              "TDD Cost column only and the cell beside this one on the overhead row is "
              "empty by design. Widening this SUM to take the overhead row in would "
              "double-count it against the portfolio summary above, which already reads "
              "the overhead out of column I. Checked on all fourteen 1.x tabs.")


def note_total_asymmetry(wb, out):
    """The H-vs-I total rows get a comment so nobody 'fixes' them."""
    from openpyxl.comments import Comment
    n = 0
    for tab in [t for t in wb.sheetnames if re.match(r"^1\.\d+ ", t)]:
        ws = wb[tab]
        for r in range(1, min(ws.max_row, 95) + 1):
            if not str(ws.cell(r, 2).value or "").strip().endswith(" Total"):
                continue
            h, i = ws.cell(r, 8), ws.cell(r, 9)
            if not (isinstance(h.value, str) and isinstance(i.value, str)):
                continue
            # the overhead row is the one this total's I range reaches and its H does not
            if h.value == i.value.replace("I", "H"):
                continue                       # symmetric block, no overhead row
            h.comment = Comment(TOTAL_NOTE, "TDD cost model")
            n += 1
    out.append(f"{n} platform total rows noted: H stops at the squads, I takes in the "
               f"Platform Overhead row, and that is deliberate")



# ------------------------------------------------ 1.13, in the shape of its two siblings
# His words: "looks nothing like the other coe tabs and doesn't tell me the budget vs
# spend". He was right on both counts. 1.11 and 1.12 read Grouping / Roles / Filled /
# Vacant / Planned spend / Budget to draw down / Variance / Cost AU / Cost NZ, with the
# budget beside the spend and the difference between them in the next column along. 1.13
# read the same nine columns except that the seventh was "Left to fund" - a funding gap,
# not a variance - and it was the only one of the three where the budget line was buried in
# a four-row bucket table with a CapEx input in the middle of it.
#
# What this pass does, and only this:
#   the seventh column becomes Variance, budget less spend, on the total row where the
#     budget sits - the same column its two siblings carry and the figure he asked for;
#   the funding block becomes two lines in 1.11's list style, the COE allocation off
#     0.2 Data Config and the total to draw down. The 0.5 CapEx input, its total row and
#     the Left to fund row come out, with every reference to them (RETIRED.md);
#   the roles table gains a cream Uplift % column - the five part-charges he set - and its
#     cost engine multiplies by (1 - uplift %), so a role that is half charged to the cyber
#     uplift programme costs the COE half of itself. The slice itself is stated in its own
#     column beside it, which is what 1.14's funding block draws its part-charge line from.
REVIEW = "REVIEW - Complete Role Mapping"
CY = "1.13 Cyber Roles"
CRSO_ALLOC = "COE - Cyber, Risk & Service Ops allocation ($m) - 0.2 Data Config"
CY_SUMMARY_H = "Variance"
UPLIFT_HDR = "Uplift %"
SLICE_HDR = "Charged to the cyber uplift programme ($m)"
UPLIFT_COL, SLICE_COL = 9, 21           # I on the face of the table, U beside the engine

# His per-role decisions, keyed on the person exactly as REVIEW columns B and C read them -
# never on a REVIEW row number (D109). Two of these keys name two people each ("Vacant |
# Operations Analyst", "Vacant | Data Risk Analyst"); both of each pair carry the same
# decision, which is why the key is safe here and why the code applies it to every match
# rather than to the first.
#
# Offshore: the COE's own offshoring levers. The two vacant Operations Analysts move off
# Hold onto Offshore - a role being offshored is not a role being cancelled.
CY_OFFSHORE = [("Jack Jenkins", "Asset Analyst"),
               ("Jas Mann", "Technology Support Engineer"),
               ("Ritika Salaria", "Configuration Analyst"),
               ("Vacant", "Lead - Asset & Configuration"),
               ("Vacant", "Lead - Service Performance & Insights"),
               ("Vacant (AKL)", "Tech Support Technician"),
               ("Vacant", "Operations Analyst")]
# Hold stays Hold: the two Data Risk Analysts, and the Sydney technician he is recruiting.
CY_HOLD = [("Vacant", "Data Risk Analyst"),
           ("Vacant (SYD)", "Tech Support Technician")]
# The part-charges to the cyber uplift programme. The role stays in the COE and this share
# of it is funded by the programme, so the COE carries the rest.
CY_UPLIFT = [("Chris Lyons", "Cyber Security Architect", 0.5),
             ("James Byrne", "Head of Cyber Strategy & Technology", 0.5),
             ("Rahul Sahni", "Cyber Security Architect", 0.4),
             ("Darshan Suvama", "Cyber Offensive Security Lead", 0.25),
             ("Vanessa Castro", "Cyber GRC Analyst", 0.0)]
ROWREF = re.compile(r"\$([A-Z]{1,2})\$(\d+)")


def _cy_rows(wb):
    """(design row, REVIEW row, name, title) for every role on 1.13's list."""
    ws, R = wb[CY], wb[REVIEW]
    hdr = next((r for r in range(1, 30)
                if str(ws.cell(r, 2).value or "").strip() == "Name"), None)
    if hdr is None:
        return None, []
    rows = []
    for r in range(hdr + 1, 90):
        v = ws.cell(r, 2).value
        if not (isinstance(v, str) and v.startswith("=")):
            if rows:
                break
            continue
        m = ROWREF.search(v)
        if not m:
            continue
        i = int(m.group(2))
        rows.append((r, i, str(R.cell(i, 2).value or "").strip(),
                     str(R.cell(i, 3).value or "").strip()))
    return hdr, rows


def cyber_summary(wb, out):
    """The seventh column is a variance, and the budget sits on the total row."""
    ws = wb[CY]
    head = next((r for r in range(1, 20)
                 if str(ws.cell(r, 2).value or "").strip() == "Grouping"), None)
    tot = next((r for r in range(1, 20)
                if str(ws.cell(r, 2).value or "").strip() == "Total"), None)
    if head is None or tot is None:
        out.append(f"{CY}: no Grouping header or Total row - summary left alone")
        return
    was = ws.cell(head, 8).value
    ws.cell(head, 8).value = CY_SUMMARY_H
    # the two grouping rows have no budget of their own - the allocation is for the COE as
    # a whole - so the variance is stated once, on the row that carries the budget
    for r in range(head + 1, tot):
        for c in (7, 8):
            ws.cell(r, c).value = None
    ws.cell(tot, 7).value = "=C12"
    ws.cell(tot, 8).value = f"=G{tot}-F{tot}"
    ws.cell(tot, 8).number_format = opts.M2
    out.append(f"{CY}!{openpyxl.utils.get_column_letter(8)}{head}: {str(was)[:24]!r} -> "
               f"{CY_SUMMARY_H!r}, budget less planned spend on the total row - the "
               f"budget-against-spend column its two siblings carry")


def cyber_funding(wb, out):
    """Two funding lines in 1.11's list style, and nothing else."""
    ws = wb[CY]
    bar = next((r for r in range(1, 30)
                if str(ws.cell(r, 2).value or "").strip().startswith("Funding")), None)
    if bar is None:
        out.append(f"{CY}: no funding block found - left alone")
        return
    ws.cell(bar, 2).value = "Funding"
    lines = [(CRSO_ALLOC, "='0.2 Data Config'!$E$7"),
             ("Total budget to draw down ($m)", f"=C{bar + 1}")]
    gone = []
    for i, (lab, f) in enumerate(lines):
        r = bar + 1 + i
        ws.cell(r, 2).value = lab
        ws.cell(r, 2).font, ws.cell(r, 2).alignment = opts.BODY, opts.LFT
        x = ws.cell(r, 3)
        x.value, x.font, x.number_format, x.alignment = f, opts.BODY, opts.M2, opts.RGT
        # the funding list is a label and a figure, the way 1.11 writes it. The
        # "Planned spend less CapEx" pair parked out at E/F on the first line goes with the
        # CapEx input it read.
        for c in range(4, 9):
            y = ws.cell(r, c)
            if y.value is not None:
                gone.append(f"{openpyxl.utils.get_column_letter(c)}{r}={y.value!r}")
            y.value = None
            y.fill, y.border, y.font = PatternFill(), Border(), opts.BODY
    # the CapEx input, the bucket total, Left to fund and the working figure beside them
    for r in range(bar + 1 + len(lines), bar + 8):
        for c in range(2, 9):
            x = ws.cell(r, c)
            if x.value is not None:
                gone.append(f"{openpyxl.utils.get_column_letter(c)}{r}={x.value!r}")
            x.value = None
            x.fill, x.border, x.font = PatternFill(), Border(), opts.BODY
    out.append(f"{CY}: funding block is two lines - {CRSO_ALLOC!r} off 0.2!E7, then the "
               f"total to draw down. Removed: {'; '.join(str(g)[:46] for g in gone[:6])}"
               + (f" (+{len(gone) - 6} more)" if len(gone) > 6 else ""))


def cyber_uplift(wb, out):
    """The levers he set, the cream Uplift % column, and the two cost engines beside it."""
    ws = wb[CY]
    hdr, rows = _cy_rows(wb)
    if hdr is None:
        out.append(f"{CY}: no roles list - levers and uplift column not written")
        return
    want = {(n, t): "Offshore" for n, t in CY_OFFSHORE}
    want.update({(n, t): "Hold" for n, t in CY_HOLD})
    pct = {(n, t): p for n, t, p in CY_UPLIFT}
    for cell, text in ((ws.cell(hdr, UPLIFT_COL), UPLIFT_HDR),
                       (ws.cell(hdr, SLICE_COL), SLICE_HDR)):
        cell.value = text
        cell._style = copy.copy(ws.cell(hdr, 8)._style)
    moved, tog, n_eng = [], [], 0
    for r, i, name, title in rows:
        state = want.get((name, title), "Onshore")
        if str(ws.cell(r, 8).value or "").strip() != state:
            moved.append(f"r{i} {name} {ws.cell(r, 8).value!r} -> {state}")
            ws.cell(r, 8).value = state
        p = pct.get((name, title))
        x = ws.cell(r, UPLIFT_COL)
        if p is None:
            x.value = None
            x.fill = PatternFill()
        else:
            x.value = p
            x.fill = opts.fl(opts.YEL)
            x.number_format = opts.PCT
            x.alignment = opts.RGT
            x.font = opts.BODY
            tog.append(f"r{i} {name} {p:.0%}")
        # the engine, extended by the uplift factor, and the slice stated beside it. Built
        # off whatever the T cell already holds so the lever lookup this file just
        # repointed at Lists is not re-typed here and cannot drift from it.
        eng = ws.cell(r, 20).value
        if not (isinstance(eng, str) and eng.startswith("=")):
            continue
        base = eng.split("*(1-N($")[0].split("*N($")[0]
        ws.cell(r, 20).value = f"{base}*(1-N($I{r}))"
        s = ws.cell(r, SLICE_COL)
        s.value = f"{base}*N($I{r})"
        s.number_format, s.font, s.alignment = opts.M2, opts.BODY, opts.RGT
        n_eng += 1
    out.append(f"{CY}!{openpyxl.utils.get_column_letter(UPLIFT_COL)}{hdr} = {UPLIFT_HDR!r}, "
               f"cream, {len(tog)} part-charges set: {', '.join(tog)}")
    out.append(f"{CY}: {len(moved)} levers set from his ruling - {'; '.join(moved)}"
               if moved else f"{CY}: every lever already reads his ruling")
    out.append(f"{CY}: {n_eng} cost engines now price cost x lever factor x (1 - uplift %), "
               f"with the uplift slice itself in "
               f"{openpyxl.utils.get_column_letter(SLICE_COL)} under {SLICE_HDR!r} - the "
               f"column 1.14's funding block totals")


def cyber_tab(wb, out):
    if CY not in wb.sheetnames or REVIEW not in wb.sheetnames:
        out.append(f"{CY} or {REVIEW} not in the workbook - the cyber restructure is NOT "
                   f"applied")
        return
    cyber_summary(wb, out)
    cyber_funding(wb, out)
    cyber_uplift(wb, out)
    # his note under the roles table states the role count and the count has changed. It is
    # corrected in post2707, not here: finish.py's prose sweep keeps that cell only while it
    # matches his own workbook byte for byte, so editing it before finish runs would delete
    # the sentence instead of correcting it.


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = (drop_names(wb) + fix_validation(wb) + fix_signs(wb)
           + drop_orphan_overheads(wb) + lever_from_lists(wb))
    note_total_asymmetry(wb, out)
    # after lever_from_lists, so the engine the uplift factor extends is the one that looks
    # its lever factor up on Lists rather than an inline 0.4
    cyber_tab(wb, out)
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
