"""Stage 13: the six outstanding defects, fixed together.

1  COST BASE - 348 of 524 cost cells were hardcoded and the rest used three different
   formula shapes. They are all the same calculation:

       salary role   = base x (1 + STI + payroll + pension + CPI) + medical
       day-rate role = day rate x days x (1 + CPI)

   That single pair reproduces 523 of the 524 stored values to the cent. Only Tim Corin
   (r172) differs - his stored $275,810.25 is a banded rate used by 26 other roles, where
   his own components give $321,135. That is a decision, not a formula, so it moves to an
   override column that is visible and editable rather than being buried as a literal.
   Days worked becomes an input on Lists instead of a 222 typed into 42 formulas.

2  BUDGET VARIANCE - the workbook stated two, $54.3m apart. 0.2 column F mixed three
   bases: design cost for the ten portfolios, net actual for two COEs, gross actual for
   Cyber, and nothing at all for EGI. It now reads actual ledger cost for every row, the
   same basis 3.1 uses, so the two agree.

3  SIGN CONVENTION - four were in use, two of them side by side in 3.1 row 6. One rule
   now: variance = actual - budget, positive = over. Every variance column relabelled.

4  OFFSHORE - the 1.x design tabs carry an Onshore/Offshore choice per squad and 0.3
   prices both (column G onshore, column H offshore at 40%). The 2.x archetype cost read
   column G unconditionally, so six squads designed offshore were priced onshore and the
   dropdown moved nothing downstream. It now follows the design.

5  COE GROSS vs NET - the same $3.6m appeared gross on 3.1/3.2 and net on 3.4/0.2/1.11/
   1.12, with notes on 1.11 and 1.12 claiming net while their formulas were gross. All
   presentation is gross; the portfolio-funded amount stays as its own column so the net
   number is still visible without a second definition of the same cost.

6  FUNDING DRAWDOWN - 3.4 reported $0.5m drawn. Two of its three formulas pointed at
   blank cells and the third was a typed 0. It now sums what the ten 1.x tabs actually
   apply, against the matching pools.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import model

REV = f"'{model.REVIEW}'"
LR = model.LAST_ROW
BOLD = Font(bold=True)
YELLOW = PatternFill("solid", start_color="FFFFFF00", end_color="FFFFFF00")
DAYS_CELL = "Lists!$AG$15"


def cost_base(wb):
    """One formula for every cost cell, with a visible override column."""
    ws = wb[model.REVIEW]
    l = wb["Lists"]
    l["AF14"] = "Cost base - input"
    l["AF14"].font = BOLD
    l["AF15"] = "Days worked per year (day-rate roles)"
    l["AG15"] = 222
    l["AG15"].fill = YELLOW
    l["AF16"] = ("Full Cost AUD = base x (1 + STI + payroll + pension + CPI) + medical, "
                 "or day rate x days x (1 + CPI). Column AB holds any agreed override.")

    ws["AU1"] = "Agreed cost override (AUD)"
    ws["AU1"].font = BOLD
    for i in range(2, LR + 1):
        ws[f"AA{i}"] = (
            f'=IF(TRIM($B{i})="","",IF(N($AU{i})>0,$AU{i},'
            f"IF(N($S{i})>0,$S{i}*{DAYS_CELL}*(1+N($Z{i})),"
            f"N($U{i})*(1+N($V{i})+N($W{i})+N($X{i})+N($Z{i}))+N($Y{i}))))"
        )
    # the one role whose stored cost is a banded rate, not its own components
    ws["AU172"] = 275810.25
    ws["AU172"].fill = YELLOW
    ws["AV172"] = ("Banded rate agreed for this role; its own components give $321,135. "
                   "Clear column AB to price it from the components.")
    return ["REVIEW AA: one formula for all 525 rows, days worked now an input on Lists, "
            "one agreed override (r172) in a yellow cell in column AU"]


def budget_one_basis(wb):
    """0.2 column F reads actual ledger cost, the same basis 3.1 uses.

    Customer, Cyber, BP&T and SA&D each draw on two budget lines. The portfolio's cost is
    charged against the first of those lines and the second shows nil, so the column
    totals the ledger exactly instead of counting a portfolio twice.
    """
    ws = wb["0.2 Data Config"]
    l = wb["Lists"]
    bmap = {}
    for r in range(2, 20):
        k, v = l.cell(r, 37).value, l.cell(r, 38).value
        if k and v:
            bmap[str(k).strip()] = str(v).strip()
    charged = set()
    for r in range(6, 26):
        lab = ws.cell(r, 2).value
        if lab is None or str(lab).strip() == "":
            continue
        pf = bmap.get(str(lab).strip())
        if pf and pf not in charged:
            charged.add(pf)
            ws.cell(r, 6).value = (f'=SUMIFS({REV}!$AA$2:$AA${LR},'
                                   f'{REV}!$AJ$2:$AJ${LR},"{pf}")/1000000')
        elif pf:
            ws.cell(r, 6).value = 0
            ws.cell(r, 8).value = f"charged on the first {pf} line above"
        else:
            ws.cell(r, 6).value = 0
        ws.cell(r, 7).value = f"=$F{r}-$E{r}"
    ws["F26"] = "=SUM(F6:F25)"
    ws["G26"] = "=$F26-$E26"
    ws["F5"] = "Actual cost ($m)"
    ws["G5"] = "Variance (actual - budget, + = over)"
    ws["H13"] = None
    ws["G14"] = None
    return ["0.2 column F reads actual ledger cost on one basis, each portfolio charged "
            "once; variance is actual - budget on every row"]


def signs(wb):
    """One convention: variance = actual - budget, positive = over."""
    out = []
    s1 = wb["3.1 Group Summary"]
    for r in range(6, 21):
        if s1.cell(r, 2).value is None:
            continue
        s1.cell(r, 5).value = f"=$D{r}-$C{r}"
    s1["E5"] = "Variance to budget (actual - budget, + = over)"
    s1["H5"] = "Variance to archetype (actual - archetype, + = over)"
    out.append("3.1: variance to budget flipped to actual - budget so both variance "
               "columns in the same row now mean the same thing")

    for tab, cells in (("1.11 BP&T", ["H6", "H7", "H8"]),
                       ("1.12 SA&D", ["I6", "I7", "I8"]),
                       ("1.13 Cyber Roles", ["H6", "H7", "H8"])):
        ws = wb[tab]
        col = cells[0][0]
        for c in cells:
            r = c[1:]
            spend = "F" if tab != "1.12 SA&D" else "G"
            budget = "G" if tab != "1.12 SA&D" else "H"
            ws[c] = f"=${spend}{r}-${budget}{r}"
        ws[f"{col}5"] = "Left to fund (spend - budget, + = over)"
    out.append("1.11 / 1.12 / 1.13: left to fund is spend - budget on all three, "
               "so adjacent COE tabs no longer carry opposite signs")

    s4 = wb["3.4 COE Summary"]
    s4["H5"] = "Left to fund (spend - budget, + = over)"
    for r in range(6, 12):
        s4[f"H{r}"] = f"=$F{r}-$G{r}"
    return out


def offshore(wb):
    """2.x archetype cost follows the Onshore/Offshore choice on the design tab."""
    from build_2xfix import DESIGN, squad_table_bounds
    A3 = "'0.3 Squad Archetypes'"
    n = 0
    for tab, dt in DESIGN.items():
        ws = wb[tab]
        lo, hi = squad_table_bounds(wb, dt)
        for r in range(6, 40):
            f = ws[f"L{r}"].value
            if not (isinstance(f, str) and A3 in f and "$G$5" in f):
                continue
            key = f'$C{r}&"|"&$R{r}'
            m = f"MATCH($S{r},'{dt}'!$B${lo}:$B${hi},0)"
            ws[f"L{r}"] = (
                f"=IFERROR(IF(INDEX('{dt}'!$E${lo}:$E${hi},{m})=\"Offshore\","
                f"INDEX({A3}!$H$5:$H$23,MATCH({key},{A3}!$A$5:$A$23,0)),"
                f'INDEX({A3}!$G$5:$G$23,MATCH({key},{A3}!$A$5:$A$23,0))),"-")'
            )
            n += 1
    return [f"2.x: {n} archetype cost cells now price offshore squads from 0.3 column H, "
            f"so the design tab's Onshore/Offshore choice reaches the summaries"]


def coe_gross(wb):
    """One definition of COE cost: gross, with the portfolio-funded amount beside it."""
    out = []
    s4 = wb["3.4 COE Summary"]
    for r in range(6, 12):
        s4[f"F{r}"] = f"=$K{r}"
    s4["F5"] = "People cost, gross ($m)"
    s4["L5"] = "Of which funded inside portfolio overheads ($m)"
    s4["B13"] = ("Cost is stated gross on every tab. Column L shows the part funded inside "
                 "portfolio overheads, so the net figure is visible without a second "
                 "definition of the same cost.")
    for tab, cell in (("1.11 BP&T", "B9"), ("1.12 SA&D", "B9")):
        wb[tab][cell] = ("Planned spend is gross people cost. The portfolio-funded amount "
                         "is shown separately below and on 3.4.")
    out.append("3.4 / 1.11 / 1.12: COE cost stated gross everywhere, funded-by-portfolios "
               "kept as its own column; the notes now match the formulas")
    return out


def drawdown(wb):
    """3.4's funding pools: sum what the ten 1.x tabs actually apply.

    The budget block sits at a different row on almost every 1.x tab, so the lines are
    found by their label in column H rather than by a row number. 3.4 previously reported
    $0.5m drawn because two of its three formulas pointed at blank cells and the third was
    a typed zero.
    """
    ws = wb["3.4 COE Summary"]
    tabs = ["1.1 Ampol Retail", "1.2 Customer", "1.3 Enterprise Data",
            "1.4 TDD Group Functions", "1.5 P&C", "1.6 Finance", "1.7 Infrastructure",
            "1.8 Energy Solutions & B2B", "1.9 Commercial Fuels", "1.10 Z Retail"]
    want = {"OpEx": ["opex initiative"], "Sig": ["significant item"], "Cap": ["capex"]}
    found = {k: [] for k in want}
    for t in tabs:
        d = wb[t]
        for r in range(1, d.max_row + 1):
            lab = str(d.cell(r, 8).value or "").strip().lower()
            if not lab:
                continue
            for k, keys in want.items():
                if any(lab.startswith(x) for x in keys):
                    # the central pool reference lines are not this portfolio's draw
                    if "central pool" in lab or "reference" in lab:
                        continue
                    found[k].append(f"IFERROR(N('{t}'!$J${r}),0)")
    for cell, k in (("D19", "OpEx"), ("D20", "Sig"), ("D21", "Cap")):
        ws[cell] = "=" + ("+".join(found[k]) if found[k] else "0")
    ws["E22"] = "=SUM(E19:E21)"
    ws["B23"] = ("Drawn is the amount each 1.x tab applies against that pool, located by "
                 "label. It previously read $0.5m because two of the three formulas "
                 "pointed at blank cells and the third was a typed zero.")
    return [f"3.4 funding pools: drawn now sums {sum(len(v) for v in found.values())} "
            f"applied lines found by label across the ten 1.x tabs"]


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = []
    for fn in (cost_base, budget_one_basis, signs, offshore, coe_gross, drawdown):
        out += fn(wb)
    wb.save(dst)
    return out


if __name__ == "__main__":
    for x in run("x3.xlsx", "f1.xlsx"):
        print("  ", x)
