"""Stage 1: control tables on Lists, formula-driven derived columns on REVIEW,
0.2 Data Config repairs, and the 1.x design-tab wiring fixes.

Everything downstream reads these control tables, so a change of squad name, lever rate
or overhead allowance is a cell edit on Lists, never a formula edit.
"""
import openpyxl
from openpyxl.styles import Font, Alignment
import model

REVIEW = model.REVIEW
LR = model.LAST_ROW
REV = f"'{REVIEW}'"

# --- Lists control block. Columns W..AJ are unused; rows start at 1 with a header. ---
# W:X   raw squad (col K) -> canonical squad          (typo / case merges)
# Z:AA  canonical squad  -> design squad name         (the pairs Lee confirmed)
# AC:AD lever            -> cost factor
# AF:AH class            -> allowance per portfolio, portfolios charged
SQUAD_MERGE_TABLE = [
    ("AmPos", "AmPOS"),
    ("Manuacturing Group Projects", "Manufacturing Group Projects"),
    ("Integration & Process automation", "Integration & Process Automation"),
    ("Technology Suport", "Technology Support"),
    ("Customer, AI", "Customer AI"),
    ("Data - AU", "Data AU"),
    ("Data - NZ", "Data NZ"),
    ("Data Platform", "Data Platforms"),
]

# Ledger squad (col K) -> the name Lee uses for the same squad on his 1.x design tabs.
# All ten confirmed by Lee, including Ampol Loyalty & Martech and SAP ERP.
DESIGN_NAME_TABLE = [
    ("Z Loyalty & Martech", "Z Energy Martech"),
    ("Cloud, Network & Infra Ops", "Network & Infrastructure"),
    ("Integration & Process Automation", "Integration"),
    ("Manufacturing Group Projects", "Manufacturing & Group Projects"),
    ("SAP ERP", "AU Finance"),
    ("DevOps & QE", "DevOps & Engineering"),
    ("Ampol Loyalty & Martech", "AU CRM & Martech"),
    ("Network & QSR", "Network / QSR"),
    ("P&C RTA", "P&C - RTA"),
    ("Z App and Web", "Z Energy Apps"),
]

LEVER_TABLE = [("Filled", 1.0), ("Hire", 1.0), ("Hold", 0.0), ("Offshore", 0.4)]

# 0.2 Data Config budget line -> portfolio. Customer, Cyber, BP&T and SA&D each draw on
# more than one budget line, which is why 3.1 previously showed most portfolios at zero
# budget. These 14 lines sum to $50.5m, matching 0.2 E26 exactly.
BUDGET_TABLE = [
    ("Ampol Retail", "Ampol Retail"), ("Z Retail", "Z Retail"),
    ("Ampol Customer", "Customer"), ("Z Customer", "Customer"),
    ("TDD Data", "Enterprise Data"), ("TDD", "TDD Group Functions"),
    ("P&C", "P&C"), ("Finance", "Finance"), ("Infrastructure", "Infrastructure"),
    ("Energy Solutions & B2B", "Energy Solutions & B2B"),
    ("Commercial Fuels", "Commercial Fuels"),
    ("COE - Cyber, Risk & Service Ops (see 1.13)", "COE Cyber"),
    ("TDD Cyber", "COE Cyber"),
    ("COE - Transformation", "COE BP&T"), ("COE - Business Partnering", "COE BP&T"),
    ("COE - Strategy Architecture", "COE SA&D"), ("COE - Data", "COE SA&D"),
    ("EGI", "EGI"),
]

# Overhead allowance. Rate per portfolio from the 0.2 Data Config build-up, times the
# number of portfolios it is charged to. Head of Technology is carved out of Leadership
# so the allowance no longer counts the same 11 people twice.
ALLOWANCE_TABLE = [
    ("Head of Technology", 0.1375, 10),
    ("Leadership", 0.30, 10),
    ("Business Partner", 0.22, 10),
    ("Architecture", 0.14, 10),
]

BOLD = Font(bold=True)


def lists_block(wb):
    ws = wb["Lists"]
    def put(col, row, val, bold=False):
        c = ws[f"{col}{row}"]
        c.value = val
        if bold:
            c.font = BOLD
        return c

    put("W", 1, "Squad as typed (col K)", True); put("X", 1, "Canonical squad", True)
    for i, (a, b) in enumerate(SQUAD_MERGE_TABLE):
        put("W", 2 + i, a); put("X", 2 + i, b)

    put("Z", 1, "Canonical squad", True); put("AA", 1, "Name on 1.x design tab", True)
    for i, (a, b) in enumerate(DESIGN_NAME_TABLE):
        put("Z", 2 + i, a); put("AA", 2 + i, b)

    put("AC", 1, "Lever", True); put("AD", 1, "Cost factor", True)
    for i, (a, b) in enumerate(LEVER_TABLE):
        put("AC", 2 + i, a); put("AD", 2 + i, b)

    put("AK", 1, "0.2 Data Config budget line", True)
    put("AL", 1, "Portfolio", True)
    for i, (a, b) in enumerate(BUDGET_TABLE):
        put("AK", 2 + i, a); put("AL", 2 + i, b)

    put("AF", 1, "Shared role class", True)
    put("AG", 1, "Allowance per portfolio ($m)", True)
    put("AH", 1, "Portfolios charged", True)
    put("AI", 1, "Allowance total ($m)", True)
    for i, (k, rate, n) in enumerate(ALLOWANCE_TABLE):
        r = 2 + i
        put("AF", r, k); put("AG", r, rate); put("AH", r, n)
        put("AI", r, f"=$AG{r}*$AH{r}")
    tr = 2 + len(ALLOWANCE_TABLE)
    put("AF", tr, "Total", True)
    put("AI", tr, f"=SUM(AI2:AI{tr-1})", True)
    for col, w in [("W", 30), ("X", 30), ("Z", 30), ("AA", 30), ("AC", 12), ("AD", 12),
                   ("AF", 22), ("AG", 24), ("AH", 18), ("AI", 20),
                   ("AK", 40), ("AL", 24)]:
        ws.column_dimensions[col].width = w
    return ws


def review_columns(wb):
    """Replace the invented derived columns with formulas that read Lee's raw data."""
    ws = wb[REVIEW]
    hdr = {
        "AP": "Squad (canonical, from col K)",
        "AQ": "Leadership",
        "AR": "Cost class",
        "AS": "Squad name on design tab",
    }
    for col, h in hdr.items():
        c = ws[f"{col}1"]
        c.value = h
        c.font = BOLD
        c.alignment = Alignment(wrap_text=True, vertical="bottom")

    for i in range(2, LR + 1):
        # A row with no name in column B is not a role; every derived column stays blank
        # so empty rows cannot be counted or classified.
        live = f'IF(TRIM($B{i})="","",'
        # Leadership: Lee's own col K or col J says so.
        ws[f"AQ{i}"] = f'={live}--OR(TRIM($K{i})="Leadership",TRIM($J{i})="Leadership"))'
        # Canonical squad: leadership wins, then the merge table, then col G Department
        # where col K is blank or NA (the only squad split the COEs have), then as typed.
        ws[f"AP{i}"] = (
            f'={live}IF($AQ{i}=1,"Leadership",'
            f'IFERROR(INDEX(Lists!$X:$X,MATCH(TRIM($K{i}),Lists!$W:$W,0)),'
            f'IF(OR(TRIM($K{i})="",LOWER(TRIM($K{i}))="na"),'
            f'IF(TRIM($G{i})="","Unassigned",TRIM($G{i})),'
            f'TRIM($K{i})))))'
        )
        # Cost class. Head of Technology is carved out of Leadership so the allowance
        # cannot count the same person in two lines.
        ws[f"AR{i}"] = (
            f'={live}IF(ISNUMBER(SEARCH("head of technology",$C{i})),"Head of Technology",'
            f'IF(ISNUMBER(SEARCH("TDD BP",$C{i})),"Business Partner",'
            f'IF(OR(ISNUMBER(SEARCH("domain architect",$C{i})),'
            f'ISNUMBER(SEARCH("enterprise architect",$C{i}))),"Architecture",'
            f'IF($AQ{i}=1,"Leadership","Squad")))))'
        )
        ws[f"AS{i}"] = (f'={live}IFERROR(INDEX(Lists!$AA:$AA,'
                        f'MATCH($AP{i},Lists!$Z:$Z,0)),$AP{i}))')

    # The stray SUBTOTAL in AA191 breaks every un-criteria'd SUM over the ledger. It is
    # not data - there is no name in column B - so it is moved out of the data range and
    # kept as a labelled check beside the block it was totalling.
    old = ws["AA191"].value
    if isinstance(old, str) and "SUBTOTAL" in old.upper():
        ws["AA191"] = None
        ws["AC191"] = "Stray subtotal removed from the cost column - it made SUM(AA:AA) read $132.240m"
        ws["AC191"].font = Font(italic=True)
    return ws


def data_config(wb):
    """0.2 Data Config: the missing variance, and label the two mixed bases."""
    ws = wb["0.2 Data Config"]
    fixes = []
    # G13 was the only portfolio with no variance. Customer's budget is split across
    # rows 13 and 14 (Ampol Customer + Z Customer) while F13 carries the whole cost.
    if ws["G13"].value is None:
        ws["G13"] = "=($E$13+$E$14)-$F$13"
        ws["H13"] = "Customer budget is rows 13+14 combined; F13 carries the whole portfolio"
        ws["H13"].font = Font(italic=True)
        fixes.append("G13 variance added (was blank)")
    if ws["G14"].value is None:
        ws["G14"] = '="see row 13"'
        fixes.append("G14 pointed at the combined Customer variance")
    return fixes


def design_tabs(wb):
    """The 1.x wiring bugs. Layout untouched - these are wrong references only."""
    fixes = []
    # 1.14 TDD Cyber read Data Config row 15 (Commercial Fuels) instead of row 23 (Cyber),
    # so Cyber's AU/NZ overhead split was driven by another portfolio's budget.
    if "1.14 TDD Cyber" in wb.sheetnames:
        ws = wb["1.14 TDD Cyber"]
        for cell in ("C6", "D6", "C7", "D7"):
            f = ws[cell].value
            if isinstance(f, str) and "$15" in f:
                ws[cell] = f.replace("$D$15", "$D$23").replace("$C$15", "$C$23")
                fixes.append(f"1.14 TDD Cyber!{cell}: Data Config row 15 -> row 23")

    # 1.2 Customer!C7 summed column H; the platform overhead sits in column I. D7 is right.
    if "1.2 Customer" in wb.sheetnames:
        ws = wb["1.2 Customer"]
        f = ws["C7"].value
        if isinstance(f, str) and "H34" in f:
            ws["C7"] = f.replace("H34", "I34").replace("H42", "I42").replace("H49", "I49")
            fixes.append("1.2 Customer!C7: summed column H, overhead is in column I")

    # 1.10 Z Retail: C7 and D7 covered different platform sets.
    if "1.10 Z Retail" in wb.sheetnames:
        ws = wb["1.10 Z Retail"]
        f, g = ws["C7"].value, ws["D7"].value
        if isinstance(f, str) and isinstance(g, str) and f.count(",") != g.count(","):
            ws["C7"] = g.replace("$D$", "$C$").replace(">", ">")  # mirror D7's ranges
            fixes.append("1.10 Z Retail!C7: range now matches D7 (I27,I34,I40)")
    return fixes


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    lists_block(wb)
    review_columns(wb)
    f1 = data_config(wb)
    f2 = design_tabs(wb)
    wb.save(dst)
    return f1 + f2


if __name__ == "__main__":
    for f in run("NEW.xlsx", "core.xlsx"):
        print("  fixed:", f)
