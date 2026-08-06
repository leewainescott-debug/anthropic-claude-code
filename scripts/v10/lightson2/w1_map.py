#!/usr/bin/env python3
"""w1 - spec Stage w1: his master mapping, and every 2.x block re-homed to it.

  python3 w1_map.py <in.xlsx> <out.xlsx>

His 5 Aug master file is the truth for every raw fact. The REVIEW raw block is
replaced with his 29 columns x 526 rows, header verbatim, values only, including
the duplicate EE Number column, the MyHR ee no column and the three cells he
typed as #N/A. The model's own derivations are rebuilt as a helper block to the
right of his data, from column AE, Role ID last, and every consumer in the book
is repointed at the new column positions.

The standing rulings ride on top as model logic, never as edits to his cells:
part time prices at Full Cost x FTE through a new effective cost helper, Holds
price at zero through the levers, the agreed person moves stay on Lists.

Then every 2.x FTE block re-homes: people join, leave and change tab, so rows
are inserted and deleted with shift_rows, which repoints every reference in the
workbook, the grid group ranges and the static SUM ranges are rebuilt from the
new layout, and 3.3 is rewritten to mirror it. Levers carry person-keyed.

Nothing is keyed on a row number: every lookup is by header, label or person.

Idempotent: handed its own output it copies it through untouched.
"""
import sys, os, re, shutil, collections

sys.path.insert(0, "/home/user/anthropic-claude-code/scripts/v10")
sys.path.insert(0, "/home/user/anthropic-claude-code/scripts/v10/update")

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter as gl
from openpyxl.utils import column_index_from_string as ci
from openpyxl.formula.translate import Translator

import wbio
from _xl import (REVIEW, Log, load, save, white, copy_style, row_style,
                 shift_rows, map_formulas, rewrite_refs, set_dv)

HIS = ("/root/.claude/uploads/e550b440-3996-5abb-87e5-bafafe598f82/"
       "0ad63df5-updates.xlsx")

Q = "'" + REVIEW + "'!"
LAST = 700                                  # every REVIEW range runs to row 700
GREY = "FFA6A6A6"

# ------------------------------------------------------------------ his schema
# his 29 raw columns, in his order, by header - read at runtime, never assumed
RAW_N = 29
HIS_HDR_HINT = ("EE Number", "Name", "EE Number", "Position Title")

# the helper block, left to right, immediately right of a one column spacer
HELPERS = ["Effective cost (AUD)", "MStatus", "Ring fenced", "MTab",
           "Leadership", "Squad (canonical, from col L)", "Overhead line",
           "Squad or overhead line", "EGI funded", "Status (PCM)",
           "Commentry (PCM)", "End date (PCM)", "Role ID"]

# his raw columns the model reads, by his header
HIS_READ = {"Name": None, "Position Title": None, "Division (GM)": None,
            "Department (GM-1)": None, "Portfolio": None, "Platform": None,
            "Squad": None, "Country": None, "FTE": None, "Type": None,
            "Full Cost \nAUD": None}

OVERHEAD_ORDER = ["Head of Technology", "Business Partner", "Domain Architect",
                  "Delivery Manager", "Technology Manager", "Program Management"]
STD_OVERHEAD = ["Head of Technology", "Delivery Manager", "Technology Manager"]

SECTIONS = [("Squads", "Archetype"),
            ("Directly funded programs and platforms", "Directly funded"),
            ("No archetype in 1.x tabs", "No figure to compare"),
            ("Overhead roles", "Overhead")]
SEC_TOTAL = {"Squads": "Squads total",
             "Directly funded programs and platforms": "Directly funded total",
             "No archetype in 1.x tabs": "No archetype total",
             "Overhead roles": "Overhead roles total"}
SEC_NAMES = [s for s, _ in SECTIONS]

# the division fallback, his words on the left
DIV_FALLBACK = [("Strategy, Architecture & Data", "COE SA&D"),
                ("Cyber, Risk & Operations", "COE Cyber"),
                ("Partnering & Transformation", "COE BP&T"),
                ("TDD Group Functions", "TDD Group Functions"),
                ("Customer", "Customer"),
                ("Ampol Retail & Z", "Ampol Retail"),
                ("TDD CFE&I", "Infrastructure"),
                ("EGI", "EGI")]
# his portfolio values that must map, spec's table
PORT_MAP = [("Retail", "Ampol Retail"), ("RETAIL", "Ampol Retail"),
            ("Z", "Z Retail"), ("Ampol Customer", "Customer"),
            ("Z ENERGY (DIGITAL)", "Customer"),
            ("Commercial Fuels", "Commercial Fuels"),
            ("B2B & Energy Solutions", "Energy Solutions & B2B"),
            ("Infrastructure", "Infrastructure"),
            ("Enterprise Data", "Enterprise Data"), ("Finance", "Finance"),
            ("P&C", "P&C"), ("P&C, Finance & Legal", "P&C"),
            ("TDD", "TDD Group Functions"), ("EGI", "EGI"),
            ("EGI Integration", "EGI")]
TDD_CYBER_SQUADS = ["Cyber Uplift", "Identity"]
# the overhead line his data alone cannot settle, keyed on the person
OH_OVERRIDES = [("Ed Tacey", "Head of AI enablement", "Squad"),
                ("Vacant", "Delivery Excellence Manager", "Delivery Manager"),
                ("Vacant", "Delivery Assurance Manager", "Delivery Manager")]

FAILS = []


def stop(msg):
    print("STOP: %s" % msg, flush=True)
    raise SystemExit(2)


def nz(v):
    return "" if v is None else str(v).strip()


def norm(s):
    return " ".join(nz(s).lower().split())


def squash(s):
    """A key that survives his en dashes, doubled spaces and lost spaces."""
    return re.sub(r"[^a-z0-9]", "", nz(s).lower())


def is_vacancy(name):
    n = norm(name)
    return n.startswith("vacant") or "ring fenced" in n or n == "remove"


def headers(ws, row=1, upto=None):
    out = {}
    for c in range(1, (upto or ws.max_column) + 1):
        h = nz(ws.cell(row, c).value)
        if h:
            out.setdefault(h, c)
    return out


# ----------------------------------------------------------------- his master


def read_his():
    """His 29 columns x 526 rows, values, plus his header row."""
    wb = openpyxl.load_workbook(HIS, data_only=True)
    ws = wb.worksheets[0]
    last = 0
    for r in range(2, ws.max_row + 1):
        if nz(ws.cell(r, 2).value):
            last = r
    hdr = [ws.cell(1, c).value for c in range(1, RAW_N + 1)]
    rows = [[ws.cell(r, c).value for c in range(1, RAW_N + 1)]
            for r in range(2, last + 1)]
    for c, want in enumerate(HIS_HDR_HINT):
        if nz(hdr[c]) != want:
            stop("his column %s reads %r, expected %r - schema moved"
                 % (gl(c + 1), nz(hdr[c]), want))
    wb.close()
    return hdr, rows


# ------------------------------------------------------- the current workbook


def tabs2x(wb):
    return [ws.title for ws in wb.worksheets if ws.title.startswith("2.")]


def fte_hdr(ws):
    for r in range(5, 200):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.endswith(" FTE"):
            return r + 1
    stop("no FTE block on %s" % ws.title)


def block_cols(ws, hdr):
    """How wide the tab's FTE block is, read off its own header row.

    Most tabs stop at G. 2.11 carries two more: the typed Uplift % and the
    charge to the cyber uplift program, and its cost after lever is net of it.
    """
    last = 7
    for c in range(8, 21):
        if nz(ws.cell(hdr, c).value):
            last = c
    return last


def read_block(ws, hdr):
    """[(group, [(role_row, id, lever)])] for the tab's FTE block, in order."""
    out, r = [], hdr + 1
    while r <= ws.max_row + 2:
        b, c, a = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 1).value
        if isinstance(c, str) and c.startswith("=COUNTIF("):
            out.append([nz(b), []])
        elif isinstance(a, str) and re.fullmatch(r"R\d{4}", a):
            if not out:
                stop("%s row %d carries an ID with no group above it"
                     % (ws.title, r))
            out[-1][1].append((r, a, ws.cell(r, 5).value))
        elif b is None and c is None and a is None:
            break
        else:
            break
        r += 1
    return out, r - 1


def grid_map(ws, hdr):
    """The tab's grid: section -> [(row, group)], plus the landmark rows."""
    sec, cur, marks = collections.OrderedDict(), None, {}
    for r in range(6, hdr - 1):
        b = ws.cell(r, 2).value
        if not isinstance(b, str) or not b.strip():
            continue
        b = b.strip()
        if b in SEC_NAMES:
            cur = b
            sec[b] = {"header": r, "rows": []}
            continue
        if b in SEC_TOTAL.values():
            if cur:
                sec[cur]["total"] = r
            continue
        if b.startswith("Portfolio overhead roles"):
            marks["pfo"] = r
            cur = None
            continue
        if b == "Total portfolio":
            marks["total"] = r
            cur = None
            continue
        if b.startswith("Control - ") or b.startswith("Business Partners and"):
            cur = None
            continue
        if b == "Funding" or "allocation" in b.lower() or b.startswith("Total budget"):
            cur = None
            continue
        if cur:
            sec[cur]["rows"].append((r, b))
    if "total" not in marks:
        stop("%s has no 'Total portfolio' row" % ws.title)
    return sec, marks


def one_x_squads(ws, sec):
    """The squad names the tab's 1.x partner knows, read off a grid formula."""
    for s in ("Squads", "Directly funded programs and platforms"):
        for r, _ in sec.get(s, {}).get("rows", []):
            f = ws.cell(r, 3).value
            if isinstance(f, str):
                m = re.search(r"MATCH\(\$B\d+,'([^']+)'!\$B\$(\d+):\$B\$(\d+)", f)
                if m:
                    return m.group(1), int(m.group(2)), int(m.group(3))
    return None


# -------------------------------------------------------------------- the map


def build_maps(wb, log):
    """Lists gets the tables his data needs; return them as python dicts."""
    ls = wb["Lists"]
    drift = []

    # portfolio -> tab
    was = {}
    for r in range(2, 22):
        k = nz(ls.cell(r, ci("T")).value)
        if k:
            was[k] = nz(ls.cell(r, ci("U")).value)
    ls.cell(1, ci("T")).value = "Portfolio (his words)"
    ls.cell(1, ci("U")).value = "Tab"
    slots = [r for r in range(2, 22) if r != 8]     # row 8 is the sheet's blank
    if len(PORT_MAP) > len(slots):
        stop("the portfolio map needs %d rows, Lists T has %d"
             % (len(PORT_MAP), len(slots)))
    for i, (a, b) in enumerate(PORT_MAP):
        ls.cell(slots[i], ci("T")).value = a
        ls.cell(slots[i], ci("U")).value = b
        if a in was and was[a] != b:
            drift.append("portfolio %r homed %r, now %r" % (a, was[a], b))
    for a in was:
        if a not in [x for x, _ in PORT_MAP]:
            drift.append("portfolio %r is no longer a value in his file, "
                         "dropped from the map (homed %r)" % (a, was[a]))
    for r in slots[len(PORT_MAP):]:
        ls.cell(r, ci("T")).value = None
        ls.cell(r, ci("U")).value = None

    # division fallback, a new table
    ls.cell(1, ci("Q")).value = "Division (GM)"
    ls.cell(1, ci("R")).value = "Tab when the Portfolio is NA or blank"
    for i, (a, b) in enumerate(DIV_FALLBACK):
        ls.cell(2 + i, ci("Q")).value = a
        ls.cell(2 + i, ci("R")).value = b

    # the name his 1.x tab gives each portfolio's EGI platform line
    egilab = egi_labels(wb)
    ls.cell(1, ci("Y")).value = "Portfolio"
    ls.cell(1, ci("Z")).value = "EGI row on the grid"
    for i, (k, v) in enumerate(sorted(egilab.items())):
        ls.cell(2 + i, ci("Y")).value = k
        ls.cell(2 + i, ci("Z")).value = v

    # the TDD Cyber squads, whatever their portfolio says
    ls.cell(1, ci("AR")).value = "Squads that home to TDD Cyber"
    for i, s in enumerate(TDD_CYBER_SQUADS):
        ls.cell(2 + i, ci("AR")).value = s

    # the overhead line overrides, aligned with the person table
    ls.cell(1, ci("AQ")).value = "Overhead line override"
    keys = {}
    free = []
    for r in range(2, 23):
        k = nz(ls.cell(r, ci("AN")).value)
        if k:
            keys[k] = r
        elif r != 8:
            free.append(r)
    for name, title, line in OH_OVERRIDES:
        k = name + " | " + title
        r = keys.get(k)
        if r is None:
            if not free:
                stop("no free row on the Lists person table for %r" % k)
            r = free.pop(0)
            ls.cell(r, ci("AN")).value = k
            keys[k] = r
        ls.cell(r, ci("AQ")).value = line

    pmap = {a.lower(): b for a, b in PORT_MAP}
    dmap = {a.lower(): b for a, b in DIV_FALLBACK}
    canon = {}
    for r in range(2, 22):
        a = nz(ls.cell(r, ci("W")).value)
        if a:
            canon[a] = nz(ls.cell(r, ci("X")).value)
    pov, sov, oov = {}, {}, {}
    for r in range(2, 23):
        k = nz(ls.cell(r, ci("AN")).value)
        if not k:
            continue
        if nz(ls.cell(r, ci("AO")).value):
            pov[k] = nz(ls.cell(r, ci("AO")).value)
        if nz(ls.cell(r, ci("AP")).value):
            sov[k] = nz(ls.cell(r, ci("AP")).value)
        if nz(ls.cell(r, ci("AQ")).value):
            oov[k] = nz(ls.cell(r, ci("AQ")).value)
    ohl = set(nz(ls.cell(r, ci("AF")).value) for r in range(2, 9))
    ohl.discard("")
    funded = [nz(ls.cell(r, ci("AU")).value) for r in range(2, 11)]
    funded = [f for f in funded if f]
    tenp = [nz(ls.cell(r, ci("AS")).value) for r in range(2, 13)]
    tenp = [p for p in tenp if p]
    return dict(pmap=pmap, dmap=dmap, canon=canon, pov=pov, sov=sov, oov=oov,
                ohl=ohl, funded=funded, tenp=tenp, egilab=egilab, drift=drift,
                keys=keys)


def egi_labels(wb):
    """portfolio -> the name his 1.x tab gives its EGI platform line.

    '1.3 Enterprise Data' calls it EGI Ent Data. A portfolio whose 1.x tab has
    no such line falls back to EGI and the portfolio's own name.
    """
    out = {}
    for t in [ws.title for ws in wb.worksheets if ws.title.startswith("1.")]:
        ws = wb[t]
        pair = None
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    m = re.search(r"'(2\.[0-9]+[^']*)'!", c.value)
                    if m:
                        pair = m.group(1)
                        break
            if pair:
                break
        if not pair:
            continue
        key = nz(wb[pair]["C3"].value)
        for r in range(1, ws.max_row + 1):
            if not nz(ws.cell(r, 2).value).startswith("Platform: EGI"):
                continue
            for rr in range(r + 1, min(r + 6, ws.max_row + 1)):
                b = nz(ws.cell(rr, 2).value)
                h = ws.cell(rr, 8).value
                if b and not b.endswith("Total") and isinstance(h, str) \
                        and h.startswith("="):
                    out[key] = b
                    break
            break
    return out


def egi_group(mtab, ap, M):
    """The grid row an EGI funded person belongs in.

    Someone whose squad already names EGI keeps it. Someone his Platform column
    marks EGI but whose squad does not - the Enterprise Data eight - joins the
    portfolio's own EGI row, so no grid row is part funded and part not.
    """
    if ap[:4].upper() == "EGI " or ap.upper() == "EGI":
        return None
    return M["egilab"].get(mtab, "EGI " + mtab)


def egi_home(ap, M):
    """A funded squad named after a portfolio homes to that portfolio.

    'EGI Customer' belongs to Customer, netted out of what Customer carries,
    not lumped onto the EGI tab. Plain 'EGI', or an EGI squad naming no
    portfolio the model knows, stays on the EGI tab.
    """
    if not ap[:4].upper() == "EGI " or len(ap) < 5:
        return None
    suf = ap[4:].strip()
    if suf in M["tenp"]:
        return suf
    return M["pmap"].get(suf.lower())


def derive(rows, cols, M):
    """Every derived value for every one of his rows, exactly as the sheet will."""
    out, fallback = [], []
    for i, row in enumerate(rows):
        def g(h):
            return nz(row[cols[h] - 1])
        name, title = g("Name"), g("Position Title")
        key = name + " | " + title
        squad, plat, dept = g("Squad"), g("Platform"), g("Department (GM-1)")
        port, div = g("Portfolio"), g("Division (GM)")
        lead = 1 if (squad == "Leadership" or plat == "Leadership") else 0
        if key in M["sov"]:
            ap = M["sov"][key]
        elif lead:
            ap = "Leadership"
        elif squad in M["canon"]:
            ap = M["canon"][squad]
        elif squad == "" or squad.lower() == "na":
            ap = dept if dept else "Unassigned"
        else:
            ap = squad
        tl = title.lower()
        if key in M["oov"]:
            ar = M["oov"][key]
        elif ap in M["ohl"]:
            ar = ap
        elif "head of " in tl:
            ar = "Head of Technology"
        elif "tdd bp" in tl:
            ar = "Business Partner"
        elif "domain architect" in tl or "enterprise architect" in tl:
            ar = "Domain Architect"
        elif "delivery man" in tl:
            ar = "Delivery Manager"
        elif any(k in tl for k in ("technology manager", "technology manger",
                                   "tech manager")):
            ar = "Technology Manager"
        else:
            ar = "Squad"
        src = "portfolio"
        eh = egi_home(ap, M)
        if key in M["pov"]:
            mt, src = M["pov"][key], "person (agreed move)"
        elif ap in TDD_CYBER_SQUADS:
            mt, src = "TDD Cyber", "TDD Cyber squad"
        elif eh:
            mt, src = eh, "funded squad names its portfolio"
        elif port == "" or port.upper() == "NA":
            mt = M["dmap"].get(div.lower(), "")
            src = "division fallback"
            fallback.append((i, name, title, port, div, plat, squad, mt))
        elif port.lower() in M["pmap"]:
            mt = M["pmap"][port.lower()]
        else:
            mt = M["dmap"].get(div.lower(), "")
            src = "division fallback (portfolio not on the map)"
            fallback.append((i, name, title, port, div, plat, squad, mt))
        eg = (egi_group(mt, ap, M) if plat.strip().upper() == "EGI" else None)
        if eg:                          # funded outside, but his squad says nothing
            at = eg
        elif ap in M["funded"]:         # a funded squad keeps its own identity
            at = ap
        elif mt.startswith("COE") or mt == "EGI":
            at = ap
        else:
            at = ar if ar != "Squad" else ap
        try:
            fte = float(g("FTE"))
        except ValueError:
            fte = 0.0
        cost = row[cols["Full Cost \nAUD"] - 1]
        cost = float(cost) if isinstance(cost, (int, float)) else 0.0
        out.append(dict(i=i, rid="R%04d" % (i + 1), name=name, title=title,
                        key=key, mtab=mt, at=at, ar=ar, ap=ap, src=src, plat=plat,
                        status="Vacant" if is_vacancy(name) else "Filled",
                        ring=1 if "ring fenced" in norm(name) else 0,
                        fte=fte, cost=cost,
                        eff=cost * fte if 0 < fte < 1 else cost,
                        country=g("Country")))
    return out, fallback


# ------------------------------------------------------------- REVIEW rewrite


def col_remap(cols, hcol):
    """old REVIEW column letter -> new column letter, one simultaneous pass."""
    m = {}
    old = ["EE Number", "Name", "Position Title", "Reports to name",
           "Reports to Position", "Division (GM)", "Department (GM-1)",
           "Team (GM-2)", "Portfolio", "Platform", "Squad", "CC", "Country",
           "Job Level", "FTE", None, "Type", "Unit", "day rate",
           "FTE Base\n Local", "FTE Base \nAUD", "STI", "Payroll", "Pensions",
           "medical", "CPI", "Full Cost \nAUD", "MyHR ee no"]
    for i, h in enumerate(old):
        src = gl(i + 1)
        if h is None:                       # his 'Vacant Column' sits here
            m[src] = gl(cols["Vacant Column"])
        elif h == "Full Cost \nAUD":
            m[src] = gl(hcol["Effective cost (AUD)"])
        else:
            m[src] = gl(cols[h])
    m["AJ"] = gl(hcol["MTab"])
    m["AK"] = gl(hcol["MStatus"])
    m["AP"] = gl(hcol["Squad (canonical, from col L)"])
    m["AQ"] = gl(hcol["Leadership"])
    m["AR"] = gl(hcol["Overhead line"])
    m["AT"] = gl(hcol["Squad or overhead line"])
    m["AW"] = gl(hcol["Status (PCM)"])
    m["AX"] = gl(hcol["Commentry (PCM)"])
    m["AY"] = gl(hcol["End date (PCM)"])
    m["AZ"] = gl(hcol["Role ID"])
    return m


def repoint_columns(wb, m):
    def fn(sh, c1, r1, c2, r2):
        if sh != REVIEW:
            return (c1, r1, c2, r2)

        def one(c):
            if c is None:
                return None
            d = "$" if c.startswith("$") else ""
            return d + m.get(c.lstrip("$"), c.lstrip("$"))
        return (one(c1), r1, one(c2), r2)

    return map_formulas(wb, lambda s, coord, f: rewrite_refs(f, s, fn))


def write_review(wb, hdr, rows, hcol):
    """His raw block, then the helper block, then nothing else."""
    rv = wb[REVIEW]
    # the book's own look for a column, keyed on the header it carried
    style = {}
    for c in range(1, rv.max_column + 1):
        h = nz(rv.cell(1, c).value)
        if h:
            style[h] = (rv.cell(2, c)._style, rv.cell(2, c).number_format,
                        rv.cell(1, c)._style)
    old_max_col = rv.max_column
    for r in range(1, max(rv.max_row, LAST) + 1):
        for c in range(1, old_max_col + 1):
            rv.cell(r, c).value = None

    n = len(rows)
    for i, h in enumerate(hdr):
        rv.cell(1, i + 1).value = h
        st = style.get(nz(h))
        if st:
            rv.cell(1, i + 1)._style = st[2]
    for h in HELPERS:
        rv.cell(1, hcol[h]).value = h
    for i, row in enumerate(rows):
        for c in range(RAW_N):
            cell = rv.cell(2 + i, c + 1)
            cell.value = row[c]
            st = style.get(nz(hdr[c]))
            if st:
                cell._style = st[0]
                cell.number_format = st[1]
    return n


def helper_formulas(hcol, cols, n):
    """The helper block's formulas for data row n (a sheet row number)."""
    def C(h):
        return "$" + gl(cols[h]) + str(n)

    def H(h):
        return "$" + gl(hcol[h]) + str(n)
    name, title = C("Name"), C("Position Title")
    key = "TRIM(%s)&\" | \"&TRIM(%s)" % (name, title)
    guard = 'IF(TRIM(%s)="","",' % name
    ov = ('IF(COUNTIFS(Lists!$AN$2:$AN$22,{k},Lists!${t}$2:${t}$22,"<>"),'
          'INDEX(Lists!${t}$2:${t}$22,MATCH({k},Lists!$AN$2:$AN$22,0)),{els})')
    f = {}
    f["Effective cost (AUD)"] = (
        guard + 'IF(AND(N(%s)>0,N(%s)<1),N(%s)*N(%s),N(%s)))'
        % (C("FTE"), C("FTE"), C("Full Cost \nAUD"), C("FTE"),
           C("Full Cost \nAUD")))
    f["MStatus"] = (
        guard + 'IF(OR(ISNUMBER(SEARCH("vacant",%s)),ISNUMBER(SEARCH('
        '"ring fenced",%s)),LOWER(TRIM(%s))="remove"),"Vacant","Filled"))'
        % (name, name, name))
    f["Ring fenced"] = guard + '--ISNUMBER(SEARCH("ring fenced",%s)))' % name
    f["EGI funded"] = guard + '--(UPPER(TRIM(%s))="EGI"))' % C("Platform")
    f["Leadership"] = (guard + '--OR(TRIM(%s)="Leadership",TRIM(%s)="Leadership"))'
                       % (C("Squad"), C("Platform")))
    f["Squad (canonical, from col L)"] = guard + ov.format(
        k=key, t="AP", els=(
            'IF(%s=1,"Leadership",IFERROR(INDEX(Lists!$X$2:$X$21,MATCH(TRIM(%s),'
            'Lists!$W$2:$W$21,0)),IF(OR(TRIM(%s)="",LOWER(TRIM(%s))="na"),'
            'IF(TRIM(%s)="","Unassigned",TRIM(%s)),TRIM(%s))))'
            % (H("Leadership"), C("Squad"), C("Squad"), C("Squad"),
               C("Department (GM-1)"), C("Department (GM-1)"), C("Squad")))) + ")"
    div = ('IFERROR(INDEX(Lists!$R$2:$R$9,MATCH(TRIM(%s),Lists!$Q$2:$Q$9,0)),'
           '"UNMAPPED")' % C("Division (GM)"))
    port = ('IF(OR(TRIM(%s)="",UPPER(TRIM(%s))="NA"),%s,IFERROR(INDEX('
            'Lists!$U$2:$U$21,MATCH(TRIM(%s),Lists!$T$2:$T$21,0)),%s))'
            % (C("Portfolio"), C("Portfolio"), div, C("Portfolio"), div))
    # a funded squad named after a portfolio homes to that portfolio
    sq = H("Squad (canonical, from col L)")
    suf = "TRIM(MID(%s,5,99))" % sq
    egi = ('IF(AND(LEFT(%s,4)="EGI ",ISNUMBER(MATCH(%s,Lists!$AS$2:$AS$12,0))),'
           '%s,IF(AND(LEFT(%s,4)="EGI ",ISNUMBER(MATCH(%s,Lists!$T$2:$T$21,0))),'
           'INDEX(Lists!$U$2:$U$21,MATCH(%s,Lists!$T$2:$T$21,0)),%s))'
           % (sq, suf, suf, sq, suf, suf, port))
    f["MTab"] = guard + ov.format(
        k=key, t="AO",
        els='IF(COUNTIF(Lists!$AR$2:$AR$3,%s),"TDD Cyber",%s)' % (sq, egi)) + ")"
    f["Overhead line"] = guard + ov.format(
        k=key, t="AQ", els=(
            'IF(COUNTIF(Lists!$AF$2:$AF$8,%s),%s,'
            'IF(ISNUMBER(SEARCH("head of ",%s)),"Head of Technology",'
            'IF(ISNUMBER(SEARCH("TDD BP",%s)),"Business Partner",'
            'IF(OR(ISNUMBER(SEARCH("domain architect",%s)),ISNUMBER(SEARCH('
            '"enterprise architect",%s))),"Domain Architect",'
            'IF(ISNUMBER(SEARCH("delivery man",%s)),"Delivery Manager",'
            'IF(OR(ISNUMBER(SEARCH("technology manager",%s)),ISNUMBER(SEARCH('
            '"technology manger",%s)),ISNUMBER(SEARCH("tech manager",%s))),'
            '"Technology Manager","Squad"))))))'
            % (H("Squad (canonical, from col L)"),
               H("Squad (canonical, from col L)"),
               title, title, title, title, title, title, title, title))) + ")"
    egirow = ('IFERROR(INDEX(Lists!$Z$2:$Z$20,MATCH(%s,Lists!$Y$2:$Y$20,0)),'
              '"EGI "&%s)' % (H("MTab"), H("MTab")))
    f["Squad or overhead line"] = (
        guard + 'IF(AND(%s=1,LEFT(%s,4)<>"EGI ",%s<>"EGI"),%s,'
        'IF(COUNTIF(Lists!$AU$2:$AU$10,%s),%s,'
        'IF(OR(LEFT(%s,3)="COE",%s="EGI"),%s,IF(%s<>"Squad",%s,%s)))))'
        % (H("EGI funded"), sq, sq, egirow, sq, sq, H("MTab"), H("MTab"), sq,
           H("Overhead line"), H("Overhead line"), sq))
    return f


# ------------------------------------------------------------- lever carrying


def carry_levers(before, new, log):
    """Old lever states follow the person; new people take the default."""
    old = []
    for tab, groups in before.items():
        for gname, roles in groups:
            for _, rid, lever, name, title, status, typed in roles:
                old.append(dict(tab=tab, name=name, title=title, lever=lever,
                                status=status, typed=typed, used=False))
    got = {}
    pending = list(new)

    def pass_(label, side, mine, keyf):
        """side picks the old rows, mine the new rows, keyf builds the key."""
        buckets = collections.defaultdict(list)
        for p in old:
            if not p["used"] and side(p):
                buckets[keyf(p)].append(p)
        rest = []
        for p in pending:
            b = buckets.get(keyf(p)) if mine(p) else None
            if b:
                src = b.pop(0)
                src["used"] = True
                got[p["rid"]] = (src, label)
            else:
                rest.append(p)
        return rest

    named = lambda p: p["status"] != "Vacant"
    vacant = lambda p: p["status"] == "Vacant"
    # a named person follows their own name, then their name alone
    pending = pass_("name and role", named, named,
                    lambda p: (squash(p["name"]), squash(p["title"])))
    pending = pass_("name", named, named, lambda p: (squash(p["name"]),))
    # a vacancy has no name, so it follows its role, on its own tab first
    pending = pass_("vacancy, same role on the same tab", vacant, vacant,
                    lambda p: (p["tab"], squash(p["title"])))
    pending = pass_("vacancy, same role", vacant, vacant,
                    lambda p: (squash(p["title"]),))
    # a vacancy that became a named person keeps the vacancy's lever slot
    pending = pass_("the vacancy this person fills", vacant, named,
                    lambda p: (p["tab"], squash(p["title"])))
    pending = pass_("the vacancy this person fills", vacant, named,
                    lambda p: (squash(p["title"]),))

    kept, defaulted, filled = [], [], []
    for p in new:
        hit = got.get(p["rid"])
        if hit is None:
            p["lever"] = "Hire" if p["status"] == "Vacant" else "Filled"
            defaulted.append(p)
            continue
        src, how = hit
        p["typed"] = src.get("typed", {})
        if src["status"] == "Vacant" and p["status"] == "Filled":
            p["lever"] = "Filled"
            filled.append((p, src, how))
        else:
            p["lever"] = src["lever"] if src["lever"] in (
                "Filled", "Hire", "Hold", "Offshore") else (
                "Hire" if p["status"] == "Vacant" else "Filled")
            kept.append((p, src, how))
    departed = [p for p in old if not p["used"]]
    return kept, defaulted, filled, departed


# --------------------------------------------------------------- 2.x rebuild


def block_ranges(layout, first):
    """group -> (first_role_row, last_role_row) for a laid out block."""
    out, r = {}, first
    for gname, roles in layout:
        out[gname] = (r + 1, r + len(roles))
        r += 1 + len(roles)
    return out, r - 1


def egi_slice(key_cell, group_cell):
    """The EGI funded slice of one grid group, straight off his Platform column."""
    return ("SUMIFS('{v}'!${e}$2:${e}${L},'{v}'!${m}$2:${m}${L},{k},'{v}'!${g}$2"
            ":${g}${L},{b},'{v}'!${p}$2:${p}${L},1)/1000000").format(
                v=REVIEW, e=EFF_COL, m=MTAB_COL, g=GROUP_COL, p=PLAT_COL,
                L=LAST, k=key_cell, b=group_cell)


def write_funded_outside(ws, r):
    """What the row costs that TDD does not carry: a flat programme number if
    Lists sets one, otherwise the row's own EGI funded people."""
    sl = egi_slice("$C$3", "$B%d" % r)
    look = ("INDEX(Lists!$AW$2:$AW$10,MATCH($B%d,Lists!$AU$2:$AU$10,0))" % r)
    ws.cell(r, 16).value = ("=IFERROR(IF(ISNUMBER(%s),%s,%s),%s)"
                            % (look, look, sl, sl))


def write_grid_group(ws, r, first, last, empty):
    write_funded_outside(ws, r)
    if empty:
        for c, v in ((6, 0), (8, 0), (9, 0), (10, 0), (11, 0), (12, 0),
                     (13, 0), (19, 0)):
            ws.cell(r, c).value = v
        ws.cell(r, 13).value = "=$F%d-$L%d" % (r, r)
        return
    a, b = first, last
    ws.cell(r, 6).value = '=COUNTIF($B$%d:$B$%d,"?*")' % (a, b)
    ws.cell(r, 8).value = '=COUNTIFS($D$%d:$D$%d,"Filled")' % (a, b)
    ws.cell(r, 9).value = '=COUNTIFS($D$%d:$D$%d,"Vacant")' % (a, b)
    ws.cell(r, 10).value = ('=COUNTIFS($D$%d:$D$%d,"Vacant",$E$%d:$E$%d,"Hire")'
                            % (a, b, a, b))
    ws.cell(r, 11).value = '=COUNTIFS($E$%d:$E$%d,"Offshore")' % (a, b)
    ws.cell(r, 12).value = '=COUNTIFS($E$%d:$E$%d,"Hold")' % (a, b)
    ws.cell(r, 13).value = "=$F%d-$L%d" % (r, r)
    ws.cell(r, 19).value = "=SUM($G$%d:$G$%d)/1000000" % (a, b)


def write_totals(ws, rows, row, kind, count_n=None):
    """A section total or the tab total over `rows` (a list of (first,last))."""
    def span(col):
        return ",".join("%s%d:%s%d" % (col, a, col, b) for a, b in rows)
    n = sum(b - a + 1 for a, b in rows)
    for c in list(range(5, 14)) + [15, 16, 17]:
        ws.cell(row, c).value = "=SUM(%s)" % span(gl(c))
    ws.cell(row, 14).value = ('=IF(COUNT(%s)=0,"",SUM(%s))'
                              % (span("N"), span("N")))
    ws.cell(row, 19).value = "=SUM(%s)" % span("S")
    if kind == "section":
        k = count_n if count_n is not None else n
        ws.cell(row, 18).value = ('=IF(COUNT(%s)=%d,ROUND($O%d-$N%d,6),"")'
                                  % (span("N"), k, row, row))
        ws.cell(row, 20).value = ('=IF(COUNT(%s)=%d,ROUND($S%d-$N%d,6),"")'
                                  % (span("N"), k, row, row))
    else:
        ws.cell(row, 18).value = ('=IF(ISNUMBER($N%d),ROUND($O%d-$N%d,6),"")'
                                  % (row, row, row))
        ws.cell(row, 20).value = ('=IF(ISNUMBER($N%d),ROUND($S%d-$N%d,6),"")'
                                  % (row, row, row))


def rehome_tab(wb, title, want, styles, log):
    """One 2.x tab: grid rows added or zeroed, block rebuilt, ranges rewired."""
    ws = wb[title]
    hdr = fte_hdr(ws)
    sec, marks = grid_map(ws, hdr)
    have = {g for s in sec for _, g in sec[s]["rows"]}
    one_x = one_x_squads(ws, sec)
    known1x = set()
    if one_x:
        t1, a1, b1 = one_x
        known1x = set(nz(wb[t1].cell(r, 2).value) for r in range(a1, b1 + 1))

    # which section does a group belong in
    def section_for(g):
        for s in sec:
            for _, gg in sec[s]["rows"]:
                if gg == g:
                    return s
        if g in FUNDED or g in ALL_EGI:
            return ("Directly funded programs and platforms"
                    if "Directly funded programs and platforms" in sec
                    else list(sec)[0])
        if g in OVERHEAD_ORDER:
            return "Overhead roles" if "Overhead roles" in sec else list(sec)[0]
        if g in known1x and "Squads" in sec:
            return "Squads"
        if "No archetype in 1.x tabs" in sec:
            return "No archetype in 1.x tabs"
        if "Squads" in sec:
            return "No archetype in 1.x tabs"
        return list(sec)[0]

    add = collections.defaultdict(list)
    for g in sorted(want):
        if g not in have:
            add[section_for(g)].append(g)

    # a new section may be needed before the overhead block
    for s in list(add):
        if s not in sec:
            anchor = sec["Overhead roles"]["header"] if "Overhead roles" in sec \
                else marks.get("pfo", marks["total"])
            shift_rows(wb, title, anchor, 1)
            ws.cell(anchor, 2).value = s
            copy_style(ws.cell(sec[list(sec)[0]]["header"], 2), ws.cell(anchor, 2))
            log("grid", title, "section %r added above the overhead roles" % s)
            hdr = fte_hdr(ws)
            sec, marks = grid_map(ws, hdr)

    for s, gs in add.items():
        for g in gs:
            rows = sec[s]["rows"]
            at = (rows[-1][0] + 1) if rows else (sec[s]["header"] + 1)
            tpl = rows[-1][0] if rows else None
            if tpl is None:                     # a section with nothing to copy
                for other in sec:
                    if sec[other]["rows"]:
                        tpl = sec[other]["rows"][0][0]
                        break
            if tpl is None:
                stop("%s has no grid row to model %r on" % (title, g))
            shift_rows(wb, title, at, 1)
            if tpl >= at:
                tpl += 1
            for c in range(2, 21):
                src = ws.cell(tpl, c)
                copy_style(src, ws.cell(at, c))
                v = src.value
                if isinstance(v, str) and v.startswith("="):
                    ws.cell(at, c).value = Translator(
                        v, origin="%s%d" % (gl(c), tpl)).translate_formula(
                            "%s%d" % (gl(c), at))
                elif not isinstance(v, str):
                    ws.cell(at, c).value = v
            ws.cell(at, 2).value = g
            log("grid", "%s!B%d" % (title, at), "%r joins %r" % (g, s))
            hdr = fte_hdr(ws)
            sec, marks = grid_map(ws, hdr)

    # the block, resized group by group so no group header is ever swallowed
    hdr = fte_hdr(ws)
    old_layout, old_last = read_block(ws, hdr)
    first = hdr + 1
    ordered = sorted(g for g in want if g not in OVERHEAD_ORDER)
    ordered += [g for g in OVERHEAD_ORDER if g in want]
    layout = [(g, want[g]) for g in ordered]
    # the tab's own row conventions, copied off its own block before it moves
    bcols = block_cols(ws, hdr)
    tpl_role = old_layout[0][1][0][0]
    tpl_grp = tpl_role - 1
    role_tpl = {c: ws.cell(tpl_role, c).value for c in range(1, bcols + 1)}
    grp_tpl = {c: ws.cell(tpl_grp, c).value for c in range(1, bcols + 1)}
    typed = [c for c in range(2, bcols + 1)
             if c != 5 and not (isinstance(role_tpl[c], str)
                                and str(role_tpl[c]).startswith("="))]
    for c in (2, 3, 4, 6, 7):
        v = role_tpl[c]
        if not (isinstance(v, str) and v.startswith("=")):
            stop("%s row %d column %s is not a formula - no role row to model on"
                 % (title, tpl_role, gl(c)))
    if RID_COL not in str(role_tpl[2]) or REVIEW not in str(role_tpl[2]):
        stop("%s row %d does not read the role mapping by Role ID"
             % (title, tpl_role))
    if bcols > 7:
        log("block", title, "block runs to column %s; every role row is copied "
            "from this tab's own row shape, so cost after lever keeps the "
            "uplift netting and the charge column keeps its own formula"
            % gl(bcols))

    have = old_last - first + 1
    for gname, roles in reversed(old_layout):
        if not roles:
            continue
        if gname not in want:
            shift_rows(wb, title, roles[0][0] - 1, -(1 + len(roles)))
            have -= 1 + len(roles)
            continue
        delta = len(want[gname]) - len(roles)
        if delta < 0:                       # the tail of the group goes
            shift_rows(wb, title, roles[-1][0] + delta + 1, delta)
        elif delta > 0:                     # room before the group's last row
            shift_rows(wb, title, roles[-1][0], delta)
        have += delta
    need = sum(1 + len(v) for _, v in layout)
    if need > have:                         # the groups that are brand new
        shift_rows(wb, title, first + have, need - have)
    elif need < have:
        stop("%s block is %d rows, %d wanted after the per group resize"
             % (title, have, need))
    ranges, last = block_ranges(layout, first)
    log("block", title, "%d role rows in %d groups, was %d in %d"
        % (sum(len(v) for _, v in layout), len(layout),
           sum(len(v) for _, v in old_layout), len(old_layout)))

    ghdr, grole = styles
    r = first
    lever_rows = []
    for gname, roles in layout:
        for c in range(1, bcols + 1):
            ws.cell(r, c)._style = ghdr[c - 1]
            ws.cell(r, c).value = None
        a, b = ranges[gname]
        ws.cell(r, 2).value = gname
        ws.cell(r, 3).value = ('=COUNTIF($B$%d:$B$%d,"?*")&IF(COUNTIF($B$%d:$B$%d'
                               ',"?*")=1," role"," roles")' % (a, b, a, b))
        for c in range(4, bcols + 1):        # F, G and any tab of its own, e.g. I
            v = grp_tpl[c]
            if isinstance(v, str) and v.startswith("=SUM("):
                ws.cell(r, c).value = "=SUM(%s%d:%s%d)" % (gl(c), a, gl(c), b)
        r += 1
        for p in roles:
            for c in range(1, bcols + 1):
                ws.cell(r, c)._style = grole[c - 1]
            ws.cell(r, 1).value = p["rid"]
            f = ws.cell(r, 1).font
            ws.cell(r, 1).font = Font(name=f.name, size=7, color=GREY)
            ws.cell(r, 5).value = p["lever"]
            for c in range(2, bcols + 1):
                if c == 5:
                    continue
                v = role_tpl[c]
                if isinstance(v, str) and v.startswith("="):
                    ws.cell(r, c).value = Translator(
                        v, origin="%s%d" % (gl(c), tpl_role)).translate_formula(
                            "%s%d" % (gl(c), r))
                else:                       # a cell someone types into
                    ws.cell(r, c).value = p.get("typed", {}).get(c, v)
            lever_rows.append(r)
            r += 1
    for rr in range(r, ws.max_row + 2):
        b = ws.cell(rr, 2).value
        if isinstance(b, str) and (b.startswith("Control - ")
                                   or b.startswith("Vacancy levers")):
            break
        for c in range(1, bcols + 1):
            ws.cell(rr, c).value = None

    # the grid reads the block again
    hdr2 = fte_hdr(ws)
    sec, marks = grid_map(ws, hdr2)
    spans = {}
    for s in sec:
        rows = []
        for gr, g in sec[s]["rows"]:
            if g in ranges:
                write_grid_group(ws, gr, ranges[g][0], ranges[g][1], False)
            else:
                write_grid_group(ws, gr, 0, 0, True)
            rows.append((gr, gr))
        spans[s] = rows
        if "total" in sec[s] and rows:
            write_totals(ws, [(rows[0][0], rows[-1][0])], sec[s]["total"],
                         "section", count_n=len(rows))
    parts = []
    for s in sec:
        if spans[s]:
            parts.append((spans[s][0][0], spans[s][-1][0]))
    if "pfo" in marks:
        parts.append((marks["pfo"], marks["pfo"]))
    write_totals(ws, parts, marks["total"], "tab")

    # the lever controls cover the whole block again
    for rr in range(last + 1, last + 12):
        b = ws.cell(rr, 2).value
        if not isinstance(b, str):
            continue
        if b.startswith("Control - every lever"):
            ws.cell(rr, 3).value = (
                '=SUMPRODUCT(($E$%d:$E$%d<>"")*($E$%d:$E$%d<>"Filled")*'
                '($E$%d:$E$%d<>"Hire")*($E$%d:$E$%d<>"Hold")*'
                '($E$%d:$E$%d<>"Offshore"))'
                % ((first, last) * 5))
        if b.startswith("Vacancy levers"):
            for c, (a1, a2) in ((10, ("Vacant", "Hire")),
                                (11, ("Vacant", "Offshore")),
                                (12, ("Vacant", "Hold")),
                                (13, ("Vacant", "Filled")),
                                (14, ("Filled", "Offshore"))):
                ws.cell(rr, c).value = (
                    '=COUNTIFS($D$%d:$D$%d,"%s",$E$%d:$E$%d,"%s")'
                    % (first, last, a1, first, last, a2))
    set_dv(ws, lever_rows)
    place = {}
    rr = first
    for gname, roles in layout:
        rr += 1
        for p in roles:
            place[p["rid"]] = rr
            rr += 1
    return ranges, sec, marks, place


# ------------------------------------------------------------------ 3.3 mirror


def rebuild_33(wb, order, grids, log):
    ws = wb["3.3 Squad Actuals to Archetype"]
    body_first = 6
    gt = None
    for r in range(body_first, ws.max_row + 1):
        if nz(ws.cell(r, 2).value) == "Group total":
            gt = r
            break
    if gt is None:
        stop("3.3 has no Group total row")
    tpl_tot = None
    for r in range(body_first, gt):
        if nz(ws.cell(r, 2).value).endswith(" total"):
            tpl_tot = r
            break
    if tpl_tot is None:
        stop("3.3 has no per tab total row to copy")
    sty_data = [ws.cell(body_first, c)._style for c in range(2, 16)]
    sty_tot = [ws.cell(tpl_tot, c)._style for c in range(2, 16)]
    need = sum(len(g["rows"]) + 1 for g in grids.values())
    have = gt - body_first
    if need != have:
        shift_rows(wb, "3.3 Squad Actuals to Archetype", body_first + 1,
                   need - have)
        log("3.3", "3.3 Squad Actuals to Archetype",
            "%d mirror rows in, %d out, %+d rows" % (need, have, need - have))
        gt += need - have
    r = body_first
    tot_rows = []
    for title in order:
        g = grids[title]
        first = r
        for gr, gname, kind in g["rows"]:
            for c in range(2, 16):
                ws.cell(r, c)._style = sty_data[c - 2]
            ws.cell(r, 2).value = g["label"]
            ws.cell(r, 3).value = kind
            for c, src in ((4, "B"), (5, "C"), (6, "D"), (7, "E"), (8, "F"),
                           (9, "H"), (10, "I"), (11, "M"), (12, "N"), (13, "O"),
                           (14, "R"), (15, "S")):
                ws.cell(r, c).value = "='%s'!$%s$%d" % (title, src, gr)
            r += 1
        for c in range(2, 16):
            ws.cell(r, c)._style = sty_tot[c - 2]
            ws.cell(r, c).value = None
        ws.cell(r, 2).value = g["label"] + " total"
        ws.cell(r, 6).value = '=""'
        for c in (7, 12):
            ws.cell(r, c).value = ('=IF(COUNT(%s%d:%s%d)=0,"",SUM(%s%d:%s%d))'
                                   % ((gl(c), first, gl(c), r - 1) * 2))
        for c in (8, 9, 10, 11, 13, 15):
            ws.cell(r, c).value = "=SUM(%s%d:%s%d)" % (gl(c), first, gl(c), r - 1)
        ws.cell(r, 14).value = ('=IF(ISNUMBER($L%d),ROUND($M%d-$L%d,6),"")'
                                % (r, r, r))
        tot_rows.append(r)
        r += 1
    for c in (7, 12):
        s = ",".join("%s%d" % (gl(c), t) for t in tot_rows)
        ws.cell(gt, c).value = '=IF(COUNT(%s)=0,"",SUM(%s))' % (s, s)
    for c in (8, 9, 10, 11, 13, 15):
        ws.cell(gt, c).value = "=" + "+".join("%s%d" % (gl(c), t)
                                              for t in tot_rows)
    ws.cell(gt, 6).value = '=""'
    ws.cell(gt, 14).value = '=""'
    return gt


# --------------------------------------------------- 3.4 and 3.2 follow suit


def read_34(ws):
    """[(key, first, last, total_row)] for every COE or EGI section, plus the
    grand total row - all read off the sheet's own labels."""
    hdr = None
    for r in range(1, ws.max_row + 1):
        if nz(ws.cell(r, 2).value) == "COE" and nz(ws.cell(r, 3).value) == "Squad":
            hdr = r
            break
    if hdr is None:
        stop("3.4 has no COE and Squad header row")
    sections, cur, grand = [], None, None
    for r in range(hdr + 1, ws.max_row + 1):
        b, c = nz(ws.cell(r, 2).value), nz(ws.cell(r, 3).value)
        if b == "COEs and EGI total":
            grand = r
            break
        if b.endswith(" total"):
            if cur:
                sections.append(cur + [r])
                cur = None
            continue
        if b and c:
            if cur and cur[0] == b:
                cur[2] = r
            else:
                if cur:
                    sections.append(cur + [None])
                cur = [b, r, r]
    if grand is None:
        stop("3.4 has no 'COEs and EGI total' row")
    for s in sections:
        if s[3] is None:
            stop("3.4 section %r has no total row" % s[0])
    return sections, grand


def fix_34(wb, derived, log):
    """Every COE and EGI section lists the squads REVIEW actually carries."""
    t = "3.4 COE Breakdown"
    ws = wb[t]
    live = collections.defaultdict(set)
    for d in derived:
        live[d["mtab"]].add(d["at"])
    sections, grand = read_34(ws)
    n = 0
    for key, first, last, total in reversed(sections):
        want = sorted(live.get(key, []))
        if not want:
            stop("3.4 section %r matches no row on the role mapping" % key)
        have = [nz(ws.cell(r, 3).value) for r in range(first, last + 1)]
        tpl = {c: ws.cell(first, c).value for c in range(4, 12)}
        sty = [ws.cell(first, c)._style for c in range(2, 12)]
        d = len(want) - len(have)
        if d > 0:
            shift_rows(wb, t, last, d)
        elif d < 0:
            shift_rows(wb, t, last + d + 1, d)
        for i, name in enumerate(want):
            r = first + i
            for c in range(2, 12):
                ws.cell(r, c)._style = sty[c - 2]
            ws.cell(r, 2).value = key
            ws.cell(r, 3).value = name
            for c in range(4, 12):
                v = tpl[c]
                if isinstance(v, str) and v.startswith("="):
                    ws.cell(r, c).value = Translator(
                        v, origin="%s%d" % (gl(c), first)).translate_formula(
                            "%s%d" % (gl(c), r))
                else:
                    ws.cell(r, c).value = v
        if have != want:
            n += 1
            log("3.4", "%s!B%d" % (t, first),
                "%s section reads %d squad(s) off the role mapping: %s (was %s)"
                % (key, len(want), ", ".join(want), ", ".join(have)))
    sections, grand = read_34(ws)
    for key, first, last, total in sections:
        for c in range(4, 12):
            ws.cell(total, c).value = "=SUM(%s%d:%s%d)" % (gl(c), first, gl(c),
                                                           last)
    for c in range(4, 12):
        ws.cell(grand, c).value = "=" + "+".join(
            "%s%d" % (gl(c), s[3]) for s in sections)
    log("3.4", "%s!B%d" % (t, grand), "the grand total sums the %d section "
        "totals at rows %s" % (len(sections),
                               ", ".join(str(s[3]) for s in sections)))
    return n


def fix_1x_egi(wb, derived, log):
    """His EGI platform line on a 1.x tab reads the portfolio's EGI people.

    Five of the six join a squad row that carries the whole slice and already
    tie. Enterprise Data's do not - its EGI people sit inside two ordinary
    squads - so that line joined a squad that does not exist and read 0.
    """
    slice_of = collections.defaultdict(float)
    for d in derived:
        if d["plat"].strip().upper() == "EGI":
            slice_of[d["mtab"]] += d["eff"]
    done = []
    for t in [ws.title for ws in wb.worksheets if ws.title.startswith("1.")]:
        ws = wb[t]
        pair = None
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    m = re.search(r"'(2\.[0-9]+[^']*)'!", c.value)
                    if m:
                        pair = m.group(1)
                        break
            if pair:
                break
        if not pair:
            continue
        key = nz(wb[pair]["C3"].value)
        want = slice_of.get(key, 0.0) / 1e6
        for r in range(1, ws.max_row + 1):
            b = nz(ws.cell(r, 2).value)
            if not b.startswith("Platform: EGI"):
                continue
            line = None
            for rr in range(r + 1, min(r + 6, ws.max_row + 1)):
                b2 = nz(ws.cell(rr, 2).value)
                h2 = ws.cell(rr, 8).value
                if b2 and not b2.endswith("Total") and isinstance(h2, str) \
                        and h2.startswith("="):
                    line = rr
                    break
            if line is None:
                continue
            h = ws.cell(line, 8).value          # does it join a live squad row?
            live = False
            if isinstance(h, str):
                m = re.search(r"'(2\.[0-9]+[^']*)'!\$[A-Z]\$(\d+)", h)
                if m:
                    grp = nz(wb[m.group(1)].cell(int(m.group(2)), 2).value)
                    live = bool(grp) and any(
                        d["mtab"] == key and d["at"] == grp
                        and d["plat"].strip().upper() == "EGI" for d in derived)
            if live or want <= 0:
                continue
            label = nz(ws.cell(line, 2).value)
            grid = None                        # the row his own name now has
            g = wb[pair]
            for gr in range(6, fte_hdr(g) - 1):
                if nz(g.cell(gr, 2).value) == label:
                    grid = gr
                    break
            if grid:
                ws.cell(line, 8).value = "='%s'!$O$%d" % (pair, grid)
                ws.cell(line, 11).value = "='%s'!$S$%d" % (pair, grid)
                done.append(t)
                log("1.x", "%s!B%d" % (t, line),
                    "%r joins its own row on %s (%.6f $m of EGI funded people)"
                    % (label, pair, want))
                continue
            hdr = fte_hdr(wb[pair])
            _, last = read_block(wb[pair], hdr)
            ws.cell(line, 8).value = ("=SUMIFS('{v}'!${e}$2:${e}${L},'{v}'!${m}"
                                      "$2:${m}${L},'{p}'!$C$3,'{v}'!${pl}$2:${pl}"
                                      "${L},1)/1000000").format(
                v=REVIEW, e=EFF_COL, m=MTAB_COL, pl=PLAT_COL, L=LAST, p=pair)
            ws.cell(line, 11).value = (
                "=SUMPRODUCT('{p}'!$G${a}:$G${b},--(COUNTIFS('{v}'!${i}$2:${i}"
                "${L},'{p}'!$A${a}:$A${b},'{v}'!${pl}$2:${pl}${L},1)>0))"
                "/1000000").format(p=pair, a=hdr + 1, b=last, v=REVIEW,
                                   i=RID_COL, pl=PLAT_COL, L=LAST)
            done.append(t)
            log("1.x", "%s!B%d" % (t, line),
                "%r joins the %s people his Platform column marks EGI "
                "(%.6f $m), not a squad of that name" % (nz(ws.cell(line, 2).value),
                                                         key, want))
    return done


def fix_32(wb, derived, log):
    """An overhead line with no roles prices no roles."""
    ws = wb["3.2 Overhead & Leadership"]
    live = collections.Counter(d["ar"] for d in derived if d["ar"] != "Squad")
    n = 0
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if not isinstance(b, str) or not b.startswith("=Lists!"):
            continue
        m = re.search(r"\$AF\$(\d+)", b)
        if not m:
            continue
        line = nz(wb["Lists"].cell(int(m.group(1)), ci("AF")).value)
        if line == "Leadership - 8 GMs":       # the GMs sit above the mapping
            continue
        e = ws.cell(r, 5).value
        if live.get(line, 0) == 0 and isinstance(e, (int, float)) and e != 0:
            ws.cell(r, 5).value = 0
            n += 1
            log("3.2", "3.2 Overhead & Leadership!E%d" % r,
                "%s carries no role in his file, so the archetype prices none "
                "(was %s)" % (line, e))
    return n


# ------------------------------------------- the vacant overhead dial on 3.5


def fix_35_dial(wb, derived, placement, log):
    """3.5's vacant overhead analysis names role rows: point it at the new ones."""
    ws = wb["3.5 TDD Lights On"]
    at = {}
    for r in range(1, ws.max_row + 1):
        b = nz(ws.cell(r, 2).value)
        if b:
            at.setdefault(b, r)
    lines = {}
    for d in derived:
        if d["status"] != "Vacant" or d["ar"] == "Squad":
            continue
        lines.setdefault(d["ar"], []).append(
            (d["tab"], placement[d["rid"]], d))
    n = 0
    for line in OVERHEAD_ORDER:
        r = at.get(line)
        if r is None:
            continue
        cells = lines.get(line, [])
        if cells:
            ws.cell(r, 3).value = "=" + "+".join(
                'COUNTIF(\'%s\'!$E$%d,"Hire")' % (t, rr) for t, rr, _ in cells)
            ws.cell(r, 4).value = "=(" + "+".join(
                "'%s'!$G$%d" % (t, rr) for t, rr, _ in cells) + ")/1000000"
        else:
            ws.cell(r, 3).value = "=0"
            ws.cell(r, 4).value = "=0"
        n += 1
        log("3.5", "3.5 TDD Lights On!B%d" % r,
            "%s reads %d vacant role rows" % (line, len(cells)))
    allv = [c for line in lines for c in lines[line]]
    r = at.get("Hold every vacant overhead role and the lights on total drops "
               "by this amount ($m)")
    if r and allv:
        ws.cell(r, 3).value = "=(" + "+".join(
            "'%s'!$G$%d" % (t, rr) for t, rr, _ in allv) + ")/1000000"
    hold = sum(1 for _, _, d in allv if d["lever"] == "Hold")
    fill = sum(1 for _, _, d in allv if d["lever"] == "Filled")
    for b, r in at.items():
        if b.startswith("Already levered:"):
            ws.cell(r, 2).value = ("Already levered: %d on Hold at zero, %d "
                                   "marked to fill" % (hold, fill))
            log("3.5", "3.5 TDD Lights On!B%d" % r, ws.cell(r, 2).value)
    return n, len(allv)


# ---------------------------------------------------------------- self checks


def chk(name, ok, detail=""):
    print("%s  %s%s" % ("PASS" if ok else "FAIL", name,
                        (" - " + detail) if detail else ""), flush=True)
    if not ok:
        FAILS.append(name)


# ---------------------------------------------------------------------- main

NAME_COL = TITLE_COL = STATUS_COL = EFF_COL = COUNTRY_COL = RID_COL = None
MTAB_COL = GROUP_COL = PLAT_COL = None
ALL_EGI = set()
FUNDED = []


def main(src, dst):
    global NAME_COL, TITLE_COL, STATUS_COL, EFF_COL, COUNTRY_COL, RID_COL, FUNDED
    global MTAB_COL, GROUP_COL, PLAT_COL, ALL_EGI
    log = Log("w1_map")
    wb = load(src)
    rv = wb[REVIEW]

    if nz(rv.cell(1, 3).value) == "EE Number" and any(
            nz(rv.cell(1, c).value) == "Effective cost (AUD)"
            for c in range(1, rv.max_column + 1)):
        print("input already carries his master mapping - copying through")
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        return

    # ---------------------------------------------------------------- W1-1
    log.head("W1-1  what the model looks like before his file lands")
    wbv = openpyxl.load_workbook(src, data_only=True)
    before, styles = {}, {}
    for t in tabs2x(wb):
        ws = wb[t]
        h = fte_hdr(ws)
        blk, _ = read_block(ws, h)
        wv = wbv[t]
        bc = block_cols(ws, h)
        grole = blk[0][1][0][0] if blk and blk[0][1] else h + 2
        typed = [c for c in range(2, bc + 1)
                 if c != 5 and not (isinstance(ws.cell(grole, c).value, str)
                                    and str(ws.cell(grole, c).value).startswith("="))]
        rows = []
        for gname, roles in blk:
            rr = []
            for row, rid, lever in roles:
                rr.append((row, rid, lever, nz(wv.cell(row, 2).value),
                           nz(wv.cell(row, 3).value), nz(wv.cell(row, 4).value),
                           {c: ws.cell(row, c).value for c in typed}))
            rows.append((gname, rr))
        before[t] = rows
        if typed:
            log.note("W1-1", "%s carries %d typed block column(s) beyond the "
                     "lever: %s - they follow the person" %
                     (t, len(typed), ", ".join(
                         "%s %r" % (gl(c), nz(ws.cell(h, c).value)) for c in typed)))
        styles[t] = ([ws.cell(h + 1, c)._style for c in range(1, bc + 1)],
                     [ws.cell(grole, c)._style for c in range(1, bc + 1)])
    b_roles = {t: sum(len(r) for _, r in v) for t, v in before.items()}
    log.note("before", "%d role rows across %d lever modelling tabs"
             % (sum(b_roles.values()), len(before)))

    vb = wbv[REVIEW]
    hdrv = headers(vb)
    aa, aj = hdrv.get("Full Cost \nAUD"), hdrv.get("MTab")
    tab_of_old = {nz(wbv[t]["C3"].value): t for t in before}
    old_cost = {t: 0.0 for t in before}
    for r in range(2, vb.max_row + 1):
        if not nz(vb.cell(r, 2).value):
            continue
        t = tab_of_old.get(nz(vb.cell(r, aj).value))
        v = vb.cell(r, aa).value
        if t and isinstance(v, (int, float)):
            old_cost[t] += v

    pcm = {}
    for h in ("Status (PCM)", "Commentry (PCM)", "End date (PCM)"):
        c = hdrv.get(h)
        if not c:
            continue
        for r in range(2, vb.max_row + 1):
            n = nz(vb.cell(r, 2).value)
            if not n or vb.cell(r, c).value is None:
                continue
            pcm.setdefault((squash(n), squash(vb.cell(r, 3).value)), {})[h] = \
                vb.cell(r, c).value

    # ---------------------------------------------------------------- W1-2
    log.head("W1-2  the maps his data needs, on Lists")
    M = build_maps(wb, log)
    FUNDED = M["funded"]
    for d in M["drift"]:
        log("Lists", "Lists!T:U", d)
    log("Lists", "Lists!T:U", "%d portfolio values map to a tab, his spellings"
        % len(PORT_MAP))
    log("Lists", "Lists!Q:R", "%d division values carry a row whose Portfolio "
        "reads NA or blank" % len(DIV_FALLBACK))
    log("Lists", "Lists!AR", "%s home to TDD Cyber whatever the portfolio says"
        % " and ".join(TDD_CYBER_SQUADS))
    log("Lists", "Lists!AQ", "%d overhead line rulings keyed on the person"
        % len(OH_OVERRIDES))

    # ---------------------------------------------------------------- W1-3
    log.head("W1-3  his 29 columns replace the REVIEW raw block")
    hdr, rows = read_his()
    if len(rows) != 526:
        log.note("W1-3", "his file carries %d data rows" % len(rows))
    cols = {nz(h): i + 1 for i, h in enumerate(hdr)}
    hcol = {h: RAW_N + 2 + i for i, h in enumerate(HELPERS)}
    for h in HIS_READ:
        if h not in cols:
            stop("his file has no %r column" % h)

    m = col_remap(cols, hcol)
    nref = repoint_columns(wb, m)
    log("refs", "workbook", "%d formulas repointed to his column positions; "
        "Full Cost now reads the effective cost helper %s, MStatus %s, MTab %s,"
        " Role ID %s" % (nref, gl(hcol["Effective cost (AUD)"]),
                         gl(hcol["MStatus"]), gl(hcol["MTab"]),
                         gl(hcol["Role ID"])))

    n = write_review(wb, hdr, rows, hcol)
    log("raw", "%s!A1:%s%d" % (REVIEW, gl(RAW_N), n + 1),
        "his header verbatim and %d data rows, values, %d columns A to %s"
        % (n, RAW_N, gl(RAW_N)))

    # his file hides no rows - the tab must read exactly as he typed it, so
    # hidden state inherited from the previous model's lock comes off
    unhid = 0
    for rd in wb[REVIEW].row_dimensions.values():
        if rd.hidden:
            rd.hidden = False
            unhid += 1
    if unhid:
        log("rows", "%s (visibility)" % REVIEW,
            "%d hidden rows inherited from the previous model unhidden; his "
            "file hides none" % unhid)

    NAME_COL = gl(cols["Name"])
    TITLE_COL = gl(cols["Position Title"])
    COUNTRY_COL = gl(cols["Country"])
    STATUS_COL = gl(hcol["MStatus"])
    EFF_COL = gl(hcol["Effective cost (AUD)"])
    RID_COL = gl(hcol["Role ID"])
    MTAB_COL = gl(hcol["MTab"])
    GROUP_COL = gl(hcol["Squad or overhead line"])
    PLAT_COL = gl(hcol["EGI funded"])

    derived, fallback = derive(rows, cols, M)
    ALL_EGI = set(M["egilab"].values())
    bad = [d for d in derived if not d["mtab"] or d["mtab"] == "UNMAPPED"]
    if bad:
        for d in bad[:20]:
            print("   unmapped: row %d %s | %s" % (d["i"] + 2, d["name"],
                                                   d["title"]))
        stop("%d of %d rows home to no tab" % (len(bad), len(derived)))

    # ---------------------------------------------------------------- W1-4
    log.head("W1-4  the helper block, rebuilt from column AE, Role ID last")
    rv = wb[REVIEW]
    for i, d in enumerate(derived):
        r = 2 + i
        f = helper_formulas(hcol, cols, r)
        for h in HELPERS:
            c = hcol[h]
            if h in f:
                rv.cell(r, c).value = "=" + f[h]
            elif h == "Role ID":
                rv.cell(r, c).value = d["rid"]
                rv.cell(r, c).font = Font(name="Calibri", size=8, color=GREY)
            else:
                v = pcm.get((squash(d["name"]), squash(d["title"])), {}).get(h)
                rv.cell(r, c).value = v
    for h in HELPERS:
        rv.cell(1, hcol[h]).font = Font(name="Calibri", size=8, color=GREY)
    rv.auto_filter.ref = "A1:%s%d" % (gl(hcol["Role ID"]), n + 1)
    log("helpers", "%s!%s:%s" % (REVIEW, gl(RAW_N + 2), gl(hcol["Role ID"])),
        "%s" % ", ".join(HELPERS))
    log("ids", "%s!%s" % (REVIEW, RID_COL),
        "R0001..R%04d down his row order" % n)

    carried = set()
    for d in derived:
        k = (squash(d["name"]), squash(d["title"]))
        if k in pcm:
            carried.add(k)
    log.note("PCM", "%d of the %d annotated people carried their PCM status, "
             "commentry and end date person-keyed; %d annotations belong to "
             "people who are not in his file"
             % (len(carried), len(pcm), len(pcm) - len(carried)))

    # the funded table has to speak his squad names
    ls = wb["Lists"]
    live = set(d["at"] for d in derived) | set(d["ap"] for d in derived)
    sq = {squash(s): s for s in live}
    ok = [nz(ls.cell(r, ci("AU")).value) for r in range(2, 11)
          if nz(ls.cell(r, ci("AU")).value) in live]
    log("Lists", "Lists!AU2:AU10",
        "funded squads that still read exactly as his Squad column: %s"
        % ", ".join(ok))
    for r in range(2, 11):
        nm = nz(ls.cell(r, ci("AU")).value)
        if not nm:
            continue
        if nm in live:
            continue
        hit = sq.get(squash(nm))
        if hit:
            ls.cell(r, ci("AU")).value = hit
            log("Lists", "Lists!AU%d" % r,
                "funded squad %r reads %r in his file, Lists realigned"
                % (nm, hit))
        else:
            log("Lists", "Lists!AU%d" % r,
                "funded squad %r matches no squad in his file - kept, it "
                "prices at 0" % nm)

    # the EGI rows this stage groups on the grid are funded outside too, so
    # the table has to name them or nothing downstream knows to leave them out
    have = {nz(ls.cell(r, ci("AU")).value)
            for r in range(2, ls.max_row + 1)}
    nxt = 11
    for lab in sorted({d["at"] for d in derived
                       if nz(d.get("plat")).upper() == "EGI" and d["at"]}
                      - have):
        while nz(ls.cell(nxt, ci("AU")).value):
            nxt += 1
        ls.cell(nxt, ci("AU")).value = lab
        log("Lists", "Lists!AU%d" % nxt,
            "the grid's own EGI row %r added to the funded table, so the "
            "support percentage and the funding lines leave it out" % lab)
        nxt += 1

    # ---------------------------------------------------------------- W1-5
    log.head("W1-5  the homing table, proved complete")
    tab_of = {}
    for t in tabs2x(wb):
        tab_of[nz(wb[t]["C3"].value)] = t
    unknown = sorted(set(d["mtab"] for d in derived) - set(tab_of))
    if unknown:
        stop("no lever modelling tab for %s" % ", ".join(unknown))
    bytab = collections.defaultdict(list)
    for d in derived:
        bytab[tab_of[d["mtab"]]].append(d)
    print("\n   %-30s %-24s %5s %5s %12s %12s"
          % ("tab", "portfolio (C3)", "was", "now", "was ($m)", "now ($m)"))
    for t in tabs2x(wb):
        print("   %-30s %-24s %5d %5d %12.3f %12.3f"
              % (t, nz(wb[t]["C3"].value), b_roles.get(t, 0), len(bytab[t]),
                 old_cost.get(t, 0) / 1e6,
                 sum(d["eff"] for d in bytab[t]) / 1e6))
    print("   %-30s %-24s %5d %5d %12.3f %12.3f"
          % ("TOTAL", "", sum(b_roles.values()),
             sum(len(v) for v in bytab.values()), sum(old_cost.values()) / 1e6,
             sum(d["eff"] for d in derived) / 1e6))
    if sum(len(v) for v in bytab.values()) != len(derived):
        stop("rows home to more than one tab")
    print("\n   fallback resolved rows (%d) - Portfolio could not settle them"
          % len(fallback))
    print("   %-4s %-26s %-42s %-24s %-30s %s"
          % ("row", "name", "role", "portfolio", "division", "tab"))
    for i, name, title, port, div, plat, squad, mt in fallback:
        print("   %-4d %-26s %-42s %-24s %-30s %s"
              % (i + 2, name[:26], title[:42], (port or "(blank)")[:24],
                 div[:30], mt))
    ov = [d for d in derived if d["src"].startswith("person")]
    cyb = [d for d in derived if d["src"] == "TDD Cyber squad"]
    print("\n   agreed person moves applied: %d" % len(ov))
    for d in ov:
        print("      %-24s %-40s -> %s / %s" % (d["name"][:24], d["title"][:40],
                                                d["mtab"], d["at"]))
    if cyb:
        print("   squad rule to TDD Cyber: %d" % len(cyb))

    print("\n   the overhead lines, each roster and total")
    byline = collections.defaultdict(list)
    for d in derived:
        if d["ar"] != "Squad":
            byline[d["ar"]].append(d)
    for line in OVERHEAD_ORDER:
        who = byline.get(line, [])
        print("   %-20s %3d roles, %10.3f $m effective"
              % (line, len(who), sum(d["eff"] for d in who) / 1e6))
        for d in sorted(who, key=lambda d: (d["mtab"], d["name"])):
            print("      %-26s %-46s %-24s %s"
                  % (d["name"][:26], d["title"][:46], d["mtab"], d["status"]))
    off = [d for d in derived if d["ar"] == "Squad" and "head of " in
           d["title"].lower()]
    for d in off:
        print("   kept off the heads line: %-24s %-40s (Squad %r)"
              % (d["name"][:24], d["title"][:40], d["ap"]))

    # ---------------------------------------------------------------- W1-6
    log.head("W1-6  levers carried person-keyed")
    for d in derived:
        d["tab"] = tab_of[d["mtab"]]
    kept, defaulted, filled, departed = carry_levers(before, derived, log)
    print("   kept          %4d levers followed their person" % len(kept))
    print("   vacancy filled%4d vacancies became a named person, lever Filled"
          % len(filled))
    print("   new           %4d people had no lever, default Vacant->Hire, "
          "Filled->Filled" % len(defaulted))
    print("   departed      %4d people left and took their lever with them"
          % len(departed))
    print("\n   people whose lever came from the vacancy they fill")
    for p, s, how in filled:
        print("      %-26s %-42s %-8s from %s" % (p["name"][:26], p["title"][:42],
                                                  p["lever"], s["tab"]))
    print("\n   new people (first 40 of %d)" % len(defaulted))
    for p in defaulted[:40]:
        print("      %-26s %-42s %-8s %s" % (p["name"][:26], p["title"][:42],
                                             p["lever"], p["tab"]))
    print("\n   departed people (first 40 of %d)" % len(departed))
    for p in departed[:40]:
        print("      %-26s %-42s %-8s %s" % (p["name"][:26], p["title"][:42],
                                             p["lever"], p["tab"]))
    moved = [(p, s) for p, s, _ in (kept + filled) if s["tab"] != p["tab"]]
    print("\n   people who changed tab: %d" % len(moved))
    for p, s in moved[:60]:
        print("      %-26s %-40s %-26s -> %s" % (p["name"][:26], p["title"][:40],
                                                 s["tab"], p["tab"]))

    # ---------------------------------------------------------------- W1-7
    log.head("W1-7  every 2.x FTE block re-homed")
    grids, order, placement = {}, [], {}
    for t in tabs2x(wb):
        want = collections.OrderedDict()
        for d in sorted(bytab[t], key=lambda d: (d["at"], d["i"])):
            want.setdefault(d["at"], []).append(d)
        ranges, sec, marks, place = rehome_tab(wb, t, want, styles[t], log)
        for rid, rr in place.items():
            placement[rid] = rr
        gone = [g for g, _ in before[t] if g not in want]
        joined = [g for g in want if g not in dict(before[t])]
        log("tab", t, "%d roles in %d groups (%d before)%s%s"
            % (sum(len(v) for v in want.values()), len(want), b_roles.get(t, 0),
               ("; groups gone: " + ", ".join(gone)) if gone else "",
               ("; groups new: " + ", ".join(joined)) if joined else ""))
        rowinfo = []
        for s in sec:
            kind = dict(SECTIONS)[s]
            for gr, g in sec[s]["rows"]:
                rowinfo.append((gr, g, kind))
        rowinfo.sort()
        grids[t] = {"label": nz(wb[t]["C3"].value), "rows": rowinfo}

    print("\n   per tab membership delta")
    print("   %-32s %-42s %5s %5s" % ("tab", "group", "was", "now"))
    for t in tabs2x(wb):
        ob = {g: len(r) for g, r in before[t]}
        nb = collections.Counter(d["at"] for d in bytab[t])
        for g in sorted(set(ob) | set(nb)):
            if ob.get(g, 0) != nb.get(g, 0):
                print("   %-32s %-42s %5d %5d" % (t, g, ob.get(g, 0), nb.get(g, 0)))

    # ---------------------------------------------------------------- W1-8
    log.head("W1-8  3.3 mirrors the new grid")
    order = []
    ws33 = wb["3.3 Squad Actuals to Archetype"]
    seen = []
    for r in range(6, ws33.max_row + 1):
        f = ws33.cell(r, 4).value
        if isinstance(f, str):
            mm = re.match(r"='([^']+)'!", f)
            if mm and mm.group(1) not in seen:
                seen.append(mm.group(1))
    order = [t for t in seen if t in grids] + [t for t in grids if t not in seen]
    rebuild_33(wb, order, grids, log)
    log("3.3", "3.3 Squad Actuals to Archetype",
        "%d squad and overhead rows across %d tabs, each with its total"
        % (sum(len(g["rows"]) for g in grids.values()), len(grids)))

    # ---------------------------------------------------------------- W1-9
    log.head("W1-9  3.4 and 3.2 read the squads and lines his file carries")
    n34 = fix_34(wb, derived, log)
    n32 = fix_32(wb, derived, log)
    n1x = fix_1x_egi(wb, derived, log)
    log.note("3.4 / 3.2 / 1.x", "%d COE or EGI section(s) rebuilt, %d overhead "
             "line(s) stood down, %d EGI platform line(s) repointed%s"
             % (n34, n32, len(n1x), (": " + ", ".join(n1x)) if n1x else ""))

    # ---------------------------------------------------------------- W1-10
    log.head("W1-10  the vacant overhead dial on 3.5 names the new role rows")
    nl, nv = fix_35_dial(wb, derived, placement, log)
    log.note("3.5", "%d overhead lines rewired over %d vacant overhead roles"
             % (nl, nv))

    # ---------------------------------------------------------------- write
    log.head("recalculating and writing the cached values back")
    tmp = dst + ".raw"
    save(wb, tmp)
    rc, st = wbio.build(tmp, dst)
    os.remove(tmp)
    d = os.path.dirname(rc)
    if os.path.basename(d).startswith("recalc_"):
        shutil.rmtree(d, ignore_errors=True)
    print("recalculated, %d formula cells populated across %d sheets"
          % (st["cells"], st["sheets"]), flush=True)

    # ---------------------------------------------------------------- checks
    log.head("self-check")
    self_check(dst, hdr, rows, derived, hcol, cols, bytab, tab_of)

    print("\n   part time people, raw against effective")
    print("   %-26s %-42s %5s %14s %14s" % ("name", "role", "FTE", "raw",
                                            "effective"))
    for d in derived:
        if 0 < d["fte"] < 1:
            print("   %-26s %-42s %5.2f %14.2f %14.2f"
                  % (d["name"][:26], d["title"][:42], d["fte"], d["cost"],
                     d["eff"]))

    if FAILS:
        print("\nw1_map: FAILED - %s" % ", ".join(FAILS))
        raise SystemExit(2)
    log.tail()
    print("wrote", dst)


def self_check(path, hdr, rows, derived, hcol, cols, bytab, tab_of):
    wf = openpyxl.load_workbook(path, data_only=False)
    wv = openpyxl.load_workbook(path, data_only=True)
    fr, vr = wf[REVIEW], wv[REVIEW]

    diffs = []
    for c in range(1, RAW_N + 1):
        if nz(vr.cell(1, c).value) != nz(hdr[c - 1]):
            diffs.append(("header", gl(c)))
    for i, row in enumerate(rows):
        for c in range(RAW_N):
            a, b = vr.cell(2 + i, c + 1).value, row[c]
            if isinstance(a, (int, float)) and isinstance(b, (int, float)):
                if abs(a - b) > 1e-9:
                    diffs.append((2 + i, gl(c + 1)))
            elif nz(a) != nz(b):
                diffs.append((2 + i, gl(c + 1)))
    chk("raw block cell identical to his file", not diffs,
        "%d cells x %d columns, %d differences%s"
        % (len(rows), RAW_N, len(diffs), (" e.g. %s" % diffs[:5]) if diffs else ""))

    hidden = [r for r, rd in fr.row_dimensions.items() if rd.hidden]
    chk("no hidden row on the tab - his file hides none", not hidden,
        "%d hidden %s" % (len(hidden), hidden[:8]))

    tail = [r for r in range(len(rows) + 2, fr.max_row + 1)
            if any(fr.cell(r, c).value is not None for c in range(1, RAW_N + 1))]
    chk("nothing left below his last row", not tail, "rows %s" % tail[:5])

    ids = [vr.cell(2 + i, hcol["Role ID"]).value for i in range(len(rows))]
    want = ["R%04d" % (i + 1) for i in range(len(rows))]
    chk("Role IDs reassigned in his row order", ids == want,
        "R0001..R%04d, %d unique" % (len(rows), len(set(ids))))

    mt = collections.Counter(nz(vr.cell(2 + i, hcol["MTab"]).value)
                             for i in range(len(rows)))
    unhomed = [k for k in mt if k in ("", "UNMAPPED")]
    chk("every row homes to exactly one tab", not unhomed,
        "%d rows over %d tabs, %d unmapped"
        % (sum(mt.values()), len([k for k in mt if k]), sum(mt[k] for k in unhomed)))

    py = collections.Counter(d["mtab"] for d in derived)
    chk("the sheet homes them where the derivation says", py == mt,
        "%d tabs agree" % len(py))

    bad = []
    for i, d in enumerate(derived):
        got = vr.cell(2 + i, hcol["Effective cost (AUD)"]).value
        if not isinstance(got, (int, float)) or abs(got - d["eff"]) > 1e-6:
            bad.append((d["name"], got, d["eff"]))
    pt = [d for d in derived if 0 < d["fte"] < 1]
    chk("part time prices at Full Cost x FTE", not bad,
        "%d people under 1.0 FTE, %d effective costs off" % (len(pt), len(bad)))

    errs, blanks = wbio.audit(path)
    real = [e for e in errs if not (e[0] == REVIEW and
                                    ci(re.match(r"([A-Z]+)", e[1]).group(1)) <= RAW_N)]
    chk("no error cells in any formula", not real,
        "%d formula cells in error%s" % (len(real), (" e.g. %s" % real[:3])
                                         if real else ""))
    refs = [e for e in real if "#REF" in str(e[2])]
    chk("no #REF anywhere", not refs, "%d" % len(refs))

    nctl = badctl = 0
    for t in wf.sheetnames:
        f, v = wf[t], wv[t]
        for row in f.iter_rows(min_col=2, max_col=2):
            for c in row:
                if isinstance(c.value, str) and "must be 0" in c.value:
                    nctl += 1
                    got = None
                    for cc in range(3, min(f.max_column, c.column + 16) + 1):
                        x = v.cell(c.row, cc).value
                        if isinstance(x, (int, float)):
                            got = x
                            break
                    if got is None or abs(got) > 1e-6:
                        badctl += 1
                        print("      control off: %s row %d = %r"
                              % (t, c.row, got))
    chk("every control on every tab reads 0", badctl == 0,
        "%d controls, %d off" % (nctl, badctl))

    at = {vr.cell(2 + i, hcol["Role ID"]).value: 2 + i for i in range(len(rows))}
    nrow = nbad = 0
    seen = collections.Counter()
    for t in [ws.title for ws in wf.worksheets if ws.title.startswith("2.")]:
        f, v = wf[t], wv[t]
        for row in f.iter_rows(min_col=1, max_col=1):
            for c in row:
                rid = c.value
                if not (isinstance(rid, str) and re.fullmatch(r"R\d{4}", rid)):
                    continue
                nrow += 1
                seen[rid] += 1
                rr = at.get(rid)
                got = (v.cell(c.row, 2).value, v.cell(c.row, 4).value,
                       v.cell(c.row, 6).value)
                wnt = (vr.cell(rr, cols["Name"]).value,
                       vr.cell(rr, hcol["MStatus"]).value,
                       vr.cell(rr, hcol["Effective cost (AUD)"]).value)
                ok = (nz(got[0]) == nz(wnt[0]) and nz(got[1]) == nz(wnt[1])
                      and isinstance(got[2], (int, float))
                      and isinstance(wnt[2], (int, float))
                      and abs(got[2] - wnt[2]) <= 1e-9)
                if not ok:
                    nbad += 1
                    if nbad <= 3:
                        print("      mismatch %s row %d %s: %r vs %r"
                              % (t, c.row, rid, got, wnt))
    dup = [k for k, v2 in seen.items() if v2 > 1]
    chk("every person sits on exactly one tab, once",
        nrow == len(rows) and not dup and nbad == 0,
        "%d role rows, %d people, %d duplicated, %d reading the wrong person"
        % (nrow, len(rows), len(dup), nbad))

    ties, worst = [], 0.0
    for t in [ws.title for ws in wf.worksheets if ws.title.startswith("2.")]:
        f, v = wf[t], wv[t]
        g = 0.0
        for row in f.iter_rows(min_col=1, max_col=1):
            for c in row:
                if isinstance(c.value, str) and re.fullmatch(r"R\d{4}", c.value):
                    x = v.cell(c.row, 7).value
                    if isinstance(x, (int, float)):
                        g += x
        grid = None
        for r in range(1, f.max_row + 1):
            if nz(f.cell(r, 2).value) == "Total portfolio":
                grid = v.cell(r, 19).value
        if not isinstance(grid, (int, float)):
            ties.append((t, "no Total portfolio after levers cell"))
            continue
        d = abs(g / 1e6 - grid)
        worst = max(worst, d)
        if d > 1e-9:
            ties.append((t, "block G %.9f, grid %.9f, off by %.9f"
                         % (g / 1e6, grid, g / 1e6 - grid)))
    for t, why in ties:
        print("      %s: %s" % (t, why))
    chk("block G ties to the grid after levers on every 2.x tab", not ties,
        "15 tabs, worst difference %.12f $m" % worst)

    LF = {"Filled": 1.0, "Hire": 1.0, "Hold": 0.0, "Offshore": 0.4}
    cty = {d["rid"]: d["country"] for d in derived}
    bad, charged, ntab, blk = [], 0.0, 0, {}
    for t in [ws.title for ws in wf.worksheets if ws.title.startswith("2.")]:
        f, v = wf[t], wv[t]
        hd = fte_hdr(f)
        if block_cols(f, hd) < 9:
            continue
        ntab += 1
        rows, last = read_block(f, hd)
        blk[t] = set(range(hd + 1, last + 1))
        for row in f.iter_rows(min_col=1, max_col=1):
            for c in row:
                rid = c.value
                if not (isinstance(rid, str) and re.fullmatch(r"R\d{4}", rid)):
                    continue
                cost = v.cell(c.row, 6).value
                after = v.cell(c.row, 7).value
                up = v.cell(c.row, 9).value
                fac = (1.0 if "WIPRO" in cty.get(rid, "").upper()
                       else LF.get(nz(f.cell(c.row, 5).value), 1.0))
                if abs((after or 0) + (up or 0) - (cost or 0) * fac) > 1e-6:
                    bad.append((t, c.row, rid))
                charged += up or 0
    hit, seen = [], 0.0
    for ws in wf.worksheets:                # whoever consumes the charge column
        if ws.title.startswith("2."):
            continue
        for row in ws.iter_rows():
            for c in row:
                v2 = c.value
                if not isinstance(v2, str) or "$I$" not in v2:
                    continue
                for t, rows in blk.items():
                    hits = [int(m) for m in re.findall(
                        r"'" + re.escape(t) + r"'!\$I\$(\d+)", v2)]
                    if hits and all(x in rows for x in hits):
                        seen = wv[ws.title][c.coordinate].value
                        hit.append("%s!%s reads %.6f" % (ws.title, c.coordinate,
                                                         seen or 0))
    ok = (not bad and (charged == 0 or (
        len(hit) == 1 and abs((seen or 0) - charged / 1e6) <= 1e-9)))
    chk("the cyber uplift charge is netted out once and charged on once", ok,
        "%d tab(s) with an uplift column, %d role rows out of step, %.6f $m "
        "charged out; %s" % (ntab, len(bad), charged / 1e6,
                             "; ".join(hit) or "no consumer"))

    f34, v34 = wf["3.4 COE Breakdown"], wv["3.4 COE Breakdown"]
    sections, grand = read_34(f34)
    keys = [s[0] for s in sections]
    want_n = sum(1 for d in derived if d["mtab"] in keys)
    want_c = sum(d["eff"] for d in derived if d["mtab"] in keys) / 1e6
    got_n, got_c = v34.cell(grand, 4).value, v34.cell(grand, 7).value
    miss = []
    for key, first, last, total in sections:
        on = sorted(set(d["at"] for d in derived if d["mtab"] == key))
        got = [nz(f34.cell(r, 3).value) for r in range(first, last + 1)]
        if on != got:
            miss.append("%s: %s not %s" % (key, on, got))
    chk("3.4 lists every squad the role mapping carries and ties to it",
        not miss and got_n == want_n
        and isinstance(got_c, (int, float)) and abs(got_c - want_c) <= 1e-9,
        "%d sections %s, %d roles and %.9f $m against %d and %.9f%s"
        % (len(sections), ", ".join(keys), got_n or 0, got_c or 0, want_n,
           want_c, ("; " + "; ".join(miss)) if miss else ""))

    f32, v32 = wf["3.2 Overhead & Leadership"], wv["3.2 Overhead & Leadership"]
    live = collections.Counter(d["ar"] for d in derived if d["ar"] != "Squad")
    lines, off = 0, []
    for r in range(1, f32.max_row + 1):
        b = f32.cell(r, 2).value
        if not (isinstance(b, str) and b.startswith("=Lists!")):
            continue
        m = re.search(r"\$AF\$(\d+)", b)
        if not m:
            continue
        name = nz(wf["Lists"].cell(int(m.group(1)), ci("AF")).value)
        lines += 1
        g, h = v32.cell(r, 7).value, v32.cell(r, 8).value
        if not isinstance(h, (int, float)) or h < -1e-9:
            off.append("%s H %r" % (name, h))
        elif not g and abs(h) > 1e-9:
            off.append("%s has no role but reads %r not applied" % (name, h))
    chk("3.2 shows no overhead line pricing a role it does not have", not off,
        "%d lines, %d out of step%s" % (lines, len(off),
                                        ("; " + "; ".join(off)) if off else ""))

    lsv = wv["Lists"]
    funded, basis = [], {}
    for r in range(2, 11):
        nm = nz(lsv.cell(r, ci("AU")).value)
        if nm:
            funded.append(nm)
            basis[nm] = lsv.cell(r, ci("AW")).value
    M2 = dict(tenp=[nz(lsv.cell(r, ci("AS")).value)
                    for r in range(2, 13) if nz(lsv.cell(r, ci("AS")).value)],
              pmap={nz(lsv.cell(r, ci("T")).value).lower():
                    nz(lsv.cell(r, ci("U")).value) for r in range(2, 22)
                    if nz(lsv.cell(r, ci("T")).value)})
    tabof = {nz(wf[t]["C3"].value): t
             for t in wf.sheetnames if t.startswith("2.")}

    off = []
    for sq in sorted(set(d["at"] for d in derived if d["at"][:4].upper() == "EGI"
                         or d["at"].upper() == "EGI")):
        want = egi_home(sq, M2) or "EGI"
        on = sorted(set(d["mtab"] for d in derived if d["at"] == sq))
        if on != [want]:
            off.append("%s on %s, not %s" % (sq, on, want))
    chk("every EGI squad sits on the tab its own name points to", not off,
        "%d EGI squads%s"
        % (len(set(d["at"] for d in derived if d["at"][:3].upper() == "EGI")),
           ("; " + "; ".join(off)) if off else ""))

    rows = []
    bad = []
    for key, t in sorted(tabof.items()):
        f, v = wf[t], wv[t]
        want = sum(d["eff"] for d in derived
                   if d["mtab"] == key and d["plat"].upper() == "EGI") / 1e6
        for g in sorted(set(d["at"] for d in derived if d["mtab"] == key)):
            b = basis.get(g)
            if isinstance(b, (int, float)):     # a flat funded programme
                want += b
        got = None
        for r in range(1, f.max_row + 1):
            if nz(f.cell(r, 2).value) == "Total portfolio":
                got = v.cell(r, 16).value
        rows.append((t, want, got))
        if not isinstance(got, (int, float)) or abs(got - want) > 1e-9:
            bad.append("%s wants %.9f, reads %r" % (t, want, got))
    chk("every tab's funded outside total is its funding lines plus its "
        "funded squads", not bad,
        "%d tabs, %.6f $m funded outside in all%s"
        % (len(rows), sum(r[1] for r in rows),
           ("; " + "; ".join(bad)) if bad else ""))

    egi = [d for d in derived if d["plat"].strip().upper() == "EGI"]
    leak = []
    labels = set(nz(lsv.cell(r, ci("Z")).value) for r in range(2, 21))
    for d in egi:
        if d["ar"] != "Squad":
            leak.append("%s reads overhead line %s" % (d["name"], d["ar"]))
        if d["at"] != d["ap"] and d["at"] not in labels:
            leak.append("%s groups as %s, neither its own squad nor its "
                        "portfolio's EGI row" % (d["name"], d["at"]))
    slice_tot = sum(d["eff"] for d in egi) / 1e6
    chk("every person his Platform column marks EGI is funded outside and on "
        "no overhead line", not leak,
        "%d roles, %.6f $m, over %d tabs%s"
        % (len(egi), slice_tot, len(set(d["mtab"] for d in egi)),
           ("; " + "; ".join(leak[:6])) if leak else ""))

    mixed, ngrp = [], 0
    for key, t in sorted(tabof.items()):
        for g in sorted(set(d["at"] for d in derived if d["mtab"] == key)):
            who = [d for d in derived if d["mtab"] == key and d["at"] == g]
            ngrp += 1
            e = sum(1 for d in who if d["plat"].strip().upper() == "EGI")
            if 0 < e < len(who):
                mixed.append("%s %s: %d of %d funded outside"
                             % (t, g, e, len(who)))
    chk("no grid row is part funded outside and part not", not mixed,
        "%d rows across the lever modelling tabs%s"
        % (ngrp, ("; " + "; ".join(mixed)) if mixed else ""))

    lines = []
    for t in [ws.title for ws in wf.worksheets if ws.title.startswith("1.")]:
        f, v = wf[t], wv[t]
        pair = None
        for row in f.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    m = re.search(r"'(2\.[0-9]+[^']*)'!", c.value)
                    if m and pair is None:
                        pair = m.group(1)
        if not pair:
            continue
        key = nz(wf[pair]["C3"].value)
        want = sum(d["eff"] for d in derived
                   if d["mtab"] == key and d["plat"].upper() == "EGI") / 1e6
        for r in range(1, f.max_row + 1):
            if not nz(f.cell(r, 2).value).startswith("Platform: EGI"):
                continue
            for rr in range(r + 1, min(r + 6, f.max_row + 1)):
                b = nz(f.cell(rr, 2).value)
                h = f.cell(rr, 8).value
                if not (b and not b.endswith("Total")
                        and isinstance(h, str) and h.startswith("=")):
                    continue
                got = v.cell(rr, 8).value
                if not isinstance(got, (int, float)) or abs(got - want) > 1e-9:
                    lines.append("%s %r reads %r, wants %.6f"
                                 % (t, b, got, want))
                break
    chk("every 1.x EGI platform line reads its portfolio's EGI slice",
        not lines, "%d tabs carry one%s"
        % (sum(1 for t in wf.sheetnames if t.startswith("1.")
               and any(nz(wf[t].cell(r, 2).value).startswith("Platform: EGI")
                       for r in range(1, wf[t].max_row + 1))),
           ("; " + "; ".join(lines)) if lines else ""))

    book = sum(d["eff"] for d in derived) / 1e6
    tot = 0.0
    for t in tabof.values():
        f, v = wf[t], wv[t]
        for r in range(1, f.max_row + 1):
            if nz(f.cell(r, 2).value) == "Total portfolio":
                tot += v.cell(r, 15).value or 0
    chk("the whole book still prices every person once", abs(tot - book) <= 1e-9,
        "%.9f $m across the tabs against %.9f on the role mapping" % (tot, book))

    levers = collections.Counter()
    for t in [ws.title for ws in wf.worksheets if ws.title.startswith("2.")]:
        f = wf[t]
        for row in f.iter_rows(min_col=1, max_col=1):
            for c in row:
                if isinstance(c.value, str) and re.fullmatch(r"R\d{4}", c.value):
                    levers[nz(f.cell(c.row, 5).value)] += 1
    chk("every lever is one of the four values",
        set(levers) <= {"Filled", "Hire", "Hold", "Offshore"},
        ", ".join("%s %d" % (k, v) for k, v in sorted(levers.items())))
    wf.close()
    wv.close()


if __name__ == "__main__":
    main(*sys.argv[1:])
