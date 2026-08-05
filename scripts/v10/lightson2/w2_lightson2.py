#!/usr/bin/env python3
"""w2 - Stage w2: the two Lights On tabs, rebuilt on his new column set.

  python3 w2_lightson2.py <in.xlsx> <out.xlsx>

'3.5 TDD Lights On' is rebuilt from scratch with his eighteen columns and the
0.2 Data Config row set: the COE pairs split, Ampol Customer and Z Customer on
their own lines, Enterprise Data where TDD Data used to read, TDD Cyber and EGI,
and no Legal line. '3.6 TDD Lights On AU NZ' repeats the table and splits the
charge between AU and NZ off each person's country.

Everything on both tabs is live. Support cost pairs each squad's after lever
cost on the 2.x grid with its support percentage on the 1.x tab. The Business
Partner, Domain Architect and GM pots divide by eleven, the ten portfolios plus
TDD Cyber. Other overheads read the row's own overhead people, and a cream
toggle scales how much of them the lights on budget carries. Every cost in the
role mapping lands in the Total People cost column, and a control row proves it.

Idempotent: handed its own output it rebuilds the same two tabs cell for cell.
"""
import sys, os, re, shutil

sys.path.insert(0, "/home/user/anthropic-claude-code/scripts/v10")
sys.path.insert(0, "/home/user/anthropic-claude-code/scripts/v10/update")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import MultiCellRange, CellRange
from openpyxl.utils import get_column_letter, column_index_from_string

from _xl import REVIEW, Log, load, save, white

TAB = "3.5 TDD Lights On"
TAB2 = "3.6 TDD Lights On AU NZ"
ANCHOR = "3.4 COE Breakdown"
CFG = "0.2 Data Config"
OVH = "3.2 Overhead & Leadership"
ARC = "3.1 Archetype to Actuals"

NAVY = "FF0F2E52"
BAND = "FFF5F4F0"
CREAM = "FFFFF2CC"
RED = "FFFF0000"
WHITE = "FFFFFFFF"
M2 = "#,##0.00;(#,##0.00)"
STEPS = ",".join("%d%%" % p for p in range(0, 101, 5))
MAXF = 8000

FACT = {"Filled": 1.0, "Hire": 1.0, "Hold": 0.0, "Offshore": 0.4}

SECT = {"Squads", "Directly funded programs and platforms", "Overhead roles",
        "No archetype in 1.x tabs"}
SECT_TOT = {"Squads total", "Directly funded total", "Overhead roles total",
            "No archetype total"}

# his eighteen columns, in his order, one column each
HEADERS = [
    ("B", "Portfolios & COEs & EGI"),
    ("C", "Total People cost"),
    ("D", "Sig items funded"),
    ("E", "Support Cost (this is the % in the 1.x tabs)"),
    ("F", "BP allocation"),
    ("G", "Domain architect allocation"),
    ("H", "GM allocation"),
    ("I", "Other overheads"),
    ("J", "Other overheads toggle"),
    ("K", "Amount of overheads charged to TDD"),
    ("L", "Total portfolio cost charged to TDD"),
    ("M", "TDD Lights On budget"),
    ("N", None),
    ("O", "Over/ Under lights on budget"),
    ("P", None),
    ("Q", "Total Cost left to be recharged to business"),
    ("R", "Amount noted in 1.x tabs"),
    ("S", "Still left to fund"),
]
# the AU NZ tab keeps C..L then adds its own block before the tail
HEAD2_MID = [("M", "AU spend"), ("N", "NZ spend"), ("O", "Total"),
             ("P", "AU budget"), ("Q", "NZ budget"), ("R", "Variance")]
HEAD2_TAIL = [("S", "TDD Lights On budget"), ("T", None),
              ("U", "Over/ Under lights on budget"), ("V", None),
              ("W", "Total Cost left to be recharged to business"),
              ("X", "Amount noted in 1.x tabs"), ("Y", "Still left to fund")]

# label, kind, 2.x prefix, 1.x prefix, 0.2 budget label(s), pair partner
#   kind: pf   portfolio, support pairs off the 1.x tab, shares, own overheads
#         cust one side of the Customer split
#         coe  a COE pair sub row
#         coe1 a whole COE tab on one line
#         cyber TDD Cyber, support pairs and shares but its own overheads
#         egi  the EGI tab, wholly funded outside
ROWS = [
    ("COE Strategy Architecture", "coe", "2.13", None,
     ["COE - Strategy Architecture"]),
    ("COE Cyber Risk & Service Ops", "coe1", "2.11", None, ["COE - Cyber"]),
    ("COE Transformation", "coe", "2.12", None, ["COE - Transformation"]),
    ("COE Business Partnering", "coe", "2.12", None,
     ["COE - Business Partnering"]),
    ("COE Data", "coe", "2.13", None, ["COE - Data"]),
    ("Ampol Retail", "pf", "2.1", "1.1", ["Ampol Retail"]),
    ("Z Retail", "pf", "2.10", "1.10", ["Z Retail"]),
    ("Ampol Customer", "cust", "2.2", "1.2", ["Ampol Customer"]),
    ("Z Customer", "cust", "2.2", "1.2", ["Z Customer"]),
    ("Commercial Fuels", "pf", "2.9", "1.9", ["Commercial Fuels"]),
    ("Energy Solutions & B2B", "pf", "2.8", "1.8", ["Energy Solutions & B2B"]),
    ("Infrastructure", "pf", "2.7", "1.7", ["Infrastructure"]),
    ("P&C", "pf", "2.5", "1.5", ["P&C"]),
    ("Finance", "pf", "2.6", "1.6", ["Finance"]),
    ("TDD Group Functions", "pf", "2.4", "1.4", ["TDD Group Functions"]),
    ("Enterprise Data", "pf", "2.3", "1.3", ["Enterprise Data"]),
    ("TDD Cyber", "cyber", "2.15", "1.14", ["TDD Cyber"]),
    ("EGI", "egi", "2.14", None, ["EGI"]),
]
# the COE pairs: row label -> (planned spend cell label on the 2.x tab, partner)
COE_PAIR = {
    "COE Strategy Architecture": ("Strategy & Architecture planned spend ($m)",
                                  "COE Data"),
    "COE Data": ("Data planned spend ($m)", "COE Strategy Architecture"),
    "COE Transformation": ("Transformation planned spend ($m)",
                           "COE Business Partnering"),
    "COE Business Partnering": ("Business Partnering planned spend ($m)",
                                "COE Transformation"),
}
# the squad on each COE tab whose cost is charged out to the portfolios
COE_POT = {"2.12": "TDD Business Partner", "2.13": "Architecture"}
CUST_SIDE = {"Ampol Customer": "AU", "Z Customer": "NZ"}

BP_POT_LABEL = "Business Partner pot after levers ($m)"
DA_POT_LABEL = "Domain Architect pot after levers ($m)"
GM_LABEL = "GM cost ($m)"
BASE_LABEL = "Overhead share base (the ten portfolios plus TDD Cyber)"


def stop(msg):
    print("STOP: %s" % msg)
    raise SystemExit(2)


def q(name):
    return "'" + name.replace("'", "''") + "'"


def norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


def money(cell):
    cell.number_format = M2


def rr(sheet, col, row):
    return "%s!$%s$%d" % (q(sheet), col, row)


def plus(terms):
    return "+".join(terms) if terms else "0"


# --------------------------------------------------------------- book reading

def find_all(ws, col, want, lo=1, hi=None, starts=False):
    hi = hi or ws.max_row
    return [r for r in range(lo, hi + 1)
            if isinstance(ws.cell(r, col).value, str) and
            (ws.cell(r, col).value == want or
             (starts and ws.cell(r, col).value.startswith(want)))]


def find_label(ws, col, want, lo=1, hi=None, starts=False):
    hits = find_all(ws, col, want, lo, hi, starts)
    if len(hits) != 1:
        stop("label %r on %s col %d: %d hits" % (want, ws.title, col, len(hits)))
    return hits[0]


def find_opt(ws, col, want, lo=1, hi=None, starts=False):
    hits = find_all(ws, col, want, lo, hi, starts)
    return hits[0] if len(hits) == 1 else None


def review_cols(ws):
    """Header -> column index, first hit wins. Names change between schemas."""
    out = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is not None and str(h).strip():
            out.setdefault(str(h).strip(), c)
    return out


def pick(cols, *names, **kw):
    pat = kw.get("pat")
    for n in names:
        if n in cols:
            return cols[n], n
    if pat:
        for k, v in cols.items():
            if re.search(pat, k, re.I):
                return v, k
    return None, None


def grid_of(ws):
    """The 2.x summary grid: header row, squad rows, overhead total, total row."""
    hdr = None
    for r in range(3, 16):
        if ws.cell(r, 2).value == "Squad" and \
                ws.cell(r, 19).value == "Squad cost after levers ($m)":
            hdr = r
            break
    if hdr is None:
        stop("no grid header on %s" % ws.title)
    g = {"hdr": hdr, "first": hdr + 1, "total": None, "ovh_total": None,
         "rows": []}
    sec = None
    for r in range(hdr + 1, hdr + 45):
        b = ws.cell(r, 2).value
        if b is None:
            continue
        if b == "Total portfolio":
            g["total"] = r
            break
        if b in SECT:
            sec = b
            continue
        if b in SECT_TOT:
            if b == "Overhead roles total":
                g["ovh_total"] = r
            continue
        if isinstance(b, str) and (b.startswith("Portfolio overhead roles") or
                                   b.startswith("Business Partners")):
            continue
        g["rows"].append({"row": r, "name": str(b).strip(), "sec": sec})
    if g["total"] is None:
        stop("no 'Total portfolio' row on %s" % ws.title)
    return g


def block_of(ws, legacy):
    """The FTE helper block: group header rows and their role rows."""
    hdr = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value == "Name" and ws.cell(r, 5).value == "Vacancy lever":
            hdr = r
            break
    if hdr is None:
        stop("no FTE block on %s" % ws.title)
    up = 8 if ws.cell(hdr, 8).value == "Uplift %" else None
    groups, cur = [], None
    for r in range(hdr + 1, ws.max_row + 1):
        a, b, c = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
        if isinstance(c, str) and c.startswith("=COUNTIF($B$"):
            cur = {"name": str(b).strip(), "row": r, "roles": []}
            groups.append(cur)
            continue
        key = None
        if isinstance(a, str) and re.fullmatch(r"R\d{3,}", a.strip()):
            key = a.strip()
        elif legacy and isinstance(b, str) and b.startswith("='" + REVIEW + "'!$B$"):
            m = re.search(r"\$B\$(\d+)$", b)
            if m:
                key = "row:" + m.group(1)
        if key is not None:
            if cur is None:
                stop("role row above every group on %s row %d" % (ws.title, r))
            cur["roles"].append({"row": r, "key": key,
                                 "lever": str(ws.cell(r, 5).value or "").strip(),
                                 "uplift": (ws.cell(r, up).value or 0) if up else 0})
            continue
        if isinstance(b, str) and b.startswith("Control"):
            break
    if not groups:
        stop("no groups in the FTE block on %s" % ws.title)
    return {"hdr": hdr, "up": up, "groups": groups}


def support_map(ws):
    """{normalised squad: (row, name, pct, au/nz)} for every 1.x Support % cell."""
    out = {}
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 7).value != "Support %":
            continue
        for rrow in range(r + 1, ws.max_row + 1):
            b = ws.cell(rrow, 2).value
            if b is None or (isinstance(b, str) and b.endswith("Total")):
                break
            if b == "Platform Overhead":
                continue
            gv = ws.cell(rrow, 7).value
            if isinstance(gv, (int, float)) and not isinstance(gv, bool):
                out[norm(b)] = (rrow, str(b).strip(), gv,
                                str(ws.cell(rrow, 6).value or "").strip())
    return out


_REFTOK = re.compile(r"(?:('(?:[^']|'')+'|[A-Za-z0-9_.]+)!)?\$?([A-Za-z]{1,3})\$?(\d+)")


def refs_in(f, own):
    out = []
    for m in _REFTOK.finditer(str(f)):
        sh = m.group(1)
        if sh is None:
            sh = own
        elif sh.startswith("'"):
            sh = sh[1:-1].replace("''", "'")
        out.append((sh, m.group(2).upper(), int(m.group(3))))
    return out


# ------------------------------------------------------------------ the build

def main(src, dst):
    log = Log("w2_lightson2")
    wb = load(src)
    wbv = openpyxl.load_workbook(src, data_only=True)

    for t in [ANCHOR, CFG, OVH, ARC, REVIEW, "Lists"]:
        if t not in wb.sheetnames:
            stop("tab %r missing" % t)

    def by_prefix(p):
        hits = [n for n in wb.sheetnames if n.startswith(p + " ")]
        if len(hits) != 1:
            stop("tab prefix %r: %d hits" % (p, len(hits)))
        return hits[0]

    T2 = {p: by_prefix(p) for p in sorted({r[2] for r in ROWS})}
    T1 = {p: by_prefix(p) for p in sorted({r[3] for r in ROWS if r[3]})}
    ALL2 = sorted([n for n in wb.sheetnames if re.match(r"^2\.\d+ ", n)],
                  key=lambda n: float(n.split()[0]))
    for n in ALL2:
        if n not in T2.values():
            stop("2.x tab %r is not on the Lights On row set" % n)

    rev, revv = wb[REVIEW], wbv[REVIEW]
    lists, listsv = wb["Lists"], wbv["Lists"]
    cfg = wb[CFG]

    # ------------------------------------------------- W0  read the schema
    log.head("W0  the schema this input carries")
    rc = review_cols(rev)
    c_cost, n_cost = pick(rc, "Effective cost \nAUD", "Effective Cost AUD",
                          "Effective cost (AUD)", "Effective cost",
                          pat=r"^effective")
    legacy = c_cost is None
    if legacy:
        c_cost, n_cost = pick(rc, "Full Cost \nAUD", "Full Cost AUD",
                              pat=r"^full cost")
    if c_cost is None:
        stop("no cost column on REVIEW")
    c_name, _ = pick(rc, "Name")
    c_stat, _ = pick(rc, "MStatus")
    c_ovh, _ = pick(rc, "Overhead line")
    c_ctry, _ = pick(rc, "Country")
    c_type, _ = pick(rc, "Type")
    c_id, _ = pick(rc, "Role ID")
    for nm, cc in (("Name", c_name), ("MStatus", c_stat),
                   ("Overhead line", c_ovh), ("Country", c_ctry),
                   ("Type", c_type)):
        if cc is None:
            stop("REVIEW column %r not found" % nm)
    log("W0", "%s!1:1" % REVIEW,
        "cost column %r%s, role ids %s" %
        (n_cost, ", the old schema so raw cost is used" if legacy else "",
         "on" if c_id else "off, role rows keyed by REVIEW row"))

    people = {}
    for r in range(2, rev.max_row + 1):
        nm = revv.cell(r, c_name).value
        if nm is None or not str(nm).strip():
            continue
        key = str(revv.cell(r, c_id).value).strip() if c_id and \
            revv.cell(r, c_id).value else "row:%d" % r
        people[key] = {
            "row": r, "name": str(nm).strip(),
            "cost": revv.cell(r, c_cost).value or 0.0,
            "status": str(revv.cell(r, c_stat).value or "").strip(),
            "ovh": str(revv.cell(r, c_ovh).value or "").strip(),
            "nz": str(revv.cell(r, c_ctry).value or "").strip().upper() == "NZ",
            "wipro": "WIPRO" in str(revv.cell(r, c_type).value or "").upper()}
        if c_id and revv.cell(r, c_id).value:
            people.setdefault("row:%d" % r, people[key])
    if not people:
        stop("no people on REVIEW; hand this stage a wbio built file")
    ovh_vals = {p["ovh"] for p in people.values() if p["ovh"]}
    if not ovh_vals or len(ovh_vals) > 14:
        stop("the Overhead line column reads like formulas, not values")

    # funded squads, their basis, and the GM input, all off Lists
    funded = {}
    au1 = find_label(lists, 47, "Squad")
    for r in range(au1 + 1, au1 + 40):
        v = lists.cell(r, 47).value
        if v is None:
            break
        funded[norm(v)] = r
    if not funded:
        stop("no funded squad table on Lists")
    gm_row = find_label(lists, 32, GM_LABEL)
    if not isinstance(lists.cell(gm_row, 33).value, (int, float)):
        stop("Lists %s does not carry a number" % GM_LABEL)

    # ----------------------------------------------- W1  the per tab model
    log.head("W1  what every lever tab holds, person by person")
    MOD = {}
    for p, t in sorted(T2.items(), key=lambda kv: float(kv[0])):
        ws, wsv = wb[t], wbv[t]
        g, blk = grid_of(ws), block_of(ws, legacy)
        gsum, gnet, gau, gnz, groles = {}, {}, {}, {}, {}
        ovh_cells, ovh_au, ovh_nz, up_cells = [], [], [], []
        tot = net = upl = 0.0
        for grp in blk["groups"]:
            s = n = a = z = 0.0
            for ro in grp["roles"]:
                pe = people.get(ro["key"])
                if pe is None:
                    stop("%s row %d points at %s which is not on REVIEW"
                         % (t, ro["row"], ro["key"]))
                f = 1.0 if pe["wipro"] else FACT.get(ro["lever"], 1.0)
                full = (pe["cost"] or 0) * f
                u = full * (ro["uplift"] or 0)
                s += full
                n += full - u
                if pe["nz"]:
                    z += full - u
                else:
                    a += full - u
                if pe["ovh"] and pe["ovh"] != "Squad":
                    ovh_cells.append(rr(t, "G", ro["row"]))
                    (ovh_nz if pe["nz"] else ovh_au).append(rr(t, "G", ro["row"]))
                ro["nz"] = pe["nz"]
                ro["after"] = full - u
            if blk["up"]:
                up_cells.append(rr(t, "I", grp["row"]))
            gsum[grp["name"]] = s
            gnet[grp["name"]] = n
            gau[grp["name"]] = a
            gnz[grp["name"]] = z
            groles[grp["name"]] = grp["roles"]
            tot += s
            net += n
            upl += s - n
        MOD[p] = {
            "tab": t, "grid": g, "blk": blk, "gsum": gsum, "gnet": gnet,
            "gau": gau, "gnz": gnz, "groles": groles, "ovh_cells": ovh_cells,
            "ovh_au": ovh_au, "ovh_nz": ovh_nz, "up_cells": up_cells,
            "total": tot / 1e6, "net": net / 1e6, "uplift": upl / 1e6,
            "gridS": wsv.cell(g["total"], 19).value or 0.0,
            "gridP": wsv.cell(g["total"], 16).value or 0.0,
            "gridO": wsv.cell(g["total"], 15).value or 0.0,
            "ovhS": (wsv.cell(g["ovh_total"], 19).value or 0.0)
                    if g["ovh_total"] else None}
        own = 0.0
        ownau = ownnz = 0.0
        for grp in blk["groups"]:
            for ro in grp["roles"]:
                pe = people[ro["key"]]
                if pe["ovh"] and pe["ovh"] != "Squad":
                    own += ro["after"]
                    if pe["nz"]:
                        ownnz += ro["after"]
                    else:
                        ownau += ro["after"]
        MOD[p]["ownovh"] = own / 1e6
        MOD[p]["ownovh_au"] = ownau / 1e6
        MOD[p]["ownovh_nz"] = ownnz / 1e6
        MOD[p]["au"] = sum(gau.values()) / 1e6
        MOD[p]["nz"] = sum(gnz.values()) / 1e6
        if abs(MOD[p]["net"] - MOD[p]["gridS"]) > 1e-6:
            stop("%s: the FTE block priced off REVIEW column %r after levers "
                 "(%.6f) does not match the grid total (%.6f). The lever tabs "
                 "must price from the same column this stage reads, so if an "
                 "effective cost column has been added the block must read it "
                 "too" % (t, n_cost, MOD[p]["net"], MOD[p]["gridS"]))
        if MOD[p]["ovhS"] is not None and \
                abs(MOD[p]["ownovh"] - MOD[p]["ovhS"]) > 1e-6:
            stop("%s: overhead coded people (%.6f) do not match the grid "
                 "overhead total (%.6f)" % (t, MOD[p]["ownovh"], MOD[p]["ovhS"]))
        log("W1", t, "%d roles after levers %.6f m, own overheads %.6f m, "
            "funded outside %.6f m%s"
            % (sum(len(v) for v in groles.values()), MOD[p]["net"],
               MOD[p]["ownovh"], MOD[p]["gridP"],
               ", charged out to a programme %.6f m" % MOD[p]["uplift"]
               if MOD[p]["uplift"] else ""))

    raw_total = sum(p["cost"] or 0 for p in
                    {id(v): v for v in people.values()}.values()) / 1e6
    cover = sum(MOD[p]["gridO"] for p in MOD)
    log.note("W1", "every cost represented: the fifteen lever tabs carry %.6f m "
             "before levers against %.6f m on the role mapping"
             % (cover, raw_total))
    if abs(cover - raw_total) > 1e-6:
        stop("the lever tabs do not carry every person on REVIEW")

    # ------------------------------------------- W2  pots and the share base
    log.head("W2  the pots and the share base, labelled on Lists")

    def lists_cell(label, formula, note):
        row = find_opt(lists, 32, label)
        if row is None:
            row = 2
            while not (lists.cell(row, 32).value is None and
                       lists.cell(row, 33).value is None and
                       lists.cell(row + 1, 32).value is None):
                row += 1
                if row > 400:
                    stop("no free rows on Lists")
            lists.cell(row, 32).value = label
        lists.cell(row, 33).value = formula
        return row

    bp_g = MOD["2.12"]["grid"]
    da_g = MOD["2.13"]["grid"]
    bp_row = lists_cell(BP_POT_LABEL,
                        '=INDEX(%s!$S$%d:$S$%d,MATCH("TDD Business Partner",'
                        "%s!$B$%d:$B$%d,0))"
                        % (q(T2["2.12"]), bp_g["first"], bp_g["total"],
                           q(T2["2.12"]), bp_g["first"], bp_g["total"]),
                        "BP pot")
    da_row = lists_cell(DA_POT_LABEL,
                        '=INDEX(%s!$S$%d:$S$%d,MATCH("Architecture",'
                        "%s!$B$%d:$B$%d,0))"
                        % (q(T2["2.13"]), da_g["first"], da_g["total"],
                           q(T2["2.13"]), da_g["first"], da_g["total"]),
                        "DA pot")
    base_row = lists_cell(BASE_LABEL, "=COUNTA(Lists!$AS$2:$AS$12)+1", "eleven")
    for r in (bp_row, da_row):
        money(lists.cell(r, 33))
    lists.cell(base_row, 33).number_format = "0"
    POT = {"bp": "Lists!$AG$%d" % bp_row, "da": "Lists!$AG$%d" % da_row,
           "gm": "Lists!$AG$%d" % gm_row}
    ELEVEN = "Lists!$AG$%d" % base_row
    n_base = sum(1 for r in range(2, 13) if listsv.cell(r, 45).value) + 1
    if n_base != 11:
        stop("the share base reads %s, his ruling is eleven" % n_base)
    log("W2", "Lists!AF%d:AG%d" % (bp_row, bp_row), "%s, live off 2.12" % BP_POT_LABEL)
    log("W2", "Lists!AF%d:AG%d" % (da_row, da_row), "%s, live off 2.13" % DA_POT_LABEL)
    log("W2", "Lists!AF%d:AG%d" % (base_row, base_row),
        "%s reads %d, his ruling that TDD Cyber carries an overhead too"
        % (BASE_LABEL, n_base))
    pot_v = {"bp": listsv.cell(bp_g["total"], 19).value,
             "da": listsv.cell(da_g["total"], 19).value}
    v_bp = MOD["2.12"]["gnet"]["TDD Business Partner"] / 1e6
    v_da = MOD["2.13"]["gnet"]["Architecture"] / 1e6
    v_gm = float(lists.cell(gm_row, 33).value)

    # -------------------------------------------- W3  Enterprise Data on 0.2
    log.head("W3  Enterprise Data, his ruling that TDD Data is not a thing")
    cfg_first, cfg_last = 5, 30
    b22 = find_opt(cfg, 2, "TDD Data", cfg_first, cfg_last)
    if b22 is not None:
        cfg.cell(b22, 2).value = "Enterprise Data"
        log("W3", "%s!B%d" % (CFG, b22), "TDD Data now reads Enterprise Data")
    else:
        find_label(cfg, 2, "Enterprise Data", cfg_first, cfg_last)
        log.note("W3", "0.2 already reads Enterprise Data")

    def budget_row(label):
        starts = label.startswith("COE - Cyber")
        return find_label(cfg, 2, label, cfg_first, cfg_last, starts=starts)

    # ------------------------------------ W4  the 1.x support and funding maps
    log.head("W4  the 1.x tabs: support percentages and funding lines")
    SUP = {p: support_map(wb[t]) for p, t in T1.items()}

    def funding_lines(p):
        """[(row, label, is_funded_slice)] for the 1.x tab's applied amounts."""
        t = T1[p]
        ws, wsv = wb[t], wbv[t]
        t2 = T2[ROWS_BY_ONE[p]]
        g2 = MOD[ROWS_BY_ONE[p]]["grid"]
        # cells this table already carries in column D, anywhere
        blocked = set()
        for gr in g2["rows"]:
            if norm(gr["name"]) in funded:
                blocked.add((t2, "P", gr["row"]))
                blocked.add((t2, "O", gr["row"]))
                blocked.add((t2, "S", gr["row"]))
        for pp, m in MOD.items():
            for cell in m["up_cells"]:
                sh, ad = cell.split("!")
                mm = re.match(r"\$([A-Z]+)\$(\d+)", ad)
                blocked.add((sh.strip("'").replace("''", "'"), mm.group(1),
                             int(mm.group(2))))
        for nm in funded:
            for rrow in range(1, ws.max_row + 1):
                if norm(ws.cell(rrow, 2).value) == nm:
                    for cl in ("H", "I", "J", "K"):
                        blocked.add((t, cl, rrow))
        head = None
        for rrow in range(1, ws.max_row + 1):
            if ws.cell(rrow, 8).value in ("Budget line", "Funding position",
                                          "Cyber uplift program funding"):
                head = (rrow, ws.cell(rrow, 8).value)
                break
        if head is None:
            stop("no funding block on %s" % t)
        stopwords = ("Total applied", "Other cost (this model)", "Left to fund",
                     "Used for cyber FTE ($m)", "Remaining for non-people ($m)",
                     "Programme funding ($m)")
        # the tab's own 'Left to fund' line names the lights on budget cells it
        # adds back, because those fund the people, not the cost left to
        # recharge. Column M already carries them, so leave them out here too.
        lights = set()
        if head[1] == "Budget line":
            lf = find_opt(ws, 8, "Left to fund", head[0], head[0] + 20)
            ta = find_opt(ws, 8, "Total applied", head[0], head[0] + 20)
            if lf and ta:
                f = ws.cell(lf, 10).value
                if isinstance(f, str) and f.startswith("="):
                    for sh, cl, rw in refs_in(f, t):
                        if sh == t and cl == "J" and rw != ta and \
                                head[0] < rw < ta:
                            lights.add(rw)
        out = []
        for rrow in range(head[0] + 1, head[0] + 18):
            lab = ws.cell(rrow, 8).value
            if lab in stopwords:
                if lab in ("Total applied", "Left to fund",
                           "Used for cyber FTE ($m)"):
                    break
                continue
            if lab is None and ws.cell(rrow, 10).value is None:
                continue
            v = ws.cell(rrow, 10).value
            if v is None:
                continue
            slice_ = any(x in blocked for x in refs_in(v, t)) \
                if isinstance(v, str) and v.startswith("=") else False
            out.append((rrow, str(lab or "").strip(),
                        slice_ or (rrow in lights),
                        "the lights on budget line" if rrow in lights
                        else "already in Sig items funded"))
        return out

    ROWS_BY_ONE = {r[3]: r[2] for r in ROWS if r[3]}
    FUND = {}
    for p in T1:
        FUND[p] = funding_lines(p)
        keep = [x for x in FUND[p] if not x[2]]
        drop = [x for x in FUND[p] if x[2]]
        log("W4", T1[p], "funding noted: %s%s"
            % (", ".join("%s J%d" % (x[1] or "line", x[0]) for x in keep) or "none",
               "; left out, " + "; ".join("%s J%d is %s" % (x[1] or "line", x[0],
                                                            x[3]) for x in drop)
               if drop else ""))

    # ---------------------------------------------------- W5  lay the tabs out
    log.head("W5  the two Lights On tabs")
    keep_prot = None
    for old in (TAB, TAB2):
        if old in wb.sheetnames:
            if keep_prot is None:
                keep_prot = wb[old].protection
            del wb[old]
    ws = wb.create_sheet(TAB, wb.sheetnames.index(ANCHOR) + 1)
    ws2 = wb.create_sheet(TAB2, wb.sheetnames.index(ANCHOR) + 2)

    HR, D1 = 4, 5
    D2 = D1 + len(ROWS) - 1
    TOT, BUD, CTL = D2 + 1, D2 + 2, D2 + 3
    RIX = {r[0]: D1 + i for i, r in enumerate(ROWS)}

    hdr_font = Font(bold=True, color=WHITE)
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    hdr_al = Alignment(wrap_text=True, horizontal="center", vertical="center")
    band = PatternFill("solid", fgColor=BAND)
    cream = PatternFill("solid", fgColor=CREAM)
    bold = Font(bold=True)

    def lay(w, heads, title, widths):
        w.sheet_view.showGridLines = False
        w.sheet_properties.tabColor = "FF002F6C"
        for col, wd in widths.items():
            w.column_dimensions[col].width = wd
        w["B2"].value = title
        w["B2"].font = Font(bold=True, size=16)
        for col, text in heads:
            c = w[col + str(HR)]
            if text is not None:
                c.value = text
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_al
        w.row_dimensions[HR].height = 60
        w.freeze_panes = "C" + str(D1)

    W1D = {"A": 2, "B": 34, "C": 13, "D": 12.5, "E": 15, "F": 11.5, "G": 13.5,
           "H": 11.5, "I": 12.5, "J": 12.5, "K": 14, "L": 14.5, "M": 13,
           "N": 2.5, "O": 14, "P": 2.5, "Q": 16, "R": 14, "S": 13}
    lay(ws, HEADERS, "TDD Lights On", W1D)
    W2D = dict(W1D)
    for col, wd in {"M": 12, "N": 12, "O": 12, "P": 12, "Q": 12, "R": 12,
                    "S": 13, "T": 2.5, "U": 14, "V": 2.5, "W": 16, "X": 14,
                    "Y": 13}.items():
        W2D[col] = wd
    lay(ws2, HEADERS[:11] + HEAD2_MID + HEAD2_TAIL,
        "TDD Lights On, AU and NZ", W2D)
    log("W5", "%s!B%d:S%d" % (TAB, HR, HR), "his eighteen column headings, verbatim")
    log("W5", "%s!B%d:Y%d" % (TAB2, HR, HR),
        "the same rows with the AU and NZ split of the charge")

    # ------------------------------------------------------ formula builders
    def gridref(p, col, row):
        return rr(T2[p], col, row)

    def gtot(p, col):
        g = MOD[p]["grid"]
        return ('INDEX(%s!$%s$%d:$%s$%d,MATCH("Total portfolio",%s!$B$%d:$B$%d,0))'
                % (q(T2[p]), col, g["first"], col, g["total"],
                   q(T2[p]), g["first"], g["total"]))

    def govh(p):
        g = MOD[p]["grid"]
        if g["ovh_total"] is None:
            return None
        return ('INDEX(%s!$S$%d:$S$%d,MATCH("Overhead roles total",'
                "%s!$B$%d:$B$%d,0))"
                % (q(T2[p]), g["first"], g["total"], q(T2[p]), g["first"],
                   g["total"]))

    def ownovh(p, side=None):
        cells = {None: "ovh_cells", "AU": "ovh_au", "NZ": "ovh_nz"}[side]
        c = MOD[p][cells]
        return "(%s)/1000000" % plus(c) if c else "0"

    def upsum(p):
        c = MOD[p]["up_cells"]
        return "(%s)/1000000" % plus(c) if c else None

    def potref(p):
        nm = COE_POT.get(p)
        if nm is None:
            return None
        g = MOD[p]["grid"]
        return ('INDEX(%s!$S$%d:$S$%d,MATCH("%s",%s!$B$%d:$B$%d,0))'
                % (q(T2[p]), g["first"], g["total"], nm, q(T2[p]),
                   g["first"], g["total"]))

    def cust_w(label):
        a, z = RIX["Ampol Customer"], RIX["Z Customer"]
        me = a if label == "Ampol Customer" else z
        return "$E$%d/($E$%d+$E$%d)" % (me, a, z)

    # which grid rows sit on which side of the Customer split
    cust_side = {}
    if "2.2" in MOD:
        sup12 = SUP["1.2"]
        for gr in MOD["2.2"]["grid"]["rows"]:
            hit = sup12.get(norm(gr["name"]))
            cust_side[gr["name"]] = hit[3] if hit and hit[3] in ("AU", "NZ") else None

    def squad_terms(p, one, side=None, country=None):
        """support cost pairs: [(2.x S cell or country sum) x 1.x support %]."""
        sup = SUP[one]
        terms, unpaired = [], []
        for gr in MOD[p]["grid"]["rows"]:
            nm = gr["name"]
            if gr["sec"] == "Overhead roles":
                continue
            if norm(nm) in funded:
                continue
            if side is not None and cust_side.get(nm) != side:
                continue
            hit = sup.get(norm(nm))
            if not hit:
                unpaired.append(nm)
                continue
            if country is None:
                cost = gridref(p, "S", gr["row"])
            else:
                cells = [rr(T2[p], "G", ro["row"])
                         for ro in MOD[p]["groles"].get(nm, [])
                         if ro["nz"] == (country == "NZ")]
                if not cells:
                    continue
                cost = "(%s)/1000000" % plus(cells)
            terms.append("%s*%s" % (cost, rr(T1[one], "G", hit[0])))
        return terms, unpaired

    def fund_terms(p):
        return [rr(T1[p], "J", x[0]) for x in FUND[p] if not x[2]]

    # ---------------------- W5b  how each COE pair tab classifies its people
    log.head("W5b  the COE pairs, squad by squad off their planned spend lines")
    PAIRS = {}
    kind_of = dict((x[0], x[1]) for x in ROWS)
    pre_of = dict((x[0], x[2]) for x in ROWS)
    for label, (comp, partner) in COE_PAIR.items():
        p = pre_of[label]
        t = T2[p]
        wst = wb[t]
        prow = find_label(wst, 2, comp, 1, wst.max_row)
        f = str(wst.cell(prow, 3).value or "")
        if not f.startswith("="):
            stop("%s row %d carries no planned spend formula" % (t, prow))
        srows, crows = [], []
        for sh, cl, rw in refs_in(f, t):
            if sh != t:
                stop("%s planned spend %r reaches off its own tab" % (t, comp))
            (srows if cl == "S" else crows).append(rw)
        if not srows or len(crows) > 1:
            stop("%s planned spend %r reads %r, which this stage cannot "
                 "classify" % (t, comp, f))
        names = {gr["row"]: gr["name"] for gr in MOD[p]["grid"]["rows"]}
        if any(x not in names for x in srows):
            stop("%s planned spend %r names a row outside the grid" % (t, comp))
        PAIRS[label] = {"p": p, "t": t, "prow": prow, "own_rows": srows,
                        "own_names": [names[x] for x in srows],
                        "net_row": crows[0] if crows else None,
                        "partner": partner}
    for p in sorted({PAIRS[k]["p"] for k in PAIRS}, key=float):
        mine = [k for k in PAIRS if PAIRS[k]["p"] == p]
        seen = [x for k in mine for x in PAIRS[k]["own_rows"]]
        allr = [gr["row"] for gr in MOD[p]["grid"]["rows"]]
        if sorted(seen) != sorted(allr) or len(set(seen)) != len(seen):
            stop("%s: its two planned spend lines do not classify every squad "
                 "exactly once, they name %s against a grid of %s"
                 % (T2[p], sorted(seen), sorted(allr)))
    for label, pp in sorted(PAIRS.items()):
        p = pp["p"]
        potname = COE_POT.get(p)
        potrow = None
        for gr in MOD[p]["grid"]["rows"]:
            if gr["name"] == potname:
                potrow = gr["row"]
        if potrow is None:
            stop("%s has no %r squad to charge out" % (T2[p], potname))
        pp["pot_row"], pp["owns_pot"] = potrow, potrow in pp["own_rows"]
        pp["potkey"] = "bp" if potname == "TDD Business Partner" else "da"
        pp["potval"] = MOD[p]["gnet"][potname] / 1e6
        cells, au, nz, val, vau, vnz, who = [], [], [], 0.0, 0.0, 0.0, []
        for nm in pp["own_names"]:
            if nm == potname:
                continue
            for ro in MOD[p]["groles"].get(nm, []):
                pe = people[ro["key"]]
                if pe["ovh"] and pe["ovh"] != "Squad":
                    cells.append(rr(T2[p], "G", ro["row"]))
                    (nz if pe["nz"] else au).append(rr(T2[p], "G", ro["row"]))
                    val += ro["after"]
                    if pe["nz"]:
                        vnz += ro["after"]
                    else:
                        vau += ro["after"]
                    who.append("%s on %s in %s" % (pe["name"], pe["ovh"], nm))
        pp["i_cells"], pp["i_au"], pp["i_nz"] = cells, au, nz
        pp["i"], pp["i_auv"], pp["i_nzv"] = val / 1e6, vau / 1e6, vnz / 1e6
        pp["c"] = sum(MOD[p]["gnet"].get(n, 0.0) for n in pp["own_names"]) / 1e6
        pp["cau"] = sum(MOD[p]["gau"].get(n, 0.0) for n in pp["own_names"]) / 1e6
        pp["cnz"] = sum(MOD[p]["gnz"].get(n, 0.0) for n in pp["own_names"]) / 1e6
        pp["potau"] = sum(x["after"] for x in MOD[p]["groles"][potname]
                          if not x["nz"]) / 1e6 if pp["owns_pot"] else 0.0
        pp["potnz"] = (pp["potval"] - pp["potau"]) if pp["owns_pot"] else 0.0
        pp["planned"] = wbv[T2[p]].cell(pp["prow"], 3).value or 0.0
        log("W5b", "%s!C%d" % (T2[p], pp["prow"]),
            "%s takes %s, cost after levers %.6f m, own overheads %.6f m%s"
            % (label, ", ".join(pp["own_names"]), pp["c"], pp["i"],
               ", and holds the %s pot of %.6f m which columns F and G charge "
               "out" % (potname, pp["potval"]) if pp["owns_pot"] else ""))
        for x in who:
            log.note("W5b", "%s own overhead: %s" % (label, x))
    # the two lines of each pair must add back to their tab, exactly
    for p in sorted({PAIRS[k]["p"] for k in PAIRS}, key=float):
        mine = [k for k in PAIRS if PAIRS[k]["p"] == p]
        cs = sum(PAIRS[k]["c"] for k in mine)
        isum = sum(PAIRS[k]["i"] for k in mine)
        pot = PAIRS[mine[0]]["potval"]
        if abs(cs - MOD[p]["net"]) > 1e-9:
            stop("%s: the pair rows carry %.9f against a tab total of %.9f"
                 % (T2[p], cs, MOD[p]["net"]))
        if abs(isum - (MOD[p]["ownovh"] - pot)) > 1e-9:
            stop("%s: the pair rows carry own overheads of %.9f against the "
                 "tab's non pot overhead total of %.9f"
                 % (T2[p], isum, MOD[p]["ownovh"] - pot))
        for k in mine:
            nr = PAIRS[k]["net_row"]
            if nr is None:
                continue
            nv = wbv[T2[p]].cell(nr, 3).value or 0.0
            if abs(-nv - pot) > 1e-9:
                stop("%s row %d nets %.9f, not the %.9f pot"
                     % (T2[p], nr, nv, pot))
            if not PAIRS[k]["owns_pot"]:
                stop("%s nets the pot but does not hold the pot squad" % k)
        log.note("W5b", "%s: the pair carries %.6f m against the tab's %.6f m, "
                 "own overheads %.6f m against %.6f m, and the %.6f m pot sits "
                 "on the line whose squads hold it"
                 % (T2[p], cs, MOD[p]["net"], isum, MOD[p]["ownovh"] - pot, pot))

    # --------------------------------------------------------- W6  every row
    log.head("W6  the eighteen rows, in his 0.2 Data Config order")
    toggles, unpaired_note, cust_note = [], {}, []
    calc = {}
    for i, (label, kind, p2, p1, blabels) in enumerate(ROWS):
        r = D1 + i
        for w in (ws, ws2):
            w.cell(r, 2).value = label
        M = MOD[p2]
        pot = potref(p2)
        up = upsum(p2)
        W = None
        pp = PAIRS.get(label)
        if kind == "cust":
            W = cust_w(label)
        side = CUST_SIDE.get(label)

        # C  everything the row's people cost after levers
        if kind == "cust":
            mine = [gridref(p2, "S", gr["row"]) for gr in M["grid"]["rows"]
                    if cust_side.get(gr["name"]) == side]
            rest = [gridref(p2, "S", gr["row"]) for gr in M["grid"]["rows"]
                    if cust_side.get(gr["name"]) is None]
            fC = "=%s+(%s)*%s" % (plus(mine), plus(rest), W)
            cust_note.append((label,
                              [gr["name"] for gr in M["grid"]["rows"]
                               if cust_side.get(gr["name"]) == side],
                              [gr["name"] for gr in M["grid"]["rows"]
                               if cust_side.get(gr["name"]) is None]))
        elif kind == "coe":
            fC = "=" + plus([gridref(p2, "S", x) for x in pp["own_rows"]])
        elif kind == "coe1":
            fC = "=%s" % gtot(p2, "S") + ("+%s" % up if up else "")
        else:
            fC = "=%s" % gtot(p2, "S") + ("+%s" % up if up else "")

        # D  the slice funded from outside TDD
        if kind == "cust":
            fD = "=" + plus([gridref(p2, "P", gr["row"])
                             for gr in M["grid"]["rows"]
                             if cust_side.get(gr["name"]) == side and
                             norm(gr["name"]) in funded])
        elif kind == "coe":
            fD = "=0"
        elif kind == "coe1":
            fD = "=%s" % (up or "0")
        else:
            fD = "=%s" % gtot(p2, "P")

        # E  support cost
        if kind in ("pf", "cyber", "cust"):
            terms, unp = squad_terms(p2, p1, side)
            if not terms:
                stop("no support pairs for %s" % label)
            fE = "=" + plus(terms)
            if unp:
                unpaired_note[label] = unp
        elif kind == "coe":
            # L is the row's planned spend, so support cost is what is left of
            # it once the row's own overhead people are taken out
            fE = "=%s-$I%d" % (rr(T2[p2], "C", pp["prow"]), r)
        elif kind == "coe1":
            fE = "=%s-%s" % (gtot(p2, "S"), ownovh(p2))
        else:
            fE = "=0"

        # F G H  the eleven way shares
        shares = kind in ("pf", "cyber", "cust")
        if shares:
            sc = "*%s" % W if kind == "cust" else ""
            fF = "=%s/%s%s" % (POT["bp"], ELEVEN, sc)
            fG = "=%s/%s%s" % (POT["da"], ELEVEN, sc)
            fH = "=%s/%s%s" % (POT["gm"], ELEVEN, sc)
        else:
            fF = fG = fH = "=0"

        # I  the row's own overhead people
        if kind == "cust":
            fI = "=(%s)*%s" % (govh(p2), W)
        elif kind == "coe":
            fI = "=(%s)/1000000" % plus(pp["i_cells"]) if pp["i_cells"] else "=0"
        elif kind == "coe1":
            fI = "=%s" % ownovh(p2)
        elif kind == "egi":
            fI = "=0"
        else:
            o = govh(p2)
            fI = "=%s" % (o if o else ownovh(p2))

        vals = {}
        for col, f in (("C", fC), ("D", fD), ("E", fE), ("F", fF), ("G", fG),
                       ("H", fH), ("I", fI)):
            for w in (ws, ws2):
                c = w.cell(r, column_index_from_string(col))
                c.value = f
                money(c)
            if len(f) > MAXF:
                stop("%s %s formula is %d characters" % (label, col, len(f)))

        # J  the toggle, only where there is something to scale
        if kind == "coe":
            own_i = pp["i"]
        elif kind == "egi":
            own_i = 0.0
        elif kind == "coe1":
            own_i = M["ownovh"]
        else:
            own_i = M["ovhS"] if M["ovhS"] is not None else M["ownovh"]
        has_i = own_i > 1e-9
        if has_i or shares:
            for w in (ws, ws2):
                c = w.cell(r, 10)
                c.value = 1
                c.fill = cream
                c.number_format = "0%"
                c.protection = Protection(locked=False)
            toggles.append(r)
            kf = "=$F%d+$G%d+$H%d+$I%d*$J%d" % (r, r, r, r, r)
        else:
            kf = "=$F%d+$G%d+$H%d+$I%d" % (r, r, r, r)
        for w in (ws, ws2):
            w.cell(r, 11).value = kf
            w.cell(r, 12).value = "=$K%d+$E%d" % (r, r)
            money(w.cell(r, 11))
            money(w.cell(r, 12))

        # the budget cells on 0.2, live, by label
        brows = [budget_row(b) for b in blabels]
        bt = "=" + plus([rr(CFG, "E", b) for b in brows])
        bau = "=" + plus([rr(CFG, "C", b) for b in brows])
        bnz = "=" + plus([rr(CFG, "D", b) for b in brows])

        ws.cell(r, 13).value = bt                      # M budget
        ws.cell(r, 15).value = "=$L%d-$M%d" % (r, r)   # O over or under
        ws.cell(r, 17).value = "=$C%d-$D%d-$L%d" % (r, r, r)   # Q to recharge
        if kind in ("pf", "cyber"):
            fR = "=" + plus(fund_terms(p1))
        elif kind == "cust":
            fR = "=(%s)*%s" % (plus(fund_terms(p1)), W)
        elif kind == "coe" and pp["owns_pot"]:
            # the pot this line holds is already funded through the allocation
            # columns, where F and G charge it to the eleven sharing rows
            fR = "=%s" % POT[pp["potkey"]]
        else:
            fR = "=0"
        ws.cell(r, 18).value = fR                              # R noted in 1.x
        ws.cell(r, 19).value = "=$Q%d-$R%d" % (r, r)           # S still to fund
        for c in (13, 15, 17, 18, 19):
            money(ws.cell(r, c))

        # the AU NZ tab
        if kind in ("pf", "cyber", "cust"):
            au = squad_terms(p2, p1, side, "AU")[0]
            nz = squad_terms(p2, p1, side, "NZ")[0]
            eau, enz = plus(au), plus(nz)
        elif kind == "coe":
            def coe_country(cc):
                own = plus([rr(T2[p2], "G", ro["row"])
                            for nm in pp["own_names"]
                            for ro in M["groles"].get(nm, [])
                            if ro["nz"] == (cc == "NZ")])
                cells = pp["i_nz"] if cc == "NZ" else pp["i_au"]
                mine = "(%s)/1000000" % plus(cells) if cells else "0"
                po = potc(MOD, p2, T2, cc) if pp["owns_pot"] else "0"
                return "(%s)/1000000-(%s)-(%s)" % (own, mine, po)
            eau, enz = coe_country("AU"), coe_country("NZ")
        elif kind == "coe1":
            eau = "%s-%s" % (tabc(MOD, p2, T2, "AU"), ownovh(p2, "AU"))
            enz = "%s-%s" % (tabc(MOD, p2, T2, "NZ"), ownovh(p2, "NZ"))
        else:
            eau = enz = "0"
        if kind == "cust":
            iau = "(%s)*%s" % (ownovh(p2, "AU"), W)
            inz = "(%s)*%s" % (ownovh(p2, "NZ"), W)
        elif kind == "coe":
            iau = "(%s)/1000000" % plus(pp["i_au"]) if pp["i_au"] else "0"
            inz = "(%s)/1000000" % plus(pp["i_nz"]) if pp["i_nz"] else "0"
        elif kind == "egi":
            iau = inz = "0"
        else:
            iau, inz = ownovh(p2, "AU"), ownovh(p2, "NZ")
        calc[label] = dict(row=r, kind=kind, p2=p2, p1=p1, side=side,
                           brows=brows, toggle=(r in toggles))
        # filled in below once the derivation block rows are known
        calc[label]["au_parts"] = (eau, iau, shares,
                                   "*%s" % W if kind == "cust" else "")
        calc[label]["nz_parts"] = (enz, inz, shares,
                                   "*%s" % W if kind == "cust" else "")
        ws2.cell(r, 16).value = bau
        ws2.cell(r, 17).value = bnz
        ws2.cell(r, 19).value = bt
        ws2.cell(r, 21).value = "=$L%d-$S%d" % (r, r)
        ws2.cell(r, 23).value = "=$C%d-$D%d-$L%d" % (r, r, r)
        ws2.cell(r, 24).value = fR
        ws2.cell(r, 25).value = "=$W%d-$X%d" % (r, r)
        for c in (16, 17, 19, 21, 23, 24, 25):
            money(ws2.cell(r, c))

        if (i % 2) == 1:
            for w, last in ((ws, 19), (ws2, 25)):
                for c in range(2, last + 1):
                    if c == 10 and r in toggles:
                        continue
                    w.cell(r, c).fill = band
        log("W6", "%s!B%d" % (TAB, r), "%s off %s%s"
            % (label, T2[p2], " and " + T1[p1] if p1 else ""))

    for label, unp in unpaired_note.items():
        log.note("W6", "%s: no support percentage for %s, so they carry no "
                 "support cost" % (label, ", ".join(unp)))
    for label, mine, rest in cust_note:
        log.note("W6", "%s takes the 1.2 Customer squads marked %s: %s. What "
                 "sits on neither side splits in proportion to the two support "
                 "costs: %s" % (label, CUST_SIDE[label], ", ".join(mine),
                                ", ".join(rest)))

    # --------------------------------- W7  totals, budget row, control row
    log.head("W7  totals, the budget line and the control row")
    for w, cols, last in ((ws, [3, 4, 5, 6, 7, 8, 9, 11, 12, 13, 15, 17, 18, 19], 19),
                          (ws2, [3, 4, 5, 6, 7, 8, 9, 11, 12, 13, 14, 15, 16,
                                 17, 18, 19, 21, 23, 24, 25], 25)):
        w.cell(TOT, 2).value = "Total"
        for c in cols:
            cl = get_column_letter(c)
            w.cell(TOT, c).value = "=SUM(%s%d:%s%d)" % (cl, D1, cl, D2)
            money(w.cell(TOT, c))
        for c in range(2, last + 1):
            w.cell(TOT, c).font = bold
            w.cell(TOT, c).fill = band

    bud_row = find_label(cfg, 2, "Budget", 20, 34)
    ws.cell(BUD, 2).value = "Budget"
    ws.cell(BUD, 13).value = "=%s" % rr(CFG, "E", bud_row)
    ws.cell(BUD, 15).value = "=$L$%d-$M$%d" % (TOT, BUD)
    ws2.cell(BUD, 2).value = "Budget"
    ws2.cell(BUD, 16).value = "=%s" % rr(CFG, "C", bud_row)
    ws2.cell(BUD, 17).value = "=%s" % rr(CFG, "D", bud_row)
    ws2.cell(BUD, 19).value = "=%s" % rr(CFG, "E", bud_row)
    ws2.cell(BUD, 21).value = "=$L$%d-$S$%d" % (TOT, BUD)
    for w, cs in ((ws, (13, 15)), (ws2, (16, 17, 19, 21))):
        for c in cs:
            money(w.cell(BUD, c))
            w.cell(BUD, c).font = bold
        w.cell(BUD, 2).font = bold
    log("W7", "%s!M%d:O%d" % (TAB, BUD, BUD),
        "the full TDD budget line, over or under shown against the total charge")

    blk_cells = []
    for p in sorted(MOD, key=float):
        for grp in MOD[p]["blk"]["groups"]:
            blk_cells.append(rr(T2[p], "G", grp["row"]))
        blk_cells += MOD[p]["up_cells"]
    o_cells = [gtot(p, "O") for p in sorted(MOD, key=float)]

    ctl = [
        ("Control, every cost in the role mapping is represented, must be 0",
         "=ROUND(%s-SUM(%s!$%s$2:$%s$%d)/1000000,6)"
         % ("+".join(o_cells), q(REVIEW), get_column_letter(c_cost),
            get_column_letter(c_cost), rev.max_row)),
        ("Control, Total People cost against the fifteen lever tabs person by "
         "person, must be 0",
         "=ROUND($C$%d-(%s)/1000000,6)" % (TOT, plus(blk_cells))),
        ("Control, the whole of TDD after levers is this table plus the GM "
         "layer, must be 0",
         "=ROUND($C$%d+%s-((%s)/1000000+%s),6)"
         % (TOT, POT["gm"], plus(blk_cells), POT["gm"])),
        ("Control, the charge to TDD against support cost plus overheads, "
         "must be 0",
         "=ROUND($L$%d-$E$%d-$K$%d,6)" % (TOT, TOT, TOT)),
        ("Control, the three pots are fully shared out, must be 0",
         "=ROUND($F$%d+$G$%d+$H$%d-(%s+%s+%s),6)"
         % (TOT, TOT, TOT, POT["bp"], POT["da"], POT["gm"])),
        ("Control, no EGI cost sits in support, shares or overheads, must be 0",
         "=ROUND($E$%d+$F$%d+$G$%d+$H$%d+$I$%d,6)"
         % tuple([RIX["EGI"]] * 5)),
    ]
    coe_rows = sorted(RIX[x[0]] for x in ROWS if x[1] in ("coe", "coe1"))
    if coe_rows != list(range(coe_rows[0], coe_rows[-1] + 1)):
        stop("the COE rows are not contiguous, the controls assume they are")
    c1, c2 = coe_rows[0], coe_rows[-1]
    ctl += [
        ("Control, the COE lines note the two charged out pots exactly once, "
         "must be 0",
         "=ROUND(SUM($R$%d:$R$%d)-(%s+%s),6)" % (c1, c2, POT["bp"], POT["da"])),
        ("Control, all that is left to fund on the COE lines is the overhead "
         "the toggles hold back, must be 0",
         "=ROUND(SUM($S$%d:$S$%d)-SUM($I$%d:$I$%d)+SUM($K$%d:$K$%d),6)"
         % (c1, c2, c1, c2, c1, c2)),
    ]
    ws.cell(CTL, 2).value = "Control, %d checks, each must read 0" % len(ctl)
    for j, (lab, f) in enumerate(ctl):
        ws.cell(CTL, 3 + j).value = f
    white(ws, *["%s%d" % (get_column_letter(c), CTL)
                for c in range(2, 3 + len(ctl))])
    ws2.cell(CTL, 2).value = "Control, the AU and NZ split adds back to the " \
                             "charge to TDD and to Total People cost, three " \
                             "checks, each must read 0"
    ws2.cell(CTL, 3).value = "=ROUND($O$%d-$L$%d,6)" % (TOT, TOT)
    ws2.cell(CTL, 4).value = "=ROUND($M$%d+$N$%d-$O$%d,6)" % (TOT, TOT, TOT)
    white(ws2, *["%s%d" % (get_column_letter(c), CTL) for c in range(2, 6)])
    log("W7", "%s!B%d:H%d" % (TAB, CTL, CTL),
        "white font control row, every cost represented and the pots shared out")
    for w, col in ((ws, "O"), (ws2, "U")):
        w.conditional_formatting.add(
            "%s%d:%s%d" % (col, D1, col, BUD),
            FormulaRule(formula=["$%s%d>0" % (col, D1)], font=Font(color=RED)))

    dv_rows = toggles
    for w in (ws, ws2):
        dv = DataValidation(type="list", formula1='"%s"' % STEPS,
                            allow_blank=False, showErrorMessage=True,
                            errorTitle="Invalid entry",
                            error="Pick a value from the list")
        w.add_data_validation(dv)
        dv.sqref = MultiCellRange([CellRange(min_col=10, max_col=10, min_row=x,
                                             max_row=x) for x in dv_rows])
    log("W7", "%s!J%d:J%d" % (TAB, D1, D2),
        "cream toggles on %d rows, 0 to 100 percent in steps of 5, default 100"
        % len(dv_rows))

    # ------------------------------- W8  the AU NZ derivation block on 3.6
    log.head("W8  3.6: where the AU and NZ split comes from")
    dr = CTL + 2
    ws2.cell(dr, 2).value = "How the AU and NZ split is derived"
    ws2.cell(dr, 2).font = bold
    for j, h in enumerate(("Lever tab", "AU after levers ($m)",
                           "NZ after levers ($m)", "Total ($m)")):
        c = ws2.cell(dr + 1, 2 + j)
        c.value = h
        c.font = bold
        c.alignment = Alignment(wrap_text=True, horizontal="center",
                                vertical="bottom")
    ws2.row_dimensions[dr + 1].height = 30
    d0 = dr + 2
    for k, p in enumerate(sorted(MOD, key=float)):
        r = d0 + k
        ws2.cell(r, 2).value = T2[p]
        ws2.cell(r, 3).value = "=%s" % tabc(MOD, p, T2, "AU", True)
        ws2.cell(r, 4).value = "=%s" % tabc(MOD, p, T2, "NZ", True)
        ws2.cell(r, 5).value = "=$C%d+$D%d" % (r, r)
        for c in (3, 4, 5):
            money(ws2.cell(r, c))
    d1 = d0 + len(MOD) - 1
    dtot = d1 + 1
    ws2.cell(dtot, 2).value = "Total"
    for c in (3, 4, 5):
        ws2.cell(dtot, c).value = "=SUM(%s%d:%s%d)" % (get_column_letter(c), d0,
                                                       get_column_letter(c), d1)
        money(ws2.cell(dtot, c))
        ws2.cell(dtot, c).font = bold
    ws2.cell(dtot, 2).font = bold
    ws2.cell(CTL, 5).value = "=ROUND($E$%d-$C$%d,6)" % (dtot, TOT)
    AUF = "$C$%d/($C$%d+$D$%d)" % (dtot, dtot, dtot)
    NZF = "$D$%d/($C$%d+$D$%d)" % (dtot, dtot, dtot)
    bp_au = "(%s)/1000000" % plus([rr(T2["2.12"], "G", ro["row"])
                                   for ro in MOD["2.12"]["groles"]["TDD Business Partner"]
                                   if not ro["nz"]])
    bp_nz = "(%s)/1000000" % plus([rr(T2["2.12"], "G", ro["row"])
                                   for ro in MOD["2.12"]["groles"]["TDD Business Partner"]
                                   if ro["nz"]])
    da_au = "(%s)/1000000" % plus([rr(T2["2.13"], "G", ro["row"])
                                   for ro in MOD["2.13"]["groles"]["Architecture"]
                                   if not ro["nz"]])
    da_nz = "(%s)/1000000" % plus([rr(T2["2.13"], "G", ro["row"])
                                   for ro in MOD["2.13"]["groles"]["Architecture"]
                                   if ro["nz"]])
    for label, cd in calc.items():
        r = cd["row"]
        for col, parts, potau, fr in ((13, cd["au_parts"], (bp_au, da_au), AUF),
                                      (14, cd["nz_parts"], (bp_nz, da_nz), NZF)):
            e, i, shares, sc = parts
            bits = ["(%s)" % e] if e != "0" else []
            if shares:
                bits.append("(%s)/%s%s" % (potau[0], ELEVEN, sc))
                bits.append("(%s)/%s%s" % (potau[1], ELEVEN, sc))
                bits.append("%s/%s*%s%s" % (POT["gm"], ELEVEN, fr, sc))
            if i != "0":
                bits.append("(%s)%s" % (i, "*$J%d" % r if cd["toggle"] else ""))
            f = "=" + (plus(bits) if bits else "0")
            if len(f) > MAXF:
                stop("%s AU NZ formula is %d characters" % (label, len(f)))
            ws2.cell(r, col).value = f
            money(ws2.cell(r, col))
        ws2.cell(r, 15).value = "=$M%d+$N%d" % (r, r)
        ws2.cell(r, 18).value = "=$O%d-($P%d+$Q%d)" % (r, r, r)
        money(ws2.cell(r, 15))
        money(ws2.cell(r, 18))
    ws2.cell(BUD, 18).value = "=$O$%d-($P$%d+$Q$%d)" % (TOT, BUD, BUD)
    money(ws2.cell(BUD, 18))
    ws2.cell(BUD, 18).font = bold
    log("W8", "%s!B%d:E%d" % (TAB2, dr, dtot),
        "the country split of every lever tab after levers, live off the "
        "role rows")

    notes2 = [
        "AU and NZ come off each person's country in the role mapping, at "
        "their cost after levers. NZ is the people whose country reads NZ, "
        "everyone else is AU.",
        "Support cost splits inside each squad. The Business Partner and "
        "Domain Architect pots split across the people in those pots, and "
        "other overheads split across the row's own overhead people.",
        "The GM layer has no roles in the role mapping, so it splits on the AU "
        "and NZ shares of the whole of TDD in the table above.",
        "Budgets are the AU and NZ columns on 0.2 Data Config. Variance is the "
        "total charge less the two budgets, so it reads the same as over or "
        "under the lights on budget."]
    for j, n in enumerate(notes2):
        ws2.cell(dtot + 2 + j, 2).value = n
    log("W8", "%s!B%d:B%d" % (TAB2, dtot + 2, dtot + 1 + len(notes2)),
        "the split basis, in plain English")

    # -------------------------------------------- W9  the analysis block
    log.head("W9  the analysis block, live")
    ovh_ws = wb[OVH]
    o_all = find_label(ovh_ws, 2, "Overheads incl. GMs")
    o_line = {}
    for row in range(3, o_all):
        m = re.fullmatch(r"=Lists!\$AF\$(\d+)",
                         str(ovh_ws.cell(row, 2).value or ""))
        if m:
            o_line[lists.cell(int(m.group(1)), 32).value] = row
    for need in ("Technology Manager", "Head of Technology",
                 "Leadership - 8 GMs", "Delivery Manager"):
        if need not in o_line:
            stop("3.2 line %r not found" % need)

    r = CTL + 2
    head_font = Font(bold=True)

    AN1 = 7            # the analysis numbers sit clear of the long labels

    def put(row, label, cells=(), fmt=M2, bold_label=False):
        ws.cell(row, 2).value = label
        if bold_label:
            ws.cell(row, 2).font = head_font
        for j, f in enumerate(cells):
            c = ws.cell(row, AN1 + j)
            c.value = f
            c.number_format = fmt if not isinstance(fmt, (list, tuple)) else fmt[j]

    put(r, "Why the lights on number is far from the archetype", bold_label=True)
    put(r + 1, "Overhead the archetype allows for, GMs included ($m)",
        ["=%s" % rr(OVH, "I", o_all)])
    put(r + 2, "What the overhead people cost after levers, GMs included ($m)",
        ["=$I$%d+%s+%s+%s" % (TOT, POT["bp"], POT["da"], POT["gm"])])
    put(r + 3, "Where the gap sits ($m over archetype):", bold_label=True)
    put(r + 4, "Technology Managers", ["=%s" % rr(OVH, "K", o_line["Technology Manager"])])
    put(r + 5, "Heads of Technology", ["=%s" % rr(OVH, "K", o_line["Head of Technology"])])
    put(r + 6, "The 8 GMs", ["=%s" % rr(OVH, "K", o_line["Leadership - 8 GMs"])])
    put(r + 7, "Delivery Managers", ["=%s" % rr(OVH, "K", o_line["Delivery Manager"])])
    put(r + 8, "Squads after levers run under archetype, which is why the total "
               "cost never showed it. The overhead people are the lights on story.")
    r += 10

    vac = {}
    for p in sorted(MOD, key=float):
        for grp in MOD[p]["blk"]["groups"]:
            for ro in grp["roles"]:
                pe = people[ro["key"]]
                if pe["ovh"] and pe["ovh"] != "Squad" and pe["status"] == "Vacant":
                    vac.setdefault(pe["ovh"], []).append(
                        (T2[p], ro["row"], ro["lever"]))
    lines = []
    for row in range(2, 12):
        v = lists.cell(row, 32).value
        if isinstance(v, str) and v in ovh_vals:
            lines.append(v)
    for v in sorted(ovh_vals - set(lines) - {"Squad"}):
        lines.append(v)
    put(r, "Vacant overheads, the biggest dial", bold_label=True)
    for j, h in enumerate(("Vacant roles", "On Hire at full price ($m)")):
        ws.cell(r, AN1 + j).value = h
        ws.cell(r, AN1 + j).font = head_font
    rowi = r + 1
    for line in lines:
        hire = [(t, br) for t, br, lv in vac.get(line, []) if lv == "Hire"]
        if hire:
            put(rowi, line,
                ["=" + plus(['COUNTIF(%s,"Hire")' % rr(t, "E", br)
                             for t, br in hire]),
                 "=(%s)/1000000" % plus([rr(t, "G", br) for t, br in hire])],
                fmt=["0", M2])
        else:
            put(rowi, line, ["=0", "=0"], fmt=["0", M2])
        rowi += 1
    vac_tot = rowi
    put(vac_tot, "All vacant overheads on Hire",
        ["=SUM(G%d:G%d)" % (r + 1, rowi - 1), "=SUM(H%d:H%d)" % (r + 1, rowi - 1)],
        fmt=["0", M2])
    for c in (2, AN1, AN1 + 1):
        ws.cell(vac_tot, c).font = head_font
    put(vac_tot + 1,
        '="The biggest vacant overhead dial is "&INDEX($B$%d:$B$%d,MATCH(MAX('
        '$H$%d:$H$%d),$H$%d:$H$%d,0))&", "&TEXT(MAX($H$%d:$H$%d),"#,##0.00")&'
        '"m of vacancies still priced on Hire."'
        % (r + 1, rowi - 1, r + 1, rowi - 1, r + 1, rowi - 1, r + 1, rowi - 1))
    all_vac = sorted({(t, br) for ms in vac.values() for t, br, _ in ms})
    cnt = {k: plus(['COUNTIF(%s,"%s")' % (rr(t, "E", br), k)
                    for t, br in all_vac]) for k in ("Hold", "Offshore", "Filled")}
    put(vac_tot + 2,
        '="Already levered: "&TEXT(%s,"0")&" on Hold at zero, "&TEXT(%s,"0")&'
        '" offshored and "&TEXT(%s,"0")&" marked to fill."'
        % (cnt["Hold"], cnt["Offshore"], cnt["Filled"]))
    claim = vac_tot + 3
    put(claim, "Hold every vacant overhead role and the lights on total drops "
               "by this amount ($m)",
        ["=(%s)/1000000" % plus([rr(t, "G", br) for t, br in all_vac])])
    r = claim + 2

    crso = RIX["COE Cyber Risk & Service Ops"]
    var_row = find_label(cfg, 2, "Variance to full TDD budget", 20, 34)
    put(r, "The dials", bold_label=True)
    put(r + 1, "Hold the vacant overhead roles ($m)", ["=$G$%d" % claim])
    put(r + 2, "Fund the GMs above the lights on budget ($m)", ["=" + POT["gm"]])
    put(r + 3, "Bring COE Cyber Risk & Service Ops back inside its allocation "
               "($m over today)", ["=$O$%d" % crso])
    put(r + 4, "The unallocated slice of the full TDD budget ($m)",
        ["=%s" % rr(CFG, "E", var_row)])
    r += 6
    arc_ws = wb[ARC]
    arc_tot = None
    for row in range(4, arc_ws.max_row + 1):
        v = arc_ws.cell(row, 2).value
        if isinstance(v, str) and "TDD total (" in v:
            arc_tot = row
            break
    if arc_tot is None:
        stop("no TDD total row on 3.1")
    put(r, "Total archetype cost against total actual cost after levers ($m)",
        ["=%s" % rr(ARC, "D", arc_tot), "=$C$%d" % TOT])
    ws.cell(r, 2).font = head_font
    put(r + 1, "Add the GM layer and the whole of TDD after levers reads ($m)",
        ["=$C$%d+%s" % (TOT, POT["gm"])])
    r += 3
    put(r, "How to read the table", bold_label=True)
    put(r + 1, "Total People cost is everything the row's people cost after "
               "levers, funded squads included, so every cost is represented. "
               "Add the GM layer and you have the whole of TDD.")
    put(r + 2, "Sig items funded is the slice paid for outside the TDD budget: "
               "the EGI squads at their funded amount, AmPOS, CTRM and the "
               "Cyber Uplift squad, plus the COE Cyber roles charged to the "
               "cyber uplift programme.")
    put(r + 3, "EGI is funded in full from outside TDD, so its whole cost sits "
               "in Sig items funded. It carries no support cost, no share of "
               "the pots and no overheads, and nothing on that line reaches "
               "the lights on budget.")
    put(r + 4, "The three pots divide by eleven, the ten portfolios plus TDD "
               "Cyber. The COEs and EGI carry no share of them.")
    put(r + 5, "Ampol Customer and Z Customer split 2.2 Customer on the AU and "
               "NZ marks on 1.2 Customer. What sits on neither side, the "
               "overhead roles and the squads with no archetype, splits in "
               "proportion to the two support costs, and so does their share "
               "of the pots.")
    put(r + 6, "The COE pairs are split by people, not by proportion. Each "
               "line takes the squads its planned spend line on the 2.x tab "
               "names, and the overhead people in those squads. The line whose "
               "squads hold the Business Partner or Domain Architect pot shows "
               "that pot as left to recharge, and notes it again in Amount "
               "noted in 1.x tabs, because columns F and G already charge it "
               "to the eleven sharing rows. Nothing is left to fund on a COE "
               "line.")
    put(r + 7, "Amount noted in 1.x tabs leaves out the funding lines that "
               "Sig items funded already carries, and the lights on budget "
               "line, which the TDD Lights On budget column already carries.")
    log("W9", "%s!B%d" % (TAB, CTL + 2), "the analysis block, every number live")

    # protection: keep whatever the input had, toggles stay editable
    if keep_prot is not None:
        import copy as _copy
        for w in (ws, ws2):
            w.protection = _copy.deepcopy(keep_prot)
        log("W9", "%s, %s" % (TAB, TAB2),
            "the protection the input carried is kept, toggle cells unlocked")

    save(wb, dst)
    log.tail()
    print("wrote", dst)

    # ------------------------------------------------------------ self-check
    print("== self-check")
    ok = [True]

    def chk(name, cond, extra=""):
        print("%s %s%s" % ("PASS" if cond else "FAIL", name,
                           "" if cond else "   << %s" % str(extra)[:300]))
        ok[0] = ok[0] and bool(cond)

    w2b = openpyxl.load_workbook(dst)
    chk("both tabs present and visible",
        all(t in w2b.sheetnames and w2b[t].sheet_state == "visible"
            for t in (TAB, TAB2)))
    chk("3.5 sits after %s and 3.6 right after 3.5" % ANCHOR,
        w2b.sheetnames.index(TAB) == w2b.sheetnames.index(ANCHOR) + 1 and
        w2b.sheetnames.index(TAB2) == w2b.sheetnames.index(TAB) + 1)
    a, b = w2b[TAB], w2b[TAB2]
    chk("his eighteen headings on 3.5, verbatim",
        all(a[c + str(HR)].value == t for c, t in HEADERS),
        [a.cell(HR, c).value for c in range(2, 20)])
    chk("3.6 keeps C to L, adds the AU and NZ block, then his tail",
        all(b[c + str(HR)].value == t for c, t in
            HEADERS[:11] + HEAD2_MID + HEAD2_TAIL))
    want = [x[0] for x in ROWS]
    for w, t in ((a, TAB), (b, TAB2)):
        chk("%s carries the eighteen 0.2 Data Config rows in order" % t,
            [w.cell(D1 + i, 2).value for i in range(len(ROWS))] == want,
            [w.cell(D1 + i, 2).value for i in range(len(ROWS))])
    chk("no Legal row and nothing reads TDD Data",
        not any(isinstance(c.value, str) and not c.value.startswith("=") and
                ("Legal" in c.value or "TDD Data" in c.value)
                for w in (a, b) for row in w.iter_rows() for c in row) and
        not find_all(w2b[CFG], 2, "TDD Data", cfg_first, cfg_last))
    for w, t, last in ((a, TAB, 19), (b, TAB2, 25)):
        tg = [x for x in range(D1, D2 + 1)
              if w.cell(x, 10).fill.patternType == "solid" and
              w.cell(x, 10).fill.fgColor.rgb == CREAM]
        chk("%s: %d cream toggles, typed 100 percent, unlocked" % (t, len(toggles)),
            tg == toggles and all(w.cell(x, 10).value == 1 and
                                  w.cell(x, 10).number_format == "0%" and
                                  w.cell(x, 10).protection.locked is False
                                  for x in tg), tg)
        dvs = [d for d in w.data_validations.dataValidation
               if d.formula1 and "5%" in str(d.formula1)]
        chk("%s: strict dropdown on every toggle" % t,
            len(dvs) == 1 and len(dvs[0].sqref.ranges) == len(toggles))
        creamf = [c.coordinate for row in w.iter_rows() for c in row
                  if c.fill.patternType == "solid" and
                  c.fill.fgColor.rgb == CREAM and
                  isinstance(c.value, str) and c.value.startswith("=")]
        chk("%s: cream cells are typed inputs only, never formulas" % t,
            not creamf, creamf)
        texty = [(c.coordinate, c.value[:60]) for row in w.iter_rows()
                 for c in row if isinstance(c.value, str) and
                 not c.value.startswith("=") and
                 re.search("[-‐-―]", c.value)]
        chk("%s: no dash of any kind in the tab's text" % t, not texty, texty)
        banned = [(c.coordinate, c.value[:60]) for row in w.iter_rows()
                  for c in row if isinstance(c.value, str) and
                  not c.value.startswith("=") and
                  re.search(r"\b(wave|waves|seat|seats|design)\b|to projects",
                            c.value, re.I)]
        chk("%s: no banned words" % t, not banned, banned)
        whole = [c.coordinate for row in w.iter_rows() for c in row
                 if isinstance(c.value, str) and c.value.startswith("=") and
                 re.search(r"(?<![A-Z0-9$:.])\$?[A-Z]{1,3}\$?:\$?[A-Z]{1,3}"
                           r"(?![A-Z0-9(])", c.value)]
        chk("%s: bounded ranges only, no whole column references" % t,
            not whole, whole)
        longf = [(c.coordinate, len(c.value)) for row in w.iter_rows()
                 for c in row if isinstance(c.value, str) and
                 c.value.startswith("=") and len(c.value) > MAXF]
        chk("%s: every formula inside the length limit" % t, not longf, longf)
        bad = [c.coordinate for row in w.iter_rows(min_row=D1, max_row=BUD)
               for c in row if 3 <= c.column <= last and c.column != 10 and
               c.value is not None and c.number_format != M2]
        chk("%s: money format on every money cell" % t, not bad, bad)
        chk("%s: freeze panes below the headings" % t, w.freeze_panes == "C%d" % D1)
        chk("%s: one red when over rule" % t,
            len([x for x in w.conditional_formatting]) == 1)
    chk("3.5 carries %d controls, all white font" % len(ctl),
        all(isinstance(a.cell(CTL, c).value, str) and
            a.cell(CTL, c).value.startswith("=ROUND")
            for c in range(3, 3 + len(ctl))) and
        all(a.cell(CTL, c).font.color and a.cell(CTL, c).font.color.rgb == WHITE
            for c in range(2, 3 + len(ctl))))
    for label, pp in sorted(PAIRS.items()):
        rw = RIX[label]
        chk("%s takes its own squads, %.6f m" % (label, pp["c"]),
            a.cell(rw, 3).value == "=" +
            plus([gridref(pp["p"], "S", x) for x in pp["own_rows"]]))
        chk("%s prices its charge off its planned spend line" % label,
            a.cell(rw, 5).value == "=%s-$I%d" % (rr(pp["t"], "C", pp["prow"]), rw))
        chk("%s notes the pot it holds" % label if pp["owns_pot"] else
            "%s holds no pot, so it notes nothing" % label,
            a.cell(rw, 18).value ==
            ("=%s" % POT[pp["potkey"]] if pp["owns_pot"] else "=0"))
    chk("the pair rows add back to their tab, cost and own overheads",
        all(abs(sum(PAIRS[k]["c"] for k in PAIRS if PAIRS[k]["p"] == p) -
                MOD[p]["net"]) < 1e-9 and
            abs(sum(PAIRS[k]["i"] for k in PAIRS if PAIRS[k]["p"] == p) -
                (MOD[p]["ownovh"] - PAIRS[[k for k in PAIRS
                                           if PAIRS[k]["p"] == p][0]]["potval"]))
            < 1e-9 for p in {PAIRS[k]["p"] for k in PAIRS}))
    chk("3.6 carries three controls, all white font",
        all(isinstance(b.cell(CTL, c).value, str) and
            b.cell(CTL, c).value.startswith("=ROUND") for c in range(3, 6)) and
        all(b.cell(CTL, c).font.color and
            b.cell(CTL, c).font.color.rgb == WHITE for c in range(2, 6)))
    chk("the share base reads eleven, live off the ten portfolios",
        w2b["Lists"].cell(base_row, 33).value == "=COUNTA(Lists!$AS$2:$AS$12)+1")
    egi_r = RIX["EGI"]
    chk("no EGI cost in support, the shares or overheads",
        all(str(a.cell(egi_r, c).value) == "=0" for c in (5, 6, 7, 8, 9)),
        [a.cell(egi_r, c).value for c in (5, 6, 7, 8, 9)])
    chk("3.6 carries the country derivation block and the split footnote",
        b.cell(dr, 2).value == "How the AU and NZ split is derived" and
        isinstance(b.cell(dtot + 2, 2).value, str))
    w2b.close()

    if os.environ.get("W2_RECALC") == "1":
        import wbio
        rp = wbio.recalc(dst)
        vv = openpyxl.load_workbook(rp, data_only=True)
        c5 = [vv[TAB].cell(CTL, c).value for c in range(3, 3 + len(ctl))]
        c6 = [vv[TAB2].cell(CTL, c).value for c in range(3, 6)]
        chk("every control on 3.5 reads 0 after a recalculation",
            all(x is not None and abs(float(x)) < 1e-6 for x in c5), c5)
        chk("every control on 3.6 reads 0 after a recalculation",
            all(x is not None and abs(float(x)) < 1e-6 for x in c6), c6)
        errs = [(t, c.coordinate, c.value) for t in (TAB, TAB2)
                for row in vv[t].iter_rows() for c in row
                if isinstance(c.value, str) and
                any(e in c.value for e in ("#REF!", "#N/A", "#VALUE!",
                                           "#DIV/0!", "#NAME?", "#NUM!"))]
        chk("no formula error anywhere on either tab", not errs, errs[:6])
        vv.close()

    if not ok[0]:
        raise SystemExit(2)
    print("self-check clean: his columns and rows, the toggles, the controls, "
          "the formats and the hygiene rules all hold")
    return dict(TOT=TOT, BUD=BUD, CTL=CTL, D1=D1, D2=D2, HR=HR, RIX=RIX,
                toggles=toggles, dtot=dtot, d0=d0)


# ------------------------------------------------- small helpers used above

def tabc(MOD, p, T2, side, full=False):
    """The tab's cost after levers for one country, off the role rows.

    full adds back whatever the tab charges out to a programme, so the block
    on 3.6 adds up to the Total People cost column.
    """
    cells = []
    qt = "'" + T2[p].replace("'", "''") + "'"
    for grp in MOD[p]["blk"]["groups"]:
        for ro in grp["roles"]:
            if ro["nz"] == (side == "NZ"):
                cells.append("%s!$G$%d" % (qt, ro["row"]))
                if full and MOD[p]["blk"]["up"]:
                    cells.append("%s!$I$%d" % (qt, ro["row"]))
    return "(%s)/1000000" % ("+".join(cells) if cells else "0")


def potc(MOD, p, T2, side):
    nm = COE_POT.get(p)
    if nm is None:
        return "0"
    cells = ["%s!$G$%d" % ("'" + T2[p].replace("'", "''") + "'", ro["row"])
             for ro in MOD[p]["groles"][nm] if ro["nz"] == (side == "NZ")]
    return "(%s)/1000000" % ("+".join(cells) if cells else "0")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        stop("usage: w2_lightson2.py <in.xlsx> <out.xlsx>")
    main(sys.argv[1], sys.argv[2])
