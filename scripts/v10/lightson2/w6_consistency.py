#!/usr/bin/env python3
"""w6 - the tabs tell one story.

  python3 w6_consistency.py <in.xlsx> <out.xlsx>

His question: "are all tabs flowing well and the logic in each tab the same?
are all numbers adding up and aligning across every single tab so there is no
confusion?"  A cold reviewer found eleven ways the answer was no.  They are
fixed here in the order they do harm: first the numbers that contradict each
other, then the shapes that differ inside a family of like for like tabs.

  A  0.2 Data Config and 3.5 TDD Lights On priced the same eighteen rows on
     two different bases and disagreed on most of them, some of them on sign.
     0.2's spend column read the 1.x tabs, which cost the support share at
     the archetype rate;
     3.5 costs it at the actual cost after levers.  0.2's spend column now
     reads 3.5's charge column row for row, so the two tabs agree by
     construction.  His labels and his typed budget numbers are untouched.
  B  One sign for variance across the book.  Over budget was negative on 0.2
     and in the 2.x funding blocks and positive on 3.5, 3.6 and the 1.x tabs,
     while every tab prints negatives in brackets, so a bracket meant good on
     one tab and bad on the next.  Over budget now reads positive everywhere
     and the headings say Over/(under).
  C  Three totals carried the same words, "cost after levers", on three
     different bases.  No number moves: each one now says what it is, and a
     line under the table names all three with live figures.
  D  3.3 leaves the portfolio overhead allowance out of its portfolio totals
     and 3.1 puts it in, so one heading gave two answers.  Both headings now
     say which basis they are on, and 3.3 carries the reason.
  E  3.4 reports before the levers without saying so.  It says so now.
  F  1.7 Infrastructure's summary block sat one row lower than its ten
     siblings.  It is aligned, everything that read the old addresses is
     repointed, and the missing lights on budget line is put back.
  G  1.5 P&C had no TDD NZ column although 0.2 gives it a NZ budget.  It has
     the column its siblings have.
  H  The 1.x budget block had two shapes and the funding block had two.  Each
     family is brought onto its majority shape, every number preserved.
  I  One total carried four labels across 1.1, 1.4, 1.6 and 1.14.  All eleven
     now read "Total to fund" over the block and "Total" on the total line.
  J  A lever tab carries an EGI row when it has EGI people and never two.  An
     EGI row left empty by the re-homing is deleted, with its mirror row on
     3.3; no EGI row is added to a tab with no EGI people.
  K  3.2 carried a stale typed note next to a line that now reads 0.

Nothing else moves.  Every cached value in the book is snapshotted before and
after and compared cell for cell through the moves and the deletions, and any
difference that is not one of these fixes stops the stage.  3.5 and 3.6 are
read, never authored: the only marks made on them are the range ends that the
row deletion moves, and the self-check proves not one of their values changed.

Nothing is keyed on a row number: every cell is found by its label.

Idempotent: handed its own output the stage recognises the finished state and
copies it through, so a second run is md5 identical to the first.
"""
import sys, os, re, shutil, collections

V10 = "/home/user/anthropic-claude-code/scripts/v10"
sys.path.insert(0, V10)
sys.path.insert(0, os.path.join(V10, "update"))

import openpyxl
from openpyxl.utils import get_column_letter as gl
from openpyxl.utils import column_index_from_string as ci

import wbio
from _xl import (Log, load, save, copy_style, map_formulas, rewrite_refs,
                 shift_rows, _scan, _sheet_of)

TOL = 1e-9
ERRS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "Err:")
MFMT = '#,##0.00;(#,##0.00)'                 # what TEXT() prints in the notes

FAILS = []
LOG = Log("w6_consistency")


def stop(msg):
    print("STOP: %s" % msg, flush=True)
    raise SystemExit(2)


# ------------------------------------------------------------------ finding

def nz(v):
    return "" if v is None else str(v)


def norm(s):
    return " ".join(nz(s).split()).strip().lower()


def nkey(s):
    """A label as a matching key: bracketed asides dropped, letters only."""
    t = re.sub(r"\([^)]*\)", " ", nz(s))
    return re.sub(r"[^a-z0-9]+", "", t.lower())


def sheet(wb, prefix):
    for s in wb.sheetnames:
        if s.startswith(prefix):
            return wb[s]
    stop("no sheet starting %r" % prefix)


def find(ws, text, col=2, lo=1, hi=None, key=False):
    """Row whose cell in `col` carries `text` (normalised), or None."""
    hi = hi or ws.max_row
    want = nkey(text) if key else norm(text)
    for r in range(lo, hi + 1):
        v = ws.cell(r, col).value
        if not isinstance(v, str):
            continue
        got = nkey(v) if key else norm(v)
        if got == want:
            return r
    return None


def find_first(ws, texts, col=2, lo=1, hi=None):
    for t in texts:
        r = find(ws, t, col=col, lo=lo, hi=hi)
        if r:
            return r, t
    return None, None


def need(ws, text, col=2, lo=1, hi=None, key=False):
    r = find(ws, text, col=col, lo=lo, hi=hi, key=key)
    if r is None:
        stop("%s: no row labelled %r in column %s" % (ws.title, text, gl(col)))
    return r


def colof(ws, row, text, lo=1, hi=None):
    hi = hi or ws.max_column
    want = norm(text)
    for c in range(lo, hi + 1):
        if norm(ws.cell(row, c).value) == want:
            return c
    return None


def needcol(ws, row, text):
    c = colof(ws, row, text)
    if c is None:
        stop("%s: no column headed %r on row %d" % (ws.title, text, row))
    return c


def onex(wb):
    return [wb[s] for s in wb.sheetnames if re.match(r"1\.\d+ ", s)]


def twox(wb):
    return [wb[s] for s in wb.sheetnames if re.match(r"2\.\d+ ", s)]


def rowref(f, col):
    """The row number out of a reference like 'sheet'!$E$21 in formula f."""
    m = re.search(r"\$%s\$(\d+)" % col, nz(f))
    return int(m.group(1)) if m else None


# --------------------------------------------------------------- change log

class Changes:
    """Every cell this stage is allowed to move, with the reason."""

    def __init__(self):
        self.why = {}

    def add(self, ws, coord, why):
        title = ws if isinstance(ws, str) else ws.title
        self.why.setdefault((title, coord), why)

    def get(self, title, coord):
        return self.why.get((title, coord))


CH = Changes()


def put(ws, coord, value, why, before=None, fmt="%s", ch=True):
    """Write one cell, log the before and after, allow it to move."""
    old = ws[coord].value
    if old == value:
        return False
    ws[coord].value = value
    if ch:
        CH.add(ws, coord, why)
    ov = None if before is None else before.get(ws.title, {}).get(coord)
    if ov is None:
        LOG(why, "%s!%s" % (ws.title, coord), "%r -> %r" % (old, value))
    else:
        LOG(why, "%s!%s" % (ws.title, coord),
            "%r -> %r  (value %s -> recalc)" % (old, value, fmt % ov
                                                if isinstance(ov, float) else ov))
    return True


# ============================================================== A  one basis

def fix_a(wb, before):
    """0.2's spend column reads 3.5's charge column, row for row."""
    LOG.head("A  0.2 Data Config prices the rows the way 3.5 prices them")
    cfg, lo = sheet(wb, "0.2"), sheet(wb, "3.5")

    hdr = need(cfg, "Portfolio")
    cSpend = needcol(cfg, hdr, "Spend")
    cBud = needcol(cfg, hdr, "Total")
    total = need(cfg, "Total", lo=hdr + 1)

    lhdr = need(lo, "Portfolios & COEs & EGI")
    cL = needcol(lo, lhdr, "Total portfolio cost charged to TDD")
    rows35, r = {}, lhdr + 1
    while r <= lo.max_row:
        v = lo.cell(r, 2).value
        if not isinstance(v, str) or not v.strip():
            break
        k = nkey(v)
        if k.startswith("total") or k.startswith("budget"):
            break
        if k in rows35:
            stop("3.5 has two rows keyed %r" % k)
        rows35[k] = r
        r += 1
    if len(rows35) < 15:
        stop("3.5 gave only %d rows to match" % len(rows35))

    used, table = set(), []
    for r in range(hdr + 1, total):
        lab = cfg.cell(r, 2).value
        if not isinstance(lab, str) or not lab.strip():
            continue
        k = nkey(lab)
        cell = "%s%d" % (gl(cSpend), r)
        ov = before.get(cfg.title, {}).get(cell)
        if k not in rows35:
            if isinstance(cfg[cell].value, str) and cfg[cell].value.startswith("="):
                stop("0.2 row %d (%s) has a spend formula but no 3.5 row" % (r, lab))
            continue
        used.add(k)
        want = "='%s'!$%s$%d" % (lo.title, gl(cL), rows35[k])
        nv = before.get(lo.title, {}).get("%s%d" % (gl(cL), rows35[k]))
        table.append((lab, ov, nv))
        put(cfg, cell, want, "A", before)
    missing = set(rows35) - used
    if missing:
        stop("3.5 rows with no 0.2 row: %s" % sorted(missing))

    # the header, and a line under the table, say which basis the column is on
    put(cfg, "%s%d" % (gl(cSpend), hdr),
        "Spend (actual cost after levers, from 3.5 TDD Lights On)", "A")
    taga = "Spend is the actual cost after levers"
    note(cfg, "%s charged to TDD, read row for row from 3.5 TDD Lights On, so "
              "the two tabs agree by construction. The 1.x tabs price the same "
              "portfolios at archetype rates in their own summary blocks, so the "
              "over/(under) lines there are answering a different question."
         % taga, "A", taga)

    diff = [t for t in table if t[1] is None or t[2] is None
            or abs(t[1] - t[2]) > 1e-9]
    flip = [t for t in table if t[1] is not None and t[2] is not None
            and (t[1] - _bud(cfg, before, t[0], cBud, hdr, total))
            * (t[2] - _bud(cfg, before, t[0], cBud, hdr, total)) < 0]
    LOG.note("A", "%d of %d rows disagreed before this fix, %d of them on sign"
             % (len(diff), len(table), len(flip)))
    for lab, ov, nv in table:
        mark = "  " if (ov is not None and nv is not None
                        and abs(ov - nv) <= 1e-9) else "->"
        LOG.note("A", "%s %-42s 0.2 %10s   3.5 %10s"
                 % (mark, lab, _f(ov), _f(nv)))
    return cSpend, cBud, hdr, total


def _f(v):
    return "" if v is None else ("%.6f" % v)


def _bud(cfg, before, label, cBud, hdr, total):
    for r in range(hdr + 1, total):
        if norm(cfg.cell(r, 2).value) == norm(label):
            return before.get(cfg.title, {}).get("%s%d" % (gl(cBud), r)) or 0.0
    return 0.0


# ================================================== B  one sign for variance

def fix_b(wb, before, aparts):
    """Over budget reads positive everywhere."""
    LOG.head("B  over budget is positive on every tab")
    cfg = sheet(wb, "0.2")
    cSpend, cBud, hdr, total = aparts
    cVar, _ = None, None
    for t in ("Variance", "Over/(under)"):
        cVar = colof(cfg, hdr, t) or cVar
    if cVar is None:
        stop("0.2: no variance column")

    put(cfg, "%s%d" % (gl(cVar), hdr), "Over/(under)", "B")
    for r in range(hdr + 1, total):
        cell = "%s%d" % (gl(cVar), r)
        cur = cfg[cell].value
        if not (isinstance(cur, str) and cur.startswith("=")):
            continue
        put(cfg, cell, "=%s%d-%s%d" % (gl(cSpend), r, gl(cBud), r), "B", before)
    CH.add(cfg, "%s%d" % (gl(cVar), total), "B")     # the SUM follows its rows

    budget = need(cfg, "Budget", lo=total + 1)
    vfull = find(cfg, "Variance to full TDD budget", lo=budget + 1)
    if vfull:
        put(cfg, "%s%d" % (gl(cVar), vfull),
            "=$%s$%d-$%s$%d" % (gl(cSpend), total, gl(cBud), budget), "B", before)

    # the reconciliation block on the right of 0.2
    rr = None
    for row in cfg.iter_rows():
        for c in row:
            if norm(c.value) == "reconciliation ($m)":
                rr, rc = c.row, c.column
    if rr:
        cAll = needcol(cfg, rr, "Allocated budget")
        cSpn = needcol(cfg, rr, "Spend against allocation")
        cOv = colof(cfg, rr, "Over/ under total") or \
            colof(cfg, rr, "Over/(under) the allocated budget")
        cTot = colof(cfg, rr, "Variance to total") or \
            colof(cfg, rr, "Over/(under) the full TDD budget")
        if cOv:
            put(cfg, "%s%d" % (gl(cOv), rr), "Over/(under) the allocated budget", "B")
        if cTot:
            put(cfg, "%s%d" % (gl(cTot), rr), "Over/(under) the full TDD budget", "B")
        for r in range(rr + 1, rr + 8):
            if cfg.cell(r, rc).value is None:
                break
            if cOv and isinstance(cfg.cell(r, cOv).value, str) \
                    and cfg.cell(r, cOv).value.startswith("="):
                put(cfg, "%s%d" % (gl(cOv), r),
                    "=%s%d-%s%d" % (gl(cSpn), r, gl(cAll), r), "B", before)
            if cTot and isinstance(cfg.cell(r, cTot).value, str) \
                    and cfg.cell(r, cTot).value.startswith("="):
                put(cfg, "%s%d" % (gl(cTot), r),
                    "=%s%d-$%s$%d" % (gl(cSpn), r, gl(cBud), budget), "B", before)

    tagb = "Over/(under) is positive when the spend is above the budget"
    note(cfg, "%s, here and on every other tab in the book. A figure in brackets "
              "is under budget everywhere." % tagb, "B", tagb)

    # the 2.x funding blocks
    for ws in twox(wb):
        row, lab = find_first(ws, ("Variance ($m)", "Over/(under) budget ($m)"))
        if not row:
            continue
        spend = find(ws, "Total planned spend ($m)") or find(ws, "Planned spend ($m)")
        draw = find(ws, "Total budget to draw down ($m)")
        if not (spend and draw):
            stop("%s: funding block has no spend or budget line" % ws.title)
        put(ws, "B%d" % row, "Over/(under) budget ($m)", "B")
        put(ws, "C%d" % row, "=$C$%d-$C$%d" % (spend, draw), "B", before)

    # the 1.x "Other Variance" line: a shortfall reads positive, like over budget
    for ws in onex(wb):
        row = find(ws, "Other Variance")
        if not row:
            continue
        cur = nz(ws["C%d" % row].value)
        if cur.startswith("=-"):
            put(ws, "C%d" % row, "=" + cur[2:], "B", before)
            tot = find(ws, "Total", lo=row + 1, hi=row + 2) or \
                find(ws, "Total over budget", lo=row + 1, hi=row + 2)
            if tot:
                CH.add(ws, "C%d" % tot, "B")


# ================================== C, D, E  a label that states its basis

NOTE_AT = {}
NOTE_STYLE = []


def set_note_style(wb):
    """Write footnotes the way 3.5 writes its own."""
    a5 = sheet(wb, "3.5")
    for r in range(a5.max_row, 1, -1):
        v = a5.cell(r, 2).value
        if isinstance(v, str) and not v.startswith("=") and len(v) > 90:
            NOTE_STYLE.append(a5.cell(r, 2))
            return


def last_used(ws):
    last = 0
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                last = max(last, c.row)
    return last


def note(ws, text, why, tag):
    """Write a footnote line under the tab, or rewrite this stage's own line."""
    row = None
    for r in range(1, last_used(ws) + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and tag in v:
            row = r
            break
    if row is None:
        row = NOTE_AT.get(ws.title) or (last_used(ws) + 2)
        NOTE_AT[ws.title] = row + 1
        if NOTE_STYLE:
            copy_style(NOTE_STYLE[0], ws.cell(row, 2))
    cur = ws.cell(row, 2).value
    if cur is not None and not (isinstance(cur, str) and tag in cur):
        stop("%s!B%d is not free for the %s note (%r)" % (ws.title, row, why, cur))
    put(ws, "B%d" % row, text, why)


def fix_cde(wb, before):
    LOG.head("C  three totals called cost after levers, each on its own basis")
    a1, a3, a5 = sheet(wb, "3.1"), sheet(wb, "3.3"), sheet(wb, "3.5")
    h1 = need(a1, "Line")
    c_after = needcol(a1, h1, "Cost after levers ($m)") if \
        colof(a1, h1, "Cost after levers ($m)") else None
    if c_after is None:
        c_after = colof(a1, h1,
                        "Cost after levers, the Business Partner and Domain "
                        "Architect pots netted out of the COE lines ($m)")
    if c_after is None:
        stop("3.1: no cost after levers column")
    c_out = colof(a1, h1, "Funded outside TDD ($m)") or \
        colof(a1, h1, "Funded outside TDD, before the cyber uplift recharge ($m)")
    c_var1 = colof(a1, h1, "Variance to archetype ($m)") or \
        colof(a1, h1, "Variance to archetype, the portfolio allowance included ($m)")
    tot1 = need(a1, "TDD total", key=False) if find(a1, "TDD total") else None
    if tot1 is None:
        for r in range(h1 + 1, a1.max_row + 1):
            v = a1.cell(r, 2).value
            if isinstance(v, str) and v.startswith("=") and "TDD total" in v:
                tot1 = r
                break
            if norm(v).startswith("tdd total"):
                tot1 = r
                break
    if tot1 is None:
        stop("3.1: no TDD total row")

    h3 = need(a3, "Portfolio")
    c_after3 = colof(a3, h3, "Cost after levers ($m)") or \
        colof(a3, h3, "Cost after levers, before the cyber uplift recharge ($m)")
    c_arch3 = colof(a3, h3, "Archetype cost ($m)") or \
        colof(a3, h3, "Archetype cost, the portfolio allowance excluded ($m)")
    c_var3 = colof(a3, h3, "Variance to archetype ($m)") or \
        colof(a3, h3, "Variance to archetype, the portfolio allowance excluded ($m)")
    tot3 = need(a3, "Group total", lo=h3 + 1)

    h5 = need(a5, "Portfolios & COEs & EGI")
    c_c5 = needcol(a5, h5, "Total People cost")
    c_d5 = needcol(a5, h5, "Sig items funded")
    tot5 = None
    for r in range(h5 + 1, a5.max_row + 1):
        if nkey(a5.cell(r, 2).value).startswith("total"):
            tot5 = r
            break
    if tot5 is None:
        stop("3.5: no total row")

    put(a1, "%s%d" % (gl(c_after), h1),
        "Cost after levers, the Business Partner and Domain Architect pots "
        "netted out of the COE lines ($m)", "C")
    put(a1, "%s%d" % (gl(c_out), h1),
        "Funded outside TDD, before the cyber uplift recharge ($m)", "C")
    put(a3, "%s%d" % (gl(c_after3), h3),
        "Cost after levers, before the cyber uplift recharge ($m)", "C")

    tag = "Cost after levers reads three ways"
    body = ('="%s. Here it is "&TEXT($%s$%d,"%s")&"m, with the Business Partner '
            'and Domain Architect pots netted out of the COE lines. On 3.3 Squad '
            'Actuals to Archetype it is "&TEXT(\'%s\'!$%s$%d,"%s")&"m, the pots '
            'left in and the cyber uplift recharge out. On 3.5 TDD Lights On it '
            'is "&TEXT(\'%s\'!$%s$%d,"%s")&"m, everything in. Funded outside TDD '
            'reads "&TEXT($%s$%d,"%s")&"m here against "&TEXT(\'%s\'!$%s$%d,"%s")'
            '&"m there, the difference being the same cyber uplift recharge."'
            % (tag, gl(c_after), tot1, MFMT, a3.title, gl(c_after3), tot3, MFMT,
               a5.title, gl(c_c5), tot5, MFMT, gl(c_out), tot1, MFMT,
               a5.title, gl(c_d5), tot5, MFMT))
    note(a1, body, "C", tag)

    body3 = ('="%s. Here it is "&TEXT($%s$%d,"%s")&"m, the Business Partner and '
             'Domain Architect pots left in and the cyber uplift recharge out. On '
             '3.1 Archetype to Actuals it is "&TEXT(\'%s\'!$%s$%d,"%s")&"m, the '
             'pots netted out of the COE lines. On 3.5 TDD Lights On it is '
             '"&TEXT(\'%s\'!$%s$%d,"%s")&"m, everything in."'
             % (tag, gl(c_after3), tot3, MFMT, a1.title, gl(c_after), tot1, MFMT,
                a5.title, gl(c_c5), tot5, MFMT))
    note(a3, body3, "C", tag)

    LOG.head("D  one heading, two bases: 3.3 leaves the portfolio allowance out")
    put(a1, "%s%d" % (gl(c_var1), h1),
        "Variance to archetype, the portfolio allowance included ($m)", "D")
    put(a3, "%s%d" % (gl(c_var3), h3),
        "Variance to archetype, the portfolio allowance excluded ($m)", "D")
    put(a3, "%s%d" % (gl(c_arch3), h3),
        "Archetype cost, the portfolio allowance excluded ($m)", "D")
    LOG.note("D", "labelled rather than merged: 3.3 is a squad level table and "
                  "the allowance is not a squad, so adding it would either "
                  "invent rows or leave a total that does not equal the rows "
                  "above it. No figure moves.")

    ar1 = find(a1, "Ampol Retail", col=2, lo=h1 + 1)
    ar3 = find(a3, "Ampol Retail total", col=2, lo=h3 + 1)
    tagd = "Variance to archetype here leaves out"
    if ar1 and ar3:
        bodyd = ('="%s the portfolio overhead allowance, the Business Partner, '
                 'Domain Architect and leadership share the archetype gives every '
                 'portfolio, which 3.1 Archetype to Actuals and the 2.x totals '
                 'carry. Ampol Retail reads "&TEXT($%s$%d,"%s")&"m here and '
                 '"&TEXT(\'%s\'!$%s$%d,"%s")&"m on 3.1. Every squad, overhead role '
                 'and funded program on the 2.x tabs is priced here; the allowance '
                 'line is not a squad, so it is not."'
                 % (tagd, gl(c_var3), ar3, MFMT, a1.title, gl(c_var1), ar1, MFMT))
        note(a3, bodyd, "D", tagd)

    LOG.head("E  3.4 reports before the levers, and now says so")
    a4 = sheet(wb, "3.4")
    h4 = need(a4, "COE")
    ren = [("Cost ($m)", "Cost before levers ($m)"),
           ("Cost - AU ($m)", "Cost before levers, AU ($m)"),
           ("Cost - NZ ($m)", "Cost before levers, NZ ($m)"),
           ("Cost of overhead roles ($m)",
            "Cost of overhead roles before levers ($m)")]
    for old, new in ren:
        c = colof(a4, h4, old)
        if c:
            put(a4, "%s%d" % (gl(c), h4), new, "E")
    tot4 = need(a4, "COEs and EGI total", lo=h4 + 1)
    cyber4 = find(a4, "Cyber, Risk & Service Operations total", lo=h4 + 1)
    cG4 = colof(a4, h4, "Cost before levers ($m)")
    tage = "Every cost on this tab is before the vacancy levers"
    if cyber4:
        bodye = ('="%s: it prices the people in the role mapping at their full '
                 'effective cost. 3.5 TDD Lights On prices the same COEs after the '
                 'levers, which is why Cyber, Risk and Service Operations reads '
                 '"&TEXT($%s$%d,"%s")&"m here and "&TEXT(\'%s\'!$%s$%d,"%s")&"m '
                 'there."' % (tage, gl(cG4), cyber4, MFMT, a5.title, gl(c_c5),
                              need(a5, "COE Cyber Risk & Service Ops", lo=h5 + 1),
                              MFMT))
        note(a4, bodye, "E", tage)



# ============================================ F  1.7 aligned with its siblings

def block_cols(a, b):
    return list(range(ci(a), ci(b) + 1))


def guard_refs(wb, title, cols, r1, r2, delta):
    """Stop if a range straddles the block being moved."""
    bad = []

    def look(own, coord, f):
        for s, e, m in _scan(f):
            if _sheet_of(m, own) != title or m.group("c2") is None:
                continue
            c1 = ci(m.group("c1").lstrip("$"))
            c2 = ci(m.group("c2").lstrip("$"))
            a = int(m.group("r1").lstrip("$"))
            b = int(m.group("r2").lstrip("$"))
            if not (set(range(c1, c2 + 1)) & set(cols)):
                continue
            ins = [r1 <= x <= r2 for x in (a, b)]
            if any(ins) and not all(ins):
                bad.append("%s!%s %s" % (own, coord, f[:70]))
        return f

    map_formulas(wb, look)
    if bad:
        stop("a range straddles the block being moved: %s" % bad[:4])


def move_block(wb, ws, cols, r1, r2, delta, why, blank_row):
    """Move a rectangle of cells by `delta` rows and repoint every reference."""
    guard_refs(wb, ws.title, cols, r1, r2, delta)
    lo, hi = min(cols), max(cols)
    merges = [m for m in list(ws.merged_cells.ranges)
              if m.min_col >= lo and m.max_col <= hi
              and r1 <= m.min_row and m.max_row <= r2]
    for m in merges:
        ws.unmerge_cells(str(m))
    order = range(r1, r2 + 1) if delta < 0 else range(r2, r1 - 1, -1)
    for r in order:
        for c in cols:
            src, dst = ws.cell(r, c), ws.cell(r + delta, c)
            dst.value = src.value
            copy_style(src, dst)
            src.value = None
            copy_style(ws.cell(blank_row, c), src)

    def fn(sh, c1, rr1, c2, rr2):
        if sh != ws.title:
            return (c1, rr1, c2, rr2)
        a = ci(c1.lstrip("$"))
        b = ci(c2.lstrip("$")) if c2 else a
        if not (lo <= a and b <= hi):
            return (c1, rr1, c2, rr2)
        n1 = rr1 + delta if r1 <= rr1 <= r2 else rr1
        n2 = None if rr2 is None else (rr2 + delta if r1 <= rr2 <= r2 else rr2)
        return (c1, n1, c2, n2)

    n = map_formulas(wb, lambda s, coord, f: rewrite_refs(f, s, fn))
    LOG(why, "%s move" % ws.title, "%s%d:%s%d moved %+d rows, %d formulas repointed"
        % (gl(lo), r1, gl(hi), r2, delta, n))
    for m in merges:
        ws.merge_cells(start_row=m.min_row + delta, end_row=m.max_row + delta,
                       start_column=m.min_col, end_column=m.max_col)
    return n


def fix_f(wb, before):
    LOG.head("F  1.7 Infrastructure's summary block aligned with its ten siblings")
    tabs = onex(wb)
    rows = {}
    for ws in tabs:
        r = find(ws, "Total Cost")
        if r:
            rows.setdefault(r, []).append(ws.title)
    target = max(rows, key=lambda k: len(rows[k]))
    odd = [(r, t) for r, ts in rows.items() if r != target for t in ts]
    if not odd:
        LOG.note("F", "every 1.x tab already carries Total Cost on row %d" % target)
        return
    for r, title in odd:
        ws = wb[title]
        delta = target - r
        if delta != -1:
            stop("%s: Total Cost is %d rows out, expected 1" % (title, -delta))
        sec = need(ws, "Portfolio Summary")
        last = need(ws, "Variance to actuals", col=5)
        kfirst = need(ws, "Actuals vs archetype", col=11)
        klast = need(ws, "Variance", col=11, lo=kfirst)
        for rr in range(sec, last + 1):
            for c in [1] + list(range(15, ws.max_column + 1)):
                if ws.cell(rr, c).value is not None:
                    stop("%s!%s is in the way of the move"
                         % (title, ws.cell(rr, c).coordinate))
        for rr in range(sec + delta, last + 1):
            for c in block_cols("B", "G"):
                CH.add(ws, "%s%d" % (gl(c), rr), "F")
        for rr in range(kfirst + delta, klast + 1):
            for c in block_cols("K", "N"):
                CH.add(ws, "%s%d" % (gl(c), rr), "F")

        move_block(wb, ws, block_cols("B", "G"), sec, last, delta, "F", sec - 1)
        move_block(wb, ws, block_cols("K", "N"), kfirst, klast, delta, "F", sec - 1)

        # the H block keeps its rows: only its section header moves up, and the
        # lights on budget line the ten siblings carry is put back underneath
        hsec = need(ws, "Budget vs TDD Cost", col=8)
        copy_style(ws.cell(hsec, 8), ws.cell(hsec + delta, 8))
        put(ws, "H%d" % (hsec + delta), ws.cell(hsec, 8).value, "F")
        put(ws, "H%d" % hsec, None, "F")
        au = need(ws, "AU Budget ($m)", col=8)
        src = None
        for other in tabs:
            if other.title != title and find(other, "TDD Lights On Budget "
                                             "(people - 0.2 Data Config)", col=8):
                src = other
                break
        if src is None:
            stop("no sibling carries the lights on budget line")
        srow = find(src, "TDD Lights On Budget (people - 0.2 Data Config)", col=8)
        cfgrow = rowref(ws["I%d" % au].value, "C")
        if cfgrow is None:
            stop("%s: cannot read the 0.2 row from %r" % (title, ws["I%d" % au].value))
        copy_style(src.cell(srow, 8), ws.cell(au - 1, 8))
        copy_style(src.cell(srow, 9), ws.cell(au - 1, 9))
        put(ws, "H%d" % (au - 1),
            "TDD Lights On Budget (people - 0.2 Data Config)", "F")
        put(ws, "I%d" % (au - 1), "='%s'!$E$%d" % (sheet(wb, "0.2").title, cfgrow),
            "F", before)

        # heights follow the block, so the tall wrapping rows stay tall
        h = {r2: (ws.row_dimensions[r2].height if r2 in ws.row_dimensions else None)
             for r2 in range(sec - 1, last + 2)}
        for r2 in range(sec, last + 1):
            ws.row_dimensions[r2 + delta].height = h.get(r2)
        ws.row_dimensions[last].height = h.get(sec - 1)

        # the mini block's TDD variance line reads the two over/(under) lines
        tv = find(ws, "TDD Variance")
        aur = need(ws, "AU over/(under) budget ($m)", col=8)
        nzr = need(ws, "NZ over/(under) budget ($m)", col=8)
        if tv:
            put(ws, "C%d" % tv, "=I%d+I%d" % (aur, nzr), "F", before)


# ================================================== G  1.5 gets its NZ column

def fix_g(wb, before):
    LOG.head("G  1.5 P&C gets the TDD NZ column its siblings have")
    for ws in onex(wb):
        hdr = find(ws, "Cost")
        if hdr is None:
            continue
        cAU = colof(ws, hdr, "TDD AU ($m)")
        cNZ = colof(ws, hdr, "TDD NZ ($m)")
        if cNZ or cAU is None:
            continue
        nzc = cAU + 1
        tot = need(ws, "Total Cost")
        for r in range(hdr, tot + 1):
            if ws.cell(r, nzc).value is not None:
                stop("%s: column %s is not free for the NZ column"
                     % (ws.title, gl(nzc)))
        copy_style(ws.cell(hdr, cAU), ws.cell(hdr, nzc))
        put(ws, "%s%d" % (gl(nzc), hdr), "TDD NZ ($m)", "G")
        for r in range(hdr + 1, tot):
            f = ws.cell(r, cAU).value
            if not (isinstance(f, str) and f.startswith("=")):
                continue
            m = re.match(r"^(=IF\(.*),0,(.*)\)$", f)
            if m:
                new = "%s,%s,0)" % (m.group(1), m.group(2))
            elif '"AU"' in f:
                new = f.replace('"AU"', '"NZ"')
            else:
                stop("%s!%s: cannot mirror %r" % (ws.title, gl(cAU) + str(r), f))
            copy_style(ws.cell(r, cAU), ws.cell(r, nzc))
            put(ws, "%s%d" % (gl(nzc), r), new, "G", before)
        copy_style(ws.cell(tot, cAU), ws.cell(tot, nzc))
        put(ws, "%s%d" % (gl(nzc), tot),
            "=SUM(%s%d:%s%d)" % (gl(nzc), hdr + 1, gl(nzc), tot - 1), "G", before)
        w = ws.column_dimensions[gl(cAU)]
        LOG.note("G", "%s: NZ column written at %s, AU column width %s kept"
                 % (ws.title, gl(nzc), w.width))


# ============================== H  one shape for the budget and funding blocks

def fix_h_budget(wb, before):
    LOG.head("H  the 1.x budget block, all eleven on the majority shape")
    tabs = onex(wb)
    shape = collections.Counter()
    for ws in tabs:
        shape[bool(find(ws, "AU over/(under) budget ($m)", col=8))] += 1
    if shape[True] < shape[False]:
        stop("the AU/NZ/TDD budget block is not the majority shape")
    for ws in tabs:
        if find(ws, "AU over/(under) budget ($m)", col=8):
            continue
        hsec = need(ws, "Budget vs TDD Cost", col=8)
        lights = find(ws, "TDD Lights On Budget (people - 0.2 Data Config)", col=8)
        if lights != hsec + 1:
            stop("%s: the lights on budget line is not under the section head"
                 % ws.title)
        tdd = find(ws, "TDD Budget ($m)", col=8)
        if tdd is None:
            stop("%s: no TDD budget line to replace" % ws.title)
        cfgrow = rowref(ws["I%d" % tdd].value, "E")
        if cfgrow is None:
            stop("%s: cannot read the 0.2 row from %r" % (ws.title, ws["I%d" % tdd].value))
        tot = need(ws, "Total Cost")
        cAU = colof(ws, need(ws, "Cost"), "TDD AU ($m)")
        cNZ = colof(ws, need(ws, "Cost"), "TDD NZ ($m)")
        cfg = sheet(wb, "0.2").title
        # the donor row for a line this tab has not carried before
        donor = tdd + 2
        lines = [(tdd, "AU Budget ($m)", "='%s'!$C$%d" % (cfg, cfgrow)),
                 (tdd + 1, "NZ Budget ($m)", "='%s'!$D$%d" % (cfg, cfgrow)),
                 (tdd + 2, "AU over/(under) budget ($m)",
                  "=%s%d-I%d" % (gl(cAU), tot, tdd)),
                 (tdd + 3, "NZ over/(under) budget ($m)",
                  "=%s%d-I%d" % (gl(cNZ), tot, tdd + 1)),
                 (tdd + 4, "TDD over/(under) budget ($m)",
                  "=I%d+I%d" % (tdd + 2, tdd + 3))]
        for r, lab, f in lines:
            for c in (8, 9):
                if ws.cell(r, c).value is None:
                    copy_style(ws.cell(donor, c), ws.cell(r, c))
            put(ws, "H%d" % r, lab, "H", before)
            put(ws, "I%d" % r, f, "H", before)
        tv = find(ws, "TDD Variance")
        if tv:
            put(ws, "C%d" % tv, "=I%d+I%d" % (tdd + 2, tdd + 3), "H", before)


def fix_h_funding(wb, before):
    LOG.head("H  the 1.x funding block, all eleven on the majority shape")
    tabs = onex(wb)
    heads = [(ws, find(ws, "Budget line", col=8)) for ws in tabs]
    if sum(1 for _, r in heads if r) < len(tabs) / 2.0:
        stop("the budget line table is not the majority shape")
    donor = next(ws for ws, r in heads if r)
    drow = find(donor, "Budget line", col=8)
    for ws in tabs:
        fp = find(ws, "Funding position", col=8)
        if fp is None:
            continue
        sec = find(ws, "Other funding", col=8)
        oth = need(ws, "Other cost (this model)", col=8)
        left = need(ws, "Left to fund", col=8)
        lines = []                       # the funding lines, in their own order
        for r in range(fp + 1, left + 1):
            lab = ws.cell(r, 8).value
            if r in (oth, left) or lab is None:
                continue
            lines.append((r, lab, ws.cell(r, 10).value))
        if not lines:
            stop("%s: funding block has no funding lines" % ws.title)
        first, last = lines[0][0], lines[-1][0]
        if first <= fp:
            stop("%s: unexpected funding block order" % ws.title)
        # section head sits directly over the table head; the lines keep their
        # rows, so anything that reads them by address still reads them
        newsec = first - 2
        newhdr = first - 1
        if newsec < sec:
            stop("%s: no room for the table head" % ws.title)
        if newsec != sec:
            put(ws, "H%d" % newsec, ws.cell(sec, 8).value, "H")
            copy_style(ws.cell(sec, 8), ws.cell(newsec, 8))
            put(ws, "H%d" % sec, None, "H")
        for c, lab in ((8, "Budget line"), (9, "Budget ($m)"),
                       (10, "Amount that can be allocated to people"),
                       (11, "Remaining for non-people ($m)")):
            copy_style(donor.cell(drow, c), ws.cell(newhdr, c))
            put(ws, "%s%d" % (gl(c), newhdr), lab, "H")
        applied, ocost, lfund = last + 1, last + 2, last + 3
        for r in (applied, ocost, lfund):
            for c in (8, 10):
                copy_style(ws.cell(last, c), ws.cell(r, c))
        put(ws, "H%d" % applied, "Total applied", "H")
        put(ws, "J%d" % applied, "=SUM(J%d:J%d)" % (first, last), "H", before)
        put(ws, "H%d" % ocost, "Other cost (this model)", "H")
        put(ws, "J%d" % ocost, "=E%d" % need(ws, "Total Cost"), "H", before)
        put(ws, "H%d" % lfund, "Left to fund", "H")
        put(ws, "J%d" % lfund, "=J%d-J%d" % (ocost, applied), "H", before)
        for r in range(lfund + 1, left + 1):
            for c in (8, 9, 10, 11):
                if ws.cell(r, c).value is not None:
                    put(ws, "%s%d" % (gl(c), r), None, "H")
        ov = find(ws, "Other Variance")
        if ov:
            put(ws, "C%d" % ov, "=J%d" % lfund, "H", before)
        LOG.note("H", "%s: funding lines kept on rows %d to %d so every "
                      "reference to them still lands" % (ws.title, first, last))


# ======================================================= I  one label, eleven

MINI = ("total to fund", "total off budget", "budget position",
        "total over budget", "total")


def fix_i(wb, before):
    LOG.head("I  one name for the block and one for its total, on all eleven")
    for ws in onex(wb):
        tv = find(ws, "TDD Variance")
        ov = find(ws, "Other Variance")
        if not (tv and ov):
            continue
        head = None
        for r in range(tv - 1, max(tv - 6, 1), -1):
            if norm(ws.cell(r, 2).value) in MINI:
                head = r
                break
        if head is None:
            stop("%s: no head over the TDD variance block" % ws.title)
        tot = None
        for r in range(ov + 1, ov + 3):
            if norm(ws.cell(r, 2).value) in MINI:
                tot = r
                break
        if tot is None:
            stop("%s: no total under the variance block" % ws.title)
        put(ws, "B%d" % head, "Total to fund", "I")
        put(ws, "B%d" % tot, "Total", "I")


# ============================================ J  an EGI row where EGI people are

def fix_j(wb, before):
    LOG.head("J  a lever tab carries an EGI row when it has EGI people, never two")
    egi_tab = sheet(wb, "2.14")
    a3 = sheet(wb, "3.3")
    kill, found = [], {}
    for ws in twox(wb):
        if ws is egi_tab:
            continue
        hdr = need(ws, "Squad")
        tot = need(ws, "Total portfolio", lo=hdr + 1)
        seen = []
        for r in range(hdr + 1, tot):
            lab = ws.cell(r, 2).value
            if not isinstance(lab, str) or not nkey(lab).startswith("egi"):
                continue
            vals = before.get(ws.title, {})
            roles = vals.get("F%d" % r) or 0
            cost = vals.get("O%d" % r) or 0
            after = vals.get("S%d" % r) or 0
            seen.append((r, lab, roles, cost, after))
        found[ws.title] = seen
        for r, lab, roles, cost, after in seen:
            empty = (not roles) and abs(cost) < TOL and abs(after) < TOL
            LOG.note("J", "%-30s %-16s roles %-4s cost %-12s %s"
                     % (ws.title, lab, roles, "%.6f" % cost,
                        "empty" if empty else "carries EGI people, kept"))
            if empty:
                kill.append((ws, r, lab))
        if len(seen) > 1 and not any(
                (not x[2]) and abs(x[3]) < TOL for x in seen):
            stop("%s carries two EGI rows and neither is empty" % ws.title)
    if not kill:
        LOG.note("J", "no empty EGI row to remove")
        return
    # the mirror row on 3.3, and anything else that reads the row
    for ws, r, lab in kill:
        mirror = None
        for rr in range(1, a3.max_row + 1):
            f = a3.cell(rr, 4).value
            if isinstance(f, str) and f.startswith("=") and \
                    re.search(r"'%s'!\$B\$%d\b" % (re.escape(ws.title), r), f):
                mirror = rr
        others = []
        for other in wb.worksheets:
            for row in other.iter_rows():
                for c in row:
                    f = c.value
                    if not (isinstance(f, str) and f.startswith("=")):
                        continue
                    for s, e, m in _scan(f):
                        if _sheet_of(m, other.title) != ws.title:
                            continue
                        a = int(m.group("r1").lstrip("$"))
                        b = int(m.group("r2").lstrip("$")) if m.group("r2") else a
                        if a != r or b != r:
                            continue
                        if other is a3 and c.row == mirror:
                            continue      # the mirror row goes with it
                        if other is ws and c.row == r:
                            continue      # the row's own cells
                        others.append("%s!%s" % (other.title, c.coordinate))
        if others:
            # whatever still reads the empty row means the EGI line, so it is
            # sent to the row that now carries the EGI people
            keep = [rr for rr, lab2, ro, co, af in found[ws.title] if rr != r]
            if len(keep) != 1:
                LOG("J", "%s r%d" % (ws.title, r),
                    "empty %r row KEPT: %s still reads it and the tab has no "
                    "other EGI row to send them to" % (lab, sorted(set(others))[0]))
                continue
            k = keep[0]

            def to_keep(sheet_name, c1, r1, c2, r2, _t=ws.title, _r=r, _k=k):
                if sheet_name != _t or c2 is not None or r1 != _r:
                    return (c1, r1, c2, r2)
                return (c1, _k, c2, r2)

            for where in sorted(set(others)):
                title, coord = where.split("!")
                cell = wb[title][coord]
                new = rewrite_refs(cell.value, title, to_keep)
                LOG("J", where, "read the empty %s row, now reads the %s row"
                    % (lab, wb[ws.title].cell(k, 2).value))
                cell.value = new
                CH.add(title, coord, "J")
                for cc in wb[title].iter_rows():
                    for c3 in cc:
                        if isinstance(c3.value, str) and c3.value.startswith("=") \
                                and coord in c3.value:
                            CH.add(title, c3.coordinate, "J")
        if mirror:
            if J_SHIFT.get(a3.title):
                stop("two rows deleted on %s, the snapshot map takes one" % a3.title)
            LOG("J", "%s r%d" % (a3.title, mirror),
                "mirror of the empty %s row, deleted with it" % lab)
            shift_rows(wb, a3.title, mirror, -1)
            CH.add(a3, "*rows", "J")
            J_SHIFT.setdefault(a3.title, []).append(mirror)
        if J_SHIFT.get(ws.title):
            stop("two rows deleted on %s, the snapshot map takes one" % ws.title)
        LOG("J", "%s r%d" % (ws.title, r), "empty %r row deleted" % lab)
        shift_rows(wb, ws.title, r, -1)
        CH.add(ws, "*rows", "J")
        J_SHIFT.setdefault(ws.title, []).append(r)


J_SHIFT = {}


# ============================================== K  the stale hand typed label

def fix_k(wb, before):
    LOG.head("K  3.2's stale note next to a line that reads 0")
    ws = sheet(wb, "3.2")
    hdr = need(ws, "Overhead roles")
    cCount = needcol(ws, hdr, "Actual number of leadership roles")
    cWhere = needcol(ws, hdr, "Where each role sits")
    for r in range(hdr + 1, ws.max_row + 1):
        v = ws.cell(r, cWhere).value
        if not isinstance(v, str) or v.startswith("=") or "in customer" not in v.lower():
            continue
        n = before.get(ws.title, {}).get("%s%d" % (gl(cCount), r))
        if n:
            LOG.note("K", "kept: %r sits against %s roles" % (v, n))
            continue
        put(ws, "%s%d" % (gl(cWhere), r), None, "K")


# ================================================================ self-check

def check(cond, text):
    print("%-4s %s" % ("PASS" if cond else "FAIL", text), flush=True)
    if not cond:
        FAILS.append(text)
    return cond


def fwd_map(title, coord):
    """Where a before-cell ends up after the moves, or None if it is gone."""
    m = re.match(r"([A-Z]+)(\d+)$", coord)
    c, r = ci(m.group(1)), int(m.group(2))
    for t, cols, r1, r2, delta in MOVES:
        if t == title and c in cols and r1 <= r <= r2:
            r += delta
    for row in sorted(J_SHIFT.get(title, []), reverse=True):
        if r == row:
            return None
        if r > row:
            r -= 1
    return "%s%d" % (gl(c), r)


MOVES = []


def boxes_read(wf):
    """{(sheet, coord): [(sheet, c1, r1, c2, r2)]} for every formula cell."""
    deps = {}
    for ws in wf.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f = c.value
                if not (isinstance(f, str) and f.startswith("=")):
                    continue
                box = []
                for s, e, m in _scan(f):
                    sh = _sheet_of(m, ws.title)
                    c1 = ci(m.group("c1").lstrip("$"))
                    r1 = int(m.group("r1").lstrip("$"))
                    c2 = ci(m.group("c2").lstrip("$")) if m.group("c2") else c1
                    r2 = int(m.group("r2").lstrip("$")) if m.group("r2") else r1
                    box.append((sh, min(c1, c2), min(r1, r2), max(c1, c2), max(r1, r2)))
                if box:
                    deps[(ws.title, c.coordinate)] = box
    return deps


def snapshot(before, after, wf):
    """Every cached value in the book, before against after.

    A value is allowed to move only if this stage wrote the cell, or if the
    cell reads one that moved for that reason.  Everything else is a failure.
    """
    LOG.head("whole book snapshot")
    deps = boxes_read(wf)
    changed, gone = [], []
    for title, cells in before.items():
        if title not in after:
            changed.append((title, "*sheet", "*sheet", "gone", None))
            continue
        for coord, ov in cells.items():
            nc = fwd_map(title, coord)
            if nc is None:
                gone.append("%s!%s" % (title, coord))
                continue
            nv = after[title].get(nc)
            same = (ov == nv) or (isinstance(ov, (int, float))
                                  and isinstance(nv, (int, float))
                                  and abs(ov - nv) <= max(TOL, abs(ov) * 1e-12))
            if not same:
                changed.append((title, coord, nc, ov, nv))
    fresh = []
    for title, cells in after.items():
        seen = set(filter(None, (fwd_map(title, c) for c in before.get(title, {}))))
        for coord, nv in cells.items():
            if coord in seen or nv is None:
                continue
            fresh.append((title, None, coord, None, nv))

    todo = changed + fresh
    explained = {}
    for t, bc, ac, ov, nv in todo:
        why = CH.get(t, bc) if bc else None
        why = why or CH.get(t, ac) or CH.get(t, "*rows")
        if why:
            explained[(t, ac)] = why
    for _ in range(12):
        grew = False
        pts = [(t, ci(re.match(r"([A-Z]+)(\d+)$", a).group(1)),
                int(re.match(r"([A-Z]+)(\d+)$", a).group(2)))
               for (t, a) in explained if re.match(r"([A-Z]+)(\d+)$", a)]
        for t, bc, ac, ov, nv in todo:
            if (t, ac) in explained:
                continue
            for sh, c1, r1, c2, r2 in deps.get((t, ac), []):
                if any(pt == sh and c1 <= pc <= c2 and r1 <= pr <= r2
                       for pt, pc, pr in pts):
                    explained[(t, ac)] = "reads a cell that moved"
                    grew = True
                    break
        if not grew:
            break
    bad = ["%s!%s %r -> %r" % (t, ac, ov, nv) for t, bc, ac, ov, nv in todo
           if (t, ac) not in explained]
    tally = collections.Counter(explained.values())
    print("   %d cached values moved, %d cells emptied by the row delete, "
          "%d new values" % (len(changed), len(gone), len(fresh)), flush=True)
    for why, n in sorted(tally.items()):
        print("      %-24s %d" % (why, n), flush=True)
    check(not bad, "every value that moved is one of the fixes or reads one: "
                   "%d unexplained %s" % (len(bad), bad[:6]))


def selfcheck(out, before, inp):
    print("", flush=True)
    LOG.head("self-check")
    wv = openpyxl.load_workbook(out, data_only=True)
    wf = openpyxl.load_workbook(out, data_only=False)
    after = {ws.title: {c.coordinate: c.value for row in ws.iter_rows()
                        for c in row if c.value is not None} for ws in wv.worksheets}
    back = {}
    for title, cells in before.items():
        for coord in cells:
            nc = fwd_map(title, coord)
            if nc:
                back[(title, nc)] = coord
    cfg, lo = sheet(wf, "0.2"), sheet(wf, "3.5")
    hdr = need(cfg, "Portfolio")
    cSpend = needcol(cfg, hdr, "Spend (actual cost after levers, from 3.5 TDD Lights On)")
    cBud = needcol(cfg, hdr, "Total")
    cVar = needcol(cfg, hdr, "Over/(under)")
    total = need(cfg, "Total", lo=hdr + 1)
    lhdr = need(lo, "Portfolios & COEs & EGI")
    cL = needcol(lo, lhdr, "Total portfolio cost charged to TDD")

    # A
    bad, n = [], 0
    for r in range(hdr + 1, total):
        f = cfg.cell(r, cSpend).value
        if not (isinstance(f, str) and f.startswith("=")):
            continue
        m = re.search(r"\$%s\$(\d+)" % gl(cL), f)
        if not m or lo.title not in f:
            bad.append("row %d: %r" % (r, f))
            continue
        n += 1
        a = after[cfg.title].get("%s%d" % (gl(cSpend), r))
        b = after[lo.title].get("%s%d" % (gl(cL), int(m.group(1))))
        if a is None or b is None or abs(a - b) > TOL:
            bad.append("row %d: %r vs %r" % (r, a, b))
    check(not bad and n >= 15,
          "A: every one of the %d priced rows on 0.2 ties to 3.5's charge at 1e-9 %s"
          % (n, bad[:3]))
    ta = after[cfg.title].get("%s%d" % (gl(cSpend), total))
    tb = after[lo.title].get("%s%d" % (gl(cL), _tot35(lo)))
    check(ta is not None and tb is not None and abs(ta - tb) <= 1e-9,
          "A: the two totals agree, 0.2 %s against 3.5 %s" % (_f(ta), _f(tb)))

    # B
    ok = True
    for r in range(hdr + 1, total):
        f = nz(cfg.cell(r, cVar).value)
        if f.startswith("=") and not f.startswith("=%s%d-%s%d"
                                                  % (gl(cSpend), r, gl(cBud), r)):
            ok = False
    check(ok, "B: 0.2's variance column is spend less budget on every row")
    gv = after[cfg.title].get("%s%d" % (gl(cVar), total))
    check(gv is not None and gv > 0 and abs(gv - (ta - after[cfg.title]
                                                  .get("%s%d" % (gl(cBud), total)))) <= 1e-9,
          "B: 0.2's total over budget reads %s, positive because spend is above budget"
          % _f(gv))
    exec_ok = True
    for ws in wf.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f = c.value
                if isinstance(f, str) and f.startswith("=") and \
                        re.search(r"'%s'!\$%s\$%d\b" % (re.escape(cfg.title),
                                                        gl(cVar), total), f):
                    v = after[ws.title].get(c.coordinate)
                    if v is None or abs(v - gv) > 1e-9:
                        exec_ok = False
    check(exec_ok, "B: every tab that reads that total now shows the same sign")
    b2 = []
    for ws in twox(wf):
        r = find(ws, "Over/(under) budget ($m)")
        if not r:
            continue
        spend = find(ws, "Total planned spend ($m)") or find(ws, "Planned spend ($m)")
        draw = find(ws, "Total budget to draw down ($m)")
        v = after[ws.title].get("C%d" % r)
        want = (after[ws.title].get("C%d" % spend, 0)
                - after[ws.title].get("C%d" % draw, 0))
        if v is None or abs(v - want) > 1e-9:
            b2.append(ws.title)
    check(not b2, "B: every 2.x funding block reads spend less budget %s" % b2)
    b3 = [ws.title for ws in onex(wf)
          if find(ws, "Other Variance")
          and nz(ws["C%d" % find(ws, "Other Variance")].value).startswith("=-")]
    check(not b3, "B: no 1.x tab negates its left to fund line %s" % b3)

    # C, D, E
    a1, a3, a4 = sheet(wf, "3.1"), sheet(wf, "3.3"), sheet(wf, "3.4")
    h1, h3, h4 = need(a1, "Line"), need(a3, "Portfolio"), need(a4, "COE")
    check(bool(colof(a1, h1, "Cost after levers, the Business Partner and Domain "
                             "Architect pots netted out of the COE lines ($m)"))
          and bool(colof(a3, h3, "Cost after levers, before the cyber uplift "
                                 "recharge ($m)")),
          "C: each cost after levers total names its basis")
    check(_has(a1, "Cost after levers reads three ways")
          and _has(a3, "Cost after levers reads three ways"),
          "C: 3.1 and 3.3 both carry the line naming all three totals")
    check(bool(colof(a1, h1, "Funded outside TDD, before the cyber uplift "
                             "recharge ($m)")),
          "C: the funded outside pair is named too")
    check(bool(colof(a1, h1, "Variance to archetype, the portfolio allowance "
                             "included ($m)"))
          and bool(colof(a3, h3, "Variance to archetype, the portfolio allowance "
                                 "excluded ($m)")),
          "D: the two variance headings say which basis they are on")
    check(bool(colof(a4, h4, "Cost before levers ($m)")),
          "E: 3.4 says it reports before the levers")

    # F, G, H, I
    rows = collections.Counter(find(ws, "Total Cost") for ws in onex(wf))
    check(len(rows) == 1, "F: every 1.x tab carries Total Cost on one row %s"
          % dict(rows))
    hs = collections.Counter()
    for ws in onex(wf):
        hs[tuple(nz(ws.cell(r, 8).value) for r in range(
            need(ws, "Budget vs TDD Cost", col=8),
            need(ws, "TDD over/(under) budget ($m)", col=8) + 1))] += 1
    check(len(hs) == 1, "F,H: all eleven budget blocks read the same lines: "
          + repr(list(hs)[0] if len(hs) == 1 else sorted(hs, key=len)))
    nzcol = [ws.title for ws in onex(wf)
             if colof(ws, need(ws, "Cost"), "TDD NZ ($m)") is None]
    check(not nzcol, "G: every 1.x tab has a TDD NZ column %s" % nzcol)
    fund = [ws.title for ws in onex(wf) if find(ws, "Other funding", col=8)
            and not find(ws, "Budget line", col=8)]
    check(not fund, "H: every 1.x funding block is a budget line table %s" % fund)
    heads, tots = set(), set()
    for ws in onex(wf):
        tv, ov = find(ws, "TDD Variance"), find(ws, "Other Variance")
        if not (tv and ov):
            continue
        for r in range(tv - 1, max(tv - 6, 1), -1):
            if norm(ws.cell(r, 2).value) in MINI:
                heads.add(norm(ws.cell(r, 2).value))
                break
        for r in range(ov + 1, ov + 3):
            if norm(ws.cell(r, 2).value) in MINI:
                tots.add(norm(ws.cell(r, 2).value))
                break
    check(heads == {"total to fund"} and tots == {"total"},
          "I: one label over the block and one on its total %s %s" % (heads, tots))

    # J
    j = []
    for ws in twox(wf):
        if ws.title.startswith("2.14"):
            continue
        hdr2 = need(ws, "Squad")
        tot2 = need(ws, "Total portfolio", lo=hdr2 + 1)
        egi = [r for r in range(hdr2 + 1, tot2)
               if isinstance(ws.cell(r, 2).value, str)
               and nkey(ws.cell(r, 2).value).startswith("egi")]
        empty = [r for r in egi if not (after[ws.title].get("F%d" % r) or 0)]
        if len(egi) > 1 and empty:
            j.append("%s carries %d EGI rows, %d of them empty"
                     % (ws.title, len(egi), len(empty)))
        for r in empty:
            read = False
            for other in wf.worksheets:
                for row in other.iter_rows():
                    for c in row:
                        f = c.value
                        if isinstance(f, str) and f.startswith("=") and \
                                other is not ws and \
                                re.search(r"'%s'!\$[A-Z]{1,2}\$%d\b"
                                          % (re.escape(ws.title), r), f):
                            read = True
            if not read:
                j.append("%s r%d is empty and nothing reads it" % (ws.title, r))
    check(not j, "J: one EGI row per lever tab, and an empty one only where a "
                 "1.x tab still reads it %s" % j)

    # K
    stale = []
    for ws in wf.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "All 3 in Customer" in c.value:
                    stale.append("%s!%s" % (ws.title, c.coordinate))
    check(not stale, "K: the stale label is gone %s" % stale)

    # the guardrails
    src = openpyxl.load_workbook(inp, data_only=False)
    prot = [ws.title for ws in src.worksheets
            if ws.title in wf.sheetnames
            and bool(ws.protection.sheet) != bool(wf[ws.title].protection.sheet)]
    check(not prot, "every sheet's protection is as it was %s" % prot)
    dv = []
    for ws in src.worksheets:
        if ws.title not in wf.sheetnames:
            continue
        a = sum(len(d.sqref.ranges) for d in ws.data_validations.dataValidation)
        b = sum(len(d.sqref.ranges) for d in
                wf[ws.title].data_validations.dataValidation)
        if a != b:
            dv.append("%s %d -> %d" % (ws.title, a, b))
    check(not dv, "every lever and toggle dropdown is still in place %s" % dv)
    cream = []
    for ws in src.worksheets:
        if ws.title not in wf.sheetnames:
            continue
        for row in ws.iter_rows():
            for c in row:
                if c.value is None or isinstance(c.value, str) and \
                        c.value.startswith("="):
                    continue
                fg = getattr(c.fill.fgColor, "rgb", None)
                if not (isinstance(fg, str) and fg in ("FFFFF2CC", "FFFFE699",
                                                       "FFFFF2CD")):
                    continue
                co = fwd_map(ws.title, c.coordinate) or c.coordinate
                if wf[ws.title][co].value == c.value:
                    continue
                if CH.get(ws.title, co) == "K":
                    LOG.note("K", "%s!%s was a cream cell: his stale note is "
                                  "cleared, the input cell stays" % (ws.title, co))
                    continue
                cream.append("%s!%s" % (ws.title, c.coordinate))
    check(not cream, "every cream typed input still reads what he typed, apart "
                     "from the stale note K clears: %d %s" % (len(cream), cream[:4]))
    src.close()

    now = was = None
    for r in range(h1 + 1, a1.max_row + 1):
        v = a1.cell(r, 2).value
        if isinstance(v, str) and "TDD total" in v:
            co = "%s%d" % (gl(colof(a1, h1, "Total roles")), r)
            now = after[a1.title].get(co)
            was = before.get(a1.title, {}).get(back.get((a1.title, co), co))
    check(now is not None and now == was, "the role count has not moved: %s" % now)
    tot31 = _tdd_total(a1, after)
    check(tot31 is not None and abs(tot31 - 114.029147297163) < 1e-6,
          "the whole book total still reads %s" % _f(tot31))
    lo_tot = after[lo.title].get("%s%d" % (gl(cL), _tot35(lo)))
    check(lo_tot is not None, "3.5's charge total reads %s" % _f(lo_tot))
    for pfx in ("3.5", "3.6"):
        try:
            ws = sheet(wf, pfx)
        except SystemExit:
            continue
        diff = [co for co, v in before.get(ws.title, {}).items()
                if _cmp(v, after[ws.title].get(fwd_map(ws.title, co) or "ZZ1"))]
        check(not diff, "%s: not one value moved %s" % (ws.title, diff[:5]))

    errs, old, blanks = [], 0, []
    for ws in wf.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f, where = c.value, "%s!%s" % (ws.title, c.coordinate)
                bc = back.get((ws.title, c.coordinate))
                was = before.get(ws.title, {}).get(bc) if bc else None
                if isinstance(f, str) and f.startswith("="):
                    v = after[ws.title].get(c.coordinate)
                    if isinstance(v, str) and any(e in v for e in ERRS):
                        errs.append("%s %s" % (where, v))
                    if v is None and (was is not None or CH.get(ws.title,
                                                                c.coordinate)):
                        blanks.append(where)
                    if re.search(r"(?<![$0-9])[A-Z]{1,3}:[A-Z]{1,3}(?![0-9])", f):
                        errs.append("%s whole column" % where)
                elif isinstance(f, str) and f.strip() in ("-", "--", "–"):
                    if isinstance(was, str) and was.strip() == f.strip():
                        old += 1
                    else:
                        errs.append("%s dash literal" % where)
    check(not errs, "nothing this stage wrote is an error cell, a dash literal or "
                    "a whole column reference: %d %s" % (len(errs), errs[:4]))
    LOG.note("hygiene", "%d dash literals were already in his Finance table and "
                        "are left alone" % old)
    check(not blanks, "no formula lost its cached value and every formula this "
                      "stage wrote carries one: %d %s" % (len(blanks), blanks[:4]))
    snapshot(before, after, wf)
    wv.close()
    wf.close()


def _cmp(a, b):
    if a == b:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) > max(TOL, abs(a) * 1e-12)
    return True


def _has(ws, tag):
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and tag in c.value:
                return True
    return False


def _tot35(lo):
    h = need(lo, "Portfolios & COEs & EGI")
    for r in range(h + 1, lo.max_row + 1):
        if nkey(lo.cell(r, 2).value).startswith("total"):
            return r
    stop("3.5 has no total row")


def _tdd_total(a1, after):
    h = need(a1, "Line")
    c = colof(a1, h, "Actual cost ($m)")
    for r in range(h + 1, a1.max_row + 1):
        v = a1.cell(r, 2).value
        if isinstance(v, str) and "TDD total" in v:
            return after[a1.title].get("%s%d" % (gl(c), r))
    return None


# ===================================================================== driver

def finished(wb):
    """The marks this stage leaves, all of them."""
    try:
        cfg = sheet(wb, "0.2")
        hdr = need(cfg, "Portfolio")
        if colof(cfg, hdr, "Spend (actual cost after levers, from 3.5 TDD Lights On)") is None:
            return False
        if colof(cfg, hdr, "Over/(under)") is None:
            return False
        if not _has(cfg, "Over/(under) is positive when the spend is above"):
            return False
        rows = {find(ws, "Total Cost") for ws in onex(wb)}
        if len(rows) != 1:
            return False
        for ws in onex(wb):
            if colof(ws, need(ws, "Cost"), "TDD NZ ($m)") is None:
                return False
            if find(ws, "Other funding", col=8) and not find(ws, "Budget line", col=8):
                return False
            if not find(ws, "AU over/(under) budget ($m)", col=8):
                return False
        if _has(sheet(wb, "3.2"), "All 3 in Customer"):
            return False
        if not _has(sheet(wb, "3.3"), "Cost after levers reads three ways"):
            return False
    except SystemExit:
        return False
    return True


def main():
    if len(sys.argv) != 3:
        stop("usage: w6_consistency.py <in.xlsx> <out.xlsx>")
    inp, out = sys.argv[1], sys.argv[2]
    if not os.path.exists(inp):
        stop("no input %s" % inp)

    wb = load(inp)
    if finished(wb):
        print("input already carries the consistency fixes - copying through",
              flush=True)
        wb.close()
        shutil.copy(inp, out)
        print("\nwrote %s" % out, flush=True)
        return
    wb.close()

    work = os.path.join(os.path.dirname(os.path.abspath(out)) or ".", "_w6")
    os.makedirs(work, exist_ok=True)
    print("reading the book as it stands (recalculating for the snapshot)",
          flush=True)
    before = wbio.harvest(wbio.recalc(inp, work))

    wb = load(inp)
    set_note_style(wb)
    aparts = fix_a(wb, before)
    fix_b(wb, before, aparts)
    fix_cde(wb, before)
    rows = collections.Counter(find(ws, "Total Cost") for ws in onex(wb))
    target = rows.most_common(1)[0][0]
    for ws in onex(wb):                  # what fix F is about to move, for the map
        r = find(ws, "Total Cost")
        if r is None or r == target:
            continue
        sec = need(ws, "Portfolio Summary")
        last = need(ws, "Variance to actuals", col=5)
        kf = need(ws, "Actuals vs archetype", col=11)
        kl = need(ws, "Variance", col=11, lo=kf)
        MOVES.append((ws.title, block_cols("B", "G"), sec, last, target - r))
        MOVES.append((ws.title, block_cols("K", "N"), kf, kl, target - r))
    fix_f(wb, before)
    fix_g(wb, before)
    fix_h_budget(wb, before)
    fix_h_funding(wb, before)
    fix_i(wb, before)
    fix_j(wb, before)
    fix_k(wb, before)
    LOG.tail()

    tmp = os.path.join(work, "edited.xlsx")
    save(wb, tmp)
    wb.close()
    print("recalculating and writing the cached values back", flush=True)
    wbio.build(tmp, out)
    selfcheck(out, before, inp)

    print("", flush=True)
    if FAILS:
        print("%d FAIL" % len(FAILS), flush=True)
        raise SystemExit(1)
    print("wrote %s" % out, flush=True)


if __name__ == "__main__":
    main()
