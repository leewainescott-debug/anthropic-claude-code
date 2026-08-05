#!/usr/bin/env python3
"""w_gate - the QA gate for build W (his master mapping + Lights On v2).

  python3 w_gate.py <final.xlsx> [--recalc] [--his PATH] [--prev PATH]

READS AND VERIFIES ONLY. The one write it can do is a recalc round-trip via
scripts/v10/wbio.py onto a scratch copy beside this script when the file
arrives without cached values (or when --recalc is passed); the input file is
never touched and the repo is never touched.

Check groups (one PASS/FAIL line each, counted at the end, non-zero exit on
any FAIL), then the ledger print:

  A raw identity   REVIEW A2:AC527 cell-identical to his master upload, his
                   header row verbatim (duplicate EE Number, MyHR ee no)
  B homing         every one of his 526 rows homes to exactly one tab, the
                   derivation re-run here from Portfolio/Division/Lists
                   overrides, matched against MTab and against real 2.x
                   membership; the reversals (Viren under EGI, Ed Tacey,
                   Sarsha Tanner, Murray Mitchell, Scott McKenzie)
  C ledger         role count, Role IDs R0001.., overhead line re-derived
                   from his titles and squads and matched cell by cell,
                   line rosters and counts, Shane Ker, the DA/DE managers
  D levers         four values only, no Paused, Hold at zero, no stale
                   Filled+Hire (BLD-18), person-keyed carry against --prev
  E part-time      the effective cost helper = Full Cost x FTE below 1, and
                   the 2.x role cost priced off it
  F lights on 3.5  tab, his 19 columns verbatim with N and P blank, the 18
                   0.2-config rows, Total and Budget rows, toggles, the
                   analysis block, Enterprise Data naming
  G recompute 3.5  every cell of C D E F G H I K L M O Q S rebuilt from
                   REVIEW + 2.x + 1.x + 0.2 + Lists and tied to 1e-6
  H shares         the pots divided by ELEVEN (ten portfolios + TDD Cyber),
                   the Customer split and the COE pair splits
  I EGI            no EGI cost anywhere in E F G H I, the EGI row itself
  J reconciliation every single cost represented: C total + the GM layer
                   against REVIEW effective cost after levers and 3.1
  K lights on 3.6  the AU NZ duplicate: same rows, C..L identical, AU + NZ
                   = Total = L, 0.2 AU and NZ budgets, Variance
  L protection     protected set exactly the 0.x and 3.x tabs, Tdd123,
                   toggles unlocked, structure locked
  M hygiene        dashes, cream formulas, banned words, parens formats,
                   the TDD Data label, BLD-19
  N controls       every control row 0, 4.0 Data QA, no error cells, the
                   netting rows, 0.2 TDD Cyber and its column total

Every numeric expectation is re-derived from the file under test or from his
master upload. The only hard-coded values are his own words (headers, row
labels), his rulings (share base 11, password Tdd123) and the 53.80 budget,
which is itself cross-read from 0.2. Nothing is pinned to a build artefact.

Structure, wbio handling and the check idiom inherited from
scripts/v10/lightson/v7_gate.py.
"""
import sys, os, re, collections, itertools, unicodedata

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
V10 = "/home/user/anthropic-claude-code/scripts/v10"
UPLOADS = "/root/.claude/uploads/e550b440-3996-5abb-87e5-bafafe598f82"
HIS_DEFAULT = os.path.join(UPLOADS, "0ad63df5-updates.xlsx")
PREV_DEFAULT = "/home/user/anthropic-claude-code/TDD_Cost_Calc.xlsx"

REVIEW = "REVIEW - Complete Role Mapping"
LO = "3.5 TDD Lights On"
LO2 = "3.6 TDD Lights On AU NZ"
LEVERS = ("Filled", "Hire", "Hold", "Offshore")
CREAM = "FFFFF2CC"
WHITE = "FFFFFFFF"
SRC_TABS = ("0.1 Budget Table (Fin)", "0.4 Presentation Pack")
ERR = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!",
       "Err:")
TOL = 1e-6
TOLC = 1e-9

PASSWORD = "Tdd123"
HIS_ROWS = 526                      # his master: 526 data rows, 29 columns
HIS_COLS = 29
SHARE_BASE = 11                     # his ruling: 10 portfolios + TDD Cyber
L_BUDGET = 53.8                     # cross-read from 0.2 Budget as well

# his Lights On columns, verbatim, at their column numbers (N=14, P=16 blank)
HEADERS = {2: "Portfolios & COEs & EGI",
           3: "Total People cost",
           4: "Sig items funded",
           5: "Support Cost (this is the % in the 1.x tabs)",
           6: "BP allocation",
           7: "Domain architect allocation",
           8: "GM allocation",
           9: "Other overheads",
           10: "Other overheads toggle",
           11: "Amount of overheads charged to TDD",
           12: "Total portfolio cost charged to TDD",
           13: "TDD Lights On budget",
           15: "Over/ Under lights on budget",
           17: "Total Cost left to be recharged to business",
           18: "Amount noted in 1.x tabs",
           19: "Still left to fund"}
BLANKCOLS = (14, 16)
COL = {"C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8, "I": 9, "J": 10,
       "K": 11, "L": 12, "M": 13, "O": 15, "Q": 17, "R": 18, "S": 19}

# his row set, in 0.2 Data Config order, Legal and EG dropped, TDD Data
# relabelled Enterprise Data
LROWS = ["COE Strategy Architecture", "COE Cyber Risk & Service Ops",
         "COE Transformation", "COE Business Partnering", "COE Data",
         "Ampol Retail", "Z Retail", "Ampol Customer", "Z Customer",
         "Commercial Fuels", "Energy Solutions & B2B", "Infrastructure",
         "P&C", "Finance", "TDD Group Functions", "Enterprise Data",
         "TDD Cyber", "EGI"]
# Lights On row -> the 2.x tab it prices
SRCTAB = {"COE Strategy Architecture": "2.13", "COE Data": "2.13",
          "COE Transformation": "2.12", "COE Business Partnering": "2.12",
          "COE Cyber Risk & Service Ops": "2.11",
          "Ampol Retail": "2.1", "Z Retail": "2.10",
          "Ampol Customer": "2.2", "Z Customer": "2.2",
          "Commercial Fuels": "2.9", "Energy Solutions & B2B": "2.8",
          "Infrastructure": "2.7", "P&C": "2.5", "Finance": "2.6",
          "TDD Group Functions": "2.4", "Enterprise Data": "2.3",
          "TDD Cyber": "2.15", "EGI": "2.14"}
# the 1.x tab each row draws its support percentages and its R amount from
SRC1X = {"Ampol Retail": "1.1", "Ampol Customer": "1.2", "Z Customer": "1.2",
         "Enterprise Data": "1.3", "TDD Group Functions": "1.4", "P&C": "1.5",
         "Finance": "1.6", "Infrastructure": "1.7",
         "Energy Solutions & B2B": "1.8", "Commercial Fuels": "1.9",
         "Z Retail": "1.10", "TDD Cyber": "1.14"}
SPLITS = {"2.2": ("Ampol Customer", "Z Customer"),
          "2.12": ("COE Business Partnering", "COE Transformation"),
          "2.13": ("COE Strategy Architecture", "COE Data")}
COE_ROWS = ("COE Strategy Architecture", "COE Cyber Risk & Service Ops",
            "COE Transformation", "COE Business Partnering", "COE Data")
NOSHARE = set(COE_ROWS) | {"EGI"}
# rows that carry one full share unit; the Customer pair shares one between
# the two of them
FULLSHARE = ["Ampol Retail", "Z Retail", "Commercial Fuels",
             "Energy Solutions & B2B", "Infrastructure", "P&C", "Finance",
             "TDD Group Functions", "Enterprise Data", "TDD Cyber"]
# Lights On row -> the 0.2 budget row label it reads (prefix match)
BUDROW = {"COE Strategy Architecture": "COE - Strategy Architecture",
          "COE Cyber Risk & Service Ops": "COE - Cyber",
          "COE Transformation": "COE - Transformation",
          "COE Business Partnering": "COE - Business Partnering",
          "COE Data": "COE - Data", "Enterprise Data": "TDD Data",
          "Ampol Customer": "Ampol Customer", "Z Customer": "Z Customer"}

# his tab-homing rules, spelled out in the spec (case-normalised)
PORTMAP = {"retail": "Ampol Retail", "z": "Z Retail",
           "ampol customer": "Customer", "z energy (digital)": "Customer",
           "commercial fuels": "Commercial Fuels",
           "b2b & energy solutions": "Energy Solutions & B2B",
           "infrastructure": "Infrastructure",
           "enterprise data": "Enterprise Data", "finance": "Finance",
           "p&c": "P&C", "p&c, finance & legal": "P&C",
           "tdd": "TDD Group Functions", "egi": "EGI",
           "egi integration": "EGI"}
DIVMAP = {"strategy, architecture & data": "COE SA&D",
          "cyber, risk & operations": "COE Cyber",
          "cyber, risk & service ops": "COE Cyber",
          "partnering & transformation": "COE BP&T",
          "tdd group functions": "TDD Group Functions",
          "customer": "Customer", "egi": "EGI"}
# homing name -> the 2.x tab prefix that must hold the person
HOME2TAB = {"Ampol Retail": "2.1", "Customer": "2.2", "Enterprise Data": "2.3",
            "TDD Group Functions": "2.4", "P&C": "2.5", "Finance": "2.6",
            "Infrastructure": "2.7", "Energy Solutions & B2B": "2.8",
            "Commercial Fuels": "2.9", "Z Retail": "2.10",
            "COE Cyber": "2.11", "COE BP&T": "2.12", "COE SA&D": "2.13",
            "EGI": "2.14", "TDD Cyber": "2.15"}
GRID_TOTALS = ("Squads total", "Directly funded total", "No archetype total",
               "Overhead roles total", "Total portfolio")
# his standing rulings on the overhead lines, people and titles he named
OFF_LINE = {"ed tacey": "Squad"}          # "Ed Tacey, let's keep him off that"
DM_TITLES = ("delivery assurance manager", "delivery excellence manager")
VACANT_NAMES = ("vacant", "ring fenced", "remove")
BANNED = (r"\bwaves?\b", r"\bseats?\b", r"design[- ]for[- ]archetype",
          r" to projects")


def norm(s):
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip()).lower()


def txt(s):
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip())


def num(x):
    return x if isinstance(x, (int, float)) and not isinstance(x, bool) else None


def near(a, b, tol=TOL):
    return a is not None and b is not None and abs(a - b) <= tol


def rgb_of(cell):
    try:
        fl = cell.fill
        if fl is None or fl.fgColor is None:
            return None
        return fl.fgColor.rgb if isinstance(fl.fgColor.rgb, str) else None
    except Exception:
        return None


# ------------------------------------------------------------------ reporting

class Gate:
    def __init__(self):
        self.groups = []

    def group(self, letter, name):
        g = {"letter": letter, "name": name, "subs": [], "notes": []}
        self.groups.append(g)
        return g

    @staticmethod
    def sub(g, ok, text):
        g["subs"].append((bool(ok), text))
        return bool(ok)

    @staticmethod
    def note(g, text):
        g["notes"].append(text)

    def report(self):
        fails = 0
        for g in self.groups:
            bad = [t for ok, t in g["subs"] if not ok]
            ok = not bad and bool(g["subs"])
            if not g["subs"]:
                bad = ["no sub-checks ran"]
            if not ok:
                fails += 1
            head = "%-5s %s %-16s" % ("PASS" if ok else "FAIL", g["letter"],
                                      g["name"])
            summary = ("%d/%d ok" % (sum(1 for s in g["subs"] if s[0]),
                                     len(g["subs"])))
            print("%s %s%s" % (head, summary,
                               "" if ok else "  -> " + "; ".join(bad[:2])),
                  flush=True)
            for okv, t in g["subs"]:
                if not okv:
                    print("        off: %s" % t, flush=True)
            for t in g["notes"]:
                print("        note: %s" % t, flush=True)
        print("\n%d groups, %d pass, %d fail"
              % (len(self.groups), len(self.groups) - fails, fails),
              flush=True)
        return fails


# ------------------------------------------------------------- layout readers

def find_tab(sheets, prefix):
    for t in sheets:
        if t == prefix or t.startswith(prefix + " "):
            return t
    return None


def header_cols(ws, upto=None):
    """{normalised header text: column} off row 1."""
    out = {}
    for c in range(1, (upto or ws.max_column) + 1):
        k = norm(ws.cell(1, c).value)
        if k and k not in out:
            out[k] = c
    return out


def find_col(hdrs, *needles):
    """first column whose header contains every needle."""
    for k, c in hdrs.items():
        if all(n in k for n in needles):
            return c
    return None


class Tab:
    """One 2.x lever modelling tab: grid, role rows, role-list sections."""

    def __init__(self, title, wf, wv, idmap):
        self.title = title
        self.wf, self.wv = wf, wv

        def cv(r, c):
            """the rendered text of a cell: cached value first, formula last"""
            x = wv.cell(r, c).value
            if x is None:
                x = wf.cell(r, c).value
                if isinstance(x, str) and x.startswith("="):
                    return ""
            return txt(x)

        self.cv = cv
        self.total_r = None
        for r in range(5, min(wf.max_row, 90) + 1):
            if cv(r, 2) == "Total portfolio":
                self.total_r = r
                break
        self.grid = []            # (row, label) squad/overhead grid lines
        self.grid_all = {}
        end = self.total_r or min(wf.max_row, 40)
        for r in range(7, end + 1):
            b = cv(r, 2)
            if not b:
                continue
            self.grid_all.setdefault(b, r)
            if b in GRID_TOTALS:
                continue
            if num(wv.cell(r, 15).value) is not None:
                self.grid.append((r, b))
        self.roles = []
        self.sections = collections.OrderedDict()   # label -> [role dicts]
        cur = None
        for r in range(1, wf.max_row + 1):
            d = wf.cell(r, 4).value
            isrole = isinstance(d, str) and d.startswith("=") and REVIEW in d
            rid = wf.cell(r, 1).value
            if not isrole and isinstance(rid, str) \
                    and re.fullmatch(r"R\d{4}", rid.strip()):
                isrole = True
            if not isrole:
                # a role-list section header: label in B, "n roles" in C, both
                # of which may arrive as formulas
                b, c = cv(r, 2), cv(r, 3)
                if b and re.match(r"^[\d.]+ roles?$", c):
                    cur = b
                    self.sections.setdefault(cur, [])
                continue
            led = None
            m = re.fullmatch(r"='[^']+'!\$[A-Z]{1,3}\$(\d+)",
                             str(d).strip()) if isinstance(d, str) else None
            if m:
                led = int(m.group(1))
            if led is None and isinstance(rid, str) and re.fullmatch(
                    r"R\d{4}", rid.strip()):
                led = idmap.get(rid.strip())
            ro = {"row": r, "ledger": led, "section": cur,
                  "id": rid.strip() if isinstance(rid, str) else None,
                  "name": wv.cell(r, 2).value, "title": wv.cell(r, 3).value,
                  "status": txt(wv.cell(r, 4).value),
                  "lever": txt(wf.cell(r, 5).value),
                  "F": num(wv.cell(r, 6).value),
                  "G": num(wv.cell(r, 7).value)}
            self.roles.append(ro)
            if cur is not None:
                self.sections.setdefault(cur, []).append(ro)

    def gval(self, label, col):
        r = self.grid_all.get(label)
        return None if r is None else num(self.wv.cell(r, col).value)

    def total(self, col=19):
        if self.total_r is None:
            return None
        return num(self.wv.cell(self.total_r, col).value)

    def after(self, roles=None):
        return sum(ro["G"] or 0 for ro in (roles or self.roles)) / 1e6

    def label_value(self, label_start, col=3):
        """value beside the first B-column label starting with label_start."""
        for r in range(1, self.wf.max_row + 1):
            if self.cv(r, 2).lower().startswith(label_start.lower()):
                return num(self.wv.cell(r, col).value), r
        return None, None


def pct_map(wf, wv):
    """{normalised squad label: support %} off a 1.x tab."""
    out, clash = {}, []
    for r in range(1, wf.max_row + 1):
        b = wf.cell(r, 2).value
        c = wf.cell(r, 3).value
        g = num(wv.cell(r, 7).value)
        if not (isinstance(b, str) and b and not b.startswith("=")):
            continue
        if not (isinstance(c, str) and c and not str(c).startswith("=")):
            continue
        if g is None or not (0 <= g <= 1):
            continue
        k = norm(b)
        if k in out and abs(out[k] - g) > 1e-12:
            clash.append(k)
        out.setdefault(k, g)
    return out, clash


def funding_amounts(wf, wv):
    """every numeric in a 1.x funding block's 'allocated to people' column."""
    vals = []
    for r in range(1, wf.max_row + 1):
        b = txt(wf.cell(r, 8).value)
        if not b:
            continue
        x = num(wv.cell(r, 10).value)
        if x is None:
            continue
        vals.append((b, x))
    return vals


def tkey(s):
    """a title as a carry key: NFKC, en and em dashes folded to a hyphen,
    case and every space dropped, so 'Lead - Asset' and 'Lead – Asset' and
    'EnterpriseProcess Analyst' all key onto their old rows."""
    s = unicodedata.normalize("NFKC", str(s if s is not None else ""))
    for d in ("–", "—", "‒", "−", "‐", "‑"):
        s = s.replace(d, "-")
    return re.sub(r"\s+", "", s).strip().lower()


def is_vacancy(name, status=""):
    n = norm(name)
    return (txt(status) == "Vacant"
            or any(n.startswith(x) or x in n for x in VACANT_NAMES))


def fmt_sections(fmt):
    out, cur, q, i = [], "", False, 0
    while i < len(fmt):
        ch = fmt[i]
        if ch == '"':
            q = not q
            cur += ch
        elif ch == "\\" and i + 1 < len(fmt):
            cur += fmt[i:i + 2]
            i += 2
            continue
        elif ch == ";" and not q:
            out.append(cur)
            cur = ""
        else:
            cur += ch
        i += 1
    out.append(cur)
    return out


DASHFMT = re.compile(r'(\\-|"-")')


# ------------------------------------------------------------------- the gate

def main(argv):
    if not argv:
        print("usage: w_gate.py <final.xlsx> [--recalc] [--his PATH] "
              "[--prev PATH]")
        return 2
    src = argv[0]
    rest = argv[1:]
    force_recalc = "--recalc" in rest

    def opt(flag, default):
        if flag in rest:
            i = rest.index(flag)
            if i + 1 < len(rest):
                return rest[i + 1]
        return default

    his_path = opt("--his", HIS_DEFAULT)
    prev_path = opt("--prev", PREV_DEFAULT)
    if not os.path.exists(src):
        print("STOP: no such file: %s" % src)
        return 2

    print("w_gate on %s" % src, flush=True)
    print("  his master: %s" % his_path, flush=True)
    f = openpyxl.load_workbook(src, data_only=False)
    v = openpyxl.load_workbook(src, data_only=True)

    # ---- cached-value probe; one wbio round-trip onto a scratch copy if bare
    blank = total = 0
    for t in [t for t in f.sheetnames if t.startswith("2.")][:6]:
        wf, wv = f[t], v[t]
        for row in wf.iter_rows():
            for cl in row:
                if isinstance(cl.value, str) and cl.value.startswith("="):
                    total += 1
                    if wv[cl.coordinate].value is None:
                        blank += 1
    bare = total > 0 and blank / total > 0.2
    if force_recalc or bare:
        print("cached values %s (%d of %d formula cells blank on the probe "
              "tabs) - one wbio recalc round-trip onto a scratch copy"
              % ("recalc forced" if force_recalc else "missing", blank, total),
              flush=True)
        sys.path.insert(0, V10)
        import wbio
        work = os.path.join(HERE, "_w_gate_recalc.xlsx")
        wbio.build(src, work)
        f = openpyxl.load_workbook(work, data_only=False)
        v = openpyxl.load_workbook(work, data_only=True)
    else:
        print("cached values present (%d of %d formula cells blank on the "
              "probe tabs) - verifying the file as handed over"
              % (blank, total), flush=True)

    gate = Gate()
    sheets = f.sheetnames
    if REVIEW not in sheets:
        print("STOP: %r missing - not the model" % REVIEW)
        return 2
    rvf, rvv = f[REVIEW], v[REVIEW]

    # ================================================== shared derived context
    hdr = header_cols(rvf)
    c_name = hdr.get("name", 2)
    c_title = find_col(hdr, "position title") or 3
    c_div = find_col(hdr, "division") or 7
    c_dept = find_col(hdr, "department") or 8
    c_port = hdr.get("portfolio")
    c_plat = hdr.get("platform")
    c_squad = hdr.get("squad")
    c_country = hdr.get("country")
    c_fte = hdr.get("fte")
    c_full = find_col(hdr, "full cost")
    c_eff = (find_col(hdr, "effective cost") or find_col(hdr, "effective")
             or None)
    c_mtab = find_col(hdr, "mtab")
    c_mstat = find_col(hdr, "mstatus")
    c_ovh = find_col(hdr, "overhead line")
    c_canon = find_col(hdr, "canonical")
    c_id = find_col(hdr, "role id")

    led = [r for r in range(2, rvf.max_row + 1)
           if txt(rvf.cell(r, c_name).value) or txt(rvv.cell(r, c_name).value)]

    def rv(r, c):
        if c is None:
            return None
        x = rvv.cell(r, c).value
        return rvf.cell(r, c).value if x is None else x

    name_of = {r: txt(rv(r, c_name)) for r in led}
    title_of = {r: txt(rv(r, c_title)) for r in led}
    port_of = {r: txt(rv(r, c_port)) for r in led}
    div_of = {r: txt(rv(r, c_div)) for r in led}
    dept_of = {r: txt(rv(r, c_dept)) for r in led}
    squad_of = {r: txt(rv(r, c_squad)) for r in led}
    plat_of = {r: txt(rv(r, c_plat)) for r in led}
    country_of = {r: txt(rv(r, c_country)) for r in led}
    fte_of = {r: num(rv(r, c_fte)) for r in led}
    full_of = {r: num(rv(r, c_full)) or 0.0 for r in led}
    eff_of = {r: num(rv(r, c_eff)) for r in led}
    ovh_of = {r: txt(rv(r, c_ovh)) for r in led}
    mtab_of = {r: txt(rv(r, c_mtab)) for r in led}
    mstat_of = {r: txt(rv(r, c_mstat)) for r in led}
    idmap = {}
    if c_id:
        for r in led:
            x = rv(r, c_id)
            if isinstance(x, str) and re.fullmatch(r"R\d{4}", x.strip()):
                idmap[x.strip()] = r

    tabs2 = [t for t in sheets if t.startswith("2.")]
    T = {t: Tab(t, f[t], v[t], idmap) for t in tabs2}

    def tabname(prefix):
        return find_tab(sheets, prefix)

    # ---- Lists, everything re-read by header
    Lf, Lv = (f["Lists"], v["Lists"]) if "Lists" in sheets else (None, None)

    def list_col(header):
        if Lf is None:
            return None
        for c in range(1, Lf.max_column + 1):
            if txt(Lf.cell(1, c).value) == header:
                return c
        return None

    def list_run(c, stop=("Total",), require=None):
        """every non-blank entry down a Lists column; blanks are gaps in his
        table, not the end of it, so only a stop word ends the run."""
        out = []
        if c is None:
            return out
        for r in range(2, min(Lf.max_row, 60) + 1):
            x = txt(Lf.cell(r, c).value)
            if not x:
                continue
            if x in stop:
                break
            if require and require not in x:
                continue
            out.append((r, x))
        return out

    ovh_labels = [b for _, b in list_run(list_col("Overhead line"))]
    person_c = list_col("Person (Name | Position Title)")
    port_ovr, squad_ovr, ovr_keys = {}, {}, []
    if person_c:
        # his own note sits in the same column, so a key must carry the pipe
        for r, key in list_run(person_c, stop=(), require="|"):
            if len(key) > 90:
                continue
            ovr_keys.append((r, key))
            po = txt(Lf.cell(r, person_c + 1).value)
            so = txt(Lf.cell(r, person_c + 2).value)
            if po:
                port_ovr[norm(key)] = po
            if so:
                squad_ovr[norm(key)] = so
    canon_c = list_col("Squad as typed (col K)") or list_col(
        "Squad as typed")
    canon_map = {}
    if canon_c:
        for r, k in list_run(canon_c, stop=()):
            canon_map[norm(k)] = txt(Lf.cell(r, canon_c + 1).value)
    fund_c = list_col("Squad")
    funded = {}
    if fund_c and txt(Lf.cell(1, fund_c + 1).value) == "Funded by":
        for r, k in list_run(fund_c, stop=()):
            funded[k] = (txt(Lf.cell(r, fund_c + 1).value),
                         Lv.cell(r, fund_c + 2).value)
    lever_factor = {}
    lc = list_col("Lever")
    if lc:
        for r, k in list_run(lc, stop=()):
            lever_factor[k] = num(Lv.cell(r, lc + 1).value)
    gm_cost = None
    if Lf is not None:
        for c in range(1, Lf.max_column + 1):
            for r in range(1, min(Lf.max_row, 40) + 1):
                if txt(Lf.cell(r, c).value) == "GM cost ($m)":
                    gm_cost = num(Lv.cell(r, c + 1).value)
    uplift_fund = None
    for k, (by, basis) in funded.items():
        if norm(k) == "cyber uplift":
            uplift_fund = num(basis)

    # ---- 0.2 Data Config
    cfg_t = tabname("0.2")
    cfgf, cfgv = (f[cfg_t], v[cfg_t]) if cfg_t else (None, None)
    cfg_rows, cfg_total_r, cfg_budget_r, cfg_var_r = [], None, None, None
    if cfgf is not None:
        for r in range(5, cfgf.max_row + 1):
            b = txt(cfgf.cell(r, 2).value)
            if not b or b == "Portfolio":
                continue
            if b == "Total":
                cfg_total_r = r
                continue
            if b == "Budget":
                cfg_budget_r = r
                continue
            if b.startswith("Variance to full TDD budget"):
                cfg_var_r = r
                continue
            if cfg_total_r is None:
                cfg_rows.append((r, b))
    cfg_val = {b: num(cfgv.cell(r, 5).value) for r, b in cfg_rows} \
        if cfgv else {}
    cfg_au = {b: num(cfgv.cell(r, 3).value) for r, b in cfg_rows} \
        if cfgv else {}
    cfg_nz = {b: num(cfgv.cell(r, 4).value) for r, b in cfg_rows} \
        if cfgv else {}
    cfg_spend = {b: num(cfgv.cell(r, 6).value) for r, b in cfg_rows} \
        if cfgv else {}

    def cfg_lookup(lab):
        """the 0.2 budget row a Lights On row draws, matched on the label
        itself first (0.2 B22 now reads Enterprise Data, not TDD Data) and on
        his older alias second, anywhere in the budget block."""
        cands = [lab] + ([BUDROW[lab]] if lab in BUDROW else [])
        for a in cands:
            for _, b in cfg_rows:
                if norm(b) == norm(a):
                    return b
        for a in cands:
            for _, b in cfg_rows:
                if norm(b).startswith(norm(a)):
                    return b
        return None

    # ---- his master
    his = {}
    his_hdr = []
    his_hidden = []
    his_err = None
    try:
        hwb = openpyxl.load_workbook(his_path, data_only=True)
        hws = hwb[hwb.sheetnames[0]]
        his_hdr = [hws.cell(1, c).value for c in range(1, HIS_COLS + 1)]
        his_hidden = sorted(r for r, d in hws.row_dimensions.items()
                            if d.hidden)
        hrows = [r for r in range(2, hws.max_row + 1)
                 if txt(hws.cell(r, 2).value)]
        for r in hrows:
            his[r] = [hws.cell(r, c).value for c in range(1, HIS_COLS + 1)]
    except Exception as e:
        his_err = repr(e)
    hnames = {r: txt(vals[1]) for r, vals in his.items()}
    hport = {r: txt(vals[9]) for r, vals in his.items()}
    hdiv = {r: txt(vals[6]) for r, vals in his.items()}
    hsq = {r: txt(vals[11]) for r, vals in his.items()}
    hplat = {r: txt(vals[10]) for r, vals in his.items()}
    htitle = {r: txt(vals[3]) for r, vals in his.items()}
    hfte = {r: num(vals[15]) for r, vals in his.items()}
    hcost = {r: num(vals[27]) for r, vals in his.items()}
    hcountry = {r: txt(vals[13]) for r, vals in his.items()}

    # ---- the homing derivation, re-run here from his columns + Lists
    def canon_squad(sq, key):
        so = squad_ovr.get(norm(key))
        if so:
            return so
        return canon_map.get(norm(sq), sq)

    def home_of(name, title, port, div, sq):
        key = "%s | %s" % (name, title)
        po = port_ovr.get(norm(key))
        if po:
            return po, "Lists person override"
        cs = canon_squad(sq, key)
        if norm(cs) in ("cyber uplift", "identity"):
            return "TDD Cyber", "TDD Cyber squad"
        t = PORTMAP.get(norm(port))
        if t:
            return t, "Portfolio"
        if norm(port) in ("", "na", "none"):
            t = DIVMAP.get(norm(div))
            if t:
                return t, "Division"
        return None, "unresolved"

    derived_home, home_basis = {}, {}
    for r in his:
        t, basis = home_of(hnames[r], htitle[r], hport[r], hdiv[r], hsq[r])
        derived_home[r] = t
        home_basis[r] = basis

    # ---- the overhead-line derivation, re-run here from his titles + squads
    def derive_line(title, sq, plat, key, name=""):
        if norm(name) in OFF_LINE:
            return OFF_LINE[norm(name)]        # his ruling, person by person
        cs = canon_squad(sq, key)
        if norm(sq) == "leadership" or norm(plat) == "leadership":
            cs = "Leadership"
        if cs in ovh_labels:
            return cs
        t = norm(title)
        if t in DM_TITLES:                     # his own Delivery Manager list
            return "Delivery Manager"
        if "head of " in t:
            return "Head of Technology"
        if "tdd bp" in t:
            return "Business Partner"
        if "domain architect" in t or "enterprise architect" in t:
            return "Domain Architect"
        if "delivery man" in t:
            return "Delivery Manager"
        if ("technology manager" in t or "technology manger" in t
                or "tech manager" in t):
            return "Technology Manager"
        return "Squad"

    derived_line = {r: derive_line(htitle[r], hsq[r], hplat[r],
                                   "%s | %s" % (hnames[r], htitle[r]),
                                   hnames[r])
                    for r in his}

    # ---- role rows across the workbook, keyed by REVIEW row
    role_by_led = {}
    dupes = []
    for t in tabs2:
        for ro in T[t].roles:
            lr = ro["ledger"]
            if lr is None:
                continue
            if lr in role_by_led:
                dupes.append((lr, role_by_led[lr][0], t))
            role_by_led.setdefault(lr, (t, ro))

    def after_of(rows):
        return sum((role_by_led[r][1]["G"] or 0) for r in rows
                   if r in role_by_led) / 1e6

    def line_rows(line):
        return [r for r in led if ovh_of.get(r) == line]

    bp_pot = after_of(line_rows("Business Partner"))
    da_pot = after_of(line_rows("Domain Architect"))
    all_after = sum((ro["G"] or 0) for t in tabs2 for ro in T[t].roles) / 1e6
    tab_after = {t: T[t].total(19) for t in tabs2}

    # the role basis: every person's effective cost after their lever, with
    # the Lists vendor-rate exemption for the WIPRO roles, and gross of the
    # 2.11 part-charges to the cyber uplift programme
    def factor_for(r, lever):
        if norm(country_of.get(r)) == "wipro":
            return 1.0
        fac = lever_factor.get(lever, 1)
        return 1.0 if fac is None else fac

    role_basis = {}
    for r in led:
        ro = role_by_led.get(r)
        if not ro:
            continue
        base = eff_of.get(r)
        if base is None:
            base = full_of[r] * (fte_of[r] if (fte_of[r] or 1) < 1 else 1)
        role_basis[r] = base * factor_for(r, ro[1]["lever"])
    role_basis_total = sum(role_basis.values()) / 1e6

    # what each tab charges out of its own roles (the 2.11 part-charges to the
    # cyber uplift programme): the tab total is net of it, the Lights On C
    # column carries it, and the row books it as a funded item
    charge_out = {}
    for t in tabs2:
        s = 0.0
        for ro in T[t].roles:
            lr = ro["ledger"]
            if lr is None:
                continue
            s += role_basis.get(lr, 0.0) - (ro["G"] or 0)
        charge_out[t] = s / 1e6

    # the cyber uplift part-charge, read off 1.14's own funding block
    uplift_charge = None
    t114 = tabname("1.14")
    if t114:
        for r in range(1, f[t114].max_row + 1):
            b = norm(f[t114].cell(r, 8).value)
            if "part-charge" in b or "part charge" in b:
                uplift_charge = num(v[t114].cell(r, 10).value)
                break

    def au_nz(rows, basis=None):
        """AU and NZ halves ($m) of a set of REVIEW rows; NZ is Country NZ,
        everyone else AU."""
        au = nz = 0.0
        for r in rows:
            if basis is not None:
                x = basis.get(r, 0.0)
            else:
                ro = role_by_led.get(r)
                x = (ro[1]["G"] or 0) if ro else 0.0
            if norm(country_of.get(r)) == "nz":
                nz += x
            else:
                au += x
        return au / 1e6, nz / 1e6

    def au_weight(rows, basis=None):
        a, n = au_nz(rows, basis)
        return None if (a + n) == 0 else a / (a + n)

    def own_overhead(prefix, roles=None):
        t = tabname(prefix)
        if t is None or t not in T:
            return None
        rs = roles if roles is not None else T[t].roles
        s = 0.0
        for ro in rs:
            lr = ro["ledger"]
            if lr is not None and ovh_of.get(lr) in ovh_labels:
                s += ro["G"] or 0
        return s / 1e6

    # ============================================================ A raw identity
    g = gate.group("A", "raw identity")
    try:
        if his_err:
            gate.sub(g, False, "his master loads: %s" % his_err)
        else:
            gate.sub(g, len(his) == HIS_ROWS,
                     "his master carries %d data rows (spec %d)"
                     % (len(his), HIS_ROWS))
            gate.sub(g, len(led) == len(his),
                     "REVIEW carries %d named rows against his %d"
                     % (len(led), len(his)))
            first = min(his) if his else 2
            gate.sub(g, led and led[0] == first and led[-1] == first
                     + len(his) - 1,
                     "REVIEW block runs rows %s..%s (his %s..%s)"
                     % (led[0] if led else None, led[-1] if led else None,
                        first, first + len(his) - 1))
            bad_h = []
            for c in range(1, HIS_COLS + 1):
                a = rvf.cell(1, c).value
                b = his_hdr[c - 1]
                if txt(a) != txt(b):
                    bad_h.append((openpyxl.utils.get_column_letter(c),
                                  txt(a), txt(b)))
            gate.sub(g, not bad_h,
                     "his header row A1:AC1 verbatim (duplicate EE Number, "
                     "MyHR ee no): %d off %s" % (len(bad_h), bad_h[:3]))
            diffs = []
            for i, hr in enumerate(sorted(his)):
                r = first + i
                for c in range(1, HIS_COLS + 1):
                    want = his[hr][c - 1]
                    got = rvv.cell(r, c).value
                    if got is None:
                        got = rvf.cell(r, c).value
                    wn, gn = num(want), num(got)
                    if wn is not None and gn is not None:
                        if abs(wn - gn) > max(TOLC, abs(wn) * 1e-12):
                            diffs.append((r, c, want, got))
                    elif txt(want) != txt(got):
                        diffs.append((r, c, want, got))
            gate.sub(g, not diffs,
                     "every cell of A2:AC%d identical to his file (text "
                     "exact, numbers 1e-9): %d off %s"
                     % (first + len(his) - 1, len(diffs),
                        [(d[0], openpyxl.utils.get_column_letter(d[1]),
                          d[2], d[3]) for d in diffs[:4]]))
            below = [r for r in range(first + len(his), rvf.max_row + 1)
                     if any(txt(rvf.cell(r, c).value)
                            for c in range(1, HIS_COLS + 1))]
            gate.sub(g, not below,
                     "nothing left below his block in A:AC: %d rows %s"
                     % (len(below), below[:4]))
            # verbatim identity covers what the tab shows, not only what it
            # holds: his file hides no row, so a REVIEW that renders blank
            # below the header is not his file
            hid = sorted(r for r, d in rvf.row_dimensions.items() if d.hidden)
            gate.sub(g, hid == his_hidden,
                     "no hidden row on REVIEW, his file hides none (%d hidden "
                     "here%s, %d in his file)"
                     % (len(hid), (": %s%s" % (hid[:8],
                                               " .." if len(hid) > 8 else ""))
                        if hid else "", len(his_hidden)))
            gate.note(g, "his 29 columns end at %s, helper block starts at %s"
                      % (openpyxl.utils.get_column_letter(HIS_COLS),
                         openpyxl.utils.get_column_letter(HIS_COLS + 2)))
            helpers_left = [openpyxl.utils.get_column_letter(c)
                            for c in (c_mtab, c_mstat, c_ovh, c_eff, c_id)
                            if c is not None and c <= HIS_COLS]
            gate.sub(g, not helpers_left,
                     "no helper column sits inside his 29: %s" % helpers_left)
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ================================================================ B homing
    g = gate.group("B", "homing")
    home_led = {}
    try:
        unres = [r for r in his if derived_home[r] is None]
        gate.sub(g, not unres,
                 "every one of his %d rows homes to exactly one tab: %d "
                 "unresolved %s"
                 % (len(his), len(unres),
                    [(r, hnames[r], hport[r], hdiv[r]) for r in unres[:4]]))
        fb = [r for r in his if home_basis[r] == "Division"]
        gate.note(g, "homed by Portfolio %d, by Division fallback %d, by "
                  "Lists override %d, by TDD Cyber squad %d"
                  % (sum(1 for r in his if home_basis[r] == "Portfolio"),
                     len(fb),
                     sum(1 for r in his
                         if home_basis[r] == "Lists person override"),
                     sum(1 for r in his
                         if home_basis[r] == "TDD Cyber squad")))
        cnt = collections.Counter(derived_home[r] for r in his)
        gate.note(g, "derived homing: %s"
                  % ", ".join("%s %d" % (k, cnt[k]) for k in sorted(cnt)
                              if k))
        # MTab helper against the derivation
        if len(led) == len(his) and c_mtab:
            first = min(his)
            offm = []
            for i, hr in enumerate(sorted(his)):
                r = first + i
                want = derived_home[hr]
                got = mtab_of.get(r)
                home_led[r] = want
                if want and norm(got) != norm(want):
                    offm.append((r, hnames[hr], got, want))
            gate.sub(g, not offm,
                     "the MTab helper matches the derivation on every row: "
                     "%d off %s" % (len(offm), offm[:4]))
        else:
            gate.sub(g, False,
                     "MTab helper present and the block aligns (col %s, %d "
                     "REVIEW rows against %d his rows)"
                     % (c_mtab, len(led), len(his)))
        # real 2.x membership
        placed = {r: t for r, (t, _) in role_by_led.items()}
        gate.sub(g, not dupes,
                 "each person sits on exactly one 2.x tab: %d duplicated %s"
                 % (len(dupes), dupes[:4]))
        missing = [r for r in led if r not in placed]
        gate.sub(g, not missing,
                 "every REVIEW row is placed on a 2.x tab: %d missing %s"
                 % (len(missing), [(r, name_of[r]) for r in missing[:4]]))
        wrong = []
        for r, t in placed.items():
            want = home_led.get(r) or mtab_of.get(r)
            pref = HOME2TAB.get(want)
            if pref and not t.startswith(pref + " ") and t != pref:
                wrong.append((r, name_of.get(r), t, want))
        gate.sub(g, not wrong,
                 "every person sits on the tab their homing says: %d off %s"
                 % (len(wrong), wrong[:4]))
        # his funded squad names against Lists
        if his:
            hsquads = {txt(hsq[r]) for r in his if txt(hsq[r])}
            drift = [k for k in funded
                     if k not in hsquads
                     and norm(k) not in {norm(x) for x in hsquads}
                     and norm(k) not in ("cyber uplift", "identity")]
            gate.sub(g, not drift,
                     "every Lists funded squad name matches one of his Squad "
                     "values: %d adrift %s" % (len(drift), drift[:5]))
            # the person-keyed override table, keyed on Name | Position Title
            hkey = collections.Counter(
                norm("%s | %s" % (hnames[r], htitle[r])) for r in his)
            dead = [k for _, k in ovr_keys if hkey.get(norm(k), 0) != 1]
            gate.sub(g, ovr_keys and not dead,
                     "all %d Lists person overrides key onto exactly one row "
                     "of his file: %d dead %s"
                     % (len(ovr_keys), len(dead), dead[:4]))
            cyb = [k for _, k in ovr_keys
                   if norm(port_ovr.get(norm(k), "")) == "tdd cyber"]
            cyb_home = [r for r in his if derived_home[r] == "TDD Cyber"]
            gate.sub(g, len(cyb_home) == len(cyb) and len(cyb) > 0,
                     "the %d TDD Cyber overrides home %d of his people to "
                     "2.15" % (len(cyb), len(cyb_home)))
        # the reversals his new file lands
        def find_person(n):
            return [r for r in his if norm(hnames[r]) == norm(n)]

        for who, want_home in (("Viren Khatri", "EGI"),
                               ("Ed Tacey", "Customer"),
                               ("Sarsha Tanner", "COE BP&T"),
                               ("Murray Mitchell", "TDD Group Functions"),
                               ("Scott McKenzie", "TDD Group Functions")):
            rs = find_person(who)
            got = derived_home.get(rs[0]) if rs else None
            wb_rows = [r for r in led if norm(name_of[r]) == norm(who)]
            wb_tab = placed.get(wb_rows[0]) if wb_rows else None
            gate.sub(g, len(rs) == 1 and got == want_home
                     and wb_tab is not None
                     and wb_tab.startswith(HOME2TAB[want_home]),
                     "%s homes to %s (his file says %s, the model puts him on "
                     "%s)" % (who, want_home, got, wb_tab))
        # the ring fenced rows are vacancies
        rf = [r for r in his if norm(hnames[r]).startswith("ring fenced")]
        first = min(his) if his else 2
        rf_wb = [first + sorted(his).index(r) for r in rf]
        badrf = [r for r in rf_wb if r in mstat_of and mstat_of[r] != "Vacant"]
        gate.sub(g, rf and not badrf,
                 "the %d ring fenced rows read as Vacant in MStatus: %d off %s"
                 % (len(rf), len(badrf), badrf[:4]))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ================================================================ C ledger
    g = gate.group("C", "ledger")
    try:
        gate.sub(g, len(led) == HIS_ROWS,
                 "roles %d (his file %d)" % (len(led), HIS_ROWS))
        if c_id:
            want_ids = ["R%04d" % (i + 1) for i in range(len(led))]
            got_ids = [txt(rv(r, c_id)) for r in led]
            off = [i for i, (a, b) in enumerate(zip(got_ids, want_ids))
                   if a != b]
            gate.sub(g, not off,
                     "Role IDs run R0001..R%04d in his row order: %d off %s"
                     % (len(led), len(off),
                        [(led[i], got_ids[i], want_ids[i])
                         for i in off[:3]]))
        else:
            gate.sub(g, False, "REVIEW carries a Role ID column")
        gate.sub(g, c_id is not None and c_id > HIS_COLS,
                 "the Role ID column sits right of his 29 (col %s)" % c_id)
        noid = ["%s!A%d" % (t, ro["row"]) for t in tabs2 for ro in T[t].roles
                if not (ro["id"] and re.fullmatch(r"R\d{4}", ro["id"]))]
        gate.sub(g, not noid,
                 "every 2.x role row carries an ID in col A: %d without %s"
                 % (len(noid), noid[:4]))
        misres = []
        for t in tabs2:
            for ro in T[t].roles:
                if not ro["id"]:
                    continue
                lr = idmap.get(ro["id"])
                if lr is None or norm(name_of.get(lr)) != norm(ro["name"]):
                    misres.append("%s!r%d %s -> %r vs %r"
                                  % (t, ro["row"], ro["id"],
                                     name_of.get(lr or -1), ro["name"]))
        gate.sub(g, not misres,
                 "every ID resolves to its own REVIEW row: %d off %s"
                 % (len(misres), misres[:3]))
        pat = re.compile(r"!\$(?:B|C|AK|AJ|AR|AT)\$\d+(?!\d)(?!:)")
        anchored = 0
        for t in tabs2:
            for ro in T[t].roles:
                for cc in (2, 3, 4):
                    fx = T[t].wf.cell(ro["row"], cc).value
                    if isinstance(fx, str) and REVIEW in fx and pat.search(fx):
                        anchored += 1
        gate.sub(g, anchored == 0,
                 "row-anchored single-cell refs left in role-row B/C/D: %d"
                 % anchored)
        # the overhead line, re-derived from his titles and squads
        if his and len(led) == len(his) and c_ovh:
            first = min(his)
            offl = []
            for i, hr in enumerate(sorted(his)):
                r = first + i
                want, got = derived_line[hr], ovh_of.get(r)
                if norm(want) != norm(got):
                    offl.append((r, hnames[hr], htitle[hr], got, want))
            gate.sub(g, not offl,
                     "the Overhead line helper matches the derivation from "
                     "his titles and squads: %d off %s"
                     % (len(offl), offl[:4]))
            dc = collections.Counter(derived_line[r] for r in his
                                     if derived_line[r] != "Squad")
            gate.note(g, "derived overhead lines: %s"
                      % ", ".join("%s %d" % (k, dc[k]) for k in sorted(dc)))
        else:
            gate.sub(g, False, "Overhead line helper present and aligned "
                     "(col %s)" % c_ovh)
        counts = collections.Counter(ovh_of[r] for r in led
                                     if ovh_of[r] in ovh_labels)
        # 3.2 holds the model's own count of each line; a line his file leaves
        # empty (Program Management on the 05/08 cut) is a truth, not a fault,
        # so the test is agreement with 3.2 rather than a floor of one role
        t32 = tabname("3.2")
        r32, w32 = {}, v[t32] if t32 else None
        if w32 is not None:
            for r in range(4, min(w32.max_row, 20) + 1):
                b = txt(w32.cell(r, 2).value)
                if b in ovh_labels:
                    r32[b] = r
        for line in ovh_labels:
            if line.startswith("Leadership"):
                continue        # the 8 GMs sit above the role mapping
            rs = line_rows(line)
            onwb = sum(1 for r in rs if r in role_by_led)
            got32 = num(w32.cell(r32[line], 7).value) if line in r32 else None
            gate.sub(g, onwb == len(rs) and got32 == len(rs),
                     "%s line: %d roles derived, %d placed on a 2.x tab, 3.2 "
                     "counts %s" % (line, len(rs), onwb, got32))
        if w32 is not None and r32:
            negH = [(b, num(w32.cell(r, 8).value)) for b, r in r32.items()
                    if (num(w32.cell(r, 8).value) or 0) < -1e-9]
            gate.sub(g, not negH,
                     "no 3.2 line shows a negative 'roles not applied in "
                     "archetype': %d off %s" % (len(negH), negH[:4]))
            emptybad = []
            for b, r in r32.items():
                if len(line_rows(b)) or b.startswith("Leadership"):
                    continue
                cells = [num(w32.cell(r, c).value) for c in (5, 6, 7, 8)]
                if any(x is None or abs(x) > 1e-9 for x in cells):
                    emptybad.append((b, cells))
            gate.sub(g, not emptybad,
                     "a line with no roles in his file reads 0 across E F G H "
                     "on 3.2: %d off %s" % (len(emptybad), emptybad[:3]))
        gate.note(g, "line counts %s"
                  % ", ".join("%s %d" % (k, counts[k])
                              for k in sorted(counts)))
        sk = [r for r in led if norm(name_of[r]) == "shane ker"]
        gate.sub(g, len(sk) == 1 and ovh_of[sk[0]] == "Technology Manager",
                 "Shane Ker on the Technology Manager line (%s)"
                 % [ovh_of[r] for r in sk])
        ed = [r for r in led if norm(name_of[r]) == "ed tacey"]
        gate.sub(g, len(ed) == 1 and ovh_of[ed[0]] != "Head of Technology",
                 "Ed Tacey off the Head of Technology line (%s)"
                 % [ovh_of[r] for r in ed])
        mgrs = [r for r in led if norm(title_of[r]) in
                ("delivery assurance manager", "delivery excellence manager")]
        gate.sub(g, len(mgrs) == 2
                 and all(ovh_of[r] == "Delivery Manager" for r in mgrs),
                 "the Delivery Assurance and Delivery Excellence managers on "
                 "the Delivery Manager line: %s"
                 % [(name_of[r], title_of[r], ovh_of[r]) for r in mgrs])
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ================================================================ D levers
    g = gate.group("D", "levers")
    try:
        bad_lever, bad_hold, holds = [], [], 0
        for t in tabs2:
            for ro in T[t].roles:
                if ro["lever"] not in LEVERS:
                    bad_lever.append("%s!E%d %r" % (t, ro["row"], ro["lever"]))
                if ro["lever"] == "Hold":
                    holds += 1
                    if ro["G"] is None or abs(ro["G"]) > 1e-9:
                        bad_hold.append("%s!r%d G=%s" % (t, ro["row"],
                                                         ro["G"]))
        gate.sub(g, not bad_lever,
                 "every lever one of the four values, no Paused: %d off %s"
                 % (len(bad_lever), bad_lever[:4]))
        gate.sub(g, not bad_hold,
                 "every Hold role costs 0 after the lever (%d traced): %d off "
                 "%s" % (holds, len(bad_hold), bad_hold[:4]))
        gate.sub(g, lever_factor.get("Hold") == 0,
                 "the Lists lever table prices Hold at 0 (%s)"
                 % lever_factor.get("Hold"))
        stale = ["%s!r%d %s" % (t, ro["row"], ro["name"])
                 for t in tabs2 for ro in T[t].roles
                 if ro["status"] == "Filled" and ro["lever"] == "Hire"]
        gate.sub(g, not stale,
                 "BLD-18: no stale Filled status on a Hire lever: %d off %s"
                 % (len(stale), stale[:4]))
        vac_not_lever = ["%s!r%d %s" % (t, ro["row"], ro["title"])
                         for t in tabs2 for ro in T[t].roles
                         if ro["status"] == "Vacant"
                         and ro["lever"] == "Filled"]
        gate.note(g, "vacancies carried at the Filled lever (his own edits): "
                  "%d %s" % (len(vac_not_lever), vac_not_lever[:3]))
        # person-keyed carry against the previous build
        if os.path.exists(prev_path) and os.path.abspath(prev_path) \
                != os.path.abspath(src):
            pf = openpyxl.load_workbook(prev_path, data_only=False)
            pv = openpyxl.load_workbook(prev_path, data_only=True)
            prev = {}
            prev_vac = collections.defaultdict(collections.Counter)
            prev_vac_any = collections.defaultdict(collections.Counter)
            for ws in pf.worksheets:
                if not ws.title.startswith("2."):
                    continue
                wsv = pv[ws.title]
                pfx = ws.title.split()[0]
                for r in range(1, ws.max_row + 1):
                    d = ws.cell(r, 4).value
                    if not (isinstance(d, str) and d.startswith("=")
                            and REVIEW in d):
                        continue
                    nm = wsv.cell(r, 2).value
                    ti = wsv.cell(r, 3).value
                    st = txt(wsv.cell(r, 4).value)
                    lv = txt(ws.cell(r, 5).value)
                    if is_vacancy(nm, st):
                        # a vacancy has no name to key on, so it carries by
                        # its title on its tab, and by its title anywhere if
                        # the role moved tab in the rehoming
                        prev_vac[(pfx, tkey(ti))][lv] += 1
                        prev_vac_any[tkey(ti)][lv] += 1
                    else:
                        prev.setdefault((norm(nm), norm(ti)), []).append(
                            (st, lv))
            pf.close()
            pv.close()
            # 'Vacant | <title>' repeats, so compare lever multisets per
            # person key, allowing only the Hire -> Filled normalisation a
            # filled vacancy brings with it
            kept = newd = dep = filled_up = off = vkept = vnew = 0
            offs = []
            cur = collections.defaultdict(list)
            cur_vac = collections.defaultdict(list)
            for t in tabs2:
                pfx = t.split()[0]
                for ro in T[t].roles:
                    if is_vacancy(ro["name"], ro["status"]):
                        cur_vac[(pfx, tkey(ro["title"]))].append(ro)
                    else:
                        cur[(norm(ro["name"]), norm(ro["title"]))].append(ro)
            # the vacancies: the typed lever carries with the title; only a
            # title his file has never had on that tab defaults to Hire
            for (pfx, k), ros in sorted(cur_vac.items()):
                pool = collections.Counter(prev_vac.get((pfx, k), {}))
                if not pool:
                    pool = collections.Counter(prev_vac_any.get(k, {}))
                if not pool:
                    vnew += len(ros)
                    for ro in ros:
                        if ro["lever"] != "Hire":
                            off += 1
                            offs.append((pfx, ro["title"], ro["lever"],
                                         "new vacancy, expected Hire"))
                    continue
                for ro in ros:
                    if pool.get(ro["lever"], 0) > 0:
                        pool[ro["lever"]] -= 1
                        vkept += 1
                    elif ro["lever"] == "Hire":
                        vnew += 1          # an extra vacancy of the same title
                    else:
                        off += 1
                        offs.append((pfx, ro["title"], ro["lever"],
                                     "not carried from %s"
                                     % dict(prev_vac.get((pfx, k))
                                            or prev_vac_any.get(k, {}))))
            for key, ros in cur.items():
                old = prev.get(key)
                if not old:
                    newd += len(ros)
                    for ro in ros:
                        want = "Hire" if ro["status"] == "Vacant" else "Filled"
                        if ro["lever"] != want:
                            off += 1
                            offs.append((ro["name"], ro["title"], ro["status"],
                                         ro["lever"], "expected " + want))
                    continue
                kept += min(len(old), len(ros))
                oldms = collections.Counter(lv for _, lv in old)
                curms = collections.Counter(ro["lever"] for ro in ros)
                if oldms == curms:
                    continue
                # a Hire that reads Filled now is a filled vacancy, allowed
                slack = min(oldms["Hire"] - curms["Hire"],
                            curms["Filled"] - oldms["Filled"])
                if slack > 0:
                    filled_up += slack
                    oldms["Hire"] -= slack
                    oldms["Filled"] += slack
                if oldms != curms and len(old) == len(ros):
                    off += 1
                    offs.append((ros[0]["name"], ros[0]["title"],
                                 dict(oldms), dict(curms)))
            dep = len([k for k in prev if k not in cur])
            gate.sub(g, off == 0,
                     "levers carry person-keyed, vacancies by title on their "
                     "tab (people kept %d, new %d defaulted, vacancy filled "
                     "%d, vacancy levers carried %d, new vacancies %d, "
                     "departed %d): %d off %s"
                     % (kept, newd, filled_up, vkept, vnew, dep, off,
                        offs[:4]))
            gate.note(g, "carry report against %s: people kept %d / "
                      "defaulted-new %d / filled %d / vacancies carried %d / "
                      "new vacancies %d / departed %d"
                      % (os.path.basename(prev_path), kept, newd, filled_up,
                         vkept, vnew, dep))
        else:
            gate.note(g, "no separate previous build to carry against (%s)"
                      % prev_path)
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ============================================================= E part-time
    g = gate.group("E", "part-time")
    try:
        gate.sub(g, c_eff is not None,
                 "REVIEW carries an effective cost helper column (col %s)"
                 % (openpyxl.utils.get_column_letter(c_eff) if c_eff
                    else None))
        pt = [r for r in led if fte_of[r] is not None and 0 < fte_of[r] < 1]
        gate.note(g, "FTE below 1: %d rows" % len(pt))
        offe, offp = [], []
        for r in led:
            fte = fte_of[r]
            want = full_of[r] * fte if (fte is not None and 0 < fte < 1) \
                else full_of[r]
            got = eff_of.get(r)
            if got is None or abs(got - want) > 1e-6 * max(1.0, abs(want)):
                if fte is not None and 0 < fte < 1:
                    offe.append((r, name_of[r], fte, full_of[r], got, want))
            ro = role_by_led.get(r)
            if ro and fte is not None and 0 < fte < 1:
                if ro[1]["F"] is None or abs(ro[1]["F"] - want) > 1e-4:
                    offp.append((r, name_of[r], ro[1]["F"], want))
        gate.sub(g, not offe,
                 "the effective cost helper reads Full Cost x FTE below 1: "
                 "%d off %s" % (len(offe), offe[:4]))
        gate.sub(g, not offp,
                 "the 2.x role cost prices off the effective cost: %d off %s"
                 % (len(offp), offp[:4]))
        for r in pt[:12]:
            gate.note(g, "  %-24s FTE %.2f raw %12.2f effective %12.2f"
                      % (name_of[r][:24], fte_of[r], full_of[r],
                         eff_of.get(r) if eff_of.get(r) is not None
                         else full_of[r] * fte_of[r]))
        if his:
            first = min(his)
            hpt = [r for r in his if hfte[r] is not None and 0 < hfte[r] < 1]
            gate.sub(g, len(hpt) == len(pt),
                     "his file's %d part-timers all carry through (%d in "
                     "REVIEW)" % (len(hpt), len(pt)))
            rawoff = []
            for hr in hpt:
                r = first + sorted(his).index(hr)
                if hcost[hr] is not None and abs(full_of.get(r, 0)
                                                 - hcost[hr]) > 1e-6:
                    rawoff.append((hnames[hr], hcost[hr], full_of.get(r)))
            gate.sub(g, not rawoff,
                     "his raw Full Cost cells stay untouched for the "
                     "part-timers: %d off %s" % (len(rawoff), rawoff[:4]))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ====================================================== the Lights On tabs
    class LOTab:
        def __init__(self, title):
            self.title = title
            self.wf, self.wv = f[title], v[title]
            self.hdr_r = None
            for r in range(1, min(self.wf.max_row, 30) + 1):
                if txt(self.wf.cell(r, 2).value) == HEADERS[2]:
                    self.hdr_r = r
                    break
            self.cols = {}
            if self.hdr_r:
                for c in range(2, self.wf.max_column + 1):
                    k = txt(self.wf.cell(self.hdr_r, c).value)
                    if k:
                        self.cols.setdefault(k, c)
            self.rows, self.labels = {}, []
            self.total_r = self.budget_r = None
            r = (self.hdr_r or 0) + 1
            while self.hdr_r and r <= self.wf.max_row and len(self.labels) < 40:
                b = txt(self.wv.cell(r, 2).value) or txt(
                    self.wf.cell(r, 2).value)
                if b == "Total":
                    self.total_r = r
                    break
                if b:
                    self.labels.append(b)
                    self.rows[b] = r
                r += 1
            if self.total_r:
                for rr in range(self.total_r + 1, self.total_r + 4):
                    if "budget" in norm(self.wv.cell(rr, 2).value):
                        self.budget_r = rr
                        break

        def colof(self, k):
            """the column a named quantity sits in on THIS tab: 3.6 shifts
            M O Q R S right of its AU NZ block, so resolve by his header."""
            c = self.cols.get(HEADERS.get(COL[k], ""))
            return c or COL[k]

        def get(self, lab, col):
            r = self.rows.get(lab)
            if r is None:
                return None
            c = self.colof(col) if isinstance(col, str) else col
            return num(self.wv.cell(r, c).value)

    lo = LOTab(LO) if LO in sheets else None
    lo2 = LOTab(LO2) if LO2 in sheets else None

    # -------------------------------------------------- the independent model
    exp = {}
    model_note = []
    try:
        # support percentages per row
        pctm = {}
        for lab, p1 in SRC1X.items():
            t1 = tabname(p1)
            if t1:
                pm, clash = pct_map(f[t1], v[t1])
                pctm[lab] = pm
                if clash:
                    model_note.append("%s support %% clash %s" % (t1,
                                                                  clash[:2]))
        # per-row people, from REVIEW homing + the 2.x placement
        rows_of = collections.defaultdict(list)
        for r in led:
            if r not in role_by_led:
                continue
            home = home_led.get(r) or mtab_of.get(r)
            if home == "Customer":
                side = "Z Customer" if norm(port_of.get(r)) in (
                    "z energy (digital)", "z customer", "z") \
                    else "Ampol Customer"
                rows_of[side].append(r)
            elif home in ("COE SA&D", "COE BP&T"):
                rows_of[home].append(r)
            else:
                lab = {"COE Cyber": "COE Cyber Risk & Service Ops"}.get(
                    home, home)
                rows_of[lab].append(r)

        def sect_after(tab, sect):
            return sum(ro["G"] or 0 for ro in T[tab].sections.get(sect, [])) \
                / 1e6

        # which side of the Customer split each 2.2 squad sits on, read off
        # the people in it
        cust_side = {}
        t22 = tabname("2.2")
        if t22 and t22 in T:
            zset = set(rows_of["Z Customer"])
            for sect, ros in T[t22].sections.items():
                lrs = [ro["ledger"] for ro in ros if ro["ledger"] is not None]
                if not lrs:
                    continue
                z = sum(1 for r in lrs if r in zset)
                cust_side[sect] = ("Z Customer" if z * 2 > len(lrs)
                                   else "Ampol Customer")

        # the COE pair splits, resolved against the tab's own planned spend
        coe_split = {}
        for prefix, (a, b) in SPLITS.items():
            if prefix == "2.2":
                continue
            t = tabname(prefix)
            if t is None or t not in T:
                continue
            comp = {}
            for lab, start in ((a, "Business Partnering planned spend")
                               if prefix == "2.12" else
                               (a, "Strategy & Architecture planned spend"),
                               (b, "Transformation planned spend")
                               if prefix == "2.12" else
                               (b, "Data planned spend")):
                val, _ = T[t].label_value(start)
                comp[lab] = val
            pot = bp_pot if prefix == "2.12" else da_pot
            potline = "Business Partner" if prefix == "2.12" \
                else "Domain Architect"
            sects = list(T[t].sections)
            potsect = {ro["section"] for ro in T[t].roles
                       if ro["ledger"] is not None
                       and ovh_of.get(ro["ledger"]) == potline}
            # every subset whose spend nets to his planned component; the pot
            # is charged out of the COE that hosts it, so a subset holding the
            # pot sections wins, and the widest such subset at that
            cands = []
            for k in range(0, len(sects) + 1):
                for combo in itertools.combinations(sects, k):
                    s = sum(sect_after(t, x) for x in combo)
                    if potsect & set(combo):
                        s -= pot
                    if comp[a] is not None and abs(s - comp[a]) <= 1e-6:
                        cands.append(set(combo))
            withpot = [c for c in cands if potsect and potsect <= c]
            found = None
            if withpot:
                found = max(withpot, key=len)
            elif cands:
                found = max(cands, key=len)
            coe_split[prefix] = {"comp": comp, "pot": pot, "potline": potline,
                                 "sections_a": found, "a": a, "b": b,
                                 "tab": t}
            if found is None:
                model_note.append("%s planned-spend split not resolvable from "
                                  "the sections" % prefix)

        for lab in LROWS:
            prefix = SRCTAB[lab]
            t = tabname(prefix)
            e = {}
            if t is None or t not in T:
                continue
            tb = T[t]
            mine = rows_of.get(lab, [])
            split = prefix in SPLITS
            if lab in ("Ampol Customer", "Z Customer"):
                # the pair splits the way the tab documents it: the squads
                # that carry a 1.x support line sit with their own side, and
                # everything shared - the squads with no 1.x line and the
                # whole overhead pool - divides on the two Support Costs.
                # C, D and I are finished once both E values are known.
                pmc = pctm.get(lab, {})
                e["side_S"] = e["shared_S"] = 0.0
                e["side_P"] = e["shared_P"] = 0.0
                for rr, lab2 in tb.grid:
                    sv = num(tb.wv.cell(rr, 19).value) or 0.0
                    pvv = num(tb.wv.cell(rr, 16).value) or 0.0
                    known = norm(lab2) in pmc and lab2 in cust_side
                    if known and cust_side[lab2] == lab:
                        e["side_S"] += sv
                        e["side_P"] += pvv
                    elif not known:
                        e["shared_S"] += sv
                        e["shared_P"] += pvv
                e["pool_S"] = tb.gval("Overhead roles total", 19) or 0.0
                e["shared_co"] = charge_out.get(t, 0.0)
                e["people"] = list(mine)
            elif prefix in ("2.12", "2.13"):
                cs = coe_split.get(prefix, {})
                comp = cs.get("comp", {}).get(lab)
                holds_pot = bool(cs.get("sections_a")) and (
                    lab == cs.get("a")) if cs.get("sections_a") else None
                mysects = cs.get("sections_a") if lab == cs.get("a") else (
                    (set(tb.sections) - (cs.get("sections_a") or set()))
                    if cs.get("sections_a") is not None else None)
                pot = cs.get("pot") or 0
                potline = cs.get("potline")
                if mysects is not None:
                    ros = [ro for ro in tb.roles if ro["section"] in mysects]
                    e["C"] = sum(ro["G"] or 0 for ro in ros) / 1e6
                    ovh = sum(ro["G"] or 0 for ro in ros
                              if ro["ledger"] is not None
                              and ovh_of.get(ro["ledger"]) in ovh_labels) / 1e6
                    haspot = any(ro["ledger"] is not None
                                 and ovh_of.get(ro["ledger"]) == potline
                                 for ro in ros)
                    e["I"] = ovh - (pot if haspot else 0)
                    e["people"] = [ro["ledger"] for ro in ros
                                   if ro["ledger"] is not None]
                    e["holds_pot"] = haspot
                    e["pot"] = pot
                elif comp is not None:
                    e["C"] = comp
                    e["I"] = None
                e["planned"] = comp
                e["potline"] = potline
            else:
                # C carries the whole people cost, the slice charged out of
                # the tab's own roles included
                e["C"] = (tb.total(19) or 0) + charge_out.get(t, 0.0)
                e["I"] = own_overhead(prefix)
                e["people"] = [ro["ledger"] for ro in tb.roles
                               if ro["ledger"] is not None]
            # D, the funded-outside slice; a charged-out slice is funded by
            # the programme that takes it, so it books here too
            co = charge_out.get(t, 0.0)
            if lab == "EGI":
                e["D"] = e.get("C")
            elif prefix in ("2.12", "2.13", "2.11"):
                e["D"] = co if abs(co) > TOL else 0.0
            elif lab in ("Ampol Customer", "Z Customer"):
                # the funded slice on the tab's own basis, by side
                e["D"] = sum((num(tb.wv.cell(rr, 16).value) or 0)
                             for rr, lab2 in tb.grid
                             if cust_side.get(lab2, lab) == lab)
            else:
                e["D"] = tb.total(16)
            # E, the support cost
            if prefix in ("2.12", "2.13", "2.11"):
                if e.get("I") is not None and e.get("planned") is not None:
                    e["E"] = e["planned"] - e["I"]
                elif prefix == "2.11" and e.get("I") is not None:
                    e["E"] = (tb.total(19) or 0) - own_overhead(prefix)
            elif lab == "EGI":
                e["E"] = 0.0
            else:
                pm = pctm.get(lab, {})
                if lab in ("Ampol Customer", "Z Customer"):
                    s = 0.0
                    for rr, lab2 in tb.grid:
                        if cust_side.get(lab2, lab) != lab:
                            continue
                        s += (num(tb.wv.cell(rr, 19).value) or 0) \
                            * pm.get(norm(lab2), 0.0)
                    e["E"] = s
                else:
                    s = 0.0
                    for rr, lab2 in tb.grid:
                        s += (num(tb.wv.cell(rr, 19).value) or 0) \
                            * pm.get(norm(lab2), 0.0)
                    e["E"] = s
            # the country weights each component splits on, mirroring the
            # tab's own footnote: support cost splits inside each squad, the
            # pots split across their own people, other overheads across the
            # row's own overhead people
            people = e.get("people") or []
            ovhp = [r for r in people if ovh_of.get(r) in ovh_labels
                    and ovh_of.get(r) != e.get("potline")]
            if prefix in ("2.12", "2.13", "2.11"):
                e["w_E"] = au_weight([r for r in people
                                      if ovh_of.get(r) not in ovh_labels])
            elif lab == "EGI":
                e["w_E"] = None
            else:
                pm = pctm.get(lab, {})
                a_num = den = 0.0
                for rr, lab2 in tb.grid:
                    if lab in ("Ampol Customer", "Z Customer") \
                            and cust_side.get(lab2, lab) != lab:
                        continue
                    pct = pm.get(norm(lab2), 0.0)
                    if not pct:
                        continue
                    sect = tb.sections.get(lab2)
                    if not sect:
                        continue
                    sa, sn = au_nz([ro["ledger"] for ro in sect
                                    if ro["ledger"] is not None])
                    a_num += sa * pct
                    den += (sa + sn) * pct
                e["w_E"] = (a_num / den) if den else None
            if lab in ("Ampol Customer", "Z Customer"):
                pool = [ro["ledger"] for ro in tb.roles
                        if ro["ledger"] is not None
                        and ovh_of.get(ro["ledger"]) in ovh_labels]
                e["w_I"] = au_weight(pool)
            else:
                e["w_I"] = au_weight(ovhp)
                if e["w_I"] is None:
                    # no overhead people left once the pot is out: fall back
                    # to the whole overhead group, then to the row itself
                    allo = [r for r in people if ovh_of.get(r) in ovh_labels]
                    e["w_I"] = au_weight(allo) or au_weight(people)
                    if (e.get("I") or 0) > TOL:
                        e["w_I_fallback"] = True
            exp[lab] = e

        # the shares, divided by eleven
        unit = {"F": (bp_pot or 0) / SHARE_BASE,
                "G": (da_pot or 0) / SHARE_BASE,
                "H": (gm_cost or 0) / SHARE_BASE}
        ea = exp.get("Ampol Customer", {}).get("E")
        ez = exp.get("Z Customer", {}).get("E")
        frac = None
        if ea is not None and ez is not None and (ea + ez) > 0:
            frac = ea / (ea + ez)
        # the Customer pair, finished on the two Support Costs: the shared
        # squads, the shared funded slice and the whole overhead pool all
        # divide on the same weight the shares use
        for lab, w in (("Ampol Customer", frac),
                       ("Z Customer", None if frac is None else 1 - frac)):
            e = exp.get(lab)
            if e is None or w is None or "side_S" not in e:
                continue
            e["C"] = e["side_S"] + (e["shared_S"] + e["shared_co"]) * w
            e["D"] = e["side_P"] + e["shared_P"] * w
            e["I"] = e["pool_S"] * w
            e["w_split"] = w
        for lab in LROWS:
            e = exp.get(lab)
            if e is None:
                continue
            if lab in NOSHARE:
                e["F"] = e["G"] = e["H"] = 0.0
            elif lab in ("Ampol Customer", "Z Customer"):
                sh = frac if lab == "Ampol Customer" else (
                    None if frac is None else 1 - frac)
                for k in ("F", "G", "H"):
                    e[k] = None if sh is None else unit[k] * sh
            else:
                for k in ("F", "G", "H"):
                    e[k] = unit[k]
        # I split for the Customer pair comes from the people themselves; the
        # pro-rata reading is checked separately in group H
        for lab in LROWS:
            e = exp.get(lab)
            if e is None:
                continue
            b = cfg_lookup(lab)
            e["M"] = cfg_val.get(b) if b else None
            e["M_au"] = cfg_au.get(b) if b else None
            e["M_nz"] = cfg_nz.get(b) if b else None
            e["budrow"] = b
    except Exception as e:
        model_note.append("model build aborted: %r" % e)

    if "--model" in rest:
        print("\nthe independent model, row by row ($m)", flush=True)
        print("        %-30s %10s %8s %9s %8s %8s %8s %9s"
              % ("row", "C", "D", "E", "F", "G", "H", "I"), flush=True)
        for lab in LROWS:
            e = exp.get(lab, {})
            print("        %-30s %s" % (lab, " ".join(
                ("%10.6f" % e[k]) if isinstance(e.get(k), (int, float))
                else "%10s" % "-" for k in ("C", "D", "E", "F", "G", "H",
                                            "I"))), flush=True)
        print("        %-30s %10.6f" % ("C total",
                                        sum(e.get("C") or 0
                                            for e in exp.values())),
              flush=True)
        for t in model_note:
            print("        note: %s" % t, flush=True)

    # ======================================================== F lights on 3.5
    g = gate.group("F", "lights on 3.5")
    try:
        if lo is None:
            gate.sub(g, False, "tab %r missing" % LO)
        else:
            i34 = sheets.index(tabname("3.4")) if tabname("3.4") else -2
            gate.sub(g, sheets.index(LO) == i34 + 1,
                     "positioned directly after 3.4 (index %d against %d)"
                     % (sheets.index(LO), i34))
            gate.sub(g, f[LO].sheet_state == "visible",
                     "tab visible (%s)" % f[LO].sheet_state)
            gate.sub(g, lo.hdr_r is not None, "header row found")
            bad_h = []
            for c, want in sorted(HEADERS.items()):
                got = txt(lo.wf.cell(lo.hdr_r or 1, c).value)
                if got != want:
                    bad_h.append((openpyxl.utils.get_column_letter(c), got,
                                  want))
            for c in BLANKCOLS:
                if txt(lo.wf.cell(lo.hdr_r or 1, c).value):
                    bad_h.append((openpyxl.utils.get_column_letter(c),
                                  "not blank", ""))
            gate.sub(g, not bad_h,
                     "his 16 headers verbatim with N and P blank: %d off %s"
                     % (len(bad_h), bad_h[:3]))
            gate.sub(g, lo.labels == LROWS,
                     "the 18 rows in the 0.2 config order: %s"
                     % ("exact" if lo.labels == LROWS
                        else "got %s" % lo.labels[:19]))
            gate.sub(g, "Legal" not in lo.labels, "no Legal row")
            gate.sub(g, lo.total_r is not None, "Total row present")
            gate.sub(g, lo.budget_r is not None, "Budget row present")
            tdd_data = []
            for ws in (lo.wf, lo2.wf if lo2 else None, cfgf):
                if ws is None:
                    continue
                for row in ws.iter_rows():
                    for cl in row:
                        if isinstance(cl.value, str) \
                                and re.search(r"\bTDD Data\b", cl.value):
                            tdd_data.append("%s!%s" % (ws.title,
                                                       cl.coordinate))
            gate.sub(g, not tdd_data,
                     "the Enterprise Data label everywhere, never TDD Data "
                     "(0.2 B22 included): %d off %s"
                     % (len(tdd_data), tdd_data[:4]))
            # toggles: derived, never counted - a row carries one when it has
            # own overheads to scale or shares to carry, and a row with
            # nothing to scale carries none rather than a dead cream input
            need = [lab for lab in lo.labels
                    if (exp.get(lab, {}).get("I") or 0) > TOL
                    or lab not in NOSHARE]
            notog = [lab for lab in need
                     if lo.get(lab, "J") != 1]
            gate.sub(g, not notog,
                     "the toggle sits at the 100%% default on the %d rows "
                     "that carry one: %d off %s"
                     % (len(need), len(notog), notog[:4]))
            dead = [lab for lab in lo.labels if lab not in need
                    and txt(lo.wf.cell(lo.rows[lab], COL["J"]).value)]
            gate.sub(g, not dead,
                     "no toggle on the %d rows with nothing to scale: %d off "
                     "%s" % (len(lo.labels) - len(need), len(dead), dead[:4]))
            uncream = [lab for lab in need
                       if rgb_of(lo.wf.cell(lo.rows[lab], COL["J"])) != CREAM]
            gate.sub(g, not uncream, "toggle cells cream: %d off %s"
                     % (len(uncream), uncream[:4]))
            cover, entries = set(), None
            for dv in lo.wf.data_validations.dataValidation:
                if dv.type != "list":
                    continue
                f1 = str(dv.formula1 or "")
                for cr in dv.sqref.ranges:
                    if cr.min_col <= COL["J"] <= cr.max_col:
                        cover |= set(range(cr.min_row, cr.max_row + 1))
                        if f1.startswith('"'):
                            entries = len(f1.strip('"').split(","))
            missing_dv = [lab for lab in need if lo.rows[lab] not in cover]
            gate.sub(g, not missing_dv,
                     "a 0-100%% list validation covers every toggle: %d off %s"
                     % (len(missing_dv), missing_dv[:4]))
            gate.sub(g, entries in (None, 21),
                     "the toggle list runs 0%%,5%%,..,100%% (21 entries, got "
                     "%s)" % entries)
            # the analysis block and the two lines he asked for
            blob = " ".join(txt(cl.value) for row in lo.wf.iter_rows()
                            for cl in row if isinstance(cl.value, str))
            gate.sub(g, "dial" in blob.lower(),
                     "the analysis block carried over (the dials)")
            gate.sub(g, re.search(r"vacant overhead", blob, re.I) is not None,
                     "a line naming the biggest vacant-overhead dial")
            gate.sub(g, "EGI" in blob and re.search(
                r"EGI[^.]{0,120}(exclud|outside|funded)", blob, re.I)
                is not None, "the EGI exclusion note on the tab")
            gate.sub(g, re.search(r"1\.x", blob) is not None,
                     "the footnote naming the 1.x basis for the R column")
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ======================================================== G recompute 3.5
    g = gate.group("G", "recompute 3.5")
    try:
        for t in model_note:
            gate.note(g, t)
        if lo is None:
            gate.sub(g, False, "tab %r missing" % LO)
        else:
            gate.sub(g, bp_pot is not None and len(line_rows(
                "Business Partner")) > 0,
                "the BP pot re-derived from the %d Business Partner roles: "
                "%.6f" % (len(line_rows("Business Partner")), bp_pot or 0))
            gate.sub(g, da_pot is not None and len(line_rows(
                "Domain Architect")) > 0,
                "the DA pot re-derived from the %d Domain Architect roles: "
                "%.6f" % (len(line_rows("Domain Architect")), da_pot or 0))
            gate.sub(g, gm_cost is not None,
                     "the GM cost off Lists: %s" % gm_cost)
            bad = []
            for lab in LROWS:
                if lab not in lo.rows:
                    bad.append("%s: no such row on the tab" % lab)
                    continue
                if lab not in exp:
                    bad.append("%s: the model could not price it" % lab)
                    continue
                e = exp[lab]
                tog = lo.get(lab, "J")
                tog = 1.0 if tog is None else tog
                if e.get("I") is not None:
                    e["K"] = ((e.get("F") or 0) + (e.get("G") or 0)
                              + (e.get("H") or 0) + e["I"] * tog)
                    if e.get("E") is not None:
                        e["L"] = e["E"] + e["K"]
                        if e.get("M") is not None:
                            e["O"] = e["L"] - e["M"]
                        if e.get("C") is not None and e.get("D") is not None:
                            e["Q"] = e["C"] - e["D"] - e["L"]
                            rr = lo.get(lab, "R")
                            if rr is not None:
                                e["S"] = e["Q"] - rr
                for k in ("C", "D", "E", "F", "G", "H", "I", "K", "L", "M",
                          "O", "Q", "S"):
                    want = e.get(k)
                    if want is None:
                        continue
                    got = lo.get(lab, k)
                    if got is None or abs(got - want) > TOL:
                        bad.append("%s!%s got %s want %.6f"
                                   % (lab, k, got, want))
            gate.sub(g, not bad,
                     "every cell of C D E F G H I K L M O Q S rebuilt from "
                     "REVIEW + 2.x + 1.x + 0.2 + Lists at 1e-6: %d off %s"
                     % (len(bad), bad[:5]))
            # totals
            if lo.total_r:
                tbad = []
                for k, c in COL.items():
                    if k == "J":
                        continue
                    got = num(lo.wv.cell(lo.total_r, c).value)
                    want = sum(lo.get(lab, k) or 0 for lab in lo.labels)
                    if got is None or abs(got - want) > TOL:
                        tbad.append("%s got %s want %.6f" % (k, got, want))
                gate.sub(g, not tbad, "the Total row sums its parts: %d off %s"
                         % (len(tbad), tbad[:4]))
                nobud = [lab for lab in LROWS if exp.get(lab, {}).get(
                    "budrow") is None]
                gate.sub(g, not nobud,
                         "every row finds its 0.2 budget line: %d without %s"
                         % (len(nobud), nobud[:4]))
                mt = num(lo.wv.cell(lo.total_r, COL["M"]).value)
                alloc = num(cfgv.cell(cfg_total_r, 5).value) \
                    if cfg_total_r else None
                gate.sub(g, near(mt, alloc),
                         "the M total %s ties to the 0.2 allocated total %s"
                         % (mt, alloc))
            if lo.budget_r:
                lb = num(lo.wv.cell(lo.budget_r, COL["M"]).value)
                e27 = num(cfgv.cell(cfg_budget_r, 5).value) \
                    if cfg_budget_r else None
                gate.sub(g, near(lb, L_BUDGET) and near(lb, e27),
                         "the budget row reads %s against 0.2 %s and his %.2f"
                         % (lb, e27, L_BUDGET))
                ob = num(lo.wv.cell(lo.budget_r, COL["O"]).value)
                lt = num(lo.wv.cell(lo.total_r, COL["L"]).value) \
                    if lo.total_r else None
                gate.sub(g, near(ob, (lt or 0) - (lb or 0)),
                         "the budget row shows O against the 53.80: %s "
                         "against %s" % (ob, None if lt is None else
                                         round(lt - (lb or 0), 6)))
            # R ties to the 1.x tab it names
            rbad = []
            for lab in LROWS:
                if lab not in lo.rows:
                    continue
                got = lo.get(lab, "R")
                p1 = SRC1X.get(lab)
                if p1 is None:
                    # no 1.x tab: the COE line that hosts a pot books the pot
                    # itself, so its Q nets to nothing; every other such row
                    # books nothing
                    e = exp.get(lab, {})
                    want = e.get("pot") or 0 if e.get("holds_pot") else 0
                    if not near(got, want):
                        rbad.append("%s R %s against %.6f (%s)"
                                    % (lab, got, want,
                                       "the pot this line holds"
                                       if e.get("holds_pot") else "no 1.x tab"))
                    continue
                t1 = tabname(p1)
                vals = [x for _, x in funding_amounts(f[t1], v[t1])] if t1 \
                    else []
                pool = set()
                for x in vals:
                    pool.add(round(x, 6))
                    for y in vals:
                        pool.add(round(x + y, 6))
                if lab in ("Ampol Customer", "Z Customer"):
                    continue
                if got is None or (round(got, 6) not in pool
                                   and abs(got) > TOL):
                    rbad.append("%s R %s not a %s funding amount" % (lab, got,
                                                                     p1))
            gate.sub(g, not rbad,
                     "every R ties to an amount on the 1.x tab it draws: %d "
                     "off %s" % (len(rbad), rbad[:4]))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ================================================================ H shares
    g = gate.group("H", "shares and splits")
    try:
        if lo is None:
            gate.sub(g, False, "tab %r missing" % LO)
        else:
            unit = {"F": (bp_pot or 0) / SHARE_BASE,
                    "G": (da_pot or 0) / SHARE_BASE,
                    "H": (gm_cost or 0) / SHARE_BASE}
            offs = []
            for lab in FULLSHARE:
                for k in ("F", "G", "H"):
                    got = lo.get(lab, k)
                    if not near(got, unit[k]):
                        offs.append("%s!%s got %s want %.6f"
                                    % (lab, k, got, unit[k]))
            gate.sub(g, not offs,
                     "the pots divided by ELEVEN on the ten portfolios and "
                     "TDD Cyber: %d off %s" % (len(offs), offs[:4]))
            tdc = [lo.get("TDD Cyber", k) for k in ("F", "G", "H")]
            gate.sub(g, all(near(x, unit[k]) for x, k in
                            zip(tdc, ("F", "G", "H"))),
                     "TDD Cyber carries its overhead share (%s)" % tdc)
            zero = []
            for lab in NOSHARE:
                for k in ("F", "G", "H"):
                    x = lo.get(lab, k)
                    if x is None or abs(x) > TOL:
                        zero.append("%s!%s=%s" % (lab, k, x))
            gate.sub(g, not zero,
                     "the COEs and EGI carry no share: %d off %s"
                     % (len(zero), zero[:4]))
            for k, pot in (("F", bp_pot), ("G", da_pot), ("H", gm_cost)):
                tot = sum(lo.get(lab, k) or 0 for lab in lo.labels)
                gate.sub(g, near(tot, pot),
                         "the %s column totals the pot: %.6f against %.6f"
                         % (k, tot, pot or 0))
            # the Customer pair
            ea, ez = lo.get("Ampol Customer", "E"), lo.get("Z Customer", "E")
            tot_e = (ea or 0) + (ez or 0)
            for k in ("F", "G", "H"):
                a, z = lo.get("Ampol Customer", k), lo.get("Z Customer", k)
                gate.sub(g, near((a or 0) + (z or 0), unit[k]),
                         "the Customer pair shares one %s unit: %.6f against "
                         "%.6f" % (k, (a or 0) + (z or 0), unit[k]))
                if tot_e > 0 and a is not None:
                    gate.sub(g, near(a, unit[k] * ea / tot_e, 1e-6),
                             "the %s unit splits pro rata Support Cost: %s "
                             "against %.6f" % (k, a, unit[k] * ea / tot_e))
            ia, iz = lo.get("Ampol Customer", "I"), lo.get("Z Customer", "I")
            pool = own_overhead("2.2")
            gate.sub(g, near((ia or 0) + (iz or 0), pool),
                     "the Customer overhead pool splits across the pair: "
                     "%.6f against 2.2's %s" % ((ia or 0) + (iz or 0), pool))
            if tot_e > 0 and pool is not None and ia is not None:
                gate.sub(g, near(ia, pool * ea / tot_e, 1e-6),
                         "the Customer overhead pool splits pro rata Support "
                         "Cost: %s against %.6f" % (ia, pool * ea / tot_e))
            # every split pair sums to its parent tab
            for prefix, (a, b) in SPLITS.items():
                t = tabname(prefix)
                if t is None or t not in T:
                    continue
                want = T[t].total(19)
                got = (lo.get(a, "C") or 0) + (lo.get(b, "C") or 0)
                gate.sub(g, near(got, want),
                         "%s and %s together carry all of %s: %.6f against %s"
                         % (a, b, t, got, want))
            # the COE pairs against the tab's own planned spend
            for prefix, (a, b) in SPLITS.items():
                if prefix == "2.2":
                    continue
                t = tabname(prefix)
                if t is None or t not in T:
                    continue
                names = (("Business Partnering planned spend",
                          "Transformation planned spend")
                         if prefix == "2.12" else
                         ("Strategy & Architecture planned spend",
                          "Data planned spend"))
                for lab, nm in zip((a, b), names):
                    want, _ = T[t].label_value(nm)
                    got = (lo.get(lab, "E") or 0) + (lo.get(lab, "I") or 0)
                    gate.sub(g, near(got, want),
                             "%s: Support Cost plus own overheads equals %s "
                             "on %s (%.6f against %s)"
                             % (lab, nm, t, got, want))
                pot = bp_pot if prefix == "2.12" else da_pot
                tot, _ = T[t].label_value("Total planned spend")
                gate.sub(g, near(tot, (T[t].total(19) or 0) - (pot or 0)),
                         "%s planned spend is its total net of the pot: %s "
                         "against %.6f"
                         % (t, tot, (T[t].total(19) or 0) - (pot or 0)))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # =================================================================== I EGI
    g = gate.group("I", "EGI exclusion")
    try:
        egi_squads = {k for k in funded if norm(k).startswith("egi")}
        egi_rows = [r for r in led
                    if norm(canon_squad(squad_of.get(r, ""),
                                        "%s | %s" % (name_of.get(r),
                                                     title_of.get(r))))
                    in {norm(x) for x in egi_squads}
                    or norm(home_led.get(r) or mtab_of.get(r)) == "egi"]
        stray = [(r, name_of[r], role_by_led[r][0]) for r in egi_rows
                 if r in role_by_led
                 and not role_by_led[r][0].startswith("2.14")]
        gate.sub(g, not stray,
                 "every EGI person sits on the EGI tab: %d elsewhere %s"
                 % (len(stray), stray[:4]))
        potnames = [r for r in line_rows("Business Partner")
                    + line_rows("Domain Architect") if r in egi_rows]
        gate.sub(g, not potnames,
                 "no EGI person inside the BP or DA pots: %d %s"
                 % (len(potnames), [name_of[r] for r in potnames[:4]]))
        if lo is not None and "EGI" in lo.rows:
            nz = []
            for k in ("E", "F", "G", "H", "I", "K", "L"):
                x = lo.get("EGI", k)
                if x is None or abs(x) > TOL:
                    nz.append("%s=%s" % (k, x))
            gate.sub(g, not nz,
                     "the EGI row is zero across E F G H I K L: %d off %s"
                     % (len(nz), nz[:5]))
            c_, d_ = lo.get("EGI", "C"), lo.get("EGI", "D")
            gate.sub(g, near(c_, d_),
                     "the EGI row books its whole cost as funded (C %s "
                     "against D %s)" % (c_, d_))
            q_ = lo.get("EGI", "Q")
            gate.sub(g, near(q_, 0.0),
                     "nothing left to recharge on the EGI row (Q %s)" % q_)
            egi_tog = lo.wf.cell(lo.rows["EGI"], COL["J"]).value
            gate.sub(g, egi_tog in (None, ""), "the EGI toggle stays blank "
                     "(%r)" % egi_tog)
        else:
            gate.sub(g, False, "EGI row on the Lights On tab")
        # funded squads never priced into a support cost
        bad = []
        for lab, p1 in SRC1X.items():
            t1, t2 = tabname(p1), tabname(SRCTAB[lab])
            if not (t1 and t2 and t2 in T):
                continue
            pm, _ = pct_map(f[t1], v[t1])
            for rr, lab2 in T[t2].grid:
                if lab2 in funded and pm.get(norm(lab2), 0.0) > 0:
                    bad.append("%s %s at %s%%" % (t2, lab2,
                                                  pm[norm(lab2)] * 100))
        gate.sub(g, not bad,
                 "no funded squad carries a support percentage: %d off %s"
                 % (len(bad), bad[:4]))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ======================================================== J reconciliation
    g = gate.group("J", "reconciliation")
    try:
        # the role basis: effective cost times the carried lever factor, the
        # WIPRO roles at their vendor rate, gross of the 2.11 part-charges
        tabsum = sum(x or 0 for x in tab_after.values())
        gate.sub(g, uplift_charge is not None,
                 "the cyber uplift part-charge reads off 1.14: %s"
                 % uplift_charge)
        gate.sub(g, near(role_basis_total, tabsum + (uplift_charge or 0)),
                 "REVIEW effective cost times the carried levers against the "
                 "%d tab totals plus the cyber uplift charge: %.6f against "
                 "%.6f plus %s" % (len(tabs2), role_basis_total, tabsum,
                                   uplift_charge))
        gate.sub(g, near(all_after, tabsum, 1e-6),
                 "every 2.x tab total sums to its own role rows: %.6f against "
                 "%.6f" % (tabsum, all_after))
        if lo is not None and lo.total_r:
            ctot = num(lo.wv.cell(lo.total_r, COL["C"]).value)
            gate.sub(g, near(ctot, role_basis_total),
                     "the C column carries every single cost, the part-charged "
                     "slice included: %s against the role basis %.6f"
                     % (ctot, role_basis_total))
            whole_tdd = (ctot or 0) + (gm_cost or 0)
            gate.sub(g, near(whole_tdd, role_basis_total + (gm_cost or 0)),
                     "C total plus the GM layer is the whole of TDD: %.6f "
                     "against %.6f" % (whole_tdd,
                                       role_basis_total + (gm_cost or 0)))
            t31 = tabname("3.1")
            if t31:
                w31 = v[t31]
                lens = None
                for r in range(4, w31.max_row + 1):
                    if txt(w31.cell(r, 2).value).startswith(
                            "Total TDD cost including the GM"):
                        for c in range(3, w31.max_column + 1):
                            if txt(w31.cell(4, c).value).startswith(
                                    "Cost after levers"):
                                lens = num(w31.cell(r, c).value)
                # 3.1's after-levers column is a different, older lens: net of
                # the two charged-out pots and of the uplift charge, plus GM
                want = (whole_tdd - (bp_pot or 0) - (da_pot or 0)
                        - (uplift_charge or 0))
                gate.sub(g, near(lens, want),
                         "3.1's after-levers total against its own lens "
                         "(whole of TDD net of the BP pot, the DA pot and the "
                         "uplift charge): %s against %.6f" % (lens, want))
                gate.note(g, "3.1 reads %s on the netted lens; the whole of "
                          "TDD is %.6f - the two are different views, not a "
                          "disagreement" % (lens, whole_tdd))
            # Q and S per row
            qbad = []
            for lab in lo.labels:
                c_, d_, l_, q_ = (lo.get(lab, "C"), lo.get(lab, "D"),
                                  lo.get(lab, "L"), lo.get(lab, "Q"))
                if None not in (c_, d_, l_, q_) \
                        and abs(q_ - (c_ - d_ - l_)) > TOL:
                    qbad.append("%s Q %s against C-D-L %.6f"
                                % (lab, q_, c_ - d_ - l_))
                r_, s_ = lo.get(lab, "R"), lo.get(lab, "S")
                if None not in (q_, r_, s_) and abs(s_ - (q_ - r_)) > TOL:
                    qbad.append("%s S %s against Q-R %.6f" % (lab, s_,
                                                              q_ - r_))
            gate.sub(g, not qbad,
                     "Q equals C less D less L and S equals Q less R on every "
                     "row: %d off %s" % (len(qbad), qbad[:4]))
            # the white control
            found = []
            for row in lo.wf.iter_rows():
                for cl in row:
                    if isinstance(cl.value, str) \
                            and cl.value.startswith("Control"):
                        val = None
                        for cc in range(cl.column + 1, lo.wf.max_column + 1):
                            x = num(lo.wv.cell(cl.row, cc).value)
                            if x is not None:
                                val = x
                                break
                        col = cl.font.color
                        found.append((cl.coordinate, val,
                                      col is not None and col.rgb == WHITE))
            gate.sub(g, bool(found) and all(
                x is not None and abs(x) <= TOL and w for _, x, w in found),
                "the white reconciliation control reads 0: %s" % found[:3])
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ======================================================== K lights on 3.6
    g = gate.group("K", "lights on 3.6")
    try:
        if lo2 is None:
            gate.sub(g, False, "tab %r missing" % LO2)
        elif lo is None:
            gate.sub(g, False, "3.5 missing, nothing to duplicate against")
        elif lo2.hdr_r is None:
            gate.sub(g, False, "no header row on %s" % LO2)
        else:
            gate.sub(g, sheets.index(LO2) == sheets.index(LO) + 1,
                     "positioned directly after 3.5")
            gate.sub(g, lo2.labels == LROWS,
                     "the same 18 rows: %s"
                     % ("exact" if lo2.labels == LROWS
                        else "got %s" % lo2.labels[:19]))
            same = []
            for lab in LROWS:
                for k in ("C", "D", "E", "F", "G", "H", "I", "J", "K", "L"):
                    a = lo.get(lab, k)
                    b = lo2.get(lab, k)
                    if (a is None) != (b is None) or (
                            a is not None and abs(a - b) > 1e-9):
                        same.append("%s!%s %s against %s" % (lab, k, b, a))
            gate.sub(g, not same,
                     "columns C to L read identically to 3.5: %d off %s"
                     % (len(same), same[:4]))
            # the AU spend, NZ spend, Total and Variance block sits right of
            # the 'Total portfolio cost charged to TDD' column
            after_l = lo2.cols.get(HEADERS[12], 0)
            au_c = nz_c = tot_c = var_c = None
            for c in range(after_l + 1, lo2.wf.max_column + 1):
                k = norm(lo2.wf.cell(lo2.hdr_r, c).value)
                if not k:
                    continue
                if au_c is None and re.match(r"^au\b", k):
                    au_c = c
                elif nz_c is None and re.match(r"^nz\b", k):
                    nz_c = c
                elif tot_c is None and re.match(r"^total\b", k):
                    tot_c = c
                elif var_c is None and re.match(r"^variance\b", k):
                    var_c = c
            gate.sub(g, None not in (au_c, nz_c, tot_c, var_c),
                     "the AU spend, NZ spend, Total and Variance columns are "
                     "there (%s %s %s %s)" % (au_c, nz_c, tot_c, var_c))
            if None not in (au_c, nz_c, tot_c):
                bad = []
                for lab in LROWS:
                    if lab not in lo2.rows:
                        continue
                    r = lo2.rows[lab]
                    a = num(lo2.wv.cell(r, au_c).value)
                    n = num(lo2.wv.cell(r, nz_c).value)
                    t_ = num(lo2.wv.cell(r, tot_c).value)
                    l_ = lo.get(lab, "L")
                    if None in (a, n, t_):
                        bad.append("%s missing AU/NZ/Total" % lab)
                        continue
                    if abs(a + n - t_) > TOL:
                        bad.append("%s AU+NZ %.6f against Total %.6f"
                                   % (lab, a + n, t_))
                    if l_ is not None and abs(t_ - l_) > TOL:
                        bad.append("%s Total %.6f against 3.5 L %.6f"
                                   % (lab, t_, l_))
                gate.sub(g, not bad,
                         "AU plus NZ equals Total equals 3.5's L on every "
                         "row: %d off %s" % (len(bad), bad[:4]))
                # the AU spend rebuilt component by component off Country, the
                # way the tab's own footnote describes it: support cost inside
                # each squad, the pots across their own people, other
                # overheads across the row's overhead people, and the GM layer
                # on the AU share of the whole role mapping
                w_gm = au_weight(list(role_basis), role_basis)
                w_bp = au_weight(line_rows("Business Partner"))
                w_da = au_weight(line_rows("Domain Architect"))
                gate.sub(g, w_gm is not None and w_bp is not None
                         and w_da is not None,
                         "the AU weights of the whole mapping and of the two "
                         "pots derive: GM %s, BP %s, DA %s"
                         % tuple(None if x is None else round(x, 6)
                                 for x in (w_gm, w_bp, w_da)))
                offs = []
                for lab in LROWS:
                    if lab not in lo2.rows:
                        continue
                    e = exp.get(lab, {})
                    comp = {}
                    for k in ("E", "F", "G", "H", "I", "J"):
                        comp[k] = lo.get(lab, k) or 0
                    want = ((comp["E"] * (e.get("w_E") or 0))
                            + comp["F"] * (w_bp or 0)
                            + comp["G"] * (w_da or 0)
                            + comp["H"] * (w_gm or 0)
                            + comp["I"] * comp["J"] * (e.get("w_I") or 0))
                    got = num(lo2.wv.cell(lo2.rows[lab], au_c).value)
                    if got is None or abs(got - want) > TOL:
                        offs.append("%s AU %s against %.6f (E %.4f x %s, I "
                                    "%.4f x %s, shares on %s/%s/%s)"
                                    % (lab, got, want, comp["E"],
                                       None if e.get("w_E") is None
                                       else round(e["w_E"], 4), comp["I"],
                                       None if e.get("w_I") is None
                                       else round(e["w_I"], 4),
                                       round(w_bp or 0, 3),
                                       round(w_da or 0, 3),
                                       round(w_gm or 0, 3)))
                gate.sub(g, not offs,
                         "every row's AU spend rebuilt from Country person by "
                         "person at 1e-6: %d off %s" % (len(offs), offs[:3]))
                fb = [lab for lab in LROWS
                      if exp.get(lab, {}).get("w_I_fallback")]
                if fb:
                    gate.note(g, "rows booking other overheads with no "
                              "overhead people left once the pot is out, "
                              "split on the whole overhead group instead: %s"
                              % fb)
            if var_c and None not in (tot_c,):
                bad = []
                for lab in LROWS:
                    if lab not in lo2.rows:
                        continue
                    r = lo2.rows[lab]
                    t_ = num(lo2.wv.cell(r, tot_c).value)
                    var = num(lo2.wv.cell(r, var_c).value)
                    b = exp.get(lab, {})
                    bud = (b.get("M_au") or 0) + (b.get("M_nz") or 0)
                    if None in (t_, var):
                        bad.append("%s missing Total or Variance" % lab)
                    elif abs(var - (t_ - bud)) > TOL:
                        bad.append("%s Variance %.6f against Total less the "
                                   "0.2 AU and NZ budgets %.6f"
                                   % (lab, var, t_ - bud))
                gate.sub(g, not bad,
                         "Variance is Total less the 0.2 AU plus NZ budgets: "
                         "%d off %s" % (len(bad), bad[:4]))
            blob = " ".join(txt(cl.value) for row in lo2.wf.iter_rows()
                            for cl in row if isinstance(cl.value, str))
            gate.sub(g, re.search(r"(country|AU|NZ)", blob) is not None
                     and len(blob) > 200,
                     "a footnote explaining the split basis")
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ============================================================ L protection
    g = gate.group("L", "protection")
    try:
        from openpyxl.utils.protection import hash_password
        want_hash = hash_password(PASSWORD).upper()
        must_lock = [ws.title for ws in f.worksheets
                     if re.match(r"^[03]\.\d", ws.title)]
        must_free = [ws.title for ws in f.worksheets
                     if ws.title == REVIEW or ws.title == "Lists"
                     or ws.title.startswith("Exec")
                     or re.match(r"^[12]\.\d", ws.title)]
        notlocked = [t for t in must_lock if not f[t].protection.sheet]
        gate.sub(g, not notlocked,
                 "every 0.x and 3.x tab protected (%d): %d off %s"
                 % (len(must_lock), len(notlocked), notlocked[:4]))
        stillock = [t for t in must_free if f[t].protection.sheet]
        gate.sub(g, not stillock,
                 "Exec, Lists, REVIEW, the 1.x and the 2.x tabs unprotected: "
                 "%d still locked %s" % (len(stillock), stillock[:6]))
        badpw = [t for t in must_lock
                 if txt(f[t].protection.password).upper() != want_hash]
        gate.sub(g, not badpw,
                 "the %s hash on every protected tab: %d off %s"
                 % (PASSWORD, len(badpw), badpw[:4]))
        others = [ws.title for ws in f.worksheets
                  if ws.title not in must_lock and ws.title not in must_free
                  and ws.protection.sheet]
        gate.note(g, "protected outside the 0.x/3.x set (4.0 and the dividers "
                  "keep their state): %s" % others)
        stray = []
        n_unlocked = 0
        allow = set()
        for t in (LO, LO2):
            if t in sheets:
                lt = lo if t == LO else lo2
                if lt:
                    for r in lt.rows.values():
                        allow.add((t, r, COL["J"]))
        for t in must_lock:
            ws = f[t]
            for row in ws.iter_rows():
                for cl in row:
                    if cl.protection.locked is not False:
                        continue
                    n_unlocked += 1
                    if (t, cl.row, cl.column) in allow:
                        continue
                    if rgb_of(cl) == CREAM:
                        continue
                    stray.append("%s!%s" % (t, cl.coordinate))
        gate.sub(g, not stray,
                 "on the protected tabs only the toggles and cream inputs are "
                 "unlocked (%d unlocked): %d stray %s"
                 % (n_unlocked, len(stray), stray[:5]))
        for t in (LO, LO2):
            lt = lo if t == LO else lo2
            if t not in sheets or lt is None:
                continue
            locked = [lab for lab, r in lt.rows.items()
                      if f[t].cell(r, COL["J"]).protection.locked is not False]
            need = [lab for lab in lt.labels
                    if lab in lt.rows and lt.get(lab, "J") is not None]
            locked = [lab for lab in locked if lab in need]
            gate.sub(g, not locked,
                     "%s toggle cells stay editable under protection: %d "
                     "locked %s" % (t, len(locked), locked[:4]))
        sec = f.security
        gate.sub(g, bool(sec is not None and sec.lockStructure),
                 "workbook structure still locked")
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # =============================================================== M hygiene
    g = gate.group("M", "hygiene")
    try:
        dash, endash, banned, creamf = [], [], [], []
        for ws in f.worksheets:
            model = ws.title not in SRC_TABS
            for row in ws.iter_rows():
                for cl in row:
                    tv = cl.value
                    isf = isinstance(tv, str) and tv.startswith("=")
                    if rgb_of(cl) == CREAM and isf:
                        creamf.append("%s!%s" % (ws.title, cl.coordinate))
                    if not model or not isinstance(tv, str):
                        continue
                    if tv.strip() in ("-", '="-"'):
                        dash.append("%s!%s" % (ws.title, cl.coordinate))
                    if isf:
                        continue
                    if ws.title == REVIEW and cl.column <= HIS_COLS:
                        continue        # his words, untouchable
                    if "–" in tv or "—" in tv:
                        endash.append("%s!%s" % (ws.title, cl.coordinate))
                    low = tv.lower()
                    for pat in BANNED:
                        if re.search(pat, low):
                            banned.append("%s!%s %r" % (ws.title,
                                                        cl.coordinate,
                                                        tv[:40]))
                            break
        gate.sub(g, not dash, "dash literals on model tabs: %d %s"
                 % (len(dash), dash[:4]))
        gate.sub(g, not creamf, "cream-filled formula cells: %d %s"
                 % (len(creamf), creamf[:4]))
        gate.sub(g, not banned, "banned words: %d %s"
                 % (len(banned), banned[:3]))
        gate.sub(g, not endash, "en and em dashes: %d %s"
                 % (len(endash), endash[:4]))
        dashfmt = []
        for ws in f.worksheets:
            if ws.title in SRC_TABS:
                continue
            for row in ws.iter_rows():
                for cl in row:
                    s = fmt_sections(cl.number_format or "")
                    if len(s) >= 3 and DASHFMT.search(s[2]):
                        dashfmt.append("%s!%s" % (ws.title, cl.coordinate))
        gate.sub(g, not dashfmt, "dash-showing zero formats: %d %s"
                 % (len(dashfmt), dashfmt[:4]))
        # BLD-19: 1.10's dash from 0.1 renders 0.00
        t110 = tabname("1.10")
        bld19 = []
        if t110:
            for row in f[t110].iter_rows():
                for cl in row:
                    val = v[t110][cl.coordinate].value
                    if isinstance(val, str) and val.strip() == "-":
                        bld19.append("%s!%s" % (t110, cl.coordinate))
                    s = fmt_sections(cl.number_format or "")
                    if len(s) >= 3 and DASHFMT.search(s[2]):
                        bld19.append("%s!%s fmt" % (t110, cl.coordinate))
        gate.sub(g, not bld19,
                 "BLD-19: nothing on 1.10 renders as a dash: %d %s"
                 % (len(bld19), bld19[:4]))
        noparen = []
        for t in tabs2:
            tb = T[t]
            if tb.total_r is None:
                continue
            for r in range(7, tb.total_r + 1):
                if num(tb.wv.cell(r, 18).value) is None:
                    continue
                s = fmt_sections(tb.wf.cell(r, 18).number_format or "")
                if len(s) >= 2 and "(" not in s[1]:
                    noparen.append("%s!R%d" % (t, r))
        t31 = tabname("3.1")
        if t31:
            for r in range(5, f[t31].max_row + 1):
                if num(v[t31].cell(r, 8).value) is None:
                    continue
                s = fmt_sections(f[t31].cell(r, 8).number_format or "")
                if len(s) >= 2 and "(" not in s[1]:
                    noparen.append("%s!H%d" % (t31, r))
        for lt in (lo, lo2):
            if lt is None:
                continue
            for lab, r in lt.rows.items():
                for k in ("O", "Q", "S"):
                    c = lt.colof(k)
                    if num(lt.wv.cell(r, c).value) is None:
                        continue
                    s = fmt_sections(lt.wf.cell(r, c).number_format or "")
                    if len(s) < 2 or "(" not in s[1]:
                        noparen.append("%s!%s%d" % (lt.title, k, r))
        gate.sub(g, not noparen,
                 "negative-capable cells carry parens formats (2.x R, 3.1 "
                 "variance, Lights On O Q S): %d off %s"
                 % (len(noparen), noparen[:4]))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ============================================================== N controls
    g = gate.group("N", "controls")
    try:
        ctrl_bad, ctrl_n = [], 0
        for ws in v.worksheets:
            for row in ws.iter_rows():
                for cl in row:
                    if not (isinstance(cl.value, str)
                            and cl.value.startswith("Control - ")):
                        continue
                    ctrl_n += 1
                    for cc in range(cl.column + 1, ws.max_column + 1):
                        x = num(ws.cell(cl.row, cc).value)
                        if x is not None:
                            if abs(x) > TOL:
                                ctrl_bad.append("%s!%s=%s"
                                                % (ws.title, cl.coordinate, x))
                            break
        gate.sub(g, ctrl_n > 10 and not ctrl_bad,
                 "every control row 0 (%d found): %d off %s"
                 % (ctrl_n, len(ctrl_bad), ctrl_bad[:4]))
        if "4.0 Data QA" in sheets:
            q = v["4.0 Data QA"]
            qa_n, qa_bad, qa_meta = 0, [], None
            for r in range(4, q.max_row + 1):
                x = num(q.cell(r, 5).value)
                if x is None:
                    continue
                if txt(q.cell(r, 2).value) == "Checks failing":
                    qa_meta = x
                    continue
                if q.cell(r, 3).value is None and q.cell(r, 4).value is None:
                    continue
                qa_n += 1
                if abs(x) > TOL:
                    qa_bad.append((r, txt(q.cell(r, 2).value)[:48], x))
            gate.sub(g, qa_n >= 8 and not qa_bad,
                     "4.0 Data QA all zero (%d checks, rollup %s): %d off %s"
                     % (qa_n, qa_meta, len(qa_bad), qa_bad[:4]))
        # his raw block carries typed #N/A text of his own (A31, C31, S525 in
        # his file); his words are untouchable, so the error scan covers the
        # helper block and every other tab, never A2:AC of the raw block
        errs, hisna = [], []
        for ws in v.worksheets:
            for row in ws.iter_rows():
                for cl in row:
                    if not (isinstance(cl.value, str)
                            and any(e in cl.value for e in ERR)):
                        continue
                    if ws.title == REVIEW and cl.column <= HIS_COLS \
                            and cl.row >= 2:
                        hisna.append(cl.coordinate)
                        continue
                    errs.append((ws.title, cl.coordinate))
        gate.sub(g, not errs, "error cells outside his raw block: %d %s"
                 % (len(errs), errs[:4]))
        gate.note(g, "his own typed #N/A text inside the raw block, left "
                  "verbatim: %d %s" % (len(hisna), hisna[:5]))

        def netting_cell(prefix):
            t = tabname(prefix)
            if t is None:
                return None
            wf2, wv2 = f[t], v[t]
            for r in range(1, wf2.max_row + 1):
                b = txt(wf2.cell(r, 2).value).lower()
                x = num(wv2.cell(r, 3).value)
                if x is None:
                    continue
                if "met by portfolio" in b or "netted out" in b:
                    return x
            return None

        n12, n13 = netting_cell("2.12"), netting_cell("2.13")
        gate.sub(g, near(n12, -(bp_pot or 0)),
                 "2.12 nets the BP pot out: %s against %.6f"
                 % (n12, -(bp_pot or 0)))
        gate.sub(g, near(n13, -(da_pot or 0)),
                 "2.13 nets the DA pot out: %s against %.6f"
                 % (n13, -(da_pot or 0)))
        t15 = tabname("2.15")
        S15 = T[t15].total(19) if t15 in T else None
        want = None if (S15 is None or uplift_fund is None) \
            else S15 - uplift_fund
        b = cfg_lookup("TDD Cyber")
        got = cfg_spend.get(b) if b else None
        gate.sub(g, near(got, want),
                 "0.2 TDD Cyber spend %s against 2.15 after levers less the "
                 "Lists uplift funding (%s less %s)" % (got, S15, uplift_fund))
        f26 = num(cfgv.cell(cfg_total_r, 6).value) if cfg_total_r else None
        colsum = sum(x for x in cfg_spend.values() if x is not None)
        gate.sub(g, near(f26, colsum),
                 "0.2 spend total %s against its column %.6f" % (f26, colsum))
        b22 = None
        for r, lab in cfg_rows:
            if r == 22:
                b22 = lab
        gate.sub(g, b22 is None or "TDD Data" not in b22,
                 "0.2 B22 relabelled off TDD Data (%r)" % b22)
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ------------------------------------------------------------- the ledger
    print("\nthe ledger (print only)", flush=True)

    def p(label, val):
        print("        %-52s %s" % (label, val), flush=True)

    try:
        p("roles in the role mapping", len(led))
        p("total cost after levers, %d tabs ($m)" % len(tabs2),
          "%.6f" % sum(x or 0 for x in tab_after.values()))
        p("the GM layer ($m)", gm_cost)
        p("whole of TDD, people plus GMs ($m)",
          "%.6f" % (sum(x or 0 for x in tab_after.values()) + (gm_cost or 0)))
        for lt, nm in ((lo, LO), (lo2, LO2)):
            if lt is None or lt.total_r is None:
                p("%s totals" % nm, "tab absent")
                continue
            # 3.6 shifts M O Q R S right of its AU NZ block, so every column
            # is resolved by his header, never by 3.5's position
            tot = {k: num(lt.wv.cell(lt.total_r, lt.colof(k)).value)
                   for k in COL}
            p("%s C total at the default toggles ($m)" % nm,
              "%s" % (None if tot["C"] is None else round(tot["C"], 6)))
            p("  L charged to TDD ($m)",
              "%s" % (None if tot["L"] is None else round(tot["L"], 6)))
            p("  M, the 0.2 allocations ($m)",
              "%s" % (None if tot["M"] is None else round(tot["M"], 6)))
            p("  over/under the 0.2 allocations ($m)",
              "%s" % (None if None in (tot["L"], tot["M"])
                      else round(tot["L"] - tot["M"], 6)))
            p("  over/under the %.2f budget ($m)" % L_BUDGET,
              "%s" % (None if tot["L"] is None
                      else round(tot["L"] - L_BUDGET, 6)))
            p("  E support cost ($m)",
              "%s" % (None if tot["E"] is None else round(tot["E"], 6)))
            p("  K overheads charged ($m)",
              "%s" % (None if tot["K"] is None else round(tot["K"], 6)))
            p("  F/G/H shares ($m)",
              "%s / %s / %s" % tuple(None if tot[k] is None
                                     else round(tot[k], 6)
                                     for k in ("F", "G", "H")))
            p("  I own overheads ($m)",
              "%s" % (None if tot["I"] is None else round(tot["I"], 6)))
            p("  D significant items funded ($m)",
              "%s" % (None if tot["D"] is None else round(tot["D"], 6)))
            p("  Q left to recharge ($m)",
              "%s" % (None if tot["Q"] is None else round(tot["Q"], 6)))
            p("  R noted in the 1.x tabs ($m)",
              "%s" % (None if tot["R"] is None else round(tot["R"], 6)))
            p("  S still left to fund ($m)",
              "%s" % (None if tot["S"] is None else round(tot["S"], 6)))
        p("BP pot ($m)", "%.6f" % (bp_pot or 0))
        p("DA pot ($m)", "%.6f" % (da_pot or 0))
        p("one share unit, pots over %d ($m)" % SHARE_BASE,
          "%.6f / %.6f / %.6f" % ((bp_pot or 0) / SHARE_BASE,
                                  (da_pot or 0) / SHARE_BASE,
                                  (gm_cost or 0) / SHARE_BASE))
        vac = collections.Counter()
        vac_cost = collections.defaultdict(float)
        for r in led:
            ro = role_by_led.get(r)
            if not ro:
                continue
            if ovh_of.get(r) in ovh_labels and mstat_of.get(r) == "Vacant" \
                    and ro[1]["lever"] == "Hire":
                vac[ovh_of[r]] += 1
                vac_cost[ovh_of[r]] += ro[1]["G"] or 0
        p("vacant overheads on Hire (count / $m)",
          "%d / %.6f" % (sum(vac.values()), sum(vac_cost.values()) / 1e6))
        for line in sorted(vac):
            p("  %s" % line, "%d / %.6f" % (vac[line],
                                            vac_cost[line] / 1e6))
        p("dial 1: hold the vacant overheads ($m)",
          "%.6f" % (sum(vac_cost.values()) / 1e6))
        p("dial 2: the GM layer above lights on ($m)", gm_cost)
        crso = cfg_lookup("COE Cyber Risk & Service Ops")
        if crso and cfgv and cfg_rows:
            rr = [r for r, b in cfg_rows if b == crso][0]
            gv = num(cfgv.cell(rr, 7).value)
            p("dial 3: COE Cyber against its allocation ($m over)",
              "%s" % (None if gv is None else round(-gv, 6)))
        if cfg_var_r:
            p("dial 4: the unallocated slice of the %.2f ($m)" % L_BUDGET,
              "%s" % num(cfgv.cell(cfg_var_r, 5).value))
        egi_t = tabname("2.14")
        egi_cost = T[egi_t].total(19) if egi_t in T else None
        p("EGI excluded from the overhead engine ($m)", egi_cost)
        if lo is not None:
            p("EGI row C / D / Q ($m)",
              "%s / %s / %s" % (lo.get("EGI", "C"), lo.get("EGI", "D"),
                                lo.get("EGI", "Q")))
        p("part-timers priced at effective cost",
          "%d" % len([r for r in led
                      if fte_of[r] is not None and 0 < fte_of[r] < 1]))
    except Exception as e:
        p("ledger print aborted", repr(e))

    print("", flush=True)
    fails = gate.report()
    return 1 if fails else 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
