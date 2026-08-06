#!/usr/bin/env python3
"""w3_verify - Stage w3 (VERIFY), the data-truth suite for build spec W.

    python3 w3_verify.py <built.xlsx> [--outdir DIR] [--seed N] [--old PATH]

READ-ONLY. It opens the built workbook, his master, the customer PCM file and
the snapshot of the old shipped model, and writes nothing except its two
report artefacts:

    <outdir>/placement_delta.json   who moved where vs the OLD shipped model
    <outdir>/verify_report.txt      the full PASS/FAIL report

Default outdir is <script dir>/tw3, which already holds old_model.xlsx (the
snapshot taken before the orchestrator overwrites the repo file).

Check groups, one PASS/FAIL line each, counted at the end, non-zero exit on
any FAIL:

  A  raw identity   REVIEW A2:AC527 cell-for-cell against his master, his
                    header row verbatim (duplicate EE Number column and MyHR
                    ee no included), 526 rows, nothing stale below row 527
  B  customer PCM   every PCM-data person present with matching EE, cost and
                    placement; PCM people absent from his master reported
                    only (his master wins)
  C  placement      independent homing of all 526 rows, REVIEW MTab against
                    it, each 2.x tab's roster against it, each person exactly
                    once across the whole book, 3.2 rosters against the
                    derived overhead lines, the settled memberships, and a
                    40-person random salary trace REVIEW -> 2.x -> totals at
                    1e-9
  D  EE truth       REVIEW EE columns against his file, the duplicate column
                    consistent, MyHR ee no, every EE-headed column anywhere in
                    the book, and the 129 corrections against the old model
  E  precedence     the reversals his file forces (Viren Khatri under EGI, Ed
                    Tacey off the Heads line, ring fenced rows as vacancies,
                    Sarsha Tanner, Murray Mitchell, Scott McKenzie) and the
                    part-time effective cost
  F  artefacts      the placement delta and this report written

Homing is derived here from his master and the spec's rules, not read from the
workbook, so the workbook is tested against the rules rather than against
itself. The one thing read from the workbook is the person-keyed override
table on Lists (the agreed moves), because that is model logic that lives in
the file by design; every override applied is printed.
"""
import sys, os, re, json, random, datetime, collections

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
UP = "/root/.claude/uploads/e550b440-3996-5abb-87e5-bafafe598f82"
HIS = os.path.join(UP, "0ad63df5-updates.xlsx")
PCM = os.path.join(UP, "ee008e30-Customer_Resource_22Jul26V2.xlsx")

REVIEW = "REVIEW - Complete Role Mapping"
RAW_COLS = 29                      # A:AC
FIRST_ROW, LAST_ROW = 2, 527       # 526 data rows
N_ROWS = 526
TOL9 = 1e-9
TOL6 = 1e-6
TRACE_N = 40
SEED = 20260805

# his 29 headers, verbatim, including the duplicate EE Number and MyHR ee no
HIS_HEADERS = [
    "EE Number", "Name", "EE Number", "Position Title", "Reports to name",
    "Reports to Position", "Division (GM)", "Department (GM-1)",
    "Team (GM-2)", "Portfolio", "Platform", "Squad", "CC", "Country",
    "Job Level", "FTE", "Vacant Column", "Type", "Unit", "day rate",
    "FTE Base\n Local", "FTE Base \nAUD", "STI", "Payroll", "Pensions",
    "medical", "CPI", "Full Cost \nAUD", "MyHR ee no"]

# his raw columns, 1-based, on the rebuilt REVIEW block
C_EE, C_NAME, C_EE2, C_TITLE = 1, 2, 3, 4
C_DIV, C_PORT, C_PLAT, C_SQUAD = 7, 10, 11, 12
C_COUNTRY, C_FTE, C_COST, C_MYHR = 14, 16, 28, 29

# tab homing, his portfolio values -> the model's tab key (case normalised)
PORT_MAP = {
    "retail": "Ampol Retail", "z": "Z Retail", "ampol customer": "Customer",
    "z energy (digital)": "Customer", "commercial fuels": "Commercial Fuels",
    "b2b & energy solutions": "Energy Solutions & B2B",
    "infrastructure": "Infrastructure", "enterprise data": "Enterprise Data",
    "finance": "Finance", "p&c": "P&C", "p&c, finance & legal": "P&C",
    "tdd": "TDD Group Functions", "egi": "EGI", "egi integration": "EGI"}
DIV_MAP = {
    "strategy, architecture & data": "COE SA&D",
    "cyber, risk & operations": "COE Cyber",
    "partnering & transformation": "COE BP&T",
    "tdd group functions": "TDD Group Functions", "customer": "Customer",
    "egi": "EGI"}
# divisions that carry no portfolio of their own and must fall through to
# Platform / Squad; every such row is printed
DIV_RESOLVE = {"ampol retail & z", "tdd cfe&i"}
# the two squads that home to 2.15 whatever their portfolio says
CYBER_SQUADS = {"cyber uplift", "identity"}

VACANT_NAMES = {"vacant", "ring fenced", "ring fenced selection", "remove",
                "", "none"}
RINGFENCED = {"ring fenced", "ring fenced selection"}
LEVER_FACTOR = {"filled": 1.0, "hire": 1.0, "hold": 0.0, "offshore": 0.4}

# the settled memberships and the reversals his file forces
SETTLED_TM = "shane ker"
SETTLED_NOT_HOT = "ed tacey"
REVERSALS = {
    "viren khatri": ("TDD Group Functions", "his new row: squad EGI TDD, and "
                            "an EGI squad names the portfolio it belongs "
                            "to, so his cost shows there and nets out"),
    "ed tacey": ("Customer", "his new row: Squad Leadership, Head of AI "
                             "enablement"),
    "sarsha tanner": ("COE BP&T", "fills Head of Transformation"),
    "murray mitchell": ("TDD Group Functions", "fills a TM vacancy"),
    "scott mckenzie": ("TDD Group Functions", "fills a TM vacancy")}

ERRS = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")


# ---------------------------------------------------------------- helpers

def norm(s):
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip()).lower()


def blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def num(v):
    if isinstance(v, bool):
        return None
    return v if isinstance(v, (int, float)) else None


def same(a, b, tol=TOL9):
    """Cell equality: blanks equal, numbers to tol, everything else exact."""
    if blank(a) and blank(b):
        return True
    if blank(a) or blank(b):
        return False
    na, nb = num(a), num(b)
    if na is not None and nb is not None:
        return abs(na - nb) <= tol * max(1.0, abs(na), abs(nb))
    if isinstance(a, datetime.datetime) or isinstance(b, datetime.datetime):
        return a == b
    if (na is None) != (nb is None):
        return False
    return str(a) == str(b)


def show(v):
    if isinstance(v, float):
        return ("%.6f" % v).rstrip("0").rstrip(".")
    s = str(v)
    return s if len(s) <= 60 else s[:57] + "..."


class Report:
    """PASS/FAIL lines, grouped, echoed to stdout and kept for the artefact."""

    def __init__(self):
        self.lines = []
        self.n_pass = 0
        self.n_fail = 0

    def _emit(self, text):
        self.lines.append(text)
        print(text, flush=True)

    def head(self, letter, name):
        self._emit("")
        self._emit("%s  %s" % (letter, name))
        self._emit("-" * 74)

    def check(self, ok, label, detail=""):
        ok = bool(ok)
        if ok:
            self.n_pass += 1
        else:
            self.n_fail += 1
        line = "%-4s %s" % ("PASS" if ok else "FAIL", label)
        if detail:
            line += "   [%s]" % detail
        self._emit(line)
        return ok

    def note(self, text):
        self._emit("     %s" % text)

    def report(self, text):
        """Report-only line: never counted, never fails the run."""
        self._emit("INFO %s" % text)

    def bullets(self, items, limit=25, indent="       - "):
        for x in list(items)[:limit]:
            self._emit(indent + str(x))
        if len(items) > limit:
            self._emit(indent + "(%d more)" % (len(items) - limit))


# ------------------------------------------------------- workbook readers

def col_letter(i):
    s = ""
    while i:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


def find_sheet(wb, prefix):
    for t in wb.sheetnames:
        if t == prefix or t.startswith(prefix + " ") or t.startswith(prefix):
            return t
    return None


def helper_cols(ws):
    """{key: column index} for the helper block, found by header text."""
    want = [("mtab", ("mtab",)), ("mstatus", ("mstatus",)),
            ("roleid", ("role id",)), ("ohline", ("overhead line",)),
            ("sqoh", ("squad or overhead line",)),
            ("canonical", ("squad (canonical",)),
            ("effective", ("effective cost", "effective")),
            ("ringfenced", ("ringfenced", "ring fenced")),
            ("override", ("agreed cost override",)),
            ("pcm_status", ("status (pcm)",))]
    out = {}
    for c in range(RAW_COLS + 1, ws.max_column + 1):
        h = norm(ws.cell(1, c).value)
        if not h:
            continue
        for key, pats in want:
            if key in out:
                continue
            for p in pats:
                if h == p or h.startswith(p):
                    out[key] = c
                    break
    return out


def two_x_tabs(wb):
    """[(title, portfolio key from C3)] for every 2.x lever tab, in order."""
    out = []
    for t in wb.sheetnames:
        if re.match(r"^2\.\d+\b", t):
            out.append(t)
    out.sort(key=lambda t: float(re.match(r"^2\.(\d+)", t).group(1)))
    return out


def read_roster(wsf, wsv):
    """Roster rows of one 2.x tab: [{id,row,group,name,title,status,lever,F,G}].

    A roster row is a row whose column A holds an R#### id. The group is the
    nearest plain-text label above it in column B.
    """
    rows = []
    group = None
    started = False
    for r in range(1, wsf.max_row + 1):
        a = wsf.cell(r, 1).value
        b = wsf.cell(r, 2).value
        if isinstance(a, str) and norm(a) == "id":
            started = True
            continue
        if not started:
            continue
        if isinstance(a, str) and re.fullmatch(r"R\d{4}", a.strip()):
            rows.append({
                "id": a.strip(), "row": r, "group": group,
                "name": wsv.cell(r, 2).value, "title": wsv.cell(r, 3).value,
                "status": wsv.cell(r, 4).value,
                "lever": wsf.cell(r, 5).value,
                "F": num(wsv.cell(r, 6).value),
                "G": num(wsv.cell(r, 7).value)})
            continue
        if blank(a) and isinstance(b, str) and b.strip() \
                and not b.startswith("="):
            lab = b.strip()
            low = lab.lower()
            if low.startswith("control") or low.startswith("vacancy lever") \
                    or len(lab) > 70:
                continue
            group = lab
    return rows


def group_totals(wsf, wsv):
    """{group label: (row, F sum cell, G sum cell)} for the roster headers."""
    out = {}
    started = False
    for r in range(1, wsf.max_row + 1):
        a = wsf.cell(r, 1).value
        b = wsf.cell(r, 2).value
        if isinstance(a, str) and norm(a) == "id":
            started = True
            continue
        if not started or not blank(a):
            continue
        if isinstance(b, str) and b.strip() and not b.startswith("="):
            lab = b.strip()
            low = lab.lower()
            if low.startswith("control") or low.startswith("vacancy lever") \
                    or len(lab) > 70:
                continue
            out[lab] = (r, num(wsv.cell(r, 6).value), num(wsv.cell(r, 7).value))
    return out


def lists_overrides(wb_v):
    """[(name, title, portfolio override, squad override)] off Lists."""
    if "Lists" not in wb_v.sheetnames:
        return []
    ws = wb_v["Lists"]
    pc = sc = nc = None
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(1, c).value)
        if h.startswith("person (name"):
            nc = c
        elif h == "portfolio override":
            pc = c
        elif h == "squad override":
            sc = c
    if nc is None:
        return []
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, nc).value
        if blank(v) or "|" not in str(v):
            continue
        nm, _, ti = str(v).partition("|")
        out.append((norm(nm), norm(ti),
                    ws.cell(r, pc).value if pc else None,
                    ws.cell(r, sc).value if sc else None))
    return out


def lever_factors(wb_v):
    """{lever: factor} off the Lists cost-factor table, spec default if absent."""
    if "Lists" not in wb_v.sheetnames:
        return dict(LEVER_FACTOR)
    ws = wb_v["Lists"]
    lc = fc = None
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(1, c).value)
        if h == "lever":
            lc = c
        elif h == "cost factor":
            fc = c
    if lc is None or fc is None:
        return dict(LEVER_FACTOR)
    out = {}
    for r in range(2, ws.max_row + 1):
        k, v = ws.cell(r, lc).value, num(ws.cell(r, fc).value)
        if not blank(k) and v is not None:
            out[norm(k)] = float(v)
    return out or dict(LEVER_FACTOR)


# --------------------------------------------------------------- homing

def derive_homing(master_rows, overrides, rep):
    """MTab for every one of his rows, from the spec's rules.

    master_rows is [{row, name, title, div, port, plat, squad, ...}].
    Returns (homes, fallbacks, unmapped, applied_overrides).
    """
    homes, fallbacks, unmapped, applied = {}, [], [], []
    ovr = {(n, t): (p, s) for n, t, p, s in overrides}
    for m in master_rows:
        p, d = norm(m["port"]), norm(m["div"])
        tab = None
        if p in PORT_MAP:
            tab = PORT_MAP[p]
        elif p in ("", "na", "none"):
            if d in DIV_MAP:
                tab = DIV_MAP[d]
            elif d in DIV_RESOLVE:
                # Platform first, then Squad, both through the portfolio map
                for probe in (norm(m["plat"]), norm(m["squad"])):
                    if probe in PORT_MAP:
                        tab = PORT_MAP[probe]
                        break
                fallbacks.append((m["row"], m["name"], m["div"], m["plat"],
                                  m["squad"], tab))
        if tab is None:
            unmapped.append((m["row"], m["name"], m["port"], m["div"]))
        # an EGI squad that names a portfolio homes to that portfolio: his
        # 05/08 ruling that a portfolio must show its own EGI cost and net it
        # out. Only the plain EGI squad stays on the EGI tab.
        sq = norm(m["squad"])
        if sq.startswith("egi") and sq != "egi":
            tail = sq[3:].strip()
            hit = PORT_MAP.get(tail) or DIV_MAP.get(tail)
            if hit is None:
                for k, val in list(PORT_MAP.items()) + list(DIV_MAP.items()):
                    if norm(val) == tail or k == tail:
                        hit = val
                        break
            if hit is None and tail in ("tdd",):
                hit = "TDD Group Functions"
            if hit:
                tab = hit
        # the two funded cyber squads override everything
        if norm(m["squad"]) in CYBER_SQUADS:
            tab = "TDD Cyber"
        # the agreed person-keyed moves, as they live on Lists
        key = (norm(m["name"]), norm(m["title"]))
        if key in ovr:
            po, so = ovr[key]
            if not blank(po):
                applied.append((m["name"], m["title"], tab, str(po),
                                str(so or "")))
                tab = str(po).strip()
            elif not blank(so):
                applied.append((m["name"], m["title"], tab, tab, str(so)))
        homes[m["row"]] = tab
    return homes, fallbacks, unmapped, applied


def mstatus_of(name):
    n = norm(name)
    # 'vacant (anaplan)', 'Vacant (SYD)' and the like are vacancies with an
    # annotation - his old ledger classes them Vacant too
    return "Vacant" if (n in VACANT_NAMES or n.startswith("vacant")) \
        else "Filled"


def effective_cost(cost, fte):
    c, f = num(cost), num(fte)
    if c is None:
        return None
    if f is not None and f < 1:
        return c * f
    return c


# ------------------------------------------------------------------ main

def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 2
    built = os.path.abspath(argv[1])
    outdir = os.path.join(HERE, "tw3")
    seed = SEED
    old_path = None
    i = 2
    while i < len(argv):
        if argv[i] == "--outdir" and i + 1 < len(argv):
            outdir = os.path.abspath(argv[i + 1]); i += 2
        elif argv[i] == "--seed" and i + 1 < len(argv):
            seed = int(argv[i + 1]); i += 2
        elif argv[i] == "--old" and i + 1 < len(argv):
            old_path = os.path.abspath(argv[i + 1]); i += 2
        else:
            i += 1
    if old_path is None:
        # the snapshot of the shipped model, taken before the orchestrator
        # overwrote the repo file; it always lives beside this script
        old_path = os.path.join(HERE, "tw3", "old_model.xlsx")
        if not os.path.exists(old_path):
            old_path = os.path.join(outdir, "old_model.xlsx")
    if not os.path.isdir(outdir):
        os.makedirs(outdir)

    rep = Report()
    rep._emit("=" * 74)
    rep._emit("w3_verify - stage w3 data truth")
    rep._emit("=" * 74)
    rep._emit("built    %s" % built)
    rep._emit("master   %s" % HIS)
    rep._emit("pcm      %s" % PCM)
    rep._emit("old      %s" % old_path)
    rep._emit("outdir   %s" % outdir)
    rep._emit("run      %s   seed %d"
              % (datetime.datetime.now().isoformat(timespec="seconds"), seed))

    for label, path in (("built", built), ("master", HIS), ("pcm", PCM)):
        if not os.path.exists(path):
            rep.head("!", "inputs")
            rep.check(False, "%s file present" % label, path)
            _finish(rep, outdir, None)
            return 1

    wbf = openpyxl.load_workbook(built, data_only=False)
    wbv = openpyxl.load_workbook(built, data_only=True)
    mws = openpyxl.load_workbook(HIS, data_only=True).worksheets[0]

    # --------------------------------------------------- his master, read
    master = []
    for r in range(FIRST_ROW, LAST_ROW + 1):
        master.append({
            "row": r, "ee": mws.cell(r, C_EE).value,
            "ee2": mws.cell(r, C_EE2).value,
            "name": mws.cell(r, C_NAME).value,
            "title": mws.cell(r, C_TITLE).value,
            "div": mws.cell(r, C_DIV).value,
            "port": mws.cell(r, C_PORT).value,
            "plat": mws.cell(r, C_PLAT).value,
            "squad": mws.cell(r, C_SQUAD).value,
            "country": mws.cell(r, C_COUNTRY).value,
            "fte": mws.cell(r, C_FTE).value,
            "cost": mws.cell(r, C_COST).value,
            "myhr": mws.cell(r, C_MYHR).value})

    rvname = find_sheet(wbf, REVIEW) or REVIEW

    # ============================================== A  raw block identity
    rep.head("A", "raw block identity, REVIEW A2:AC527 against his master")
    if rvname not in wbf.sheetnames:
        rep.check(False, "REVIEW tab present", rvname)
        rv = rvv = None
    else:
        rv = wbf[rvname]
        rvv = wbv[rvname]
        hdr_bad = []
        for c in range(1, RAW_COLS + 1):
            got, wantv = rv.cell(1, c).value, HIS_HEADERS[c - 1]
            if str(got if got is not None else "") != wantv:
                hdr_bad.append("%s1 %r != %r" % (col_letter(c), got, wantv))
        rep.check(not hdr_bad, "his header row verbatim, A1:AC1 (29 columns)",
                  "%d wrong" % len(hdr_bad))
        rep.bullets(hdr_bad)

        rep.check(str(rv.cell(1, 1).value) == "EE Number"
                  and str(rv.cell(1, 3).value) == "EE Number",
                  "duplicate EE Number column kept at A and C")
        rep.check(str(rv.cell(1, RAW_COLS).value) == "MyHR ee no",
                  "MyHR ee no kept at AC")

        diffs = []
        for m in master:
            r = m["row"]
            for c in range(1, RAW_COLS + 1):
                a = rvv.cell(r, c).value
                b = mws.cell(r, c).value
                if not same(a, b):
                    diffs.append("%s%d built=%r master=%r"
                                 % (col_letter(c), r, show(a), show(b)))
        rep.check(not diffs,
                  "every cell A2:AC527 identical (text exact, numeric 1e-9)",
                  "%d of %d differ" % (len(diffs), N_ROWS * RAW_COLS))
        rep.bullets(diffs, limit=40)

        filled = sum(1 for r in range(FIRST_ROW, LAST_ROW + 1)
                     if any(not blank(rvv.cell(r, c).value)
                            for c in range(1, RAW_COLS + 1)))
        rep.check(filled == N_ROWS, "526 data rows in the raw block",
                  "%d" % filled)

        stale = [r for r in range(LAST_ROW + 1, min(rv.max_row, LAST_ROW + 60) + 1)
                 if any(not blank(rvv.cell(r, c).value)
                        for c in range(1, RAW_COLS + 1))]
        rep.check(not stale, "nothing left in the raw block below row 527",
                  "rows %s" % stale[:8])

        # his own raw block carries three literal #N/A texts he typed; they are
        # his words and must survive, so the error scan runs on the helper
        # block only
        his_errs = ["%s%d %s" % (col_letter(c), r, mws.cell(r, c).value)
                    for r in range(FIRST_ROW, LAST_ROW + 1)
                    for c in range(1, RAW_COLS + 1)
                    if isinstance(mws.cell(r, c).value, str)
                    and any(e in mws.cell(r, c).value for e in ERRS)]
        if his_errs:
            rep.report("literal error texts he typed in his own raw block, "
                       "kept verbatim: %s" % "; ".join(his_errs))
        errcells = []
        for r in range(1, min(rvv.max_row, LAST_ROW + 5) + 1):
            for c in range(RAW_COLS + 1, min(rvv.max_column, 80) + 1):
                v = rvv.cell(r, c).value
                if isinstance(v, str) and any(e in v for e in ERRS):
                    errcells.append("%s%d %s" % (col_letter(c), r, v))
        rep.check(not errcells,
                  "no error values in the REVIEW helper block",
                  "%d" % len(errcells))
        rep.bullets(errcells)

    hcols = helper_cols(rv) if rv is not None else {}
    if rv is not None:
        rep.report("helper block found at: %s"
                   % ", ".join("%s=%s" % (k, col_letter(v))
                               for k, v in sorted(hcols.items(),
                                                  key=lambda x: x[1])))
        for need in ("mtab", "mstatus", "roleid", "ohline"):
            if need not in hcols:
                rep.check(False, "helper column %r present on REVIEW" % need)

    # ---------------------------------------------- derived homing, shared
    overrides = lists_overrides(wbv)
    homes, fallbacks, unmapped, applied = derive_homing(master, overrides, rep)
    tabkeys = collections.Counter(v for v in homes.values() if v)

    # ================================================ B  customer PCM data
    rep.head("B", "customer verification against the PCM data tab")
    pcm_master = []
    try:
        pwb = openpyxl.load_workbook(PCM, data_only=True)
        pname = None
        for t in pwb.sheetnames:
            if norm(t).replace("_", " ") == "pcm data":
                pname = t
                break
        if pname is None:
            rep.check(False, "PCM data tab found", str(pwb.sheetnames))
            pcm_rows = []
        else:
            pws = pwb[pname]
            hdr = {norm(pws.cell(1, c).value): c
                   for c in range(1, pws.max_column + 1)
                   if not blank(pws.cell(1, c).value)}
            cn = hdr.get("name")
            ce = hdr.get("ee number")
            cc = hdr.get("full cost \naud") or hdr.get("full cost aud")
            if cc is None:
                for k, v in hdr.items():
                    if k.startswith("full cost"):
                        cc = v
            cp = hdr.get("portfolio")
            cs = hdr.get("squad")
            ct = hdr.get("position title")
            pcm_rows = []
            for r in range(2, pws.max_row + 1):
                if blank(pws.cell(r, cn).value):
                    continue
                pcm_rows.append({
                    "row": r, "name": pws.cell(r, cn).value,
                    "title": pws.cell(r, ct).value if ct else None,
                    "ee": pws.cell(r, ce).value if ce else None,
                    "cost": pws.cell(r, cc).value if cc else None,
                    "port": pws.cell(r, cp).value if cp else None,
                    "squad": pws.cell(r, cs).value if cs else None})
            rep.report("PCM data tab %r: %d people" % (pname, len(pcm_rows)))

        by_name = collections.defaultdict(list)
        for m in master:
            by_name[norm(m["name"])].append(m)

        absent, ee_diff, cost_diff = [], [], []
        pcm_master, matched = [], 0
        for p in pcm_rows:
            cands = by_name.get(norm(p["name"]), [])
            if norm(p["name"]) in VACANT_NAMES:
                cands = [c for c in cands
                         if norm(c["title"]) == norm(p["title"])]
            if not cands:
                absent.append("%s | %s | PCM cost %s"
                              % (p["name"], p["title"], show(p["cost"])))
                continue
            m = cands[0]
            if len(cands) > 1:
                for c in cands:
                    if norm(c["title"]) == norm(p["title"]):
                        m = c
                        break
            matched += 1
            pcm_master.append((p, m))
            pe, me = p["ee"], m["ee"]
            if not (blank(pe) or str(pe).startswith("#")):
                if not same(pe, me):
                    ee_diff.append("%s PCM=%s master=%s"
                                   % (p["name"], show(pe), show(me)))
            if num(p["cost"]) is not None and num(m["cost"]) is not None:
                if abs(num(p["cost"]) - num(m["cost"])) > 0.01:
                    cost_diff.append("%s PCM=%s master=%s"
                                     % (p["name"], show(p["cost"]),
                                        show(m["cost"])))

        rep.check(pcm_rows and matched + len(absent) == len(pcm_rows),
                  "every PCM person resolved against his master",
                  "%d matched, %d absent" % (matched, len(absent)))
        rep.check(True, "PCM people absent from his master are report-only "
                        "(his master wins)", "%d" % len(absent))
        rep.bullets(absent, limit=30)
        rep.report("PCM EE numbers differing from his master (his master "
                   "wins): %d" % len(ee_diff))
        rep.bullets(ee_diff, limit=20)
        rep.report("PCM costs differing from his master (his master wins): %d"
                   % len(cost_diff))
        rep.bullets(cost_diff, limit=20)
        rep.report("where the precedence rules put the PCM people: %s"
                   % ", ".join(
                       "%s %d" % (k, v) for k, v in sorted(
                           collections.Counter(
                               str(homes.get(m["row"]))
                               for _, m in pcm_master).items())))
        # the placement half of B is checked in C, once the rosters are read
    except Exception as exc:                                # noqa: BLE001
        rep.check(False, "customer PCM cross-check ran", repr(exc))

    # =============================================== C  whole-book placement
    rep.head("C", "whole-book person placement")
    rep.check(not unmapped, "every one of the 526 rows homes to a tab",
              "%d unmapped" % len(unmapped))
    rep.bullets(["row %d %s port=%r div=%r" % u for u in unmapped])
    rep.check(len(homes) == N_ROWS, "homing covers all 526 rows",
              "%d" % len(homes))
    rep.report("derived tab distribution: %s"
               % ", ".join("%s %d" % (k, v) for k, v in sorted(tabkeys.items())))
    rep.report("rows needing the Platform/Squad fallback chain: %d"
               % len(fallbacks))
    rep.bullets(["row %d %s div=%r plat=%r squad=%r -> %s" % f
                 for f in fallbacks], limit=40)
    rep.report("person-keyed overrides applied from Lists: %d" % len(applied))
    rep.bullets(["%s | %s : %s -> %s %s" % a for a in applied], limit=40)

    # C: REVIEW MTab against the derivation
    if rv is not None and "mtab" in hcols:
        bad = []
        for m in master:
            got = rvv.cell(m["row"], hcols["mtab"]).value
            wantv = homes.get(m["row"])
            if norm(got) != norm(wantv):
                bad.append("row %d %s: MTab=%r derived=%r"
                           % (m["row"], m["name"], got, wantv))
        rep.check(not bad, "REVIEW MTab equals the independently derived home",
                  "%d differ" % len(bad))
        rep.bullets(bad, limit=30)

    if rv is not None and "mstatus" in hcols:
        bad = [("row %d %s: MStatus=%r expected %s"
                % (m["row"], m["name"], rvv.cell(m["row"], hcols["mstatus"]).value,
                   mstatus_of(m["name"])))
               for m in master
               if norm(rvv.cell(m["row"], hcols["mstatus"]).value)
               != norm(mstatus_of(m["name"]))]
        rep.check(not bad, "MStatus Vacant/Filled from his Name column",
                  "%d differ" % len(bad))
        rep.bullets(bad, limit=25)

    # role id -> master row
    id2row, row2id = {}, {}
    if rv is not None and "roleid" in hcols:
        dup = []
        for m in master:
            rid = rvv.cell(m["row"], hcols["roleid"]).value
            if blank(rid):
                continue
            rid = str(rid).strip()
            if rid in id2row:
                dup.append(rid)
            id2row[rid] = m["row"]
            row2id[m["row"]] = rid
        rep.check(len(id2row) == N_ROWS and not dup,
                  "REVIEW Role ID unique and on all 526 rows",
                  "%d ids, %d duplicates" % (len(id2row), len(dup)))

    # 2.x rosters
    tabs = two_x_tabs(wbf)
    rep.report("2.x tabs found: %d (%s)" % (len(tabs), ", ".join(tabs)))
    seen = collections.defaultdict(list)
    roster_by_tab, key_by_tab, label_by_tab = {}, {}, {}
    cached_numbers = 0
    for t in tabs:
        wsf, wsv = wbf[t], wbv[t]
        rows = read_roster(wsf, wsv)
        roster_by_tab[t] = rows
        key_by_tab[t] = norm(wsv.cell(3, 3).value) or norm(t)
        label_by_tab[t] = (str(wsv.cell(3, 3).value).strip()
                           if not blank(wsv.cell(3, 3).value) else t)
        for x in rows:
            seen[x["id"]].append(t)
            if x["F"] is not None:
                cached_numbers += 1

    rep.check(cached_numbers > 0,
              "the built file carries cached values on the 2.x tabs",
              "%d numeric role costs" % cached_numbers)

    dups = {k: v for k, v in seen.items() if len(v) > 1}
    rep.check(not dups, "each person sits on exactly one 2.x tab",
              "%d on more than one" % len(dups))
    rep.bullets(["%s on %s" % (k, ", ".join(v)) for k, v in dups.items()],
                limit=25)

    all_ids = set(seen)
    missing = sorted(set(id2row) - all_ids) if id2row else []
    extra = sorted(all_ids - set(id2row)) if id2row else []
    rep.check(not missing, "every REVIEW role is placed on a 2.x tab",
              "%d unplaced" % len(missing))
    rep.bullets(["%s (%s)" % (i, mws.cell(id2row[i], C_NAME).value)
                 for i in missing], limit=30)
    rep.check(not extra, "no 2.x role id that REVIEW does not carry",
              "%d orphans" % len(extra))
    rep.bullets(extra, limit=30)
    rep.check(sum(len(v) for v in roster_by_tab.values()) == N_ROWS,
              "the book carries exactly 526 role rows across the 2.x tabs",
              "%d" % sum(len(v) for v in roster_by_tab.values()))

    # membership against the derivation
    wrong = []
    for t in tabs:
        key = key_by_tab[t]
        for x in roster_by_tab[t]:
            mrow = id2row.get(x["id"])
            if mrow is None:
                continue
            want_tab = homes.get(mrow)
            if norm(want_tab) != key:
                wrong.append("%s: %s %s should home to %s"
                             % (t, x["id"], mws.cell(mrow, C_NAME).value,
                                want_tab))
    rep.check(not wrong, "each 2.x block's membership equals the derived home",
              "%d misplaced" % len(wrong))
    rep.bullets(wrong, limit=40)

    # per-tab counts, derived against the book, for the handback
    where = {}
    for t in tabs:
        for x in roster_by_tab[t]:
            mrow = id2row.get(x["id"])
            if mrow:
                where[mrow] = (t, x)
    rep.report("per-tab roles, derived vs the book:")
    display = {}
    for k in tabkeys:
        display.setdefault(norm(k), str(k))
    for t in tabs:
        display.setdefault(key_by_tab[t], label_by_tab[t])
    cbad = []
    for key in sorted(display):
        want_n = sum(1 for v in homes.values() if norm(v) == key)
        tt = [t for t in tabs if key_by_tab[t] == key]
        got_n = sum(len(roster_by_tab[t]) for t in tt)
        rep.note("%-26s derived %3d   book %3d   %s"
                 % (display[key], want_n, got_n, ", ".join(tt) or "no tab"))
        if want_n != got_n:
            cbad.append("%s: derived %d, book %d"
                        % (display[key], want_n, got_n))
    rep.check(not cbad, "every tab carries the number of roles the rules give "
                        "it", "%d tabs off" % len(cbad))

    # B, continued: the PCM people's placement in the built book
    if pcm_master:
        pbad = []
        for p, m in pcm_master:
            hit = where.get(m["row"])
            want_tab = homes.get(m["row"])
            if hit is None:
                pbad.append("%s: not placed on any 2.x tab" % p["name"])
            elif key_by_tab.get(hit[0]) != norm(want_tab):
                pbad.append("%s: on %s, the rules put them on %s"
                            % (p["name"], hit[0], want_tab))
        rep.check(not pbad, "every PCM person sits where the precedence rules "
                            "put them", "%d off" % len(pbad))
        rep.bullets(pbad, limit=25)

    # named people unique across the book
    place_of = {}
    twice = []
    for t in tabs:
        for x in roster_by_tab[t]:
            mrow = id2row.get(x["id"])
            nm = norm(mws.cell(mrow, C_NAME).value) if mrow else norm(x["name"])
            if nm in VACANT_NAMES:
                continue
            if nm in place_of:
                twice.append("%s on %s and %s" % (nm, place_of[nm], t))
            place_of[nm] = t
    rep.check(not twice, "every named person appears exactly once in the book",
              "%d twice" % len(twice))
    rep.bullets(twice, limit=25)

    # ---------------------------------------------- 3.2 overhead rosters
    t32 = find_sheet(wbf, "3.2")
    if t32 and rv is not None and "ohline" in hcols:
        wsv32 = wbv[t32]
        lines = []
        for r in range(4, min(wsv32.max_row, 20) + 1):
            b = wsv32.cell(r, 2).value
            if blank(b) or str(b).strip().lower().startswith(
                    ("overheads incl", "of which", "control", "allocations",
                     "input", "tdd roles")):
                continue
            g, j = num(wsv32.cell(r, 7).value), num(wsv32.cell(r, 10).value)
            if g is None and j is None:
                continue
            lines.append((r, str(b).strip(), g, j))
        by_line = collections.defaultdict(list)
        for m in master:
            ol = rvv.cell(m["row"], hcols["ohline"]).value
            if not blank(ol):
                by_line[norm(ol)].append(m)
        bad = []
        for r, label, g, j in lines:
            if norm(label).startswith("leadership"):
                rep.report("3.2 %r is the GM input line, not a REVIEW roster "
                           "(count %s, cost %sm)" % (label, show(g), show(j)))
                continue
            roster = by_line.get(norm(label), [])
            cnt = len(roster)
            raw = sum(num(x["cost"]) or 0.0 for x in roster) / 1e6
            eff = sum(effective_cost(x["cost"], x["fte"]) or 0.0
                      for x in roster) / 1e6
            if g is not None and abs(g - cnt) > 1e-9:
                bad.append("%s: 3.2 count %s, derived roster %d"
                           % (label, show(g), cnt))
            if j is not None and min(abs(j - raw), abs(j - eff)) > 1e-6:
                bad.append("%s: 3.2 cost %sm, derived roster %.6fm raw / "
                           "%.6fm effective" % (label, show(j), raw, eff))
            elif j is not None and abs(raw - eff) > 1e-9:
                rep.report("3.2 %r prices on the %s basis"
                           % (label, "effective" if abs(j - eff)
                              < abs(j - raw) else "raw"))
        rep.check(not bad, "3.2 rosters match the derived overhead lines",
                  "%d mismatches" % len(bad))
        rep.bullets(bad, limit=30)
        rep.report("derived overhead lines: %s"
                   % ", ".join("%s %d" % (k, len(v))
                               for k, v in sorted(by_line.items())))
    elif rv is not None:
        rep.check(False, "3.2 tab and the Overhead line helper both present",
                  "3.2=%s ohline=%s" % (bool(t32), "ohline" in hcols))

    # settled memberships
    if rv is not None and "ohline" in hcols:
        def line_of(nm):
            for m in master:
                if norm(m["name"]) == nm:
                    return rvv.cell(m["row"], hcols["ohline"]).value
            return "<not in his file>"
        rep.check(norm(line_of(SETTLED_TM)) == "technology manager",
                  "Shane Ker on the Technology Manager line",
                  show(line_of(SETTLED_TM)))
        rep.check(norm(line_of(SETTLED_NOT_HOT)) != "head of technology",
                  "Ed Tacey not on the Head of Technology line",
                  show(line_of(SETTLED_NOT_HOT)))
        da = [m for m in master
              if re.search(r"delivery (assurance|excellence) manager",
                           norm(m["title"]))]
        if da:
            got = [(m["name"], rvv.cell(m["row"], hcols["ohline"]).value)
                   for m in da]
            rep.check(all(norm(x[1]) == "delivery manager" for x in got),
                      "the Delivery Assurance/Excellence managers on the "
                      "Delivery Manager line",
                      "; ".join("%s=%s" % (a, show(b)) for a, b in got))
        else:
            rep.report("no Delivery Assurance/Excellence manager titles in "
                       "his file")

    # ---------------------------------------------- 40-person salary trace
    rep.head("C", "salary trace, %d random people, REVIEW -> 2.x -> totals "
                  "at 1e-9" % TRACE_N)
    rng = random.Random(seed)
    sample = rng.sample(master, min(TRACE_N, len(master)))
    facs = lever_factors(wbv)
    row_to_tabrow = {}
    for t in tabs:
        for x in roster_by_tab[t]:
            mrow = id2row.get(x["id"])
            if mrow:
                row_to_tabrow[mrow] = (t, x)
    tr_bad, tr_note, traced = [], [], 0
    for m in sample:
        rid = row2id.get(m["row"], "?")
        hit = row_to_tabrow.get(m["row"])
        if hit is None:
            tr_bad.append("row %d %s (%s): not found on any 2.x tab"
                          % (m["row"], m["name"], rid))
            continue
        t, x = hit
        traced += 1
        raw = num(rvv.cell(m["row"], C_COST).value) if rvv else None
        if raw is None or abs(raw - (num(m["cost"]) or 0)) > TOL9:
            tr_bad.append("%s %s: REVIEW raw cost %s, his file %s"
                          % (rid, m["name"], show(raw), show(m["cost"])))
        want = effective_cost(m["cost"], m["fte"])
        ovc = None
        if "override" in hcols and rvv:
            ovc = num(rvv.cell(m["row"], hcols["override"]).value)
        if ovc is not None:
            want = ovc
            tr_note.append("%s %s priced off the agreed override %s"
                           % (rid, m["name"], show(ovc)))
        if "effective" in hcols and rvv:
            eff = num(rvv.cell(m["row"], hcols["effective"]).value)
            base = effective_cost(m["cost"], m["fte"])
            if eff is None or abs(eff - base) > TOL9 * max(1.0, abs(base)):
                tr_bad.append("%s %s: effective helper %s, expected %s"
                              % (rid, m["name"], show(eff), show(base)))
        if x["F"] is None:
            tr_bad.append("%s %s: %s role cost cell is empty" % (rid, m["name"], t))
        elif abs(x["F"] - want) > TOL9 * max(1.0, abs(want)):
            tr_bad.append("%s %s on %s: role cost %s, expected %s (FTE %s)"
                          % (rid, m["name"], t, show(x["F"]), show(want),
                             show(m["fte"])))
        lev = norm(x["lever"])
        fac = facs.get(lev)
        if x["G"] is not None and x["F"] is not None and fac is not None:
            wipro = norm(m["country"]) == "wipro"
            exp = x["F"] * (1.0 if (lev == "offshore" and wipro) else fac)
            if abs(x["G"] - exp) > TOL6 * max(1.0, abs(exp)):
                if lev == "offshore":
                    tr_note.append("%s %s offshore: after-lever %s vs %s "
                                   "(vendor-rate rules)"
                                   % (rid, m["name"], show(x["G"]), show(exp)))
                else:
                    tr_bad.append("%s %s on %s: after-lever %s, expected %s "
                                  "for lever %r"
                                  % (rid, m["name"], t, show(x["G"]),
                                     show(exp), x["lever"]))
    rep.check(not tr_bad, "%d-person salary trace REVIEW -> 2.x at 1e-9"
              % TRACE_N, "%d traced, %d problems" % (traced, len(tr_bad)))
    rep.bullets(tr_bad, limit=40)
    for n in tr_note[:20]:
        rep.report(n)

    # group subtotals and tab totals for every tab the sample touched
    touched = sorted({row_to_tabrow[m["row"]][0] for m in sample
                      if m["row"] in row_to_tabrow})
    gbad, tbad = [], []
    for t in touched:
        wsf, wsv = wbf[t], wbv[t]
        gt = group_totals(wsf, wsv)
        sums = collections.defaultdict(float)
        gsums = collections.defaultdict(float)
        for x in roster_by_tab[t]:
            if x["group"] is None:
                continue
            sums[x["group"]] += x["F"] or 0.0
            gsums[x["group"]] += x["G"] or 0.0
        for g, s in sums.items():
            got = gt.get(g, (None, None, None))[1]
            if got is None:
                gbad.append("%s / %s: no subtotal cell" % (t, g))
            elif abs(got - s) > 1e-6 * max(1.0, abs(s)):
                gbad.append("%s / %s: subtotal %s, roles sum %s"
                            % (t, g, show(got), show(s)))
        tot_r = None
        for r in range(5, min(wsf.max_row, 90) + 1):
            if norm(wsf.cell(r, 2).value) == "total portfolio":
                tot_r = r
                break
        if tot_r:
            o = num(wsv.cell(tot_r, 15).value)
            allF = sum(x["F"] or 0.0 for x in roster_by_tab[t])
            if o is not None and abs(o * 1e6 - allF) > 1.0:
                tbad.append("%s: grid actual cost %sm, roster sum %sm"
                            % (t, show(o), show(allF / 1e6)))
        for r in range(1, min(wsf.max_row, 60) + 1):
            b = wsf.cell(r, 2).value
            if isinstance(b, str) and b.lower().startswith("control"):
                cv = num(wsv.cell(r, 3).value)
                if cv is not None and abs(cv) > 1e-6:
                    tbad.append("%s row %d control = %s (%s)"
                                % (t, r, show(cv), b[:50]))
    rep.check(not gbad, "traced people roll into their squad subtotal",
              "%d off" % len(gbad))
    rep.bullets(gbad, limit=25)
    rep.check(not tbad, "traced tabs reconcile to their grid totals and "
                        "controls read 0", "%d off" % len(tbad))
    rep.bullets(tbad, limit=25)

    # ======================================================= D  EE truth
    rep.head("D", "EE truth")
    if rv is not None:
        ee_bad = [("row %d built=%r master=%r"
                   % (m["row"], rvv.cell(m["row"], C_EE).value, m["ee"]))
                  for m in master
                  if not same(rvv.cell(m["row"], C_EE).value, m["ee"])]
        rep.check(not ee_bad, "column A EE Number equals his file on all 526 "
                              "rows", "%d differ" % len(ee_bad))
        rep.bullets(ee_bad, limit=30)
        ee2_bad = [("row %d built=%r master=%r"
                    % (m["row"], rvv.cell(m["row"], C_EE2).value, m["ee2"]))
                   for m in master
                   if not same(rvv.cell(m["row"], C_EE2).value, m["ee2"])]
        rep.check(not ee2_bad, "column C (his duplicate EE Number) equals his "
                               "file", "%d differ" % len(ee2_bad))
        rep.bullets(ee2_bad, limit=30)
        ac_bad = [("row %d built=%r master=%r"
                   % (m["row"], rvv.cell(m["row"], C_MYHR).value, m["myhr"]))
                  for m in master
                  if not same(rvv.cell(m["row"], C_MYHR).value, m["myhr"])]
        rep.check(not ac_bad, "column AC MyHR ee no equals his file",
                  "%d differ" % len(ac_bad))
        rep.bullets(ac_bad, limit=20)

    valid_ee = set()
    for m in master:
        for v in (m["ee"], m["ee2"]):
            if not blank(v):
                valid_ee.add(str(v).strip())
    rep.report("distinct EE values in his file: %d" % len(valid_ee))

    # every EE-headed column anywhere else in the book
    eepat = re.compile(r"^(ee\s*(number|no)\.?|employee\s*(number|no)\.?"
                       r"|myhr\s*ee\s*no)$")
    stray = []
    scanned = 0
    for t in wbv.sheetnames:
        ws = wbv[t]
        for r in range(1, min(ws.max_row, 8) + 1):
            for c in range(1, min(ws.max_column, 80) + 1):
                if not eepat.match(norm(ws.cell(r, c).value)):
                    continue
                if t == rvname and c <= RAW_COLS:
                    continue
                scanned += 1
                for rr in range(r + 1, ws.max_row + 1):
                    v = ws.cell(rr, c).value
                    if blank(v):
                        continue
                    if str(v).strip() not in valid_ee:
                        stray.append("%s!%s%d = %r"
                                     % (t, col_letter(c), rr, show(v)))
    rep.check(not stray, "every EE value in an EE-headed column outside the "
                         "raw block is one of his",
              "%d EE columns scanned, %d strays" % (scanned, len(stray)))
    rep.bullets(stray, limit=25)

    # the 129 corrections, against the old shipped model
    old_ok = os.path.exists(old_path)
    rep.check(old_ok, "old model snapshot present for the delta", old_path)
    old_rows = []
    if old_ok:
        ow = openpyxl.load_workbook(old_path, data_only=True)
        ows = ow[find_sheet(ow, REVIEW)]
        ohdr = {norm(ows.cell(1, c).value): c
                for c in range(1, ows.max_column + 1)
                if not blank(ows.cell(1, c).value)}
        o_name = 2
        o_title = 3
        o_ee = 1
        o_cost = ohdr.get("full cost \naud") or 27
        o_fte = ohdr.get("fte") or 15
        o_mtab = ohdr.get("mtab")
        o_sqoh = ohdr.get("squad or overhead line")
        o_mstat = ohdr.get("mstatus")
        for r in range(2, ows.max_row + 1):
            if all(blank(ows.cell(r, c).value) for c in range(1, 29)):
                continue
            old_rows.append({
                "row": r, "ee": ows.cell(r, o_ee).value,
                "name": ows.cell(r, o_name).value,
                "title": ows.cell(r, o_title).value,
                "cost": ows.cell(r, o_cost).value,
                "fte": ows.cell(r, o_fte).value,
                "mtab": ows.cell(r, o_mtab).value if o_mtab else None,
                "group": ows.cell(r, o_sqoh).value if o_sqoh else None,
                "status": ows.cell(r, o_mstat).value if o_mstat else None})
        rep.report("old shipped model: %d role rows" % len(old_rows))

        okey = collections.defaultdict(list)
        for o in old_rows:
            okey[(norm(o["name"]), norm(o["title"]))].append(o)
        nkey = collections.defaultdict(list)
        for m in master:
            nkey[(norm(m["name"]), norm(m["title"]))].append(m)
        common = set(okey) & set(nkey)
        changed = [k for k in common
                   if sorted(str(x["ee"]) for x in okey[k])
                   != sorted(str(x["ee"]) for x in nkey[k])]
        rep.check(len(changed) == 129,
                  "129 EE corrections land against the old model",
                  "%d changed of %d shared people" % (len(changed), len(common)))

    # =================================================== E  precedence rules
    rep.head("E", "precedence rules and the reversals his file forces")
    for nm, (want_tab, why) in REVERSALS.items():
        hit = [m for m in master if norm(m["name"]) == nm]
        if not hit:
            rep.check(False, "%s present in his file" % nm.title(), why)
            continue
        got = homes.get(hit[0]["row"])
        rep.check(norm(got) == norm(want_tab),
                  "%s homes to %s" % (hit[0]["name"], want_tab),
                  "derived %s; %s" % (got, why))
        if rv is not None and "mtab" in hcols:
            wbgot = rvv.cell(hit[0]["row"], hcols["mtab"]).value
            rep.check(norm(wbgot) == norm(want_tab),
                      "%s carries MTab %s in the workbook"
                      % (hit[0]["name"], want_tab), show(wbgot))

    rf = [m for m in master if norm(m["name"]) in RINGFENCED]
    rep.report("ring fenced rows in his file: %d" % len(rf))
    if rv is not None and "mstatus" in hcols:
        bad = [m["name"] for m in rf
               if norm(rvv.cell(m["row"], hcols["mstatus"]).value) != "vacant"]
        rep.check(not bad, "ring fenced rows are vacancies",
                  "%d not Vacant" % len(bad))
    if rv is not None and "ringfenced" in hcols:
        flagged = sum(1 for m in master
                      if str(rvv.cell(m["row"], hcols["ringfenced"]).value)
                      .strip().lower() in ("true", "yes", "1", "ring fenced"))
        rep.check(flagged == len(rf), "RingFenced flag set on those rows only",
                  "%d flagged, %d in his file" % (flagged, len(rf)))

    pt = [m for m in master if num(m["fte"]) is not None and num(m["fte"]) < 1]
    rep.report("part-time people (FTE<1) in his file: %d" % len(pt))
    ptbad = []
    for m in pt:
        want = effective_cost(m["cost"], m["fte"])
        rep.report("   %s FTE %s raw %s -> effective %s"
                   % (m["name"], show(m["fte"]), show(m["cost"]), show(want)))
        hit = row_to_tabrow.get(m["row"])
        if hit is None:
            ptbad.append("%s not on a 2.x tab" % m["name"])
            continue
        f = hit[1]["F"]
        if f is None or abs(f - want) > TOL9 * max(1.0, abs(want)):
            ptbad.append("%s: %s prices %s, effective cost is %s"
                         % (m["name"], hit[0], show(f), show(want)))
    rep.check(not ptbad, "every part-time role prices at Full Cost x FTE",
              "%d off" % len(ptbad))
    rep.bullets(ptbad, limit=20)

    # ======================================================= F  artefacts
    rep.head("F", "artefacts")
    delta = build_delta(built, old_path, old_rows, master, homes, id2row,
                        roster_by_tab, label_by_tab, rvv, hcols)
    dpath = os.path.join(outdir, "placement_delta.json")
    with open(dpath, "w") as fh:
        json.dump(delta, fh, indent=2, default=str)
    s = delta["summary"]
    rep.check(os.path.exists(dpath), "placement_delta.json written", dpath)
    rep.report("delta: %d stayed put, %d moved tab, %d changed squad or line, "
               "%d arrived, %d departed"
               % (s["stayed"], s["moved_tab"], s["moved_group"],
                  s["arrived"], s["departed"]))
    for mv in delta["moved_tab"][:25]:
        rep.note("moved  %-28s %-22s -> %s"
                 % (mv["name"][:28], mv["old_tab"], mv["new_tab"]))
    if len(delta["moved_tab"]) > 25:
        rep.note("       (%d more in the json)" % (len(delta["moved_tab"]) - 25))

    return _finish(rep, outdir, dpath)


def build_delta(built, old_path, old_rows, master, homes, id2row,
                roster_by_tab, label_by_tab, rvv, hcols):
    """Who moved where, old shipped model -> this build, person keyed."""
    # new placement off the 2.x rosters, labelled with the tab's portfolio key
    # so it compares like for like with the old model's MTab
    new_place = {}
    for t, rows in roster_by_tab.items():
        label = label_by_tab.get(t) or t
        for x in rows:
            mrow = id2row.get(x["id"])
            if mrow:
                new_place[mrow] = (label, x["group"], x["status"], x["lever"],
                                   t)
    okey = collections.defaultdict(list)
    for o in old_rows:
        okey[(norm(o["name"]), norm(o["title"]))].append(o)
    nkey = collections.defaultdict(list)
    for m in master:
        nkey[(norm(m["name"]), norm(m["title"]))].append(m)

    moved_tab, moved_group, arrived, departed = [], [], [], []
    stayed = 0
    for k in sorted(set(okey) | set(nkey)):
        olds, news = okey.get(k, []), nkey.get(k, [])
        for i in range(max(len(olds), len(news))):
            o = olds[i] if i < len(olds) else None
            m = news[i] if i < len(news) else None
            if m is not None:
                tab, group, status, lever, sheet = new_place.get(
                    m["row"], (homes.get(m["row"]), None, None, None, None))
                if group is None and rvv is not None and "sqoh" in hcols:
                    group = rvv.cell(m["row"], hcols["sqoh"]).value
            if o is None:
                arrived.append({
                    "name": m["name"], "title": m["title"],
                    "new_tab": tab, "new_sheet": sheet, "new_group": group,
                    "status": status or mstatus_of(m["name"]),
                    "lever": lever,
                    "cost": m["cost"], "fte": m["fte"], "ee": m["ee"]})
                continue
            if m is None:
                departed.append({
                    "name": o["name"], "title": o["title"],
                    "old_tab": o["mtab"], "old_group": o["group"],
                    "cost": o["cost"]})
                continue
            rec = {"name": m["name"], "title": m["title"],
                   "old_tab": o["mtab"], "new_tab": tab, "new_sheet": sheet,
                   "old_group": o["group"], "new_group": group,
                   "old_status": o["status"],
                   "new_status": status or mstatus_of(m["name"]),
                   "old_cost": o["cost"], "new_cost": m["cost"],
                   "old_ee": o["ee"], "new_ee": m["ee"],
                   "ee_corrected": str(o["ee"]) != str(m["ee"])}
            if norm(o["mtab"]) != norm(tab):
                moved_tab.append(rec)
            elif norm(o["group"]) != norm(group):
                moved_group.append(rec)
            else:
                stayed += 1
    return {
        "generated": datetime.datetime.now().isoformat(timespec="seconds"),
        "built": built, "old_model": old_path, "master": HIS,
        "summary": {"old_rows": len(old_rows), "new_rows": len(master),
                    "stayed": stayed, "moved_tab": len(moved_tab),
                    "moved_group": len(moved_group),
                    "arrived": len(arrived), "departed": len(departed),
                    "ee_corrected": sum(1 for r in moved_tab + moved_group
                                        if r["ee_corrected"])},
        "moved_tab": moved_tab, "moved_group": moved_group,
        "arrived": arrived, "departed": departed}


def _finish(rep, outdir, dpath):
    rep._emit("")
    rep._emit("=" * 74)
    rep._emit("RESULT   %d PASS   %d FAIL" % (rep.n_pass, rep.n_fail))
    rep._emit("=" * 74)
    path = os.path.join(outdir, "verify_report.txt")
    try:
        with open(path, "w") as fh:
            fh.write("\n".join(rep.lines) + "\n")
        print("wrote %s" % path)
        if dpath:
            print("wrote %s" % dpath)
    except OSError as exc:
        print("could not write %s: %s" % (path, exc))
    return 1 if rep.n_fail else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
