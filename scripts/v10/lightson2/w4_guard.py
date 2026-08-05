#!/usr/bin/env python3
"""w4 - spec stage w4: protection re-scoped, the two small fixes, nothing else.

  python3 w4_guard.py <in.xlsx> <out.xlsx>

His 05/08 ruling supersedes the whole-book lock: protection stays on the tabs
nobody types into and comes off everything a GM works in.

  G1  protection on the 0.x and 3.x tabs only, password Tdd123.  Everything
      else - the executive summary, the role mapping, Lists, the 1.x funding
      tabs, the 2.x lever tabs, the QA tab, the section dividers - carries no
      sheet protection at all.  The workbook structure stays locked, so tabs
      still cannot be moved, renamed or unhidden.
  G2  on the tabs that stay protected the typed inputs stay unlocked: every
      cream cell, and on 3.5 / 3.6 the overhead toggles, which have to be
      editable while the tab around them is locked.
  BLD-18  a filled person carrying the lever Hire is stale state - Filled and
      Hire price identically (Lists cost factor 1 either way), so the levers
      are set to Filled and the cost cannot move.  Every pair found is listed.
  BLD-19  1.10 row 17 reads a cell his Finance budget table holds as the text
      "-", so the model tab printed a dash.  The reference is wrapped in N(),
      which reads text as zero, and the cell takes the money format the rest
      of its column carries.  His 0.1 cell is not touched.

Idempotent: every edit is a no-op the second time, and handed its own output
the script recognises the finished state and copies it through untouched.
"""
import sys, os, re, shutil, collections

V10 = "/home/user/anthropic-claude-code/scripts/v10"
sys.path.insert(0, V10)
sys.path.insert(0, os.path.join(V10, "update"))

import openpyxl
from openpyxl.styles import Protection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.workbook.protection import WorkbookProtection

import wbio
from _xl import REVIEW, LEVERS, Log, load, save

try:
    from openpyxl.utils.protection import hash_password
except ImportError:                                   # older layout
    from openpyxl.worksheet.protection import hash_password

PW = "Tdd123"
PWH = hash_password(PW)

# the tabs that stay protected: his input tabs and the summary tabs
KEEP = re.compile(r"^[03]\.\d")
LIGHTS = "Lights On"
CLASSES = ("cream", "toggle")

# blocked (True) and allowed (False) actions on a protected sheet
OPTIONS = dict(sheet=True, objects=True, scenarios=True,
               formatCells=True, formatColumns=True, formatRows=True,
               insertColumns=True, insertRows=True, insertHyperlinks=True,
               deleteColumns=True, deleteRows=True,
               sort=True, autoFilter=True, pivotTables=True,
               selectLockedCells=False, selectUnlockedCells=False)

# BLD-19: a bare reference into his Finance budget table, which holds typed
# dashes in the cells that have no number
FIN = "0.1 Budget Table (Fin)"
BARE = re.compile(r"^=\s*'%s'!\$?[A-Za-z]{1,3}\$?\d+\s*$" % re.escape(FIN))
MONEY = r"#,##0.00;\(#,##0.00\);#,##0.00"
DASHY = ("-", "–", "—")

TOL = 1e-9


# his raw block carries typed #N/A text of his own (A31, C31, S525 in his
# file); his words are untouchable, so the error scan exempts exactly the
# raw-block addresses his file holds as #N/A - and nothing else
HIS = ("/root/.claude/uploads/e550b440-3996-5abb-87e5-bafafe598f82/"
       "0ad63df5-updates.xlsx")
_HISNA = None


def his_na_cells():
    global _HISNA
    if _HISNA is None:
        hb = openpyxl.load_workbook(HIS, read_only=True, data_only=True)
        hs = hb[hb.sheetnames[0]]
        _HISNA = {c.coordinate for row in hs.iter_rows() for c in row
                  if isinstance(c.value, str) and c.value.strip() == "#N/A"}
        hb.close()
    return _HISNA


def real_errors(err):
    """Split the audit into (real, his own #N/A raw-block cells)."""
    keep, his = [], []
    for sn, coord, val, formula in err:
        if sn == REVIEW and coord in his_na_cells() and val == "#N/A":
            his.append(coord)
        else:
            keep.append((sn, coord, val, formula))
    return keep, his


# ------------------------------------------------------------------ helpers

def protected(title):
    return bool(KEEP.match(title))


def is_cream(cell):
    f = cell.fill
    if f is None or f.patternType != "solid":
        return False
    rgb = getattr(f.fgColor, "rgb", None)
    return isinstance(rgb, str) and rgb.endswith("FFF2CC")


def dv_cells(ws, pick):
    """Coordinates covered by a list validation whose formula pick() accepts."""
    out = set()
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list" or not pick(str(dv.formula1)):
            continue
        for cr in dv.sqref.ranges:
            for r in range(cr.min_row, cr.max_row + 1):
                for c in range(cr.min_col, cr.max_col + 1):
                    out.add(ws.cell(r, c).coordinate)
    return out


def cream_cells(ws):
    return {c.coordinate for row in ws.iter_rows() for c in row if is_cream(c)}


def toggles(ws):
    """The overhead toggle cells on a Lights On tab, wherever the column sits.

    The percentage dropdown names them; if the tab is ever rebuilt without one,
    the cream that marks a typed input names them instead.
    """
    if LIGHTS not in ws.title:
        return set()
    return dv_cells(ws, lambda f: "%" in f and "100%" in f) or cream_cells(ws)


def allowmap(wb):
    """{sheet: {coord: class}} - what stays unlocked on the protected tabs."""
    out = {}
    for ws in wb.worksheets:
        if not protected(ws.title):
            continue
        tg = toggles(ws)
        allow = {co: "toggle" for co in tg}
        for co in cream_cells(ws):
            if co not in allow:
                allow[co] = "cream"
        out[ws.title] = allow
    return out


def lever_cells(wb):
    """[(sheet, row, coord)] for every cell carrying a vacancy lever."""
    out = []
    for ws in wb.worksheets:
        cov = dv_cells(ws, lambda f: "Filled" in f and "Hire" in f)
        for co in sorted(cov, key=lambda c: (ws[c].column, ws[c].row)):
            out.append((ws.title, ws[co].row, co))
    return out


def stray_levers(wb, known):
    """Lever-valued cells no dropdown covers - reported, never edited."""
    seen = {(s, c) for s, _, c in known}
    out = []
    for ws in wb.worksheets:
        if not ws.title.startswith("2."):
            continue
        for row in ws.iter_rows(min_col=5, max_col=5):
            for c in row:
                if c.value in LEVERS and (ws.title, c.coordinate) not in seen:
                    out.append((ws.title, c.coordinate, c.value))
    return out


def cached(path):
    """{sheet: {coord: value}} of the values the file already carries."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {ws.title: {c.coordinate: c.value for row in ws.iter_rows()
                      for c in row if c.value is not None}
           for ws in wb.worksheets}
    wb.close()
    return out


def values_for(src, wb):
    """Cached values for src, recalculating first if the cache is not there."""
    vals = cached(src)
    want = [(s, "D%d" % r) for s, r, _ in lever_cells(wb)]
    miss = [k for k in want if vals.get(k[0], {}).get(k[1]) is None]
    if not miss:
        return vals, False
    print("cache is thin (%d of %d status cells empty) - recalculating the "
          "input first" % (len(miss), len(want)), flush=True)
    return cached(wbio.recalc(src)), True


def stale_pairs(wb, vals):
    """[(sheet, row, coord, name, title)] filled people carrying Hire."""
    out = []
    for sheet, r, co in lever_cells(wb):
        if wb[sheet][co].value != "Hire":
            continue
        v = vals.get(sheet, {})
        if v.get("D%d" % r) != "Filled":
            continue
        out.append((sheet, r, co, v.get("B%d" % r), v.get("C%d" % r)))
    return out


def dash_pullers(wb, vals):
    """[(sheet, coord, formula)] 1.x cells that print his typed dash."""
    out = []
    for ws in wb.worksheets:
        if not ws.title.startswith("1."):
            continue
        v = vals.get(ws.title, {})
        for row in ws.iter_rows():
            for c in row:
                f = c.value
                if not (isinstance(f, str) and BARE.match(f)):
                    continue
                cv = v.get(c.coordinate)
                if isinstance(cv, str) and cv.strip() in DASHY:
                    out.append((ws.title, c.coordinate, f))
    return out


def already(wb, vals):
    """True when the file is already in the finished state."""
    want = {ws.title for ws in wb.worksheets if protected(ws.title)}
    got = {ws.title for ws in wb.worksheets
           if ws.protection.sheet and ws.protection.password == PWH}
    if want != got:
        return False
    if any(ws.protection.sheet for ws in wb.worksheets
           if not protected(ws.title)):
        return False
    sec = wb.security
    if not (sec is not None and sec.lockStructure and
            sec.workbookPassword == PWH):
        return False
    return not stale_pairs(wb, vals) and not dash_pullers(wb, vals)


# --------------------------------------------------------------------- main

def main(src, dst):
    log = Log("w4_guard")
    wb = load(src)
    vals, recalced = values_for(src, wb)

    if already(wb, vals):
        print("input is already guarded - copying through")
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        return

    keep = [ws.title for ws in wb.worksheets if protected(ws.title)]
    drop = [ws.title for ws in wb.worksheets if not protected(ws.title)]

    # ------------------------------------------------- G1  protection scope
    log.head("G1  protection on the 0.x and 3.x tabs only - his 05/08 ruling")
    log.note("G1", "the before picture comes from %s"
             % ("a fresh recalculation of the input" if recalced
                else "the values the input already carries"))
    for ws in wb.worksheets:
        was = bool(ws.protection.sheet)
        if protected(ws.title):
            p = ws.protection
            p.set_password(PW)
            for k, v in OPTIONS.items():
                setattr(p, k, v)
            if not was:
                log("G1", ws.title, "protected with the password - select "
                                    "locked and unlocked allowed, sort and "
                                    "filter off")
        else:
            ws.protection = SheetProtection(sheet=False)
            if was:
                log("G1", ws.title,
                    "protection removed - every cell on the tab is typeable")
    log.note("G1", "%d tabs protected (%s), %d unprotected"
             % (len(keep), ", ".join(keep), len(drop)))

    # --------------------------------------------------- G2  the input cells
    log.head("G2  on the protected tabs the typed inputs stay unlocked")
    allow = allowmap(wb)
    for title in keep:
        ws, a = wb[title], allow[title]
        relocked = unlocked = 0
        for row in ws.iter_rows():
            for c in row:
                want = c.coordinate not in a
                if bool(c.protection.locked) != want:
                    c.protection = Protection(locked=want,
                                              hidden=c.protection.hidden)
                    unlocked += 0 if want else 1
                    relocked += 1 if want else 0
        by = collections.Counter(a.values())
        log("G2", title, "%d cells unlocked (%s)%s"
            % (len(a), ", ".join("%d %s" % (by[k], k) for k in CLASSES
                                 if by[k]) or "none",
               "" if not (unlocked or relocked) else
               " - %d newly unlocked, %d relocked" % (unlocked, relocked)))

    # ------------------------------------------------ G3  workbook structure
    log.head("G3  the workbook structure stays locked with the same password")
    sec = WorkbookProtection(lockStructure=True)
    sec.workbookPassword = PW
    wb.security = sec
    log("G3", "workbook", "structure locked - tabs cannot move, be renamed or "
                          "be unhidden")

    # ------------------------------------------------------------- BLD-18
    log.head("BLD-18  a filled person does not carry the lever Hire")
    pairs = stale_pairs(wb, vals)
    for sheet, r, co, name, title in pairs:
        wb[sheet][co].value = "Filled"
        log("BLD-18", "%s!%s" % (sheet, co),
            "%s, %s - filled, lever Hire -> Filled (both price at factor 1, "
            "so the cost cannot move)" % (name, title))
    if not pairs:
        log.note("BLD-18", "no filled person carries Hire - nothing to do")
    for sheet, co, v in stray_levers(wb, lever_cells(wb)):
        log.note("BLD-18", "%s!%s holds %r with no dropdown over it - reported, "
                           "not edited" % (sheet, co, v))

    # ------------------------------------------------------------- BLD-19
    log.head("BLD-19  the cell that pulled a dash off his budget table reads 0.00")
    pulls = dash_pullers(wb, vals)
    for sheet, co, f in pulls:
        c = wb[sheet][co]
        c.value = "=N(" + f[1:].strip() + ")"
        old = c.number_format
        if c.number_format != MONEY:
            c.number_format = MONEY
        log("BLD-19", "%s!%s" % (sheet, co),
            "%s -> %s, format %s%s - his 0.1 cell is left exactly as he typed it"
            % (f, c.value, MONEY, "" if old == MONEY else " (was %s)" % old))
    if not pulls:
        log.note("BLD-19", "no 1.x cell prints his typed dash - nothing to do")

    # ------------------------------------------------------ build and check
    tmp = dst + ".raw"
    save(wb, tmp)
    log.head("recalculating and writing the cached values back")
    rc, st = wbio.build(tmp, dst)
    os.remove(tmp)
    print("recalculated, %d formula cells populated across %d sheets"
          % (st["cells"], st["sheets"]), flush=True)
    err, _ = wbio.audit(dst)
    err, hisna = real_errors(err)
    if hisna:
        print("his own typed #N/A text inside the raw block, left verbatim: "
              "%d %s" % (len(hisna), sorted(hisna)))
    if err:
        print("STOP: %d error cells, e.g. %r" % (len(err), err[:5]))
        raise SystemExit(2)

    self_check(dst, vals, pairs, pulls)
    log.tail()
    print("wrote", dst)


# --------------------------------------------------------------- self-check

def self_check(dst, before, pairs, pulls):
    print("\n== self-check (everything recomputed from the written file)",
          flush=True)
    wb = openpyxl.load_workbook(dst)
    after = cached(dst)
    checks = []

    def ck(name, ok, detail):
        checks.append(ok)
        print("%s  %-54s %s" % ("PASS" if ok else "FAIL", name, detail),
              flush=True)

    # 1 - the scope, listed tab by tab
    want = sorted(ws.title for ws in wb.worksheets if protected(ws.title))
    got = sorted(ws.title for ws in wb.worksheets if ws.protection.sheet)
    ck("protected tabs are exactly the 0.x and 3.x set", want == got,
       "%d protected: %s" % (len(got), ", ".join(got) or "none"))
    print("       inventory, every tab in the book:", flush=True)
    for ws in wb.worksheets:
        p = ws.protection
        print("       %-34s %-9s %s" % (
            ws.title, ws.sheet_state,
            "PROTECTED password %s" % p.password if p.sheet
            else "unprotected%s" % ("" if p.password is None
                                    else " BUT PASSWORD %s" % p.password)),
              flush=True)

    bad = [ws.title for ws in wb.worksheets if protected(ws.title)
           and not (ws.protection.password == PWH and ws.protection.sort
                    and ws.protection.autoFilter
                    and not ws.protection.selectLockedCells
                    and not ws.protection.selectUnlockedCells)]
    ck("every protected tab carries the password Tdd123", not bad,
       "hash %s, sort and filter off, selecting cells allowed%s"
       % (PWH, "" if not bad else "; wrong on %r" % bad))

    loose = [ws.title for ws in wb.worksheets if not protected(ws.title)
             and (ws.protection.sheet or ws.protection.password is not None)]
    named = [t for t in ("Exec Summary", REVIEW, "Lists", "4.0 Data QA")
             if t in wb.sheetnames and not wb[t].protection.sheet]
    ck("everything else carries no protection and no password", not loose,
       "%d unprotected tabs, %s free%s"
       % (sum(1 for ws in wb.worksheets if not ws.protection.sheet),
          " / ".join(named), "" if not loose else "; still set on %r" % loose))

    sec = wb.security
    ck("workbook structure still locked",
       sec is not None and bool(sec.lockStructure)
       and sec.workbookPassword == PWH,
       "lockStructure with the same password")

    # 2 - what stays typeable behind the protection
    allow = allowmap(wb)
    mism = []
    for title in want:
        a = allow[title]
        un = {c.coordinate for row in wb[title].iter_rows() for c in row
              if c.protection.locked is False}
        if un != set(a):
            mism.append((title, sorted(un - set(a))[:4],
                         sorted(set(a) - un)[:4]))
        by = collections.Counter(a.values())
        print("       %-34s %s = %d unlocked"
              % (title, "  ".join("%s %d" % (k, by[k]) for k in CLASSES
                                  if by[k]) or "no typed inputs", len(un)),
              flush=True)
    ck("unlocked cells on the protected tabs are the typed inputs", not mism,
       "cream and toggles only%s" % ("" if not mism else "; %r" % mism[:2]))

    lo = [ws for ws in wb.worksheets if LIGHTS in ws.title]
    rows = []
    ok = bool(lo)
    for ws in lo:
        tg = toggles(ws)
        lk = [co for co in sorted(tg) if ws[co].protection.locked is not False]
        ok = ok and bool(tg) and not lk and ws.protection.sheet \
            and not ws.protection.selectUnlockedCells
        rows.append("%s: %d toggles %s, %d locked"
                    % (ws.title, len(tg),
                       ",".join(sorted(tg, key=lambda c: ws[c].row)[:3]) + "...",
                       len(lk)))
    ck("the Lights On toggles are editable on the protected tab", ok,
       "; ".join(rows) or "no Lights On tab in the book")

    # 3 - BLD-18
    left = stale_pairs(wb, after)
    ck("no filled person carries the lever Hire", not left,
       "%d pairs normalised, %d left%s"
       % (len(pairs), len(left), "" if not left else " %r" % left[:3]))
    for sheet, r, co, name, title in pairs:
        print("       %-30s %-5s %-26s %-40s now %s"
              % (sheet, co, name, title, wb[sheet][co].value), flush=True)
    ck("the normalised levers read Filled in the written file",
       all(wb[s][c].value == "Filled" for s, _, c, _, _ in pairs),
       "%d cells" % len(pairs))

    # 4 - BLD-19
    good = []
    for sheet, co, f in pulls:
        c, v = wb[sheet][co], after.get(sheet, {}).get(co)
        good.append(c.value == "=N(" + f[1:].strip() + ")"
                    and isinstance(v, (int, float)) and abs(v) < TOL
                    and c.number_format == MONEY)
        print("       %-30s %-5s %s -> %r, format %s"
              % (sheet, co, c.value, v, c.number_format), flush=True)
    ck("the dash cell holds a number and prints 0.00", all(good),
       "%d cell(s) wrapped in N(), value 0, zero section prints 0.00"
       % len(pulls))

    # 5 - nothing else moved: the only cells allowed to differ from the input
    # are the levers this stage set and the cell it wrapped
    moved, nums = [], 0
    expect = {(s, c) for s, c, _ in pulls} | {(s, c) for s, _, c, _, _ in pairs}
    for sheet, d in after.items():
        b = before.get(sheet, {})
        for co, v in d.items():
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                nums += 1
            if (sheet, co) in expect:
                continue
            o = b.get(co)
            if isinstance(v, (int, float)) and isinstance(o, (int, float)):
                if abs(v - o) > 1e-6:
                    moved.append((sheet, co, o, v))
            elif o != v:
                moved.append((sheet, co, o, v))
        for co in set(b) - set(d):
            moved.append((sheet, co, b[co], None))
    ck("zero cost effect - every other value in the book is unchanged",
       not moved, "%d cells compared, %d of them numbers; only the %d levers "
                  "and the %d wrapped cell differ from the input%s"
       % (sum(len(d) for d in after.values()), nums, len(pairs), len(pulls),
          "" if not moved else "; %d others moved e.g. %r"
          % (len(moved), moved[:4])))

    err, _ = wbio.audit(dst)
    err, hisna = real_errors(err)
    ck("no error cells beyond his own typed #N/A raw-block text", not err,
       "%d errors; his verbatim #N/A cells: %s" % (len(err), sorted(hisna)))

    wb.close()
    if not all(checks):
        print("\nself-check FAILED")
        raise SystemExit(2)
    print("\nself-check clean: %d/%d PASS" % (sum(checks), len(checks)),
          flush=True)


if __name__ == "__main__":
    main(*sys.argv[1:3])
