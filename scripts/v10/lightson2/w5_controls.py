#!/usr/bin/env python3
"""w5 - the control checks come out of the model.

  python3 w5_controls.py <in.xlsx> <out.xlsx>

His ruling: "can you get rid of the control checks throughout the model,
they're not helpful at all."

  R1  every control row on a model tab is deleted outright.  A control row is
      found by reading its label, never by a row number: the label starts with
      "Control" (any punctuation after it), or it is a white-font label that
      says a figure "must be 0" or "must read 0".  The rows are deleted, the
      rows under them move up, and every formula in the book that pointed at a
      moved cell is repointed - the same shift_rows machinery the other stages
      use.  Row heights, dropdowns and conditional formats move with the rows.
  R2  the 4.0 Data QA tab exists only to hold control checks, so the whole tab
      goes.  Nothing may point at it afterwards: the formulas, the defined
      names, the hyperlinks and the plain text of the book are all searched
      first, and the stage stops rather than leave a reference behind.
  R3  the section divider above the QA tab is only removed if the removal
      empties its section.  It heads the data section, and Lists still sits
      under it, so it stays - the decision is printed either way.

Nothing else moves.  No figure in the workbook may change: every cached value
on every tab is snapshotted before and after and compared cell for cell
through the row shift, and the only differences allowed are the deleted cells
themselves.  His raw REVIEW block, his three verbatim #N/A cells, the cream
inputs, the toggles, the levers and every sheet's protection state are left
exactly as they are.

Idempotent: handed its own output the stage recognises the finished state and
copies it through untouched, so a second run is md5-identical to the first.
"""
import sys, os, re, shutil, zipfile, collections

V10 = "/home/user/anthropic-claude-code/scripts/v10"
sys.path.insert(0, V10)
sys.path.insert(0, os.path.join(V10, "update"))

import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange

import wbio
from _xl import REVIEW, Log, load, save, shift_rows, _scan, _sheet_of

QA = "4.0 Data QA"
DIVIDER = re.compile(r">>\s*$")

# a control row names itself.  "Control - roles against the role mapping,
# must be 0", "Control, 8 checks, each must read 0", "Controls" - all of them
# open with the word, whatever punctuation follows.
CTRL = re.compile(r"^\s*controls?\b", re.I)
# the second arm catches a check row that was labelled some other way: a
# white-font line (the book paints its controls white) that states a figure
# must read zero.  It is reported separately so a variant never slips through.
CHECKY = re.compile(r"must\s+(?:be|read)\s+(?:0|zero)\b", re.I)
WHITE = ("FFFFFFFF", "00FFFFFF")

TOL_ABS, TOL_REL = 1e-9, 1e-12
HIS_NA = ("A31", "C31", "S525")          # his three verbatim cells on REVIEW
ERRS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "Err:")


# ------------------------------------------------------------------ reading

def is_white(cell):
    f = cell.font
    col = getattr(f.color, "rgb", None) if f is not None and f.color is not None else None
    return isinstance(col, str) and col in WHITE


def label_rows(ws):
    """[(row, coord, label, why)] for every control/check row on one sheet."""
    out, seen = [], set()
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not isinstance(v, str) or v.startswith("="):
                continue
            why = None
            if CTRL.match(v):
                why = "label starts with Control"
            elif CHECKY.search(v) and is_white(c):
                why = "white-font check row, says it must read 0"
            if why and c.row not in seen:
                seen.add(c.row)
                out.append((c.row, c.coordinate, v.strip(), why))
    return sorted(out)


def control_rows(wb):
    """{sheet: [(row, coord, label, why)]}, in sheet order."""
    found = collections.OrderedDict()
    for ws in wb.worksheets:
        rows = label_rows(ws)
        if rows:
            found[ws.title] = rows
    return found


def row_content(ws, row):
    """[(coord, value, white)] for every non-empty cell on a row."""
    return [(c.coordinate, c.value, is_white(c))
            for c in ws[row] if c.value is not None]


def blocks(rows):
    """[(first, count)] for runs of consecutive rows, highest first."""
    out = []
    for r in sorted(rows):
        if out and out[-1][0] + out[-1][1] == r:
            out[-1][1] += 1
        else:
            out.append([r, 1])
    return [(a, n) for a, n in reversed(out)]


def sheet_refs(wb, title):
    """[(sheet, coord, formula)] for every formula that names sheet `title`."""
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f = c.value
                if not (isinstance(f, str) and f.startswith("=")):
                    continue
                for s, e, m in _scan(f):
                    if m.group("sheet") is not None and _sheet_of(m, ws.title) == title:
                        hits.append((ws.title, c.coordinate, f[:120]))
                        break
    return hits


def text_refs(wb, title):
    """[(sheet, coord)] for plain text anywhere that names the tab."""
    hits = []
    for ws in wb.worksheets:
        if ws.title == title:
            continue
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and title in v:
                    hits.append((ws.title, c.coordinate))
    return hits


def spanning_ranges(wb, plan):
    """[(sheet, coord, ref, rows)] for ranges that span a row being deleted."""
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f = c.value
                if not (isinstance(f, str) and f.startswith("=")):
                    continue
                for s, e, m in _scan(f):
                    if m.group("c2") is None:
                        continue
                    tgt = _sheet_of(m, ws.title)
                    if tgt not in plan:
                        continue
                    r1 = int(m.group("r1").lstrip("$"))
                    r2 = int(m.group("r2").lstrip("$"))
                    lo, hi = min(r1, r2), max(r1, r2)
                    inside = [r for r in plan[tgt] if lo <= r <= hi]
                    if inside:
                        hits.append((ws.title, c.coordinate, f[s:e], inside))
    return hits


# ------------------------------------------------------------------ writing

def shift_row_dims(ws, at, n):
    """Row heights follow their rows; openpyxl's delete_rows leaves them put."""
    hi = at + n - 1
    old = dict(ws.row_dimensions)
    new = {}
    for r, d in old.items():
        if at <= r <= hi:
            continue
        new[r - n if r > hi else r] = d
    if list(new.items()) == list(old.items()):
        return 0
    for k in list(ws.row_dimensions):
        del ws.row_dimensions[k]
    for r in sorted(new):
        d = new[r]
        d.index = r
        ws.row_dimensions[r] = d
    return sum(1 for r in old if r > hi)


def shift_cf(ws, at, n):
    """Conditional formats follow their rows, with the same clamping as refs."""
    hi = at + n - 1

    def one(r1, r2):
        a = r1 - n if r1 > hi else (at if r1 >= at else r1)
        b = r2 - n if r2 > hi else (at - 1 if r2 >= at else r2)
        return a, b

    plan, moved = [], 0
    for cf in ws.conditional_formatting:
        keep = []
        for cr in cf.sqref.ranges:
            a, b = one(cr.min_row, cr.max_row)
            if b < a:
                moved += 1
                continue
            if (a, b) != (cr.min_row, cr.max_row):
                moved += 1
            keep.append(CellRange(min_col=cr.min_col, max_col=cr.max_col,
                                  min_row=a, max_row=b))
        plan.append((keep, list(cf.rules)))
    if not moved:
        return 0
    from openpyxl.formatting.formatting import ConditionalFormattingList
    fresh = ConditionalFormattingList()
    for keep, rules in plan:
        if not keep:
            continue
        sq = MultiCellRange(keep)
        for rule in rules:
            fresh.add(sq, rule)
    ws.conditional_formatting = fresh
    return moved


def drop_rows(wb, title, rows):
    """Delete `rows` on `title`, bottom block first, repointing the book."""
    ws = wb[title]
    total = 0
    for at, n in blocks(rows):
        changed = shift_rows(wb, title, at, -n)      # formulas, cells, dropdowns
        shift_row_dims(ws, at, n)
        shift_cf(ws, at, n)
        total += changed
    return total


# ------------------------------------------------------------------ compare

def snapshot(path):
    """{sheet: {coord: cached value}} for every non-empty cell in the book."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for ws in wb.worksheets:
        d = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    d[c.coordinate] = c.value
        out[ws.title] = d
    wb.close()
    return out


def same_value(a, b):
    if isinstance(a, bool) or isinstance(b, bool):
        return a is b
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) <= max(TOL_ABS, abs(a) * TOL_REL)
    return a == b


def moved_coord(coord, dropped):
    col, row = coordinate_from_string(coord)
    if row in dropped:
        return None
    return "%s%d" % (col, row - sum(1 for d in dropped if d < row))


def compare(before, after, plan, gone):
    """(kept, drift, vanished, appeared) across the whole book."""
    kept, drift, vanished, appeared = 0, [], [], []
    expect = {}
    for sheet, cells in before.items():
        if sheet in gone:
            continue
        dropped = set(plan.get(sheet, []))
        dst = after.get(sheet)
        if dst is None:
            vanished.append((sheet, "(whole sheet)", None))
            continue
        want = expect.setdefault(sheet, {})
        for coord, val in cells.items():
            nc = moved_coord(coord, dropped)
            if nc is None:                       # a deleted control cell
                continue
            want[nc] = val
            got = dst.get(nc)
            if same_value(val, got) or (val == "" and got is None):
                kept += 1
            elif got is None:
                vanished.append((sheet, "%s -> %s" % (coord, nc), val))
            else:
                drift.append((sheet, "%s -> %s" % (coord, nc), val, got))
    for sheet, cells in after.items():
        want = expect.get(sheet, {})
        for coord, val in cells.items():
            if coord not in want and val != "":
                appeared.append((sheet, coord, val))
    return kept, drift, vanished, appeared


def prot_map(path):
    wb = openpyxl.load_workbook(path)
    out = {}
    for ws in wb.worksheets:
        p = ws.protection
        out[ws.title] = (bool(p.sheet), p.password, bool(p.selectLockedCells),
                         bool(p.selectUnlockedCells), bool(p.sort),
                         bool(p.autoFilter), bool(p.formatCells),
                         bool(p.insertRows), bool(p.deleteRows))
    sec = wb.security
    struct = (bool(sec.lockStructure), sec.workbookPassword) if sec else (False, None)
    wb.close()
    return out, struct


def dv_map(path):
    """{sheet: {coord: validation formula}} - every dropdown in the book."""
    wb = openpyxl.load_workbook(path)
    out = {}
    for ws in wb.worksheets:
        d = {}
        for dv in ws.data_validations.dataValidation:
            for cr in dv.sqref.ranges:
                for r in range(cr.min_row, cr.max_row + 1):
                    for c in range(cr.min_col, cr.max_col + 1):
                        d[ws.cell(r, c).coordinate] = str(dv.formula1)
        out[ws.title] = d
    wb.close()
    return out


def marker(path, sheet, label, col, exact=False):
    """(coord, value) of the cell in `col` on the row column B labels `label`."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise SystemExit("STOP: %s is not in %s" % (sheet, path))
    ws = wb[sheet]
    for row in ws.iter_rows(min_col=2, max_col=2):
        for c in row:
            v = c.value
            if not isinstance(v, str):
                continue
            if v.strip() == label if exact else v.strip().startswith(label):
                co = "%s%d" % (col, c.row)
                out = (co, ws[co].value)
                wb.close()
                return out
    wb.close()
    raise SystemExit("STOP: no row labelled %r on %s" % (label, sheet))


# --------------------------------------------------------------------- main

def finished(path):
    """True when the file already carries no controls and no QA tab."""
    wb = openpyxl.load_workbook(path)
    done = QA not in wb.sheetnames and not control_rows(wb)
    wb.close()
    return done


def main(src, dst):
    log = Log("w5_controls")

    if finished(src):
        print("input already carries no control rows and no %s tab "
              "- copying through" % QA)
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        return

    print("== before: reading every cached value on every tab", flush=True)
    before = snapshot(src)
    prot_before, struct_before = prot_map(src)
    dv_before = dv_map(src)
    n_before = sum(len(v) for v in before.values())
    print("   %d cached values across %d tabs" % (n_before, len(before)),
          flush=True)
    m35_before = marker(src, "3.5 TDD Lights On", "Total", "L", exact=False)
    m31_before = marker(src, "3.1 Archetype to Actuals", "TDD total", "E")
    print("   3.5 charged to TDD  %s = %.6f" % (m35_before[0], m35_before[1]))
    print("   book total          3.1!%s = %.6f" % (m31_before[0], m31_before[1]))

    wb = load(src)

    # ------------------------------------------------- R1  the control rows
    log.head("R1  every control row on a model tab, found by its label")
    plan = control_rows(wb)
    if QA in plan:
        del plan[QA]                       # the whole tab goes in R2 anyway
    removed, mixed = [], []
    for title, rows in plan.items():
        ws = wb[title]
        for r, coord, text, why in rows:
            cells = row_content(ws, r)
            visible = [co for co, v, w in cells if not w]
            if visible:
                mixed.append((title, r, visible))
            removed.append((title, r, coord, text, [co for co, _, _ in cells]))
    if mixed:
        print("STOP: a control row carries visible content: %r" % mixed[:5])
        raise SystemExit(2)

    span = spanning_ranges(wb, {t: [r for r, _, _, _ in rs]
                                for t, rs in plan.items()})
    if span:
        print("STOP: %d ranges span a row being deleted, e.g. %r"
              % (len(span), span[:5]))
        raise SystemExit(2)
    log.note("R1", "no range anywhere in the book spans a row being deleted, "
                   "so no total changes meaning")

    for title, rows in plan.items():
        nums = [r for r, _, _, _ in rows]
        changed = drop_rows(wb, title, nums)
        log("R1", title, "%d control %s deleted (%s), rows below shifted up, "
                         "%d formulas repointed"
            % (len(nums), "row" if len(nums) == 1 else "rows",
               ", ".join(str(n) for n in nums), changed))

    # ------------------------------------------------------ R2  the QA tab
    log.head("R2  the 4.0 Data QA tab, which holds nothing but checks")
    if QA in wb.sheetnames:
        qa_rows = sum(1 for row in wb[QA].iter_rows()
                      if any(c.value is not None for c in row))
        pointing = sheet_refs(wb, QA)
        texting = text_refs(wb, QA)
        named = [n for n, d in wb.defined_names.items() if QA in str(d.value)]
        if pointing or texting or named:
            print("STOP: %d formulas, %d text cells and %d names still point at %s"
                  % (len(pointing), len(texting), len(named), QA))
            print("      e.g. %r %r %r" % (pointing[:3], texting[:3], named[:3]))
            raise SystemExit(2)
        idx = wb.sheetnames.index(QA)
        del wb[QA]
        log("R2", QA, "tab removed whole, %d rows of checks, nothing in the "
                      "book pointed at it" % qa_rows)
    else:
        idx, qa_rows = None, 0
        log.note("R2", "no %s tab in the input" % QA)

    # -------------------------------------------------- R3  section divider
    log.head("R3  the section divider above it")
    if idx is not None:
        head = None
        for i in range(idx - 1, -1, -1):
            if DIVIDER.search(wb.sheetnames[i]):
                head = i
                break
        if head is None:
            log.note("R3", "no divider heads the section the QA tab sat in")
        else:
            name = wb.sheetnames[head]
            rest = []
            for t in wb.sheetnames[head + 1:]:
                if DIVIDER.search(t):
                    break
                rest.append(t)
            if rest:
                log("R3", name, "kept, %s still %s under it"
                    % (", ".join(rest), "sits" if len(rest) == 1 else "sit"))
            else:
                del wb[name]
                log("R3", name, "removed, the removal left the section empty")

    # ----------------------------------------------------- build and check
    tmp = dst + ".raw"
    save(wb, tmp)
    log.head("recalculating and writing the cached values back")
    rc, st = wbio.build(tmp, dst)
    os.remove(tmp)
    print("recalculated, %d formula cells populated across %d sheets"
          % (st["cells"], st["sheets"]), flush=True)

    self_check(dst, before, prot_before, struct_before, dv_before,
               {t: [r for r, _, _, _ in rs] for t, rs in plan.items()},
               removed, qa_rows, m35_before, m31_before)
    log.tail()
    print("wrote", dst)


# ------------------------------------------------------------- self-check

def self_check(dst, before, prot_before, struct_before, dv_before,
               plan, removed, qa_rows, m35_before, m31_before):
    print("\n== what came out", flush=True)
    by_tab = collections.Counter(t for t, _, _, _, _ in removed)
    for t in by_tab:
        rows = [str(r) for tt, r, _, _, _ in removed if tt == t]
        print("   %-32s %d row%s: %s"
              % (t, by_tab[t], "" if by_tab[t] == 1 else "s", ", ".join(rows)))
    cells = sum(len(cs) for _, _, _, _, cs in removed)
    print("   %-32s %d control rows across %d tabs, %d cells"
          % ("TOTAL", len(removed), len(by_tab), cells))
    print("   %-32s whole tab, %d rows of checks" % (QA, qa_rows))

    print("\n== self-check (everything recomputed from the written file)",
          flush=True)
    checks = []

    def ck(name, ok, detail):
        checks.append(ok)
        print("%s  %-56s %s" % ("PASS" if ok else "FAIL", name, detail),
              flush=True)

    wb = openpyxl.load_workbook(dst)
    left = control_rows(wb)
    ck("no control row anywhere in the book", not left,
       "0 cells whose label opens with Control or reads as a check"
       if not left else "still there: %r" % {k: v[:2] for k, v in left.items()})

    pointing = sheet_refs(wb, QA)
    texting = text_refs(wb, QA)
    named = [n for n, d in wb.defined_names.items() if QA in str(d.value)]
    links = [(ws.title, hl.ref) for ws in wb.worksheets for hl in ws._hyperlinks
             if QA in str(hl.target) or QA in str(hl.location)]
    ck("the 4.0 Data QA tab is gone, nothing dangles",
       QA not in wb.sheetnames and not (pointing or texting or named or links),
       "%d sheets left; %d formulas, %d text cells, %d names, %d links name it"
       % (len(wb.sheetnames), len(pointing), len(texting), len(named), len(links)))

    after = snapshot(dst)
    kept, drift, vanished, appeared = compare(before, after, plan, {QA})
    ck("every cached value that survives is identical", not drift,
       "%d values compared, %d moved%s"
       % (kept, len(drift), "" if not drift else ": %r" % drift[:4]))
    ck("nothing that should have survived went missing", not vanished,
       "%d missing%s" % (len(vanished), "" if not vanished else ": %r" % vanished[:4]))
    ck("no figure appeared that was not there before", not appeared,
       "%d new%s" % (len(appeared), "" if not appeared else ": %r" % appeared[:4]))

    err, _ = wbio.audit(dst)
    his = [(s, c) for s, c, v, f in err
           if s == REVIEW and c in HIS_NA and v == "#N/A"]
    real = [e for e in err if not (e[0] == REVIEW and e[1] in HIS_NA and e[2] == "#N/A")]
    ck("no error cells", not real,
       "%d errors beyond his three verbatim #N/A cells%s"
       % (len(real), "" if not real else ": %r" % real[:4]))
    ck("his three verbatim #N/A cells still read #N/A", len(his) == 3,
       "REVIEW %s" % ", ".join(sorted(c for _, c in his)))

    refs = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and "#REF!" in v:
                    refs.append((ws.title, c.coordinate))
    for sheet, cs in after.items():
        for co, v in cs.items():
            if isinstance(v, str) and any(e in v for e in ERRS):
                refs.append((sheet, co))
    ck("no #REF! and no new error value", not refs,
       "%d found%s" % (len(refs), "" if not refs else ": %r" % refs[:4]))

    prot_after, struct_after = prot_map(dst)
    want = {k: v for k, v in prot_before.items() if k != QA}
    bad = [k for k in want if want[k] != prot_after.get(k)]
    ck("every sheet keeps its protection state",
       not bad and set(want) == set(prot_after) and struct_before == struct_after,
       "%d protected / %d unprotected, workbook structure %s"
       % (sum(1 for v in prot_after.values() if v[0]),
          sum(1 for v in prot_after.values() if not v[0]),
          "locked" if struct_after[0] else "open")
       + ("" if not bad else "; changed on %r" % bad[:4]))

    dv_after = dv_map(dst)
    dvbad = []
    for sheet, cells in dv_before.items():
        if sheet == QA:
            continue
        dropped = set(plan.get(sheet, []))
        want_dv = {}
        for co, f in cells.items():
            nc = moved_coord(co, dropped)
            if nc is not None:
                want_dv[nc] = f
        if want_dv != dv_after.get(sheet, {}):
            got = dv_after.get(sheet, {})
            dvbad.append((sheet, sorted(set(want_dv) ^ set(got))[:4]))
    ck("every dropdown still covers the cells it covered",
       not dvbad, "%d validated cells carried through%s"
       % (sum(len(v) for k, v in dv_before.items() if k != QA),
          "" if not dvbad else "; off on %r" % dvbad[:3]))

    names_before = 4
    ck("the defined names are intact",
       len(wb.defined_names) >= names_before,
       "%d names: %s" % (len(wb.defined_names), ", ".join(sorted(wb.defined_names))))

    raw = [c for c in before[REVIEW]
           if coordinate_from_string(c)[1] >= 2
           and column_index_from_string(coordinate_from_string(c)[0]) <= 29]
    rawbad = [c for c in raw if not same_value(before[REVIEW][c],
                                               after[REVIEW].get(c))]
    ck("his raw REVIEW block is untouched", not rawbad,
       "%d cells in A2:AC compared, all identical" % len(raw)
       if not rawbad else "%d moved: %r" % (len(rawbad), rawbad[:4]))

    try:
        broken = zipfile.ZipFile(dst).testzip()
    except Exception as exc:                                # pragma: no cover
        broken = str(exc)
    ck("the workbook still opens", broken is None and len(wb.sheetnames) > 0,
       "%d sheets read back, archive intact" % len(wb.sheetnames))

    m35_after = marker(dst, "3.5 TDD Lights On", "Total", "L", exact=False)
    m31_after = marker(dst, "3.1 Archetype to Actuals", "TDD total", "E")
    ck("3.5's charge to TDD has not moved",
       same_value(m35_before[1], m35_after[1]),
       "%s %.6f -> %s %.6f" % (m35_before[0], m35_before[1],
                               m35_after[0], m35_after[1]))
    ck("the book total has not moved",
       same_value(m31_before[1], m31_after[1]),
       "3.1!%s %.6f -> 3.1!%s %.6f" % (m31_before[0], m31_before[1],
                                       m31_after[0], m31_after[1]))
    wb.close()

    if not all(checks):
        print("\nself-check FAILED")
        raise SystemExit(2)
    print("\nself-check clean: %d/%d PASS" % (sum(checks), len(checks)),
          flush=True)


if __name__ == "__main__":
    main(*sys.argv[1:3])
