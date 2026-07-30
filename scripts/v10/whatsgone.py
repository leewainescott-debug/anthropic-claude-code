"""What of his did not make it into the build.

He asked a fair question - "what else have i not requested that you've got rid of?" - and
the only honest way to answer it is mechanically, not from memory.

Method: take every non-empty cell he typed in his own workbooks, and ask whether that
content still appears anywhere on the same tab of the build. Not the same cell - rows
moved, tables were relocated, the 1.x actuals block went from the foot of the tab to the
top - so a cell-by-cell diff would report hundreds of false losses. Content survival is
the question that matters: is what he wrote still in the file.

Typed content only. Formulas are the model's plumbing and are rewritten by design; a
formula that changed is a decision, and the decisions are in DECISIONS.md. A label or a
number he typed that is simply gone is a loss, and that is what this prints.

    python3 whatsgone.py base_2707.xlsx cand_A.xlsx
    python3 whatsgone.py base_2707.xlsx cand_A.xlsx --later rev.xlsx

With --later, a second workbook of his that came after the first, the output splits in two:
what his own later book already replaced, which is not a loss, and what the build changed,
which is the list he actually wants.
"""
import re
import sys
from collections import defaultdict

import openpyxl

# tabs the chain builds from scratch: their labels are the model's, not his
BUILT = re.compile(r"^(2\.\d+ |3\.\d+ |4\.0|Exec |Lists$|REVIEW)")

# Content of his that a ruling of his has since replaced. Every entry names the cell, quotes
# enough of what he typed to find it, and says which ruling took it out. Splitting these off
# is the point: "green" on this report means every loss is declared, not that the list is
# empty - his own decisions remove his own earlier content, and a report that cannot tell
# those two apart is a report nobody reads twice.
#
# Wave M, his cyber uplift and Customer rulings:
DECLARED = [
    ("0.2 Data Config", "I7", "Separate out and incl. cyber uplift",
     "his own action note, and this is the pass that actions it - the COE and TDD Cyber "
     "are funded on two rows now, and his offshoring note has moved onto the COE row"),
    ("0.2 Data Config", "B23", "TDD Cyber incl. COE",
     "the line is split in two: row 7 is the COE, row 23 is TDD Cyber, so the label no "
     "longer says 'incl. COE'"),
    ("0.2 Data Config", "C23", 2.5,
     "his 3.5 TDD Cyber allocation splits 2.0 to the COE row and 1.5 to TDD Cyber; the "
     "total he allocated is unchanged at 50.5"),
    ("0.2 Data Config", "D23", 1,
     "same split - the NZ half of his 3.5 allocation now sits on two rows"),
    ("1.2 Customer", "I54", 2.21,
     "his 27/07 typed figure for EGI Customer. He has since ruled that the EGI squads are "
     "funded by EGI at the actual cost of their roles, so the cell is his review "
     "workbook's formula again and the actual flows to Significant Items EGI"),
    ("1.13 Cyber Roles", "C13", 0.5,
     "the Cyber CapEx input, removed on his ruling with the bucket table it sat in"),
    ("1.13 Cyber Roles", "B13", "Cyber CapEx - Monitoring",
     "the label of the input above"),
    ("1.13 Cyber Roles", "B12", "TDD Cyber budget ($m)",
     "1.13 draws on the COE allocation only now - TDD Cyber is funded on 1.14, and a "
     "budget read on both tabs was the double count his ruling removes"),
    ("1.13 Cyber Roles", "B15", "Left to fund ($m)",
     "the seventh column of the summary is a Variance now, budget less spend, which is the "
     "figure he asked for; a funding gap stated twice was what made it two different things"),
    ("1.13 Cyber Roles", "H5", "Left to fund ($m)",
     "same ruling - the column heading its two sibling tabs already carry"),
    ("1.13 Cyber Roles", "B10", "Funding buckets to draw down",
     "there are two funding lines now, not four buckets"),
    ("1.13 Cyber Roles", "E11", "Planned spend less CapEx",
     "the label of the CapEx working figure, retired with it"),
    ("1.13 Cyber Roles", "B73", "52 roles",
     "the role count inside his own sentence. Nine of the 52 moved to TDD Cyber on his "
     "ruling, so the sentence would otherwise contradict the table above it - the same "
     "typo-class exception D114 lists, and the only part of the sentence that changes"),
]


def declared(tab, coord, value):
    """The ruling that removed this, or None."""
    for t, ref, want, why in DECLARED:
        if t != tab or ref != coord:
            continue
        if isinstance(want, str):
            if str(want).lower() in str(value).lower():
                return why
        elif norm(want) == norm(value):
            return why
    return None


def typed(ws, fws=None):
    """Every value he typed on this tab, as (coordinate, value).

    With fws - the same tab read formulas-first - a cell whose formula view is a formula
    is dropped: its value is something the model computed, not something he wrote, and it
    is meant to move when the data moves."""
    out = []
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if fws is not None:
                f = fws[c.coordinate].value
                if isinstance(f, str) and f.startswith("="):
                    continue
            if isinstance(v, str):
                v = v.strip()
                if not v or v.startswith("="):
                    continue
            out.append((c.coordinate, v))
    return out


def norm(v):
    """Compare the way a reader would: case and spacing do not count, an en dash and a
    hyphen are the same character to anyone reading the page, and a number is its value."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return round(float(v), 4)
    s = str(v).lower()
    for d in "‐‑‒–—―":
        s = s.replace(d, "-")
    return re.sub(r"\s+", " ", s).strip()


def run(his_path, build_path):
    a = openpyxl.load_workbook(his_path, data_only=True)
    af = openpyxl.load_workbook(his_path)
    b = openpyxl.load_workbook(build_path, data_only=True)
    gone = defaultdict(list)
    tabs_gone = []
    for t in a.sheetnames:
        if t not in b.sheetnames:
            tabs_gone.append(t)
            continue
        if BUILT.match(t):
            continue
        # the build's side keeps computed values: his typed 5.5 is not lost if the build
        # now computes 5.5 in the same place
        have = {norm(v) for _, v in typed(b[t])}
        for coord, v in typed(a[t], af[t]):
            if norm(v) not in have:
                gone[t].append((coord, v))
    return gone, tabs_gone


def later_has(later_path, build_path):
    """Every value typed in his later workbook, per tab, for the split below."""
    lat = openpyxl.load_workbook(later_path, data_only=True)
    bld = openpyxl.load_workbook(build_path, read_only=True)
    out = {}
    for t in lat.sheetnames:
        if t in bld.sheetnames:
            out[t] = {norm(v) for _, v in typed(lat[t])}
    return out


def report(gone, build, later=None, cap=40):
    """One line per distinct value, most-repeated first, with what the build says now."""
    b = openpyxl.load_workbook(build, data_only=True)
    seen = defaultdict(list)
    for t, items in gone.items():
        for coord, v in items:
            seen[str(v)].append((t, coord))
    mine, superseded, ruled = [], [], []
    for v, places in sorted(seen.items(), key=lambda kv: -len(kv[1])):
        # a ruling of his took it out -> declared, and named
        why = next((w for t, c in places if (w := declared(t, c, v))), None)
        if why:
            ruled.append((v, places, why))
            continue
        # his later book still says it and the build does not -> the build changed it
        bucket = mine
        if later is not None and not any(norm(v) in later.get(t, ()) for t, _ in places):
            bucket = superseded
        bucket.append((v, places))
    return mine, superseded, ruled, b


if __name__ == "__main__":
    his, build = sys.argv[1], sys.argv[2]
    later = None
    if "--later" in sys.argv:
        later = later_has(sys.argv[sys.argv.index("--later") + 1], build)
    gone, tabs_gone = run(his, build)
    n = sum(len(v) for v in gone.values())
    print(f"{his} -> {build}")
    print(f"{n} typed values on his own tabs are not in the build, across {len(gone)} tabs")
    if tabs_gone:
        print(f"tabs of his that are not in the build: {tabs_gone}")
    cap = 40
    mine, superseded, ruled, b = report(gone, build, later)
    print(f"\n=== a ruling of his replaced it - declared, not a loss: "
          f"{sum(len(pl) for _v, pl, _w in ruled)} cells, {len(ruled)} distinct ===")
    for v, places, why in ruled:
        t, coord = places[0]
        print(f"{len(places):>3} x  {str(v)[:44]:<44} | {t}!{coord}\n          {why}")
    for title, group in (("the build changed it", mine),
                         ("his own later workbook replaced it - not a loss", superseded)):
        if later is None and title.startswith("his"):
            continue
        print(f"\n=== {title}: {sum(len(p) for _, p in group)} cells, "
              f"{len(group)} distinct ===")
        for v, places in group[:cap]:
            t, coord = places[0]
            print(f"{len(places):>3} x  {v[:62]:<62} | {t}!{coord} -> "
                  f"{str(b[t][coord].value)[:36]!r}")
        if len(group) > cap:
            print(f"     ... {len(group) - cap} more distinct")
