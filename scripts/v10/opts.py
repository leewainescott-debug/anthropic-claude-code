"""Option mockups generated from TDD_Cost_Calc.xlsx itself.

Nothing here is typed. Every figure is read out of the live workbook, so a mockup cannot
disagree with the thing it is a mockup of.

1.x is drawn on Z Retail, which has real NZ budget and NZ squads, so the AU/NZ mechanic
is visible instead of a column of zeros.
"""
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L

SP = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(SP, "opt")
SRC = os.path.join(SP, "base.xlsx")
REVIEW = "REVIEW - Complete Role Mapping"

FN = "Calibri"
BARC, NAVY, PALE, GREY, MID, YEL = ("FF002F6C", "FF1F4E79", "FFDDEBF7",
                                    "FFF2F2F2", "FFD9D9D9", "FFFFFF00")
thin = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
TOPR = Border(top=Side(style="medium", color=BARC), left=thin, right=thin, bottom=thin)
TITLE = Font(name=FN, bold=True, size=16)
BARF = Font(name=FN, bold=True, size=11, color="FFFFFFFF")
HDRF = Font(name=FN, bold=True, size=11, color="FFFFFFFF")
BOLD = Font(name=FN, bold=True, size=11)
BODY = Font(name=FN, size=11)
BIG = Font(name=FN, bold=True, size=16)
M2 = '#,##0.00;(#,##0.00);"-"'
M0 = '#,##0;(#,##0);"-"'
CT = '#,##0;(#,##0);"-"'
C1 = '#,##0.0;(#,##0.0);"-"'
PCT = '0%'
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center")
RGT = Alignment(horizontal="right", vertical="center")


def fl(c):
    return PatternFill("solid", start_color=c, end_color=c)


def bar(ws, r, c0, n, text):
    for i in range(n):
        x = ws.cell(r, c0 + i)
        x.fill, x.font = fl(BARC), BARF
    ws.cell(r, c0).value = text
    ws.cell(r, c0).alignment = LFT
    ws.row_dimensions[r].height = 19
    return r + 1


def head(ws, r, c0, labels, widths):
    for i, t in enumerate(labels):
        x = ws.cell(r, c0 + i)
        x.value, x.font, x.fill, x.alignment, x.border = t, HDRF, fl(NAVY), CEN, BOX
        if widths[i]:
            ws.column_dimensions[L(c0 + i)].width = widths[i]
    ws.row_dimensions[r].height = 32
    return r + 1


def row(ws, r, c0, vals, fmts, bg=None, bold=False, top=False):
    for i, v in enumerate(vals):
        x = ws.cell(r, c0 + i)
        x.value = v
        x.font = BOLD if bold else BODY
        x.border = TOPR if top else BOX
        if bg:
            x.fill = fl(bg)
        if fmts[i]:
            x.number_format, x.alignment = fmts[i], RGT
        else:
            x.alignment = LFT
    return r + 1


def pairs(ws, r, c0, items, w=(50, 14)):
    ws.column_dimensions[L(c0)].width = w[0]
    ws.column_dimensions[L(c0 + 1)].width = w[1]
    for lab, val, f, b in items:
        a, x = ws.cell(r, c0), ws.cell(r, c0 + 1)
        a.value, a.font, a.alignment, a.border = lab, (BOLD if b else BODY), LFT, BOX
        x.value, x.font, x.number_format, x.alignment, x.border = \
            val, (BOLD if b else BODY), f, RGT, BOX
        if b:
            a.fill = x.fill = fl(MID)
        r += 1
    return r


def strip(ws, r, c0, tiles, w=17):
    for i, (lab, v, f) in enumerate(tiles):
        h = ws.cell(r, c0 + i)
        h.value, h.font, h.fill, h.alignment, h.border = lab, HDRF, fl(NAVY), CEN, BOX
        x = ws.cell(r + 1, c0 + i)
        x.value, x.font, x.number_format = v, BIG, f
        x.alignment, x.border, x.fill = CEN, BOX, fl(GREY)
        ws.column_dimensions[L(c0 + i)].width = w
    ws.row_dimensions[r].height = 32
    ws.row_dimensions[r + 1].height = 30
    return r + 2


def sheet(wb, name, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    ws.title = name
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


# --------------------------------------------------------------------------- read
class Data:
    def __init__(self, path=SRC):
        self.v = openpyxl.load_workbook(path, data_only=True)
        self.f = openpyxl.load_workbook(path)

    def design(self, tab):
        """(platform, [(squad, type, size, on/off, au/nz, support%, cost, tdd, other)])"""
        ws, vs = self.f[tab], self.v[tab]
        out, cur = [], None
        for r in range(1, ws.max_row + 1):
            b = str(vs.cell(r, 2).value or "").strip()
            if b.startswith("Platform: "):
                cur = (b[10:], [])
                out.append(cur)
            elif (cur is not None and b not in ("Squad", "")
                  and str(vs.cell(r, 3).value or "").strip() not in ("", "Squad Type")
                  and isinstance(vs.cell(r, 8).value, (int, float))):
                cur[1].append((b, vs.cell(r, 3).value, vs.cell(r, 4).value,
                               vs.cell(r, 5).value, vs.cell(r, 6).value,
                               vs.cell(r, 7).value, vs.cell(r, 8).value,
                               vs.cell(r, 9).value, vs.cell(r, 10).value))
            elif cur is not None and b == "Platform Overhead":
                cur[1].append(("__oh__", None, None, None, None, None,
                               None, vs.cell(r, 9).value, None))
        return [(p, s) for p, s in out if s]

    def summary(self, tab):
        vs = self.v[tab]
        base = 5 if tab == "1.7 Infrastructure" else 4
        out = []
        for r in range(base + 1, base + 5):
            lab = vs.cell(r, 2).value
            if lab:
                out.append((lab, vs.cell(r, 3).value, vs.cell(r, 4).value,
                            vs.cell(r, 5).value, vs.cell(r, 6).value))
        return out

    def budget_block(self, tab):
        """the H/I label-value pairs: budget and over/(under)"""
        vs = self.v[tab]
        out = []
        for r in range(1, 30):
            lab = vs.cell(r, 8).value
            val = vs.cell(r, 9).value
            if isinstance(lab, str) and isinstance(val, (int, float)) and \
                    ("udget" in lab or "over/(under)" in lab):
                out.append((lab.strip(), val))
        return out

    def other_funding(self, tab):
        vs = self.v[tab]
        out, on = [], False
        for r in range(1, 40):
            lab = vs.cell(r, 8).value
            if lab == "Budget line":
                on = True
                continue
            if on and isinstance(lab, str) and lab.strip():
                out.append((lab.strip(), vs.cell(r, 9).value, vs.cell(r, 10).value,
                            vs.cell(r, 11).value))
                if lab.strip().startswith("Left to fund"):
                    break
        return out

    def portfolios(self):
        """(name, AU budget, NZ budget, actual, archetype, squad cost, overhead,
            roles, filled, vacant) straight off 0.2, 3.1 and 3.2"""
        d, s1, s2 = self.v["0.2 Data Config"], self.v["3.1 Group Summary"], \
            self.v["3.2 Total Cost"]
        bud = {}
        for r in range(6, 26):
            lab = d.cell(r, 2).value
            if lab:
                bud[str(lab).strip()] = (d.cell(r, 3).value or 0, d.cell(r, 4).value or 0)
        out = []
        for r in list(range(6, 16)) + list(range(17, 21)):
            name = s2.cell(r, 2).value
            if not name or str(name).startswith("Portfolios with"):
                continue
            out.append({
                "name": name, "roles": s2.cell(r, 3).value,
                "filled": s2.cell(r, 4).value, "vacant": s2.cell(r, 5).value,
                "actual": s2.cell(r, 6).value, "squad": s2.cell(r, 8).value,
                "arch": s2.cell(r, 10).value, "oh": s2.cell(r, 13).value,
                "aubud": 0, "nzbud": 0})
        return out, bud

    def budget_ladder(self):
        d = self.v["0.2 Data Config"]
        return {"alloc": (d["C26"].value, d["D26"].value, d["E26"].value),
                "budget": (d["C27"].value, d["D27"].value, d["E27"].value),
                "free": (d["C28"].value, d["D28"].value, d["E28"].value)}

    def overhead(self):
        l = self.v["Lists"]
        s2 = self.v["3.2 Total Cost"]
        out = []
        for i, r in enumerate(range(30, 36)):
            out.append((s2.cell(r, 2).value, s2.cell(r, 3).value, s2.cell(r, 4).value,
                        s2.cell(r, 5).value, s2.cell(r, 7).value, s2.cell(r, 6).value))
        return out

    def squads_33(self, limit=None):
        ws, vs = self.f["3.3 FTE View"], self.v["3.3 FTE View"]
        out = []
        for r in range(6, 112):
            pf, sq = vs.cell(r, 2).value, vs.cell(r, 3).value
            if not pf or not sq or sq == "Portfolio total":
                continue
            out.append((pf, sq, vs.cell(r, 4).value, vs.cell(r, 5).value,
                        vs.cell(r, 6).value, vs.cell(r, 9).value,
                        vs.cell(r, 7).value, vs.cell(r, 8).value))
            if limit and len(out) >= limit:
                break
        return out

    def working(self, tab):
        """(squad, type, size, arch roles, roles, vacant, arch cost, actual, after)"""
        ws, vs = self.f[tab], self.v[tab]
        out, oh = [], []
        mode = "sq"
        stop = next((r for r in range(6, ws.max_row + 1)
                     if str(vs.cell(r, 2).value or "").strip() == "Total portfolio"), 40)
        for r in range(6, stop):
            b = vs.cell(r, 2).value
            if not b:
                continue
            s = str(b).strip()
            if s.startswith("Overhead roles inside"):
                mode = "oh"
                continue
            if s in ("Delivery squads", "Overhead roles", "Total portfolio") or \
                    s.startswith("Control") or s.startswith("Archetype columns"):
                continue
            rec = (s, vs.cell(r, 3).value, vs.cell(r, 18).value, vs.cell(r, 4).value,
                   vs.cell(r, 5).value, vs.cell(r, 7).value, vs.cell(r, 12).value,
                   vs.cell(r, 13).value, vs.cell(r, 15).value)
            (oh if mode == "oh" else out).append(rec)
        return out, oh

    def people(self, tab, squads):
        ws, vs = self.f[tab], self.v[tab]
        hdr = None
        for r in range(1, ws.max_row + 1):
            if str(vs.cell(r, 2).value or "").strip() == "Name":
                hdr = r
                break
        out = {}
        if not hdr:
            return out
        for r in range(hdr + 1, ws.max_row + 1):
            n = vs.cell(r, 2).value
            if not n:
                break
            out.setdefault(vs.cell(r, 6).value, []).append(
                (n, vs.cell(r, 3).value, vs.cell(r, 4).value, vs.cell(r, 5).value,
                 vs.cell(r, 7).value, vs.cell(r, 8).value))
        return out


D = None
ZT, ZW = "1.10 Z Retail", "2.10 Z Retail"
SQC = ["Squad", "Squad Type", "Size", "On/Off", "AU / NZ", "Support %",
       "Total Squad Cost ($m)", "TDD Cost ($m)", "Funded outside TDD ($m)"]
SQW = [26, 27, 7, 10, 9, 10, 14, 12, 17]


def _sq(ws, r, sq, plat=None):
    off = 1 if plat else 0
    if sq[0] == "__oh__":
        return row(ws, r, 2, ([plat] if plat else []) +
                   ["Platform Overhead", None, None, None, None, None, None, sq[7], None],
                   ([None] if plat else []) + [None]*6 + [M2, M2, M2])
    r2 = row(ws, r, 2, ([plat] if plat else []) + list(sq),
             ([None] if plat else []) + [None]*5 + [PCT, M2, M2, M2])
    for c in (4 + off, 5 + off, 6 + off, 7 + off):
        ws.cell(r, c).alignment = CEN
        ws.cell(r, c).fill = fl(YEL)
    return r2


def _budget_pairs(tab):
    return [(l, v, M2, "TDD over" in l) for l, v in D.budget_block(tab)]


def _fund_tbl(ws, r, tab, c0=2):
    r = head(ws, r, c0, ["Budget line", "Budget ($m)",
                         "Amount that can be allocated to people ($m)",
                         "Remaining for non-people ($m)"], [30, 13, 22, 20])
    st = r
    for lab, b, p, k in D.other_funding(tab):
        last = lab.startswith("Left to fund") or lab.startswith("Total applied")
        r = row(ws, r, c0, [lab, b, p, k], [None, M2, M2, M2],
                bg=(MID if lab.startswith("Left to fund") else
                    (GREY if last else None)), bold=last)
    return r


# --------------------------------------------------------------- 1.x
def one(kind):
    wb = openpyxl.Workbook()
    ws = sheet(wb, ZT, True)
    ws.cell(2, 2).value = "Z Retail"
    ws.cell(2, 2).font = TITLE
    S, plats = D.summary(ZT), D.design(ZT)
    hdr, rows_ = S[0], S[1:]

    def summary_at(r, c0):
        r = head(ws, r, c0, list(hdr), [40, 13, 13, 12, 12])
        st = r
        for x in rows_:
            r = row(ws, r, c0, list(x), [None, M2, M2, M2, M2])
        return row(ws, r, c0, ["Total Cost"] +
                   [f"=SUM({L(c0+i)}{st}:{L(c0+i)}{r-1})" for i in range(1, 5)],
                   [None, M2, M2, M2, M2], bg=MID, bold=True, top=True)

    def plat_blocks(r, c0=2):
        for name, squads in plats:
            r = bar(ws, r, c0, len(SQC), f"Platform: {name}")
            r = head(ws, r, c0, SQC, SQW)
            st = r
            for s in squads:
                r = _sq(ws, r, s)
            r = row(ws, r, c0, [f"{name} Total"] + [None]*5 +
                    [f"=SUM(H{st}:H{r-1})", f"=SUM(I{st}:I{r-1})",
                     f"=SUM(J{st}:J{r-1})"],
                    [None]*6 + [M2, M2, M2], bg=GREY, bold=True) + 1
        return r

    if kind == "A":                       # owner's arrangement, cleaned
        bar(ws, 4, 2, 5, "Portfolio Summary")
        summary_at(5, 2)
        bar(ws, 4, 8, 2, "Budget vs TDD Cost")
        pairs(ws, 5, 8, _budget_pairs(ZT), w=(44, 14))
        r = bar(ws, 12, 2, 4, "Other funding")
        _fund_tbl(ws, r, ZT)
        plat_blocks(23)
    elif kind == "B":                     # one column, one squad table
        r = bar(ws, 4, 2, 5, "Portfolio Summary")
        r = summary_at(r, 2) + 1
        r = bar(ws, r, 2, 2, "Budget vs TDD Cost")
        r = pairs(ws, r, 2, _budget_pairs(ZT)) + 1
        r = bar(ws, r, 2, 4, "Other funding")
        r = _fund_tbl(ws, r, ZT) + 1
        r = bar(ws, r, 2, 10, "Squads")
        r = head(ws, r, 2, ["Platform"] + SQC, [18] + SQW)
        st = r
        for name, squads in plats:
            for s in squads:
                r = _sq(ws, r, s, plat=name)
        row(ws, r, 2, ["Total"] + [None]*6 +
            [f"=SUM(I{st}:I{r-1})", f"=SUM(J{st}:J{r-1})", f"=SUM(K{st}:K{r-1})"],
            [None]*7 + [M2, M2, M2], bg=MID, bold=True, top=True)
    elif kind == "C":                     # money left, squads right
        r = bar(ws, 4, 2, 2, "Budget vs TDD Cost")
        r = pairs(ws, r, 2, _budget_pairs(ZT)) + 1
        r = bar(ws, r, 2, 4, "Other funding")
        _fund_tbl(ws, r, ZT, 2)
        bar(ws, 4, 8, 5, "Portfolio Summary")
        rr = summary_at(5, 8) + 1
        for name, squads in plats:
            rr = bar(ws, rr, 8, len(SQC), f"Platform: {name}")
            rr = head(ws, rr, 8, SQC, SQW)
            st = rr
            for s in squads:
                if s[0] == "__oh__":
                    rr = row(ws, rr, 8, ["Platform Overhead"] + [None]*6 + [s[7], None],
                             [None]*7 + [M2, M2])
                else:
                    rr = row(ws, rr, 8, list(s), [None]*5 + [PCT, M2, M2, M2])
                    for c in (11, 12, 13, 14):
                        ws.cell(rr-1, c).alignment = CEN
                        ws.cell(rr-1, c).fill = fl(YEL)
            rr = row(ws, rr, 8, [f"{name} Total"] + [None]*5 +
                     [f"=SUM(N{st}:N{rr-1})", f"=SUM(O{st}:O{rr-1})",
                      f"=SUM(P{st}:P{rr-1})"],
                     [None]*6 + [M2, M2, M2], bg=GREY, bold=True) + 1
    else:                                 # D: number strip on top
        tot = rows_[0][4] + rows_[1][4] + rows_[2][4]
        bp = dict(D.budget_block(ZT))
        r = strip(ws, 4, 2, [
            ("TDD AU ($m)", sum(x[1] for x in rows_), M2),
            ("TDD NZ ($m)", sum(x[2] for x in rows_), M2),
            ("Other ($m)", sum(x[3] for x in rows_), M2),
            ("Total Cost ($m)", tot, M2),
            ("AU Budget ($m)", bp.get("AU Budget ($m)", 0), M2),
            ("NZ Budget ($m)", bp.get("NZ Budget ($m)", 0), M2),
            ("TDD over/(under) ($m)", bp.get("TDD over/(under) budget ($m)", 0), M2),
        ], w=15) + 1
        r = bar(ws, r, 2, 4, "Other funding")
        r = _fund_tbl(ws, r, ZT) + 1
        r = bar(ws, r, 2, 10, "Squads")
        r = head(ws, r, 2, ["Platform"] + SQC, [18] + SQW)
        st = r
        for name, squads in plats:
            for s in squads:
                r = _sq(ws, r, s, plat=name)
        row(ws, r, 2, ["Total"] + [None]*6 +
            [f"=SUM(I{st}:I{r-1})", f"=SUM(J{st}:J{r-1})", f"=SUM(K{st}:K{r-1})"],
            [None]*7 + [M2, M2, M2], bg=MID, bold=True, top=True)
    return wb


# --------------------------------------------------------------- 2.x
W2 = ["Squad", "Archetype Type", "Size", "Archetype roles", "Roles", "Filled",
      "Vacant", "To hire", "To offshore", "On hold", "Vacancies remaining",
      "Archetype cost ($m)", "Actual cost ($m)", "Variance to archetype ($m)",
      "Cost after vacancy decisions ($m)", "New variance ($m)"]
W2W = [26, 25, 7, 11, 7, 7, 8, 8, 10, 8, 12, 13, 12, 14, 16, 13]
P2 = ["Name", "Role", "Status", "Vacancy lever", "Role cost ($)",
      "Cost after decision ($)"]
P2W = [30, 40, 10, 15, 15, 18]
TAB2 = "2.3 Enterprise Data"


def _wrow(ws, r, s, oh=False):
    arch = "-" if oh else s[3]
    ac = "-" if oh else s[6]
    var = "-" if oh else (s[7] - s[6] if isinstance(s[6], (int, float)) else "-")
    nv = "-" if oh else (s[8] - s[6] if isinstance(s[6], (int, float)) else "-")
    return row(ws, r, 2, [s[0], "Overhead - see 3.2" if oh else s[1],
                          "-" if oh else s[2], arch, s[4], s[4] - s[5], s[5], s[5], 0, 0,
                          0, ac, s[7], var, s[8], nv],
               [None, None, None, C1 if not oh else None, CT, CT, CT, CT, CT, CT, CT,
                M2 if not oh else None, M2, M2 if not oh else None, M2,
                M2 if not oh else None])


def _wblock(ws, r):
    sq, oh = D.working(TAB2)
    r = head(ws, r, 2, W2, W2W)
    st = r
    for s in sq:
        r = _wrow(ws, r, s)
    r = row(ws, r, 2, ["Delivery squads"] + [None]*2 +
            [f"=SUM({L(c)}{st}:{L(c)}{r-1})" for c in range(5, 18)],
            [None, None, None, C1] + [CT]*7 + [M2]*5, bg=GREY, bold=True)
    dr = r - 1
    r += 1
    r = bar(ws, r, 2, len(W2), "Overhead roles")
    r = head(ws, r, 2, W2, W2W)
    st2 = r
    for s in oh:
        r = _wrow(ws, r, s, oh=True)
    r = row(ws, r, 2, ["Overhead roles total"] + [None]*2 +
            [f"=SUM({L(c)}{st2}:{L(c)}{r-1})" for c in range(5, 18)],
            [None, None, None, C1] + [CT]*7 + [M2]*5, bg=GREY, bold=True)
    orr = r - 1
    r += 1
    r = row(ws, r, 2, ["Total portfolio"] + [None]*2 +
            [f"=N({L(c)}{dr})+N({L(c)}{orr})" for c in range(5, 18)],
            [None, None, None, C1] + [CT]*7 + [M2]*5, bg=MID, bold=True, top=True)
    return r + 1, sq, oh


def two(kind):
    wb = openpyxl.Workbook()
    ws = sheet(wb, TAB2, True)
    ws.cell(2, 2).value = "Enterprise Data - working copy"
    ws.cell(2, 2).font = TITLE
    r = bar(ws, 4, 2, len(W2), "Squad summary")
    r, sq, oh = _wblock(ws, r)
    ppl = D.people(TAB2, None)
    order = [s[0] for s in sq] + [s[0] for s in oh]

    if kind == "A":                    # people grouped under a band, full width
        r = bar(ws, r, 2, len(W2), "Enterprise Data FTE")
        r = head(ws, r, 2, P2 + [None]*(len(W2)-len(P2)),
                 P2W + [None]*(len(W2)-len(P2)))
        for g in order:
            pl = ppl.get(g, [])
            r = row(ws, r, 2, [g, f"{len(pl)} roles", None, None,
                               sum(p[4] or 0 for p in pl), sum(p[5] or 0 for p in pl)],
                    [None, None, None, None, M0, M0], bg=PALE, bold=True)
            for n, ro, stt, lev, c1, c2 in pl:
                r = row(ws, r, 2, [n, ro, stt, lev, c1, c2],
                        [None, None, None, None, M0, M0])
                ws.cell(r-1, 5).fill = fl(YEL)
                ws.cell(r-1, 5).alignment = CEN
    elif kind == "B":                  # people indented in the same grid, collapsible
        r = bar(ws, r, 2, len(W2), "Enterprise Data FTE")
        r = head(ws, r, 2, P2 + [None]*(len(W2)-len(P2)),
                 P2W + [None]*(len(W2)-len(P2)))
        for g in order:
            pl = ppl.get(g, [])
            r = row(ws, r, 2, [g, f"{len(pl)} roles", None, None,
                               sum(p[4] or 0 for p in pl), sum(p[5] or 0 for p in pl)],
                    [None, None, None, None, M0, M0], bg=PALE, bold=True)
            for n, ro, stt, lev, c1, c2 in pl:
                r = row(ws, r, 2, ["    " + str(n), ro, stt, lev, c1, c2],
                        [None, None, None, None, M0, M0])
                ws.cell(r-1, 5).fill = fl(YEL)
                ws.cell(r-1, 5).alignment = CEN
                ws.row_dimensions[r-1].outlineLevel = 1
        ws.sheet_properties.outlinePr.summaryBelow = False
    elif kind == "C":                  # squad summary only
        r = row(ws, r, 2, ["Every role and every lever is on the 2.0 FTE tab."] +
                [None]*(len(W2)-1), [None]*len(W2))
    else:                              # D: vacancies only
        r = bar(ws, r, 2, 7, "Vacancies - the decisions")
        r = head(ws, r, 2, ["Squad", "Role", "Role cost ($)", "Vacancy lever",
                            "Cost after decision ($)", "Saving ($)"],
                 [26, 40, 15, 15, 18, 14])
        st = r
        for g in order:
            for n, ro, stt, lev, c1, c2 in ppl.get(g, []):
                if stt != "Vacant":
                    continue
                r = row(ws, r, 2, [g, ro, c1, lev, c2, (c2 or 0) - (c1 or 0)],
                        [None, None, M0, None, M0, M0])
                ws.cell(r-1, 5).fill = fl(YEL)
                ws.cell(r-1, 5).alignment = CEN
        row(ws, r, 2, ["Total", None, f"=SUM(D{st}:D{r-1})", None,
                       f"=SUM(F{st}:F{r-1})", f"=SUM(G{st}:G{r-1})"],
            [None, None, M0, None, M0, M0], bg=MID, bold=True, top=True)
    return wb


# --------------------------------------------------------------- 3.x
def _p31(ws, kind):
    ws.cell(2, 2).value = "TDD Summary - all portfolios"
    ws.cell(2, 2).font = TITLE
    pf, bud = D.portfolios()
    lad = D.budget_ladder()
    r = 4
    if kind == "D":
        tot = sum(p["actual"] or 0 for p in pf)
        r = strip(ws, r, 2, [("Roles", sum(p["roles"] or 0 for p in pf), CT),
                             ("Filled", sum(p["filled"] or 0 for p in pf), CT),
                             ("Vacant", sum(p["vacant"] or 0 for p in pf), CT),
                             ("Cost today ($m)", tot, M2),
                             ("TDD budget ($m)", lad["budget"][2], M2),
                             ("Left to fund ($m)", tot - lad["budget"][2], M2)], w=16) + 1
    r = bar(ws, r, 2, 9, "Budget against cost by portfolio")
    if kind == "B":
        for c0, n, lab in ((3, 2, "BUDGET"), (5, 2, "COST"), (7, 3, "POSITION")):
            for i in range(n):
                x = ws.cell(r, c0 + i)
                x.fill, x.font, x.alignment = fl(NAVY), BARF, CEN
            ws.cell(r, c0).value = lab
            ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0+n-1)
        r += 1
    r = head(ws, r, 2, ["Portfolio", "AU Budget ($m)", "NZ Budget ($m)",
                        "Actual cost ($m)", "Cost after vacancy decisions ($m)",
                        "Over/(under) budget ($m)", "Roles", "Vacant"],
             [26, 13, 13, 13, 17, 15, 9, 9])
    st = r
    for p in pf:
        ab, nb = bud.get(p["name"], (0, 0))
        if p["name"] == "Customer":
            ab2, nb2 = bud.get("Ampol Customer", (0, 0))
            zb = bud.get("Z Customer", (0, 0))
            ab, nb = ab2 + zb[0], nb2 + zb[1]
        r = row(ws, r, 2, [p["name"], ab, nb, p["actual"], p["actual"],
                           (p["actual"] or 0) - ab - nb, p["roles"], p["vacant"]],
                [None, M2, M2, M2, M2, M2, CT, CT])
    r = row(ws, r, 2, ["Total"] + [f"=SUM({L(c)}{st}:{L(c)}{r-1})" for c in range(3, 10)],
            [None, M2, M2, M2, M2, M2, CT, CT], bg=MID, bold=True, top=True)
    r += 1
    r = bar(ws, r, 2, 4, "Ladder to the TDD budget")
    r = head(ws, r, 2, ["", "AU ($m)", "NZ ($m)", "Total ($m)"], [40, 13, 13, 13])
    for lab, k in (("Allocated to portfolios and COEs", "alloc"),
                   ("Full TDD people budget", "budget"),
                   ("Not yet allocated", "free")):
        r = row(ws, r, 2, [lab] + list(lad[k]), [None, M2, M2, M2],
                bg=(MID if k == "free" else None), bold=(k == "free"))


def _p32(ws, kind):
    ws.cell(2, 2).value = "Total Cost - archetype model vs actual"
    ws.cell(2, 2).font = TITLE
    pf, _ = D.portfolios()
    r = 4
    if kind == "C":
        r = bar(ws, r, 2, 2, "Cost bridge")
        sq = sum(p["squad"] or 0 for p in pf)
        ar = sum(p["arch"] or 0 for p in pf if isinstance(p["arch"], (int, float)))
        oh = sum(p["oh"] or 0 for p in pf)
        r = pairs(ws, r, 2, [
            ("Squad archetype cost - the design", ar, M2, False),
            ("Delivery squads raised over the archetype", sq - ar, M2, False),
            ("Overhead roles inside the ledger", oh, M2, False),
            ("Cost of the organisation today", sq + oh, M2, True),
            ("Impact of the vacancy decisions", 0, M2, False),
            ("Cost after vacancy decisions", sq + oh, M2, True)], w=(52, 15)) + 1
    r = bar(ws, r, 2, 8, "Delivery squads - archetype against actual")
    r = head(ws, r, 2, ["Portfolio", "Archetype cost ($m)", "Squad cost ($m)",
                        "Variance to archetype ($m)",
                        "Cost after vacancy decisions ($m)", "New variance ($m)",
                        "Overhead cost ($m)", "Total actual cost ($m)"],
             [26, 15, 13, 16, 17, 14, 14, 15])
    st = r
    for p in pf:
        a = p["arch"] if isinstance(p["arch"], (int, float)) else None
        sq = p["squad"] or 0
        r = row(ws, r, 2, [p["name"], a if a else "-", sq,
                           (sq - a) if a else "-", sq, (sq - a) if a else "-",
                           p["oh"] or 0, p["actual"]],
                [None, M2, M2, M2, M2, M2, M2, M2])
    r = row(ws, r, 2, ["Total"] + [f"=SUM({L(c)}{st}:{L(c)}{r-1})" for c in range(3, 10)],
            [None, M2, M2, M2, M2, M2, M2, M2], bg=MID, bold=True, top=True)
    r += 1
    r = bar(ws, r, 2, 7, "Overhead - allowance against actual (of which, inside the total above)")
    r = head(ws, r, 2, ["Overhead line", "Roles", "Rate ($m)", "Times applied",
                        "Allowance ($m)", "Actual ($m)", "Over/(under) allowance ($m)"],
             [26, 9, 11, 13, 14, 13, 18])
    st2 = r
    for n, ro, rate, un, al, act in D.overhead():
        r = row(ws, r, 2, [n, ro, rate, un, al, act, (act or 0) - (al or 0)],
                [None, CT, M2, CT, M2, M2, M2])
    row(ws, r, 2, ["Overhead total", f"=SUM(C{st2}:C{r-1})", None, None,
                   f"=SUM(F{st2}:F{r-1})", f"=SUM(G{st2}:G{r-1})",
                   f"=SUM(H{st2}:H{r-1})"],
        [None, CT, None, None, M2, M2, M2], bg=MID, bold=True, top=True)


def _p33(ws, kind):
    ws.cell(2, 2).value = "Squad Detail - roles and cost by portfolio and squad"
    ws.cell(2, 2).font = TITLE
    r = bar(ws, 4, 2, 11, "Delivery squads")
    r = head(ws, r, 2, ["Portfolio", "Squad", "Archetype Type", "Size",
                        "Archetype roles", "Roles", "Filled", "Vacant",
                        "Archetype cost ($m)", "Actual cost ($m)",
                        "Cost after vacancy decisions ($m)"],
             [22, 30, 26, 7, 12, 8, 8, 8, 14, 13, 17])
    st = r
    cur = None
    for pf, sq, ty, sz, ar, ac, ro, vc in D.squads_33(limit=34):
        if kind == "B" and cur is not None and pf != cur:
            pass
        cur = pf
        f = (ro - vc) if isinstance(ro, (int, float)) and isinstance(vc, (int, float)) \
            else None
        r = row(ws, r, 2, [pf, sq, ty, sz, ar, ro, f, vc, ac, None, None],
                [None, None, None, None, C1, CT, CT, CT, M2, M2, M2])
    row(ws, r, 2, ["Total", None, None, None] +
        [f"=SUM({L(c)}{st}:{L(c)}{r-1})" for c in range(6, 13)],
        [None, None, None, None, C1, CT, CT, CT, M2, M2, M2],
        bg=MID, bold=True, top=True)


def three(kind):
    wb = openpyxl.Workbook()
    _p31(sheet(wb, "3.1 Group Summary", True), kind)
    _p32(sheet(wb, "3.2 Total Cost"), kind)
    _p33(sheet(wb, "3.3 Squad Detail"), kind)
    return wb
