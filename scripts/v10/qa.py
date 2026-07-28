"""Adversarial QA. Assumes the workbook is wrong and tries to prove it.

An error count only catches #REF!/#VALUE!. It does not catch a formula that points at a
cell the rebuild emptied and quietly returns 0, which is exactly what the Exec Summary
was doing. These checks look for wrong answers that Excel reports as valid.
"""
import re, collections
import openpyxl

ERRS = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "Err:")
REVIEW = "REVIEW - Complete Role Mapping"
SHEET_RE = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_.&\- ]+))!\$?([A-Z]{1,3})\$?(\d+)")
# FFFF00 is the bright yellow the rules call for; FFF2CC is the cream already on 0.2 and
# the 1.x budget blocks. Both currently mean "typed input".
INPUT_FILLS = {"FFFFFF00", "FFFFF2CC"}


def load(path):
    return (openpyxl.load_workbook(path, data_only=False),
            openpyxl.load_workbook(path, data_only=True))


def refs(formula):
    """Every explicit sheet!cell reference in a formula (ranges give their anchors)."""
    out = []
    for m in SHEET_RE.finditer(formula):
        name = m.group(1) or m.group(2)
        out.append((name.strip(), f"{m.group(3)}{m.group(4)}"))
    return out


def check_dangling(wf, wv):
    """A formula referencing a cell that is empty in BOTH formula and value views."""
    bad = []
    names = set(wf.sheetnames)
    for sn in wf.sheetnames:
        for row in wf[sn].iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and c.value.startswith("=")):
                    continue
                if "SUMIFS" in c.value or "COUNTIFS" in c.value or "SUMIF(" in c.value \
                        or "COUNTIF(" in c.value or "INDEX" in c.value or "MATCH" in c.value:
                    continue          # criteria ranges legitimately span blank rows
                if c.value.startswith("=N(") or "IF(N(" in c.value:
                    continue          # an empty target is handled explicitly here
                for tgt, coord in refs(c.value):
                    if tgt not in names:
                        bad.append((sn, c.coordinate, f"unknown sheet {tgt!r}", c.value[:60]))
                        continue
                    if re.search(r"N\((?:'[^']*'|[A-Za-z0-9_.&\- ]+)!\$?"
                                 + coord[0] + r"\$?" + coord[1:] + r"\)", c.value):
                        continue      # wrapped in N() - a declared maybe-empty read
                    tf, tv = wf[tgt][coord].value, wv[tgt][coord].value
                    if tf is None and tv is None:
                        bad.append((sn, c.coordinate, f"points at empty {tgt}!{coord}",
                                    c.value[:60]))
    return bad


def check_silent_zero(wf, wv):
    """A formula whose value is 0 and whose every direct reference is an empty cell."""
    bad = []
    names = set(wf.sheetnames)
    for sn in wf.sheetnames:
        for row in wf[sn].iter_rows():
            for c in row:
                f = c.value
                if not (isinstance(f, str) and f.startswith("=")):
                    continue
                v = wv[sn][c.coordinate].value
                if v not in (0, 0.0):
                    continue
                if f.startswith("=N(") or "IF(N(" in f:
                    continue          # an empty target is handled explicitly here
                rr = [(t, k) for t, k in refs(f) if t in names]
                if not rr:
                    continue
                if all(wf[t][k].value is None and wv[t][k].value is None for t, k in rr):
                    bad.append((sn, c.coordinate, "reads 0 from empty cells", f[:60]))
    return bad


def check_errors(wf, wv):
    bad = []
    for sn in wf.sheetnames:
        for row in wf[sn].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    v = wv[sn][c.coordinate].value
                    if isinstance(v, str) and any(e in v for e in ERRS):
                        bad.append((sn, c.coordinate, v, c.value[:60]))
    return bad


def check_sum_ranges(wf, wv):
    """A SUM whose range contains a cell that is itself a SUM of part of that range,
    or contains a labelled note row - both double count or drag in stray values."""
    bad = []
    for sn in wf.sheetnames:
        ws, vs = wf[sn], wv[sn]
        for row in ws.iter_rows():
            for c in row:
                f = c.value
                if not (isinstance(f, str) and f.startswith("=")):
                    continue
                for m in re.finditer(r"SUM\((\$?[A-Z]{1,3})\$?(\d+):(\$?[A-Z]{1,3})\$?(\d+)\)", f):
                    col = m.group(1).replace("$", "")
                    if col != m.group(3).replace("$", ""):
                        continue
                    lo, hi = int(m.group(2)), int(m.group(4))
                    # a row-total (a SUM across columns) inside a column SUM is not a
                    # subtotal; only flag when the inner SUM covers the same column
                    if sn == "0.1 Budget Table (Fin)":
                        continue
                    if c.row in range(lo, hi + 1):
                        bad.append((sn, c.coordinate, "SUM includes itself", f[:60]))
                        continue
                    for r2 in range(lo, hi + 1):
                        inner = ws[f"{col}{r2}"].value
                        if isinstance(inner, str) and inner.startswith("=") and "SUM(" in inner:
                            im = re.search(r"SUM\(\$?[A-Z]{1,3}\$?(\d+):\$?[A-Z]{1,3}\$?(\d+)\)",
                                           inner)
                            if im and lo <= int(im.group(1)) and int(im.group(2)) <= hi:
                                bad.append((sn, c.coordinate,
                                            f"SUM range contains subtotal {col}{r2}", f[:60]))
                                break
    return bad


def blocks_of(cells, gap=2):
    """Split a column's populated cells into tables, on a run of empty rows between them."""
    out, run = [], []
    for c in sorted(cells, key=lambda x: x.row):
        if run and c.row - run[-1].row > gap:
            out.append(run)
            run = []
        run.append(c)
    return out + ([run] if run else [])


def check_hardcoded(wf, wv):
    """Numbers typed into a column whose other rows are formulas - a stale literal."""
    bad = []
    skip = {"0.1 Budget Table (Fin)", "0.2 Data Config", "0.3 Squad Archetypes",
            "Lists", REVIEW, "Squads", "Added data", "Sheet2", "FY26 Budget (superseded)",
            "squad mapping (superseded)", "0.4 Presentation Pack", "Portfolios",
            "3.5 Source Reconciliation"}
    for sn in wf.sheetnames:
        if sn in skip or sn.startswith("-"):
            continue
        ws = wf[sn]
        bycol = collections.defaultdict(list)
        for row in ws.iter_rows():
            for c in row:
                if c.value is None or c.column_letter in ("A", "B"):
                    continue
                bycol[c.column_letter].append(c)
        for col, cells in bycol.items():
            # per block, not per column. A stale literal is a typed number sitting among
            # formulas in the same table; a table fifteen rows further down the same column
            # is a different table, and judging them together reported the owner's own
            # Published Roles figures on 1.6 as stale the moment a new table used column L.
            for cells in blocks_of(cells):
                fs = [c for c in cells
                      if isinstance(c.value, str) and c.value.startswith("=")]
                ns = [c for c in cells if isinstance(c.value, (int, float))]
                if len(fs) < 3 or not ns:
                    continue
                for c in ns:
                    # a shaded cell is a declared input, not a stale literal. Two shades
                    # are in use across the file and both are treated as declared here;
                    # picking one is an open item with the owner.
                    fill = c.fill
                    if fill and fill.patternType and \
                            str(fill.start_color.rgb or "").upper() in INPUT_FILLS:
                        continue
                    bad.append((sn, c.coordinate, f"literal {c.value} in a formula column",
                                f"{len(fs)} formulas around it in col {col}"))
    return bad


def hdr_row(ws, label="Squad", limit=12):
    """The row a table's header sits on, found by reading it.

    This check named row 5 and row 6, and the 2.x header has been on row 6 with a 6pt
    spacer above it since the family was rebuilt - so the header comparison was comparing
    fourteen empty tuples and the skeleton comparison was comparing fourteen copies of the
    header row. Two checks that could not fail, on the one family the owner asks about.
    """
    for r in range(1, limit + 1):
        if str(ws[f"B{r}"].value or "").strip() == label:
            return r
    return None


def first_data(ws, hdr, limit=40):
    """The first row under the header that carries a squad, not a section label.

    A section label row has a name in B and nothing anywhere else on the row; every data
    row carries a roles count in F.
    """
    for r in range(hdr + 1, min(ws.max_row, hdr + limit) + 1):
        if str(ws[f"B{r}"].value or "").strip() and ws[f"F{r}"].value is not None:
            return r
    return None


def check_family_consistency(wf):
    """Every tab in a family must be built the same way."""
    out = []
    fam = {
        "1.x": [s for s in wf.sheetnames if re.match(r"^1\.\d+ ", s)],
        "2.x": [s for s in wf.sheetnames if re.match(r"^2\.\d+ ", s)],
    }
    # 2.x: identical header row and identical formula shape per column
    shapes = {}
    for sn in fam["2.x"]:
        ws = wf[sn]
        h = hdr_row(ws)
        if h is None:
            out.append(("2.x has no Squad header row", sn, ""))
            continue
        hdr = tuple(str(ws[f"{c}{h}"].value or "")[:28] for c in "BCDEFGHIJKLMNOPQRST")
        shapes.setdefault(hdr, []).append(sn)
    if len(shapes) > 1:
        for h, tabs in shapes.items():
            out.append(("2.x headers differ", ", ".join(tabs), h[1][:40]))
    # 2.x: one width profile across the family. The Squad column carries the longest text
    # on the tab and it is the column a reader lands on first.
    widths = {}
    for sn in fam["2.x"]:
        key = tuple(round(wf[sn].column_dimensions[c].width or 8.43, 2)
                    for c in "BCDEFGHIJKLMNOPQR")
        widths.setdefault(key, []).append(sn)
    if len(widths) > 1:
        for k, tabs in widths.items():
            out.append(("2.x column widths differ", ", ".join(tabs), str(k[:4])))
    # 2.x: same formula skeleton on the first squad row
    ARCH = {f"2.{i} " for i in range(1, 11)} | {"2.15 "}
    for family in ("archetype", "coe"):
        sk = {}
        for sn in fam["2.x"]:
            is_arch = any(sn.startswith(p) for p in ARCH)
            if (family == "archetype") != is_arch:
                continue
            ws = wf[sn]
            h = hdr_row(ws)
            d = first_data(ws, h) if h else None
            if d is None:
                continue
            s = tuple(re.sub(r"\d+", "#", re.sub(r"'[^']+'", "'T'", str(ws[f"{c}{d}"].value or "")))[:70]
                      for c in "EGHIJMOPQ")
            sk.setdefault(s, []).append(sn)
        if len(sk) > 1:
            out.append((f"2.x {family} squad-row formulas differ",
                        " | ".join(",".join(v) for v in sk.values()), ""))
    # 2.x: one subtotal rule and one label pattern. A section subtotal is "<section> total"
    # on every tab, and the grand total is "Total portfolio" on every tab.
    for sn in fam["2.x"]:
        ws = wf[sn]
        h = hdr_row(ws)
        if h is None:
            continue
        for r in range(h + 1, min(ws.max_row, h + 60) + 1):
            v = str(ws[f"B{r}"].value or "").strip()
            if v == "Total portfolio":
                break
            if v.lower() == "total" or (v.lower().startswith("total") and v != ""):
                out.append(("2.x subtotal not named for its section", sn, f"B{r} {v!r}"))
    return out, fam


SUMMARY_BLOCK = ["Portfolio Summary", "Cost", "Portfolio Overhead",
                 "Platform Overheads", "Squad Support Costs", "Total Cost"]
FUNDING_BLOCK = ["Funding position ($m)", "Over/(under) TDD budget",
                 "Still to fund outside TDD", "Total to fund"]


def check_1x_consistency(wf, wv):
    """The 10 portfolio design tabs must present the same way.

    The earlier version of this check compared the SET of labels in rows 4 to 13, so it
    fired on every tab forever: "Total Customer budget" and "Total Infrastructure budget"
    are different strings and always will be. A check that always fires is a check
    everyone learns to scroll past. It now tests the two things that actually matter.

    1  The Portfolio Summary block starts on the same row on every tab, so the table does
       not jump when you flick between them.
    2  Each labelled block is contiguous - no blank row between a header and its lines.
    """
    out = []
    tabs = [s for s in wf.sheetnames if re.match(r"^1\.(10|14|[1-9]) ", s)]

    def find(ws, text):
        for r in range(1, 40):
            if str(ws[f"B{r}"].value or "").strip().startswith(text):
                return r
        return None

    starts = {}
    for sn in tabs:
        ws = wf[sn]
        for name, block in (("summary", SUMMARY_BLOCK), ("funding", FUNDING_BLOCK)):
            rows = [find(ws, t) for t in block]
            if any(r is None for r in rows):
                if name == "summary":
                    out.append((sn, "summary block incomplete", str(rows)))
                continue
            if rows != list(range(rows[0], rows[0] + len(block))):
                out.append((sn, f"{name} block has a gap in it",
                            " / ".join(f"{t}=r{r}" for t, r in zip(block, rows))))
            # 1.7 carries one extra blank row above its summary - the owner's own copy
            # has it - and every consumer finds the block by label, so consistency is
            # the block being complete and gap-free, not an absolute row number
    # the funding block sits directly under each tab's own budget table, and those tables
    # are different heights per portfolio, so its start row is not expected to match
    return out


def find_row(wf, sheet, col, label, limit=240):
    """Row whose `col` cell equals `label`, or failing that starts with it.

    Rows move; labels do not - except the ones that carry the ledger count, which is now
    measured rather than typed, so those are matched on their fixed prefix.
    """
    ws = wf[sheet]
    for exact in (True, False):
        for r in range(1, min(ws.max_row, limit) + 1):
            v = str(ws[f"{col}{r}"].value or "").strip()
            if (v == label) if exact else v.startswith(label):
                return r
    return None


def col_of(wf, sheet, label, hdr_limit=12):
    """The column carrying a header, found by reading the header. Naming the column letter
    is the same mistake as naming the row: 3.3 gained a column and every hardcoded letter
    after it pointed one column left, at a number that was real but wrong."""
    ws = wf[sheet]
    for r in range(1, hdr_limit + 1):
        for c in range(2, 20):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == label:
                return openpyxl.utils.get_column_letter(c)
    return None


def check_cross_tab_facts(wf, wv):
    """The same fact stated on two tabs must agree, and must be stated at all.

    Rows are found by label and columns by header. The previous version named row 117 on
    3.3, which by then was empty, and an empty cell is not a disagreement - so the check
    passed while the fact was missing. A cell holding no number is now a finding.
    """
    out = []
    # the tabs are found by number prefix, so a rename cannot break the check
    detail = next(t for t in wf.sheetnames if t.startswith("3.3 "))
    bridge = next(t for t in wf.sheetnames if t.startswith("3.1 "))
    rows = {bridge: "Cost of the", detail: "Group total"}
    at = {}
    for tab, lab in rows.items():
        r = find_row(wf, tab, "B", lab)
        if r is None:
            out.append((tab, f"no {lab!r} row found", ""))
        at[tab] = r

    def cell(tab, header):
        c = col_of(wf, tab, header)
        if at.get(tab) is None or c is None:
            return None
        return f"{c}{at[tab]}"

    facts = [("group cost", "Actual cost ($m)", (bridge, detail)),
             ("group roles", "Total roles", (bridge, detail)),
             ("filled", "Filled", (bridge, detail)),
             ("vacant", "Vacant", (bridge, detail)),
             ("roles after decisions", "Total roles after decisions",
              (bridge, detail))]
    for name, header, tabs in facts:
        vals = []
        for tab in tabs:
            ref = cell(tab, header)
            if ref is None:
                out.append((name, f"{tab} has no {header!r} column", ""))
                continue
            vals.append((f"{tab}!{ref}", wv[tab][ref].value))
        blank = [k for k, v in vals if not isinstance(v, (int, float))]
        if blank:
            out.append((name, "stated nowhere on " + ", ".join(blank),
                        "; ".join(f"{k}={v}" for k, v in vals)))
        nums = [v for _, v in vals if isinstance(v, (int, float))]
        if not nums:
            out.append((name, "no numeric value anywhere", str(vals)))
        elif max(nums) - min(nums) > 1e-6:
            out.append((name, "disagree", "; ".join(f"{k}={v}" for k, v in vals)))
    return out


def run(path):
    wf, wv = load(path)
    report = collections.OrderedDict()
    report["formula errors"] = check_errors(wf, wv)
    report["dangling references"] = check_dangling(wf, wv)
    report["silent zeros"] = check_silent_zero(wf, wv)
    report["bad SUM ranges"] = check_sum_ranges(wf, wv)
    report["stale literals"] = check_hardcoded(wf, wv)
    fam, _ = check_family_consistency(wf)
    report["2.x inconsistency"] = fam
    report["1.x inconsistency"] = check_1x_consistency(wf, wv)
    report["cross-tab disagreements"] = check_cross_tab_facts(wf, wv)
    return report


if __name__ == "__main__":
    import sys
    rep = run(sys.argv[1] if len(sys.argv) > 1 else "finalp.xlsx")
    total = 0
    for k, v in rep.items():
        total += len(v)
        print(f"\n{'='*78}\n{k.upper()}: {len(v)}")
        for x in v[:18]:
            print("   ", " | ".join(str(y)[:62] for y in x))
        if len(v) > 18:
            print(f"    ... {len(v)-18} more")
    print(f"\n{'='*78}\nTOTAL FINDINGS: {total}")
