"""Stage 16: your squad assignments, and the design tabs renamed to REVIEW's names.

Your raw columns are not touched. The assignments live in a small table on Lists keyed by
REVIEW row, and the grouping column reads that table first. So REVIEW stays your data and
the corrections are visible in one place you can edit.

Your calls:
  109, 123, 126, 131, 177   Customer, stay as Leadership
  283, 313                  Data COE - your column F already says Strategy, Architecture
                            & Data, department Group Data, so the portfolio moves too
  528                       AU Finance, which REVIEW calls SAP ERP
  364                       an Infrastructure squad - your data says department and team
                            are both "Fuel Infrastructure", which is not one of the four
                            Infrastructure squads, so this one is still open

Design tabs now take REVIEW's squad names, not the other way round. REVIEW is the source
of truth, so where the two differed the design tab is renamed.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import model

BOLD = Font(bold=True)
YELLOW = PatternFill("solid", start_color="FFFFFF00", end_color="FFFFFF00")

# REVIEW row -> (portfolio override or None, squad override)
ASSIGN = {
    283: ("COE SA&D", "Group Data"),
    313: ("COE SA&D", "Group Data"),
    528: (None, "SAP ERP"),
}

# design tab squad name -> the name REVIEW uses. REVIEW wins.
RENAME = {
    "1.1 Ampol Retail":            {"Network / QSR": "Network & QSR"},
    "1.2 Customer":                {"Z Energy Martech": "Z Loyalty & Martech",
                                    "AU CRM & Martech": "Ampol Loyalty & Martech",
                                    "Z Energy Apps": "Z App and Web"},
    "1.4 TDD Group Functions":     {"Network & Infrastructure": "Cloud, Network & Infra Ops",
                                    "DevOps & Engineering": "DevOps & QE",
                                    "Integration": "Integration & Process Automation"},
    "1.5 P&C":                     {"P&C - RTA": "P&C RTA"},
    "1.6 Finance":                 {"AU Finance": "SAP ERP"},
    "1.7 Infrastructure":          {"Manufacturing & Group Projects":
                                    "Manufacturing Group Projects"},
}


def lists_assign(wb):
    ws = wb["Lists"]
    ws["AN1"] = "REVIEW row"; ws["AO1"] = "Portfolio override"; ws["AP1"] = "Squad override"
    for c in ("AN", "AO", "AP"):
        ws[f"{c}1"].font = BOLD
    for i, (row, (pf, sq)) in enumerate(sorted(ASSIGN.items())):
        r = 2 + i
        ws[f"AN{r}"] = row; ws[f"AO{r}"] = pf; ws[f"AP{r}"] = sq
        for c in ("AN", "AO", "AP"):
            ws[f"{c}{r}"].fill = YELLOW
    n = 2 + len(ASSIGN)
    ws[f"AN{n}"] = None
    ws[f"AO{n+1}"] = ("Agreed assignments. REVIEW's own columns are left as they are; "
                      "these override the grouping only.")
    return [f"Lists AN:AP holds {len(ASSIGN)} agreed assignments, your raw columns untouched"]


def review_cols(wb):
    """AJ and AT respect the assignment table."""
    ws = wb[model.REVIEW]
    LR = model.LAST_ROW
    lo, hi = 2, 1 + len(ASSIGN)
    for i in range(2, LR + 1):
        # portfolio: assignment table first, then the Lists T:U map on your column I
        # an override cell left blank must fall through to your column I, not blank the row
        ovr = (f'IFERROR(INDEX(Lists!$AO${lo}:$AO${hi},'
               f'MATCH(ROW(),Lists!$AN${lo}:$AN${hi},0)),"")')
        mapped = f'IFERROR(INDEX(Lists!$U:$U,MATCH(TRIM($I{i}),Lists!$T:$T,0)),TRIM($I{i}))'
        ws[f"AJ{i}"] = f'=IF(TRIM($B{i})="","",IF({ovr}<>"",{ovr},{mapped}))'
        
        # squad or overhead line: assignment table first
        sq_ovr = (f'IFERROR(INDEX(Lists!$AP${lo}:$AP${hi},'
                  f'MATCH(ROW(),Lists!$AN${lo}:$AN${hi},0)),"")')
        base = (f'IF(OR(LEFT($AJ{i},3)="COE",$AJ{i}="EGI"),$AP{i},'
                f'IF($AR{i}<>"Squad",$AR{i},$AP{i}))')
        ws[f"AT{i}"] = f'=IF(TRIM($B{i})="","",IF({sq_ovr}<>"",{sq_ovr},{base}))'
        
    return ["REVIEW AJ and AT read the assignment table first, then your columns"]


def rename_design(wb):
    out = 0
    for tab, m in RENAME.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for r in range(1, ws.max_row + 1):
            v = str(ws.cell(r, 2).value or "").strip()
            if v in m:
                ws.cell(r, 2).value = m[v]
                out += 1
    return [f"{out} design-tab squad names renamed to match REVIEW"]


def drop_name_map(wb):
    """The lookup that translated REVIEW names into design names is gone - it let the
    design tab override your data. The design tabs now carry REVIEW's names directly."""
    ws = wb["Lists"]
    for r in range(1, 20):
        ws.cell(r, 26).value = None      # Z
        ws.cell(r, 27).value = None      # AA
    ws["Z1"] = "Squad name map retired - design tabs now use REVIEW's names"
    for tab in model.TAB_PORTFOLIO:
        if tab not in wb.sheetnames:
            continue
        t = wb[tab]
        for r in range(6, 44):
            if isinstance(t[f"S{r}"].value, str) and "Lists!$AA" in t[f"S{r}"].value:
                t[f"S{r}"] = f"=$B{r}"
    return ["Lists Z:AA name map retired; 2.x reads the squad name straight through"]


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = []
    for fn in (lists_assign, review_cols, rename_design, drop_name_map):
        out += fn(wb)
    wb.save(dst)
    return out


if __name__ == "__main__":
    for x in run("c6.xlsx", "a1.xlsx"):
        print("  ", x)
