"""Rewire everything that read the old 3.x row positions, and rebuild Exec and 4.0.

Rebuilding 3.1, 3.2 and 3.3 moves every row on them. Anchors come from the JSON the
builders write rather than from a label search, because two rows on 3.1 now begin with the
same words - a pale section label and the grey subtotal under it - and a search would find
the wrong one.

4.0 used to test one thing thirty-seven ways: roles and cost against the ledger. It could
not fail on a design number, on an allowance, or on a control cell reading -3 somewhere
else in the file, and one of those was live. It now tests the design side too.
"""
import json
import re

import openpyxl
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation

import final2x as f2
import opts

REVIEW = f2.REVIEW
REV = f2.REV
LAST = None                             # refreshed from f2 after _boot_last
S = f2.S
A3 = "'0.3 Squad Archetypes'"
G1, G2, G3 = "'3.1 Group Summary'", "'3.2 Total Cost'", "'3.3 FTE View'"
# 3.3 columns after the "How it is funded" column was added
C33 = dict(pf="B", kind="C", squad="D", aroles="G", roles="H", filled="I", vacant="J",
           rafter="K", acost="L", actual="M", var="N", after="O")
# 3.1 is a bridge now: the line name is in B, the portfolio in C, and the figures start at D
C31 = dict(name="B", pf="C", acost="D", actual="E", var="F", after="G", roles="H",
           filled="I", vacant="J", rafter="K")
# 3.2 is one row per overhead line reading left to right: what the allowance applies to the
# portfolios, what the organisation actually carries, and the two gaps between them. The
# old per-portfolio / outside-the-portfolios pairs are gone - the split they used to spell
# out across four columns is two band rows now, and the roles and cost columns are the
# organisation's.
# his layout: roles read together, then costs. B line, C applied-to, D rate,
# E times applied (his), F roles priced for, G actual roles, H roles gap,
# I archetype cost, J actual cost, K variance, L where they sit, M allocation applied
C32 = dict(line="B", basis="C", rate="D", times="E", appfte="F", roles="G", rgap="H",
           applied="I", cost="J", cgap="K", where="L")


def find_row(ws, label, col=2, limit=200, exact=False):
    for r in range(1, min(ws.max_row, limit) + 1):
        v = ws.cell(r, col).value
        if not isinstance(v, str):
            continue
        if (v.strip() == label) if exact else v.strip().startswith(label):
            return r
    raise KeyError(f"{ws.title}: no row matching {label!r}")


def fte_rows(ws):
    """First and last data row of a 2.x tab's FTE block.

    anchors_final.json does not carry these - final2x returns anchors for the squad table
    and its sections but the FTE block's bounds stay local to the builder - so they are
    found here off the block's own header row, which is the next best thing to an anchor
    and is still not a literal: the header is written by the same builder, in one place,
    and a tab that grew or shrank moves the answer with it.

    The header is the only row on the tab carrying "Name" in B and "Status" in D. The
    block is the last thing on the tab, so it runs to the last row carrying a person -
    read off column B, which holds either a name reference or a squad band label on every
    row of the block and nothing below it.
    """
    hdr = None
    for r in range(1, ws.max_row + 1):
        if (str(ws.cell(r, 2).value or "").strip() == "Name"
                and str(ws.cell(r, 4).value or "").strip() == "Status"):
            hdr = r
            break
    if hdr is None:
        raise KeyError(f"{ws.title}: no FTE block header (B='Name', D='Status')")
    last = hdr
    for r in range(hdr + 1, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(2, 8)):
            last = r
    # an empty block - a portfolio designed before it is staffed - still needs a range
    # Excel accepts, and one row of header is a range that counts nothing
    return hdr + 1, max(last, hdr + 1)


def anchors(wb):
    a3 = json.load(open("anchors_final3.json"))
    s3 = wb[G3.strip("'")]
    a = dict(a3["3.1"])
    # 3.2's four band rows come from the builder rather than from a label search. Two of
    # them now begin with the same four words, and the ledger row's label is a formula -
    # it states the counts it is built from, so it cannot be a literal - and a search of
    # the formula workbook would find neither.
    a.update({k: v for k, v in a3["3.2"].items() if k.endswith("32")})
    a["g33"] = find_row(s3, "Group total")
    a["first33"] = find_row(s3, "Portfolio") + 1
    return a


def repoint(wb, a):
    """Lists!K prices the archetype roles per portfolio out of 3.3.

    3.3 gained a column, so archetype roles moved from F to G, and the old formula excluded
    total rows by testing the Squad column for '*total'. The total row's label now sits in
    the portfolio column, so that test matched nothing and the totals were being counted
    twice. Selecting on the new 'How it is funded' column is both narrower and honest: only
    archetype rows carry an archetype.
    """
    lo, hi = a["first33"], a["g33"] - 1
    new = (f"=SUMIFS({G3}!${C33['aroles']}${lo}:${C33['aroles']}${hi},"
           f"{G3}!${C33['pf']}${lo}:${C33['pf']}${hi},$J{{r}},"
           f'{G3}!${C33["kind"]}${lo}:${C33["kind"]}${hi},"Archetype")')
    l = wb["Lists"]
    n = 0
    for r in range(2, 20):
        if not str(l.cell(r, 10).value or "").strip():
            continue
        if isinstance(l.cell(r, 11).value, str) and "3.3 FTE View" in l.cell(r, 11).value:
            l.cell(r, 11).value = new.format(r=r)
            n += 1
    return n


# ------------------------------------------------------------------- Exec Summary
def build_exec(wb, a, a2):
    ws = wb["Exec Summary"]
    f2.wipe(ws)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 18
    gt = a["total"]
    ws.cell(2, 2).value = "TDD operating model - executive summary"
    ws.cell(2, 2).font = opts.TITLE

    def block(r, heading, lines):
        r = opts.bar(ws, r, 2, 2, heading)
        for lab, f, nf in lines:
            ws.cell(r, 2).value = lab
            ws.cell(r, 2).font = opts.BODY
            ws.cell(r, 2).alignment = opts.LFT
            ws.cell(r, 2).border = opts.BOX
            x = ws.cell(r, 3)
            x.value, x.number_format = f, nf
            x.alignment, x.font, x.border = opts.RGT, opts.BODY, opts.BOX
            r += 1
        return r + 1

    r = block(4, "The organisation today", [
        ("Roles in the ledger", f"={G1}!${C31['roles']}${gt}", opts.CT),
        ("Filled", f"={G1}!${C31['filled']}${gt}", opts.CT),
        ("Vacant", f"={G1}!${C31['vacant']}${gt}", opts.CT),
        (f"Cost of the {opts.ledger_count(wb)} roles in the ledger ($m)", f"={G1}!${C31['actual']}${gt}", opts.M2),
        ("Of which filled roles ($m)",
         f'=SUMIFS({REV}!$AA$2:$AA${LAST},{REV}!$AK$2:$AK${LAST},"Filled")/1000000',
         opts.M2),
        ("Of which vacancies not yet filled ($m)",
         f'=SUMIFS({REV}!$AA$2:$AA${LAST},{REV}!$AK$2:$AK${LAST},"Vacant")/1000000',
         opts.M2),
        ("The 8 GMs, outside the ledger ($m)", "=N(Lists!$AG$12)", opts.M2),
        ("Cost today including the GM layer ($m)", f"={G1}!${C31['actual']}${a['grand']}", opts.M2)])

    r = block(r, "Against the archetype", [
        ("Squads priced by an archetype - archetype cost ($m)",
         f"={G1}!${C31['acost']}${a['arch']}", opts.M2),
        ("Squads priced by an archetype - actual ($m)",
         f"={G1}!${C31['actual']}${a['arch']}", opts.M2),
        ("Squads priced by an archetype - over/(under) ($m)",
         f"={G1}!${C31['var']}${a['arch']}", opts.M2),
        ("Directly funded programmes - over/(under) funded ($m)",
         f"={G1}!${C31['var']}${a['direct']}", opts.M2),
        # This line used to read "COEs and EGI - over/(under) their 1.x planned spend
        # ($m)" and point at 3.1's variance column for the COE step, which is the literal
        # string "-". A dash under a $m heading on an over/(under) line reads as "nothing
        # in it", and the four groups behind it cost 28.68 - a quarter of TDD.
        #
        # There is no such over/(under) to state. EGI has no 1.x tab; the three that do
        # carry a "planned spend" that is their own ledger cost after decisions, net of an
        # allowance the actual column is gross of. The note above the COE block in
        # final3x.build_31 works all four through cell by cell. So the line names what the
        # figure actually is - the cost - and promises no comparison.
        ("COEs and EGI - actual cost, no plan to compare against ($m)",
         f"={G1}!${C31['actual']}${a['coe']}", opts.M2),
        ("Overhead roles - not covered by the allowance ($m)",
         f"={G1}!${C31['var']}${a['overhead']}", opts.M2),
        # without this line the four components above summed to 6.378 under a total of
        # 8.478, because the total includes the GM layer and nothing listed it
        ("The 8 GMs - over/(under) their allowance ($m)",
         f"={G1}!${C31['var']}${a['gm']}", opts.M2),
        ("Total over/(under) archetype, everything comparable ($m)",
         f"={G1}!${C31['var']}${a['comparable']}+{G1}!${C31['var']}${a['gm']}",
         opts.M2),
        ("Groups with no archetype and no funded figure ($m)",
         f"={G1}!${C31['actual']}${a['nofig']}", opts.M2)])

    # The five vacancy lines are a TRUE partition of the vacant count, each one
    # status-qualified. Summing the 2.x lever columns instead only added to 145 by
    # coincidence: the owner's r431 vacancy set to fill fell out of all three buckets
    # while Stevani Kho - a filled role he offshored - fell in, and the two cancelled.
    #
    # These were whole-column COUNTIFS - '2.1 Ampol Retail'!$D:$D - and they were right
    # only by luck. On a 2.x tab columns D and E hold "Squad Size" and "Archetype roles"
    # in the squad table at the top; Status and Vacancy lever are D and E of the FTE
    # block further down. The counts came out right because no squad-table cell happens
    # to contain the word "Vacant" or "Hire" today, which is a property of the data and
    # not of the model. Each range is bounded to its tab's own FTE block now, found by
    # that block's header row rather than typed, so the meaning of the count does not
    # depend on what a squad is called.
    fte = {t: fte_rows(wb[t]) for t in a2}

    def _lev(status, lever):
        return "=" + "+".join(
            f"COUNTIFS('{t}'!$D${fte[t][0]}:$D${fte[t][1]},\"{status}\","
            f"'{t}'!$E${fte[t][0]}:$E${fte[t][1]},\"{lever}\")"
            for t in a2)

    r = block(r, "The vacancy decision", [
        ("Vacant roles", f"={G1}!${C31['vacant']}${gt}", opts.CT),
        ("Vacancies set to hire", _lev("Vacant", "Hire"), opts.CT),
        ("Vacancies set to offshore", _lev("Vacant", "Offshore"), opts.CT),
        ("Vacancies put on hold", _lev("Vacant", "Hold"), opts.CT),
        ("Vacancies set to fill as they are", _lev("Vacant", "Filled"), opts.CT),
        ("Filled roles offshored", _lev("Filled", "Offshore"), opts.CT),
        ("Roles after the decisions set today", f"={G1}!${C31['rafter']}${gt}", opts.CT),
        ("Cost after the decisions set today ($m)", f"={G1}!${C31['after']}${gt}", opts.M2),
        ("Impact of those decisions ($m)", f"={G1}!${C31['after']}${gt}-{G1}!${C31['actual']}${gt}", opts.M2)])

    # portfolio drill-down. The name list is read out of 3.1's archetype block, which
    # carries one row per portfolio.
    r = opts.bar(ws, r, 2, 2, "Portfolio drill-down")
    sel = r
    ws.cell(r, 2).value = "Pick a portfolio"
    ws.cell(r, 2).font = opts.BOLD
    ws.cell(r, 2).alignment = opts.LFT
    ws.cell(r, 2).border = opts.BOX
    pick = ws.cell(r, 3)
    pick.value = "Ampol Retail"
    pick.fill, pick.border = opts.fl(opts.YEL), opts.BOX
    pick.font, pick.alignment = opts.BODY, opts.CEN
    l = wb["Lists"]
    names = [str(l.cell(x, 45).value) for x in range(2, 12) if l.cell(x, 45).value]
    dv = DataValidation(type="list", formula1='"' + ",".join(names) + '"',
                        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(pick)
    r += 1
    # Read out of 3.3, which carries every squad of every portfolio tagged by how it is
    # funded. Reading it off 3.1 instead would have lost the portfolio's overhead cost,
    # because the overhead allowance is a group figure and sits on one row for all ten.
    lo3, hi3 = a["first33"], a["g33"] - 1
    def _sumifs(col, kind=None):
        f = (f"SUMIFS({G3}!${col}${lo3}:${col}${hi3},"
             f"{G3}!${C33['pf']}${lo3}:${C33['pf']}${hi3},$C${sel}")
        if kind:
            f += f',{G3}!${C33["kind"]}${lo3}:${C33["kind"]}${hi3},"{kind}"'
        return f + ")"

    def s33(col, kind=None):
        return f"=ROUND({_sumifs(col, kind)},6)"

    def var33(kind):
        """A variance on the same route 3.1 takes: difference first, rounded once.

        Adding up 3.3's variance column instead added up figures each already rounded to
        six places, and the rounding error survived the total: Ampol Retail's archetype
        variance read (0.766475) here and (0.766476) on 3.1!F6, for the same fact, three
        clicks apart. 3.1 differences the two totals and rounds the difference once, so
        this does too. Same correction the count and cost lines above already carry.

        The third criterion is 3.1's gate, not an extra one: on 3.1 the directly funded
        subtotal covers only the rows that have a funded figure, and the rows that do not
        are shown separately below the comparison. A row with no figure carries "-" in the
        archetype column, which is text, so ">=0" keeps exactly those rows out of BOTH
        sums - archetype costs are never negative. Without it this line would set EGI
        TDD's 0.2997 of actual against a funded figure of nothing and call the difference
        overspend, which is the one thing this tab is built not to do.
        """
        gate = (f',{G3}!${C33["acost"]}${lo3}:${C33["acost"]}${hi3},">=0"')
        return (f"=ROUND({_sumifs(C33['actual'], kind)[:-1]}{gate})-"
                f"{_sumifs(C33['acost'], kind)[:-1]}{gate}),6)")

    for lab, f, nf in (
            ("Roles", s33(C33["roles"]), opts.CT),
            ("Filled", s33(C33["filled"]), opts.CT),
            ("Vacant", s33(C33["vacant"]), opts.CT),
            ("Roles after decisions", s33(C33["rafter"]), opts.CT),
            ("Actual cost ($m)", s33(C33["actual"]), opts.M2),
            ("Of which overhead roles ($m)", s33(C33["actual"], "Overhead"), opts.M2),
            ("Squads priced by an archetype - over/(under) archetype ($m)",
             var33("Archetype"), opts.M2),
            ("Directly funded programmes - over/(under) funded ($m)",
             var33("Directly funded"), opts.M2),
            ("Cost after vacancy decisions ($m)", s33(C33["after"]), opts.M2)):
        ws.cell(r, 2).value = lab
        ws.cell(r, 2).font = opts.BODY
        ws.cell(r, 2).alignment = opts.LFT
        ws.cell(r, 2).border = opts.BOX
        x = ws.cell(r, 3)
        x.value, x.number_format, x.alignment = f, nf, opts.RGT
        x.font, x.border = opts.BODY, opts.BOX
        r += 1
    return r


# ---------------------------------------------------------------------- 4.0 Data QA
def coe_control(wb):
    """The 'must be 0' cells the COE tabs carry, found by their own label."""
    out = []
    for tab in ("1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles"):
        ws = wb[tab]
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if isinstance(v, str) and v.strip().lower().startswith("check"):
                for c in range(3, 12):
                    if isinstance(ws.cell(r, c).value, str) and \
                            ws.cell(r, c).value.startswith("="):
                        out.append((tab, f"{L(c)}{r}"))
                        break
                break
    return out


def build_qa(wb, a, a2):
    ws = wb["4.0 Data QA"]
    f2.wipe(ws)
    ws.column_dimensions["A"].width = 2
    ws.cell(2, 2).value = "Data QA - every difference must read zero"
    ws.cell(2, 2).font = opts.TITLE
    HDR = 4
    opts.head(ws, HDR, 2, ["Check", "Model", "Expected", "Difference"], [72, 16, 16, 14])
    gt, g33 = a["total"], a["g33"]
    lo33, hi33 = a["first33"], g33 - 1
    # "<>Squad" also matches the two empty ledger rows, whose AR is a blank string, so
    # the count came to 65 against 63 real overhead roles. The name column is the guard.
    oh = f'{REV}!$AR$2:$AR${LAST},"<>Squad",{REV}!$B$2:$B${LAST},"<>"'
    checks = [
        # ---- the ledger ----
        ("Roles on 3.1 against the ledger", f"={G1}!${C31['roles']}${gt}",
         f"=COUNTA({REV}!$B$2:$B${LAST})", opts.CT),
        ("Filled on 3.1 against the ledger", f"={G1}!${C31['filled']}${gt}",
         f'=COUNTIFS({REV}!$AK$2:$AK${LAST},"Filled")', opts.CT),
        ("Vacant on 3.1 against the ledger", f"={G1}!${C31['vacant']}${gt}",
         f'=COUNTIFS({REV}!$AK$2:$AK${LAST},"Vacant")', opts.CT),
        ("Filled plus vacant against roles on 3.1", f"={G1}!${C31['filled']}${gt}+{G1}!${C31['vacant']}${gt}",
         f"={G1}!${C31['roles']}${gt}", opts.CT),
        ("Cost on 3.1 against the ledger ($m)", f"={G1}!${C31['actual']}${gt}",
         f"=SUM({REV}!$AA$2:$AA${LAST})/1000000", opts.M2),
        ("Filled plus vacant cost against cost today ($m)",
         f'=(SUMIFS({REV}!$AA$2:$AA${LAST},{REV}!$AK$2:$AK${LAST},"Filled")'
         f'+SUMIFS({REV}!$AA$2:$AA${LAST},{REV}!$AK$2:$AK${LAST},"Vacant"))/1000000',
         f"={G1}!${C31['actual']}${gt}", opts.M2),
        # ---- summary against summary, by a different route ----
        ("Roles on 3.3 against 3.1", f"={G3}!${C33['roles']}${g33}",
         f"={G1}!${C31['roles']}${gt}", opts.CT),
        ("Cost on 3.3 against 3.1 ($m)", f"={G3}!${C33['actual']}${g33}",
         f"={G1}!${C31['actual']}${gt}", opts.M2),
        ("Total including the GM layer against the ledger plus the GM input ($m)",
         f"={G1}!${C31['actual']}${a['grand']}",
         f"=SUM({REV}!$AA$2:$AA${LAST})/1000000+N(Lists!$AG$12)", opts.M2),
        ("Archetype variance - the comparable steps against the comparable subtotal ($m)",
         # the COEs came out of the comparable subtotal when their archetype column stopped
         # restating the actual, so they come out of the check that reconciles to it.
         # Built off each step's own two cells rather than off its rounded variance: three
         # figures rounded to six places and then added differ from one difference rounded
         # once, and the check was reporting that $1 as a failure.
         "=ROUND(" + "+".join(
             f"(N({G1}!${C31['actual']}${a[k]})-N({G1}!${C31['acost']}${a[k]}))"
             for k in ("arch", "direct", "overhead")) + ",6)",
         f"={G1}!${C31['var']}${a['comparable']}", opts.M2),
        ("Roles including the GM layer against the ledger plus the GM count",
         f"={G1}!${C31['roles']}${a['grand']}",
         f"=COUNTA({REV}!$B$2:$B${LAST})+N(Lists!$AG$11)", opts.CT),
        # ---- the design side ----
        ("Archetype cost, squad by squad on 3.3, against 3.1 ($m)",
         f"=SUMIFS({G3}!${C33['acost']}${lo33}:${C33['acost']}${hi33},"
         f'{G3}!${C33["kind"]}${lo33}:${C33["kind"]}${hi33},"Archetype")',
         f"={G1}!${C31['acost']}${a['arch']}", opts.M2),
        ("Directly funded amount, squad by squad on 3.3, against 3.1 ($m)",
         f"=SUMIFS({G3}!${C33['acost']}${lo33}:${C33['acost']}${hi33},"
         f'{G3}!${C33["kind"]}${lo33}:${C33["kind"]}${hi33},"Directly funded")',
         f"={G1}!${C31['acost']}${a['direct']}", opts.M2),
        ("Archetype roles on 3.3 against the priced-per-portfolio list on Lists",
         f"=SUMIFS({G3}!${C33['aroles']}${lo33}:${C33['aroles']}${hi33},"
         f'{G3}!${C33["kind"]}${lo33}:${C33["kind"]}${hi33},"Archetype")',
         # Lists!J12 carries an eleventh portfolio and K12 prices it, so a range stopping
         # at row 11 tested ten of eleven rows and the eleventh could go wrong unseen
         "=SUM(Lists!$K$2:$K$12)", opts.C1),
        ("Offshore archetype against 40% of onshore, first archetype",
         f"=ROUND({A3}!$H$5/{A3}!$G$5,6)", f"=ROUND({A3}!$K$5,6)", opts.C1),
        # ---- the overhead allowance ----
        # the applied columns are driven by the owner's own Times applied cells, so this
        # proves the table's arithmetic - his count times the rate, line by line, against
        # the total he reads - rather than comparing his figure to the derived one. That
        # comparison is stated on 3.2 itself, under the bands, where he is typing.
        ("Applied to the portfolios on 3.2 against its own lines ($m)",
         f"={G2}!${C32['applied']}${a['ohtot32']}",
         f"=ROUND(SUMPRODUCT({G2}!${C32['times']}${a['first32']}:"
         f"${C32['times']}${a['last32']},{G2}!${C32['rate']}${a['first32']}:"
         f"${C32['rate']}${a['last32']}),6)", opts.M2),
        ("Applied where the people sit against the lines that draw it ($m)",
         "=N(Lists!$AJ$9)", '=SUMIF(Lists!$AM$2:$AM$7,"Yes",Lists!$AJ$2:$AJ$7)',
         opts.M2),
        # 3.2's roles and cost columns are the organisation's - the portfolios, the COEs
        # and EGI together - and the only overhead people the ledger does not carry are
        # the GMs, whose count and cost are typed on Lists. Take those out and what is
        # left has to be every ledger role carrying an overhead line, exactly.
        ("Roles in the organisation on 3.2, less the GM layer, against every overhead "
         "role in the ledger",
         f"={G2}!${C32['roles']}${a['ohtot32']}-N(Lists!$AG$11)",
         f"=COUNTIFS({oh})", opts.CT),
        ("Cost in the organisation on 3.2, less the GM layer, against every overhead "
         "role in the ledger ($m)",
         f"={G2}!${C32['cost']}${a['ohtot32']}-N(Lists!$AG$12)",
         f"=SUMIFS({REV}!$AA$2:$AA${LAST},{oh})/1000000", opts.M2),
        # 3.2 states the ledger itself on one row - every role in the organisation, all
        # lines and squads, each counted once. The tab carries its own control on the row
        # under it; this one proves the same fact from 4.0, where a reader looks for it.
        ("Roles in the organisation, all lines and squads, on 3.2 against the ledger",
         f"={G2}!${C32['roles']}${a['all32']}",
         f"=COUNTA({REV}!$B$2:$B${LAST})", opts.CT),
        # 3.2 counts the people who sit in a portfolio off REVIEW rather than reading them
        # back off 3.1, so this compares two independent routes to the same figure.
        ("Overhead on 3.1 against the cost that sits in the portfolios on 3.2 ($m)",
         f"={G1}!${C31['actual']}${a['overhead']}",
         f"={G2}!${C32['cost']}${a['ohpf32']}", opts.M2),
        # the two tabs stated two different allowances for the same 43 overhead roles -
        # 3.2 read Lists, which priced the per-platform lines over a typed 30 platforms,
        # and 3.1 read what the ten design tabs actually allow. Lists now counts the
        # platforms off the design tabs, and this is the check that keeps it that way.
        # both sides derived, so this stays a tie between the two tabs whatever the owner
        # types into Times applied: Lists holds the count the model carries and 3.1's
        # overhead step is what the ten design tabs actually draw.
        ("The allowance the model carries against the overhead step on 3.1 ($m)",
         "=N(Lists!$AJ$9)",
         f"={G1}!${C31['acost']}${a['overhead']}", opts.M2),
        ("The two 'of which' bands on 3.2 against the overheads total above them ($m)",
         f"={G2}!${C32['applied']}${a['ohpf32']}"
         f"+{G2}!${C32['applied']}${a['ohout32']}",
         f"={G2}!${C32['applied']}${a['ohtot32']}", opts.M2),
        # the working tabs price each portfolio at what its own design tab says, which
        # includes the Business Partner, Domain Architect and Leadership allowance whose
        # people sit in the COEs and above the ledger. 3.1 carries those in its COE step
        # and its GM line instead, so the two tie only once that 6.60 is taken out - and
        # this check is what proves it still does.
        ("Archetype on the working tabs, less what sits outside them, against 3.1 ($m)",
         "=" + "+".join(f"N('{t}'!${L(S['acost'])}${i['total_row']})"
                        for t, i in a2.items() if i.get("elsewhere_row"))
         + "-" + "-".join(f"N('{t}'!${L(S['acost'])}${i['elsewhere_row']})"
                          for t, i in a2.items() if i.get("elsewhere_row")),
         f"={G1}!${C31['acost']}${gt}", opts.M2),
        ("Overhead allowance on 3.1 against the working tabs ($m)",
         f"={G1}!${C31['acost']}${a['overhead']}",
         "=" + "+".join(f"N('{t}'!${L(S['acost'])}${i['overhead_row']})"
                        for t, i in a2.items() if i["overhead_row"]), opts.M2),
        # ---- the lever ----
        ("Roles after decisions against roles less anything on hold",
         f"={G1}!${C31['rafter']}${gt}",
         f"={G1}!${C31['roles']}${gt}-" + "-".join(
             f"N('{t}'!${L(S['hold'])}${i['total_row']})" for t, i in a2.items()),
         opts.CT),
        # the file ships with the owner's decisions set - five cyber holds, three SA&D
        # holds, two BP&T offshores, Stevani Kho offshored - so after-decisions no longer
        # equals today. The identity that always holds: the bridge's decision impact
        # equals the sum of the working tabs' own impacts.
        ("Decision impact on 3.1 against the working tabs ($m)",
         f"={G1}!${C31['actual']}${gt}-{G1}!${C31['after']}${gt}",
         "=" + "+".join(
             f"N('{t}'!${L(S['actual'])}${i['total_row']})"
             f"-N('{t}'!${L(S['after'])}${i['total_row']})" for t, i in a2.items()),
         opts.M2),
    ]
    # the COE tabs price their own planned spend off their own roles list. Where that list
    # is short of the ledger the planned spend is short too, and nothing else in the file
    # notices: 1.12 was missing three roles worth $747,896.
    # the COE tabs now state planned spend net of portfolio funding and after the
    # owner's own levers, so a gross-spend tie no longer exists on them. The structural
    # tie that survives his redesign is the roles count.
    for tab, pf, cell in (("1.11 BP&T", "COE BP&T", "$C$6+'1.11 BP&T'!$C$7"),
                          ("1.12 SA&D", "COE SA&D", "$C$6+'1.12 SA&D'!$C$7"),
                          ("1.13 Cyber Roles", "COE Cyber",
                           "$C$6+'1.13 Cyber Roles'!$C$7")):
        checks.append((f"{tab} roles against the ledger",
                       f"='{tab}'!{cell}",
                       f'=COUNTIFS({REV}!$AJ$2:$AJ${LAST},"{pf}",'
                       f'{REV}!$B$2:$B${LAST},"<>")', opts.CT))
    for tab, cell in coe_control(wb):
        checks.append((f"{tab} own control - roles listed against roles counted",
                       f"='{tab}'!{cell}", "=0", opts.CT))
    # The archetype total on a working tab has to be the Total Cost on its own design tab -
    # squads plus the portfolio and platform overhead the archetype allows for. The working
    # tabs used to state the squads only, so 2.8 read 7.90 where 1.8 reads 9.03, and every
    # portfolio's archetype was short by its own overhead allowance with nothing to catch it.
    from build_2xfix import DESIGN
    for tab, inf in a2.items():
        d = DESIGN.get(tab)
        if not d:
            continue
        tc = next((r for r in range(1, 16)
                   if str(wb[d].cell(r, 2).value or "").strip() == "Total Cost"), None)
        if tc:
            checks.append((f"{tab} archetype total against Total Cost on {d} ($m)",
                           f"='{tab}'!${L(S['acost'])}${inf['total_row']}",
                           f"='{d}'!$F${tc}", opts.M2))
    for tab, inf in a2.items():
        t = inf["total_row"]
        checks.append((f"{tab} roles against the ledger",
                       f"='{tab}'!${L(S['roles'])}${t}",
                       f"=COUNTIFS({REV}!$AJ$2:$AJ${LAST},\"{inf['pf']}\")", opts.CT))
        checks.append((f"{tab} cost against the ledger ($m)",
                       f"='{tab}'!${L(S['actual'])}${t}",
                       f"=SUMIFS({REV}!$AA$2:$AA${LAST},"
                       f"{REV}!$AJ$2:$AJ${LAST},\"{inf['pf']}\")/1000000", opts.M2))
    # The check column holds a sentence, not a heading, and the longest of them is the
    # portfolio-by-portfolio archetype tie, which names two tabs: 96 characters against a
    # 72-wide column, so eight of the checks read "...against Total Cost on 1.4 TDD Group
    # Functions (" and the reader could not tell which check had failed. The column is as
    # wide as its own longest label - the rest of the tab is one line per row and stays
    # that way, which is what this tab's style is.
    ws.column_dimensions["B"].width = max(72, max(len(lab) for lab, _, _, _ in checks) + 2)
    r = HDR + 1
    for lab, m, e, nf in checks:
        ws.cell(r, 2).value = lab
        ws.cell(r, 2).font = opts.BODY
        ws.cell(r, 2).alignment = opts.LFT
        ws.cell(r, 2).border = opts.BOX
        for col, f in ((3, m), (4, e), (5, f"=ROUND($C{r}-$D{r},6)")):
            x = ws.cell(r, col)
            x.value = f
            x.number_format = nf if col < 5 else (
                opts.CTL_C if nf in (opts.CT, opts.C1) else opts.CTL_M)
            x.alignment, x.font, x.border = opts.RGT, opts.BODY, opts.BOX
        r += 1
    opts.row(ws, r, 2, ["Checks failing", None, None, None], [None] * 4,
             bg=opts.MID, bold=True, top=True)
    ws.cell(r, 2).alignment = opts.LFT
    x = ws.cell(r, 5)
    x.value = f'=COUNTIF($E{HDR+1}:$E{r-1},"<>0")'
    x.number_format, x.alignment = opts.CTL_C, opts.RGT
    return len(checks)


# The retired raw-data tabs keep whatever marks the owner put on them. Everything in the
# live model uses one input colour.
RETIRED = {"Squads", "Added data", "Sheet2", "FY26 Budget (superseded)",
           "squad mapping (superseded)"}


def cream(wb):
    """One input colour across the live model. The file had bright yellow and cream both
    meaning 'typed input', side by side on the same tabs."""
    n = 0
    for ws in wb.worksheets:
        if ws.title in RETIRED:
            continue
        for row in ws.iter_rows():
            for c in row:
                fl = c.fill
                try:
                    rgb = str(fl.start_color.rgb or "").upper() \
                        if fl and fl.patternType else ""
                except Exception:
                    rgb = ""
                if rgb == "FFFFFF00":
                    c.fill = opts.fl("FFFFF2CC")
                    n += 1
    return n


def run(src, dst):
    global LAST
    f2._boot_last(src)
    LAST = f2.LAST
    wb = openpyxl.load_workbook(src)
    a = anchors(wb)
    a2 = json.load(open("anchors_final.json"))
    n = repoint(wb, a)
    build_exec(wb, a, a2)
    k = build_qa(wb, a, a2)
    nc = cream(wb)
    wb.save(dst)
    return [f"Lists archetype-roles lookup repointed on {n} portfolios",
            f"{nc} inputs recoloured to cream across the live model",
            "Exec Summary rebuilt on design against actual, with a portfolio drill-down",
            f"4.0 Data QA rebuilt: {k} checks, model / expected / difference",
            f"anchors: 3.1 ledger r{a['total']} grand r{a['grand']}, "
            f"3.2 overhead total r{a['ohtot32']}, 3.3 total r{a['g33']}"]


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
