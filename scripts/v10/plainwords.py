"""Say it the way the owner says it.

"Ledger" is not his word. It is mine, and it reached 87 labels - every control row on
every working tab, the Exec Summary's headline lines, half of 4.0 Data QA. He calls the
same thing the role mapping, and the tab is literally named "REVIEW - Complete Role
Mapping". A technology GM opening this file has a role mapping in front of them, not a
ledger.

Everything else the label scan turned up is his own vocabulary and stays: archetype is
his word and he ruled on it in as many words ("Do not call it design. It is the
archetype."); COE, EGI, BP&T, SA&D and Lights On are Ampol's; "Vacancy lever" is his.
Replacing those would be me overwriting his language with mine, which is the mistake this
file exists to undo.

Run last, on the finished workbook, so it catches every label whoever wrote it.
"""
import re
import sys

import openpyxl

# ordered: the longest phrase first, so a short rule cannot eat a long one
SAY = [
    ("the 531-role ledger", "the 531 roles in the role mapping"),
    ("531-role ledger", "531 roles in the role mapping"),
    ("Cost of the 531 roles in the ledger", "Cost of the 531 roles in the role mapping"),
    ("roles in the ledger carry", "roles in the role mapping carry"),
    ("Roles in the ledger", "Roles in the role mapping"),
    ("roles in the ledger", "roles in the role mapping"),
    ("outside the ledger", "outside the role mapping"),
    ("Above the ledger", "Above the role mapping"),
    ("above the ledger", "above the role mapping"),
    ("against the ledger", "against the role mapping"),
    ("in the ledger", "in the role mapping"),
    ("the ledger plus", "the role mapping plus"),
    ("the ledger", "the role mapping"),
    # "Budget to draw down" was on this list too, swapped for "Budget available". It should
    # not have been. "Ledger" is mine and he never wrote it; "Budget to draw down" is his,
    # it is in his own review workbook on 1.11, 1.12 and 1.13, and D83 settled that his
    # labels win. Rewording his own heading to one I preferred is the exact mistake this
    # file was written to undo, committed inside the file undoing it.
    # applied after the swap above: three labels that the longer wording pushes past
    # their column. The count they drop is stated on the same page either way.
    ("outside the 531 roles in the role mapping", "outside the role mapping"),
    ("above the 531 roles in the role mapping", "above the role mapping"),
    ("against every overhead role in the role mapping",
     "against overhead roles in the role mapping"),
    # one spelling. His own squad type is "Strategic Programs"; the built prose said
    # "programmes" in some places and "program" in others, so the same thing wore two
    # spellings on one page. His spelling wins, as always.
    ("programmes", "programs"),
    ("programme", "program"),
]


# His exact wording, put back where a pass normalised it for no reason but tidiness.
# tab -> {cell: (what the build now says, what he wrote)}. The build's text is stated so a
# rename that lands somewhere else fails loudly instead of overwriting the wrong cell.
HIS = {
    "0.2 Data Config": {
        # his own abbreviation, on his own config tab. M5 four rows up says "Allocation %"
        # in his book too - the two headings differ because he wrote them that way.
        "M13": ("Allocation %", "Alloc %"),
    },
}


def restore(wb, out):
    for tab, cells in HIS.items():
        if tab not in wb.sheetnames:
            out.append(f"{tab} is not in the workbook - nothing restored")
            continue
        for cell, (now, his) in cells.items():
            cur = wb[tab][cell].value
            if cur == his:
                continue
            if cur != now:
                out.append(f"{tab}!{cell} holds {str(cur)[:40]!r}, expected {now!r} - "
                           f"left alone")
                continue
            wb[tab][cell] = his
            out.append(f"{tab}!{cell} {now!r} -> {his!r} - his wording")


# ---------------------------------------------------------------- the no-dash sweep
# His wave-M ruling, in as many words: no dashes anywhere in any output cell. Three things
# put one on the page and all three are swept here, on the finished workbook, so nothing
# a later pass writes can survive:
#
#   a typed "-" in a cell            -> blank. A cell with nothing in it is blank.
#   a "-" fallback inside a formula  -> "", which renders blank for the same reason.
#   the third section of a number    -> dropped. That section is what a zero renders as,
#   format                              and a zero is 0.00 or 0, not a dash.
#
# The writers were changed at source as well (opts.py's five formats, the IFERROR
# fallbacks in final2x/final3x/final4x/actuals). This sweep is the belt to that braces: it
# is the only pass that sees the whole finished file, and it is what makes "no dash
# anywhere" checkable in one place rather than trusted across eighteen scripts.
#
# The owner's own source tabs are exempt from the wording sweep above and from this one:
# 0.1, 0.3 and 0.4 are his and the chain is locked out of them (D101).
SOURCE_TABS = ("0.1 Budget Table (Fin)", "0.4 Presentation Pack",
               "0.3 Squad Archetypes")
DASHES = "-‐‑‒–—―−"
FMT_DASH = re.compile(r';(?:"[-‐-―−]"|_\(?[-‐-―−]_?\)?)\s*$')


def no_dashes(wb, out):
    vals = fmls = fmts = 0
    for ws in wb.worksheets:
        if ws.title in SOURCE_TABS:
            continue
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.strip() and all(ch in DASHES + " " for ch in v):
                    c.value = None
                    vals += 1
                elif isinstance(v, str) and v.startswith("="):
                    new = v
                    for d in DASHES:
                        new = new.replace(f'"{d}"', '""')
                    if new != v:
                        c.value = new
                        fmls += 1
                nf = c.number_format or ""
                if FMT_DASH.search(nf):
                    c.number_format = FMT_DASH.sub("", nf)
                    fmts += 1
    out.append(f"no-dash sweep: {vals} typed dashes blanked, {fmls} formula fallbacks now "
               f'return "", {fmts} number formats no longer render a zero as a dash')


# "seat" and "design" are banned words in cell text - he ruled on both ("The word 'seat' is
# never used"; "Do not call it design. It is the archetype."). En and em dashes are banned
# in prose for the same reason they are banned in figures: hyphens only. This does not
# rewrite anything - a banned word in a label is a build defect, not a formatting slip - it
# reports, so the gate can fail on it and the sentence can be rewritten at source.
BANNED = (("seat", re.compile(r"\bseats?\b", re.I)),
          ("design", re.compile(r"\bdesign(s|ed|ing)?\b", re.I)))


def banned_words(wb, out):
    hits = []
    for ws in wb.worksheets:
        if ws.title in SOURCE_TABS or ws.title.startswith("REVIEW"):
            continue
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                for word, pat in BANNED:
                    if pat.search(v):
                        hits.append(f"{ws.title}!{c.coordinate} says {word!r}: {v[:56]}")
                if any(d in v for d in "–—"):
                    hits.append(f"{ws.title}!{c.coordinate} carries an en/em dash: "
                                f"{v[:56]}")
    out.append(f"banned words and en/em dashes in cell text: {len(hits)}")
    for h in hits[:12]:
        out.append(f"  {h}")
    if len(hits) > 12:
        out.append(f"  ... {len(hits) - 12} more")
    return hits


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    n = 0
    hit = []
    for ws in wb.worksheets:
        # the owner's own source tabs are never reworded
        if ws.title in ("0.1 Budget Table (Fin)", "0.4 Presentation Pack",
                        "0.3 Squad Archetypes"):
            continue
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                new = v
                for old, rep in SAY:
                    if old in new:
                        new = new.replace(old, rep)
                if new != v:
                    c.value = new
                    n += 1
                    if len(hit) < 6:
                        hit.append(f"{ws.title}!{c.coordinate}: {new[:60]}")
    # and inside the formula-built labels, which are strings the reader still reads
    nf = 0
    for ws in wb.worksheets:
        if ws.title in ("0.1 Budget Table (Fin)", "0.4 Presentation Pack",
                        "0.3 Squad Archetypes"):
            continue
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=") and '"' in v):
                    continue
                new = v
                for old, rep in SAY:
                    if f'"{old}' in new or f'{old}"' in new or f' {old} ' in new:
                        new = new.replace(old, rep)
                if new != v:
                    c.value = new
                    nf += 1
    back = []
    restore(wb, back)
    # last, after every reword, so a sentence the sweep above rebuilt cannot smuggle a dash
    # back in and the banned-word report is of the file as it ships
    no_dashes(wb, back)
    banned_words(wb, back)
    wb.save(dst)
    return [f"{n} labels and {nf} formula-built labels now say role mapping, not ledger",
            *[f"  e.g. {h}" for h in hit], *back]


if __name__ == "__main__":
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
