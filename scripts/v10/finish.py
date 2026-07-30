"""The last presentation pass, from an independent review of the rendered tabs.

Every item here was found by looking at the sheets, not the cells, and several of them had
survived earlier passes because a test was subtly wrong rather than absent:

  the [Red] number-format strip tested for "[Red]" and the file writes "[RED]", so 240 cells
  across 0.2, 1.11, 1.12, 1.13 and REVIEW still printed negatives in red

  0.1 Budget Table (Fin) and 0.4 Presentation Pack are raw pastes from Finance and the deck -
  a full red / amber / green traffic-light grid, 741 [RED] formats, Arial and Aptos, four
  extra blues, hidden columns. They are the source and they stay in the file, but they cannot
  ship visible

  0.2 Data Config used the reserved subtotal grey as a zebra stripe, in patches, and one data
  row was 76pt tall against 14.25 for the rest, which dragged the second table's header band
  to a different depth from the first

  fifteen sentences were sitting in data cells, one of them the only Cambria cell in the
  workbook, and one of them a 122-character section label I had written myself

0.3 Squad Archetypes is exempt from the two sweeps here that would otherwise reach it - the
orphan-cream strip and the Calibri sweep. It is the owner's cost library, a source tab like
0.1 and 0.4 rather than a built one, and the whole chain now leaves it alone; regress2707
proves it cell-for-cell against rev.xlsx.

The sentence sweep is the most destructive thing in this file and it had no idea whose
sentences it was clearing. Ten of them were the owner's: his two open questions on 1.8,
his notes on 1.11, 1.12 and 1.13, his two on 1.2. It now reads rev.xlsx first and will not
clear a cell his own workbook writes at the same address. That one rule is what makes the
sweep safe, and it is a rule about provenance rather than a longer list of exceptions, so
it holds for whatever he writes next.
"""
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter as L

import opts

NONE_FILL = PatternFill()
NO_BORDER = Border()
HIDE = ["0.1 Budget Table (Fin)", "0.4 Presentation Pack"]
# The owner's source tabs, exempt from the presentation sweeps below. 0.1 and 0.4 were
# already exempt in practice - they are hidden by the time those sweeps run, and both skip
# hidden tabs. 0.3 Squad Archetypes is a source tab too, his archetype price table, but it
# ships visible, so it has to be named. no_red stays workbook-wide and still reaches it:
# the shipped workbook is gated on there being no [Red] format anywhere, and 0.3 carries
# none, so that sweep is a no-op on his tab rather than an exception to it.
SOURCE = {"0.1 Budget Table (Fin)", "0.4 Presentation Pack", "0.3 Squad Archetypes"}
REVIEW = "REVIEW - Complete Role Mapping"

# sentences sitting in a data cell. Cleared, not reworded: the fact each one carried is
# either on the face of the model now or belongs in the decisions log.
PROSE = {"1.5 P&C": ["B11"], "1.11 BP&T": ["B9", "B47", "B49"],
         "1.12 SA&D": ["B9", "B53", "B55"], "1.2 Customer": ["L15", "L18"],
         "1.8 Energy Solutions & B2B": ["E17", "E18", "B21"],
         # H10 came off this list with D118: the chain now writes his Data COE note
         # there (repair_design.add_data_coe_note), and this sweep was eating it four
         # passes after it was written
         "0.2 Data Config": ["H9", "H14"], "1.13 Cyber Roles": ["B73"]}
# labels that were too long to sit on a row carrying figures
SHORTEN = {"3.1 Cost Bridge": {
    "Directly funded programmes and platforms - no archetype prices them, so the "
    "comparison is the amount funded on the 1.x tab":
        "Directly funded programmes and platforms - funded on the 1.x tab",
    "Overhead roles in the portfolios - the allowance is the comparison, line by line "
    "on 3.2": "Overhead roles in the portfolios - the allowance is on 3.2",
    "COEs and EGI - priced off the planned spend on their own 1.x tabs":
        "COEs and EGI - planned spend on their own 1.x tabs",
    "Squads priced by an archetype - detail on 3.3":
        "Squads priced by an archetype - detail on 3.3"}}
TITLES = {"3.2 Overhead & Leadership":
              "Archetype Overhead & Actual Leadership comparison",
          "1.13 Cyber Roles": "Cyber, Risk & Service Operations - roles and funding"}
# one width profile for the three COE design tabs, wide enough for all three
COE_W = {"B": 46, "C": 38, "D": 36, "E": 11, "F": 11, "G": 15, "H": 15, "I": 15, "J": 15,
         "K": 15}
COE_TABS = ["1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles"]
REV = "rev.xlsx"


def rev_literals(path=REV):
    """Every literal the owner's review workbook holds, by sheet and coordinate.

    Read-only and read once. A formula is not his content in the sense that matters here -
    the chain rebuilds formulas by design - so only typed values are collected.

    If his workbook cannot be read this stops. The sweep below is only safe because this
    map exists; running it without one is how ten of his notes were lost last time.
    """
    if not os.path.exists(path):
        near = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            os.path.basename(path))
        path = near if os.path.exists(near) else path
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception as e:                                  # noqa: BLE001
        raise SystemExit(f"finish: cannot read {path} ({e}). The sentence sweep needs "
                         f"the owner's workbook to know whose sentences it is clearing.")
    keep = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None or (isinstance(v, str)
                                 and (v.startswith("=") or not v.strip())):
                    continue
                keep[(ws.title, c.coordinate)] = v
    wb.close()
    return keep, f"rev whitelist: {len(keep):,} of his own literals, from {path}"


def his(keep, tab, ref, value):
    """Is this cell the owner's writing?

    His value at his address is the plain case. The chain also corrects a handful of his
    notes on purpose - a tab it renamed, a reference that went stale, a typed slip - so a
    literal still standing where he wrote one is his too, corrected. What it will not
    protect is an address he left empty and something else later wrote a sentence into,
    which is exactly what the entries this sweep still clears are.
    """
    if keep is None or (tab, ref) not in keep:
        return None
    if keep[(tab, ref)] == value:
        return "his, unchanged"
    if isinstance(value, str) and not value.startswith("="):
        return "his, as the chain corrected it"
    return None


def no_red(wb):
    """The strip that missed 240 cells because Excel writes [RED], not [Red]."""
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                nf = c.number_format or ""
                if "[red]" not in nf.lower():
                    continue
                c.number_format = (opts.M2 if ".00" in nf else
                                   (opts.M3 if ".000" in nf else
                                    (opts.PCT if "%" in nf else opts.CT)))
                n += 1
    return [f"[RED] stripped from {n} number formats, case-insensitively this time"]


def hide_sources(wb):
    out = []
    for t in HIDE:
        if t in wb.sheetnames and wb[t].sheet_state == "visible":
            wb[t].sheet_state = "hidden"
            out.append(f"{t}: hidden - it is a raw paste, not a built tab")
    return out


def drop_prose(wb, keep):
    n = 0
    held = []
    for tab, cells in PROSE.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for ref in cells:
            v = ws[ref].value
            if v is None:
                continue
            why = his(keep, tab, ref, v)
            if why:
                held.append(f"{tab}!{ref} kept - {why}")
                continue
            ws[ref].value = None
            ws[ref].font = opts.BODY
            n += 1
    for tab, m in SHORTEN.items():
        ws = wb[tab]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip() in m:
                    c.value = m[c.value.strip()]
    return [f"{n} sentences cleared from data cells, and the long section labels on 3.1 "
            f"cut to fit their row"] + held


def fix_02(wb):
    """0.2 Data Config: one row height, no zebra, a real total row."""
    ws = wb["0.2 Data Config"]
    out = []
    for r in range(1, 40):
        h = ws.row_dimensions[r].height
        # a row carrying one of his Actions notes keeps the height the wrapped note
        # needs - design2707 sizes those, and flattening them clips his own words
        if isinstance(ws.cell(r, 9).value, str) and ws.cell(r, 9).value.strip():
            continue
        if h and h > 50:
            ws.row_dimensions[r].height = 14.25
            out.append(f"0.2 row {r} height {h} -> 14.25")
    # the platform-overhead rate block foots on the page: at two decimals 0.084 + 0.081
    # displayed as 0.08 + 0.08 under a subtotal displaying 0.17, and that column is the
    # first thing a finance reader foots on the input tab
    for cell in ("N14", "N15", "N16"):
        if isinstance(ws[cell].value, (int, float, str)) and ws[cell].value is not None:
            ws[cell].number_format = "0.000"
    out.append("0.2!N14:N16 shown at three decimals, so the per-platform rate foots")
    # every header row on the tab, sized by the same wrap arithmetic the builders use,
    # so the two overhead bands come out the same depth as each other and nothing clips
    for r in range(1, 30):
        cells = [c for c in range(2, 14) if _fill(ws.cell(r, c)) == opts.NAVY
                 and isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.strip()]
        if len(cells) < 3:
            continue
        lines = 1
        for c in cells:
            ws.cell(r, c).alignment = opts.CEN
            w = ws.column_dimensions[L(c)].width or 8.43
            lines = max(lines, opts.wrap_lines(ws.cell(r, c).value, w))
        ws.row_dimensions[r].height = max(32, 14 * lines + 6)
    # the subtotal grey is reserved. Strip it from every row that is not a total.
    n = 0
    for r in range(6, 30):
        lab = str(ws.cell(r, 2).value or "").strip()
        istotal = lab.startswith(("Total", "Budget", "Variance"))
        for c in range(2, 8):
            x = ws.cell(r, c)
            if _fill(x) != opts.GREY:
                continue
            if istotal:
                continue
            x.fill = NONE_FILL
            n += 1
    # and the total row gets the block-total grey across the table
    for r in range(6, 30):
        if str(ws.cell(r, 2).value or "").strip() == "Total":
            for c in range(2, 8):
                ws.cell(r, c).fill = opts.fl(opts.MID)
                ws.cell(r, c).font = opts.BOLD
            out.append(f"0.2 row {r} is the total row, shaded as one")
    out.append(f"0.2: subtotal grey removed from {n} data cells")
    return out


def _fill(c):
    try:
        return str(c.fill.start_color.rgb or "").upper() if c.fill and c.fill.patternType \
            else ""
    except Exception:
        return ""


def bars_and_totals(wb):
    """Bars that stop short of their table, and total rows shaded in two greys."""
    out = []
    # a bar in column H over an H..K table, merged on 1.2 and not on 1.1
    ws = wb["1.1 Ampol Retail"]
    for r in range(1, 30):
        v = str(ws.cell(r, 8).value or "").strip()
        if v.startswith("Other funding") and _fill(ws.cell(r, 8)) == opts.BARC:
            ws.cell(r, 8).value = "Other funding"          # it had a trailing space
            for c in range(9, 12):
                ws.cell(r, c).fill = opts.fl(opts.BARC)
                ws.cell(r, c).font = opts.BARF
            out.append(f"1.1!H{r} bar extended across H..K")
    # three empty outlined boxes hanging off the end of 1.13's Roles bar
    ws = wb["1.13 Cyber Roles"]
    for r in range(1, 30):
        if str(ws.cell(r, 2).value or "").strip() == "Roles":
            for c in range(8, 12):
                x = ws.cell(r, c)
                if x.value is None:
                    x.border, x.fill = NO_BORDER, NONE_FILL
            out.append(f"1.13 row {r}: empty bordered cells past the bar cleared")
    # every total row in one grey, across the cells that carry the table
    for tab in [t for t in wb.sheetnames if re.match(r"^1\.\d+ ", t)]:
        ws = wb[tab]
        for r in range(1, min(ws.max_row, 90) + 1):
            greys = [c for c in range(2, 13)
                     if _fill(ws.cell(r, c)) in (opts.GREY, opts.MID)]
            if not greys:
                continue
            has = any(ws.cell(r, c).value is not None for c in range(2, 13))
            if not has:                                     # an empty shaded row
                for c in greys:
                    ws.cell(r, c).fill = NONE_FILL
                    ws.cell(r, c).border = NO_BORDER
                continue
            if len({_fill(ws.cell(r, c)) for c in greys}) > 1:
                for c in range(min(greys), max(greys) + 1):
                    ws.cell(r, c).fill = opts.fl(opts.MID)
    out.append("total rows shaded in one grey, empty shaded rows cleared")
    return out


def empty_inputs(wb):
    """Cream on an empty cell reads as a box waiting for a number that nothing wants.

    And cream on a formula is a worse lie: cream is the file's one promise that a cell is
    a typed input the reader may change, so a cream cell computing =0.5*I14 both hides a
    rule and invites the reader to destroy it by typing over it. The formula keeps its
    value; only the input paint comes off."""
    n = m = 0
    for ws in wb.worksheets:
        if ws.sheet_state != "visible" or ws.title in SOURCE:
            continue
        for row in ws.iter_rows():
            for c in row:
                if _fill(c) != opts.YEL:
                    continue
                if c.value is None:
                    # keep it where the row it sits on carries a label; otherwise it is
                    # an orphan
                    if not any(isinstance(ws.cell(c.row, k).value, str)
                               and ws.cell(c.row, k).value.strip()
                               for k in range(2, c.column)):
                        c.fill, c.border = NONE_FILL, NO_BORDER
                        n += 1
                elif isinstance(c.value, str) and c.value.startswith("="):
                    c.fill = NONE_FILL
                    m += 1
    return [f"{n} orphan cream cells cleared",
            f"{m} formula cells stripped of the input colour - cream is for typed "
            f"inputs only"]


def review_font(wb):
    """Calibri everywhere a reader can see. The strays were the ledger's header row, two
    notes in its spare column, and four empty cells that had kept the theme default."""
    n = 0
    for ws in wb.worksheets:
        if ws.sheet_state != "visible" or ws.title in SOURCE:
            continue
        for row in ws.iter_rows():
            for c in row:
                f = c.font
                if f and (f.name or opts.FN) != opts.FN:
                    c.font = Font(name=opts.FN, size=max(f.size or 11, 10), bold=f.bold,
                                  color=(f.color.rgb if f.color and isinstance(
                                      getattr(f.color, "rgb", None), str) else None))
                    n += 1
    return [f"{n} cells set to Calibri across every visible tab"]


def _unused(wb):
    ws = wb[REVIEW]
    n = 0
    for c in ws[1]:
        f = c.font
        if f and (f.name or "") != opts.FN:
            c.font = Font(name=opts.FN, size=max(f.size or 11, 10), bold=f.bold,
                          color=(f.color.rgb if f.color and
                                 isinstance(getattr(f.color, "rgb", None), str)
                                 else None))
            n += 1
    return [f"{n} header cells on the ledger set to Calibri"]


def coe_widths(wb):
    for t in COE_TABS:
        for k, v in COE_W.items():
            wb[t].column_dimensions[k].width = v
    return [f"one width profile across {', '.join(COE_TABS)}"]


def titles(wb, keep=None):
    n = 0
    out = []
    for tab, t in TITLES.items():
        if tab not in wb.sheetnames:
            continue
        x = wb[tab].cell(2, 2)
        # a title he wrote himself stands - 1.13's is his, and his newest wording wins
        w = his(keep, tab, "B2", x.value)
        if w:
            out.append(f"{tab}!B2 title left alone - {w}")
            continue
        x.value, x.font = t, opts.TITLE
        n += 1
    return out + [f"{n} titles matched to their tab name"]


def strays(wb, keep):
    out = []
    # a naked ratio with no label, in an otherwise blank area
    for tab, ref in (("1.1 Ampol Retail", "E18"), ("1.9 Commercial Fuels", "E17"),
                     ("1.10 Z Retail", "E16")):
        ws = wb[tab]
        if his(keep, tab, ref, ws[ref].value):
            out.append(f"{tab}!{ref} left alone - his")
            continue
        if isinstance(ws[ref].value, str) and ws[ref].value.startswith("="):
            ws[ref].value = None
            out.append(f"{tab}!{ref} unlabelled ratio cleared")
    return out


def review_cream(wb):
    """The input colour on the ledger marks cells nobody types into.

    `cream()` recolours every bright-yellow cell in the live model, and REVIEW had four -
    including a column header. The one place an input genuinely belongs on the ledger is the
    cost-override column, so that is the only place it is kept.
    """
    ws = wb[REVIEW]
    n = 0
    for row in ws.iter_rows():
        for c in row:
            if _fill(c) == opts.YEL and c.column != 47:
                c.fill = NONE_FILL
                n += 1
    return [f"{n} stray input-coloured cells cleared from the ledger"]


def align_3x(wb):
    """The four summary tabs are meant to read as one family; their label columns were
    34 / 26 / 22 / 16 wide, so the left edge of the table landed somewhere different on
    every one of them."""
    n = 0
    for t in [x for x in wb.sheetnames if re.match(r"^3\.\d ", x)]:
        wb[t].column_dimensions["B"].width = 34
        n += 1
    return [f"column B set to 34 on all {n} summary tabs, so the left edge lines up"]


def snapshot_flag(wb):
    """His hand-typed role-review columns on 1.4, 1.5 and 1.6, named as a snapshot.

    The blocks are his - "Nbr Archetype Roles", "Published Roles", "Vacant Now" - and
    they disagree with the live counts on the same tabs on five of the eight squads they
    cover, because the model has moved since he typed them. His numbers are not touched;
    the model says what they are beside them, which is the rule for content of his the
    model disagrees with. Undated and unlabelled they read as a second, contradicting set
    of live figures."""
    out = []
    for t in ("1.4 TDD Group Functions", "1.5 P&C", "1.6 Finance"):
        if t not in wb.sheetnames:
            continue
        ws = wb[t]
        for r in range(1, min(ws.max_row, 60) + 1):
            for c in range(18, 26):
                if str(ws.cell(r, c).value or "").strip() == "Nbr Archetype Roles":
                    tgt = ws.cell(r - 1, c)
                    if tgt.value is None:
                        tgt.value = ("Role review notes - a hand-typed snapshot, not "
                                     "live model figures")
                        tgt.font = openpyxl.styles.Font(name="Calibri", size=10,
                                                        italic=True)
                        out.append(f"{t}: review block at {L(c)}{r} named as a snapshot")
                    break
            else:
                continue
            break
    return out or ["no review blocks found to flag"]


def qa_bar(wb):
    """4.0 was the only built tab with a header row and no bar over it.

    Not the tab title's own sentence again: the title two rows up already says every
    difference must read zero, and a bar repeating it word for word read as a mistake."""
    ws = wb["4.0 Data QA"]
    if _fill(ws.cell(3, 2)) != opts.BARC:
        opts.bar(ws, 3, 2, 4, "Live checks - both sides computed from this file")
    elif str(ws.cell(3, 2).value or "") == "Every difference must read zero":
        ws.cell(3, 2).value = "Live checks - both sides computed from this file"
    return ["4.0 given a section bar, like every other built tab"]


def run(src, dst, rev=REV):
    wb = openpyxl.load_workbook(src)
    keep, note = rev_literals(rev)
    out = ([note] + no_red(wb) + hide_sources(wb) + drop_prose(wb, keep) + fix_02(wb)
           + bars_and_totals(wb) + empty_inputs(wb) + review_font(wb) + coe_widths(wb)
           + titles(wb, keep) + strays(wb, keep) + review_cream(wb) + align_3x(wb)
           + snapshot_flag(wb) + qa_bar(wb))
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(*sys.argv[1:]):
        print("  ", x)
