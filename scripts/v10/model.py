"""Canonical derivations for the TDD Cost Calculator, from Lee's raw REVIEW columns only.

Every grouping here traces to columns A..AB of 'REVIEW - Complete Role Mapping'.
Nothing is invented. Columns AC..AQ of the existing workbook are a previous session's
guesses and are rebuilt from these rules.

Rules
-----
Portfolio  col I, case/label-normalised (PORTFOLIO_MAP). 17 raw values -> 14 tabs.
Squad      col K, typo-merged (SQUAD_MERGE). Blank squad -> falls back per BLANK_SQUAD.
Status     col B: Vacant iff the name contains "vacant"; else Filled.
Leadership col K == "Leadership" OR col J == "Leadership".
Cost       col AA, untouched.
Lever      Filled/Hire = 1.0x, Hold = 0.0x, Offshore = 0.4x  (0.3 Squad Archetypes K5).
"""

# --- col I -> portfolio tab.  Mirrors Lists!T:U, which is the mapping already in the
# --- workbook; the only judgement calls are the case variants and P&C/Finance & Legal.
PORTFOLIO_MAP = {
    "Ampol Retail": "Ampol Retail",
    "Retail": "Ampol Retail",                              # [confirm] case variant
    "RETAIL": "Ampol Retail",                              # [confirm] case variant
    "Z": "Z Retail",
    "Customer": "Customer",
    "Ampol Customer": "Customer",
    "Enterprise Data": "Enterprise Data",
    "TDD": "TDD Group Functions",
    "Infrastructure": "Infrastructure",
    "B2B & Energy Solutions": "Energy Solutions & B2B",
    "Commercial Fuels": "Commercial Fuels",
    "Finance": "Finance",
    "P&C": "P&C",
    "P&C, Finance & Legal": "P&C",                          # [confirm] 2 roles
    "EGI": "EGI",
    "COE - Cyber, Risk & Operations": "COE Cyber",
    "COE - Partnering & Transformation": "COE BP&T",
    "COE - Strategy, Architecture, Data": "COE SA&D",
}

# --- col K squad typo/case merges.  Each pair is the same squad written two ways.
SQUAD_MERGE = {
    "AmPos": "AmPOS",                                       # 9 + 1  = 10 roles
    "Manuacturing Group Projects": "Manufacturing Group Projects",   # 2 + 7 = 9
    "Integration & Process automation": "Integration & Process Automation",  # 9 + 2 = 11
    "Technology Suport": "Technology Support",              # spelling
    "Customer, AI": "Customer AI",                          # stray comma
    "Data - AU": "Data AU",
    "Data - NZ": "Data NZ",
    "Network & QSR": "Network & QSR",
    "P&C RTA": "P&C RTA",
    "Data Platform": "Data Platforms",
    "Trading & Shipping Data": "Trading & Shipping Data",
}

# Roles whose col K is blank or a placeholder. 'NA' is what COE Cyber and COE BP&T
# carry for every role, so those two portfolios have no squad split in the source at all.
BLANK_SQUAD_LABEL = "Unassigned"
NA_SQUAD_LABELS = {"", "na", "n/a", "none"}

LEVERS = {"Filled": 1.0, "Hire": 1.0, "Hold": 0.0, "Offshore": 0.4}
OFFSHORE_RATE = 0.4

# Tab -> portfolio label, for the 2.x working copies and 1.x design tabs.
TAB_PORTFOLIO = {
    "2.1 Ampol Retail": "Ampol Retail",
    "2.2 Customer": "Customer",
    "2.3 Enterprise Data": "Enterprise Data",
    "2.4 TDD Group Functions": "TDD Group Functions",
    "2.5 P&C": "P&C",
    "2.6 Finance": "Finance",
    "2.7 Infrastructure": "Infrastructure",
    "2.8 Energy Solutions & B2B": "Energy Solutions & B2B",
    "2.9 Commercial Fuels": "Commercial Fuels",
    "2.10 Z Retail": "Z Retail",
    "2.11 TDD Cyber": "COE Cyber",
    "2.12 BP&T": "COE BP&T",
    "2.13 SA&D": "COE SA&D",
    "2.14 EGI": "EGI",
    # TDD Cyber is a portfolio in the model before it is one in the ledger: 1.14 prices it
    # and 2.15 is its working copy, and no role in REVIEW carries it yet.
    "2.15 TDD Cyber": "TDD Cyber",
}

REVIEW = "REVIEW - Complete Role Mapping"
LAST_ROW = 528          # real data ends at 528 (Rob Struthers, Jens Tom); 191 is a stray subtotal
STRAY_SUBTOTAL_ROW = 191


def norm_portfolio(raw):
    return PORTFOLIO_MAP.get(str(raw or "").strip())


def norm_squad(raw):
    s = str(raw or "").strip()
    if s.lower() in NA_SQUAD_LABELS:
        return BLANK_SQUAD_LABEL
    return SQUAD_MERGE.get(s, s)


def is_vacant(name):
    return "vacant" in str(name or "").lower()


def load(path):
    """Read the ledger into plain dicts using the rules above."""
    import openpyxl
    wv = openpyxl.load_workbook(path, data_only=True)
    ws = wv[REVIEW]
    out = []
    for i in range(2, LAST_ROW + 1):
        if i == STRAY_SUBTOTAL_ROW:
            continue
        nm = ws.cell(i, 2).value
        if nm is None or str(nm).strip() == "":
            continue
        cost = ws.cell(i, 27).value
        out.append(dict(
            row=i,
            name=str(nm).strip(),
            title=str(ws.cell(i, 3).value or "").strip(),
            portfolio=norm_portfolio(ws.cell(i, 9).value),
            raw_portfolio=str(ws.cell(i, 9).value or "").strip(),
            department=str(ws.cell(i, 7).value or "").strip(),
            platform=str(ws.cell(i, 10).value or "").strip(),
            squad=norm_squad(ws.cell(i, 11).value),
            raw_squad=str(ws.cell(i, 11).value or "").strip(),
            country=str(ws.cell(i, 13).value or "").strip(),
            cost=cost if isinstance(cost, (int, float)) else 0.0,
            vacant=is_vacant(nm),
        ))
    for r in out:
        r["leadership"] = (r["squad"].lower() == "leadership"
                           or r["platform"].strip().lower() == "leadership")
        # A role with no squad in col K but flagged Leadership in col J IS leadership.
        if r["squad"] == BLANK_SQUAD_LABEL and r["leadership"]:
            r["squad"] = "Leadership"
        # The three COEs carry 'NA' in col K for every role, so they have no squad
        # split in the source at all. Col G (Department) is Lee's own grouping and
        # foots exactly, so it stands in as the squad for those tabs.
        elif r["squad"] == BLANK_SQUAD_LABEL and r["department"]:
            d = r["department"]
            r["squad"] = BLANK_SQUAD_LABEL if d.strip().lower() in NA_SQUAD_LABELS else d
        r["lever"] = "Hold" if r["vacant"] else "Filled"
        r["effective"] = r["cost"] * LEVERS[r["lever"]]
    wv.close()
    return out
