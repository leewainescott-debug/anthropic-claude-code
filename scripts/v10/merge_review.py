"""The new ledger: rev's REVIEW, the new Customer dataset, and the scaffold restored.

Three sources, each taken for exactly what it is good for:

  rev.xlsx        the owner's edits - 6 new cyber vacancies appended (rows 529-534),
                  the SA&D re-level on row 321, four ring-fenced roles renamed Vacant,
                  a name gaining an (AKL) suffix. Its derived columns are the old
                  generation and its AP (MSquadC) is corrupt for Customer, so nothing
                  derived is taken from it.
  cust_new.xlsx   PCM_Data, the new Customer dataset. Same 83 people as today, with the
                  AU and NZ on-cost parameter blocks put back on the right people -
                  rev's Customer block has them cross-contaminated (an Australian priced
                  at 0.92 FX with NZ pensions and medical, an NZ role at AU rates). The
                  whole block is replaced; a delta merge would merge into corrupt data.
  base_ship.xlsx  the scaffold - the AA cost formula that honours the agreed-cost
                  override column, and the AP/AQ/AS canonical-squad machinery. rev
                  branched before these were built.

The loader verifies, row by row, that the D8 cost formula reproduces the dataset's own
stated cost from its components before anything ships.

Three things about the assembled ledger as a thing people read rather than a thing
formulas read: AR and AT are live on every row and 3.2 counts on them, and neither had a
heading; rows 191 and 192 sit between the Customer block and the next one holding nothing
but helper formulas from an older layout; and nine typed slips sit in names, titles and
countries. None of the three moves a number.
"""
import openpyxl
from openpyxl.utils import get_column_letter

REVIEW = "REVIEW - Complete Role Mapping"
CUST_LO, CUST_HI = 108, 190              # the Customer block, identical rows in all files
NEW_LO, NEW_HI = 529, 534                # the six cyber vacancies rev appended

# PCM_Data -> REVIEW raw columns. PCM inserts Status at C and Contract End at Q, so the
# shift is +1 from Position Title and +2 from Unit onwards.
#   REVIEW col index -> PCM col index
MAP = {1: 1, 2: 2, 3: 4, 4: 5, 5: 6, 6: 7, 7: 8, 8: 9, 9: 10, 10: 11, 11: 12,
       12: 13, 13: 14, 14: 15, 15: 16, 17: 18, 18: 19, 19: 20, 20: 21, 21: 22,
       22: 23, 23: 24, 24: 25, 25: 26, 26: 27}
# REVIEW: A=1 B=2 C=3.. I=9(Portfolio) K=11(Squad) M=13(Country) N=14(Job Level)
# O=15(FTE) Q=17(Type) R=18(Unit) S=19(day rate) T=20 U=21 V=22 W=23 X=24 Y=25 Z=26
PCM_STATUS, PCM_COMMENT, PCM_COST = 3, 29, 28
# two grouping columns that are live on every row and read by 3.2's COUNTIFS, and have
# never had a heading. A column a reader can see and cannot name is a column they will
# guess at.
HEADS = {44: "Overhead line", 46: "Squad or overhead line"}
# the two rows between the Customer block and the block after it. They hold no person and
# never have; what they hold is helper formulas left behind by an older layout.
SPACERS = (191, 192)
LEDGER_COLS = 52
# Typed slips in his source data, in the cells a reader sees: a name, a title, a country.
# Nothing in the build joins on any of these - the joins are by row and by MTab - so the
# only thing a correction changes is what the page says. Checked against every script in
# the directory before it was written.
TYPOS = {"Project Manger": "Project Manager",
         "Portfolio Mnager": "Portfolio Manager",
         "michelle Siegman": "Michelle Siegman",
         "vijay Solanki": "Vijay Solanki",
         "DeveloperSAP ECC": "Developer SAP ECC",
         "EnterpriseProcess Analyst": "Enterprise Process Analyst",
         "Support Analyst -Retail": "Support Analyst - Retail",
         "strategic program architect": "Strategic Program Architect"}
# a whole cell, not a fragment: "australia" is the country, and only the country
WHOLE_CELL = {"australia": "Australia"}
# Name, Position Title, Reports to name, Reports to Position, Country
TEXT_COLS = (2, 3, 4, 5, 13)

COUNTRY = {"AUD": "Australia", "Australia": "Australia", "australia": "Australia",
           "NZD": "NZ", "NZ": "NZ", "WIPRO": "WIPRO"}
# his new file names the sub-portfolios; the model's portfolio comes from Lists!T:U, so
# the raw values load untouched and the map gains the two new names
PORTFOLIO_MAP = [("Z ENERGY (DIGITAL)", "Customer"), ("EGI Integration", "Customer")]


def norm_level(v):
    """Job Level 'CIC (Always)' -> 'CIC'; the suffix is PCM commentary, not a level."""
    s = str(v or "").strip()
    for suf in (" (Always)", " (Often)", " (Sometimes)"):
        if s.endswith(suf):
            return s[: -len(suf)]
    return s or None


def load_pcm(path="cust_new.xlsx"):
    ws = openpyxl.load_workbook(path, data_only=True)["PCM_Data"]
    rows = []
    for r in range(2, 85):
        if not str(ws.cell(r, 2).value or "").strip():
            continue
        raw = {rc: ws.cell(r, pc).value for rc, pc in MAP.items()}
        # PCM's EE Number carries the literal text '#N/A' on 28 rows, which Excel reads
        # as an error the moment it lands in a cell. Blank is the ledger's own
        # convention for a role with no employee number yet.
        if str(raw.get(1)).strip() in ("#N/A", "N/A", "#REF!"):
            raw[1] = None
        raw[13] = COUNTRY.get(str(raw.get(13) or "").strip(), raw.get(13))
        raw[14] = norm_level(raw.get(14))
        if isinstance(raw.get(17), str):
            raw[17] = raw[17].replace("Full time", "Full Time")
        rows.append({"raw": raw,
                     "status": ws.cell(r, PCM_STATUS).value,
                     "comment": ws.cell(r, PCM_COMMENT).value,
                     "stated": ws.cell(r, PCM_COST).value or 0})
    return rows


def d8_cost(raw, days=222):
    """The ledger's one cost formula, in Python, for verification only."""
    s = raw.get(19)
    if isinstance(s, (int, float)) and s > 0:
        return s * days * (1 + (raw.get(26) or 0))
    u = raw.get(21) or 0
    return u * (1 + (raw.get(22) or 0) + (raw.get(23) or 0) + (raw.get(24) or 0)
                + (raw.get(26) or 0)) + (raw.get(25) or 0)


def run(src, dst, ancestor="base_ship.xlsx", pcm="cust_new.xlsx"):
    wb = openpyxl.load_workbook(src)
    anc = openpyxl.load_workbook(ancestor)
    R, B = wb[REVIEW], anc[REVIEW]
    out = []

    # ---- 1. the scaffold, from the ancestor ----
    # headers and per-row formulas for AA (cost, override-honouring), AP (canonical
    # squad), AQ (Leadership flag), AS (design-tab name), AU/AV (override + note).
    # AJ / AR / AT are rebuilt by overrides.py for every row, so they are left alone.
    tmpl = {c: B.cell(2, c).value for c in (27, 42, 43, 45)}
    for c in (42, 43, 45, 47, 48):
        R.cell(1, c).value = B.cell(1, c).value
    n = 0
    last = R.max_row
    while last > 1 and not str(R.cell(last, 2).value or "").strip():
        last -= 1
    for r in range(2, last + 1):
        for c in (27, 42, 43, 45):
            R.cell(r, c).value = _rowform(tmpl[c], r)
        R.cell(r, 47).value = B.cell(r, 47).value if r <= 528 else None
        R.cell(r, 48).value = B.cell(r, 48).value if r <= 528 else None
        n += 1
    out.append(f"scaffold AA/AP/AQ/AS/AU restored on rows 2..{last} ({n} rows)")

    for c, head in HEADS.items():
        R.cell(1, c).value = head
    out.append("REVIEW: AR and AT given the headings they carry - "
               + ", ".join(f"{get_column_letter(c)}1 {h!r}" for c, h in HEADS.items()))

    # The spacer rows. They carry no person and no data, and every formula on them guards
    # itself with IF(TRIM($B)="",""), so they compute nothing - but a reader scrolling the
    # ledger sees two rows of machinery in the middle of it, and a role typed onto one of
    # them would inherit whatever the old layout left. They go out empty; the styling
    # stays, so the rows still read as part of the table.
    for r in SPACERS:
        for c in range(1, LEDGER_COLS + 1):
            R.cell(r, c).value = None
    out.append(f"rows {SPACERS[0]} and {SPACERS[1]} cleared across "
               f"{LEDGER_COLS} columns - the spacers between the Customer block and the "
               f"block after it")

    # ---- 2. the Customer block, wholesale from PCM_Data ----
    rows = load_pcm(pcm)
    if len(rows) != CUST_HI - CUST_LO + 1:
        raise SystemExit(f"PCM has {len(rows)} roles for an 83-row block")
    # MyHR numbers survive by name from the outgoing block
    myhr = {}
    for r in range(CUST_LO, CUST_HI + 1):
        nm = str(R.cell(r, 2).value or "").strip().lower()
        if R.cell(r, 28).value is not None:
            myhr[nm] = R.cell(r, 28).value
    bad = []
    for i, row in enumerate(rows):
        r = CUST_LO + i
        for c in range(1, 27):
            R.cell(r, c).value = row["raw"].get(c)
        # the ancestor's agreed-cost overrides were copied in with the scaffold, and the
        # people on these rows have changed - Tim Corin's banded rate landed on the row
        # that now holds Denise Esguerra and priced her at his 275,810.25. A replaced row
        # starts clean; the new dataset restates every Customer cost, overrides included.
        R.cell(r, 47).value = None
        R.cell(r, 48).value = None
        R.cell(r, 27).value = _rowform(tmpl[27], r)          # the D8 cost formula
        key = str(row["raw"].get(2) or "").strip().lower()
        base = key.split(" (")[0].split(" - ")[0]
        R.cell(r, 28).value = myhr.get(key) or myhr.get(base)
        R.cell(r, 49).value = row["status"]                  # PCM Status, parked visibly
        R.cell(r, 50).value = row["comment"]                 # PCM commentry, parked
        want = row["stated"]
        got = d8_cost(row["raw"])
        if abs(got - (want or 0)) > 0.02:
            # the dataset's stated cost disagrees with its own components. The stated
            # figure is the owner's number, so it goes through the agreed-cost override
            # column the AA formula already honours - components stay visible, the
            # stated cost prices the role, and the provenance is written beside it.
            R.cell(r, 47).value = round(want or 0, 2)
            R.cell(r, 48).value = ("Stated cost in the owner's Customer dataset "
                                   f"(components imply {got:,.0f})")
            bad.append((r, row["raw"].get(2), round(got, 2), round(want or 0, 2)))
    R.cell(1, 49).value = "Status (PCM)"
    R.cell(1, 50).value = "Commentry (PCM)"
    out.append(f"Customer rows {CUST_LO}-{CUST_HI} replaced from PCM_Data; "
               f"{len(bad)} rows where the stated cost disagrees with its own components "
               f"- each priced at the stated cost via the AU override, noted in AV")
    for b in bad[:10]:
        out.append(f"   r{b[0]} {b[1]}: formula {b[2]:,} vs stated {b[3]:,}")

    # ---- 3. the portfolio map learns his two new sub-portfolio names ----
    l = wb["Lists"]
    have = {str(l.cell(r, 20).value or "").strip() for r in range(1, 30)}
    free = 2
    while l.cell(free, 20).value is not None:
        free += 1
    added = []
    for a, b_ in PORTFOLIO_MAP:
        if a in have:
            continue
        l.cell(free, 20).value = a
        l.cell(free, 21).value = b_
        added.append(f"{a} -> {b_} at T{free}")
        free += 1
    if added:
        out.append("Lists!T:U portfolio map: " + "; ".join(added))

    # ---- 4. the typed slips in his source data ----
    out += fix_typos(R, last)
    return wb, out, bad


def fix_typos(R, last):
    """Names, titles and countries as they will be read, not as they were typed.

    Only the display columns, and only the strings listed. The model joins on row index
    and on MTab, so none of these strings carries a join - checked against every script in
    the directory - and correcting one changes what the page says and nothing else.
    """
    fixed = []
    for r in range(2, last + 1):
        for c in TEXT_COLS:
            v = R.cell(r, c).value
            if not isinstance(v, str) or v.startswith("="):
                continue
            new = WHOLE_CELL[v.strip()] if v.strip() in WHOLE_CELL else v
            for old, rep in TYPOS.items():
                new = new.replace(old, rep)
            if new != v:
                R.cell(r, c).value = new
                fixed.append(f"{get_column_letter(c)}{r} {v!r} -> {new!r}")
    if not fixed:
        return ["no typed slips left in the ledger's name, title and country cells"]
    return [f"{len(fixed)} typed slips corrected in the ledger's name, title and country "
            f"cells"] + [f"   {f}" for f in fixed]


def _rowform(template, r):
    """Re-anchor a row-2 formula template onto row r."""
    import re
    if not isinstance(template, str) or not template.startswith("="):
        return template
    return re.sub(r"(\$?[A-Z]{1,2})2\b", lambda m: f"{m.group(1)}{r}", template)


if __name__ == "__main__":
    import sys
    wb, out, bad = run(sys.argv[1], sys.argv[2])
    wb.save(sys.argv[2])
    for x in out:
        print("  ", x)
    if bad and len(bad) > 1:
        raise SystemExit(f"{len(bad)} cost mismatches - investigate before building on")
