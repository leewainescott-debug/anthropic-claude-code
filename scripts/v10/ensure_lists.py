"""Recreate the Lists scaffolding the review workbook's branch never had.

rev.xlsx forked before the model's Lists machinery was built, so everything from column W
rightward is absent: the squad-name fold table, the lever cost factors, the overhead
allowance table with the GM layer and the days-per-year input. Five chain scripts and the
2.x after-decision formulas read those cells, so they are copied back in from the
ancestor - values and formulas alike - before anything else builds.

What is NOT copied: the three fold rows the owner's own data overruled last round
(Customer, AI / Z Energy Martech / AU CRM & Martech are real squads, not typos), and the
retired Z:AA name map, which stays retired.

His own Lists edits are already in the file this runs on and are left alone: Hybrid in the
On/Off list, 0.15 in the Support % list, and the portfolio list renaming cyber.
"""
import openpyxl

UNFOLD = {"Customer, AI", "Z Energy Martech", "AU CRM & Martech"}
# (col range, rows) blocks to copy from the ancestor's Lists
BLOCKS = [
    (23, 24, 1, 13),                    # W:X   squad-name fold table (typos only)
    (29, 30, 1, 6),                     # AC:AD lever cost factors
    (31, 36, 1, 16),                    # AE:AJ overhead allowance + GM layer + days
]


def run(src, dst, ancestor="base_ship.xlsx"):
    wb = openpyxl.load_workbook(src)
    anc = openpyxl.load_workbook(ancestor)
    l, la = wb["Lists"], anc["Lists"]
    out = []
    for c0, c1, r0, r1 in BLOCKS:
        n = 0
        for r in range(r0, r1 + 1):
            if c0 == 23 and str(la.cell(r, 23).value or "").strip() in UNFOLD:
                continue
            for c in range(c0, c1 + 1):
                src_cell = la.cell(r, c)
                if src_cell.value is None:
                    continue
                x = l.cell(r, c)
                v = src_cell.value
                # the ancestor's rates read 0.2's old I:L layout; his 0.2 moved the
                # tables two columns right to K:N
                if isinstance(v, str) and "'0.2 Data Config'!$" in v:
                    v = v.replace("'0.2 Data Config'!$L$", "'0.2 Data Config'!$N$") \
                         .replace("'0.2 Data Config'!$K$", "'0.2 Data Config'!$M$")
                x.value = v
                x.font = openpyxl.styles.Font(name="Calibri", size=10)
                n += 1
        cols = f"{openpyxl.utils.get_column_letter(c0)}:{openpyxl.utils.get_column_letter(c1)}"
        out.append(f"Lists {cols}: {n} cells restored from the ancestor")
    # One offshore rate, not two. AD5 drives the vacancy lever on all fifteen working
    # tabs; 0.3!K5 is the owner's own Offshore rate and drives the archetype prices. As
    # two typed 0.4s, retyping his K5 moved the archetype side only, and every variance
    # column became a comparison of two different offshore assumptions with no control
    # that fires. AD5 now reads his cell, so his one input drives both sides.
    if l["AD5"].value in (0.4, "='0.3 Squad Archetypes'!$K$5"):
        l["AD5"] = "='0.3 Squad Archetypes'!$K$5"
        out.append("Lists!AD5 = 0.3!K5 - the lever's offshore rate is his archetype "
                   "Offshore rate, one input driving both sides")
    else:
        out.append(f"Lists!AD5 holds {l['AD5'].value!r} - left alone")
    # and the note beside the GM layer still said 525 and "ledger"
    if isinstance(l["AF13"].value, str) and "525" in l["AF13"].value:
        l["AF13"] = ("The 8 GMs are the only overhead line with no role in REVIEW, so "
                     "their cost is entered here and sits above the role mapping.")
        out.append("Lists!AF13: 525/ledger wording brought up to date")
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
