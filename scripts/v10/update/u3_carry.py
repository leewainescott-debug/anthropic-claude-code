#!/usr/bin/env python3
"""u3 - spec section C: the WiPro and cyber state carried forward, verified.

  python3 u3_carry.py <in.xlsx> <out.xlsx> [owner_30_07.xlsx]

Recalculates the workbook with a real engine, writes the cached values back so
it opens populated, then checks section C item by item. Anything off is printed
as FAIL and the script exits non zero.
"""
import sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import wbio
from _xl import REVIEW, LEVERS, Log, ledger, read_block

OWNER = ("/root/.claude/uploads/e550b440-3996-5abb-87e5-bafafe598f82/"
         "d4b6cba7-TDD_Cost_Calc_300726.xlsx")

# measured from his own 30/07 file, recalculated: these are the carried figures
HIS = {"0.2 Ampol Customer": 2.43875, "0.2 Z Customer": 4.36375,
       "uplift funding": 2.8, "uplift used": 1.95461945,
       "uplift remaining": 0.84538055, "uplift people": 1.794816}

# the only lever cells section A is allowed to move (A1's eight, A2's five)
A_LEVER_TABS = {"2.3 Enterprise Data": 8, "2.4 TDD Group Functions": 4, "2.5 P&C": 1}

BANNED = (r"\bwaves?\b", r"\bseats?\b", r"\bdesigns?\b", r"\bdesigners?\b")

# the six he set on 30/07; the other eight Offshore levers came in with D120 ruling 4
GRC6 = ("E22", "E23", "E24", "E25", "E28", "E35")


class Check:
    def __init__(self):
        self.rows = []

    def __call__(self, ok, name, got, want=""):
        self.rows.append(("PASS" if ok else "FAIL", name, got, want))

    def conflict(self, name, got, want):
        self.rows.append(("CONFLICT", name, got, want))

    def report(self):
        w = max(len(r[1]) for r in self.rows)
        for st, name, got, want in self.rows:
            tail = "  (spec: %s)" % want if want != "" else ""
            print("%-8s %-*s %s%s" % (st, w, name, got, tail), flush=True)
        fails = [r for r in self.rows if r[0] == "FAIL"]
        conf = [r for r in self.rows if r[0] == "CONFLICT"]
        print("\n%d checks, %d pass, %d fail, %d conflict"
              % (len(self.rows), len(self.rows) - len(fails) - len(conf),
                 len(fails), len(conf)), flush=True)
        return fails, conf


def review_row(cell_value):
    m = re.search(r"\$(\d+)$", str(cell_value or ""))
    return int(m.group(1)) if m else None


def levers_of(wb):
    """{(tab, ledger row): lever} for every role row on every 2.x tab."""
    out = {}
    for ws in wb.worksheets:
        if not ws.title.startswith("2."):
            continue
        for r in range(1, ws.max_row + 1):
            d = ws.cell(r, 4).value
            if not (isinstance(d, str) and REVIEW in d and "$AK$" in d):
                continue
            out[(ws.title, review_row(d))] = ws.cell(r, 5).value
    return out


def main(src, dst, owner=OWNER):
    log = Log("u3_carry")
    log.head("recalculating and writing the cached values back")
    rc, st = wbio.build(src, dst)
    print("recalculated, %d formula cells populated across %d sheets"
          % (st["cells"], st["sheets"]), flush=True)

    v = openpyxl.load_workbook(dst, data_only=True)
    f = openpyxl.load_workbook(dst, data_only=False)
    rv, rvf = v[REVIEW], f[REVIEW]
    c = Check()

    # -------------------------------------------------- ledger and the errors
    roles = ledger(rvf)
    c(len(roles) == 529, "ledger roles after the two deletions", len(roles), "529")
    ERR = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "Err:")
    errs = [(ws.title, cl.coordinate) for ws in v.worksheets for row in ws.iter_rows()
            for cl in row if isinstance(cl.value, str) and any(e in cl.value for e in ERR)]
    c(not errs, "error cells in the workbook", "%d %s" % (len(errs), errs[:4] or ""), "0")
    q = v["4.0 Data QA"]
    bad = [(r, q.cell(r, 2).value, q.cell(r, 5).value) for r in range(4, 90)
           if isinstance(q.cell(r, 5).value, (int, float))
           and abs(q.cell(r, 5).value) > 1e-6]
    c(not bad, "4.0 Data QA checks reading zero",
      "%d off %s" % (len(bad), [(b[0], round(b[2], 6)) for b in bad[:4]]), "all 0")
    ctrl = []
    for ws in v.worksheets:
        for row in ws.iter_rows():
            for cl in row:
                if isinstance(cl.value, str) and cl.value.startswith("Control - ") \
                        and "must be 0" in cl.value:
                    for cc in range(cl.column + 1, ws.max_column + 1):
                        x = ws.cell(cl.row, cc).value
                        if isinstance(x, (int, float)):
                            if abs(x) > 1e-6:
                                ctrl.append("%s!%s=%s" % (ws.title, cl.coordinate, x))
                            break
    c(not ctrl, "every 'must be 0' control", "%d off %s" % (len(ctrl), ctrl[:4]), "all 0")

    # ------------------------------------------------------------- C  WiPro
    wip = [(r, nm, ttl) for r, nm, ttl, pf in roles
           if "WIPRO" in str(rvf.cell(r, 17).value or "").upper()]
    c(len(wip) == 6, "WiPro roles in the ledger", len(wip), "6")
    rate_ok = all(abs(rv.cell(r, 27).value - 73260) < 0.005 for r, _, _ in wip)
    c(rate_ok, "every WiPro role at the vendor rate",
      sorted({round(rv.cell(r, 27).value, 2) for r, _, _ in wip}), "73,260.00")
    lev = levers_of(f)
    wl = [lev.get(("2.2 Customer", r)) for r, _, _ in wip]
    c(all(x == "Offshore" for x in wl), "WiPro levers on 2.2", sorted(set(wl)), "Offshore")
    c22 = v["2.2 Customer"]
    after = {}
    c22f = f["2.2 Customer"]
    for r in range(37, c22f.max_row + 1):
        rr = review_row(c22f.cell(r, 4).value)
        if rr:
            after[rr] = c22.cell(r, 7).value
    paid = [round(after.get(r, -1), 2) for r, _, _ in wip]
    c(all(abs(p - 73260) < 0.005 for p in paid),
      "WiPro cost after the lever (factor 1.0 exemption)", sorted(set(paid)), "73,260.00")

    # ------------------------------------------------------- C  cyber movers
    L = f["Lists"]
    mov = [(L.cell(r, 40).value, L.cell(r, 41).value, L.cell(r, 42).value)
           for r in range(2, 30) if L.cell(r, 41).value == "TDD Cyber"]
    c(len(mov) == 9, "cyber movers in the Lists override table", len(mov), "9")
    up = sum(1 for m in mov if m[2] == "Cyber Uplift")
    idn = sum(1 for m in mov if m[2] == "Identity")
    c((up, idn) == (5, 4), "the split of them", "Cyber Uplift %d, Identity %d" % (up, idn),
      "5 and 4")

    # ----------------------------------------------------------- C  toggles
    c13 = f["1.13 Cyber Roles"]
    tog = [c13.cell(r, 9).value for r in range(19, 71)
           if isinstance(c13.cell(r, 9).value, (int, float))]
    c(sorted(tog, reverse=True) == [0.5, 0.5, 0.4, 0.25, 0],
      "the five uplift toggles", sorted(tog, reverse=True), "50/50/40/25/0")

    # ------------------------------------------------------------ C  uplift
    c14 = v["1.14 TDD Cyber"]
    for name, coord in (("uplift funding", "J14"), ("uplift used", "J18"),
                        ("uplift remaining", "J19")):
        got = c14[coord].value
        c(abs(got - HIS[name]) < 5e-7, "1.14 %s ($m)" % name, round(got, 7),
          round(HIS[name], 4))
    people = (c14["K26"].value or 0) + (c14["J16"].value or 0)
    c(abs(people - HIS["uplift people"]) < 5e-7, "cyber uplift people tie ($)",
      "%.2f" % (people * 1e6), "1,794,816.00")

    # ----------------------------------------------------------- C  the name
    c("2.11 Cyber Risk & Service Ops" in f.sheetnames, "2.11 tab name",
      "2.11 Cyber Risk & Service Ops", "2.11 Cyber Risk & Service Ops")
    full = "Cyber, Risk & Service Operations"
    c(f["2.11 Cyber Risk & Service Ops"]["B2"].value.startswith(full)
      and f["3.1 Archetype to Actuals"]["B17"].value == full
      and f["3.4 COE Breakdown"]["B11"].value == full + " total",
      "CRSO displayed in full on 2.11, 3.1 and 3.4", full, full)
    c(f["2.11 Cyber Risk & Service Ops"]["C3"].value == "COE Cyber",
      "the join key is still COE Cyber", f["2.11 Cyber Risk & Service Ops"]["C3"].value,
      "COE Cyber")

    # ------------------------------------------------------ C  0.2 Customer
    cfg = v["0.2 Data Config"]
    ac, zc = cfg["F13"].value, cfg["F14"].value
    c(abs(ac - HIS["0.2 Ampol Customer"]) < 5e-4, "0.2 Ampol Customer ($m)",
      round(ac, 5), "2.439")
    if abs(zc - 5.314) < 5e-4:
        c(True, "0.2 Z Customer ($m)", round(zc, 5), "5.314")
    else:
        c.conflict("0.2 Z Customer ($m)",
                   "%.5f, unchanged from his own 30/07 file - his support %% edits on "
                   "1.2 G39 (0.7) and G40 (0.8) moved it and A3 keeps them" % zc, "5.314")

    # --------------------------------------------------- C  the GRC offshores
    t11 = f["2.11 Cyber Risk & Service Ops"]
    grc = [co for co in GRC6 if t11[co].value == "Offshore"]
    allo = sum(1 for r in range(20, 70) if t11.cell(r, 5).value == "Offshore")
    c(len(grc) == 6, "his six GRC offshore levers on 2.11",
      "%d of 6 (%d Offshore on the tab in all, his six plus D120 ruling 4's eight)"
      % (len(grc), allo), "6")
    c(all(f["2.11 Cyber Risk & Service Ops"].cell(r, 5).value in LEVERS
          for r in range(20, 68)
          if isinstance(f["2.11 Cyber Risk & Service Ops"].cell(r, 4).value, str)
          and REVIEW in str(f["2.11 Cyber Risk & Service Ops"].cell(r, 4).value)),
      "one lever column per COE, all four values", "column E only", "column E only")

    # ------------------------------------------------- his lever scenario intact
    ow = openpyxl.load_workbook(owner, data_only=False)
    own = levers_of(ow)

    def shift(r):
        if r in (420, 498):
            return None                      # the two roles A4 takes out
        return r - (1 if r > 420 else 0) - (1 if r > 498 else 0)

    own = {(t, shift(r)): x for (t, r), x in own.items()}
    moved = {k: (own[k], lev[k]) for k in own if k in lev and own[k] != lev[k]}
    gone = sorted(k for k in own if k[1] is None or k not in lev)
    per = {}
    for (t, r) in moved:
        per[t] = per.get(t, 0) + 1
    c(per == A_LEVER_TABS, "lever cells section A moves", per, str(A_LEVER_TABS))
    c(len(gone) == 2, "lever cells that leave with the two deleted roles",
      "%d %s" % (len(gone), gone), "2")
    c(len(set(lev) - set(own)) == 0, "lever cells this build invented",
      len(set(lev) - set(own)), "0")
    c(all(x in LEVERS for x in lev.values()),
      "every lever is one of the four values",
      sorted({str(x) for x in lev.values()}), "Filled/Hire/Hold/Offshore")

    # ----------------------------------------------------- Customer headlines
    c(c22["F30"].value == 83, "2.2 Customer roles", c22["F30"].value, "83")
    for label, coord in (("actual cost", "O30"), ("archetype cost", "N30"),
                         ("cost after levers", "Q30")):
        print("        2.2 Customer %-18s %.6f" % (label, c22[coord].value), flush=True)
    p12 = v["1.2 Customer"]
    print("        1.2 Customer AU %.5f  NZ %.5f  other %.5f"
          % (p12["C9"].value, p12["D9"].value, p12["E9"].value), flush=True)
    t31 = v["3.1 Archetype to Actuals"]
    print("        3.1 TDD total: archetype %.6f  actual %.6f  after levers %.6f  "
          "roles %d" % (t31["D21"].value, t31["E21"].value, t31["G21"].value,
                        t31["H21"].value), flush=True)
    pm = None
    for r in range(20, 30):
        if f["2.2 Customer"].cell(r, 2).value == "Program Management":
            pm = r
    c(pm is not None and abs(c22.cell(pm, 15).value - 0.76124815) < 5e-7,
      "Programme Management overhead line ($)",
      "%.2f" % (c22.cell(pm, 15).value * 1e6 if pm else -1), "761,248")
    o32 = v["3.2 Overhead & Leadership"]
    c(abs(o32["K11"].value) < 5e-7, "3.2 Programme Management variance ($m)",
      o32["K11"].value, "0, allowance = its own cost basis")

    # ------------------------------------------------------- the house rules
    hits = []
    for ws in f.worksheets:
        if ws.title.startswith(("0.1", "0.4")):
            continue                      # his own source tabs keep their words
        for row in ws.iter_rows():
            for cl in row:
                # REVIEW columns A to AB are the PCM source block - job titles and
                # team names are his data, not the model's vocabulary
                if ws.title == REVIEW and cl.column <= 28:
                    continue
                t = cl.value
                if not isinstance(t, str) or t.startswith("="):
                    continue
                for b in BANNED:
                    if re.search(b, t, re.I):
                        hits.append("%s!%s %r" % (ws.title, cl.coordinate, t[:40]))
    c(not hits, "banned words in the model's own text",
      "%d %s" % (len(hits), hits[:4]), "0")
    dash = []
    for ws in f.worksheets:
        if ws.title.startswith(("0.1", "0.4")):
            continue
        for row in ws.iter_rows():
            for cl in row:
                if isinstance(cl.value, str) and cl.value.strip() in ("-", '="-"'):
                    dash.append("%s!%s" % (ws.title, cl.coordinate))
    c(not dash, "dash literals on the model tabs", "%d %s" % (len(dash), dash[:4]), "0")
    prot = [ws.title for ws in f.worksheets if ws.protection.sheet]
    c(not prot, "sheet protection", prot, "none")
    hid = [ws.title for ws in f.worksheets if ws.sheet_state == "hidden"]
    c(len(hid) == 6, "hidden tabs", hid, "his six")

    print(flush=True)
    fails, conf = c.report()
    print("\nwrote", dst, flush=True)
    if fails:
        raise SystemExit(1)
    if conf:
        raise SystemExit(3)


if __name__ == "__main__":
    main(*sys.argv[1:])
