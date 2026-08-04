#!/usr/bin/env python3
"""u7 - the gate suite for base_u6: u3's checks re-derived, plus D, F and H.

  python3 u7_carry.py <in.xlsx> <out.xlsx> [owner_30_07.xlsx] [pre_u6.xlsx]

This is u3_carry.py carried onto the new layout. Every adaptation is marked
"ADAPTED" below so the change log can quote them:

 1. ledger roles 529 -> 528 - integrator ruling 1 takes his third REMOVE out
 2. the owner-row map drops 476 as well as 420 and 498, and the lever cells
    that leave with the deleted roles go from 2 to 3
 3. the 2.x column letters move: P and Q are the new funded outside pair, so
    "cost after levers" reads S and "variance to archetype" reads R
 4. 3.1's column letters move the same way: after levers is I, roles is J,
    and F / G are the new funded outside pair
 5. the five uplift toggles are read off 2.11's own Uplift % column, not the
    1.13 tab that no longer exists
 6. his six GRC offshore levers are found by ledger row off his own file
    rather than by the 2.11 cell coordinates, which the funding block moved
 7. the 2.11 lever column span is derived from the helper block, not typed
 8. 0.2 Z Customer and the overhead above allowance are re-derived from the
    file - his 30/07 support percentages and reclassifications moved both,
    and neither is pinned to a pre-30/07 constant
 9. the 1.11 / 1.12 / 1.13 checks become checks on the consolidated blocks
10. new sections: D funded outside, F consolidation, H hardening, and an
    independent tie-check off REVIEW alone

Anything off is printed as FAIL and the script exits non zero.
"""
import sys, os, re, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import wbio
from _xl import REVIEW, LEVERS, Log, ledger

HERE = os.path.dirname(os.path.abspath(__file__))
OWNER = ("/root/.claude/uploads/e550b440-3996-5abb-87e5-bafafe598f82/"
         "d4b6cba7-TDD_Cost_Calc_300726.xlsx")
PRE = os.path.join(HERE, "base_u5.xlsx")      # the state H3 rewrote

CYBER = "2.11 Cyber Risk & Service Ops"
DEAD = ("1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles")
CREAM = "FFFFF2CC"
SRC = ("0.1 Budget Table (Fin)", "0.4 Presentation Pack")

HIS = {"0.2 Ampol Customer": 2.43875, "uplift funding": 2.8,
       "uplift used": 1.95461945, "uplift remaining": 0.84538055,
       "uplift people": 1.794816}

# ADAPTED 1/2 - the three roles the owner typed REMOVE over, in his row numbers
REMOVED = (420, 476, 498)
ROLES = 528
A_LEVER_TABS = {"2.3 Enterprise Data": 8, "2.4 TDD Group Functions": 4,
                "2.5 P&C": 1}
GRC6 = ("22", "23", "24", "25", "28", "35")   # his 2.11 rows, in his file

BANNED = (r"\bwaves?\b", r"\bseats?\b", r"\bdesigns?\b", r"\bdesigners?\b")

EGI_SQUADS = ("EGI Retail", "EGI TDD", "EGI Customer", "EGI", "EGI P&C",
              "EGI Finance")
PROGRAMMES = {"CTRM": 3.8, "AmPOS": 1.404, "Cyber Uplift": 1.2998}

TOTALS = {}
DASH = re.compile(r'(\\-|"-")')


class Check:
    def __init__(self):
        self.rows = []

    def __call__(self, ok, name, got, want=""):
        self.rows.append(("PASS" if ok else "FAIL", name, got, want))

    def note(self, name, got, want=""):
        self.rows.append(("NOTE", name, got, want))

    def report(self):
        w = max(len(r[1]) for r in self.rows)
        for st, name, got, want in self.rows:
            tail = "  (spec: %s)" % want if want != "" else ""
            print("%-8s %-*s %s%s" % (st, w, name, got, tail), flush=True)
        fails = [r for r in self.rows if r[0] == "FAIL"]
        notes = [r for r in self.rows if r[0] == "NOTE"]
        print("\n%d checks, %d pass, %d fail, %d note"
              % (len(self.rows), len(self.rows) - len(fails) - len(notes),
                 len(fails), len(notes)), flush=True)
        return fails


def review_row(cell_value):
    m = re.search(r"\$(\d+)$", str(cell_value or ""))
    return int(m.group(1)) if m else None


def levers_of(wb):
    """{(tab, ledger row): lever} for every role row on every 2.x tab."""
    out = {}
    for ws in wb.worksheets:
        if not ws.title.startswith("2."):
            continue
        for r in range(1, ws.max_row + 1):
            d = ws.cell(r, 4).value
            if not (isinstance(d, str) and REVIEW in d and "$AK$" in d):
                continue
            out[(ws.title, review_row(d))] = ws.cell(r, 5).value
    return out


def total_row(ws):
    for r in range(5, 60):
        if ws.cell(r, 2).value == "Total portfolio":
            return r
    raise SystemExit("STOP: no 'Total portfolio' row on %s" % ws.title)


def role_rows(ws):
    """ADAPTED 7 - derived, never typed."""
    return [r for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 4).value, str)
            and REVIEW in str(ws.cell(r, 4).value)
            and "$AK$" in str(ws.cell(r, 4).value)]


def sections(fmt):
    out, cur, q, i = [], "", False, 0
    while i < len(fmt):
        ch = fmt[i]
        if ch == '"':
            q = not q
            cur += ch
        elif ch == "\\" and i + 1 < len(fmt):
            cur += fmt[i:i + 2]
            i += 2
            continue
        elif ch == ";" and not q:
            out.append(cur)
            cur = ""
        else:
            cur += ch
        i += 1
    out.append(cur)
    return out


def main(src, dst, owner=OWNER, pre=None):
    log = Log("u7_carry")
    log.head("recalculating and writing the cached values back")
    rc, st = wbio.build(src, dst)
    print("recalculated, %d formula cells populated across %d sheets"
          % (st["cells"], st["sheets"]), flush=True)

    v = openpyxl.load_workbook(dst, data_only=True)
    f = openpyxl.load_workbook(dst, data_only=False)
    rv, rvf = v[REVIEW], f[REVIEW]
    c = Check()
    tabs = [ws.title for ws in f.worksheets if ws.title.startswith("2.")]
    for t in tabs:
        TOTALS[t] = total_row(f[t])

    # -------------------------------------------------- ledger and the errors
    roles = ledger(rvf)
    c(len(roles) == ROLES, "ledger roles after the three deletions",
      len(roles), str(ROLES))
    ERR = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!",
           "Err:")
    errs = [(ws.title, cl.coordinate) for ws in v.worksheets
            for row in ws.iter_rows() for cl in row
            if isinstance(cl.value, str) and any(e in cl.value for e in ERR)]
    c(not errs, "error cells in the workbook",
      "%d %s" % (len(errs), errs[:4] or ""), "0")
    q = v["4.0 Data QA"]
    bad = [(r, q.cell(r, 2).value, q.cell(r, 5).value) for r in range(4, 95)
           if isinstance(q.cell(r, 5).value, (int, float))
           and abs(q.cell(r, 5).value) > 1e-6]
    c(not bad, "4.0 Data QA checks reading zero",
      "%d off %s" % (len(bad), [(b[0], round(b[2], 6)) for b in bad[:4]]),
      "all 0")
    ctrl, notwhite = [], []
    for ws in v.worksheets:
        for row in ws.iter_rows():
            for cl in row:
                if not (isinstance(cl.value, str)
                        and cl.value.startswith("Control - ")
                        and "must be 0" in cl.value):
                    continue
                for cc in range(cl.column + 1, ws.max_column + 1):
                    x = ws.cell(cl.row, cc).value
                    if isinstance(x, (int, float)):
                        if abs(x) > 1e-6:
                            ctrl.append("%s!%s=%s"
                                        % (ws.title, cl.coordinate, x))
                        break
                if f[ws.title].sheet_state == "visible":
                    col = f[ws.title][cl.coordinate].font.color
                    if col is None or col.rgb != "FFFFFFFF":
                        notwhite.append("%s!%s" % (ws.title, cl.coordinate))
    c(not ctrl, "every 'must be 0' control", "%d off %s" % (len(ctrl), ctrl[:4]),
      "all 0")
    c(not notwhite, "controls on visible tabs in white font",
      "%d visible %s" % (len(notwhite), notwhite[:4]), "all invisible")

    # ------------------------------------------------------------- C  WiPro
    wip = [(r, nm, ttl) for r, nm, ttl, pf in roles
           if "WIPRO" in str(rvf.cell(r, 17).value or "").upper()]
    c(len(wip) == 6, "WiPro roles in the ledger", len(wip), "6")
    rate_ok = all(abs(rv.cell(r, 27).value - 73260) < 0.005 for r, _, _ in wip)
    c(rate_ok, "every WiPro role at the vendor rate",
      sorted({round(rv.cell(r, 27).value, 2) for r, _, _ in wip}), "73,260.00")
    lev = levers_of(f)
    wl = [lev.get(("2.2 Customer", r)) for r, _, _ in wip]
    c(all(x == "Offshore" for x in wl), "WiPro levers on 2.2",
      sorted(set(wl)), "Offshore")
    c22, c22f = v["2.2 Customer"], f["2.2 Customer"]
    after = {}
    for r in role_rows(c22f):
        after[review_row(c22f.cell(r, 4).value)] = c22.cell(r, 7).value
    paid = [round(after.get(r, -1), 2) for r, _, _ in wip]
    c(all(abs(p - 73260) < 0.005 for p in paid),
      "WiPro cost after the lever (factor 1.0 exemption)", sorted(set(paid)),
      "73,260.00")

    # ------------------------------------------------------- C  cyber movers
    L = f["Lists"]
    mov = [(L.cell(r, 40).value, L.cell(r, 41).value, L.cell(r, 42).value)
           for r in range(2, 30) if L.cell(r, 41).value == "TDD Cyber"]
    c(len(mov) == 9, "cyber movers in the Lists override table", len(mov), "9")
    up = sum(1 for m in mov if m[2] == "Cyber Uplift")
    idn = sum(1 for m in mov if m[2] == "Identity")
    c((up, idn) == (5, 4), "the split of them",
      "Cyber Uplift %d, Identity %d" % (up, idn), "5 and 4")

    # ------------------------------- C  toggles (ADAPTED 5 - now 2.11 column H)
    t11, t11v = f[CYBER], v[CYBER]
    rr11 = role_rows(t11)
    tog = [t11.cell(r, 8).value for r in rr11
           if isinstance(t11.cell(r, 8).value, (int, float))
           and t11.cell(r, 8).value > 0]
    zeros = sum(1 for r in rr11 if t11.cell(r, 8).value == 0)
    c(sorted(tog, reverse=True) == [0.5, 0.5, 0.4, 0.25],
      "the uplift toggles on 2.11's own column",
      "%s plus %d explicit zeros" % (sorted(tog, reverse=True), zeros),
      "50/50/40/25/0")

    # ------------------------------------------------------------ C  uplift
    c14 = v["1.14 TDD Cyber"]
    for name, coord in (("uplift funding", "J14"), ("uplift used", "J18"),
                        ("uplift remaining", "J19")):
        got = c14[coord].value
        c(abs(got - HIS[name]) < 5e-7, "1.14 %s ($m)" % name, round(got, 7),
          round(HIS[name], 4))
    people = (c14["K26"].value or 0) + (c14["J16"].value or 0)
    c(abs(people - HIS["uplift people"]) < 5e-7, "cyber uplift people tie ($)",
      "%.2f" % (people * 1e6), "1,794,816.00")

    # ----------------------------------------------------------- C  the name
    c(CYBER in f.sheetnames, "2.11 tab name", CYBER, CYBER)
    full = "Cyber, Risk & Service Operations"
    c(t11["B2"].value.startswith(full)
      and f["3.1 Archetype to Actuals"]["B17"].value == full
      and f["3.4 COE Breakdown"]["B11"].value == full + " total",
      "CRSO displayed in full on 2.11, 3.1 and 3.4", full, full)
    c(t11["C3"].value == "COE Cyber", "the join key is still COE Cyber",
      t11["C3"].value, "COE Cyber")

    # ------------------------------------------------------ C  0.2 Customer
    cfg, cfgf = v["0.2 Data Config"], f["0.2 Data Config"]
    ac, zc = cfg["F13"].value, cfg["F14"].value
    c(abs(ac - HIS["0.2 Ampol Customer"]) < 5e-4, "0.2 Ampol Customer ($m)",
      round(ac, 5), "2.439")
    # ADAPTED 8 - Z Customer re-derived from 1.2's own NZ rows, never pinned
    p12 = v["1.2 Customer"]
    zder = sum(p12.cell(r, 4).value or 0 for r in (6, 7, 8))
    c(abs(zc - zder) < 5e-9,
      "0.2 Z Customer ($m), re-derived from 1.2's NZ rows",
      "%.6f against %.6f - portfolio overhead, platform overheads and squad "
      "support, the last set by his 30/07 G39/G40 percentages" % (zc, zder),
      "the file, not a constant")

    # ------------------- C  the GRC offshores (ADAPTED 6 - keyed on ledger row)
    ow = openpyxl.load_workbook(owner, data_only=False)
    owg = ow[CYBER]

    def shift(r):
        if r in REMOVED:
            return None
        return r - sum(1 for x in REMOVED if r > x)

    grc = [shift(review_row(owg["D%s" % n].value)) for n in GRC6]
    got = [lev.get((CYBER, g)) for g in grc]
    allo = sum(1 for r in rr11 if t11.cell(r, 5).value == "Offshore")
    c(all(x == "Offshore" for x in got), "his six GRC offshore levers on 2.11",
      "%d of 6 (%d Offshore on the tab in all, his six plus D120 ruling 4's "
      "eight)" % (sum(1 for x in got if x == "Offshore"), allo), "6")
    c(all(t11.cell(r, 5).value in LEVERS for r in rr11),
      "one lever column per COE, all four values", "column E only",
      "column E only")

    # ------------------------------------------- his lever scenario intact
    own_raw = levers_of(ow)
    own = {(t, shift(r)): x for (t, r), x in own_raw.items()}
    moved = {k: (own[k], lev[k]) for k in own if k in lev and own[k] != lev[k]}
    gone = sorted(k for k in own if k[1] is None or k not in lev)
    per = collections.Counter(t for (t, r) in moved)
    c(dict(per) == A_LEVER_TABS, "lever cells section A moves", dict(per),
      str(A_LEVER_TABS))
    c(len(gone) == 3, "lever cells that leave with the deleted roles",
      "%d %s" % (len(gone), gone), "3")
    c(len(set(lev) - set(own)) == 0, "lever cells this build invented",
      len(set(lev) - set(own)), "0")
    c(all(x in LEVERS for x in lev.values()),
      "every lever is one of the four values",
      sorted({str(x) for x in lev.values()}), "Filled/Hire/Hold/Offshore")

    # ----------------------------------------------------- Customer headlines
    c(c22["F30"].value == 83, "2.2 Customer roles", c22["F30"].value, "83")
    pm = None
    for r in range(20, 30):
        if c22f.cell(r, 2).value == "Program Management":
            pm = r
    c(pm is not None and abs(c22.cell(pm, 15).value - 0.76124815) < 5e-7,
      "Programme Management overhead line ($)",
      "%.2f" % (c22.cell(pm, 15).value * 1e6 if pm else -1), "761,248")
    o32 = v["3.2 Overhead & Leadership"]
    c(abs(o32["K11"].value) < 5e-7, "3.2 Programme Management variance ($m)",
      o32["K11"].value, "0, allowance = its own cost basis")
    # ADAPTED 8 - overhead above allowance re-derived, never pinned to 7.0155
    ex = v["Exec Summary"]
    oh = 0.0
    for t in tabs:
        wf, wv = f[t], v[t]
        for r in range(5, TOTALS[t] + 1):
            if wf.cell(r, 2).value == "Overhead roles total":
                oh += wv.cell(r, 18).value or 0
    c(abs(ex["C20"].value - oh) < 5e-6,
      "overhead in the portfolios above the allowance ($m), re-derived",
      "%.6f against the lever modelling tabs' %.6f, his reclassifications "
      "moved it off 7.0155" % (ex["C20"].value, oh),
      "the file, not a constant")
    c(abs(o32["K13"].value - ex["C20"].value) < 5e-6,
      "3.2's portfolio share of the overhead variance against the Exec line",
      "%.6f and %.6f" % (o32["K13"].value, ex["C20"].value), "equal")

    # ======================================================= D funded outside
    Lf = f["Lists"]
    table = {}
    for r in range(2, 40):
        sq = Lf.cell(r, 47).value
        if not sq:
            break
        table[sq] = (Lf.cell(r, 48).value, Lf.cell(r, 49).value)
    c(set(table) == set(EGI_SQUADS) | set(PROGRAMMES),
      "D1 Lists funded outside table", sorted(table), "the 9 squad keys")
    c(all(table[s][1] == "actual" for s in EGI_SQUADS)
      and all(abs(table[s][1] - PROGRAMMES[s]) < 1e-9 for s in PROGRAMMES),
      "D1 the bases", {s: table[s][1] for s in sorted(table)},
      "EGI at actual, CTRM 3.8, AmPOS 1.404, Cyber Uplift 1.2998")

    ties, outside, per_squad = [], 0.0, {}
    for t in tabs:
        wf, wv = f[t], v[t]
        tot = TOTALS[t]
        for r in range(7, tot + 1):
            o, p, qq = (wv.cell(r, 15).value, wv.cell(r, 16).value,
                        wv.cell(r, 17).value)
            if not isinstance(o, (int, float)):
                continue
            if abs((p or 0) + (qq or 0) - o) > 5e-9:
                ties.append("%s!%d" % (t, r))
            b = wf.cell(r, 2).value
            if r < tot and isinstance(b, str) and not b.startswith("=") \
                    and b in table:
                per_squad.setdefault(b, []).append((t, r, p, o))
        outside += wv.cell(tot, 16).value or 0
    c(not ties, "every 2.x row: funded outside plus TDD-funded is the actual",
      "%d off %s" % (len(ties), ties[:4]), "0")
    egi = sum(p for s in EGI_SQUADS for (_, _, p, _) in per_squad.get(s, []))
    c(abs(egi - 11.880489) < 5e-7, "the EGI family funded outside ($m)",
      round(egi, 6), "11.880489")
    prog = {}
    for s, want in PROGRAMMES.items():
        for (t, r, p, o) in per_squad.get(s, []):
            prog[s] = (round(p, 6), abs(p - want) < 5e-9)
    c(all(x[1] for x in prog.values()),
      "CTRM, AmPOS and Cyber Uplift funded outside",
      {s: x[0] for s, x in prog.items()}, "3.8, 1.404, 1.2998")
    c(all(abs(p - o) < 5e-9 for s in EGI_SQUADS
          for (_, _, p, o) in per_squad.get(s, [])),
      "every EGI squad funded outside at its actual cost",
      "%d rows" % sum(len(per_squad.get(s, [])) for s in EGI_SQUADS), "all")
    c(abs(outside - (egi + sum(PROGRAMMES.values()))) < 5e-7,
      "the funded outside total ($m)", round(outside, 6),
      round(11.880489 + sum(PROGRAMMES.values()), 6))
    t14 = v["2.14 EGI"]
    c(abs(t14["P9"].value - t14["O9"].value) < 5e-9
      and abs(t14["Q9"].value) < 5e-9,
      "E: 2.14 EGI funded outside is its whole cost, TDD-funded 0",
      "%.6f and %.6f" % (t14["P9"].value, t14["Q9"].value), "whole cost and 0")
    t31, t31f = v["3.1 Archetype to Actuals"], f["3.1 Archetype to Actuals"]
    c(abs(t31["F21"].value - outside) < 5e-7,
      "3.1's funded outside total against the tabs",
      "%.6f and %.6f" % (t31["F21"].value, outside), "equal")
    c(abs(t31["G22"].value - t31["G21"].value) < 5e-9
      and abs(t31["G21"].value - (t31["E21"].value - t31["F21"].value)) < 5e-7,
      "3.1's TDD-funded total row", "%.6f" % t31["G22"].value,
      "actual less funded outside")
    c(t31f["B22"].value == "TDD-funded total", "3.1's new row label",
      t31f["B22"].value, "TDD-funded total")
    disc = [("1.1 Ampol Retail", "H66", "2.1 Ampol Retail", 18),
            ("1.2 Customer", "H54", "2.2 Customer", 17),
            ("1.4 TDD Group Functions", "H30", "2.4 TDD Group Functions", 14),
            ("1.5 P&C", "H32", "2.5 P&C", 12),
            ("1.6 Finance", "H33", "2.6 Finance", 12)]
    off = [d[0] for d in disc
           if f[d[0]][d[1]].value != "='%s'!$O$%d" % (d[2], d[3])]
    c(not off, "D4 the five 1.x funding disclosures read the actual cell",
      "%d off %s" % (len(off), off), "all on $O")
    c(f["2.1 Ampol Retail"]["N19"].value == "=$O19",
      "D4 2.1's EGI TDD archetype", f["2.1 Ampol Retail"]["N19"].value, "=$O19")

    # ============================================== F  the COE consolidation
    c(not [t for t in DEAD if t in f.sheetnames], "F3 the three tabs are gone",
      [t for t in DEAD if t in f.sheetnames], "none")
    named = ["%s!%s" % (ws.title, cl.coordinate) for ws in f.worksheets
             for row in ws.iter_rows() for cl in row
             if isinstance(cl.value, str) and any(d in cl.value for d in DEAD)]
    c(not named, "nothing anywhere names the three dead tabs",
      "%d %s" % (len(named), named[:4]), "0")
    reads = {"F6": "2.13 COE SA&D", "F7": CYBER, "F8": "2.12 COE BP&T",
             "F9": "2.12 COE BP&T", "F10": "2.13 COE SA&D"}
    c(all(reads[co] in str(cfgf[co].value) for co in reads),
      "F3 0.2 F6 to F10 read the consolidated blocks",
      "; ".join("%s -> %s" % (co, str(cfgf[co].value)) for co in
                ("F6", "F7", "F8", "F9", "F10")), "2.11 / 2.12 / 2.13")
    c(CYBER in str(f["1.14 TDD Cyber"]["J16"].value),
      "F3 1.14's uplift reads 2.11's slice column",
      str(f["1.14 TDD Cyber"]["J16"].value)[:70], "2.11's column I")
    c("(see %s)" % CYBER in cfgf["B7"].value,
      "F3 the COE note on 0.2 B7 names the consolidated tab",
      cfgf["B7"].value, "(see %s)" % CYBER)
    blocks = {}
    for t in (CYBER, "2.12 COE BP&T", "2.13 COE SA&D"):
        wf = f[t]
        blocks[t] = [r for r in range(1, wf.max_row + 1)
                     if wf.cell(r, 2).value == "Funding"]
    c(all(len(rr) == 1 for rr in blocks.values()),
      "F1 one funding block on each COE tab", dict(blocks), "one each")
    stray = []
    for t in (CYBER, "2.12 COE BP&T", "2.13 COE SA&D"):
        wf = f[t]
        for r in role_rows(wf):
            for col in range(6, wf.max_column + 1):
                if wf.cell(r, col).value in ("Onshore", "Offshore", "Hold",
                                             "Paused"):
                    stray.append("%s!r%d c%d" % (t, r, col))
    c(not stray, "F2 a single lever column per COE",
      "%d stray %s" % (len(stray), stray[:4]), "column E only")

    # ================================================================= H
    dvs = []
    for ws in f.worksheets:
        for dv in ws.data_validations.dataValidation:
            if dv.type != "list":
                continue
            dvs.append((ws.title, str(dv.sqref)[:22], dv.showErrorMessage,
                        dv.errorTitle, dv.error))
    loose = [d for d in dvs if not (d[2] and d[3] == "Invalid entry"
                                    and d[4] == "Pick a value from the list")]
    c(not loose, "H2 every list validation is strict",
      "%d loose of %d %s" % (len(loose), len(dvs), loose[:3]), "0 loose")
    uncovered = []
    for t in tabs:
        wf = f[t]
        cover = set()
        for dv in wf.data_validations.dataValidation:
            if dv.formula1 and "Filled" in str(dv.formula1):
                for cr in dv.sqref.ranges:
                    if cr.min_col == 5:
                        cover |= set(range(cr.min_row, cr.max_row + 1))
        miss = [r for r in role_rows(wf) if r not in cover]
        if miss:
            uncovered.append((t, miss[:3]))
    c(not uncovered, "H2 the lever dropdown covers every role row",
      "%d tabs %s" % (len(uncovered), uncovered[:3]), "all covered")
    updv = [dv for dv in f[CYBER].data_validations.dataValidation
            if dv.formula1 and "5%" in str(dv.formula1)]
    upcov = set()
    for dv in updv:
        for cr in dv.sqref.ranges:
            upcov |= set(range(cr.min_row, cr.max_row + 1))
    c(len(updv) == 1 and set(rr11) <= upcov
      and str(updv[0].formula1).count("%") == 21,
      "H2 the Uplift column is a strict 0 to 100 in 5s dropdown",
      "%d validation over %d cells" % (len(updv), len(upcov)),
      "every role row")
    cream_ok = all(t11.cell(r, 8).fill.fgColor.rgb == CREAM for r in rr11)
    c(cream_ok, "the Uplift column is cream (typed input)", cream_ok, "cream")
    exdv = list(f["Exec Summary"].data_validations.dataValidation)
    names = str(exdv[0].formula1).strip('"').split(",") if exdv else []
    t33 = v["3.3 Squad Actuals to Archetype"]
    groups = []
    for r in range(6, 122):
        g = t33.cell(r, 2).value
        if isinstance(g, str) and g and not g.endswith(" total") \
                and g not in groups:
            groups.append(g)
    c(len(exdv) == 1 and sorted(names) == sorted(groups),
      "H2 the Exec drill-down lists every group",
      "%d of %d groups" % (len(names), len(groups)), "15 of 15")
    guard = [r for r in range(37, 46)
             if "Not a group" in str(f["Exec Summary"].cell(r, 3).value)]
    c(len(guard) == 9, "H2 the drill-down's not-a-group guard",
      "%d of 9 cells" % len(guard), "9")

    if pre is None and os.path.exists(PRE):
        print("recalculating the pre-H3 state for the 6dp comparison",
              flush=True)
        pre = wbio.recalc(PRE)
    if pre:
        pv = openpyxl.load_workbook(pre, data_only=True)
        diff, n = [], 0
        for ws in pv.worksheets:
            if ws.title not in v.sheetnames:
                continue
            wv = v[ws.title]
            for r in range(1, ws.max_row + 1):
                a, b = ws.cell(r, 8).value, wv.cell(r, 8).value
                if isinstance(a, (int, float)) and isinstance(b, (int, float)):
                    n += 1
                    if round(a, 6) != round(b, 6):
                        diff.append((ws.title, "H%d" % r, a, b))
        c(not diff, "H3 helper columns: every H value identical to 6dp",
          "%d cells compared, %d differ %s" % (n, len(diff), diff[:3]), "0")
        exd = [(co, pv["Exec Summary"][co].value, ex[co].value)
               for co in ("C26", "C27", "C28", "C29", "C30")
               if pv["Exec Summary"][co].value != ex[co].value]
        c(not exd, "H3 the Exec vacancy counts identical before and after",
          "%d differ %s" % (len(exd), exd), "0")
    else:
        c.note("H3 the 6dp before/after comparison", "no pre-H3 file given")
    longest = max((len(cl.value), ws.title, cl.coordinate)
                  for ws in f.worksheets if ws.sheet_state == "visible"
                  for row in ws.iter_rows() for cl in row
                  if isinstance(cl.value, str) and cl.value.startswith("="))
    c(longest[0] < 700, "H3 the longest formula on a visible tab",
      "%d chars at %s!%s" % longest, "under 700")

    dashfmt = []
    for ws in f.worksheets:
        if ws.title in SRC:
            continue
        for row in ws.iter_rows():
            for cl in row:
                s = sections(cl.number_format or "")
                if len(s) >= 3 and DASH.search(s[2]):
                    dashfmt.append("%s!%s" % (ws.title, cl.coordinate))
    c(not dashfmt, "H4 dash-showing zero formats on the model tabs",
      "%d %s" % (len(dashfmt), dashfmt[:4]), "0")
    creamf = []
    for ws in f.worksheets:
        for row in ws.iter_rows():
            for cl in row:
                fl = cl.fill
                if fl is not None and fl.fgColor is not None \
                        and fl.fgColor.rgb == CREAM \
                        and isinstance(cl.value, str) \
                        and cl.value.startswith("="):
                    creamf.append("%s!%s" % (ws.title, cl.coordinate))
    c(not creamf, "cream on a formula", "%d %s" % (len(creamf), creamf[:4]), "0")
    summ = [t for t in tabs
            if not any(f[t].cell(r, 2).value
                       == "Vacancy levers on this tab, for the executive summary"
                       for r in range(1, f[t].max_row + 1))]
    c(not summ, "H3 a summary row on every lever modelling tab",
      "%d without %s" % (len(summ), summ), "all 15")

    # ==================================== I2  re-derived from REVIEW alone
    aa = {r: (rv.cell(r, 27).value or 0) for r, _, _, _ in roles}
    tot_actual = sum(aa.values()) / 1e6
    c(abs(t31["E21"].value - tot_actual) < 5e-7,
      "I2 total actual re-derived from REVIEW",
      "%.6f and %.6f" % (t31["E21"].value, tot_actual), "equal")
    byp = collections.defaultdict(float)
    for r, _, _, _ in roles:
        byp[str(rv.cell(r, 36).value or "")] += aa[r] / 1e6
    offp = []
    for t in tabs:
        key = f[t]["C3"].value
        gotv = v[t].cell(TOTALS[t], 15).value
        if abs(gotv - byp.get(key, 0)) > 5e-7:
            offp.append((t, round(gotv, 6), round(byp.get(key, 0), 6)))
    c(not offp, "I2 every portfolio actual re-derived from REVIEW",
      "%d off %s" % (len(offp), offp), "0")
    egi_rev = sum(aa[r] for r, _, _, _ in roles
                  if str(rv.cell(r, 46).value or "") in EGI_SQUADS) / 1e6
    c(abs(egi_rev - egi) < 5e-7,
      "I2 the EGI funded outside re-derived from REVIEW's squad column",
      "%.6f and %.6f" % (egi_rev, egi), "equal")
    offr = v["0.3 Squad Archetypes"]["K5"].value
    factor = {"Filled": 1.0, "Hire": 1.0, "Hold": 0.0, "Offshore": offr}
    scen = 0.0
    for (t, r), x in lev.items():
        wipro = "WIPRO" in str(rvf.cell(r, 17).value or "").upper()
        scen += (rv.cell(r, 27).value or 0) * (1.0 if wipro else factor[x])
    upl = sum((t11v.cell(r, 9).value or 0) for r in rr11)
    c(abs((scen - upl) / 1e6 - t31["I21"].value) < 5e-6,
      "I2 the lever scenario cost re-derived from REVIEW",
      "%.6f and %.6f" % ((scen - upl) / 1e6, t31["I21"].value), "equal")

    # ------------------------------------------------------- the house rules
    hits = []
    for ws in f.worksheets:
        if ws.title.startswith(("0.1", "0.4")):
            continue
        for row in ws.iter_rows():
            for cl in row:
                if ws.title == REVIEW and cl.column <= 28:
                    continue
                t = cl.value
                if not isinstance(t, str) or t.startswith("="):
                    continue
                for b in BANNED:
                    if re.search(b, t, re.I):
                        hits.append("%s!%s %r"
                                    % (ws.title, cl.coordinate, t[:40]))
    c(not hits, "banned words in the model's own text",
      "%d %s" % (len(hits), hits[:4]), "0")
    dash = []
    for ws in f.worksheets:
        if ws.title.startswith(("0.1", "0.4")):
            continue
        for row in ws.iter_rows():
            for cl in row:
                if isinstance(cl.value, str) \
                        and cl.value.strip() in ("-", '="-"'):
                    dash.append("%s!%s" % (ws.title, cl.coordinate))
    c(not dash, "dash literals on the model tabs",
      "%d %s" % (len(dash), dash[:4]), "0")
    prot = [ws.title for ws in f.worksheets if ws.protection.sheet]
    c(not prot, "sheet protection", prot, "none")
    hid = [ws.title for ws in f.worksheets if ws.sheet_state == "hidden"]
    c(len(hid) == 6, "hidden tabs", hid, "his six")

    # -------------------------------------------------------- the headlines
    print("\nheadline numbers", flush=True)
    print("        roles                           %d" % t31["J21"].value)
    print("        actual cost ($m)                %.6f" % t31["E21"].value)
    print("        funded outside TDD ($m)         %.6f" % t31["F21"].value)
    print("        TDD-funded cost ($m)            %.6f" % t31["G22"].value)
    print("        archetype cost ($m)             %.6f" % t31["D21"].value)
    print("        cost after levers ($m)          %.6f" % t31["I21"].value)
    print("        roles after levers              %d" % t31["M21"].value)
    print("        0.2 spend ($m)                  %.6f" % cfg["F26"].value)
    print("        0.2 allocated budget ($m)       %.6f" % cfg["E26"].value)
    print("        0.2 Ampol Customer / Z Customer %.5f / %.5f" % (ac, zc))
    print("        1.2 Customer AU %.5f  NZ %.5f  other %.5f"
          % (p12["C9"].value, p12["D9"].value, p12["E9"].value))

    print(flush=True)
    fails = c.report()
    print("\nwrote", dst, flush=True)
    if fails:
        raise SystemExit(1)


if __name__ == "__main__":
    main(*sys.argv[1:])
