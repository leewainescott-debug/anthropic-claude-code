"""Shared helpers for the update pipeline (u1/u2/u3).

The owner's 30/07 file is the base. Everything here is surgical: we rewrite the
cells a spec item names, and when a row has to appear or disappear we move the
rows AND repoint every reference in the workbook, because openpyxl does neither.

Nothing in here writes a dash or any of the banned words.
"""
import re
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

REVIEW = "REVIEW - Complete Role Mapping"
WHITE = "FFFFFFFF"
LEVERS = ("Filled", "Hire", "Hold", "Offshore")

GFMT = ('=$F{r}*IF(ISNUMBER(SEARCH("WIPRO",\'' + REVIEW + '\'!$Q${rr})),1,'
        'IFERROR(INDEX(Lists!$AD$2:$AD$5,MATCH($E{r},Lists!$AC$2:$AC$5,0)),1))')

# ---------------------------------------------------------------- change log


class Log:
    """One printed line per edit, and a running count."""

    def __init__(self, script):
        self.script = script
        self.n = 0
        self.items = []

    def __call__(self, item, where, what):
        self.n += 1
        line = "%-5s %-34s %s" % (item, where, what)
        self.items.append((item, where, what))
        print(line, flush=True)

    def note(self, item, text):
        self.items.append((item, "", text))
        print("%-5s %-34s %s" % (item, "(note)", text), flush=True)

    def head(self, text):
        print("\n== %s" % text, flush=True)

    def tail(self):
        print("\n%s: %d edits\n" % (self.script, self.n), flush=True)


# ------------------------------------------------------- formula ref scanner

_REF = re.compile(
    r"(?:(?P<sheet>'(?:[^']|'')+'|[A-Za-z0-9_.]+)!)?"
    r"(?P<c1>\$?[A-Za-z]{1,3})(?P<r1>\$?\d+)"
    r"(?::(?P<c2>\$?[A-Za-z]{1,3})(?P<r2>\$?\d+))?"
)
_IDENT = re.compile(r"[A-Za-z0-9_.]")


def _scan(f):
    """Yield (start, end, match) for every A1 reference token in formula f.

    Skips string literals, function names, and anything glued to an identifier.
    """
    i, n = 0, len(f)
    while i < n:
        ch = f[i]
        if ch == '"':
            i += 1
            while i < n:
                if f[i] == '"':
                    if i + 1 < n and f[i + 1] == '"':
                        i += 2
                        continue
                    i += 1
                    break
                i += 1
            continue
        if ch == "'":
            # a sheet name in quotes only ever precedes a '!' - let the regex try
            pass
        m = _REF.match(f, i)
        if m:
            s, e = m.start(), m.end()
            before = f[s - 1] if s else ""
            after = f[e] if e < n else ""
            ok = True
            if before and (_IDENT.match(before) or before in "$!'"):
                ok = False
            if after == "(":            # function name, e.g. LOG10(
                ok = False
            if after and (_IDENT.match(after) or after == "$"):
                ok = False
            if ok:
                yield s, e, m
                i = e
                continue
        i += 1


def _sheet_of(m, own):
    s = m.group("sheet")
    if s is None:
        return own
    if s.startswith("'"):
        return s[1:-1].replace("''", "'")
    return s


def _quote(name):
    if re.fullmatch(r"[A-Za-z0-9_.]+", name) and not re.fullmatch(
            r"\$?[A-Za-z]{1,3}\$?\d+", name):
        return name
    return "'" + name.replace("'", "''") + "'"


def rewrite_refs(f, own_sheet, fn):
    """Rebuild formula f, passing every ref through fn.

    fn(sheet, col1, row1, col2, row2) -> (col1, row1, col2, row2) or None for
    #REF!.  Column strings keep their '$'; rows are ints and the '$' is kept
    from the original text.
    """
    out, last = [], 0
    for s, e, m in _scan(f):
        sheet = _sheet_of(m, own_sheet)
        c1, r1 = m.group("c1"), m.group("r1")
        c2, r2 = m.group("c2"), m.group("r2")
        a1 = r1.startswith("$")
        a2 = r2.startswith("$") if r2 else False
        n1 = int(r1.lstrip("$"))
        n2 = int(r2.lstrip("$")) if r2 else None
        res = fn(sheet, c1, n1, c2, n2)
        if res is None:
            out.append(f[last:s])
            out.append("#REF!")
            last = e
            continue
        nc1, nr1, nc2, nr2 = res
        if (nc1, nr1, nc2, nr2) == (c1, n1, c2, n2):
            continue
        txt = ""
        if m.group("sheet") is not None:
            txt += _quote(sheet) + "!"
        txt += nc1 + ("$" if a1 else "") + str(nr1)
        if nc2 is not None:
            txt += ":" + nc2 + ("$" if a2 else "") + str(nr2)
        out.append(f[last:s])
        out.append(txt)
        last = e
    out.append(f[last:])
    return "".join(out)


def map_formulas(wb, fn):
    """Apply fn(sheet_title, coord, formula) -> formula over every formula cell."""
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    nv = fn(ws.title, c.coordinate, v)
                    if nv != v:
                        c.value = nv
                        changed += 1
    return changed


# ------------------------------------------------------------- row shifting


def shift_rows(wb, sheet, at, n, move_cells=True):
    """Insert (n>0) or delete (n<0) n rows at `at` on `sheet`, repointing all refs.

    Returns the number of formulas rewritten.  Refs that land inside a deleted
    block become #REF! (they should have been removed by the caller first).
    """
    dele = -n if n < 0 else 0
    hi = at + dele - 1

    def one(r):
        if n > 0:
            return r + n if r >= at else r
        if r < at:
            return r
        if r <= hi:
            return None
        return r - dele

    def fn(sh, c1, r1, c2, r2):
        if sh != sheet:
            return (c1, r1, c2, r2)
        if c2 is None:
            m = one(r1)
            return None if m is None else (c1, m, None, None)
        a, b = one(r1), one(r2)
        if a is None and b is None:
            return None
        if a is None:
            a = at
        if b is None:
            b = at - 1
        if a > b:
            return None
        return (c1, a, c2, b)

    changed = map_formulas(wb, lambda s, coord, f: rewrite_refs(f, s, fn))

    if move_cells:
        ws = wb[sheet]
        if n > 0:
            ws.insert_rows(at, n)
        else:
            ws.delete_rows(at, dele)
        _shift_dv(ws, at, n)
    return changed


def shift_cols(wb, sheet, at, n):
    """Insert n columns at column index `at` on `sheet`, repointing all refs.

    Same contract as shift_rows but on the other axis. Inserts only (n > 0):
    the funded outside architecture only ever widens a tab, and a column delete
    would need the range clamping shift_rows does for rows.
    """
    if n <= 0:
        raise ValueError("shift_cols inserts only")

    def one(cs):
        d = "$" if cs.startswith("$") else ""
        i = column_index_from_string(cs.lstrip("$"))
        return cs if i < at else d + get_column_letter(i + n)

    def fn(sh, c1, r1, c2, r2):
        if sh != sheet:
            return (c1, r1, c2, r2)
        return (one(c1), r1, None if c2 is None else one(c2), r2)

    changed = map_formulas(wb, lambda s, coord, f: rewrite_refs(f, s, fn))

    ws = wb[sheet]
    ws.insert_cols(at, n)
    _shift_dv_cols(ws, at, n)
    dims = dict(ws.column_dimensions)
    for key in list(ws.column_dimensions):
        del ws.column_dimensions[key]
    for key, dim in dims.items():
        i = column_index_from_string(key)
        ni = i + n if i >= at else i
        dim.min = dim.max = ni
        ws.column_dimensions[get_column_letter(ni)] = dim
    src = ws.column_dimensions.get(get_column_letter(at + n))
    for i in range(at, at + n):
        d = ws.column_dimensions[get_column_letter(i)]
        if src is not None and src.width:
            d.width = src.width
    return changed


def swap_col(f, sheet, src, dst):
    """Rewrite formula f so every ref to column `src` on `sheet` reads `dst`."""
    def one(c):
        if c.lstrip("$") != src:
            return c
        return ("$" if c.startswith("$") else "") + dst

    def fn(sh, c1, r1, c2, r2):
        if sh != sheet:
            return (c1, r1, c2, r2)
        return (one(c1), r1, None if c2 is None else one(c2), r2)

    return rewrite_refs(f, sheet, fn)


def drop_sheet(wb, title):
    """Remove a sheet; every formula that named it becomes #REF!."""
    def rewrite(own, coord, f):
        out, last = [], 0
        for s, e, m in _scan(f):
            if m.group("sheet") is None or _sheet_of(m, own) != title:
                continue
            out.append(f[last:s])
            out.append("#REF!")
            last = e
        out.append(f[last:])
        return "".join(out)

    del wb[title]
    return map_formulas(wb, rewrite)


def _shift_dv_cols(ws, at, n):
    from openpyxl.worksheet.cell_range import MultiCellRange, CellRange
    for dv in ws.data_validations.dataValidation:
        keep = []
        for cr in dv.sqref.ranges:
            c1 = cr.min_col + n if cr.min_col >= at else cr.min_col
            c2 = cr.max_col + n if cr.max_col >= at else cr.max_col
            keep.append(CellRange(min_col=c1, max_col=c2,
                                  min_row=cr.min_row, max_row=cr.max_row))
        dv.sqref = MultiCellRange(keep)


def _shift_dv(ws, at, n):
    """Keep the lever dropdowns pointing at the rows they were on."""
    from openpyxl.worksheet.cell_range import MultiCellRange, CellRange
    dele = -n if n < 0 else 0
    hi = at + dele - 1
    for dv in ws.data_validations.dataValidation:
        keep = []
        for cr in dv.sqref.ranges:
            r1, r2 = cr.min_row, cr.max_row
            if n > 0:
                r1 = r1 + n if r1 >= at else r1
                r2 = r2 + n if r2 >= at else r2
            else:
                if r1 > hi:
                    r1 -= dele
                elif r1 >= at:
                    r1 = at
                if r2 > hi:
                    r2 -= dele
                elif r2 >= at:
                    r2 = at - 1
                if r2 < r1:
                    continue
            keep.append(CellRange(min_col=cr.min_col, max_col=cr.max_col,
                                  min_row=r1, max_row=r2))
        dv.sqref = MultiCellRange(keep)


# ---------------------------------------------------------- sheet repointing


def repoint_sheet(wb, old, new):
    """Every formula that names sheet `old` names `new` instead."""
    def fn(sh, c1, r1, c2, r2):
        return (c1, r1, c2, r2)

    hit = [0]

    def rewrite(own, coord, f):
        out, last = [], 0
        for s, e, m in _scan(f):
            if m.group("sheet") is None:
                continue
            if _sheet_of(m, own) != old:
                continue
            out.append(f[last:s])
            out.append(_quote(new) + "!" + f[m.start("c1"):e])
            last = e
            hit[0] += 1
        out.append(f[last:])
        return "".join(out)

    map_formulas(wb, rewrite)
    return hit[0]


# ---------------------------------------------------- REVIEW range extension


def extend_review(wb, new_last=700):
    """G3: every REVIEW range runs to row `new_last`.

    Only whole-column-style data ranges are touched: a range whose start row is
    2 and whose end row is at or below the ledger's old extent.
    """
    def fn(sh, c1, r1, c2, r2):
        if sh != REVIEW or c2 is None:
            return (c1, r1, c2, r2)
        if r1 == 2 and r2 < new_last:
            return (c1, r1, c2, new_last)
        return (c1, r1, c2, r2)

    return map_formulas(wb, lambda s, coord, f: rewrite_refs(f, s, fn))


# ------------------------------------------------------------------ styling


def white(ws, *coords):
    """Controls are functional and invisible: white font, fill left alone."""
    for co in coords:
        c = ws[co]
        f = c.font
        c.font = Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                      vertAlign=f.vertAlign, underline=f.underline,
                      strike=f.strike, color=WHITE)


def copy_style(src, dst):
    dst._style = src._style.copy() if hasattr(src._style, "copy") else src._style


def row_style(ws, src_row, dst_row, c1, c2):
    for c in range(c1, c2 + 1):
        copy_style(ws.cell(src_row, c), ws.cell(dst_row, c))


# ------------------------------------------------- 2.x FTE helper blocks

def read_block(ws, start, end):
    """[(kind, name, review_row, lever)] for an FTE helper block."""
    out = []
    for r in range(start, end + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if b is None and c is None:
            continue
        if isinstance(c, str) and c.startswith("=COUNTIF("):
            out.append(("group", b, None, None))
            continue
        rr = None
        for cell in (ws.cell(r, 3), ws.cell(r, 2), ws.cell(r, 6)):
            v = cell.value
            if isinstance(v, str) and REVIEW in v:
                import re
                m = re.search(r"\$(\d+)$", v)
                if m:
                    rr = int(m.group(1))
                    break
        out.append(("role", b, rr, ws.cell(r, 5).value))
    return out


def write_block(ws, start, groups, style_rows, clear_to):
    """Lay out group headers and role rows; return {group: (first, last)}."""
    ghdr, grole = style_rows
    ranges, r = {}, start
    for name, rows in groups:
        row_style(ws, ghdr, r, 2, 7)
        ws.cell(r, 2).value = name
        first = r + 1
        last = r + len(rows)
        ws.cell(r, 3).value = ('=COUNTIF($B${a}:$B${b},"?*")&IF(COUNTIF($B${a}:$B${b},'
                               '"?*")=1," role"," roles")').format(a=first, b=last)
        ws.cell(r, 6).value = "=SUM(F{a}:F{b})".format(a=first, b=last)
        ws.cell(r, 7).value = "=SUM(G{a}:G{b})".format(a=first, b=last)
        for c in (4, 5):
            ws.cell(r, c).value = None
        r += 1
        for rr, lever in rows:
            row_style(ws, grole, r, 2, 7)
            q = "'" + REVIEW + "'!$%s$%d"
            ws.cell(r, 2).value = "=" + q % ("B", rr)
            ws.cell(r, 3).value = "=" + q % ("C", rr)
            ws.cell(r, 4).value = "=" + q % ("AK", rr)
            ws.cell(r, 5).value = lever
            ws.cell(r, 6).value = "=" + q % ("AA", rr)
            ws.cell(r, 7).value = GFMT.format(r=r, rr=rr)
            r += 1
        ranges[name] = (first, last)
    for rr in range(r, clear_to + 1):
        for c in range(2, 8):
            ws.cell(rr, c).value = None
    return ranges, r - 1


def set_dv(ws, rows):
    from openpyxl.worksheet.cell_range import MultiCellRange, CellRange
    for dv in ws.data_validations.dataValidation:
        if dv.formula1 and "Filled" in str(dv.formula1):
            dv.sqref = MultiCellRange([CellRange(min_col=5, max_col=5,
                                                 min_row=r, max_row=r)
                                       for r in rows])


def wire_summary(ws, row, first, last):
    """Point one summary line's counts and sums at its helper group."""
    ws.cell(row, 6).value = '=COUNTIF($B${a}:$B${b},"?*")'.format(a=first, b=last)
    ws.cell(row, 8).value = '=COUNTIFS($D${a}:$D${b},"Filled")'.format(a=first, b=last)
    ws.cell(row, 9).value = '=COUNTIFS($D${a}:$D${b},"Vacant")'.format(a=first, b=last)
    ws.cell(row, 10).value = ('=COUNTIFS($D${a}:$D${b},"Vacant",$E${a}:$E${b},'
                              '"Hire")').format(a=first, b=last)
    ws.cell(row, 11).value = '=COUNTIFS($E${a}:$E${b},"Offshore")'.format(a=first, b=last)
    ws.cell(row, 12).value = '=COUNTIFS($E${a}:$E${b},"Hold")'.format(a=first, b=last)
    ws.cell(row, 17).value = "=SUM($G${a}:$G${b})/1000000".format(a=first, b=last)


# ------------------------------------------------------------------- ledger


def ledger(ws, last=None):
    """[(row, name, title, portfolio)] for every non-blank ledger row."""
    last = last or ws.max_row
    out = []
    for r in range(2, last + 1):
        n = ws.cell(r, 2).value
        if n is None or (isinstance(n, str) and not n.strip()):
            continue
        out.append((r, str(n).strip(), str(ws.cell(r, 3).value or "").strip(),
                    str(ws.cell(r, 9).value or "").strip()))
    return out


def role_count(ws, last=None):
    return len(ledger(ws, last))


def load(path):
    return openpyxl.load_workbook(path, data_only=False)


def save(wb, path):
    wb.save(path)
    return path
