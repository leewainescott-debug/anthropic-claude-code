#!/usr/bin/env python3
"""u6 - spec sections G3, G4 and H: the model made hard to break.

  python3 u6_harden.py <in.xlsx> <out.xlsx> [recalculated_in.xlsx]

Strict dropdowns instead of sheet protection, the long formulas broken into
helper columns that say what they mean, no format anywhere that prints a dash
for a zero, and every control still reading 0 in white font.

The optional third argument is a recalculated copy of the input; the format
sweep needs to know which cells actually hold a number. Without it the script
recalculates the input itself.

Idempotent: handed its own output it copies it through untouched.
"""
import sys, os, re, shutil, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import MultiCellRange, CellRange

import wbio
from _xl import REVIEW, LEVERS, Log, load, save, shift_rows, white, copy_style

EXEC = "Exec Summary"
QA = "4.0 Data QA"
T33 = "3.3 Squad Actuals to Archetype"
SRC = ("0.1 Budget Table (Fin)", "0.4 Presentation Pack")
# a column on the QA grid and on the executive summary deliberately mixes
# counts with dollars, so the in-column format rule does not apply there
NO_MODAL = SRC + (EXEC, QA, REVIEW, "Lists")

ARCH = "'0.3 Squad Archetypes'"
HELP = {27: ("Onshore rate ($m)", "$G$5:$G$23"),
        28: ("Offshore rate ($m)", "$H$5:$H$23"),
        29: ("Archetype roles", "$F$5:$F$23")}
SHORT = ('=IFERROR(IF($E{r}="Onshore",$AA{r}*1,IF($E{r}="Hybrid",'
         '$AA{r}*MIN({a}!$K$8,$AC{r})/$AC{r}+$AB{r}*($AC{r}-MIN({a}!$K$8,$AC{r}))'
         '/$AC{r},$AB{r}*1)),"check size")')

# a squad somebody else pays for is measured against what they fund, at actual;
# every other squad is measured against what TDD is left carrying
FIRST, LAST = 2, 10
VAR_PICK = "IF(COUNTIF(Lists!$AU${a}:$AU${z},$B{r}),$O{r},$Q{r})"

OLD_DEFAULT = ",Lists!$AC$2:$AC$5,0)),1)"
NEW_DEFAULT = ",Lists!$AC$2:$AC$5,0)),NA())"

# Exec's five vacancy chains, in the order they sit on the tab
COUNTS = [("J", '=COUNTIFS($D${a}:$D${z},"Vacant",$E${a}:$E${z},"Hire")'),
          ("K", '=COUNTIFS($D${a}:$D${z},"Vacant",$E${a}:$E${z},"Offshore")'),
          ("L", '=COUNTIFS($D${a}:$D${z},"Vacant",$E${a}:$E${z},"Hold")'),
          ("M", '=COUNTIFS($D${a}:$D${z},"Vacant",$E${a}:$E${z},"Filled")'),
          ("N", '=COUNTIFS($D${a}:$D${z},"Filled",$E${a}:$E${z},"Offshore")')]
EXEC_ROWS = {26: "J", 27: "K", 28: "L", 29: "M", 30: "N"}

# the independent QA pass: one dropped owner edit and nine finish defects
D1_ROW = 464                       # his 2.5 B39 overtype, in the ledger
D7_TAB, D7_LEDGER, D7_GROUP = "2.7 Infrastructure", 372, "Technology Manager"
WORDS = [("Cost after decision ($)", "Cost after lever ($)"),
         ("The vacancy decision", "The vacancy levers"),
         ("Decision impact on 3.1", "Lever impact on 3.1"),
         ("cost after decisions", "cost after levers")]
ORPHAN_FILLS = [("2.2 Customer", ["H55", "I55", "J55", "K55", "L55", "M55"]),
                ("2.2 Customer", ["A79", "H79", "I79", "J79", "K79", "L79"]),
                ("2.6 Finance", ["B44", "C44", "D44", "H44"])]

LEVER_CTRL = ("Control - every lever on this tab is one of the four values, "
              "must be 0")
SUMMARY_LBL = "Vacancy levers on this tab, for the executive summary"

DASH = re.compile(r'(\\-|"-")')


# ----------------------------------------------------------------- formats

def sections(fmt):
    out, cur, q, i = [], "", False, 0
    while i < len(fmt):
        ch = fmt[i]
        if ch == '"':
            q = not q
            cur += ch
        elif ch == "\\" and i + 1 < len(fmt):
            cur += fmt[i:i + 2]
            i += 2
            continue
        elif ch == ";" and not q:
            out.append(cur)
            cur = ""
        else:
            cur += ch
        i += 1
    out.append(cur)
    return out


def undash(fmt):
    """A zero section that prints a dash prints the number instead."""
    if not fmt or fmt == "General":
        return fmt
    s = sections(fmt)
    if len(s) >= 3 and DASH.search(s[2]):
        s[2] = s[0]
        return ";".join(s)
    return fmt


# ------------------------------------------------------------- tab helpers

def tabs2x(wb):
    return [ws.title for ws in wb.worksheets if ws.title.startswith("2.")]


def total_row(ws):
    for r in range(5, 60):
        if ws.cell(r, 2).value == "Total portfolio":
            return r
    raise SystemExit("STOP: no 'Total portfolio' row on %s" % ws.title)


def fte_hdr(ws):
    for r in range(5, 120):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.endswith(" FTE"):
            return r + 1
    raise SystemExit("STOP: no FTE block on %s" % ws.title)


def block_span(ws):
    """(first row of the helper block, last role row)."""
    hdr = fte_hdr(ws)
    first = last = None
    for r in range(hdr + 1, ws.max_row + 1):
        c, d = ws.cell(r, 3).value, ws.cell(r, 4).value
        if isinstance(c, str) and c.startswith("=COUNTIF("):
            first = first or r
        elif isinstance(d, str) and REVIEW in d and "$AK$" in d:
            first = first or r
            last = r
    return first, last


def role_rows(ws):
    out = []
    for r in range(1, ws.max_row + 1):
        d = ws.cell(r, 4).value
        if isinstance(d, str) and REVIEW in d and "$AK$" in d:
            out.append(r)
    return out


def main(src, dst, pre=None):
    log = Log("u6_harden")
    wb = load(src)

    a21 = wb["2.1 Ampol Retail"]
    if any(a21.cell(r, 2).value == SUMMARY_LBL
           for r in range(1, a21.max_row + 1)):
        print("input is already hardened - copying through")
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        return

    if pre is None:
        pre = wbio.recalc(src)
    vals = openpyxl.load_workbook(pre, data_only=True)

    # ----------------------------------------------------------------- D2
    log.head("D2  the variance compares like with like, per row kind")
    for title in tabs2x(wb):
        ws = wb[title]
        tot = total_row(ws)
        n_row = n_tot = 0
        for r in range(7, tot + 1):
            var = ws.cell(r, 18).value
            if not isinstance(var, str) or ("$Q%d" % r) not in var:
                continue
            o = ws.cell(r, 15).value
            if isinstance(o, str) and o.startswith("=SUM("):
                # a total row carries both kinds, so it compares the two totals
                ws.cell(r, 18).value = var.replace("$Q%d" % r, "$O%d" % r)
                n_tot += 1
            else:
                ws.cell(r, 18).value = var.replace(
                    "$Q%d" % r, VAR_PICK.format(r=r, a=FIRST, z=LAST))
                n_row += 1
        log("D2", title,
            "%d squad rows read the actual against the funded figure when the "
            "squad is funded outside, the TDD-funded cost when it is not; "
            "%d total rows compare the two totals" % (n_row, n_tot))

    # ---------------------------------------------------------------- G4a
    log.head("G4  a lever the model cannot price shows #N/A, never full price")
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and OLD_DEFAULT in v:
                    c.value = v.replace(OLD_DEFAULT, NEW_DEFAULT)
                    n += 1
    log("G4", "workbook", "%d lever cost formulas default to NA() not 1" % n)

    # ---------------------------------------------------------------- G4b/H3
    log.head("G4 / H3  a control and a summary row on every lever modelling tab")
    exec_at = {}
    for title in tabs2x(wb):
        ws = wb[title]
        a, z = block_span(ws)
        ctrl, summ = z + 2, z + 3
        ws.cell(ctrl, 2).value = LEVER_CTRL
        ws.cell(ctrl, 3).value = (
            '=SUMPRODUCT(($E${a}:$E${z}<>"")*($E${a}:$E${z}<>"Filled")'
            '*($E${a}:$E${z}<>"Hire")*($E${a}:$E${z}<>"Hold")'
            '*($E${a}:$E${z}<>"Offshore"))').format(a=a, z=z)
        ws.cell(ctrl, 3).number_format = "0"
        ws.cell(summ, 2).value = SUMMARY_LBL
        for col, f in COUNTS:
            ws[col + str(summ)].value = f.format(a=a, z=z)
            ws[col + str(summ)].number_format = "0"
        white(ws, "B%d" % ctrl, "C%d" % ctrl, "B%d" % summ,
              *[c + str(summ) for c, _ in COUNTS])
        exec_at[title] = summ
        log("G4", "%s!B%d:N%d" % (title, ctrl, summ),
            "lever control and the executive summary row, white font")

    log.head("H3  the executive summary's five chains become short sums")
    ex = wb[EXEC]
    for row, col in EXEC_ROWS.items():
        was = len(ex.cell(row, 3).value)
        f = "=" + "+".join("N('%s'!$%s$%d)" % (t, col, exec_at[t])
                           for t in tabs2x(wb))
        ex.cell(row, 3).value = f
        log("H3", "%s!C%d" % (EXEC, row),
            "%d chars -> %d, reading the per tab summary rows" % (was, len(f)))

    # ---------------------------------------------------------------- H3
    log.head("H3  the archetype formula becomes three helper columns")
    done = collections.Counter()
    for ws in wb.worksheets:
        rows = [c.row for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and len(c.value) > 900
                and c.value.startswith("=IFERROR(IF($E")
                and "0.3 Squad Archetypes" in c.value and c.column == 8]
        if not rows:
            continue
        heads = set()
        for r in rows:
            for col, (lbl, rng) in HELP.items():
                ws.cell(r, col).value = ("=IFERROR(INDEX(%s!%s,MATCH($C%d&\"|\"&$D%d,"
                                         "%s!$A$5:$A$23,0)),\"\")"
                                         % (ARCH, rng, r, r, ARCH))
                copy_style(ws.cell(r, 8), ws.cell(r, col))
                ws.cell(r, col).number_format = "#,##0.00" if col < 29 else "0"
            ws.cell(r, 8).value = SHORT.format(r=r, a=ARCH)
            h = r - 1
            while h > 1 and ws.cell(h, 3).value != "Squad Type":
                h -= 1
            heads.add(h)
            done[ws.title] += 1
        for h in heads:
            for col, (lbl, rng) in HELP.items():
                copy_style(ws.cell(h, 8), ws.cell(h, col))
                ws.cell(h, col).value = lbl
            white(ws, *["%s%d" % (openpyxl.utils.get_column_letter(c), h)
                        for c in HELP])
        log("H3", ws.title,
            "%d archetype formulas at 1,011 chars -> %d, helper columns AA:AC"
            % (done[ws.title], len(SHORT.format(r=99, a=ARCH))))

    # ---------------------------------------------------------------- H2
    log.head("H2  every dropdown a GM types into is strict")
    n = 0
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            if dv.type != "list":
                continue
            dv.showErrorMessage = True
            dv.errorTitle = "Invalid entry"
            dv.error = "Pick a value from the list"
            n += 1
    log("H2", "workbook",
        "%d list validations now refuse anything off the list" % n)
    groups = []
    t33 = wb[T33]
    for r in range(6, 122):
        v = t33.cell(r, 2).value
        if isinstance(v, str) and v and not v.endswith(" total") \
                and v not in groups:
            groups.append(v)
    if len(groups) != 15:
        print("STOP: 3.3 carries %d groups, not 15: %r" % (len(groups), groups))
        raise SystemExit(2)
    ex.data_validations.dataValidation = []
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(groups),
                        allow_blank=False, showErrorMessage=True,
                        errorTitle="Invalid entry",
                        error="Pick a value from the list")
    ex.add_data_validation(dv)
    dv.sqref = MultiCellRange([CellRange("C36")])
    log("H2", "%s!C36" % EXEC,
        "drill-down picker lists all %d groups (was 10 of 15)" % len(groups))
    guard = ("=IF(COUNTIF('%s'!$B$6:$B$121,$C$36)=0,\"Not a group\","
             % T33)
    n = 0
    for r in range(37, 46):
        v = ex.cell(r, 3).value
        if isinstance(v, str) and v.startswith("=") and not v.startswith(guard[:8]):
            ex.cell(r, 3).value = guard + v[1:] + ")"
            n += 1
    log("H2", "%s!C37:C45" % EXEC,
        "%d drill-down cells read 'Not a group' rather than 0 for a name 3.3 "
        "does not carry" % n)

    # ---------------------------------------------------------------- H5
    log.head("H5  the counts in his sentences come off the live model")
    o32 = wb["3.2 Overhead & Leadership"]
    o32["L9"].value = '="All "&TEXT($G9,"0")&" in the portfolios"'
    o32["L9"].fill = PatternFill(fill_type=None)
    log("H5", "3.2 Overhead & Leadership!L9",
        "'All 24 in the portfolios' -> his words, the count live off $G9; the "
        "cream comes off, it is a formula now")
    b8 = ex["B8"].value
    ex["B8"].value = ('="Cost of the "&TEXT(COUNTA(\'' + REVIEW +
                      "'!$B$2:$B$700),\"0\")&\" roles in the role mapping ($m)\"")
    log("H5", "%s!B8" % EXEC, "%r -> the live count (A4: never hardcode)" % b8)

    # ---------------------------------------------------------------- G3
    log.head("G3  the overflow control for the two spare ledger rows")
    q = wb[QA]
    last = None
    for r in range(5, 100):
        if isinstance(q.cell(r, 2).value, str) and \
                q.cell(r, 2).value == "Checks failing":
            last = r
    if last is None:
        print("STOP: no 'Checks failing' row on 4.0")
        raise SystemExit(2)
    shift_rows(wb, QA, last, 1)
    q = wb[QA]
    for c in range(2, 6):
        copy_style(q.cell(last - 1, c), q.cell(last, c))
    q.cell(last, 2).value = ("Roles typed below the range the model reads - "
                             "REVIEW rows 701 and beyond")
    q.cell(last, 3).value = ("=COUNTA('%s'!$B$2:$B$1000)" % REVIEW)
    q.cell(last, 4).value = ("=COUNTA('%s'!$B$2:$B$700)" % REVIEW)
    q.cell(last, 5).value = "=ROUND($C%d-$D%d,6)" % (last, last)
    q.cell(last + 1, 5).value = '=COUNTIF($E$5:$E$%d,"<>0")' % last
    log("G3", "%s row %d" % (QA, last),
        "overflow control: COUNTA to row 1000 less COUNTA to row 700, must be 0")

    # ---------------------------------------------------------------- H4
    log.head("H4  no format on a model tab prints a dash for a zero")
    n = 0
    for ws in wb.worksheets:
        if ws.title in SRC:
            continue
        for row in ws.iter_rows():
            for c in row:
                f = undash(c.number_format)
                if f != c.number_format:
                    c.number_format = f
                    n += 1
    log("H4", "workbook", "%d cells lose their dash zero section" % n)

    moved = 0
    for ws in wb.worksheets:
        if ws.title in NO_MODAL:
            continue
        wv = vals[ws.title] if ws.title in vals.sheetnames else None
        if wv is None:
            continue
        for col in range(1, ws.max_column + 1):
            run = []
            for r in range(1, ws.max_row + 2):
                v = ws.cell(r, col).value if r <= ws.max_row else None
                live = v is not None and (isinstance(v, (int, float))
                                          or (isinstance(v, str) and v.startswith("=")))
                if live:
                    run.append(r)
                    continue
                num = [x for x in run
                       if isinstance(wv.cell(x, col).value, (int, float))
                       and not isinstance(wv.cell(x, col).value, bool)]
                if len(num) >= 3:
                    fm = collections.Counter(ws.cell(x, col).number_format
                                             for x in num)
                    mode = fm.most_common(1)[0][0]
                    for x in num:
                        if ws.cell(x, col).number_format != mode:
                            ws.cell(x, col).number_format = mode
                            moved += 1
                run = []
    log("H4", "the grid tabs",
        "%d cells take the format the rest of their column carries" % moved)

    log.head("the house rule: a control on a visible tab is invisible")
    n = 0
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        for row in ws.iter_rows():
            for cl in row:
                t = cl.value
                if not (isinstance(t, str) and "must be 0" in t
                        and t.startswith(("Control - ", "Check - "))):
                    continue
                cells = [cl.coordinate]
                for cc in range(cl.column + 1, ws.max_column + 1):
                    if ws.cell(cl.row, cc).value is not None:
                        cells.append(ws.cell(cl.row, cc).coordinate)
                white(ws, *cells)
                n += 1
    log("rule", "the visible tabs",
        "%d control rows in white font - functional, reads 0, invisible" % n)

    # ------------------------------------------------------------ QA finish
    log.head("QA  the dropped owner edit and the finish defects")

    # D1 - his 2.5 B39 overtype was name to vacancy: honour it in the ledger
    rv = wb[REVIEW]
    if rv["B%d" % D1_ROW].value != "Vacant":
        log("D1", "REVIEW!B%d" % D1_ROW,
            "%r -> 'Vacant' (his 2.5 B39 overtype, in the proper cell); his "
            "Hire lever on 2.5 E39 stays" % rv["B%d" % D1_ROW].value)
        rv["B%d" % D1_ROW].value = "Vacant"

    # D4 - the levers sweep finished
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cl in row:
                t = cl.value
                if not isinstance(t, str) or t.startswith("="):
                    continue
                new = t
                for a, b in WORDS:
                    new = new.replace(a, b)
                if new != t:
                    cl.value = new
                    n += 1
    log("D4", "workbook",
        "%d labels move off 'decision' onto his lever wording" % n)

    # D5 / D6
    t33 = wb[T33]
    t33["B4"].value = t33["B4"].value.replace("working tab", "lever modelling tab")
    log("D5", "%s!B4" % T33, "%r" % t33["B4"].value)
    for r in range(2, rv.max_row + 1):
        v3 = rv.cell(r, 3).value
        if isinstance(v3, str) and "(ios)" in v3:
            rv.cell(r, 3).value = v3.replace("(ios)", "(iOS)")
            log("D6", "REVIEW!C%d" % r, "(ios) -> (iOS)")

    # D9 - one rounding, not fifteen
    ex["C20"].value = "='3.2 Overhead & Leadership'!$K$13"
    log("D9", "%s!C20" % EXEC,
        "reads 3.2's once-rounded portfolio share, so the 2e-6 drift goes")

    # D3 - the highlights the relocated notes left behind
    for tab, cells in ORPHAN_FILLS:
        w = wb[tab]
        for co in cells:
            w[co].fill = PatternFill(fill_type=None)
        log("D3", "%s!%s" % (tab, ",".join(cells)),
            "highlight cleared, the note lives in REVIEW Commentry")

    # D10 - one FTE format down the column, the total row included
    n = 0
    for title in tabs2x(wb):
        ws = wb[title]
        tot = total_row(ws)
        live = [r for r in range(7, tot + 1)
                if ws.cell(r, 7).value is not None
                and ws.cell(r, 7).value != '=""']
        fm = collections.Counter(ws.cell(r, 7).number_format for r in live)
        mode = fm.most_common(1)[0][0]
        for r in live:
            if ws.cell(r, 7).number_format != mode:
                ws.cell(r, 7).number_format = mode
                n += 1
    log("D10", "the lever modelling tabs",
        "%d FTE cells take the one decimal the column carries" % n)

    # D2 - the bands run unbroken across the two new columns
    n = 0
    for title in tabs2x(wb):
        ws = wb[title]
        for r in range(1, total_row(ws) + 3):
            src = ws.cell(r, 15)
            if not src.fill.patternType:
                continue
            for col in (16, 17):
                band = ws.cell(r, col)
                if (band.fill.patternType != src.fill.patternType
                        or band.fill.fgColor.rgb != src.fill.fgColor.rgb):
                    copy_style(src, ws.cell(r, col))
                    n += 1
    w31 = wb["3.1 Archetype to Actuals"]
    for col in (6, 7):
        copy_style(w31.cell(5, 5), w31.cell(5, col))
    log("D2", "the lever modelling tabs and 3.1",
        "%d band cells filled across P and Q, and 3.1 F5:G5 - the title row "
        "and the section bands run unbroken again" % (n + 2))

    # D8 - the group totals stop rendering as hashes
    for title in tabs2x(wb):
        wb[title].column_dimensions["F"].width = 12
    log("D8", "the lever modelling tabs", "column F set to width 12")

    # D7 - his C372 repair reclassified the person; the helper grouping follows
    w27 = wb[D7_TAB]
    src = None
    for r in range(1, w27.max_row + 1):
        d = w27.cell(r, 4).value
        if isinstance(d, str) and REVIEW in d and d.endswith("$AK$%d" % D7_LEDGER):
            src = r
    if src is None:
        print("STOP: no helper row for ledger %d on %s" % (D7_LEDGER, D7_TAB))
        raise SystemExit(2)
    lever = w27.cell(src, 5).value
    shift_rows(wb, D7_TAB, src, -1)
    w27 = wb[D7_TAB]
    hdr = last = None
    for r in range(1, w27.max_row + 1):
        c3 = w27.cell(r, 3).value
        if (w27.cell(r, 2).value == D7_GROUP and isinstance(c3, str)
                and c3.startswith("=COUNTIF(")):
            hdr = r
        elif hdr and last is None:
            d = w27.cell(r, 4).value
            if not (isinstance(d, str) and REVIEW in d and "$AK$" in d):
                last = r - 1
    shift_rows(wb, D7_TAB, last, 1)
    w27 = wb[D7_TAB]
    donor = last + 1
    for col in range(2, 8):
        copy_style(w27.cell(donor, col), w27.cell(last, col))
    q = "='" + REVIEW + "'!$%s$%d"
    w27.cell(last, 2).value = q % ("B", D7_LEDGER)
    w27.cell(last, 3).value = q % ("C", D7_LEDGER)
    w27.cell(last, 4).value = q % ("AK", D7_LEDGER)
    w27.cell(last, 5).value = lever
    w27.cell(last, 6).value = q % ("AA", D7_LEDGER)
    g = w27.cell(donor, 7).value
    g = re.sub(r"\$F\d+", "$F%d" % last, g)
    g = re.sub(r"\$E\d+", "$E%d" % last, g)
    g = re.sub(r"(\$Q\$)\d+", lambda m: m.group(1) + str(D7_LEDGER), g)
    w27.cell(last, 7).value = g
    for dv in w27.data_validations.dataValidation:
        if dv.formula1 and "Filled" in str(dv.formula1):
            keep = list(dv.sqref.ranges)
            keep.append(CellRange(min_col=5, max_col=5, min_row=last,
                                  max_row=last))
            dv.sqref = MultiCellRange(keep)
    log("D7", "%s row %d" % (D7_TAB, last),
        "the helper row for ledger %d joins the %s group, so its actual and "
        "its cost after levers sit on the same row" % (D7_LEDGER, D7_GROUP))

    cfg = wb["0.2 Data Config"]
    for co in ("K22", "L22", "M22", "N22", "O22"):
        c = cfg[co]
        fo = c.font
        c.font = Font(name=fo.name, size=fo.size, bold=fo.bold, italic=True,
                      color="FF808080")
    log("H4", "0.2 Data Config!K22:O22",
        "his 'Position on 23/07' block reads as the note it is, text kept")

    tmp = dst + ".raw"
    save(wb, tmp)
    log.head("recalculating and writing the cached values back")
    rc, st = wbio.build(tmp, dst)
    os.remove(tmp)
    print("recalculated, %d formula cells populated across %d sheets"
          % (st["cells"], st["sheets"]), flush=True)
    err, blank = wbio.audit(dst)
    if err:
        print("STOP: %d error cells, e.g. %r" % (len(err), err[:5]))
        raise SystemExit(2)
    log.tail()
    print("wrote", dst)


if __name__ == "__main__":
    main(*sys.argv[1:])
