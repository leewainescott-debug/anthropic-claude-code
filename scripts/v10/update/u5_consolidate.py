#!/usr/bin/env python3
"""u5 - spec section F: one tab per COE.

  python3 u5_consolidate.py <in.xlsx> <out.xlsx>

1.11, 1.12 and 1.13 said the same thing as 2.12, 2.13 and 2.11 in a second
language, with a second lever column and, on 1.12, an inverted sign. The
funding story moves onto the lever modelling tab, the cyber uplift toggles
become a column on 2.11, and the three tabs go.

Idempotent: handed its own output it copies it through untouched.
"""
import sys, os, re, shutil
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import MultiCellRange, CellRange

from _xl import (REVIEW, Log, load, save, shift_rows, drop_sheet, copy_style,
                 row_style)

DEAD = ["1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles"]
CYBER = "2.11 Cyber Risk & Service Ops"
BPT = "2.12 COE BP&T"
SAD = "2.13 COE SA&D"

UPLIFT_STEPS = ",".join("%d%%" % p for p in range(0, 101, 5))

# his own wording, carried across from the tabs that go
M2 = "#,##0.00"

BLOCKS = {
    CYBER: [
        ("Funding", None, None),
        ("COE - Cyber, Risk & Service Ops allocation ($m) - 0.2 Data Config",
         "='0.2 Data Config'!$E$7", M2),
        ("Total budget to draw down ($m)", "=$C${r1}", M2),
        ("Planned spend ($m)", "=$S${tot}", M2),
        ("Variance ($m)", "=$C${r2}-$C${r3}", M2),
    ],
    BPT: [
        ("Funding", None, None),
        ("Portfolios funded (3.1 Archetype to Actuals)",
         "=COUNTA(Lists!$AS$2:$AS$12)", "0"),
        ("Business Partner allocation per portfolio (FTE - 0.2 Data Config)",
         "='0.2 Data Config'!$M$7", "0.00"),
        ("Business Partner FTEs funded by portfolio overheads",
         "=$C${r1}*$C${r2}", "0.00"),
        ("Business Partner funding from portfolio overheads ($m)",
         "=COUNTA(Lists!$AS$2:$AS$12)*'0.2 Data Config'!$N$7", M2),
        ("COE - Business Partnering allocation ($m) - 0.2 Data Config",
         "='0.2 Data Config'!$E$9", M2),
        ("COE - Transformation allocation ($m) - 0.2 Data Config",
         "='0.2 Data Config'!$E$8", M2),
        ("Total budget to draw down ($m)", "=$C${r5}+$C${r6}", M2),
        ("Business Partnering planned spend ($m)", "=$S$8+$S$9-$C${r4}", M2),
        ("Transformation planned spend ($m)", "=$S$10", M2),
        ("Total planned spend ($m)", "=$C${r8}+$C${r9}", M2),
        ("Variance ($m)", "=$C${r7}-$C${r10}", M2),
        ("Business Partner funding met by portfolio overheads, netted out of "
         "planned spend ($m)", "=-$C${r4}", M2),
        ("Planned spend is net of the Business Partner FTEs funded inside "
         "portfolio overheads (row {r4}); the COE draws down its own "
         "allocation only.", None, None),
    ],
    SAD: [
        ("Funding", None, None),
        ("Portfolios funded (3.1 Archetype to Actuals)",
         "=COUNTA(Lists!$AS$2:$AS$12)", "0"),
        ("Domain Architect allocation per portfolio (FTE - 0.2 Data Config)",
         "='0.2 Data Config'!$M$8", "0.00"),
        ("Number of Domain Architects funded by Portfolios",
         "=$C${r1}*$C${r2}", "0.00"),
        ("Domain Architect funding from portfolio overheads ($m)",
         "=COUNTA(Lists!$AS$2:$AS$12)*'0.2 Data Config'!$N$8", M2),
        ("COE - Strategy Architecture allocation ($m) - 0.2 Data Config",
         "='0.2 Data Config'!$E$6", M2),
        ("COE - Data allocation ($m) - 0.2 Data Config",
         "='0.2 Data Config'!$E$10", M2),
        ("Total budget to draw down ($m)", "=$C${r5}+$C${r6}", M2),
        ("Strategy & Architecture planned spend ($m)",
         "=$S$8+$S$11+$S$12-$C${r4}", M2),
        ("Data planned spend ($m)", "=$S$9+$S$10", M2),
        ("Total planned spend ($m)", "=$C${r8}+$C${r9}", M2),
        ("Variance ($m)", "=$C${r7}-$C${r10}", M2),
        ("Domain Architect funding met by portfolio overheads, netted out of "
         "planned spend ($m)", "=-$C${r4}", M2),
        ("Planned spend is net of the Domain Architect FTEs funded inside "
         "portfolio overheads (row {r4}); COEs draw down on their own "
         "allocation only.", None, None),
    ],
}


def total_row(ws):
    for r in range(5, 60):
        if ws.cell(r, 2).value == "Total portfolio":
            return r
    raise SystemExit("STOP: no 'Total portfolio' row on %s" % ws.title)


def fte_title_row(ws):
    for r in range(5, 80):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.endswith(" FTE"):
            return r
    raise SystemExit("STOP: no FTE block on %s" % ws.title)


def block_rows(ws, hdr):
    """(group header rows, role rows) of the FTE helper block below `hdr`."""
    groups, roles = [], []
    for r in range(hdr + 1, ws.max_row + 1):
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        if isinstance(c, str) and c.startswith("=COUNTIF("):
            groups.append(r)
        elif isinstance(d, str) and REVIEW in d and "$AK$" in d:
            roles.append(r)
    return groups, roles


def put_block(wb, log, title, tot):
    """Lay the funding block in below the controls, above the FTE block."""
    ws = wb[title]
    at = fte_title_row(ws)
    spec = BLOCKS[title]
    n = len(spec) + 1                     # one blank line before the FTE block
    shift_rows(wb, title, at, n)
    ws = wb[title]
    rows = {"r%d" % i: at + i for i in range(len(spec))}
    rows["tot"] = tot
    for i, (label, formula, fmt) in enumerate(spec):
        r = at + i
        copy_style(ws.cell(7 if i == 0 else 8, 2), ws.cell(r, 2))
        copy_style(ws.cell(8, 3), ws.cell(r, 3))
        ws.cell(r, 2).value = label.format(**rows)
        if formula:
            ws.cell(r, 3).value = formula.format(**rows)
            ws.cell(r, 3).number_format = fmt
    log("F1", "%s!B%d:C%d" % (title, at, at + len(spec) - 1),
        "funding block moved onto the tab: allocation, budget to draw down, "
        "planned spend, variance%s"
        % (", and the portfolio funded FTE rows" if title != CYBER else ""))
    return at


def main(src, dst):
    log = Log("u5_consolidate")
    wb = load(src)

    if not any(t in wb.sheetnames for t in DEAD):
        print("input already carries the consolidated COEs - copying through")
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        return

    # -------------------------------------------------------------- F1 (2.11)
    log.head("F1  2.11 takes the cyber Uplift %% column and the uplift slice")
    c11 = wb[CYBER]
    cream = wb["1.13 Cyber Roles"]["I22"]
    hdr11 = fte_title_row(c11) + 1
    groups, roles = block_rows(c11, hdr11)
    c13 = wb["1.13 Cyber Roles"]
    carried = 0
    for r in roles:
        g = c11.cell(r, 7).value
        m = re.search(r"\(1-N\('1\.13 Cyber Roles'!\$I\$(\d+)\)\)", g)
        if not m:
            print("STOP: %s!G%d does not carry a 1.13 uplift toggle" % (CYBER, r))
            raise SystemExit(2)
        raw = c13.cell(int(m.group(1)), 9).value
        pct = raw or 0
        copy_style(cream, c11.cell(r, 8))
        c11.cell(r, 8).value = pct
        c11.cell(r, 8).number_format = "0%"
        c11.cell(r, 9).value = g[:m.start()] + "N($H%d)" % r
        copy_style(c11.cell(r, 7), c11.cell(r, 9))
        c11.cell(r, 7).value = g[:m.start()] + "(1-N($H%d))" % r
        if raw is not None:
            carried += 1
    for i, r in enumerate(groups):
        end = groups[i + 1] - 1 if i + 1 < len(groups) else max(roles)
        mine = [x for x in roles if r < x <= end]
        copy_style(c11.cell(r, 7), c11.cell(r, 9))
        c11.cell(r, 9).value = "=SUM(I%d:I%d)" % (min(mine), max(mine))
    copy_style(c11.cell(hdr11, 7), c11.cell(hdr11, 8))
    copy_style(c11.cell(hdr11, 7), c11.cell(hdr11, 9))
    c11.cell(hdr11, 8).value = "Uplift %"
    c11.cell(hdr11, 9).value = "Charged to the cyber uplift program ($)"
    dv = DataValidation(type="list", formula1='"%s"' % UPLIFT_STEPS,
                        allow_blank=False, showErrorMessage=True,
                        errorTitle="Invalid entry",
                        error="Pick a value from the list")
    c11.add_data_validation(dv)
    dv.sqref = MultiCellRange([CellRange(min_col=8, max_col=8, min_row=r,
                                         max_row=r) for r in roles])
    log("F1", "%s!H%d:I%d" % (CYBER, hdr11, max(roles)),
        "Uplift %% column (cream, dropdown 0%% to 100%% in 5%% steps) and the "
        "uplift slice column; %d toggles carried off 1.13" % carried)
    log("F2", CYBER,
        "column G now reads the local Uplift %% column, not 1.13's")

    # ---------------------------------------------------------- F1 (the blocks)
    log.head("F1  the funding story moves onto each COE tab")
    slice_groups = list(groups)
    tots = {t: total_row(wb[t]) for t in (CYBER, BPT, SAD)}
    at = {}
    for t in (CYBER, BPT, SAD):
        at[t] = put_block(wb, log, t, tots[t])
    # the uplift slice rows moved down with the FTE block on 2.11
    moved = len(BLOCKS[CYBER]) + 1
    slice_groups = [r + moved for r in slice_groups]

    # ---------------------------------------------------------------- F3
    log.head("F3  every reference re-points, then the three tabs go")
    cfg = wb["0.2 Data Config"]
    reads = [("F6", SAD, 8), ("F7", CYBER, 3), ("F8", BPT, 9), ("F9", BPT, 8),
             ("F10", SAD, 9)]
    for coord, tab, off in reads:
        new = "='%s'!$C$%d" % (tab, at[tab] + off)
        log("F3", "0.2 Data Config!%s" % coord,
            "%r -> %s" % (cfg[coord].value, new))
        cfg[coord].value = new
    old = cfg["B7"].value
    cfg["B7"].value = old.replace("(see 1.13 Cyber Roles)", "(see %s)" % CYBER)
    log("F3", "0.2 Data Config!B7",
        "%r -> %r (the COE note points at the consolidated tab)"
        % (old, cfg["B7"].value))

    c14 = wb["1.14 TDD Cyber"]
    new = "=SUM(%s)/1000000" % ",".join("'%s'!$I$%d" % (CYBER, r)
                                        for r in slice_groups)
    log("F3", "1.14 TDD Cyber!J16", "%r -> %s" % (c14["J16"].value, new))
    c14["J16"].value = new
    for coord in ("H11", "H16"):
        t = c14[coord].value
        c14[coord].value = t.replace("1.13 Cyber Roles", CYBER)
        log("F3", "1.14 TDD Cyber!%s" % coord, "label names %s" % CYBER)

    q = wb["4.0 Data QA"]
    checks = [
        (33, BPT, "2.12 COE BP&T groupings against the role mapping",
         "='%s'!$F$8+'%s'!$F$9+'%s'!$F$10" % (BPT, BPT, BPT), "COE BP&T"),
        (34, SAD, "2.13 COE SA&D groupings against the role mapping",
         "='%s'!$F$8+'%s'!$F$9+'%s'!$F$10+'%s'!$F$11+'%s'!$F$12"
         % (SAD, SAD, SAD, SAD, SAD), "COE SA&D"),
        (35, CYBER, "2.11 Cyber Risk & Service Ops groupings against the role "
                    "mapping",
         "='%s'!$F$8+'%s'!$F$9+'%s'!$F$10+'%s'!$F$11+'%s'!$F$12"
         % (CYBER, CYBER, CYBER, CYBER, CYBER), "COE Cyber"),
    ]
    for row, tab, label, c, key in checks:
        q.cell(row, 2).value = label
        q.cell(row, 3).value = c
        q.cell(row, 4).value = ("=COUNTIFS('%s'!$AJ$2:$AJ$700,\"%s\",'%s'"
                                "!$B$2:$B$700,\"<>\")" % (REVIEW, key, REVIEW))
        log("F3", "4.0 Data QA row %d" % row, "%s" % label)
    spend = [(36, BPT, "2.12 COE BP&T funding block against its cost after "
                       "levers ($m)", 10, 4),
             (37, SAD, "2.13 COE SA&D funding block against its cost after "
                       "levers ($m)", 10, 4),
             (38, CYBER, "2.11 Cyber Risk & Service Ops funding block against "
                         "its cost after levers ($m)", 3, None)]
    for row, tab, label, off, net in spend:
        q.cell(row, 2).value = label
        c = "='%s'!$C$%d" % (tab, at[tab] + off)
        if net:
            c += "+'%s'!$C$%d" % (tab, at[tab] + net)
        q.cell(row, 3).value = c
        q.cell(row, 4).value = "='%s'!$S$%d" % (tab, tots[tab])
        log("F3", "4.0 Data QA row %d" % row, "%s" % label)
    q["B84"].value = q["B84"].value.replace("1.13 Cyber Roles", CYBER)
    log("F3", "4.0 Data QA!B84", "the uplift fact names %s" % CYBER)

    left = []
    for ws in wb.worksheets:
        if ws.title in DEAD:
            continue
        for r in ws.iter_rows():
            for cl in r:
                if isinstance(cl.value, str) and any(d in cl.value for d in DEAD):
                    left.append("%s!%s" % (ws.title, cl.coordinate))
    if left:
        print("STOP: %d references to the three tabs are still live: %r"
              % (len(left), left[:8]))
        raise SystemExit(2)
    for t in DEAD:
        drop_sheet(wb, t)
        log("F3", t, "tab removed, nothing points at it")
    log.note("F2", "one lever column per COE: 1.11/1.12/1.13's On/Off column "
                   "died with the tabs and their 15 blank lever cells with it; "
                   "every lever state was already carried on the 2.x tabs, his "
                   "six GRC offshores included")
    log.note("F4", "3.4 keeps reading the ledger and names none of the three")

    save(wb, dst)
    log.tail()
    print("wrote", dst)


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
