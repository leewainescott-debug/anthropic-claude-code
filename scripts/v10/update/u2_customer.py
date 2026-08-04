#!/usr/bin/env python3
"""u2 - spec section B: the Customer V2 load through the alias layer.

  python3 u2_customer.py <in.xlsx> <out.xlsx> [Customer_Resource_22Jul26V2.xlsx]

PCM_Data only. The four overhead names never become 1.2 squad rows: they load
through Lists W:X onto overhead lines. Programme Management becomes a
Customer only overhead line of its own.
"""
import sys, os, re, shutil
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from _xl import (REVIEW, Log, load, save, shift_rows, white, row_style,
                 copy_style, ledger, read_block, write_block, set_dv,
                 wire_summary)

V2_DEFAULT = ("/root/.claude/uploads/e550b440-3996-5abb-87e5-bafafe598f82/"
              "ee008e30-Customer_Resource_22Jul26V2.xlsx")

COUNTRY = {"AUD": "Australia", "NZD": "NZ", "WIPRO": "WIPRO"}

ALIASES = [("Heads of Customer Portfolio", "Head of Technology"),
           ("Program Management", "Program Management"),
           ("Technology Manager", "Technology Manager"),
           ("Delivery Manager", "Delivery Manager")]

PM = "Program Management"

ORDER = ["Ampol App", "Ampol Loyalty & Martech", "Ampol Web", "Customer, AI",
         "Digital Operations", "EGI Customer", "Energy", "Leadership",
         "Z App and Web", "Z Loyalty & Martech",
         "Head of Technology", "Delivery Manager", "Technology Manager", PM]

SUMMARY = {"Ampol App": 8, "Ampol Loyalty & Martech": 9, "Ampol Web": 10,
           "Customer, AI": 11, "Digital Operations": 12, "Z App and Web": 13,
           "Z Loyalty & Martech": 14, "EGI Customer": 17, "Energy": 19,
           "Leadership": 20, "Head of Technology": 23, "Delivery Manager": 24,
           "Technology Manager": 25, PM: 26}

OVERHEAD_TITLES = [("head of ", "Head of Technology"), ("tdd bp", "Business Partner"),
                   ("domain architect", "Domain Architect"),
                   ("enterprise architect", "Domain Architect"),
                   ("delivery man", "Delivery Manager"),
                   ("technology manager", "Technology Manager"),
                   ("technology manger", "Technology Manager"),
                   ("tech manager", "Technology Manager")]


def norm(s):
    s = str(s or "").replace("–", "-").replace("—", "-")
    return re.sub(r"\s+", " ", s).strip().casefold()


def resolve_at(k, j, alias, lines):
    """Mirror the ledger's AP / AQ / AR / AT chain for a Customer role."""
    lead = (k.strip() == "Leadership" or j.strip() == "Leadership")
    ap = "Leadership" if lead else alias.get(k.strip(), k.strip())
    return ap


def resolve_ar(title, ap, lines):
    if ap in lines:
        return ap
    t = title.casefold()
    for needle, line in OVERHEAD_TITLES:
        if needle in t:
            return line
    return "Squad"


def main(src, dst, v2path=V2_DEFAULT):
    log = Log("u2_customer")
    wb = load(src)
    rv = wb[REVIEW]
    L = wb["Lists"]

    if L["AF8"].value == PM:
        print("input already carries the Customer V2 load - copying through")
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        return

    # ---------------------------------------------------------------- B4/B1
    log.head("B4  Programme Management becomes a Customer only overhead line")
    shift_rows(wb, "Lists", 8, 1)
    L = wb["Lists"]
    row_style(L, 7, 8, 32, 39)
    L["AF8"] = PM
    L["AG8"] = ("=SUMIFS('%s'!$AA$2:$AA$700,'%s'!$AT$2:$AT$700,$AF8)/1000000"
                % (REVIEW, REVIEW))
    L["AH8"] = 1
    L["AI8"] = "portfolio"
    L["AJ8"] = "=$AG8*$AH8"
    L["AM8"] = "Yes"
    log("B4", "Lists!AF8:AM8",
        "overhead line added, rate = its own cost basis, times applied 1, drawn in the portfolios")
    L["AJ9"] = "=SUM(AJ2:AJ8)"
    L["AJ10"] = '=SUMIF($AM$2:$AM$8,"Yes",$AJ$2:$AJ$8)'
    log("B4", "Lists!AJ9 AJ10", "overhead total and allowance drawn widened to the new line")

    log.head("B1  the alias layer")
    have = {str(L.cell(r, 23).value).strip() for r in range(2, 25)
            if L.cell(r, 23).value}
    r = 13
    for w_, x_ in ALIASES:
        if w_ in have:
            continue
        while L.cell(r, 23).value is not None:
            r += 1
        row_style(L, 11, r, 23, 24)
        L.cell(r, 23).value = w_
        L.cell(r, 24).value = x_
        log("B1", "Lists!W%d:X%d" % (r, r), "%r -> %r" % (w_, x_))
    keep = [(str(L.cell(rr, 23).value), str(L.cell(rr, 24).value))
            for rr in (11, 12)]
    log.note("B1", "historical aliases kept: %r" % (keep,))
    alias = {str(L.cell(rr, 23).value).strip(): str(L.cell(rr, 24).value).strip()
             for rr in range(2, 25) if L.cell(rr, 23).value}
    lines = [str(L.cell(rr, 32).value).strip() for rr in range(2, 9)]

    # ------------------------------------------- the ledger recognises the line
    log.head("B4  a canonical squad that is itself an overhead line IS that line")
    n = 0
    for rr, nm, ttl, pf in ledger(rv):
        cur = rv.cell(rr, 44).value
        pfx = '=IF(TRIM($B%d)="","",' % rr
        if not isinstance(cur, str) or not cur.startswith(pfx):
            continue
        if "COUNTIF(Lists!$AF" in cur:
            continue
        rv.cell(rr, 44).value = (
            pfx + "IF(COUNTIF(Lists!$AF$2:$AF$8,$AP%d),$AP%d," % (rr, rr)
            + cur[len(pfx):-1] + "))")
        n += 1
    log("B4", "REVIEW!AR2:AR%d" % (rr), "%d rows: overhead line follows the alias" % n)

    # ------------------------------------------------------- B3 / B4 / B5 load
    log.head("B3 / B4 / B5  PCM_Data loads")
    v2 = openpyxl.load_workbook(v2path, data_only=True)["PCM_Data"]
    pool = {}
    n_v2 = 0
    for r2 in range(2, v2.max_row + 1):
        nm = v2.cell(r2, 2).value
        if nm in (None, ""):
            continue
        pool.setdefault((norm(nm), norm(v2.cell(r2, 4).value)), []).append(r2)
        n_v2 += 1
    cust = [(rr, nm, ttl) for rr, nm, ttl, pf in ledger(rv)
            if pf in ("Ampol Customer", "Z ENERGY (DIGITAL)", "EGI Integration")]
    if len(cust) != n_v2:
        log("B", "PCM_Data", "STOP: %d ledger rows against %d V2 rows"
            % (len(cust), n_v2))
        raise SystemExit(2)
    # a handful of vacancies share a name and a title, so pair them on the squad
    # and platform they sit in today before falling back to file order
    pair = {}
    for key, rows2 in pool.items():
        mine = [rr for rr, nm, ttl in cust if (norm(nm), norm(ttl)) == key]
        if len(mine) != len(rows2):
            log("B", "PCM_Data", "STOP: %r is %d ledger rows against %d V2 rows"
                % (key, len(mine), len(rows2)))
            raise SystemExit(2)
        if len(mine) == 1:
            pair[mine[0]] = rows2[0]
            continue
        cand = []
        for rr in mine:
            k = alias.get(str(rv.cell(rr, 11).value or "").strip(),
                          str(rv.cell(rr, 11).value or "").strip())
            j = str(rv.cell(rr, 10).value or "").strip()
            for r2 in rows2:
                s = (2 * (norm(k) == norm(v2.cell(r2, 12).value))
                     + (norm(j) == norm(v2.cell(r2, 11).value)))
                cand.append((-s, rr, r2))
        used_l, used_r = set(), set()
        for _, rr, r2 in sorted(cand):
            if rr in used_l or r2 in used_r:
                continue
            pair[rr] = r2
            used_l.add(rr)
            used_r.add(r2)
    moved, recoded, dated = 0, 0, 0
    at_before, at_after = {}, {}
    for rr, nm, ttl in cust:
        if rr not in pair:
            log("B", "REVIEW row %d" % rr, "STOP: %r / %r not in PCM_Data" % (nm, ttl))
            raise SystemExit(2)
        r2 = pair[rr]
        old = [str(rv.cell(rr, c).value or "").strip() for c in (9, 10, 11)]
        at_before[rr] = resolve_ar(
            ttl, resolve_at(old[2], old[1], alias, lines), lines)
        at_before[rr] = (at_before[rr] if at_before[rr] != "Squad"
                         else resolve_at(old[2], old[1], alias, lines))
        new = [str(v2.cell(r2, c).value or "").strip() for c in (10, 11, 12)]
        for c, o, nv in zip((9, 10, 11), old, new):
            if o != nv and nv:
                rv.cell(rr, c).value = nv
                moved += 1
                log("B3", "REVIEW!%s%d" % ("_IJK"[c - 8], rr),
                    "%s: %r -> %r" % (nm, o, nv))
        ap = resolve_at(new[2], new[1], alias, lines)
        ar = resolve_ar(ttl, ap, lines)
        at_after[rr] = ar if ar != "Squad" else ap
        cc = COUNTRY.get(str(v2.cell(r2, 14).value or "").strip())
        if cc and rv.cell(rr, 13).value != cc:
            log("B5", "REVIEW!M%d" % rr,
                "%s: %r -> %r" % (nm, rv.cell(rr, 13).value, cc))
            rv.cell(rr, 13).value = cc
            recoded += 1
        end = v2.cell(r2, 30).value
        if end is not None:
            rv.cell(rr, 51).value = end
            rv.cell(rr, 51).number_format = "dd/mm/yyyy"
            dated += 1
    copy_style(rv.cell(1, 50), rv.cell(1, 51))
    rv.cell(1, 51).value = "End date (PCM)"
    log("B5", "REVIEW!AY1:AY%d" % rv.max_row,
        "typed End Date column beside Commentry, %d dates (display only)" % dated)
    log("B5", "REVIEW!M", "%d country values recoded (AUD/NZD/WIPRO)" % recoded)
    log.note("B5", "Contingent Staff and both WIPRO sheets not loaded (rulings 8/scope)")
    changed = sorted(r for r in at_before if at_before[r] != at_after[r])
    for r in changed:
        log("B", "REVIEW row %d" % r, "%s: line %r -> %r"
            % (rv.cell(r, 2).value, at_before[r], at_after[r]))
    log.note("B2", "Annabel Phu reprices off the day rate 860 x 222 = 190,920 as the "
                   "model already does; the V2 source total understates her")
    log.note("B7", "the four 'Move to Z Customer' people are not added (ruling 10)")
    log.note("B8", "the 13/07 EGI portfolio moves are not adopted, its archetype grid "
                   "is out of scope, its Vacant Column helper is not carried (rulings 11/12)")

    # ------------------------------------------------------------- 2.2 rebuild
    log.head("B4  2.2 Customer gains the line and its FTE block is relaid")
    c22 = wb["2.2 Customer"]
    old_block = read_block(c22, 36, 131)
    lev = {rr: lv for k, nm, rr, lv in old_block if k == "role"}
    shift_rows(wb, "2.2 Customer", 26, 1)
    c22 = wb["2.2 Customer"]
    row_style(c22, 25, 26, 2, 18)
    c22["B26"] = PM
    c22["C26"] = '=""'
    c22["D26"] = '=""'
    c22["E26"] = "=Lists!$AH$8"
    c22["G26"] = ("=ROUND(SUMIFS('%s'!$O$2:$O$700,'%s'!$AJ$2:$AJ$700,$C$3,"
                  "'%s'!$AT$2:$AT$700,$B26),2)" % (REVIEW, REVIEW, REVIEW))
    c22["M26"] = "=$F26-$L26"
    c22["N26"] = "=Lists!$AJ$8"
    c22["O26"] = ("=SUMIFS('%s'!$AA$2:$AA$700,'%s'!$AJ$2:$AJ$700,$C$3,"
                  "'%s'!$AT$2:$AT$700,$B26)/1000000" % (REVIEW, REVIEW, REVIEW))
    c22["P26"] = "=ROUND($O26-$N26,6)"
    c22["R26"] = "=ROUND($Q26-$N26,6)"
    log("B4", "2.2 Customer!B26:R26",
        "Programme Management overhead line, archetype = its own allowance")

    at = {}
    for rr, nm, ttl, pf in ledger(rv):
        if pf not in ("Ampol Customer", "Z ENERGY (DIGITAL)", "EGI Integration"):
            continue
        at[rr] = at_after[rr]
    groups, empty = [], []
    for name in ORDER:
        rows = sorted(r for r in at if at[r] == name)
        if not rows:
            empty.append(name)
            continue
        groups.append((name, [(r, lev.get(r, "Filled")) for r in rows]))
    spare = sorted(set(at) - {r for _, rows in groups for r, _ in rows})
    if spare:
        log("B4", "2.2 Customer", "STOP: %d roles land outside the 14 lines: %r"
            % (len(spare), [(r, at[r]) for r in spare[:5]]))
        raise SystemExit(2)
    ranges, endrow = write_block(c22, 37, groups, (37, 38), 140)
    for name, row in SUMMARY.items():
        if name in empty:
            # his own shape for a line with no roles in this portfolio (see 2.6)
            for col in "FHIJKL":
                c22["%s%d" % (col, row)] = "=0"
            c22["M%d" % row] = "=$F%d-$L%d" % (row, row)
            c22["Q%d" % row] = "=0"
            log("B4", "2.2 Customer!row %d" % row,
                "%s keeps its line and reads 0 - every role has left it" % name)
            continue
        a, z = ranges[name]
        wire_summary(c22, row, a, z)
    log("B4", "2.2 Customer!B37:G%d" % endrow,
        "FTE block relaid, %d groups, %d roles, his lever states carried"
        % (len(groups), len(at)))
    set_dv(c22, sorted(r for a, z in ranges.values() for r in range(a, z + 1)))
    for col in "EFGHIJKLMOQ":
        c22["%s27" % col] = "=SUM({c}23:{c}26)".format(c=col)
    c22["N27"] = '=IF(COUNT(N23:N26)=0,"",SUM(N23:N26))'
    c22["P27"] = '=IF(COUNT(N23:N26)=4,ROUND($O27-$N27,6),"")'
    c22["R27"] = '=IF(COUNT(N23:N26)=4,ROUND($Q27-$N27,6),"")'
    picks = "E8:E14,E17:E17,E19:E20,E23:E26,E28"
    for col in "EFGHIJKLMOQ":
        c22["%s30" % col] = "=SUM(%s)" % picks.replace("E", col)
    c22["N30"] = ('=IF(COUNT(%s)=0,"",SUM(%s))'
                  % (picks.replace("E", "N"), picks.replace("E", "N")))
    c22["P30"] = '=IF(ISNUMBER($N30),ROUND($O30-$N30,6),"")'
    c22["R30"] = '=IF(ISNUMBER($N30),ROUND($Q30-$N30,6),"")'
    log("B4", "2.2 Customer!row 27 and row 30",
        "overhead total and portfolio total take in the new line")

    # ------------------------------------------------------------------- 3.2
    log.head("B4  3.2 carries the line, times applied 1")
    o32 = wb["3.2 Overhead & Leadership"]
    shift_rows(wb, "3.2 Overhead & Leadership", 11, 1)
    o32 = wb["3.2 Overhead & Leadership"]
    row_style(o32, 10, 11, 2, 13)
    o32["B11"] = "=Lists!$AF$8"
    o32["C11"] = "=Lists!$AI$8"
    o32["D11"] = "=Lists!$AG$8"
    o32["E11"] = 1
    o32["F11"] = "=$E11"
    o32["G11"] = ('=COUNTIFS(\'%s\'!$AR$2:$AR$700,$B11,\'%s\'!$B$2:$B$700,"<>")'
                  % (REVIEW, REVIEW))
    o32["H11"] = "=ROUND($G11-$F11,6)"
    o32["I11"] = "=ROUND($E11*$D11,6)"
    o32["J11"] = ('=SUMIFS(\'%s\'!$AA$2:$AA$700,\'%s\'!$AR$2:$AR$700,$B11,'
                  '\'%s\'!$B$2:$B$700,"<>")/1000000' % (REVIEW, REVIEW, REVIEW))
    o32["K11"] = "=ROUND($J11-$I11,6)"
    o32["L11"] = "All 3 in Customer"
    o32["M11"] = '=TEXT(1,"0%")&" across "&TEXT($E11,"0")&" "&LOWER($C11)'
    log("B4", "3.2 Overhead & Leadership!B11:M11",
        "Programme Management line, allowance = actual so it adds no variance")
    for col in "FGHIJK":
        o32["%s12" % col] = "=SUM({c}5:{c}11)".format(c=col)
    o32["F13"] = '=SUMIF(Lists!$AM$2:$AM$8,"Yes",$F5:$F11)'
    o32["I13"] = '=ROUND(SUMIF(Lists!$AM$2:$AM$8,"Yes",$I5:$I11),6)'
    for co in ("G13", "J13"):
        o32[co] = str(o32[co].value).replace("$B$5:$B$10", "$B$5:$B$11") \
                                    .replace("$B5:$B10", "$B5:$B11")
    o32["I15"] = ("=ROUND($I12-SUMPRODUCT(Lists!$AG$2:$AG$8,Lists!$AH$2:$AH$8),6)")
    log("B4", "3.2 Overhead & Leadership!rows 12 to 15",
        "totals, the portfolio split and the control widened to seven lines")

    # ------------------------------------------------------------------- 3.3
    log.head("B4  3.3 carries every line on every lever modelling tab")
    r33 = wb["3.3 Squad Actuals to Archetype"]
    tag = "='2.2 Customer'!$B$26"
    if not any(r33.cell(r, 4).value == tag for r in range(20, 40)):
        shift_rows(wb, "3.3 Squad Actuals to Archetype", 33, 1)
        r33 = wb["3.3 Squad Actuals to Archetype"]
        row_style(r33, 32, 33, 2, 15)
        r33.cell(33, 2).value = "Customer"
        r33.cell(33, 3).value = "Overhead"
        for col, s in zip(range(4, 16),
                          ["B", "C", "D", "E", "F", "H", "I", "M", "N", "O", "P", "Q"]):
            r33.cell(33, col).value = "='2.2 Customer'!$%s$26" % s
        for col in range(7, 16):
            f = r33.cell(34, col).value
            if isinstance(f, str):
                r33.cell(34, col).value = f.replace("20:", "20:").replace("32)", "33)") \
                                           .replace("32,", "33,")
        log("B4", "3.3 Squad Actuals to Archetype rows 33 and 34",
            "Customer / Overhead / Programme Management line added and totalled")

    # ------------------------------------------------------------------- 4.0
    log.head("B4  the gate suite follows the new line")
    q = wb["4.0 Data QA"]
    q["D21"] = ("=ROUND(SUMPRODUCT('3.2 Overhead & Leadership'!$E$5:$E$11,"
                "'3.2 Overhead & Leadership'!$D$5:$D$11),6)")
    q["D22"] = '=SUMIF(Lists!$AM$2:$AM$8,"Yes",Lists!$AJ$2:$AJ$8)'
    log("B4", "4.0 Data QA!D21 D22", "3.2 and Lists checks widened to seven lines")
    c40 = "='2.2 Customer'!$N$30+N('1.2 Customer'!$H$41)"
    d40 = "='1.2 Customer'!$F$9+N('2.2 Customer'!$N$26)"
    if q["C40"].value != c40 or q["D40"].value != d40:
        q["C40"], q["D40"] = c40, d40
        log("B4", "4.0 Data QA!C40 D40",
            "identity restated: 2.2 carries a Customer only overhead line 1.2 has no row for")

    # ------------------------------------------------------------------- B6
    log.head("B6  Digital Support NZ prices at the archetype")
    c12 = wb["1.2 Customer"]
    note = "CPI actuals; pull through (0.217)"
    if str(c12["M41"].value or "").strip() != note:
        log("B6", "1.2 Customer!M41",
            "%r -> %r" % (c12["M41"].value, note))
        c12["M41"] = note

    save(wb, dst)
    log.tail()
    print("wrote", dst)


if __name__ == "__main__":
    main(*sys.argv[1:])
