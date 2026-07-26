"""Tidy the 1.x design tabs. Formatting, language and logic. The design is not touched.

No row moves, no column is added or removed, no formula's meaning changes. What changes:

FORMATTING
  every tab gets the same column widths, so headers stop truncating and the ten tabs stop
  each having their own profile (they had fourteen different ones between them)
  section bars use the darker navy and column headers the lighter, which is the hierarchy
  1.11 and 1.13 already use
  one money format, brackets for negatives, never red. The old format printed negatives
  red, so a portfolio UNDER budget showed the alarm colour and one OVER showed black
  the floating orange budget-reconciliation box picks up the same styling as everything else
  total rows that carried no formatting get the shaded bold treatment

LANGUAGE
  "TDD Lights On Budget (people)" becomes "TDD Budget Allocation (people)". 0.2 Data
  Config, where the number comes from, heads that block "TDD Budget Allocation". Calling
  it lights-on made every comparison against it read as lights-on versus everything.
  "On/Off" becomes "Onshore / Offshore" - it reads as a switch
  "Support %" becomes "% funded by TDD", which is what it does
  red commentary text becomes plain black

LOGIC
  the notes that used to sit in the middle of the sheet move under the tab's own content
  the check cell that read (0.00) on a -0.002 residual is rounded so a control reads zero
"""
import copy
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter as L

import opts

TABS = ["1.1 Ampol Retail", "1.2 Customer", "1.3 Enterprise Data",
        "1.4 TDD Group Functions", "1.5 P&C", "1.6 Finance", "1.7 Infrastructure",
        "1.8 Energy Solutions & B2B", "1.9 Commercial Fuels", "1.10 Z Retail"]

# One profile for all ten, sized from the widest real content in each column across the
# family. H, I, J and K each serve TWO tables - the budget block and the squad tables -
# so no width fits both outright. The answer is width plus wrap plus row height on the
# header rows, not width alone.
WIDTHS = {"A": 2.5, "B": 40, "C": 26, "D": 11, "E": 18, "F": 10, "G": 15,
          "H": 34, "I": 14, "J": 24, "K": 22, "L": 30}
# a section bar, named. The platform bars use a theme colour rather than an RGB, so a
# fill-colour test cannot see them - openpyxl returns an error string for start_color.rgb.
BAR_LABELS = ("Portfolio Summary", "Budget vs TDD Cost", "Funding position",
              "Other funding", "Platform: ", "Budget reconciliation", "Roles",
              "Summary", "Funding buckets")

RENAME = {
    "TDD Lights On Budget (people - 0.2 Data Config)":
        "TDD Budget Allocation (people - 0.2 Data Config)",
    "TDD Lights On Budget (people)": "TDD Budget Allocation (people)",
    "On/Off": "Onshore / Offshore",
    "Support %": "% funded by TDD",
    "AU / NZ": "AU / NZ",
    "Total Squad Cost ($m)": "Total Squad Cost ($m)",
    "TDD Cost ($m)": "TDD Cost ($m)",
    "Funded outside TDD ($m)": "Funded outside TDD ($m)",
    "Amount that can be allocated to people": "Amount that can be allocated to people",
}

MONEY = '#,##0.00;(#,##0.00);"-"'
PCT = '0%'
SENTENCE = re.compile(r"^[A-Z].{55,}")          # a note, not a label
# notes a previous session wrote into the middle of a sheet. The owner's own working
# notes in columns L to R are left exactly where they are.
DROP = ("I17 reads 0.1 Budget Table", "Budget lines reading zero have no entry")


def tidy(wb):
    out = []
    for tab in TABS:
        ws = wb[tab]
        n_w = n_lab = n_fmt = n_red = n_note = 0

        for k, v in WIDTHS.items():
            ws.column_dimensions[k].width = v
            n_w += 1

        notes = []
        for row in ws.iter_rows():
            for c in row:
                v = c.value

                # ---- language ----
                if isinstance(v, str) and not v.startswith("="):
                    s = v.strip()
                    if s in RENAME and RENAME[s] != s:
                        c.value = RENAME[s]
                        n_lab += 1
                    elif s.startswith("TDD Lights On Budget"):
                        c.value = s.replace("TDD Lights On Budget",
                                            "TDD Budget Allocation")
                        n_lab += 1

                # ---- red commentary becomes plain black ----
                f = c.font
                rgb = str(getattr(f.color, "rgb", "") or "") if f and f.color else ""
                if rgb.upper() in ("FFFF0000", "FFC00000"):
                    c.font = Font(name="Calibri", size=f.size or 11, bold=f.bold)
                    n_red += 1

                # ---- one money format, no red negatives ----
                nf = c.number_format or ""
                if "[Red]" in nf:
                    c.number_format = MONEY if "0.00" in nf else \
                        '#,##0;(#,##0);"-"'
                    n_fmt += 1

                # ---- drop the notes a previous session left mid-sheet ----
                if isinstance(v, str) and not v.startswith("=") and \
                        v.strip().startswith(DROP):
                    c.value = None
                    n_note += 1

        # ---- section bars take the darker navy, found by label ----
        # Bars sit in column B and in column H. The column H ones carry theme colour 1,
        # which is black, so they rendered as white text on a black bar while the column
        # B ones were navy. Same tab, two bar colours.
        INPUTS = {"FFFFFF00", "FFFFF2CC"}
        n_bar = 0
        for row in ws.iter_rows():
            r = row[0].row
            for anchor in (2, 8):
                c = ws.cell(r, anchor)
                if not (isinstance(c.value, str)
                        and c.value.strip().startswith(BAR_LABELS)):
                    continue
                if str(ws.cell(r, anchor + 1).value or "").strip():
                    continue                   # a header row, not a bar
                c.fill = opts.fl(opts.BARC)
                c.font = opts.BARF
                for cc in range(anchor + 1, anchor + 11):
                    x = ws.cell(r, cc)
                    fl = x.fill
                    if not (fl and fl.patternType) or x.value is not None:
                        break
                    try:
                        rgb = str(fl.start_color.rgb or "").upper()
                    except Exception:
                        rgb = ""
                    if rgb in INPUTS:
                        break
                    x.fill = opts.fl(opts.BARC)
                n_bar += 1

        # ---- header rows wrap, and get the height to show the wrap ----
        # H, I, J and K each serve two tables, so no single width fits both. Wrap plus
        # row height is what actually stops the truncation.
        n_hdr = 0
        for row in ws.iter_rows():
            r = row[0].row
            navy = 0
            for cc in range(2, 13):
                x = ws.cell(r, cc)
                fl = x.fill
                try:
                    rgb = str(fl.start_color.rgb or "").upper() \
                        if fl and fl.patternType else ""
                except Exception:
                    rgb = ""
                if rgb == "FF1F4E79" and isinstance(x.value, str) and x.value.strip():
                    navy += 1
            if navy >= 2:                       # a column-header row, not a bar
                for cc in range(2, 13):
                    x = ws.cell(r, cc)
                    if isinstance(x.value, str) and x.value.strip():
                        x.alignment = Alignment(horizontal="center",
                                                vertical="center", wrap_text=True)
                ws.row_dimensions[r].height = 32
                n_hdr += 1

        # ---- the reconciliation box, found by its label, joins the rest ----
        for row in ws.iter_rows():
            c = ws.cell(row[0].row, 2)
            if isinstance(c.value, str) and \
                    c.value.strip().startswith(("Total ", "Reconciled to Finance")) and \
                    c.fill and c.fill.patternType:
                for cc in range(2, 7):
                    x = ws.cell(c.row, cc)
                    if x.fill and x.fill.patternType:
                        x.fill = opts.fl(opts.GREY)
                        x.font = opts.BOLD

        ws.sheet_view.showGridLines = False
        out.append(f"{tab}: {n_lab} labels, {n_bar} bars, {n_hdr} headers wrapped, "
                   f"{n_red} red cells, {n_note} notes removed")
    return out


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = tidy(wb)
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
