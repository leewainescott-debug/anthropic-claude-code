"""Three agreed role moves, made through the visible override table on Lists.

The owner's raw columns on REVIEW are never edited. The grouping columns read an override
table first and fall back to the raw data, so every reassignment is one visible row a
reader can check, and the ledger still totals $115,113,262.27.

  r136 Jasper Na       squad Energy -> Ampol Web
        Associate Engineer - BE reporting to Jin Zhong, Lead Engineer - BE, who sits in
        Ampol Web; the only other person under Jin Zhong is also Ampol Web. "Energy" was
        a one-person squad that appears on no design tab.

Two further moves were proposed and withdrawn on the owner's instruction: Viren Khatri
stays in Ampol Retail, which is his home, and Vikram Chhahira stays in the EGI P&C squad,
which is a squad in its own right. Neither is in the table.

The table was three rows wide and hardcoded at $AN$2:$AN$4. It is now ten.

It is keyed on the person - "Name | Position Title", the two columns a reader can see in
REVIEW B and C - and not on a REVIEW row number. Row numbers were the worst trap in the
file: insert a row anywhere above 283 and all three agreed moves apply to three different
people, silently, with every total still balancing. Name alone is not enough (143 rows are
called "Vacant"), so the title is part of the key, and run() refuses to build if a key
matches anything other than exactly one ledger row.

It carried a fourth column, an overhead-line override read by REVIEW!AR. It was empty in
every build, so the branch in AR could never fire and was costing 250 characters and two
IF levels in a formula on 531 rows. Both are gone - the column and the branch - rather
than leave a control on Lists that no formula reads.
"""
import re

import openpyxl
from openpyxl.styles import Border, PatternFill
from openpyxl.utils import get_column_letter as L

import opts

REVIEW = "REVIEW - Complete Role Mapping"
LAST = None                             # measured from the ledger in run()
N = 11                                  # override window runs Lists!$AN$2:$AN$11

# The three moves the owner instructed (D7). They lived as typed rows in the Lists
# override table; the review workbook's branch predates that table, so they are
# reinstated here or the roles silently snap back to their raw homes.
NEW = [(283, "COE SA&D", "Group Data", None),
       (313, "COE SA&D", "Group Data", None),
       (528, None, "SAP ERP", None)]

# Folds to take back out of Lists!W:X. The table is meant to fix typing - AmPos to AmPOS,
# Manuacturing to Manufacturing - and three rows in it were doing something else:
#
#   Customer, AI      -> Customer AI              a rename, not a typo. The owner's column
#                                                 K has the comma and so does his own list.
#   Z Energy Martech  -> Z Loyalty & Martech      two real squads merged into one
#   AU CRM & Martech  -> Ampol Loyalty & Martech  two real squads merged into one
#
# Merging them hid two of Customer's twelve squads completely: the owner's list has
# AU CRM & Martech at 2.0 FTE and Z Energy Martech at 2.0, and neither appeared anywhere in
# the workbook. Undoing the merge also puts Ampol Loyalty & Martech back to 6.8 and
# Z Loyalty & Martech back to 12.6, which is what his column K says.
UNFOLD = ["Customer, AI", "Z Energy Martech", "AU CRM & Martech"]

# The move of Jasper Na off the one-person Energy squad was mine, not the owner's, and his
# own reconciliation lists Energy at 1.0 FTE. It comes out. The three moves he did instruct -
# r283 and r313 to the Data COE, r528 to AU Finance - stay.
DROP_OVERRIDE = [136]

# A filled role priced at zero. REVIEW row 491, Nidhi Aggarwal, Snr Engineer - Boomi, NZ,
# FTE 1.0, status Filled. Column T carries her local base of 150,000 but U is empty, and the
# cost formula prices off U, so she counted in the 525 headcount and contributed nothing to
# the total. She is the only role in the ledger with T populated and U empty, and no check in
# the workbook could see it: all fifty-six reconcile TO the ledger, and a zero inside the
# ledger reconciles perfectly.
#
# The figure is derived from her own cohort, not invented. Every one of the thirteen other NZ
# roles in TDD Group Functions prices the same way: U = T x 0.92, then STI 0.15, pensions
# 0.05, CPI 0.03, no payroll component. 150,000 x 0.92 x 1.23 = 169,740. It goes in the
# cost-override column the formula already honours, so her raw columns stay untouched.
COST = {491: (169740.0,
              "Priced from her local base of 150,000 at the 0.92 rate her cohort uses, "
              "then STI 0.15, pensions 0.05, CPI 0.03 - the pattern all thirteen other NZ "
              "roles in this portfolio use. Column U was blank, so the formula read zero.")}
# build scaffolding and stray input colour left in the ledger
CLEAR_NOTES = [(191, 29), (491, 29)]

# ---- the live control sitting on the ledger ----
# REVIEW ships with an AutoFilter carrying active criteria - column F held to "Strategy,
# Architecture & Data" and column G to "Architecture" - which hides 519 of the 531 role
# rows, and a sortState beside it. Both stop at row 528 and column AH. The ledger runs to
# row 534 and out to column AX, so one sort through that control reorders A:AH for rows
# 2-528 and leaves the ten cost overrides in AU, and every row past 528, exactly where they
# were. The names, the costs and the groupings come apart, every total still adds up, and
# all fifty-six controls still read 0. It is the worst shape a defect can have in this file.
#
# A filter on the ledger is worth having - it is the one tab a reader filters - so REVIEW
# keeps one, criteria-free and spanning the whole ledger, and every other tab loses its
# filter outright. The owner's two source tabs are never touched.
FILTER_KEEP = REVIEW
FILTER_LASTCOL = 50                     # AX, the last column the ledger uses
FILTER_SOURCE = ("0.1 Budget Table (Fin)", "0.4 Presentation Pack")

PORTFOLIOS = ["Ampol Retail", "Customer", "Enterprise Data", "TDD Group Functions",
              "P&C", "Finance", "Infrastructure", "Energy Solutions & B2B",
              "Commercial Fuels", "Z Retail"]

# The keyword chain that classifies a position title into an overhead line. This is the
# whole of REVIEW!AR now - the override wrapper that used to sit in front of it is gone.
#
# "technology manger" is not a typo in this file. It is a typo in the owner's ledger:
# at least one Position Title is spelt that way, and without the extra SEARCH that person
# classifies as "Squad" and lands in a delivery block instead of on the Technology Manager
# overhead line. Correcting his raw column is not ours to do, so the chain catches it.
# Delete this branch and a real person moves.
KEYWORD = ('IF(ISNUMBER(SEARCH("head of ",$C{r})),"Head of Technology",'
           'IF(ISNUMBER(SEARCH("TDD BP",$C{r})),"Business Partner",'
           'IF(OR(ISNUMBER(SEARCH("domain architect",$C{r})),'
           'ISNUMBER(SEARCH("enterprise architect",$C{r}))),"Domain Architect",'
           'IF(ISNUMBER(SEARCH("delivery man",$C{r})),"Delivery Manager",'
           'IF(OR(ISNUMBER(SEARCH("technology manager",$C{r})),'
           'ISNUMBER(SEARCH("technology manger",$C{r})),'
           'ISNUMBER(SEARCH("tech manager",$C{r}))),"Technology Manager","Squad"))))')


# Where a design tab and the ledger disagree on a squad name, the design tab is renamed -
# the ledger is the source of truth. This has to happen before the working tabs are built:
# they decide which section a squad belongs in by looking its name up on the design tab, and
# renaming afterwards left "Customer, AI" filed under "no archetype" with a 0.80 archetype
# sitting beside it.
SQUAD_RENAME = {"1.2 Customer": {"Customer AI": "Customer, AI"}}


def rename_squads(wb):
    out = []
    for tab, pairs in SQUAD_RENAME.items():
        ws = wb[tab]
        for r in range(1, min(ws.max_row, 95) + 1):
            v = ws.cell(r, 2).value
            if isinstance(v, str) and v.strip() in pairs:
                new = pairs[v.strip()]
                ws.cell(r, 2).value = new
                out.append(f"{tab}!B{r} {v.strip()!r} -> {new!r}, to match the ledger")
    return out


def sweep_filters(wb, last):
    """Take every AutoFilter and sortState out of the file, and unhide what they hid.

    REVIEW keeps a filter because a reader wants one there, but a criteria-free one over
    the whole ledger - A1:AX{last} - so it can never sort a subset of the columns or a
    subset of the rows. Everywhere else the filter goes: none of those tabs is a list a
    reader filters, and each one is the same hazard waiting for a click.

    Rows hidden by a filter are unhidden here as well. Removing the control without
    unhiding leaves the rows invisible with nothing on the tab to bring them back.
    """
    out = []
    for ws in wb.worksheets:
        ref = ws.auto_filter.ref
        crit = len(ws.auto_filter.filterColumn or [])
        sort = ws.auto_filter.sortState is not None
        hidden = [r for r, d in ws.row_dimensions.items() if d.hidden]
        if not (ref or crit or sort or (hidden and ws.title == FILTER_KEEP)):
            continue
        if ws.title in FILTER_SOURCE:
            out.append(f"{ws.title}: autoFilter {ref} left where it is - the owner's own "
                       f"tab, and nothing downstream reads it")
            continue
        ws.auto_filter.filterColumn = []
        ws.auto_filter.sortState = None
        if ws.title == FILTER_KEEP:
            ws.auto_filter.ref = f"A1:{L(FILTER_LASTCOL)}{last}"
            what = f"filter reset to A1:{L(FILTER_LASTCOL)}{last}, no criteria, no sort"
        else:
            ws.auto_filter.ref = None
            what = "filter and sort removed"
        for r in hidden:
            ws.row_dimensions[r].hidden = False
        out.append(f"{ws.title}: {what} (was {ref}, {crit} criteria, "
                   f"{'a' if sort else 'no'} sortState, {len(hidden)} rows hidden)")
    return out


def key_expr(r):
    """The cell expression that identifies a person, for matching against the table.

    Name alone is not a key: 143 of the 531 rows are called "Vacant". Name and position
    title together are unique for every row the table names, and build_key() below proves
    that at build time rather than trusting it.
    """
    return f'TRIM($B{r})&" | "&TRIM($C{r})'


def build_key(name, title):
    """The same key, as text, for the table cell."""
    return f"{str(name or '').strip()} | {str(title or '').strip()}"


def ovr(col, r):
    """One override lookup for one column of the table, evaluated once.

    The old shape computed IFERROR(INDEX(...)) to test it and then computed the identical
    expression again to use it - the same lookup twice in one cell, in a formula already
    seven IFs deep. The COUNTIFS is the test: it is zero both when nobody in the table is
    this person and when the table names them but leaves this column empty, which are the
    two cases that have to fall through to the raw data. Where it is non-zero there is a
    value to take, and the INDEX runs once.

    Note the key: the person, not their row number. MATCH(ROW(), ...) against a table of
    typed row numbers is wrong the moment anybody inserts a row in the ledger - the three
    agreed moves would land on whoever slid into rows 283, 313 and 528.
    """
    k = key_expr(r)
    test = f'COUNTIFS(Lists!$AN$2:$AN${N},{k},Lists!${col}$2:${col}${N},"<>")'
    take = f'INDEX(Lists!${col}$2:${col}${N},MATCH({k},Lists!$AN$2:$AN${N},0))'
    return test, take


def formulas(r):
    t, x = ovr("AO", r)
    aj = (f'=IF(TRIM($B{r})="","",IF({t},{x},'
          f'IFERROR(INDEX(Lists!$U:$U,MATCH(TRIM($I{r}),Lists!$T:$T,0)),TRIM($I{r}))))')
    # AR carries no override branch. It used to wrap the keyword chain in a lookup against
    # Lists!AQ, which is empty and always has been, so that branch could never fire - it
    # was 250 characters and two extra IF levels of dead weight in front of the only thing
    # the cell actually does, which is read the position title. The override column went
    # with it rather than leave a control on Lists that nothing reads.
    ar = f'=IF(TRIM($B{r})="","",{KEYWORD.format(r=r)})'
    t, x = ovr("AP", r)
    at = (f'=IF(TRIM($B{r})="","",IF({t},{x},'
          f'IF(OR(LEFT($AJ{r},3)="COE",$AJ{r}="EGI"),$AP{r},'
          f'IF($AR{r}<>"Squad",$AR{r},$AP{r}))))')
    return {36: aj, 44: ar, 46: at}


# REVIEW columns that nothing in the workbook reads. AL (MAUNZ), AM (MRank) and AN (MKey)
# are a previous session's scaffolding: AM is read only by AN, and AN by nothing at all.
# AS looked up Lists!Z:AA, which is empty, so it always returned AP unchanged, and nothing
# read AS either. Every formula in the built workbook was grepped for each of them, on the
# REVIEW sheet and from every other sheet, before this list was written.
# (Lists!AN, the override table, is a different column on a different sheet - untouched.)
DEAD = {38: "AL", 39: "AM", 40: "AN", 45: "AS"}


def run(src, dst):
    global LAST
    wb = openpyxl.load_workbook(src)
    LAST = opts.ledger_last(wb)
    l = wb["Lists"]
    out = []

    # The columns the new table needs must be empty, or already hold this table from a
    # previous run - the build has to be re-runnable against its own output, or the second
    # run stops on the header the first one wrote.
    OURS = {45: "Portfolios (10)"}
    for c, mine in OURS.items():
        if l.cell(1, c).value in (None, mine):
            continue
        raise SystemExit(f"Lists!{L(c)}1 holds {l.cell(1, c).value!r}, not "
                         f"{mine!r} - pick another column")

    # the table already carries the earlier agreed moves (r283 and r313 to the Data COE,
    # r528 to AU Finance). Append, never overwrite - writing over row 2 silently pulled
    # two roles back out of COE SA&D and moved a third out of SAP ERP.
    # ---- take the three merges back out of the fold table ----
    gone = []
    for r in range(2, 20):
        if str(l.cell(r, 23).value or "").strip() in UNFOLD:
            gone.append(f"{l.cell(r, 23).value} -> {l.cell(r, 24).value}")
            l.cell(r, 23).value = None
            l.cell(r, 24).value = None
    if gone:
        out.append("fold removed: " + "; ".join(gone))

    # ---- the table, re-keyed on the person ----
    # It used to hold typed REVIEW row numbers - 283, 313, 528 - and the grouping columns
    # found their row with MATCH(ROW(), ...). Insert one row anywhere above 283 in the
    # ledger and all three agreed moves apply to three different people, quietly, with
    # every total still adding up. The key is now the person: "Name | Position Title",
    # which is what a reader checking the row would look for anyway.
    R = wb[REVIEW]
    key_of, rows_for = {}, {}
    for r in range(2, LAST + 1):
        if not str(R.cell(r, 2).value or "").strip():
            continue
        k = build_key(R.cell(r, 2).value, R.cell(r, 3).value)
        key_of[r] = k
        rows_for.setdefault(k, []).append(r)

    def as_key(v):
        """Accept either shape in the table on Lists: a typed row number from an older
        build, or a name key this build already wrote."""
        if isinstance(v, (int, float)):
            return key_of.get(int(v)), int(v)
        s = str(v or "").strip()
        if not s:
            return None, None
        return s, (rows_for.get(s) or [None])[0]

    keep, seen = [], set()
    for r in range(2, 40):
        k, row = as_key(l.cell(r, 40).value)
        if k is None:
            continue
        if row in DROP_OVERRIDE:
            out.append(f"override removed: {k} -> {l.cell(r, 42).value!r}")
            continue
        keep.append((k, l.cell(r, 41).value, l.cell(r, 42).value))
        seen.add(k)
    for row, pf, sq, _oh in NEW:
        k = key_of.get(row)
        if k is None:
            raise SystemExit(f"REVIEW row {row} carries no name - the override table "
                             f"cannot be keyed on it")
        if k not in seen:
            keep.append((k, pf, sq))
            seen.add(k)
    moves = keep
    if len(moves) + 1 > N:
        raise SystemExit(f"{len(moves)} moves will not fit in {N - 1} slots")

    # a key that names two people is a key that moves two people. The whole point of
    # dropping the row numbers is that the new key is unambiguous, so it is checked, not
    # assumed - "Vacant" on its own matches 143 rows, which is why the title is in the key.
    for k, _pf, _sq in moves:
        hit = rows_for.get(k, [])
        if len(hit) != 1:
            raise SystemExit(f"override key {k!r} matches {len(hit)} ledger rows "
                             f"({hit}) - it must match exactly one")
    out.append("keyed on the person: " + "; ".join(
        f"{k} (row {rows_for[k][0]} today)" for k, _p, _s in moves))

    HEAD = {40: "Person (Name | Position Title)", 41: "Portfolio override",
            42: "Squad override"}
    for c, h in HEAD.items():
        x = l.cell(1, c)
        x.value = h
        x.font, x.fill, x.alignment = opts.HDRF, opts.fl(opts.NAVY), opts.CEN
        l.column_dimensions[L(c)].width = 34 if c == 40 else 24
    l.cell(1, 43).value = None                       # the dead overhead-line override
    for i, (k, pf, sq) in enumerate(moves):
        r = 2 + i
        for c, v in ((40, k), (41, pf), (42, sq), (43, None)):
            x = l.cell(r, c)
            x.value = v
            x.font, x.border = opts.BODY, opts.BOX
            x.fill = opts.fl(opts.YEL) if c != 43 else PatternFill()
            x.alignment = opts.LFT
    for r in range(2 + len(moves), N + 1):              # spare, declared and empty
        for c in (40, 41, 42):
            x = l.cell(r, c)
            x.value = None
            x.font, x.border, x.fill = opts.BODY, opts.BOX, opts.fl(opts.YEL)
        x = l.cell(r, 43)
        x.value, x.border, x.fill = None, Border(), PatternFill()
    l.cell(N + 2, 40).value = (
        "Agreed moves, keyed on the person - Name | Position Title, exactly as they read "
        "in REVIEW columns B and C - so the moves follow the person if rows are inserted "
        "or deleted. REVIEW's own columns are untouched; these override the grouping only.")
    l.cell(N + 2, 40).font = opts.BODY
    out.append(f"Lists override table: {len(moves)} moves, {N - 1} slots, name-keyed")

    # a clean list of the ten portfolios, so the allowance stops counting a block of
    # cells on 3.1 whose row numbers move every time the tab is rebuilt
    l.cell(1, 45).value = "Portfolios (10)"
    l.cell(1, 45).font = opts.HDRF
    l.cell(1, 45).fill, l.cell(1, 45).alignment = opts.fl(opts.NAVY), opts.CEN
    l.column_dimensions["AS"].width = 26
    for i, p in enumerate(PORTFOLIOS):
        x = l.cell(2 + i, 45)
        x.value, x.font, x.border = p, opts.BODY, opts.BOX
        x.alignment = opts.LFT
    out.append("Lists!AS2:AS11: the ten portfolios, named")

    # ---- repoint every reader of the old three-row window ----
    # every column of the table, not only the key column. Widening AN alone left
    # INDEX(Lists!$AO$2:$AO$4, MATCH(ROW(), Lists!$AN$2:$AN$11, 0)): a match at position
    # four indexes past the end of a three-row range, so an override in one of the new
    # slots silently returned nothing.
    old = re.compile(r"\$(AN|AO|AP|AQ)\$2:\$(AN|AO|AP|AQ)\$4\b")
    hits = 0
    for s in wb.sheetnames:
        for row in wb[s].iter_rows():
            for c in row:
                if isinstance(c.value, str) and old.search(c.value):
                    c.value = old.sub(lambda m: f"${m.group(1)}$2:${m.group(2)}${N}",
                                      c.value)
                    hits += 1
    out.append(f"widened override window in {hits} formulas")

    # ---- rebuild the three grouping columns ----
    R = wb[REVIEW]
    n = 0
    # every populated row. The spacer rows the Customer reload left behind (191-192) are
    # cleared upstream by merge_review, and a formula written onto an empty row is exactly
    # the stray the audit flagged - so blank-name rows are skipped, not refreshed. The old
    # stale-range risk this loop once guarded against is gone: the spacers arrive empty,
    # and an empty cell cannot carry an old mismatched range.
    for r in range(2, LAST + 1):
        if not str(R.cell(r, 2).value or "").strip():
            continue
        for col, f in formulas(r).items():
            R.cell(r, col).value = f
        n += 1
    out.append(f"REVIEW AJ / AR / AT rebuilt on {n} rows")

    # ---- a filled role priced at zero, and the scaffolding around it ----
    for row, (cost, why) in COST.items():
        R.cell(row, 47).value = cost                     # AU, the cost override
        R.cell(row, 47).number_format = '#,##0'
        R.cell(row, 48).value = why                      # AV, its provenance
        R.cell(row, 48).font = opts.BODY
        out.append(f"REVIEW row {row}: cost override {cost:,.0f} - {why[:48]}...")
    for row, col in CLEAR_NOTES:
        R.cell(row, col).value = None
    # ---- the columns nothing reads ----
    # Cleared, not left to rot: four columns of formulas on 531 rows each, every one of
    # them a thing a reader has to work out is dead before they can ignore it. See DEAD.
    dead = 0
    for c in DEAD:
        R.cell(1, c).value = None
        for r in range(2, LAST + 1):
            if R.cell(r, c).value is not None:
                R.cell(r, c).value = None
                dead += 1
    out.append("REVIEW: build notes removed; dead columns "
               + ", ".join(DEAD.values()) + f" cleared ({dead} cells)")
    out += rename_squads(wb)

    # ---- every fixed window over the ledger follows the measured extent ----
    # The design tabs the owner edits carry SUMIFS/COUNTIFS over REVIEW with the end row
    # typed in - $AJ$2:$AJ$528, $AA$2:$AA$530 and friends. A ledger that grows past the
    # typed end silently drops the new rows out of every one of those formulas, which is
    # the worst failure mode this workbook has: wrong and green. Any range anchored at
    # row 2 whose typed end sits in 500..999 is a ledger window and follows LAST.
    win = re.compile(r"(\$[A-Z]{1,2})\$2:(\$[A-Z]{1,2})\$(5\d\d|[6-9]\d\d)\b")
    wided = 0
    for s_ in wb.sheetnames:
        for row in wb[s_].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("=") and win.search(c.value):
                    new = win.sub(lambda m: f"{m.group(1)}$2:{m.group(2)}${LAST}", c.value)
                    if new != c.value:
                        c.value = new
                        wided += 1
    out.append(f"ledger windows widened to row {LAST} in {wided} formulas")

    # ---- the allowance stops counting rows on a summary tab ----
    fixed = 0
    pat = re.compile(r"COUNTA\('3\.1 Group Summary'!\$B\$\d+:\$B\$\d+\)")
    for s in wb.sheetnames:
        for row in wb[s].iter_rows():
            for c in row:
                if isinstance(c.value, str) and pat.search(c.value):
                    c.value = pat.sub("COUNTA(Lists!$AS$2:$AS$11)", c.value)
                    fixed += 1
    out.append(f"portfolio count keyed off the named list in {fixed} formulas")

    # ---- the filters, last, so REVIEW's filter follows the measured extent ----
    out += sweep_filters(wb, LAST)
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
