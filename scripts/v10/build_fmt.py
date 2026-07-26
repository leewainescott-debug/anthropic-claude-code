"""Stage 17: make the 3.x tabs look like the 2.x tabs.

The 2.x family is uniform - one header style, one set of number formats. The 3.x family
was not: 3.2, 3.3 and the Exec Summary had no header fill at all, 3.4 had four different
number formats in its body, 4.0 was half filled and half not. Same styling rules now
apply everywhere so a reader moving between tabs sees one workbook.
"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HDR_FILL = PatternFill("solid", start_color="FF1F4E79", end_color="FF1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFFFF")
CNT = '#,##0;[Red](#,##0);"-"'
MON = '#,##0.00;[Red](#,##0.00);"-"'
PCT = '0.0%'

TABS = {
    "3.1 Group Summary":   (5, "BCDEFGHIJKL",  {"J": CNT, "K": CNT}),
    "3.2 Total Cost":      (5, "BCDEFGHIJKLMN", {"C": CNT, "D": CNT, "E": CNT,
                                                 "G": CNT, "I": CNT, "L": CNT}),
    "3.3 FTE View":        (5, "BCDEFGHIJKL",  {"F": CNT, "G": CNT, "H": CNT,
                                                "I": CNT, "J": CNT, "K": CNT, "L": PCT}),
    "3.4 COE Summary":     (5, "BCDEFGHIJKL",  {"C": CNT, "D": CNT, "E": CNT}),
    "3.5 Source Reconciliation": (5, "BCDEFGHIJK", {c: CNT for c in "CDEFGHIJK"}),
    "4.0 Data QA":         (5, "BCDE",         {}),
}


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = []
    for tab, (hdr, cols, overrides) in TABS.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for c in cols:
            cell = ws[f"{c}{hdr}"]
            if cell.value is None:
                continue
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="bottom")
            cell.number_format = "General"
        for r in range(hdr + 1, ws.max_row + 1):
            for c in cols:
                cell = ws[f"{c}{r}"]
                if cell.value is None:
                    continue
                if c in overrides:
                    cell.number_format = overrides[c]
                elif c not in ("B",):
                    cell.number_format = MON
        out.append(f"{tab}: header and number formats aligned to the 2.x style")
    # the Exec Summary is a narrative page, so it gets the section headings only
    ex = wb["Exec Summary"]
    for r in (4, 10, 17, 22, 29, 35, 48, 61):
        c = ex.cell(r, 2)
        if c.value:
            c.font = Font(bold=True, color="FF1F4E79")
    for r in range(5, 80):
        v = ex.cell(r, 3).value
        if v is None:
            continue
        lab = str(ex.cell(r, 2).value or "")
        ex.cell(r, 3).number_format = (PCT if "rate" in lab.lower()
                                       else MON if "$m" in lab else CNT)
    out.append("Exec Summary: section headings styled, every figure formatted by its label")
    wb.save(dst)
    return out


if __name__ == "__main__":
    for x in run("a6.xlsx", "a7.xlsx"):
        print("  ", x)
