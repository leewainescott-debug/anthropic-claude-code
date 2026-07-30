"""The owner's 2707 adoptions that need the finished blocks in place.

Four jobs, each one of his edits generalised the way he applied it:

1. The Actuals column. On 1.7, 1.8 and 1.9 he added "Actuals" to the Portfolio Summary
   table, reading the tab's own actual-after-decisions total, with a "Variance to
   actuals" line under the block (summary Total minus Actuals). Applied to all ten
   portfolio tabs, wired by label to the block actuals.py just built - by label, and to
   the whole tab rather than to column B, because that block has since moved from the foot
   of the tab to the top and been redrawn from five lines to three, and the wiring had to
   follow it both times without being told where it went.
2. His live levers. The COE design tabs now carry his On/Off states (Hold and Offshore),
   and 2.7 carries two hand-set levers. The working tabs' lever cells are set to match,
   by person, so his decisions price through the whole cascade.
3. The 0.2 COE spend cells. His rev formulas read the old '3.4 COE Summary'; the intended
   figures are the COE groupings' net planned spend, which live on the COE tabs
   themselves. Each is repointed at the cell whose cached value his own file shows.
4. The 1.13 bar, extended over his new On/Off column and the wave-M Uplift % column
   beside it - done here because the bar normaliser earlier in the chain repaints it to the
   old width. His note under the roles table is corrected here too, for the same reason in
   reverse: finish.py's prose sweep keeps that sentence only while it matches his own
   workbook exactly, so the role count in it can only be corrected after finish has run.
"""
import json
import re

import openpyxl
from openpyxl.utils import get_column_letter as L

import opts

REVIEW = "REVIEW - Complete Role Mapping"
# the line of the 1.x actuals table this column quotes, and the head of the column its
# figure sits in. Both are the table's own words, which is the only part of that block that
# has stayed put through two redesigns and one move from the foot of the tab to the top.
ACTUALS_ROW = "Actual portfolio"
COST_HEAD = "Cost ($m)"


# ---------------------------------------------------------------- 1. the Actuals column
def actuals_column(wb, wv, out):
    a = json.load(open("anchors_final.json"))
    for one in [t for t in wb.sheetnames if re.match(r"^1\.(10|14|[1-9]) ", t)]:
        ws, wsv = wb[one], wv[one]
        # The actuals table's own actual line and the column its cost sits in, both found by
        # what they say rather than by where they are. The table used to be at the foot of
        # the tab with its labels in column B; it is now at the top beside the budget box
        # with them in K, and its five decomposition lines are now three - Actual portfolio,
        # Archetype portfolio, Variance - because the owner mocked it that way. Wiring by
        # label is what let it move and then change shape without this step being rewritten
        # twice, and is why the scan is over every column rather than over column B.
        foot = lab_col = act_col = None
        for r in range(1, ws.max_row + 1):
            for c in range(2, 21):
                if str(ws.cell(r, c).value or "").strip() == ACTUALS_ROW:
                    foot, lab_col = r, c
        if foot is None:
            out.append(f"{one}: no {ACTUALS_ROW!r} line on the actuals table - skipped")
            continue
        # the cost column is the one the table's own header calls Cost ($m), read upwards
        # from the actual line; the old rule - the rightmost cell on the row that reads a
        # working tab - is kept as the fallback if that head is ever reworded
        for k in range(foot - 1, max(foot - 8, 0), -1):
            hit = next((c for c in range(lab_col, 21)
                        if str(ws.cell(k, c).value or "").strip() == COST_HEAD), None)
            if hit:
                act_col = hit
                break
        if act_col is None:
            for c in range(lab_col + 1, 30):
                v = ws.cell(foot, c).value
                if isinstance(v, str) and v.startswith("=") and "'2." in v:
                    act_col = c
        # the Portfolio Summary table: bar, header row, Total Cost row
        bar = hdr = tot = None
        for r in range(1, 20):
            b = str(wsv.cell(r, 2).value or "").strip()
            if b == "Portfolio Summary":
                bar = r
            elif b == "Cost" and bar and r > bar:
                hdr = r
            elif b == "Total Cost" and hdr:
                tot = r
                break
        if not (hdr and tot and act_col):
            out.append(f"{one}: no Portfolio Summary table - skipped")
            continue
        if bar:
            ws.cell(bar, 7).fill = opts.fl(opts.BARC)
            ws.cell(bar, 7).font = opts.BARF
        h = ws.cell(hdr, 7)
        h.value = "Actuals"
        h.font, h.fill, h.alignment, h.border = (opts.HDRF, opts.fl(opts.NAVY),
                                                 opts.CEN, opts.BOX)
        x = ws.cell(tot, 7)
        x.value = f"=${L(act_col)}${foot}"
        x.number_format, x.alignment, x.font, x.border = (opts.M2, opts.RGT, opts.BOLD,
                                                          opts.BOX)
        # the variance line, on the first free row under the summary block
        placed = False
        for r in range(tot + 1, tot + 9):
            if all(wsv.cell(r, c).value is None and wb[one].cell(r, c).value is None
                   for c in (5, 6, 7)):
                ws.cell(r, 5).value = "Variance to actuals"
                ws.cell(r, 5).font = opts.BOLD
                ws.cell(r, 5).alignment = opts.LFT
                v = ws.cell(r, 6)
                # actual less archetype, the same way round as the actuals table's own
                # Variance line and 3.1's F column. The first cut was F-G (archetype less
                # actual), which put this cell and the table beside it on the same screen
                # with the same number and opposite signs.
                v.value = f"=$G${tot}-$F${tot}"
                v.number_format, v.alignment, v.font = opts.M2, opts.RGT, opts.BOLD
                placed = True
                break
        out.append(f"{one}: Actuals column on the summary (G{tot} <- {L(act_col)}{foot}, "
                   f"the actuals table's {ACTUALS_ROW} cost)"
                   f"{', variance line placed' if placed else ''}")


# ---------------------------------------------------------------- 2. his live levers
# The ledger row behind an FTE line, recovered from the formula that names the person.
# Two shapes are live: the direct reference the 2.x tabs now use - 'REVIEW…'!$B$36 - and
# the older INDEX($B:$B,36) form, which is still what a workbook built before the join
# was made insert-safe will carry. Matching both is what stops this step going quietly
# blind and dropping ten of the owner's levers.
_NAME_REF = re.compile(r"'REVIEW - Complete Role Mapping'!\$B(?::\$B,|\$)(\d+)")


def _fte_rows(wb, tab):
    """REVIEW row -> FTE-block row, read out of the name formulas.

    The name cells are live INDEX formulas into the ledger, so the ledger row each one
    points at is in the formula text - the one join that survives every rename and every
    cache strip.
    """
    ws = wb[tab]
    rows = {}
    for r in range(1, ws.max_row + 1):
        f = ws.cell(r, 2).value
        if isinstance(f, str):
            m = _NAME_REF.search(f)
            if m and ws.cell(r, 5).value is not None:
                rows.setdefault(int(m.group(1)), []).append(r)
    return rows


def _tab(wb, num):
    """A working tab by its number. The names change - 2.11 is renamed twice in the chain -
    and the number is the one part of a tab's name the model treats as its identity."""
    return next((s for s in wb.sheetnames if s.split(" ", 1)[0] == num), None)


def _design_levers(wb, wv, tab, hdr_col=8):
    """(REVIEW row, state, design row) off a COE design tab's own lever column.

    Every state, not only Hold and Offshore. His demand is that the per-role On/Off column
    on 1.11 / 1.12 / 1.13 is the single source of lever state and that the working tab
    matches it role by role. Reading two of the three states meant a lever he moved BACK to
    Onshore left the working tab holding the old Offshore, and the two tabs could disagree
    with every control on both of them still reading zero - the counts and the totals are
    each computed from their own tab.
    """
    ws = wb[tab]
    out = []
    pat = re.compile(r"\$AA\$?(\d+)")
    for r in range(1, ws.max_row + 1):
        f = ws.cell(r, 20).value                        # column T, the cost engine
        state = str(wv[tab].cell(r, hdr_col).value or "").strip()
        if isinstance(f, str) and REVIEW in f and state:
            m = pat.search(f)
            if m:
                out.append((int(m.group(1)), state, r))
    return out


# The design tabs speak Onshore / Offshore / Hold; the working tabs speak Filled / Hire /
# Offshore / Hold, because a working-tab lever also says whether a vacancy is being filled.
# Onshore therefore maps to whichever of Filled and Hire the role's own status makes it, and
# the two tabs agree on the one thing they both state: is this role offshored, held, or
# neither.
NEUTRAL = ("Filled", "Hire")


def sync_levers(wb, wv, out):
    R = wv[REVIEW]
    # design tab -> its working tab
    # by tab number, not by tab name: 2.11 is renamed to "2.11 Cyber Risk & Service Ops" in
    # the first chain and this step runs in the second, so a hardcoded name here would look
    # for a tab that no longer exists and silently sync nothing.
    for dtab, num in (("1.11 BP&T", "2.12"), ("1.12 SA&D", "2.13"),
                      ("1.13 Cyber Roles", "2.11")):
        wtab = _tab(wb, num)
        if wtab is None:
            out.append(f"no {num} tab in the workbook - {dtab}'s levers are NOT synced")
            continue
        pairs = _design_levers(wb, wv, dtab)
        fte = _fte_rows(wb, wtab)
        n = 0
        for rev_row, state, _drow in pairs:
            for r in fte.get(rev_row, []):
                cur = str(wb[wtab].cell(r, 5).value or "")
                if state in ("Hold", "Offshore"):
                    want = state
                elif cur in NEUTRAL:
                    want = cur                    # already neutral, and status decided it
                else:
                    want = ("Filled"
                            if str(R.cell(rev_row, 37).value or "") == "Filled" else "Hire")
                if cur != want:
                    wb[wtab].cell(r, 5).value = want
                    n += 1
                break
        out.append(f"{wtab}: {n} levers set from {dtab} ({len(pairs)} states on the tab) - "
                   f"the design tab's On/Off column is the single source of lever state")
    uplift_factor(wb, wv, out)
    # his two hand-set levers on 2.7, found by person in the ledger
    fte = _fte_rows(wb, "2.7 Infrastructure")
    targets = []
    # the portfolio column is a formula whose cache does not survive an openpyxl save,
    # so membership of 2.7 is proven by the FTE block itself: only 2.7's own ledger rows
    # are in its name-formula map
    for i in range(2, R.max_row + 1):
        nm = str(R.cell(i, 2).value or "").strip().lower()
        ro = str(R.cell(i, 3).value or "").strip().lower()
        if nm == "stevani kho" and ro == "delivery lead":
            targets.append((i, "Offshore", "Stevani Kho, Delivery Lead"))
        elif i == 431 and nm == "vacant" and ro == "quality assurance":
            # his 2707 set exactly one of Infrastructure's vacant QA roles - the name
            # formula on his own edited row points at ledger row 431
            targets.append((i, "Filled", "the vacant Quality Assurance role (r431)"))
    for rev_row, state, who in targets:
        rows = fte.get(rev_row, [])
        if rows:
            wb["2.7 Infrastructure"].cell(rows[0], 5).value = state
            out.append(f"2.7 Infrastructure: {who} -> {state}")
        else:
            out.append(f"2.7 Infrastructure: {who} (ledger row {rev_row}) not in the "
                       f"FTE block - NOT set")


# The cyber uplift part-charge, carried through to the working tab.
#
# His toggles on 1.13 charge a share of five COE roles to the cyber uplift programme. The
# role stays in the COE and the COE carries the rest of it, so 1.13's planned spend is net
# of the share - and 2.11's cost after decisions has to be the same figure, or the tab a GM
# pulls levers on disagrees with the tab the levers live on. 1.13!F8 and 2.11!Q13 are the
# two cells that have to tie, and they tie because both are the sum of the same per-role
# arithmetic: cost x lever factor x (1 - uplift %).
#
# The factor is a direct reference to the Uplift % cell on 1.13, so his cream cell drives
# both tabs from one place. It is written on every cyber role, not only the five he has
# toggled, so the column is one formula and a sixth toggle needs no code.
UPLIFT_TAB = "1.13 Cyber Roles"
UPLIFT_COL = 9                            # column I on 1.13, his cream Uplift % column
UPLIFT_SUFFIX = "*(1-N('{tab}'!$I${row}))"


def uplift_factor(wb, wv, out):
    dtab, wtab = UPLIFT_TAB, _tab(wb, "2.11")
    if dtab not in wb.sheetnames or wtab is None:
        out.append(f"{dtab} or {wtab} is not in the workbook - the uplift part-charge is "
                   f"NOT carried through to the working tab")
        return
    pairs = _design_levers(wb, wv, dtab)
    fte = _fte_rows(wb, wtab)
    ws = wb[wtab]
    n, missing = 0, []
    for rev_row, _state, drow in pairs:
        rows = fte.get(rev_row, [])
        if not rows:
            missing.append(rev_row)
            continue
        r = rows[0]
        f = ws.cell(r, 7).value                   # G, cost after decision
        if not (isinstance(f, str) and f.startswith("=")):
            continue
        suffix = UPLIFT_SUFFIX.format(tab=dtab, row=drow)
        base = f.split("*(1-N('")[0]
        if base + suffix == f:
            continue
        ws.cell(r, 7).value = base + suffix
        n += 1
    out.append(f"{wtab}: {n} cost-after cells now carry his 1.13 Uplift % as a factor, so "
               f"2.11's cost after decisions is 1.13's planned spend role for role"
               + (f" - {len(missing)} design rows have no FTE row on the working tab "
                  f"({missing[:6]})" if missing else ""))


# ---------------------------------------------------------------- 3. the 0.2 COE cells
# His rev formulas read the retired '3.4 COE Summary', whose F column was planned spend
# per COE grouping. Those figures live on the COE tabs' own grouping rows, so each cell
# is wired straight to its source and moves with the data - the mapping is his own, read
# off the old 3.4's labels: F6 Strategy & Architecture, F8 Transformation, F9 Business
# Partnering, F10 Data.
CFG_COE = {6: "='1.12 SA&D'!$G$6", 8: "='1.11 BP&T'!$F$7",
           9: "='1.11 BP&T'!$F$6", 10: "='1.12 SA&D'!$G$7"}


def repoint_02(wb, wv, rev_path, out):
    ws = wb["0.2 Data Config"]
    for r, f in CFG_COE.items():
        cur = ws.cell(r, 6).value
        if isinstance(cur, str) and "'3.4 COE " in cur:
            ws.cell(r, 6).value = f
            out.append(f"0.2!F{r} -> {f[1:]} (the grouping's planned spend, live)")
        else:
            out.append(f"0.2!F{r} holds {str(cur)[:40]!r} - not the 3.4 read, left alone")


# ------------------------------------------------- 4. the 1.13 bar and his note
ROWREF = re.compile(r"\$([A-Z]{1,2})\$(\d+)")


def _cy_rows(wb):
    """(design row, REVIEW row, name, title) for every role on 1.13's list."""
    ws, R = wb[UPLIFT_TAB], wb[REVIEW]
    hdr = next((r for r in range(1, 30)
                if str(ws.cell(r, 2).value or "").strip() == "Name"), None)
    if hdr is None:
        return None, []
    rows = []
    for r in range(hdr + 1, 90):
        v = ws.cell(r, 2).value
        if not (isinstance(v, str) and v.startswith("=")):
            if rows:
                break
            continue
        m = ROWREF.search(v)
        if not m:
            continue
        i = int(m.group(2))
        rows.append((r, i, str(R.cell(i, 2).value or "").strip(),
                     str(R.cell(i, 3).value or "").strip()))
    return hdr, rows


def cyber_note(wb, out):
    """His note under the roles table states the role count, and the count has changed."""
    ws = wb[UPLIFT_TAB]
    hdr, rows = _cy_rows(wb)
    if hdr is None:
        return
    n = len(rows)
    for r in range(hdr, min(ws.max_row, hdr + 80) + 1):
        v = ws.cell(r, 2).value
        if not (isinstance(v, str) and "roles)" in v and "come straight from" in v):
            continue
        new = re.sub(r"\b\d+ roles\)", f"{n} roles)", v)
        if new != v:
            ws.cell(r, 2).value = new
            out.append(f"{UPLIFT_TAB}!B{r}: the role count in his note reads {n}, not the count "
                       f"before the cyber uplift roles moved to 1.14 - a figure in a "
                       f"sentence that would otherwise contradict the table above it")
        return



def cyber_bar(wb, out):
    ws = wb["1.13 Cyber Roles"]
    for r in range(1, 30):
        if str(ws.cell(r, 2).value or "").strip() == "Roles" \
                and (ws.cell(r, 2).fill.patternType or ws.cell(r, 3).fill.patternType):
            # B to I: his On/Off column and, since wave M, the Uplift % column beside it
            for c in range(2, 10):
                ws.cell(r, c).fill = opts.fl(opts.BARC)
                ws.cell(r, c).font = opts.BARF
            out.append(f"1.13!B{r} bar extended over the On/Off and Uplift % columns")
            return


def freeze_32_counts(wb, out, done="cand.xlsx"):
    """3.2's Times applied cells become typed cream numbers, seeded from the final count.

    They are the owner's to set (D103): cream, and seeded so the tab is right on the day
    it is built. final3x cannot know the final platform count - it runs before fix1x
    takes the empty platforms off the 1.x tabs, so the count at its stage reads 30 where
    the finished file says 22 - so it writes =Lists!AH placeholders and this step, whose
    chain input is the fully built and recalculated cand.xlsx, freezes them to the real
    numbers. A live formula here was wrong twice over: cream promised a typed input, and
    a cell tracking the model's own count made the control under the bands (his count
    against the model's) incapable of ever firing.
    """
    t32 = "3.2 Overhead & Leadership"
    if t32 not in wb.sheetnames:
        return
    try:
        lists = openpyxl.load_workbook(done, data_only=True)["Lists"]
    except Exception as e:                                    # noqa: BLE001
        out.append(f"3.2 Times applied left as formulas - cannot read {done}: {e}")
        return
    ws, n = wb[t32], 0
    for row in ws.iter_rows(min_col=5, max_col=5):
        for c in row:
            m = re.match(r"^=Lists!\$AH\$(\d+)$", str(c.value or ""))
            if not m:
                continue
            v = lists.cell(int(m.group(1)), 34).value
            if not isinstance(v, (int, float)):
                out.append(f"3.2!{c.coordinate} left as formula - Lists!AH{m.group(1)} "
                           f"holds {v!r}")
                continue
            v = round(v, 6)
            c.value = int(v) if abs(v - round(v)) < 1e-6 else v
            # cream goes back on here: finish.py strips the input colour from every
            # formula (rightly - these were formulas when it ran), and the freeze is
            # the moment the cell becomes the typed input the colour promises
            c.fill = opts.fl(opts.YEL)
            n += 1
    out.append(f"3.2 Times applied: {n} cells frozen to the model's final counts, "
               f"typed and cream, his to retype")


def run(src, dst, rev_path="rev.xlsx"):
    wb = openpyxl.load_workbook(src)
    before = {t: {} for t in wb.sheetnames}
    for t in wb.sheetnames:
        for row in wb[t].iter_rows():
            for c in row:
                if c.value is not None:
                    before[t][c.coordinate] = c.value
    wv = openpyxl.load_workbook(src, data_only=True)
    out = []
    actuals_column(wb, wv, out)
    sync_levers(wb, wv, out)
    repoint_02(wb, wv, rev_path, out)
    cyber_bar(wb, out)
    cyber_note(wb, out)
    freeze_32_counts(wb, out)
    # the manifest of every cell this step wrote, so the shipped-workbook diff can tell a
    # declared edit from an accident
    touched = []
    for t in wb.sheetnames:
        for row in wb[t].iter_rows():
            for c in row:
                if c.value != before[t].get(c.coordinate):
                    touched.append([t, c.coordinate])
    json.dump(touched, open("post2707_manifest.json", "w"))
    out.append(f"{len(touched)} cells declared in post2707_manifest.json")
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
