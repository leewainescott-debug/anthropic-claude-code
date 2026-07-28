"""Logic fixes on the design tabs. No column is added or removed and no squad is repriced.

1.12 SA&D's roles list was three roles short of the ledger. Its own control cell said so -
"Check - roles listed vs counted (must be 0)" reading -3 - and nothing else in the file
noticed, because 4.0 tested roles and cost against the ledger and never read a control
cell on another tab. Two of the three are the roles DECISIONS D7 moved into COE SA&D; the
third was never written. The counts and the planned spend on that tab read the whole ledger
directly and were always right, but the AU / NZ split sums the roles list, so the split came
to 5.7816 against a planned spend of 6.5295 on the same row.

The rest of that gap is not the offshore toggle. The On/Off lever prices through the T
column, and planned spend, the AU / NZ split and the working tabs all read T, so the
discount is inside both sides of the subtraction and cannot be what is left over. What is
left over is the Business Partner and Domain Architect funding the portfolios carry, which
his own note on row 9 says is netted out of planned spend. That is what the line now says,
beside its label, on the two tabs where it exists. 1.13 nets nothing and gets no line.

His On/Off notes were being rewritten on the old finding and are not any more: they
describe what these tabs do. 1.11's lever, though, offered Onshore and Offshore only and
had no Hold leg in its T column, so a Hold set on 2.12 had nowhere to land. It has both
now, as 1.12 and 1.13 already did.

The strategic-programme note on 1.4 summed the ledger by squad name alone, and "EGI TDD"
is also a squad name behind Ampol Retail, so it reported 1.32m where the tab's own share
is 1.02m. It now carries the portfolio its working tab joins on.

Three platform blocks were empty shells: a navy platform bar, a full column-header row, no
squad row, a blank overhead line and a grey total reading zero. One of them, EGI Data on
1.3, DECISIONS D6a records as removed. They are collapsed to the single line that carries
the fact, so the tabs stop showing tables that are not finished.
"""
import re

import openpyxl
from openpyxl.styles import Border, PatternFill
from openpyxl.utils import get_column_letter as L

import opts

REVIEW = "REVIEW - Complete Role Mapping"
REV = f"'{REVIEW}'"
LAST = None                             # measured from the ledger in run()
NONE_FILL = PatternFill()
NO_BORDER = Border()

# Only blocks with no roles behind them at all. 1.5's EGI P&C block was in this list and
# should not have been: its funded input is blank, which made it look empty, but the squad
# is real and has a role in the ledger. The owner's instruction is that EGI P&C is a squad
# and stays as it is, so the block stays as it is and the blank input is his to set.
SHELLS = [("1.1 Ampol Retail", "Platform: Pricing & WFM"),
          ("1.3 Enterprise Data", "Platform: EGI Data")]
SHELL_NOTE = {"Platform: Pricing & WFM": "Platform: Pricing & WFM - combined into Above "
                                         "Store, no squad of its own",
              "Platform: EGI Data": "Platform: EGI Data - removed, no roles",
              }
# tab, the portfolio its roles list covers, the list's first row
COE = [("1.11 BP&T", "COE BP&T"), ("1.12 SA&D", "COE SA&D"),
       ("1.13 Cyber Roles", "COE Cyber")]
ROWREF = re.compile(r"\$([A-Z]{1,2})\$(\d+)")
# the lever column, by the two headings the owner has used for it
ONOFF_HDR = ("On/Off", "Onshore / Offshore")
# the columns that identify the person. A literal here is the role's own identity and is
# copied as it stands; a literal anywhere else on the template row is that role's decision
# or that role's note, and belongs to that role alone.
IDENTITY_HDR = ("Name", "Position Title", "Department", "Country")


def block_rows(ws, label, limit=90):
    start = None
    for r in range(1, min(ws.max_row, limit) + 1):
        v = str(ws.cell(r, 2).value or "").strip()
        if start is None:
            if v.startswith(label):
                start = r
            continue
        if v.startswith(("Platform:", "Total ", "Reconciled")):
            return start, r - 1
        if v.endswith(" Total"):
            return start, r
    return (start, min(ws.max_row, limit)) if start else (None, None)


def collapse_shells(wb, wv):
    out = []
    for tab, label in SHELLS:
        ws, wsv = wb[tab], wv[tab]
        lo, hi = block_rows(ws, label)
        if lo is None:
            out.append(f"{tab}: no {label!r} block")
            continue
        # a real block has a squad row: a name that is not a header, an overhead label or
        # a total, carrying a cost in column H
        real = any(
            (n := str(wsv.cell(r, 2).value or "").strip())
            and not n.startswith(("Platform Overhead", "Squad"))
            and not n.endswith(" Total")
            and isinstance(wsv.cell(r, 8).value, (int, float))
            and wsv.cell(r, 8).value
            for r in range(lo + 1, hi + 1))
        if real:
            out.append(f"{tab}: {label!r} has a squad in it, left alone")
            continue
        for m in [m for m in list(ws.merged_cells.ranges)
                  if m.min_row >= lo and m.max_row <= hi]:
            ws.unmerge_cells(str(m))
        for r in range(lo, hi + 1):
            for c in range(2, 13):
                x = ws.cell(r, c)
                x.value = None
                x.fill, x.border, x.font = NONE_FILL, NO_BORDER, opts.BODY
            ws.row_dimensions[r].height = None
        x = ws.cell(lo, 2)
        x.value, x.font, x.alignment = SHELL_NOTE[label], opts.BOLD, opts.LFT
        out.append(f"{tab}: {label!r} collapsed to one line, rows {lo}-{hi}")
    return out


def list_bounds(ws):
    """The roles list: the header row carrying 'Name', and the last row its formulas
    reference. The SUMIFS windows above run to row 50, so the list can grow into it."""
    hdr = next((r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").strip() == "Name"), None)
    if hdr is None:
        return None, None, None
    last = hdr
    for r in range(hdr + 1, ws.max_row + 1):
        if isinstance(ws.cell(r, 2).value, str) and ws.cell(r, 2).value.startswith("="):
            last = r
        elif str(ws.cell(r, 2).value or "").strip():
            break
    win = 50
    for row in ws.iter_rows(min_row=1, max_row=hdr):
        for c in row:
            if isinstance(c.value, str) and "$50" in c.value:
                win = 50
    return hdr, last, win


def list_columns(ws, hdr):
    """The lever column and the identity columns, found by their headings."""
    onoff, identity = None, set()
    for c in range(2, 30):
        h = str(ws.cell(hdr, c).value or "").strip()
        if h in ONOFF_HDR:
            onoff = c
        elif h in IDENTITY_HDR:
            identity.add(c)
    return onoff, identity


def extend_lists(wb, lg):
    """Write into each COE roles list any role the ledger has and the list does not.

    The row is copied off the last row of the list, because that row carries the formulas
    and the styling the list is built from. What it must not carry across is the owner's
    own decision about the person who happens to sit on it: 1.12's template row holds his
    Hold lever and his one-word note beside it, and copying those wrote three fabricated
    Hold decisions into roles he has never ruled on - $747,896.05, two of them filled
    people, zeroed by a copy-paste. A new role arrives Onshore, which is the list's own
    default, and arrives with no note against it.
    """
    out = []
    for tab, pf in COE:
        ws = wb[tab]
        hdr, last, win = list_bounds(ws)
        if hdr is None:
            out.append(f"{tab}: no roles list found")
            continue
        listed = set()
        for r in range(hdr + 1, last + 1):
            v = ws.cell(r, 2).value
            if isinstance(v, str):
                m = ROWREF.search(v)
                if m:
                    listed.add(int(m.group(2)))
        want = [i for i in range(2, LAST + 1)
                if str(lg.cell(i, 36).value or "") == pf]
        missing = [i for i in want if i not in listed]
        if not missing:
            out.append(f"{tab}: roles list complete, {len(listed)} roles")
            continue
        if last + len(missing) > win:
            out.append(f"{tab}: {len(missing)} roles will not fit before row {win}")
            continue
        cols = [c for c in range(2, 30) if ws.cell(last, c).value is not None]
        onoff, identity = list_columns(ws, hdr)
        if onoff is None:
            out.append(f"{tab}: no On/Off column on the header row - list left alone")
            continue
        r = last + 1
        cleared = []
        for i in missing:
            for c in cols:
                new, old = ws.cell(r, c), ws.cell(last, c)
                new._style = old._style
                new.number_format = old.number_format
                v = old.value
                if isinstance(v, str) and v.startswith("="):
                    # the reference carries the REVIEW row; the local row number moves too
                    v = ROWREF.sub(
                        lambda m: f"${m.group(1)}${i}" if int(m.group(2)) in listed
                        or int(m.group(2)) > 200 else m.group(0), v)
                    v = re.sub(r"(?<![$\d])" + str(last) + r"(?![\d])", str(r), v)
                elif c == onoff:
                    # the lever is a decision about the person on the template row
                    v = "Onshore"
                elif c not in identity and v is not None:
                    # a note, or a number typed against that person - neither is scaffold
                    cleared.append(f"{L(c)}{r} ({v!r} on the template row)")
                    v = None
                new.value = v
            if ws.cell(r, onoff).value is None:
                ws.cell(r, onoff).value = "Onshore"
                ws.cell(r, onoff)._style = ws.cell(last, onoff)._style
            r += 1
        out.append(f"{tab}: {len(missing)} roles written in "
                   f"(REVIEW rows {', '.join(str(i) for i in missing)}), "
                   f"each set Onshore in column {L(onoff)}")
        if cleared:
            out.append(f"{tab}: not carried across from the template row - "
                       f"{'; '.join(cleared)}")
    return out


# the gap between planned spend and the AU / NZ split, tab by tab, in the owner's terms.
# It is not an offshore discount: the On/Off lever prices through the T column on both
# sides of that subtraction, so the discount is already inside both figures. What is left
# is the funding the portfolios carry for this COE, which his own note on row 9 says is
# netted out of planned spend and which row 13 states in full.
FUND_LINE = {
    "1.11 BP&T": "Business Partner funding met by portfolio overheads, netted out of "
                 "planned spend ($m)",
    "1.12 SA&D": "Domain Architect funding met by portfolio overheads, netted out of "
                 "planned spend ($m)",
}
# a negative reads in brackets, the way every other funding-block figure does
FUND_FMT = "#,##0.00;(#,##0.00)"


def funding_line(wb):
    """Name the gap between planned spend and the AU / NZ split, and put it beside its
    label.

    1.13 gets no line at all. Its planned spend is not netted against a portfolio-funded
    line, so the figure would be a permanent zero under a heading that implies otherwise -
    and with no funding block to sit under it landed at B9, in the white space over the
    roles table.

    The figure stays a formula, spend less AU less NZ, so it goes on checking itself
    against the two columns it reconciles rather than restating row 13.
    """
    out = []
    for tab, _ in COE:
        if tab not in FUND_LINE:
            out.append(f"{tab}: planned spend is not netted against a funded line, "
                       f"no funding line written")
            continue
        ws = wb[tab]
        tot = next((r for r in range(1, 30)
                    if str(ws.cell(r, 2).value or "").strip() == "Total"), None)
        if tot is None:
            out.append(f"{tab}: no Total row on the summary block")
            continue
        # the planned spend and the two split columns, found by their headers
        hdr = tot - 3
        while hdr > 1 and str(ws.cell(hdr, 2).value or "").strip() != "Grouping":
            hdr -= 1
        cols = {}
        for c in range(3, 14):
            v = str(ws.cell(hdr, c).value or "").strip()
            if v.startswith("Planned spend"):
                cols["spend"] = c
            elif v.startswith("Cost - AU"):
                cols["au"] = c
            elif v.startswith("Cost - NZ"):
                cols["nz"] = c
        if len(cols) < 3:
            out.append(f"{tab}: could not find the spend and split columns")
            continue
        r = tot + 1
        while str(ws.cell(r, 2).value or "").strip():
            r += 1
        x = ws.cell(r, 2)
        x.value = FUND_LINE[tab]
        x.font, x.alignment = opts.BODY, opts.LFT
        y = ws.cell(r, 3)                       # column C, beside its label
        y.value = (f"={L(cols['spend'])}{tot}-{L(cols['au'])}{tot}"
                   f"-{L(cols['nz'])}{tot}")
        y.font, y.number_format, y.alignment = opts.BODY, FUND_FMT, opts.RGT
        out.append(f"{tab}: portfolio-funded line stated at B{r}, figure in C{r}")
    return out


# the lever engine as 1.12 and 1.13 write it. The Offshore leg on its own is the older
# form, and it is the only form 1.11 has.
OFFSHORE_ONLY = re.compile(r'\*IF\((\$[A-Z]{1,2}\d+)="Offshore",0\.4,1\)')


def hold_lever(wb):
    """Give 1.11 the Hold branch its two sibling tabs already have.

    Hold is a real decision on 1.12 and 1.13 - the role stays on the list and drops out of
    planned spend - and 2.12 mirrors whatever 1.11's lever says. On 1.11 the dropdown
    offered Onshore and Offshore only, and the T column had no Hold leg to price it with,
    so the one lever the reader can pull on the working tab had nowhere to land.
    """
    out = []
    for tab, _ in COE:
        ws = wb[tab]
        hdr, last, _ = list_bounds(ws)
        if hdr is None:
            out.append(f"{tab}: no roles list found, lever left alone")
            continue
        n = 0
        for r in range(hdr + 1, last + 1):
            for c in range(2, 30):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.startswith("=") and OFFSHORE_ONLY.search(v):
                    ws.cell(r, c).value = OFFSHORE_ONLY.sub(
                        lambda m: f'*IF({m.group(1)}="Hold",0,'
                                  f'IF({m.group(1)}="Offshore",0.4,1))', v)
                    n += 1
        out.append(f"{tab}: Hold branch added to {n} lever formulas, rows {hdr + 1}-{last}"
                   if n else f"{tab}: lever formulas already carry the Hold branch")
        d = 0
        for dv in ws.data_validations.dataValidation:
            f1 = dv.formula1
            if (dv.type == "list" and isinstance(f1, str) and "Onshore" in f1
                    and "Hold" not in f1 and f1.endswith('"')):
                dv.formula1 = f1[:-1] + ',Hold"'
                d += 1
        out.append(f"{tab}: {d} lever dropdown(s) now offer Hold" if d else
                   f"{tab}: lever dropdown already offers Hold")
    return out


# The two On/Off notes used to be rewritten here, on the finding that the lever moved the
# AU / NZ split and nothing else. That was true of the generation this file was written
# for. It is not true of this one: the lever prices through the T column, which is what
# planned spend, the split and the working tabs all read, so his notes describe what the
# tabs do and they pass through untouched.
NOTE_FIX = {
    # 38 characters in an 11-wide column: it rendered as "spend - budg" clipped at both
    # ends on all three COE tabs
    "Left to fund (spend - budget, + = over)": "Left to fund ($m, + = over budget)",
    "2.0 Group Summary": "3.1 Group Summary",
    "(1.3 / 4.3)": "(1.3 / 2.3)",
    "the yellow cell": "the cream cell",
    "yellow cell": "cream cell",
    # two typed slips in his own text: a section heading on 1.3 and a note on 0.2
    "Siginificant": "Significant",
    "acorss": "across",
    # 1.5's summary header lost its country word; nine siblings read "TDD AU ($m)"
    "TDD  ($m)": "TDD AU ($m)",
}


# the five strategic-programme notes are built by formula, so the text swaps have to reach
# inside a formula string too - that is why "Set the agreed cost in the yellow cell" was
# still on 1.1, 1.4, 1.5 and 1.6 after the recolour. The same five formulas also read
# REVIEW rows 529 and 530, which are outside the 528-row window every other formula uses,
# and grouped on column AP instead of AT, the grouping column the rest of the model joins on.
IN_FORMULA = {"yellow cell": "cream cell"}


# the strategic-programme note: "People in this program today cost ..." summed off the
# ledger by squad name alone
NOTE_SUMIFS = re.compile(
    r"SUMIFS\('([^']+)'!\$AA\$2:\$AA\$(\d+),"
    r"'([^']+)'!\$(AP|AT)\$2:\$(?:AP|AT)\$(\d+),\$B(\d+)\)")


def _col_by_header(lg, name):
    for c in range(1, lg.max_column + 1):
        if str(lg.cell(1, c).value or "").strip() == name:
            return c
    return None


def _portfolio_cell(wb, tab):
    """The portfolio this design tab covers, taken from its own working tab.

    Not hardcoded and not read off the design tab: the working tab is where the model
    states a tab's portfolio, and it is the value every 2.x formula already joins the
    ledger on, so the note lands on exactly the rows its own tab reports.
    """
    num = tab.split(" ", 1)[0]
    if not num.startswith("1."):
        return None, None
    want = "2." + num[2:]
    sib = next((s for s in wb.sheetnames if s.split(" ", 1)[0] == want), None)
    if sib is None:
        return None, None
    ws = wb[sib]
    for r in range(1, 12):
        if str(ws.cell(r, 2).value or "").strip() == "Portfolio":
            v = ws.cell(r, 3).value
            if isinstance(v, str) and v.strip() and not v.startswith("="):
                return f"'{sib}'!$C${r}", v.strip()
    return None, None


def fix_double_count(wb, lg):
    """A programme name that belongs to two portfolios read both of them.

    The note under a strategic programme sums the ledger by squad name and nothing else.
    "EGI TDD" is a squad name on 1.4 and it is also a squad name behind Ampol Retail, so
    1.4's note reported 1.32m of people against a programme whose share of them is 1.02m -
    a figure a reader would check the funding against.

    The criterion added is the tab's own portfolio, read from its working tab, so the note
    counts the same rows the rest of the tab counts. Only the notes that actually double
    count are touched; the other four name a programme that sits in one portfolio, and a
    criterion there would be noise.
    """
    mtab = _col_by_header(lg, "MTab")
    if mtab is None:
        return ["MTab column not found on the ledger - programme notes left alone"]
    mt_l = L(mtab)
    out, n = [], 0
    for ws in wb.worksheets:
        pref, pval = _portfolio_cell(wb, ws.title)
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                m = NOTE_SUMIFS.search(v)
                if not m:
                    continue
                rev, end = m.group(1), m.group(2)
                grp_col, brow = m.group(4), int(m.group(6))
                name = ws.cell(brow, 2).value
                if not isinstance(name, str) or name.startswith("="):
                    continue
                gc = _col_by_header(lg, "Squad (canonical, from col K)") \
                    if grp_col == "AP" else None
                gc = gc or (42 if grp_col == "AP" else 46)
                spread = {str(lg.cell(i, mtab).value or "")
                          for i in range(2, LAST + 1)
                          if str(lg.cell(i, gc).value or "").strip() == name.strip()}
                if len(spread) < 2:
                    continue
                if pref is None:
                    out.append(f"{ws.title}!{c.coordinate}: {name!r} spans "
                               f"{sorted(spread)} but the tab's portfolio could not be "
                               f"read from its working tab - left alone")
                    continue
                c.value = v.replace(
                    m.group(0),
                    f"SUMIFS('{rev}'!$AA$2:$AA${end},"
                    f"'{m.group(3)}'!${grp_col}$2:${grp_col}${m.group(5)},$B{brow},"
                    f"'{rev}'!${mt_l}$2:${mt_l}${end},{pref})")
                n += 1
                out.append(f"{ws.title}!{c.coordinate}: {name!r} also sits under "
                           f"{sorted(spread - {pval})}, so the note now reads "
                           f"{pval!r} only ({pref})")
    out.append(f"{n} programme notes given their own portfolio as a criterion")
    return out


def fix_formula_notes(wb):
    # the window repairs are built at run time off the measured extent, so a ledger that
    # grows past 528 does not resurrect the off-window bug this map was written to fix
    IN_FORMULA["$AA$2:$AA$530"] = f"$AA$2:$AA${LAST}"
    IN_FORMULA["$AP$2:$AP$530"] = f"$AT$2:$AT${LAST}"
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                new = v
                for old, rep in IN_FORMULA.items():
                    new = new.replace(old, rep)
                if new != v:
                    c.value = new
                    n += 1
    return [f"{n} formula-built notes corrected - the cream wording, the 528-row window "
            f"and the AT grouping column"]


def fix_notes(wb):
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                new = v
                for old, rep in NOTE_FIX.items():
                    if old in new:
                        new = new.replace(old, rep)
                if new != v:
                    c.value = new
                    n += 1
    return [f"{n} labels corrected - two stale tab references, the yellow-cell wording "
            f"the recolour made wrong, and two typed slips"]


def run(src, dst, ledger="w1r.xlsx"):
    global LAST
    wb = openpyxl.load_workbook(src)
    LAST = opts.ledger_last(wb)
    wv = openpyxl.load_workbook(src, data_only=True)
    lg = openpyxl.load_workbook(ledger, data_only=True)[REVIEW]
    out = (extend_lists(wb, lg) + funding_line(wb) + hold_lever(wb)
           + collapse_shells(wb, wv) + fix_notes(wb) + fix_double_count(wb, lg)
           + fix_formula_notes(wb))
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(*sys.argv[1:]):
        print("  ", x)
