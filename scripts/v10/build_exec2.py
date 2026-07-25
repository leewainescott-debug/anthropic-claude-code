"""Stage 11: Exec Summary onto the split geometry, with labels that match the numbers."""
import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
import model, build_3x2

ITAL=Font(italic=True); REV=f"'{model.REVIEW}'"; LR=model.LAST_ROW
G32="'3.2 Total Cost'"; G33="'3.3 FTE View'"; G31="'3.1 Group Summary'"
SUB,TOT,OHT,ORG,GT,T31 = 16,21,36,37,117,20

def run(src,dst):
    wb=openpyxl.load_workbook(src); ws=wb["Exec Summary"]
    order=list(dict.fromkeys(model.TAB_PORTFOLIO.values()))
    VAC=f'SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AK$2:$AK${LR},"Vacant")/1000000'
    FIL=f'SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AK$2:$AK${LR},"Filled")/1000000'
    lines={
      5:("1. The contract: what the squad archetypes allow (roles)",f"={G32}!$I${SUB}"),
      6:("2. What happened: roles actually raised (all 525)",f"={G32}!$C${TOT}"),
      7:("3. What the organisation costs today ($m)",f"={G32}!$F${TOT}"),
      8:("4. Delivery squads over the archetype by ($m)",f"={G32}!$K${SUB}"),
      9:("5. The decision: value on the table - cost of every vacancy ($m)",f"={VAC}"),
      19:("Total TDD people budget ($m)","='0.2 Data Config'!$E$27"),
      20:("Allocated to portfolios and COEs ($m)",f"={G31}!$C${T31}"),
      21:("Not yet allocated ($m)",f"='0.2 Data Config'!$E$27-{G31}!$C${T31}"),
      23:("Squad archetype cost - the ten portfolios with an archetype ($m)",f"={G32}!$J${SUB}"),
      24:("Overhead allowance - all six lines from 0.2 Data Config ($m)","=Lists!$AJ$8"),
      25:("Total designed cost ($m)",f"={G32}!$J${SUB}+Lists!$AJ$8"),
      26:("Actual cost of the organisation today ($m)",f"={G32}!$F${TOT}"),
      27:("Delivery squads: actual over archetype ($m) - like for like",f"={G32}!$K${SUB}"),
      28:("Overhead: actual over allowance ($m) - like for like",f"={G32}!$H${OHT}"),
      30:("All roles today - filled plus vacant ($m)",f"={G32}!$F${TOT}"),
      31:("of which filled ($m)",f"={FIL}"),
      32:("of which vacant ($m)",f"={VAC}"),
      33:("Delivery squad cost ($m)",f"={G32}!$H${TOT}"),
      36:("Roles the archetypes allow - delivery squads",f"={G32}!$I${SUB}"),
      37:("Delivery squad roles actually raised",f"={G32}!$G${SUB}"),
      38:("Delivery squad roles beyond the archetype",f"={G32}!$G${SUB}-{G32}!$I${SUB}"),
      39:("Roles in the COEs and EGI - measured against budget",f"={G32}!$C${TOT}-{G32}!$C${SUB}"),
      40:("All org roles",f"={G33}!$I${GT}"),
      41:("Filled - people in roles today",f"={G33}!$G${GT}"),
      42:("Vacant - raised, not yet hired",f"={G33}!$H${GT}"),
      45:("Vacancy rate",f"={G33}!$L${GT}"),
      49:("Today's filled roles cost ($m)",f"={FIL}"),
      50:("Delivery squad cost over/(under) the archetype ($m)",f"={G32}!$K${SUB}"),
      51:("Hiring every vacancy would add ($m)",f"={VAC}"),
      52:("of which delivery squad roles ($m)",
          f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AR$2:$AR${LR},"Squad",'
          f'{REV}!$AK$2:$AK${LR},"Vacant")/1000000'),
      54:("Overhead roles inside the ledger ($m)",f"={G32}!$M${TOT}"),
      57:("Roles with no squad in the ledger",f'=COUNTIFS({REV}!$AP$2:$AP${LR},"Unassigned")'),
      59:("COEs - left to fund after budgets, see 3.4 ($m)","='3.4 COE Summary'!$H$12"),
    }
    for r,(lab,f) in lines.items():
        ws.cell(r,2).value=lab; ws.cell(r,3).value=f
    # drill-down onto the rebuilt 3.1 columns
    RNG=f"{G31}!$B$6:$B$19"
    dd=[(64,"TDD budget ($m)","C"),(65,"Actual cost of the portfolio ($m)","D"),
        (66,"Variance to budget ($m)","E"),(67,"Delivery squad cost ($m)","F"),
        (68,"Squad archetype cost ($m)","G"),(69,"Variance to archetype ($m)","H"),
        (70,"Overhead cost inside the portfolio ($m)","I"),(71,"Roles","J")]
    for r,lab,col in dd:
        ws.cell(r,2).value=lab
        ws.cell(r,3).value=(f"=IFERROR(INDEX({G31}!${col}$6:${col}$19,"
                            f'MATCH($C$63,{RNG},0)),"-")')
    for r in range(72,76):
        v=ws.cell(r,3).value
        if isinstance(v,str) and "$530" in v: ws.cell(r,3).value=v.replace("$530",f"${LR}")
    for dv in list(ws.data_validations.dataValidation):
        if "C63" in str(dv.sqref): ws.data_validations.dataValidation.remove(dv)
    dv=DataValidation(type="list",formula1=f'"{",".join(order)}"',allow_blank=False)
    ws.add_data_validation(dv); dv.add("C63")
    if str(ws["C63"].value or "") not in order: ws["C63"]=order[0]
    ws["B12"]=("Overhead is compared to its allowance, never added to a portfolio. The 8 GMs "
               "are the only overhead line with no role in the ledger.")
    ws["B13"]=("Strategic Programs squads carry no rate in 0.3, so they show no archetype "
               "comparison until one is set.")
    ws["B15"]=("Delivery squads and overhead roles are separated on every 2.x tab, so "
               "archetype and actual always compare the same people.")
    wb.save(dst); return len(lines)+len(dd)

if __name__=="__main__":
    print("rewired",run("s2.xlsx","s3.xlsx"),"Exec lines")
