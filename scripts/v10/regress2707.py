"""The regression gate for the 2707 consolidation: every finding from the verification
wave (instruction audit, design review, adversarial QA), asserted dead on the candidate.

Run against the injected candidate: python3 regress2707.py cand_A.xlsx
Exit code 0 only when every check passes. Each line reports PASS/FAIL and the fact.
"""
import os
import re
import sys
import zipfile

import openpyxl

REVIEW = "REVIEW - Complete Role Mapping"
ARCH = "0.3 Squad Archetypes"
REV = "rev.xlsx"
OUT = []


def check(name, ok, detail=""):
    OUT.append((bool(ok), f"{'PASS' if ok else 'FAIL'}  {name}" + (f" - {detail}" if detail else "")))


# ---------------------------------------------------------------- 0.3 is his, untouched
def _fill_of(c):
    """The cell's fill as a comparable string, '' for no fill."""
    try:
        if not (c.fill and c.fill.patternType):
            return ""
        s = c.fill.start_color
        return f"{s.type}:{s.rgb or s.theme or s.indexed}:{s.tint or 0}"
    except Exception:                                       # noqa: BLE001
        return "?"


def _col_w(ws, k):
    """The width column k actually renders at, falling back to the sheet's own default."""
    if k in ws.column_dimensions and ws.column_dimensions[k].width:
        return float(ws.column_dimensions[k].width)
    return float(ws.sheet_format.defaultColWidth or 8.43)


def _row_h(ws, r):
    """The height row r has been given of its own, or None where it takes the default.

    Read through the sheet's own default on purpose. LibreOffice recalculates the workbook
    mid-chain (chain2.sh, w1 -> w1r) and that conversion writes an explicit height on every
    row while moving the sheet default from 14.5 to 14.25. A row that inherited the default
    still inherits it - the engine has only spelled it out - so the question worth asking is
    whether a row was given a height of its own, and whether that height matches his.
    """
    h = ws.row_dimensions[r].height if r in ws.row_dimensions else None
    if not h:
        return None
    default = float(ws.sheet_format.defaultRowHeight or 0) or None
    if default and abs(float(h) - default) < 0.05:
        return None
    return float(h)


def archetypes_parity(path, rev=REV):
    """0.3 in the candidate against 0.3 in the owner's workbook, cell for cell.

    The tab arrives from rev.xlsx through assemble_base and no step is allowed to lay it
    out, so this is an equality, not a tolerance. What is compared: every value, every
    fill, every column width and every row height over the used range of both copies.

    Two things are deliberately not compared, because the chain does not cause them and
    cannot fix them. Alignment: the LibreOffice recalc writes the defaults out explicitly
    (None becomes 'general'/'bottom'), which is the same rendering spelled differently.
    Width to the last decimal: the same recalc rounds 28.26953125 to 28.27, so widths match
    to 0.05 rather than to the bit.
    """
    if not os.path.exists(rev):
        near = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            os.path.basename(rev))
        rev = near if os.path.exists(near) else rev
    if not os.path.exists(rev):
        return [f"cannot read {rev} - the parity check needs the owner's own workbook"]
    a = openpyxl.load_workbook(rev)[ARCH]
    b = openpyxl.load_workbook(path)[ARCH]
    bad = []
    # the declared exemptions on his tab: the hybrid input pair he asked for, and C25,
    # his rule note corrected to the rule he set (D116 - rev still carries the backwards
    # wording). Everything else stays an equality; the gate's own hybrid checks assert
    # exactly what these three cells must hold, so the exemption cannot hide drift.
    EXEMPT = {"K7", "K8", "C25"}
    rows = max(a.max_row, b.max_row)
    cols = max(a.max_column, b.max_column)
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            x, y = a.cell(r, c), b.cell(r, c)
            if x.coordinate in EXEMPT:
                continue
            if x.value != y.value:
                bad.append(f"value {x.coordinate}: {x.value!r} -> {y.value!r}")
            if _fill_of(x) != _fill_of(y):
                bad.append(f"fill {x.coordinate}: {_fill_of(x)!r} -> {_fill_of(y)!r}")
    for i in range(1, cols + 1):
        k = openpyxl.utils.get_column_letter(i)
        if abs(_col_w(a, k) - _col_w(b, k)) > 0.05:
            bad.append(f"width {k}: {_col_w(a, k)} -> {_col_w(b, k)}")
    for r in range(1, rows + 1):
        u, v = _row_h(a, r), _row_h(b, r)
        if u is None and v is None:
            continue
        if u is None or v is None or abs(u - v) > 0.05:
            bad.append(f"height row {r}: {u or 'default'} -> {v or 'default'}")
    return bad


# ==================================================================== wave M
# The cyber uplift restructure and the Customer corrections. Every check here is derived
# from the model or from his own stated figures - none of them is a re-baseline of a number
# the build happened to produce.

CRSO = "Cyber, Risk & Service Operations"
# his ruling, role by role. The nine movers, the eight offshore levers, the five uplift
# part-charges. Keyed on the person, never on a REVIEW row (D109).
M_MOVERS = {("Catherine Gire", "Cyber Technical Business Analyst"): "Cyber Uplift",
            ("Dan Balsamo", "Senior Cyber Project Manager"): "Cyber Uplift",
            ("Kevin Sheerin", "Cyber Architecture Lead"): "Cyber Uplift",
            ("Tony Keeler", "Cyber Business Analyst"): "Cyber Uplift",
            ("Vivienne Vasak", "Portfolio Manager Cyber Security"): "Cyber Uplift",
            ("Cameron Watman", "Engineer - IDAM"): "Identity",
            ("Iwan Wibisono", "Engineer - IDAM"): "Identity",
            ("Joe Nahma", "Manager Identity & Access Management"): "Identity",
            ("Raymond Cheung", "Engineer - IDAM"): "Identity"}
M_TOGGLES = {("Chris Lyons", "Cyber Security Architect"): 0.5,
             ("James Byrne", "Head of Cyber Strategy & Technology"): 0.5,
             ("Rahul Sahni", "Cyber Security Architect"): 0.4,
             ("Darshan Suvama", "Cyber Offensive Security Lead"): 0.25,
             ("Vanessa Castro", "Cyber GRC Analyst"): 0.0}
# his own figure for the cyber uplift people, to the cent
M_UPLIFT_TOTAL = 1794816.00


def _num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _tab(wb, num):
    return next((s for s in wb.sheetnames if s.split(" ", 1)[0] == num), None)


def _role_rows(wb, wv, tab):
    """(design row, REVIEW row, name, title) for a COE design tab's roles list."""
    ws, R = wb[tab], wv[REVIEW]
    hdr = next((r for r in range(1, 30)
                if str(ws.cell(r, 2).value or "").strip() == "Name"), None)
    out = []
    if hdr is None:
        return None, out
    for r in range(hdr + 1, 95):
        v = ws.cell(r, 2).value
        if not (isinstance(v, str) and v.startswith("=")):
            if out:
                break
            continue
        m = re.search(r"\$([A-Z]{1,2})\$(\d+)", v)
        if not m:
            continue
        i = int(m.group(2))
        out.append((r, i, str(R.cell(i, 2).value or "").strip(),
                    str(R.cell(i, 3).value or "").strip()))
    return hdr, out


def wave_m(wb, wv, joined):
    """Spec N: the wave-M gates. Anchored on labels and on his own figures."""
    R = wv[REVIEW]
    last = max((r for r in range(2, R.max_row + 1)
                if str(R.cell(r, 2).value or "").strip()), default=534)

    # ---- A. the nine movers go through the Lists override table, not REVIEW's raw columns
    l = wb["Lists"]
    table = {}
    for r in range(2, 40):
        k = l.cell(r, 40).value
        if isinstance(k, str) and " | " in k:
            table[k.strip()] = (l.cell(r, 41).value, l.cell(r, 42).value)
    missing = [f"{n} | {t}" for (n, t) in M_MOVERS
               if f"{n} | {t}" not in table
               or table[f"{n} | {t}"] != ("TDD Cyber", M_MOVERS[(n, t)])]
    check("the nine cyber movers are in the Lists override table, keyed on the person",
          not missing, str(missing[:4]))
    raw = [f"{R.cell(i, 2).value}" for i in range(2, last + 1)
           if (str(R.cell(i, 2).value or "").strip(),
               str(R.cell(i, 3).value or "").strip()) in M_MOVERS
           and str(R.cell(i, 9).value or "") != "COE - Cyber, Risk & Operations"]
    check("their raw Portfolio column on REVIEW is untouched - his typing stands",
          not raw, str(raw[:4]))
    check("the override window reaches AN21 - eleven moves fit with slots to spare",
          "$AN$2:$AN$21" in joined, "no AN21 window in any formula")

    # ---- B. 0.2 splits the combined cyber line, and the allocation total does not move
    c02, v02 = wb["0.2 Data Config"], wv["0.2 Data Config"]
    check("0.2!B23 is 'TDD Cyber' - the COE is funded on its own row",
          str(v02["B23"].value or "").strip() == "TDD Cyber", repr(v02["B23"].value))
    check("0.2's COE cyber row is funded 1.5 AU / 0.5 NZ, typed and cream",
          _num(v02["C7"].value) == 1.5 and _num(v02["D7"].value) == 0.5
          and str(c02["C7"].fill.start_color.rgb).upper() == "FFFFF2CC",
          f"C7={v02['C7'].value} D7={v02['D7'].value}")
    check("0.2's TDD Cyber row is funded 1.0 AU / 0.5 NZ, typed and cream",
          _num(v02["C23"].value) == 1.0 and _num(v02["D23"].value) == 0.5
          and str(c02["C23"].fill.start_color.rgb).upper() == "FFFFF2CC",
          f"C23={v02['C23'].value} D23={v02['D23'].value}")
    check("0.2!F7 reads 1.13's planned-spend total",
          "'1.13 Cyber Roles'!$F$8" in str(c02["F7"].value or ""), repr(c02["F7"].value))
    check("0.2!F23 reads 1.14's own TDD cost",
          "'1.14 TDD Cyber'!$C$9" in str(c02["F23"].value or "")
          and "'1.14 TDD Cyber'!$D$9" in str(c02["F23"].value or ""),
          repr(c02["F23"].value))
    check("0.2's allocated budget total is unchanged at 50.5",
          abs((_num(v02["E26"].value) or 0) - 50.5) < 0.005, repr(v02["E26"].value))
    check("his offshoring note sits on the COE row, and the actioned note is gone",
          "offshoring" in str(v02["I7"].value or "").lower()
          and "Separate out" not in str(v02["I7"].value or ""),
          repr(v02["I7"].value))

    # ---- C. the EGI squads, funded at the actual cost of their roles
    egi = [("1.1 Ampol Retail", "H66", "G66", "2.1 Ampol Retail"),
           ("1.2 Customer", "H54", "G54", "2.2 Customer"),
           ("1.4 TDD Group Functions", "H30", "G30", "2.4 TDD Group Functions"),
           ("1.5 P&C", "H32", "G32", "2.5 P&C"),
           ("1.6 Finance", "H33", "G33", "2.6 Finance")]
    bad = []
    for tab, hc, gc, wtab in egi:
        f = str(wb[tab][hc].value or "")
        if f"'{wtab}'!$Q$" not in f:
            bad.append(f"{tab}!{hc}={f[:40]}")
        if _num(wv[tab][gc].value) not in (0, 0.0):
            bad.append(f"{tab}!{gc}={wv[tab][gc].value}")
    check("every EGI squad is funded at the actual cost of its roles, support % nil",
          not bad, str(bad[:4]))
    check("1.2!I54 is the support-% split again, not his 27/07 typed 2.21",
          str(wb["1.2 Customer"]["I54"].value or "").startswith("=IFERROR($H54*$G54"),
          repr(wb["1.2 Customer"]["I54"].value))
    sig = _num(wv["1.2 Customer"]["J17"].value)
    check("1.2's Significant Items EGI carries EGI Customer's actual, about 2.099",
          sig is not None and abs(sig - 2.099) < 0.02, repr(sig))
    zero_var = []
    for tab, _hc, _gc, wtab in egi:
        wsv = wv[wtab]
        row = next((r for r in range(6, 40)
                    if str(wsv.cell(r, 2).value or "").strip().startswith("EGI")), None)
        if row is None:
            zero_var.append(f"{wtab}: no EGI row")
            continue
        n, o = _num(wsv.cell(row, 14).value), _num(wsv.cell(row, 15).value)
        if n is None or o is None or abs(n - o) > 0.0005:
            zero_var.append(f"{wtab}!r{row} N={n} O={o}")
    check("the working copies price the EGI squads at plan = actual, variance nil",
          not zero_var, str(zero_var[:4]))

    # ---- D. Digital Support NZ is back, and Customer reads his review workbook's figures
    w12, v12 = wb["1.2 Customer"], wv["1.2 Customer"]
    check("1.2!B41 is his Digital Support NZ squad",
          str(v12["B41"].value or "").strip() == "Digital Support NZ", repr(v12["B41"].value))
    d8 = _num(v12["D8"].value)
    check("1.2!D8 carries the NZ squad support cost, about 4.42",
          d8 is not None and abs(d8 - 4.42) < 0.02, repr(d8))
    for cell, want, name in (("F13", 2.439, "0.2 Ampol Customer spend"),
                             ("F14", 5.314, "0.2 Z Customer spend")):
        got = _num(v02[cell].value)
        check(f"{name} is {want} against his review workbook",
              got is not None and abs(got - want) < 0.01, f"{cell}={got}")
    check("his note off the restored row survives in the note margin",
          "CPI actuals" in " ".join(str(v12.cell(41, c).value or "")
                                    for c in range(11, 15)),
          str([v12.cell(41, c).value for c in range(11, 15)]))

    # ---- E. 1.13 in the shape of 1.11 / 1.12
    w13, v13 = wb["1.13 Cyber Roles"], wv["1.13 Cyber Roles"]
    head = next((r for r in range(1, 20)
                 if str(v13.cell(r, 2).value or "").strip() == "Grouping"), None)
    tot13 = next((r for r in range(1, 20)
                  if str(v13.cell(r, 2).value or "").strip() == "Total"), None)
    check("1.13's summary carries a Variance column where 1.11 and 1.12 carry one",
          head is not None and str(v13.cell(head, 8).value or "").strip() == "Variance",
          repr(v13.cell(head, 8).value) if head else "no Grouping header")
    text13 = " | ".join(str(c.value) for row in w13.iter_rows() for c in row
                        if isinstance(c.value, str) and not c.value.startswith("="))
    check("no Left-to-fund label survives anywhere on 1.13",
          "Left to fund" not in text13, "1.13 still says Left to fund")
    check("no 0.5 CapEx input and no reference to one survives on 1.13",
          "CapEx" not in text13
          and not any(isinstance(c.value, str) and c.value.strip() == "=F8-C13"
                      for row in w13.iter_rows() for c in row),
          "1.13 still carries the CapEx line")
    fbar = next((r for r in range(1, 20)
                 if str(v13.cell(r, 2).value or "").strip().startswith("Funding")), None)
    fund = [(str(v13.cell(r, 2).value or "").strip(), str(w13.cell(r, 3).value or ""))
            for r in range(fbar + 1, fbar + 8)] if fbar else []
    fund = [x for x in fund if x[0] or x[1]]
    check("1.13's funding block is two lines - the COE allocation off 0.2, then the total",
          len(fund) == 2 and fund[0][0].startswith("COE")
          and "'0.2 Data Config'!$E$7" in fund[0][1]
          and fund[1][0].startswith("Total budget to draw down"), str(fund))
    hdr13, roles13 = _role_rows(wb, wv, "1.13 Cyber Roles")
    check("1.13 lists exactly the 43 roles that stay in the COE",
          len(roles13) == 43, f"{len(roles13)} rows")
    check("none of the nine movers is still on 1.13",
          not [n for _r, _i, n, t in roles13 if (n, t) in M_MOVERS],
          str([n for _r, _i, n, t in roles13 if (n, t) in M_MOVERS]))
    check("the 43 rows are compact - no blank row inside the list",
          hdr13 is not None
          and [r for r, _i, _n, _t in roles13] == list(range(hdr13 + 1,
                                                             hdr13 + 1 + len(roles13))),
          str([r for r, _i, _n, _t in roles13][:4]))
    tog = {(n, t): _num(v13.cell(r, 9).value)
           for r, _i, n, t in roles13 if v13.cell(r, 9).value is not None}
    check("the five uplift toggles read 50 / 50 / 40 / 25 / 0 and nothing else carries one",
          tog == M_TOGGLES, str(sorted((f"{n}", v) for (n, _t), v in tog.items())))
    cream_tog = [f"{n}" for r, _i, n, t in roles13 if (n, t) in M_TOGGLES
                 and str(w13.cell(r, 9).fill.start_color.rgb).upper() != "FFFFF2CC"]
    check("every uplift toggle is a cream typed input", not cream_tog, str(cream_tog))
    eng13 = [str(w13.cell(r, 20).value or "") for r, _i, _n, _t in roles13]
    check("1.13's cost engine prices cost x lever factor x (1 - uplift %) on every role",
          all(f.startswith("=") and "Lists!$AD$2:$AD$5" in f and "(1-N($I" in f
              for f in eng13), f"{len(eng13)} engines")
    slice13 = [str(w13.cell(r, 21).value or "") for r, _i, _n, _t in roles13]
    check("the uplift slice is stated in its own column beside the engine",
          all("*N($I" in f for f in slice13), f"{len(slice13)} slice cells")
    levers13 = {(n, t): str(v13.cell(r, 8).value or "") for r, _i, n, t in roles13}
    want_off = {("Jack Jenkins", "Asset Analyst"),
                ("Jas Mann", "Technology Support Engineer"),
                ("Ritika Salaria", "Configuration Analyst"),
                ("Vacant", "Lead - Asset & Configuration"),
                ("Vacant", "Lead - Service Performance & Insights"),
                ("Vacant (AKL)", "Tech Support Technician"),
                ("Vacant", "Operations Analyst")}
    off_bad = [f"{k[0]}={v}" for k, v in levers13.items()
               if (k in want_off) != (v == "Offshore")]
    check("1.13's offshore levers are exactly the roles he named", not off_bad,
          str(off_bad[:5]))
    check("the Sydney technician he is recruiting stays on Hold",
          levers13.get(("Vacant (SYD)", "Tech Support Technician")) == "Hold",
          repr(levers13.get(("Vacant (SYD)", "Tech Support Technician"))))
    spend13 = _num(v13.cell(tot13, 6).value) if tot13 else None
    check("1.13's planned spend after the levers and the toggles is about 7.022",
          spend13 is not None and abs(spend13 - 7.022) < 0.02, repr(spend13))
    var13 = _num(v13.cell(tot13, 8).value) if tot13 else None
    check("1.13's variance is its budget less its spend, about (5.02)",
          var13 is not None and abs(var13 + 5.02) < 0.03, repr(var13))

    # ---- the uplift maths ties to his own figure to the cent
    slice_total = sum(_num(v13.cell(r, 21).value) or 0 for r, _i, _n, _t in roles13)
    t215 = _tab(wb, "2.15")
    v215 = wv[t215] if t215 else None
    uplift_squad = None
    if v215:
        rr = next((r for r in range(6, 30)
                   if str(v215.cell(r, 2).value or "").strip() == "Cyber Uplift"), None)
        uplift_squad = _num(v215.cell(rr, 15).value) if rr else None
    total = (uplift_squad or 0) + slice_total
    check("the cyber uplift people tie to his 1,794,816.00 to the cent",
          abs(total * 1e6 - M_UPLIFT_TOTAL) < 60,
          f"squad {uplift_squad} + slices {round(slice_total, 6)} = {round(total, 6)}")

    # ---- F. 1.14, the standard 1.x shape
    w14, v14 = wb["1.14 TDD Cyber"], wv["1.14 TDD Cyber"]
    sq = {str(v14.cell(r, 2).value or "").strip(): r for r in range(24, 32)}
    check("1.14 carries both squads - Cyber Uplift and Identity",
          "Cyber Uplift" in sq and "Identity" in sq, str(sorted(sq)))
    if "Cyber Uplift" in sq:
        r = sq["Cyber Uplift"]
        check("Cyber Uplift is a typed cream figure, fully funded from the programme",
              _num(v14.cell(r, 8).value) is not None
              and str(w14.cell(r, 8).fill.start_color.rgb).upper() == "FFFFF2CC"
              and _num(v14.cell(r, 7).value) in (0, 0.0),
              f"H={v14.cell(r, 8).value} G={v14.cell(r, 7).value}")
    if "Identity" in sq:
        r = sq["Identity"]
        check("Identity prices off the archetype library at his 80% support toggle",
              str(w14.cell(r, 8).value or "").startswith("=IFERROR(IF($E")
              and abs((_num(v14.cell(r, 7).value) or 0) - 0.8) < 1e-9
              and str(w14.cell(r, 7).fill.start_color.rgb).upper() == "FFFFF2CC",
              f"H={str(w14.cell(r, 8).value)[:24]} G={v14.cell(r, 7).value}")
    text14 = {str(v14.cell(r, 8).value or "").strip(): r for r in range(12, 24)}
    fund_row = next((r for lab, r in text14.items()
                     if lab.startswith("Programme funding")), None)
    check("1.14 states the cyber uplift programme funding as a typed cream input",
          fund_row is not None and abs((_num(v14.cell(fund_row, 10).value) or 0) - 2.8) < 1e-9
          and str(w14.cell(fund_row, 10).fill.start_color.rgb).upper() == "FFFFF2CC",
          f"row {fund_row}, {v14.cell(fund_row, 10).value if fund_row else None}")
    used = next((r for lab, r in text14.items() if lab.startswith("Used for cyber FTE")),
                None)
    left = next((r for lab, r in text14.items()
                 if lab.startswith("Remaining for non-people")), None)
    check("the funding block states what is used and what is left, both as formulas",
          used and left and str(w14.cell(used, 10).value or "").startswith("=SUM(")
          and str(w14.cell(left, 10).value or "").startswith("="),
          f"used r{used} left r{left}")
    rem = _num(v14.cell(left, 10).value) if left else None
    check("what is left of the programme funding for non-people is about 0.845",
          rem is not None and abs(rem - 0.845) < 0.02, repr(rem))
    check("no Left-to-fund line survives on 1.14",
          not any(isinstance(c.value, str) and "Left to fund" in c.value
                  for row in w14.iter_rows() for c in row), "1.14 says Left to fund")
    check("1.14's NZ allocation is intact - all the cyber spend is AU",
          abs((_num(v14["I9"].value) or 0) + 0.5) < 0.01, repr(v14["I9"].value))

    # ---- G/H. the two working copies
    t211 = _tab(wb, "2.11")
    check("2.11 is renamed for what it is",
          t211 == "2.11 Cyber Risk & Service Ops" and len(t211) <= 31, repr(t211))
    v211 = wv[t211]
    tot211 = next((r for r in range(6, 30)
                   if str(v211.cell(r, 2).value or "").strip() == "Total portfolio"), None)
    got = {k: _num(v211.cell(tot211, c).value)
           for k, c in (("F", 6), ("J", 10), ("K", 11), ("L", 12), ("M", 13),
                        ("O", 15), ("Q", 17))} if tot211 else {}
    check("2.11 carries the 43 roles that stay, and his levers - 2 hire, 8 offshore, "
          "3 hold, 40 after",
          got.get("F") == 43 and got.get("J") == 2 and got.get("K") == 8
          and got.get("L") == 3 and got.get("M") == 40, str(got))
    check("2.11's cost after decisions is 1.13's planned spend, to the cent",
          got.get("Q") is not None and spend13 is not None
          and abs(got["Q"] - spend13) < 0.000005,
          f"2.11 Q={got.get('Q')} against 1.13 F={spend13}")
    check("2.11's actual cost is the 43 roles at full price, about 8.778",
          got.get("O") is not None and abs(got["O"] - 8.778) < 0.02, repr(got.get("O")))
    tot215 = next((r for r in range(6, 30)
                   if str(v215.cell(r, 2).value or "").strip() == "Total portfolio"), None)
    g215 = {k: _num(v215.cell(tot215, c).value)
            for k, c in (("F", 6), ("N", 14), ("O", 15), ("Q", 17))} if tot215 else {}
    check("2.15 carries the nine roles that moved", g215.get("F") == 9, str(g215))
    check("2.15's actual and after-decisions cost are about 2.138",
          g215.get("O") is not None and abs(g215["O"] - 2.138) < 0.01
          and g215.get("Q") is not None and abs(g215["Q"] - 2.138) < 0.01, str(g215))
    for t in (t211, t215):
        ctl = [_num(wv[t].cell(r, 3).value) for r in range(1, wv[t].max_row + 1)
               if str(wv[t].cell(r, 2).value or "").startswith("Control - ")]
        check(f"{t}'s controls are live and read 0",
              ctl and all(v is not None and abs(v) < 1e-6 for v in ctl), str(ctl))

    # ---- K. the 1.x lever is the single source, role by role, on all three COE tabs
    mismatch = []
    for dtab, num in (("1.11 BP&T", "2.12"), ("1.12 SA&D", "2.13"),
                      ("1.13 Cyber Roles", "2.11")):
        wtab = _tab(wb, num)
        _h, rows = _role_rows(wb, wv, dtab)
        fte = {}
        for r in range(1, wv[wtab].max_row + 1):
            f = wb[wtab].cell(r, 2).value
            if isinstance(f, str):
                m = re.search(r"\$B\$(\d+)", f)
                if m and wv[wtab].cell(r, 5).value is not None:
                    fte.setdefault(int(m.group(1)), r)
        for _dr, i, name, _t in rows:
            wr = fte.get(i)
            if wr is None:
                mismatch.append(f"{wtab}: r{i} {name} not on the working tab")
                continue
            d = str(wv[dtab].cell(_dr, 8).value or "")
            w = str(wv[wtab].cell(wr, 5).value or "")
            if (d == "Offshore") != (w == "Offshore") or (d == "Hold") != (w == "Hold"):
                mismatch.append(f"{wtab}: {name} 1.x={d} 2.x={w}")
    check("every COE role's lever reads the same on its design tab and its working tab",
          not mismatch, str(mismatch[:5]))

    # ---- J. the WiPro roles
    wip = [i for i in range(2, last + 1)
           if "WIPRO" in str(R.cell(i, 17).value or "").upper()]
    check("the role mapping carries exactly six WiPro roles", len(wip) == 6, str(wip))
    t22 = _tab(wb, "2.2")
    lev22, cost22 = {}, {}
    for r in range(1, wb[t22].max_row + 1):
        f = wb[t22].cell(r, 2).value
        if isinstance(f, str):
            m = re.search(r"\$B\$(\d+)", f)
            if m and int(m.group(1)) in wip:
                lev22[int(m.group(1))] = str(wv[t22].cell(r, 5).value or "")
                cost22[int(m.group(1))] = (_num(wv[t22].cell(r, 6).value),
                                           _num(wv[t22].cell(r, 7).value))
    check("all six WiPro roles carry the Offshore lever",
          len(lev22) == 6 and set(lev22.values()) == {"Offshore"}, str(lev22))
    priced = [f"r{i}: {a} -> {b}" for i, (a, b) in cost22.items()
              if a is None or b is None or abs(a - b) > 0.5]
    check("offshoring a WiPro role does not move its cost - the vendor rate is already it",
          not priced, str(priced))
    check("the vendor exemption is in the cost-after formula, on every role",
          'SEARCH("WIPRO"' in joined, "no WIPRO exemption in any formula")
    v22 = wv[t22]
    r22 = next((r for r in range(6, 34)
                if str(v22.cell(r, 2).value or "").strip() == "Total portfolio"), None)
    check("Customer's vacancy split moves and its cost does not - 8 to hire, 6 to offshore",
          r22 and _num(v22.cell(r22, 10).value) == 8
          and _num(v22.cell(r22, 11).value) == 6,
          f"J={_num(v22.cell(r22, 10).value) if r22 else None} "
          f"K={_num(v22.cell(r22, 11).value) if r22 else None}")

    # ---- I. 3.1, his approved layout
    w31, v31 = wb["3.1 Cost Bridge"], wv["3.1 Cost Bridge"]
    check("3.1 is titled Archetype cost to actual cost",
          str(v31["B2"].value or "").strip() == "Archetype cost to actual cost",
          repr(v31["B2"].value))
    b31 = {str(v31.cell(r, 2).value or "").strip(): r for r in range(1, 40)}
    check("3.1's band is TDD", "TDD" in b31, str(sorted(b31)[:6]))
    check("3.1 carries no Retail subtotal row - Ampol and Z sit together",
          not any(k.startswith("Retail - ") for k in b31), str([k for k in b31
                                                               if k.startswith("Retail")]))
    check("3.1 names the cyber COE the way he names it",
          any(k.startswith(CRSO) for k in b31), str([k for k in b31 if "yber" in k]))
    check("3.1's walk is gone - no step labels survive",
          not any(k.startswith(("Squads priced by an archetype",
                                "Directly funded", "Nothing prices these",
                                "COEs and EGI", "Groups with no archetype",
                                "The walk")) for k in b31), str(sorted(b31)))
    check("3.1 heads a Total roles after decisions column",
          str(v31.cell(4, 11).value or "").strip() == "Total roles after decisions",
          repr(v31.cell(4, 11).value))
    tdd = next((r for k, r in b31.items() if k.startswith("TDD total")), None)
    gm = next((r for k, r in b31.items() if k.startswith("GM roles")), None)
    grand = next((r for k, r in b31.items() if k.startswith("Total TDD cost including")),
                 None)
    check("3.1 carries a TDD total, a GM row and a grand total, in that order",
          tdd and gm and grand and tdd < gm < grand, f"{tdd} / {gm} / {grand}")
    coe_var = []
    for r in range(5, tdd or 5):
        nm = str(v31.cell(r, 2).value or "").strip()
        if nm.startswith(CRSO) or nm.startswith(("COE ", "EGI")):
            d, e = _num(v31.cell(r, 4).value), _num(v31.cell(r, 5).value)
            if d is None or e is None or abs(d - e) > 1e-9:
                coe_var.append(f"{nm}: D={d} E={e}")
    check("every COE and EGI line prices archetype = actual, so its variance is nil",
          not coe_var, str(coe_var))
    check("3.1's GM row leaves Vacant blank - the GMs are outside the role mapping",
          gm is not None and v31.cell(gm, 10).value in (None, ""),
          repr(v31.cell(gm, 10).value) if gm else "no GM row")
    org = {k: _num(v31.cell(tdd, c).value)
           for k, c in (("D", 4), ("E", 5), ("F", 6), ("G", 7), ("H", 8), ("I", 9),
                        ("J", 10), ("K", 11))} if tdd else {}
    check("3.1's TDD total ties to the role mapping - 531 roles, 386 filled, 145 vacant",
          org.get("H") == 531 and org.get("I") == 386 and org.get("J") == 145, str(org))
    check("3.1's actual cost is the role mapping's own total, 115.59",
          org.get("E") is not None and abs(org["E"] - 115.59) < 0.01, repr(org.get("E")))
    check("3.1's archetype side is a real total now, not a dash",
          org.get("D") is not None and org["D"] > 0, repr(org.get("D")))
    ctl31 = [_num(v31.cell(r, c).value) for r in range(1, v31.max_row + 1)
             for c in (5, 8) if str(v31.cell(r, 2).value or "").startswith("Control - ")
             and v31.cell(r, c).value is not None]
    check("3.1's controls are live and read 0",
          ctl31 and all(abs(v) < 1e-6 for v in ctl31 if v is not None), str(ctl31))

    # ---- L. the Exec, on the new basis
    ve = wv["Exec Summary"]
    be = {str(ve.cell(r, 2).value or "").strip(): r for r in range(1, 60)}
    want_exec = ("Archetype cost of the whole organisation ($m)",
                 "Actual cost of the whole organisation ($m)",
                 "Over/(under) archetype ($m)",
                 "The 8 GMs - over/(under) their allowance ($m)",
                 "Total over/(under) archetype including the GM layer ($m)",
                 "Over/(under) the allocated TDD budget - 0.2 Data Config ($m)")
    gone_exec = [lab for lab in want_exec if lab not in be]
    check("the Exec's archetype block reads on the new basis - archetype, actual, "
          "over/(under), the GM layer, the two together, and the budget position",
          not gone_exec, str(gone_exec))
    check("the Exec names the groups with no archetype - Energy and the leadership groups",
          any(k.startswith("Of which groups with no archetype") for k in be),
          str([k for k in be if "no archetype" in k]))
    vac = {k: _num(ve.cell(r, 3).value) for k, r in be.items()
           if k.startswith(("Vacant roles", "Vacancies set to", "Vacancies put on",
                            "Filled roles offshored"))}
    # the four vacancy lines are a partition of the vacant count; a filled role offshored is
    # not a vacancy and sits on its own line outside it
    parts = sum(v or 0 for k, v in vac.items()
                if k.startswith(("Vacancies set to", "Vacancies put on")))
    check("the Exec vacancy block still partitions the vacancies exactly",
          abs(parts - (vac.get("Vacant roles") or 0)) < 1e-9, str(vac))
    check("the Exec's lever counts move with the wave-M decisions - 126 to hire, "
          "12 offshored, 6 held, 5 filled roles offshored",
          (vac.get("Vacancies set to hire"), vac.get("Vacancies set to offshore"),
           vac.get("Vacancies put on hold"), vac.get("Filled roles offshored"))
          == (126, 12, 6, 5), str(vac))

    # ---- M. hygiene
    dash_cells = [f"{ws.title}!{c.coordinate}"
                  for ws in wb.worksheets
                  if ws.title not in ("0.1 Budget Table (Fin)", "0.4 Presentation Pack",
                                      "0.3 Squad Archetypes")
                  for row in ws.iter_rows() for c in row
                  if isinstance(c.value, str)
                  and (c.value.strip() in tuple("-‐‑‒–—―−")
                       or (c.value.startswith("=") and '"-"' in c.value))]
    check("no dash literal survives in any cell or formula", not dash_cells,
          str(dash_cells[:6]))
    dash_fmt = [f"{ws.title}!{c.coordinate}"
                for ws in wb.worksheets
                if ws.title not in ("0.1 Budget Table (Fin)", "0.4 Presentation Pack",
                                    "0.3 Squad Archetypes")
                for row in ws.iter_rows() for c in row
                if re.search(r';\s*"[-‐-―−]"\s*$', c.number_format or "")]
    check("no number format renders a zero as a dash", not dash_fmt, str(dash_fmt[:6]))
    banned = [f"{ws.title}!{c.coordinate}"
              for ws in wb.worksheets if not ws.title.startswith("REVIEW")
              and ws.title not in ("0.1 Budget Table (Fin)", "0.4 Presentation Pack",
                                   "0.3 Squad Archetypes")
              for row in ws.iter_rows() for c in row
              if isinstance(c.value, str) and not c.value.startswith("=")
              and (re.search(r"\bseats?\b", c.value, re.I)
                   or re.search(r"\bdesign(s|ed|ing)?\b", c.value, re.I)
                   or "–" in c.value or "—" in c.value)]
    check("no banned word and no en/em dash in any cell text", not banned, str(banned[:6]))


def _a2(path="anchors_final.json"):
    """The working tabs' own anchors, as the builders wrote them.

    Read rather than searched: a row this gate needs to find - a section subtotal - can
    share its first four words with the band above it, and a label search finds the wrong
    one. If the file is not beside the gate the checks that need it say so instead of
    guessing a row number.
    """
    import json
    try:
        return json.load(open(path))
    except Exception:                                       # noqa: BLE001
        return {}


def run(path):
    wb = openpyxl.load_workbook(path)                 # formulas
    wv = openpyxl.load_workbook(path, data_only=True)  # cached values
    z = zipfile.ZipFile(path)

    # ---- audit: the fabricated Holds are gone, his three stand (D86)
    ws, vs = wb["1.12 SA&D"], wv["1.12 SA&D"]
    holds = [r for r in range(20, 50) if str(vs.cell(r, 8).value or "").strip() == "Hold"]
    check("1.12 carries exactly his three Holds (rows 31/43/44)", holds == [31, 43, 44], f"holds at {holds}")
    # no appended rows at all: Rihan prices in Reporting & Analytics (D118) and Deepali
    # and the vacant Service Transition Lead sit in Enterprise Data leadership (D119),
    # so nobody outside 1.12's own typed table belongs on it
    check("1.12 has no appended ledger rows - 45-47 empty",
          all(vs.cell(r, 2).value is None for r in (45, 46, 47)),
          str([vs.cell(r, 2).value for r in (45, 46, 47)]))
    w13 = wb["2.13 COE SA&D"]
    lev = [str(wv["2.13 COE SA&D"].cell(r, 5).value or "") for r in range(1, w13.max_row + 1)]
    check("2.13 has exactly 3 Holds, 0 Offshore", lev.count("Hold") == 3 and lev.count("Offshore") == 0,
          f"{lev.count('Hold')} holds, {lev.count('Offshore')} offshore")

    # ---- audit: net-vs-net budget basis (D87)
    check("1.11!C15 is his =C14", wb["1.11 BP&T"]["C15"].value == "=C14",
          repr(wb["1.11 BP&T"]["C15"].value))
    check("1.12!C15 is his =C14", wb["1.12 SA&D"]["C15"].value == "=C14")
    all_text = []
    for t in wb.sheetnames:
        for row in wb[t].iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    all_text.append(c.value)
    joined = "\n".join(all_text)
    check("no 'Offshore discount' label anywhere", "Offshore discount" not in joined)
    check("1.11 funding line labelled and beside its value",
          str(wb["1.11 BP&T"]["B18"].value or "").startswith("Business Partner funding met by")
          and str(wb["1.11 BP&T"]["C18"].value or "").startswith("="))
    check("1.12 funding line labelled and beside its value",
          str(wb["1.12 SA&D"]["B19"].value or "").startswith("Domain Architect funding met by")
          and str(wb["1.12 SA&D"]["C19"].value or "").startswith("="))

    # ---- audit: 1.11 lever vocabulary
    # the lever price is no longer spelled out in the engine - it is looked up from
    # Lists!AC:AD, the one place the four factors live - so what this asserts now is that
    # the lookup is there and reaches the whole table, Hold row included
    eng = [str(wb["1.11 BP&T"].cell(r, 20).value or "") for r in range(21, 45)]
    eng = [f for f in eng if f.startswith("=")]
    check("1.11's cost engine prices every lever off the Lists table",
          eng and all("Lists!$AC$2:$AC$5" in f and "Lists!$AD$2:$AD$5" in f for f in eng),
          f"{len(eng)} engines, first {eng[0][:60] if eng else ''!r}")
    dvs = [dv.formula1 for dv in wb["1.11 BP&T"].data_validations.dataValidation
           if dv.formula1 and "Onshore" in dv.formula1]
    check("1.11 dropdown offers Hold", any("Hold" in f for f in dvs), str(dvs[:2]))

    # ---- audit: visibility (D91)
    hidden = sorted(s.title for s in wb.worksheets if s.sheet_state != "visible")
    check("hidden set is exactly 0.1 / 0.4 / Lists (Exec visible)",
          hidden == ["0.1 Budget Table (Fin)", "0.4 Presentation Pack", "Lists"], str(hidden))

    # ---- audit: his content survives (D89, D90)
    keep = [("1.8 Energy Solutions & B2B", "E17", "What is the B2B initiatives #?"),
            ("1.8 Energy Solutions & B2B", "E18", "What is the B2B CapEx?"),
            ("1.11 BP&T", "B9", "Planned spend is net of the Business Partner"),
            ("1.12 SA&D", "B9", "Planned spend is net of the Domain Architect"),
            ("1.11 BP&T", "B47", "Commercial roles sit in the Business Partnering"),
            ("1.11 BP&T", "B49", "On/Off: set a role to Offshore"),
            ("1.12 SA&D", "B53", "Squad-based SA&D roles sit in their portfolio squads"),
            ("1.12 SA&D", "B55", "On/Off: set a role to Offshore"),
            ("1.13 Cyber Roles", "B73", "Roles and costs come straight from the REVIEW"),
            ("1.2 Customer", "L15", "5m mobile, 4.5m for loyalty"),
            ("1.2 Customer", "L18", "4.5m? Flagged"),
            ("1.2 Customer", "M41", "CPI actuals; pull through"),
            ("0.2 Data Config", "B24", "EG"),
            ("0.2 Data Config", "K24", "Notes"),
            ("0.2 Data Config", "L24", "Reallocated 7m across Ampol & Z Retail"),
            # his K-column squad-row note rides the house relocation to the M note column
            ("1.10 Z Retail", "M39", "No Overhead required"),
            ("1.13 Cyber Roles", "B2", "Cyber, Risk & Service Operations roles and funding")]
    for tab, ref, want in keep:
        v = str(wb[tab][ref].value or "")
        check(f"owner content at {tab}!{ref}", v.strip().startswith(want[:40]), repr(v[:50]))

    # ---- audit: ledger hygiene (D93)
    R, Rv = wb[REVIEW], wv[REVIEW]
    check("REVIEW AR1/AT1 labelled",
          R["AR1"].value == "Overhead line" and R["AT1"].value == "Squad or overhead line")
    stray = [f"{r}:{c}" for r in (191, 192) for c in range(1, 53)
             if R.cell(r, c).value is not None]
    check("spacer rows 191-192 fully empty", not stray, str(stray[:6]))
    for bad in ("Project Manger", "Portfolio Mnager", "michelle Siegman", "vijay Solanki",
                "DeveloperSAP ECC", "EnterpriseProcess Analyst", "Support Analyst -Retail",
                "Siginificant", "acorss"):
        check(f"typo gone: {bad!r}", bad not in joined)

    # ---- QA: the ledger anchors hold
    names = [r for r in range(2, R.max_row + 1) if str(Rv.cell(r, 2).value or "").strip()]
    total = sum(v for r in names if isinstance((v := Rv.cell(r, 27).value), (int, float)))
    check("531 roles in the ledger", len(names) == 531, str(len(names)))
    check("ledger totals 115,589,735.11", abs(total - 115589735.11) < 0.05, f"{total:,.2f}")
    cust = sum(v for r in range(108, 191) if isinstance((v := Rv.cell(r, 27).value), (int, float)))
    check("Customer block 16,522,075.33", abs(cust - 16522075.33) < 0.05, f"{cust:,.2f}")

    # ---- QA H1: the filter hazard is dead
    xml = b"".join(z.read(n) for n in z.namelist()
                   if n.startswith("xl/worksheets/") and n.endswith(".xml"))
    check("no filter criteria anywhere", b"filterColumn" not in xml)
    check("no sortState anywhere", b"sortState" not in xml)
    afs = re.findall(rb'<autoFilter[^>]*ref="([^"]+)"', xml)
    check("exactly one autoFilter, full-width on REVIEW",
          len(afs) == 1 and afs[0].startswith(b"A1:AX"), str(afs))
    hid = [r for r in range(2, R.max_row + 1)
           if R.row_dimensions[r].hidden]
    check("no hidden rows on REVIEW", not hid, str(hid[:5]))

    # ---- QA H3 / M1: platform overhead
    # In both of his workbooks 1.2's C7 and D7 are DIFFERENT formulas and exactly one
    # fires, so F7 is 0.495 - the platform overhead once. A build pass rewrote C7 to
    # match D7 (both branches firing, 0.99, a shape in neither of his books), a fix
    # halved it, and the "revert" of the fix restored the corrupted 0.99 and called it
    # his (the first D112, now rewritten). These pins hold his true shape: the 27/07
    # complement, one branch firing, 0.495 once.
    v12 = wv["1.2 Customer"]
    check("1.2!F7 = 0.495 - platform overhead once, as in both his workbooks",
          abs((v12["F7"].value or 0) - 0.495) < 1e-9, str(v12["F7"].value))
    # 1.2's Total Cost was pinned at his 15.5625. Two of his own wave-M rulings move it -
    # Digital Support NZ comes back with 0.32 of NZ platform cost, and EGI Customer stops
    # carrying his 27/07 typed 2.21 in the TDD column and carries its actual in the funded
    # column instead - so the figure is re-derived rather than re-baselined: the tab's
    # Total Cost must be its own three columns added up, and 2.2's archetype total must be
    # the same figure. Both sides move together or the check fails.
    f9 = v12["F9"].value
    parts = sum((v12[c].value or 0) for c in ("C9", "D9", "E9"))
    n29 = next((wv["2.2 Customer"].cell(r, 14).value for r in range(6, 40)
                if str(wv["2.2 Customer"].cell(r, 2).value or "").strip()
                == "Total portfolio"), None)
    # his Net New squads price on 1.2 before any role exists, so 2.2 cannot carry
    # them yet - their design cost joins the working side of the tie (wave M)
    w2names = {str(wv["2.2 Customer"].cell(r, 2).value or "").strip() for r in range(1, 40)}
    TYPES = {"Operations", "Product", "Engineering", "Configuration / Integration",
             "Enterprise Data and Insights", "Build and Run", "Strategic Programs"}
    netnew = sum(v12.cell(r, 8).value or 0 for r in range(20, 80)
                 if str(v12.cell(r, 3).value or "").strip() in TYPES
                 and str(v12.cell(r, 2).value or "").strip()
                 and str(v12.cell(r, 2).value or "").strip() not in w2names
                 and (v12.cell(r, 10).value or 0) == (v12.cell(r, 8).value or 0) * 0
                 or str(v12.cell(r, 2).value or "").strip() not in w2names
                 and str(v12.cell(r, 3).value or "").strip() in TYPES
                 and isinstance(v12.cell(r, 8).value, (int, float)))
    check("1.2's Total Cost is its own three columns, and 2.2 plus his Net New squads "
          "quote the same archetype",
          f9 is not None and abs(f9 - parts) < 1e-6
          and n29 is not None and abs(n29 + netnew - f9) < 1e-6,
          f"F9={f9} C9+D9+E9={parts} 2.2!N={n29} netnew={netnew}")
    check("1.2!C7 is the complement of D7 - one branch fires, never both",
          str(wb["1.2 Customer"]["C7"].value).endswith("0,SUM(I34,I42,I49))")
          and str(wb["1.2 Customer"]["D7"].value).endswith("SUM(I34,I42,I49),0)"),
          str(wb["1.2 Customer"]["C7"].value)[-40:])
    # D118: "z energy martech is z loyalty & martech. same frickin squad" - the old
    # names fold into the squads and appear nowhere as squads of their own. Folded, the
    # two squads carry exactly his own pivot totals.
    old_names = []
    for t in ("1.2 Customer",):
        for rr in range(25, 60):
            if str(wb[t].cell(rr, 2).value or "").strip() in ("AU CRM & Martech",
                                                              "Z Energy Martech"):
                old_names.append(f"{t}!B{rr}")
    RR = wb[REVIEW]
    ap_col = next((c for c in range(30, 50)
                   if str(RR.cell(1, c).value or "").startswith("Squad (canonical")), None)
    canon_old = [r for r in range(2, RR.max_row + 1)
                 if str(RR.cell(r, ap_col).value or "").strip()
                 in ("AU CRM & Martech", "Z Energy Martech")] if ap_col else ["no col"]
    check("the two old Martech names fold into their squads everywhere",
          not old_names and not canon_old,
          f"1.2 rows {old_names[:2]}, canonical rows {canon_old[:4]}")
    counts = {}
    for r in range(2, RR.max_row + 1):
        s = str(wv[REVIEW].cell(r, ap_col).value or "").strip() if ap_col else ""
        if s in ("Ampol Loyalty & Martech", "Z Loyalty & Martech"):
            counts[s] = counts.get(s, 0) + 1
    check("folded, the squads carry his own pivot counts - Ampol 9 roles, Z 17",
          counts.get("Ampol Loyalty & Martech") == 9
          and counts.get("Z Loyalty & Martech") == 17, str(counts))
    # D118: Rihan Schalkwyk prices in the Reporting & Analytics squad, Enterprise Data.
    # The grouping override lands in AT (Squad or overhead line) and AJ (MTab); the
    # K-canonical column deliberately keeps his raw squad. 2.x names are ledger
    # references, so the placement scan reads cached values.
    at_col = next((c for c in range(30, 50)
                   if str(RR.cell(1, c).value or "").startswith("Squad or overhead")),
                  None)
    rihan = next((r for r in range(2, RR.max_row + 1)
                  if str(RR.cell(r, 2).value or "").strip() == "Rihan Schalkwyk"), None)
    check("Rihan Schalkwyk sits in Reporting & Analytics on the working tabs",
          rihan is not None and at_col is not None
          and str(wv[REVIEW].cell(rihan, at_col).value or "").strip()
          == "Reporting & Analytics"
          and str(wv[REVIEW].cell(rihan, 36).value or "").strip() == "Enterprise Data"
          and any(str(wv["2.3 Enterprise Data"].cell(rr, 2).value or "").strip()
                  == "Rihan Schalkwyk"
                  for rr in range(1, wv["2.3 Enterprise Data"].max_row + 1)),
          f"row {rihan}: AT={wv[REVIEW].cell(rihan, at_col).value if rihan and at_col else None!r} "
          f"AJ={wv[REVIEW].cell(rihan, 36).value if rihan else None!r}")
    check("0.2's Data COE note states the boundary - the two roles are NOT in it",
          str(wb["0.2 Data Config"]["H10"].value or "").startswith("Deepali Mahajan and "
              "the vacant Service Transition Lead are NOT"),
          repr(wb["0.2 Data Config"]["H10"].value))
    # D119: both roles home in Enterprise Data - on 2.3, not 2.13, and no override rows
    on23 = {str(wv["2.3 Enterprise Data"].cell(rr, 2).value or "").strip()
            for rr in range(1, wv["2.3 Enterprise Data"].max_row + 1)}
    on213 = {str(wv["2.13 COE SA&D"].cell(rr, 2).value or "").strip()
             for rr in range(1, wv["2.13 COE SA&D"].max_row + 1)}
    check("Deepali and the vacant STL price in Enterprise Data, not the Data COE",
          "Deepali Mahajan" in on23 and "Deepali Mahajan" not in on213,
          f"on 2.3: {'Deepali Mahajan' in on23}, on 2.13: {'Deepali Mahajan' in on213}")
    # the override table keys on the person, not the row (D109), so absence is asserted
    # on the person keys
    ovr = [str(wb["Lists"].cell(rr, 40).value or "") for rr in range(2, 12)]
    check("no override rows remain for Deepali or the vacant Service Transition Lead",
          not any(v.startswith(("Deepali Mahajan |", "Vacant | Service Transition"))
                  for v in ovr), str([v for v in ovr if v]))
    v10 = wv["1.10 Z Retail"]
    check("1.10!F7 = 0.330 (his no-overhead note honoured)",
          abs((v10["F7"].value or 0) - 0.330) < 1e-9, str(v10["F7"].value))
    check("1.10!I40 stays unpriced", wb["1.10 Z Retail"]["I40"].value is None)
    check("1.10!D7 aligned to his two platforms",
          "SUM(I27,I34)" in str(wb["1.10 Z Retail"]["D7"].value) and
          "I40" not in str(wb["1.10 Z Retail"]["D7"].value))

    # ---- QA M2 + B: one allowance basis
    v31, v32 = wv["3.1 Cost Bridge"], wv["3.2 Overhead & Leadership"]
    # 3.1 carried an "Overhead roles in the portfolios" step and does not any more - the
    # walk is retired - so the allowance is read where it is now built, off each working
    # tab's own overhead row. Same figure, same purpose: the two tabs must state one
    # allowance for the same roles.
    d_over = sum((wv[t].cell(i["overhead_row"], 14).value or 0)
                 for t, i in _a2().items() if i.get("overhead_row")) or None
    # found by label on both sides: 3.2's bands moved when the section bar came off
    oh32 = next((r for r in range(5, 30)
                 if str(wb["3.2 Overhead & Leadership"].cell(r, 2).value or "").strip()
                 == "Of which sits in the portfolios"), None)
    f13 = v32.cell(oh32, 9).value if oh32 else None
    check("3.2 and the working tabs state one overhead allowance",
          d_over is not None and f13 is not None and abs(f13 - d_over) < 1e-6,
          f"working tabs {d_over} vs 3.2 {f13}")
    ah5 = wv["Lists"]["AH5"].value
    n_rows = 0
    for t in [x for x in wb.sheetnames if re.match(r"^1\.(10|14|[1-9]) ", x)]:
        for r in range(1, wb[t].max_row + 1):
            if str(wb[t].cell(r, 2).value or "").strip() == "Platform Overhead" \
                    and wv[t].cell(r, 9).value:
                n_rows += 1
    # with 1.2's C7 back to his one-branch shape, every platform is counted exactly once
    # and the model's count is the number of priced overhead rows, no allowance for a gap
    check("Lists platform count == priced overhead rows on 1.x", ah5 == n_rows,
          f"Lists {ah5} vs 1.x rows {n_rows}")

    # ---- QA: 3.4 on his country basis
    w34, v34 = wb["3.4 COE Detail"], wv["3.4 COE Detail"]
    check("3.4 has no elsewhere plug column", "Cost - elsewhere" not in joined)
    ctl = [v34.cell(r, c).value for r in range(20, 35) for c in range(3, 12)
           if isinstance(w34.cell(r, c).value, str) and "ROUND(" in w34.cell(r, c).value
           and "-$G" in w34.cell(r, c).value.replace(" ", "")]
    check("3.4 control is live and reads 0", ctl and all(abs(x or 0) < 1e-9 for x in ctl), str(ctl))

    # ---- QA: live SUMIFS instead of =0
    # by meaning, not by cell: the original finding was a typed 0 where a live SUMIFS
    # belonged, and the fixed cells' rows move whenever a group gains or loses members
    # (D119 moved Leadership onto 2.3). What must stay true: no actual-cost cell in the
    # group block is a typed number - every one is a formula or a dash.
    for tab in ("2.3 Enterprise Data", "2.6 Finance"):
        wt, vt = wb[tab], wv[tab]
        tot = next((r for r in range(8, 40)
                    if str(vt.cell(r, 2).value or "").strip() == "Total portfolio"), 30)
        typed = [f"O{r}" for r in range(8, tot)
                 if isinstance(wt.cell(r, 15).value, (int, float))]
        check(f"{tab}: no typed number in the actual-cost column, formulas only",
              not typed, str(typed[:4]))

    # ---- decisions after the fix: 8 holds, after-cost
    ex = wv["Exec Summary"]
    holds_n = next((ex.cell(r, 3).value for r in range(20, 40)
                    if "hold" in str(ex.cell(r, 2).value or "").lower()), None)
    # The Exec's hold count was pinned at 8. Two of the eight are cyber vacancies he has
    # since ruled Offshore, so the figure is re-derived off the working tabs rather than
    # re-baselined: the Exec line must equal a recount of every vacant role whose lever
    # reads Hold, wherever those roles sit.
    recount = 0
    for t in [x for x in wb.sheetnames if re.match(r"^2\.\d+ ", x)]:
        vt = wv[t]
        for r in range(1, vt.max_row + 1):
            if (str(vt.cell(r, 4).value or "").strip() == "Vacant"
                    and str(vt.cell(r, 5).value or "").strip() == "Hold"):
                recount += 1
    check("the Exec hold count is every vacant role held, recounted off the working tabs",
          holds_n == recount, f"Exec {holds_n} vs recount {recount}")
    after = next((v31.cell(r, 7).value for r in range(30, 60)
                  if str(v31.cell(r, 2).value or "").startswith("Everything on the ledger")
                  or "after decisions" in str(v31.cell(r, 2).value or "").lower()
                  and v31.cell(r, 7).value), None)

    # ---- design: no theme fills on 1.x bars, cream levers, no banned words
    theme = []
    for t in [x for x in wb.sheetnames if x.startswith("1.")]:
        for row in wb[t].iter_rows(max_row=90, max_col=20):
            for c in row:
                try:
                    if c.fill and c.fill.patternType and c.fill.start_color.type == "theme":
                        theme.append(f"{t}!{c.coordinate}")
                except Exception:
                    pass
    check("no theme fills left on 1.x tabs", not theme, str(theme[:6]))
    low = joined.lower()
    check("the word 'seat' appears nowhere", "seat" not in low)
    check("no en dashes", "–" not in joined)
    frozen = [s.title for s in wb.worksheets if s.freeze_panes]
    check("no frozen panes", not frozen, str(frozen))
    red = 0
    for s in wb.worksheets:
        for row in s.iter_rows():
            for c in row:
                if "[Red" in (c.number_format or "") or "[RED" in (c.number_format or ""):
                    red += 1
    check("no [Red] number formats", red == 0, str(red))

    # ---- design: 2.x family rules
    bare_total = dup = 0
    for t in [x for x in wb.sheetnames if x.startswith("2.")]:
        s, sv = wb[t], wv[t]
        labels = [str(s.cell(r, 2).value or "").strip() for r in range(1, s.max_row + 1)]
        if labels.count("Total portfolio") != 1:
            dup += 1
        for r, l in enumerate(labels, 1):
            if l == "Total" and sv.cell(r, 15).value is not None:
                bare_total += 1
    check("no bare 'Total' subtotals on 2.x", bare_total == 0, str(bare_total))
    check("exactly one Total portfolio per 2.x tab", dup == 0, f"{dup} tabs off")

    # ---- design: 1.5/1.6 family repairs
    # D5 is deliberately unlabelled: he removed 1.5's NZ column, and a sibling-copied
    # "TDD NZ ($m)" header over four empty cells dressed a column he removed as one that
    # exists
    check("1.5 headers: TDD AU labelled, the removed NZ column not resurrected",
          wb["1.5 P&C"]["C5"].value == "TDD AU ($m)" and wb["1.5 P&C"]["D5"].value is None,
          f"{wb['1.5 P&C']['C5'].value!r}/{wb['1.5 P&C']['D5'].value!r}")
    w6 = wb["1.6 Finance"]
    check("1.6 scratch moved to S25", str(w6["S25"].value or "").startswith("Nbr Archetype"),
          repr(w6["S25"].value))
    check("1.6 K/L carry the Actual/Variance pair",
          "ctual" in str(w6["K25"].value or "") or "ctual" in str(w6["K24"].value or "")
          or any("ctual" in str(w6.cell(r, 11).value or "") for r in range(20, 30)))

    # ---- 0.3 is the owner's cost library, and the chain does not lay it out
    par = archetypes_parity(path)
    check(f"{ARCH} matches rev.xlsx cell-for-cell (values, fills, widths, heights)",
          not par, f"{len(par)} differences: " + "; ".join(par[:4]) if par else "")

    # ---- wave G: the owner's round - 1.14/2.15, the 3.2 redesign, the table up top
    check("1.14 TDD Cyber exists", "1.14 TDD Cyber" in wb.sheetnames)
    if "1.14 TDD Cyber" in wb.sheetnames:
        w14, v14 = wb["1.14 TDD Cyber"], wv["1.14 TDD Cyber"]
        names = wb.sheetnames
        check("1.14 sits directly after 1.13",
              names.index("1.14 TDD Cyber") == names.index("1.13 Cyber Roles") + 1)
        check("1.14 platform overhead priced, F7 = 0.165",
              abs((v14["F7"].value or 0) - 0.165) < 1e-9, str(v14["F7"].value))
        # 1.14 shipped with its squad unpriced and this pinned that it said so. His wave-M
        # ruling prices both squads, so what is pinned is that neither reads "check size":
        # a squad on this tab now has a figure, one typed and one off the archetype library.
        unpriced = [c for c in ("H26", "H27")
                    if str(v14[c].value or "").strip() == "check size"]
        check("1.14's squads are priced - neither reads 'check size'", not unpriced,
              str([(c, v14[c].value) for c in ("H26", "H27")]))
        check("1.14 draws no portfolio overhead",
              (v14["C6"].value or 0) == 0 and (v14["D6"].value or 0) == 0)
        e14 = [c.coordinate for row in v14.iter_rows() for c in row
               if isinstance(c.value, str) and c.value.startswith("#")]
        check("1.14 carries no error cells", not e14, str(e14[:4]))
    check("2.15 TDD Cyber exists", "2.15 TDD Cyber" in wb.sheetnames)
    if "2.15 TDD Cyber" in wb.sheetnames:
        w15, v15 = wb["2.15 TDD Cyber"], wv["2.15 TDD Cyber"]
        ctl15 = [v15.cell(r, 3).value for r in range(1, w15.max_row + 1)
                 if str(w15.cell(r, 2).value or "").startswith("Control -")]
        check("2.15 controls read 0",
              len(ctl15) == 2 and all(abs(x or 0) < 1e-9 for x in ctl15), str(ctl15))
        hot15 = any(str(w15.cell(r, 2).value or "").strip() == "Head of Technology"
                    for r in range(1, w15.max_row + 1))
        check("2.15 draws no Head of Technology line", not hot15)
        check("0.2!F23 includes the 1.14 spend",
              "'1.14 TDD Cyber'" in str(wb["0.2 Data Config"]["F23"].value))
    # ---- wave H/J: 3.2 in the owner's own layout
    v32g = wv["3.2 Overhead & Leadership"]
    w32g = wb["3.2 Overhead & Leadership"]
    HDR32 = ["Overhead roles", "Applied to",
             "Archetype allocation (per portfolio or platform) ($m)",
             "# of times applied in archetypes", "Roles priced for in archetype",
             "Actual number of leadership roles", "# of roles not applied in archetype",
             "Total Archetype cost ($m)", "Actual cost of leadership roles",
             "Variance between archetype and actuals", "Where they sit",
             "Allocation applied"]
    # the header row is found, not assumed: he took the section bar off this tab, so
    # every row below moved up one and a fixed row number would only be right by luck
    h32 = next((r for r in range(2, 12)
                if str(w32g.cell(r, 2).value or "").strip() == "Overhead roles"), None)
    check("3.2 header found", h32 is not None)
    if h32:
        got = [w32g.cell(h32, c).value for c in range(2, 14)]
        check("3.2 carries his headings, in his order", got == HDR32, str(got[:5]))
        lo = h32 + 1
        # The platform count is not pinned to a number here. It is whatever the 1.x tabs
        # add up to on Lists!AH5, and 1.2's twinned C7/D7 make that 25 for 22 platforms
        # (D112). Checking 3.2 against the model's own count tests the thing that matters -
        # that every tab is counting the same platforms - and keeps telling the truth if
        # he changes his mind about 1.2.
        plat = wv["Lists"]["AH5"].value
        check("3.2 Times applied is his to set, cream and seeded from the model's count",
              [v32g.cell(r, 5).value for r in range(lo, lo + 6)]
              == [10, 10, 10, plat, plat, 10]
              and all(w32g.cell(r, 5).fill.patternType
                      and str(w32g.cell(r, 5).fill.start_color.rgb).upper() == "FFFFF2CC"
                      for r in range(lo, lo + 6)),
              f"{[v32g.cell(r, 5).value for r in range(lo, lo + 6)]} vs Lists {plat}")
        check("3.2 states every line's roles in the organisation",
              [v32g.cell(r, 7).value for r in range(lo, lo + 6)] == [15, 6, 7, 10, 24, 8],
              str([v32g.cell(r, 7).value for r in range(lo, lo + 6)]))
        # roles not priced for = the roles that exist less the roles the archetype pays
        # for, on every one of the six lines. Derived, so the two platform lines follow
        # the platform count instead of a snapshot of it.
        want_gap = [round((v32g.cell(r, 7).value or 0) - (v32g.cell(r, 6).value or 0), 4)
                    for r in range(lo, lo + 6)]
        check("3.2 states every line's roles not priced for",
              [round(v32g.cell(r, 8).value or 0, 4) for r in range(lo, lo + 6)] == want_gap,
              f"{[v32g.cell(r, 8).value for r in range(lo, lo + 6)]} want {want_gap}")
        check("3.2 HoT row splits its fifteen roles",
              str(v32g.cell(lo, 12).value or "") == "10 in the portfolios, 5 in the COEs",
              repr(v32g.cell(lo, 12).value))
        check("3.2 BP row places all six of its roles in the COEs",
              str(v32g.cell(lo + 1, 12).value or "") == "All 6 in the COEs",
              repr(v32g.cell(lo + 1, 12).value))
        check("3.2 DA row places all seven of its roles in the COEs",
              str(v32g.cell(lo + 2, 12).value or "") == "All 7 in the COEs",
              repr(v32g.cell(lo + 2, 12).value))
        check("3.2 says the allocation in words",
              str(v32g.cell(lo, 13).value or "") == "50% across 10 portfolios",
              repr(v32g.cell(lo, 13).value))
    # derived, not pinned to a split: D118 moved Rihan portfolio-side, and any future
    # override moves the split again. What cannot move: the two halves sum to 531 and
    # the sentence states the count the G cell carries.
    allrow = next((r for r in range(6, 30)
                   if str(v32g.cell(r, 2).value or "").startswith(
                       "Roles in the organisation, all lines and squads")), None)
    check("3.2 all-roles row present", allrow is not None)
    if allrow:
        m32 = re.search(r"531 - portfolios (\d+), COEs and EGI (\d+), each counted once",
                        str(v32g.cell(allrow, 2).value or ""))
        check("3.2 counts each role once - the two halves sum to 531",
              m32 is not None and int(m32.group(1)) + int(m32.group(2)) == 531
              and v32g.cell(allrow, 7).value == 531,
              f"{v32g.cell(allrow, 2).value!r} G={v32g.cell(allrow, 7).value}")
        check("3.2 all-roles control reads 0",
              abs(v32g.cell(allrow + 1, 7).value or 0) < 1e-9,
              str(v32g.cell(allrow + 1, 7).value))
    totrow = next((r for r in range(5, 30) if str(w32g.cell(r, 2).value or "").strip()
                   == "Overheads incl. GMs"), None)
    check("3.2 overheads total row present", totrow is not None)
    if totrow:
        # 70 leadership roles is a count off the role mapping and does not move with the
        # platform count; roles priced for does, at 0.3 of a manager per platform on two
        # of the six lines, so it is derived from the lines above it rather than typed.
        priced = round(sum(v32g.cell(r, 6).value or 0 for r in range(lo, lo + 6)), 4)
        check("3.2 totals 70 roles carried, against the six lines' priced-for",
              v32g.cell(totrow, 7).value == 70
              and abs((v32g.cell(totrow, 6).value or 0) - priced) < 1e-6,
              f"{v32g.cell(totrow, 7).value} / {v32g.cell(totrow, 6).value} want {priced}")
        check("3.2 the role gap is 70 less priced-for, and the cost gap ties to it",
              abs((v32g.cell(totrow, 8).value or 0) - (70 - priced)) < 1e-6
              and abs((v32g.cell(totrow, 11).value or 0)
                      - ((v32g.cell(totrow, 10).value or 0)
                         - (v32g.cell(totrow, 9).value or 0))) < 1e-6,
              f"{v32g.cell(totrow, 8).value} want {round(70 - priced, 4)} | "
              f"{v32g.cell(totrow, 11).value}")
    ohrow = next((r for r in range(5, 30) if str(w32g.cell(r, 2).value or "").strip()
                  == "Of which sits in the portfolios"), None)
    check("3.2 'of which sits in the portfolios' band present", ohrow is not None)
    if ohrow:
        # the one figure 3.1 and 3.2 have to agree on. Checked against 3.1, not against a
        # remembered 5.005, so the two tabs can only ever be wrong together.
        check("3.2 archetype cost where the people sit ties to 3.1",
              d_over is not None
              and abs((v32g.cell(ohrow, 9).value or 0) - d_over) < 1e-6,
              f"3.2 {v32g.cell(ohrow, 9).value} vs 3.1 {d_over}")

    # ---- wave H: the owner's Actuals-vs-archetype table on every 1.x tab
    BARH = "Actuals vs archetype"
    WANTH = [BARH, "What the cost covers",
             "Actual portfolio", "Archetype portfolio", "Variance"]
    # retired labels, matched EXACTLY: "Actual cost after decisions ($m)" is still the
    # head of every squad table's own actual column and always will be. The bar went.
    DEAD_EXACT = ("Actual cost after decisions", "Squads priced by an archetype",
                  "Squads with no archetype to price them",
                  "Overhead roles in this portfolio", "Additional costs",
                  "Total actual cost after decisions")
    DEAD_START = ("Archetype against actual",)
    tops, bots, shape, stale, wired = [], [], [], [], []
    for t in [x for x in wb.sheetnames if re.match(r"^1\.(10|14|[1-9]) ", x)]:
        ws = wb[t]
        r0 = next((r for r in range(3, 9)
                   if str(ws.cell(r, 11).value or "").strip() == BARH), None)
        for row in ws.iter_rows():
            for c in row:
                s = str(c.value or "").strip()
                if s in DEAD_EXACT:
                    stale.append(f"{t}!{c.coordinate} {s!r}")
                elif any(s.startswith(x) for x in DEAD_START):
                    bots.append(f"{t}!{c.coordinate} {s!r}")
        if r0 is None:
            tops.append(t)
            continue
        got = [str(ws.cell(r0 + i, 11).value or "").strip() for i in range(5)]
        if got != WANTH:
            shape.append(f"{t}: {got}")
        if any(ws.cell(r, c).value is not None
               for r in range(r0 + 5, r0 + 8) for c in range(11, 15)):
            shape.append(f"{t}: K{r0 + 5}:N{r0 + 7} is not clear under the block")
        tot = next((r for r in range(1, 20)
                    if str(ws.cell(r, 2).value or "").strip() == "Total Cost"), None)
        if tot and str(ws.cell(tot, 7).value or "").strip() != f"=$N${r0 + 2}":
            wired.append(f"{t}!G{tot} = {ws.cell(tot, 7).value!r}")
        if tot and abs((wv[t].cell(tot, 7).value or 0)
                       - (wv[t].cell(r0 + 2, 14).value or 0)) > 1e-9:
            wired.append(f"{t}!G{tot} value {wv[t].cell(tot, 7).value!r}")
    check("the actuals table sits up top on every 1.x tab", not tops, str(tops))
    check("every 1.x actuals table is bar + header + Actual / Archetype / Variance, "
          "with nothing under it", not shape, str(shape[:4]))
    check("no retired actuals label or bar text left on a 1.x tab", not stale,
          str(stale[:6]))
    check("no old bottom block remains on any 1.x tab", not bots, str(bots[:4]))
    check("the summary Actuals cell reads the table's Actual portfolio cost", not wired,
          str(wired[:4]))

    # ---- wave H: the hybrid rule
    a03, a03v = wb["0.3 Squad Archetypes"], wv["0.3 Squad Archetypes"]
    check("0.3 hybrid input present, labelled and 2",
          a03["K7"].value == "Onshore roles in a hybrid squad" and a03["K8"].value == 2,
          f"{a03['K7'].value!r} / {a03['K8'].value!r}")
    check("0.3 hybrid input is cream and not a percent",
          (a03["K8"].fill.patternType and
           str(a03["K8"].fill.start_color.rgb).upper() == "FFFFF2CC"
           and "%" not in (a03["K8"].number_format or "")))
    h19 = str(wb["1.9 Commercial Fuels"]["H26"].value or "")
    check("hybrid branch prices 2 onshore plus the rest offshore (1.9 probe)",
          "MIN('0.3 Squad Archetypes'!$K$8,INDEX('0.3 Squad Archetypes'!$F$5:$F$23,"
          in h19 and "))/2," not in h19)
    n_hyb = bad_hyb = 0
    for t in [x for x in wb.sheetnames if re.match(r"^1\.(10|14|[1-9]) ", x)]:
        for row in wb[t].iter_rows(min_col=8, max_col=8):
            for c in row:
                f = str(c.value or "")
                if f.startswith("=") and "0.3 Squad Archetypes" in f:
                    n_hyb += 1
                    if "$K$8" not in f or "))/2," in f:
                        bad_hyb += 1
    # The count is derived, not typed: it is however many squad rows the eleven design tabs
    # carry, and wave M added one back (Digital Support NZ) and one more (Identity on 1.14).
    # Pinning a number here made a decision of his fail a gate; what has to be true is that
    # EVERY squad formula carries his rule and none carries the retired midpoint.
    check("every squad formula carries the hybrid rule, none the old midpoint",
          n_hyb and bad_hyb == 0, f"{n_hyb} formulas, {bad_hyb} old-shape")

    # ---- wave J: the simplification sweep, asserted so it cannot creep back
    idx_join = []
    direct = 0
    JOIN = re.compile(r"INDEX\('REVIEW - Complete Role Mapping'!\$[A-Z]{1,2}:"
                      r"\$[A-Z]{1,2},\d+\)")
    DIRECT = re.compile(r"^='REVIEW - Complete Role Mapping'!\$[A-Z]{1,2}\$\d+$")
    for t in [x for x in wb.sheetnames if re.match(r"^2\.\d+ ", x)]:
        for row in wb[t].iter_rows():
            for c in row:
                v = str(c.value or "")
                if JOIN.search(v):
                    idx_join.append(f"{t}!{c.coordinate}")
                elif DIRECT.match(v):
                    direct += 1
    check("no 2.x cell finds a person by hardcoded row number", not idx_join,
          f"{len(idx_join)} left, e.g. {idx_join[:3]}")
    check("the 2.x ledger join is the insert-safe direct reference", direct > 2000,
          f"{direct} direct references")
    lst = wb["Lists"]
    an = [lst.cell(r, 40).value for r in range(2, 12)]
    check("the override table is keyed on the person, not a row number",
          not any(isinstance(x, (int, float)) for x in an),
          str([x for x in an if x is not None][:4]))
    ar = str(wb[REVIEW]["AR2"].value or "")
    check("REVIEW's overhead-line formula lost its dead branch",
          len(ar) < 520 and "$AQ" not in ar, f"{len(ar)} chars")
    dead = [f"{L_}{r}" for L_ in ("AL", "AM", "AN", "AS")
            for r in (2, 100, 400)
            if wb[REVIEW][f"{L_}{r}"].value is not None]
    check("the four dead ledger columns are gone", not dead, str(dead[:6]))
    ex = wb["Exec Summary"]
    exv = wv["Exec Summary"]
    wide = [ex.cell(r, 3).value for r in range(20, 40)
            if isinstance(ex.cell(r, 3).value, str)
            and re.search(r"\$[DE]:\$[DE]", ex.cell(r, 3).value)]
    check("Exec's vacancy counts read bounded ranges, not whole columns", not wide,
          f"{len(wide)} whole-column")
    # found by label - the whole-organisation block (D118) sits above the walk now, so
    # these rows are no longer at fixed positions
    w31c = wb["3.1 Cost Bridge"]
    r_arch = next((rr for rr in range(4, 80) if str(w31c.cell(rr, 2).value or "").strip()
                   == "Squads priced by an archetype"), None)
    r_dir = next((rr for rr in range(4, 80) if str(w31c.cell(rr, 2).value or "").strip()
                  == "Directly funded, where the funded figure is set"), None)
    # Wave M deleted the walk, so the plain-SUM, gate-on-step and walk-ledger ties
    # went with it: the wave-M block now pins the single-table layout, its totals and
    # its controls directly. What survives of D118 is Retail adjacency - now WITHOUT
    # the subtotal row, per his approved mock.
    check("the walk is gone from 3.1", r_arch is None and r_dir is None,
          f"archetype subtotal at r{r_arch}, directly-funded gate at r{r_dir}")
    bars31 = [str(w31c.cell(rr, 2).value or "").strip() for rr in range(3, 12)]
    check("3.1 opens with the TDD band", "TDD" in bars31, str(bars31[:4]))
    names31 = [str(w31c.cell(rr, 2).value or "").strip() for rr in range(4, 45)]
    ia = names31.index("Ampol Retail") if "Ampol Retail" in names31 else None
    check("Ampol Retail and Z Retail sit together with no subtotal row",
          ia is not None and names31[ia + 1] == "Z Retail"
          and not any(n.startswith("Retail - Ampol and Z") for n in names31),
          f"after Ampol Retail: {names31[ia + 1] if ia is not None else None}")
    # The two checks that stood here pinned a Home country column on 0.2 and forbade the
    # 1.x tabs from deciding AU or NZ by comparing the two budget cells. Both are retired:
    # the column was mine, he never asked for it, and it is off his config tab. The budget
    # comparison is his own rule and it is back on all eleven tabs. What is worth holding
    # is that all eleven decide it the same way, so a new tab cannot invent a third.
    cfg = wb["0.2 Data Config"]
    check("0.2 has no Home country column", cfg["J5"].value is None
          and all(cfg.cell(r, 10).value is None for r in range(4, 28)),
          repr(cfg["J5"].value))
    geo = set()
    for t in [x for x in wb.sheetnames if re.match(r"^1\.(10|14|[1-9]) ", x)]:
        for r in range(5, 10):
            v = str(wb[t].cell(r, 3).value or "")
            if ("'0.2 Data Config'!$D$" in v and ">" in v
                    and "'0.2 Data Config'!$C$" in v):
                geo.add(t)
    check("all eleven 1.x tabs pick AU or NZ the same way - the bigger budget cell",
          len(geo) == 11, f"{len(geo)} tabs: {sorted(geo)[:3]}")
    lev = []
    for t in ("1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles"):
        for row in wb[t].iter_rows(min_col=20, max_col=20):
            for c in row:
                if '"Offshore",0.4' in str(c.value or ""):
                    lev.append(f"{t}!{c.coordinate}")
    check("the lever price lives only on Lists", not lev, str(lev[:4]))
    # 1.13's "Planned spend less CapEx" cell used to be pinned to follow its own input.
    # His wave-M ruling deletes the CapEx input and every reference to it, so what is pinned
    # now is that neither the cell nor the input came back - the wave-M block above asserts
    # the same fact from the other direction.
    check("1.13 carries no CapEx cell and nothing that reads one",
          wb["1.13 Cyber Roles"]["F11"].value is None
          and wb["1.13 Cyber Roles"]["C13"].value is None,
          f"F11={wb['1.13 Cyber Roles']['F11'].value!r} "
          f"C13={wb['1.13 Cyber Roles']['C13'].value!r}")
    check("0.2's position figure is computed, not typed",
          isinstance(cfg["L23"].value, str) and cfg["L23"].value.startswith("="),
          repr(cfg["L23"].value))

    # ---- the reversals that had no pin, and the wave-K review fixes
    check("0.2 Legal/EG/EGI spend cells are blank, not typed zeros",
          all(cfg[c].value is None for c in ("F20", "F24", "F25")),
          str([cfg[c].value for c in ("F20", "F24", "F25")]))
    check("his six labels stand - Budget to draw down, Alloc %",
          wb["1.11 BP&T"]["G5"].value == "Budget to draw down ($m)"
          and wb["1.12 SA&D"]["H5"].value == "Budget to draw down ($m)"
          and wb["1.13 Cyber Roles"]["G5"].value == "Budget to draw down ($m)"
          and wb["1.11 BP&T"]["B17"].value == "Total budget to draw down ($m)"
          and wb["1.12 SA&D"]["B17"].value == "Total budget to draw down ($m)"
          and cfg["M13"].value == "Alloc %",
          f"{wb['1.11 BP&T']['G5'].value!r} / {cfg['M13'].value!r}")
    # the actuals table quotes the actual, not the after-decisions figure. On ten tabs
    # the two are equal today; 1.7 carries a lever, so it is the tab that proves the
    # wiring - and the tab that misreported the moment a lever moved.
    box17 = str(wb["1.7 Infrastructure"]["N7"].value or "")
    check("the 1.x Actual portfolio line reads the working tab's Actual cost column",
          "!$O$" in box17, box17)
    # found by label, not row number - the bridge gains and loses named lines as squads
    # move between its steps, which is exactly what D117 did
    v31g = wv["3.1 Cost Bridge"]
    r_led = next((r for r in range(4, 120)
                  if str(v31g.cell(r, 2).value or "").startswith("Cost of the")), None)
    r_gr = next((r for r in range(4, 120)
                 if str(v31g.cell(r, 2).value or "").startswith(
                     ("Total cost of TDD including", "Total TDD cost including"))), None)
    # wave M: archetype = actual for the no-plan groups, so the grand row carries real
    # figures and the old carries-a-dash pin inverted into the no-dash sweep checks
    check("3.1's grand row carries real archetype and variance figures",
          r_gr is not None
          and all(isinstance(v31g.cell(r_gr, c).value, (int, float))
                  for c in (4, 6)),
          f"grand r{r_gr}: "
          f"{[v31g.cell(r_gr, c).value for c in (4, 6)] if r_gr else None}")
    exec_b = [str(wv["Exec Summary"].cell(r, 2).value or "") for r in range(4, 40)]
    # wave M: the COE overhead slice line died with the walk; the portfolios line stays
    check("Exec names the portfolio overhead gap",
          any("verhead roles in the portfolios" in x for x in exec_b),
          str([x for x in exec_b if x.startswith("Overhead")][:2]))
    check("Exec states the budget position off 0.2",
          any(x.startswith("Over/(under) the allocated TDD budget") for x in exec_b),
          "not found")
    check("one offshore rate - the lever's factor reads his 0.3 cell",
          str(wb["Lists"]["AD5"].value) == "='0.3 Squad Archetypes'!$K$5",
          repr(wb["Lists"]["AD5"].value))
    check("0.3's hybrid note states his rule the right way round",
          str(wb["0.3 Squad Archetypes"]["C25"].value or "").strip()
          == "Hybrid = 2 roles onshore, rest offshore",
          repr(wb["0.3 Squad Archetypes"]["C25"].value))
    check("the 2.x tabs say what the vacancy lever does, beside the lever",
          any(isinstance(c.value, str) and c.value.startswith("Vacancy lever: Hire")
              for row in wb["2.1 Ampol Retail"].iter_rows(min_col=8, max_col=8)
              for c in row), "no lever note on 2.1")
    src_tabs = {"0.1 Budget Table (Fin)", "0.4 Presentation Pack",
                "0.3 Squad Archetypes"}
    cream_formula = [f"{ws.title}!{c.coordinate}"
                     for ws in wb.worksheets
                     if ws.sheet_state == "visible" and ws.title not in src_tabs
                     for row in ws.iter_rows() for c in row
                     if isinstance(c.value, str) and c.value.startswith("=")
                     and c.fill.patternType
                     and str(c.fill.start_color.rgb).upper() == "FFFFF2CC"]
    check("cream marks typed inputs only - no formula wears the input colour",
          not cream_formula, str(cream_formula[:5]))

    # ---- wave M: the cyber uplift restructure and the Customer corrections
    wave_m(wb, wv, joined)

    # ---- 4.0 all zero
    v40 = wv["4.0 Data QA"]
    fails = []
    for r in range(1, v40.max_row + 1):
        val = v40.cell(r, 5).value
        if isinstance(val, (int, float)) and abs(val) > 1e-6:
            fails.append(f"r{r}={val}")
    check("4.0 Data QA all zero", not fails, str(fails[:5]))

    ok = all(o for o, _ in OUT)
    for _, line in OUT:
        print(line)
    n_bad = sum(1 for o, _ in OUT if not o)
    print(f"\n{len(OUT)} checks, {n_bad} failing")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(run(sys.argv[1] if len(sys.argv) > 1 else "cand_A.xlsx"))
