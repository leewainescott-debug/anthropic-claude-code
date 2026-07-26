"""Build 3.1, 3.2 and 3.3 from scratch, to docs/LAYOUT_SPEC.md.

Each tab has one job and states no fact the other two state.

  3.1  the budget story      budget, what it costs, what is left to fund
  3.2  the archetype story   archetype, actual, variance, impact, total
  3.3  the squad detail      every squad, roles and cost

They all read the 2.x working tabs, so a lever pulled on a working tab moves all three.
The old 3.1 read 128 cells from 3.2, which is why the two tabs said the same thing twice
and disagreed whenever one of them was edited.
"""
import json

import openpyxl
from openpyxl.utils import get_column_letter as L

import fmt
import wb2

REVIEW = "REVIEW - Complete Role Mapping"
REV = f"'{REVIEW}'"
LAST = 528


def clean(wb, name):
    ws = wb[name]
    wb2.wipe(ws)
    ws.column_dimensions["A"].width = 2.5
    return ws


def order(anchors):
    by = {a["pf"]: t for t, a in anchors.items()}
    return [(p, by[p], anchors[by[p]]) for p in wb2.PORTFOLIO_ORDER if p in by]


# ------------------------------------------------------------------ 3.1 budget story
H31 = ["Portfolio", "TDD lights-on budget ($m)", "Actual cost ($m)",
       "Over/(under) budget ($m)", "Total cost after decisions ($m)",
       "Left to fund after decisions ($m)"]
W31 = [30, 17, 14, 16, 17, 18]


def build_31(wb, anchors):
    ws = clean(wb, "3.1 Group Summary")
    wb2.title(ws, 2, "Group summary - budget against cost")
    wb2.title(ws, 4, "One line per portfolio. Budget is the lights-on allocation on "
                     "0.2 Data Config.", 12)
    ws.cell(4, 2).font = fmt.BODY
    HDR = 6
    fmt.header(ws, HDR, 2, H31, W31)
    r = HDR + 1
    first = r
    for pf, tab, a in order(anchors):
        t = a["total_row"]
        ws.cell(r, 2).value = pf
        ws.cell(r, 2).font = fmt.BODY
        ws.cell(r, 2).alignment = fmt.LEFT
        fmt.money(ws, r, 3, f"=IFERROR(INDEX('0.2 Data Config'!$E$6:$E$25,"
                            f"MATCH($B{r},Lists!$AL$2:$AL$19,0)),0)")
        fmt.money(ws, r, 4, f"='{tab}'!${L(wb2.S_ACTUAL)}${t}")
        fmt.money(ws, r, 5, f"=$D{r}-$C{r}")
        fmt.money(ws, r, 6, f"='{tab}'!${L(wb2.S_TOTAL)}${t}")
        fmt.money(ws, r, 7, f"=$F{r}-$C{r}")
        r += 1
    last = r - 1
    fmt.band(ws, r, 2, len(H31), fmt.TOT_FILL, line=True)
    ws.cell(r, 2).value = "Total"
    ws.cell(r, 2).alignment = fmt.LEFT
    for c in range(3, 8):
        x = ws.cell(r, c)
        x.value = f"=SUM({L(c)}{first}:{L(c)}{last})"
        x.number_format = fmt.MONEY_M
        x.alignment = fmt.RIGHT
    ws.freeze_panes = f"C{HDR + 1}"
    return {"first": first, "last": last, "total": r}


# --------------------------------------------------------------- 3.2 archetype story
H32 = ["Portfolio", "Archetype cost ($m)", "Actual cost ($m)",
       "Variance to archetype ($m)", "Impact of decisions ($m)",
       "Total cost after decisions ($m)", "New variance to archetype ($m)"]
W32 = [30, 15, 14, 16, 15, 17, 17]
H32B = ["Overhead line", "Roles", "Rate ($m)", "Units", "Allowance ($m)",
        "Actual cost ($m)", "Over/(under) allowance ($m)"]
W32B = [30, 9, 11, 9, 14, 14, 18]


def build_32(wb, anchors):
    ws = clean(wb, "3.2 Total Cost")
    wb2.title(ws, 2, "Total cost - archetype against actual")
    ws.cell(4, 2).value = ("Archetype prices delivery squads only. Overhead is priced on "
                           "Lists and is stated in the second table.")
    ws.cell(4, 2).font = fmt.BODY
    HDR = 6
    fmt.header(ws, HDR, 2, H32, W32)
    r = HDR + 1
    first = r
    for pf, tab, a in order(anchors):
        t, d = a["total_row"], a["delivery_row"]
        ws.cell(r, 2).value = pf
        ws.cell(r, 2).font = fmt.BODY
        ws.cell(r, 2).alignment = fmt.LEFT
        fmt.money(ws, r, 3, f"='{tab}'!${L(wb2.S_ACOST)}${d}")
        fmt.money(ws, r, 4, f"='{tab}'!${L(wb2.S_ACTUAL)}${t}")
        fmt.money(ws, r, 5, f"=$D{r}-$C{r}")
        fmt.money(ws, r, 6, f"='{tab}'!${L(wb2.S_IMPACT)}${t}")
        fmt.money(ws, r, 7, f"='{tab}'!${L(wb2.S_TOTAL)}${t}")
        fmt.money(ws, r, 8, f"=$G{r}-$C{r}")
        r += 1
    last = r - 1
    fmt.band(ws, r, 2, len(H32), fmt.TOT_FILL, line=True)
    ws.cell(r, 2).value = "Total"
    ws.cell(r, 2).alignment = fmt.LEFT
    for c in range(3, 9):
        x = ws.cell(r, c)
        x.value = f"=SUM({L(c)}{first}:{L(c)}{last})"
        x.number_format = fmt.MONEY_M
        x.alignment = fmt.RIGHT
    tot32 = r

    r += 2
    wb2.title(ws, r, "Overhead - allowance against actual", 12)
    r += 1
    HDR2 = r
    fmt.header(ws, HDR2, 2, H32B, W32B)
    r += 1
    f2 = r
    for i in range(2, 8):
        ws.cell(r, 2).value = f"=Lists!$AF${i}"
        ws.cell(r, 2).font = fmt.BODY
        ws.cell(r, 2).alignment = fmt.LEFT
        fmt.money(ws, r, 3,
                  f"=COUNTIFS({REV}!$AR$2:$AR${LAST},$B{r})", fmt.COUNT)
        fmt.money(ws, r, 4, f"=Lists!$AG${i}", '#,##0.000;(#,##0.000);"-"')
        fmt.money(ws, r, 5, f"=Lists!$AH${i}", fmt.COUNT)
        fmt.money(ws, r, 6, f"=Lists!$AJ${i}")
        fmt.money(ws, r, 7,
                  f"=IF($B{r}=\"Leadership - 8 GMs\",Lists!$AG$12,"
                  f"SUMIFS({REV}!$AA$2:$AA${LAST},{REV}!$AR$2:$AR${LAST},$B{r})/1000000)")
        fmt.money(ws, r, 8, f"=$G{r}-$F{r}")
        r += 1
    l2 = r - 1
    fmt.band(ws, r, 2, len(H32B), fmt.TOT_FILL, line=True)
    ws.cell(r, 2).value = "Overhead total"
    ws.cell(r, 2).alignment = fmt.LEFT
    for c, nf in ((3, fmt.COUNT), (6, fmt.MONEY_M), (7, fmt.MONEY_M), (8, fmt.MONEY_M)):
        x = ws.cell(r, c)
        x.value = f"=SUM({L(c)}{f2}:{L(c)}{l2})"
        x.number_format = nf
        x.alignment = fmt.RIGHT
    r += 2
    ws.cell(r, 2).value = ("The 8 GMs are the only overhead line with no role in the "
                           "ledger, so their cost is the input on Lists.")
    ws.cell(r, 2).font = fmt.BODY
    ws.freeze_panes = f"C{HDR + 1}"
    return {"first": first, "last": last, "total": tot32}


# ------------------------------------------------------------------- 3.3 squad detail
H33 = ["Portfolio", "Squad", "Archetype type", "Archetype size", "Archetype roles",
       "Roles", "Filled", "Vacant", "Archetype cost ($m)", "Actual cost ($m)",
       "Variance to archetype ($m)", "Total cost after decisions ($m)"]
W33 = [24, 32, 26, 12, 12, 8, 8, 8, 13, 13, 15, 17]


def build_33(wb, anchors):
    ws = clean(wb, "3.3 FTE View")
    wb2.title(ws, 2, "Squad detail - roles and cost, squad by squad")
    HDR = 5
    fmt.header(ws, HDR, 2, H33, W33)
    r = HDR + 1
    pf_rows = []
    for pf, tab, a in order(anchors):
        start = r
        for g in a["squads"] + a["overhead"]:
            s = a["srow"][g]
            ws.cell(r, 2).value = pf
            ws.cell(r, 2).font = fmt.BODY
            ws.cell(r, 2).alignment = fmt.LEFT
            for col, src, nf in (
                    (3, wb2.S_SQUAD, None), (4, wb2.S_TYPE, None),
                    (5, wb2.S_SIZE, None), (6, wb2.S_AROLES, fmt.COUNT1),
                    (7, wb2.S_ROLES, fmt.COUNT), (9, wb2.S_VAC, fmt.COUNT),
                    (10, wb2.S_ACOST, fmt.MONEY_M), (11, wb2.S_ACTUAL, fmt.MONEY_M),
                    (12, wb2.S_VAR, fmt.MONEY_M), (13, wb2.S_TOTAL, fmt.MONEY_M)):
                x = ws.cell(r, col)
                x.value = f"='{tab}'!${L(src)}${s}"
                x.font = fmt.BODY
                if nf:
                    x.number_format = nf
                    x.alignment = fmt.RIGHT
                else:
                    x.alignment = fmt.LEFT
            fmt.money(ws, r, 8, f"=$G{r}-$I{r}", fmt.COUNT)
            r += 1
        fmt.band(ws, r, 2, len(H33), fmt.SUB_FILL)
        ws.cell(r, 2).value = f"{pf} total"
        ws.cell(r, 2).alignment = fmt.LEFT
        for c in range(6, 14):
            x = ws.cell(r, c)
            x.value = f"=SUM({L(c)}{start}:{L(c)}{r - 1})"
            x.number_format = (fmt.MONEY_M if c >= 10 else
                               (fmt.COUNT1 if c == 6 else fmt.COUNT))
            x.alignment = fmt.RIGHT
        pf_rows.append(r)
        r += 1
    fmt.band(ws, r, 2, len(H33), fmt.TOT_FILL, line=True)
    ws.cell(r, 2).value = "Group total"
    ws.cell(r, 2).alignment = fmt.LEFT
    for c in range(6, 14):
        x = ws.cell(r, c)
        x.value = "=" + "+".join(f"{L(c)}{p}" for p in pf_rows)
        x.number_format = (fmt.MONEY_M if c >= 10 else
                           (fmt.COUNT1 if c == 6 else fmt.COUNT))
        x.alignment = fmt.RIGHT
    grp = r
    r += 2
    ws.cell(r, 2).value = "Control - roles against the ledger, must be 0"
    ws.cell(r, 2).font = fmt.BODY
    fmt.money(ws, r, 7,
              f"=$G${grp}-COUNTA({REV}!$B$2:$B${LAST})", fmt.COUNT)
    ws.cell(r + 1, 2).value = "Control - cost against the ledger ($m), must be 0"
    ws.cell(r + 1, 2).font = fmt.BODY
    fmt.money(ws, r + 1, 11,
              f"=ROUND($K${grp}-SUM({REV}!$AA$2:$AA${LAST})/1000000,6)")
    ws.freeze_panes = f"D{HDR + 1}"
    return {"group_total": grp}


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    anchors = json.load(open("anchors2.json"))
    a31 = build_31(wb, anchors)
    a32 = build_32(wb, anchors)
    a33 = build_33(wb, anchors)
    json.dump({"3.1": a31, "3.2": a32, "3.3": a33}, open("anchors3.json", "w"), indent=1)
    wb.save(dst)
    return [f"3.1 Group Summary: {a31['last'] - a31['first'] + 1} portfolios, "
            f"total row {a31['total']}",
            f"3.2 Total Cost: archetype table plus the overhead allowance block",
            f"3.3 Squad Detail: every squad with roles and cost, "
            f"group total row {a33['group_total']}"]


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
