"""3.4 COE detail and 3.5 Source Reconciliation, on the same layout as the rest of 3.x.

3.4 carried two definitions of the same number - column F was literally `=$K6`, headed
"People cost, gross ($m)" against K's "Gross people cost ($m)" - and priced the COEs
against a "budget to draw down" the owner ruled out: 12.00 against 27.77 actual is not a
comparison, it is two unrelated figures on one row.

What 3.4 is for now is the thing no other tab shows: where COE cost sits geographically,
and which COE roles carry an overhead title. It reads the same squad grouping as every
other tab rather than a second cut built off the department column.

3.5 Source Reconciliation is gone. Its only job was reconciling the retired Squads tab to
REVIEW, and REVIEW is the only source of truth, so there is nothing to reconcile it against.
"""
import json

import openpyxl
from openpyxl.utils import get_column_letter as L

import final2x as f2
import opts

REV = f2.REV
LAST = None                             # refreshed from f2 after _boot_last
S = f2.S
COE_ORDER = ["COE Cyber", "COE BP&T", "COE SA&D", "EGI"]

# AU is every role whose country is not NZ, which is how the owner's own COE tabs define
# it: 1.11, 1.12 and 1.13 all split their planned spend on Country <> "NZ" against Country
# = "NZ". This tab used to read Country = "Australia" for AU and the derived AU/NZ column
# for NZ, and carry a third "Cost - elsewhere" column holding the remainder as a plug -
# which put Neil Reilly (REVIEW 264, Singapore, 666,087.50) in "elsewhere" here and in AU
# on 1.11, so two tabs published two different AU figures for the same group. Worse, the
# control under them summed AU plus NZ plus a plug against the cost the plug was derived
# from, so it was zero by construction and could never fail.
#
# Two columns now, on the owner's definition, with nothing between them and the total. The
# control is AU plus NZ against cost, which is a check: the two SUMIFS have to partition
# the same rows the cost column adds up, and if either criterion or either range drifts it
# says so.
H34 = ["COE", "Squad", "Roles", "Filled", "Vacant", "Cost ($m)", "Cost - AU ($m)",
       "Cost - NZ ($m)", "Roles carrying an overhead title", "Cost of those roles ($m)"]
# the Squad column has to hold the longest squad name on a COE tab - "Technology Strategy
# & AI Capability" is 35 characters and the column was 34, so it lost its last letter with
# a roles count in the cell beside it
W34 = [16, 37, 8, 8, 8, 13, 13, 13, 15, 15]
F34 = [None, None, opts.CT, opts.CT, opts.CT, opts.M2, opts.M2, opts.M2,
       opts.CT, opts.M2]
SUMS = (4, 5, 6, 7, 8, 9, 10, 11)


def build_34(wb, anchors):
    ws = wb["3.4 COE Summary"]
    f2.wipe(ws)
    ws.column_dimensions["A"].width = 2
    ws.cell(2, 2).value = "COE detail - where the cost sits"
    ws.cell(2, 2).font = opts.TITLE
    by = {a["pf"]: a for a in anchors.values()}
    r = opts.bar(ws, 4, 2, len(H34), "The COEs and EGI, squad by squad")
    r = opts.head(ws, r, 2, H34, W34)
    subs = []
    for pf in COE_ORDER:
        a = by[pf]
        st = r
        for sq in a["direct"] + a["squads"] + a["overhead"]:
            base = (f'{REV}!$AA$2:$AA${LAST},{REV}!$AJ$2:$AJ${LAST},$B{r},'
                    f"{REV}!$AT$2:$AT${LAST},$C{r}")
            cnt = (f'{REV}!$AJ$2:$AJ${LAST},$B{r},{REV}!$AT$2:$AT${LAST},$C{r}')
            for c, v in ((2, pf), (3, sq)):
                x = ws.cell(r, c)
                x.value, x.font, x.alignment = v, opts.BODY, opts.LFT
            f2._m(ws, r, 4, f"=COUNTIFS({cnt})", opts.CT)
            f2._m(ws, r, 5, f'=COUNTIFS({cnt},{REV}!$AK$2:$AK${LAST},"Filled")', opts.CT)
            f2._m(ws, r, 6, f'=COUNTIFS({cnt},{REV}!$AK$2:$AK${LAST},"Vacant")', opts.CT)
            f2._m(ws, r, 7, f"=SUMIFS({base})/1000000")
            # the owner's two-way split, off the Country column both sides, so the pair
            # covers every role in the group exactly once
            f2._m(ws, r, 8, f'=SUMIFS({base},{REV}!$M$2:$M${LAST},"<>NZ")/1000000')
            f2._m(ws, r, 9, f'=SUMIFS({base},{REV}!$M$2:$M${LAST},"NZ")/1000000')
            f2._m(ws, r, 10,
                  f'=COUNTIFS({cnt},{REV}!$AR$2:$AR${LAST},"<>Squad")', opts.CT)
            f2._m(ws, r, 11,
                  f'=SUMIFS({base},{REV}!$AR$2:$AR${LAST},"<>Squad")/1000000')
            r += 1
        opts.row(ws, r, 2, [f"{pf} total"] + [None] * (len(H34) - 1),
                 [None] * len(H34), bg=opts.GREY, bold=True)
        ws.cell(r, 2).alignment = opts.LFT
        for c in SUMS:
            x = ws.cell(r, c)
            x.value = f"=SUM({L(c)}{st}:{L(c)}{r-1})"
            x.number_format, x.alignment = F34[c - 2], opts.RGT
        subs.append(r)
        r += 1
    opts.row(ws, r, 2, ["COEs and EGI total"] + [None] * (len(H34) - 1),
             [None] * len(H34), bg=opts.MID, bold=True, top=True)
    ws.cell(r, 2).alignment = opts.LFT
    for c in SUMS:
        x = ws.cell(r, c)
        x.value = "=" + "+".join(f"{L(c)}{p}" for p in subs)
        x.number_format, x.alignment = F34[c - 2], opts.RGT
    gt = r
    r += 2
    # The four COE names as one array constant, so each control is one call over four
    # criteria rather than four calls spelled out and added. The cost control was the
    # worst of them - four SUMIFS, identical apart from a name, on one line - and the
    # reader had to check all four to see that the set was complete. It is COE_ORDER
    # itself now, which is the list the tab above was built from, so the control and the
    # table cannot come to cover different groups.
    coes = "{" + ",".join(f'"{p}"' for p in COE_ORDER) + "}"
    for lab, col, f, nf in (
            ("Control - roles against the ledger, must be 0", 4,
             f'=$D{gt}-SUMPRODUCT(COUNTIFS({REV}!$AJ$2:$AJ${LAST},{coes}))',
             opts.CTL_C),
            ("Control - cost against the ledger ($m), must be 0", 7,
             f"=ROUND($G{gt}-SUMPRODUCT(SUMIFS({REV}!$AA$2:$AA${LAST},"
             f"{REV}!$AJ$2:$AJ${LAST},{coes}))/1000000,6)", opts.CTL_M),
            ("Control - AU plus NZ against cost ($m), must be 0", 9,
             f"=ROUND($H{gt}+$I{gt}-$G{gt},6)", opts.CTL_M)):
        ws.cell(r, 2).value = lab
        ws.cell(r, 2).font = opts.BODY
        ws.cell(r, 2).alignment = opts.LFT
        f2._m(ws, r, col, f, nf)
        r += 1
    ws.freeze_panes = "D6"
    return {"total": gt}


H35 = ["Portfolio", "Squads roles", "Squads filled", "Squads vacant", "REVIEW roles",
       "REVIEW filled", "REVIEW vacant", "Difference - roles", "Difference - filled",
       "Difference - vacant"]
W35 = [34, 12, 12, 12, 12, 12, 12, 13, 13, 13]


def build_35(wb):
    """Relabel and restyle. The population logic is left exactly as it is - it is the
    evidence that the retired Squads tab was replaced by REVIEW, and rewriting it would
    destroy the thing it exists to prove."""
    ws = wb["3.5 Source Reconciliation"]
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.cell(2, 2).value = "Source Reconciliation - the retired Squads tab against REVIEW"
    ws.cell(2, 2).font = opts.TITLE
    ws.cell(3, 2).value = None                       # a sentence in a data area
    opts.bar(ws, 4, 2, len(H35), "Where the two sources disagree, role by role")
    opts.head(ws, 5, 2, H35, W35)
    last = 20
    for r in range(6, last + 1):
        band = opts.GREY if r % 2 == 1 else None
        for c in range(2, 12):
            x = ws.cell(r, c)
            x.border = opts.BOX
            x.font = opts.BOLD if r == last else opts.BODY
            x.alignment = opts.LFT if c == 2 else opts.RGT
            if c > 2:
                x.number_format = opts.CT
            if r == last:
                x.fill = opts.fl(opts.MID)
            elif band:
                x.fill = opts.fl(band)
    ws.freeze_panes = "C6"
    return last


def run(src, dst):
    global LAST
    f2._boot_last(src)
    LAST = f2.LAST
    wb = openpyxl.load_workbook(src)
    anchors = json.load(open("anchors_final.json"))
    a34 = build_34(wb, anchors)
    wb.save(dst)
    return [f"3.4 COE detail rebuilt: squad by squad, AU / NZ on the owner's own basis "
            f"(country <> NZ), overhead titles, total row {a34['total']}"]


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
