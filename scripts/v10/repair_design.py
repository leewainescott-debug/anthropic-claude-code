"""Re-apply the design-tab fixes the review workbook's branch predates.

rev.xlsx forked before two documented rounds of work, so its 1.x tabs still carry:

  the pre-D4 squad names - the ledger is the source of truth and the design tabs were
  renamed to follow it ("Z Energy Apps" is the ledger's "Z App and Web", and so on);
  the D6a zero-people squads - "We cannot have squads with 0 people in them" (owner) -
  which charge archetype cost against nobody. The new Customer dataset still has no
  people in any of them, so the rule still holds.

Renaming a row keeps every figure the owner typed on it, which is the point: his squad
cost for "Z Energy Martech" belongs to the squad the ledger calls "Z Loyalty & Martech",
and the rename carries it there. A rename now follows the name onto the two rows that
frame the block - "Platform: {name}" above it and "{name} Total" below - because matching
column B exactly left 1.1 with a platform bar and a total row still calling the squad
"Network / QSR" over a squad row renamed "Network & QSR".

A removed row loses its figures, not his writing: anything he typed in K..N moves along the
row to M and stays. Digital Support NZ used to be on the removal list and came off it in
wave M on his ruling - the squad is on his own review workbook and its 0.32 of NZ platform
cost is his - so the row stays and only his note moves, out of what is now a live column.

Also here: the one-cell typo his 1.2 edit introduced (platform overhead summing the
Total Squad Cost column instead of the TDD Cost column).

--- the "formulas that make no sense" round -------------------------------------------

Three more things on the design side that computed the right number for the wrong reason.

--- wave M -----------------------------------------------------------------------------

His three rulings that land on the tabs this file owns: 0.2's combined cyber line splits in
two, one row for the COE and one for TDD Cyber, each reading its own tab; the EGI squads are
funded by EGI at the actual cost of their roles, so their Total Squad Cost reads the working
tab and their support % is nil; and Digital Support NZ comes back.


LEFT TO FUND.  1.1's "Left to fund" enumerated the funding lines it subtracts -
`=SUM(E9-J15-J16-J17-J18)` - where its nine siblings all say `=E9-(J19-J14)`: the applied
total, less the Lights On line, which funds people and not the Other column. Same answer
(0.173), but the enumerated form silently misses any line inserted into the block and
wraps a subtraction in a pointless SUM(). It now reads the way the other nine read.

1.13's CYBER CAPEX was repaired here to follow its own input instead of a typed 0.5. His
wave-M ruling deletes the CapEx input from 1.13 altogether, so there is nothing left to
follow: fixcoe.py rebuilds that block without it and RETIRED.md carries the old shape.

0.2!L23.  A typed 50.5 sitting beside a computed 50.5 (E26), in a reconciliation table
whose next cell along, M23, already reads its total from F26. It reads E26 now.

The H-vs-I asymmetry on the platform total rows - H stops at the last squad row, I runs
one further to take in the Platform Overhead line - was checked across all 29 total rows
on all fourteen tabs and is right, consistent and deliberate: an overhead is not a squad
and has no archetype price, so it exists in the TDD Cost column only, and the H cell on
every overhead row is empty. Those total rows carry a cell comment saying so - written in
fixcoe.py, after the empty platform blocks are collapsed - because it reads like a typo
and it is not.
"""
import copy
import re

import openpyxl

RENAME = {
    "1.1 Ampol Retail": {"Network / QSR": "Network & QSR"},
    "1.2 Customer": {"Z Energy Apps": "Z App and Web",
                     "Z Energy Martech": "Z Loyalty & Martech",
                     "AU CRM & Martech": "Ampol Loyalty & Martech",
                     "Customer AI": "Customer, AI"},
    "1.4 TDD Group Functions": {"Network & Infrastructure": "Cloud, Network & Infra Ops",
                                "DevOps & Engineering": "DevOps & QE",
                                "Integration": "Integration & Process Automation"},
    "1.5 P&C": {"P&C - RTA": "P&C RTA"},
    "1.6 Finance": {"AU Finance": "SAP ERP"},
    "1.7 Infrastructure": {"Manufacturing & Group Projects": "Manufacturing Group Projects"},
}
# D6a: design rows with nobody in them, confirmed still empty in the new dataset.
#
# Digital Support NZ came off this list in wave M on his ruling. It is a squad on his own
# review workbook (row 41, "Net New" in the gutter beside it), it is the NZ half of the
# Customer story, and removing it took 0.32 of NZ platform cost off 1.2 - which is what
# put Z Customer at 4.99 against his own review workbook's 5.314. D6a still holds for a
# squad with no people AND no line of his behind it; this one has his line behind it.
REMOVE = {
    "1.3 Enterprise Data": ["EGI Data", "Enterprise Data Delivery"],
}
# A squad row of his that survives, but whose note has to move: K used to be the note
# margin on the 1.x squad tables and is now "Actual cost after decisions ($m)", a live
# column actuals.py writes. His writing outlives the column it sat in - it moves along the
# row to the note margin at M/N, the same rule REMOVE uses, and the cell it came from is
# freed for the figure that belongs there.
RESTORED = {"1.2 Customer": [("Digital Support NZ", 41)]}
# 1.2 C7. In both of his workbooks C7 and D7 are different formulas and only one fires:
# his review book's C7 sums column H, which is empty on the platform-overhead rows, and
# his 27/07 book's C7 is the exact complement of D7 (the 0 and the SUM swap branches).
# Either way F7 = 0.495 - the overhead counted once. An earlier pass here "corrected"
# C7's H columns to I, which made C7 byte-identical to D7 and both branches fire: the
# overhead counted twice, a shape that exists in NEITHER of his books. That rewrite was
# the defect. C7 is now set to his 27/07 shape - the complement, his newest statement of
# the cell, robust in both countries - and the 0.495 stands.
FIX = {
    "1.2 Customer": [("C7", None, None)],
    # his 1.5 edit removed the NZ column, so 0.2's P&C spend read wraps the dead D9 the
    # house way
    "0.2 Data Config": [("F18", "=('1.5 P&C'!$C$9+'1.5 P&C'!$D$9)",
                         "='1.5 P&C'!$C$9+N('1.5 P&C'!$D$9)"),
                        # every other row of this table reads its 1.x tab absolutely; two
                        # read relatively, which is a copy away from pointing at the wrong
                        # row
                        ("F13", "='1.2 Customer'!C9", "='1.2 Customer'!$C$9"),
                        ("F14", "='1.2 Customer'!D9", "='1.2 Customer'!$D$9"),
                        ("F15", "='1.9 Commercial Fuels'!C9",
                         "='1.9 Commercial Fuels'!$C$9")],
    # row 23 used to read both 1.13 and 1.14 because its own label said "TDD Cyber incl.
    # COE". His wave-M ruling splits that line in two - row 7 is the COE, row 23 is TDD
    # Cyber - so the combined read is retired and cyber_budget_split() below writes both
    # rows. RETIRED.md carries the old shape.
    # 1.10's Data NZ platform draws no overhead - the owner's own note on the squad row,
    # 1.10!K39 "No Overhead required", says so, and his C7 already prices two platforms.
    # Only D7 still drags the dead I40 reference from before he removed the charge, so the
    # two branches of the same switch priced different sets. D7 is aligned to his C7; his
    # note, his empty I40 and his B40 label are not touched.
    "1.10 Z Retail": [("D7", "=IF(('0.2 Data Config'!$D$12)>('0.2 Data Config'!$C$12),"
                             "SUM(I27,I34,I40),0)",
                       "=IF(('0.2 Data Config'!$D$12)>('0.2 Data Config'!$C$12),"
                       "SUM(I27,I34),0)")],
}

# His Legal, EG and EGI rows on 0.2 carry no Spend and no Variance. I filled them with
# zeros to make the column read consistently; he did not ask for that, and a typed zero is
# a statement he did not make. Left blank, as he has them.
CFG_GAPS = []


# his typed inputs on tabs the chain does not repaint - declared cream so the
# stale-literal police reads them as inputs, which they are.
#
# 1.2!I54 is off this list from wave M. It held his typed 2.21 for EGI Customer; he ruled
# that the EGI squads are funded by EGI at the ACTUAL cost of their roles, so the cell is a
# formula again (his own review workbook's shape, =IFERROR($H54*$G54,"")) and cream never
# sits on a formula.
CREAM = [("1.8 Energy Solutions & B2B", "E12"),
         ("1.8 Energy Solutions & B2B", "I14"), ("1.8 Energy Solutions & B2B", "I15")]

# ---------------------------------------------------------------- wave M, his rulings
# 1. The EGI squads are funded by EGI at the ACTUAL cost of their roles. Support % is 0,
#    so no part of them is a TDD cost, and the whole of them shows as funded outside. The
#    Total Squad Cost cell reads the working tab's own after-decisions total for that
#    squad, which is the actual - so the line moves with the ledger instead of holding a
#    figure typed on one day. None of it lands in Squad Support Costs; it lands in the
#    portfolio's Other column and so in its total cost, which is where he wants it.
#    (tab, support % cell, total-squad-cost cell, the working tab's own total)
EGI_ACTUAL = [("1.1 Ampol Retail", None, "H66", "'2.1 Ampol Retail'!$Q$18"),
              ("1.2 Customer", "G54", "H54", "'2.2 Customer'!$Q$17"),
              ("1.4 TDD Group Functions", "G30", "H30",
               "'2.4 TDD Group Functions'!$Q$14"),
              ("1.5 P&C", "G32", "H32", "'2.5 P&C'!$Q$12"),
              ("1.6 Finance", "G33", "H33", "'2.6 Finance'!$Q$12")]
# 1.1's EGI Retail support % is already 0 and is his own cell, so it is left alone; the
# other four are set to 0 here because the ruling is that none of the EGI cost is TDD's.
CYBER_NOTE_FROM, CYBER_NOTE_TO = "I23", "I7"


def wrap_empty_budget_reads(wb, out):
    """His zeroing rewires point budget cells at empty 0.1 cells; N() is the house style
    for a read that may legitimately be blank, and it is what keeps the adversarial pass
    able to tell a deliberate empty read from a broken one."""
    import re
    pat = re.compile(r"^='0\.1 Budget Table \(Fin\)'!([A-Z]{1,2}\d+)$")
    fin = wb["0.1 Budget Table (Fin)"]
    n = 0
    for t in [s_ for s_ in wb.sheetnames if re.match(r"^1\.\d+ ", s_)]:
        ws = wb[t]
        for row in ws.iter_rows(min_row=1, max_row=30, min_col=8, max_col=11):
            for c in row:
                if isinstance(c.value, str):
                    m = pat.match(c.value)
                    if m and fin[m.group(1)].value is None:
                        c.value = f"=N('0.1 Budget Table (Fin)'!{m.group(1)})"
                        n += 1
    out.append(f"{n} budget reads of empty 0.1 cells wrapped in N()")


def close_config_gaps(wb, out):
    """0.2's Spend and Variance cells, on the rows carrying neither."""
    ws = wb["0.2 Data Config"]
    for r in CFG_GAPS:
        f_, g_ = ws.cell(r, 6), ws.cell(r, 7)
        if f_.value is not None or g_.value is not None:
            out.append(f"0.2!F{r}/G{r} hold {f_.value!r} / {g_.value!r} - left alone")
            continue
        f_.value = 0
        g_.value = f"=E{r}-F{r}"
        out.append(f"0.2!{ws.cell(r, 2).value!r}: F{r} = 0 and G{r} = E{r}-F{r}, the pair "
                   f"every other row of this table carries")


# ---------------------------------------------------------------- the rest of the round
# cell -> (the text it must currently hold, what it says instead, why)
# 1.13's F11 ("planned spend less CapEx") used to be repaired here to follow its C13 input
# instead of a typed 0.5. His wave-M ruling removes the 0.5 CapEx input from 1.13 outright
# - "no 0.5 CapEx input (delete it and every reference)" - so there is no input left to
# follow and no cell left to repair. fixcoe.py rebuilds the block without either.
HONEST = [
    ("1.1 Ampol Retail", "J21", "=SUM(E9-J15-J16-J17-J18)", "=E9-(J19-J14)",
     "Left to fund: the applied total less the Lights On line, the shape its nine "
     "siblings use, instead of enumerating the four lines in between"),
    ("0.2 Data Config", "L23", 50.5, "=E26",
     "the allocated-budget total the row beside it already computes"),
]


def cyber_budget_split(wb, out):
    """0.2 Data Config: the COE line and the TDD Cyber line, each funded on its own row.

    His ruling, and his own action note on row 7 said it first: "Separate out and incl.
    cyber uplift funding bucket". One row was carrying two things - row 7 read zero against
    a COE spending ten, and row 23 was labelled "TDD Cyber incl. COE" and summed 1.13 and
    1.14 together - so neither line could be read against its own budget.

    After the split: row 7 is the COE (Cyber, Risk & Service Operations) at 1.5 AU / 0.5 NZ
    reading 1.13's planned spend, row 23 is TDD Cyber at 1.0 AU / 0.5 NZ reading 1.14's own
    TDD cost. The two allocations still add to the 3.5 he had on row 23, so E26 is 50.5 as
    it was. The four figures are cream: they are his to retype.

    His offshoring note moves from row 23 to row 7 with the COE it is about. His row-7
    action note is the one this pass actions, so it comes off - the thing it asks for is
    what the row now does.
    """
    ws = wb["0.2 Data Config"]
    COE_ROW, CYB_ROW = 7, 23
    for row, au, nz in ((COE_ROW, 1.5, 0.5), (CYB_ROW, 1.0, 0.5)):
        for col, v in (("C", au), ("D", nz)):
            x = ws[f"{col}{row}"]
            x.value = v
            x.fill = _cream()
            x.number_format = "0.00"
        ws[f"E{row}"] = f"=C{row}+D{row}"
    ws[f"F{COE_ROW}"] = "='1.13 Cyber Roles'!$F$8"
    ws[f"G{COE_ROW}"] = f"=E{COE_ROW}-F{COE_ROW}"
    # 1.14's own TDD cost, AU plus NZ - the pair every other portfolio row on this table
    # reads. The platform overhead is inside it, which is right: 1.14 draws one.
    ws[f"F{CYB_ROW}"] = "='1.14 TDD Cyber'!$C$9+'1.14 TDD Cyber'!$D$9"
    ws[f"G{CYB_ROW}"] = f"=SUM(E{CYB_ROW}-F{CYB_ROW})"
    ws[f"B{CYB_ROW}"] = "TDD Cyber"
    out.append("0.2!B7/C7/D7: the COE - Cyber, Risk & Service Ops line funded on its own "
               "row (1.5 AU / 0.5 NZ, cream), spend off 1.13's planned-spend total")
    out.append("0.2!B23: 'TDD Cyber incl. COE - Cyber, Risk & Service Ops' -> 'TDD Cyber' "
               "(1.0 AU / 0.5 NZ, cream), spend off 1.14's own TDD cost - the two rows "
               "still allocate the 3.5 the combined row held, so E26 stays 50.5")
    note = ws[CYBER_NOTE_FROM].value
    if isinstance(note, str) and note.strip():
        ws[CYBER_NOTE_TO].value = note
        ws[CYBER_NOTE_TO]._style = copy.copy(ws[CYBER_NOTE_FROM]._style)
        ws[CYBER_NOTE_FROM].value = None
        out.append(f"0.2!{CYBER_NOTE_FROM} -> {CYBER_NOTE_TO}: his offshoring note follows "
                   f"the COE line it is about ({note.strip()[:44]!r})")
    else:
        out.append(f"0.2!{CYBER_NOTE_FROM} holds nothing to move - note already at "
                   f"{CYBER_NOTE_TO}")


def egi_at_actual(wb, out):
    """The EGI squads, funded by EGI at the actual cost of their roles (his ruling).

    Support % 0, Total Squad Cost = the working tab's own after-decisions total for that
    squad, TDD Cost therefore 0 and Funded outside therefore the whole of it. The cell that
    used to hold a typed figure holds a live reference, so cream comes off it - a formula is
    never an input.
    """
    for tab, gcell, hcell, ref in EGI_ACTUAL:
        ws = wb[tab]
        if gcell is not None:
            ws[gcell] = 0
            ws[gcell].fill = _cream()
            ws[gcell].number_format = "0%"
        was = ws[hcell].value
        ws[hcell] = f"={ref}"
        ws[hcell].fill = openpyxl.styles.PatternFill()
        # his own review workbook's shape for the TDD Cost cell beside it: the split, not a
        # typed figure. His 27/07 book typed 2.21 into 1.2!I54 and he has now ruled on it.
        icell = "I" + hcell[1:]
        ws[icell] = f'=IFERROR(${hcell}*${gcell or ("G" + hcell[1:])},"")'
        ws[icell].fill = openpyxl.styles.PatternFill()
        out.append(f"{tab}!{hcell} = {ref} (was {str(was)[:20]!r}) and {icell} back to the "
                   f"support-% split - EGI funds this squad at the actual cost of its roles")


def restore_squad_rows(wb, out):
    """Squad rows of his that come back, with his note moved clear of a live column."""
    for tab, rows in RESTORED.items():
        ws = wb[tab]
        for name, r in rows:
            here = str(ws.cell(r, 2).value or "").strip()
            if here != name:
                out.append(f"{tab}!B{r} holds {here!r}, not {name!r} - NOT restored, "
                           f"left exactly as it is")
                continue
            x = ws.cell(r, 11)                       # K, now a live data column
            if isinstance(x.value, str) and x.value.strip() and not x.value.startswith("="):
                free = next((k for k in (13, 14) if ws.cell(r, k).value is None), None)
                if free is None:
                    out.append(f"{tab}!K{r} {x.value.strip()!r} kept where it is - M and N "
                               f"are both taken")
                else:
                    ws.cell(r, free).value = x.value
                    ws.cell(r, free)._style = copy.copy(x._style)
                    out.append(f"{tab}!K{r} {x.value.strip()!r} moved to "
                               f"{openpyxl.utils.get_column_letter(free)}{r} - his note "
                               f"outlives the column it sat in")
                    x.value = None
            out.append(f"{tab}!B{r} {name!r} stays - his own review workbook prices this "
                       f"squad and the NZ platform cost is his")


def _cream():
    import opts as _o
    return _o.fl(_o.YEL)


# ------------------------------------------------- the two squads that never were
# D117 added "AU CRM & Martech" and "Z Energy Martech" as squad rows on 1.2. D118
# reversed it on his ruling: "z energy martech is z loyalty & martech. same frickin
# squad, not a new squad". The old names in his dataset's pivot are old names for the
# renamed squads - folded, the counts land on exactly his own pivot totals (Ampol
# Loyalty & Martech 9 roles / 8.8 FTE, Z Loyalty & Martech 17 / 16.6). The W:X fold
# rows are back (ensure_lists / overrides UNFOLD keeps only "Customer, AI") and no rows
# are added here. The insertion machinery D117 used is gone with it; if a squad row
# ever needs adding again, the D117 commit holds the working pattern and its guards.


def add_data_coe_note(wb, out):
    """His note on the config tab, saying where the two roles actually sit.

    D119: "we literally said deepali and service transition lead DO NOT sit in COE-
    data, they sit in Ent Data." Both count in Enterprise Data's leadership group -
    their own raw rows, no override - and the note on the COE - Data row states the
    boundary, because anyone reconciling the COE against an org chart will look for
    them here and must be told to look in Enterprise Data instead."""
    ws = wb["0.2 Data Config"]
    note = ("Deepali Mahajan and the vacant Service Transition Lead are NOT in this "
            "line - they sit in Enterprise Data's leadership group.")
    cur = ws["H10"].value
    if cur is None or str(cur).startswith(("Incl. Deepali", "Deepali Mahajan")):
        ws["H10"] = note
        ws["H10"].font = openpyxl.styles.Font(name="Calibri", size=10)
        out.append("0.2!H10: Data COE boundary note - the two roles sit in Enterprise "
                   "Data leadership, not here (D119)")
    else:
        out.append(f"0.2!H10 holds {str(cur)[:40]!r} - his note, left alone; the Data "
                   f"COE explanation was NOT written")


def honest_cells(wb, out):
    for tab, cell, expect, repl, why in HONEST:
        cur = wb[tab][cell].value
        if cur == expect:
            wb[tab][cell] = repl
            out.append(f"{tab}!{cell} {expect!r} -> {repl!r} - {why}")
        else:
            out.append(f"{tab}!{cell} holds {str(cur)[:50]!r}, expected {expect!r} - "
                       f"left alone")


def run(src, dst):
    import opts as _o
    wb = openpyxl.load_workbook(src)
    out = []
    wrap_empty_budget_reads(wb, out)
    for tab, cell in CREAM:
        wb[tab][cell].fill = _o.fl(_o.YEL)
    out.append(f"{len(CREAM)} of his typed inputs declared cream")
    # his new 0.2 headers and 1.x note headers run past their columns; wrap them the way
    # every other header in the file wraps
    from openpyxl.styles import Alignment
    WRAP = [("0.2 Data Config", x) for x in ("N5", "N13", "N21", "O21", "L21", "M21")]
    for t in wb.sheetnames:
        import re as _re
        if _re.match(r"^1\.\d+ ", t):
            for r_ in range(10, 16):
                for c_ in (10, 11):
                    v = wb[t].cell(r_, c_).value
                    if isinstance(v, str) and len(v) > 24:
                        WRAP.append((t, wb[t].cell(r_, c_).coordinate))
    n_w = 0
    for t, cell in WRAP:
        x = wb[t][cell]
        if isinstance(x.value, str):
            x.alignment = Alignment(horizontal=x.alignment.horizontal or "left",
                                    vertical="center", wrap_text=True)
            n_w += 1
    out.append(f"{n_w} long headers set to wrap")
    # his new On/Off column widened the 1.13 roles table to H; the bar above follows it
    ws13 = wb["1.13 Cyber Roles"]
    for r_ in range(1, 30):
        v = str(ws13.cell(r_, 2).value or "")
        if v.strip() == "Roles":
            # B to I: his On/Off column, and the Uplift % column fixcoe adds beside it
            for c_ in range(2, 10):
                ws13.cell(r_, c_).fill = _o.fl(_o.BARC)
                ws13.cell(r_, c_).font = _o.BARF
            out.append(f"1.13!B{r_} bar extended across the On/Off and Uplift % columns")
            break
    for tab, pairs in RENAME.items():
        ws = wb[tab]
        # a squad's name appears three times in its block: on the platform bar above it,
        # on the squad row, and on the total row below. All three follow the ledger.
        wide = {}
        for old, new in pairs.items():
            wide[old] = new
            wide[f"Platform: {old}"] = f"Platform: {new}"
            wide[f"{old} Total"] = f"{new} Total"
        for r in range(1, min(ws.max_row, 95) + 1):
            v = ws.cell(r, 2).value
            if isinstance(v, str) and v.strip() in wide:
                new = wide[v.strip()]
                # whatever padding he typed around the name is his and stays
                lead = v[: len(v) - len(v.lstrip())]
                trail = v[len(v.rstrip()):]
                ws.cell(r, 2).value = lead + new + trail
                out.append(f"{tab}!B{r} {v.strip()!r} -> {new!r} (ledger name)")
    for tab, names in REMOVE.items():
        ws = wb[tab]
        for r in range(1, min(ws.max_row, 95) + 1):
            v = str(ws.cell(r, 2).value or "").strip()
            if v in names:
                dropped = [f"{openpyxl.utils.get_column_letter(c)}={ws.cell(r, c).value!r}"
                           for c in range(2, 11) if ws.cell(r, c).value is not None]
                for c in range(2, 11):              # B..J, the figures and the levers
                    ws.cell(r, c).value = None
                # K..N is the note margin. His writing survives the row it was sitting on:
                # it moves to M, where the design tabs carry a note, and the cell it came
                # from is cleared.
                for c in range(11, 15):
                    x = ws.cell(r, c)
                    if not (isinstance(x.value, str) and x.value.strip()
                            and not x.value.startswith("=")):
                        continue
                    if c in (13, 14):               # already in the note margin
                        continue
                    free = next((k for k in (13, 14) if ws.cell(r, k).value is None), None)
                    if free is None:
                        out.append(f"{tab}!{openpyxl.utils.get_column_letter(c)}{r} "
                                   f"{x.value.strip()!r} kept where it is - M and N are "
                                   f"both taken")
                        continue
                    ws.cell(r, free).value = x.value
                    ws.cell(r, free)._style = x._style
                    out.append(f"{tab}!{openpyxl.utils.get_column_letter(c)}{r} "
                               f"{x.value.strip()!r} moved to "
                               f"{openpyxl.utils.get_column_letter(free)}{r} - his note "
                               f"outlives the row it sat on")
                    x.value = None
                out.append(f"{tab}!B{r} {v!r} removed - zero people in the ledger; "
                           f"carried {'; '.join(dropped[:4])}")
    for tab, fixes in FIX.items():
        ws = wb[tab]
        for cell, expect, repl in fixes:
            cur = ws[cell].value
            if expect is None:
                # C7 arrives in his review-book shape (SUM over the empty H columns) and
                # leaves in his 27/07 shape: the complement of D7, one branch firing
                if isinstance(cur, str) and cur.endswith("SUM(H34,H42,H49),0)"):
                    ws[cell] = cur.replace("SUM(H34,H42,H49),0)",
                                           "0,SUM(I34,I42,I49))")
                    out.append(f"{tab}!{cell}: his 27/07 complement shape - platform "
                               f"overhead prices once, whichever country is home")
                elif isinstance(cur, str) and cur.endswith("0,SUM(I34,I42,I49))"):
                    out.append(f"{tab}!{cell} already his 27/07 shape")
                else:
                    out.append(f"{tab}!{cell} holds {str(cur)[:60]!r} - LEFT ALONE, "
                               f"expected his review or 27/07 shape")
                continue
            if cur == expect:
                ws[cell] = repl
                out.append(f"{tab}!{cell} -> {repl}")
            else:
                out.append(f"{tab}!{cell} holds {str(cur)[:50]!r} - left alone")
    close_config_gaps(wb, out)
    add_data_coe_note(wb, out)
    honest_cells(wb, out)
    # wave M, in his order: the two cyber budget lines, the EGI squads at actual, and the
    # squad row of his that comes back
    cyber_budget_split(wb, out)
    egi_at_actual(wb, out)
    restore_squad_rows(wb, out)
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
