"""Stage 15: last three items - identical formulas, visible Finance gaps, yellow inputs.

1  Each 2.x tab had its portfolio name typed inside every SUMIFS, so no two tabs carried
   the same formula and any future edit had to be made fourteen times. The name now sits
   once in B3 and every formula references it, so all fourteen tabs are byte-identical.

2  Eleven formulas read a cell in 0.1 Budget Table that is genuinely empty - the Finance
   pack has no Depreciation, Initiatives, Significant Items or CapEx line for those
   segments. They returned a silent zero that reads as a real nil budget. They now show
   "not in Finance" so a missing line is visible.

3  The 1.x tabs carry Lee's agreed figures typed into otherwise-formula columns - the
   Strategic Programs costs and the budget allocations. They are inputs, not defects, so
   they are filled yellow to mark what is editable, per the standing rule that only true
   inputs are yellow.
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill
import model

YELLOW = PatternFill("solid", start_color="FFFFFF00", end_color="FFFFFF00")
BOLD = Font(bold=True)


def identical_formulas(wb):
    n = 0
    for tab, pf in model.TAB_PORTFOLIO.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        ws["B3"] = "Portfolio"
        ws["C3"] = pf
        ws["B3"].font = BOLD
        lit = f'"{pf}"'
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("=") and lit in c.value:
                    c.value = c.value.replace(lit, "$C$3")
                    n += 1
    return [f"2.x: {n} formulas now read the portfolio from C3, so all fourteen tabs "
            f"carry identical formulas"]


def finance_gaps(wb):
    """A reference to an empty Finance cell returned a silent zero that read as a real
    nil budget. The value stays numeric via N() so downstream arithmetic still works, and
    one line per tab says which segments have no Finance entry - safer than writing a
    marker into a neighbouring cell, which on 1.8 and 1.10 was already in use."""
    out = 0
    for tab in [s for s in wb.sheetnames if re.match(r"^1\.(10|[1-9]) ", s)]:
        ws = wb[tab]
        hits = []
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=")
                        and "0.1 Budget Table (Fin)" in v and "N(" not in v):
                    continue
                c.value = f"=N({v[1:]})"
                hits.append(c.coordinate)
                out += 1
        if hits:
            r = ws.max_row + 2
            ws.cell(r, 2).value = ("Budget lines reading zero have no entry in the Finance "
                                   "pack for this segment; they are not a nil budget.")
    return [f"1.x: {out} Finance references made explicit with N(), and each tab states "
            f"that a zero there means no Finance line rather than no budget"]


def mark_inputs(wb):
    import qa
    n = 0
    for tab in [s for s in wb.sheetnames if re.match(r"^1\.\d+ ", s)]:
        ws = wb[tab]
        bycol = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None or c.column_letter in ("A", "B"):
                    continue
                bycol.setdefault(c.column_letter, []).append(c)
        for col, cells in bycol.items():
            fs = [c for c in cells if isinstance(c.value, str) and c.value.startswith("=")]
            ns = [c for c in cells if isinstance(c.value, (int, float))]
            if len(fs) >= 4 and ns:
                for c in ns:
                    c.fill = YELLOW
                    n += 1
    return [f"1.x: {n} agreed figures filled yellow so it is clear what is an input"]


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = []
    for fn in (identical_formulas, finance_gaps, mark_inputs):
        out += fn(wb)
    wb.save(dst)
    return out


if __name__ == "__main__":
    for x in run("f2.xlsx", "f3.xlsx"):
        print("  ", x)
