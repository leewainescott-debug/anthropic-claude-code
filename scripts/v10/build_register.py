import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
NAVY,INK,MUT,WHITE="FF0F2E52","FF1A1A1A","FF6B6B66","FFFFFFFF"
BAND,CREAM,RED,GREEN,AMBER="FFF5F4F0","FFFFF2CC","FFC0392B","FF1B7A3D","FFB9770E"
F=lambda c:PatternFill("solid",fgColor=c)
R=json.load(open("register_data.json"))
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Register"
ws.sheet_view.showGridLines=False
for k,v in {"A":1.5,"B":7,"C":21,"D":56,"E":6,"F":44,"G":11,"H":50,"I":2}.items():
    ws.column_dimensions[k].width=v
def put(c,v,sz=9,b=False,col=INK,fill=None,al=None,it=False,wrap=False):
    x=ws[c]; x.value=v
    x.font=Font(name="Calibri",size=sz,bold=b,color=col,italic=it)
    if fill: x.fill=F(fill)
    x.alignment=Alignment(horizontal=al or "left",vertical="top",wrap_text=wrap)
put("B2","TDD Cost Model - change register",16,True,NAVY)
put("B3","Everything since the 30/07 workbook. Status is verified against the shipped file, not from memory. Set the Status column as things land.",9,False,MUT)
ws.merge_cells("B3:H3")
r=5
for cl,h in zip("BCDEFGH",["ID","Area","What","Pri","Where it stands","Status","Your words"]):
    put(f"{cl}{r}",h,8,True,WHITE,fill=NAVY)
r+=1
PRI={"HIGH":RED,"MEDIUM":AMBER,"LOW":MUT,"":MUT}
def section(title,colr,rows,done=False):
    global r
    put(f"B{r}",title,10,True,colr,fill=BAND)
    for cl in "CDEFGH": ws[f"{cl}{r}"].fill=F(BAND)
    ws.merge_cells(f"B{r}:H{r}"); r+=1
    for row in rows:
        band=BAND if r%2 else None
        if done:
            rid,area,what,pri,stands,quote=f"DNE-{rows.index(row)+1:02d}",row[0],row[1],"",row[2],""
        else:
            rid,area,what,pri,stands,quote=row
        put(f"B{r}",rid,8,False,MUT,fill=band)
        put(f"C{r}",area,9,False,INK,fill=band,wrap=True)
        put(f"D{r}",what,9,False,INK,fill=band,wrap=True)
        put(f"E{r}",pri,8,True,PRI.get(pri,MUT),fill=band,al="center")
        put(f"F{r}",stands,8,False,MUT,fill=band,wrap=True)
        put(f"G{r}","",9,False,INK,fill=(band if done else CREAM))
        put(f"H{r}",quote,8,False,MUT,fill=band,it=True,wrap=True)
        r+=1
section(f"NEEDS YOUR DECISION  ({len(R['decisions'])})",RED,R["decisions"])
section(f"AGREED, NOT BUILT YET  ({len(R['build'])})",AMBER,R["build"])
section(f"DONE AND VERIFIED IN THE SHIPPED FILE  ({len(R['done'])})",GREEN,R["done"],done=True)
dv=DataValidation(type="list",formula1='"Open,In progress,Done,Parked"',allow_blank=True,showErrorMessage=True)
ws.add_data_validation(dv); dv.add(f"G6:G{r}")
put(f"B{r+1}",f"{len(R['decisions'])} decisions, {len(R['build'])} to build, {len(R['done'])} done. Nothing has been built since the file you have.",9,True,NAVY)
ws.merge_cells(f"B{r+1}:H{r+1}")
ws.freeze_panes="B6"
wb.save("TDD_Model_Register.xlsx")
print("rows",len(R['decisions'])+len(R['build'])+len(R['done']))
