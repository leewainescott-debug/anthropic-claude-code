"""3.4 column L had four typed zeros in a formula column.

Column L is "Of which funded inside portfolio overheads ($m)". Two of the six COE lines
drew a formula (Business Partnering read 1.11 C13, Strategy & Architecture read 1.12 C13)
and the other four were typed zeros.

Zero is the right answer for those four - Transformation, Data, Cyber and EGI have no
portfolio overhead line funding them. But typing the answer means the column stops being
a calculation. Add an overhead line for Transformation tomorrow and the zero stays a zero.

The overhead allowance table on Lists AF:AJ already prices all six lines from
0.2 Data Config. It just did not say which COE each line lands in. That mapping becomes a
column on the same table, and all six cells in 3.4 column L then run one formula: the
allowance for the overhead lines that land in me, nil if none do.

Values are unchanged: 2.2, 0, 1.4, 0, 0, 0, totalling 3.6.
"""
import openpyxl
from openpyxl.styles import Font
import wbio

BOLD = Font(bold=True)

# overhead line on Lists AF -> the 3.4 COE line it is funded into ("" = stays in the
# portfolios and never lands in a COE)
LANDS_IN = {
    "Head of Technology": "",
    "Business Partner": "Business Partnering",
    "Domain Architect": "Strategy & Architecture",
    "Delivery Manager": "",
    "Technology Manager": "",
    "Leadership - 8 GMs": "",
}


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    l = wb["Lists"]

    if l["AE1"].value not in (None, ""):
        raise RuntimeError(f"Lists AE1 is not free, it holds {l['AE1'].value!r}")
    l["AE1"] = "Funded into (3.4 COE line)"
    l["AE1"].font = BOLD
    for r in range(2, 8):
        line = str(l.cell(r, 32).value or "").strip()      # AF
        if line not in LANDS_IN:
            raise RuntimeError(f"Lists AF{r} reads {line!r}, not one of the six lines")
        l.cell(r, 31).value = LANDS_IN[line]               # AE

    s4 = wb["3.4 COE Summary"]
    seen = []
    for r in range(6, 12):
        name = str(s4[f"B{r}"].value or "").strip()
        if not name:
            raise RuntimeError(f"3.4 B{r} is empty")
        seen.append(name)
        s4[f"L{r}"] = ("=SUMIFS(Lists!$AJ$2:$AJ$7,Lists!$AE$2:$AE$7,$B{})".format(r))
    s4["L12"] = "=SUM(L6:L11)"

    # every mapped target must actually be a row on 3.4, or the mapping silently drops
    for line, target in LANDS_IN.items():
        if target and target not in seen:
            raise RuntimeError(f"{line!r} maps to {target!r}, which is not a 3.4 COE line")

    wb.save(dst)
    return [f"Lists AE: each overhead line now states which COE it is funded into",
            f"3.4 column L: all six cells derive the allowance from the Lists overhead "
            f"table; four were typed zeros"]


if __name__ == "__main__":
    import sys
    src, mid, dst = sys.argv[1], sys.argv[2], sys.argv[3]
    for x in run(src, mid):
        print("  ", x)
    rc, st = wbio.build(mid, dst)
    print("  injected", st)
    wv = openpyxl.load_workbook(dst, data_only=True)
    s4 = wv["3.4 COE Summary"]
    got = [s4[f"L{r}"].value for r in range(6, 13)]
    want = [2.2, 0, 1.4, 0, 0, 0, 3.6]
    ok = all(abs((g or 0) - w) < 1e-9 for g, w in zip(got, want))
    print("  3.4 column L:", got, "->", "unchanged" if ok else f"CHANGED, expected {want}")
