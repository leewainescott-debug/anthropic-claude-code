"""3.1, 3.2 and 3.3, reading the rebuilt 2.x tabs.

3.1 is layout 3D, which is what the owner picked and what the last build did not deliver: a
cost bridge. It starts at the archetype cost of the squads an archetype prices and walks,
one named line at a time, to what TDD actually costs. Every directly funded programme is on
the page by name - AmPOS, CTRM, the six EGI programmes, the two Leadership groups - because
a single "Directly funded 10.44" line tells a reader nothing about what the 10.44 is.
Per-portfolio and per-squad detail lives on 3.3, which is where 3D puts it.

3.2 was "Total Cost" and restated 3.1's subtotals. It is now overhead and leadership, which
is the one thing no other tab in the workbook states.

Design cost against actual cost, no budget anywhere. The design side is built four ways,
because the organisation is funded four ways, and each block states which one it is:

    squads priced by an archetype    the archetype library on 0.3
    directly funded programmes       the amount funded on the design tab's platform block
    COEs and EGI                     the planned spend on their own 1.x tabs
    overhead roles                   the allowance on Lists

The old build put the archetype cost of the squads that have one against the actual cost of
every squad in one subtotal. The group read +11.488 over design; the squads the archetype
actually prices are 0.559 under it, and the difference was the directly funded programmes
being charged against a zero. They are their own block now, priced against what they are
funded, so nothing is compared to nothing and nothing is left out.
"""
import json

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter as L

import final2x as f2
import opts

# a left-aligned wrapped body cell. opts.LFT does not wrap and opts.CEN centres, and the
# one column on 3.2 that now holds a sentence rather than a word needs both.
LFTW = Alignment(horizontal="left", vertical="center", wrap_text=True)

REVIEW = f2.REVIEW
REV = f2.REV
LAST = None                             # refreshed from f2 after _boot_last
S = f2.S

# The reading order of the portfolios, not a list of what exists. A name can sit here
# before the ledger rows and the working tab that will carry it do - order() skips any
# name with nothing behind it yet - so the tab can be added without a second edit here.
PF_ORDER = ["Ampol Retail", "Customer", "Enterprise Data", "TDD Group Functions", "P&C",
            "Finance", "Infrastructure", "Energy Solutions & B2B", "Commercial Fuels",
            "Z Retail", "TDD Cyber"]
COE_ORDER = ["COE Cyber", "COE BP&T", "COE SA&D", "EGI"]
# A COE's design cost is the planned spend on its own 1.x tab, built from its real roles
# rather than from an archetype. Reading the 1.x tab rather than the actual keeps it an
# independent number: where the 1.x roles list is short of the ledger the variance shows it
# instead of hiding it behind a cell that points at itself.
COE_DESIGN = {"COE BP&T": "='1.11 BP&T'!$F$6+'1.11 BP&T'!$F$7",
              "COE SA&D": "='1.12 SA&D'!$G$6+'1.12 SA&D'!$G$7",
              "COE Cyber": "='1.13 Cyber Roles'!$F$6+'1.13 Cyber Roles'!$F$7"}

# Lists!AF2:AJ7 is the overhead allowance. Three of the six lines are drawn in the
# portfolios; the Business Partner and Domain Architect allowances are drawn inside the
# COEs, where those people sit, and the 8 GMs sit above the 525-role ledger entirely.
# 3.1's overhead row therefore compares the portfolio-drawn allowance to the portfolio
# overhead cost. Netting all six against the portfolio cost only was what made the same
# allowance read -1.271 on one row of 3.2 and +9.996 seven rows below it.
# Yes where the line's people sit in the portfolios and its allowance is therefore the one
# the portfolio overhead cost should be measured against. The Business Partner and Domain
# Architect allowances are drawn inside the COEs, where all thirteen of those people sit,
# and the 8 GMs sit above the 525-role ledger, so none of the three belongs in the figure
# 3.1 compares the portfolio overhead cost to.
DRAWN = ["Yes", "No", "No", "Yes", "Yes", "No"]


def platforms(wb):
    """How many platforms the ten design tabs actually carry, as a formula.

    Each design tab's "Platform Overheads" line is the per-platform rate times that tab's
    own platform count, so dividing the ten of them by the rate gives the count back. Lists
    carried a typed 30 here. The design tabs do not carry 30, so the two per-platform
    allowances on Lists were built on a number the model does not use anywhere else: 3.2
    stated 6.325 of allowance for the 43 overhead roles in the portfolios and 3.1 stated
    5.335 for the same 43 roles, and neither tab said which one the reader should believe.
    Counting the platforms off the design tabs puts both tabs on the same basis - what the
    ten portfolios are actually allowed - and the two figures become one figure.
    """
    parts = []
    for design in sorted(set(f2.DESIGN.values())):
        if design not in wb.sheetnames:
            continue
        _, plat = f2.oh_rows(wb, design)
        if plat:
            parts.append(f"N('{design}'!$F${plat})")
    if not parts:
        return None
    return f"=ROUND(({'+'.join(parts)})/{f2.CFG}!$N$16,6)"


def patch_lists(wb):
    """Say where each overhead line's cost sits, and total the portfolio-drawn allowance."""
    l = wb["Lists"]
    n_plat = platforms(wb)
    if n_plat:
        for i in range(2, 8):
            if str(l.cell(i, 35).value or "").strip() == "platforms":
                l.cell(i, 34).value = n_plat
                l.cell(i, 34).number_format = opts.C1
    col = next(c for c in range(39, 70)
               if all(l.cell(r, c).value is None for r in range(1, 40)))
    cl = L(col)
    x = l.cell(1, col)
    x.value, x.font, x.fill, x.alignment = ("Allowance drawn in the portfolios",
                                            opts.HDRF, opts.fl(opts.NAVY), opts.CEN)
    l.column_dimensions[cl].width = 28
    for i, w in enumerate(DRAWN):
        c = l.cell(2 + i, col)
        c.value, c.font, c.border, c.alignment = w, opts.BODY, opts.BOX, opts.LFT
    a = l.cell(9, 32)
    a.value, a.font = "Allowance drawn in the portfolios", opts.BOLD
    b = l.cell(9, 36)
    b.value = f'=SUMIF(${cl}$2:${cl}$7,"Yes",$AJ$2:$AJ$7)'
    b.font, b.number_format, b.alignment = opts.BOLD, opts.M2, opts.RGT
    return (f"Lists: overhead lines tagged in {cl}, portfolio-drawn allowance at AJ9, "
            f"per-platform lines counted off the design tabs"), cl


H31 = ["Line", "Portfolio", "Archetype cost ($m)", "Actual cost ($m)",
       "Variance to archetype ($m)", "Cost after decisions ($m)", "Total roles",
       "Filled", "Vacant", "Total roles after decisions"]
W31 = [34, 24, 15, 14, 17, 17, 9, 8, 8, 14]
NUM = {4: opts.M2, 5: opts.M2, 6: opts.M2, 7: opts.M2,
       8: opts.CT, 9: opts.CT, 10: opts.CT, 11: opts.CT}
FIRST, LASTC = 4, 11                                     # D .. K


def order(anchors, wb=None):
    """PF_ORDER and COE_ORDER against what the workbook actually carries.

    A name on either list is skipped in silence when nothing stands behind it yet: no
    ledger rows, so no anchor was written for it, or an anchor whose working tab is not
    in this workbook. The lists are the reading order the owner wants, and a portfolio
    can be named here before the tab that will hold it exists.
    """
    by = {a["pf"]: t for t, a in anchors.items()}

    def pick(names):
        out = []
        for p in names:
            t = by.get(p)
            if t is None or (wb is not None and t not in wb.sheetnames):
                continue
            out.append((p, t, anchors[t]))
        return out

    return pick(PF_ORDER), pick(COE_ORDER)


def build_31(wb, anchors):
    """The cost bridge. One line per step, every step named.

    Option 3D, which is what the owner picked: start at the archetype cost of the squads
    the archetype prices and walk to what the organisation actually costs, naming each
    step. Every directly funded programme is on its own line by name - AmPOS, CTRM, the
    EGI programmes - because "Directly funded, 10.44" tells a reader nothing about what
    the 10.44 is. Per-portfolio and per-squad detail is on 3.3.
    """
    ws = wb["3.1 Group Summary"]
    n_roles = opts.ledger_count(wb)
    f2.wipe(ws)
    ws.column_dimensions["A"].width = 2
    ws.cell(2, 2).value = "TDD cost bridge - archetype cost to actual cost"
    ws.cell(2, 2).font = opts.TITLE
    pfs, coes = order(anchors, wb)

    r = opts.head(ws, 4, 2, H31, W31)

    def line(rw, name, pf, tab, src, design):
        ws.cell(rw, 2).value = name
        ws.cell(rw, 3).value = pf
        for c in (2, 3):
            ws.cell(rw, c).font = opts.BODY
            ws.cell(rw, c).alignment = opts.LFT
        f2._m(ws, rw, 4, design)
        f2._m(ws, rw, 5, f"='{tab}'!${L(S['actual'])}${src}")
        f2._m(ws, rw, 6, f'=IFERROR(ROUND($E{rw}-$D{rw},6),"-")')
        f2._m(ws, rw, 7, f"='{tab}'!${L(S['after'])}${src}")
        for i, k in enumerate(("roles", "filled", "vacant", "rafter")):
            f2._m(ws, rw, 8 + i, f"='{tab}'!${L(S[k])}${src}", opts.CT)

    def sub(rw, text, r0, r1, blank=(), rows=None, priced_only=False):
        opts.row(ws, rw, 2, [text] + [None] * (len(H31) - 1), [None] * len(H31),
                 bg=opts.GREY, bold=True)
        ws.cell(rw, 2).alignment = opts.LFT
        pick = rows if rows is not None else list(range(r0, r1 + 1))
        for c in range(FIRST, LASTC + 1):
            x = ws.cell(rw, c)
            # the variance on a total is actual less archetype, never the sum of the row
            # variances: a row with no archetype carries "-" in that column and drops out
            # of a SUM, so the directly funded subtotal read 0.13 against a real 1.44 and
            # the group read 5.07 against 6.38
            if c in blank:
                x.value = '="-"'
            elif c == 6:
                x.value = f'=IF(ISNUMBER($D{rw}),ROUND($E{rw}-$D{rw},6),"-")'
            elif c == 4 and priced_only:
                # a portfolio can sit in this section before its squad is priced - 1.14
                # TDD Cyber ships with its archetype unset - so the total covers the
                # priced members and the unpriced line waits with a dash. SUM skips text.
                x.value = f'=IF(COUNT($D{r0}:$D{r1})=0,"-",SUM($D{r0}:$D{r1}))'
            elif c == 4:
                # the archetype side is a total only if every row under it has a figure.
                # SUM over a block of dashes is 0, which would put a whole step's actual
                # against an archetype of nothing and call the difference overspend.
                cells = ",".join(f"$D{k}" for k in pick)
                x.value = f'=IF(COUNT({cells})={len(pick)},SUM({cells}),"-")'
            elif priced_only:
                # both sides of the comparison cover the same members: a line whose
                # archetype is not a number stays out of the actual and the counts too,
                # and joins every column at once the moment it is priced. Same mechanism
                # split() uses. If an unpriced portfolio ever carries real roles, the
                # ledger ties on 4.0 go loudly non-zero rather than absorbing it.
                x.value = (f"=SUMPRODUCT(--ISNUMBER($D{r0}:$D{r1}),"
                           f"{L(c)}{r0}:{L(c)}{r1})")
            else:
                x.value = "=" + "+".join(f"N({L(c)}{k})" for k in pick)
            x.number_format, x.alignment = NUM[c], opts.RGT

    def label(rw, text):
        opts.row(ws, rw, 2, [text] + [None] * (len(H31) - 1), [None] * len(H31),
                 bg=opts.PALE, bold=True)
        ws.cell(rw, 2).alignment = opts.LFT

    def split(rw, text, r0, r1, priced):
        """One half of a block: the rows that carry a figure to compare, or the rest.

        Both halves are driven off ISNUMBER of the archetype column rather than off a list
        worked out at build time, so the split follows the workbook. Set a funded figure on
        a 1.x tab and that programme moves up a line on its own.
        """
        opts.row(ws, rw, 2, [text] + [None] * (len(H31) - 1), [None] * len(H31),
                 bg=opts.GREY, bold=True)
        ws.cell(rw, 2).alignment = opts.LFT
        pick = f"--{'' if priced else 'NOT('}ISNUMBER($D{r0}:$D{r1}){'' if priced else ')'}"
        for c in range(FIRST, LASTC + 1):
            x = ws.cell(rw, c)
            if c == 4:
                x.value = (f'=IF(COUNT($D{r0}:$D{r1})=0,"-",SUM($D{r0}:$D{r1}))'
                           if priced else '="-"')
            elif c == 6:
                x.value = (f'=IF(ISNUMBER($D{rw}),ROUND($E{rw}-$D{rw},6),"-")'
                           if priced else '="-"')
            else:
                x.value = f"=SUMPRODUCT({pick},{L(c)}{r0}:{L(c)}{r1})"
            x.number_format, x.alignment = NUM[c], opts.RGT

    # ---- step 1: the squads an archetype prices, one line per portfolio ----
    label(r, "Squads priced by an archetype - detail on 3.3")
    r += 1
    st = r
    for pf, tab, a in pfs:
        # a portfolio with no archetyped squad has no row to read - the section is not on
        # its tab at all - so it does not get a line here either
        if not a["delivery_row"]:
            continue
        line(r, pf, pf, tab, a["delivery_row"],
             f"='{tab}'!${L(S['acost'])}${a['delivery_row']}")
        r += 1
    sub(r, "Squads priced by an archetype", st, r - 1, priced_only=True)
    s1 = r
    r += 1

    # ---- step 2: every directly funded programme, by name ----
    label(r, "Directly funded programmes and platforms - no archetype prices them, so "
             "the comparison is the amount funded on the 1.x tab")
    r += 1
    d_st = r
    for pf, tab, a in pfs:
        for g in a["direct"]:
            line(r, g, pf, tab, a["srow"][g],
                 f"='{tab}'!${L(S['acost'])}${a['srow'][g]}")
            r += 1
    d_en = r - 1
    # Two subtotals, not one. Some of these programmes have a funded figure set against
    # them on the 1.x tab and some do not, and a single subtotal put an archetype side
    # covering two of eight rows against an actual side covering all eight - the imbalance
    # the owner picked up on the 2.x overhead lines, one level up. The split is by formula,
    # so a figure typed into a 1.x tab tomorrow moves its programme to the top line by
    # itself.
    #
    # Only the half with a funded figure sits here, directly under the rows it draws from.
    # The other half has nothing to compare to, so it belongs below the comparison with the
    # rest of what cannot be compared - see the block after the subtotal.
    split(r, "Directly funded, where the funded figure is set", d_st, d_en, True)
    s2 = r
    r += 1

    # ---- step 3: overhead in the portfolios, against the allowance ----
    # One row. The allowance is built per portfolio and per platform at group level, so
    # there is no per-portfolio allowance to set beside a per-portfolio cost - ten rows of
    # dashes under a subtotal carrying 6.325 is a subtotal that is not the sum of its rows.
    # The detail by portfolio is on 3.3 and on each working tab, line by line on 3.2.
    oh = [(p, t, a) for p, t, a in pfs if a["overhead_row"]]
    opts.row(ws, r, 2, ["Overhead roles in the portfolios - the allowance is on 3.2"] +
             [None] * (len(H31) - 1), [None] * len(H31), bg=opts.GREY, bold=True)
    ws.cell(r, 2).alignment = opts.LFT
    # the allowance the ten design tabs actually give these lines, read off the working
    # tabs. Lists used to price the two per-platform lines over a typed 30 platforms while
    # the design tabs carried fewer, so 3.2 stated one allowance for these 43 roles and 3.1
    # stated another for the same 43 roles. patch_lists now counts the platforms off the
    # design tabs, so both tabs are on this basis and the two figures are the same figure.
    f2._m(ws, r, 4, "=" + "+".join(
        f"N('{t}'!${L(S['acost'])}${a['overhead_row']})" for _, t, a in oh))
    for c, k in ((5, "actual"), (7, "after")):
        f2._m(ws, r, c, "=" + "+".join(
            f"N('{t}'!${L(S[k])}${a['overhead_row']})" for _, t, a in oh))
    f2._m(ws, r, 6, f"=ROUND($E{r}-$D{r},6)")
    for i, k in enumerate(("roles", "filled", "vacant", "rafter")):
        f2._m(ws, r, 8 + i, "=" + "+".join(
            f"N('{t}'!${L(S[k])}${a['overhead_row']})" for _, t, a in oh), opts.CT)
    for c in range(FIRST, LASTC + 1):
        ws.cell(r, c).font = opts.BOLD
    s4 = r
    r += 1
    # Only the steps whose archetype side covers the whole of their own actual side. The
    # COEs came out because their figure was the actual restated, and the directly funded
    # programmes with no funded figure came out for the same reason. Adding a step whose
    # archetype is a dash counts it as zero on one side and in full on the other, which is
    # the difference between a variance and a subtraction.
    opts.row(ws, r, 2, ["Everything with a figure to compare"] + [None] * (len(H31) - 1),
             [None] * len(H31), bg=opts.MID, bold=True, top=True)
    ws.cell(r, 2).alignment = opts.LFT
    for c in range(FIRST, LASTC + 1):
        x = ws.cell(r, c)
        x.value = (f"=ROUND($E{r}-$D{r},6)" if c == 6
                   else "=" + "+".join(f"N({L(c)}{p})" for p in (s1, s2, s4)))
        x.number_format, x.alignment = NUM[c], opts.RGT
    cmp_row = r
    r += 1

    # ---- everything the subtotal above does not reach, after it and not before it ----
    # The three steps below carry a dash in the archetype column: nothing prices them, so
    # they cannot be inside a comparison. They used to be printed above the subtotal, which
    # left a bold grey total sitting under two grey rows it did not include - 83.19 drawn
    # under a 1.81 and a 28.68 that are not in it. Every row above the subtotal is now in
    # it and every row below it is not, which is the only arrangement a reader can check.
    label(r, "Nothing prices these, so they sit below the comparison")
    r += 1
    split(r, "Directly funded, where no funded figure is set yet", d_st, d_en, False)
    s2c = r
    r += 1

    # ---- the COEs and EGI ----
    # No figure prices these independently of what they cost. The column used to read the
    # planned spend off their own 1.x tabs, and planned spend is the same SUMIFS over the
    # ledger that the actual column is - EGI read the identical cell on both sides. Four
    # lines of 27.77 against 27.77 padded the comparable total by a quarter of the group
    # and moved nothing. The comparison is a dash and the step sits below the line.
    label(r, "COEs and EGI - nothing prices these apart from what they cost")
    r += 1
    st = r
    for pf, tab, a in coes:
        line(r, pf, pf, tab, a["total_row"], '="-"')
        r += 1
    sub(r, "COEs and EGI", st, r - 1, blank=(4, 6))
    s3 = r
    r += 1

    # ---- groups with nothing to compare against ----
    nofig = [(p, t, a) for p, t, a in pfs if a["nofig"]]
    # the label and the subtotal must not read identically, or anything looking the row
    # up by name finds the label - which carries no figures
    label(r, "Groups with no archetype and no funded figure - nothing to compare to")
    r += 1
    st = r
    for pf, tab, a in nofig:
        for g in a["nofig"]:
            line(r, g, pf, tab, a["srow"][g], '="-"')
            r += 1
    sub(r, "Groups with no archetype and no funded figure", st, r - 1, blank=(4, 6))
    s2b = r
    r += 1

    # ---- the ledger total ----
    # Its variance is a dash, not a number. The archetype column prices four of the five
    # steps above it and the actual column covers all five, so no single figure on this row
    # can be true. The comparable subtotal two rows up is the one that can.
    opts.row(ws, r, 2, [f"Cost of the {n_roles} roles in the ledger"] + [None] * (len(H31) - 1),
             [None] * len(H31), bg=opts.MID, bold=True, top=True)
    ws.cell(r, 2).alignment = opts.LFT
    for c in range(FIRST, LASTC + 1):
        x = ws.cell(r, c)
        # The ledger row carries the comparison, on the owner's instruction: the archetype
        # total against the actual total, and the difference. The archetype side prices
        # three of the six steps above it, so the difference is everything the archetype
        # does not reach - the COEs, the programmes with no funded figure set, Leadership -
        # plus the overspend on what it does reach. Every one of those steps is a named
        # line above with a dash in this column, so the figure cannot be read as anything
        # else. Same treatment as the Total portfolio row on every 2.x tab.
        steps = (s1, s2, s2c, s2b, s3, s4)
        if c == 4:
            cells = ",".join(f"{L(c)}{p}" for p in steps)
            x.value = f'=IF(COUNT({cells})=0,"-",SUM({cells}))'
        elif c == 6:
            x.value = f'=IF(ISNUMBER($D{r}),ROUND($E{r}-$D{r},6),"-")'
        else:
            x.value = "=" + "+".join(f"N({L(c)}{p})" for p in steps)
        x.number_format, x.alignment = NUM[c], opts.RGT
    gt = r
    r += 1

    # ---- the GM layer, which has no role in the ledger to price ----
    # A data row, styled as one. It carried the pale fill this tab uses for a section label,
    # so the one line on the bridge that is neither a heading nor a subtotal was wearing the
    # heading's colour and a reader had to work out from the figures that it was a line.
    opts.row(ws, r, 2, [f"Leadership - the 8 GMs, outside the {n_roles}-role ledger"] +
             [None] * (len(H31) - 1), [None] * len(H31))
    ws.cell(r, 2).alignment = opts.LFT
    f2._m(ws, r, 4, "=N(Lists!$AJ$7)")
    f2._m(ws, r, 5, "=N(Lists!$AG$12)")
    f2._m(ws, r, 6, f"=ROUND($E{r}-$D{r},6)")
    f2._m(ws, r, 7, "=N(Lists!$AG$12)")
    for c in (8, 9, 11):
        f2._m(ws, r, c, "=N(Lists!$AG$11)", opts.CT)
    f2._m(ws, r, 10, "=0", opts.CT)
    gm = r
    r += 1
    opts.row(ws, r, 2, ["Total cost of TDD including the GM layer"] +
             [None] * (len(H31) - 1), [None] * len(H31), bg=opts.MID, bold=True, top=True)
    ws.cell(r, 2).alignment = opts.LFT
    for c in range(FIRST, LASTC + 1):
        x = ws.cell(r, c)
        x.value = (f'=IF(ISNUMBER($D{r}),ROUND($E{r}-$D{r},6),"-")' if c == 6
                   else f"=N({L(c)}{gt})+N({L(c)}{gm})")
        x.number_format, x.alignment = NUM[c], opts.RGT
    grand = r
    r += 2

    for lab, col, f, nf in (
            ("Control - roles against the ledger, must be 0", 8,
             f"=$H${gt}-COUNTA({REV}!$B$2:$B${LAST})", opts.CTL_C),
            ("Control - cost against the ledger ($m), must be 0", 5,
             f"=ROUND($E${gt}-SUM({REV}!$AA$2:$AA${LAST})/1000000,6)", opts.CTL_M)):
        ws.cell(r, 2).value = lab
        ws.cell(r, 2).font = opts.BODY
        f2._m(ws, r, col, f, nf)
        r += 1
    return {"total": gt, "grand": grand, "gm": gm, "arch": s1, "direct": s2,
            "direct_unpriced": s2c, "nofig": s2b, "coe": s3, "overhead": s4,
            "comparable": cmp_row}


# 3.2 was "Total Cost" and restated 3.1's four subtotals in a second table. It gave a
# reader nothing 3.1 did not already give them, and the owner deleted it. What it did carry
# that nothing else does is the overhead allowance, line by line, and the GM layer - so
# that is what the tab is now, and only that.
#
# The layout below is the owner's own brief, in his words: "the whole purpose is to show
# how much we've applied to portfolios and how many roles there are in the organisation,
# the discrepancy between those two numbers and then the discrepancy between the cost."
# So the table reads left to right as exactly that sentence - what the allowance grants a
# portfolio, what the organisation actually carries, and the two gaps - and the two gap
# columns are the point of the tab, so they are the ones set in bold.
#
# What went, and why. "Roles in the portfolios" and "Cost in the portfolios" answered half
# the question and the other half sat four columns away under "outside the portfolios",
# so a reader had to add two cells to learn how many people are on a line. The roles and
# the cost columns are the organisation's now - portfolios, COEs and EGI together - which
# is the number he asked for. "Portfolio cost less allowance" is the cost gap, on the same
# basis as the roles gap beside it. The Yes/No column and the sentence that replaced it
# both answered "is the allowance drawn here", which is a question about the allowance,
# not about the people; "Where they sit" answers the question a reader actually has. The
# separate "All roles" column stated one figure on one row and a dash on every other, so
# it is a row now and not a column.
H32 = ["Overhead line", "Basis", "Rate ($m)", "Applied to portfolios - roles",
       "Applied to portfolios ($m)", "Roles in the organisation",
       "Cost in the organisation ($m)", "Roles gap", "Cost gap ($m)",
       "Where they sit"]
# the bar on row 4 and the header on row 5 are both drawn off len(H32), so a column is
# added by adding to these two lists and to nothing else
W32 = [26, 13, 11, 16, 16, 15, 17, 11, 13, 44]
# the last column of the table: B is 2, so B .. K is 2 .. 2 + len(H32) - 1
C32K = 2 + len(H32) - 1
# The allocation each overhead line grants per application, on 0.2 Data Config, keyed by
# the line's row on Lists. Lists!AH holds how many times the line is applied and 0.2's M
# column holds the fraction of a role each application buys, so the two multiplied are the
# FTE the allowance grants the portfolios. The rate block at the foot of this tab reads the
# same six cells for its allocation column, which is what keeps the two halves of the tab
# on one set of inputs.
ALLOC32 = {2: "$M$6", 3: "$M$7", 4: "$M$8", 5: "$M$14", 6: "$M$15", 7: "$M$9"}
# the one line whose people have no role in the ledger at all
GM32 = "Leadership - 8 GMs"


def build_32(wb, wcol):
    """Overhead and leadership. The one thing no other tab states.

    Every figure is computed. A role carrying an overhead title inside a COE has AT set to
    its COE squad rather than to the overhead line, so AR = AT selects the roles that sit
    in a portfolio. That test is what the "Where they sit" column and the two "of which"
    bands are built on; the roles and cost columns themselves do not use it, because the
    question they answer is how many of these people the organisation has in total.
    """
    ws = wb["3.2 Total Cost"]
    f2.wipe(ws)
    ws.column_dimensions["A"].width = 2
    ws.cell(2, 2).value = ("Overhead and leadership - what the portfolios are allowed "
                           "against what the organisation has")
    ws.cell(2, 2).font = opts.TITLE
    r = opts.bar(ws, 4, 2, len(H32), "Every overhead line - applied to the portfolios, "
                                     "carried by the organisation, and the gap")
    r = opts.head(ws, r, 2, H32, W32)
    st = r
    n_roles = opts.ledger_count(wb)
    for i in range(2, 8):
        # every role in the ledger carrying this line, wherever the person sits, which is
        # the figure the columns state; and the same count narrowed to the people who sit
        # in a portfolio, which only the "Where they sit" sentence and the bands use
        both = f"{REV}!$AR$2:$AR${LAST},$B{r},{REV}!$B$2:$B${LAST},\"<>\""
        pf_only = f"COUNTIFS({both},{REV}!$AT$2:$AT${LAST},$B{r})"
        gm = '=IF($B{r}="' + GM32 + '",{v},{e})'
        ws.cell(r, 2).value = f"=Lists!$AF${i}"
        ws.cell(r, 3).value = f"=Lists!$AI${i}"
        for c in (2, 3):
            ws.cell(r, c).font = opts.BODY
            ws.cell(r, c).alignment = opts.LFT
        f2._m(ws, r, 4, f"=Lists!$AG${i}", opts.M3)
        # the FTE the allowance grants: times applied x the allocation per application.
        # Half a Head of Technology over ten portfolios is five people's worth of Head of
        # Technology, and that is the number the 15 in the next column but one is short of.
        f2._m(ws, r, 5,
              f"=Lists!$AH${i}*N('0.2 Data Config'!{ALLOC32[i]})", opts.C1)
        f2._m(ws, r, 6, f"=Lists!$AJ${i}")
        # the 8 GMs carry no role in the ledger, so their count and cost are the input
        f2._m(ws, r, 7, gm.format(r=r, v="N(Lists!$AG$11)",
                                  e=f"COUNTIFS({both})"), opts.CT)
        f2._m(ws, r, 8, gm.format(
            r=r, v="N(Lists!$AG$12)",
            e=f"SUMIFS({REV}!$AA$2:$AA${LAST},{both})/1000000"))
        # the two gaps. They are the tab: what the organisation carries less what the
        # portfolios are allowed, in people and in dollars, on every line.
        for c, f, nf in ((9, f"=ROUND($G{r}-$E{r},6)", opts.C1),
                         (10, f"=ROUND($H{r}-$F{r},6)", opts.M2)):
            f2._m(ws, r, c, f, nf).font = opts.BOLD
        # One plain-English cell instead of a Yes/No and a sentence about the allowance.
        # Built off the same AR = AT test the bands below use, so it cannot go stale: move
        # a Head of Technology out of a COE and this cell says so on the next calculation.
        ws.cell(r, 11).value = (
            f'=IF($B{r}="{GM32}","Above the ledger",'
            f'IF($G{r}=0,"No roles in the ledger carry this line",'
            f'IF({pf_only}=0,'
            f'"All "&TEXT($G{r},"0")&IF($G{r}=1," in a COE"," in the COEs"),'
            f'IF({pf_only}=$G{r},'
            f'"All "&TEXT($G{r},"0")'
            f'&IF($G{r}=1," in a portfolio"," in the portfolios"),'
            f'TEXT({pf_only},"0")'
            f'&IF({pf_only}=1," in a portfolio, "," in the portfolios, ")'
            f'&TEXT($G{r}-{pf_only},"0")'
            f'&IF($G{r}-{pf_only}=1," in a COE"," in the COEs")))))')
        ws.cell(r, 11).font = opts.BODY
        ws.cell(r, 11).alignment = LFTW
        # the column is as wide as its own longest sentence, so every line renders on one;
        # the height is two lines' worth anyway, for a line that ever states more
        ws.row_dimensions[r].height = 30
        r += 1
    lines = list(range(st, r))

    def dash(rw, skip):
        """Every column the row does not state carries a dash, so a band reads as one row
        across the whole table rather than one that gives up halfway."""
        for c in range(4, C32K + 1):
            if c in skip:
                continue
            x = ws.cell(rw, c)
            x.value = '="-"'
            x.alignment = opts.LFT if c == C32K else opts.RGT

    opts.row(ws, r, 2, ["Overheads incl. GMs"] + [None] * (len(H32) - 1),
             [None] * len(H32), bg=opts.MID, bold=True, top=True)
    ws.cell(r, 2).alignment = opts.LFT
    adds = ((5, opts.C1), (6, opts.M2), (7, opts.CT), (8, opts.M2),
            (9, opts.C1), (10, opts.M2))
    for c, nf in adds:
        x = ws.cell(r, c)
        x.value = f"=SUM({L(c)}{st}:{L(c)}{r-1})"
        x.number_format, x.alignment = nf, opts.RGT
    # a rate cannot be added across six lines on six different bases, and the last column
    # holds words. They say so rather than sitting blank inside the band.
    dash(r, {c for c, _ in adds})
    tot32 = r

    def band(rw, text, cells):
        # the fill and the border run B to K, the full width of the header above.
        # C stays empty: the label in B is longer than B and has to run into it.
        opts.row(ws, rw, 2, [text] + [None] * (len(H32) - 1), [None] * len(H32),
                 bg=opts.GREY, bold=True)
        ws.cell(rw, 2).alignment = opts.LFT
        for c, (f, nf) in cells.items():
            f2._m(ws, rw, c, f, nf)
        dash(rw, set(cells))
        for c in range(4, C32K + 1):
            ws.cell(rw, c).font = opts.BOLD

    # The two bands split every column above them on one test - where the people are. The
    # allowance splits with them: three of the six lines are drawn where their people sit,
    # and the other three are allowed for per portfolio while their people sit in a COE or
    # above the ledger entirely. That is why the ten portfolios draw 5.005 of an 11.605
    # allowance, and it is the fact the old tab left a reader to work out.
    # The roles and cost are counted here off REVIEW, line by line, and not read back off
    # 3.1 - 3.1's overhead step is measured against this band on 4.0, and a band that read
    # 3.1 would have made that check compare a cell to itself.
    pf_n = "+".join(f"COUNTIFS({REV}!$AR$2:$AR${LAST},$B{k},"
                    f"{REV}!$B$2:$B${LAST},\"<>\","
                    f"{REV}!$AT$2:$AT${LAST},$B{k})" for k in lines)
    pf_c = "+".join(f"SUMIFS({REV}!$AA$2:$AA${LAST},{REV}!$AR$2:$AR${LAST},$B{k},"
                    f"{REV}!$AT$2:$AT${LAST},$B{k})" for k in lines)
    r += 1
    ohpf = r
    band(r, "Of which sits in the portfolios",
         {5: (f'=SUMIF(Lists!${wcol}$2:${wcol}$7,"Yes",$E{st}:$E{lines[-1]})', opts.C1),
          6: ("=N(Lists!$AJ$9)", opts.M2),
          7: (f"={pf_n}", opts.CT),
          8: (f"=({pf_c})/1000000", opts.M2),
          9: (f"=ROUND($G{r}-$E{r},6)", opts.C1),
          10: (f"=ROUND($H{r}-$F{r},6)", opts.M2)})
    r += 1
    ohout = r
    band(r, f"Of which sits in the COEs and EGI, or above the {n_roles}-role ledger",
         {5: (f"=ROUND($E{tot32}-$E{ohpf},6)", opts.C1),
          6: (f"=ROUND($F{tot32}-$F{ohpf},6)", opts.M2),
          7: (f"=$G{tot32}-$G{ohpf}", opts.CT),
          8: (f"=ROUND($H{tot32}-$H{ohpf},6)", opts.M2),
          9: (f"=ROUND($G{r}-$E{r},6)", opts.C1),
          10: (f"=ROUND($H{r}-$F{r},6)", opts.M2)})
    r += 1

    # ---- the whole model, each role counted once ----
    # The bands above split the overhead lines. This one row states the ledger itself,
    # which is the question a reader arrives with: how many people are in the organisation
    # at all, how many of them sit in a portfolio, how many sit in a COE or EGI. It is one
    # sentence and one figure, not a set of columns, because it is one fact. The ten
    # portfolios are named off Lists rather than tested as "not a COE": a role whose
    # portfolio is blank or mistyped then falls out of both halves and the control under
    # the row says so, instead of being swept into whichever half carries the negation.
    allrow = r
    pfs32 = "Lists!$AS$2:$AS$11"
    # TDD Cyber is a portfolio that sits outside the ten-name Lists range on purpose
    # (joining the range would inflate the funding maths), so its future roles are
    # counted by name here - without this term they would fall out of both halves and
    # the control would fire the day the first role lands
    pf_roles = (f"(SUMPRODUCT(COUNTIFS({REV}!$AJ$2:$AJ${LAST},{pfs32},"
                f'{REV}!$B$2:$B${LAST},"<>"))'
                f'+COUNTIFS({REV}!$AJ$2:$AJ${LAST},"TDD Cyber",'
                f'{REV}!$B$2:$B${LAST},"<>"))')
    coe_roles = (f'COUNTIFS({REV}!$AJ$2:$AJ${LAST},"COE*",{REV}!$B$2:$B${LAST},"<>")'
                 f'+COUNTIFS({REV}!$AJ$2:$AJ${LAST},"EGI",{REV}!$B$2:$B${LAST},"<>")')
    all_roles = f"COUNTA({REV}!$B$2:$B${LAST})"
    opts.row(ws, r, 2, [None] * len(H32), [None] * len(H32), bg=opts.GREY)
    # the sentence is built from the three counts rather than typed beside them, so the
    # row cannot state one number and mean another. It is left unbolded on purpose: it
    # runs across B to F and bold text at this length would not fit the five columns.
    ws.cell(r, 2).value = (
        f'="Roles in the organisation, all lines and squads: "&TEXT({all_roles},"0")'
        f'&" - portfolios "&TEXT({pf_roles},"0")'
        f'&", COEs and EGI "&TEXT({coe_roles},"0")&", each counted once"')
    ws.cell(r, 2).font = opts.BODY
    ws.cell(r, 2).alignment = opts.LFT
    # C to F stay empty so the sentence has the room to render; the count itself sits in
    # the column that is headed for it, which is what 4.0 reads
    f2._m(ws, r, 7, f"={all_roles}", opts.CT).font = opts.BOLD
    r += 1
    ws.cell(r, 2).value = ("Control - the portfolios plus the COEs and EGI against the "
                           "ledger, must be 0")
    ws.cell(r, 2).font = opts.BODY
    f2._m(ws, r, 7, f"=({pf_roles})+({coe_roles})-{all_roles}", opts.CTL_C)
    r += 2

    # ---- what the allowance is built from, in full, on the page ----
    # The rate on the block above is an allocated share: half a Head of Technology per
    # portfolio, 30% of a manager per platform. The actual is whole heads. That is the
    # whole reason 11.61 of allowance sits against 23.29 of cost, and a reader cannot judge
    # it unless the full role cost and the allocation are both stated, so they are.
    r = opts.bar(ws, r, 2, 5, "What the allocated rate is built from")
    # rate, times applied and allowance are already on the block above; repeating them here
    # was half a table restating the other half
    r = opts.head(ws, r, 2,
                  ["Input", "Where it is set on 0.2 Data Config", "Full role cost ($m)",
                   "Allocation %", "Allocated rate ($m)"],
                  [26, 42, 14, 12, 16])
    # 0.2's overhead tables moved two columns right (K:N) for the owner's Notes and
    # Actions columns, so cost sits in L and allocation in M
    for lab, where, full, pct, i in (
            ("Head of Technology", "Portfolio Overhead, Head of Tech", "$L$6", "$M$6", 2),
            ("Business Partner", "Portfolio Overhead, Business Partner", "$L$7", "$M$7", 3),
            ("Domain Architect", "Portfolio Overhead, Domain Architect", "$L$8", "$M$8", 4),
            ("Delivery Manager", "Platform Overhead, Delivery Manager", "$L$14", "$M$14",
             5),
            ("Technology Manager", "Platform Overhead, Tech Manager", "$L$15", "$M$15", 6),
            ("Leadership - 8 GMs", "Portfolio Overhead, Leadership Overhead", "$L$9",
             "$M$9", 7)):
        ws.cell(r, 2).value = lab
        ws.cell(r, 3).value = where
        for c in (2, 3):
            ws.cell(r, c).font = opts.BODY
            ws.cell(r, c).alignment = opts.LFT
        f2._m(ws, r, 4, f"='0.2 Data Config'!{full}", opts.M3)
        f2._m(ws, r, 5, f"='0.2 Data Config'!{pct}", opts.PCT)
        f2._m(ws, r, 6, f"=Lists!$AG${i}", opts.M3)
        r += 1
    r += 1
    r += 1
    ws.cell(r, 2).value = ("Allocated rate = cost x allocation. "
                           "Cost is roles in role mapping")
    ws.cell(r, 2).font = opts.BODY
    ws.cell(r, 2).alignment = opts.LFT
    # The rate block above reset the widths of B to F, which the main table shares, so the
    # header of the main table is measured again against the widths it will actually be
    # rendered at. Sizing it when it was written measured columns that no longer exist.
    hdr = st - 1
    lines_hi = max(opts.wrap_lines(t, ws.column_dimensions[L(2 + n)].width or 8.43)
                   for n, t in enumerate(H32))
    ws.row_dimensions[hdr].height = max(32, 14 * lines_hi + 6)
    # every row 4.0 points at, by name. The keys ending in 32 are the ones final4x lifts
    # straight into its anchor map, so a row that moves here moves the check with it.
    return {"first32": st, "last32": lines[-1], "ohtot32": tot32, "ohpf32": ohpf,
            "ohout32": ohout, "all32": allrow}


H33 = ["Portfolio", "How it is funded", "Squad", "Archetype Type", "Squad Size",
       "Archetype roles", "Total roles", "Filled", "Vacant",
       "Total roles after decisions", "Archetype cost ($m)", "Actual cost ($m)",
       "Variance to archetype ($m)", "Cost after decisions ($m)"]
# "How it is funded" holds one of four words, and the longest of them - "No figure to
# compare" - is 20 characters against a 17-wide column with the squad name in the cell
# beside it, so eleven rows on the tab read "No figure to compa". The column is as wide as
# its own longest value.
W33 = [22, 22, 30, 24, 7, 11, 7, 7, 8, 13, 13, 13, 15, 19]
# one entry per column B..N, indexed F33[c - 2]
F33 = [None, None, None, None, None, opts.C1, opts.CT, opts.CT, opts.CT, opts.CT,
       opts.M2, opts.M2, opts.M2, opts.M2]
# the totals cover the columns that can be added across every kind of row: roles, filled,
# vacant, actual cost and cost after decisions.
T33 = (8, 9, 10, 11, 13, 15)
# Archetype roles, archetype cost and the variance are the other three. Only some rows carry
# an archetype, so these were left as a flat dash - which printed "-" against a portfolio
# whose squads above it carry 9.88 of archetype and a real variance, and a reader has no way
# to tell that dash from "there is nothing here". They now add the rows that do carry a
# figure and state a dash only where no row above them carries one at all. The variance is
# actual less archetype on the total row itself, never the sum of the row variances: a row
# with a dash in the archetype column drops out of a SUM and would leave a variance covering
# some of the block measured against an actual covering all of it.
A33 = (7, 12, 14)


def build_33(wb, anchors):
    ws = wb["3.3 FTE View"]
    f2.wipe(ws)
    ws.column_dimensions["A"].width = 2
    ws.cell(2, 2).value = "Squad Detail - roles and cost, squad by squad"
    ws.cell(2, 2).font = opts.TITLE
    pfs, coes = order(anchors, wb)
    r = opts.bar(ws, 4, 2, len(H33), "Every squad on every working tab")
    r = opts.head(ws, r, 2, H33, W33)
    pf_rows = []
    for pf, tab, a in pfs + coes:
        st = r
        for kind, names in (("Archetype", a["squads"]),
                            ("Directly funded", a["direct"]),
                            ("No figure to compare", a["nofig"]),
                            ("Overhead", a["overhead"])):
            for gname in names:
                # a group named on the anchor with no row on the working tab has nothing
                # to read, so it is skipped rather than pointing every column at a
                # row number that does not exist
                s = a["srow"].get(gname)
                if not s:
                    continue
                ws.cell(r, 2).value = pf
                ws.cell(r, 3).value = kind
                for c in (2, 3):
                    ws.cell(r, c).font = opts.BODY
                    ws.cell(r, c).alignment = opts.LFT
                for i, k in enumerate(("squad", "type", "size", "aroles", "roles",
                                       "filled", "vacant", "rafter", "acost", "actual",
                                       "var", "after")):
                    c = 4 + i
                    x = ws.cell(r, c)
                    x.value = f"='{tab}'!${L(S[k])}${s}"
                    x.font = opts.BODY
                    if F33[c - 2]:
                        x.number_format, x.alignment = F33[c - 2], opts.RGT
                    else:
                        x.alignment = opts.LFT
                r += 1
        if r == st:
            # a portfolio the workbook carries no squad row for gets no total row either:
            # SUM over a range that runs backwards is not a zero, it is an error
            continue
        opts.row(ws, r, 2, [f"{pf} total"] + [None] * (len(H33) - 1),
                 [None] * len(H33), bg=opts.GREY, bold=True)
        ws.cell(r, 2).alignment = opts.LFT
        for c in range(6, 2 + len(H33)):
            x = ws.cell(r, c)
            if c in T33:
                x.value = f"=SUM({L(c)}{st}:{L(c)}{r-1})"
            elif c == 14:
                x.value = (f'=IF(ISNUMBER($L{r}),ROUND($M{r}-$L{r},6),"-")')
            elif c in A33:
                x.value = (f'=IF(COUNT({L(c)}{st}:{L(c)}{r-1})=0,"-",'
                           f"SUM({L(c)}{st}:{L(c)}{r-1}))")
            else:
                x.value = '="-"'
            x.number_format, x.alignment = F33[c - 2] or opts.M2, opts.RGT
        pf_rows.append(r)
        r += 1
    opts.row(ws, r, 2, ["Group total"] + [None] * (len(H33) - 1), [None] * len(H33),
             bg=opts.MID, bold=True, top=True)
    ws.cell(r, 2).alignment = opts.LFT
    for c in range(6, 2 + len(H33)):
        x = ws.cell(r, c)
        cells = ",".join(f"{L(c)}{p}" for p in pf_rows)
        if c in T33:
            x.value = "=" + "+".join(f"{L(c)}{p}" for p in pf_rows)
        elif c == 14:
            x.value = f'=IF(ISNUMBER($L{r}),ROUND($M{r}-$L{r},6),"-")'
        elif c in A33:
            # SUM ignores the portfolio totals that carry a dash, so this is the archetype
            # where there is one - the same rule the row above it uses one level down
            x.value = f'=IF(COUNT({cells})=0,"-",SUM({cells}))'
        else:
            x.value = '="-"'
        x.number_format, x.alignment = F33[c - 2] or opts.M2, opts.RGT
    gt = r
    r += 2
    for lab, col, f, nf in (
            ("Control - roles against the ledger, must be 0", 8,
             f"=$H${gt}-COUNTA({REV}!$B$2:$B${LAST})", opts.CTL_C),
            ("Control - cost against the ledger ($m), must be 0", 13,
             f"=ROUND($M${gt}-SUM({REV}!$AA$2:$AA${LAST})/1000000,6)", opts.CTL_M)):
        ws.cell(r, 2).value = lab
        ws.cell(r, 2).font = opts.BODY
        f2._m(ws, r, col, f, nf)
        r += 1
    return {"group_total": gt}


def run(src, dst):
    global LAST
    f2._boot_last(src)
    LAST = f2.LAST
    wb = openpyxl.load_workbook(src)
    anchors = json.load(open("anchors_final.json"))
    msg, wcol = patch_lists(wb)
    out = [msg]
    a31 = build_31(wb, anchors)
    a32 = build_32(wb, wcol)
    a33 = build_33(wb, anchors)
    json.dump({"3.1": a31, "3.2": a32, "3.3": a33},
              open("anchors_final3.json", "w"), indent=1)
    wb.save(dst)
    return out + [
        f"3.1: cost bridge, every directly funded programme named, ledger row "
        f"{a31['total']}, grand total row {a31['grand']}",
        "3.2: overhead and leadership, line by line, plus what the allowance is built from",
        f"3.3: every squad by portfolio, total row {a33['group_total']}"]


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
