"""Layout options that keep the owner's model, not just his colours.

The first set of mockups dropped three things he built deliberately, so all four 1.x
options were wrong for the same reason:

  AU / NZ            every squad routes its TDD cost to an AU or an NZ column, and the
                     budget is held and compared separately for each
  TDD vs Other       Total squad cost x Support % is the TDD lights-on cost. The
                     remainder is funded outside TDD - recharged elsewhere. Both are
                     stated on every squad row and roll up to the portfolio
  the budget table   Budget line / Budget / Amount that can be allocated to people /
                     Remaining for non-people, reconciled to the Finance pack

Every option below carries all three. They differ only in where the blocks sit.

Each 3.x option shows all three summary tabs, because a summary set has to be judged as a
set.
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L

SP = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(SP, "mocks2")

# two navies, taken from 1.11 BP&T: the section bar is the darker one, the column
# header the lighter. Collapsing them into one is what made the mockups read as a slab.
BARC = "FF002F6C"
NAVY, PALE, GREY, MID, YEL = "FF1F4E79", "FFDDEBF7", "FFF2F2F2", "FFD9D9D9", "FFFFFF00"
# openpyxl's default theme font is Cambria, a serif. The owner's tabs are Calibri, so
# every Font below names it explicitly or the render comes out in the wrong typeface.
FN = "Calibri"
thin = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
TOPR = Border(top=Side(style="medium", color="FF1F4E79"), left=thin, right=thin,
              bottom=thin)
TITLE = Font(name=FN, bold=True, size=16)
BARF = Font(name=FN, bold=True, size=11, color="FFFFFFFF")
HDRF = Font(name=FN, bold=True, size=11, color="FFFFFFFF")
BOLD = Font(name=FN, bold=True, size=11)
BODY = Font(name=FN, size=11)
NOTE = Font(name=FN, size=11)
M2 = '#,##0.00;(#,##0.00);"-"'
M0 = '#,##0;(#,##0);"-"'
CT = '#,##0;(#,##0);"-"'
C1 = '#,##0.0;(#,##0.0);"-"'
PC = '0%'
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center")
RGT = Alignment(horizontal="right", vertical="center")


def fill(c):
    return PatternFill("solid", start_color=c, end_color=c)


def bar(ws, r, c0, n, text):
    for i in range(n):
        x = ws.cell(r, c0 + i)
        x.fill = fill(BARC)
        x.font = BARF
    ws.cell(r, c0).value = text
    ws.cell(r, c0).alignment = LFT
    ws.row_dimensions[r].height = 18


def head(ws, r, c0, labels, widths):
    for i, t in enumerate(labels):
        x = ws.cell(r, c0 + i)
        x.value, x.font, x.fill, x.alignment, x.border = t, HDRF, fill(NAVY), CEN, BOX
        if widths[i]:
            ws.column_dimensions[L(c0 + i)].width = widths[i]
    ws.row_dimensions[r].height = 30


def row(ws, r, c0, vals, fmts, bg=None, bold=False, top=False):
    for i, v in enumerate(vals):
        x = ws.cell(r, c0 + i)
        x.value = v
        x.font = BOLD if bold else BODY
        x.border = TOPR if top else BOX
        if bg:
            x.fill = fill(bg)
        f = fmts[i]
        if f:
            x.number_format, x.alignment = f, RGT
        else:
            x.alignment = LFT
    return r + 1


def pairs(ws, r, c0, items, w=(48, 14)):
    ws.column_dimensions[L(c0)].width = w[0]
    ws.column_dimensions[L(c0 + 1)].width = w[1]
    for lab, val, f, b in items:
        a, x = ws.cell(r, c0), ws.cell(r, c0 + 1)
        a.value, a.font, a.alignment, a.border = lab, (BOLD if b else BODY), LFT, BOX
        x.value, x.font, x.number_format, x.alignment, x.border = \
            val, (BOLD if b else BODY), f, RGT, BOX
        if b:
            a.fill = x.fill = fill(MID)
        r += 1
    return r


def note(ws, r, c0, t):
    ws.cell(r, c0).value, ws.cell(r, c0).font, ws.cell(r, c0).alignment = t, NOTE, LFT
    return r + 1


def newbook(name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return wb, ws


def addsheet(wb, name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


# ------------------------------------------------------------------------ 1.x content
SUMMARY = [("Portfolio overhead (see 0.2 Data Config)", 0.7975, 0, 0),
           ("Platform overheads", 0.66, 0, 0),
           ("Squad support costs", 4.622, 0, 8.182)]
BUDGET = [("TDD lights-on budget - people ($m)", 5.50), ("AU budget ($m)", 5.50),
          ("NZ budget ($m)", 0.00)]
OVERUNDER = [("AU over/(under) budget ($m)", 0.5795), ("NZ over/(under) budget ($m)", 0.0),
             ("TDD over/(under) budget ($m)", 0.5795)]
FUNDING = [("Over/(under) TDD budget", 0.5795), ("Still to fund outside TDD", -0.002),
           ("Total to fund", 0.5775)]
OTHERFUND = [("Retail Lights On", 13.9, 3.4), ("OpEx initiatives", 1.3, 0.0),
             ("Retail CapEx", 11.8, 5.26), ("Significant items", 6.8, 1.404),
             ("Significant items EGI", 0.0, 1.52)]
# platform -> [(squad, type, size, on/off, AU/NZ, support%, total cost)]
PLATFORMS = [
    ("Store Operations", [("POS", "Configuration / Integration", "M", "Onshore", "AU", 0.7, 2.10),
                          ("Payments", "Configuration / Integration", "S", "Onshore", "AU", 0.2, 1.40),
                          ("Store Operations", "Operations", "M", "Onshore", "AU", 1.0, 1.40),
                          ("Deployment", "Operations", "S", "Onshore", "AU", 0.2, 0.80)]),
    ("Above Store", [("Above Store", "Configuration / Integration", "M", "Onshore", "AU", 0.2, 2.10)]),
    ("AmPOS", [("AmPOS", "Strategic Programs", "-", "Onshore", "AU", 0.0, 1.404)]),
    ("Network / QSR", [("Network & QSR", "Configuration / Integration", "S", "Onshore", "AU", 0.2, 1.40)]),
    ("Data AU", [("Data AU", "Enterprise Data and Insights", "M", "Offshore", "AU", 0.9, 0.68)]),
]
PLAT_OH = 0.165

SQCOLS = ["Squad", "Squad type", "Size", "On/Off", "AU / NZ", "Support %",
          "Total squad cost ($m)", "TDD cost ($m)", "Funded outside TDD ($m)"]
SQW = [26, 27, 7, 10, 9, 10, 14, 12, 17]


def squad_rows(ws, r, squads, plat=None, platcol=False):
    st = r
    for s in squads:
        vals = ([plat] if platcol else []) + [s[0], s[1], s[2], s[3], s[4], s[5],
                                              s[6], s[6] * s[5], s[6] * (1 - s[5])]
        r = row(ws, r, 2, vals,
                ([None] if platcol else []) +
                [None, None, None, None, None, PC, M2, M2, M2])
        base = 3 if platcol else 2
        for c in (base + 2, base + 3, base + 4):
            ws.cell(r - 1, c).alignment = CEN
            ws.cell(r - 1, c).fill = fill(YEL)
    return r, st


# =========================================================== 1.x  four options
def _plat_blocks(ws, r):
    tot = []
    for name, squads in PLATFORMS:
        bar(ws, r, 2, len(SQCOLS), f"Platform: {name}"); r += 1
        head(ws, r, 2, SQCOLS, SQW); r += 1
        r, st = squad_rows(ws, r, squads)
        r = row(ws, r, 2, ["Platform overhead", None, None, None, None, None,
                           None, PLAT_OH, None],
                [None]*6 + [M2, M2, M2])
        r = row(ws, r, 2, [f"{name} total", None, None, None, None, None,
                           f"=SUM(H{st}:H{r-2})", f"=SUM(I{st}:I{r-1})",
                           f"=SUM(J{st}:J{r-2})"],
                [None]*6 + [M2, M2, M2], bg=GREY, bold=True)
        tot.append(r - 1); r += 1
    return r, tot


def _summary_block(ws, r, c0=2):
    head(ws, r, c0, ["Cost", "TDD AU ($m)", "TDD NZ ($m)", "Other ($m)", "Total ($m)"],
         [40, 12, 12, 12, 12]); r += 1
    st = r
    for lab, au, nz, ot in SUMMARY:
        r = row(ws, r, c0, [lab, au, nz, ot, au + nz + ot], [None, M2, M2, M2, M2])
    r = row(ws, r, c0, ["Total cost"] +
            [f"=SUM({L(c0+i)}{st}:{L(c0+i)}{r-1})" for i in range(1, 5)],
            [None, M2, M2, M2, M2], bg=MID, bold=True)
    return r


def _otherfund_block(ws, r, c0):
    head(ws, r, c0, ["Budget line", "Budget ($m)",
                     "Allocated to people ($m)", "Non-people ($m)"],
         [28, 14, 20, 16]); r += 1
    st = r
    for lab, b, p in OTHERFUND:
        r = row(ws, r, c0, [lab, b, p, b - p], [None, M2, M2, M2])
    r = row(ws, r, c0, ["Total applied"] +
            [f"=SUM({L(c0+i)}{st}:{L(c0+i)}{r-1})" for i in range(1, 4)],
            [None, M2, M2, M2], bg=GREY, bold=True)
    r = row(ws, r, c0, ["Other cost (this model)", None, 8.182, None],
            [None, M2, M2, M2])
    r = row(ws, r, c0, ["Left to fund", None, -0.002, None],
            [None, M2, M2, M2], bg=MID, bold=True)
    return r


def one_A():
    """The owner's own arrangement, cleaned. Summary beside budget, then platforms."""
    wb, ws = newbook("1.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - squad design"
    ws.cell(2, 2).font = TITLE
    bar(ws, 4, 2, 5, "Portfolio summary")
    _summary_block(ws, 5)
    bar(ws, 4, 8, 2, "Budget against TDD cost")
    r = pairs(ws, 5, 8, [(l, v, M2, False) for l, v in BUDGET] +
              [(l, v, M2, i == 2) for i, (l, v) in enumerate(OVERUNDER)], w=(44, 14))
    bar(ws, 11, 2, 2, "Funding position")
    pairs(ws, 12, 2, [(l, v, M2, i == 2) for i, (l, v) in enumerate(FUNDING)])
    bar(ws, 11, 8, 4, "Other funding")
    _otherfund_block(ws, 12, 8)
    bar(ws, 22, 2, 4, "Budget reconciliation")
    pairs(ws, 23, 2, [("Total Retail budget ($m)", 33.8, M2, False),
                      ("Reconciled to Finance ($m)", 33.8, M2, True)])
    r = 27
    r, tot = _plat_blocks(ws, r)
    note(ws, r, 2, "Yellow cells are inputs. Support % splits total squad cost into TDD "
                   "cost and cost funded outside TDD. AU / NZ routes the TDD cost.")
    return wb


def one_B():
    """Blocks stacked in one column, then one squad table with platform as a column."""
    wb, ws = newbook("1.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - squad design"
    ws.cell(2, 2).font = TITLE
    r = 4
    bar(ws, r, 2, 5, "Portfolio summary"); r += 1
    r = _summary_block(ws, r) + 1
    bar(ws, r, 2, 2, "Budget against TDD cost"); r += 1
    r = pairs(ws, r, 2, [(l, v, M2, False) for l, v in BUDGET] +
              [(l, v, M2, i == 2) for i, (l, v) in enumerate(OVERUNDER)] +
              [(l, v, M2, i == 2) for i, (l, v) in enumerate(FUNDING)]) + 1
    bar(ws, r, 2, 4, "Other funding"); r += 1
    r = _otherfund_block(ws, r, 2) + 1
    bar(ws, r, 2, 10, "Squads"); r += 1
    head(ws, r, 2, ["Platform"] + SQCOLS, [20] + SQW); r += 1
    st = r
    for name, squads in PLATFORMS:
        r, _ = squad_rows(ws, r, squads, plat=name, platcol=True)
    r = row(ws, r, 2, ["Total", None, None, None, None, None, None,
                       f"=SUM(I{st}:I{r-1})", f"=SUM(J{st}:J{r-1})",
                       f"=SUM(K{st}:K{r-1})"],
            [None]*7 + [M2, M2, M2], bg=MID, bold=True, top=True)
    note(ws, r + 1, 2, "Platform overheads are in the portfolio summary above rather than "
                       "repeated per platform.")
    return wb


def one_C():
    """Two columns the whole way down: money on the left, squads on the right."""
    wb, ws = newbook("1.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - squad design"
    ws.cell(2, 2).font = TITLE
    bar(ws, 4, 2, 2, "Budget against TDD cost")
    r = pairs(ws, 5, 2, [(l, v, M2, False) for l, v in BUDGET] +
              [(l, v, M2, i == 2) for i, (l, v) in enumerate(OVERUNDER)])
    r += 1
    bar(ws, r, 2, 2, "Funding position"); r += 1
    r = pairs(ws, r, 2, [(l, v, M2, i == 2) for i, (l, v) in enumerate(FUNDING)]) + 1
    bar(ws, r, 2, 4, "Other funding"); r += 1
    _otherfund_block(ws, r, 2)
    # right: summary then platforms
    bar(ws, 4, 7, 5, "Portfolio summary")
    rr = _summary_block(ws, 5, 7) + 2
    for name, squads in PLATFORMS:
        bar(ws, rr, 7, len(SQCOLS), f"Platform: {name}"); rr += 1
        head(ws, rr, 7, SQCOLS, SQW); rr += 1
        st = rr
        for s in squads:
            rr = row(ws, rr, 7, [s[0], s[1], s[2], s[3], s[4], s[5], s[6],
                                 s[6]*s[5], s[6]*(1-s[5])],
                     [None, None, None, None, None, PC, M2, M2, M2])
            for c in (10, 11, 12):
                ws.cell(rr-1, c).alignment = CEN
                ws.cell(rr-1, c).fill = fill(YEL)
        rr = row(ws, rr, 7, [f"{name} total", None, None, None, None, None,
                             f"=SUM(M{st}:M{rr-1})", f"=SUM(N{st}:N{rr-1})",
                             f"=SUM(O{st}:O{rr-1})"],
                 [None]*6 + [M2, M2, M2], bg=GREY, bold=True)
        rr += 1
    return wb


def one_D():
    """A number strip on top carrying AU, NZ and Other, then budget, then platforms."""
    wb, ws = newbook("1.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - squad design"
    ws.cell(2, 2).font = TITLE
    tiles = [("TDD AU ($m)", 6.08), ("TDD NZ ($m)", 0.00), ("Other ($m)", 8.18),
             ("Total cost ($m)", 14.26), ("TDD budget ($m)", 5.50),
             ("Over/(under) budget ($m)", 0.58)]
    c = 2
    for lab, v in tiles:
        h = ws.cell(4, c)
        h.value, h.font, h.fill, h.alignment, h.border = lab, HDRF, fill(NAVY), CEN, BOX
        x = ws.cell(5, c)
        x.value, x.font, x.number_format = v, Font(name=FN, bold=True, size=16), M2
        x.alignment, x.border, x.fill = CEN, BOX, fill(GREY)
        ws.column_dimensions[L(c)].width = 19
        c += 1
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 32
    r = 7
    bar(ws, r, 2, 4, "Other funding"); r += 1
    r = _otherfund_block(ws, r, 2) + 1
    bar(ws, r, 2, 10, "Squads"); r += 1
    head(ws, r, 2, ["Platform"] + SQCOLS, [20] + SQW); r += 1
    st = r
    for name, squads in PLATFORMS:
        r, _ = squad_rows(ws, r, squads, plat=name, platcol=True)
    r = row(ws, r, 2, ["Total", None, None, None, None, None, None,
                       f"=SUM(I{st}:I{r-1})", f"=SUM(J{st}:J{r-1})",
                       f"=SUM(K{st}:K{r-1})"],
            [None]*7 + [M2, M2, M2], bg=MID, bold=True, top=True)
    return wb


# =========================================================== 3.x  four options
# every option renders 3.1, 3.2 and 3.3 so the set can be judged as a set
PF = [("Ampol Retail", 5.50, 0.00, 6.08, 0.00, 8.18, 9.88, 70, 22),
      ("Customer", 2.50, 4.00, 9.31, 2.16, 5.66, 12.06, 83, 13),
      ("Enterprise Data", 3.50, 0.00, 4.29, 0.00, 2.13, 4.30, 28, 9),
      ("TDD Group Functions", 4.50, 1.00, 7.32, 1.50, 1.21, 5.70, 46, 14),
      ("P&C", 1.00, 1.00, 2.97, 0.00, 1.21, 3.50, 18, 10),
      ("Finance", 1.00, 1.00, 2.63, 0.00, 1.21, 2.80, 17, 8),
      ("Infrastructure", 2.50, 0.00, 3.05, 0.00, 4.70, 6.40, 36, 12),
      ("Energy Solutions & B2B", 2.50, 0.00, 3.29, 0.00, 3.16, 7.90, 33, 9),
      ("Commercial Fuels", 2.50, 0.00, 5.49, 0.00, 4.71, 4.88, 42, 9),
      ("Z Retail", 0.00, 6.50, 0.00, 4.63, 2.71, 6.78, 39, 7),
      ("COE Cyber", 2.50, 1.00, 8.84, 1.06, 0.00, None, 46, 4),
      ("COE BP&T", 3.00, 1.00, 5.43, 0.74, 0.00, None, 24, 9),
      ("COE SA&D", 3.00, 1.00, 5.60, 0.93, 0.00, None, 26, 9),
      ("EGI", 0.00, 0.00, 4.94, 0.00, 0.00, None, 17, 0)]
SQ33 = [("Ampol Retail", "Store Operations", "POS", "Config / Integration", "M", 7.5,
         11, 4, 2.10, 2.47, 1.47), 
        ("Ampol Retail", "Store Operations", "Payments", "Config / Integration", "S", 4.0,
         4, 2, 1.40, 0.89, 0.28),
        ("Ampol Retail", "Above Store", "Above Store", "Config / Integration", "M", 7.5,
         8, 1, 2.10, 1.58, 0.42),
        ("Customer", "Ampol Digital", "Ampol Web", "Product", "M", 8.5, 8, 0, 1.90, 1.59,
         1.59),
        ("Customer", "Ampol Digital", "Ampol App", "Product", "L", 12.0, 14, 3, 2.60,
         2.94, 2.94),
        ("Enterprise Data", "Group Data", "Data Platforms", "Ent Data & Insights", "M",
         7.5, 9, 6, 1.70, 1.79, 1.79)]


def _t31(ws, mode):
    ws.cell(2, 2).value = "3.1 Group summary - budget against cost"
    ws.cell(2, 2).font = TITLE
    r = 4
    if mode == "tiles":
        for i, (lab, v, f) in enumerate([("Roles", 525, CT), ("Vacant", 135, CT),
                                         ("TDD AU ($m)", 69.24, M2),
                                         ("TDD NZ ($m)", 11.02, M2),
                                         ("Other ($m)", 34.85, M2),
                                         ("Total ($m)", 115.11, M2)]):
            h = ws.cell(4, 2 + i)
            h.value, h.font, h.fill, h.alignment, h.border = lab, HDRF, fill(NAVY), CEN, BOX
            x = ws.cell(5, 2 + i)
            x.value, x.font, x.number_format = v, Font(name=FN, bold=True, size=15), f
            x.alignment, x.border, x.fill = CEN, BOX, fill(GREY)
            ws.column_dimensions[L(2 + i)].width = 15
        ws.row_dimensions[5].height = 28
        r = 8
    bar(ws, r, 2, 10, "Budget against cost by portfolio"); r += 1
    if mode == "groups":
        for c0, n, lab in ((3, 2, "BUDGET"), (5, 3, "TDD COST"), (8, 2, "OVER/(UNDER)")):
            for i in range(n):
                x = ws.cell(r, c0 + i)
                x.fill, x.font, x.alignment = fill(NAVY), BARF, CEN
            ws.cell(r, c0).value = lab
            ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0+n-1)
        r += 1
        head(ws, r, 2, ["Portfolio", "AU ($m)", "NZ ($m)", "AU ($m)", "NZ ($m)",
                        "Other ($m)", "AU ($m)", "NZ ($m)", "Left to fund ($m)"],
             [26, 11, 11, 11, 11, 11, 11, 11, 14])
    else:
        head(ws, r, 2, ["Portfolio", "AU budget ($m)", "NZ budget ($m)",
                        "TDD AU ($m)", "TDD NZ ($m)", "Other ($m)",
                        "AU over/(under) ($m)", "NZ over/(under) ($m)",
                        "Left to fund ($m)"],
             [26, 13, 13, 12, 12, 12, 15, 15, 14])
    r += 1
    st = r
    for p, ab, nb, au, nz, ot, arch, ro, vc in PF:
        r = row(ws, r, 2, [p, ab, nb, au, nz, ot, au - ab, nz - nb,
                           (au - ab) + (nz - nb)],
                [None, M2, M2, M2, M2, M2, M2, M2, M2])
    r = row(ws, r, 2, ["Total"] + [f"=SUM({L(2+i)}{st}:{L(2+i)}{r-1})"
                                   for i in range(1, 9)],
            [None] + [M2]*8, bg=MID, bold=True, top=True)
    note(ws, r + 1, 2, "Budget is the lights-on people allocation on 0.2 Data Config. "
                       "Other is cost recharged outside TDD.")


def _t32(ws, mode):
    ws.cell(2, 2).value = "3.2 Total cost - design against actual"
    ws.cell(2, 2).font = TITLE
    if mode == "bridge":
        r = 4
        bar(ws, r, 2, 2, "Cost bridge"); r += 1
        r = pairs(ws, r, 2, [
            ("Squad archetype cost - the design", 64.20, M2, False),
            ("Delivery squads raised over the archetype", 39.26, M2, False),
            ("Overhead roles inside the ledger", 11.65, M2, False),
            ("Cost of the organisation today", 115.11, M2, True),
            ("Impact of the vacancy decisions", 0.00, M2, False),
            ("Cost after decisions", 115.11, M2, True)], w=(52, 14)) + 1
        bar(ws, r, 2, 2, "Split of the cost after decisions"); r += 1
        pairs(ws, r, 2, [("TDD lights on - AU", 69.24, M2, False),
                         ("TDD lights on - NZ", 11.02, M2, False),
                         ("Funded outside TDD (recharged)", 34.85, M2, False),
                         ("Total", 115.11, M2, True)], w=(52, 14))
        return
    r = 4
    bar(ws, r, 2, 8, "Archetype against actual by portfolio"); r += 1
    if mode == "groups":
        for c0, n, lab in ((3, 1, "DESIGN"), (4, 2, "TODAY"), (6, 3, "AFTER DECISIONS")):
            for i in range(n):
                x = ws.cell(r, c0 + i)
                x.fill, x.font, x.alignment = fill(NAVY), BARF, CEN
            ws.cell(r, c0).value = lab
            if n > 1:
                ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0+n-1)
        r += 1
    head(ws, r, 2, ["Portfolio", "Archetype cost ($m)", "Actual cost ($m)",
                    "Variance to archetype ($m)", "Impact of decisions ($m)",
                    "Total after decisions ($m)", "New variance ($m)"],
         [26, 15, 14, 16, 15, 16, 14]); r += 1
    st = r
    for p, ab, nb, au, nz, ot, arch, ro, vc in PF:
        act = au + nz + ot
        r = row(ws, r, 2, [p, arch if arch else "-", act,
                           (act - arch) if arch else "-", 0.0, act,
                           (act - arch) if arch else "-"],
                [None, M2, M2, M2, M2, M2, M2])
    r = row(ws, r, 2, ["Total"] + [f"=SUM({L(2+i)}{st}:{L(2+i)}{r-1})"
                                   for i in range(1, 7)],
            [None] + [M2]*6, bg=MID, bold=True, top=True)
    r += 1
    bar(ws, r, 2, 7, "Overhead - allowance against actual"); r += 1
    head(ws, r, 2, ["Overhead line", "Roles", "Rate ($m)", "Units", "Allowance ($m)",
                    "Actual ($m)", "Over/(under) ($m)"],
         [26, 9, 11, 9, 14, 13, 16]); r += 1
    st2 = r
    for n, ro, rate, un, act in [("Head of Technology", 15, 0.1375, 10, 4.87),
                                 ("Business Partner", 6, 0.22, 10, 2.31),
                                 ("Domain Architect", 7, 0.14, 10, 1.66),
                                 ("Delivery Manager", 10, 0.084, 30, 2.67),
                                 ("Technology Manager", 24, 0.081, 30, 6.30),
                                 ("Leadership - 8 GMs", 8, 0.30, 10, 5.10)]:
        r = row(ws, r, 2, [n, ro, rate, un, rate * un, act, act - rate * un],
                [None, CT, '#,##0.000', CT, M2, M2, M2])
    row(ws, r, 2, ["Overhead total", f"=SUM(C{st2}:C{r-1})", None, None,
                   f"=SUM(F{st2}:F{r-1})", f"=SUM(G{st2}:G{r-1})",
                   f"=SUM(H{st2}:H{r-1})"],
        [None, CT, None, None, M2, M2, M2], bg=MID, bold=True, top=True)


def _t33(ws, mode):
    ws.cell(2, 2).value = "3.3 Squad detail - roles and cost, squad by squad"
    ws.cell(2, 2).font = TITLE
    r = 4
    bar(ws, r, 2, 12, "Delivery squads"); r += 1
    head(ws, r, 2, ["Portfolio", "Platform", "Squad", "Archetype type", "Size",
                    "Archetype roles", "Roles", "Vacant", "Archetype cost ($m)",
                    "Actual cost ($m)", "TDD cost ($m)", "Variance ($m)"],
         [20, 20, 24, 22, 7, 12, 8, 8, 13, 13, 12, 12]); r += 1
    st = r
    for pf, pl, sq, ty, sz, ar, ro, vc, ac, act, tdd in SQ33:
        r = row(ws, r, 2, [pf, pl, sq, ty, sz, ar, ro, vc, ac, act, tdd, act - ac],
                [None, None, None, None, None, C1, CT, CT, M2, M2, M2, M2])
    row(ws, r, 2, ["Total"] + [None]*4 +
        [f"=SUM({L(2+i)}{st}:{L(2+i)}{r-1})" for i in range(5, 12)],
        [None, None, None, None, None, C1, CT, CT, M2, M2, M2, M2],
        bg=MID, bold=True, top=True)
    r += 2
    note(ws, r, 2, "Overhead lines are on 3.2 against their allowance, because they are "
                   "not priced by a squad archetype.")


def _three(mode):
    wb, ws = newbook("3.1 Group Summary")
    _t31(ws, mode)
    _t32(addsheet(wb, "3.2 Total Cost"), mode)
    _t33(addsheet(wb, "3.3 Squad Detail"), mode)
    return wb


def three_A():
    return _three("plain")


def three_B():
    return _three("groups")


def three_C():
    return _three("bridge")


def three_D():
    return _three("tiles")


OPTIONS = {"1A": one_A, "1B": one_B, "1C": one_C, "1D": one_D,
           "3A": three_A, "3B": three_B, "3C": three_C, "3D": three_D}

if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    for k, fn in OPTIONS.items():
        fn().save(os.path.join(OUT, f"{k}.xlsx"))
        print("  wrote", k)
