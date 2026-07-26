"""Twelve rendered mockups: four layout options for 1.x, four for 2.x, four for 3.x.

Real numbers from the ledger so they read like the finished thing. These are standalone
files for looking at. Nothing here touches TDD_Cost_Calc.xlsx.
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L

SP = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(SP, "mocks")

NAVY = "FF1F4E79"
PALE = "FFDDEBF7"
GREY = "FFF2F2F2"
MID = "FFD9D9D9"
YEL = "FFFFFF00"
thin = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
TOP = Border(top=Side(style="medium", color="FF1F4E79"))

TITLE = Font(bold=True, size=14)
BARF = Font(bold=True, size=11, color="FFFFFFFF")
HDRF = Font(bold=True, size=10, color="FFFFFFFF")
BOLD = Font(bold=True, size=10)
BODY = Font(size=10)
NOTE = Font(size=9, color="FF595959")
M2 = '#,##0.00;(#,##0.00);"-"'
M0 = '#,##0;(#,##0);"-"'
CT = '#,##0;(#,##0);"-"'
C1 = '#,##0.0;(#,##0.0);"-"'
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center")
RGT = Alignment(horizontal="right", vertical="center")


def bar(ws, r, c0, n, text):
    for i in range(n):
        x = ws.cell(r, c0 + i)
        x.fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
        x.font = BARF
    ws.cell(r, c0).value = text
    ws.cell(r, c0).alignment = LFT
    ws.row_dimensions[r].height = 18


def head(ws, r, c0, labels, widths, align=None):
    for i, t in enumerate(labels):
        x = ws.cell(r, c0 + i)
        x.value = t
        x.font = HDRF
        x.fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
        x.alignment = CEN
        x.border = BOX
        ws.column_dimensions[L(c0 + i)].width = widths[i]
    ws.row_dimensions[r].height = 30


def row(ws, r, c0, vals, fmts, fill=None, bold=False, line=False):
    for i, v in enumerate(vals):
        x = ws.cell(r, c0 + i)
        x.value = v
        x.font = BOLD if bold else BODY
        x.border = BOX if not line else Border(left=thin, right=thin, top=thin,
                                               bottom=thin)
        if fill:
            x.fill = PatternFill("solid", start_color=fill, end_color=fill)
        f = fmts[i]
        if f:
            x.number_format = f
            x.alignment = RGT
        else:
            x.alignment = LFT


def pairs(ws, r, c0, items, w=(46, 12)):
    ws.column_dimensions[L(c0)].width = w[0]
    ws.column_dimensions[L(c0 + 1)].width = w[1]
    for lab, val, f, bold in items:
        a, b = ws.cell(r, c0), ws.cell(r, c0 + 1)
        a.value = lab
        a.font = BOLD if bold else BODY
        a.alignment = LFT
        a.border = BOX
        b.value = val
        b.font = BOLD if bold else BODY
        b.number_format = f
        b.alignment = RGT
        b.border = BOX
        if bold:
            for x in (a, b):
                x.fill = PatternFill("solid", start_color=MID, end_color=MID)
        r += 1
    return r


def note(ws, r, c0, text):
    ws.cell(r, c0).value = text
    ws.cell(r, c0).font = NOTE
    ws.cell(r, c0).alignment = LFT
    return r + 1


def newbook(name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return wb, ws


# ------------------------------------------------------------------ data for the mocks
SQUADS = [("Store Operations", "Configuration / Integration", "M", 7.5, 11, 4, 2.10, 2.47),
          ("Above Store", "Configuration / Integration", "M", 7.5, 8, 1, 2.10, 1.58),
          ("AmPOS", "Strategic Programs", "-", 0.0, 10, 2, 1.40, 2.12),
          ("Network & QSR", "Configuration / Integration", "S", 4.0, 6, 2, 1.40, 1.21),
          ("Data AU", "Enterprise Data and Insights", "M", 7.5, 8, 6, 0.68, 1.11),
          ("EGI Retail", "Strategic Programs", "-", 0.0, 6, 0, 1.52, 0.79),
          ("Payments", "Configuration / Integration", "S", 4.0, 4, 2, 1.40, 0.89),
          ("Deployment", "Operations", "S", 4.0, 5, 1, 0.60, 1.22)]
OH = [("Head of Technology", 1, 0, 0.19), ("Delivery Manager", 1, 0, 0.30),
      ("Technology Manager", 3, 0, 0.77)]
PEOPLE = [("Store Operations", [("Andrew Whitford", "Field Systems Operator", "Filled", 164033),
                                ("Ephraim Adan", "Manager - AU Retail", "Filled", 77708),
                                ("Geoff Goddard", "Field Systems Operator", "Filled", 167360),
                                ("Vacant", "Support Analyst - Field", "Vacant", 202853)]),
          ("Above Store", [("Vijay Patel", "Systems Analyst", "Filled", 193575),
                           ("Zhan Zhang", "Solution Architect", "Filled", 243638),
                           ("Vacant", "Quality Assurance", "Vacant", 170502)])]
PORTFOLIOS = [("Ampol Retail", 5.50, 14.01, 9.88, 70, 22),
              ("Customer", 6.50, 17.13, 12.06, 83, 13),
              ("Enterprise Data", 3.50, 6.42, 4.30, 28, 9),
              ("TDD Group Functions", 5.50, 10.03, 5.70, 46, 14),
              ("P&C", 2.00, 4.18, 3.50, 18, 10),
              ("Finance", 2.00, 3.84, 2.80, 17, 8),
              ("Infrastructure", 2.50, 7.75, 6.40, 36, 12),
              ("Energy Solutions & B2B", 2.50, 6.45, 7.90, 33, 9),
              ("Commercial Fuels", 2.50, 10.20, 4.88, 42, 9),
              ("Z Retail", 6.50, 7.34, 6.78, 39, 7),
              ("COE Cyber", 3.50, 9.90, 0, 46, 4),
              ("COE BP&T", 4.00, 6.40, 0, 24, 9),
              ("COE SA&D", 4.00, 6.53, 0, 26, 9),
              ("EGI", 0.00, 4.94, 0, 17, 0)]


# =========================================================== 1.x  four options
def one_A():
    """Stacked blocks. Every section a navy bar, one column down the page."""
    wb, ws = newbook("1.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - squad design"
    ws.cell(2, 2).font = TITLE
    r = 4
    bar(ws, r, 2, 2, "Portfolio summary"); r += 1
    r = pairs(ws, r, 2, [("Portfolio overhead ($m)", 0.80, M2, False),
                         ("Platform overheads ($m)", 0.66, M2, False),
                         ("Squad support costs ($m)", 12.80, M2, False),
                         ("Total design cost ($m)", 14.26, M2, True)])
    r += 1
    bar(ws, r, 2, 2, "Budget position"); r += 1
    r = pairs(ws, r, 2, [("TDD lights-on budget ($m)", 5.50, M2, False),
                         ("Design cost ($m)", 14.26, M2, False),
                         ("Over/(under) budget ($m)", 8.76, M2, True)])
    r += 2
    bar(ws, r, 2, 9, "Squads"); r += 1
    head(ws, r, 2, ["Squad", "Archetype type", "Size", "On/Off", "Support %",
                    "Squad cost ($m)", "TDD cost ($m)", "Funded outside TDD ($m)"],
         [30, 30, 8, 10, 11, 14, 13, 18]); r += 1
    st = r
    for s in SQUADS:
        row(ws, r, 2, [s[0], s[1], s[2], "Onshore", 0.7, s[6], s[6] * 0.7, s[6] * 0.3],
            [None, None, None, None, "0%", M2, M2, M2])
        ws.cell(r, 5).alignment = CEN
        ws.cell(r, 6).alignment = CEN
        for c in (5, 6):
            ws.cell(r, c).fill = PatternFill("solid", start_color=YEL, end_color=YEL)
        r += 1
    row(ws, r, 2, ["Total", None, None, None, None,
                   f"=SUM(G{st}:G{r-1})", f"=SUM(H{st}:H{r-1})", f"=SUM(I{st}:I{r-1})"],
        [None, None, None, None, None, M2, M2, M2], fill=MID, bold=True)
    r += 2
    note(ws, r, 2, "Yellow cells are inputs. Support % and On/Off drive the split "
                   "between TDD cost and cost funded outside TDD.")
    return wb


def one_B():
    """Two columns at the top: summary beside budget. Squad table full width below."""
    wb, ws = newbook("1.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - squad design"
    ws.cell(2, 2).font = TITLE
    bar(ws, 4, 2, 2, "Portfolio summary")
    pairs(ws, 5, 2, [("Portfolio overhead ($m)", 0.80, M2, False),
                     ("Platform overheads ($m)", 0.66, M2, False),
                     ("Squad support costs ($m)", 12.80, M2, False),
                     ("Total design cost ($m)", 14.26, M2, True)], w=(36, 12))
    bar(ws, 4, 6, 2, "Budget position")
    pairs(ws, 5, 6, [("TDD lights-on budget ($m)", 5.50, M2, False),
                     ("Design cost ($m)", 14.26, M2, False),
                     ("Over/(under) budget ($m)", 8.76, M2, True)], w=(30, 12))
    r = 11
    bar(ws, r, 2, 9, "Squads"); r += 1
    head(ws, r, 2, ["Squad", "Archetype type", "Size", "On/Off", "Support %",
                    "Squad cost ($m)", "TDD cost ($m)", "Funded outside TDD ($m)"],
         [30, 30, 8, 10, 11, 14, 13, 18]); r += 1
    st = r
    for s in SQUADS:
        row(ws, r, 2, [s[0], s[1], s[2], "Onshore", 0.7, s[6], s[6]*0.7, s[6]*0.3],
            [None, None, None, None, "0%", M2, M2, M2])
        for c in (5, 6):
            ws.cell(r, c).fill = PatternFill("solid", start_color=YEL, end_color=YEL)
            ws.cell(r, c).alignment = CEN
        r += 1
    row(ws, r, 2, ["Total", None, None, None, None, f"=SUM(G{st}:G{r-1})",
                   f"=SUM(H{st}:H{r-1})", f"=SUM(I{st}:I{r-1})"],
        [None, None, None, None, None, M2, M2, M2], fill=MID, bold=True)
    note(ws, r + 2, 2, "Yellow cells are inputs.")
    return wb


def one_C():
    """No separate platform blocks - platform becomes a column on one squad table."""
    wb, ws = newbook("1.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - squad design"
    ws.cell(2, 2).font = TITLE
    r = 4
    bar(ws, r, 2, 10, "Squads by platform"); r += 1
    head(ws, r, 2, ["Platform", "Squad", "Archetype type", "Size", "On/Off", "Support %",
                    "Squad cost ($m)", "TDD cost ($m)", "Funded outside TDD ($m)"],
         [22, 28, 28, 8, 10, 11, 14, 13, 18]); r += 1
    plats = ["Store Operations", "Above Store", "AmPOS", "Network / QSR",
             "Data AU", "EGI Retail", "Store Operations", "Store Operations"]
    st = r
    for p, s in zip(plats, SQUADS):
        row(ws, r, 2, [p, s[0], s[1], s[2], "Onshore", 0.7, s[6], s[6]*0.7, s[6]*0.3],
            [None, None, None, None, None, "0%", M2, M2, M2])
        for c in (6, 7):
            ws.cell(r, c).fill = PatternFill("solid", start_color=YEL, end_color=YEL)
            ws.cell(r, c).alignment = CEN
        r += 1
    row(ws, r, 2, ["Total", None, None, None, None, None, f"=SUM(H{st}:H{r-1})",
                   f"=SUM(I{st}:I{r-1})", f"=SUM(J{st}:J{r-1})"],
        [None, None, None, None, None, None, M2, M2, M2], fill=MID, bold=True)
    r += 2
    bar(ws, r, 2, 2, "Overheads and budget"); r += 1
    pairs(ws, r, 2, [("Portfolio overhead ($m)", 0.80, M2, False),
                     ("Platform overheads ($m)", 0.66, M2, False),
                     ("Total design cost ($m)", 14.26, M2, True),
                     ("TDD lights-on budget ($m)", 5.50, M2, False),
                     ("Over/(under) budget ($m)", 8.76, M2, True)])
    return wb


def one_D():
    """A number strip across the top, then one squad table."""
    wb, ws = newbook("1.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - squad design"
    ws.cell(2, 2).font = TITLE
    tiles = [("Design cost", 14.26, M2), ("Lights-on budget", 5.50, M2),
             ("Over budget", 8.76, M2), ("Squads", 8, CT), ("Designed roles", 47.5, C1)]
    c = 2
    for lab, v, f in tiles:
        ws.cell(4, c).value = lab
        ws.cell(4, c).font = HDRF
        ws.cell(4, c).fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
        ws.cell(4, c).alignment = CEN
        ws.cell(4, c).border = BOX
        x = ws.cell(5, c)
        x.value = v
        x.font = Font(bold=True, size=16)
        x.number_format = f
        x.alignment = CEN
        x.border = BOX
        x.fill = PatternFill("solid", start_color=GREY, end_color=GREY)
        ws.column_dimensions[L(c)].width = 17
        ws.row_dimensions[5].height = 30
        c += 1
    r = 8
    bar(ws, r, 2, 9, "Squads"); r += 1
    head(ws, r, 2, ["Squad", "Archetype type", "Size", "On/Off", "Support %",
                    "Squad cost ($m)", "TDD cost ($m)", "Funded outside TDD ($m)"],
         [30, 30, 8, 10, 11, 14, 13, 18]); r += 1
    st = r
    for s in SQUADS:
        row(ws, r, 2, [s[0], s[1], s[2], "Onshore", 0.7, s[6], s[6]*0.7, s[6]*0.3],
            [None, None, None, None, "0%", M2, M2, M2])
        for cc in (5, 6):
            ws.cell(r, cc).fill = PatternFill("solid", start_color=YEL, end_color=YEL)
            ws.cell(r, cc).alignment = CEN
        r += 1
    row(ws, r, 2, ["Total", None, None, None, None, f"=SUM(G{st}:G{r-1})",
                   f"=SUM(H{st}:H{r-1})", f"=SUM(I{st}:I{r-1})"],
        [None, None, None, None, None, M2, M2, M2], fill=MID, bold=True)
    return wb


# =========================================================== 2.x  four options
COLS2 = ["Squad", "Archetype type", "Size", "Arch roles", "Roles", "Vacant",
         "Hire/offshore", "Hold", "Roles after", "Arch cost ($m)", "Actual ($m)",
         "Variance ($m)", "Impact ($m)", "Total after ($m)"]
W2 = [28, 27, 7, 10, 7, 8, 12, 7, 11, 12, 11, 12, 11, 13]


def _squad_block(ws, r, with_oh=True):
    head(ws, r, 2, COLS2, W2); r += 1
    st = r
    for s in SQUADS:
        row(ws, r, 2, [s[0], s[1], s[2], s[3], s[4], s[5], s[5], 0, s[4],
                       s[6], s[7], s[7] - s[6], 0.0, s[7]],
            [None, None, None, C1, CT, CT, CT, CT, CT, M2, M2, M2, M2, M2])
        r += 1
    row(ws, r, 2, ["Delivery squads"] + [None]*2 +
        [f"=SUM({L(c)}{st}:{L(c)}{r-1})" for c in range(5, 16)],
        [None, None, None, C1, CT, CT, CT, CT, CT, M2, M2, M2, M2, M2],
        fill=GREY, bold=True)
    dr = r; r += 2
    if with_oh:
        bar(ws, r, 2, len(COLS2), "Overhead roles"); r += 1
        ot = r
        for o in OH:
            row(ws, r, 2, [o[0], "Overhead", "-", "-", o[1], o[2], o[2], 0, o[1],
                           "-", o[3], "-", 0.0, o[3]],
                [None, None, None, None, CT, CT, CT, CT, CT, None, M2, None, M2, M2])
            r += 1
        row(ws, r, 2, ["Overhead roles total"] + [None]*2 +
            [f"=SUM({L(c)}{ot}:{L(c)}{r-1})" for c in range(5, 16)],
            [None, None, None, C1, CT, CT, CT, CT, CT, M2, M2, M2, M2, M2],
            fill=GREY, bold=True)
        orr = r; r += 2
    else:
        orr = None
    row(ws, r, 2, ["Total portfolio"] + [None]*2 +
        [f"=N({L(c)}{dr})" + (f"+N({L(c)}{orr})" if orr else "") for c in range(5, 16)],
        [None, None, None, C1, CT, CT, CT, CT, CT, M2, M2, M2, M2, M2],
        fill=MID, bold=True)
    for i in range(len(COLS2)):
        ws.cell(r, 2 + i).border = TOP
    return r + 2


def two_A():
    """Squad summary on top, people grouped under a band per squad below."""
    wb, ws = newbook("2.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - working copy"
    ws.cell(2, 2).font = TITLE
    r = 4
    bar(ws, r, 2, len(COLS2), "Squad summary"); r += 1
    r = _squad_block(ws, r)
    bar(ws, r, 2, 6, "Ampol Retail FTE"); r += 1
    head(ws, r, 2, ["Name", "Role", "Status", "Vacancy lever", "Cost if hired ($)",
                    "Cost after decision ($)"], [28, 34, 10, 15, 16, 18]); r += 1
    for sq, ppl in PEOPLE:
        row(ws, r, 2, [sq, f"{len(ppl)} roles", None, None,
                       sum(p[3] for p in ppl), sum(p[3] for p in ppl)],
            [None, None, None, None, M0, M0], fill=PALE, bold=True)
        r += 1
        for n, ro, stt, cost in ppl:
            row(ws, r, 2, [n, ro, stt, "Filled" if stt == "Filled" else "Hire",
                           cost, cost], [None, None, None, None, M0, M0])
            ws.cell(r, 5).fill = PatternFill("solid", start_color=YEL, end_color=YEL)
            ws.cell(r, 5).alignment = CEN
            r += 1
    note(ws, r + 1, 2, "The squad band carries that squad's totals. "
                       "Yellow is the only cell you change.")
    return wb


def two_B():
    """One table. People indented under their squad, collapsible."""
    wb, ws = newbook("2.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - working copy"
    ws.cell(2, 2).font = TITLE
    r = 4
    bar(ws, r, 2, 10, "Squads and people"); r += 1
    head(ws, r, 2, ["Squad / person", "Role", "Status", "Vacancy lever", "Roles",
                    "Vacant", "Arch cost ($m)", "Actual ($m)", "Total after ($m)"],
         [32, 34, 10, 15, 8, 8, 13, 12, 14]); r += 1
    for sq, ppl in PEOPLE:
        s = next(x for x in SQUADS if x[0] == sq)
        row(ws, r, 2, [sq, s[1], None, None, s[4], s[5], s[6], s[7], s[7]],
            [None, None, None, None, CT, CT, M2, M2, M2], fill=PALE, bold=True)
        r += 1
        for n, ro, stt, cost in ppl:
            row(ws, r, 2, ["    " + n, ro, stt,
                           "Filled" if stt == "Filled" else "Hire",
                           None, None, None, cost / 1e6, cost / 1e6],
                [None, None, None, None, None, None, None, M2, M2])
            ws.cell(r, 5).fill = PatternFill("solid", start_color=YEL, end_color=YEL)
            ws.cell(r, 5).alignment = CEN
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = False
            r += 1
    for s in SQUADS[2:]:
        row(ws, r, 2, [s[0], s[1], None, None, s[4], s[5], s[6], s[7], s[7]],
            [None, None, None, None, CT, CT, M2, M2, M2], fill=PALE, bold=True)
        r += 1
    row(ws, r, 2, ["Total portfolio", None, None, None, 70, 22, 9.88, 14.01, 14.01],
        [None, None, None, None, CT, CT, M2, M2, M2], fill=MID, bold=True)
    ws.sheet_properties.outlinePr.summaryBelow = False
    note(ws, r + 2, 2, "Click the minus in the margin to collapse a squad to one line.")
    return wb


def two_C():
    """Squad summary only. Everyone lives on one FTE tab."""
    wb, ws = newbook("2.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - working copy"
    ws.cell(2, 2).font = TITLE
    r = 4
    bar(ws, r, 2, len(COLS2), "Squad summary"); r += 1
    r = _squad_block(ws, r)
    note(ws, r, 2, "Every role sits on the 2.0 FTE tab, filtered to this portfolio. "
                   "The vacancy levers are set there.")
    return wb


def two_D():
    """Filled roles are not decisions. Only the vacancies get a lever."""
    wb, ws = newbook("2.1 Ampol Retail")
    ws.cell(2, 2).value = "Ampol Retail - working copy"
    ws.cell(2, 2).font = TITLE
    r = 4
    bar(ws, r, 2, len(COLS2), "Squad summary"); r += 1
    r = _squad_block(ws, r)
    bar(ws, r, 2, 6, "Vacancies - the decisions"); r += 1
    head(ws, r, 2, ["Squad", "Role", "Cost if hired ($)", "Vacancy lever",
                    "Cost after decision ($)", "Saving ($)"],
         [28, 34, 16, 15, 18, 14]); r += 1
    vac = [("Store Operations", "Support Analyst - Field", 202853),
           ("Above Store", "Quality Assurance", 170502),
           ("Data AU", "Engineer - Data", 202853),
           ("Data AU", "Data Analyst", 201585),
           ("Payments", "Systems Analyst", 170502)]
    st = r
    for sq, ro, c in vac:
        row(ws, r, 2, [sq, ro, c, "Hire", c, 0], [None, None, M0, None, M0, M0])
        ws.cell(r, 5).fill = PatternFill("solid", start_color=YEL, end_color=YEL)
        ws.cell(r, 5).alignment = CEN
        r += 1
    row(ws, r, 2, ["Total", None, f"=SUM(D{st}:D{r-1})", None,
                   f"=SUM(F{st}:F{r-1})", f"=SUM(G{st}:G{r-1})"],
        [None, None, M0, None, M0, M0], fill=MID, bold=True)
    note(ws, r + 2, 2, "Filled roles are not a decision, so they are not listed here. "
                       "Their cost is in the squad summary above.")
    return wb


# =========================================================== 3.x  four options
def three_A():
    """One table per tab, stacked blocks. 3.1 shown."""
    wb, ws = newbook("3.1 Group Summary")
    ws.cell(2, 2).value = "Group summary - budget against cost"
    ws.cell(2, 2).font = TITLE
    r = 4
    bar(ws, r, 2, 6, "Budget position by portfolio"); r += 1
    head(ws, r, 2, ["Portfolio", "Lights-on budget ($m)", "Actual cost ($m)",
                    "Over/(under) budget ($m)", "Total after decisions ($m)",
                    "Left to fund ($m)"], [28, 16, 14, 16, 17, 14]); r += 1
    st = r
    for p, b, a, ar, ro, v in PORTFOLIOS:
        row(ws, r, 2, [p, b, a, a - b, a, a - b],
            [None, M2, M2, M2, M2, M2]); r += 1
    row(ws, r, 2, ["Total"] + [f"=SUM({L(c)}{st}:{L(c)}{r-1})" for c in range(3, 8)],
        [None, M2, M2, M2, M2, M2], fill=MID, bold=True)
    for i in range(6):
        ws.cell(r, 2 + i).border = TOP
    note(ws, r + 2, 2, "Budget is the lights-on allocation on 0.2 Data Config. "
                       "It is not a total people budget.")
    return wb


def three_B():
    """Column groups: Design | Today | After decisions."""
    wb, ws = newbook("3.2 Total Cost")
    ws.cell(2, 2).value = "Total cost - design against actual"
    ws.cell(2, 2).font = TITLE
    r = 4
    for c0, n, lab in ((3, 2, "DESIGN"), (5, 2, "TODAY"), (7, 2, "AFTER DECISIONS")):
        for i in range(n):
            x = ws.cell(r, c0 + i)
            x.fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
            x.font = BARF
            x.alignment = CEN
        ws.cell(r, c0).value = lab
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + n - 1)
    r += 1
    head(ws, r, 2, ["Portfolio", "Roles", "Cost ($m)", "Roles", "Cost ($m)",
                    "Roles", "Cost ($m)", "Variance to design ($m)"],
         [28, 10, 12, 10, 12, 10, 12, 16]); r += 1
    st = r
    for p, b, a, ar, ro, v in PORTFOLIOS:
        row(ws, r, 2, [p, ro * 0.7 if ar else "-", ar if ar else "-", ro, a, ro, a,
                       a - ar if ar else "-"],
            [None, C1, M2, CT, M2, CT, M2, M2]); r += 1
    row(ws, r, 2, ["Total"] + [f"=SUM({L(c)}{st}:{L(c)}{r-1})" for c in range(3, 10)],
        [None, C1, M2, CT, M2, CT, M2, M2], fill=MID, bold=True)
    for i in range(8):
        ws.cell(r, 2 + i).border = TOP
    return wb


def three_C():
    """A vertical bridge - the whole gap explained in lines."""
    wb, ws = newbook("3.2 Total Cost")
    ws.cell(2, 2).value = "Total cost - how the design becomes the actual"
    ws.cell(2, 2).font = TITLE
    r = 4
    bar(ws, r, 2, 2, "Cost bridge"); r += 1
    r = pairs(ws, r, 2, [
        ("Squad archetype cost - the design", 64.20, M2, False),
        ("Delivery squads raised over the archetype", 39.26, M2, False),
        ("Overhead roles inside the ledger", 11.65, M2, False),
        ("Cost of the organisation today", 115.11, M2, True),
        ("Impact of the vacancy decisions", 0.00, M2, False),
        ("Cost after decisions", 115.11, M2, True)], w=(52, 14))
    r += 2
    bar(ws, r, 2, 2, "Against the budget"); r += 1
    r = pairs(ws, r, 2, [("TDD lights-on budget", 50.50, M2, False),
                         ("Cost after decisions", 115.11, M2, False),
                         ("Left to fund", 64.61, M2, True)], w=(52, 14))
    r += 2
    note(ws, r, 2, "Per-portfolio detail is on 3.1. Squad detail is on 3.3.")
    return wb


def three_D():
    """Number tiles across the top, table underneath."""
    wb, ws = newbook("3.1 Group Summary")
    ws.cell(2, 2).value = "Group summary"
    ws.cell(2, 2).font = TITLE
    tiles = [("Roles", 525, CT), ("Filled", 390, CT), ("Vacant", 135, CT),
             ("Cost today ($m)", 115.11, M2), ("Budget ($m)", 50.50, M2),
             ("Left to fund ($m)", 64.61, M2)]
    c = 2
    for lab, v, f in tiles:
        h = ws.cell(4, c)
        h.value = lab; h.font = HDRF; h.alignment = CEN; h.border = BOX
        h.fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
        x = ws.cell(5, c)
        x.value = v; x.font = Font(bold=True, size=16); x.number_format = f
        x.alignment = CEN; x.border = BOX
        x.fill = PatternFill("solid", start_color=GREY, end_color=GREY)
        ws.column_dimensions[L(c)].width = 16
        c += 1
    ws.row_dimensions[5].height = 32
    r = 8
    bar(ws, r, 2, 6, "By portfolio"); r += 1
    head(ws, r, 2, ["Portfolio", "Budget ($m)", "Actual ($m)", "Over/(under) ($m)",
                    "Roles", "Vacant"], [28, 14, 13, 15, 9, 9]); r += 1
    st = r
    for p, b, a, ar, ro, v in PORTFOLIOS:
        row(ws, r, 2, [p, b, a, a - b, ro, v], [None, M2, M2, M2, CT, CT]); r += 1
    row(ws, r, 2, ["Total"] + [f"=SUM({L(c)}{st}:{L(c)}{r-1})" for c in range(3, 8)],
        [None, M2, M2, M2, CT, CT], fill=MID, bold=True)
    for i in range(6):
        ws.cell(r, 2 + i).border = TOP
    return wb


OPTIONS = {"1A": one_A, "1B": one_B, "1C": one_C, "1D": one_D,
           "2A": two_A, "2B": two_B, "2C": two_C, "2D": two_D,
           "3A": three_A, "3B": three_B, "3C": three_C, "3D": three_D}

if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    for k, fn in OPTIONS.items():
        p = os.path.join(OUT, f"{k}.xlsx")
        fn().save(p)
        print("  wrote", k)
