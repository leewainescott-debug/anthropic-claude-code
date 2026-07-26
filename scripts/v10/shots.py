"""Render each tab to a PNG so the design can be looked at rather than read as cells.

Every layout mistake in this file so far came from inspecting cell values and never
seeing the sheet. This loads the cached values, isolates one sheet at a time, and puts
LibreOffice's own rendering on screen: fills, borders, fonts, column widths, exactly what
opens in Excel.
"""
import os
import re
import shutil
import subprocess
import sys

import openpyxl

SP = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(SP, "shots")
LOHOME = os.path.join(SP, "lohome")


def one_sheet(src, title, path):
    wb = openpyxl.load_workbook(src, data_only=True)
    for s in list(wb.sheetnames):
        if s != title:
            del wb[s]
    ws = wb[title]
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = False
    wb.save(path)


def to_png(xlsx, stem):
    os.makedirs(LOHOME, exist_ok=True)
    env = dict(os.environ, HOME=LOHOME)
    d = os.path.dirname(xlsx)
    subprocess.run(["soffice", f"-env:UserInstallation=file://{LOHOME}/.lo4",
                    "--headless", "--norestore", "--convert-to", "pdf",
                    "--outdir", d, xlsx],
                   capture_output=True, text=True, timeout=300, env=env)
    pdf = xlsx.replace(".xlsx", ".pdf")
    if not os.path.exists(pdf):
        return None
    import fitz
    doc = fitz.open(pdf)
    final = os.path.join(OUT, stem + ".png")
    doc[0].get_pixmap(dpi=110).save(final)
    doc.close()
    os.remove(pdf)
    return final


def run(src, pattern):
    os.makedirs(OUT, exist_ok=True)
    wb = openpyxl.load_workbook(src, read_only=True)
    names = [s for s in wb.sheetnames if re.match(pattern, s)]
    wb.close()
    tmp = os.path.join(SP, "_shot.xlsx")
    made = []
    for n in names:
        stem = re.sub(r"[^A-Za-z0-9]+", "_", n).strip("_")
        one_sheet(src, n, tmp)
        p = to_png(tmp, stem)
        made.append((n, p))
        print(("  ok   " if p else "  FAIL ") + n)
    if os.path.exists(tmp):
        os.remove(tmp)
    return made


if __name__ == "__main__":
    run(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else r"^[123]\.\d")
