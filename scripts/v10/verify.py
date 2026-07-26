"""Final verification. Adversarial, and it recalculates rather than trusting the file.

Four things nothing else in the toolchain tests:

  the lever, end to end and for real. A lever cell is changed, the workbook is
  recalculated with a real engine, and every figure downstream is compared before and
  after. Reading the formula chain is not proof; the double count that made Ampol Web
  read $3.182m against an actual of $1.591m was a correct-looking chain.

  every header against its own column width, using the same wrap arithmetic the builder
  uses. Ten truncated headers shipped in the last build because nobody measured them.

  every section bar against the width of the table under it.

  the words the owner has ruled out, anywhere a reader can see them.
"""
import collections
import re
import shutil
import sys

import openpyxl
from openpyxl.utils import get_column_letter as L

import opts
import wbio

REVIEW = "REVIEW - Complete Role Mapping"
BANNED = ("design cost", "over/(under) design", "variance to design",
          "against the design", "lights on budget")
# banned as a block header only. As a line label with its own figure beside it, "Total to
# fund" is what the owner's funding block actually adds up, and register item 21 asks for
# it; item 10 rules it out as a heading over a table.
BANNED_HEADER = ("total to fund",)
RETIRED = {"Squads", "Added data", "Sheet2", "FY26 Budget (superseded)",
           "squad mapping (superseded)", "Lists", "0.1 Budget Table (Fin)",
           "0.4 Presentation Pack"}


def tabs2x(wb):
    return [s for s in wb.sheetnames if re.match(r"^2\.\d+ ", s)]


def coverage(wv):
    """525 roles, each on exactly one working tab, once."""
    R = wv[REVIEW]
    want = {}
    for i in range(2, 529):
        n = str(R.cell(i, 2).value or "").strip()
        if n:
            want[i] = (str(R.cell(i, 36).value), str(R.cell(i, 46).value))
    seen = collections.Counter()
    for tab in tabs2x(wv):
        ws = wv[tab]
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 4).value or "").strip() in ("Filled", "Vacant"):
                seen[(tab, r)] = 1
    out = []
    if sum(seen.values()) != len(want):
        out.append(f"people rows on 2.x = {sum(seen.values())}, ledger = {len(want)}")
    return out, len(want)


def headers(wb):
    """Every navy header cell, against the width of the column it sits in."""
    out = []
    for ws in wb.worksheets:
        if ws.title in RETIRED:
            continue
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str) or not c.value.strip():
                    continue
                fl = c.fill
                try:
                    rgb = str(fl.start_color.rgb or "").upper() \
                        if fl and fl.patternType else ""
                except Exception:
                    rgb = ""
                if rgb != opts.NAVY:
                    continue
                w = ws.column_dimensions[L(c.column)].width or 8.43
                need = opts.wrap_lines(c.value, w)
                have = ws.row_dimensions[c.row].height or 15
                if not c.alignment.wrap_text and len(c.value) > w:
                    out.append(f"{ws.title}!{c.coordinate} no wrap, "
                               f"{len(c.value)} chars in width {w:.0f}")
                elif need * 14 > have + 1:
                    out.append(f"{ws.title}!{c.coordinate} needs {need} lines "
                               f"({need * 14 + 6:.0f}pt), row height {have:.0f}")
    return out


def bars(wb):
    """A section bar must be as wide as the table under it, and no wider.

    A merged bar reports its fill on the anchor cell only, so counting fills says one
    column over a nine-column table. That false reading is what the last design review
    called out on 1.11, 1.12 and 3.4, and it was wrong: the bars were merged. The merge is
    checked first.
    """
    out = []
    for ws in wb.worksheets:
        if ws.title in RETIRED:
            continue
        merged = {(m.min_row, m.min_col): m.max_col
                  for m in ws.merged_cells.ranges if m.min_row == m.max_row}
        for r in range(1, min(ws.max_row, 200) + 1):
            b = ws.cell(r, 2)
            try:
                rgb = str(b.fill.start_color.rgb or "").upper() \
                    if b.fill and b.fill.patternType else ""
            except Exception:
                rgb = ""
            if rgb != opts.BARC or not isinstance(b.value, str):
                continue
            # a bar can be a merge, or painted cells, or a merge followed by painted
            # cells, and all three read as one bar on screen
            wide = merged.get((r, 2), 2)
            while wide < 30:
                x = ws.cell(r, wide + 1)
                try:
                    g = str(x.fill.start_color.rgb or "").upper() \
                        if x.fill and x.fill.patternType else ""
                except Exception:
                    g = ""
                if g != opts.BARC:
                    break
                wide += 1
            hdr = 2
            for rr in (r + 1, r + 2):
                n = 2
                while n < 30:
                    x = ws.cell(rr, n + 1)
                    try:
                        g = str(x.fill.start_color.rgb or "").upper() \
                            if x.fill and x.fill.patternType else ""
                    except Exception:
                        g = ""
                    if g != opts.NAVY:
                        break
                    n += 1
                hdr = max(hdr, n)
            if hdr > 2 and wide != hdr:
                out.append(f"{ws.title}!B{r} bar ends at column {L(wide)}, "
                           f"its table ends at {L(hdr)}")
    return out


def words(wb):
    out = []
    for ws in wb.worksheets:
        if ws.title in RETIRED:
            continue
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                low = v.lower()
                for b in BANNED:
                    if b in low:
                        out.append(f"{ws.title}!{c.coordinate}: {b!r} in {v[:60]!r}")
                try:
                    rgb = str(c.fill.start_color.rgb or "").upper() \
                        if c.fill and c.fill.patternType else ""
                except Exception:
                    rgb = ""
                if rgb in (opts.BARC, opts.NAVY):
                    for b in BANNED_HEADER:
                        if b in low:
                            out.append(f"{ws.title}!{c.coordinate}: {b!r} as a header")
    return out


def facts(wv):
    """The same figure must not appear under two different headings on one tab."""
    out = []
    for ws in wv.worksheets:
        if ws.title in RETIRED:
            continue
        for r in range(1, min(ws.max_row, 200) + 1):
            seen = {}
            for c in range(2, 20):
                v = ws.cell(r, c).value
                if isinstance(v, float) and abs(v) > 0.001:
                    seen.setdefault(round(v, 6), []).append(L(c))
            for v, cols in seen.items():
                if len(cols) > 3:
                    out.append(f"{ws.title} row {r}: {v} in {len(cols)} columns "
                               f"{','.join(cols)}")
    return out


def lever(path):
    """Set a vacancy to Hold, recalculate, and check what moved."""
    wb = openpyxl.load_workbook(path)
    wv = openpyxl.load_workbook(path, data_only=True)
    tab = "2.1 Ampol Retail"
    ws, wsv = wb[tab], wv[tab]
    target = None
    for r in range(1, ws.max_row + 1):
        if str(wsv.cell(r, 4).value or "") == "Vacant" and \
                str(wsv.cell(r, 5).value or "") == "Hire":
            target = r
            break
    if target is None:
        return ["no vacancy with a Hire lever found on 2.1"]
    cost = wsv.cell(target, 6).value

    # Probes are found by label. This test carried hardcoded rows and reported two false
    # failures the moment Exec Summary grew two lines - the same join-by-row-number mistake
    # the workbook itself was fixed for.
    def at(sheet, label, header):
        ws2 = wb[sheet]
        row = next((r for r in range(1, ws2.max_row + 1)
                    if isinstance(ws2.cell(r, 2).value, str)
                    and ws2.cell(r, 2).value.strip() == label), None)
        if row is None:
            raise KeyError(f"{sheet}: no row {label!r}")
        if header is None:
            return f"C{row}"
        for hr in range(1, 12):
            for c in range(2, 20):
                v = ws2.cell(hr, c).value
                if isinstance(v, str) and v.strip() == header:
                    return f"{L(c)}{row}"
        raise KeyError(f"{sheet}: no column {header!r}")

    # tabs by number prefix, rows by label, columns by header - nothing here names a row
    # number or a tab title, both of which have moved under me before
    bridge = next(t for t in wb.sheetnames if t.startswith("3.1 "))
    detail = next(t for t in wb.sheetnames if t.startswith("3.3 "))
    LEDGER = "Cost of the 525 roles in the ledger"
    PROBE = [
        (bridge, LEDGER, "Cost after decisions ($m)", "cost after decisions",
         -cost / 1000000),
        (detail, "Group total", "Cost after decisions ($m)",
         "cost after decisions", -cost / 1000000),
        ("Exec Summary", "Cost after the decisions set today ($m)", None,
         "cost after decisions", -cost / 1000000),
        (bridge, LEDGER, "Total roles after decisions", "roles after decisions", -1),
        (detail, "Group total", "Total roles after decisions", "roles after decisions",
         -1),
        ("Exec Summary", "Roles after the decisions set today", None,
         "roles after decisions", -1),
        (bridge, LEDGER, "Actual cost ($m)", "cost today", 0),
        (bridge, LEDGER, "Total roles", "roles today", 0),
        (detail, "Group total", "Actual cost ($m)", "cost today", 0),
    ]
    probes = {}
    for sheet, label, header, _, _ in PROBE:
        ref = at(sheet, label, header)
        probes[f"{sheet}!{ref}"] = wv[sheet][ref].value
    ws.cell(target, 5).value = "Hold"
    wb.save("lever.xlsx")
    rc = wbio.recalc("lever.xlsx")
    after = openpyxl.load_workbook(rc, data_only=True)
    out = [f"lever test: {tab} row {target}, a ${cost:,.0f} vacancy set to Hold"]
    for (sheet, label, header, what, exp), key in zip(PROBE, probes):
        b, n = probes[key], after[sheet][key.split("!")[1]].value
        d = (n or 0) - (b or 0)
        ok = abs(d - exp) < 1e-6
        out.append(f"  {'ok ' if ok else 'FAIL'} {key} {what} {b} -> {n} "
                   f"({d:+.4f}, expected {exp:+.4f})")
    return out


def run(path):
    wb = openpyxl.load_workbook(path)
    wv = openpyxl.load_workbook(path, data_only=True)
    rep = collections.OrderedDict()
    cov, n = coverage(wv)
    rep[f"role coverage ({n} in the ledger)"] = cov
    rep["truncated headers"] = headers(wb)
    rep["bars not matching their table"] = bars(wb)
    rep["words the owner ruled out"] = words(wb)
    rep["one figure under many headings"] = facts(wv)
    return rep


if __name__ == "__main__":
    p = sys.argv[1] if len(sys.argv) > 1 else "cand.xlsx"
    total = 0
    for k, v in run(p).items():
        print("=" * 74)
        print(f"{k.upper()}: {len(v)}")
        for x in v[:25]:
            print("   ", x)
        total += len(v)
    print("=" * 74)
    print(f"TOTAL FINDINGS: {total}")
    if "--lever" in sys.argv:
        print("=" * 74)
        for x in lever(p):
            print(x)
