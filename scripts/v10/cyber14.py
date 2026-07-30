"""1.14 TDD Cyber - the design tab for the one-platform cyber portfolio.

The owner's instruction: "build a tdd cyber 1.x tab ... exactly the same [as] the 1.x tabs
for portfolios, however, it just has one platform and one squad. platform is called tdd
cyber, squad is cyber uplift."

Wave M gave the platform its second squad. His cyber uplift ruling moves nine roles out of
the COE onto this portfolio: five staff Cyber Uplift, four staff Identity. Cyber Uplift is
fully funded by the programme and carries a typed cost; Identity is a real operations squad
priced off his archetype library, with 80% of it supported by TDD and 20% charged to the
programmes through his own support % toggle. The tab also carries the programme's funding
block, which states what the 2.8 is paying for line by line - the squad, the part-charges
off 1.13's toggles, Identity's share - and what is left for non-people.

The tab that carried this name until now was a dead copy of 1.9 Commercial Fuels - its own
title said TDD Cyber while every figure on it, and every cell reference under it, read
Commercial Fuels' budget row and Commercial Fuels' three platforms. It is rebuilt here
from scratch rather than patched, because nothing on it was right except the name.

Two templates, both read at runtime rather than transcribed, so the new tab cannot drift
away from the family the next time a sibling restyles it:

  1.3 Enterprise Data   the block layout - Portfolio Summary, the budget box, Other
                        funding, Total to fund - and every style, fill, border and number
                        format those blocks carry. 1.3 is the closest sibling in shape:
                        one funded platform, the reduced Other funding table, no budget
                        reconciliation block.
  1.9 Commercial Fuels  the squad row itself - the archetype formula, the TDD/Funded split
                        beside it, the cream input cells and their four dropdowns. Its own
                        squad row is row 26 and so is Cyber Uplift's, so that formula copies
                        over with no renumbering; Identity sits on row 27 and its copy is
                        renumbered by exact match on the row it came from.

Two decisions the shape states rather than hides:

  the portfolio overhead is NOT drawn. TDD Cyber is not one of the ten funded portfolios
  on 0.2 Data Config, so it draws no Head of Technology / Business Partner / Domain
  Architect / Leadership allowance. Row 6 is zero on every column and the reason is
  written at B11, in words, under the block it explains.
  the platform overhead IS drawn - one platform, one 0.2!N16 - and it is the whole of the
  tab's cost until the owner sizes the squad.

What this tab deliberately does NOT carry: the actuals footer and the K/L columns
(actuals.py builds those), and the Actuals figure at G9 with the variance line under it
(post2707.py writes those). The header cell at G5 and the bar over it are here so those
two steps have a column to write into.

Shipped priced: F7 reads 0.165 (the platform overhead), the two squads read 1.2998 and
0.8, F9 reads 2.2648, and no cell on the tab is an error. The three cream cells are his -
the Cyber Uplift cost, Identity's support %, and the programme funding.

The empty 2.15 TDD Cyber shell is created here too, so the pair exists from this point in
the chain onward; final2x.py builds its contents.
"""
import copy
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

TAB = "1.14 TDD Cyber"
WORK = "2.15 TDD Cyber"
T3 = "1.3 Enterprise Data"          # block layout and styles
T9 = "1.9 Commercial Fuels"         # the squad row: formulas, cream cells, dropdowns
CFG = "'0.2 Data Config'"
CFGROW = 23                         # 0.2's "TDD Cyber" line. Wave M split the COE off it
                                    # onto row 7, so this row is TDD Cyber and nothing else.

# The shipped column profile, the same on all thirteen design tabs. K and L are the two
# columns actuals.py writes into and are sized for it here so the tab does not resize
# under the reader later.
WIDTHS = {"A": 2.0, "B": 40.0, "C": 27.0, "D": 9.5, "E": 17.8, "F": 10.0, "G": 11.0,
          "H": 43.5, "I": 14.0, "J": 45.0, "K": 24.8, "L": 25.5}

PLATFORM = "TDD Cyber"
# Two squads, his ruling. Cyber Uplift is the programme's own squad and is fully funded by
# it: support % 0, so none of it is a TDD cost, and its Total Squad Cost is a typed cream
# figure rather than an archetype price - no archetype in his library prices a cyber uplift
# programme, and he set the number. Identity is a real operations squad and prices off the
# archetype library like every other squad in the file, with 80% of it supported by TDD and
# the remaining 20% charged to the programmes through his support % toggle.
SQUAD = "Cyber Uplift"
SQUAD2 = "Identity"
R_UPLIFT, R_IDENTITY, R_PLATOH, R_TOTAL = 26, 27, 28, 29
UPLIFT_COST = 1.2998            # his figure for the Cyber Uplift squad, cream and typed
IDENTITY_SUPPORT = 0.8          # his 80% toggle: 80% TDD, 20% to the programmes
PROGRAMME_FUNDING = 2.8         # the cyber uplift programme's own funding, cream and typed

WHY_NO_PF_OVERHEAD = (
    "TDD Cyber does not draw a portfolio overhead - it is not one of the ten funded "
    "portfolios on 0.2 Data Config. The platform overhead is drawn. Cyber Uplift is fully "
    "funded from the cyber uplift programme and carries no overhead cost. Identity is "
    "priced on the archetype library like every other squad; 20% of it is charged to the "
    "programmes through the support % toggle.")
WHY_BUDGET = ("This is the TDD Cyber allocation on 0.2 Data Config. The COE - Cyber, Risk "
              "& Service Operations is funded separately, on its own line, and its roles "
              "and funding are on 1.13 Cyber Roles.")
HOW_TO_PRICE = ("Cyber Uplift is a typed figure - no archetype prices a programme. Identity "
                "prices itself off its Squad Type and Size")

# ---- the cyber uplift programme's funding block ----
# It replaces the family's "Other funding" table on this tab, because on this tab there is
# no EGI line and no Lights On line to reconcile: there is one bucket, the programme, and
# what the programme is paying for. Funding is his to type; every use of it is a formula
# reading the cell that already carries that figure, so the block cannot drift from the
# squad table above it or from 1.13's part-charges.
FUND_BAR = "Cyber uplift programme funding"
FUND_LINES = [
    ("Programme funding ($m)", PROGRAMME_FUNDING, True),
    ("Cyber Uplift squad - fully funded ($m)", f"=$H${R_UPLIFT}", False),
    ("Part-charges from the COE roles - 1.13 Cyber Roles toggles ($m)",
     "=SUM('1.13 Cyber Roles'!$U$19:$U$70)", False),
    ("Identity share charged to the programmes ($m)", f"=$J${R_IDENTITY}", False),
    ("Used for cyber FTE ($m)", None, False),          # the sum of the three above
    ("Remaining for non-people ($m)", None, False),    # funding less used
]

# the four family dropdowns plus the country list, and the cells each one lands on. Cyber
# Uplift takes the On/Off, country and Support % lists; its Squad Type and Size are not set
# from the library, because the library does not price it.
CREAM = PatternFill("solid", fgColor="FFFFF2CC")
DVS = [("SquadTypes", f"C{R_IDENTITY}"), ("SquadSizes", f"D{R_IDENTITY}"),
       ("OnOff", f"E{R_UPLIFT}"), ("OnOff", f"E{R_IDENTITY}"),
       ("SupportPct", f"G{R_UPLIFT}"), ("SupportPct", f"G{R_IDENTITY}"),
       ('"AU,NZ"', f"F{R_UPLIFT}"), ('"AU,NZ"', f"F{R_IDENTITY}")]


def wipe(ws):
    """Everything off the tab. The old 1.14 was a Commercial Fuels copy down to its
    merges and its column widths; anything left behind would be a figure from another
    portfolio sitting on this one."""
    blank = copy.copy(openpyxl.cell.cell.Cell(ws)._style)
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    for row in ws.iter_rows():
        for c in row:
            c.value = None
            c._style = copy.copy(blank)
    for k in list(ws.column_dimensions):
        del ws.column_dimensions[k]
    for k in list(ws.row_dimensions):
        del ws.row_dimensions[k]
    ws.conditional_formatting._cf_rules.clear()
    ws.data_validations.dataValidation = []
    ws.freeze_panes = None
    ws.sheet_view.showGridLines = False


def build(wb, out):
    ws, t3, t9 = wb[TAB], wb[T3], wb[T9]
    wipe(ws)
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.sheet_properties.tabColor = copy.copy(t3.sheet_properties.tabColor)

    def put(coord, src, src_coord, value=None):
        """One cell: the template's style, and a value where this tab has one."""
        ws[coord]._style = copy.copy(src[src_coord]._style)
        if value is not None:
            ws[coord] = value

    def like(coord, src, src_coord):
        """The template's style and the template's own text - a label the family owns."""
        put(coord, src, src_coord, src[src_coord].value)

    def note(coord, text):
        c = ws[coord]
        c.value = text
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")

    def bar(coord, value=None):
        put(coord, t3, "B4", value)      # 1.3!B4 is the one true bar: FF002F6C, white bold

    # ---- title -------------------------------------------------------------------
    put("B2", t3, "B2", PLATFORM)
    ws.row_dimensions[2].height = t3.row_dimensions[2].height

    # ---- Portfolio Summary ---------------------------------------------------------
    # The bar runs the width of the table under it, which on this tab is B to G: the
    # Actuals column is part of the summary here, so the strip covers F and G too rather
    # than stopping at the merge.
    like("B4", t3, "B4")
    for c in "CDEFG":
        bar(f"{c}4")
    bar("H4", t3["H4"].value)
    bar("I4")
    ws.row_dimensions[4].height = 19

    for c in "BCDEF":
        like(f"{c}5", t3, f"{c}5")
    put("G5", t3, "F5", "Actuals")       # post2707 rewrites the value; the column is here
    ws.row_dimensions[5].height = t3.row_dimensions[5].height

    # row 6 - no portfolio overhead. Zero in each column rather than a blank, so the row
    # is a stated nil and F6 stays a live sum of three cells like its ten siblings.
    like("B6", t3, "B6")
    for c, f in (("C", "=0"), ("D", "=0"), ("E", "=0"), ("F", "=C6+D6+E6")):
        put(f"{c}6", t3, f"{c}6", f)

    # row 7 - one platform, so one I-cell in the SUM. The AU/NZ branch is the family's,
    # written exactly the way its ten siblings write it: if the portfolio's NZ budget on
    # 0.2 is the bigger of the two, the cost lands in the NZ column, else in AU.
    like("B7", t3, "B7")
    put("C7", t3, "C7",
        f'=IF({CFG}!$D${CFGROW}>{CFG}!$C${CFGROW},0,SUM(I{R_PLATOH}))')
    put("D7", t3, "D7",
        f'=IF({CFG}!$D${CFGROW}>{CFG}!$C${CFGROW},SUM(I{R_PLATOH}),0)')
    put("E7", t3, "E7", "=0")
    put("F7", t3, "F7", "=C7+D7+E7")

    # row 8 - two squads, so the SUMIF spans both squad rows on each side
    lo, hi = R_UPLIFT, R_IDENTITY
    like("B8", t3, "B8")
    put("C8", t3, "C8", f'=SUMIF(F{lo}:F{hi},"AU",I{lo}:I{hi})')
    put("D8", t3, "D8", f'=SUMIF(F{lo}:F{hi},"NZ",I{lo}:I{hi})')
    put("E8", t3, "E8", f"=SUM(J{lo}:J{hi})")
    put("F8", t3, "F8", "=C8+D8+E8")

    like("B9", t3, "B9")
    for c in "CDEF":
        put(f"{c}9", t3, f"{c}9", f"=SUM({c}6:{c}8)")
    # G9 and the E10/F10 variance line are post2707's - left empty, not left wrong

    # ---- Budget vs TDD Cost --------------------------------------------------------
    like("H5", t3, "H5")
    put("I5", t3, "I5", f"={CFG}!$E${CFGROW}")
    like("H6", t3, "H6")
    put("I6", t3, "I6", f"={CFG}!$C${CFGROW}")
    like("H7", t3, "H7")
    put("I7", t3, "I7", f"={CFG}!$D${CFGROW}")
    like("H8", t3, "H8")
    put("I8", t3, "I8", "=C9-I6")
    like("H9", t3, "H9")
    put("I9", t3, "I9", "=D9-I7")
    like("H10", t3, "H10")
    put("I10", t3, "I10", "=I8+I9")

    note("B11", WHY_NO_PF_OVERHEAD)
    note("H11", WHY_BUDGET)

    # ---- the cyber uplift programme's funding -------------------------------------
    # Not the family's "Other funding" table: there is no EGI line and no Lights On line on
    # this tab to reconcile against. There is one bucket - the programme - and the block
    # states what it is funding. Only the funding figure is typed; everything under it reads
    # the cell that already carries the figure, so no line here can disagree with the squad
    # table above it or with 1.13's part-charges.
    bar("H13", FUND_BAR)
    bar("I13")
    bar("J13")
    ws.row_dimensions[13].height = 19
    r_fund = 14
    for i, (lab, val, is_input) in enumerate(FUND_LINES):
        r = r_fund + i
        for c in "HIJ":
            put(f"{c}{r}", t3, f"{c}{'13' if i == 0 else '15'}")
        ws[f"H{r}"] = lab
        # 1.3's funding table paints its amount column cream, because on that tab every one
        # of those cells is a figure he types. Here only the first line is. Cream is the
        # file's one promise that a cell is a typed input, so it comes off the five that
        # compute themselves before they are ever written.
        ws[f"J{r}"].fill = CREAM if is_input else PatternFill()
        if val is None:
            continue
        ws[f"J{r}"] = val
    r_used = r_fund + 4
    ws[f"J{r_used}"] = f"=SUM(J{r_fund + 1}:J{r_fund + 3})"
    ws[f"J{r_used + 1}"] = f"=J{r_fund}-J{r_used}"
    for r in (r_fund, r_used, r_used + 1):
        ws[f"J{r}"].number_format = "#,##0.00;(#,##0.00)"

    # ---- Budget position -----------------------------------------------------------
    # The family's "Total to fund" block, with its own name: on this tab it is the budget
    # position and nothing else, because the programme funding block above answers the
    # other-cost question in its own words.
    like("B16", t3, "B15")
    ws["B16"] = "Budget position"
    put("C16", t3, "C15")
    put("B17", t3, "B16")
    put("C17", t3, "C16")
    like("B18", t3, "B17")
    put("C18", t3, "C17", "=I8+I9")
    like("B19", t3, "B18")
    # Nothing is left to fund. Every dollar of Other cost on this tab is the programme's own
    # spend and the programme funds all of it, which the block above states line by line. A
    # stated nil, written the way row 6 states its nil, not a blank.
    put("C19", t3, "C18", "=0")
    like("B20", t3, "B19")
    put("C20", t3, "C19", "=C18+C19")

    # ---- Platform: TDD Cyber -------------------------------------------------------
    for c in "BCDEFGHIJ":
        bar(f"{c}24")
    ws["B24"] = f"Platform: {PLATFORM}"
    ws.row_dimensions[24].height = 19

    for c in "BCDEFGHIJ":
        like(f"{c}25", t9, f"{c}25")     # K25/L25 stay empty - actuals.py heads them

    # Cyber Uplift. Its Total Squad Cost is a typed cream figure, not an archetype lookup:
    # no archetype in his library prices a cyber uplift programme, and the figure is his.
    # Support % 0 means none of it is a TDD cost, so the whole of it shows in Funded outside
    # TDD - which is what "fully funded from the uplift programme" means on this table.
    for c in "BCDEFGHIJ":
        put(f"{c}{R_UPLIFT}", t9, f"{c}26")
    ws[f"B{R_UPLIFT}"] = SQUAD
    ws[f"C{R_UPLIFT}"] = "Strategic Programs"
    ws[f"D{R_UPLIFT}"] = None
    ws[f"E{R_UPLIFT}"] = "Onshore"
    ws[f"F{R_UPLIFT}"] = "AU"
    ws[f"G{R_UPLIFT}"] = 0
    ws[f"H{R_UPLIFT}"] = UPLIFT_COST
    for c in ("G", "H"):
        ws[f"{c}{R_UPLIFT}"].fill = CREAM
    ws[f"I{R_UPLIFT}"] = f'=IFERROR($H{R_UPLIFT}*$G{R_UPLIFT},"")'
    ws[f"J{R_UPLIFT}"] = f'=IFERROR($H{R_UPLIFT}*(1-$G{R_UPLIFT}),"")'
    note(f"M{R_UPLIFT}", HOW_TO_PRICE)

    # Identity. A real operations squad, priced off the archetype library exactly the way
    # 1.9's own squad row prices itself - the formula copies across with the row it came
    # from - with his 80% support toggle splitting it 0.64 TDD / 0.16 programmes.
    for c in "BCDEFGHIJ":
        put(f"{c}{R_IDENTITY}", t9, f"{c}26")
    ws[f"B{R_IDENTITY}"] = SQUAD2
    ws[f"C{R_IDENTITY}"] = "Operations"
    ws[f"D{R_IDENTITY}"] = "S"
    ws[f"E{R_IDENTITY}"] = "Onshore"
    ws[f"F{R_IDENTITY}"] = "AU"
    ws[f"G{R_IDENTITY}"] = IDENTITY_SUPPORT
    ws[f"G{R_IDENTITY}"].fill = CREAM
    ws[f"G{R_IDENTITY}"].number_format = "0%"
    for c in ("H", "I", "J"):
        ws[f"{c}{R_IDENTITY}"] = re.sub(r"(?<![$\d])26(?![\d])", str(R_IDENTITY),
                                        t9[f"{c}26"].value)

    for c in "BCDEFGHIJ":
        put(f"{c}{R_PLATOH}", t9, f"{c}28")
    ws[f"B{R_PLATOH}"] = t9["B28"].value           # "Platform Overhead"
    ws[f"I{R_PLATOH}"] = t9["I28"].value           # ='0.2 Data Config'!$N$16

    for c in "BCDEFGHIJ":
        put(f"{c}{R_TOTAL}", t9, f"{c}29")
    # 1.9's own total row leaves F unfilled, a hole in the grey band that the polish step
    # closes later. The band is written closed here rather than shipped broken.
    ws[f"F{R_TOTAL}"]._style = copy.copy(t9["E29"]._style)
    ws[f"B{R_TOTAL}"] = f"{PLATFORM} Total"
    ws[f"H{R_TOTAL}"] = f"=SUM(H{R_UPLIFT}:H{R_IDENTITY})"
    ws[f"I{R_TOTAL}"] = f"=SUM(I{R_UPLIFT}:I{R_PLATOH})"
    ws[f"J{R_TOTAL}"] = f"=SUM(J{R_UPLIFT}:J{R_IDENTITY})"

    # ---- merges, last, so no cell is written through a MergedCell -------------------
    for rng in ("B4:E4", "H4:I4", "H13:J13", "B24:J24"):
        ws.merge_cells(rng)

    # ---- the four dropdowns, copied off 1.9 rather than re-declared -----------------
    src = {dv.formula1: dv for dv in t9.data_validations.dataValidation}
    for f1, cell in DVS:
        s = src.get(f1)
        if s is None:
            out.append(f"{TAB}: 1.9 carries no {f1} validation - {cell} left without one")
            continue
        d = DataValidation(type=s.type, operator=s.operator, formula1=s.formula1,
                           allow_blank=s.allow_blank, showDropDown=s.showDropDown,
                           showErrorMessage=s.showErrorMessage,
                           showInputMessage=s.showInputMessage,
                           errorTitle=s.errorTitle, error=s.error, errorStyle=s.errorStyle,
                           promptTitle=s.promptTitle, prompt=s.prompt)
        ws.add_data_validation(d)
        d.add(cell)
    out.append(f"{TAB} rebuilt: one platform, two squads ({SQUAD} at a typed {UPLIFT_COST} "
               f"fully funded by the programme, {SQUAD2} on the Operations|S archetype at "
               f"{IDENTITY_SUPPORT:.0%} support), platform overhead drawn, portfolio "
               f"overhead not, programme funding {PROGRAMME_FUNDING} stated line by line - "
               f"layout off {T3}, squad rows off {T9}")


def pair(wb, out):
    """The working copy's shell. final2x.py writes its contents; it only needs the sheet
    to exist, and it has to exist before the chain recalculates."""
    if WORK in wb.sheetnames:
        out.append(f"{WORK} already present - left for final2x")
        return
    prev = "2.14 EGI"
    idx = wb.sheetnames.index(prev) + 1 if prev in wb.sheetnames else None
    ws = wb.create_sheet(WORK, idx)
    ws.sheet_view.showGridLines = False
    if prev in wb.sheetnames:
        ws.sheet_properties.tabColor = copy.copy(wb[prev].sheet_properties.tabColor)
    out.append(f"{WORK} created empty after {prev} - final2x builds it")


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = []
    if TAB not in wb.sheetnames:
        wb.create_sheet(TAB, wb.sheetnames.index(T9) + 1 if T9 in wb.sheetnames else None)
        out.append(f"{TAB} did not exist - created")
    build(wb, out)
    pair(wb, out)
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys

    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
