"""Hybrid squads price as 2 roles onshore and the rest offshore, and the 2 is an input.

The owner's rule, in his words: "assume 2 roles onshore and the rest offshore for hybrid
squads, any squad any size, per-FTE derived by dividing squad cost by # of roles,
settable". His own note at 0.3!C25 already says so - "Hybrid = 2 roles offshore, rest
offshore" - and the model did not: every Hybrid squad on the 1.x tabs priced at the plain
midpoint of the onshore and offshore archetype costs, (G+H)/2, which is only the right
answer for a squad whose roles happen to split down the middle.

What the shipped formula did (1.9!H26, the family shape on all forty squad rows):

    IF($E26="Hybrid", (INDEX(0.3!$G$5:$G$23,MATCH(key))
                     + INDEX(0.3!$H$5:$H$23,MATCH(key)))/2, ...)

What it does now, with n the archetype's own "# of roles" (0.3 column F, the same MATCH)
and k the new input cell:

    G * MIN(k,n)/n  +  H * (n-MIN(k,n))/n

so the per-role price is the squad price divided by the archetype's role count on each
side of the shore, exactly as he described it. Two properties the shape buys:

  MIN(k,n)  the 1-role XS archetype (Configuration / Integration|XS, F8 = 1) cannot buy
            two onshore roles out of one, so it prices fully onshore - 0.4, not 0.4 plus
            a role of offshore change that does not exist.
  n as read the two fractional archetypes (Engineering|S and Product|S, 4.5 and 5.5 roles)
            need no special case: 2/4.5 of the onshore price plus 2.5/4.5 of the offshore
            price is what the division gives.

The 2 is a cream input on 0.3, not a constant in forty formulas, so he can retype it and
the whole model reprices. It is written next to his existing Offshore rate pair -
K7 label, K8 value - which are the only two cells this chain ever writes on that tab.
0.3 is his source tab: assemble_base carries it over from rev.xlsx untouched, polish,
finish and design2707 all name it as out of scope, and regress2707 proves it against
rev.xlsx cell for cell. The exemptions are three cells: the K7/K8 input pair, and C25.

C25 is his note stating the hybrid rule, and as typed it reads "Hybrid = 2 roles offshore,
rest offshore" - the opposite of the rule he set in as many words ("assume 2 roles would
be onshore and the rest of the roles offshore", D104). His newer instruction outranks his
older note, the same way his newest workbook's labels outrank an earlier ruling (D83). It
is the only written statement of the rule a reader can find, on the tab that defines the
archetypes, so left as typed it teaches every GM the rule backwards. Corrected to what he
ruled, and disclosed (D116) so one word from him reverses it.

Placement: K4/K5 sit label-over-value at the right of the archetype table, so the new pair
sits the same way one blank row below at K7/K8, in the same column, with K4's and K5's own
styles copied cell for cell. It does not go in L4/L5: L4 is the empty cell K4's "Offshore
rate" label currently spills into, and filling it would clip his label. The one property
not copied from K5 is its number format - K5 is a rate and renders 0.4 as 40%; a count of
roles renders as a count.

Runs directly after cyber14.py so 1.14 TDD Cyber, whose squad row is copied off 1.9 at
build time, is swept with the other thirteen tabs rather than keeping the old midpoint.
Idempotent: a formula already carrying the new shape has no (G+H)/2 branch left to match.
"""
import copy
import re

import openpyxl

ARCH = "0.3 Squad Archetypes"
A3 = f"'{ARCH}'"                      # how every 1.x formula spells the tab
DESIGN = re.compile(r"^1\.\d+ ")      # the design tabs, 1.1-1.14
HCOL = 8                              # column H, Total Squad Cost

# the new input, and the pair it is modelled on
LABEL_AT, VALUE_AT = "K7", "K8"
LIKE_LABEL, LIKE_VALUE = "K4", "K5"   # 'Offshore rate' / 0.4
LABEL = "Onshore roles in a hybrid squad"
VALUE = 2
KREF = f"{A3}!${VALUE_AT[0]}${VALUE_AT[1:]}"     # '0.3 Squad Archetypes'!$K$8

# The Hybrid branch as shipped, with the row's own lookup key captured so the replacement
# can be built around it. The key is $C26&"|"&$D26 on every row and carries no comma or
# bracket of its own, so [^,()]+ ends exactly where the MATCH's next argument begins; the
# backreference then requires the second INDEX to look the same squad up.
_S = re.escape(A3)
HYBRID = re.compile(
    r"\(INDEX\(" + _S + r"!\$G\$5:\$G\$23,MATCH\((?P<key>[^,()]+)," + _S
    + r"!\$A\$5:\$A\$23,0\)\)"
    r"\+INDEX\(" + _S + r"!\$H\$5:\$H\$23,MATCH\((?P=key)," + _S
    + r"!\$A\$5:\$A\$23,0\)\)\)/2")
# a formula that has already been through this step
DONE = re.compile(r"MIN\(" + re.escape(KREF) + r",")


def branch(key):
    """The 2-onshore price for one squad row, built around that row's own MATCH key."""
    m = f"MATCH({key},{A3}!$A$5:$A$23,0)"
    g = f"INDEX({A3}!$G$5:$G$23,{m})"          # onshore squad cost
    h = f"INDEX({A3}!$H$5:$H$23,{m})"          # offshore squad cost
    n = f"INDEX({A3}!$F$5:$F$23,{m})"          # # of roles in the archetype
    on = f"MIN({KREF},{n})"                    # onshore roles actually available
    return f"({g}*{on}/{n}+{h}*({n}-{on})/{n})"


def input_cell(wb, out):
    """The settable 2: two cells on the owner's tab, styled off his own input pair."""
    ws = wb[ARCH]
    lab, val = ws[LABEL_AT], ws[VALUE_AT]
    if lab.value not in (None, LABEL) or val.value not in (None, VALUE):
        raise RuntimeError(
            f"{ARCH}!{LABEL_AT}/{VALUE_AT} are not free ({lab.value!r}/{val.value!r}) - "
            "the input pair would overwrite the owner's own cells")
    lab._style = copy.copy(ws[LIKE_LABEL]._style)
    lab.value = LABEL
    val._style = copy.copy(ws[LIKE_VALUE]._style)
    val.value = VALUE
    val.number_format = "General"     # K5 is a rate and formats 0.4 as 40%; this is a count
    out.append(f"{ARCH}!{LABEL_AT} = {LABEL!r}, {ARCH}!{VALUE_AT} = {VALUE} "
               f"(cream input, styled off {LIKE_LABEL}/{LIKE_VALUE})")
    # his C25 note states the rule this file implements, backwards - see the module
    # docstring. His D104 instruction wins over his older note, and only over the exact
    # text he typed: anything else in the cell is left alone and reported.
    old = ("Hybrid = 2 roles offshore, rest offshore ",
           "Hybrid = 2 roles offshore, rest offshore")
    new = "Hybrid = 2 roles onshore, rest offshore"
    if ws["C25"].value in old:
        ws["C25"] = new
        out.append(f"{ARCH}!C25 -> {new!r} - his rule (D104), stated the right way "
                   "round (D116)")
    elif ws["C25"].value != new:
        out.append(f"{ARCH}!C25 holds {str(ws['C25'].value)[:50]!r} - left alone")


def sweep(wb, out):
    """Every H-column archetype formula on the 1.x tabs, Hybrid branch rewritten."""
    done = skipped = 0
    for name in wb.sheetnames:
        if not DESIGN.match(name):
            continue
        ws = wb[name]
        rows, already, missed = [], [], []
        for r in range(1, ws.max_row + 1):
            c = ws.cell(r, HCOL)
            v = c.value
            if not (isinstance(v, str) and v.startswith("=") and ARCH in v):
                continue
            new, n = HYBRID.subn(lambda m: branch(m.group("key")), v)
            if n:
                c.value = new
                rows.append(f"{c.coordinate}x{n}" if n > 1 else c.coordinate)
            elif DONE.search(v):
                already.append(c.coordinate)
            else:
                missed.append(c.coordinate)
        done += len(rows)
        skipped += len(already)
        if rows or already or missed:
            note = f"{name}: {len(rows)} squad rows repriced"
            if already:
                note += f", {len(already)} already carried the rule"
            if missed:
                note += (f", {len(missed)} archetype formulas MATCHED NOTHING "
                         f"({', '.join(missed)}) - check the shape")
            out.append(note + (f" [{', '.join(rows)}]" if rows else ""))
    out.append(f"{done} Hybrid branches rewritten to "
               f"G*MIN(k,n)/n + H*(n-MIN(k,n))/n, k = {KREF}"
               + (f"; {skipped} were already rewritten" if skipped else ""))
    return done


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = []
    if ARCH not in wb.sheetnames:
        raise RuntimeError(f"{ARCH} is not in {src} - nothing to price against")
    input_cell(wb, out)
    sweep(wb, out)
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys

    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
