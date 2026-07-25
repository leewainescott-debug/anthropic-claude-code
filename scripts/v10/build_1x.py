"""Stage 5: normalise the ten 1.x portfolio design tabs so they read the same way.

Three of these are regressions from my own earlier pass and are fixed first.

REGRESSIONS I INTRODUCED
  1.10 C7  I rewrote it as IF(C12>C12, ...) - a cell compared to itself, always FALSE,
           so the AU platform overhead is hard-wired to zero. The original was correct.
  1.5  D7  the NZ branch I added sums I25 (a squad cost, 0.63) and I31 (a text header)
           instead of the overhead cell I27. Dormant only because NZ budget = AU budget.
  1.5  D8  the SUMIF range I added spans rows 25-32, swallowing the Platform Overhead
           row 27 and the "P&C Total" row 28.

PRE-EXISTING, LIVE
  1.2  C7/D7  byte-identical branches, so the platform overhead is counted on BOTH the
              AU and NZ side: F7 reads 0.99 where the true figure is 0.495. This flows
              into F9 and into 0.2 Data Config F13, moving Customer's variance by
              $0.495m.
  1.2  C6/D6  byte-identical, wrong branch on AU, and carry an undocumented *0.5. The
              only tab that splits the portfolio overhead 50/50. If AU ever exceeded NZ
              both branches would go to zero and the whole $0.7975m would vanish.
  1.10 C8     the AU branch omits the "Data NZ" platform block entirely, so $0.612m can
              never land in the AU column.

SIGN CONVENTIONS (numbers right, meanings different - these cannot be added together)
  TDD Variance    nine tabs use spend - budget; 1.4 uses budget - spend.
  Other Variance  seven tabs use +left-to-fund; 1.4, 1.9 and 1.10 negate it.
Both are normalised to the majority convention: positive = over budget / still to fund.
"""
import json
import re

import openpyxl
from openpyxl.styles import Font

BOLD = Font(bold=True)
ITAL = Font(italic=True)
DC = "'0.2 Data Config'"

# tab -> (0.2 budget row, portfolio-overhead row, platform-overhead row, list of
#         Platform Overhead I-cells, list of (first_squad_row, last_squad_row) blocks)
GEOM = {
    "1.1 Ampol Retail":            (11, 6, 7, ["I29", "I36", "I49", "I55", "I61"],
                                    [(27, 28), (34, 35), (47, 48), (53, 54), (59, 60), (64, 68)]),
    "1.2 Customer":                (13, 6, 7, ["I34", "I42", "I49"],
                                    [(31, 33), (39, 41), (47, 48), (54, 54)]),
    "1.3 Enterprise Data":         (22, 6, 7, ["I31", "I37"], [(27, 30), (36, 36)]),
    "1.4 TDD Group Functions":     (21, 6, 7, ["I25"], [(21, 24), (30, 30)]),
    "1.5 P&C":                     (18, 6, 7, ["I27"], [(25, 26), (32, 32)]),
    "1.6 Finance":                 (19, 6, 7, ["I28"], [(25, 27), (33, 33)]),
    "1.7 Infrastructure":          (17, 7, 8, ["I27", "I34", "I40"],
                                    [(25, 26), (32, 33), (39, 39)]),
    "1.8 Energy Solutions & B2B":  (16, 6, 7, ["I30", "I36"], [(27, 29), (35, 35)]),
    "1.9 Commercial Fuels":        (15, 6, 7, ["I28", "I34", "I40"],
                                    [(26, 27), (33, 33), (39, 39)]),
    "1.10 Z Retail":               (12, 6, 7, ["I27", "I34", "I40"],
                                    [(26, 26), (32, 33), (39, 39)]),
}

# tab -> (TDD Variance row, Other Variance row, Total row, left-to-fund cell)
VAR = {
    "1.1 Ampol Retail": (12, 13, 14, "J21"),
    "1.2 Customer": (17, 18, 19, "J26"),
    "1.3 Enterprise Data": (17, 18, 19, "J17"),
    "1.4 TDD Group Functions": (15, 16, 17, "J15"),
    "1.5 P&C": (16, 17, 18, "J21"),
    "1.6 Finance": (17, 18, 19, "J22"),
    "1.7 Infrastructure": (17, 18, 19, "J21"),
    "1.8 Energy Solutions & B2B": (16, 17, 18, "J23"),
    "1.9 Commercial Fuels": (16, 17, 18, "J21"),
    "1.10 Z Retail": (17, 18, 19, "J21"),
}


def derive(wb):
    """Read each tab's real platform blocks rather than trusting a hand-written map.

    Hand-coding these was wrong on three tabs - 1.1's EGI Retail range ran to the
    platform TOTAL row, so the total would have counted itself.
    """
    out = {}
    for sn in GEOM:
        ws = wb[sn]
        blocks, oh = [], []
        r = 1
        while r <= ws.max_row:
            if str(ws.cell(r, 2).value or "").strip().startswith("Platform:"):
                first = r + 2
                last, rr = None, first
                while rr <= ws.max_row:
                    lab = str(ws.cell(rr, 2).value or "").strip()
                    if lab == "Platform Overhead":
                        oh.append(f"I{rr}"); last = rr - 1; break
                    if lab.endswith("Total"):
                        last = rr - 1; break
                    rr += 1
                if last is None:
                    last = first
                while last >= first and ws.cell(last, 2).value in (None, ""):
                    last -= 1
                if last >= first:
                    blocks.append((first, last))
                r = rr
            r += 1
        out[sn] = (blocks, oh)
    return out


def normalise(wb):
    out = []
    derived = derive(wb)
    for tab, (brow, po, plo, _oh, _bl) in GEOM.items():
        blocks, oh_cells = derived[tab]
        ws = wb[tab]
        au, nz = f"{DC}!$C${brow}", f"{DC}!$D${brow}"
        if tab == "1.2 Customer":
            # Customer's budget spans rows 13 (Ampol) and 14 (Z)
            au = f"({DC}!$C$13+{DC}!$C$14)"
            nz = f"({DC}!$D$13+{DC}!$D$14)"

        # 1. portfolio overhead: one convention, mirrored, no multiplier
        before_c, before_d = ws[f"C{po}"].value, ws[f"D{po}"].value
        ws[f"C{po}"] = f"=IF({nz}>{au},0,{DC}!$L$10)"
        ws[f"D{po}"] = f"=IF({nz}>{au},{DC}!$L$10,0)"
        if str(before_c) != str(ws[f"C{po}"].value):
            out.append(f"{tab}!C{po}/D{po} portfolio overhead normalised")

        # 2. platform overhead: identical range on both branches, every block's cell
        rng = ",".join(oh_cells)
        ws[f"C{plo}"] = f"=IF({nz}>{au},0,SUM({rng}))"
        ws[f"D{plo}"] = f"=IF({nz}>{au},SUM({rng}),0)"
        out.append(f"{tab}!C{plo}/D{plo} platform overhead = SUM({rng}) on both branches")

        # 3. squad support: one SUMIF per block, bounded to that block's squad rows only,
        #    so no overhead row or platform total can leak in, and AU/NZ are identical
        #    apart from the tag
        def terms(tag):
            return "+".join(f'SUMIF($F${a}:$F${b},"{tag}",$I${a}:$I${b})' for a, b in blocks)
        ws[f"C{plo+1}"] = "=" + terms("AU")
        ws[f"D{plo+1}"] = "=" + terms("NZ")
        ws[f"E{plo+1}"] = "=" + "+".join(f"SUM($J${a}:$J${b})" for a, b in blocks)
        out.append(f"{tab}!C{plo+1}/D{plo+1}/E{plo+1} squad support bounded to squad rows")

    # 4. variance conventions: positive = over budget / still to fund, on all ten
    for tab, (tv, ov, tot, ltf) in VAR.items():
        ws = wb[tab]
        po = GEOM[tab][1]
        ws[f"C{tv}"] = f"=$I${po+2}+$I${po+3}"          # AU over + NZ over
        ws[f"C{ov}"] = f"=${ltf}"                        # never negated
        ws[f"C{tot}"] = f"=$C{tv}+$C{ov}"
        ws[f"B{tv}"] = "TDD Variance (positive = over budget)"
        ws[f"B{ov}"] = "Other Variance (positive = still to fund)"
    out.append("TDD and Other Variance normalised to one sign convention on all ten tabs")

    # 5. the labels that were missing beside a live number
    for tab in ("1.5 P&C", "1.8 Energy Solutions & B2B", "1.9 Commercial Fuels"):
        ws = wb[tab]
        po = GEOM[tab][1]
        if ws[f"H{po+4}"].value is None:
            ws[f"H{po+4}"] = "TDD over/(under) budget ($m)"
    out.append("1.5/1.8/1.9 H10 label restored beside the live I10 figure")

    # 6. flag, do not silently patch, the two data gaps
    ws = wb["1.8 Energy Solutions & B2B"]
    ws["B21"] = ("Budget block does not reconcile: E12 totals $8.9m against $7.2m in "
                 "Finance. I14 and I15 are hand-typed 2.5s in a column that otherwise "
                 "references 0.1 Budget Table (Fin).")
    ws["B21"].font = ITAL
    ws = wb["1.10 Z Retail"]
    ws["B23"] = ("I17 reads 0.1 Budget Table (Fin) O6, which holds the text \"-\" rather "
                 "than a number; IFERROR masks it as zero.")
    ws["B23"].font = ITAL
    out.append("1.8 budget gap and 1.10 text-budget cell flagged on the tabs")
    return out


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = normalise(wb)
    wb.save(dst)
    return out


if __name__ == "__main__":
    for x in run("oh.xlsx", "n1x.xlsx"):
        print("  ", x)
