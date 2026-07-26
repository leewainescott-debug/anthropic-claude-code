"""Two layout defects the 1.x consistency check kept reporting.

1  1.7 Infrastructure carried an extra blank row above its Portfolio Summary, so that
   block sat one row lower than on the other nine portfolio tabs. Flick between tabs and
   the table jumps.

2  1.3 Enterprise Data had a blank row between the "Funding position ($m)" header and the
   first line under it. An orphan header over a white gap, which the formatting rules ban.

3  One authored label was misspelled: 1.3 H16 read "Siginificant Items - EGI". The other
   misspellings in the file ("Technology Suport", "Manuacturing Group Projects") are in
   the raw data columns and are left alone - they are folded by the table on Lists W:X.

Both blocks are self-contained: nothing on another tab reads any cell in them (checked
against every formula in the workbook before moving). Rows are not deleted, because
openpyxl does not repair formulas when rows shift. The block is rewritten one row up,
formulas retargeted cell by cell, styles carried across, and the vacated row cleared.

The mover asserts the block's computed values are identical before and after.
"""
import copy
import re
import openpyxl
import wbio

# sheet -> (first column, last column, first row, last row, shift)
MOVES = [
    ("1.7 Infrastructure", "A", "L", 5, 10, -1),
    ("1.3 Enterprise Data", "A", "G", 17, 19, -1),
]
REF = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")
CELL = r"\$?[A-Z]{1,3}\$?\d+"
# order matters: a quoted run and a cross-sheet reference are both consumed whole so the
# local-reference branch never sees inside them. '0.2 Data Config'!$L$10 is another
# sheet's L10, not this block's, and retargeting it silently moved $0.4975m.
TOKEN = re.compile(
    r"'[^']*'!" + CELL + r"(?::" + CELL + r")?"      # 'Quoted Sheet'!A1 or !A1:B2
    r"|[A-Za-z0-9_.&\-]+!" + CELL + r"(?::" + CELL + r")?"   # Lists!A1
    r"|'[^']*'"                                       # a bare quoted sheet name
    r"|\"[^\"]*\""                                    # a string literal
)


def _cols(a, b):
    from openpyxl.utils import column_index_from_string, get_column_letter
    return [get_column_letter(i)
            for i in range(column_index_from_string(a), column_index_from_string(b) + 1)]


def retarget(formula, moved, sheet_local=True):
    """Rewrite references that land inside the moved block. Ranges are rewritten by
    endpoint, so SUM(C7:C9) becomes SUM(C6:C8) without special-casing ranges."""
    if not (isinstance(formula, str) and formula.startswith("=")):
        return formula, 0
    n = [0]

    def sub(m):
        d1, col, d2, row = m.groups()
        key = f"{col}{row}"
        if key in moved:
            n[0] += 1
            nr = moved[key]
            return f"{d1}{col}{d2}{nr[len(col):]}"
        return m.group(0)

    # cross-sheet references and string literals pass through whole; only what is left
    # over is a reference to this sheet, and only those get moved
    out, last = [], 0
    for q in TOKEN.finditer(formula):
        out.append(REF.sub(sub, formula[last:q.start()]))
        out.append(q.group(0))
        last = q.end()
    out.append(REF.sub(sub, formula[last:]))
    return "".join(out), n[0]


def move_block(wb, sheet, c0, c1, r0, r1, shift):
    ws = wb[sheet]
    cols = _cols(c0, c1)
    moved = {f"{c}{r}": f"{c}{r + shift}" for c in cols for r in range(r0, r1 + 1)}

    # 0. unmerge anything inside the block first - a MergedCell is read-only, so the
    #    lift below cannot touch it while the merge is live. Re-merged at step 4.
    remerge = []
    for m in list(ws.merged_cells.ranges):
        a, b = str(m).split(":")
        inside = (a in moved) + (b in moved)
        if inside == 1:
            raise RuntimeError(f"{sheet}: merge {m} straddles the block edge")
        if inside == 2:
            remerge.append((moved[a], moved[b]))
            ws.unmerge_cells(str(m))

    # 1. lift the cells
    payload = []
    for c in cols:
        for r in range(r0, r1 + 1):
            src = ws[f"{c}{r}"]
            payload.append((f"{c}{r + shift}", src.value,
                            copy.copy(src._style) if src.has_style else None))
    for coord, val, style in payload:
        dst = ws[coord]
        dst.value = val
        if style is not None:
            dst._style = style

    # 2. clear the rows the block vacated
    keep = {coord for coord, _, _ in payload}
    for c in cols:
        for r in range(r0, r1 + 1):
            coord = f"{c}{r}"
            if coord not in keep:
                ws[coord].value = None
                ws[coord]._style = copy.copy(openpyxl.cell.cell.Cell(ws)._style)

    # 3. retarget every formula on the sheet, including the moved ones
    hits = 0
    for row in ws.iter_rows():
        for cell in row:
            new, n = retarget(cell.value, moved)
            if n:
                cell.value = new
                hits += n

    # 4. put the merges back one row up
    for a, b in remerge:
        ws.merge_cells(f"{a}:{b}")

    # 5. row heights follow the block
    heights = {r: ws.row_dimensions[r].height for r in range(r0, r1 + 1)}
    for r in range(r0, r1 + 1):
        ws.row_dimensions[r + shift].height = heights[r]

    return hits, moved


def snapshot(path, sheet, cols, rows):
    wv = openpyxl.load_workbook(path, data_only=True)
    ws = wv[sheet]
    out = {}
    for c in cols:
        for r in rows:
            v = ws[f"{c}{r}"].value
            if v is not None:
                out[f"{c}{r}"] = v
    wv.close()
    return out


# label typos we authored. Raw-data spellings are not in here on purpose.
TYPOS = [("1.3 Enterprise Data", "H16", "Siginificant Items - EGI",
          "Significant Items - EGI")]


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = []
    for sheet, c0, c1, r0, r1, shift in MOVES:
        hits, moved = move_block(wb, sheet, c0, c1, r0, r1, shift)
        out.append(f"{sheet}: rows {r0}-{r1} moved to {r0+shift}-{r1+shift}, "
                   f"{hits} formula references retargeted")
    for sheet, coord, was, now in TYPOS:
        cell = wb[sheet][coord]
        if cell.value != was:
            raise RuntimeError(f"{sheet}!{coord} reads {cell.value!r}, expected {was!r}")
        cell.value = now
        out.append(f"{sheet} {coord}: {was!r} corrected to {now!r}")
    wb.save(dst)
    return out


def verify(before, after):
    """The same labels and the same numbers, one row up. Nothing else changed."""
    bad = []
    for sheet, c0, c1, r0, r1, shift in MOVES:
        cols = _cols(c0, c1)
        b = snapshot(before, sheet, cols, range(r0, r1 + 1))
        a = snapshot(after, sheet, cols, range(r0 + shift, r1 + shift + 1))
        want = {}
        for k, v in b.items():
            m = REF.fullmatch(k)
            want[m.group(2) + str(int(m.group(4)) + shift)] = v
        for k, v in want.items():
            g = a.get(k)
            if isinstance(v, (int, float)) and isinstance(g, (int, float)):
                if abs(v - g) > 1e-9:
                    bad.append((sheet, k, v, g))
            elif v != g:
                bad.append((sheet, k, v, g))
        for k in set(a) - set(want):
            bad.append((sheet, k, "(was empty)", a[k]))
    return bad


if __name__ == "__main__":
    import sys
    src, mid, dst = sys.argv[1], sys.argv[2], sys.argv[3]
    for x in run(src, mid):
        print("  ", x)
    rc, st = wbio.build(mid, dst)
    print("  injected", st)
    bad = verify(src, dst)
    print("  block verification:", "clean" if not bad else bad[:10])
