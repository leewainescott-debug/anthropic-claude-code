"""Cost after decisions counted every filled person twice.

The column read:

    SUMIFS(REVIEW cost WHERE status = "Filled" AND squad = this squad)
  + SUMIFS(H, F = this squad)

which was right when only vacant roles carried a lever. Filled roles now carry one too
(at a factor of 1.0), so column H already contains every person on the tab, filled and
vacant. Adding the filled total on top counted them a second time.

Ampol Web has eight people, all filled, and no vacancies. It cost $1.591m and reported
$3.182m after decisions - a decision surface showing a change of $1.591m when nobody had
touched a dropdown.

Column H is the whole answer: role cost times its lever factor, every role, one row each.

Also removes the three design squads that have nobody in REVIEW. They contributed
archetype cost against zero actual people, which is the one thing archetype-versus-actual
must not do.
"""
import re
import openpyxl
import wbio

DOUBLE = re.compile(
    r"^=\(SUMIFS\('[^']+'!\$AA\$\d+:\$AA\$\d+,[^)]*?\"Filled\"\)"
    r"\+(SUMIFS\(\$H\$\d+:\$H\$\d+,\$F\$\d+:\$F\$\d+,\$B\d+\))\)/(\d+)$")

EMPTY_SQUADS = [("1.2 Customer", "Digital Support NZ"),
                ("1.3 Enterprise Data", "EGI Data"),
                ("1.3 Enterprise Data", "Enterprise Data Delivery")]


def fix_after(wb):
    n = 0
    for sn in [s for s in wb.sheetnames if re.match(r"^2\.\d+ ", s)]:
        ws = wb[sn]
        for row in ws.iter_rows():
            for c in row:
                if c.column_letter != "O":
                    continue
                m = DOUBLE.match(str(c.value or ""))
                if m:
                    c.value = f"={m.group(1)}/{m.group(2)}"
                    n += 1
    return [f"2.x column O: {n} squad rows no longer add the filled cost on top of the "
            f"lever column that already contains it"]


def drop_empty(wb):
    """Blank the squad row, leaving the row in place so nothing below it shifts."""
    out = []
    for tab, squad in EMPTY_SQUADS:
        ws = wb[tab]
        hit = None
        for r in range(1, ws.max_row + 1):
            if str(ws[f"B{r}"].value or "").strip() == squad:
                hit = r
                break
        if hit is None:
            raise RuntimeError(f"{tab}: no row for {squad!r}")
        for col in "BCDEFGHIJKL":
            ws[f"{col}{hit}"].value = None
        out.append(f"{tab}: removed design squad {squad!r} (row {hit}), nobody in REVIEW "
                   f"carries that name")
    return out


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = fix_after(wb) + drop_empty(wb)
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    src, mid, dst = sys.argv[1], sys.argv[2], sys.argv[3]
    for x in run(src, mid):
        print("  ", x)
    rc, st = wbio.build(mid, dst)
    print("  injected", st)
