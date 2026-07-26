"""Exec Summary, Data QA, and the housekeeping the owner called out.

- Exec Summary was 77 rows with the commentary written into the cells. It is now the
  story in plain lines with a portfolio drill-down, reading the rebuilt 2.x and 3.x.
- 4.0 Data QA repointed at the new anchors and stated as check / expected / difference.
- 1.14 TDD Cyber deleted. The brief says cyber appears once; it appeared twice.
- Tab order follows the flow. 1.10 sorted after 1.9 instead of between 1.1 and 1.2.
- Every explanatory sentence left inside a cell on the owner's own tabs is removed, and
  the red commentary with it.
"""
import json
import re

import openpyxl
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation

import fmt
import wb2

REVIEW = "REVIEW - Complete Role Mapping"
REV = f"'{REVIEW}'"
LAST = 528


def build_exec(wb, a2, a3):
    ws = wb["Exec Summary"]
    wb2.wipe(ws)
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    t31, t32 = a3["3.1"]["total"], a3["3.2"]["total"]
    g33 = a3["3.3"]["group_total"]

    wb2.title(ws, 2, "TDD operating model - executive summary")

    def block(r, heading, lines):
        wb2.title(ws, r, heading, 12)
        fmt.band(ws, r, 2, 3, fmt.GRP_FILL)
        ws.cell(r, 2).alignment = fmt.LEFT
        r += 1
        for lab, f, nf in lines:
            ws.cell(r, 2).value = lab
            ws.cell(r, 2).font = fmt.BODY
            ws.cell(r, 2).alignment = fmt.LEFT
            fmt.money(ws, r, 3, f, nf)
            r += 1
        return r + 1

    r = 4
    r = block(r, "The organisation today", [
        ("Roles in the ledger", f"='3.3 FTE View'!$G${g33}", fmt.COUNT),
        ("Filled", f"='3.3 FTE View'!$H${g33}", fmt.COUNT),
        ("Vacant", f"='3.3 FTE View'!$I${g33}", fmt.COUNT),
        ("Cost today ($m)", f"='3.1 Group Summary'!$D${t31}", fmt.MONEY_M)])

    r = block(r, "Against the archetype", [
        ("Archetype cost - delivery squads ($m)", f"='3.2 Total Cost'!$C${t32}",
         fmt.MONEY_M),
        ("Actual cost ($m)", f"='3.2 Total Cost'!$D${t32}", fmt.MONEY_M),
        ("Over/(under) the archetype ($m)", f"='3.2 Total Cost'!$E${t32}", fmt.MONEY_M)])

    r = block(r, "Against the budget", [
        ("TDD lights-on budget ($m)", f"='3.1 Group Summary'!$C${t31}", fmt.MONEY_M),
        ("Actual cost ($m)", f"='3.1 Group Summary'!$D${t31}", fmt.MONEY_M),
        ("Over/(under) budget ($m)", f"='3.1 Group Summary'!$E${t31}", fmt.MONEY_M)])

    r = block(r, "The vacancy decision", [
        ("Vacant roles", f"='3.3 FTE View'!$I${g33}", fmt.COUNT),
        ("Cost of hiring every vacancy ($m)",
         f"=SUMIFS({REV}!$AA$2:$AA${LAST},{REV}!$AK$2:$AK${LAST},\"Vacant\")/1000000",
         fmt.MONEY_M),
        ("Impact of the decisions set today ($m)", f"='3.2 Total Cost'!$F${t32}",
         fmt.MONEY_M),
        ("Total cost after those decisions ($m)", f"='3.2 Total Cost'!$G${t32}",
         fmt.MONEY_M)])

    # ---- portfolio drill-down ----
    wb2.title(ws, r, "Portfolio drill-down", 12)
    fmt.band(ws, r, 2, 3, fmt.GRP_FILL)
    ws.cell(r, 2).alignment = fmt.LEFT
    r += 1
    sel = r
    ws.cell(r, 2).value = "Pick a portfolio"
    ws.cell(r, 2).font = fmt.BOLD
    ws.cell(r, 2).alignment = fmt.LEFT
    pick = ws.cell(r, 3)
    pick.value = "Ampol Retail"
    pick.fill = fmt.IN_FILL
    pick.border = fmt.BOX
    pick.font = fmt.BODY
    pick.alignment = fmt.CENTRE
    names = [p for p in wb2.PORTFOLIO_ORDER]
    dv = DataValidation(type="list", formula1='"' + ",".join(names) + '"',
                        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(pick)
    r += 1
    f31, l31 = a3["3.1"]["first"], a3["3.1"]["last"]
    f32, l32 = a3["3.2"]["first"], a3["3.2"]["last"]
    for lab, sheet, col, lo, hi, nf in [
            ("TDD lights-on budget ($m)", "3.1 Group Summary", "C", f31, l31, fmt.MONEY_M),
            ("Actual cost ($m)", "3.1 Group Summary", "D", f31, l31, fmt.MONEY_M),
            ("Over/(under) budget ($m)", "3.1 Group Summary", "E", f31, l31, fmt.MONEY_M),
            ("Archetype cost ($m)", "3.2 Total Cost", "C", f32, l32, fmt.MONEY_M),
            ("Variance to archetype ($m)", "3.2 Total Cost", "E", f32, l32, fmt.MONEY_M),
            ("Impact of decisions ($m)", "3.2 Total Cost", "F", f32, l32, fmt.MONEY_M),
            ("Total cost after decisions ($m)", "3.2 Total Cost", "G", f32, l32,
             fmt.MONEY_M)]:
        ws.cell(r, 2).value = lab
        ws.cell(r, 2).font = fmt.BODY
        ws.cell(r, 2).alignment = fmt.LEFT
        fmt.money(ws, r, 3,
                  f"=IFERROR(INDEX('{sheet}'!${col}${lo}:${col}${hi},"
                  f"MATCH($C${sel},'{sheet}'!$B${lo}:$B${hi},0)),\"-\")", nf)
        r += 1
    for lab, col in (("Roles", "G"), ("Filled", "H"), ("Vacant", "I")):
        ws.cell(r, 2).value = lab
        ws.cell(r, 2).font = fmt.BODY
        ws.cell(r, 2).alignment = fmt.LEFT
        fmt.money(ws, r, 3,
                  f"=SUMIFS('3.3 FTE View'!${col}$6:${col}${g33 - 1},"
                  f"'3.3 FTE View'!$B$6:$B${g33 - 1},$C${sel},"
                  f"'3.3 FTE View'!$C$6:$C${g33 - 1},\"<>\")/2", fmt.COUNT)
        r += 1
    ws.freeze_panes = "C4"
    return r


def build_qa(wb, a2, a3):
    ws = wb["4.0 Data QA"]
    wb2.wipe(ws)
    ws.column_dimensions["A"].width = 2.5
    wb2.title(ws, 2, "Data QA - every check must read zero")
    HDR = 4
    fmt.header(ws, HDR, 2, ["Check", "Model", "Expected", "Difference"],
               [64, 16, 16, 14])
    t31 = a3["3.1"]["total"]
    t32 = a3["3.2"]["total"]
    g33 = a3["3.3"]["group_total"]
    checks = [
        ("Roles on 3.3 against the ledger", f"='3.3 FTE View'!$G${g33}",
         f"=COUNTA({REV}!$B$2:$B${LAST})", fmt.COUNT),
        ("Filled on 3.3 against the ledger", f"='3.3 FTE View'!$H${g33}",
         f'=COUNTIFS({REV}!$AK$2:$AK${LAST},"Filled")', fmt.COUNT),
        ("Vacant on 3.3 against the ledger", f"='3.3 FTE View'!$I${g33}",
         f'=COUNTIFS({REV}!$AK$2:$AK${LAST},"Vacant")', fmt.COUNT),
        ("Cost on 3.1 against the ledger ($m)", f"='3.1 Group Summary'!$D${t31}",
         f"=SUM({REV}!$AA$2:$AA${LAST})/1000000", fmt.MONEY_M),
        ("Cost on 3.2 against the ledger ($m)", f"='3.2 Total Cost'!$D${t32}",
         f"=SUM({REV}!$AA$2:$AA${LAST})/1000000", fmt.MONEY_M),
        ("Cost on 3.3 against the ledger ($m)", f"='3.3 FTE View'!$K${g33}",
         f"=SUM({REV}!$AA$2:$AA${LAST})/1000000", fmt.MONEY_M),
        ("3.1 and 3.2 state the same actual cost ($m)", f"='3.1 Group Summary'!$D${t31}",
         f"='3.2 Total Cost'!$D${t32}", fmt.MONEY_M),
        ("3.1 and 3.2 state the same cost after decisions ($m)",
         f"='3.1 Group Summary'!$F${t31}", f"='3.2 Total Cost'!$G${t32}", fmt.MONEY_M),
    ]
    r = HDR + 1
    for tab, inf in a2.items():
        t = inf["total_row"]
        checks.append((f"{tab} roles against the ledger",
                       f"='{tab}'!${L(wb2.S_ROLES)}${t}",
                       f"=COUNTIFS({REV}!$AJ$2:$AJ${LAST},\"{inf['pf']}\")", fmt.COUNT))
    for lab, m, e, nf in checks:
        ws.cell(r, 2).value = lab
        ws.cell(r, 2).font = fmt.BODY
        ws.cell(r, 2).alignment = fmt.LEFT
        fmt.money(ws, r, 3, m, nf)
        fmt.money(ws, r, 4, e, nf)
        fmt.money(ws, r, 5, f"=ROUND($C{r}-$D{r},6)", nf)
        r += 1
    fmt.band(ws, r, 2, 4, fmt.TOT_FILL, line=True)
    ws.cell(r, 2).value = "Checks failing"
    ws.cell(r, 2).alignment = fmt.LEFT
    fmt.money(ws, r, 5, f'=COUNTIF($E{HDR + 1}:$E{r - 1},"<>0")', fmt.COUNT)
    ws.freeze_panes = "C5"
    return r - HDR


SLOP = re.compile(
    r"^(Delivery squads are compared|Archetype columns|Cost is stated gross|"
    r"Drawn is the amount|Planned spend is gross|The 8 GMs are the only|"
    r"Full Cost AUD =|Banded rate agreed|Group view\.|One cost statement|"
    r"Data QA\.|Vacant = open roles|Vacant roles are priced|Org roles include|"
    r"Role counts and dollars|Yellow cell = dropdown|This tab was a copy|"
    r"Squads are priced|Overhead is compared|Strategic Programs squads|"
    r"TDD Cyber is priced|Delivery squads and overhead|The main lever|"
    r"Resizing the archetypes|charged on the first|Budget block does not|"
    r"P&C has a \$1\.0m NZ budget|No cost in column|Agreed assignments|"
    r"Squad name map retired)")


def housekeeping(wb):
    out = []
    # 1.14 - cyber twice
    if "1.14 TDD Cyber" in wb.sheetnames:
        del wb["1.14 TDD Cyber"]
        out.append("1.14 TDD Cyber deleted - the brief says cyber appears once")

    # sentences left inside cells, and any red commentary
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                red = ""
                if c.font and c.font.color and c.font.color.rgb:
                    red = str(c.font.color.rgb).upper()
                if SLOP.match(v.strip()) or (red in ("FFFF0000", "FFC00000")
                                             and len(v) > 45):
                    c.value = None
                    n += 1
                elif red in ("FFFF0000", "FFC00000"):
                    c.font = fmt.BODY
    out.append(f"{n} explanatory sentences removed from cells, red commentary cleared")

    # tab order: inputs, 1.x, 2.x, 3.x, evidence
    def key(name):
        m = re.match(r"^(\d+)\.(\d+) ", name)
        if m:
            return (2 + int(m.group(1)), int(m.group(2)), name)
        order = {"Exec Summary": (0, 0, ""), "- INPUTS -": (1, 0, ""),
                 REVIEW: (1, 8, ""), "Portfolios": (1, 9, ""),
                 "- WORKING -": (3, -1, ""), "- SUMMARIES -": (5, -1, ""),
                 "- EVIDENCE -": (6, -1, "")}
        if name in order:
            return order[name]
        return (9, 0, name)
    wb._sheets.sort(key=lambda s: key(s.title))
    for sep, before in (("- WORKING -", "2.1 Ampol Retail"),
                        ("- SUMMARIES -", "3.1 Group Summary")):
        if sep not in wb.sheetnames and before in wb.sheetnames:
            s = wb.create_sheet(sep)
            wb.move_sheet(s, offset=wb.sheetnames.index(before) - wb.sheetnames.index(sep))
    if "ACTUAL WORKBOOKS" in wb.sheetnames:
        del wb["ACTUAL WORKBOOKS"]
    wb._sheets.sort(key=lambda s: key(s.title))
    out.append("tab order set to the flow: inputs, 1.x, 2.x, 3.x, evidence")
    return out


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    a2 = json.load(open("anchors2.json"))
    a3 = json.load(open("anchors3.json"))
    build_exec(wb, a2, a3)
    n = build_qa(wb, a2, a3)
    out = [f"Exec Summary rebuilt as the story plus a portfolio drill-down",
           f"4.0 Data QA rebuilt: {n} checks, each stated as model / expected / difference"]
    out += housekeeping(wb)
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
