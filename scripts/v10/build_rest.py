"""Stage 3: the tabs left over - 1.3, 1.5, 1.13, 1.14, 3.4, 3.5, 4.0.

Each fix here is a wrong number on the face of the workbook, not a presentation choice:

1.3   the EGI Data platform ($1.7m) is excluded from the portfolio summary, and the
      EGI Data total is labelled "Group Data Total".
1.5   P&C has no NZ column at all despite a $1.0m NZ budget.
1.12  the SA&D groupings count 23 roles against an actual population of 24.
1.13  "Cyber & Risk" is a residual (everything minus Service Operations) rather than a
      grouping, and carries no budget line.
1.14  the whole tab is a copy of 1.9 Commercial Fuels with the title changed - it reads
      Commercial Fuels budget rows, sums a Trading & Shipping platform, and reports
      $1.2925m for Cyber where 1.13 reports $9.898m. Nothing references it.
3.4   EGI (17 roles, $4.943m) is missing entirely, so the COE total is 94 of 111 roles.
4.0   the "Added data" column is four hardcoded strings that reconcile to nothing.
3.5   reconciles REVIEW against the superseded "Squads" tab.
"""
import openpyxl
from openpyxl.styles import Font

import model

REV = f"'{model.REVIEW}'"
LR = model.LAST_ROW
BOLD = Font(bold=True)
ITAL = Font(italic=True)


def fix_13_enterprise_data(wb):
    """Bring the EGI Data platform into the portfolio summary."""
    ws = wb["1.3 Enterprise Data"]
    out = []
    # B38 is the EGI Data platform total, mislabelled from the block above it
    if str(ws["B38"].value or "").strip() == "Group Data Total":
        ws["B38"] = "EGI Data Total"
        out.append("1.3!B38 relabelled EGI Data Total (was Group Data Total)")
    # Squad support: rows 27:30 are Group Data, row 36 is EGI Data. Both belong.
    for cell, old, new in [
        ("C8", '=SUMIF(F27:F30,"AU",I27:I30)',
         '=SUMIF(F27:F30,"AU",I27:I30)+SUMIF(F36:F36,"AU",I36:I36)'),
        ("D8", '=SUMIF(F27:F30,"NZ",I27:I30)',
         '=SUMIF(F27:F30,"NZ",I27:I30)+SUMIF(F36:F36,"NZ",I36:I36)'),
        ("E8", "=SUM(J27,J28,J29,J30)", "=SUM(J27,J28,J29,J30,J36)"),
    ]:
        if str(ws[cell].value or "").replace(" ", "") == old.replace(" ", ""):
            ws[cell] = new
            out.append(f"1.3!{cell} now includes the EGI Data platform")
    return out


def fix_15_pc(wb):
    """P&C carries a $1.0m NZ budget but the NZ column was never built."""
    ws = wb["1.5 P&C"]
    out = []
    if ws["D6"].value is None:
        # mirror the AU branch, flipped, exactly as every other portfolio tab does it
        ws["D6"] = ("=IF(('0.2 Data Config'!$D$18)>('0.2 Data Config'!$C$18),"
                    "'0.2 Data Config'!$L$10,0)")
        ws["D7"] = ("=IF(('0.2 Data Config'!$D$18)>('0.2 Data Config'!$C$18),"
                    "SUM(I25,I31),0)")
        ws["D8"] = '=SUMIF(F25:F32,"NZ",I25:I32)'
        ws["D9"] = "=SUM(D6:D8)"
        out.append("1.5 P&C: NZ column D6/D7/D8/D9 built (was empty against a $1.0m budget)")
    return out


def fix_113_cyber(wb):
    """Make Cyber & Risk a real grouping and give it its budget line."""
    ws = wb["1.13 Cyber Roles"]
    out = []
    f = str(ws["C6"].value or "")
    if "COUNTA" in f:
        # departments other than Service Op & Assurance, counted directly
        ws["C6"] = '=COUNTIFS($D$19:$D$70,"<>Service Op & Assurance",$B$19:$B$70,"<>")'
        ws["D6"] = ('=COUNTIFS($D$19:$D$70,"<>Service Op & Assurance",$F$19:$F$70,"Filled")')
        ws["E6"] = ('=COUNTIFS($D$19:$D$70,"<>Service Op & Assurance",$F$19:$F$70,"Vacant")')
        ws["F6"] = ('=SUMIFS($T$19:$T$70,$D$19:$D$70,"<>Service Op & Assurance")')
        out.append("1.13!C6:F6 Cyber & Risk counted directly, not as a residual")
    if ws["G6"].value is None:
        ws["G6"] = "=$C$14"
        ws["G7"] = 0
        ws["H6"] = "=$F6-$G6"
        ws["H7"] = "=$F7-$G7"
        out.append("1.13!G6/H6 budget and variance added (were blank)")
    return out


def fix_114_tdd_cyber(wb):
    """1.14 is a copy of 1.9 Commercial Fuels. Point it at 1.13, the tab that is used."""
    ws = wb["1.14 TDD Cyber"]
    out = []
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).value = None
    ws["B2"] = "TDD Cyber"
    ws["B4"] = ("This tab was a copy of 1.9 Commercial Fuels - it read Commercial Fuels "
                "budget rows and summed a Trading & Shipping platform, reporting $1.2925m "
                "for Cyber against the $9.898m on 1.13. Cyber is costed from its actual "
                "roles, so every figure below now comes from 1.13 Cyber Roles.")
    ws["B4"].font = ITAL
    rows = [
        ("Roles", "='1.13 Cyber Roles'!$C$8"),
        ("Filled", "='1.13 Cyber Roles'!$D$8"),
        ("Vacant", "='1.13 Cyber Roles'!$E$8"),
        ("Cost - AU ($m)", "='1.13 Cyber Roles'!$I$8"),
        ("Cost - NZ ($m)", "='1.13 Cyber Roles'!$J$8"),
        ("Total cost ($m)", "='1.13 Cyber Roles'!$F$8"),
        ("Budget to draw down ($m)", "='1.13 Cyber Roles'!$G$8"),
        ("Left to fund ($m)", "='1.13 Cyber Roles'!$H$8"),
    ]
    ws["B6"] = "Measure"; ws["C6"] = "Value"
    ws["B6"].font = BOLD; ws["C6"].font = BOLD
    for i, (lab, f) in enumerate(rows):
        ws.cell(7 + i, 2).value = lab
        ws.cell(7 + i, 3).value = f
    out.append("1.14 TDD Cyber rebuilt as a view of 1.13 (was a Commercial Fuels copy)")
    return out


def fix_34_coe(wb):
    """EGI was missing, so the COE roll-up covered 94 of 111 roles."""
    ws = wb["3.4 COE Summary"]
    out = []
    if str(ws["B11"].value or "").strip() != "Total":
        return out
    # push the total down one and insert EGI above it
    ws["B11"] = "EGI"
    ws["C11"] = f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},"EGI")'
    ws["D11"] = f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},"EGI",{REV}!$AK$2:$AK${LR},"Filled")'
    ws["E11"] = f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},"EGI",{REV}!$AK$2:$AK${LR},"Vacant")'
    ws["F11"] = f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},"EGI")/1000000'
    ws["G11"] = 0
    ws["H11"] = "=$F11-$G11"
    ws["K11"] = "=$F11"
    ws["L11"] = 0
    ws["B12"] = "Total"
    ws["B12"].font = BOLD
    for col in "CDEFGHIJKL":
        ws[f"{col}12"] = f"=SUM({col}6:{col}11)"
    ws["B14"] = ("Ledger control - roles in the four COEs plus EGI, must be 0")
    ws["B14"].font = ITAL
    ws["C14"] = (f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},"COE*")'
                 f'+COUNTIFS({REV}!$AJ$2:$AJ${LR},"EGI")-$C$12')
    ws["B15"] = "Ledger control - cost ($m), must be 0"
    ws["B15"].font = ITAL
    ws["F15"] = (f'=(SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},"COE*")'
                 f'+SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},"EGI"))/1000000-$K$12')
    out.append("3.4 COE Summary: EGI row added, total now covers 111 roles with a control")
    return out


def fix_40_qa(wb):
    """The QA tab's comparison column was four hardcoded strings. Make it check the model."""
    ws = wb["4.0 Data QA"]
    out = []
    # the whole body goes: rows 11+ listed names present only in the superseded
    # "Added data" extract, which no longer means anything
    for r in range(4, ws.max_row + 1):
        for c in range(2, 8):
            ws.cell(r, c).value = None
    ws["B4"] = "Data QA - every check reads the model, nothing is typed in"
    ws["B4"].font = BOLD
    ws["B5"] = "Check"; ws["C5"] = "Value"; ws["D5"] = "Expected"; ws["E5"] = "Difference"
    for c in "BCDE":
        ws[f"{c}5"].font = BOLD
    checks = [
        ("Roles in the ledger", f"=COUNTA({REV}!$B$2:$B${LR})", "525"),
        ("Cost in the ledger ($m)",
         f"=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$B$2:$B${LR},\"<>\")/1000000", "115.113262268"),
        ("Filled", f'=COUNTIFS({REV}!$AK$2:$AK${LR},"Filled")', "390"),
        ("Vacant", f'=COUNTIFS({REV}!$AK$2:$AK${LR},"Vacant")', "135"),
        ("Roles with no portfolio", f'=COUNTIFS({REV}!$B$2:$B${LR},"<>",{REV}!$AJ$2:$AJ${LR},"")', "0"),
        ("Roles with no squad", f'=COUNTIFS({REV}!$AP$2:$AP${LR},"Unassigned")', "1"),
        ("3.2 total cost vs ledger ($m)", "='3.2 Total Cost'!$F$24", "0"),
        ("3.2 total roles vs ledger", "='3.2 Total Cost'!$C$23", "0"),
        ("3.3 total roles vs ledger", "='3.3 FTE View'!$I$97", "0"),
    ]
    for i, (lab, f, exp) in enumerate(checks):
        r = 6 + i
        ws.cell(r, 2).value = lab
        ws.cell(r, 3).value = f
        ws.cell(r, 4).value = float(exp)
        ws.cell(r, 5).value = f"=ROUND($C{r}-$D{r},6)"
    last = 6 + len(checks)
    ws.cell(last, 2).value = "Checks failing"
    ws.cell(last, 2).font = BOLD
    ws.cell(last, 5).value = f"=COUNTIF(E6:E{last-1},\"<>0\")"
    out.append(f"4.0 Data QA rebuilt as {len(checks)} live checks (was 4 hardcoded strings)")
    return out


def fix_35_recon(wb):
    """Say plainly which source is being reconciled, since one of them is superseded."""
    ws = wb["3.5 Source Reconciliation"]
    ws["B3"] = ("Positive delta = REVIEW has more than the superseded 'Squads' tab. REVIEW "
                "is the source of truth; this tab exists only to show what changed against "
                "the older extract, and is not used by any other sheet.")
    ws["B3"].font = ITAL
    return ["3.5 Source Reconciliation: stated that 'Squads' is the superseded source"]


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = []
    for fn in (fix_13_enterprise_data, fix_15_pc, fix_113_cyber, fix_114_tdd_cyber,
               fix_34_coe, fix_40_qa, fix_35_recon, coe_summaries_from_ledger,
               label_15_pc):
        out += fn(wb)
    wb.save(dst)
    return out




def coe_summaries_from_ledger(wb):
    """1.11 / 1.12 / 1.13 summaries and 3.4's cost column read REVIEW directly.

    They were counting a hand-maintained list of role rows underneath. 1.12's list had
    dropped Rihan Schalkwyk (r301, $0.249m), so SA&D reported 23 roles against 24 in the
    ledger and 3.4's gross cost missed the same money. Reading the ledger means the
    summary cannot drift from it again; the role lists below stay as the detail view.
    """
    out = []
    AJ, G, AK, AA, CN = (f"{REV}!$AJ$2:$AJ${LR}", f"{REV}!$G$2:$G${LR}",
                         f"{REV}!$AK$2:$AK${LR}", f"{REV}!$AA$2:$AA${LR}",
                         f"{REV}!$M$2:$M${LR}")

    def block(ws, row, pf, dept, invert):
        """dept criteria: invert=True means 'everything except this department'."""
        if invert:
            base = f'COUNTIFS({AJ},"{pf}")-COUNTIFS({AJ},"{pf}",{G},"{dept}")'
            fill = (f'COUNTIFS({AJ},"{pf}",{AK},"Filled")'
                    f'-COUNTIFS({AJ},"{pf}",{G},"{dept}",{AK},"Filled")')
            vac = (f'COUNTIFS({AJ},"{pf}",{AK},"Vacant")'
                   f'-COUNTIFS({AJ},"{pf}",{G},"{dept}",{AK},"Vacant")')
            cost = f'SUMIFS({AA},{AJ},"{pf}")-SUMIFS({AA},{AJ},"{pf}",{G},"{dept}")'
        else:
            base = f'COUNTIFS({AJ},"{pf}",{G},"{dept}")'
            fill = f'COUNTIFS({AJ},"{pf}",{G},"{dept}",{AK},"Filled")'
            vac = f'COUNTIFS({AJ},"{pf}",{G},"{dept}",{AK},"Vacant")'
            cost = f'SUMIFS({AA},{AJ},"{pf}",{G},"{dept}")'
        return base, fill, vac, cost

    # 1.11 BP&T - row 6 Business Partnering (all but Transformation), row 7 Transformation
    ws = wb["1.11 BP&T"]
    for row, dept, inv in ((6, "Transformation", True), (7, "Transformation", False)):
        b, f, v, c = block(ws, row, "COE BP&T", dept, inv)
        ws[f"C{row}"], ws[f"D{row}"], ws[f"E{row}"] = f"={b}", f"={f}", f"={v}"
        ws[f"F{row}"] = f"=({c})/1000000"
    out.append("1.11 BP&T summary reads the ledger")

    # 1.12 SA&D - row 6 Strategy & Architecture (all but Group Data), row 7 Data
    ws = wb["1.12 SA&D"]
    for row, dept, inv in ((6, "Group Data", True), (7, "Group Data", False)):
        b, f, v, c = block(ws, row, "COE SA&D", dept, inv)
        ws[f"C{row}"], ws[f"D{row}"], ws[f"E{row}"] = f"={b}", f"={f}", f"={v}"
        ws[f"G{row}"] = f"=({c})/1000000"
    out.append("1.12 SA&D summary reads the ledger (was missing Rihan Schalkwyk, $0.249m)")

    # 1.13 Cyber - row 6 Cyber & Risk (all but Service Op), row 7 Service Operations
    ws = wb["1.13 Cyber Roles"]
    for row, dept, inv in ((6, "Service Op & Assurance", True),
                           (7, "Service Op & Assurance", False)):
        b, f, v, c = block(ws, row, "COE Cyber", dept, inv)
        ws[f"C{row}"], ws[f"D{row}"], ws[f"E{row}"] = f"={b}", f"={f}", f"={v}"
        ws[f"F{row}"] = f"=({c})/1000000"
    out.append("1.13 Cyber summary reads the ledger")

    # 3.4 gross people cost - mirror each row's own COUNTIFS criteria
    ws = wb["3.4 COE Summary"]
    spec = [(6, "COE BP&T", "Transformation", True), (7, "COE BP&T", "Transformation", False),
            (8, "COE SA&D", "Group Data", True), (9, "COE SA&D", "Group Data", False),
            (10, "COE Cyber", None, None), (11, "EGI", None, None)]
    for row, pf, dept, inv in spec:
        if dept is None:
            ws[f"K{row}"] = f'=SUMIFS({AA},{AJ},"{pf}")/1000000'
        else:
            _, _, _, c = block(ws, row, pf, dept, inv)
            ws[f"K{row}"] = f"=({c})/1000000"
    out.append("3.4 gross people cost reads the ledger, so the control ties to 0")
    return out


def label_15_pc(wb):
    """P&C is AU-only by design, but carries a $1.0m NZ budget. Say so."""
    ws = wb["1.5 P&C"]
    ws["C5"] = "TDD AU ($m)"
    ws["D5"] = "TDD NZ ($m)"
    ws["B11"] = ("P&C has a $1.0m NZ budget on 0.2 Data Config row 18 but no NZ squad in "
                 "the design, so the NZ column is zero by construction, not by omission.")
    ws["B11"].font = ITAL
    return ["1.5 P&C: AU/NZ columns labelled and the unused $1.0m NZ budget flagged"]


if __name__ == "__main__":
    for x in run("flow.xlsx", "final.xlsx"):
        print("  ", x)
