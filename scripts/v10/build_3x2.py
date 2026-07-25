"""Stage 10: rewire 3.1 / 3.2 / 3.3 onto the split 2.x geometry.

Archetype now covers delivery squads only, so every variance on the summaries compares the
same population on both sides. 3.2 gains a squad/overhead split that foots to the ledger.
"""
import openpyxl
from openpyxl.styles import Alignment, Font
import model, build_split

REV = f"'{model.REVIEW}'"; LR = model.LAST_ROW
BOLD = Font(bold=True); ITAL = Font(italic=True)

def hdr(ws,row,m):
    for c,h in m.items():
        x=ws[f"{c}{row}"]; x.value=h; x.font=BOLD
        x.alignment=Alignment(wrap_text=True,vertical="bottom")

def run(src, dst):
    roles = build_split.load_roles(src)
    wb = openpyxl.load_workbook(src)
    geom = build_split.build(wb, roles)
    order = list(dict.fromkeys(model.TAB_PORTFOLIO.values()))
    tab_of = {v:k for k,v in model.TAB_PORTFOLIO.items()}
    ARCH = [t for t in build_split.DESIGN]

    # ---------------- 3.2 ----------------
    ws = wb["3.2 Total Cost"]
    for r in range(4, ws.max_row+1):
        for c in range(1,17):
            if ws.cell(r,c).value is not None: ws.cell(r,c).value=None
    ws["B4"]=("One cost statement. Delivery squads are compared to the squad archetype; "
              "overhead is compared to the allowance. Both together are the ledger.")
    ws["B4"].font=ITAL
    hdr(ws,5,{"B":"Portfolio","C":"Roles","D":"Filled","E":"Vacant",
              "F":"Actual cost ($m)","G":"Squad roles","H":"Squad cost ($m)",
              "I":"Archetype roles","J":"Archetype cost ($m)",
              "K":"Variance to archetype ($m)","L":"Overhead roles","M":"Overhead cost ($m)",
              "N":"Cost after decisions ($m)"})
    r0=6
    for i,pf in enumerate(order):
        r=r0+i; t=tab_of[pf]; g=geom[t]
        ws[f"B{r}"]=pf
        ws[f"C{r}"]=f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r})'
        ws[f"D{r}"]=f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r},{REV}!$AK$2:$AK${LR},"Filled")'
        ws[f"E{r}"]=f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r},{REV}!$AK$2:$AK${LR},"Vacant")'
        ws[f"F{r}"]=f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},$B{r})/1000000'
        ws[f"G{r}"]=f"='{t}'!$E${g['asub']}"
        ws[f"H{r}"]=f"='{t}'!$M${g['asub']}"
        # a centre of excellence has no overhead block, so its overhead is nil by
        # construction rather than a reference to a cell that does not exist
        ws[f"L{r}"]=(f"='{t}'!$E${g['bsub']}" if g['ohs'] else "=0")
        ws[f"M{r}"]=(f"='{t}'!$M${g['bsub']}" if g['ohs'] else "=0")
        ws[f"N{r}"]=f"='{t}'!$O${g['tot']}"
        if t in ARCH:
            ws[f"I{r}"]=f"='{t}'!$D${g['asub']}"
            ws[f"J{r}"]=f"='{t}'!$L${g['asub']}"
            ws[f"K{r}"]=f"=$H{r}-$J{r}"
        else:
            ws[f"I{r}"]='="-"'; ws[f"J{r}"]='="-"'; ws[f"K{r}"]='="-"'
    n=len(ARCH); sub=r0+n
    ws[f"B{sub}"]="Portfolios with a squad archetype"; ws[f"B{sub}"].font=BOLD
    for c in "CDEFGHIJKLMN": ws[f"{c}{sub}"]=f"=SUM({c}{r0}:{c}{sub-1})"
    for i,pf in enumerate(order[n:]):
        r=sub+1+i; t=tab_of[pf]; g=geom[t]
        ws[f"B{r}"]=pf
        ws[f"C{r}"]=f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r})'
        ws[f"D{r}"]=f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r},{REV}!$AK$2:$AK${LR},"Filled")'
        ws[f"E{r}"]=f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r},{REV}!$AK$2:$AK${LR},"Vacant")'
        ws[f"F{r}"]=f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},$B{r})/1000000'
        ws[f"G{r}"]=f"='{t}'!$E${g['asub']}"; ws[f"H{r}"]=f"='{t}'!$M${g['asub']}"
        ws[f"L{r}"]=(f"='{t}'!$E${g['bsub']}" if g['ohs'] else "=0")
        ws[f"M{r}"]=(f"='{t}'!$M${g['bsub']}" if g['ohs'] else "=0")
        ws[f"N{r}"]=f"='{t}'!$O${g['tot']}"
        ws[f"I{r}"]='="-"'; ws[f"J{r}"]='="-"'; ws[f"K{r}"]='="-"'
    tr=sub+1+len(order)-n
    ws[f"B{tr}"]="Total"; ws[f"B{tr}"].font=BOLD
    for c in "CDEFGHLMN": ws[f"{c}{tr}"]=f"=${c}{sub}+SUM({c}{sub+1}:{c}{tr-1})"
    for c in "IJK": ws[f"{c}{tr}"]=f"=${c}{sub}"
    ws[f"B{tr+1}"]=("Archetype columns cover the ten portfolios with a squad archetype; "
                    "the COEs and EGI are measured against budget on 3.1.")
    ws[f"B{tr+1}"].font=ITAL
    c1,c2,c3=tr+3,tr+4,tr+5
    ws[f"B{c1}"]="Control - roles vs the ledger, must be 0"
    ws[f"C{c1}"]=f"=COUNTA({REV}!$B$2:$B${LR})-$C${tr}"
    ws[f"B{c2}"]="Control - cost vs the ledger ($m), must be 0"
    ws[f"F{c2}"]=(f'=ROUND(SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$B$2:$B${LR},"<>")/1000000'
                  f"-$F${tr},6)")
    ws[f"B{c3}"]="Control - squad + overhead vs total ($m), must be 0"
    ws[f"F{c3}"]=f"=ROUND($H${tr}+$M${tr}-$F${tr},6)"
    for r in (c1,c2,c3): ws[f"B{r}"].font=ITAL

    # overhead allowance block, unchanged in logic, moved below the controls
    ob=c3+2
    ws[f"B{ob}"]=("Overhead - allowance against actual. The 8 GMs are the only line with no "
                  "role in the ledger, so their cost sits above the 525.")
    ws[f"B{ob}"].font=ITAL
    hdr(ws,ob+1,{"B":"Overhead line","C":"Roles","D":"Rate ($m)","E":"Units",
                 "F":"Actual ($m)","G":"Allowance ($m)","H":"Variance ($m)"})
    for i in range(6):
        r=ob+2+i; lr=2+i
        ws[f"B{r}"]=f"=Lists!$AF${lr}"; ws[f"D{r}"]=f"=Lists!$AG${lr}"
        ws[f"E{r}"]=f"=Lists!$AH${lr}"; ws[f"G{r}"]=f"=Lists!$AJ${lr}"
        if i==5:
            ws[f"C{r}"]="=Lists!$AG$11"; ws[f"F{r}"]="=Lists!$AG$12"
        else:
            ws[f"C{r}"]=f'=COUNTIFS({REV}!$AR$2:$AR${LR},$B{r})'
            ws[f"F{r}"]=f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AR$2:$AR${LR},$B{r})/1000000'
        ws[f"H{r}"]=f"=$F{r}-$G{r}"
    ot=ob+8
    ws[f"B{ot}"]="Overhead total"; ws[f"B{ot}"].font=BOLD
    for c in "CFGH": ws[f"{c}{ot}"]=f"=SUM({c}{ob+2}:{c}{ot-1})"
    ws[f"B{ot+1}"]="Total organisation including the GM layer ($m)"; ws[f"B{ot+1}"].font=BOLD
    ws[f"F{ot+1}"]=f"=$F${tr}+Lists!$AG$12"
    g32=dict(r0=r0,sub=sub,tr=tr,ot=ot,org=ot+1)

    # ---------------- 3.3 ----------------
    ws=wb["3.3 FTE View"]
    for r in range(3,ws.max_row+1):
        for c in range(1,17):
            if ws.cell(r,c).value is not None: ws.cell(r,c).value=None
    ws["B4"]="One FTE statement. Every squad and every overhead line, counts only."
    ws["B4"].font=ITAL
    hdr(ws,5,{"B":"Portfolio","C":"Squad or overhead line","D":"Archetype type",
              "E":"Archetype size","F":"Archetype roles","G":"Filled","H":"Vacant",
              "I":"Total roles","J":"Variance to archetype","K":"Roles after decisions",
              "L":"Vacancy %"})
    r=6; starts={}
    for pf in order:
        t=tab_of[pf]; g=geom[t]; starts[pf]=r
        for i in range(len(g['squads'])+len(g['ohs'])):
            src = g['a0']+i if i<len(g['squads']) else g['b0']+i-len(g['squads'])
            ws[f"B{r}"]=pf
            for col,s in (("C","B"),("D","C"),("E","R"),("F","D"),("K","J")):
                ws[f"{col}{r}"]=f"='{t}'!${s}${src}"
            ws[f"G{r}"]=f"='{t}'!$E${src}-'{t}'!$G${src}"
            ws[f"H{r}"]=f"='{t}'!$G${src}"
            ws[f"I{r}"]=f"='{t}'!$E${src}"
            ws[f"J{r}"]=f'=IFERROR($I{r}-$F{r},"-")'
            ws[f"L{r}"]=f'=IFERROR($H{r}/$I{r},"-")'
            r+=1
        ws[f"B{r}"]=pf; ws[f"C{r}"]="Portfolio total"; ws[f"C{r}"].font=BOLD
        for c in "GHIK": ws[f"{c}{r}"]=f"=SUM({c}{starts[pf]}:{c}{r-1})"
        ws[f"F{r}"]=f'=IF(COUNT(F{starts[pf]}:F{r-1})=0,"-",SUM(F{starts[pf]}:F{r-1}))'
        ws[f"J{r}"]=f'=IF(COUNT(F{starts[pf]}:F{r-1})=0,"-",SUM(J{starts[pf]}:J{r-1}))'
        ws[f"L{r}"]=f'=IFERROR($H{r}/$I{r},"-")'
        starts[pf+"_t"]=r; r+=1
    gt=r
    ws[f"C{gt}"]="Group total"; ws[f"C{gt}"].font=BOLD
    for c in "GHIK": ws[f"{c}{gt}"]="="+"+".join(f"${c}${starts[p+'_t']}" for p in order)
    for c in ("F","J"):
        ws[f"{c}{gt}"]=f'=SUMIF($C$6:$C${gt-1},"<>Portfolio total",{c}6:{c}{gt-1})'
    ws[f"L{gt}"]=f'=IFERROR($H{gt}/$I{gt},"-")'
    ws[f"B{gt+2}"]="Control - roles vs the ledger, must be 0"; ws[f"B{gt+2}"].font=ITAL
    ws[f"I{gt+2}"]=f"=COUNTA({REV}!$B$2:$B${LR})-$I${gt}"

    # ---------------- 3.1 ----------------
    ws=wb["3.1 Group Summary"]
    for r in range(4,ws.max_row+1):
        for c in range(1,15):
            if ws.cell(r,c).value is not None: ws.cell(r,c).value=None
    ws["B4"]="Group view. Every figure is pulled from 3.2 - nothing is restated here."
    ws["B4"].font=ITAL
    hdr(ws,5,{"B":"Portfolio","C":"Budget ($m)","D":"Actual cost ($m)",
              "E":"Variance to budget ($m)","F":"Squad cost ($m)",
              "G":"Archetype cost ($m)","H":"Variance to archetype ($m)",
              "I":"Overhead cost ($m)","J":"Roles","K":"Vacant",
              "L":"Cost after decisions ($m)"})
    for i,pf in enumerate(order):
        r=6+i; s=g32["r0"]+i+(1 if i>=n else 0)
        ws[f"B{r}"]=f"='3.2 Total Cost'!$B${s}"
        ws[f"C{r}"]=(f"=SUMPRODUCT(('0.2 Data Config'!$B$6:$B$25<>\"\")*"
                     f"(IFERROR(INDEX(Lists!$AL$2:$AL$19,MATCH('0.2 Data Config'!$B$6:$B$25,"
                     f"Lists!$AK$2:$AK$19,0)),\"\")=$B{r})*'0.2 Data Config'!$E$6:$E$25)")
        for c,sc in (("D","F"),("F","H"),("G","J"),("H","K"),("I","M"),("J","C"),
                     ("K","E"),("L","N")):
            ws[f"{c}{r}"]=f"='3.2 Total Cost'!${sc}${s}"
        ws[f"E{r}"]=f"=$C{r}-$D{r}"
    t1=6+len(order)
    ws[f"B{t1}"]="Total"; ws[f"B{t1}"].font=BOLD
    for c in "CDEFIJKL": ws[f"{c}{t1}"]=f"=SUM({c}6:{c}{t1-1})"
    for c in "GH": ws[f"{c}{t1}"]=f"='3.2 Total Cost'!${'J' if c=='G' else 'K'}${g32['tr']}"
    wb.save(dst)
    return geom,g32

if __name__=="__main__":
    g,g32=run("v4p.xlsx","s2.xlsx"); print("3.2 tr row",g32["tr"],"org",g32["org"])
