"""One pass over the shipped workbook. Everything mechanical, nothing subjective.

Recomputes every figure from the ledger, ties every cross-tab statement of the same fact,
and applies the model-review checks a finance reviewer would run by hand: column-formula
consistency, constants buried in formulas, inputs that are not marked as inputs, rounding
inside a chain that is later summed, controls that cannot fail, whole-column references.

Run: python3 audit.py cand_A.xlsx
Exit 0 only when nothing is outstanding. Findings print grouped, worst first.
"""
import re
import sys
from collections import defaultdict

import openpyxl

REVIEW = "REVIEW - Complete Role Mapping"
CREAM = "FFFFF2CC"
OUT = defaultdict(list)


def add(kind, where, msg):
    OUT[kind].append((where, msg))


def fill_of(c):
    try:
        return str(c.fill.start_color.rgb or "").upper() if c.fill.patternType else ""
    except Exception:                                       # noqa: BLE001
        return ""


# ---------------------------------------------------------------- the ledger, recomputed
def ledger(wv):
    """Every person, from the calculated ledger, with the fields the model joins on."""
    R = wv[REVIEW]
    people = []
    for r in range(2, R.max_row + 1):
        name = str(R.cell(r, 2).value or "").strip()
        if not name:
            continue
        people.append({
            "row": r, "name": name, "role": str(R.cell(r, 3).value or "").strip(),
            "fte": R.cell(r, 15).value, "cost": R.cell(r, 27).value,
            "tab": str(R.cell(r, 36).value or "").strip(),
            "status": str(R.cell(r, 37).value or "").strip(),
            "ohline": str(R.cell(r, 44).value or "").strip(),
            "group": str(R.cell(r, 46).value or "").strip(),
        })
    return people


def num(x):
    return x if isinstance(x, (int, float)) else 0.0


def close(a, b, tol=1e-6):
    return abs(num(a) - num(b)) <= tol


# ---------------------------------------------------------------- 1. figures vs the ledger
def check_figures(wb, wv, people):
    tot_cost = sum(num(p["cost"]) for p in people)
    tot_fte = sum(num(p["fte"]) for p in people)
    filled = sum(1 for p in people if p["status"] == "Filled")
    vacant = sum(1 for p in people if p["status"] == "Vacant")
    add("ANCHORS", "ledger", f"{len(people)} roles, {tot_fte:.2f} FTE, "
                             f"${tot_cost:,.2f}, {filled} filled, {vacant} vacant")

    # every working tab's total against the ledger slice it claims
    bypf = defaultdict(list)
    for p in people:
        bypf[p["tab"]].append(p)
    for t in [s for s in wb.sheetnames if re.match(r"^2\.\d+ ", s)]:
        ws, vs = wb[t], wv[t]
        pf = str(vs["C3"].value or "").strip()
        mine = bypf.get(pf, [])
        tr = next((r for r in range(1, ws.max_row + 1)
                   if str(ws.cell(r, 2).value or "").strip() == "Total portfolio"), None)
        if tr is None:
            add("TIE", t, "no 'Total portfolio' row")
            continue
        want_n, want_c = len(mine), sum(num(p["cost"]) for p in mine) / 1e6
        want_f = sum(num(p["fte"]) for p in mine)
        got_n, got_c, got_f = (vs.cell(tr, 6).value, vs.cell(tr, 15).value,
                               vs.cell(tr, 7).value)
        if not close(got_n, want_n, 0.5):
            add("TIE", f"{t}!F{tr}", f"roles {got_n} vs ledger {want_n}")
        if not close(got_c, want_c, 1e-6):
            add("TIE", f"{t}!O{tr}", f"cost {got_c} vs ledger {want_c:.9f}")
        if not close(got_f, want_f, 1e-6):
            add("TIE", f"{t}!G{tr}", f"FTE {got_f} vs ledger {want_f}")

    # every person on exactly one working tab, joined the insert-safe way
    ref = re.compile(r"'" + re.escape(REVIEW) + r"'!\$[A-Z]{1,2}\$(\d+)")
    seen = defaultdict(list)
    for t in [s for s in wb.sheetnames if re.match(r"^2\.\d+ ", s)]:
        ws = wb[t]
        for r in range(1, ws.max_row + 1):
            m = ref.search(str(ws.cell(r, 2).value or ""))
            if m:
                seen[int(m.group(1))].append(f"{t}!B{r}")
    rows = {p["row"] for p in people}
    missing = sorted(rows - set(seen))
    dupes = {k: v for k, v in seen.items() if len(v) > 1}
    extra = sorted(set(seen) - rows)
    if missing:
        add("COVERAGE", "2.x", f"{len(missing)} ledger rows on no working tab: {missing[:8]}")
    if dupes:
        add("COVERAGE", "2.x", f"{len(dupes)} people on more than one tab: "
                               f"{list(dupes.items())[:4]}")
    if extra:
        add("COVERAGE", "2.x", f"{len(extra)} rows referenced that carry no person: {extra[:8]}")
    if not (missing or dupes or extra):
        add("ANCHORS", "2.x", f"all {len(rows)} people appear exactly once across the "
                              f"working tabs")

    # placement: the tab a person sits on must equal the tab the ledger assigns
    wrong = []
    for t in [s for s in wb.sheetnames if re.match(r"^2\.\d+ ", s)]:
        pf = str(wv[t]["C3"].value or "").strip()
        ws = wb[t]
        for r in range(1, ws.max_row + 1):
            m = ref.search(str(ws.cell(r, 2).value or ""))
            if m:
                p = next((x for x in people if x["row"] == int(m.group(1))), None)
                if p and p["tab"] != pf:
                    wrong.append(f"{t}!B{r} {p['name']} belongs to {p['tab']}")
    if wrong:
        add("COVERAGE", "2.x", f"{len(wrong)} people on the wrong tab: {wrong[:5]}")
    else:
        add("ANCHORS", "2.x", "every person sits on the tab the ledger assigns")


# ---------------------------------------------------------------- 2. controls
def check_controls(wb, wv):
    n = bad = 0
    for t in wb.sheetnames:
        ws, vs = wb[t], wv[t]
        for r in range(1, min(ws.max_row, 200) + 1):
            lab = str(ws.cell(r, 2).value or "").strip()
            if not lab.lower().startswith("control"):
                continue
            hit = None
            for c in range(3, 20):
                if isinstance(vs.cell(r, c).value, (int, float)):
                    hit = c
                    break
            if hit is None:
                add("CONTROL", f"{t}!B{r}", "control row states no figure")
                continue
            n += 1
            if abs(vs.cell(r, hit).value) > 1e-6:
                bad += 1
                add("CONTROL", f"{t}!{ws.cell(r, hit).coordinate}",
                    f"reads {vs.cell(r, hit).value}, must be 0")
    add("ANCHORS", "controls", f"{n} control cells, {bad} not zero")


def check_tautologies(wb, wv):
    """A check whose two sides come from the same cell can never fail."""
    ws, vs = wb["4.0 Data QA"], wv["4.0 Data QA"]
    cell = re.compile(r"'([^']+)'!\$?([A-Z]{1,2})\$?(\d+)")
    n = 0
    for r in range(1, ws.max_row + 1):
        lab = str(ws.cell(r, 2).value or "").strip()
        c_, d_ = str(ws.cell(r, 3).value or ""), str(ws.cell(r, 4).value or "")
        if not (c_.startswith("=") and d_.startswith("=")):
            continue
        n += 1
        left, right = set(cell.findall(c_)), set(cell.findall(d_))
        if left and left == right:
            add("CONTROL", f"4.0!E{r}", f"both sides read the same cells - cannot fail "
                                        f"({lab[:60]})")
        if d_.strip() in ("=0", "0"):
            add("CONTROL", f"4.0!D{r}", f"expected side is a literal zero ({lab[:60]})")
    add("ANCHORS", "4.0", f"{n} checks with two computed sides")


# ---------------------------------------------------------------- 3. modelling standards
LITERAL = re.compile(r"(?<![A-Z0-9$!:.])(\d+\.?\d*)(?![0-9$])")
OK_CONST = {"0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "1000000", "100",
            "0.5", "2.0"}


def normalise(f):
    """A formula with its row numbers removed, so two rows of one block compare equal."""
    f = re.sub(r"(?<=[A-Z$])\d+", "#", f)
    return re.sub(r",\s*\d+\s*\)", ",#)", f)


def check_standards(wb, wv):
    hard = incons = 0
    for t in wb.sheetnames:
        ws = wb[t]
        if t in ("0.1 Budget Table (Fin)", "0.4 Presentation Pack", "0.3 Squad Archetypes"):
            continue                                        # owner source tabs
        # column consistency: inside a contiguous run of formula cells, one cell whose
        # shape differs from a clear majority is an outlier worth reporting. A shape that
        # simply changes and stays changed is a new block - a total under its members,
        # a different section - and is not a defect.
        for c in range(2, min(ws.max_column, 26) + 1):
            run = []
            for r in range(1, min(ws.max_row, 200) + 2):
                v = ws.cell(r, c).value if r <= ws.max_row else None
                f = v if isinstance(v, str) and v.startswith("=") else None
                if f:
                    run.append((r, normalise(f)))
                    continue
                if len(run) >= 4:
                    # a subtotal legitimately differs from the rows it sums, and a
                    # control legitimately differs from everything. Judge only the
                    # data rows against each other.
                    def _is_data(rr):
                        lab = str(ws.cell(rr, 2).value or "").strip().lower()
                        return not (lab.endswith("total") or lab.startswith("control")
                                    or lab.startswith("of which") or lab.startswith("all "))
                    run = [(rr, sh) for rr, sh in run if _is_data(rr)]
                    tally = defaultdict(list)
                    for rr, sh in run:
                        tally[sh].append(rr)
                    if len(tally) > 1:
                        big = max(tally.values(), key=len)
                        if len(big) >= len(run) - 1 and len(big) >= 3:
                            for sh, rows_ in tally.items():
                                if rows_ is big:
                                    continue
                                for rr in rows_:
                                    incons += 1
                                    add("CONSISTENCY", f"{t}!{ws.cell(rr, c).coordinate}",
                                        f"one row in a block of {len(big)} identical "
                                        f"formulas does not match them")
                run = []
        # constants buried in a formula
        for row in ws.iter_rows(max_row=min(ws.max_row, 200)):
            for c in row:
                f = c.value
                if not (isinstance(f, str) and f.startswith("=")):
                    continue
                body = re.sub(r"'[^']*'", "", f)          # sheet names first
                body = re.sub(r'"[^"]*"', "", body)         # then text arguments
                body = re.sub(r"\$?[A-Z]{1,2}\$?\d+", "", body)
                body = re.sub(r"\b[A-Z]{1,2}:\$?[A-Z]{1,2}\b", "", body)
                body = re.sub(r"ROUND\([^,]*,\s*\d+\)", "", body)
                for lit in LITERAL.findall(body):
                    if lit in OK_CONST or lit.startswith("0.00"):
                        continue
                    hard += 1
                    add("HARDCODE", f"{t}!{c.coordinate}",
                        f"constant {lit} inside a formula: {f[:70]}")
                    break
    add("ANCHORS", "standards", f"{incons} column-shape breaks, {hard} buried constants")


def check_inputs(wb):
    """Cream means typed. A cream formula, or a typed number that is not cream, misleads."""
    cf = tn = 0
    for t in wb.sheetnames:
        if t in ("0.1 Budget Table (Fin)", "0.4 Presentation Pack", "0.3 Squad Archetypes",
                 "Lists", REVIEW):
            continue
        ws = wb[t]
        for row in ws.iter_rows(max_row=min(ws.max_row, 200)):
            for c in row:
                if c.value is None:
                    continue
                cream = fill_of(c) == CREAM
                isf = isinstance(c.value, str) and c.value.startswith("=")
                if cream and isf:
                    cf += 1
                    add("INPUTS", f"{t}!{c.coordinate}",
                        f"marked as a typed input but holds a formula: {str(c.value)[:60]}")
                elif (not cream) and isinstance(c.value, (int, float)) and c.value not in (0, 1):
                    tn += 1
                    add("INPUTS", f"{t}!{c.coordinate}",
                        f"typed number {c.value} not marked as an input")
    add("ANCHORS", "inputs", f"{cf} cream formulas, {tn} unmarked typed numbers")


def check_rounding(wb):
    """ROUND inside a chain that is then summed compounds the error it hides."""
    n = 0
    for t in wb.sheetnames:
        ws = wb[t]
        rounded = set()
        for row in ws.iter_rows(max_row=min(ws.max_row, 200)):
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("=ROUND("):
                    rounded.add(c.coordinate)
        for row in ws.iter_rows(max_row=min(ws.max_row, 200)):
            for c in row:
                f = c.value
                if not (isinstance(f, str) and f.startswith("=SUM(")):
                    continue
                m = re.match(r"=SUM\(\$?([A-Z]{1,2})\$?(\d+):\$?([A-Z]{1,2})\$?(\d+)\)", f)
                if not m:
                    continue
                col, lo, hi = m.group(1), int(m.group(2)), int(m.group(4))
                hits = [f"{col}{r}" for r in range(lo, hi + 1) if f"{col}{r}" in rounded]
                if len(hits) >= 3:
                    n += 1
                    add("ROUNDING", f"{t}!{c.coordinate}",
                        f"sums {len(hits)} already-rounded cells ({col}{lo}:{col}{hi})")
    add("ANCHORS", "rounding", f"{n} sums over rounded chains")


def check_levers_agree(wv):
    """The COE tabs carry the lever twice, and the two entries must say the same thing.

    1.11/1.12/1.13 hold an On/Off column the owner types (Onshore / Offshore / Hold);
    the model prices off the Vacancy lever column on 2.12/2.13/2.11. Neither references
    the other, so a lever flipped on the tab in front of a GM would not move the bridge.
    This is the divergence alarm: per person, "Offshore" must meet Offshore, "Hold" must
    meet Hold, and "Onshore" must meet Filled or Hire.
    """
    # the working tab by its number, not by its name: 2.11 is renamed in the chain, and a
    # hardcoded name here would skip the tab in silence - which is exactly the divergence
    # this alarm exists to catch
    def _by_num(num):
        return next((t for t in wv.sheetnames if t.split(" ", 1)[0] == num), num)

    PAIRS = [("1.11 BP&T", _by_num("2.12")), ("1.12 SA&D", _by_num("2.13")),
             ("1.13 Cyber Roles", _by_num("2.11"))]
    OK = {"Offshore": {"Offshore"}, "Hold": {"Hold"}, "Onshore": {"Filled", "Hire"}}
    for one, two in PAIRS:
        if one not in wv.sheetnames or two not in wv.sheetnames:
            continue
        d, w = wv[one], wv[two]
        # the design tab's roles table: Name in B under a "Name" header; the On/Off
        # column is the one headed On/Off on the same row
        hdr = onoff = None
        for r in range(1, min(d.max_row, 40) + 1):
            if str(d.cell(r, 2).value or "").strip() == "Name":
                hdr = r
                for c in range(3, 24):
                    if str(d.cell(r, c).value or "").strip() == "On/Off":
                        onoff = c
                break
        if hdr is None or onoff is None:
            continue
        design = {}
        for r in range(hdr + 1, d.max_row + 1):
            nm, lv = d.cell(r, 2).value, d.cell(r, onoff).value
            if isinstance(nm, str) and nm.strip() and isinstance(lv, str):
                design.setdefault(f"{nm.strip()}|{d.cell(r, 3).value}", lv.strip())
        whdr = wlev = None
        for r in range(1, min(w.max_row, 60) + 1):
            if str(w.cell(r, 2).value or "").strip() == "Name":
                whdr = r
                for c in range(3, 12):
                    if str(w.cell(r, c).value or "").strip() == "Vacancy lever":
                        wlev = c
                break
        if whdr is None or wlev is None:
            continue
        for r in range(whdr + 1, w.max_row + 1):
            nm, lv = w.cell(r, 2).value, w.cell(r, wlev).value
            if not (isinstance(nm, str) and nm.strip() and isinstance(lv, str)):
                continue
            key = f"{nm.strip()}|{w.cell(r, 3).value}"
            his = design.get(key)
            if his and his in OK and lv.strip() not in OK[his]:
                add("LEVER2", f"{one} vs {two}",
                    f"{nm.strip()}: the design tab says {his!r}, the working tab prices "
                    f"{lv.strip()!r} - the model follows the working tab")


def check_refs(wb):
    whole = errs = 0
    for t in wb.sheetnames:
        ws = wb[t]
        for row in ws.iter_rows():
            for c in row:
                f = c.value
                if not isinstance(f, str):
                    continue
                if f.startswith("=") and re.search(r"!\$[A-Z]{1,2}:\$[A-Z]{1,2}", f):
                    whole += 1
                    add("REFS", f"{t}!{c.coordinate}", f"whole-column reference: {f[:70]}")
                if f.strip() in ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?",
                                 "#NULL!", "#NUM!"):
                    errs += 1
                    add("ERROR", f"{t}!{c.coordinate}", f"error value {f}")
    add("ANCHORS", "refs", f"{whole} whole-column references, {errs} error cells")


# ---------------------------------------------------------------- 4. labels, for reading
def dump_labels(wb, wv, path="labels.txt"):
    n = 0
    with open(path, "w") as fh:
        for t in wb.sheetnames:
            ws, vs = wb[t], wv[t]
            fh.write(f"\n===== {t}\n")
            for r in range(1, min(ws.max_row, 200) + 1):
                for c in range(1, min(ws.max_column, 26) + 1):
                    v = vs.cell(r, c).value
                    if isinstance(v, str) and v.strip() and not v.startswith("="):
                        fh.write(f"  {ws.cell(r, c).coordinate}: {v.strip()}\n")
                        n += 1
    add("ANCHORS", "labels", f"{n} labels written to {path} for the plain-English read")


def run(path):
    wb = openpyxl.load_workbook(path)
    wv = openpyxl.load_workbook(path, data_only=True)
    people = ledger(wv)
    check_figures(wb, wv, people)
    check_controls(wb, wv)
    check_tautologies(wb, wv)
    check_standards(wb, wv)
    check_inputs(wb)
    check_rounding(wb)
    check_refs(wb)
    check_levers_agree(wv)
    dump_labels(wb, wv)

    order = ["ANCHORS", "ERROR", "TIE", "COVERAGE", "CONTROL", "CONSISTENCY", "HARDCODE",
             "INPUTS", "ROUNDING", "REFS", "LEVER2"]
    bad = 0
    for k in order:
        rows = OUT.get(k, [])
        if not rows:
            continue
        print(f"\n{'=' * 78}\n{k}: {len(rows)}")
        for where, msg in rows[:25]:
            print(f"    {where}: {msg}")
        if len(rows) > 25:
            print(f"    ... {len(rows) - 25} more")
        if k not in ("ANCHORS",):
            bad += len(rows)
    print(f"\n{'=' * 78}\nOUTSTANDING: {bad}")
    return 0 if bad == 0 else 1


if __name__ == "__main__":
    sys.exit(run(sys.argv[1] if len(sys.argv) > 1 else "cand_A.xlsx"))
