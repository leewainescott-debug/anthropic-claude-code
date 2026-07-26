"""Workbook-wide presentation sweep. Nothing here changes a number.

Everything in this file is something a reader sees before they read a figure, and every one
of them was raised by a reviewer looking at the rendered tabs rather than at the cells:

  the tab strip contradicted the numbering - 1.11 to 1.14 sat after the "- SUMMARIES -"
  divider, filed under Summaries, and 3.5 sat inside the working-tab group with no colour

  1.14 TDD Cyber was a dead tab: a copy of 1.9 that reported $1.2925m for Cyber against the
  $9.898m on 1.13, three black bars from a theme fill, twenty styled rows with no values,
  and a 276-character build changelog in its title bar. No formula anywhere referenced it

  239 cells carried a [Red] negative format, so two tabs bled red on screen and a portfolio
  UNDER its design showed the alarm colour

  204 cells on the design tabs carried hardcoded green or blue font as judgement colouring

  three tabs carried bright cyan note blocks with white text, front of house

  Lists was visible; the tabs with no colour had none because nothing had ever set them
"""
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter as L

import opts

DEAD = ["1.14 TDD Cyber"]
RENAME = {"Portfolios": "- PORTFOLIOS -", "ACTUAL WORKBOOKS": "- WORKING -"}
# tabs whose name contradicted their own title. 3.3 is headed "Squad Detail", 3.4 "COE
# detail", and 2.11 carries "COE Cyber" in its title and portfolio cell while the tab said
# TDD Cyber - the ledger calls it COE Cyber everywhere else.
RETITLE = {"3.3 FTE View": "3.3 Squad Detail", "3.4 COE Summary": "3.4 COE Detail",
           "2.11 TDD Cyber": "2.11 COE Cyber"}
ORDER = ["Exec Summary", "- INPUTS -", "0.1 Budget Table (Fin)", "0.2 Data Config",
         "0.3 Squad Archetypes", "0.4 Presentation Pack",
         "REVIEW - Complete Role Mapping", "- PORTFOLIOS -",
         "1.1 Ampol Retail", "1.2 Customer", "1.3 Enterprise Data",
         "1.4 TDD Group Functions", "1.5 P&C", "1.6 Finance", "1.7 Infrastructure",
         "1.8 Energy Solutions & B2B", "1.9 Commercial Fuels", "1.10 Z Retail",
         "1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles", "- WORKING -",
         "2.1 Ampol Retail", "2.2 Customer", "2.3 Enterprise Data",
         "2.4 TDD Group Functions", "2.5 P&C", "2.6 Finance", "2.7 Infrastructure",
         "2.8 Energy Solutions & B2B", "2.9 Commercial Fuels", "2.10 Z Retail",
         "2.11 COE Cyber", "2.12 BP&T", "2.13 SA&D", "2.14 EGI", "- SUMMARIES -",
         "3.1 Group Summary", "3.2 Total Cost", "3.3 Squad Detail", "3.4 COE Detail",
         "- EVIDENCE -", "4.0 Data QA"]
GREY, DESIGN, WORK, SUMM, EVID = ("FF808080", "FF1F4E79", "FFBF8F00", "FF002F6C",
                                  "FF375623")
HIDE = ["Lists"]
# the owner's own source tabs. Their palette is whatever Finance and the deck arrived in,
# and repainting evidence is not tidying.
SOURCE = {"0.1 Budget Table (Fin)", "0.4 Presentation Pack"}
RETIRED = set()          # the retired sources are deleted, not skipped
DIVIDER = re.compile(r"^- .* -$")

JUDGE = {"FFFF0000", "FFC00000", "FF008000", "FF00B050", "FF0000FF", "FF00B0F0",
         "FF92D050", "FFFFC000"}
CYAN = "FF00B0F0"
JUNK = [("1.1 Ampol Retail", "K52"), ("1.9 Commercial Fuels", "S13")]


def live(wb):
    return [ws for ws in wb.worksheets
            if ws.title not in RETIRED | SOURCE and not DIVIDER.match(ws.title)]


def drop_dead(wb):
    out = []
    for name in DEAD:
        if name not in wb.sheetnames:
            continue
        hits = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
                   if isinstance(c.value, str) and name in c.value)
        if hits:
            wb[name].sheet_state = "hidden"
            out.append(f"{name}: {hits} references, hidden rather than removed")
        else:
            del wb[name]
            out.append(f"{name}: removed, nothing referenced it")
    return out


def retitle(wb):
    """Rename a tab and repoint everything that named it - formulas and labels both.

    A reference is rewritten before the sheet is renamed, because openpyxl does not follow
    a rename into formula text: the tab would be called one thing and every formula would
    still point at the other, which is a #REF! on open.
    """
    out = []
    for old, new in RETITLE.items():
        if old not in wb.sheetnames or new in wb.sheetnames:
            continue
        n = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if not isinstance(v, str) or old not in v:
                        continue
                    c.value = (v.replace(f"'{old}'", f"'{new}'").replace(old, new)
                               if v.startswith("=") else v.replace(old, new))
                    n += 1
        wb[old].title = new
        out.append(f"{old} -> {new}, {n} references and labels repointed")
    return out


def order_and_colour(wb):
    for old, new in RENAME.items():
        if old in wb.sheetnames and new not in wb.sheetnames:
            wb[old].title = new
    want = [n for n in ORDER if n in wb.sheetnames]
    rest = [n for n in wb.sheetnames if n not in want]
    wb._sheets = [wb[n] for n in want + rest]
    band, n = None, 0
    for name in want:
        ws = wb[name]
        if DIVIDER.match(name):
            band = {"- INPUTS -": GREY, "- PORTFOLIOS -": DESIGN, "- WORKING -": WORK,
                    "- SUMMARIES -": SUMM, "- EVIDENCE -": EVID}.get(name, band)
        ws.sheet_properties.tabColor = SUMM if name == "Exec Summary" else (band or GREY)
        n += 1
    for name in HIDE:
        if name in wb.sheetnames:
            wb[name].sheet_state = "hidden"
    return [f"{n} tabs ordered to match the numbering and coloured by group",
            f"hidden: {', '.join(HIDE)}"]


def no_red_formats(wb):
    """A [Red] negative format put the alarm colour on a portfolio under its design."""
    n = 0
    for ws in wb.worksheets:
        if ws.title in RETIRED:
            continue
        for row in ws.iter_rows():
            for c in row:
                nf = c.number_format or ""
                if "[Red]" not in nf:
                    continue
                c.number_format = (opts.M2 if ".00" in nf else
                                   (opts.M3 if ".000" in nf else opts.CT))
                n += 1
    return [f"[Red] stripped from {n} number formats"]


def no_judgement_colour(wb):
    """Hardcoded green, blue and red font, and the cyan note blocks."""
    nf = nb = 0
    for ws in live(wb):
        for row in ws.iter_rows():
            for c in row:
                f = c.font
                rgb = str(getattr(f.color, "rgb", "") or "") if f and f.color else ""
                if rgb.upper() in JUDGE:
                    c.font = Font(name=opts.FN, size=max(f.size or 11, 10),
                                  bold=f.bold, italic=False)
                    nf += 1
                elif (f.size or 11) < 10:
                    c.font = Font(name=opts.FN, size=10, bold=f.bold,
                                  color=rgb or None)
                    nf += 1
                fl = c.fill
                try:
                    fill = str(fl.start_color.rgb or "").upper() \
                        if fl and fl.patternType else ""
                except Exception:
                    fill = ""
                if fill == CYAN:
                    c.fill = opts.fl(opts.GREY)
                    c.font = opts.BOLD
                    nb += 1
    return [f"{nf} fonts set to plain black Calibri, nothing under 10pt",
            f"{nb} cyan note-block cells restyled to the standard grey"]


def strays(wb):
    out = []
    for tab, cell in JUNK:
        if tab in wb.sheetnames and wb[tab][cell].value is not None:
            wb[tab][cell].value = None
            out.append(f"{tab}!{cell} cleared")
    # 0.2 carried a row keyed 'EG', a broken duplicate of the EGI row directly below it
    ws = wb["0.2 Data Config"]
    for r in range(1, 40):
        if str(ws.cell(r, 2).value or "").strip() == "EG":
            for c in range(2, 14):
                ws.cell(r, c).value = None
            out.append(f"0.2 Data Config row {r} cleared - a broken duplicate of EGI")
    return out


def en_dash(wb):
    """Register item 85. They reach 1.13 through formulas reading the retired Sheet2."""
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "–" in c.value:
                    c.value = c.value.replace("–", "-")
                    n += 1
    return [f"{n} en dashes replaced with a hyphen"]


def gutters_and_grid(wb):
    n = 0
    for ws in live(wb):
        ws.sheet_view.showGridLines = False
        if ws.title != "REVIEW - Complete Role Mapping":
            ws.column_dimensions["A"].width = 2
        n += 1
    return [f"gridlines off and a 2-wide gutter on {n} tabs"]


def freeze(wb):
    """Every long table gets a frozen header. It was on nineteen tabs and off on the
    twenty with the longest lists, including the 528-row ledger."""
    n = 0
    for ws in live(wb):
        if ws.max_row < 25 or ws.freeze_panes:
            continue
        hdr = None
        for r in range(1, 12):
            filled = sum(1 for c in range(2, 16)
                         if isinstance(ws.cell(r, c).value, str)
                         and ws.cell(r, c).value.strip())
            if filled >= 3:
                hdr = r
                break
        if hdr:
            ws.freeze_panes = f"C{hdr + 1}"
            n += 1
    R = wb["REVIEW - Complete Role Mapping"]
    if not R.freeze_panes:
        R.freeze_panes = "C2"
        n += 1
    return [f"frozen header rows added on {n} tabs"]


SENTENCE = 90


def build_notes(wb):
    """Build commentary in a data area. A note that explains what a cell used to read
    belongs in the decisions log, not on an executive's tab."""
    n = 0
    for ws in live(wb):
        for r in range(1, min(ws.max_row, 200) + 1):
            v = ws.cell(r, 2).value
            if not isinstance(v, str) or v.startswith("="):
                continue
            s = v.strip()
            if len(s) > SENTENCE and any(
                    k in s for k in ("previously read", "It previously", "used to",
                                     "was a copy", "does not reconcile",
                                     "hand-typed", "typed zero", "pointed at blank")):
                ws.cell(r, 2).value = None
                n += 1
    return [f"{n} build-commentary sentences removed from data areas"]


# The COE tabs and 0.2 were never built by a family builder, so their bars are one cell
# wide over eleven-column tables and their headers are 39 characters in an 11-wide column.
# The ten portfolio design tabs are left alone: they have their own profile, and a generic
# repaint over them once painted seven of the owner's cream inputs navy.
# Two tables share these columns - a summary block whose column C counts roles, and a
# roles list whose column C holds a position title - so a width that suits one clips the
# other. Measured from the content of both, which is the only thing that fits both.
MEASURED = ("1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles", "0.2 Data Config",
            "0.3 Squad Archetypes")
CAP = {"B": 46, "C": 38, "D": 24}


def measure(ws, vs, first=2, last=13, floor=9, cap=20):
    """ws for the fills, vs for the values.

    Measuring the formula workbook measured nothing: every position title on a COE roles
    list is =INDEX(REVIEW...), so a length test on the formula text skipped all of them and
    the column fell back to its floor. The cached values are what the reader sees.
    """
    out = {}
    for c in range(first, last + 1):
        k = L(c)
        need = floor
        for r in range(1, min(ws.max_row, 120) + 1):
            v = vs.cell(r, c).value if vs else ws.cell(r, c).value
            if isinstance(v, str) and not v.startswith("=") and v.strip():
                try:
                    rgb = str(ws.cell(r, c).fill.start_color.rgb or "").upper() \
                        if ws.cell(r, c).fill.patternType else ""
                except Exception:
                    rgb = ""
                # a bar or a header is wrapped, so it does not drive the width; a value in
                # the body of the table is not wrapped and does
                if rgb in (opts.BARC, opts.NAVY):
                    continue
                need = max(need, min(len(v.strip()) + 2, CAP.get(k, cap)))
        out[k] = need
    return out


def table_width(ws, r, limit=30):
    """How many columns the table under a bar occupies, from its header row."""
    best = 0
    for rr in (r + 1, r + 2):
        n = 0
        for c in range(2, limit):
            x = ws.cell(rr, c)
            try:
                rgb = str(x.fill.start_color.rgb or "").upper() \
                    if x.fill and x.fill.patternType else ""
            except Exception:
                rgb = ""
            if rgb != opts.NAVY:
                break
            n = c
        best = max(best, n)
    return best


def lone_headers(wb):
    """A navy label that is not part of a header row - a single input caption such as
    0.3's "Offshore rate" - still has to fit the row it sits in."""
    n = 0
    for ws in live(wb):
        for row in ws.iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and c.value.strip()):
                    continue
                try:
                    rgb = str(c.fill.start_color.rgb or "").upper() \
                        if c.fill and c.fill.patternType else ""
                except Exception:
                    rgb = ""
                if rgb != opts.NAVY:
                    continue
                w = ws.column_dimensions[L(c.column)].width or 8.43
                lines = opts.wrap_lines(c.value, w)
                have = ws.row_dimensions[c.row].height or 15
                if lines * 14 > have + 1:
                    c.alignment = opts.CEN
                    ws.row_dimensions[c.row].height = max(have, 14 * lines + 6)
                    n += 1
    return [f"{n} header cells given the row height their text needs"]


def widen_bars(wb):
    """A one-cell navy bar over a nine-column table reads as a small navy tag with
    clipped text. It happens because the previous pass only painted cells that already
    carried a fill, and the cell beside a platform bar carries none.

    Only empty cells are painted. A pass that walked the row unconditionally once turned
    seven of the owner's cream input cells navy.
    """
    n = 0
    for ws in live(wb):
        for r in range(1, min(ws.max_row, 200) + 1):
            b = ws.cell(r, 2)
            try:
                rgb = str(b.fill.start_color.rgb or "").upper() \
                    if b.fill and b.fill.patternType else ""
            except Exception:
                rgb = ""
            if rgb != opts.BARC or not isinstance(b.value, str) or not b.value.strip():
                continue
            w = table_width(ws, r)
            if w < 3:
                continue
            merged = {m.min_col: m.max_col for m in ws.merged_cells.ranges
                      if m.min_row == m.max_row == r}
            for c in range(max(3, merged.get(2, 2) + 1), w + 1):
                x = ws.cell(r, c)
                if x.value is not None:
                    break
                x.fill, x.font = opts.fl(opts.BARC), opts.BARF
            # and nothing beyond it. A fixed profile width painted 1.11 and 1.13 one
            # column past their tables, so the bar over-ran the table it headed.
            for c in range(w + 1, 30):
                x = ws.cell(r, c)
                try:
                    g = str(x.fill.start_color.rgb or "").upper() \
                        if x.fill and x.fill.patternType else ""
                except Exception:
                    g = ""
                if g != opts.BARC or x.value is not None:
                    break
                x.fill = PatternFill()
            ws.row_dimensions[r].height = 19
            n += 1
    return [f"{n} section bars widened to the table under them"]


def bars_and_headers(wb, vals):
    nh = 0
    for tab in MEASURED:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        widths = measure(ws, vals.get(tab))
        for k, v in widths.items():
            ws.column_dimensions[k].width = v
        wide = 2 + len(widths) - 1
        for r in range(1, min(ws.max_row, 120) + 1):
            navy = 0
            for c in range(2, wide + 1):
                x = ws.cell(r, c)
                fl = x.fill
                try:
                    rgb = str(fl.start_color.rgb or "").upper() \
                        if fl and fl.patternType else ""
                except Exception:
                    rgb = ""
                if rgb == opts.NAVY and isinstance(x.value, str) and x.value.strip():
                    navy += 1
            if navy >= 3:                                  # a column-header row
                # sized by the same wrap arithmetic the family builders use: widen the
                # column first, then make the row tall enough for what is left. A fixed
                # 32pt over a three-line header pushes the first line off the top, which
                # is how "Budget to draw down ($m)" rendered as "udget to draw down ($m".
                lines = 1
                for c in range(2, wide + 1):
                    x = ws.cell(r, c)
                    if not (isinstance(x.value, str) and x.value.strip()):
                        continue
                    x.alignment = opts.CEN
                    k = L(c)
                    w = ws.column_dimensions[k].width or 8.43
                    while opts.wrap_lines(x.value, w) > 3 and w < 24:
                        w += 1
                    ws.column_dimensions[k].width = w
                    lines = max(lines, opts.wrap_lines(x.value, w))
                ws.row_dimensions[r].height = max(32, 14 * lines + 6)
                nh += 1
                continue
    return [f"{nh} header rows wrapped and re-widthed on the COE and input tabs"]


# A summary a GM prints must not put its total row on page two.
ONE_PAGE = {"Exec Summary", "3.1 Group Summary", "3.2 Total Cost", "3.4 COE Summary",
            "3.5 Source Reconciliation", "4.0 Data QA", "0.2 Data Config",
            "0.3 Squad Archetypes"}


def print_setup(wb):
    n = 0
    for ws in live(wb):
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1 if ws.title in ONE_PAGE else 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = False
        n += 1
    return [f"print setup on {n} tabs, landscape and fitted to the page width"]


def run(src, dst, ledger=None):
    wb = openpyxl.load_workbook(src)
    vals = {}
    if ledger:
        lw = openpyxl.load_workbook(ledger, data_only=True)
        vals = {t: lw[t] for t in lw.sheetnames if t in wb.sheetnames}
    out = (drop_dead(wb) + retitle(wb) + order_and_colour(wb) + no_red_formats(wb)
           + no_judgement_colour(wb) + strays(wb) + en_dash(wb)
           + gutters_and_grid(wb) + bars_and_headers(wb, vals) + widen_bars(wb)
           + lone_headers(wb) + build_notes(wb)
           + freeze(wb) + print_setup(wb))
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(*sys.argv[1:]):
        print("  ", x)
