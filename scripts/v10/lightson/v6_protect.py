#!/usr/bin/env python3
"""v6 - spec stage v6: protection.

  python3 v6_protect.py <in.xlsx> <out.xlsx>

Every cell on every sheet locked, except the cells a GM is meant to type in:
  lever   the 2.x FTE-block vacancy lever cells (col E under the
          Filled/Hire/Hold/Offshore dropdown)
  uplift  the 2.11 Uplift % cells (col H under the percent dropdown)
  cream   every FFF2CC input cell anywhere in the book
  toggle  the Lights On toggle column I cells (cream carries them too)
REVIEW - Complete Role Mapping stays fully locked, cream or not.

Sheet protection goes on with the password on every sheet - selecting locked
and unlocked cells stays allowed, sort and filter are off - and the workbook
structure is locked with the same password.

Idempotent: handed its own output it copies it through untouched.
"""
import sys, os, shutil, collections

V10 = "/home/user/anthropic-claude-code/scripts/v10"
sys.path.insert(0, V10)
sys.path.insert(0, os.path.join(V10, "update"))

import openpyxl
from openpyxl.styles import Protection

import wbio
from _xl import REVIEW, Log, load, save

try:
    from openpyxl.utils.protection import hash_password
except ImportError:                                   # older layout
    from openpyxl.worksheet.protection import hash_password

PW = "Tdd123"
LEVER_LIST = "Filled,Hire,Hold,Offshore"
CLASSES = ("lever", "uplift", "cream", "toggle")

# blocked (True) and allowed (False) actions under sheet protection
OPTIONS = dict(sheet=True, objects=True, scenarios=True,
               formatCells=True, formatColumns=True, formatRows=True,
               insertColumns=True, insertRows=True, insertHyperlinks=True,
               deleteColumns=True, deleteRows=True,
               sort=True, autoFilter=True, pivotTables=True,
               selectLockedCells=False, selectUnlockedCells=False)


def is_cream(cell):
    f = cell.fill
    if f is None or f.patternType != "solid":
        return False
    rgb = getattr(f.fgColor, "rgb", None)
    return isinstance(rgb, str) and rgb.endswith("FFF2CC")


def dv_cells(ws, want_col, pick):
    """Coordinates in column want_col covered by a validation pick() accepts."""
    out = set()
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list" or not pick(str(dv.formula1)):
            continue
        for cr in dv.sqref.ranges:
            for r in range(cr.min_row, cr.max_row + 1):
                for c in range(cr.min_col, cr.max_col + 1):
                    if c == want_col:
                        out.add(ws.cell(r, c).coordinate)
    return out


def allowmap(wb):
    """{sheet: {coord: class}} - the only cells the book leaves unlocked."""
    out = {}
    for ws in wb.worksheets:
        allow = {}
        if ws.title != REVIEW:
            if ws.title.startswith("2."):
                for co in dv_cells(ws, 5, lambda f: LEVER_LIST in f):
                    allow[co] = "lever"
                for co in dv_cells(ws, 8, lambda f: f.startswith('"0%')):
                    allow[co] = "uplift"
            if "Lights On" in ws.title:
                for co in dv_cells(ws, 9, lambda f: "%" in f):
                    allow[co] = "toggle"
            for row in ws.iter_rows():
                for c in row:
                    if c.coordinate in allow:
                        continue
                    if is_cream(c):
                        allow[c.coordinate] = ("toggle" if "Lights On" in
                                               ws.title and c.column == 9
                                               else "cream")
        out[ws.title] = allow
    return out


def main(src, dst):
    log = Log("v6_protect")
    wb = load(src)

    if wb.security is not None and wb.security.lockStructure \
            and all(ws.protection.sheet for ws in wb.worksheets):
        print("input is already protected - copying through")
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        return

    allow = allowmap(wb)

    # ------------------------------------------------------- P1  cell locks
    log.head("P1  every cell locked except the input allowlist")
    for ws in wb.worksheets:
        a = allow[ws.title]
        relocked = unlocked = 0
        for row in ws.iter_rows():
            for c in row:
                want = c.coordinate not in a
                if bool(c.protection.locked) != want:
                    c.protection = Protection(locked=want,
                                              hidden=c.protection.hidden)
                    if want:
                        relocked += 1
                    else:
                        unlocked += 1
        by = collections.Counter(a.values())
        log("P1", ws.title,
            "%d cells unlocked (%s), %d relocked"
            % (unlocked, ", ".join("%d %s" % (by[k], k) for k in CLASSES
                                   if by[k]) or "none", relocked))

    # ------------------------------------------------ P2  sheet protection
    log.head("P2  sheet protection on with the password, every sheet")
    for ws in wb.worksheets:
        p = ws.protection
        p.set_password(PW)
        for k, v in OPTIONS.items():
            setattr(p, k, v)
        log("P2", ws.title,
            "protected - select locked/unlocked allowed, sort/filter off")

    # -------------------------------------------- P3  workbook structure
    log.head("P3  workbook structure locked with the same password")
    from openpyxl.workbook.protection import WorkbookProtection
    sec = WorkbookProtection(lockStructure=True)
    sec.workbookPassword = PW
    wb.security = sec
    log("P3", "workbook", "structure locked, sheets cannot move or unhide")

    # ------------------------------------------------------- build and check
    tmp = dst + ".raw"
    save(wb, tmp)
    log.head("recalculating and writing the cached values back")
    rc, st = wbio.build(tmp, dst)
    os.remove(tmp)
    print("recalculated, %d formula cells populated across %d sheets"
          % (st["cells"], st["sheets"]), flush=True)
    err, _ = wbio.audit(dst)
    if err:
        print("STOP: %d error cells, e.g. %r" % (len(err), err[:5]))
        raise SystemExit(2)

    self_check(dst)
    log.tail()
    print("wrote", dst)


# ------------------------------------------------------------- self-check

def self_check(dst):
    print("\n== self-check (inventory recomputed from the file)", flush=True)
    wb = openpyxl.load_workbook(dst)
    checks = []

    def ck(name, ok, detail):
        checks.append(ok)
        print("%s  %-52s %s" % ("PASS" if ok else "FAIL", name, detail),
              flush=True)

    want_hash = hash_password(PW)
    bad = [ws.title for ws in wb.worksheets
           if not (ws.protection.sheet and ws.protection.password == want_hash
                   and ws.protection.sort and ws.protection.autoFilter
                   and not ws.protection.selectLockedCells
                   and not ws.protection.selectUnlockedCells)]
    ck("sheet protection enabled on all %d sheets" % len(wb.worksheets),
       not bad, "password set, sort/filter off%s"
       % ("" if not bad else "; missing on %r" % bad))

    sec = wb.security
    ck("workbook structure locked",
       sec is not None and bool(sec.lockStructure)
       and sec.workbookPassword == want_hash,
       "lockStructure with the same password")

    allow = allowmap(wb)
    mism, table = [], []
    tot = collections.Counter()
    for ws in wb.worksheets:
        a = allow[ws.title]
        unlocked = set()
        for row in ws.iter_rows():
            for c in row:
                if c.protection.locked is False:
                    unlocked.add(c.coordinate)
        if unlocked != set(a):
            mism.append((ws.title, sorted(unlocked - set(a))[:5],
                         sorted(set(a) - unlocked)[:5]))
        by = collections.Counter(a.values())
        tot.update(by)
        if unlocked or a:
            table.append("       %-32s %s  = %d unlocked"
                         % (ws.title,
                            "  ".join("%s %d" % (k, by[k]) for k in CLASSES
                                      if by[k]), len(unlocked)))
    ck("unlocked cells match the allowlist exactly", not mism,
       "%s" % ("classes tie on every sheet" if not mism else
               "mismatches %r" % mism[:3]))
    print("\n".join(table), flush=True)
    print("       %-32s %s  = %d unlocked across the book"
          % ("TOTAL", "  ".join("%s %d" % (k, tot[k]) for k in CLASSES),
             sum(tot.values())), flush=True)

    rvu = [c.coordinate for row in wb[REVIEW].iter_rows() for c in row
           if c.protection.locked is False]
    ck("REVIEW fully locked", not rvu, "%d unlocked cells" % len(rvu))

    err, _ = wbio.audit(dst)
    ck("no error cells after recalc", not err, "%d errors" % len(err))

    wb.close()
    if not all(checks):
        print("\nself-check FAILED")
        raise SystemExit(2)
    print("\nself-check clean: %d/%d PASS" % (sum(checks), len(checks)),
          flush=True)


if __name__ == "__main__":
    main(*sys.argv[1:3])
