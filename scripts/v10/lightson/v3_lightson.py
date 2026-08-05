#!/usr/bin/env python3
"""v3 - Stage v3: the Lights On tab.

  python3 v3_lightson.py <in.xlsx> <out.xlsx>

One new visible tab, '3.5 TDD Lights On', after '3.4 COE Breakdown'. Fifteen
rows in 0.2 Data Config order, his column headings verbatim, everything live:
support cost pairs each squad's after lever cost on the 2.x grid with its
Support %% on the 1.x tab, the BP, Domain Architect and GM shares split the
pots equally across the ten portfolios, other overheads read each tab's
Overhead roles total, and a cream toggle scales how much of those overheads
the lights on budget carries. COE rows net the charged out pots, TDD Cyber
nets the uplift programme funding, EGI carries only its funded cost. Below
the table sits the analysis block, live numbers, and a white font control row.

Idempotent: handed its own output it copies it through untouched.
"""
import sys, os, re, shutil

sys.path.insert(0, "/home/user/anthropic-claude-code/scripts/v10")
sys.path.insert(0, "/home/user/anthropic-claude-code/scripts/v10/update")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import MultiCellRange, CellRange
from openpyxl.utils import get_column_letter

from _xl import REVIEW, Log, load, save, white

TAB = "3.5 TDD Lights On"
ANCHOR = "3.4 COE Breakdown"
CFG = "0.2 Data Config"
OVH = "3.2 Overhead & Leadership"
ARC = "3.1 Archetype to Actuals"

NAVY = "FF0F2E52"
BAND = "FFF5F4F0"
CREAM = "FFFFF2CC"
RED = "FFFF0000"
WHITE = "FFFFFFFF"
M2 = "#,##0.00;(#,##0.00)"
STEPS = ",".join("%d%%" % p for p in range(0, 101, 5))

# his columns, verbatim, one column each
HEADERS = [
    ("B", "Portfolios & COEs & EGI"),
    ("C", "Cost (Total Cost)"),
    ("D", "Support Cost (this is the % in the 1.x tabs)"),
    ("E", "BP allocation"),
    ("F", "Domain architect allocation"),
    ("G", "GM allocation"),
    ("H", "Other overheads"),
    ("I", "Other overheads toggle"),
    ("J", "Amount of overheads charged to TDD"),
    ("K", "Total portfolio cost charged to TDD"),
    ("L", "TDD Lights On budget"),
    ("M", "Over/ Under lights on budget"),
    ("N", None),
    ("O", "Total Cost left to be recharged to business"),
    ("P", "Amount allocated in 1.x tabs"),
]

# row, kind, 2.x tab, 1.x tab, 0.2 budget line labels (row order is 0.2 order)
ROWS = [
    ("COE SA&D", "coe", "2.13 COE SA&D", None,
     ["COE - Strategy Architecture", "COE - Data"]),
    ("COE Cyber Risk & Service Ops", "coe", "2.11 Cyber Risk & Service Ops",
     None, ["COE - Cyber"]),
    ("COE BP&T", "coe", "2.12 COE BP&T", None,
     ["COE - Transformation", "COE - Business Partnering"]),
    ("Ampol Retail", "pf", "2.1 Ampol Retail", "1.1 Ampol Retail",
     ["Ampol Retail"]),
    ("Z Retail", "pf", "2.10 Z Retail", "1.10 Z Retail", ["Z Retail"]),
    ("Ampol & Z Customer", "pf", "2.2 Customer", "1.2 Customer",
     ["Ampol Customer", "Z Customer"]),
    ("Commercial Fuels", "pf", "2.9 Commercial Fuels", "1.9 Commercial Fuels",
     ["Commercial Fuels"]),
    ("Energy Solutions & B2B", "pf", "2.8 Energy Solutions & B2B",
     "1.8 Energy Solutions & B2B", ["Energy Solutions & B2B"]),
    ("Infrastructure", "pf", "2.7 Infrastructure", "1.7 Infrastructure",
     ["Infrastructure"]),
    ("P&C", "pf", "2.5 P&C", "1.5 P&C", ["P&C"]),
    ("Finance", "pf", "2.6 Finance", "1.6 Finance", ["Finance"]),
    ("TDD Group Functions", "pf", "2.4 TDD Group Functions",
     "1.4 TDD Group Functions", ["TDD Group Functions"]),
    ("TDD Data", "pf", "2.3 Enterprise Data", "1.3 Enterprise Data",
     ["TDD Data"]),
    ("TDD Cyber", "cyber", "2.15 TDD Cyber", "1.14 TDD Cyber", ["TDD Cyber"]),
    ("EGI", "egi", "2.14 EGI", None, ["EGI"]),
]

BP_POT_LABEL = "Business Partner pot after levers ($m)"
DA_POT_LABEL = "Domain Architect pot after levers ($m)"
GM_LABEL = "GM cost ($m)"

SECTIONS = {"Squads", "Directly funded programs and platforms",
            "Overhead roles", "No archetype in 1.x tabs"}
SEC_TOTALS = {"Squads total", "Directly funded total", "Overhead roles total",
              "No archetype total"}
SQUAD_SECTIONS = {"Squads", "Directly funded programs and platforms"}


def stop(msg):
    print("STOP: %s" % msg)
    raise SystemExit(2)


def q(name):
    return "'" + name.replace("'", "''") + "'"


def norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


def grid(ws):
    """The 2.x grid: header row, data extent, squad rows, marker rows."""
    hdr = None
    for r in range(4, 12):
        if ws.cell(r, 2).value == "Squad" and \
           ws.cell(r, 19).value == "Squad cost after levers ($m)":
            hdr = r
            break
    if hdr is None:
        stop("no grid header on %s" % ws.title)
    g = {"hdr": hdr, "first": hdr + 1, "total": None, "ovh_total": None,
         "squads": []}
    sec = None
    for r in range(hdr + 1, hdr + 40):
        b = ws.cell(r, 2).value
        if b is None:
            continue
        if b == "Total portfolio":
            g["total"] = r
            break
        if b in SECTIONS:
            sec = b
            continue
        if b in SEC_TOTALS:
            if b == "Overhead roles total":
                g["ovh_total"] = r
            continue
        if isinstance(b, str) and (b.startswith("Portfolio overhead roles") or
                                   b.startswith("Business Partners")):
            continue
        if sec in SQUAD_SECTIONS:
            g["squads"].append((r, str(b).strip(), sec))
    if g["total"] is None:
        stop("no 'Total portfolio' row on %s" % ws.title)
    return g


def support_map(ws):
    """{normalized squad: (row, name)} for every 1.x Support %% cell."""
    out = {}
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 7).value != "Support %":
            continue
        rr = r + 1
        while rr <= ws.max_row:
            b = ws.cell(rr, 2).value
            if b is None or (isinstance(b, str) and b.endswith("Total")):
                break
            if b == "Platform Overhead":
                rr += 1
                continue
            gv = ws.cell(rr, 7).value
            if isinstance(gv, (int, float)) and not isinstance(gv, bool):
                out[norm(b)] = (rr, str(b).strip())
            rr += 1
    return out


def find_label(ws, col, want, lo=1, hi=None, starts=False):
    hi = hi or ws.max_row
    hits = []
    for r in range(lo, hi + 1):
        v = ws.cell(r, col).value
        if not isinstance(v, str):
            continue
        if v == want or (starts and v.startswith(want)):
            hits.append(r)
    if len(hits) != 1:
        stop("label %r on %s col %d: %d hits" % (want, ws.title, col, len(hits)))
    return hits[0]


def review_cols(ws):
    """Column letters for the REVIEW columns the tab reads."""
    want = {"Name": None, "Full Cost \nAUD": None, "MStatus": None,
            "Overhead line": None}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h in want:
            want[h] = get_column_letter(c)
    for k, v in want.items():
        if v is None:
            stop("REVIEW column %r not found" % k)
    return want


def block_role_rows(ws):
    """[(tab row, REVIEW row)] for every FTE block role row on a 2.x tab."""
    out = []
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if isinstance(b, str) and b.startswith("='" + REVIEW + "'!$B$"):
            m = re.search(r"\$B\$(\d+)$", b)
            if m:
                out.append((r, int(m.group(1))))
    return out


def money(cell):
    cell.number_format = M2


def main(src, dst):
    log = Log("v3_lightson")
    wb = load(src)

    if TAB in wb.sheetnames:
        print("input already carries the Lights On tab - copying through")
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        return

    for t in [ANCHOR, CFG, OVH, ARC, REVIEW, "Lists"] + [r[2] for r in ROWS] \
            + [r[3] for r in ROWS if r[3]]:
        if t not in wb.sheetnames:
            stop("tab %r missing" % t)

    rev = wb[REVIEW]
    rc = review_cols(rev)
    lists = wb["Lists"]
    cfg = wb[CFG]

    # ---------------------------------------------------- what the book holds
    grids = {r[2]: grid(wb[r[2]]) for r in ROWS}

    # the Overhead line column is derived by formula, so its values come off
    # the cached copy the pipeline always carries (wbio builds every stage)
    wbv = openpyxl.load_workbook(src, data_only=True)
    revv = wbv[REVIEW]
    ar_of = {}                       # REVIEW row -> Overhead line
    st_of = {}                       # REVIEW row -> MStatus
    arc_col = stat_col = None
    for c in range(1, rev.max_column + 1):
        if rev.cell(1, c).value == "Overhead line":
            arc_col = c
        if rev.cell(1, c).value == "MStatus":
            stat_col = c
    for r in range(2, 701):
        v = revv.cell(r, arc_col).value
        if v is not None and str(v).strip():
            ar_of[r] = str(v).strip()
        s = revv.cell(r, stat_col).value
        if s is not None and str(s).strip():
            st_of[r] = str(s).strip()
    wbv.close()
    if not ar_of:
        stop("the input carries no cached Overhead line values; hand this "
             "stage a wbio built file")
    if len(set(ar_of.values())) > 12:
        stop("the Overhead line column reads like formulas, not values")

    # every vacant overhead role, classified by its lever as the input stands:
    # {line: [(tab, block row, lever)]}, membership frozen, values live
    vac_of = {}
    for t2 in dict.fromkeys(r[2] for r in ROWS):
        for br, rr in block_role_rows(wb[t2]):
            line = ar_of.get(rr)
            if line and line != "Squad" and st_of.get(rr) == "Vacant":
                vac_of.setdefault(line, []).append(
                    (t2, br, str(wb[t2].cell(br, 5).value or "").strip()))

    # the funded squads table on Lists (no Support %% row is fine for these)
    au1 = find_label(lists, 47, "Squad")            # AU header row
    funded = set()
    au_end = au1
    for r in range(au1 + 1, au1 + 30):
        v = lists.cell(r, 47).value
        if v is None:
            break
        funded.add(norm(v))
        au_end = r

    # GM cost and the uplift funding, already labelled on Lists
    gm_row = find_label(lists, 32, GM_LABEL)
    if not isinstance(lists.cell(gm_row, 33).value, (int, float)):
        stop("Lists %s row %d does not carry a number" % (GM_LABEL, gm_row))
    up_row = None
    for r in range(au1 + 1, au_end + 1):
        if norm(lists.cell(r, 47).value) == norm("Cyber Uplift"):
            up_row = r
    if up_row is None:
        stop("no Cyber Uplift line in the Lists funded table")

    # -------------------------------------- V1  pot source cells on Lists
    log.head("V1  the BP and Domain Architect pots, labelled on Lists")
    free = None
    for r in range(gm_row + 1, 60):
        if lists.cell(r, 32).value is None and lists.cell(r, 33).value is None \
                and lists.cell(r + 1, 32).value is None \
                and lists.cell(r + 1, 33).value is None:
            free = r
            break
    if free is None:
        stop("no free rows on Lists for the pot source cells")
    g12 = grids["2.12 COE BP&T"]
    g13 = grids["2.13 COE SA&D"]
    bp_grid = find_label(wb["2.12 COE BP&T"], 2, "TDD Business Partner",
                         g12["first"], g12["total"])
    da_grid = find_label(wb["2.13 COE SA&D"], 2, "Architecture",
                         g13["first"], g13["total"])
    bp_row, da_row = free, free + 1
    for row, label, tab, g in ((bp_row, BP_POT_LABEL, "2.12 COE BP&T", g12),
                               (da_row, DA_POT_LABEL, "2.13 COE SA&D", g13)):
        lists.cell(row, 32).value = label
        lists.cell(row, 33).value = (
            "=INDEX(%s!$S$%d:$S$%d,MATCH(%s,%s!$B$%d:$B$%d,0))"
            % (q(tab), g["first"], g["total"],
               '"TDD Business Partner"' if tab == "2.12 COE BP&T"
               else '"Architecture"',
               q(tab), g["first"], g["total"]))
        money(lists.cell(row, 33))
        log("V1", "Lists!AF%d:AG%d" % (row, row), "%s reads the %s grid live"
            % (label, tab))
    POT = {"bp": "Lists!$AG$%d" % bp_row, "da": "Lists!$AG$%d" % da_row,
           "gm": "Lists!$AG$%d" % gm_row}
    TEN = "COUNTA(Lists!$AS$2:$AS$12)"

    # ------------------------------------------------- V2  the tab, laid out
    log.head("V2  the Lights On tab")
    ws = wb.create_sheet(TAB, wb.sheetnames.index(ANCHOR) + 1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF002F6C"
    widths = {"A": 2, "B": 52, "C": 13.5, "D": 14.5, "E": 12, "F": 13,
              "G": 12, "H": 13, "I": 12, "J": 14, "K": 14.5, "L": 13.5,
              "M": 13.5, "N": 2.5, "O": 16.5, "P": 15.5}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws["B2"].value = "TDD Lights On"
    ws["B2"].font = Font(bold=True, size=16)

    HR = 4                              # header row
    D1 = HR + 1                         # first data row
    D2 = D1 + len(ROWS) - 1
    TOT = D2 + 1
    BUD = TOT + 1
    CTL = BUD + 1

    hdr_font = Font(bold=True, color=WHITE)
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    hdr_al = Alignment(wrap_text=True, horizontal="center", vertical="center")
    for col, text in HEADERS:
        c = ws[col + str(HR)]
        c.value = text
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_al
    ws.row_dimensions[HR].height = 58
    ws.freeze_panes = "A" + str(D1)
    log("V2", "%s!B%d:P%d" % (TAB, HR, HR),
        "his fifteen column headings, verbatim, navy and white")

    band = PatternFill("solid", fgColor=BAND)
    cream = PatternFill("solid", fgColor=CREAM)

    # ------------------------------------------------ V3  the fifteen rows
    log.head("V3  fifteen rows in 0.2 Data Config order, all live")
    toggle_rows = []
    pairs_note = {}
    for i, (label, kind, t2, t1, blabels) in enumerate(ROWS):
        r = D1 + i
        g = grids[t2]
        w2 = wb[t2]
        rng = lambda col: "%s!$%s$%d:$%s$%d" % (q(t2), col, g["first"],
                                                col, g["total"])
        tot_s = ('INDEX(%s,MATCH("Total portfolio",%s,0))'
                 % (rng("S"), rng("B")))
        ws.cell(r, 2).value = label

        # C  the row's whole after lever cost
        if kind == "pf":
            ws.cell(r, 3).value = "=%s+$E%d+$F%d+$G%d" % (tot_s, r, r, r)
        elif kind == "coe" and label == "COE BP&T":
            ws.cell(r, 3).value = "=%s-%s" % (tot_s, POT["bp"])
        elif kind == "coe" and label == "COE SA&D":
            ws.cell(r, 3).value = "=%s-%s" % (tot_s, POT["da"])
        else:                            # CRSO gross, TDD Cyber gross, EGI
            ws.cell(r, 3).value = "=" + tot_s

        # D  support cost
        if kind == "pf":
            sup = support_map(wb[t1])
            terms, missing = [], []
            for gr, name, sec in g["squads"]:
                hit = sup.get(norm(name))
                if hit:
                    terms.append("%s!$S$%d*%s!$G$%d" % (q(t2), gr, q(t1),
                                                        hit[0]))
                elif norm(name) in funded or sec == "No archetype in 1.x tabs":
                    missing.append(name)
                else:
                    stop("%s squad %r has no Support %% row on %s and is not "
                         "in the funded table" % (t2, name, t1))
            if not terms:
                stop("no support pairs for %s" % t2)
            ws.cell(r, 4).value = "=" + "+".join(terms)
            pairs_note[label] = (len(terms), missing)
        elif kind == "coe":
            ws.cell(r, 4).value = "=$C%d-$H%d" % (r, r)
        elif kind == "cyber":
            ws.cell(r, 4).value = (
                "=$C%d-INDEX(Lists!$AW$%d:$AW$%d,MATCH(\"Cyber Uplift\","
                "Lists!$AU$%d:$AU$%d,0))" % (r, au1 + 1, au_end, au1 + 1,
                                             au_end))
        else:
            ws.cell(r, 4).value = "=0"

        # E F G  the equal shares
        if kind == "pf":
            ws.cell(r, 5).value = "=%s/%s" % (POT["bp"], TEN)
            ws.cell(r, 6).value = "=%s/%s" % (POT["da"], TEN)
            ws.cell(r, 7).value = "=%s/%s" % (POT["gm"], TEN)
        else:
            for c in (5, 6, 7):
                ws.cell(r, c).value = "=0"

        # H  other overheads
        if kind == "pf":
            if g["ovh_total"] is None:
                stop("no 'Overhead roles total' row on %s" % t2)
            ws.cell(r, 8).value = ('=INDEX(%s,MATCH("Overhead roles total",'
                                   "%s,0))" % (rng("S"), rng("B")))
        elif kind == "coe":
            own = []
            for br, rr in block_role_rows(w2):
                line = ar_of.get(rr)
                if line and line != "Squad":
                    own.append("%s!$G$%d" % (q(t2), br))
            if not own:
                stop("no overhead coded people found on %s" % t2)
            f = "=(%s)/1000000" % "+".join(own)
            if label == "COE BP&T":
                f += "-%s" % POT["bp"]
            elif label == "COE SA&D":
                f += "-%s" % POT["da"]
            ws.cell(r, 8).value = f
        else:
            ws.cell(r, 8).value = "=0"

        # I  the toggle
        if kind in ("pf", "coe"):
            c = ws.cell(r, 9)
            c.value = 1
            c.fill = cream
            c.number_format = "0%"
            toggle_rows.append(r)

        # J K M O  one shape everywhere
        ws.cell(r, 10).value = "=$E%d+$F%d+$G%d+$H%d*$I%d" % (r, r, r, r, r)
        ws.cell(r, 11).value = "=$D%d+$J%d" % (r, r)

        # L  the 0.2 budget, live
        parts = []
        for bl in blabels:
            starts = bl.startswith("COE - Cyber")
            br = find_label(cfg, 2, bl, 5, 26, starts=starts)
            parts.append("%s!$E$%d" % (q(CFG), br))
        ws.cell(r, 12).value = "=" + "+".join(parts)

        ws.cell(r, 13).value = "=$K%d-$L%d" % (r, r)
        ws.cell(r, 15).value = "=$C%d-$K%d" % (r, r)

        # P  amount allocated in the 1.x tabs
        if kind == "egi":
            ws.cell(r, 16).value = ('=INDEX(%s,MATCH("Total portfolio",%s,0))'
                                    % (rng("P"), rng("B")))
        elif t1 is None:
            ws.cell(r, 16).value = "=0"
        else:
            w1 = wb[t1]
            pr = None
            for rr in range(1, w1.max_row + 1):
                if w1.cell(rr, 8).value == "Total applied":
                    pr = "=%s!$J$%d" % (q(t1), rr)
                    break
            if pr is None and kind == "cyber":
                rr = find_label(w1, 8, "Used for cyber FTE ($m)")
                pr = "=%s!$J$%d" % (q(t1), rr)
            if pr is None:
                for rr in range(1, w1.max_row + 1):
                    if w1.cell(rr, 8).value == "Budget line":
                        last = rr
                        for r2 in range(rr + 1, rr + 15):
                            h = w1.cell(r2, 8).value
                            if h is None or h in ("Total applied",
                                                  "Other cost (this model)",
                                                  "Left to fund"):
                                break
                            last = r2
                        pr = "=SUM(%s!$J$%d:$J$%d)" % (q(t1), rr + 1, last)
                        break
            if pr is None:
                for rr in range(1, w1.max_row + 1):
                    if w1.cell(rr, 8).value == "Funding position":
                        cells = []
                        for r2 in range(rr + 1, rr + 8):
                            h = w1.cell(r2, 8).value
                            if h is None or h == "Left to fund":
                                break
                            if h == "Other cost (this model)":
                                continue
                            cells.append("%s!$J$%d" % (q(t1), r2))
                        pr = "=" + "+".join(cells)
                        break
            if pr is None:
                stop("no funding block found on %s" % t1)
            ws.cell(r, 16).value = pr

        for c in (3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 15, 16):
            money(ws.cell(r, c))
        if (i % 2) == 1:
            for c in range(2, 17):
                if c == 9 and kind in ("pf", "coe"):
                    continue
                ws.cell(r, c).fill = band
        log("V3", "%s!B%d" % (TAB, r), "%s (%s)" % (label, t2))

    for label, (n, miss) in pairs_note.items():
        if miss:
            log.note("V3", "%s support pairs: %d squads paired, no Support %% "
                     "for %s so their support is 0" % (label, n,
                                                       ", ".join(miss)))

    # ------------------------------------------- V4  toggles get their list
    log.head("V4  the toggle column")
    dv = DataValidation(type="list", formula1='"%s"' % STEPS,
                        allow_blank=False, showErrorMessage=True,
                        errorTitle="Invalid entry",
                        error="Pick a value from the list")
    ws.add_data_validation(dv)
    dv.sqref = MultiCellRange([CellRange(min_col=9, max_col=9, min_row=r,
                                         max_row=r) for r in toggle_rows])
    log("V4", "%s!I%d:I%d" % (TAB, D1, D2),
        "cream toggles on %d rows, dropdown 0%% to 100%% in 5%% steps, "
        "default 100%%" % len(toggle_rows))

    # ------------------------------------- V5  total, budget, control rows
    log.head("V5  total, budget and the white control row")
    bold = Font(bold=True)
    ws.cell(TOT, 2).value = "Total"
    for c in (3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 15, 16):
        cl = get_column_letter(c)
        ws.cell(TOT, c).value = "=SUM(%s%d:%s%d)" % (cl, D1, cl, D2)
        money(ws.cell(TOT, c))
    for c in range(2, 17):
        ws.cell(TOT, c).font = bold
        ws.cell(TOT, c).fill = band

    bud_row = find_label(cfg, 2, "Budget", 20, 30)
    ws.cell(BUD, 2).value = "Budget"
    ws.cell(BUD, 12).value = "=%s!$E$%d" % (q(CFG), bud_row)
    ws.cell(BUD, 13).value = "=$K$%d-$L$%d" % (TOT, BUD)
    money(ws.cell(BUD, 12))
    money(ws.cell(BUD, 13))
    ws.cell(BUD, 2).font = bold
    ws.cell(BUD, 12).font = bold
    ws.cell(BUD, 13).font = bold
    log("V5", "%s!L%d:M%d" % (TAB, BUD, BUD),
        "the 53.80 budget row, over or under shown against total K")

    tot_terms = []
    for _, _, t2, _, _ in ROWS:
        g = grids[t2]
        tot_terms.append('INDEX(%s!$S$%d:$S$%d,MATCH("Total portfolio",'
                         "%s!$B$%d:$B$%d,0))" % (q(t2), g["first"], g["total"],
                                                 q(t2), g["first"], g["total"]))
    ws.cell(CTL, 2).value = ("Control, must read 0: total cost against the "
                             "fifteen lever tabs plus the GM layer, K against "
                             "D plus J, and O against C minus K")
    ws.cell(CTL, 3).value = "=ROUND($C$%d-(%s+%s),6)" % (
        TOT, "+".join(tot_terms), POT["gm"])
    ws.cell(CTL, 4).value = "=ROUND($K$%d-$D$%d-$J$%d,6)" % (TOT, TOT, TOT)
    ws.cell(CTL, 5).value = "=ROUND($O$%d-($C$%d-$K$%d),6)" % (TOT, TOT, TOT)
    white(ws, "B%d" % CTL, "C%d" % CTL, "D%d" % CTL, "E%d" % CTL)
    log("V5", "%s!B%d:E%d" % (TAB, CTL, CTL),
        "white font control row ties the table to the lever tabs")

    ws.conditional_formatting.add(
        "M%d:M%d" % (D1, BUD),
        FormulaRule(formula=["$M%d>0" % D1], font=Font(color=RED)))
    log("V5", "%s!M%d:M%d" % (TAB, D1, BUD),
        "red font only when genuinely over, parentheses when under")

    # ------------------------------------------------ V6  the analysis block
    log.head("V6  the analysis block, live numbers")
    r = CTL + 2
    head_font = Font(bold=True)

    def put(row, label, cells=(), fmt=M2, bold_label=False):
        ws.cell(row, 2).value = label
        if bold_label:
            ws.cell(row, 2).font = head_font
        for j, f in enumerate(cells):
            c = ws.cell(row, 3 + j)
            c.value = f
            c.number_format = fmt if not isinstance(fmt, (list, tuple)) \
                else fmt[j]

    ovh_ws = wb[OVH]
    o_all = find_label(ovh_ws, 2, "Overheads incl. GMs")
    o_line = {}                  # 3.2's line rows point at Lists AF by formula
    for rr in range(3, o_all):
        m = re.fullmatch(r"=Lists!\$AF\$(\d+)",
                         str(ovh_ws.cell(rr, 2).value or ""))
        if m:
            o_line[lists.cell(int(m.group(1)), 32).value] = rr
    for need in ("Technology Manager", "Head of Technology",
                 "Leadership - 8 GMs", "Delivery Manager"):
        if need not in o_line:
            stop("3.2 line %r not found" % need)
    o_tm = o_line["Technology Manager"]
    o_hot = o_line["Head of Technology"]
    o_gm = o_line["Leadership - 8 GMs"]
    o_dm = o_line["Delivery Manager"]

    put(r, "Why the lights on number is far from the archetype",
        bold_label=True)
    put(r + 1, "Overhead the archetype allows for, GMs included ($m)",
        ["=%s!$I$%d" % (q(OVH), o_all)])
    put(r + 2, "What the overhead people cost after levers, GMs included ($m)",
        ["=$H$%d+%s+%s+%s" % (TOT, POT["bp"], POT["da"], POT["gm"])])
    put(r + 3, "The gap sits in the leadership lines ($m over archetype): "
               "Technology Managers, then Heads of Technology, then GMs, "
               "then Delivery Managers",
        ["=%s!$K$%d" % (q(OVH), o_tm), "=%s!$K$%d" % (q(OVH), o_hot),
         "=%s!$K$%d" % (q(OVH), o_gm), "=%s!$K$%d" % (q(OVH), o_dm)])
    put(r + 4, "Squads after levers run under archetype, which is why the "
               "total cost never showed it. The overhead people are the "
               "lights on story.")
    log("V6", "%s!B%d" % (TAB, r), "why the number is far from the archetype")
    r += 6

    lines = []
    for rr in range(2, 9):
        v = lists.cell(rr, 32).value
        if isinstance(v, str) and v in set(ar_of.values()):
            lines.append(v)
    for v in sorted(set(ar_of.values()) - set(lines) - {"Squad"}):
        lines.append(v)
    put(r, "Vacant overheads, the biggest dial", bold_label=True)
    ws.cell(r, 3).value = "Vacant roles"
    ws.cell(r, 3).font = head_font
    ws.cell(r, 4).value = "On Hire at full price ($m)"
    ws.cell(r, 4).font = head_font
    rr = r + 1
    for line in lines:
        hire = [(t, br) for t, br, lv in vac_of.get(line, []) if lv == "Hire"]
        if hire:
            put(rr, line,
                ["=" + "+".join('COUNTIF(%s!$E$%d,"Hire")' % (q(t), br)
                                for t, br in hire),
                 "=(%s)/1000000" % "+".join("%s!$G$%d" % (q(t), br)
                                            for t, br in hire)],
                fmt=["0", M2])
        else:
            put(rr, line, ["=0", "=0"], fmt=["0", M2])
        rr += 1
    vac_tot = rr
    put(vac_tot, "All vacant overheads on Hire",
        ["=SUM(C%d:C%d)" % (r + 1, rr - 1), "=SUM(D%d:D%d)" % (r + 1, rr - 1)],
        fmt=["0", M2])
    ws.cell(vac_tot, 2).font = head_font
    ws.cell(vac_tot, 3).font = head_font
    ws.cell(vac_tot, 4).font = head_font
    levered = [(t, br, lv) for ms in vac_of.values() for t, br, lv in ms
               if lv != "Hire"]
    n_hold = sum(1 for x in levered if x[2] == "Hold")
    n_off = sum(1 for x in levered if x[2] == "Offshore")
    n_fill = sum(1 for x in levered if x[2] == "Filled")
    parts = []
    if n_hold:
        parts.append("%d on Hold at zero" % n_hold)
    if n_off:
        parts.append("%d offshored" % n_off)
    if n_fill:
        parts.append("%d marked to fill" % n_fill)
    put(vac_tot + 1, "Already levered: " + (", ".join(parts) or "none"))
    all_vac = [(t, br) for ms in vac_of.values() for t, br, lv in ms]
    claim_row = vac_tot + 2
    put(claim_row, "Hold every vacant overhead role and the lights on total "
                   "drops by this amount ($m)",
        ["=(%s)/1000000" % "+".join("%s!$G$%d" % (q(t), br)
                                    for t, br in sorted(all_vac))])
    log("V6", "%s!B%d" % (TAB, r),
        "vacant overheads by line, on Hire at full price, off the block "
        "cells; %d vacant overhead roles in all, %d still on Hire"
        % (len(all_vac), sum(len([1 for _, _, lv in ms if lv == "Hire"])
                             for ms in vac_of.values())))
    r = claim_row + 2

    crso = D1 + [x[0] for x in ROWS].index("COE Cyber Risk & Service Ops")
    var_row = find_label(cfg, 2, "Variance to full TDD budget", 20, 30)
    put(r, "The dials", bold_label=True)
    put(r + 1, "Hold the vacant overhead roles ($m)", ["=$C$%d" % claim_row])
    put(r + 2, "Fund the GMs above the lights on budget ($m)", ["=" + POT["gm"]])
    put(r + 3, "Bring COE Cyber Risk & Service Ops back inside its allocation "
               "($m over today)", ["=$M$%d" % crso])
    put(r + 4, "The unallocated slice of the full TDD budget ($m)",
        ["=%s!$E$%d" % (q(CFG), var_row)])
    log("V6", "%s!B%d" % (TAB, r), "the dials, every number live")
    r += 6

    arc_ws = wb[ARC]
    arc_tot = None
    for rr in range(4, arc_ws.max_row + 1):
        v = arc_ws.cell(rr, 2).value
        if isinstance(v, str) and "TDD total (" in v:
            arc_tot = rr
            break
    if arc_tot is None:
        stop("no TDD total row on 3.1")
    put(r, "Total archetype cost against total actual cost after levers ($m)",
        ["=%s!$D$%d" % (q(ARC), arc_tot), "=$C$%d-%s" % (TOT, POT["gm"])])
    ws.cell(r, 2).font = head_font
    log("V6", "%s!B%d" % (TAB, r), "the archetype comparison line")

    save(wb, dst)
    log.tail()
    print("wrote", dst)

    # ------------------------------------------------------------ self-check
    print("== self-check")
    ok = True

    def chk(name, cond):
        nonlocal ok
        print("%s %s" % ("PASS" if cond else "FAIL", name))
        ok = ok and cond

    wb2 = openpyxl.load_workbook(dst)
    chk("tab present and visible",
        TAB in wb2.sheetnames and wb2[TAB].sheet_state == "visible")
    chk("tab sits after %s" % ANCHOR,
        wb2.sheetnames.index(TAB) == wb2.sheetnames.index(ANCHOR) + 1)
    w = wb2[TAB]
    heads = all(w[col + str(HR)].value == text for col, text in HEADERS
                if text is not None)
    chk("his fifteen headings verbatim", heads and w["N" + str(HR)].value is None)
    chk("fifteen data rows in 0.2 order",
        [w.cell(D1 + i, 2).value for i in range(15)] == [x[0] for x in ROWS])
    tg = [rr for rr in range(D1, D2 + 1)
          if w.cell(rr, 9).fill.patternType == "solid"
          and w.cell(rr, 9).fill.fgColor.rgb == CREAM]
    chk("13 cream toggles, value 100 percent, format 0%",
        tg == toggle_rows and all(w.cell(rr, 9).value == 1 and
                                  w.cell(rr, 9).number_format == "0%"
                                  for rr in tg))
    dvs = [d for d in w.data_validations.dataValidation
           if d.formula1 and "5%" in str(d.formula1)]
    chk("strict dropdown on every toggle",
        len(dvs) == 1 and len(dvs[0].sqref.ranges) == len(toggle_rows))
    bad_fmt = [c.coordinate for row in w.iter_rows(min_row=D1, max_row=TOT)
               for c in row if c.column in (3, 4, 5, 6, 7, 8, 10, 11, 12, 13,
                                            15, 16)
               and c.value is not None and c.number_format != M2]
    chk("money format %r on every money cell" % M2, not bad_fmt)
    whole = []
    for row in w.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                if re.search(r"(?<![A-Z0-9$:.])\$?[A-Z]{1,3}\$?:\$?[A-Z]{1,3}"
                             r"(?![A-Z0-9(])", c.value):
                    whole.append(c.coordinate)
    chk("bounded ranges only, no whole column references", not whole)
    texty = [c.coordinate for row in w.iter_rows() for c in row
             if isinstance(c.value, str) and not c.value.startswith("=")
             and ("-" in c.value or "–" in c.value or
                  "—" in c.value)]
    chk("no dash of any kind in the tab's text", not texty)
    banned = [c.coordinate for row in w.iter_rows() for c in row
              if isinstance(c.value, str) and re.search(
                  r"\b(wave|seat|seats|design)\b|to projects", c.value, re.I)]
    chk("no banned words", not banned)
    creamf = [c.coordinate for row in w.iter_rows() for c in row
              if c.fill.patternType == "solid"
              and c.fill.fgColor.rgb == CREAM
              and isinstance(c.value, str) and c.value.startswith("=")]
    chk("cream cells are typed inputs only, never formulas", not creamf)
    chk("freeze panes below the header", w.freeze_panes == "A" + str(D1))
    chk("control row is white font",
        all(w.cell(CTL, c).font.color and
            w.cell(CTL, c).font.color.rgb == WHITE for c in (2, 3, 4, 5)))
    cf = [x for x in w.conditional_formatting]
    chk("red when over rule on M", len(cf) == 1 and
        str(cf[0].sqref) == "M%d:M%d" % (D1, BUD))
    wb2.close()
    if not ok:
        raise SystemExit(2)
    print("self-check clean: the tab, toggles, formats and controls hold")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        stop("usage: v3_lightson.py <in.xlsx> <out.xlsx>")
    main(sys.argv[1], sys.argv[2])
