#!/usr/bin/env python3
"""v7_gate - the QA gate for the Lights On build (v6 output).

  python3 v7_gate.py <final.xlsx> [--recalc]

READS AND VERIFIES ONLY. The one write it can do is a recalc round-trip via
scripts/v10/wbio.py onto a scratch copy when the file arrives without cached
values (or when --recalc is passed); the input file is never touched.

Check groups (one PASS/FAIL line each, counted at the end, non-zero exit on
any FAIL):
  A ledger      528 roles, overhead line counts, Ed Tacey off the lines,
                Shane Ker on TM, the two Assurance/Excellence managers
  B levers      his 11 lever edits, person-keyed multisets against HIS file
  C viren/egi   Viren on 2.4, absent 2.1, the 1.4 / 1.1 / 1.3 EGI lines
  D part-time   FTE<1 scaling evidence, plausible base, unscaled heuristic
  E lights on   tab, verbatim headers, rows, every cell of C D E F G H J K L
                M O independently recomputed from REVIEW + 2.x + 1.x + 0.2 +
                Lists cached values and tied to 1e-6, toggles cream with the
                0-100% validation, L total 50.5, budget row 53.8
  F exclusions  every Hold role at 0 after the lever, Hold factor 0, no
                Paused lever, EGI row zeros except C O P
  G netting     2.12 = -(BP pot), 2.13 = -(DA pot), 3.1 netted to match
  H 0.2         TDD Cyber row = 2.15 after levers less the Lists uplift
                funding, spend total = the column
  I controls    every control row 0, 4.0 Data QA all zero, the Lights On
                control 0 and white, no error cells
  J hygiene     dash literals, cream formulas, banned words, en/em dashes,
                parens formats on the negative-capable columns
  K protection  every sheet locked with the Tdd123 hash, REVIEW fully locked,
                unlocked cells only in the allowed classes, structure locked
  L role id     REVIEW Role ID column R0001.., 2.x col A IDs resolve, zero
                row-anchored $B$/$C$/$AK$ refs left in role-row B/C/D
  M ledger      the final numbers block (print only, not a check)

Every numeric expectation is re-derived from the file itself (grids, REVIEW,
Lists, 0.2) or is one of his spec constants (528 roles, line counts, 50.5,
53.8, Tdd123). Nothing is pinned to a stale build artefact.

Check idiom inherited from scripts/v10/update/u7_carry.py.
"""
import sys, os, re, json, collections

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SCRATCH = os.path.dirname(HERE)
V10 = "/home/user/anthropic-claude-code/scripts/v10"
HIS = ("/root/.claude/uploads/e550b440-3996-5abb-87e5-bafafe598f82/"
       "ea5ee9ff-Copy_of_TDD_Cost_Calc_300726_old_version_w_edits.xlsx")
EDITS = os.path.join(SCRATCH, "his_lever_edits.json")

REVIEW = "REVIEW - Complete Role Mapping"
LO = "3.5 TDD Lights On"
LEVERS = ("Filled", "Hire", "Hold", "Offshore")
CREAM = "FFFFF2CC"
WHITE = "FFFFFFFF"
SRC_TABS = ("0.1 Budget Table (Fin)", "0.4 Presentation Pack")
ERR = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!",
       "Err:")
TOL = 1e-6

# his spec constants (acceptance targets, not derived state)
ROLES = 528
LINES = {"Head of Technology": 15, "Technology Manager": 25,
         "Delivery Manager": 12, "Business Partner": 6,
         "Domain Architect": 7, "Program Management": 3}
L_TOTAL = 50.5
L_BUDGET = 53.8
PASSWORD = "Tdd123"

# the Lights On tab as he specified it, verbatim
HEADERS = {2: "Portfolios & COEs & EGI",
           3: "Cost (Total Cost)",
           4: "Support Cost (this is the % in the 1.x tabs)",
           5: "BP allocation",
           6: "Domain architect allocation",
           7: "GM allocation",
           8: "Other overheads",
           9: "Other overheads toggle",
           10: "Amount of overheads charged to TDD",
           11: "Total portfolio cost charged to TDD",
           12: "TDD Lights On budget",
           13: "Over/ Under lights on budget",
           15: "Total Cost left to be recharged to business",
           16: "Amount allocated in 1.x tabs"}
LROWS = ["COE SA&D", "COE Cyber Risk & Service Ops", "COE BP&T",
         "Ampol Retail", "Z Retail", "Ampol & Z Customer",
         "Commercial Fuels", "Energy Solutions & B2B", "Infrastructure",
         "P&C", "Finance", "TDD Group Functions", "TDD Data", "TDD Cyber",
         "EGI"]
TABOF = {"COE SA&D": "2.13", "COE Cyber Risk & Service Ops": "2.11",
         "COE BP&T": "2.12", "Ampol Retail": "2.1", "Z Retail": "2.10",
         "Ampol & Z Customer": "2.2", "Commercial Fuels": "2.9",
         "Energy Solutions & B2B": "2.8", "Infrastructure": "2.7",
         "P&C": "2.5", "Finance": "2.6", "TDD Group Functions": "2.4",
         "TDD Data": "2.3", "TDD Cyber": "2.15", "EGI": "2.14"}
COES = ("COE SA&D", "COE Cyber Risk & Service Ops", "COE BP&T")
NOSHARE = set(COES) | {"TDD Cyber", "EGI"}
# 0.2 budget rows feeding each Lights On budget cell (prefix match)
BUDMAP = {"COE SA&D": ["COE - Strategy Architecture", "COE - Data"],
          "COE Cyber Risk & Service Ops": ["COE - Cyber"],
          "COE BP&T": ["COE - Transformation", "COE - Business Partnering"],
          "Ampol & Z Customer": ["Ampol Customer", "Z Customer"],
          "TDD Data": ["TDD Data"]}

GRID_TOTALS = ("Squads total", "Directly funded total",
               "Overhead roles total", "Total portfolio")


def norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def num(x):
    return x if isinstance(x, (int, float)) and not isinstance(x, bool) else None


# ------------------------------------------------------------------ reporting

class Gate:
    def __init__(self):
        self.groups = []

    def group(self, letter, name):
        g = {"letter": letter, "name": name, "subs": [], "notes": []}
        self.groups.append(g)
        return g

    @staticmethod
    def sub(g, ok, text):
        g["subs"].append((bool(ok), text))
        return bool(ok)

    @staticmethod
    def note(g, text):
        g["notes"].append(text)

    def report(self):
        fails = 0
        for g in self.groups:
            bad = [t for ok, t in g["subs"] if not ok]
            ok = not bad and bool(g["subs"])
            if not g["subs"]:
                bad = ["no sub-checks ran"]
            if not ok:
                fails += 1
            head = "%-5s %s %-12s" % ("PASS" if ok else "FAIL", g["letter"],
                                      g["name"])
            summary = ("%d/%d ok" % (sum(1 for s in g["subs"] if s[0]),
                                     len(g["subs"])))
            print("%s %s%s" % (head, summary,
                               "" if ok else "  -> " + "; ".join(bad[:3])),
                  flush=True)
            for okv, t in g["subs"]:
                if not okv:
                    print("        off: %s" % t, flush=True)
            for t in g["notes"]:
                print("        note: %s" % t, flush=True)
        print("\n%d groups, %d pass, %d fail"
              % (len(self.groups), len(self.groups) - fails, fails),
              flush=True)
        return fails


# ------------------------------------------------------------- layout readers

def find_tab(sheets, prefix):
    for t in sheets:
        if t == prefix or t.startswith(prefix + " "):
            return t
    return None


def total_row(wf):
    for r in range(5, min(wf.max_row, 90) + 1):
        if wf.cell(r, 2).value == "Total portfolio":
            return r
    return None


class Tab:
    """One 2.x lever modelling tab, both loads, with derived structure."""

    def __init__(self, title, wf, wv, idcol_map):
        self.title = title
        self.wf, self.wv = wf, wv
        self.total_r = total_row(wf)
        self.grid = []            # (row, label) with numeric O, non-total
        self.grid_all = {}        # label -> row for every plain-text B row
        end = self.total_r or min(wf.max_row, 40)
        for r in range(7, (self.total_r or end) + 1):
            b = wf.cell(r, 2).value
            if not (isinstance(b, str) and b and not b.startswith("=")):
                continue
            self.grid_all.setdefault(b, r)
            if b in GRID_TOTALS:
                continue
            if num(wv.cell(r, 15).value) is not None:
                self.grid.append((r, b))
        self.roles = []
        for r in range(1, wf.max_row + 1):
            d = wf.cell(r, 4).value
            if not (isinstance(d, str) and d.startswith("=") and REVIEW in d):
                continue
            led = None
            # the old direct form only ( ='REVIEW...'!$AK$36 ), never the
            # $AK$2:$AK$531 range inside an INDEX/MATCH rewrite
            m = re.fullmatch(r"='[^']+'!\$AK\$(\d+)", d.strip())
            if m:
                led = int(m.group(1))
            rid = wf.cell(r, 1).value
            if led is None and isinstance(rid, str) and re.fullmatch(
                    r"R\d{4}", rid.strip()):
                led = idcol_map.get(rid.strip())
            self.roles.append({
                "row": r, "ledger": led,
                "id": rid if isinstance(rid, str) else None,
                "name": wv.cell(r, 2).value, "title": wv.cell(r, 3).value,
                "status": wv.cell(r, 4).value,
                "lever": wf.cell(r, 5).value,
                "F": num(wv.cell(r, 6).value),
                "G": num(wv.cell(r, 7).value)})

    def gval(self, label, col):
        r = self.grid_all.get(label)
        if r is None:
            return None
        return num(self.wv.cell(r, col).value)

    def total(self, col=19):
        if self.total_r is None:
            return None
        return num(self.wv.cell(self.total_r, col).value)


def pct_map(wf, wv):
    """{normalized squad label: support %} off a 1.x tab."""
    out, clash = {}, []
    for r in range(1, wf.max_row + 1):
        b = wf.cell(r, 2).value
        c = wf.cell(r, 3).value
        g = num(wv.cell(r, 7).value)
        if not (isinstance(b, str) and b and not b.startswith("=")):
            continue
        if not (isinstance(c, str) and c and not str(c).startswith("=")):
            continue
        if g is None or not (0 <= g <= 1):
            continue
        k = norm(b)
        if k in out and abs(out[k] - g) > 1e-12:
            clash.append(k)
        out.setdefault(k, g)
    return out, clash


def ledger_rows(rvf):
    out = []
    for r in range(2, rvf.max_row + 1):
        n = rvf.cell(r, 2).value
        if n is None or (isinstance(n, str) and not n.strip()):
            continue
        out.append(r)
    return out


def role_levers_of(path):
    """{(tab, norm name, norm title): sorted lever multiset} for a workbook."""
    wf = openpyxl.load_workbook(path, data_only=False)
    wv = openpyxl.load_workbook(path, data_only=True)
    out = collections.defaultdict(list)
    for ws in wf.worksheets:
        if not ws.title.startswith("2."):
            continue
        wsv = wv[ws.title]
        for r in range(1, ws.max_row + 1):
            d = ws.cell(r, 4).value
            if not (isinstance(d, str) and d.startswith("=") and REVIEW in d):
                continue
            out[(ws.title, norm(wsv.cell(r, 2).value),
                 norm(wsv.cell(r, 3).value))].append(ws.cell(r, 5).value)
    wf.close()
    wv.close()
    return {k: sorted(str(x) for x in v) for k, v in out.items()}


def fmt_sections(fmt):
    out, cur, q, i = [], "", False, 0
    while i < len(fmt):
        ch = fmt[i]
        if ch == '"':
            q = not q
            cur += ch
        elif ch == "\\" and i + 1 < len(fmt):
            cur += fmt[i:i + 2]
            i += 2
            continue
        elif ch == ";" and not q:
            out.append(cur)
            cur = ""
        else:
            cur += ch
        i += 1
    out.append(cur)
    return out


DASHFMT = re.compile(r'(\\-|"-")')


# ------------------------------------------------------------------- the gate

def main(argv):
    if not argv:
        print("usage: v7_gate.py <final.xlsx> [--recalc]")
        return 2
    src = argv[0]
    force_recalc = "--recalc" in argv[1:]
    if not os.path.exists(src):
        print("STOP: no such file: %s" % src)
        return 2

    print("v7_gate on %s" % src, flush=True)
    f = openpyxl.load_workbook(src, data_only=False)
    v = openpyxl.load_workbook(src, data_only=True)

    # ---- cached-value probe; one wbio round-trip onto a scratch copy if bare
    blank = total = 0
    probe_tabs = [t for t in f.sheetnames if t.startswith("2.")][:6]
    for t in probe_tabs:
        wf, wv = f[t], v[t]
        for row in wf.iter_rows():
            for cl in row:
                if isinstance(cl.value, str) and cl.value.startswith("="):
                    total += 1
                    if wv[cl.coordinate].value is None:
                        blank += 1
    bare = total > 0 and blank / total > 0.2
    if force_recalc or bare:
        print("cached values %s (%d of %d formula cells blank on the probe "
              "tabs) - one wbio recalc round-trip onto a scratch copy"
              % ("recalc forced" if force_recalc else "missing", blank,
                 total), flush=True)
        sys.path.insert(0, V10)
        import wbio
        work = os.path.join(HERE, "_gate_recalc.xlsx")
        wbio.build(src, work)
        f = openpyxl.load_workbook(work, data_only=False)
        v = openpyxl.load_workbook(work, data_only=True)
    else:
        print("cached values present (%d of %d formula cells blank on the "
              "probe tabs) - verifying the file as handed over"
              % (blank, total), flush=True)

    gate = Gate()
    sheets = f.sheetnames
    if REVIEW not in sheets:
        print("STOP: %r missing - not the model" % REVIEW)
        return 2
    rvf, rvv = f[REVIEW], v[REVIEW]

    # ---- shared context
    led = ledger_rows(rvf)
    name_of = {r: str(rvf.cell(r, 2).value or "").strip() for r in led}
    title_of = {r: str(rvf.cell(r, 3).value or "").strip() for r in led}
    ar_of = {r: str(rvv.cell(r, 44).value or "").strip() for r in led}
    status_of = {r: str(rvv.cell(r, 37).value or "").strip() for r in led}
    mtab_of = {r: str(rvv.cell(r, 36).value or "").strip() for r in led}
    fte_of = {r: num(rvv.cell(r, 15).value) for r in led}
    cost_of = {r: num(rvv.cell(r, 27).value) or 0.0 for r in led}

    # Role ID column (first header 'Role ID' beyond AY=51, else anywhere)
    idcol = None
    for c in range(52, rvf.max_column + 1):
        if str(rvf.cell(1, c).value or "").strip() == "Role ID":
            idcol = c
            break
    if idcol is None:
        for c in range(1, rvf.max_column + 1):
            if str(rvf.cell(1, c).value or "").strip() == "Role ID":
                idcol = c
                break
    idmap = {}
    if idcol:
        for r in led:
            val = rvv.cell(r, idcol).value or rvf.cell(r, idcol).value
            if isinstance(val, str) and re.fullmatch(r"R\d{4}", val.strip()):
                idmap[val.strip()] = r

    tabs2 = [t for t in sheets if t.startswith("2.")]
    T = {}
    for t in tabs2:
        T[t] = Tab(t, f[t], v[t], idmap)

    def tab_for(label):
        return find_tab(sheets, TABOF[label])

    # Lists lookups, all re-derived
    Lf, Lv = f["Lists"], v["Lists"]
    ovh_labels = []
    for c in range(1, Lf.max_column + 1):
        if str(Lf.cell(1, c).value or "").strip() == "Overhead line":
            for r in range(2, 20):
                x = Lf.cell(r, c).value
                if not isinstance(x, str) or x == "Total":
                    break
                ovh_labels.append(x)
            break
    gm_cost = None
    for c in range(1, Lf.max_column + 1):
        for r in range(1, min(Lf.max_row, 40) + 1):
            if str(Lf.cell(r, c).value or "").strip() == "GM cost ($m)":
                gm_cost = num(Lv.cell(r, c + 1).value)
    lever_factor = {}
    for c in range(1, Lf.max_column + 1):
        if str(Lf.cell(1, c).value or "").strip() == "Lever":
            for r in range(2, 10):
                k = Lf.cell(r, c).value
                if not isinstance(k, str) or not k:
                    break
                lever_factor[k] = num(Lv.cell(r, c + 1).value)
            break
    uplift_fund = None
    for c in range(1, Lf.max_column + 1):
        if str(Lf.cell(1, c).value or "").strip() == "Squad" \
                and str(Lf.cell(1, c + 1).value or "").strip() == "Funded by":
            for r in range(2, 20):
                if str(Lf.cell(r, c).value or "").strip() == "Cyber Uplift":
                    uplift_fund = num(Lv.cell(r, c + 2).value)
            break

    # 0.2 grid map
    cfg_t = find_tab(sheets, "0.2")
    cfgf, cfgv = f[cfg_t], v[cfg_t]
    cfg_rows, cfg_total_r, cfg_budget_r, cfg_var_r = [], None, None, None
    for r in range(6, min(cfgf.max_row, 60) + 1):
        b = cfgf.cell(r, 2).value
        if not isinstance(b, str) or not b:
            continue
        if b == "Total":
            cfg_total_r = r
            continue
        if b == "Budget":
            cfg_budget_r = r
            continue
        if b.startswith("Variance to full TDD budget"):
            cfg_var_r = r
            continue
        if cfg_total_r is None:
            cfg_rows.append((r, b))
    cfg_E = {b: num(cfgv.cell(r, 5).value) for r, b in cfg_rows}
    cfg_F = {b: num(cfgv.cell(r, 6).value) for r, b in cfg_rows}

    def cfg_match(alias):
        hits = [b for _, b in cfg_rows if b == alias or b.startswith(alias)]
        return hits

    # the pots, ledger-based with a grid cross-read
    def pot(tab_prefix, line):
        t = find_tab(sheets, tab_prefix)
        if t is None or t not in T:
            return None, None
        s = 0.0
        n = 0
        for ro in T[t].roles:
            lr = ro["ledger"]
            if lr is not None and ar_of.get(lr) == line and ro["G"] is not None:
                s += ro["G"]
                n += 1
        return s / 1e6, n

    bp_pot, bp_n = pot("2.12", "Business Partner")
    da_pot, da_n = pot("2.13", "Domain Architect")

    def own_overhead(tab_prefix):
        t = find_tab(sheets, tab_prefix)
        if t is None or t not in T:
            return None
        s = 0.0
        for ro in T[t].roles:
            lr = ro["ledger"]
            if lr is not None and ar_of.get(lr) in ovh_labels \
                    and ro["G"] is not None:
                s += ro["G"]
        return s / 1e6

    # ============================================================== A ledger
    g = gate.group("A", "ledger")
    try:
        gate.sub(g, len(led) == ROLES, "roles %d (spec %d)" % (len(led), ROLES))
        counts = collections.Counter(ar_of[r] for r in led
                                     if ar_of[r] in LINES)
        for line, want in sorted(LINES.items()):
            gate.sub(g, counts.get(line, 0) == want,
                     "%s %d (spec %d)" % (line, counts.get(line, 0), want))
        ed = [r for r in led if norm(name_of[r]) == "ed tacey"]
        ed_ok = len(ed) == 1 and ar_of[ed[0]] not in ovh_labels
        gate.sub(g, ed_ok,
                 "Ed Tacey overhead line %r (must not be one of %s)"
                 % ([ar_of[r] for r in ed], ovh_labels))
        sk = [r for r in led if norm(name_of[r]) == "shane ker"]
        gate.sub(g, len(sk) == 1 and ar_of[sk[0]] == "Technology Manager",
                 "Shane Ker line %r (spec Technology Manager)"
                 % [ar_of[r] for r in sk])
        mgrs = [r for r in led if norm(title_of[r]) in
                ("delivery assurance manager", "delivery excellence manager")]
        ok = (len(mgrs) == 2
              and all(ar_of[r] == "Delivery Manager" for r in mgrs)
              and all(mtab_of[r] == "COE BP&T" for r in mgrs))
        gate.sub(g, ok,
                 "Assurance/Excellence managers on the DM line with MTab COE "
                 "BP&T: %s"
                 % [(title_of[r], ar_of[r], mtab_of[r]) for r in mgrs])
        dmcost = sum(cost_of[r] for r in led
                     if ar_of[r] == "Delivery Manager")
        gate.note(g, "Delivery Manager line cost %.2f (derived)" % dmcost)
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ============================================================== B levers
    g = gate.group("B", "his levers")
    try:
        edits = json.load(open(EDITS))
        his = role_levers_of(HIS)
        # the gate's own view of the file under test, keyed the same way
        cur = {}
        for t in tabs2:
            for ro in T[t].roles:
                cur.setdefault((t, norm(ro["name"]), norm(ro["title"])),
                               []).append(str(ro["lever"]))
        cur = {k: sorted(vv) for k, vv in cur.items()}
        gate.sub(g, len(edits) == 11, "11 keys in his_lever_edits.json (%d)"
                 % len(edits))
        for e in edits:
            k = (e["tab"], norm(e["name"]), norm(e["role"]))
            hm = his.get(k, [])
            cm = cur.get(k, [])
            jt = sorted(str(x) for x in e["to"])
            gate.sub(g, hm == jt,
                     "his file matches the diff for %s / %s / %s: his=%s "
                     "json=%s" % (e["tab"], e["name"], e["role"], hm, jt))
            gate.sub(g, cm == hm and len(cm) > 0,
                     "%s / %s / %s levers %s (his %s)"
                     % (e["tab"], e["name"], e["role"], cm, hm))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # =========================================================== C viren/egi
    g = gate.group("C", "viren/egi")
    try:
        t24 = find_tab(sheets, "2.4")
        t21 = find_tab(sheets, "2.1")
        names24 = {norm(ro["name"]) for ro in T[t24].roles}
        names21 = {norm(ro["name"]) for ro in T[t21].roles}
        gate.sub(g, "viren khatri" in names24, "Viren Khatri on %s" % t24)
        gate.sub(g, "viren khatri" not in names21,
                 "Viren Khatri absent from %s" % t21)
        f24 = T[t24].gval("EGI TDD", 6)
        gate.sub(g, f24 == 5, "%s grid EGI TDD roles %s (spec 5)" % (t24, f24))
        gate.sub(g, "EGI TDD" not in T[t21].grid_all,
                 "%s grid has no EGI TDD line" % t21)
        vk = [r for r in led if norm(name_of[r]) == "viren khatri"]
        gate.sub(g, vk and mtab_of[vk[0]] == "TDD Group Functions",
                 "Viren MTab %r (spec TDD Group Functions)"
                 % [mtab_of[r] for r in vk])
        egi_tdd_funded = T[t24].gval("EGI TDD", 16)
        # 1.4 Significant Items EGI reads the 2.4 EGI TDD funded total
        t14 = find_tab(sheets, "1.4")
        got = None
        for r in range(1, f[t14].max_row + 1):
            for c in range(2, f[t14].max_column + 1):
                if str(f[t14].cell(r, c).value or "").strip() \
                        == "Significant Items EGI":
                    for cc in range(c + 1, f[t14].max_column + 1):
                        x = num(v[t14].cell(r, cc).value)
                        if x is not None:
                            got = x
                            break
        gate.sub(g, got is not None and egi_tdd_funded is not None
                 and abs(got - egi_tdd_funded) <= TOL,
                 "1.4 'Significant Items EGI' %s against 2.4 EGI TDD funded "
                 "%s" % (got, egi_tdd_funded))
        # 1.1 Significant Items EGI reads the 2.1 EGI Retail funded total
        t11x = find_tab(sheets, "1.1")
        egi_retail = T[t21].gval("EGI Retail", 16)
        got11 = None
        for r in range(1, f[t11x].max_row + 1):
            for c in range(2, f[t11x].max_column + 1):
                if str(f[t11x].cell(r, c).value or "").strip() \
                        == "Significant Items EGI":
                    for cc in range(c + 1, f[t11x].max_column + 1):
                        x = num(v[t11x].cell(r, cc).value)
                        if x is not None:
                            got11 = x
                            break
        gate.sub(g, got11 is not None and egi_retail is not None
                 and abs(got11 - egi_retail) <= TOL,
                 "1.1 EGI line %s against 2.1 EGI Retail funded %s"
                 % (got11, egi_retail))
        # 1.3 EGI Ent Data platform line reads 2.3's directly funded line
        t13 = find_tab(sheets, "1.3")
        t23 = find_tab(sheets, "2.3")
        egi_23 = T[t23].gval("EGI", 16)
        got13 = None
        for r in range(1, f[t13].max_row + 1):
            if str(f[t13].cell(r, 2).value or "").strip() == "EGI Ent Data":
                # the row carries Support % 0 before the funded amount; the
                # line's value is the first numeric that is not the 0% cell
                for cc in range(3, f[t13].max_column + 1):
                    x = num(v[t13].cell(r, cc).value)
                    if x is not None and abs(x) > 1e-9:
                        got13 = x
                        break
                break
        gate.sub(g, got13 is not None and egi_23 is not None
                 and abs(got13 - egi_23) <= TOL,
                 "1.3 'EGI Ent Data' line %s against 2.3 EGI funded %s"
                 % (got13, egi_23))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ============================================================ D part-time
    g = gate.group("D", "part-time")
    try:
        full_by_title = collections.defaultdict(list)
        for r in led:
            if fte_of[r] == 1:
                full_by_title[norm(title_of[r])].append(cost_of[r])
        pt = [r for r in led
              if fte_of[r] is not None and 0 < fte_of[r] < 1]
        gate.note(g, "FTE<1 rows: %d" % len(pt))
        for r in pt:
            fte, cost = fte_of[r], cost_of[r]
            base = cost / fte
            in_band = 1e5 <= base <= 7e5
            twins = [x for x in full_by_title[norm(title_of[r])]
                     if abs(x - cost) < 0.005]
            aa = rvf.cell(r, 27).value
            au = num(rvv.cell(r, 47).value)
            typed = num(aa) is not None
            frm = isinstance(aa, str) and aa.startswith("=")
            refs_fte = frm and re.search(r"\$O\$?%d\b" % r, aa) is not None
            evidence = typed or (au is not None and au > 0) or refs_fte
            gate.sub(g, in_band,
                     "%s (%s) FTE %.2f cost %.2f implied base %.0f in "
                     "100k-700k" % (name_of[r], title_of[r], fte, cost, base))
            gate.sub(g, not twins,
                     "%s cost %.2f equals a same-title full-timer while "
                     "FTE<1 (unscaled heuristic): twins %s"
                     % (name_of[r], cost, twins[:2]))
            gate.sub(g, evidence,
                     "%s Full Cost carries scaling evidence (typed override, "
                     "AU override or an FTE-referencing formula); AA %s AU %s"
                     % (name_of[r],
                        "typed" if typed else ("formula" if frm else repr(aa)),
                        au))
            gate.note(g, "  %-24s FTE %.2f cost %12.2f base %9.0f %s"
                      % (name_of[r][:24], fte, cost, base,
                         "OVERRIDE" if (typed or au) else "formula"))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ============================================================ E lights on
    g = gate.group("E", "lights on")
    lo_ctx = {}
    try:
        if LO not in sheets:
            gate.sub(g, False, "tab %r missing" % LO)
        else:
            lof, lov = f[LO], v[LO]
            i34 = sheets.index(find_tab(sheets, "3.4")) if find_tab(
                sheets, "3.4") else -2
            gate.sub(g, sheets.index(LO) == i34 + 1,
                     "positioned directly after 3.4 (index %d against %d)"
                     % (sheets.index(LO), i34))
            gate.sub(g, f[LO].sheet_state == "visible", "tab visible (%s)"
                     % f[LO].sheet_state)
            hdr_r = None
            for r in range(1, min(lof.max_row, 30) + 1):
                if lof.cell(r, 2).value == HEADERS[2]:
                    hdr_r = r
                    break
            gate.sub(g, hdr_r is not None, "header row found (B %r)"
                     % HEADERS[2])
            if hdr_r is None:
                raise RuntimeError("no header row")
            bad_h = []
            for c, want in HEADERS.items():
                got = lof.cell(hdr_r, c).value
                if str(got or "").strip() != want:
                    bad_h.append((openpyxl.utils.get_column_letter(c), got))
            got_n = lof.cell(hdr_r, 14).value
            if got_n not in (None, ""):
                bad_h.append(("N", got_n))
            gate.sub(g, not bad_h, "his 15 headers verbatim, N blank: %d off "
                     "%s" % (len(bad_h), bad_h[:3]))
            rows = {}
            r = hdr_r + 1
            labels = []
            total_r_lo = None
            while r <= lof.max_row and len(labels) < 40:
                b = lov.cell(r, 2).value
                if isinstance(b, str) and b.strip() == "Total":
                    total_r_lo = r
                    break
                if b not in (None, ""):
                    labels.append(str(b).strip())
                    rows[str(b).strip()] = r
                r += 1
            gate.sub(g, labels == LROWS,
                     "the 15 rows in 0.2 order: %s"
                     % ("exact" if labels == LROWS else
                        "got %s" % labels[:16]))
            gate.sub(g, total_r_lo is not None, "Total row present")
            budget_r_lo = None
            if total_r_lo:
                for rr in range(total_r_lo + 1, total_r_lo + 4):
                    if "budget" in str(lov.cell(rr, 2).value or "").lower():
                        budget_r_lo = rr
                        break
            gate.sub(g, budget_r_lo is not None, "Budget row present")
            lo_ctx = {"hdr": hdr_r, "rows": rows, "total": total_r_lo,
                      "budget": budget_r_lo, "labels": labels}

            # ---- independent recompute of C D E F G H J K L M O
            gate.sub(g, bp_pot is not None and bp_n == LINES[
                "Business Partner"],
                "BP pot from %d Business Partner roles on 2.12: %s"
                % (bp_n or 0, bp_pot))
            gate.sub(g, da_pot is not None and da_n == LINES[
                "Domain Architect"],
                "DA pot from %d Domain Architect roles on 2.13: %s"
                % (da_n or 0, da_pot))
            gate.sub(g, gm_cost is not None,
                     "GM cost on Lists: %s" % gm_cost)
            shares = ((bp_pot or 0) + (da_pot or 0) + (gm_cost or 0)) / 10.0

            exp = {}
            for lab in LROWS:
                t = tab_for(lab)
                tb = T.get(t)
                if tb is None:
                    gate.sub(g, False, "no 2.x tab for %r" % lab)
                    continue
                S = tb.total(19)
                ovhS = tb.gval("Overhead roles total", 19)
                e_ = {}
                if lab in NOSHARE:
                    e_["E"] = e_["F"] = e_["G"] = 0.0
                else:
                    e_["E"] = (bp_pot or 0) / 10.0
                    e_["F"] = (da_pot or 0) / 10.0
                    e_["G"] = (gm_cost or 0) / 10.0
                if lab == "COE BP&T":
                    e_["C"] = S - (bp_pot or 0)
                    e_["H"] = (own_overhead("2.12") or 0) - (bp_pot or 0)
                elif lab == "COE SA&D":
                    e_["C"] = S - (da_pot or 0)
                    e_["H"] = (own_overhead("2.13") or 0) - (da_pot or 0)
                elif lab == "COE Cyber Risk & Service Ops":
                    e_["C"] = S
                    e_["H"] = own_overhead("2.11") or 0
                elif lab == "TDD Cyber":
                    e_["C"] = S
                    e_["H"] = 0.0
                elif lab == "EGI":
                    e_["C"] = S
                    e_["H"] = 0.0
                else:
                    e_["C"] = (S or 0) + shares
                    e_["H"] = ovhS if ovhS is not None else 0.0
                # D
                if lab in ("COE BP&T", "COE SA&D",
                           "COE Cyber Risk & Service Ops"):
                    e_["D"] = e_["C"] - e_["H"]
                elif lab == "TDD Cyber":
                    e_["D"] = e_["C"] - (uplift_fund or 0)
                elif lab == "EGI":
                    e_["D"] = 0.0
                else:
                    m = re.match(r"2\.(\d+)", TABOF[lab])
                    p1 = find_tab(sheets, "1.%s" % m.group(1))
                    pm, clash = ({}, []) if p1 is None else pct_map(f[p1],
                                                                   v[p1])
                    if clash:
                        gate.note(g, "%s support %% label clash: %s"
                                  % (p1, clash[:3]))
                    d = 0.0
                    for rr, lab2 in tb.grid:
                        s2 = num(tb.wv.cell(rr, 19).value) or 0.0
                        d += s2 * pm.get(norm(lab2), 0.0)
                    e_["D"] = d
                exp[lab] = e_

            offs = []
            tolbad = []
            for lab in LROWS:
                if lab not in rows or lab not in exp:
                    continue
                r = rows[lab]
                e_ = exp[lab]
                tog = num(lov.cell(r, 9).value)
                tog_eff = tog if tog is not None else (
                    1.0 if e_["H"] or e_["E"] or e_["F"] or e_["G"] else 0.0)
                e_["J"] = e_["E"] + e_["F"] + e_["G"] + e_["H"] * tog_eff
                e_["K"] = e_["D"] + e_["J"]
                # L off the 0.2 budget rows
                aliases = BUDMAP.get(lab, [lab])
                hits = []
                for a in aliases:
                    hits += cfg_match(a)
                e_["L"] = sum(cfg_E.get(b) or 0 for b in hits)
                e_["M"] = e_["K"] - e_["L"]
                e_["O"] = e_["C"] - e_["K"]
                for col, cc in (("C", 3), ("D", 4), ("E", 5), ("F", 6),
                                ("G", 7), ("H", 8), ("J", 10), ("K", 11),
                                ("L", 12), ("M", 13), ("O", 15)):
                    got = num(lov.cell(r, cc).value)
                    want = e_[col]
                    if got is None or abs(got - want) > TOL:
                        tolbad.append("%s!%s got %s want %.6f"
                                      % (lab, col, got, want))
            gate.sub(g, not tolbad,
                     "every cell of C D E F G H J K L M O recomputed from "
                     "REVIEW + 2.x + 1.x + 0.2 + Lists and tied to 1e-6: %d "
                     "off %s" % (len(tolbad), tolbad[:4]))

            # totals row ties
            if total_r_lo:
                tot_bad = []
                for col, cc in (("C", 3), ("D", 4), ("E", 5), ("F", 6),
                                ("G", 7), ("H", 8), ("J", 10), ("K", 11),
                                ("L", 12), ("M", 13), ("O", 15), ("P", 16)):
                    got = num(lov.cell(total_r_lo, cc).value)
                    want = sum(num(lov.cell(rows[lab], cc).value) or 0
                               for lab in labels if lab in rows)
                    if got is None or abs(got - want) > TOL:
                        tot_bad.append("%s got %s want %.6f"
                                       % (col, got, want))
                gate.sub(g, not tot_bad, "Total row sums its parts: %d off %s"
                         % (len(tot_bad), tot_bad[:4]))
                lt = num(lov.cell(total_r_lo, 12).value)
                gate.sub(g, lt is not None and abs(lt - L_TOTAL) <= TOL,
                         "L total %s (spec %.1f)" % (lt, L_TOTAL))
                e26 = num(cfgv.cell(cfg_total_r, 5).value) \
                    if cfg_total_r else None
                gate.sub(g, lt is not None and e26 is not None
                         and abs(lt - e26) <= TOL,
                         "L total ties to 0.2 allocated total %s" % e26)
            if budget_r_lo:
                lb = num(lov.cell(budget_r_lo, 12).value)
                e27 = num(cfgv.cell(cfg_budget_r, 5).value) \
                    if cfg_budget_r else None
                gate.sub(g, lb is not None and abs(lb - L_BUDGET) <= TOL
                         and (e27 is None or abs(lb - e27) <= TOL),
                         "Budget row L %s (spec %.1f, 0.2 budget %s)"
                         % (lb, L_BUDGET, e27))

            # toggles: cream with the 0-100% validation, default 100%
            need_tog = [lab for lab in LROWS if lab in rows and lab in exp
                        and (exp[lab]["H"] > TOL or lab not in NOSHARE)]
            notog = [lab for lab in need_tog
                     if num(lov.cell(rows[lab], 9).value) != 1]
            gate.sub(g, not notog,
                     "toggles at the 100%% default on the %d rows that carry "
                     "one: %d off %s" % (len(need_tog), len(notog), notog[:4]))
            uncream = [lab for lab in need_tog
                       if lof.cell(rows[lab], 9).fill.fgColor.rgb != CREAM]
            gate.sub(g, not uncream, "toggle cells cream: %d off %s"
                     % (len(uncream), uncream[:4]))
            egi_tog = lof.cell(rows["EGI"], 9).value if "EGI" in rows else None
            gate.sub(g, egi_tog in (None, ""), "EGI toggle blank (%r)"
                     % egi_tog)
            dvs = [dv for dv in lof.data_validations.dataValidation
                   if dv.type == "list"]
            cover = set()
            entries = None
            for dv in dvs:
                f1 = str(dv.formula1 or "")
                if "%" in f1 or "0.05" in f1 or "1" in f1:
                    for cr in dv.sqref.ranges:
                        if cr.min_col <= 9 <= cr.max_col:
                            cover |= set(range(cr.min_row, cr.max_row + 1))
                            if f1.startswith('"'):
                                entries = len(f1.strip('"').split(","))
            missing_dv = [lab for lab in need_tog
                          if rows[lab] not in cover]
            gate.sub(g, not missing_dv,
                     "a 0-100%% list validation covers every toggle: %d off "
                     "%s" % (len(missing_dv), missing_dv[:4]))
            gate.sub(g, entries in (None, 21),
                     "the toggle list runs 0%%,5%%,..,100%% (21 entries, got "
                     "%s)" % entries)
            if entries is None and not missing_dv:
                gate.note(g, "toggle validation uses a range source, entry "
                          "count not checked")
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # =========================================================== F exclusions
    g = gate.group("F", "exclusions")
    try:
        holds, bad_hold = 0, []
        bad_lever = []
        for t in tabs2:
            for ro in T[t].roles:
                lv = ro["lever"]
                if lv not in LEVERS:
                    bad_lever.append("%s!r%d %r" % (t, ro["row"], lv))
                if lv == "Hold":
                    holds += 1
                    if ro["G"] is None or abs(ro["G"]) > 1e-9:
                        bad_hold.append("%s!r%d G=%s" % (t, ro["row"],
                                                         ro["G"]))
        gate.sub(g, not bad_hold,
                 "every Hold role costs 0 after the lever (%d traced): %d "
                 "off %s" % (holds, len(bad_hold), bad_hold[:4]))
        gate.sub(g, lever_factor.get("Hold") == 0,
                 "Lists lever table prices Hold at 0 (%s)"
                 % lever_factor.get("Hold"))
        gate.sub(g, not bad_lever,
                 "every lever one of the four values, no Paused: %d off %s"
                 % (len(bad_lever), bad_lever[:4]))
        if LO in sheets and lo_ctx.get("rows", {}).get("EGI"):
            lov = v[LO]
            r = lo_ctx["rows"]["EGI"]
            nz = []
            for col, cc in (("D", 4), ("E", 5), ("F", 6), ("G", 7),
                            ("H", 8), ("J", 10), ("K", 11), ("L", 12),
                            ("M", 13)):
                x = num(lov.cell(r, cc).value)
                if x is None or abs(x) > TOL:
                    nz.append("%s=%s" % (col, x))
            gate.sub(g, not nz,
                     "EGI row zeros outside C O P: %d off %s"
                     % (len(nz), nz[:5]))
            cop = [num(lov.cell(r, cc).value) for cc in (3, 15, 16)]
            gate.note(g, "EGI row C/O/P: %s" % cop)
        else:
            gate.sub(g, False, "EGI row zeros: Lights On row not available")
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ============================================================== G netting
    g = gate.group("G", "netting")
    try:
        def netting_cell(tab_prefix):
            t = find_tab(sheets, tab_prefix)
            wf, wv = f[t], v[t]
            fr = None
            for r in range(1, wf.max_row + 1):
                if wf.cell(r, 2).value == "Funding":
                    fr = r
                    break
            if fr is None:
                return None, None
            met, neg = [], []
            for r in range(fr + 1, min(fr + 20, wf.max_row) + 1):
                b = str(wf.cell(r, 2).value or "")
                x = num(wv.cell(r, 3).value)
                if x is None:
                    continue
                if "met by portfolio" in b.lower():
                    met.append((r, x))
                elif x < -1e-9:
                    neg.append((r, x))
            if met:
                return met[0]
            if len(neg) == 1:
                return neg[0]
            return None, None

        r12, n12 = netting_cell("2.12")
        gate.sub(g, n12 is not None and bp_pot is not None
                 and abs(n12 + bp_pot) <= TOL,
                 "2.12 netting cell %s against -(BP pot %s)" % (n12, bp_pot))
        r13, n13 = netting_cell("2.13")
        gate.sub(g, n13 is not None and da_pot is not None
                 and abs(n13 + da_pot) <= TOL,
                 "2.13 netting cell %s against -(DA pot %s)" % (n13, da_pot))
        t31 = find_tab(sheets, "3.1")
        w31f, w31v = f[t31], v[t31]
        alcol = None
        hdr31 = None
        for r in range(1, 10):
            for c in range(2, w31f.max_column + 1):
                if str(w31f.cell(r, c).value or "").startswith(
                        "Cost after levers"):
                    alcol, hdr31 = c, r
        rows31 = {}
        for r in range((hdr31 or 4) + 1, w31f.max_row + 1):
            b = w31v.cell(r, 2).value
            if isinstance(b, str) and b:
                rows31.setdefault(b.strip(), r)
        for lab, prefix, potv in (("COE BP&T", "2.12", bp_pot),
                                  ("COE SA&D", "2.13", da_pot)):
            t = find_tab(sheets, prefix)
            S = T[t].total(19)
            want = None if (S is None or potv is None) else S - potv
            r31 = rows31.get(lab)
            got = num(w31v.cell(r31, alcol).value) if (r31 and alcol) else None
            gate.sub(g, got is not None and want is not None
                     and abs(got - want) <= TOL,
                     "3.1 %s after levers %s against netted %s"
                     % (lab, got, None if want is None else round(want, 6)))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ================================================================= H 0.2
    g = gate.group("H", "0.2 cyber")
    try:
        t15 = find_tab(sheets, "2.15")
        S15 = T[t15].total(19) if t15 in T else None
        want = None if (S15 is None or uplift_fund is None) \
            else S15 - uplift_fund
        got = cfg_F.get("TDD Cyber")
        gate.sub(g, got is not None and want is not None
                 and abs(got - want) <= TOL,
                 "0.2 TDD Cyber spend %s against 2.15 after levers %s less "
                 "the Lists uplift funding %s = %s"
                 % (got, S15, uplift_fund,
                    None if want is None else round(want, 6)))
        f26 = num(cfgv.cell(cfg_total_r, 6).value) if cfg_total_r else None
        colsum = sum(x for x in cfg_F.values() if x is not None)
        gate.sub(g, f26 is not None and abs(f26 - colsum) <= TOL,
                 "0.2 spend total %s against the F column %s"
                 % (f26, round(colsum, 6)))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ============================================================= I controls
    g = gate.group("I", "controls")
    try:
        ctrl_bad, ctrl_n = [], 0
        for ws in v.worksheets:
            for row in ws.iter_rows():
                for cl in row:
                    if not (isinstance(cl.value, str)
                            and cl.value.startswith("Control - ")):
                        continue
                    ctrl_n += 1
                    for cc in range(cl.column + 1, ws.max_column + 1):
                        x = num(ws.cell(cl.row, cc).value)
                        if x is not None:
                            if abs(x) > TOL:
                                ctrl_bad.append("%s!%s=%s" % (ws.title,
                                                              cl.coordinate,
                                                              x))
                            break
        gate.sub(g, ctrl_n > 10 and not ctrl_bad,
                 "every control row 0 (%d found): %d off %s"
                 % (ctrl_n, len(ctrl_bad), ctrl_bad[:4]))
        q = v["4.0 Data QA"]
        qa_n, qa_bad, qa_meta = 0, [], None
        for r in range(4, q.max_row + 1):
            x = num(q.cell(r, 5).value)
            if x is None:
                continue
            if str(q.cell(r, 2).value or "").strip() == "Checks failing":
                qa_meta = x        # the rollup counter, not a check itself
                continue
            if q.cell(r, 3).value is None and q.cell(r, 4).value is None:
                continue           # not a Model/Expected comparison row
            qa_n += 1
            if abs(x) > TOL:
                qa_bad.append((r, str(q.cell(r, 2).value or "")[:48], x))
        gate.sub(g, qa_n >= 8 and not qa_bad,
                 "4.0 Data QA all zero (%d checks, rollup %s): %d off %s"
                 % (qa_n, qa_meta, len(qa_bad), qa_bad[:4]))
        if LO in sheets:
            lof, lov = f[LO], v[LO]
            found = []
            for row in lof.iter_rows():
                for cl in row:
                    if isinstance(cl.value, str) \
                            and cl.value.startswith("Control"):
                        val = None
                        for cc in range(cl.column + 1, lof.max_column + 1):
                            x = num(lov.cell(cl.row, cc).value)
                            if x is not None:
                                val = x
                                break
                        col = cl.font.color
                        white = col is not None and col.rgb == WHITE
                        found.append((cl.coordinate, val, white))
            ok = bool(found) and all(vv is not None and abs(vv) <= TOL
                                     and w for _, vv, w in found)
            gate.sub(g, ok, "Lights On white control reads 0: %s" % found[:3])
        else:
            gate.sub(g, False, "Lights On control: tab missing")
        errs = [(ws.title, cl.coordinate) for ws in v.worksheets
                for row in ws.iter_rows() for cl in row
                if isinstance(cl.value, str)
                and any(e in cl.value for e in ERR)]
        gate.sub(g, not errs, "error cells: %d %s" % (len(errs), errs[:4]))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ============================================================== J hygiene
    g = gate.group("J", "hygiene")
    try:
        dash, endash, banned, creamf = [], [], [], []
        for ws in f.worksheets:
            model = ws.title not in SRC_TABS
            for row in ws.iter_rows():
                for cl in row:
                    tv = cl.value
                    isf = isinstance(tv, str) and tv.startswith("=")
                    fl = cl.fill
                    if fl is not None and fl.fgColor is not None \
                            and fl.fgColor.rgb == CREAM and isf:
                        creamf.append("%s!%s" % (ws.title, cl.coordinate))
                    if not model or not isinstance(tv, str):
                        continue
                    if tv.strip() in ("-", '="-"'):
                        dash.append("%s!%s" % (ws.title, cl.coordinate))
                    if isf:
                        continue
                    if ws.title == REVIEW and cl.column <= 28:
                        continue
                    if "–" in tv or "—" in tv:
                        endash.append("%s!%s" % (ws.title, cl.coordinate))
                    low = tv.lower()
                    if re.search(r"\bwaves?\b", low) \
                            or re.search(r"\bseats?\b", low) \
                            or " to projects" in low:
                        banned.append("%s!%s %r" % (ws.title, cl.coordinate,
                                                    tv[:40]))
        gate.sub(g, not dash, "dash literals on model tabs: %d %s"
                 % (len(dash), dash[:4]))
        gate.sub(g, not creamf, "cream-filled formula cells: %d %s"
                 % (len(creamf), creamf[:4]))
        gate.sub(g, not banned,
                 "banned words (wave, seat, ' to projects'): %d %s"
                 % (len(banned), banned[:3]))
        gate.sub(g, not endash, "en/em dashes: %d %s"
                 % (len(endash), endash[:4]))
        dashfmt = []
        for ws in f.worksheets:
            if ws.title in SRC_TABS:
                continue
            for row in ws.iter_rows():
                for cl in row:
                    s = fmt_sections(cl.number_format or "")
                    if len(s) >= 3 and DASHFMT.search(s[2]):
                        dashfmt.append("%s!%s" % (ws.title, cl.coordinate))
        gate.sub(g, not dashfmt, "dash-showing zero formats: %d %s"
                 % (len(dashfmt), dashfmt[:4]))
        noparen = []
        for t in tabs2:
            tb = T[t]
            if tb.total_r is None:
                continue
            for r in range(7, tb.total_r + 1):
                cl = tb.wf.cell(r, 18)
                if num(tb.wv.cell(r, 18).value) is None:
                    continue
                s = fmt_sections(cl.number_format or "")
                if len(s) >= 2 and "(" not in s[1]:
                    noparen.append("%s!R%d" % (t, r))
        t31 = find_tab(sheets, "3.1")
        if t31:
            for r in range(5, f[t31].max_row + 1):
                cl = f[t31].cell(r, 8)
                if num(v[t31].cell(r, 8).value) is None:
                    continue
                s = fmt_sections(cl.number_format or "")
                if len(s) >= 2 and "(" not in s[1]:
                    noparen.append("%s!H%d" % (t31, r))
        if LO in sheets and lo_ctx.get("rows"):
            for lab, r in lo_ctx["rows"].items():
                for cc, nm2 in ((13, "M"), (15, "O")):
                    cl = f[LO].cell(r, cc)
                    if num(v[LO].cell(r, cc).value) is None:
                        continue
                    s = fmt_sections(cl.number_format or "")
                    if len(s) < 2 or "(" not in s[1]:
                        noparen.append("%s!%s%d" % (LO, nm2, r))
        gate.sub(g, not noparen,
                 "negative-capable cells carry parens formats (2.x R, 3.1 "
                 "variance, Lights On M and O): %d off %s"
                 % (len(noparen), noparen[:4]))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # =========================================================== K protection
    g = gate.group("K", "protection")
    try:
        from openpyxl.utils.protection import hash_password
        want_hash = hash_password(PASSWORD).upper()
        unprot = [ws.title for ws in f.worksheets if not ws.protection.sheet]
        gate.sub(g, not unprot, "sheet protection enabled everywhere: %d off "
                 "%s" % (len(unprot), unprot[:4]))
        badpw = [ws.title for ws in f.worksheets
                 if ws.protection.sheet
                 and str(ws.protection.password or "").upper() != want_hash]
        gate.sub(g, not unprot and not badpw,
                 "every sheet carries the %s hash %s: %d unprotected, %d "
                 "wrong hash %s"
                 % (PASSWORD, want_hash, len(unprot), len(badpw), badpw[:4]))
        rv_unlocked = ["%s" % cl.coordinate for row in rvf.iter_rows()
                       for cl in row if cl.protection.locked is False]
        gate.sub(g, not rv_unlocked, "REVIEW fully locked: %d unlocked %s"
                 % (len(rv_unlocked), rv_unlocked[:4]))
        allow = set()
        for t in tabs2:
            for ro in T[t].roles:
                allow.add((t, ro["row"], 5))
        t11 = find_tab(sheets, "2.11")
        if t11 in T:
            for ro in T[t11].roles:
                allow.add((t11, ro["row"], 8))
        if LO in sheets and lo_ctx.get("rows"):
            for r in lo_ctx["rows"].values():
                allow.add((LO, r, 9))
        stray = []
        n_unlocked = 0
        for ws in f.worksheets:
            if ws.title == REVIEW:
                continue
            for row in ws.iter_rows():
                for cl in row:
                    if cl.protection.locked is not False:
                        continue
                    n_unlocked += 1
                    if (ws.title, cl.row, cl.column) in allow:
                        continue
                    fl = cl.fill
                    if fl is not None and fl.fgColor is not None \
                            and fl.fgColor.rgb == CREAM:
                        continue
                    stray.append("%s!%s" % (ws.title, cl.coordinate))
        gate.sub(g, not stray,
                 "unlocked cells (%d) only levers, 2.11 uplift, cream and "
                 "the toggles: %d stray %s"
                 % (n_unlocked, len(stray), stray[:5]))
        sec = f.security
        lock = bool(sec is not None and sec.lockStructure)
        whash = ""
        if sec is not None:
            whash = str(getattr(sec, "workbookPasswordHash", None)
                        or getattr(sec, "workbookPassword", None) or "")
        gate.sub(g, lock, "workbook structure locked (%r)" % lock)
        gate.sub(g, whash.upper() == want_hash,
                 "workbook password hash %r against %s" % (whash, want_hash))
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ============================================================== L role id
    g = gate.group("L", "role id")
    try:
        gate.sub(g, idcol is not None and idcol >= 52,
                 "REVIEW Role ID column beyond AY (col %s)" % idcol)
        if idcol:
            want_ids = ["R%04d" % (i + 1) for i in range(len(led))]
            got_ids = [str((rvv.cell(r, idcol).value
                            or rvf.cell(r, idcol).value) or "").strip()
                       for r in led]
            off = [i for i, (a, b) in enumerate(zip(got_ids, want_ids))
                   if a != b]
            gate.sub(g, not off,
                     "IDs run R0001..R%04d in row order: %d off (first %s)"
                     % (len(led), len(off),
                        [(led[i], got_ids[i], want_ids[i])
                         for i in off[:3]]))
        else:
            gate.sub(g, False, "IDs run R0001..: no Role ID column")
        noid, misres = [], []
        anchored = 0
        anchored_fg = 0
        # single-cell row-anchored refs only - a $B$2:$B$531 range inside an
        # INDEX/MATCH rewrite must not count
        pat = re.compile(r"!\$(?:B|C|AK)\$\d+(?!\d)(?!:)")
        pat_fg = re.compile(r"!\$(?:AA|Q)\$\d+(?!\d)(?!:)")
        for t in tabs2:
            for ro in T[t].roles:
                rid = ro["id"]
                if not (isinstance(rid, str)
                        and re.fullmatch(r"R\d{4}", rid.strip())):
                    noid.append("%s!A%d" % (t, ro["row"]))
                else:
                    lr = idmap.get(rid.strip())
                    if lr is None or norm(name_of.get(lr)) != norm(ro["name"]):
                        misres.append("%s!r%d %s -> %s vs %r"
                                      % (t, ro["row"], rid,
                                         name_of.get(lr or -1), ro["name"]))
                for cc in (2, 3, 4):
                    fx = T[t].wf.cell(ro["row"], cc).value
                    if isinstance(fx, str) and REVIEW in fx and pat.search(fx):
                        anchored += 1
                for cc in (6, 7):
                    fx = T[t].wf.cell(ro["row"], cc).value
                    if isinstance(fx, str) and REVIEW in fx \
                            and pat_fg.search(fx):
                        anchored_fg += 1
        gate.sub(g, not noid, "every 2.x role row carries an ID in col A: %d "
                 "without %s" % (len(noid), noid[:4]))
        gate.sub(g, not misres,
                 "every ID resolves to its own REVIEW row: %d off %s"
                 % (len(misres), misres[:3]))
        gate.sub(g, anchored == 0,
                 "row-anchored $B$/$C$/$AK$ refs left in role-row B/C/D: %d"
                 % anchored)
        gate.note(g, "row-anchored $AA$/$Q$ refs in role-row F/G: %d "
                  "(outside this check's scope)" % anchored_fg)
    except Exception as e:
        gate.sub(g, False, "exception: %r" % e)

    # ------------------------------------------------------- M  final ledger
    print("\nM     final ledger (print only)", flush=True)

    def p(label, val):
        print("        %-46s %s" % (label, val), flush=True)

    try:
        tot_after = sum(T[t].total(19) or 0 for t in tabs2)
        p("total cost after levers, 15 tabs ($m)", "%.6f" % tot_after)
        t31 = find_tab(sheets, "3.1")
        if t31:
            w31v = v[t31]
            for r in range(5, w31v.max_row + 1):
                b = str(w31v.cell(r, 2).value or "")
                if b.startswith("TDD total"):
                    p("3.1 TDD total after levers ($m)",
                      "%s" % w31v.cell(r, 9).value)
                if b.startswith("Total TDD cost including the GM"):
                    p("3.1 total incl GM layer after levers ($m)",
                      "%s" % w31v.cell(r, 9).value)
        if LO in sheets and lo_ctx.get("total"):
            lov = v[LO]
            ktot = num(lov.cell(lo_ctx["total"], 11).value)
            dtot = num(lov.cell(lo_ctx["total"], 4).value)
            jtot = num(lov.cell(lo_ctx["total"], 10).value)
            p("Lights On K total at default toggles ($m)",
              "%s" % (None if ktot is None else round(ktot, 6)))
            p("  over/under the 50.5 allocation ($m)",
              "%s" % (None if ktot is None else round(ktot - L_TOTAL, 6)))
            p("  over/under the 53.8 budget ($m)",
              "%s" % (None if ktot is None else round(ktot - L_BUDGET, 6)))
            p("support total D ($m)",
              "%s" % (None if dtot is None else round(dtot, 6)))
            p("overhead charged total J ($m)",
              "%s" % (None if jtot is None else round(jtot, 6)))
        else:
            p("Lights On totals", "tab absent")
        vac = collections.Counter()
        vac_cost = collections.defaultdict(float)
        lever_by_row = {}
        for t in tabs2:
            for ro in T[t].roles:
                if ro["ledger"] is not None:
                    lever_by_row[ro["ledger"]] = ro["lever"]
        for r in led:
            if ar_of[r] in ovh_labels and status_of[r] == "Vacant" \
                    and lever_by_row.get(r) == "Hire":
                vac[ar_of[r]] += 1
                vac_cost[ar_of[r]] += cost_of[r]
        p("vacant overheads on Hire (count / $m)",
          "%d / %.6f" % (sum(vac.values()),
                         sum(vac_cost.values()) / 1e6))
        for line in sorted(vac):
            p("  %s" % line, "%d / %.6f" % (vac[line],
                                            vac_cost[line] / 1e6))
        p("dial 1: hold the vacant overheads ($m)",
          "%.6f" % (sum(vac_cost.values()) / 1e6))
        p("dial 2: GM layer above lights on ($m)", "%s" % gm_cost)
        crso = [b for _, b in cfg_rows if b.startswith("COE - Cyber")]
        if crso:
            gvar = num(cfgv.cell([r for r, b in cfg_rows
                                  if b == crso[0]][0], 7).value)
            p("dial 3: CRSO spend against its 2.00 ($m over)",
              "%s" % (None if gvar is None else round(-gvar, 6)))
        p("  0.2 TDD Cyber spend (2.15 net of uplift, $m)",
          "%s" % cfg_F.get("TDD Cyber"))
        if cfg_var_r:
            p("dial 4: unallocated slice of the 53.8 ($m)",
              "%s" % cfgv.cell(cfg_var_r, 5).value)
    except Exception as e:
        p("ledger print aborted", repr(e))

    print("", flush=True)
    fails = gate.report()
    if fails:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
