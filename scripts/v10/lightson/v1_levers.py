#!/usr/bin/env python3
"""v1 - Stage v1: his lever edits.

  python3 v1_levers.py <in.xlsx> <out.xlsx>

His 11 lever edits (scratchpad/his_lever_edits.json, person-keyed multiset
changes) go onto the base. Rows on the 2.x FTE blocks are keyed by the col-B
formula ='REVIEW - Complete Role Mapping'!$B$n; identity is the cached
(name col B, role col C) pair, en dash and whitespace normalized, falling back
to the REVIEW row the key names when a cached value is absent. Where a group
of identical rows changes partially, the minimum number of rows change,
earliest on the tab first. Only the lever cell in column E moves - never
column H uplift values, statuses, or anything else.
Idempotent: handed its own output it copies it through untouched.
"""
import json, os, re, shutil, sys, unicodedata
from collections import Counter

for p in ("/home/user/anthropic-claude-code/scripts/v10",
          "/home/user/anthropic-claude-code/scripts/v10/update"):
    if p not in sys.path:
        sys.path.insert(0, p)

import openpyxl
from _xl import REVIEW, LEVERS, Log, load, save, ledger

HERE = os.path.dirname(os.path.abspath(__file__))
EDITS_JSON = os.path.join(os.path.dirname(HERE), "his_lever_edits.json")
HIS = ("/root/.claude/uploads/e550b440-3996-5abb-87e5-bafafe598f82/"
       "ea5ee9ff-Copy_of_TDD_Cost_Calc_300726_old_version_w_edits.xlsx")
ROLES = 528

KEYED = re.compile(r"^='" + re.escape(REVIEW) + r"'!\$B\$(\d+)$")


def norm(s):
    """His text, comparable: NFKC, dashes plain, whitespace single, casefold."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    for d in ("\u2013", "\u2014", "\u2212", "\u2011"):   # en, em, minus, nb hyphen
        s = s.replace(d, "-")
    return " ".join(s.split()).casefold()


# ------------------------------------------------- 2.x FTE block extraction


def _extract(wf, wv):
    """{(tab, name, role): [(sheet_row, lever), ...]} in tab order.

    wf is the formula load (the $B$n key and the lever value live there), wv
    the cached-value load (identity). An openpyxl-saved file carries no cached
    values, so identity falls back to the REVIEW row the key points at.
    """
    rvv, rvf = wv[REVIEW], wf[REVIEW]
    out = {}
    for ws in wf.worksheets:
        if not ws.title.startswith("2."):
            continue
        vs = wv[ws.title]
        for r in range(1, ws.max_row + 1):
            f = ws.cell(r, 2).value
            m = KEYED.match(f) if isinstance(f, str) else None
            if not m:
                continue
            rr = int(m.group(1))
            name = vs.cell(r, 2).value
            role = vs.cell(r, 3).value
            if name is None:
                name = rvv.cell(rr, 2).value or rvf.cell(rr, 2).value
            if role is None:
                role = rvv.cell(rr, 3).value or rvf.cell(rr, 3).value
            k = (ws.title, norm(name), norm(role))
            out.setdefault(k, []).append((r, ws.cell(r, 5).value))
    return out


def keyrows(path):
    return _extract(load(path), openpyxl.load_workbook(path, data_only=True))


def plan(rows, frm, to):
    """The minimum row edits taking the group's multiset from `frm` to `to`.

    rows come in tab order; the earliest rows still carrying a lever the `to`
    multiset has too many of are the ones that change. New values are handed
    out in the order his `to` list gives them.
    """
    need = Counter(frm) - Counter(to)
    have = Counter(frm)
    adds = []
    for v in to:
        if have[v] > 0:
            have[v] -= 1
        else:
            adds.append(v)
    picked = []
    for r, lev in rows:
        if need.get(lev, 0) > 0:
            need[lev] -= 1
            picked.append((r, lev))
    if len(picked) != len(adds):
        raise SystemExit("STOP: multiset plan does not balance: %r vs %r" % (picked, adds))
    return [(r, old, new) for (r, old), new in zip(picked, adds)]


# ------------------------------------------------------------------- script


def main(src, dst):
    log = Log("v1_levers")
    if not os.path.exists(EDITS_JSON):
        print("STOP: his edits file is missing: %s" % EDITS_JSON)
        raise SystemExit(2)
    with open(EDITS_JSON, encoding="utf-8") as fh:
        edits = json.load(fh)

    got = keyrows(src)
    plans, stops = [], []
    for e in edits:
        k = (e["tab"], norm(e["name"]), norm(e["role"]))
        rows = got.get(k, [])
        cur = sorted(l for _, l in rows)
        if cur == sorted(e["to"]) and cur != sorted(e["from"]):
            plans.append((e, []))                    # his edit already carried
        elif cur == sorted(e["from"]):
            plans.append((e, plan(rows, e["from"], e["to"])))
        else:
            stops.append("%s / %s / %s: rows carry %r, his json says %r -> %r"
                         % (e["tab"], e["name"], e["role"], cur,
                            sorted(e["from"]), sorted(e["to"])))
    if stops:
        print("STOP: input matches neither the base nor his edited state:")
        for s in stops:
            print("  " + s)
        raise SystemExit(2)

    expected = set()
    if all(not ch for _, ch in plans):
        print("input already carries his 11 lever edits - copying through untouched")
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        selfcheck(src, dst, edits, expected)
        return

    log.head("V1  his 11 lever edits, person-keyed, col E only")
    wb = load(src)
    for e, changes in plans:
        keytxt = "%s / %s" % (e["name"], e["role"])
        if not changes:
            log.note("V1", "%s / %s already at his multiset" % (e["tab"], keytxt))
            continue
        ws = wb[e["tab"]]
        for r, old, new in changes:
            if new not in LEVERS:
                print("STOP: %r is not one of the four levers" % new)
                raise SystemExit(2)
            if ws.cell(r, 5).value != old:
                print("STOP: %s!E%d moved under us (%r)"
                      % (e["tab"], r, ws.cell(r, 5).value))
                raise SystemExit(2)
            ws.cell(r, 5).value = new
            expected.add((e["tab"], "E%d" % r))
            log("V1", "%s!E%d" % (e["tab"], r), "%r -> %r (%s)" % (old, new, keytxt))
    log.note("V1", "column H uplift values untouched (his file predates them)")

    save(wb, dst)
    log.tail()
    print("wrote", dst)
    selfcheck(src, dst, edits, expected)


# --------------------------------------------------------------- self-check


def cellmap(path):
    """{(sheet, coord): value} over every non-empty cell, formulas as text."""
    w = openpyxl.load_workbook(path, data_only=False)
    m = {}
    for ws in w.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    m[(ws.title, c.coordinate)] = c.value
    return m, list(w.sheetnames)


def selfcheck(src, dst, edits, expected):
    print("== V1 self-check", flush=True)
    fails = []

    def check(name, ok, extra=""):
        print("%s  %s%s" % ("PASS" if ok else "FAIL", name,
                            " " + extra if extra else ""), flush=True)
        if not ok:
            fails.append(name)

    wf = load(dst)
    wv = openpyxl.load_workbook(dst, data_only=True)
    keys = _extract(wf, wv)

    # 1. every key sits at his multiset
    for e in edits:
        k = (e["tab"], norm(e["name"]), norm(e["role"]))
        cur = sorted(l for _, l in keys.get(k, []))
        check("lever multiset %s / %s / %s" % (e["tab"], e["name"], e["role"]),
              cur == sorted(e["to"]), "= %r" % cur)

    # 2. person-keyed re-diff against his file, zero differences over the 11
    if os.path.exists(HIS):
        his = keyrows(HIS)
        bad = []
        for e in edits:
            k = (e["tab"], norm(e["name"]), norm(e["role"]))
            if sorted(l for _, l in keys.get(k, [])) != \
                    sorted(l for _, l in his.get(k, [])):
                bad.append(k)
        check("person-keyed re-diff vs his file over %d keys" % len(edits),
              not bad, "zero differences" if not bad else "%r differ" % bad)
    else:
        print("note  his upload not on disk here; his multisets checked from the json above",
              flush=True)

    # 3. the ledger is untouched
    n_src = len(ledger(load(src)[REVIEW]))
    n_dst = len(ledger(wf[REVIEW]))
    check("total role count %d unchanged" % ROLES,
          n_src == n_dst == ROLES, "(src %d, dst %d)" % (n_src, n_dst))

    # 4. no other cell changed: full-workbook formula+value diff vs input
    ma, sa = cellmap(src)
    mb, sb = cellmap(dst)
    check("sheet list unchanged", sa == sb)
    changed = sorted(k for k in set(ma) | set(mb) if ma.get(k) != mb.get(k))
    for sheet, coord in changed:
        print("      %s!%s %r -> %r" % (sheet, coord, ma.get((sheet, coord)),
                                        mb.get((sheet, coord))), flush=True)
    check("changed cells are exactly the edited lever cells",
          set(changed) == expected,
          "(%d cells)" % len(changed) if set(changed) == expected else
          "unexpected %r / missing %r" % (sorted(set(changed) - expected),
                                          sorted(expected - set(changed))))

    if fails:
        print("\nv1_levers self-check: FAIL (%d)" % len(fails), flush=True)
        raise SystemExit(2)
    print("\nv1_levers self-check: all PASS", flush=True)


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
