#!/usr/bin/env python3
"""v2 - Stage v2: role mapping corrections.

  python3 v2_mapping.py <in.xlsx> <out.xlsx>

On v1 -> v2, in REVIEW - Complete Role Mapping (join by name or title, never
row number):
  M1  Ed Tacey's overhead coding reverts to the exact original 30/07 values
      for 'Leadership', 'Overhead line', 'Squad or overhead line'.
  M2  the vacant Delivery Assurance / Delivery Excellence Manager rows recode
      'Overhead line' -> 'Delivery Manager'.
  M3  Viren Khatri retags into TDD Group Functions (Division (GM), Portfolio,
      MTab copied from Anjali Patra), and his role row moves from the 2.1 FTE
      block (EGI TDD group) to the 2.4 FTE block (EGI TDD group) with
      shift_rows so every workbook reference shifts with Excel semantics.
  M4  part-time roles (FTE < 1, cost still at full rate) get 'Full Cost AUD'
      multiplied by FTE.
  M5  Ed Tacey's 2.2 Customer block row follows his reverted coding: out of
      the Head of Technology group, into a Leadership group in the
      No-archetype block area, and the Leadership grid row rewires from its
      static zeros onto that group (shift_rows for the delete and the
      insert, the u6 D7 pattern, so every SUM/COUNTIF range and validation
      adjusts with Excel semantics).

Idempotent: handed its own output it copies through untouched.  No recalc
here; the self-check emulates the SUMIFS/COUNTIFS logic from raw cells, never
cached numbers.  Ends with a labelled self-check block printing PASS/FAIL
lines; exits non-zero on any FAIL.
"""
import os
import re
import shutil
import sys

sys.path.insert(0, "/home/user/anthropic-claude-code/scripts/v10")
sys.path.insert(0, "/home/user/anthropic-claude-code/scripts/v10/update")

import openpyxl
from _xl import (REVIEW, LEVERS, Log, load, save, shift_rows, rewrite_refs,
                 copy_style, ledger, _scan, _sheet_of)

ORIG = ("/root/.claude/uploads/e550b440-3996-5abb-87e5-bafafe598f82/"
        "d4b6cba7-TDD_Cost_Calc_300726.xlsx")
T21 = "2.1 Ampol Retail"
T22 = "2.2 Customer"
T24 = "2.4 TDD Group Functions"
T33 = "3.3 Squad Actuals to Archetype"

# the three overhead-coding columns the Ed Tacey ruling names, plus the
# person-keyed fields the Viren ruling names - resolved by header, never index
ED_COLS = ("Leadership", "Overhead line", "Squad or overhead line")
VIREN_COLS = ("Division (GM)", "Portfolio", "MTab")

BREF = re.compile(r"^='" + re.escape(REVIEW) + r"'!\$B\$(\d+)$")


def norm(s):
    return " ".join(str(s).split())


def headers(ws):
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            out[norm(v)] = c
    return out


def need(cols, names, where):
    missing = [n for n in names if n not in cols]
    if missing:
        print("STOP: %s is missing headers %r" % (where, missing))
        raise SystemExit(2)


def rows_where(ws, col, value):
    return [r for r in range(2, ws.max_row + 1)
            if norm(ws.cell(r, col).value or "") == value]


def one_row(ws, col, value, what):
    hits = rows_where(ws, col, value)
    if len(hits) != 1:
        print("STOP: expected exactly one %s, found rows %r" % (what, hits))
        raise SystemExit(2)
    return hits[0]


def label_row(ws, label):
    hits = [r for r in range(1, ws.max_row + 1)
            if norm(ws.cell(r, 2).value or "") == label]
    if len(hits) != 1:
        print("STOP: %s: expected one row labelled %r, found %r"
              % (ws.title, label, hits))
        raise SystemExit(2)
    return hits[0]


def grid_egi_row(ws):
    """2.1's EGI TDD grid row: between the directly funded labels, or None."""
    a = label_row(ws, "Directly funded programs and platforms")
    b = label_row(ws, "Directly funded total")
    hits = [r for r in range(a + 1, b)
            if norm(ws.cell(r, 2).value or "") == "EGI TDD"]
    if len(hits) > 1:
        print("STOP: 2.1 carries %d EGI TDD grid rows" % len(hits))
        raise SystemExit(2)
    return hits[0] if hits else None


def block_ref_rows(ws):
    """{sheet row -> REVIEW row} for every FTE-block col-B REVIEW ref."""
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str):
            m = BREF.match(v)
            if m:
                out[r] = int(m.group(1))
    return out


def group_span(ws, hdr_row):
    """(first, last) role rows a block group header's COUNTIF covers."""
    c = str(ws.cell(hdr_row, 3).value or "")
    m = re.search(r"COUNTIF\(\$B\$(\d+):\$B\$(\d+)", c)
    if not m:
        print("STOP: %s row %d is not a group header" % (ws.title, hdr_row))
        raise SystemExit(2)
    return int(m.group(1)), int(m.group(2))


def block_row_of(ws, ledger_row):
    """The FTE-block row carrying a ledger row: col-B ref, or the D-column
    $AK$ idiom u6's D7 move keyed on.  None when absent."""
    for r, rr in block_ref_rows(ws).items():
        if rr == ledger_row:
            return r
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 4).value
        if isinstance(d, str) and REVIEW in d \
                and d.endswith("$AK$%d" % ledger_row):
            return r
    return None


def group_header_above(ws, r):
    """The nearest block group header at or above row r, or None."""
    for h in range(r - 1, 1, -1):
        c3 = ws.cell(h, 3).value
        if isinstance(c3, str) and c3.startswith("=COUNTIF("):
            return h
    return None


def grid_leadership_row(ws):
    """2.2's Leadership grid row, inside the No-archetype section."""
    a = label_row(ws, "No archetype in 1.x tabs")
    b = label_row(ws, "No archetype total")
    hits = [r for r in range(a + 1, b)
            if norm(ws.cell(r, 2).value or "") == "Leadership"]
    if len(hits) != 1:
        print("STOP: 2.2 No-archetype section carries %d Leadership rows"
              % len(hits))
        raise SystemExit(2)
    return hits[0]


# ------------------------------------------------ REVIEW formula emulation
# The self-check never trusts cached values: it recomputes the derived
# columns (AQ/AP/AR/AJ/AT) and Full Cost AUD from raw cells, honouring any
# literal this stage wrote, exactly as Excel will after the recalc.

def emulator(wb):
    rv = wb[REVIEW]
    cols = headers(rv)
    need(cols, ("Name", "Position Title", "Department (GM-1)", "Platform",
                "Squad", "Portfolio", "FTE", "Full Cost AUD", "day rate",
                "FTE Base AUD", "STI", "Payroll", "Pensions", "medical",
                "CPI", "MTab", "Leadership", "Overhead line",
                "Squad or overhead line", "Agreed cost override (AUD)"),
         "REVIEW")
    L = wb["Lists"]
    AF = [L.cell(r, 32).value for r in range(2, 9)]
    W2X = {norm(L.cell(r, 23).value): L.cell(r, 24).value
           for r in range(2, 22) if L.cell(r, 23).value is not None}
    T2U = {norm(L.cell(r, 20).value): L.cell(r, 21).value
           for r in range(2, 22) if L.cell(r, 20).value is not None}
    EXC = {norm(L.cell(r, 40).value): (L.cell(r, 41).value,
                                       L.cell(r, 42).value)
           for r in range(2, 23) if L.cell(r, 40).value is not None}
    ag16 = L.cell(16, 33).value

    def N(x):
        return x if isinstance(x, (int, float)) else 0

    def is_f(v):
        return isinstance(v, str) and v.startswith("=")

    def cell(r, name):
        return rv.cell(r, cols[name]).value

    def full_cost_formula(r):
        """The AA formula's own arithmetic from raw typed inputs."""
        au, s = cell(r, "Agreed cost override (AUD)"), cell(r, "day rate")
        u = cell(r, "FTE Base AUD")
        v, w = cell(r, "STI"), cell(r, "Payroll")
        x, y, z = cell(r, "Pensions"), cell(r, "medical"), cell(r, "CPI")
        if N(au) > 0:
            return au
        if N(s) > 0:
            if not isinstance(ag16, (int, float)):
                print("STOP: Lists!AG16 is not a number; day-rate row %d" % r)
                raise SystemExit(2)
            return s * ag16 * (1 + N(z))
        return N(u) * (1 + N(v) + N(w) + N(x) + N(z)) + N(y)

    def role(r):
        b = norm(cell(r, "Name") or "")
        if not b:
            return None
        c = norm(cell(r, "Position Title") or "")
        cl = c.lower()
        g = norm(cell(r, "Department (GM-1)") or "")
        i = norm(cell(r, "Portfolio") or "")
        j = norm(cell(r, "Platform") or "")
        k = norm(cell(r, "Squad") or "")
        aq_c = cell(r, "Leadership")
        aq = (1 if (k == "Leadership" or j == "Leadership") else 0) \
            if is_f(aq_c) else N(aq_c)
        if aq == 1:
            ap = "Leadership"
        elif k in W2X:
            ap = W2X[k]
        elif k == "" or k.lower() == "na":
            ap = g if g else "Unassigned"
        else:
            ap = k
        ar_c = cell(r, "Overhead line")
        if not is_f(ar_c):
            ar = ar_c
        elif ap in AF:
            ar = ap
        elif "head of " in cl:
            ar = "Head of Technology"
        elif "tdd bp" in cl:
            ar = "Business Partner"
        elif "domain architect" in cl or "enterprise architect" in cl:
            ar = "Domain Architect"
        elif "delivery man" in cl:
            ar = "Delivery Manager"
        elif ("technology manager" in cl or "technology manger" in cl
              or "tech manager" in cl):
            ar = "Technology Manager"
        else:
            ar = "Squad"
        key = "%s | %s" % (b, c)
        aj_c = cell(r, "MTab")
        if not is_f(aj_c):
            aj = aj_c
        elif key in EXC and EXC[key][0] is not None:
            aj = EXC[key][0]
        else:
            aj = T2U.get(i, i)
        at_c = cell(r, "Squad or overhead line")
        if not is_f(at_c):
            at = at_c
        elif key in EXC and EXC[key][1] is not None:
            at = EXC[key][1]
        elif str(aj).startswith("COE") or aj == "EGI":
            at = ap
        else:
            at = ar if ar != "Squad" else ap
        aa_c = cell(r, "Full Cost AUD")
        aa = aa_c if not is_f(aa_c) else full_cost_formula(r)
        return dict(row=r, name=b, title=c, ar=ar, aj=aj, at=at,
                    aa=N(aa), aa_unscaled=full_cost_formula(r),
                    aa_literal=not is_f(aa_c), fte=cell(r, "FTE"))

    return [d for d in (role(r) for r in range(2, rv.max_row + 1)) if d]


# ---------------------------------------------------------------- the edits

def orig_ed_values():
    """Ed Tacey's cached coding out of the original 30/07 file, by header."""
    owb = openpyxl.load_workbook(ORIG, data_only=True)
    orv = owb[REVIEW]
    ocols = headers(orv)
    need(ocols, ("Name",) + ED_COLS, "original 30/07 REVIEW")
    r = one_row(orv, ocols["Name"], "Ed Tacey", "Ed Tacey in the 30/07 file")
    vals = {}
    for name in ED_COLS:
        v = orv.cell(r, ocols[name]).value
        if v is None or (isinstance(v, str) and v.startswith("=")):
            print("STOP: 30/07 cached %r for Ed Tacey is %r" % (name, v))
            raise SystemExit(2)
        vals[name] = int(v) if isinstance(v, bool) else v
    return vals


def find_people(rv, cols):
    ed = one_row(rv, cols["Name"], "Ed Tacey", "Ed Tacey")
    viren = one_row(rv, cols["Name"], "Viren Khatri", "Viren Khatri")
    anjali = one_row(rv, cols["Name"], "Anjali Patra", "Anjali Patra")
    dam = one_row(rv, cols["Position Title"], "Delivery Assurance Manager",
                  "Delivery Assurance Manager")
    dem = one_row(rv, cols["Position Title"], "Delivery Excellence Manager",
                  "Delivery Excellence Manager")
    return ed, viren, anjali, dam, dem


def part_time_rows(rv, cols):
    """[(row, fte, cell)] for FTE < 1 - the ruling's criterion."""
    out = []
    for r, _, _, _ in ledger(rv):
        fte = rv.cell(r, cols["FTE"]).value
        if isinstance(fte, (int, float)) and not isinstance(fte, bool) \
                and fte < 1:
            out.append((r, fte, rv.cell(r, cols["Full Cost AUD"])))
    return out


def applied(wb, rv, cols, ed, viren, anjali, dam, dem, want):
    """True only when every M1..M4 edit is already in the file."""
    lit = lambda v: not (isinstance(v, str) and v.startswith("="))
    for name in ED_COLS:
        v = rv.cell(ed, cols[name]).value
        if not lit(v) or v != want[name]:
            return False
    for r in (dam, dem):
        if rv.cell(r, cols["Overhead line"]).value != "Delivery Manager":
            return False
    for name in ("Division (GM)", "Portfolio"):
        if rv.cell(viren, cols[name]).value != \
                rv.cell(anjali, cols[name]).value:
            return False
    if grid_egi_row(wb[T21]) is not None:
        return False
    if viren in block_ref_rows(wb[T21]).values():
        return False
    if viren not in block_ref_rows(wb[T24]).values():
        return False
    for r, fte, cell in part_time_rows(rv, cols):
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return False
    c22 = wb[T22]
    er = block_row_of(c22, ed)
    if er is None:
        return False
    hdr = group_header_above(c22, er)
    if hdr is None or norm(c22.cell(hdr, 2).value or "") != "Leadership":
        return False
    if not str(c22.cell(grid_leadership_row(c22), 6).value or "") \
            .startswith("=COUNTIF"):
        return False
    return True


def main(src, dst):
    log = Log("v2_mapping")
    wb = load(src)
    if REVIEW not in wb.sheetnames:
        print("STOP: no %r sheet in %s" % (REVIEW, src))
        raise SystemExit(2)
    rv = wb[REVIEW]
    cols = headers(rv)
    need(cols, ("Name", "Position Title", "FTE", "Full Cost AUD",
                "Agreed cost override (AUD)") + ED_COLS + VIREN_COLS,
         "REVIEW")
    if len(ledger(rv)) != 528:
        print("STOP: input carries %d roles, not the 528 this stage expects"
              % len(ledger(rv)))
        raise SystemExit(2)
    ed, viren, anjali, dam, dem = find_people(rv, cols)
    want = orig_ed_values()

    if applied(wb, rv, cols, ed, viren, anjali, dam, dem, want):
        print("input already carries the v2 role mapping corrections - "
              "copying through untouched")
        shutil.copy(src, dst)
        self_check(dst)
        log.tail()
        print("wrote", dst)
        return

    # ------------------------------------------------------------------ M1
    log.head("M1  Ed Tacey's overhead coding reverts to the original 30/07")
    for name in ED_COLS:
        cell = rv.cell(ed, cols[name])
        if cell.value != want[name] or (isinstance(cell.value, str)
                                        and cell.value.startswith("=")):
            was = ("formula" if isinstance(cell.value, str)
                   and cell.value.startswith("=") else repr(cell.value))
            log("M1", "REVIEW!%s%d (%s)" % (cell.column_letter, ed, name),
                "%s -> %r (his 30/07 value)" % (was, want[name]))
            cell.value = want[name]
    log.note("M1", "3.2 reads these via SUMIFS/COUNTIFS on the Overhead line "
                   "column - no other edit needed")

    # ------------------------------------------------------------------ M2
    log.head("M2  the two vacant Delivery Manager recodes")
    for r, ttl in ((dam, "Delivery Assurance Manager"),
                   (dem, "Delivery Excellence Manager")):
        cell = rv.cell(r, cols["Overhead line"])
        if cell.value != "Delivery Manager":
            log("M2", "REVIEW!%s%d (%s)" % (cell.column_letter, r,
                                            rv.cell(r, cols["Name"]).value),
                "%s -> 'Delivery Manager' (%s)"
                % ("formula" if isinstance(cell.value, str)
                   and str(cell.value).startswith("=") else repr(cell.value),
                   ttl))
            cell.value = "Delivery Manager"

    # ------------------------------------------------------------------ M3
    log.head("M3  Viren Khatri retags into TDD Group Functions and moves "
             "2.1 -> 2.4")
    for name in ("Division (GM)", "Portfolio"):
        src_v = rv.cell(anjali, cols[name]).value
        cell = rv.cell(viren, cols[name])
        if cell.value != src_v:
            log("M3", "REVIEW!%s%d (%s)" % (cell.column_letter, viren, name),
                "%r -> %r (copied from Anjali Patra)" % (cell.value, src_v))
            cell.value = src_v
    # MTab: Anjali's cell row-remapped onto Viren's row; both carry the same
    # row-relative formula, so this is a write only if they ever diverge
    aj_src = rv.cell(anjali, cols["MTab"]).value
    aj_cell = rv.cell(viren, cols["MTab"])
    if isinstance(aj_src, str) and aj_src.startswith("="):
        def remap(sh, c1, r1, c2, r2):
            f = lambda r: viren if (sh == REVIEW and r == anjali) else r
            return (c1, f(r1), c2, None if r2 is None else f(r2))
        aj_new = rewrite_refs(aj_src, REVIEW, remap)
    else:
        aj_new = aj_src
    if aj_cell.value != aj_new:
        log("M3", "REVIEW!%s%d (MTab)" % (aj_cell.column_letter, viren),
            "repointed to Anjali Patra's coding, row-remapped")
        aj_cell.value = aj_new
    else:
        log.note("M3", "MTab already carries Anjali Patra's coding "
                       "(same row-relative formula); Division (GM) and "
                       "Portfolio now make it read TDD Group Functions")

    # the 2.1 -> 2.4 row move, every step derived from the file
    g21 = wb[T21]
    refs21 = block_ref_rows(g21)
    vrows = [r for r, rr in refs21.items() if rr == viren]
    egi_grid = grid_egi_row(g21)
    lever = None

    if vrows or egi_grid is not None:
        if len(vrows) != 1 or egi_grid is None:
            print("STOP: 2.1 is part-moved: block rows %r, grid row %r"
                  % (vrows, egi_grid))
            raise SystemExit(2)
        vb = vrows[0]
        hb = vb - 1
        if norm(g21.cell(hb, 2).value or "") != "EGI TDD":
            print("STOP: the row above Viren's 2.1 row is %r, not the EGI "
                  "TDD group header" % g21.cell(hb, 2).value)
            raise SystemExit(2)
        a, b = group_span(g21, hb)
        if (a, b) != (vb, vb):
            print("STOP: 2.1's EGI TDD group spans %d:%d - not only Viren"
                  % (a, b))
            raise SystemExit(2)
        lever = g21.cell(vb, 5).value
        if lever not in LEVERS:
            print("STOP: Viren's 2.1 lever is %r" % lever)
            raise SystemExit(2)

        # 3.3 carries every 2.1 grid row; its EGI TDD line leaves with it
        r33 = wb[T33]
        t33 = [r for r in range(2, r33.max_row + 1)
               if str(r33.cell(r, 4).value or "")
               == "='%s'!$B$%d" % (T21, egi_grid)]
        if len(t33) == 1:
            if (norm(r33.cell(t33[0], 2).value or "") != "Ampol Retail"
                    or norm(r33.cell(t33[0], 3).value or "")
                    != "Directly funded"):
                print("STOP: 3.3 row %d is not the Ampol Retail directly "
                      "funded EGI TDD line" % t33[0])
                raise SystemExit(2)
            shift_rows(wb, T33, t33[0], -1)
            log("M3", "%s row %d" % (T33, t33[0]),
                "Ampol Retail / Directly funded / EGI TDD line deleted "
                "with 2.1's grid row")
        elif t33:
            print("STOP: %d rows on 3.3 reference 2.1's EGI TDD grid row"
                  % len(t33))
            raise SystemExit(2)

        # single-cell refs into the doomed rows, from surviving cells:
        # the 4.0 Data QA identity carries a +N('2.1 ...'!$N$grid) term
        doomed = {egi_grid, hb, vb}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    if ws.title == T21 and cell.row in doomed:
                        continue
                    hit = None
                    for s, e, m in _scan(v):
                        if _sheet_of(m, ws.title) != T21 or m.group("c2"):
                            continue
                        if int(m.group("r1").lstrip("$")) in doomed:
                            hit = v[s:e]
                    if hit is None:
                        continue
                    term = "+N('%s'!$N$%d)" % (T21, egi_grid)
                    if term in v:
                        cell.value = v.replace(term, "")
                        log("M3", "%s!%s" % (ws.title, cell.coordinate),
                            "the departed grid row's %s term dropped "
                            "(2.1's archetype total no longer carries "
                            "EGI TDD)" % term)
                    else:
                        print("STOP: %s!%s still reads the doomed 2.1 row "
                              "%s: %r" % (ws.title, cell.coordinate, hit, v))
                        raise SystemExit(2)

        # the grid row leaves; the directly funded total's ranges clamp to
        # AmPOS + EGI Retail on their own
        shift_rows(wb, T21, egi_grid, -1)
        log("M3", "%s row %d" % (T21, egi_grid),
            "EGI TDD grid row deleted; directly funded total now sums "
            "AmPOS + EGI Retail")

        # the block row and its now-empty group header leave together
        g21 = wb[T21]
        refs21 = block_ref_rows(g21)
        vb = [r for r, rr in refs21.items() if rr == viren][0]
        hb = vb - 1
        if norm(g21.cell(hb, 2).value or "") != "EGI TDD":
            print("STOP: EGI TDD header lost during the grid delete")
            raise SystemExit(2)
        shift_rows(wb, T21, hb, -2)
        log("M3", "%s rows %d-%d" % (T21, hb, vb),
            "Viren's FTE-block row and the now-empty EGI TDD group "
            "header deleted")

    # the directly funded total's presence guard counts two rows now
    dft = label_row(wb[T21], "Directly funded total")
    for c in range(2, 21):
        cell = wb[T21].cell(dft, c)
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            m = re.search(r"COUNT\(N(\d+):N(\d+)\)=3", v)
            if m and int(m.group(2)) - int(m.group(1)) == 1:
                cell.value = v.replace(m.group(0), m.group(0)[:-1] + "2")
                log("M3", "%s!%s%d" % (T21, cell.column_letter, dft),
                    "variance guard COUNT(...)=3 -> =2 (two directly "
                    "funded rows remain)")

    # 2.4: a role row into the EGI TDD group's FTE block
    g24 = wb[T24]
    if viren not in block_ref_rows(g24).values():
        if lever is None:
            print("STOP: Viren sits on neither 2.1 nor 2.4 - the input is "
                  "not a state this stage knows")
            raise SystemExit(2)
        hdr24 = [r for r in range(2, g24.max_row + 1)
                 if norm(g24.cell(r, 2).value or "") == "EGI TDD"
                 and str(g24.cell(r, 3).value or "").startswith("=COUNTIF")]
        if len(hdr24) != 1:
            print("STOP: 2.4 EGI TDD group header not found: %r" % hdr24)
            raise SystemExit(2)
        a, b = group_span(g24, hdr24[0])
        shift_rows(wb, T24, b, 1)          # inside the group, so every
        g24 = wb[T24]                      # $a:$b range takes the new row
        nb = b + 1                         # the style and formula pattern
        rr_n = block_ref_rows(g24).get(nb)
        if rr_n is None:
            print("STOP: 2.4 neighbour row %d lost its REVIEW ref" % nb)
            raise SystemExit(2)

        def remap(sh, c1, r1, c2, r2):
            def f(r):
                if sh == REVIEW and r == rr_n:
                    return viren
                if sh == T24 and r == nb:
                    return b
                return r
            return (c1, f(r1), c2, None if r2 is None else f(r2))

        for c in (2, 3, 4, 6, 7):
            src_c = g24.cell(nb, c)
            dst_c = g24.cell(b, c)
            copy_style(src_c, dst_c)
            dst_c.value = rewrite_refs(str(src_c.value), T24, remap)
        copy_style(g24.cell(nb, 5), g24.cell(b, 5))
        g24.cell(b, 5).value = lever
        if g24.row_dimensions[nb].height is not None:
            g24.row_dimensions[b].height = g24.row_dimensions[nb].height
        log("M3", "%s row %d" % (T24, b),
            "Viren's role row into the EGI TDD group (pattern of the "
            "adjacent row, repointed to his REVIEW row; lever %r carried)"
            % lever)

        from openpyxl.worksheet.cell_range import MultiCellRange, CellRange
        for dv in g24.data_validations.dataValidation:
            if dv.formula1 and "Filled" in str(dv.formula1):
                covered = any(cr.min_col <= 5 <= cr.max_col
                              and cr.min_row <= b <= cr.max_row
                              for cr in dv.sqref.ranges)
                if not covered:
                    dv.sqref = MultiCellRange(list(dv.sqref.ranges) +
                                              [CellRange(min_col=5, max_col=5,
                                                         min_row=b,
                                                         max_row=b)])
                    log("M3", "%s!E%d" % (T24, b),
                        "lever dropdown extended over the new row")
        log.note("M3", "2.4's EGI TDD grid row is SUMIFS/COUNTIFS-driven "
                       "(MTab + squad) - it picks him up on recalc, nothing "
                       "hardcoded")

    # ------------------------------------------------------------------ M4
    log.head("M4  part-time roles: Full Cost AUD scales by FTE")
    roles = {d["row"]: d for d in emulator(wb)}
    pt = part_time_rows(rv, cols)
    scaled, skipped, total = [], [], 0.0
    for r, fte, cell in pt:
        d = roles[r]
        if not (isinstance(cell.value, str) and cell.value.startswith("=")):
            skipped.append((r, d["name"], "already a typed value"))
            continue
        au = rv.cell(r, cols["Agreed cost override (AUD)"]).value
        if isinstance(au, (int, float)) and au > 0:
            skipped.append((r, d["name"], "agreed cost override in place"))
            continue
        before = d["aa_unscaled"]
        after = round(before * fte, 2)
        cell.value = after
        total += before - after
        scaled.append((d["name"], fte, before, after))
        log("M4", "REVIEW!%s%d" % (cell.column_letter, r),
            "%s FTE %.2f: %.2f -> %.2f" % (d["name"], fte, before, after))
    for r, nm, why in skipped:
        log.note("M4", "row %d %s left alone (%s)" % (r, nm, why))
    log.note("M4", "%d part-time roles scaled, cost drops %.2f (%.4fm); "
                   "the ruling expected roughly 7 and roughly 0.36"
             % (len(scaled), total, total / 1e6))

    # ------------------------------------------------------------------ M5
    log.head("M5  Ed Tacey's 2.2 block row follows his coding into the "
             "Leadership group")
    c22 = wb[T22]
    erow = block_row_of(c22, ed)
    if erow is None:
        print("STOP: no 2.2 Customer block row carries Ed Tacey (REVIEW %d)"
              % ed)
        raise SystemExit(2)
    ghdr = group_header_above(c22, erow)
    glabel = norm(c22.cell(ghdr, 2).value or "") if ghdr else ""
    if glabel == "Leadership":
        log.note("M5", "Ed Tacey's 2.2 row already sits in the Leadership "
                       "group - nothing to move")
    elif glabel != "Head of Technology":
        print("STOP: Ed Tacey's 2.2 row sits under %r - not a state this "
              "stage knows" % glabel)
        raise SystemExit(2)
    else:
        # nothing outside his own row may read it cell-by-cell
        for ws in wb.worksheets:
            for row_ in ws.iter_rows():
                for cell in row_:
                    v = cell.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    if ws.title == T22 and cell.row == erow:
                        continue
                    for s, e, m in _scan(v):
                        if _sheet_of(m, ws.title) != T22 or m.group("c2"):
                            continue
                        if int(m.group("r1").lstrip("$")) == erow:
                            print("STOP: %s!%s reads 2.2 row %d directly"
                                  % (ws.title, cell.coordinate, erow))
                            raise SystemExit(2)
        kept = {c: c22.cell(erow, c).value for c in (2, 3, 4, 6, 7)}
        lever22 = c22.cell(erow, 5).value
        if lever22 not in LEVERS:
            print("STOP: Ed Tacey's 2.2 lever is %r" % lever22)
            raise SystemExit(2)
        shift_rows(wb, T22, erow, -1)
        log("M5", "%s row %d" % (T22, erow),
            "Ed Tacey's row leaves the Head of Technology group; the "
            "group's SUM($G$...) and COUNTIF ranges shrink with it")

        # the Leadership group goes into the No-archetype block area, right
        # after the group of the grid row above it in that section
        c22 = wb[T22]
        glead = grid_leadership_row(c22)
        prev_label = None
        for r in range(glead - 1,
                       label_row(c22, "No archetype in 1.x tabs"), -1):
            if norm(c22.cell(r, 2).value or ""):
                prev_label = norm(c22.cell(r, 2).value)
                break
        phdr = [r for r in range(2, c22.max_row + 1)
                if norm(c22.cell(r, 2).value or "") == prev_label
                and str(c22.cell(r, 3).value or "").startswith("=COUNTIF")]
        if prev_label is None or len(phdr) != 1:
            print("STOP: no block group for the %r grid row to anchor the "
                  "Leadership group after" % prev_label)
            raise SystemExit(2)
        phdr = phdr[0]
        pa, pb = group_span(c22, phdr)
        at = pb + 1
        shift_rows(wb, T22, at, 2)
        c22 = wb[T22]
        hnew, rnew = at, at + 1

        def gmap(sh, c1, r1, c2, r2):
            f = lambda r: rnew if (sh == T22 and r in (pa, pb)) else r
            return (c1, f(r1), c2, None if r2 is None else f(r2))

        def emap(sh, c1, r1, c2, r2):
            f = lambda r: rnew if (sh == T22 and r == erow) else r
            return (c1, f(r1), c2, None if r2 is None else f(r2))

        for c in range(2, 8):
            copy_style(c22.cell(phdr, c), c22.cell(hnew, c))
            copy_style(c22.cell(pa, c), c22.cell(rnew, c))
        c22.cell(hnew, 2).value = "Leadership"
        for c in (3, 6, 7):
            c22.cell(hnew, c).value = rewrite_refs(
                str(c22.cell(phdr, c).value), T22, gmap)
        for c in (2, 3, 4, 6, 7):
            c22.cell(rnew, c).value = rewrite_refs(str(kept[c]), T22, emap)
        c22.cell(rnew, 5).value = lever22
        if c22.row_dimensions[phdr].height is not None:
            c22.row_dimensions[hnew].height = c22.row_dimensions[phdr].height
        if c22.row_dimensions[pa].height is not None:
            c22.row_dimensions[rnew].height = c22.row_dimensions[pa].height
        log("M5", "%s rows %d-%d" % (T22, hnew, rnew),
            "Leadership group header and Ed Tacey's role row, after the %s "
            "group (his row pattern kept, repointed refs intact, lever %r "
            "carried)" % (prev_label, lever22))

        from openpyxl.worksheet.cell_range import MultiCellRange, CellRange
        for dvn in c22.data_validations.dataValidation:
            if dvn.formula1 and "Filled" in str(dvn.formula1):
                if not any(cr.min_col <= 5 <= cr.max_col
                           and cr.min_row <= rnew <= cr.max_row
                           for cr in dvn.sqref.ranges):
                    dvn.sqref = MultiCellRange(
                        list(dvn.sqref.ranges) +
                        [CellRange(min_col=5, max_col=5,
                                   min_row=rnew, max_row=rnew)])
                    log("M5", "%s!E%d" % (T22, rnew),
                        "lever dropdown extended over the new row")

    # the Leadership grid row reads its group, not the static zeros; the
    # wiring pattern comes from the section's template grid row
    c22 = wb[T22]
    erow = block_row_of(c22, ed)
    ghdr = group_header_above(c22, erow)
    if norm(c22.cell(ghdr, 2).value or "") != "Leadership":
        print("STOP: Ed Tacey's 2.2 row still not under the Leadership "
              "group header")
        raise SystemExit(2)
    la, lb = group_span(c22, ghdr)
    glead = grid_leadership_row(c22)
    tmpl = None
    for r in range(glead - 1, label_row(c22, "No archetype in 1.x tabs"), -1):
        if norm(c22.cell(r, 2).value or "") \
                and str(c22.cell(r, 6).value or "").startswith("=COUNTIF"):
            tmpl = r
            break
    if tmpl is None:
        print("STOP: no wired grid row in the No-archetype section to take "
              "the pattern from")
        raise SystemExit(2)
    ta, tb = (int(x) for x in re.search(
        r"COUNTIF\(\$B\$(\d+):\$B\$(\d+)", str(c22.cell(tmpl, 6).value))
        .groups())

    def wmap(sh, c1, r1, c2, r2):
        f = lambda r: {ta: la, tb: lb}.get(r, r) if sh == T22 else r
        return (c1, f(r1), c2, None if r2 is None else f(r2))

    for c in (6, 8, 9, 10, 11, 12, 19):
        want_f = rewrite_refs(str(c22.cell(tmpl, c).value), T22, wmap)
        cell = c22.cell(glead, c)
        if cell.value != want_f:
            log("M5", "%s!%s%d" % (T22, cell.column_letter, glead),
                "%r -> the %s grid pattern over the Leadership group "
                "rows %d:%d" % (cell.value,
                                norm(c22.cell(tmpl, 2).value), la, lb))
            cell.value = want_f

    save(wb, dst)
    self_check(dst)
    log.tail()
    print("wrote", dst)


# ---------------------------------------------------------------- self-check

def check(name, ok, detail=""):
    print("%s %s%s" % ("PASS" if ok else "FAIL", name,
                       (" - " + detail) if detail else ""), flush=True)
    return ok


def self_check(path):
    print("\n== v2 self-check (formula logic emulated from raw cells, "
          "no cached numbers) ==")
    wb = load(path)
    rv = wb[REVIEW]
    cols = headers(rv)
    roles = emulator(wb)
    by_ar = {}
    for d in roles:
        by_ar.setdefault(d["ar"], []).append(d)
    ok = True

    ok &= check("528 roles in the ledger", len(roles) == 528,
                "%d" % len(roles))

    hot = by_ar.get("Head of Technology", [])
    ok &= check("Head of Technology line 15 roles / 4,908,911.68",
                len(hot) == 15
                and abs(sum(d["aa"] for d in hot) - 4908911.68) < 0.02,
                "%d / %.2f" % (len(hot), sum(d["aa"] for d in hot)))

    dm = by_ar.get("Delivery Manager", [])
    dm_now = sum(d["aa"] for d in dm)
    # the line's mapping-only cost: any part-time literal at its full rate
    dm_pre = sum(d["aa_unscaled"] if (d["aa_literal"]
                 and isinstance(d["fte"], (int, float)) and d["fte"] < 1)
                 else d["aa"] for d in dm)
    ok &= check("Delivery Manager line 12 roles", len(dm) == 12,
                "%d" % len(dm))
    ok &= check("Delivery Manager cost 3,332,474.76 before part-time "
                "scaling", abs(dm_pre - 3332474.76) < 0.02,
                "%.2f" % dm_pre)
    print("INFO Delivery Manager line after part-time scaling (Vanessa "
          "Allen FTE 0.7 sits on it): %.2f = 3,332,474.76 - %.2f"
          % (dm_now, dm_pre - dm_now))

    tm = by_ar.get("Technology Manager", [])
    ok &= check("Technology Manager line 25 roles / 6,776,685.52",
                len(tm) == 25
                and abs(sum(d["aa"] for d in tm) - 6776685.52) < 0.02,
                "%d / %.2f" % (len(tm), sum(d["aa"] for d in tm)))

    viren = one_row(rv, cols["Name"], "Viren Khatri", "Viren Khatri")
    dv = [d for d in roles if d["row"] == viren][0]
    ok &= check("Viren's MTab reads TDD Group Functions",
                dv["aj"] == "TDD Group Functions", repr(dv["aj"]))
    ok &= check("Viren's squad stays EGI TDD", dv["at"] == "EGI TDD",
                repr(dv["at"]))

    g21, g24 = wb[T21], wb[T24]
    refs21, refs24 = block_ref_rows(g21), block_ref_rows(g24)
    ok &= check("Viren on 2.4's FTE block via col-B REVIEW ref",
                viren in refs24.values())
    ok &= check("Viren absent from 2.1's FTE block",
                viren not in refs21.values())
    ok &= check("2.1 carries no EGI TDD grid row and no EGI TDD group",
                not rows_where(g21, 2, "EGI TDD"))

    hdr24 = [r for r in range(2, g24.max_row + 1)
             if norm(g24.cell(r, 2).value or "") == "EGI TDD"
             and str(g24.cell(r, 3).value or "").startswith("=COUNTIF")]
    a, b = group_span(g24, hdr24[0])
    in_group = [refs24[r] for r in range(a, b + 1) if r in refs24]
    ok &= check("2.4 EGI TDD group counts 5 role rows and spans Viren's",
                len(in_group) == 5 and viren in in_group,
                "rows %d:%d -> REVIEW %r" % (a, b, sorted(in_group)))

    egi24 = [d for d in roles
             if d["aj"] == "TDD Group Functions" and d["at"] == "EGI TDD"]
    ok &= check("2.4's grid SUMIFS basis (MTab + squad) resolves 5 roles",
                len(egi24) == 5,
                "%d roles / %.4fm funded and actual"
                % (len(egi24), sum(d["aa"] for d in egi24) / 1e6))
    grid24 = [r for r in range(2, g24.max_row + 1)
              if norm(g24.cell(r, 2).value or "") == "EGI TDD"
              and r not in range(a - 1, b + 1)]
    gfml = str(g24.cell(grid24[0], 15).value or "") if grid24 else ""
    ok &= check("2.4 grid EGI TDD row stays SUMIFS-driven on $C$3 + label",
                bool(grid24) and "SUMIFS" in gfml and "$C$3" in gfml
                and "$AT$" in gfml)
    m = re.search(r"COUNTIF\(\$B\$(\d+):\$B\$(\d+)",
                  str(g24.cell(grid24[0], 6).value or "")) if grid24 else None
    ok &= check("2.4 grid EGI TDD role count reads the whole block group",
                m is not None and (int(m.group(1)), int(m.group(2))) == (a, b))

    ah = label_row(g21, "Directly funded programs and platforms")
    at_ = label_row(g21, "Directly funded total")
    labels = [norm(g21.cell(r, 2).value or "") for r in range(ah + 1, at_)]
    fsum = str(g21.cell(at_, 6).value or "")
    ok &= check("2.1 directly funded total sums only AmPOS + EGI Retail",
                labels == ["AmPOS", "EGI Retail"]
                and fsum == "=SUM(F%d:F%d)" % (ah + 1, at_ - 1),
                "%r, F total %r" % (labels, fsum))

    # M5: Ed Tacey's 2.2 placement and the four grid lines it moves
    c22 = wb[T22]
    ed = one_row(rv, cols["Name"], "Ed Tacey", "Ed Tacey")
    er = block_row_of(c22, ed)
    eh = group_header_above(c22, er) if er else None
    ok &= check("Ed Tacey's 2.2 block row sits in the Leadership group",
                er is not None and eh is not None
                and norm(c22.cell(eh, 2).value or "") == "Leadership")
    la, lb = group_span(c22, eh) if eh else (0, -1)
    lead_refs = [r for r in range(la, lb + 1)
                 if r in block_ref_rows(c22)]
    ok &= check("2.2 Leadership group carries exactly his row",
                lead_refs == [er], "rows %d:%d" % (la, lb))
    hot_h = [r for r in range(2, c22.max_row + 1)
             if norm(c22.cell(r, 2).value or "") == "Head of Technology"
             and str(c22.cell(r, 3).value or "").startswith("=COUNTIF")]
    ha, hb = group_span(c22, hot_h[0])
    hot_refs = [block_ref_rows(c22)[r] for r in range(ha, hb + 1)
                if r in block_ref_rows(c22)]
    ok &= check("2.2 Head of Technology group is 2 rows without Ed",
                len(hot_refs) == 2 and ed not in hot_refs,
                "rows %d:%d" % (ha, hb))
    cust = [d for d in roles if d["aj"] == "Customer"]
    csum = lambda t: sum(d["aa"] for d in cust if d["at"] == t) / 1e6
    ccnt = lambda t: sum(1 for d in cust if d["at"] == t)
    ok &= check("2.2 grid Head of Technology line = 2 roles / 0.5754",
                ccnt("Head of Technology") == 2
                and abs(csum("Head of Technology") - 0.5754) < 0.001,
                "%d / %.4f" % (ccnt("Head of Technology"),
                               csum("Head of Technology")))
    ok &= check("2.2 grid Leadership line = 1 role / 0.3208",
                ccnt("Leadership") == 1
                and abs(csum("Leadership") - 0.3208) < 0.001,
                "%d / %.4f" % (ccnt("Leadership"), csum("Leadership")))
    ovh = [norm(c22.cell(r, 2).value or "")
           for r in range(label_row(c22, "Overhead roles") + 1,
                          label_row(c22, "Overhead roles total"))
           if norm(c22.cell(r, 2).value or "")]
    ovh_m = sum(csum(t) for t in ovh)
    ok &= check("2.2 Overhead roles total = 2.9005 (%s)" % ", ".join(ovh),
                abs(ovh_m - 2.9005) < 0.001, "%.4f" % ovh_m)
    noarch = [norm(c22.cell(r, 2).value or "")
              for r in range(label_row(c22, "No archetype in 1.x tabs") + 1,
                             label_row(c22, "No archetype total"))
              if norm(c22.cell(r, 2).value or "")]
    na_m = sum(csum(t) for t in noarch)
    ok &= check("2.2 No archetype total = 0.4576 (%s)" % ", ".join(noarch),
                abs(na_m - 0.4576) < 0.001, "%.4f" % na_m)
    glead = grid_leadership_row(c22)
    ok &= check("2.2 Leadership grid row wired onto the group "
                "(COUNTIF + SUM($G$...) ranges)",
                str(c22.cell(glead, 6).value)
                == '=COUNTIF($B$%d:$B$%d,"?*")' % (la, lb)
                and str(c22.cell(glead, 19).value)
                == "=SUM($G$%d:$G$%d)/1000000" % (la, lb))
    hot_g = [r for r in range(2, c22.max_row + 1)
             if norm(c22.cell(r, 2).value or "") == "Head of Technology"
             and str(c22.cell(r, 19).value or "").startswith("=SUM")]
    ok &= check("2.2 Head of Technology grid row's after-lever sum spans "
                "the 2-row group",
                bool(hot_g) and str(c22.cell(hot_g[0], 19).value)
                == "=SUM($G$%d:$G$%d)/1000000" % (ha, hb))
    grid_labels = [norm(c22.cell(r, 2).value or "")
                   for r in range(2, label_row(c22, "Total portfolio"))
                   if "SUMIFS" in str(c22.cell(r, 15).value or "")]
    tied = abs(sum(csum(t) for t in grid_labels)
               - sum(d["aa"] for d in cust) / 1e6) < 1e-6
    ok &= check("2.2 cost-control basis ties: every Customer role lands on "
                "a grid line", tied,
                "%.6f vs %.6f" % (sum(csum(t) for t in grid_labels),
                                  sum(d["aa"] for d in cust) / 1e6))

    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "#REF" in v:
                    bad.append("%s!%s" % (ws.title, cell.coordinate))
    ok &= check("no #REF! anywhere in the workbook's formulas", not bad,
                "; ".join(bad[:6]))

    for tab, mtab in ((T21, "Ampol Retail"), (T22, "Customer"),
                      (T24, "TDD Group Functions")):
        blk = len(block_ref_rows(wb[tab]))
        cnt = sum(1 for d in roles if d["aj"] == mtab)
        ok &= check("%s FTE block rows equal the role mapping's %s count"
                    % (tab, mtab), blk == cnt, "%d vs %d" % (blk, cnt))

    scaled = [d for d in roles if d["aa_literal"]
              and isinstance(d["fte"], (int, float)) and d["fte"] < 1]
    good = all(abs(d["aa"] - round(d["aa_unscaled"] * d["fte"], 2)) < 0.01
               for d in scaled)
    ok &= check("part-time Full Cost AUD equals full rate x FTE for the "
                "scaled set", bool(scaled) and good, "%d roles" % len(scaled))
    print("INFO part-time set: " + "; ".join(
        "%s %.2f x %.1f -> %.2f" % (d["name"], d["aa_unscaled"], d["fte"],
                                    d["aa"]) for d in scaled))
    print("INFO part-time reduction total: %.2f (%.4fm)"
          % (sum(d["aa_unscaled"] - d["aa"] for d in scaled),
             sum(d["aa_unscaled"] - d["aa"] for d in scaled) / 1e6))

    if not ok:
        print("\nv2 self-check: FAIL")
        raise SystemExit(1)
    print("\nv2 self-check: all PASS")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
