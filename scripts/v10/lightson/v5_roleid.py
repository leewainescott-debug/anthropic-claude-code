#!/usr/bin/env python3
"""v5 - spec Stage v5: Role ID rekey.

  python3 v5_roleid.py <in.xlsx> <out.xlsx>

The 2.x FTE blocks stop naming REVIEW rows by number. The role mapping gets a
Role ID column (R0001.. down the data rows, plain values, quiet grey); every
role row on a 2.x tab stores its ID in column A, and every single-cell
='REVIEW - Complete Role Mapping'!$X$n pull on those rows becomes INDEX/MATCH
keyed on that ID. Re-sorting the role mapping can then never silently repoint
a row: the model follows the ID wherever the row goes.

Value-neutral: no number and no text changes anywhere except the new ID cells
and their headers.

Idempotent: handed its own output it copies it through untouched.
"""
import sys, os, re, shutil, collections
sys.path.insert(0, "/home/user/anthropic-claude-code/scripts/v10")
sys.path.insert(0, "/home/user/anthropic-claude-code/scripts/v10/update")

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import wbio
from _xl import REVIEW, Log, load, save, white

ID_HDR = "Role ID"
GREY = "FFA6A6A6"                    # light grey, legible but quiet

# a single-cell, fully anchored pull off the role mapping: 'REVIEW - ...'!$X$n
# (the range starts in SUMIFS/COUNTIFS carry a trailing ':' and never match)
SINGLE = re.compile(r"'" + re.escape(REVIEW) + r"'!\$([A-Z]{1,2})\$(\d+)(?![0-9:])")
# a role row is a row whose status cell reads the ledger's AK by row number
DPAT = re.compile(r"^='" + re.escape(REVIEW) + r"'!\$AK\$(\d+)$")

IM = "INDEX('{v}'!${c}$2:${c}$700,MATCH($A{r},'{v}'!${i}$2:${i}$700,0))"


def tabs2x(wb):
    return [ws.title for ws in wb.worksheets if ws.title.startswith("2.")]


def fte_hdr(ws):
    for r in range(5, 120):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.endswith(" FTE"):
            return r + 1
    raise SystemExit("STOP: no FTE block on %s" % ws.title)


def data_rows(rv):
    """Ledger rows that carry a role: column B typed and non-blank."""
    out = []
    for r in range(2, rv.max_row + 1):
        v = rv.cell(r, 2).value
        if v is not None and str(v).strip():
            out.append(r)
    return out


def last_used_col(rv):
    last = 0
    for c in range(1, rv.max_column + 1):
        for r in range(1, rv.max_row + 1):
            if rv.cell(r, c).value is not None:
                last = c
                break
    return last


def grey(cell, size):
    f = cell.font
    cell.font = Font(name=f.name, size=size, bold=f.bold, italic=f.italic,
                     color=GREY)


def main(src, dst):
    log = Log("v5_roleid")
    wb = load(src)
    rv = wb[REVIEW]

    if any(rv.cell(1, c).value == ID_HDR for c in range(1, rv.max_column + 1)):
        print("input is already rekeyed - copying through")
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        return

    # ------------------------------------------------------------- the IDs
    log.head("Role ID column on the role mapping, first free column")
    rows = data_rows(rv)
    idc = last_used_col(rv) + 1
    idl = get_column_letter(idc)
    rv.cell(1, idc).value = ID_HDR
    grey(rv.cell(1, idc), 8)
    id_of = {}
    for i, rr in enumerate(rows, 1):
        rid = "R%04d" % i
        rv.cell(rr, idc).value = rid
        grey(rv.cell(rr, idc), 8)
        id_of[rr] = rid
    log("IDs", "%s!%s" % (REVIEW, idl),
        "R0001..R%04d down the %d data rows in current order, plain values, "
        "grey 8pt (last used column %s)" % (len(rows), len(rows),
                                            get_column_letter(idc - 1)))

    # ------------------------------------------- the 2.x rekey, row by row
    log.head("every 2.x role row keyed on its ID, not the ledger row number")
    per_tab = collections.Counter()
    left = []                       # (tab, coord, ref, reason) not rewritten
    for title in tabs2x(wb):
        ws = wb[title]
        roles = []
        for row in ws.iter_rows(min_col=4, max_col=4):
            for c in row:
                if isinstance(c.value, str):
                    m = DPAT.match(c.value)
                    if m:
                        roles.append((c.row, int(m.group(1))))
        for r, n in roles:
            if n not in id_of:
                print("STOP: %s row %d reads ledger row %d, which carries "
                      "no role" % (title, r, n))
                raise SystemExit(2)
            ws.cell(r, 1).value = id_of[n]
            grey(ws.cell(r, 1), 7)
            for c in next(ws.iter_rows(min_row=r, max_row=r)):
                v = c.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue

                def sub(m):
                    if int(m.group(2)) != n:
                        left.append((title, c.coordinate, m.group(0),
                                     "points at ledger row %s, not this "
                                     "row's role" % m.group(2)))
                        return m.group(0)
                    per_tab[title] += 1
                    return IM.format(v=REVIEW, c=m.group(1), r=r, i=idl)

                nv = SINGLE.sub(sub, v)
                if nv != v:
                    c.value = nv
        hdr = fte_hdr(ws)
        ws.cell(hdr, 1).value = "ID"
        white(ws, "A%d" % hdr)
        ws.column_dimensions["A"].width = 5.5
        log("rekey", title,
            "%d role rows carry their ID in column A (grey 7pt, width 5.5, "
            "header white); %d row-anchored refs now INDEX/MATCH on the ID"
            % (len(roles), per_tab[title]))

    # ------------------------------------------------- the sweep, workbook
    log.head("sweep - row-anchored role mapping refs left anywhere")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                for m in SINGLE.finditer(v):
                    known = any(t == ws.title and co == c.coordinate
                                and ref == m.group(0)
                                for t, co, ref, _ in left)
                    if not known:
                        left.append((ws.title, c.coordinate, m.group(0),
                                     "not on an FTE-block role row - no ID "
                                     "cell to key on"))
    for t, co, ref, why in left:
        log("left", "%s!%s" % (t, co), "%s kept: %s" % (ref, why))
    log.note("sweep", "%d refs rewritten across %d tabs, %d deliberately left"
             % (sum(per_tab.values()), len(per_tab), len(left)))

    tmp = dst + ".raw"
    save(wb, tmp)
    log.head("recalculating and writing the cached values back")
    rc, st = wbio.build(tmp, dst)
    os.remove(tmp)
    rcdir = os.path.dirname(rc)
    if os.path.basename(rcdir).startswith("recalc_"):
        shutil.rmtree(rcdir, ignore_errors=True)
    print("recalculated, %d formula cells populated across %d sheets"
          % (st["cells"], st["sheets"]), flush=True)

    # ---------------------------------------------------------- self-check
    log.head("self-check")
    fails = []

    def chk(name, ok, detail):
        print("%s  %s - %s" % ("PASS" if ok else "FAIL", name, detail),
              flush=True)
        if not ok:
            fails.append(name)

    err, _ = wbio.audit(dst)
    chk("no error cells", not err,
        "%d error cells%s" % (len(err), (", e.g. %r" % err[:3]) if err else ""))

    wf = openpyxl.load_workbook(dst, data_only=False)
    wv = openpyxl.load_workbook(dst, data_only=True)
    fr, vr = wf[REVIEW], wv[REVIEW]
    ids = [(r, fr.cell(r, idc).value) for r in range(2, fr.max_row + 1)
           if fr.cell(r, idc).value is not None]
    seq = ["R%04d" % i for i in range(1, len(ids) + 1)]
    chk("ID column complete", [v for _, v in ids] == seq and len(ids) == 528,
        "%d IDs, unique and sequential R0001..R%04d, 528 expected"
        % (len(ids), len(ids)))

    n_left = 0
    for ws in wf.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    n_left += len(SINGLE.findall(c.value))
    chk("row-anchored refs gone", n_left == len(left),
        "%d single-cell row-anchored refs remain, %d deliberately left"
        % (n_left, len(left)))

    at = {v: r for r, v in ids}
    n_rows = n_bad = 0
    for title in tabs2x(wf):
        f, v = wf[title], wv[title]
        for row in f.iter_rows(min_col=1, max_col=1):
            for c in row:
                rid = c.value
                if not (isinstance(rid, str) and re.fullmatch(r"R\d{4}", rid)):
                    continue
                n_rows += 1
                rr = at.get(rid)
                got = (v.cell(c.row, 2).value, v.cell(c.row, 4).value,
                       v.cell(c.row, 6).value)
                want = (vr.cell(rr, 2).value, vr.cell(rr, 37).value,
                        vr.cell(rr, 27).value)
                ok = (got[0] == want[0] and got[1] == want[1]
                      and isinstance(got[2], (int, float))
                      and isinstance(want[2], (int, float))
                      and abs(got[2] - want[2]) <= 1e-9)
                if not ok:
                    n_bad += 1
                    if n_bad <= 3:
                        print("  mismatch %s row %d %s: %r vs ledger %r"
                              % (title, c.row, rid, got, want))
    chk("role rows read their own person", n_rows == 528 and n_bad == 0,
        "%d role rows keyed, %d read a name, status or cost different from "
        "their ledger row" % (n_rows, n_bad))

    n_ctl = bad_ctl = 0
    for title in tabs2x(wf):
        f, v = wf[title], wv[title]
        for row in f.iter_rows(min_col=2, max_col=2):
            for c in row:
                if isinstance(c.value, str) and "must be 0" in c.value:
                    n_ctl += 1
                    cv = v.cell(c.row, 3).value
                    if not (isinstance(cv, (int, float)) and abs(cv) <= 1e-9):
                        bad_ctl += 1
    chk("controls read 0", n_ctl == 45 and bad_ctl == 0,
        "%d control rows on the lever modelling tabs, %d non-zero"
        % (n_ctl, bad_ctl))
    wf.close(); wv.close()

    if fails:
        print("\nv5_roleid: FAILED - %s" % ", ".join(fails))
        raise SystemExit(2)
    log.tail()
    print("wrote", dst)


if __name__ == "__main__":
    main(*sys.argv[1:])
