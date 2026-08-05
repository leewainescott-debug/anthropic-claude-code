#!/usr/bin/env python3
"""v4 - spec stage v4: model fixes.

  python3 v4_fixes.py <in.xlsx> <out.xlsx>

Six fixes, every location derived by label, never by row number:
  1  COE netting at actual - 2.12's netting cell reads minus the TDD Business
     Partner group's cost after levers off its own grid, 2.13's reads minus the
     Domain Architect group's; the planned spend lines consume the netting cell.
  2  3.1 consistency - the two COE 'Cost after levers' cells show gross less
     the netting, so 3.1 says what the funding blocks say.
  3  0.2 TDD Cyber spend at actual - 2.15's total after levers less the Lists
     cyber uplift funding, not the 1.14 budget split.
  4  1.3 carries the EGI Ent Data platform line live off 2.3's directly funded
     grid row; the funding block reads it and Left to fund nets it off.
  5  1.4's pool line becomes 'Significant Items EGI' reading 2.4's EGI TDD
     funded amount off the grid.
  6  1.1's 'Significant Items EGI' line verified live off 2.1's EGI Retail
     block; not edited.
Plus: the 4.0 Data QA ties restated on the same basis, and the language sweep.

Idempotent: handed its own output it copies it through untouched.
"""
import sys, os, re, shutil

V10 = "/home/user/anthropic-claude-code/scripts/v10"
sys.path.insert(0, V10)
sys.path.insert(0, os.path.join(V10, "update"))

import openpyxl
from openpyxl.utils import get_column_letter

import wbio
from _xl import REVIEW, Log, load, save, copy_style

T212 = "2.12 COE BP&T"
T213 = "2.13 COE SA&D"
T31 = "3.1 Archetype to Actuals"
T02 = "0.2 Data Config"
T215 = "2.15 TDD Cyber"
T21 = "2.1 Ampol Retail"
T23 = "2.3 Enterprise Data"
T24 = "2.4 TDD Group Functions"
T11 = "1.1 Ampol Retail"
T13 = "1.3 Enterprise Data"
T14 = "1.4 TDD Group Functions"
QA = "4.0 Data QA"

MARK = "Platform: EGI Ent Data"      # the new 1.3 platform header, and the guard
LINE = "EGI Ent Data"
POOL_LBL = "Funded from TDD Corporate pool (3.4 COE Breakdown)"
SIG_LBL = "Significant Items EGI"

NOTE = ("Planned spend is net of the %s cost met inside portfolio overheads, "
        "at the actual cost after levers; %s.")
NOTE_TAIL = {T212: "the COE draws down its own allocation only",
             T213: "COEs draw down on their own allocation only"}
NOTE_ROLE = {T212: "Business Partner", T213: "Domain Architect"}

# words his standing rules ban from the workbook; nothing v4 writes may carry
# one, and 'to projects' must stay at zero across the whole book
BAN = re.compile(r"\bwaves?\b|\bseats?\b|\bdesign\b|to projects", re.I)
DASHES = ("–", "—")


def sq(name):
    return "'" + name.replace("'", "''") + "'"


def put(ws, r, c, value):
    if isinstance(value, str):
        if BAN.search(value) or any(d in value for d in DASHES):
            print("STOP: banned wording in %r" % value)
            raise SystemExit(2)
    ws.cell(r, c).value = value


def brow(ws, text, exact=True, lo=1, hi=None):
    """First row whose column B carries `text` (exact or as a prefix)."""
    hi = hi or ws.max_row
    for r in range(lo, hi + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and (v == text if exact else v.startswith(text)):
            return r
    print("STOP: no row labelled %r on %s" % (text, ws.title))
    raise SystemExit(2)


def hcol_row(ws, col, text, exact=True):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if isinstance(v, str) and (v == text if exact else text in v):
            return r
    print("STOP: no row labelled %r in col %d on %s" % (text, col, ws.title))
    raise SystemExit(2)


def grid(ws):
    """(header_row, total_row, {header label: col}) for a 2.x grid."""
    hdr = brow(ws, "Squad")
    tot = brow(ws, "Total portfolio")
    cols = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(hdr, c).value
        if isinstance(v, str) and v:
            cols[v] = c
    for need in ("Actual cost ($m)", "Funded outside TDD ($m)",
                 "Squad cost after levers ($m)"):
        if need not in cols:
            print("STOP: %s grid has no %r column" % (ws.title, need))
            raise SystemExit(2)
    return hdr, tot, cols


def grid_row(ws, name):
    hdr, tot, _ = grid(ws)
    for r in range(hdr + 1, tot):
        if ws.cell(r, 2).value == name:
            return r
    print("STOP: no grid row %r on %s" % (name, ws.title))
    raise SystemExit(2)


def da_group(ws, rv):
    """The FTE-block group carrying the Domain Architect roles, by title."""
    hdr = None
    for r in range(5, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.endswith(" FTE"):
            hdr = r
            break
    if hdr is None:
        print("STOP: no FTE block on %s" % ws.title)
        raise SystemExit(2)
    hits, group = {}, None
    for r in range(hdr + 1, ws.max_row + 1):
        b, c = ws.cell(r, 2).value, ws.cell(r, 3).value
        if isinstance(c, str) and c.startswith("=COUNTIF("):
            group = b
            continue
        if group and isinstance(c, str) and REVIEW in c:
            m = re.search(r"\$C\$(\d+)", c)
            if m:
                title = str(rv.cell(int(m.group(1)), 3).value or "")
                if "Domain Architect" in title:
                    hits[group] = hits.get(group, 0) + 1
    if not hits:
        print("STOP: no Domain Architect titles behind %s" % ws.title)
        raise SystemExit(2)
    return max(hits, key=hits.get)


def fund_span(ws):
    """(first, last) row of the consolidated COE funding block."""
    fund = brow(ws, "Funding")
    last = fund
    for r in range(fund + 1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.endswith(" FTE"):
            break
        last = r
    return fund + 1, last


def main(src, dst):
    log = Log("v4_fixes")
    wb = load(src)

    w13 = wb[T13]
    if any(w13.cell(r, 2).value == MARK for r in range(1, w13.max_row + 1)):
        print("input already carries the v4 fixes - copying through")
        shutil.copy(src, dst)
        log.tail()
        print("wrote", dst)
        return

    rv = wb[REVIEW]
    qa = wb[QA]

    # ------------------------------------------------- 1  netting at actual
    log.head("1  COE netting at actual on 2.12 and 2.13")
    nets = {}          # tab -> (netting row, group grid row, total planned row)
    for tab in (T212, T213):
        ws = wb[tab]
        group = ("TDD Business Partner" if tab == T212 else da_group(ws, rv))
        grow = grid_row(ws, group)
        _, _, cols = grid(ws)
        sc = get_column_letter(cols["Squad cost after levers ($m)"])
        nrow = hcol_row(ws, 2, "netted out of planned spend", exact=False)
        tprow = brow(ws, "Total planned spend ($m)")
        nets[tab] = (nrow, grow, tprow)
        f = ws.cell(nrow, 3).value
        m = re.fullmatch(r"=-\$C\$(\d+)", str(f))
        if m is None:
            print("STOP: %s netting cell reads %r, not the allowance" % (tab, f))
            raise SystemExit(2)
        allow = int(m.group(1))
        if "funding from portfolio overheads" not in str(ws.cell(allow, 2).value):
            print("STOP: %s C%d is not the allowance line" % (tab, allow))
            raise SystemExit(2)
        put(ws, nrow, 3, "=-$%s$%d" % (sc, grow))
        log("1", "%s!C%d" % (tab, nrow),
            "netting reads -$%s$%d (the %s group after levers), was -$C$%d "
            "(the allowance)" % (sc, grow, group, allow))
        a, z = fund_span(ws)
        pat = re.compile(r"-\$C\$%d(?!\d)" % allow)
        n = 0
        for r in range(a, z + 1):
            if r == nrow:
                continue
            v = ws.cell(r, 3).value
            if isinstance(v, str) and pat.search(v):
                ws.cell(r, 3).value = pat.sub("+$C$%d" % nrow, v)
                log("1", "%s!C%d" % (tab, r),
                    "planned spend consumes the netting cell C%d, so it nets "
                    "the actual" % nrow)
                n += 1
        if n == 0:
            print("STOP: no planned spend consumed the %s allowance" % tab)
            raise SystemExit(2)
        for r in (nrow + 1, nrow + 2):
            v = ws.cell(r, 2).value
            if isinstance(v, str) and v.startswith("Planned spend is net"):
                put(ws, r, 2, NOTE % (NOTE_ROLE[tab], NOTE_TAIL[tab]))
                log("1", "%s!B%d" % (tab, r),
                    "the note drops its allowance row reference, states the "
                    "actual basis")
                break

    # --------------------------------------------------- 2  3.1 consistency
    log.head("2  3.1 shows the netted COE cost after levers")
    t31 = wb[T31]
    lrow = brow(t31, "Line")
    icol = next(c for c in range(2, t31.max_column + 1)
                if isinstance(t31.cell(lrow, c).value, str)
                and "Cost after levers" in t31.cell(lrow, c).value)
    for tab, label in ((T212, "COE BP&T"), (T213, "COE SA&D")):
        r = hcol_row(t31, 3, label)
        _, tot, cols = grid(wb[tab])
        sc = get_column_letter(cols["Squad cost after levers ($m)"])
        nrow = nets[tab][0]
        cur = str(t31.cell(r, icol).value)
        want_gross = "=%s!$%s$%d" % (sq(tab), sc, tot)
        if cur != want_gross:
            print("STOP: 3.1 %s after-levers reads %r, expected %r"
                  % (label, cur, want_gross))
            raise SystemExit(2)
        put(t31, r, icol, "%s+%s!$C$%d" % (want_gross, sq(tab), nrow))
        log("2", "%s!%s%d" % (T31, get_column_letter(icol), r),
            "gross $S$%d plus the (negative) netting C%d - the netted number"
            % (tot, nrow))

    # ------------------------------------------------ 3  0.2 cyber accurate
    log.head("3  0.2's TDD Cyber spend on the actuals basis")
    t215 = wb[T215]
    _, tot215, cols215 = grid(t215)
    pcol = cols215["Funded outside TDD ($m)"]
    pf = None
    for r in range(7, tot215):
        v = t215.cell(r, pcol).value
        if isinstance(v, str) and "Lists!" in v:
            pf = v
            break
    m = re.search(r"INDEX\(Lists!\$([A-Z]{1,2})\$2:\$\1\$(\d+),"
                  r"MATCH\(\$B\d+,Lists!\$([A-Z]{1,2})\$2", pf or "")
    if m is None:
        print("STOP: cannot derive the Lists funded-outside columns")
        raise SystemExit(2)
    wcol, wlast, lcol = m.group(1), int(m.group(2)), m.group(3)
    lists = wb["Lists"]
    li = openpyxl.utils.column_index_from_string
    uprow = next((r for r in range(2, wlast + 1)
                  if lists.cell(r, li(lcol)).value == "Cyber Uplift"), None)
    if uprow is None or not isinstance(lists.cell(uprow, li(wcol)).value,
                                       (int, float)):
        print("STOP: no numeric Cyber Uplift funding amount on Lists")
        raise SystemExit(2)
    cfg = wb[T02]
    hdr02 = brow(cfg, "Portfolio")
    fcol = next(c for c in range(2, 10) if cfg.cell(hdr02, c).value == "Spend")
    cyrow = brow(cfg, "TDD Cyber", lo=hdr02 + 1)
    cur = str(cfg.cell(cyrow, fcol).value)
    if "1.14 TDD Cyber" not in cur:
        print("STOP: 0.2 TDD Cyber spend reads %r, not the 1.14 split" % cur)
        raise SystemExit(2)
    scol = get_column_letter(cols215["Squad cost after levers ($m)"])
    put(cfg, cyrow, fcol, "=%s!$%s$%d-Lists!$%s$%d"
        % (sq(T215), scol, tot215, wcol, uprow))
    log("3", "%s!%s%d" % (T02, get_column_letter(fcol), cyrow),
        "2.15's total after levers less the Lists Cyber Uplift funding "
        "(Lists!%s%d), was the 1.14 C9+D9 budget split" % (wcol, uprow))

    # ------------------------------------- 4  1.3's EGI Ent Data platform
    log.head("4  the EGI Ent Data platform line on 1.3, live off 2.3")
    w23 = wb[T23]
    _, tot23, cols23 = grid(w23)
    egi23 = grid_row(w23, "EGI")
    o23 = get_column_letter(cols23["Actual cost ($m)"])
    s23 = get_column_letter(cols23["Squad cost after levers ($m)"])
    prow = brow(w13, "Platform: EGI Data", exact=False)
    hrow, srow = prow + 1, prow + 2
    w14 = wb[T14]
    d_t = brow(w14, "Platform: EGI TDD")            # style donors, his block
    gd_t = brow(w13, "Platform: Group Data")        # labels and row heights
    covered = any(cr.min_row <= srow <= cr.max_row and cr.min_col == 7
                  for dv in w13.data_validations.dataValidation
                  if str(dv.formula1) == "SupportPct"
                  for cr in dv.sqref.ranges)
    if not covered:
        print("STOP: 1.3 row %d has no Support %% dropdown - not the old "
              "EGI block's squad row" % srow)
        raise SystemExit(2)
    for c in range(2, 14):
        copy_style(w14.cell(d_t, c), w13.cell(prow, c))
        copy_style(w14.cell(d_t + 1, c), w13.cell(hrow, c))
        copy_style(w14.cell(d_t + 2, c), w13.cell(srow, c))
        w13.cell(hrow, c).value = w13.cell(gd_t + 1, c).value
    put(w13, prow, 2, MARK)
    put(w13, srow, 2, LINE)
    put(w13, srow, 3, "Strategic Programs")
    put(w13, srow, 5, "Onshore")
    put(w13, srow, 7, 0)
    put(w13, srow, 8, "=%s!$%s$%d" % (sq(T23), o23, egi23))
    put(w13, srow, 9, '=IFERROR($H%d*$G%d,"")' % (srow, srow))
    put(w13, srow, 10, '=IFERROR($H%d*(1-$G%d),"")' % (srow, srow))
    put(w13, srow, 11, "=%s!$%s$%d" % (sq(T23), s23, egi23))
    put(w13, srow, 12, "=IF(AND(ISNUMBER($H%d),ISNUMBER($K%d)),"
        'ROUND($K%d-$H%d,6),"")' % (srow, srow, srow, srow))
    for r_from, r_to in ((gd_t, prow), (gd_t + 1, hrow), (gd_t + 2, srow)):
        h = w13.row_dimensions[r_from].height
        if h:
            w13.row_dimensions[r_to].height = h
    log("4", "%s!B%d:L%d" % (T13, prow, srow),
        "platform header %r and one squad row %r - Total Squad Cost live "
        "from %s!%s%d, Support %% 0, funded outside carries the lot"
        % (MARK, LINE, T23, o23, egi23))

    other = hcol_row(w13, 8, "Other cost (this model)")
    pool = hcol_row(w13, 8, POOL_LBL)
    sig = hcol_row(w13, 8, "Significant Items - EGI")
    ltf = hcol_row(w13, 8, "Left to fund")
    put(w13, sig, 10, "=J%d" % srow)
    put(w13, ltf, 10, "=J%d-J%d-J%d" % (other, pool, sig))
    log("4", "%s!J%d,J%d" % (T13, sig, ltf),
        "'Significant Items - EGI' reads the new line's funded amount J%d; "
        "'Left to fund' nets it off, so the EGI slice reads 0" % srow)

    ssc = brow(w13, "Squad Support Costs")
    ev = str(w13.cell(ssc, 5).value)
    if not (ev.startswith("=SUM(") and ev.endswith(")")):
        print("STOP: 1.3 other-cost sum reads %r" % ev)
        raise SystemExit(2)
    w13.cell(ssc, 5).value = ev[:-1] + ",J%d)" % srow
    for col, mkt in ((3, "AU"), (4, "NZ")):
        w13.cell(ssc, col).value = (str(w13.cell(ssc, col).value)
                                    + '+SUMIF(F%d:F%d,"%s",I%d:I%d)'
                                    % (srow, srow, mkt, srow, srow))
    log("4", "%s!C%d:E%d" % (T13, ssc, ssc),
        "the cost rows count the new line, as 1.4 counts its EGI TDD line")

    # --------------------------------------- 5  1.4 Significant Items EGI
    log.head("5  1.4's pool line becomes Significant Items EGI, off 2.4")
    w24 = wb[T24]
    _, _, cols24 = grid(w24)
    egi24 = grid_row(w24, "EGI TDD")
    p24 = get_column_letter(cols24["Funded outside TDD ($m)"])
    r = hcol_row(w14, 8, POOL_LBL)
    put(w14, r, 8, SIG_LBL)
    put(w14, r, 10, "=%s!$%s$%d" % (sq(T24), p24, egi24))
    log("5", "%s!H%d:J%d" % (T14, r, r),
        "%r -> %r, reading %s!%s%d, the EGI TDD grid row's funded amount"
        % (POOL_LBL, SIG_LBL, T24, p24, egi24))

    # ------------------------------------------- 6  1.1 verified, not edited
    log.head("6  1.1's Significant Items EGI verified live, left alone")
    w11 = wb[T11]
    r = hcol_row(w11, 8, SIG_LBL)
    jf = str(w11.cell(r, 10).value)
    m = re.fullmatch(r"=J(\d+)", jf)
    if m is None or w11.cell(int(m.group(1)), 2).value != "EGI Retail":
        print("STOP: 1.1 %r reads %r, not the EGI Retail block" % (SIG_LBL, jf))
        raise SystemExit(2)
    amp = hcol_row(w11, 8, "Significant Items")
    log("6", "%s!J%d" % (T11, r),
        "reads J%s, the EGI Retail platform line; the AmPOS line (H%d) stays"
        % (m.group(1), amp))

    # ------------------------------------------ QA ties on the same basis
    log.head("QA  the 4.0 ties restated on the actual-netting basis")
    for tab in (T212, T213):
        nrow, _, tprow = nets[tab]
        _, tot, _ = grid(wb[tab])
        r = hcol_row(qa, 2, "%s funding block against its cost" % tab,
                     exact=False)
        put(qa, r, 3, "=%s!$C$%d-%s!$C$%d" % (sq(tab), tprow, sq(tab), nrow))
        log("QA", "%s!C%d" % (QA, r),
            "planned spend less the netting equals the gross after levers")
        r32 = hcol_row(qa, 2, "Lever impact on 3.1", exact=False)
        term = "-N(%s!$S$%d)" % (sq(tab), tot)
        d = str(qa.cell(r32, 4).value)
        if term not in d:
            print("STOP: QA lever-impact tie carries no %r" % term)
            raise SystemExit(2)
        qa.cell(r32, 4).value = d.replace(
            term, term + "-N(%s!$C$%d)" % (sq(tab), nrow))
        log("QA", "%s!D%d" % (QA, r32),
            "%s nets the COE pot out of 3.1's after-levers column, the tie "
            "follows" % tab)
    r41 = hcol_row(qa, 2, "%s archetype total against Total Cost" % T23,
                   exact=False)
    old = "+N(%s!$N$%d)" % (sq(T23), egi23)
    d = str(qa.cell(r41, 4).value)
    if old not in d:
        print("STOP: QA 1.3 tie carries no EGI compensation term %r" % old)
        raise SystemExit(2)
    qa.cell(r41, 4).value = d.replace(old, "")
    log("QA", "%s!D%d" % (QA, r41),
        "1.3's model now carries the EGI line itself, the compensation "
        "term goes")

    # -------------------------------------------------- the language sweep
    log.head("sweep  'to projects' anywhere in the book (expect none)")
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and not v.startswith("=") \
                        and "to projects" in v.lower():
                    c.value = re.sub("to projects", "rechargeable", v,
                                     flags=re.I)
                    log("sweep", "%s!%s" % (ws.title, c.coordinate),
                        "'to projects' -> 'rechargeable'")
                    n += 1
    log.note("sweep", "%d 'to projects' phrasings found" % n)

    # ------------------------------------------------------- build and check
    tmp = dst + ".raw"
    save(wb, tmp)
    log.head("recalculating and writing the cached values back")
    rc, st = wbio.build(tmp, dst)
    os.remove(tmp)
    print("recalculated, %d formula cells populated across %d sheets"
          % (st["cells"], st["sheets"]), flush=True)
    err, blank = wbio.audit(dst)
    if err:
        print("STOP: %d error cells, e.g. %r" % (len(err), err[:5]))
        raise SystemExit(2)

    self_check(src, dst, nets)
    log.tail()
    print("wrote", dst)


# ------------------------------------------------------------- self-check

def self_check(src, dst, nets):
    print("\n== self-check (every target recomputed from the file)", flush=True)
    wf = openpyxl.load_workbook(dst, data_only=False)
    wv = openpyxl.load_workbook(dst, data_only=True)
    checks = []

    def ck(name, ok, detail):
        checks.append(ok)
        print("%s  %-52s %s" % ("PASS" if ok else "FAIL", name, detail),
              flush=True)

    def num(tab, r, c):
        v = wv[tab].cell(r, c).value
        return v if isinstance(v, (int, float)) else 0.0

    # 1  netting at actual
    for tab, group in ((T212, "TDD Business Partner"), (T213, None)):
        ws, nrow = wf[tab], nets[tab][0]
        grow, tprow = nets[tab][1], nets[tab][2]
        _, tot, cols = grid(ws)
        scol = cols["Squad cost after levers ($m)"]
        pot = num(tab, grow, scol)
        got = num(tab, nrow, 3)
        ck("%s netting cell" % tab, abs(got + pot) < 1e-6
           and wf[tab].cell(nrow, 3).value == "=-$S$%d" % grow,
           "reads %.4f = -(group after levers %.4f)" % (got, pot))
        planned = num(tab, tprow, 3)
        gross = num(tab, tot, scol)
        ck("%s planned spend nets the actual" % tab,
           abs(planned - got - gross) < 1e-6,
           "total planned %.4f less netting %.4f = gross %.4f"
           % (planned, got, gross))

    # 2  3.1 netted
    t31v = wv[T31]
    lrow = brow(wf[T31], "Line")
    icol = next(c for c in range(2, wf[T31].max_column + 1)
                if isinstance(wf[T31].cell(lrow, c).value, str)
                and "Cost after levers" in wf[T31].cell(lrow, c).value)
    for tab, label in ((T212, "COE BP&T"), (T213, "COE SA&D")):
        r = hcol_row(wf[T31], 3, label)
        _, tot, cols = grid(wf[tab])
        want = num(tab, tot, cols["Squad cost after levers ($m)"]) \
            + num(tab, nets[tab][0], 3)
        got = t31v.cell(r, icol).value
        ck("3.1 %s after levers netted" % label, abs(got - want) < 1e-6,
           "%.4f = gross less the pot" % got)

    # 3  0.2 cyber
    cfg = wf[T02]
    hdr02 = brow(cfg, "Portfolio")
    fcol = next(c for c in range(2, 10) if cfg.cell(hdr02, c).value == "Spend")
    cyrow = brow(cfg, "TDD Cyber", lo=hdr02 + 1)
    _, tot215, cols215 = grid(wf[T215])
    f = str(cfg.cell(cyrow, fcol).value)
    m = re.search(r"Lists!\$([A-Z]{1,2})\$(\d+)", f)
    lref = num("Lists", int(m.group(2)), openpyxl.utils.
               column_index_from_string(m.group(1))) if m else None
    want = num(T215, tot215, cols215["Squad cost after levers ($m)"]) - (lref or 0)
    got = wv[T02].cell(cyrow, fcol).value
    ck("0.2 TDD Cyber spend accurate", m is not None and T215 in f
       and abs(got - want) < 1e-6,
       "%.4f = 2.15 after levers less the Lists uplift funding %.4f"
       % (got, lref if lref is not None else -1))
    totrow02 = brow(cfg, "Total")
    tot02 = wv[T02].cell(totrow02, fcol).value
    parts = sum(num(T02, r, fcol) for r in range(hdr02 + 1, totrow02))
    ck("0.2 total spend consistent", abs(tot02 - parts) < 1e-6,
       "Total spend %.4f = the column above it" % tot02)

    # 4  1.3
    w13 = wf[T13]
    prow = brow(w13, MARK)
    srow = prow + 2
    egi23 = grid_row(wf[T23], "EGI")
    _, _, cols23 = grid(wf[T23])
    o = num(T23, egi23, cols23["Actual cost ($m)"])
    ck("1.3 platform line live", w13.cell(srow, 2).value == LINE
       and abs(num(T13, srow, 8) - o) < 1e-6,
       "%r Total Squad Cost %.4f = 2.3's EGI grid row" % (LINE, num(T13, srow, 8)))
    ck("1.3 line split", num(T13, srow, 9) == 0
       and abs(num(T13, srow, 10) - o) < 1e-6,
       "TDD Cost 0, funded outside %.4f" % num(T13, srow, 10))
    other = hcol_row(w13, 8, "Other cost (this model)")
    pool = hcol_row(w13, 8, POOL_LBL)
    sig = hcol_row(w13, 8, "Significant Items - EGI")
    ltf = hcol_row(w13, 8, "Left to fund")
    ck("1.3 Significant Items - EGI reads the line",
       abs(num(T13, sig, 10) - o) < 1e-6, "%.4f" % num(T13, sig, 10))
    want = num(T13, other, 10) - num(T13, pool, 10) - num(T13, sig, 10)
    ck("1.3 Left to fund nets the EGI slice",
       abs(num(T13, ltf, 10) - want) < 1e-6,
       "%.4f = other cost %.4f less the %.4f funded - the slice reads 0"
       % (num(T13, ltf, 10), num(T13, other, 10), num(T13, sig, 10)))

    # 5  1.4
    w14 = wf[T14]
    r = hcol_row(w14, 8, SIG_LBL)
    egi24 = grid_row(wf[T24], "EGI TDD")
    _, _, cols24 = grid(wf[T24])
    pc = cols24["Funded outside TDD ($m)"]
    jf = str(w14.cell(r, 10).value)
    ck("1.4 Significant Items EGI", "$%s$%d" %
       (get_column_letter(pc), egi24) in jf and T24 in jf
       and abs(num(T14, r, 10) - num(T24, egi24, pc)) < 1e-6,
       "label moved, reads %.4f live off the 2.4 grid" % num(T14, r, 10))

    # 6  1.1
    w11 = wf[T11]
    r = hcol_row(w11, 8, SIG_LBL)
    egi21 = grid_row(wf[T21], "EGI Retail")
    _, _, cols21 = grid(wf[T21])
    got = num(T11, r, 10)
    ck("1.1 Significant Items EGI verified",
       abs(got - num(T21, egi21, cols21["Actual cost ($m)"])) < 1e-6,
       "reads %.4f, the EGI Retail block; untouched" % got)

    # controls and hygiene
    bad = []
    for ws in wf.worksheets:
        for row in ws.iter_rows():
            for c in row:
                t = c.value
                if isinstance(t, str) and "must be 0" in t:
                    for cc in range(c.column + 1, ws.max_column + 1):
                        v = wv[ws.title].cell(c.row, cc).value
                        if isinstance(v, (int, float)) and abs(v) > 1e-9:
                            bad.append((ws.title, c.row, cc, v))
    ck("every control reads 0", not bad, "%d exceptions %r" % (len(bad), bad[:4]))
    qv = wv[QA]
    fails = [r for r in range(5, 81)
             if isinstance(qv.cell(r, 5).value, (int, float))
             and abs(qv.cell(r, 5).value) > 1e-9]
    ck("4.0 Data QA all ties hold", not fails and
       (qv.cell(hcol_row(wf[QA], 2, "Checks failing"), 5).value or 0) == 0,
       "failing rows: %r" % fails)

    err, blank = wbio.audit(dst)
    _, blank_in = wbio.audit(src)
    new_blank = set((s, co) for s, co, _ in blank) \
        - set((s, co) for s, co, _ in blank_in)
    ck("no error cells, no new blank formulas", not err and not new_blank,
       "%d errors; %d blanks, all carried over from the input (%d there)"
       % (len(err), len(blank), len(blank_in)) if not new_blank else
       "%d errors; new blanks %r" % (len(err), sorted(new_blank)[:6]))
    n_dash = n_proj = 0
    for ws in wf.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    if any(d in c.value for d in DASHES):
                        n_dash += 1
                    if "to projects" in c.value.lower():
                        n_proj += 1
    ck("no en or em dashes, no 'to projects'", n_dash == 0 and n_proj == 0,
       "%d dashes, %d phrasings" % (n_dash, n_proj))
    zdash = [(T13, prow + 2, c) for c in (8, 9, 10, 11, 12)
             if re.search(r'(\\-|"-")',
                          ";".join(wf[T13].cell(prow + 2, c).number_format
                                   .split(";")[2:3]))]
    ck("no dash-for-zero format introduced", not zdash, "%r" % zdash)

    wf.close(); wv.close()
    if not all(checks):
        print("\nself-check FAILED")
        raise SystemExit(2)
    print("\nself-check clean: %d/%d PASS" % (sum(checks), len(checks)),
          flush=True)


if __name__ == "__main__":
    main(*sys.argv[1:3])
