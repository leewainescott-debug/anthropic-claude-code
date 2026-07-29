"""Actual cost against archetype cost, on the 1.x design tabs. Two designs.

The design tabs price squads from the archetype library and never say what those squads
actually cost, so the comparison could only be read on 2.x or 3.x. This puts it on the design
tab itself, driven by formula off the working tab's cost-after-decisions column, so it moves
the moment a lever is pulled.

Design A - inline. Two columns are appended to every squad table and every platform total
row: the actual cost after decisions, and the variance to the archetype. You read the
comparison on the row you are already reading. Nothing moves; K and L were empty on every
squad table. The portfolio's own actual-against-archetype table sits at the TOP of the tab,
beside the Budget vs TDD Cost box and on the box's own rows - see top_block below for why it
moved there from the foot of the tab and what changed about its shape.

Design B - one comparison table. Every existing table is left exactly as it is, and one
"Archetype against actual" table is added at the foot of each tab: every squad on the tab,
grouped by platform, with platform subtotals, the portfolio total, and the overhead roles
stated separately so the figure ties back to the working tab.

Both put the same three facts on the portfolio block: what the archetype prices, what those
squads actually cost after decisions, and the difference. Neither touches a single existing
formula.
"""
import copy
import json
import re

import openpyxl
from openpyxl.styles import Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter as L

import opts
from build_2xfix import DESIGN

NOTE_AL = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 1.x tab -> its working tab
PAIR = {v: k for k, v in DESIGN.items()}
# Every column on a 1.x tab already belongs to a table. The squad tables run B:J, the budget
# block above them runs I:N, and setting a width for one shrinks a column of the other - C
# went 26 to 16 and truncated "Configuration / Integration", K went 22 to 20 and squeezed the
# budget block's own last two columns. Nothing here sets a width. The three figures go under
# H, K and L, which is where the archetype cost already sits and where the squad tables have
# nothing, so every figure of a kind stays in one column all the way down the tab.
ARCH, ACT, VAR = 8, 11, 12                          # H, K, L
LAST = 12                                           # the band runs B:L
A_HDR = {ACT: "Actual cost after decisions ($m)", VAR: "Variance to archetype ($m)"}
# The comparison belongs in K and L on all ten tabs, so a reader flicking between them finds
# it in the same place every time. free_pair only moves off K when K or L is still carrying
# somebody else's content on a row this writes to - 1.6 carried the owner's own five scratch
# columns there - and the fallback is reported rather than made quietly, because a tab whose
# comparison sits in P and Q is not the same tab as its nine siblings.
HOME = (ACT, VAR)


def blocks(ws, wsv, limit=95):
    """Every platform block on a design tab: its squad rows and its total row.

    A block is a "Platform: x" bar, a header row whose column B reads "Squad", the squad
    rows, an optional platform overhead line, and a total row. A block with no squad row -
    a platform folded into another, or one that was removed - is skipped.
    """
    out = []
    r = 1
    while r <= min(ws.max_row, limit):
        b = str(wsv.cell(r, 2).value or "").strip()
        if not b.startswith("Platform:"):
            r += 1
            continue
        name = b[9:].strip()
        hdr = r + 1
        if str(wsv.cell(hdr, 2).value or "").strip() != "Squad":
            r += 1
            continue
        squads, total = [], None
        k = hdr + 1
        while k <= min(ws.max_row, limit):
            v = str(wsv.cell(k, 2).value or "").strip()
            if v.endswith(" Total"):
                total = k
                break
            if v.startswith("Platform:"):
                break
            # a squad row carries a squad type. Testing for a cost instead skipped the
            # EGI P&C row on 1.5, whose funded input the owner has not set yet, and pushed
            # its 0.24 into the residual line as though the squad had no row at all.
            # A squad whose type the owner has not set yet - 1.14's Cyber Uplift ships
            # with cream inputs empty - still has the archetype formula in H, and that is
            # what makes it a squad row rather than a stray note.
            if v and v != "Platform Overhead" and \
                    (str(wsv.cell(k, 3).value or "").strip()
                     or str(ws.cell(k, 8).value or "").startswith("=")):
                squads.append(k)
            k += 1
        if squads and total:
            # a block is comparable only if an archetype prices every squad in it. EGI P&C
            # on 1.5 has no size set, so its block has no archetype side at all and cannot
            # be added to one that has.
            comparable = all(isinstance(wsv.cell(s, 8).value, (int, float)) for s in squads)
            out.append({"name": name, "hdr": hdr, "squads": squads, "total": total,
                        "names": [str(wsv.cell(s, 2).value or "").strip() for s in squads],
                        "comparable": comparable})
        r = k + 1
    return out


def lookup(tab, lo, hi, row, after):
    """The squad's cost after decisions on its working tab, found by name."""
    return (f"=IFERROR(INDEX('{tab}'!${after}${lo}:${after}${hi},"
            f"MATCH($B{row},'{tab}'!$B${lo}:$B${hi},0)),\"-\")")


def diff(a, b):
    """b less a, stated only when both sides are figures.

    EGI P&C on 1.5 carries no archetype - the owner has not set its size - so K-H read the
    whole 0.24 as an overspend against nothing. A variance needs two figures on the same
    basis or it is not a variance, and the honest answer is a dash.
    """
    return f'=IF(AND(ISNUMBER({a}),ISNUMBER({b})),ROUND({b}-{a},6),"-")'


def priced(cells):
    """TRUE when every squad in a block carries an archetype cost."""
    return f"COUNT({','.join(cells)})={len(cells)}"


def _span(ws, r, spec, font, fill, al, height=None):
    """One row of a table, merged where a label needs more than its own column.

    Widths are never touched, so a label that will not fit in one column is given two or
    five. The bands run B:L on every table here, which is what the squad tables above
    already span once design A has added its two columns.
    """
    for c0, c1, text in spec:
        if c1 > c0:
            ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c1)
        x = ws.cell(r, c0)
        x.value = text
        for c in range(c0, c1 + 1):
            y = ws.cell(r, c)
            y.font, y.border = font, opts.BOX
            if fill:
                y.fill = opts.fl(fill)
        x.alignment = al
    if height:
        ws.row_dimensions[r].height = height
    return r + 1


def _head(ws, r, spec):
    """A header row that widens nothing and grows its own height instead."""
    lines = max(opts.wrap_lines(t, sum(ws.column_dimensions[L(c)].width or 8.43
                                       for c in range(c0, c1 + 1)))
                for c0, c1, t in spec if t)
    return _span(ws, r, spec, opts.HDRF, opts.NAVY, opts.CEN, max(32, 14 * lines + 6))


def free_pair(ws, wsv, blks, start=ACT, limit=40):
    """K and L, or the first adjacent pair after them that design A can write to.

    The comparison belongs in K and L on all ten tabs and the search starts there, so it
    lands there whenever K and L are clear. It only moves when they are not: the squad
    tables are not all the same width, 1.4, 1.5 and 1.6 carry five more columns the owner
    added - Nbr Archetype Roles, Published Roles, Review Outcome, Vacant Now, FY27 - and
    writing at a fixed K and L destroyed two of them on 1.6 without a single check noticing,
    because a typed number overwritten by a formula is still just a number. Moving is the
    guard, not the aim, and run() reports every tab that has to move.
    """
    rows, note = set(), set()
    for b in blks:
        rows.update([b["hdr"] - 1, b["hdr"], b["total"]] + b["squads"])
        edge = max((c for c in range(2, limit) if wsv.cell(b["hdr"], c).value is not None),
                   default=2)
        # a free-floating note past the table's own last column will be moved along the row,
        # so it must not push the columns further out. Counting it as occupied left K empty
        # and the figures one column adrift from the table on three tabs.
        note.update((r, c) for r in range(b["hdr"] + 1, b["total"] + 1)
                    for c in range(edge + 1, limit)
                    if isinstance(wsv.cell(r, c).value, str))
    for c in range(start, limit):
        if all((wsv.cell(r, k).value is None or (r, k) in note)
               for r in rows for k in (c, c + 1)):
            return c, c + 1
    raise SystemExit("no free column pair on this tab")


def clipped(ws, wsv, r, c):
    """A note to the left of column c that would stop being readable if c were filled.

    The owner's notes sit in a single cell and run across the empty columns beside them -
    "People in this program today cost 0.24m. Set the agreed cost in the cream cell." on
    1.6 needs four columns to show. Putting a figure in the fourth truncates it to "People
    in this program", which loses the instruction without touching a cell.
    """
    for k in range(c - 1, 1, -1):
        v = wsv.cell(r, k).value
        if v is None:
            continue
        if not isinstance(v, str):
            return None
        room = sum(ws.column_dimensions[L(x)].width or 8.43 for x in range(k, c))
        return k if opts.wrap_lines(v, room) > 1 else None
    return None


def shift_notes(ws, wsv, blks, act, var):
    """Move a note the new columns would truncate to the far side of them.

    Nothing is lost and nothing is reworded - the note stays on its own row, and it is the
    same move a reader would make by hand after widening a table under an annotation. The
    alternative, placing the columns beyond the widest note, left eleven blank columns
    between the table and its own figures on three tabs.

    Only a squad row is considered, and only a cell past the right-hand edge of that
    block's own header. Walking left from the new column on any row instead picked up
    "Funded outside TDD ($m)" - a header, not a note - and moved the table's own last
    column out of the table on six tabs.
    """
    out = []
    for b in blks:
        edge = max((c for c in range(2, act) if wsv.cell(b["hdr"], c).value is not None),
                   default=2)
        for r in range(b["hdr"] + 1, b["total"] + 1):
            # a note the columns will sit on top of has to move whether it would have been
            # clipped or not. Every row of the block, not only the squad rows: 1.9 carries
            # one on its overhead line and another on its total.
            k = next((c for c in range(edge + 1, var + 1)
                      if isinstance(wsv.cell(r, c).value, str)), None)
            if k is None:
                k = clipped(ws, wsv, r, act)
            if k is None or k <= edge:
                continue
            to = var + 1
            while wsv.cell(r, to).value is not None or ws.cell(r, to).value is not None:
                to += 1
            src, dst = ws.cell(r, k), ws.cell(r, to)
            dst.value, dst.font, dst.alignment = src.value, copy.copy(src.font), opts.LFT
            src.value = None
            out.append(f"note on row {r} moved {L(k)} -> {L(to)}, past the new columns")
    return out


def empty_col(ws, wsv, c, limit=95):
    """A column with nothing anywhere on the tab, so its width is nobody else's."""
    return all(wsv.cell(r, c).value is None
               for r in range(1, min(ws.max_row, limit) + 1))


def design_a(ws, wsv, blks, tab, lo, hi, after):
    act, var = free_pair(ws, wsv, blks)
    moved = shift_notes(ws, wsv, blks, act, var)
    for c in (act, var):
        # a width may only be set on a column no other table on this tab is using
        if empty_col(ws, wsv, c) and (ws.column_dimensions[L(c)].width or 8.43) < 20:
            ws.column_dimensions[L(c)].width = 22
    for b in blks:
        # the table just got two columns wider, so its bar has to follow it. The platform
        # bars are merged B:J, and a merge followed by painted cells reads as one bar.
        for c in (act, var):
            x = ws.cell(b["hdr"] - 1, c)
            x.fill, x.font = opts.fl(opts.BARC), opts.BARF
        _head(ws, b["hdr"], [(act, act, "Actual cost after decisions ($m)"),
                             (var, var, "Variance to archetype ($m)")])
        for r in b["squads"]:
            _m(ws, r, act, lookup(tab, lo, hi, r, after))
            _m(ws, r, var, diff(f"${L(ARCH)}{r}", f"${L(act)}{r}"))
        t = b["total"]
        # named cells, not a range: a block whose overhead line sits between its squads
        # would otherwise be summed with the overhead in it
        _m(ws, t, act, "=SUM(" + ",".join(f"${L(act)}{r}" for r in b["squads"]) + ")",
           bold=True)
        # the block total compares only when every squad in it is priced by an archetype,
        # or the archetype side is short a squad the actual side is carrying
        _m(ws, t, var, f'=IF({priced([f"${L(ARCH)}{r}" for r in b["squads"]])},'
                       f'{diff(f"${L(ARCH)}{t}", f"${L(act)}{t}")[1:]},"-")', bold=True)
        for c in (act, var):
            ws.cell(t, c).fill = copy.copy(ws.cell(t, ARCH).fill)
    return act, var, moved


# B:C the platform, D:G the squad, then the squad tables' own three cost columns and the
# two new ones, each under the column it already lives in further up the tab
B_SPEC = [(2, 3, "Platform"), (4, 7, "Squad"), (8, 8, "Total Squad Cost ($m)"),
          (9, 9, "TDD Cost ($m)"), (10, 10, "Funded outside TDD ($m)"),
          (11, 11, "Actual cost after decisions ($m)"),
          (12, 12, "Variance to archetype ($m)")]


def design_b(ws, blks, tab, lo, hi, after, anchor):
    """One table at the foot of the tab. Returns the row its portfolio block starts on."""
    r = (ws.max_row or 1) + 3
    r = opts.bar(ws, r, 2, LAST - 1, "Archetype against actual - every squad on this tab")
    r = _head(ws, r, B_SPEC)
    subs, unpriced = [], []
    for b in blks:
        st = r
        for src in b["squads"]:
            _span(ws, r, [(2, 3, b["name"]), (4, 7, f"=$B{src}")], opts.BODY, None, opts.LFT)
            # a blank archetype cell reads back as 0 through "=$H", which would print
            # $0.00m as though the archetype priced this squad at nothing
            for c in (ARCH, 9, 10):
                _m(ws, r, c, f'=IF({L(c)}{src}="","-",{L(c)}{src})')
            _m(ws, r, ACT, lookup(tab, lo, hi, src, after).replace(f"$B{src}", f"$D{r}"))
            _m(ws, r, VAR, diff(f"${L(ARCH)}{r}", f"${L(ACT)}{r}"))
            r += 1
        _span(ws, r, [(2, 7, f"{b['name']} total")] + [(c, c, None) for c in range(8, 13)],
              opts.BOLD, opts.GREY, opts.LFT)
        for c in (ARCH, 9, 10):
            cells = [f"${L(c)}{x}" for x in range(st, r)]
            _m(ws, r, c, f'=IF({priced(cells)},SUM({",".join(cells)}),"-")', bold=True)
        _m(ws, r, ACT, f"=SUM(${L(ACT)}{st}:${L(ACT)}{r-1})", bold=True)
        _m(ws, r, VAR, diff(f"${L(ARCH)}{r}", f"${L(ACT)}{r}"), bold=True)
        (subs if b["comparable"] else unpriced).append((f"${L(ARCH)}{r}", f"${L(ACT)}{r}"))
        r += 1
    return _foot(ws, r, subs, unpriced, tab, anchor, (ARCH, ACT, VAR),
                 [n for b in blks if not b["comparable"] for n in b["names"]])


def _line(ws, r, text, cols, arch, act, var, band=False):
    ca, cb, cv = cols
    _span(ws, r, [(2, ca - 1, text)] + [(c, c, None) for c in range(ca, cv + 1)],
          opts.BOLD if band else opts.BODY, opts.MID if band else None, opts.LFT)
    if band:
        for c in range(2, cv + 1):
            ws.cell(r, c).border = opts.TOPR
    for c, f in ((ca, arch), (cb, act), (cv, var)):
        _m(ws, r, c, f, bold=band)
    return r + 1


def _foot(ws, r, subs, unpriced, tab, anchor, cols, unpriced_names=()):
    """Design B's portfolio lines, at the foot of its own one big table.

    Design A no longer comes through here. Its block moved to the top of the tab and is now
    the three-line actual-against-archetype table - top_block below - because that is what
    the owner asked for. This is left as design B's ending because in design B the lines
    genuinely are the foot of the table above them, and B is not the design that ships.

    One label set, one order, one rule.

    Order, on all ten tabs and never varied:

        Squads priced by an archetype
        Squads with no archetype to price them
        Overhead roles in this portfolio
        Additional costs
        Total actual cost after decisions - ties to the working tab

    The rule for what appears: the first line and the total are always printed, because a
    portfolio always has squads and always has a total. Each of the three lines between them
    is printed only when its own figure is not zero - a line reading 0.00 states nothing the
    total does not already state, and the owner took the nil ones off himself. The test is
    the same test for all three, taken from the working tab's own cached figures, so a tab
    cannot drop a line its siblings keep for any reason other than having nothing to put on
    it. Every line the archetype does not price states a dash rather than a figure, so
    nothing on this block adds two different bases together.
    """
    A, oh = L(anchor["cols"]["after"]), anchor["overhead_row"]
    ca, cb, _ = cols
    K = L(cb)
    add = []
    r = _line(ws, r, "Squads priced by an archetype", cols,
              "=" + "+".join(f"N({a})" for a, _ in subs),
              "=" + "+".join(f"N({b})" for _, b in subs),
              diff(f"${L(ca)}{r}", f"${K}{r}"), band=True)
    add.append(r - 1)
    if unpriced and abs(_after(tab, anchor, unpriced_names)) > 1e-6:
        r = _line(ws, r, "Squads with no archetype to price them", cols, '="-"',
                  "=" + "+".join(f"N({b})" for _, b in unpriced), '="-"')
        add.append(r - 1)
    if oh and abs(_after(tab, anchor, rows=[oh])) > 1e-6:
        r = _line(ws, r, "Overhead roles in this portfolio", cols,
                  '="-"', f"=N('{tab}'!${A}${oh})", '="-"')
        add.append(r - 1)
    # squads the working tab carries with no row on this tab - the owner renamed the line
    # "Additional costs" and dropped it where it is nil. The zero-check control moved off
    # the tab with his edit; qa1x recomputes the same arithmetic in Python on every run.
    resid = f"=ROUND(N('{tab}'!${A}${anchor['total_row']})-" \
            + "-".join(f"${K}{x}" for x in add) + ",6)"
    if abs(_residual(ws.parent, tab, anchor, add, K)) > 1e-6:
        r = _line(ws, r, "Additional costs", cols, '="-"', resid, '="-"')
        add.append(r - 1)
    r = _line(ws, r, "Total actual cost after decisions - ties to the working tab", cols,
              '="-"', f"='{tab}'!${A}${anchor['total_row']}", '="-"', band=True)
    return add[0]


_WV = None
_NAMES = None


def _after(tab, anchor, names=(), rows=()):
    """After-decisions cost on the working tab, for named groups or for named rows.

    The footer rows written this run have no cached value yet, so what a footer line is
    going to say is read off the working tab's own cached figures instead. Same source for
    every line, so the suppress-when-nil rule is one test rather than three.
    """
    A = anchor["cols"]["after"]
    got = list(rows) + [anchor["srow"][g] for g in names if g in anchor["srow"]]
    tot = 0.0
    for rw in got:
        v = _WV[tab].cell(rw, A).value
        tot += v if isinstance(v, (int, float)) else 0
    return tot


def _residual(wb, tab, anchor, add_rows, K):
    """The residual the "Additional costs" line would carry, from cached values.

    It is the after-decisions cost of every working-tab group that has no squad row on
    this design tab - a one-person programme, a Leadership group. The footer rows written
    this run have no cached values yet, so the figure is derived from the working tab's
    own cached group rows instead.
    """
    A = anchor["cols"]["after"]
    resid = 0.0
    for g in anchor["squads"] + anchor["direct"] + anchor["nofig"]:
        if g in (_NAMES or set()):
            continue
        v = _WV[tab].cell(anchor["srow"][g], A).value
        resid += v if isinstance(v, (int, float)) else 0
    return resid


def _m(ws, r, c, f, fmt=None, bold=False):
    x = ws.cell(r, c)
    x.value = f
    x.number_format = fmt or opts.M2
    x.alignment, x.font = opts.RGT, (opts.BOLD if bold else opts.BODY)
    x.border = opts.BOX
    return x


# ------------------------------------------------- design A's portfolio block, up top
# The owner, on the block this used to write at the foot of every 1.x tab: "table at bottom
# of 1.x tabs needs to be up top. it's also not a clean table." Both halves are answered
# here, and the second is the harder one.
#
# Up top. The block is written into the empty band beside the Budget vs TDD Cost box, on the
# box's own rows, so the two read as a matched pair: same bar row, same header row. Nothing
# is inserted and nothing is moved. Dozens of absolute references point into these tabs -
# 0.2 reads C9 and D9, the 2.x tabs read the design H rows, 4.0 reads F9 - and one inserted
# row would break them silently, so this is a relocation into cells that are proved empty
# first and never an insertion.
#
# A clean table. The first answer to that was a five-line decomposition of the actual cost -
# squads priced by an archetype, squads with none, overhead, additional costs, total. It was
# a clean table and it answered the wrong question: five ways of splitting one number, and
# the archetype the whole tab is built on nowhere on it. The owner mocked what he wanted
# over the shipped file and it is three lines, not five:
#
#     Actuals vs archetype
#     What the cost covers            Roles    Cost ($m)
#     Actual portfolio                   70        14.01
#     Archetype portfolio                52        14.26
#     Variance                           18        (0.26)
#
# One question, asked once. Both sides come off the working tab's own "Total portfolio" row -
# the actual from the cost-after-decisions column, the archetype from the archetype-cost
# column that 4.0 already ties to the design tab's own F9 - so the block moves the moment a
# lever is pulled and there is nothing on it to keep in step by hand. The roles column does
# the same in headcount: total roles against archetype roles, which is the fact a cost
# figure on its own never gave a reader.
#
# The decomposition is not lost. It is on the tab already, one line per platform, in K and L
# beside every squad table, with the portfolio's own split on the Portfolio Summary where
# post2707 sets Total Cost, Actuals and the variance between them side by side.
OLD_LINES = ("Squads priced by an archetype",
             "Squads with no archetype to price them",
             "Overhead roles in this portfolio",
             "Additional costs",
             "Total actual cost after decisions")
LINES = ("Actual portfolio", "Archetype portfolio", "Variance")
# K:L for the label, M the roles count, N the cost. K and L are the only two columns whose
# width is the same on all ten tabs - the family profile sets them - and merged they are
# wide enough for the longest line on one line. M and N take whatever width the tab already
# gives them: both carry the owner's notes further down, and this file does not set widths.
TOP_C0, TOP_LAB, TOP_ROLES, TOP_COST = 11, 12, 13, 14
TOP_HEADS = ("What the cost covers", "Roles", "Cost ($m)")
TOP_BAR = "Actuals vs archetype"
TOP_ROWS = 2 + len(LINES)                       # the bar, the header, the three lines
# the family's header depth. Nine of the ten tabs already have it on this row - it is the
# budget box's own header row - and 1.7 has 28.5, so stating it makes all ten the same
# table. A row is only ever raised here, never lowered: raising one can clip nothing.
HDR_H = 34
# every bar this block has ever shipped under. A re-run has to recognise the shape that is
# on the tab, not the shape it is about to write: the band is refused when it is occupied,
# so a five-line table left in place would stop the three-line one being written at all and
# the tab would keep the old design with no error anywhere.
TOP_BARS = ("Actual cost after decisions", TOP_BAR)
OLD_BAR = "Archetype against actual - this portfolio"
# the working tab's own headers for the four figures this block quotes. The anchors carry a
# column map, but it was written before the archetype side of this block existed and a map
# is not evidence that the column is on the tab: the header is read first and the anchor is
# the fallback, so a reworded header is reported rather than silently quoting column E of
# whatever happens to be there.
W_HEADS = {"roles": "Total roles", "aroles": "Archetype roles",
           "acost": "Archetype cost ($m)", "after": "Squad cost after decisions ($m)",
           "actual": "Actual cost ($m)"}


def budget_anchor(ws, wsv, limit=13):
    """The row the Budget vs TDD Cost bar sits on: 4 on nine tabs, 5 on 1.7.

    Found by the label rather than by a row number, which is what carries 1.7's one-row
    offset across without this file having to know 1.7 exists.
    """
    for r in range(1, limit):
        for c in range(7, 13):
            if str(wsv.cell(r, c).value or ws.cell(r, c).value or "").strip() == \
                    "Budget vs TDD Cost":
                return r
    return None


def band_free(ws, wsv, r0, rows, c0, c1):
    """Everything this block would write on, that is not already empty.

    Both workbooks are read: a cell can be empty of formula and still be carrying a cached
    value, and a note of his read one way and not the other is still his note. Merges count
    too - writing into half of one is how a table ends up with a cell that cannot be typed
    in.
    """
    busy = []
    for r in range(r0, r0 + rows):
        for c in range(c0, c1 + 1):
            if ws.cell(r, c).value is not None or wsv.cell(r, c).value is not None:
                busy.append(f"{L(c)}{r}")
    for m in ws.merged_cells.ranges:
        if m.min_row <= r0 + rows - 1 and m.max_row >= r0 \
                and m.min_col <= c1 and m.max_col >= c0:
            busy.append(f"merge {m}")
    return busy


def _wipe(sheets, r, c0, c1):
    """One row of a block, cleared on every workbook that is carrying it.

    Both the formula sheet and the cached-value sheet: band_free reads both, so a cell
    stripped of its formula and left with last run's cached number still reads as occupied
    and the band is refused. That is the whole failure this function exists to stop, and
    clearing one sheet of the two reproduces it exactly.
    """
    for ws in sheets:
        for c in range(c0, c1 + 1):
            x = ws.cell(r, c)
            x.value, x.fill, x.border, x.font = None, PatternFill(), Border(), opts.BODY
            x.number_format, x.alignment = "General", Alignment()


def _unmerge(sheets, r0, r1, c0, c1):
    """Every merge wholly inside the rectangle, released on both workbooks.

    Both again, and for a harder reason than _wipe's: openpyxl will not let a value be set
    on a cell that a merge has swallowed, so a merge left in place on the cached-value
    sheet stops that sheet being cleared at all.
    """
    for ws in sheets:
        for m in [str(x) for x in ws.merged_cells.ranges
                  if r0 <= x.min_row and x.max_row <= r1
                  and c0 <= x.min_col and x.max_col <= c1]:
            ws.unmerge_cells(m)


def clear_old_block(ws, wsv):
    """Strip whichever shape of this block an earlier run of this script left behind.

    Two shapes have shipped and both are looked for. The first lived at the foot of the tab
    under an "Archetype against actual - this portfolio" bar. The second is the five-line
    table up top under an "Actual cost after decisions" bar, and the three-line table this
    file writes now goes in exactly the cells that one occupies - so leaving it in place
    does not double the block, it stops the new one being written at all. band_free refuses
    an occupied band and says so, which is the right answer to somebody else's content and
    the wrong answer to this file's own last output.

    Nothing in chain2 writes either shape, so on a clean build there is nothing here to
    find. It matters the moment anyone re-runs chainA2 over its own output, or runs this
    file over a workbook built by the version of it that shipped before the owner's mock.

    The top block is cleared in its own four columns and nowhere else. Its rows are shared -
    the Portfolio Summary is in B:F on them and the budget box in H:I - so a clearer that
    swept the row, or reset the row height, would take out two tables to remove one.
    """
    out, sheets = [], (ws, wsv)
    top = next((r for r in range(1, ws.max_row + 1)
                if str(ws.cell(r, 2).value or "").strip().startswith(OLD_BAR)), None)
    if top is not None:
        end = top
        for r in range(top, min(ws.max_row, top + 12) + 1):
            lab = str(ws.cell(r, 2).value or "").strip()
            if lab.startswith(OLD_BAR) or lab == "Line" \
                    or any(lab.startswith(x) for x in OLD_LINES):
                end = r
        _unmerge(sheets, top, end, 1, ws.max_column + 1)
        for r in range(top, end + 1):
            _wipe(sheets, r, 2, LAST)
            ws.row_dimensions[r].height = None
        out.append(f"{ws.title}: the old block at the foot of the tab removed from rows "
                   f"{top}-{end} - values, labels, fills and borders")
    # the block up top, found by its bar wherever the bar happens to sit: 1.7 is one row
    # lower than its nine siblings and the block has to be recognised there too
    found = next(((r, c) for r in range(1, min(ws.max_row, 30) + 1)
                  for c in range(2, 21)
                  if str(ws.cell(r, c).value or "").strip() in TOP_BARS), None)
    if found is None:
        return out
    r0, c0 = found
    c1 = c0 + TOP_COST - TOP_C0
    labels = set(LINES) | set(OLD_LINES) | set(TOP_HEADS)
    end = r0
    for r in range(r0 + 1, min(ws.max_row, r0 + 2 + len(OLD_LINES)) + 1):
        if str(ws.cell(r, c0).value or "").strip() in labels:
            end = r
    _unmerge(sheets, r0, end, c0, c1)
    for r in range(r0, end + 1):
        _wipe(sheets, r, c0, c1)
    out.append(f"{ws.title}: the block already up top removed from "
               f"{L(c0)}{r0}:{L(c1)}{end} - its own four columns only, so the summary and "
               f"the budget box on those rows are untouched")
    return out


def w_col(wsw, anchor, key, title, out):
    """A column on the working tab, by the header it carries, with the anchor behind it."""
    hdr, fallback = anchor.get("header_row"), anchor["cols"][key]
    want = W_HEADS[key]
    if hdr:
        for c in range(2, 26):
            if str(wsw.cell(hdr, c).value or "").strip() == want:
                if c != fallback:
                    out.append(f"  {title}: {wsw.title} heads {L(c)} {want!r} where the "
                               f"anchors say {L(fallback)} - the header is followed")
                return c
    out.append(f"  {title}: no column headed {want!r} on {wsw.title} row {hdr} - "
               f"the anchor's {L(fallback)} is used instead")
    return fallback


def top_block(ws, wsv, tab, anchor, out):
    """Actual against archetype, at the top of the tab beside the budget box.

    Three lines and never any other number of them, on all eleven tabs. Every figure is a
    live read of one cell on the working tab's "Total portfolio" row, so the table cannot
    drift from the tab it is quoting and there is nothing on it to keep in step by hand.
    """
    r0 = budget_anchor(ws, wsv)
    if r0 is None:
        out.append(f"  {ws.title}: NO ANCHOR - nothing on this tab is labelled 'Budget vs "
                   f"TDD Cost', so the actuals table was not written")
        return None
    busy = band_free(ws, wsv, r0, TOP_ROWS, TOP_C0, TOP_COST)
    if busy:
        out.append(f"  {ws.title}: NOT UNIFORM - "
                   f"{L(TOP_C0)}{r0}:{L(TOP_COST)}{r0 + TOP_ROWS - 1} is carrying "
                   f"{', '.join(busy[:6])}, so the actuals table was not written here. "
                   f"Nothing was overwritten.")
        return None

    wsw = _WV[tab]
    tot = anchor["total_row"]
    # the Actual line quotes the ACTUAL cost column, not the after-decisions one. A row
    # labelled "Actual portfolio" that silently rewrites itself when a lever is pulled
    # stops meaning what it says - the after-decisions figure lives under labels that say
    # "after decisions" (the squad K columns, 3.1's G, Exec's decisions block).
    A = L(w_col(wsw, anchor, "actual", ws.title, out))    # O, actual cost
    F = L(w_col(wsw, anchor, "roles", ws.title, out))     # F, total roles
    E = L(w_col(wsw, anchor, "aroles", ws.title, out))    # E, archetype roles
    NC = L(w_col(wsw, anchor, "acost", ws.title, out))    # N, archetype cost

    r = opts.bar(ws, r0, TOP_C0, TOP_COST - TOP_C0 + 1, TOP_BAR)

    hdr = r
    ws.merge_cells(start_row=hdr, start_column=TOP_C0, end_row=hdr, end_column=TOP_LAB)
    for c, t in zip((TOP_C0, TOP_ROLES, TOP_COST), TOP_HEADS):
        ws.cell(hdr, c).value = t
    for c in range(TOP_C0, TOP_COST + 1):
        y = ws.cell(hdr, c)
        y.fill, y.font, y.border, y.alignment = (opts.fl(opts.NAVY), opts.HDRF, opts.BOX,
                                                 opts.CEN)
    # one fixed depth, so the eleven tabs are the same table. 1.7's row is 28.5 where the
    # others are already 34; raising a row can clip nothing, lowering one can.
    ws.row_dimensions[hdr].height = max(ws.row_dimensions[hdr].height or 0, HDR_H)

    at = {name: hdr + 1 + i for i, name in enumerate(LINES)}
    r1, r2, r3 = (at[n] for n in LINES)
    K, M = L(TOP_COST), L(TOP_ROLES)
    # the reads are direct, not wrapped in N(): a working tab that states a dash where it
    # has no figure has to reach this table as a dash. N() would turn it into a zero, the
    # money format would print that zero as "-", and a variance would then be struck
    # against a number nobody wrote.
    cost = {r1: f"='{tab}'!${A}${tot}",
            r2: f"='{tab}'!${NC}${tot}",
            r3: diff(f"${K}${r2}", f"${K}${r1}")}
    roles = {r1: f"='{tab}'!${F}${tot}",
             r2: f"='{tab}'!${E}${tot}",
             r3: diff(f"${M}${r2}", f"${M}${r1}")}
    for name, rr in at.items():
        last = rr == r3
        ws.merge_cells(start_row=rr, start_column=TOP_C0, end_row=rr, end_column=TOP_LAB)
        ws.cell(rr, TOP_C0).value = name
        # one decimal, not integers: the archetype side is FTE (31.6, 37.5), and a Roles
        # column that rounds to whole numbers stops adding on the page - 33 - 31 printing
        # as 3 where the truth is 2.5 is exactly the kind of gap the owner checks for
        _m(ws, rr, TOP_ROLES, roles[rr], fmt=opts.C1, bold=last)
        _m(ws, rr, TOP_COST, cost[rr], bold=last)
        for c in range(TOP_C0, TOP_COST + 1):
            y = ws.cell(rr, c)
            y.font = opts.BOLD if last else opts.BODY
            # the variance is the line the table exists to state, so it is banded and ruled
            # off the way every other table on these tabs bands the line that answers it
            y.border = opts.TOPR if last else opts.BOX
            if last:
                y.fill = opts.fl(opts.MID)
        ws.cell(rr, TOP_C0).alignment = opts.LFT
    return r0


COE_NOTE = ("These are centres of excellence, funded by allocation rather than priced by an "
            "archetype, so there is no archetype cost to compare against. The actual cost "
            "of these groups is on 3.4 COE Detail; the cost after the decisions set today "
            "is on 3.1 Cost Bridge and the working tabs.")


# a sentence under the summary block, as against a row label with figures beside it
NOTE_MIN = 40


def own_note(ws, wsv, r, last=15):
    """True when row r under the summary is somebody's own written note.

    A note is a literal sentence in B with nothing beside it on its own row. A row label -
    "COE - Cyber allocation ($m) - 0.2 Data Config" is 44 characters - always has its figure
    in the cell next to it, so the figure is what tells the two apart, not the length.
    """
    v = ws.cell(r, 2).value
    if not isinstance(v, str) or v.startswith("=") or len(v.strip()) < NOTE_MIN:
        return False
    return all(ws.cell(r, c).value is None and wsv.cell(r, c).value is None
               for c in range(3, last + 1))


def coe_note(wb, wv):
    """The three COE tabs say why they carry no comparison.

    They group by department - Business Partnering is Commercial plus TDD Business Partner -
    while the working tab groups by squad, so the split cannot be taken from either tab
    without inventing a mapping. Saying so is better than a silent gap on three of thirteen
    tabs.

    Where the owner has written his own basis note in that slot - "Planned spend is net of
    the Business Partner FTEs funded inside the portfolios" on 1.11 and 1.12 - his sentence
    is the answer and this one is not written at all. Two notes under one summary block, one
    of them generic, is the tab answering the same question twice.
    """
    out = []
    for one in ("1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles"):
        if one not in wb.sheetnames:
            continue
        ws, wsv = wb[one], wv[one]
        tot = next((r for r in range(1, min(ws.max_row, 30) + 1)
                    if str(wsv.cell(r, 2).value or "").strip() == "Total"), None)
        if tot is None:
            out.append(f"{one}: no Total row on the summary table")
            continue
        own = next((k for k in range(tot + 1, tot + 9) if own_note(ws, wsv, k)), None)
        if own is not None:
            out.append(f"{one}: the basis note already on row {own} is kept - "
                       f"the generic note is not written")
            continue
        r = next((k for k in range(tot + 1, tot + 9)
                  if not str(wsv.cell(k, 2).value or "").strip()
                  and not str(ws.cell(k, 2).value or "").strip()), None)
        if r is None:                                        # never write over a live row
            out.append(f"{one}: no free row under the summary")
            continue
        ws.cell(r, 2).value = COE_NOTE
        ws.cell(r, 2).font = opts.BODY
        ws.cell(r, 2).alignment = NOTE_AL
        ws.row_dimensions[r].height = 14 * opts.wrap_lines(COE_NOTE, 60) + 4
        out.append(f"{one}: note on row {r} - no archetype prices a COE")
    return out


def run(src, dst, variant="A", anchors="anchors_final.json"):
    global _WV
    wb = openpyxl.load_workbook(src)
    wv = openpyxl.load_workbook(src, data_only=True)
    _WV = wv
    a = json.load(open(anchors))
    # the working tabs were renamed after the anchors were written, so map by portfolio
    live = {str(wv[t]["C3"].value): t for t in wb.sheetnames if t.startswith("2.")}
    out = []
    for one in [t for t in wb.sheetnames if re.match(r"^1\.\d+ ", t) and t in PAIR]:
        anchor = next(x for x in a.values() if PAIR[one] == x["tab"])
        tab = live[anchor["pf"]]
        lo, hi = anchor["first_squad"], anchor["total_row"]
        after = L(anchor["cols"]["after"])
        ws, wsv = wb[one], wv[one]
        blks = blocks(ws, wsv)
        if not blks:
            out.append(f"{one}: no platform block with a squad in it")
            continue
        global _NAMES
        _NAMES = {str(wsv.cell(s_, 2).value or "").strip()
                  for b in blks for s_ in b["squads"]}
        if variant == "A":
            act, var, moved = design_a(ws, wsv, blks, tab, lo, hi, after)
            out += clear_old_block(ws, wsv)
            r0 = top_block(ws, wsv, tab, anchor, out)
            where = f"in {L(act)} and {L(var)}, the actuals table at " \
                    f"{L(TOP_C0)}{r0}:{L(TOP_COST)}{r0 + TOP_ROWS - 1}" if r0 \
                else f"in {L(act)} and {L(var)}, NO actuals table"
            out += [f"  {one}: {m}" for m in moved]
            if (act, var) != HOME:
                out.append(f"  {one}: NOT UNIFORM - {L(HOME[0])} and {L(HOME[1])} are "
                           f"still carrying other content on the squad-table rows, so the "
                           f"comparison fell back to {L(act)} and {L(var)}. Nothing was "
                           f"overwritten.")
        else:
            design_b(ws, blks, tab, lo, hi, after, anchor)
            where = "in one table at the foot"
        out.append(f"{one}: {len(blks)} platform blocks, "
                   f"{sum(len(b['squads']) for b in blks)} squads, actual against "
                   f"archetype from {tab}, {where}")
    out += coe_note(wb, wv)
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else "A"):
        print("  ", x)
