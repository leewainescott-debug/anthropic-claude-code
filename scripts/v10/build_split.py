"""Stage 9: separate delivery squads from overhead on every 2.x tab, so archetype vs
actual is a like-for-like comparison and no dollar is left over.

The problem this fixes. Until now every 2.x tab grouped all of a portfolio's roles by
squad, including a "Leadership" row holding Technology Managers, Delivery Managers and
Heads of Technology. The archetype columns then compared a squad's designed cost against
a population that also contained overhead roles, and the Leadership row had no archetype
at all - so the tab's archetype total and its actual total were measuring different sets
of people. That is the leftover-dollars problem.

The structure now. Every role is classified once in REVIEW column AR as either "Squad" or
one of five overhead lines, and each 2.x tab shows two blocks:

  Block A  delivery squads   467 roles  $98.779m  compared to the squad archetype
  Block B  overhead lines     58 roles  $16.335m  compared to the overhead allowance
  Total                      525 roles $115.113m  ties to the ledger

Block A is now pure: the roles in a squad row are exactly the roles the archetype prices.
Block B names the overhead sitting inside the portfolio instead of hiding it in a squad.
The 8 GMs are the only overhead with no role in the ledger and stay on 3.2 alone.

Vacancy behaviour, per Lee: a vacancy that is offshored is still a vacancy and still a
role. Offshore changes cost only - it never moves a headcount. Hold is the only lever that
removes a role from the after-decisions count.
"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import model
from build_2xfix import (CNT, HDR_FILL, HDR_FONT, MON, SUB_FILL, DESIGN,
                         squad_table_bounds)

REV = f"'{model.REVIEW}'"
LR = model.LAST_ROW
A3 = "'0.3 Squad Archetypes'"
BOLD = Font(bold=True)
ITAL = Font(italic=True)
OH_LINES = ["Head of Technology", "Business Partner", "Domain Architect",
            "Delivery Manager", "Technology Manager"]

HDRS = {
    "B": "Squad", "C": "Archetype type", "D": "Archetype roles",
    "E": "Roles (incl. vacant)", "F": "Variance to archetype (roles)",
    "G": "Vacant", "H": "Vacancies to hire or offshore", "I": "Vacancies on hold",
    "J": "Roles after decisions", "K": "Flag",
    "L": "Archetype cost ($m)", "M": "Actual cost ($m)",
    "N": "Variance to archetype ($m)", "O": "Cost after decisions ($m)",
    "P": "Change from decisions ($m)", "Q": "Vacant cost ($m)", "R": "Archetype size",
    "S": "Squad name on design tab",
}
ROLE_HDRS = {"B": "Name", "C": "Role", "D": "Status", "E": "Vacancy lever",
             "F": "Squad or overhead line", "G": "Cost if hired ($)",
             "H": "Effective cost ($)"}


def build(wb, roles):
    out = {}
    for tab, pf in model.TAB_PORTFOLIO.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        mine = [r for r in roles if r["portfolio"] == pf]
        squads = sorted({r["squad"] for r in mine if r["line"] == "Squad"},
                        key=lambda s: (-sum(1 for r in mine
                                            if r["line"] == "Squad" and r["squad"] == s), s))
        ohs = [l for l in OH_LINES if any(r["line"] == l for r in mine)]

        title = ws.cell(2, 2).value
        for r in range(4, ws.max_row + 1):
            for c in range(1, max(ws.max_column, 20) + 1):
                if ws.cell(r, c).value is not None:
                    ws.cell(r, c).value = None
        for dv in list(ws.data_validations.dataValidation):
            ws.data_validations.dataValidation.remove(dv)

        hdr = 5
        a0 = 6
        a1 = a0 + len(squads) - 1
        asub = a1 + 1                      # delivery squad subtotal
        b0 = asub + 2
        bhdr = b0 - 1
        b1 = b0 + len(ohs) - 1
        bsub = b1 + 1                      # overhead subtotal
        tot = bsub + 2
        m1, m2 = tot + 2, tot + 3
        rl_hdr = m2 + 3
        rl0, rl1 = rl_hdr + 1, rl_hdr + len(mine)
        RL = lambda c: f"${c}${rl0}:${c}${rl1}"

        ws["B4"] = ("Delivery squads are compared to the squad archetype; overhead roles "
                    "are compared to the allowance on Lists. The two blocks together are "
                    "the whole portfolio, so every role and every dollar lands once.")
        ws["B4"].font = ITAL
        for col, h in HDRS.items():
            c = ws[f"{col}{hdr}"]
            c.value = h
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(wrap_text=True, vertical="bottom")

        dt = DESIGN.get(tab)
        lo = hi = None
        if dt:
            lo, hi = squad_table_bounds(wb, dt)

        def common(r):
            ws[f"E{r}"] = f"=COUNTIFS({RL('F')},$B{r})"
            ws[f"G{r}"] = f'=COUNTIFS({RL("F")},$B{r},{RL("D")},"Vacant")'
            # a vacancy stays a vacancy whether it is hired or offshored
            ws[f"H{r}"] = (f'=COUNTIFS({RL("F")},$B{r},{RL("D")},"Vacant",{RL("E")},"Hire")'
                           f'+COUNTIFS({RL("F")},$B{r},{RL("D")},"Vacant",{RL("E")},"Offshore")')
            ws[f"I{r}"] = f'=COUNTIFS({RL("F")},$B{r},{RL("D")},"Vacant",{RL("E")},"Hold")'
            # Hold is the only lever that removes a role; Offshore keeps the headcount
            ws[f"J{r}"] = f'=$E{r}-COUNTIFS({RL("F")},$B{r},{RL("E")},"Hold")'
            ws[f"M{r}"] = f"=SUMIFS({RL('G')},{RL('F')},$B{r})/1000000"
            ws[f"O{r}"] = f"=SUMIFS({RL('H')},{RL('F')},$B{r})/1000000"
            ws[f"P{r}"] = f"=$O{r}-$M{r}"
            ws[f"Q{r}"] = f'=SUMIFS({RL("G")},{RL("F")},$B{r},{RL("D")},"Vacant")/1000000'

        # ---- Block A: delivery squads, compared to the archetype ----
        for i, sq in enumerate(squads):
            r = a0 + i
            ws[f"B{r}"] = sq
            ws[f"S{r}"] = f"=IFERROR(INDEX(Lists!$AA:$AA,MATCH($B{r},Lists!$Z:$Z,0)),$B{r})"
            if dt:
                m = f"MATCH($S{r},'{dt}'!$B${lo}:$B${hi},0)"
                ws[f"C{r}"] = (f"=IFERROR(INDEX('{dt}'!$C${lo}:$C${hi},{m}),"
                               f'"No archetype for this squad")')
                ws[f"R{r}"] = (f"=IFERROR(IF(INDEX('{dt}'!$D${lo}:$D${hi},{m})=\"\",\"-\","
                               f"INDEX('{dt}'!$D${lo}:$D${hi},{m})),\"-\")")
            else:
                ws[f"C{r}"] = '="COE - measured against budget"'
                ws[f"R{r}"] = '="-"'
            key = f'$C{r}&"|"&$R{r}'
            ws[f"D{r}"] = f'=IFERROR(INDEX({A3}!$F$5:$F$23,MATCH({key},{A3}!$A$5:$A$23,0)),"-")'
            ws[f"L{r}"] = f'=IFERROR(INDEX({A3}!$G$5:$G$23,MATCH({key},{A3}!$A$5:$A$23,0)),"-")'
            common(r)
            ws[f"F{r}"] = f'=IFERROR($E{r}-$D{r},"-")'
            ws[f"N{r}"] = f'=IFERROR($M{r}-$L{r},"-")'
            ws[f"K{r}"] = (f'=IF(NOT(ISNUMBER($D{r})),"No archetype - see 0.3",'
                           f'IF($J{r}>$D{r},"Over archetype after decisions",'
                           f'IF($J{r}<$D{r},"Under archetype","On archetype")))')
        ws[f"B{asub}"] = "Delivery squads"
        ws[f"B{asub}"].font = BOLD
        for col in "DEFGHIJLMNOPQ":
            ws[f"{col}{asub}"] = f"=SUM({col}{a0}:{col}{a1})"

        # ---- Block B: overhead roles that sit inside this portfolio ----
        ws[f"B{bhdr}"] = "Overhead roles inside this portfolio - compared to the allowance"
        ws[f"B{bhdr}"].font = ITAL
        for i, l in enumerate(ohs):
            r = b0 + i
            ws[f"B{r}"] = l
            ws[f"C{r}"] = '="Overhead - see 3.2"'
            ws[f"R{r}"] = '="-"'
            ws[f"D{r}"] = '="-"'
            ws[f"L{r}"] = '="-"'
            common(r)
            ws[f"F{r}"] = '="-"'
            ws[f"N{r}"] = '="-"'
            ws[f"K{r}"] = '="Overhead"'
        ws[f"B{bsub}"] = "Overhead roles"
        ws[f"B{bsub}"].font = BOLD
        for col in "EGHIJMOPQ":
            # 2.11 and 2.14 carry no overhead roles at all; an empty block must read 0,
            # not SUM over an inverted range
            ws[f"{col}{bsub}"] = (f"=SUM({col}{b0}:{col}{b1})" if ohs else 0)

        # ---- Total: the two blocks are the whole portfolio ----
        ws[f"B{tot}"] = "Total portfolio"
        ws[f"B{tot}"].font = BOLD
        for col in "EGHIJMOPQ":
            ws[f"{col}{tot}"] = f"=${col}{asub}+${col}{bsub}"
        for col in "DL":
            ws[f"{col}{tot}"] = f"=${col}{asub}"
        ws[f"F{tot}"] = f'=IFERROR($E{asub}-$D{asub},"-")'
        ws[f"N{tot}"] = f'=IFERROR($M{asub}-$L{asub},"-")'
        ws[f"B{tot+1}"] = ("Archetype columns D, F, L and N cover the delivery squads only - "
                           "overhead is not priced by a squad archetype.")
        ws[f"B{tot+1}"].font = ITAL

        ws[f"B{m1}"] = "Control - roles vs the ledger, must be 0"
        ws[f"E{m1}"] = f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},"{pf}")-$E{tot}'
        ws[f"B{m2}"] = "Control - cost vs the ledger ($m), must be 0"
        ws[f"M{m2}"] = (f'=ROUND(SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},"{pf}")'
                        f"/1000000-$M{tot},6)")
        for rr in (m1, m2):
            ws[f"B{rr}"].font = ITAL

        # ---- role list ----
        ws.cell(rl_hdr - 1, 2).value = f"{pf} roles - one row per named role, from REVIEW"
        ws.cell(rl_hdr - 1, 2).font = ITAL
        for col, h in ROLE_HDRS.items():
            c = ws[f"{col}{rl_hdr}"]
            c.value = h
            c.font = BOLD
            c.fill = SUB_FILL
            c.alignment = Alignment(wrap_text=True, vertical="bottom")
        ordered = sorted(mine, key=lambda x: (x["line"] != "Squad",
                                              [*squads, *ohs].index(x["squad"] if x["line"]
                                                                    == "Squad" else x["line"]),
                                              x["vacant"], x["name"]))
        for i, rec in enumerate(ordered):
            r, src = rl0 + i, rec["row"]
            ws[f"B{r}"] = f"=INDEX({REV}!$B:$B,{src})"
            ws[f"C{r}"] = f"=INDEX({REV}!$C:$C,{src})"
            ws[f"D{r}"] = f"=INDEX({REV}!$AK:$AK,{src})"
            ws[f"E{r}"] = rec["lever"]
            # one row per role, landing in its squad OR its overhead line, never both
            ws[f"F{r}"] = (f'=IF(INDEX({REV}!$AR:$AR,{src})="Squad",'
                           f"INDEX({REV}!$AP:$AP,{src}),INDEX({REV}!$AR:$AR,{src}))")
            ws[f"G{r}"] = f"=INDEX({REV}!$AA:$AA,{src})"
            ws[f"H{r}"] = (f"=$G{r}*IFERROR(INDEX(Lists!$AD:$AD,"
                           f"MATCH($E{r},Lists!$AC:$AC,0)),1)")
            for col in "GH":
                ws[f"{col}{r}"].number_format = "#,##0"
        for r in range(a0, tot + 1):
            for col in "DEFGHIJ":
                ws[f"{col}{r}"].number_format = CNT
            for col in "LMNOPQ":
                ws[f"{col}{r}"].number_format = MON

        dv = DataValidation(type="list", formula1="=Lists!$AC$2:$AC$5", allow_blank=False)
        dv.errorTitle, dv.error = "Vacancy lever", "Choose a lever from Lists AC2:AC5."
        ws.add_data_validation(dv)
        dv.add(f"E{rl0}:E{rl1}")
        ws.cell(2, 2).value = title
        out[tab] = dict(pf=pf, a0=a0, a1=a1, asub=asub, b0=b0, b1=b1, bsub=bsub,
                        tot=tot, rl0=rl0, rl1=rl1, squads=squads, ohs=ohs)
    return out


def load_roles(path):
    """Roles with their overhead line, straight from the REVIEW derived columns."""
    wv = openpyxl.load_workbook(path, data_only=True)
    ws = wv[model.REVIEW]
    out = []
    for i in range(2, LR + 1):
        nm = ws.cell(i, 2).value
        if nm is None or str(nm).strip() == "":
            continue
        cost = ws.cell(i, 27).value
        out.append(dict(row=i, name=str(nm).strip(),
                        portfolio=ws.cell(i, 36).value,        # AJ
                        status=ws.cell(i, 37).value,           # AK
                        squad=ws.cell(i, 42).value,            # AP
                        line=ws.cell(i, 44).value,             # AR
                        cost=cost if isinstance(cost, (int, float)) else 0.0,
                        vacant=str(ws.cell(i, 37).value) == "Vacant"))
    for r in out:
        r["lever"] = "Hold" if r["vacant"] else "Filled"
    wv.close()
    return out


def run(src, dst):
    roles = load_roles(src)
    wb = openpyxl.load_workbook(src)
    geom = build(wb, roles)
    wb.save(dst)
    return geom


if __name__ == "__main__":
    g = run("v4p.xlsx", "s1.xlsx")
    print(f"rebuilt {len(g)} tabs")
    for t, d in list(g.items())[:3]:
        print(f"   {t}: {len(d['squads'])} squads + {len(d['ohs'])} overhead lines, "
              f"roles {d['rl0']}-{d['rl1']}")
