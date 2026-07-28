"""Design pass over the 1.x, 0.2 and 0.3 tabs, from a rendered review of the candidate.

Nothing here changes a number. Every item was raised by someone looking at the rendered
tabs, and the biggest of them could only be seen that way:

  dozens of bar and section cells on the portfolio tabs carried a THEME fill rather than
  FF002F6C. A theme fill resolves against whatever theme the reader's Excel is using, and
  in this workbook theme1 is dk1 - so a navy bar rendered BLACK, next to the rebuilt cells
  that carry real navy. Every affected bar was two-tone

  bars that stopped short of the table under them, subtotal bands with a hole where the
  AU / NZ column sits, header rows with blank navy cells hanging off the end, and ten tabs
  that are meant to read as one family sized to ten different column profiles

  the owner's own working blocks - the role-review scratch tables on 1.4, 1.5 and 1.6 -
  were never formatted, and the one on 1.6 sat inside the squad table and punched a hole
  through its header bar

Everything is located by label and content, never by a remembered row number: the tabs
move under this script every time the chain runs. A cell that does not look the way the
review described it is left alone and reported, so a stale line in the fix list cannot
silently repaint something that has since been corrected.
"""
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import column_index_from_string as CI
from openpyxl.utils import get_column_letter as L

import opts

CREAM = opts.YEL                      # FFFFF2CC, the one input colour
BAND = opts.MID                       # FFD9D9D9, the subtotal band
SOFT = opts.GREY                      # FFF2F2F2, the lighter band
NONE_FILL = PatternFill()
NO_BORDER = Border()
WRAPL = Alignment(horizontal="left", vertical="center", wrap_text=True)

DESIGN = re.compile(r"^1\.\d+ ")      # the portfolio and COE design tabs
PORTFOLIO = re.compile(r"^1\.(?:[1-9]|10|14) ")   # the squad tabs, 1.14 included
COE = ("1.11 BP&T", "1.12 SA&D", "1.13 Cyber Roles")
CONFIG, ARCH = "0.2 Data Config", "0.3 Squad Archetypes"

# the money format the family uses. The review asked for "#,##0.00;(#,##0.00)"; opts.M2 is
# that plus the house's dash for zero, which is what every neighbouring cell in these same
# blocks already renders, so the block comes out consistent rather than nearly consistent.
MONEY = opts.M2
SUMMARY_BAR = "Portfolio Summary"
FUND_HEADS = ("Budget line", "Funding position")
DEFAULT_W = 8.43


# ----------------------------------------------------------------- small readers

def txt(ws, r, c):
    v = ws.cell(r, c).value
    return v.strip() if isinstance(v, str) else ""


def label(ws, r, c):
    """The cell's own text, only when it is text and not a formula."""
    t = txt(ws, r, c)
    return "" if t.startswith("=") else t


def rgb(cell):
    f = cell.fill
    if f is None or not f.fill_type:
        return ""
    g = f.fgColor
    return str(g.rgb or "").upper() if g.type == "rgb" else ""


def themed(cell):
    f = cell.fill
    return bool(f is not None and f.fill_type and f.fgColor.type == "theme")


def white(cell):
    co = cell.font.color if cell.font else None
    return bool(co is not None and co.type == "rgb" and str(co.rgb or "").upper()
                in ("FFFFFFFF", "00FFFFFF"))


def width(ws, c):
    return ws.column_dimensions[L(c)].width or DEFAULT_W


def height(ws, r):
    return ws.row_dimensions[r].height or 15.0


def empty(ws, r, c0, c1):
    return all(ws.cell(r, c).value is None for c in range(c0, c1 + 1))


def last_labelled(ws, r, c0, c1):
    cols = [c for c in range(c0, min(c1, ws.max_column) + 1) if label(ws, r, c)]
    return max(cols) if cols else None


def runway(ws, r, c, span=6):
    """How many width units the text in (r, c) has before a neighbour cuts it off."""
    w = width(ws, c)
    for k in range(c + 1, min(c + span, ws.max_column) + 1):
        if ws.cell(r, k).value is not None:
            return w
        w += width(ws, k)
    return w


def paint(ws, r, c, fill=None, font=None, align=None, nf=None, border=None):
    """Style one cell. Cells swallowed by a merge take style but not value."""
    x = ws.cell(r, c)
    if fill is not None:
        x.fill = fill
    if font is not None:
        x.font = font
    if align is not None:
        x.alignment = align
    if nf is not None:
        x.number_format = nf
    if border is not None:
        x.border = border
    return x


def write(ws, r, c, value):
    """Set a value unless the cell is inside a merge, where openpyxl forbids it."""
    x = ws.cell(r, c)
    if isinstance(x, openpyxl.cell.cell.MergedCell):
        return False
    x.value = value
    return True


def deref(wb, ws, r, c):
    """The value behind ='Some tab'!$X$9 - the COE lists are built entirely of these.

    One hop, and only a bare reference. A cell that resolves to another formula is
    reported as unknown rather than measured as if the formula text were its value: the
    ledger's Status column is itself a formula, and treating that as a 55-character
    string had this widening two columns that render six characters.
    """
    v = ws.cell(r, c).value
    if not isinstance(v, str) or not v.startswith("="):
        return v
    m = re.fullmatch(r"='([^']+)'!\$?([A-Z]{1,3})\$?(\d+)", v.strip())
    if not m or m.group(1) not in wb.sheetnames:
        return None
    got = wb[m.group(1)].cell(int(m.group(3)), CI(m.group(2))).value
    return None if isinstance(got, str) and got.startswith("=") else got


def tabs(wb, rx):
    return [ws for ws in wb.worksheets if rx.match(ws.title)]


def merged_span(ws, r, c):
    for m in ws.merged_cells.ranges:
        if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
            return m.min_col, m.max_col
    return c, c


# ----------------------------------------------------------------- table finders

def squad_blocks(ws):
    """Every squad table: its platform bar, header row, total row and last column.

    The last column is the last navy header cell walking right from B, not the last
    labelled cell: on 1.4, 1.5 and 1.6 the owner's own scratch headers sit in the same
    row, grey, and counting those made the squad table five columns wider than it is.
    """
    out = []
    for r in range(1, ws.max_row + 1):
        if label(ws, r, 2) != "Squad" or rgb(ws.cell(r, 2)) != opts.NAVY:
            continue
        last = 2
        for c in range(3, min(ws.max_column, 20) + 1):
            if rgb(ws.cell(r, c)) == opts.NAVY and label(ws, r, c):
                last = c
            else:
                break
        total = None
        for k in range(r + 1, min(ws.max_row, r + 40) + 1):
            t = label(ws, k, 2)
            if t.endswith("Total"):
                total = k
                break
            if t.startswith("Platform:"):     # the next bar. "Platform Overhead" is a row
                break
        bar = r - 1 if label(ws, r - 1, 2).startswith("Platform:") else None
        out.append({"hdr": r, "total": total, "last": last, "bar": bar})
    return out


def summary_block(ws):
    """The Portfolio Summary: bar row, header row, last column, Total Cost row."""
    for r in range(1, min(ws.max_row, 12) + 1):
        if label(ws, r, 2) != SUMMARY_BAR:
            continue
        hdr = r + 1
        if label(ws, hdr, 2) != "Cost":
            return None
        last = 2                    # navy header cells only: H on this row is the
        for c in range(3, 11):      # budget box's first label, not a summary column
            if rgb(ws.cell(hdr, c)) == opts.NAVY:
                last = c
            else:
                break
        total = next((k for k in range(hdr + 1, min(ws.max_row, hdr + 12) + 1)
                      if label(ws, k, 2) == "Total Cost"), None)
        return {"bar": r, "hdr": hdr, "last": last, "total": total}
    return None


def funding_block(ws):
    """The Other funding table: its header row, column extent and existing bar row."""
    for r in range(1, min(ws.max_row, 30) + 1):
        first = next((c for c in range(8, min(ws.max_column, 14) + 1) if label(ws, r, c)),
                     None)
        if first is None or label(ws, r, first) not in FUND_HEADS:
            continue
        heads = [c for c in range(first, min(ws.max_column, 14) + 1) if label(ws, r, c)]
        if len(heads) > 1:
            last = max(heads)
        else:                                  # Funding position: one label, data to its right
            last = first
            for k in range(r + 1, min(ws.max_row, r + 14) + 1):
                if not label(ws, k, first):
                    break
                cols = [c for c in range(first, min(ws.max_column, 14) + 1)
                        if ws.cell(k, c).value is not None]
                last = max([last] + cols)
        return {"hdr": r, "c0": first, "last": last}
    return None


def role_list(ws):
    """A COE tab's roles list: bar row, header row, first and last data row.

    A role row is one whose name is pulled from the ledger, so the list ends where the
    formulas end - not at the first blank, which would swallow the check row underneath.
    """
    for r in range(1, ws.max_row + 1):
        if label(ws, r, 2) != "Name" or rgb(ws.cell(r, 2)) != opts.NAVY:
            continue
        last = last_labelled(ws, r, 2, 14) or 8
        lo, hi = r + 1, r
        for k in range(lo, ws.max_row + 1):
            if str(ws.cell(k, 2).value or "").startswith("="):
                hi = k
        return {"hdr": r, "last": last, "lo": lo, "hi": hi,
                "bar": r - 1 if label(ws, r - 1, 2) == "Roles" else None}
    return None


# ----------------------------------------------------------------- 1  theme fills

def theme_bars(wb):
    """A bar cell holding a theme fill renders black. Give it the real FF002F6C.

    Rule from the review: on a 1.x tab, a theme-filled cell that functions as a bar or a
    section header - bold white text, or a cell in a row that is a platform bar or a table
    header - becomes an explicit BARC fill with the bar font. Anything else theme-filled is
    left alone and reported; the peach budget blocks are handled separately below.
    """
    out, fixed, skipped = [], 0, []
    for ws in tabs(wb, DESIGN):
        for r in range(1, ws.max_row + 1):
            bar_row = (label(ws, r, 2).startswith("Platform")
                       or any(rgb(ws.cell(r, c)) in (opts.BARC, opts.NAVY)
                              for c in range(2, min(ws.max_column, 14) + 1)))
            for c in range(1, ws.max_column + 1):
                x = ws.cell(r, c)
                if not themed(x):
                    continue
                tint = x.fill.fgColor.tint or 0
                if abs(tint) > 0.01:              # the peach budget block, not a bar
                    continue
                if not ((x.font.b and white(x)) or bar_row):
                    skipped.append(f"{ws.title}!{L(c)}{r}")
                    continue
                c0, c1 = merged_span(ws, r, c)
                for k in range(c0, c1 + 1):
                    paint(ws, r, k, fill=opts.fl(opts.BARC), font=opts.BARF)
                if label(ws, r, c):
                    paint(ws, r, c, align=opts.LFT)   # every bar reads from the left
                fixed += 1
    out.append(f"{fixed} theme-filled bar cells on the 1.x tabs given the explicit "
               f"FF002F6C they were meant to have - they were rendering black")
    if skipped:
        out.append(f"theme fills left alone, not bar cells: {', '.join(skipped[:8])}")
    return out


def orange_block(wb):
    """The peach budget block: the typed number is an input, the label is a label."""
    out, n, cream = [], 0, 0
    for ws in tabs(wb, PORTFOLIO):
        rows = sorted({r for r in range(1, ws.max_row + 1)
                       for c in range(2, min(ws.max_column, 8) + 1)
                       if themed(ws.cell(r, c)) and abs(ws.cell(r, c).fill.fgColor.tint
                                                        or 0) > 0.01})
        for r in rows:
            for c in range(2, min(ws.max_column, 8) + 1):
                x = ws.cell(r, c)
                if themed(x):
                    c0, c1 = merged_span(ws, r, c)
                    for k in range(c0, c1 + 1):
                        paint(ws, r, k, fill=NONE_FILL, align=opts.LFT)
                    n += 1
            val = next((c for c in range(2, min(ws.max_column, 8) + 1)
                        if ws.cell(r, c).value is not None and not label(ws, r, c)), None)
            if val and isinstance(ws.cell(r, val).value, (int, float)):
                paint(ws, r, val, fill=opts.fl(CREAM))   # a typed number is an input
                cream += 1
        if len(rows) > 1 and len({height(ws, r) for r in rows}) > 1:
            # levelled up to the tallest, but only where no row of the block is still a
            # plain single line: on most tabs the top row of the block also carries the
            # funding table's header, and raising a 14.25 data row to match it would
            # invent a double-height row rather than tidy one
            h = max(height(ws, r) for r in rows)
            if min(height(ws, r) for r in rows) > 15:
                for r in rows:
                    ws.row_dimensions[r].height = h
                out.append(f"{ws.title}: budget block rows {rows} levelled to {h}")
            else:
                out.append(f"{ws.title}: budget block rows {rows} are "
                           f"{[height(ws, r) for r in rows]} - one of them is a single "
                           f"line and the other carries the funding table's header row, "
                           f"so the difference is not the block's to level; reported")
    out.append(f"orange budget blocks: {n} label cells returned to plain, {cream} typed "
               f"inputs painted cream")
    return out


# ----------------------------------------------------------------- 2  bar spans

def summary_strip(wb):
    """The Portfolio Summary bar is a merge plus loose cells, so it showed a seam."""
    out = 0
    for ws in tabs(wb, PORTFOLIO):
        b = summary_block(ws)
        if not b:
            continue
        for c in range(2, b["last"] + 1):
            paint(ws, b["bar"], c, fill=opts.fl(opts.BARC), font=opts.BARF)
        paint(ws, b["bar"], 2, align=opts.LFT)
        out += 1
    return [f"Portfolio Summary bar made one uniform strip on {out} tabs"]


def funding_bar(wb):
    """One bar over the Other funding table, the width of the table, on every tab."""
    made, widened, missing = [], [], []
    for ws in tabs(wb, PORTFOLIO):
        f = funding_block(ws)
        if not f:
            missing.append(ws.title)
            continue
        r = f["hdr"] - 1
        c0, c1 = f["c0"], f["last"]
        has = any(txt(ws, r, c).startswith("Other funding") for c in range(c0, c1 + 1))
        if not has and not empty(ws, r, c0, c1):
            missing.append(f"{ws.title} row {r} not free for a bar")
            continue
        for c in range(c0, c1 + 1):
            paint(ws, r, c, fill=opts.fl(opts.BARC), font=opts.BARF)
        if write(ws, r, c0, "Other funding"):
            paint(ws, r, c0, align=opts.LFT)
        ws.row_dimensions[r].height = max(height(ws, r), 19)
        (widened if has else made).append(f"{ws.title}!{L(c0)}{r}:{L(c1)}{r}")
    out = []
    if made:
        out.append(f"Other funding bar added where it was missing: {', '.join(made)}")
    if widened:
        out.append(f"Other funding bar squared to its table: {', '.join(widened)}")
    if missing:
        out.append(f"no Other funding table found on: {', '.join(missing)}")
    return out


def platform_bars(wb):
    """A platform bar has to run the width of the table it heads, with no hole in it."""
    n = 0
    for ws in tabs(wb, PORTFOLIO):
        for b in squad_blocks(ws):
            if not b["bar"]:
                continue
            for c in range(2, b["last"] + 1):
                paint(ws, b["bar"], c, fill=opts.fl(opts.BARC), font=opts.BARF)
            paint(ws, b["bar"], 2, align=opts.LFT)
            ws.row_dimensions[b["bar"]].height = max(height(ws, b["bar"]), 19)
            n += 1
    return [f"{n} platform bars run the full width of their squad table, unbroken"]


def bucket_bar(wb):
    """1.13's funding bucket bar covered one column over a two-column table."""
    if "1.13 Cyber Roles" not in wb.sheetnames:
        return []
    ws = wb["1.13 Cyber Roles"]
    for r in range(1, min(ws.max_row, 30) + 1):
        if not label(ws, r, 2).startswith("Funding buckets"):
            continue
        body = []
        for k in range(r + 1, min(ws.max_row, r + 12) + 1):
            if not label(ws, k, 2) or rgb(ws.cell(k, 2)):
                break                          # the table ends at the next blank or bar
            body.append(k)
        last = 2
        for c in range(3, 8):                  # the table is as wide as it is unbroken
            if not any(ws.cell(k, c).value is not None for k in body):
                break
            last = c
        for c in range(2, last + 1):
            paint(ws, r, c, fill=opts.fl(opts.BARC), font=opts.BARF)
        paint(ws, r, 2, align=opts.LFT)
        return [f"1.13!B{r} bar spans B:{L(last)}, the width of its own table"]
    return ["1.13: no 'Funding buckets to draw down' bar found - skipped"]


# ----------------------------------------------------------------- 3  blank navy

ACTUALS_HEAD = "What the cost covers"


def footer_header(wb):
    """The actuals table's header row: labelled or unfilled, nothing blank between.

    This was written for the block actuals.py used to put at the foot of each tab, and it
    found that block by looking anywhere on the tab for "Archetype cost ($m)" or "Actual
    cost after decisions ($m)". Both of those are gone from the block - it moved to the top
    of the tab and became a label / Roles / Cost table - and the second is still the head of
    every squad table, so the old finder had quietly stopped dressing a footer and started
    matching the first squad table's header row instead.

    It now looks for the actuals table's own first head, and it skips the cells inside a
    merge: the label column is K:L merged, so L is blank navy on purpose and clearing it
    would put a hole in the header. On a normal build actuals.py writes the row square and
    this reports that there was nothing to do.
    """
    out, done = [], []
    for ws in tabs(wb, PORTFOLIO):
        found = next(((k, c) for k in range(1, min(ws.max_row, 30) + 1)
                      for c in range(2, min(ws.max_column, 20) + 1)
                      if label(ws, k, c) == ACTUALS_HEAD), None)
        if found is None:
            continue
        r, c0 = found
        last = last_labelled(ws, r, c0, 20)
        held = {c for m in ws.merged_cells.ranges if m.min_row == m.max_row == r
                for c in range(m.min_col, m.max_col + 1)}
        blank = [c for c in range(c0, last)
                 if not label(ws, r, c) and c not in held
                 and rgb(ws.cell(r, c)) == opts.NAVY]
        for c in blank:
            paint(ws, r, c, fill=NONE_FILL, border=NO_BORDER)
        if blank:
            done.append(f"{ws.title} row {r}: {','.join(L(c) for c in blank)}")
    if done:
        out.append("blank navy cleared from the actuals table header rows: "
                   + "; ".join(done))
    else:
        out.append("the actuals table header rows are square - a label in every navy cell "
                   "that is not part of the K:L merge, so there is nothing to clear")
    return out


def empty_navy_rows(wb):
    """A run of navy cells with no text in any of them is a band of colour saying nothing.

    Run by run rather than row by row: 1.3 carries an empty navy strip in B:C on a row
    that has a live funding line out in H:J, so the row is not empty and the strip is.
    """
    out = []
    for ws in tabs(wb, DESIGN):
        for r in range(1, ws.max_row + 1):
            run = []
            for c in range(2, min(ws.max_column, 20) + 2):
                if rgb(ws.cell(r, c)) in (opts.NAVY, opts.BARC):
                    run.append(c)
                    continue
                if run and not any(label(ws, r, k) for k in run):
                    for k in run:
                        paint(ws, r, k, fill=NONE_FILL, border=NO_BORDER)
                    out.append(f"{ws.title}!{L(run[0])}{r}:{L(run[-1])}{r}: a navy strip "
                               f"with no text in it, fill removed")
                run = []
    return out or ["no empty navy strips found"]


def funding_header(wb):
    """The Funding position header: navy across its own table and no further."""
    out = []
    for ws in tabs(wb, PORTFOLIO):
        f = funding_block(ws)
        if not f or label(ws, f["hdr"], f["c0"]) != "Funding position":
            continue
        r, c0, c1 = f["hdr"], f["c0"], f["last"]
        beyond = [L(c) for c in range(c1 + 1, min(ws.max_column, 14) + 1)
                  if rgb(ws.cell(r, c)) == opts.NAVY and not label(ws, r, c)]
        for c in range(c1 + 1, min(ws.max_column, 14) + 1):
            if rgb(ws.cell(r, c)) == opts.NAVY and not label(ws, r, c):
                paint(ws, r, c, fill=NONE_FILL, border=NO_BORDER)
        for c in range(c0, c1 + 1):
            paint(ws, r, c, fill=opts.fl(opts.NAVY), font=opts.HDRF)
        paint(ws, r, c0, align=opts.LFT)
        if beyond:
            out.append(f"{ws.title} row {r}: blank navy past the table cleared at "
                       f"{','.join(beyond)}; header squared to {L(c0)}:{L(c1)}")
    return out or ["no stray navy on a Funding position header"]


def summary_head_labels(wb):
    """A blank navy cell in the Portfolio Summary header, where nine siblings agree."""
    seen, out = {}, []
    for ws in tabs(wb, PORTFOLIO):
        b = summary_block(ws)
        if not b:
            continue
        for c in range(2, b["last"] + 1):
            t = label(ws, b["hdr"], c)
            if t:
                seen.setdefault(c, {}).setdefault(t, 0)
                seen[c][t] += 1
    for ws in tabs(wb, PORTFOLIO):
        b = summary_block(ws)
        if not b:
            continue
        for c in range(2, b["last"] + 1):
            if label(ws, b["hdr"], c) or rgb(ws.cell(b["hdr"], c)) != opts.NAVY:
                continue
            best = max(seen.get(c, {}).items(), key=lambda kv: kv[1], default=None)
            if best and best[1] >= 2 and write(ws, b["hdr"], c, best[0]):
                paint(ws, b["hdr"], c, font=opts.HDRF, align=opts.CEN)
                out.append(f"{ws.title}!{L(c)}{b['hdr']} was blank navy; labelled "
                           f"'{best[0]}', which {best[1]} sibling tabs carry there")
    return out or ["no blank cell in a Portfolio Summary header row"]


# ----------------------------------------------------------------- 4  grey bands

def subtotal_bands(wb):
    """A subtotal band with a hole in it reads as two bands and one missing figure."""
    filled, strays = 0, []
    for ws in tabs(wb, PORTFOLIO):
        for b in squad_blocks(ws):
            if not b["total"]:
                continue
            r = b["total"]
            shade = next((rgb(ws.cell(r, c)) for c in range(2, b["last"] + 1)
                          if rgb(ws.cell(r, c)) in (BAND, SOFT)), BAND)
            for c in range(2, b["last"] + 1):
                if rgb(ws.cell(r, c)) != shade:
                    paint(ws, r, c, fill=opts.fl(shade))
                    filled += 1
            for k in range(b["hdr"] + 1, r):     # a lone banded cell on a data row
                lone = [c for c in range(2, b["last"] + 1)
                        if rgb(ws.cell(k, c)) in (BAND, SOFT)]
                if len(lone) == 1 and ws.cell(k, lone[0]).value is None:
                    paint(ws, k, lone[0], fill=NONE_FILL)
                    strays.append(f"{ws.title}!{L(lone[0])}{k}")
    out = [f"{filled} cells added to squad subtotal bands so each runs the full width "
           f"of its table - column F (AU / NZ) was unfilled on every one of them"]
    if strays:
        out.append(f"stray band fill cleared from data rows: {', '.join(strays)}")
    return out


# ----------------------------------------------------------------- 5  widths

def width_profile(wb):
    """One column profile for 1.1 - 1.10, derived from what the family already uses.

    Per column: the widths that more than one tab uses are the family's; take the widest of
    them, then make sure the column's own longest header still fits, and cap at 45 so no
    single column pushes the tab off a landscape page.
    """
    fit = {"B": "Portfolio Overhead (see 0.2 Data Config)",
           "C": "Configuration / Integration", "H": "TDD Lights On Budget ($m)"}
    two_line = {"L": "Variance to archetype ($m)"}
    sheets = tabs(wb, PORTFOLIO)
    prof = {}
    for k in "BCDEFGHIJKL":
        vals = [ws.column_dimensions[k].width for ws in sheets
                if k in ws.column_dimensions and ws.column_dimensions[k].width]
        if not vals:
            continue
        common = [v for v in set(vals) if vals.count(v) > 1] or list(set(vals))
        w = max(common)
        if k in fit:
            w = max(w, len(fit[k]))
        if k in two_line:
            while opts.wrap_lines(two_line[k], w) > 2 and w < 45:
                w += 1
        prof[k] = round(min(w, 45.0), 1)
    return prof


def harmonise_widths(wb):
    prof = width_profile(wb)
    n = 0
    for ws in tabs(wb, PORTFOLIO):
        for k, v in prof.items():
            ws.column_dimensions[k].width = v
        n += 1
    return [f"one column profile across {n} portfolio tabs: "
            + ", ".join(f"{k}={v}" for k, v in sorted(prof.items(),
                                                      key=lambda kv: CI(kv[0])))]


def wrap_to_fit(wb):
    """Wrap and raise the rows whose label no longer fits the harmonised column."""
    done = []
    for ws in tabs(wb, PORTFOLIO):
        targets = []
        b = summary_block(ws)
        if b:
            for r in range(b["hdr"] + 1, (b["total"] or b["hdr"] + 5) + 1):
                if label(ws, r, 2).startswith("Portfolio Overhead"):
                    targets.append((r, 2))
        for r in range(1, ws.max_row + 1):
            if label(ws, r, 5) == "Variance to actuals":
                targets.append((r, 5))
            if label(ws, r, 8) and not rgb(ws.cell(r, 8)) in (opts.NAVY, opts.BARC):
                targets.append((r, 8))
        for r, c in targets:
            t = label(ws, r, c)
            room = runway(ws, r, c)
            if len(t) <= room:
                continue
            lines = opts.wrap_lines(t, width(ws, c))
            need = 14 * lines + 6
            if height(ws, r) + 1 < need:
                ws.row_dimensions[r].height = need
            paint(ws, r, c, align=WRAPL)
            done.append(f"{ws.title}!{L(c)}{r}")
    return [f"{len(done)} labels wrapped and given the row height their text needs "
            f"({', '.join(done[:8])}{', ...' if len(done) > 8 else ''})"]


# ----------------------------------------------------------------- 6  numbers

def summary_formats(wb):
    """A bare 0.00 puts a minus sign in front of a negative. The family uses brackets."""
    n = 0
    for ws in tabs(wb, PORTFOLIO):
        b = summary_block(ws)
        if not b or not b["total"]:
            continue
        for r in range(b["hdr"] + 1, b["total"] + 1):
            for c in range(3, b["last"] + 1):
                x = ws.cell(r, c)
                if x.number_format == "0.00" or (x.number_format == "General"
                                                 and x.value is not None):
                    x.number_format = MONEY
                    n += 1
    return [f"{n} Portfolio Summary cells moved off a bare 0.00 onto the family's "
            f"bracketed money format"]


def funding_formats(wb):
    """One cell in the allocation column rendered a bare 0 where its siblings show -."""
    done = []
    for ws in tabs(wb, PORTFOLIO):
        f = funding_block(ws)
        if not f:
            continue
        col = next((c for c in range(f["c0"], f["last"] + 1)
                    if "allocated to people" in label(ws, f["hdr"], c).lower()), None)
        if col is None:
            continue
        for r in range(f["hdr"] + 1, min(ws.max_row, f["hdr"] + 20) + 1):
            if rgb(ws.cell(r, f["c0"])) in (opts.NAVY, opts.BARC):
                break                          # the next table's header row
            if not label(ws, r, f["c0"]) and ws.cell(r, col).value is None:
                break                          # past the bottom of this table
            x = ws.cell(r, col)
            if x.value is not None and x.number_format == "General":
                x.number_format = MONEY
                done.append(f"{ws.title}!{L(col)}{r}")
    return [f"allocation-column cells given the family's money format: "
            f"{', '.join(done) if done else 'none needed'}"]


TRIM = [(CONFIG, "Actions"), (CONFIG, "Over/ under total"), (CONFIG, "Variance to total"),
        (CONFIG, "Position today"), (CONFIG, "CPI actuals; pull through")]


def trim_labels(wb):
    """Trailing spaces in a label are invisible until a header centres itself."""
    done, missing = [], []
    for tab, want in TRIM:
        if tab not in wb.sheetnames:
            missing.append(f"{tab} (no such tab)")
            continue
        ws = wb[tab]
        hit = False
        for row in ws.iter_rows():
            for x in row:
                if isinstance(x.value, str) and x.value != want and x.value.strip() == want:
                    x.value = want
                    done.append(f"{tab}!{x.coordinate}")
                    hit = True
        if not hit:
            missing.append(f"'{want}' on {tab}")
    for ws in tabs(wb, PORTFOLIO):
        for row in ws.iter_rows():
            for x in row:
                if isinstance(x.value, str) and x.value.strip() == "Other funding" \
                        and x.value != "Other funding":
                    x.value = "Other funding"
                    done.append(f"{ws.title}!{x.coordinate}")
    out = [f"{len(done)} labels trimmed of trailing space: {', '.join(done)}"]
    if missing:
        out.append(f"nothing to trim (label not present as described): "
                   f"{'; '.join(missing)}")
    return out


# ----------------------------------------------------------------- 7  0.2 Data Config

def config_actions(ws):
    """Every Actions note is clipped: the column is 20 wide and the row is one line.

    The ceiling is not mine: finish.py's fix_02 flattens any 0.2 row over 50pt back to a
    single line, so a note that needs more than three lines is widened as far as the
    column can take it and then reported rather than given a row that will not survive.
    """
    col = next((c for r in range(1, 10) for c in range(2, ws.max_column + 1)
                if label(ws, r, c) == "Actions"), None)
    if col is None:
        return ["0.2: no Actions column found - skipped"]
    notes = [(r, label(ws, r, col)) for r in range(6, ws.max_row + 1)
             if label(ws, r, col)]
    if not notes:
        return [f"0.2: no populated Actions rows in column {L(col)} - skipped"]
    w = width(ws, col)
    while w < 34 and max(opts.wrap_lines(t, w) + t.count("\n") for _, t in notes) > 3:
        w += 1
    ws.column_dimensions[L(col)].width = round(w, 1)
    n = 0
    for r, t in notes:
        lines = opts.wrap_lines(t, w) + t.count("\n")
        need = 14 * lines + 6
        # finish.py's fix_02 flattens tall 0.2 rows but skips any row whose Actions cell
        # holds text, so the full height a note needs survives the chain
        if abs(max(height(ws, r), need) - height(ws, r)) > 1:
            ws.row_dimensions[r].height = max(height(ws, r), need)
        paint(ws, r, col, align=WRAPL)
        n += 1
    return [f"0.2 Actions column {L(col)} widened to {ws.column_dimensions[L(col)].width} "
            f"and {n} notes wrapped, each row raised to the lines its note needs"]


def config_recon(ws):
    """The reconciliation block: a title on its corner cell and one money format."""
    r = next((k for k in range(1, ws.max_row + 1)
              if label(ws, k, 11) == "" and rgb(ws.cell(k, 11)) == opts.NAVY
              and label(ws, k, 12)), None)
    out = []
    if r is None:
        r = next((k for k in range(1, ws.max_row + 1)
                  if label(ws, k, 11) == "Reconciliation ($m)"), None)
        if r is None:
            return ["0.2: reconciliation block corner cell not as described - skipped"]
        out.append("0.2: reconciliation title already in place")
    else:
        write(ws, r, 11, "Reconciliation ($m)")
        out.append(f"0.2!K{r} was an empty navy corner; titled 'Reconciliation ($m)'")
    last = last_labelled(ws, r, 11, 16) or 15
    for c in range(11, last + 1):
        paint(ws, r, c, fill=opts.fl(opts.NAVY), font=opts.HDRF, align=opts.CEN)
    n = 0
    for k in range(r + 1, min(ws.max_row, r + 4) + 1):
        for c in range(12, last + 1):
            x = ws.cell(k, c)
            if x.value is not None and x.number_format != MONEY:
                x.number_format, x.alignment = MONEY, opts.RGT
                n += 1
    out.append(f"0.2 reconciliation block K{r}:{L(last)}{r + 3}: title bar squared and "
               f"{n} figures put on one bracketed money format")
    return out


def config_overheads(ws):
    """The two overhead tables are the same table twice; they must read the same."""
    out = []
    heads = [r for r in range(1, ws.max_row + 1) if label(ws, r, 11) == "Role"]
    want = {}
    for r in heads:
        for c in range(11, 16):
            t = label(ws, r, c)
            if t:
                want.setdefault(c, []).append(t)
    for c, seen in want.items():
        full = max(seen, key=len)
        for r in heads:
            was = label(ws, r, c)
            if was and was != full and full.startswith(was[:5]):
                write(ws, r, c, full)
                out.append(f"0.2!{L(c)}{r} '{was}' -> '{full}', so both overhead tables "
                           f"use one word for one thing")
    n = 0
    for r in range(1, ws.max_row + 1):
        if "subtotal" not in label(ws, r, 11).lower():
            continue
        last = max([c for c in range(11, 16) if ws.cell(r, c).value is not None] + [14])
        for c in range(11, last + 1):
            paint(ws, r, c, fill=opts.fl(SOFT), font=opts.BOLD)
            x = ws.cell(r, c)
            if isinstance(x.value, str) and x.value.startswith("=") \
                    and x.number_format != MONEY:
                x.number_format, x.alignment = MONEY, opts.RGT
        n += 1
    out.append(f"0.2: {n} overhead subtotal rows banded like every other subtotal and "
               f"put on two decimals")
    return out


def config_bars(ws):
    """Two bars left, one centred. A bar is a title, and a title starts at the left."""
    n = 0
    for r in range(1, ws.max_row + 1):
        for c in range(2, min(ws.max_column, 14) + 1):
            if rgb(ws.cell(r, c)) == opts.BARC and label(ws, r, c):
                if ws.cell(r, c).alignment.horizontal != "left":
                    paint(ws, r, c, align=opts.LFT)
                    n += 1
    return [f"0.2: {n} of the tab's section bars re-aligned to the left, so all of them "
            f"read the same way"]


def config_shape(ws):
    """Column B wide enough for its own labels, one height for a data row, nothing tall
    without a reason, and every data cell sitting on the same line as its neighbours."""
    out = []
    labs = [label(ws, r, 2) for r in range(1, ws.max_row + 1) if label(ws, r, 2)]
    longest = max(labs, key=len) if labs else ""
    if longest and len(longest) + 1 > width(ws, 2):
        ws.column_dimensions["B"].width = round(min(len(longest) + 1, 58.0), 1)
        out.append(f"0.2 column B widened to {ws.column_dimensions['B'].width} to fit "
                   f"'{longest}'")
    bw = width(ws, 2)
    heads = {r for r in range(1, ws.max_row + 1)
             if sum(1 for c in range(2, 16)
                    if rgb(ws.cell(r, c)) == opts.NAVY and label(ws, r, c)) >= 2}
    bars = {r for r in range(1, ws.max_row + 1)
            if any(rgb(ws.cell(r, c)) == opts.BARC for c in range(2, 16))}
    blank, flat, held = [], [], []
    for r in range(6, ws.max_row + 1):
        if r in heads or r in bars:
            continue
        used = any(ws.cell(r, c).value is not None for c in range(2, 16))
        if not used:
            if height(ws, r) > 15:
                ws.row_dimensions[r].height = 14.25
                blank.append(r)
            for c in range(2, 16):
                if rgb(ws.cell(r, c)):
                    paint(ws, r, c, fill=NONE_FILL, border=NO_BORDER)
            continue
        if any(label(ws, r, c) for c in range(8, 10)):
            continue                            # a Notes or Actions row; sized below
        # a data row wraps and doubles only where its own label needs the room
        fits = len(label(ws, r, 2)) + 1 <= bw
        if fits and height(ws, r) > 20:
            ws.row_dimensions[r].height = 14.25
            flat.append(r)
        elif not fits:
            held.append(r)
    # the left table's own cells read the same way on every row, header row or not: one
    # row on 0.2 carries a data row and another table's header side by side
    odd = []
    for r in range(6, ws.max_row + 1):
        if r in bars:
            continue
        for c in range(2, 8):
            x = ws.cell(r, c)
            if rgb(x) in (opts.NAVY, opts.BARC) or x.value is None:
                continue
            a = x.alignment
            if not a.wrap_text and a.horizontal in (None, "left", "right", "general"):
                continue
            if c == 2 and len(label(ws, r, 2)) + 1 > bw:
                continue                        # this one genuinely needs to wrap
            x.alignment = Alignment(horizontal="left" if c == 2 else None,
                                    vertical="center")
            odd.append(f"{L(c)}{r}")
    if odd:
        out.append(f"0.2: data cells {', '.join(odd)} were centred and wrapped where "
                   f"every other row on the table is left and plain - normalised")
    for r in range(6, ws.max_row + 1):
        for c in range(2, 16):
            x = ws.cell(r, c)
            if x.value is None and not rgb(x):
                continue
            a = x.alignment
            if a.vertical != "center":
                x.alignment = Alignment(horizontal=a.horizontal, vertical="center",
                                        wrap_text=a.wrap_text)
    if blank:
        out.append(f"0.2: tall blank rows {blank} collapsed to a standard row and their "
                   f"stray banding cleared")
    if flat:
        out.append(f"0.2: double-height data rows {flat} put back on one line - the label "
                   f"fits column B now, so nothing needed the second line")
    if held:
        out.append(f"0.2: rows {held} keep their extra height; the label in B is longer "
                   f"than the column can be without pushing the tab off the page")
    out.append("0.2: every data cell in B:O vertically centred")
    return out


def config_notes_band(ws):
    """The Notes column is banded past the table it belongs to."""
    col = next((c for c in range(2, ws.max_column + 1)
                for r in range(1, 10) if label(ws, r, c) == "Notes"), None)
    if col is None:
        return ["0.2: no Notes column found - skipped"]
    body = [r for r in range(6, ws.max_row + 1)
            if any(ws.cell(r, c).value is not None for c in range(2, 8))]
    if not body:
        return ["0.2: Notes column has no table under it - skipped"]
    hi, cleared = max(body), []
    for r in range(6, ws.max_row + 1):
        if r in body or not rgb(ws.cell(r, col)):
            continue
        paint(ws, r, col, fill=NONE_FILL, border=NO_BORDER)
        cleared.append(r)
    where = ("cleared on rows " + ",".join(map(str, cleared))) if cleared else \
        f"it already ended at the last table row ({hi})"
    return [f"0.2 Notes column {L(col)}: band trimmed to the table's own rows, {where}"]


def config_ceiling(ws):
    """No row on this tab may leave here taller than 50pt.

    finish.py's fix_02 flattens any 0.2 row over 50 back to a single line, so a row left
    at 52 loses everything it was given the height for. Capped here, deliberately.
    """
    over = [r for r in range(1, ws.max_row + 1) if height(ws, r) > 50]
    for r in over:
        ws.row_dimensions[r].height = 50.0
    return [f"0.2 rows {over} held at 50pt - finish.py flattens anything taller, so a "
            f"row above the ceiling would come out on one line"] if over else []


def data_config(wb):
    if CONFIG not in wb.sheetnames:
        return [f"{CONFIG} not in this workbook - skipped"]
    ws = wb[CONFIG]
    return (config_recon(ws) + config_overheads(ws) + config_bars(ws)
            + config_shape(ws) + config_notes_band(ws) + config_actions(ws)
            + config_ceiling(ws))


# ----------------------------------------------------------------- 8  scratch blocks

SCRATCH_HEAD = "Nbr Archetype Roles"


def scratch_at(ws):
    """The owner's role-review block: header row and column range, wherever it sits."""
    for r in range(1, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            if label(ws, r, c) == SCRATCH_HEAD:
                last = c
                for k in range(c + 1, min(ws.max_column, c + 10) + 1):
                    if label(ws, r, k):
                        last = k
                    else:
                        break
                hi = r
                for k in range(r + 1, min(ws.max_row, r + 12) + 1):
                    if any(ws.cell(k, x).value is not None for x in range(c, last + 1)):
                        hi = k
                    else:
                        break
                return {"hdr": r, "c0": c, "c1": last, "hi": hi}
    return None


def move_scratch(ws, blk, to):
    """Move the block verbatim - values and number formats - and clear where it was."""
    keep = []
    for r in range(blk["hdr"], blk["hi"] + 1):
        row = []
        for c in range(blk["c0"], blk["c1"] + 1):
            x = ws.cell(r, c)
            row.append((x.value, x.number_format))
        keep.append(row)
    for r in range(blk["hdr"], blk["hi"] + 1):
        for c in range(blk["c0"], blk["c1"] + 1):
            x = ws.cell(r, c)
            if not isinstance(x, openpyxl.cell.cell.MergedCell):
                x.value = None
            x.fill, x.font, x.border = NONE_FILL, opts.BODY, NO_BORDER
            x.number_format, x.alignment = "General", Alignment()
    for i, row in enumerate(keep):
        for j, (v, nf) in enumerate(row):
            x = ws.cell(blk["hdr"] + i, to + j)
            if not isinstance(x, openpyxl.cell.cell.MergedCell):
                x.value = v
            x.number_format = nf
    return {"hdr": blk["hdr"], "c0": to, "c1": to + blk["c1"] - blk["c0"],
            "hi": blk["hi"]}


def dress_scratch(ws, blk, cap=50.0):
    """A grey bold mini-header, columns that fit, rows tall enough for what wraps."""
    for c in range(blk["c0"], blk["c1"] + 1):
        paint(ws, blk["hdr"], c, fill=opts.fl(SOFT), font=opts.BOLD, align=WRAPL)
        need = max([len(str(ws.cell(r, c).value)) for r in range(blk["hdr"], blk["hi"] + 1)
                    if ws.cell(r, c).value is not None] + [4])
        ws.column_dimensions[L(c)].width = round(min(max(need + 1, 9.0), cap), 1)
    for r in range(blk["hdr"], blk["hi"] + 1):
        lines = 1
        for c in range(blk["c0"], blk["c1"] + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v:
                if r > blk["hdr"]:
                    paint(ws, r, c, align=WRAPL)
                lines = max(lines, opts.wrap_lines(v, width(ws, c)))
        if lines > 1:
            ws.row_dimensions[r].height = max(height(ws, r), 14 * lines + 6)


def scratch_blocks(wb):
    """The owner's working tables. Every literal is preserved; only the shape changes."""
    out = []
    for ws in tabs(wb, PORTFOLIO):
        blk = scratch_at(ws)
        if not blk:
            continue
        last_tbl = max([b["last"] for b in squad_blocks(ws)] + [10])
        if blk["c0"] <= last_tbl + 2 and blk["c0"] < 19:
            before = f"{L(blk['c0'])}{blk['hdr']}:{L(blk['c1'])}{blk['hi']}"
            blk = move_scratch(ws, blk, 19)
            out.append(f"{ws.title}: role-review block moved verbatim from {before} to "
                       f"{L(blk['c0'])}{blk['hdr']}:{L(blk['c1'])}{blk['hi']} - it sat "
                       f"inside the squad table and broke its header bar")
        dress_scratch(ws, blk)
        out.append(f"{ws.title}: role-review block {L(blk['c0'])}:{L(blk['c1'])} given a "
                   f"grey header row, fitted widths and rows that hold their own text")
    return out or ["no role-review scratch blocks found"]


def bar_continuity(wb):
    """After the move, no bar may have a hole in it or a theme fill left in it."""
    bad = []
    for ws in tabs(wb, PORTFOLIO):
        for b in squad_blocks(ws):
            if not b["bar"]:
                continue
            for c in range(2, b["last"] + 1):
                x = ws.cell(b["bar"], c)
                if themed(x) or rgb(x) != opts.BARC:
                    bad.append(f"{ws.title}!{L(c)}{b['bar']}")
    how = "all platform bars solid FF002F6C end to end" if not bad \
        else "still broken at " + ", ".join(bad)
    return [f"bar continuity check: {how}"]


# ----------------------------------------------------------------- 9  COE tabs

def coe_levers(wb):
    """The On/Off cells are inputs. Two tabs said so in cream; the third did not."""
    out = []
    for t in COE:
        if t not in wb.sheetnames:
            continue
        ws = wb[t]
        lst = role_list(ws)
        if not lst:
            out.append(f"{t}: no roles list found - lever cells skipped")
            continue
        col = next((c for c in range(2, lst["last"] + 1)
                    if label(ws, lst["hdr"], c) == "On/Off"), None)
        if col is None:
            out.append(f"{t}: no On/Off column in the roles header - skipped")
            continue
        n = 0
        for r in range(lst["lo"], lst["hi"] + 1):
            if rgb(ws.cell(r, col)) != CREAM:
                paint(ws, r, col, fill=opts.fl(CREAM))
                n += 1
        if n:
            out.append(f"{t}!{L(col)}{lst['lo']}:{L(col)}{lst['hi']}: {n} lever cells "
                       f"painted cream, the input colour its siblings already use")
    return out or ["COE lever cells already cream"]


def coe_helper_header(wb):
    """The per-role cost column is unlabelled on all three tabs."""
    out = []
    for t in COE:
        if t not in wb.sheetnames:
            continue
        ws = wb[t]
        lst = role_list(ws)
        if not lst:
            continue
        rows = range(lst["lo"], lst["hi"] + 1)
        # the helper is the column that carries a figure on essentially every role, not
        # the one with four stray words left in it
        counts = {c: sum(1 for r in rows if ws.cell(r, c).value is not None)
                  for c in range(lst["last"] + 1, min(ws.max_column, 24) + 1)}
        col = max((c for c, n in counts.items() if n >= 0.8 * len(rows)),
                  key=lambda c: counts[c], default=None)
        if col is None:
            out.append(f"{t}: no per-role helper column past the roles table - skipped")
            continue
        if label(ws, lst["hdr"], col) == "Cost after On/Off ($)":
            continue
        if label(ws, lst["hdr"], col):
            out.append(f"{t}!{L(col)}{lst['hdr']} already carries "
                       f"'{label(ws, lst['hdr'], col)}' - left alone")
            continue
        model = ws.cell(lst["hdr"], lst["last"])
        if write(ws, lst["hdr"], col, "Cost after On/Off ($)"):
            paint(ws, lst["hdr"], col, fill=opts.fl(rgb(model) or opts.NAVY),
                  font=opts.HDRF, align=opts.CEN, border=opts.BOX)
            out.append(f"{t}!{L(col)}{lst['hdr']} headed 'Cost after On/Off ($)' - the "
                       f"column held per-role costs with nothing to say so")
    return out


def coe_list_shape(wb):
    """Rows that clip their own wrapped text, and empty ruled rows at the foot."""
    out = []
    for t in COE:
        if t not in wb.sheetnames:
            continue
        ws = wb[t]
        lst = role_list(ws)
        if not lst:
            continue
        raised, wide = [], {}
        for r in range(lst["lo"], lst["hi"] + 1):
            lines = 1
            for c in range(2, lst["last"] + 1):
                v = deref(wb, ws, r, c)
                if not isinstance(v, str) or not v.strip():
                    continue
                if len(v) <= runway(ws, r, c):
                    continue                    # it fits as it stands
                if ws.cell(r, c).alignment.wrap_text:
                    lines = max(lines, opts.wrap_lines(v, width(ws, c)))
                else:
                    wide[c] = max(wide.get(c, 0), len(v) + 1)
            if lines > 1 and height(ws, r) + 1 < 14 * lines + 6:
                ws.row_dimensions[r].height = 14 * lines + 6
                raised.append(r)
        for c, need in wide.items():
            ws.column_dimensions[L(c)].width = round(
                max(width(ws, c), min(need, 40.0)), 1)
        if raised:
            out.append(f"{t}: rows {raised} raised so their wrapped title is not clipped")
        if wide:
            out.append(f"{t}: columns {','.join(L(c) for c in sorted(wide))} widened to "
                       f"their longest value - NOTE finish.py's coe_widths resets the COE "
                       f"width profile afterwards, so this needs COE_W widened to hold")
        # empty ruled rows between the last role and whatever follows the list
        stop = next((r for r in range(lst["hi"] + 1, min(ws.max_row, lst["hi"] + 12) + 1)
                     if label(ws, r, 2)), ws.max_row + 1)
        cleared = []
        for r in range(lst["hi"] + 1, stop):
            if any(ws.cell(r, c).value is not None for c in range(2, lst["last"] + 1)):
                continue
            touched = False
            for c in range(2, lst["last"] + 1):
                x = ws.cell(r, c)
                if rgb(x) or any(s and s.style for s in (x.border.left, x.border.right,
                                                         x.border.top, x.border.bottom)):
                    x.fill, x.border = NONE_FILL, NO_BORDER
                    touched = True
            if touched:
                cleared.append(r)
        if cleared:
            out.append(f"{t}: empty ruled rows {cleared} at the foot of the roles list "
                       f"stripped of their borders and fill")
    return out


def coe_labels(wb):
    """Funding labels in column B that run under the figure beside them."""
    out = []
    for t in COE:
        if t not in wb.sheetnames:
            continue
        ws = wb[t]
        lst = role_list(ws)
        hi = (lst["bar"] or lst["hdr"]) - 1 if lst else ws.max_row
        done = []
        for r in range(3, hi + 1):
            lab = label(ws, r, 2)
            if not lab or ws.cell(r, 3).value is None:
                continue
            if len(lab) <= runway(ws, r, 2):
                continue
            lines = opts.wrap_lines(lab, width(ws, 2))
            ws.row_dimensions[r].height = max(height(ws, r), 14 * lines + 6)
            paint(ws, r, 2, align=WRAPL)
            done.append(r)
        if done:
            out.append(f"{t}: column B labels on rows {done} wrapped and their rows "
                       f"raised - they were being cut off by the figure beside them")
    return out


def coe_odds(wb):
    """1.12's accounting format, and 1.13's unlabelled working figures."""
    out = []
    if "1.12 SA&D" in wb.sheetnames:
        ws = wb["1.12 SA&D"]
        for r in range(1, ws.max_row + 1):
            if not label(ws, r, 2).startswith("Domain Architect funding from portfolio"):
                continue
            x = ws.cell(r, 3)
            if "$" in x.number_format:
                sibs = [ws.cell(k, 3).number_format for k in range(r - 3, r + 4)
                        if 0 < k <= ws.max_row and k != r
                        and ws.cell(k, 3).value is not None
                        and "$" not in ws.cell(k, 3).number_format]
                use = max(set(sibs), key=sibs.count) if sibs else MONEY
                x.number_format, x.alignment = use, opts.RGT
                out.append(f"1.12!C{r} carried an accounting format that printed a lone "
                           f"'$' in the value column; put on '{use}' like its siblings")
    if "1.13 Cyber Roles" in wb.sheetnames:
        ws = wb["1.13 Cyber Roles"]
        head = next((r for r in range(1, 20) if label(ws, r, 2) == "Grouping"), None)
        bar = next((r for r in range(1, 20)
                    if label(ws, r, 2).startswith("Funding buckets")), None)
        spend = next((c for c in range(3, 12)
                      if head and label(ws, head, c).startswith("Planned spend")), None)
        if head and bar and spend:
            body = []
            for k in range(bar + 1, min(ws.max_row, bar + 10) + 1):
                if not label(ws, k, 2) or rgb(ws.cell(k, 2)):
                    break
                body.append(k)
            wide = max([c for c in range(2, spend)
                        if any(ws.cell(k, c).value is not None for k in body)] + [2])
            for r in body:
                if ws.cell(r, spend).value is None:
                    continue
                if any(ws.cell(r, c).value is not None for c in range(wide + 1, spend)):
                    continue                     # something already sits beside it
                # the bucket table beside it ends at column `wide`, so this figure sits
                # in the summary table's spend column with nothing on its own row to name
                # it. Its formula is the spend total less the CapEx input below it.
                write(ws, r, spend - 1, "Planned spend less CapEx ($m)")
                paint(ws, r, spend - 1, font=opts.BODY, align=opts.RGT)
                out.append(f"1.13!{L(spend - 1)}{r} now labels the working figure in "
                           f"{L(spend)}{r} ({ws.cell(r, spend).value}) - it sat under the "
                           f"'Planned spend' column with no label anywhere on its row")
            lone = [r for r in range(head + 1, bar)
                    if ws.cell(r, spend).value is not None and label(ws, r, 2)
                    and all(ws.cell(r, c).value is None for c in range(3, spend))]
            for r in lone:
                out.append(f"1.13!{L(spend)}{r} renders '-' a long way from its label in "
                           f"B{r} ('{label(ws, r, 2)}'), but it is labelled - left alone")
    return out or ["no COE odds and ends to fix"]


# ----------------------------------------------------------------- 10  small items

def spacer_column(wb):
    """Text parked in the two-wide gutter column bleeds across the table beside it."""
    out = []
    for ws in tabs(wb, PORTFOLIO):
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is None or isinstance(v, str) and not v.strip():
                continue
            to = 14                                   # column N, clear of the table
            while to < 30 and ws.cell(r, to).value is not None:
                to += 1
            if ws.cell(r, to).value is not None:
                out.append(f"{ws.title}!A{r} has nowhere free to move to - left alone")
                continue
            src = ws.cell(r, 1)
            write(ws, r, to, src.value)
            paint(ws, r, to, font=opts.BODY, align=opts.LFT, fill=NONE_FILL)
            src.value, src.fill, src.border = None, NONE_FILL, NO_BORDER
            out.append(f"{ws.title}: '{v}' moved out of the gutter column A{r} to "
                       f"{L(to)}{r}")
    return out or ["nothing parked in the gutter column"]


def notes_header(wb):
    """The Notes column header sat in a data row on three tabs and in the header on one."""
    out = []
    for ws in tabs(wb, PORTFOLIO):
        f = funding_block(ws)
        if not f:
            continue
        for r in range(f["hdr"], min(ws.max_row, f["hdr"] + 3) + 1):
            for c in range(f["c0"], min(ws.max_column, f["last"] + 3) + 1):
                if label(ws, r, c).strip() != "Notes":
                    continue
                if r == f["hdr"]:
                    write(ws, r, c, "Notes")
                    paint(ws, r, c, fill=opts.fl(opts.NAVY), font=opts.HDRF,
                          align=opts.CEN)
                    continue
                if ws.cell(f["hdr"], c).value is not None:
                    out.append(f"{ws.title}!{L(c)}{f['hdr']} is occupied - Notes header "
                               f"left where it is")
                    continue
                write(ws, f["hdr"], c, "Notes")
                paint(ws, f["hdr"], c, fill=opts.fl(opts.NAVY), font=opts.HDRF,
                      align=opts.CEN)
                src = ws.cell(r, c)
                src.value, src.fill = None, NONE_FILL
                out.append(f"{ws.title}: Notes header moved out of the data row {r} up "
                           f"into the table's header row {f['hdr']}, navy like 1.5")
    return out or ["Notes headers already sit in their header row"]


SUMS = re.compile(r"^=\s*I(\d+)\s*\+\s*I(\d+)\s*$")


def orphan_variance(wb):
    """A figure at the foot of the budget box with no label on its row.

    The label is not invented: on this tab the cell adds the two rows above it, and eight
    sibling tabs put the same words in front of the same sum, so those are the words used.
    """
    votes = {}
    for ws in tabs(wb, PORTFOLIO):
        for r in range(1, min(ws.max_row, 14) + 1):
            if SUMS.match(str(ws.cell(r, 9).value or "")) and label(ws, r, 8):
                votes[label(ws, r, 8)] = votes.get(label(ws, r, 8), 0) + 1
    want = max(votes, key=votes.get) if votes else None
    out = []
    for ws in tabs(wb, PORTFOLIO):
        for r in range(1, min(ws.max_row, 14) + 1):
            if label(ws, r, 8) or not SUMS.match(str(ws.cell(r, 9).value or "")):
                continue
            if any(ws.cell(r, c).value is not None for c in range(2, 8)):
                continue
            if not want or votes[want] < 2:
                out.append(f"{ws.title}!I{r} is unlabelled and no sibling agrees on a "
                           f"name for it - left alone")
                continue
            model = next((ws.cell(k, 8) for k in range(r - 1, max(r - 4, 0), -1)
                          if label(ws, k, 8)), None)
            write(ws, r, 8, want)
            paint(ws, r, 8, align=opts.LFT,
                  font=Font(name=opts.FN, size=(model.font.size if model else 10) or 10,
                            bold=bool(model.font.b) if model else True))
            out.append(f"{ws.title}!H{r} labelled '{want}' ({votes[want]} sibling tabs "
                       f"use it for the same sum) - the figure in I{r} was the orphan "
                       f"the review saw at the foot of the budget box")
    return out or ["no orphan figures in the budget boxes"]


def floating_cream(wb):
    """Cream on an empty cell that nothing reads is a box waiting for nothing."""
    refs = set()
    rx = re.compile(r"'([^']+)'!\$?([A-Z]{1,3})\$?(\d+)")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for x in row:
                if isinstance(x.value, str) and x.value.startswith("="):
                    for m in rx.finditer(x.value):
                        refs.add((m.group(1), m.group(2), int(m.group(3))))
    out, cleared, kept = [], [], []
    for ws in tabs(wb, PORTFOLIO):
        f = funding_block(ws)
        lo = f["hdr"] if f else 1
        for r in range(lo, min(ws.max_row, lo + 16) + 1):
            for c in range(9, min(ws.max_column, 12) + 1):
                x = ws.cell(r, c)
                if x.value is not None or rgb(x) != CREAM:
                    continue
                if label(ws, r, f["c0"] if f else 8):
                    continue                       # it belongs to a labelled line
                if (ws.title, L(c), r) in refs:
                    kept.append(f"{ws.title}!{L(c)}{r}")
                    continue
                x.fill, x.border = NONE_FILL, NO_BORDER
                cleared.append(f"{ws.title}!{L(c)}{r}")
    if cleared:
        out.append(f"floating cream cells cleared - empty, unlabelled and read by no "
                   f"formula anywhere in the workbook: {', '.join(cleared)}")
    if kept:
        out.append(f"floating cream kept, a formula reads it: {', '.join(kept)}")
    return out or ["no floating cream cells on the portfolio tabs"]


# ----------------------------------------------------------------- 11  0.3

def squad_archetypes(wb):
    """Nothing. 0.3 Squad Archetypes is the owner's, and this pass keeps its hands off it.

    This used to move the doubled title on row 2 into a section bar and widen the columns
    to their own text - the treatment 0.2 gets. It was wrong about whose tab it was. 0.3 is
    his cost library, the archetype price table every 1.x tab INDEX/MATCHes into; it arrives
    from his review workbook through assemble_base and it is not a built tab. He asked why
    the chain had changed it, and the answer the workbook now gives is that it does not.

    The function is kept rather than deleted so the run log still says so on every pass, and
    so the next person looking for the 0.3 dressing finds this note instead of nothing.
    """
    if ARCH not in wb.sheetnames:
        return [f"{ARCH} not in this workbook"]
    return [f"{ARCH}: left exactly as it arrives - it is the owner's cost library, a "
            f"source tab like 0.1 and 0.4, and nothing in this pass touches it. "
            f"regress2707 proves it cell-for-cell against rev.xlsx."]


# ----------------------------------------------------------------- run

def budget_bar(wb):
    """The budget box's own bar covers its two columns on every tab, not one on some.

    Row 4 (5 on 1.7) carries two side-by-side tables: the Portfolio Summary to G and the
    Budget vs TDD Cost box at H:I. Six tabs painted the box's bar on H only, four on H:I,
    so the same strip ended in two different places across the family.
    """
    out = []
    n = 0
    for t in [x for x in wb.sheetnames if re.match(r"^1\.(10|14|[1-9]) ", x)]:
        ws = wb[t]
        for r in range(3, 8):
            for c in range(7, 13):
                x = ws.cell(r, c)
                if str(x.value or "").strip() == "Budget vs TDD Cost":
                    for cc in (c, c + 1):
                        y = ws.cell(r, cc)
                        y.fill, y.font = opts.fl(opts.BARC), opts.BARF
                    n += 1
                    break
    out.append(f"budget-box bar squared to its two columns on {n} tabs")
    return out


def run(src, dst):
    import json
    wb = openpyxl.load_workbook(src)
    before = {t: {c.coordinate: c.value for row in wb[t].iter_rows() for c in row
                  if c.value is not None} for t in wb.sheetnames}
    out = (theme_bars(wb) + orange_block(wb)
           + notes_header(wb) + summary_strip(wb) + budget_bar(wb)
           + funding_bar(wb) + platform_bars(wb)
           + bucket_bar(wb)
           + footer_header(wb) + empty_navy_rows(wb) + funding_header(wb)
           + summary_head_labels(wb)
           + subtotal_bands(wb)
           + harmonise_widths(wb) + wrap_to_fit(wb)
           + summary_formats(wb) + funding_formats(wb) + trim_labels(wb)
           + data_config(wb)
           + scratch_blocks(wb) + bar_continuity(wb)
           + coe_levers(wb) + coe_helper_header(wb) + coe_list_shape(wb) + coe_labels(wb)
           + coe_odds(wb)
           + spacer_column(wb) + orphan_variance(wb) + floating_cream(wb)
           + squad_archetypes(wb))
    # every cell this step wrote or cleared, declared - same contract as post2707, so the
    # shipped-workbook diff can tell a declared edit from an accident
    touched = []
    for t in wb.sheetnames:
        seen = set()
        for row in wb[t].iter_rows():
            for c in row:
                seen.add(c.coordinate)
                if c.value != before[t].get(c.coordinate):
                    touched.append([t, c.coordinate])
        for coord in set(before[t]) - seen:
            touched.append([t, coord])
    json.dump(touched, open("design2707_manifest.json", "w"))
    out.append(f"{len(touched)} cells declared in design2707_manifest.json")
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(*sys.argv[1:]):
        print("  ", x)
