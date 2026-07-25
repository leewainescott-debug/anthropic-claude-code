"""Stage 4: overhead and leadership logic, plus the two breaks the rebuild introduced.

The model, as Lee has defined it:

  "Leadership overhead" is the 8 GMs priced at $5.1m - nothing else. The 53 roles that
  carry "Leadership" in REVIEW column K are Heads of Technology, Technology Managers and
  Delivery Managers, and those are priced by their own overhead lines at a 30% allocation
  to one platform. No line is a 1:1 headcount rule.

So every line in the 0.2 Data Config build-up gets its matching named population, and the
workbook shows allowance against actual for each:

  line                 rate    x units          allowance   actual        variance
  Head of Technology   0.1375  x 10 portfolios      1.375   11   3.383      +2.008
  Business Partner     0.22    x 10 portfolios      2.200    6   2.314      +0.114
  Domain Architect     0.14    x 10 portfolios      1.400    7   1.665      +0.265
  Delivery Manager     0.084   x 30 platforms       2.520   10   2.675      +0.155
  Technology Manager   0.081   x 30 platforms       2.430   24   6.298      +3.868
  Leadership (8 GMs)   0.30    x 10 portfolios      3.000    8   5.100      +2.100
                                                   12.925       21.435      +8.510

The 8 GMs are the only layer with no role in the ledger, so their $5.1m sits ABOVE the
$115.113m rather than inside it. Every other line is people already counted, which is why
overhead is compared here and never added to a portfolio's cost.
"""
import openpyxl
from openpyxl.styles import Font

import model

REV = f"'{model.REVIEW}'"
LR = model.LAST_ROW
BOLD = Font(bold=True)
ITAL = Font(italic=True)

# label, 0.2 rate cell, units, unit basis, title match for the actual population
OVERHEAD = [
    ("Head of Technology", "'0.2 Data Config'!$L$6", 10, "portfolios"),
    ("Business Partner", "'0.2 Data Config'!$L$7", 10, "portfolios"),
    ("Domain Architect", "'0.2 Data Config'!$L$8", 10, "portfolios"),
    ("Delivery Manager", "'0.2 Data Config'!$L$14", 30, "platforms"),
    ("Technology Manager", "'0.2 Data Config'!$L$15", 30, "platforms"),
    ("Leadership - 8 GMs", "'0.2 Data Config'!$L$9", 10, "portfolios"),
]
GM_COUNT, GM_COST = 8, 5.1


def review_class(wb):
    """AR now names the overhead line a role belongs to, so each line has a population."""
    ws = wb[model.REVIEW]
    ws["AR1"] = "Overhead line"
    ws["AR1"].font = BOLD
    for i in range(2, LR + 1):
        ws[f"AR{i}"] = (
            f'=IF(TRIM($B{i})="","",'
            f'IF(ISNUMBER(SEARCH("head of technology",$C{i})),"Head of Technology",'
            f'IF(ISNUMBER(SEARCH("TDD BP",$C{i})),"Business Partner",'
            f'IF(OR(ISNUMBER(SEARCH("domain architect",$C{i})),'
            f'ISNUMBER(SEARCH("enterprise architect",$C{i}))),"Domain Architect",'
            f'IF(ISNUMBER(SEARCH("delivery man",$C{i})),"Delivery Manager",'
            f'IF(OR(ISNUMBER(SEARCH("technology manager",$C{i})),'
            f'ISNUMBER(SEARCH("technology manger",$C{i})),'
            f'ISNUMBER(SEARCH("tech manager",$C{i}))),"Technology Manager",'
            f'"Squad"))))))'
        )
    return ["REVIEW AR: overhead line per role (6 lines + Squad)"]


def lists_allowance(wb):
    """Rebuild the allowance table so each line carries its rate, units and basis."""
    ws = wb["Lists"]
    for r in range(1, 12):
        for c in range(32, 36):          # AF..AI
            ws.cell(r, c).value = None
    hdr = ["Overhead line", "Rate ($m)", "Units", "Basis", "Allowance ($m)"]
    for i, h in enumerate(hdr):
        c = ws.cell(1, 32 + i)
        c.value = h
        c.font = BOLD
    for i, (lab, rate_ref, units, basis) in enumerate(OVERHEAD):
        r = 2 + i
        ws.cell(r, 32).value = lab
        ws.cell(r, 33).value = f"={rate_ref}"
        ws.cell(r, 34).value = units
        ws.cell(r, 35).value = basis
        ws.cell(r, 36).value = f"=$AG{r}*$AH{r}"
    tr = 2 + len(OVERHEAD)
    ws.cell(tr, 32).value = "Total"
    ws.cell(tr, 32).font = BOLD
    ws.cell(tr, 36).value = f"=SUM(AJ2:AJ{tr-1})"
    ws.cell(tr, 36).font = BOLD
    for col, w in [("AF", 22), ("AG", 12), ("AH", 10), ("AI", 12), ("AJ", 16)]:
        ws.column_dimensions[col].width = w
    return ["Lists AF:AJ allowance table rebuilt (rate x units, 6 lines)"]


def fix_lists_k(wb):
    """Lists K2:K12 still pointed at 3.3's pre-rebuild portfolio rows."""
    ws = wb["Lists"]
    order = list(dict.fromkeys(model.TAB_PORTFOLIO.values()))
    for i, pf in enumerate(order[:10]):
        r = 2 + i
        # archetype roles for the portfolio, read from 3.2 rather than a 3.3 row number
        ws.cell(r, 11).value = (f"=SUMIFS('3.3 FTE View'!$F$6:$F$94,"
                                f"'3.3 FTE View'!$B$6:$B$94,$J{r},"
                                f"'3.3 FTE View'!$C$6:$C$94,\"<>Portfolio total\")")
    ws["K12"] = None
    ws["J12"] = None
    return ["Lists K2:K11 repointed (were reading pre-rebuild 3.3 rows); K12 cleared"]


def fix_34_leftovers(wb):
    """The EGI row still carried the old total formulas in columns I and J."""
    ws = wb["3.4 COE Summary"]
    ws["I11"] = (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},"EGI",'
                 f'{REV}!$M$2:$M${LR},"Australia")/1000000')
    ws["J11"] = (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},"EGI",'
                 f'{REV}!$M$2:$M${LR},"NZ")/1000000')
    return ["3.4 I11/J11 were leftover =SUM(I6:I10) totals, double counting into row 12"]


def overhead_block(wb):
    """One overhead statement, on 3.2, under the portfolio table."""
    ws = wb["3.2 Total Cost"]
    # find the existing shared-role block and clear it
    start = None
    for r in range(20, 60):
        if str(ws.cell(r, 2).value or "").startswith("Shared roles"):
            start = r
            break
    if start is None:
        start = 26
    for r in range(start, start + 16):
        for c in range(2, 10):
            ws.cell(r, c).value = None

    ws.cell(start, 2).value = ("Overhead - allowance against actual. Every line below is "
                               "priced in 0.2 Data Config and staffed by named people in "
                               "REVIEW, so overhead is compared here and never added to a "
                               "portfolio's cost.")
    ws.cell(start, 2).font = ITAL
    h = start + 1
    for col, lab in (("B", "Overhead line"), ("C", "Roles"), ("D", "Rate ($m)"),
                     ("E", "Units"), ("F", "Actual ($m)"), ("G", "Allowance ($m)"),
                     ("H", "Variance ($m)")):
        ws[f"{col}{h}"] = lab
        ws[f"{col}{h}"].font = BOLD
    for i, (lab, _, _, _) in enumerate(OVERHEAD):
        r = h + 1 + i
        lr = 2 + i
        ws[f"B{r}"] = f"=Lists!$AF${lr}"
        ws[f"D{r}"] = f"=Lists!$AG${lr}"
        ws[f"E{r}"] = f"=Lists!$AH${lr}"
        ws[f"G{r}"] = f"=Lists!$AJ${lr}"
        if lab.startswith("Leadership"):
            # the 8 GMs are the one layer with no role in the ledger
            ws[f"C{r}"] = GM_COUNT
            ws[f"F{r}"] = GM_COST
        else:
            ws[f"C{r}"] = f'=COUNTIFS({REV}!$AR$2:$AR${LR},$B{r})'
            ws[f"F{r}"] = (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AR$2:$AR${LR},$B{r})'
                           f"/1000000")
        ws[f"H{r}"] = f"=$F{r}-$G{r}"
    tot = h + 1 + len(OVERHEAD)
    ws[f"B{tot}"] = "Overhead total"
    ws[f"B{tot}"].font = BOLD
    for col in "CFGH":
        ws[f"{col}{tot}"] = f"=SUM({col}{h+1}:{col}{tot-1})"

    sq = tot + 1
    ws[f"B{sq}"] = "Squad (delivery) roles"
    ws[f"C{sq}"] = f'=COUNTIFS({REV}!$AR$2:$AR${LR},"Squad")'
    ws[f"F{sq}"] = f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AR$2:$AR${LR},"Squad")/1000000'

    led = sq + 1
    ws[f"B{led}"] = "Ledger total - the 525 named roles"
    ws[f"B{led}"].font = BOLD
    ws[f"C{led}"] = f"=$C{tot}+$C{sq}-{GM_COUNT}"
    ws[f"F{led}"] = f"=$F{tot}+$F{sq}-{GM_COST}"
    ws[f"B{led+1}"] = "Ledger control - must be 0"
    ws[f"B{led+1}"].font = ITAL
    ws[f"C{led+1}"] = f"=COUNTA({REV}!$B$2:$B${LR})-$C{led}"
    ws[f"F{led+1}"] = (f'=ROUND(SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$B$2:$B${LR},"<>")'
                       f"/1000000-$F{led},6)")

    org = led + 3
    ws[f"B{org}"] = "Total organisation including the GM layer ($m)"
    ws[f"B{org}"].font = BOLD
    ws[f"F{org}"] = f"=$F{led}+{GM_COST}"
    ws[f"B{org+1}"] = (f"The {GM_COUNT} GMs are the only overhead layer with no role in the "
                       f"ledger, so their ${GM_COST}m sits above the 525 rather than inside "
                       f"them. Every other line is people already counted.")
    ws[f"B{org+1}"].font = ITAL
    return [f"3.2 overhead block rebuilt at rows {start}-{org+1} (6 lines, GM layer separate)"]


def fix_exec_allowance(wb):
    """Exec Summary referenced Lists!$AI$6, the old allowance total. AI is now the
    unit basis (text), so the designed-cost lines errored. Point at the new total."""
    ws = wb["Exec Summary"]
    n = 0
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "Lists!$AI$6" in c.value:
                c.value = c.value.replace("Lists!$AI$6", "Lists!$AJ$8")
                n += 1
    ws["B24"] = "Overhead allowance - all six lines from 0.2 Data Config ($m)"
    return [f"Exec Summary: {n} designed-cost formulas repointed to the new allowance total"]


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    out = []
    for fn in (review_class, lists_allowance, fix_lists_k, fix_34_leftovers,
               overhead_block, fix_exec_allowance, fix_lists_portfolios,
               fix_exec_drilldown, gm_inputs_to_lists):
        out += fn(wb)
    wb.save(dst)
    return out




def fix_exec_drilldown(wb):
    """Exec Summary rows 61-77 are a portfolio drill-down that survived the 3.1 rebuild.

    3.1's columns were repurposed, so four lines read the wrong quantity under the old
    label - row 67 said "Funded outside TDD ($m)" and returned 22, which is Ampol
    Retail's VACANT ROLE COUNT. Rows 69 and 70 read columns K and M, which no longer
    exist, and returned a silent 0. The dropdown also offered "TDD Cyber" where 3.1 now
    says "COE Cyber", and omitted BP&T, SA&D and EGI.
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    ws = wb["Exec Summary"]
    order = list(dict.fromkeys(model.TAB_PORTFOLIO.values()))
    G31 = "'3.1 Group Summary'"
    RNG = "$B$6:$B$19"

    def idx(col):
        return f"=IFERROR(INDEX({G31}!${col}$6:${col}$19,MATCH($C$63,{G31}!{RNG},0)),\"-\")"

    lines = [
        (64, "TDD budget ($m)", idx("C")),
        (65, "Actual cost of the portfolio ($m)", idx("D")),
        (66, "Variance to budget ($m)", idx("E")),
        (67, "Archetype cost ($m)", idx("F")),
        (68, "Variance to archetype ($m)", idx("G")),
        (69, "Cost after decisions ($m)", idx("J")),
        (70, "Roles", idx("H")),
        (71, "Archetype squad roles allowed",
         f"=IFERROR(INDEX(Lists!$K$2:$K$15,MATCH($C$63,Lists!$J$2:$J$15,0)),\"-\")"),
    ]
    for r, lab, f in lines:
        ws.cell(r, 2).value = lab
        ws.cell(r, 3).value = f
    # the four COUNTIFS below ran to row 530; the ledger ends at 528
    for r in range(72, 76):
        v = ws.cell(r, 3).value
        if isinstance(v, str) and "$530" in v:
            ws.cell(r, 3).value = v.replace("$530", f"${LR}")

    for dv in list(ws.data_validations.dataValidation):
        if "C63" in str(dv.sqref):
            ws.data_validations.dataValidation.remove(dv)
    dv = DataValidation(type="list", formula1=f'"{",".join(order)}"', allow_blank=False)
    dv.errorTitle, dv.error = "Portfolio", "Pick a portfolio from the list."
    ws.add_data_validation(dv)
    dv.add("C63")
    if str(ws["C63"].value or "") not in order:
        ws["C63"] = order[0]
    return ["Exec Summary drill-down remapped to the rebuilt 3.1 columns; dropdown now "
            "lists all 14 portfolios with the names 3.1 uses"]


def fix_lists_portfolios(wb):
    """Lists J/K carried 12 portfolios under the old naming. Make it the real 14."""
    ws = wb["Lists"]
    order = list(dict.fromkeys(model.TAB_PORTFOLIO.values()))
    for r in range(2, 20):
        ws.cell(r, 10).value = None
        ws.cell(r, 11).value = None
    for i, pf in enumerate(order):
        r = 2 + i
        ws.cell(r, 10).value = pf
        ws.cell(r, 11).value = (f"=SUMIFS('3.3 FTE View'!$F$6:$F$94,"
                                f"'3.3 FTE View'!$B$6:$B$94,$J{r},"
                                f"'3.3 FTE View'!$C$6:$C$94,\"<>Portfolio total\")")
    return [f"Lists J2:K{1+len(order)} now lists all {len(order)} portfolios"]




def gm_inputs_to_lists(wb):
    """The GM count and cost were typed into 3.2. Make them named inputs on Lists.

    They are the one overhead line with no role in the ledger, so they have to be entered
    somewhere - but an input belongs on a control tab where it is labelled and editable,
    not buried as a literal in the middle of a summary.
    """
    ws = wb["Lists"]
    ws["AF10"] = "GM layer - input"
    ws["AF10"].font = BOLD
    ws["AF11"] = "Number of GMs"
    ws["AG11"] = GM_COUNT
    ws["AF12"] = "GM cost ($m)"
    ws["AG12"] = GM_COST
    ws["AF13"] = ("The 8 GMs are the only overhead line with no role in REVIEW, so their "
                  "cost is entered here and sits above the 525-role ledger.")
    ws["AF13"].font = ITAL

    g = wb["3.2 Total Cost"]
    for r in range(26, 46):
        if str(g.cell(r, 2).value or "").startswith("=Lists!$AF$7") or \
           "Leadership" in str(g.cell(r, 2).value or ""):
            if isinstance(g.cell(r, 3).value, (int, float)):
                g.cell(r, 3).value = "=Lists!$AG$11"
            if isinstance(g.cell(r, 6).value, (int, float)):
                g.cell(r, 6).value = "=Lists!$AG$12"
        lab = str(g.cell(r, 2).value or "")
        if lab.startswith("Ledger total"):
            g.cell(r, 3).value = g.cell(r, 3).value.replace("-8", "-Lists!$AG$11")
            g.cell(r, 6).value = g.cell(r, 6).value.replace("-5.1", "-Lists!$AG$12")
        if lab.startswith("Total organisation"):
            g.cell(r, 6).value = g.cell(r, 6).value.replace("+5.1", "+Lists!$AG$12")
    return ["GM count and cost moved from 3.2 literals to Lists AG11/AG12"]


if __name__ == "__main__":
    for x in run("final.xlsx", "oh.xlsx"):
        print("  ", x)
