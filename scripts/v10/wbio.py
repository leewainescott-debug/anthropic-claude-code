"""Safe build harness for TDD_Cost_Calc.

The trap this avoids: openpyxl.save() writes formulas but strips every cached value,
so the file opens blank in Excel until the user hits recalc. Fix: recalculate with a
real engine, then inject the resulting values back into the openpyxl output as <v>
elements, and set fullCalcOnLoad so Excel recomputes anyway on open.

Pipeline:  edit (openpyxl) -> recalc (LibreOffice) -> inject values -> assert
"""
import os, re, shutil, subprocess, zipfile, tempfile
import openpyxl

SP = os.path.dirname(os.path.abspath(__file__))
LOHOME = os.path.join(os.path.dirname(SP), "lohome")
ERRS = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "Err:")


def recalc(src, outdir=None):
    """Full-recalculate src with LibreOffice; return path to the recalculated copy."""
    outdir = outdir or tempfile.mkdtemp(prefix="recalc_", dir=SP)
    os.makedirs(LOHOME, exist_ok=True)
    env = dict(os.environ, HOME=LOHOME)
    r = subprocess.run(
        ["soffice", f"-env:UserInstallation=file://{LOHOME}/.lo3", "--headless",
         "--norestore", "--convert-to", "xlsx", "--outdir", outdir, src],
        capture_output=True, text=True, timeout=2400, env=env)
    out = os.path.join(outdir, os.path.basename(src))
    if not os.path.exists(out):
        raise RuntimeError(f"recalc failed: {r.stdout[-800:]} {r.stderr[-800:]}")
    return out


def harvest(path):
    """{sheet: {coord: value}} of every cached value in a recalculated workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for ws in wb.worksheets:
        d = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    d[c.coordinate] = c.value
        out[ws.title] = d
    wb.close()
    return out


def _sheet_part_map(path):
    """sheet title -> xl/worksheets/sheetN.xml, via workbook.xml + rels."""
    z = zipfile.ZipFile(path)
    wbx = z.read("xl/workbook.xml").decode("utf8")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf8")
    # attribute order varies by writer (Excel puts Id first, openpyxl puts it last)
    rid2t = {}
    for tag in re.findall(r"<Relationship\b[^>]*/?>", rels):
        rid = re.search(r'\bId="([^"]+)"', tag)
        tgt = re.search(r'\bTarget="([^"]+)"', tag)
        if rid and tgt:
            rid2t[rid.group(1)] = tgt.group(1)
    out = {}
    for tag in re.findall(r"<sheet\b[^>]*/?>", wbx):
        nm = re.search(r'\bname="([^"]+)"', tag)
        ri = re.search(r'\br:id="([^"]+)"', tag)
        if not (nm and ri):
            continue
        name, rid = nm.group(1), ri.group(1)
        t = rid2t.get(rid, "")
        if t.startswith("/xl/"):
            t = t[1:]
        elif not t.startswith("xl/"):
            t = "xl/" + t.lstrip("/")
        out[_unesc(name)] = t
    z.close()
    return out


def _unesc(s):
    return (s.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
             .replace("&quot;", '"').replace("&apos;", "'"))


def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def inject(target, values, set_full_calc=True):
    """Write cached values into target's XML for every cell that holds a formula.

    values: {sheet: {coord: value}} from harvest().
    Formulas keep their <f>; we add/replace the sibling <v> (and t= for strings/bools).
    """
    parts = _sheet_part_map(target)
    z = zipfile.ZipFile(target)
    names = z.namelist()
    blobs = {n: z.read(n) for n in names}
    z.close()

    stats = {"cells": 0, "sheets": 0}
    for sheet, part in parts.items():
        vals = values.get(sheet)
        if not vals or part not in blobs:
            continue
        xml = blobs[part].decode("utf8")
        changed = [0]

        def fix(m):
            whole = m.group(0)
            if whole.endswith("/>"):          # self-closing empty cell: nothing to inject
                return whole
            attrs, body = m.group(1), m.group(2)
            rm = re.search(r'\br="([A-Z]+\d+)"', attrs)
            if not rm or "<f" not in body:
                return whole
            if rm.group(1) not in vals:
                return whole
            v = vals[rm.group(1)]
            fm = re.search(r"<f[^>]*/>|<f[^>]*>.*?</f>", body, re.S)
            if not fm:
                return whole
            fpart = fm.group(0)
            a = re.sub(r'\s+t="[^"]*"', "", attrs)
            if isinstance(v, bool):
                a += ' t="b"'
                vx = "<v>%d</v>" % (1 if v else 0)
            elif isinstance(v, (int, float)):
                vx = "<v>%r</v>" % v if isinstance(v, float) else "<v>%d</v>" % v
            elif isinstance(v, str) and any(e in v for e in ERRS):
                a += ' t="e"'
                vx = "<v>%s</v>" % _esc(v)
            else:
                a += ' t="str"'
                vx = "<v>%s</v>" % _esc(v)
            changed[0] += 1
            return "<c%s>%s%s</c>" % (a, fpart, vx)

        # self-closing alternative MUST come first, else an empty <c/> swallows the
        # following cell and its formula is silently left without a cached value
        xml = re.sub(r"<c\b[^>]*/>|<c\b([^>]*)>(.*?)</c>", fix, xml, flags=re.S)
        blobs[part] = xml.encode("utf8")
        stats["cells"] += changed[0]
        stats["sheets"] += 1

    if set_full_calc:
        wbx = blobs["xl/workbook.xml"].decode("utf8")
        if "<calcPr" in wbx:
            wbx = re.sub(r"<calcPr[^>]*/>",
                         '<calcPr calcId="191029" fullCalcOnLoad="1"/>', wbx)
        else:
            wbx = wbx.replace("</workbook>",
                              '<calcPr calcId="191029" fullCalcOnLoad="1"/></workbook>')
        blobs["xl/workbook.xml"] = wbx.encode("utf8")

    tmp = target + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for n in names:
            out.writestr(n, blobs[n])
    os.replace(tmp, target)
    return stats


def build(src, dst):
    """src already carries the edited formulas (openpyxl-saved). Populate dst.

    Asserts injection completeness: every formula cell the engine gave a value must
    end up with that value in dst. A silent miss here ships a blank cell to Lee.
    """
    shutil.copy(src, dst)
    rc = recalc(dst)
    vals = harvest(rc)
    st = inject(dst, vals)

    wf = openpyxl.load_workbook(dst, data_only=False)
    wv = openpyxl.load_workbook(dst, data_only=True)
    missed = []
    for sn in wf.sheetnames:
        want = vals.get(sn, {})
        f, v = wf[sn], wv[sn]
        for row in f.iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and c.value.startswith("=")):
                    continue
                exp = want.get(c.coordinate)
                if exp is None:
                    continue
                got = v[c.coordinate].value
                if got is None:
                    missed.append((sn, c.coordinate, exp))
                elif isinstance(exp, (int, float)) and isinstance(got, (int, float)):
                    if abs(exp - got) > max(1e-9, abs(exp) * 1e-12):
                        missed.append((sn, c.coordinate, f"{exp} != {got}"))
    wf.close(); wv.close()
    if missed:
        raise AssertionError(
            f"injection incomplete: {len(missed)} cells, e.g. {missed[:8]}")
    st["verified"] = True
    return rc, st


def audit(path):
    """Return (errors, blank_formula_cells) for a populated workbook."""
    wf = openpyxl.load_workbook(path, data_only=False)
    wv = openpyxl.load_workbook(path, data_only=True)
    errors, blanks = [], []
    for sn in wf.sheetnames:
        f, v = wf[sn], wv[sn]
        for row in f.iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and c.value.startswith("=")):
                    continue
                cv = v[c.coordinate].value
                if isinstance(cv, str) and any(e in cv for e in ERRS):
                    errors.append((sn, c.coordinate, cv, c.value[:90]))
                elif cv is None:
                    blanks.append((sn, c.coordinate, c.value[:90]))
    wf.close(); wv.close()
    return errors, blanks
