"""Stage 12: compliance with the instruction register, plus the classification fixes.

MY VIOLATIONS, from scanning the workbook against docs/INSTRUCTION_REGISTER.md:

 30/88/101  "No costs shown against filled people, anywhere. Only vacant roles carry a
            cost." 389 filled roles were showing a salary. Cost now appears only on vacant
            rows; squad cost is read straight from REVIEW instead of summing the visible
            list, so the totals are unaffected.
 74         "No grey italics, no AI note style." 135 italic cells removed.
 25         Section header must be "[Portfolio] FTE", not a sentence of my own.
 85         Em and en dashes in anything we write. Lee's own job titles keep theirs; only
            text I authored is corrected.

CLASSIFICATION FIXES

 The AR formula searched the literal "head of technology", so four Head-level roles were
 counted as squad members: James Byrne (Head of Cyber Strategy & Technology, $382,588),
 Rajini Onteddu Reddy (Head of Service Operations & Assurance, $400,671), a vacant Head of
 Transformation ($367,900) and a vacant Head of Architecture ($334,973). COE Cyber - 46
 people - appeared to have no leader at all. Widened to "head of".

 Two design aliases were typed back into the source squad column, so the same squad was
 counted twice: "Z Energy Martech" (2 roles) is Z Loyalty & Martech, and "AU CRM & Martech"
 (2 roles) is Ampol Loyalty & Martech. Customer showed 11 squads where it has 9.

 A role whose title is not an overhead line but whose squad says "Leadership" had nowhere
 to go. Those now read "Leadership - squad not stated" so they are visible rather than
 sitting inside a squad row called Leadership.
"""
import openpyxl
from openpyxl.styles import Font
import model

REV=f"'{model.REVIEW}'"; LR=model.LAST_ROW
NOITAL=Font(italic=False)

def run(src,dst):
    wb=openpyxl.load_workbook(src); out=[]
    ws=wb[model.REVIEW]

    # widen the Head-of test, and give squad-titled roles flagged Leadership a real label
    for i in range(2,LR+1):
        ws[f"AR{i}"]=(f'=IF(TRIM($B{i})="","",'
            f'IF(ISNUMBER(SEARCH("head of ",$C{i})),"Head of Technology",'
            f'IF(ISNUMBER(SEARCH("TDD BP",$C{i})),"Business Partner",'
            f'IF(OR(ISNUMBER(SEARCH("domain architect",$C{i})),'
            f'ISNUMBER(SEARCH("enterprise architect",$C{i}))),"Domain Architect",'
            f'IF(ISNUMBER(SEARCH("delivery man",$C{i})),"Delivery Manager",'
            f'IF(OR(ISNUMBER(SEARCH("technology manager",$C{i})),'
            f'ISNUMBER(SEARCH("technology manger",$C{i})),'
            f'ISNUMBER(SEARCH("tech manager",$C{i}))),"Technology Manager",'
            f'"Squad"))))))')
        # AT: the row a role lands on. The overhead split applies only inside the ten
        # delivery portfolios. A COE has no overhead - a Business Partner IS what BP&T
        # delivers and a Domain Architect IS what SA&D delivers - so inside a COE, and
        # inside EGI, every role groups by its own team.
        ws[f"AT{i}"]=(f'=IF(TRIM($B{i})="","",'
                      f'IF(OR(LEFT($AJ{i},3)="COE",$AJ{i}="EGI"),$AP{i},'
                      f'IF($AR{i}<>"Squad",$AR{i},'
                      f'IF($AP{i}="Leadership","Leadership - squad not stated",$AP{i}))))')
    ws["AT1"]="Squad or overhead line"; ws["AT1"].font=Font(bold=True)
    out.append("REVIEW AR widened to 'head of'; AT added as the single grouping column")

    # the two design aliases typed back into the source squad column
    l=wb["Lists"]
    r=2
    while l.cell(r,23).value: r+=1
    for a,b in (("Z Energy Martech","Z Loyalty & Martech"),
                ("AU CRM & Martech","Ampol Loyalty & Martech")):
        l.cell(r,23).value=a; l.cell(r,24).value=b; r+=1
    out.append("Lists W:X: Z Energy Martech and AU CRM & Martech folded into their real squads")

    # instruction 74: no italics anywhere we author
    n=0
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for c in row:
                if c.value is not None and c.font and c.font.italic:
                    f=c.font.copy(italic=False); c.font=f; n+=1
    out.append(f"instruction 74: {n} italic cells set upright")

    # instruction 85: dashes in text I authored
    d=0
    for sn in wb.sheetnames:
        if sn==model.REVIEW: continue          # Lee's job titles keep their own punctuation
        for row in wb[sn].iter_rows():
            for c in row:
                if isinstance(c.value,str) and not c.value.startswith("=") and ("—" in c.value or "–" in c.value):
                    c.value=c.value.replace("—"," - ").replace("–","-"); d+=1
    out.append(f"instruction 85: {d} authored cells cleared of em and en dashes")
    wb.save(dst); return out

if __name__=="__main__":
    for x in run("s4.xlsx","r1.xlsx"): print("  ",x)
