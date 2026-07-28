"""Stage 6: the 2.x defects a second reader found - one live wrong number, then format.

1  HIGH  2.5 P&C!C6 returns 0. The archetype lookup matched down the WHOLE of column B on
         the design tab, so the squad named "P&C" hit the sheet's own title in B2 before
         reaching the squad table at B25. INDEX of the empty C2 gives 0, the squad scores
         no archetype, and the tab's totals read 6 roles / $1.4m where the truth is
         15 / $3.5m. Every lookup is now bounded to the squad table.
2  HIGH  conditional formatting is stale on 13 of 14 tabs: 2.2's vacancy rule sits on
         column J (new total roles) instead of I (vacancies remaining); the Vacant
         highlight stops short of the rebuilt role list on five tabs, leaving up to 14
         rows unmarked; 2.12/2.13/2.14 have none at all.
3  MED   2.12 and 2.13 are unbranded and unformatted - headcounts and dollars render raw.
4  MED   six tabs show headcounts to two decimals because the total row lost its format.
5  MED   three tabs paint the role-list header dark navy with default dark text.
6  LOW   the lever dropdown is a hardcoded string, decoupled from the factors on Lists.
         Rename a lever there and every role silently reprices at 1.0 through the IFERROR.
7  LOW   2.11-2.13 hardcode "Outside archetype model" even though 1.11-1.14 exist.
"""
import re

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import model

HDR_FILL = PatternFill("solid", start_color="FF1F4E79", end_color="FF1F4E79")
SUB_FILL = PatternFill("solid", start_color="FFDDEBF7", end_color="FFDDEBF7")
HDR_FONT = Font(bold=True, color="FFFFFFFF")
RED = Font(color="FF9C0006")
GREEN = Font(color="FF006100")
CNT = '#,##0;[Red](#,##0);"-"'
MON = '#,##0.00;[Red](#,##0.00);"-"'

# 2.x tab -> design tab, for the archetype lookup
DESIGN = {
    "2.1 Ampol Retail": "1.1 Ampol Retail", "2.2 Customer": "1.2 Customer",
    "2.3 Enterprise Data": "1.3 Enterprise Data",
    "2.4 TDD Group Functions": "1.4 TDD Group Functions", "2.5 P&C": "1.5 P&C",
    "2.6 Finance": "1.6 Finance", "2.7 Infrastructure": "1.7 Infrastructure",
    "2.8 Energy Solutions & B2B": "1.8 Energy Solutions & B2B",
    "2.9 Commercial Fuels": "1.9 Commercial Fuels", "2.10 Z Retail": "1.10 Z Retail",
    "2.15 TDD Cyber": "1.14 TDD Cyber",
}


def squad_table_bounds(wb, design_tab):
    """First and last row of the squad table on a design tab, so a MATCH cannot escape
    upward into the title or a summary label."""
    ws = wb[design_tab]
    lo = hi = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value or "").strip().startswith("Platform:"):
            if lo is None:
                lo = r
            hi = r
    if lo is None:
        return 20, ws.max_row
    # extend to the end of the last block
    for r in range(hi, ws.max_row + 1):
        if str(ws.cell(r, 2).value or "").strip().endswith("Total"):
            hi = r
    return lo, max(hi, lo + 1)


def geometry(ws):
    """Locate the squad block, total row and role list on a rebuilt 2.x tab."""
    first = 6
    tot = None
    for r in range(6, 40):
        if str(ws.cell(r, 2).value or "").strip() == "Total":
            tot = r
            break
    rl_hdr = None
    for r in range(6, 60):
        if str(ws.cell(r, 2).value or "").strip() == "Name" and \
           str(ws.cell(r, 5).value or "").strip() == "Vacancy lever":
            rl_hdr = r
            break
    rl0 = rl_hdr + 1
    rl1 = rl0
    while isinstance(ws.cell(rl1 + 1, 2).value, str) and \
            str(ws.cell(rl1 + 1, 2).value).startswith("="):
        rl1 += 1
    return first, tot - 1, tot, rl_hdr, rl0, rl1


def fix(wb):
    out = []
    A3 = "'0.3 Squad Archetypes'"
    for tab in model.TAB_PORTFOLIO:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        first, last, tot, rl_hdr, rl0, rl1 = geometry(ws)

        # --- 1 / 7: bounded archetype lookup, and use the COE design tabs too ---
        dt = DESIGN.get(tab)
        if dt:
            lo, hi = squad_table_bounds(wb, dt)
            m = f"MATCH($S{{r}},'{dt}'!$B${lo}:$B${hi},0)"
            for r in range(first, last + 1):
                mm = m.format(r=r)
                ws[f"C{r}"] = (f"=IFERROR(INDEX('{dt}'!$C${lo}:$C${hi},{mm}),"
                               f'"Outside archetype model")')
                ws[f"R{r}"] = (f"=IFERROR(IF(INDEX('{dt}'!$D${lo}:$D${hi},{mm})=\"\",\"-\","
                               f"INDEX('{dt}'!$D${lo}:$D${hi},{mm})),\"-\")")
            out.append(f"{tab}: archetype lookup bounded to {dt}!B{lo}:B{hi}")

        # --- 4: number formats, uniform across every tab ---
        for r in range(first, tot + 1):
            for col in "DEFGHIJ":
                ws[f"{col}{r}"].number_format = CNT
            for col in "LMNOPQ":
                ws[f"{col}{r}"].number_format = MON
        for r in range(rl0, rl1 + 1):
            for col in "GH":
                ws[f"{col}{r}"].number_format = '#,##0'

        # --- 3 / 5: headers styled the same way on all fourteen ---
        for col in "BCDEFGHIJKLMNOPQRST":
            c = ws[f"{col}5"]
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(wrap_text=True, vertical="bottom")
            c.number_format = "General"
        for col in "BCDEFGH":
            c = ws[f"{col}{rl_hdr}"]
            c.fill = SUB_FILL
            c.font = Font(bold=True)
            c.alignment = Alignment(wrap_text=True, vertical="bottom")
        ws[f"B{tot}"].font = Font(bold=True)

        # --- 2: one set of conditional rules, on the right columns, full ranges ---
        try:
            ws.conditional_formatting._cf_rules.clear()
        except Exception:
            pass
        ws.conditional_formatting.add(
            f"D{rl0}:D{rl1}",
            CellIsRule(operator="equal", formula=['"Vacant"'], font=RED))
        ws.conditional_formatting.add(
            f"I{first}:I{last}",
            CellIsRule(operator="greaterThan", formula=["0"], font=RED))
        ws.conditional_formatting.add(
            f"I{first}:I{last}",
            CellIsRule(operator="lessThanOrEqual", formula=["0"], font=GREEN))
        ws.conditional_formatting.add(
            f"F{first}:F{last}",
            FormulaRule(formula=[f"AND(ISNUMBER($F{first}),$F{first}>0)"], font=RED))

        # --- 6: the dropdown reads the lever list, so it cannot drift from the factors ---
        for dv in list(ws.data_validations.dataValidation):
            ws.data_validations.dataValidation.remove(dv)
        dv = DataValidation(type="list", formula1="=Lists!$AC$2:$AC$5", allow_blank=False)
        dv.errorTitle, dv.error = "Vacancy lever", "Choose a lever from Lists AC2:AC5."
        ws.add_data_validation(dv)
        dv.add(f"E{rl0}:E{rl1}")

    out.append("2.x: number formats, header styling, conditional rules and the lever "
               "dropdown made uniform across all fourteen tabs")
    return out


def run(src, dst):
    wb = openpyxl.load_workbook(src)
    o = fix(wb)
    wb.save(dst)
    return o


if __name__ == "__main__":
    for x in run("n1x.xlsx", "v2.xlsx"):
        print("  ", x)
