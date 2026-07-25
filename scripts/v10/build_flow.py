"""Stage 2: rebuild the 2.x working copies and the 3.x summaries as one flow.

Chain, end to end, all formula-driven:

  REVIEW (Lee's cols B/I/J/K/AA)
    -> AP canonical squad, AQ leadership, AR cost class, AS design-tab squad name
      -> 2.x squad rows: archetype from 1.x design tab + 0.3 rate, actual from the ledger
        -> 3.1 / 3.2 / 3.3 portfolio and class summaries
          -> Exec Summary

No tab re-states a number another tab already owns. Overhead is never added to actual
cost: shared roles are classified once in REVIEW and compared to the allowance on Lists.
"""
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

import model
from build_2x import HDRS, ROLE_HDRS, DESIGN, LEVER_LIST  # DESIGN = the 10 archetype-based tabs

REV = f"'{model.REVIEW}'"
LR = model.LAST_ROW
A3 = "'0.3 Squad Archetypes'"
BOLD = Font(bold=True)
ITAL = Font(italic=True)
CLASSES = ["Head of Technology", "Leadership", "Business Partner", "Architecture"]

# 2.x tabs get the BP + Architecture allowance only: those two populations sit in the
# COEs, outside the portfolio, so they are the one part of the allowance a portfolio
# consumes without already carrying. Leadership and Head of Technology are already
# named people inside the portfolio's own cost, so adding them would double count.
ADDITIVE = "(Lists!$AG$4+Lists!$AG$5)"


def hdr_row(ws, row, mapping):
    for col, h in mapping.items():
        c = ws[f"{col}{row}"]
        c.value = h
        c.font = BOLD
        c.alignment = Alignment(wrap_text=True, vertical="bottom")


def build_2x(wb, roles):
    out = {}
    for tab, pf in model.TAB_PORTFOLIO.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        mine = [r for r in roles if r["portfolio"] == pf]

        def order(s):
            tail = 2 if s == model.BLANK_SQUAD_LABEL else (1 if s == "Leadership" else 0)
            return (tail, -sum(1 for r in mine if r["squad"] == s), s)
        squads = sorted({r["squad"] for r in mine}, key=order)

        title = ws.cell(2, 2).value
        for r in range(4, ws.max_row + 1):
            for c in range(1, max(ws.max_column, 19) + 1):
                if ws.cell(r, c).value is not None:
                    ws.cell(r, c).value = None
        for dv in list(ws.data_validations.dataValidation):
            ws.data_validations.dataValidation.remove(dv)

        hdr, first = 5, 6
        last = first + len(squads) - 1
        tot = last + 1
        m1, m2, m3 = tot + 2, tot + 3, tot + 4
        rl_hdr = m3 + 3
        rl0, rl1 = rl_hdr + 1, rl_hdr + len(mine)
        RL = lambda c: f"${c}${rl0}:${c}${rl1}"

        ws.cell(4, 2).value = ("Position by squad - archetype vs actual. Roles, status and "
                               "cost come from REVIEW; the archetype comes from the design tab.")
        ws.cell(4, 2).font = ITAL
        hdr_row(ws, hdr, HDRS)
        hdr_row(ws, hdr, {"S": "Squad name on design tab",
                          "T": "Archetype basis"})

        dt = DESIGN.get(tab, (None,))[0]
        for i, sq in enumerate(squads):
            r = first + i
            ws[f"B{r}"] = sq
            # the confirmed ledger -> design name map, so an archetype can be found at all
            ws[f"S{r}"] = f"=IFERROR(INDEX(Lists!$AA:$AA,MATCH($B{r},Lists!$Z:$Z,0)),$B{r})"
            if dt:
                ws[f"C{r}"] = (f"=IFERROR(INDEX('{dt}'!$C:$C,MATCH($S{r},'{dt}'!$B:$B,0)),"
                               f'"Outside archetype model")')
                ws[f"R{r}"] = (f"=IFERROR(IF(INDEX('{dt}'!$D:$D,"
                               f'MATCH($S{r},\'{dt}\'!$B:$B,0))="","-",'
                               f"INDEX('{dt}'!$D:$D,MATCH($S{r},'{dt}'!$B:$B,0))),\"-\")")
            else:
                ws[f"C{r}"] = '="Outside archetype model"'
                ws[f"R{r}"] = '="-"'
            key = f'$C{r}&"|"&$R{r}'
            ws[f"D{r}"] = (f"=IFERROR(INDEX({A3}!$F$5:$F$23,"
                           f'MATCH({key},{A3}!$A$5:$A$23,0)),"-")')
            ws[f"L{r}"] = (f"=IFERROR(INDEX({A3}!$G$5:$G$23,"
                           f'MATCH({key},{A3}!$A$5:$A$23,0)),"-")')
            # Why a squad has no archetype matters: a shared role is correct, a
            # Strategic Programs squad is waiting on a rate 0.3 marks TBC, and a COE
            # is measured against budget. One catch-all label hid all three.
            ws[f"T{r}"] = (
                f'=IF($B{r}="Leadership","Shared role - compared to the allowance",'
                f'IF($C{r}="Strategic Programs","Strategic Programs - no rate in 0.3 (TBC)",'
                f'IF(NOT(ISNUMBER($D{r})),"No archetype for this squad","Archetype applies")))')

            ws[f"E{r}"] = f"=COUNTIFS({RL('F')},$B{r})"
            ws[f"F{r}"] = f'=IFERROR($E{r}-$D{r},"-")'
            ws[f"G{r}"] = f'=COUNTIFS({RL("F")},$B{r},{RL("D")},"Vacant")'
            ws[f"H{r}"] = (f'=COUNTIFS({RL("F")},$B{r},{RL("D")},"Vacant",{RL("E")},"Hire")'
                           f'+COUNTIFS({RL("F")},$B{r},{RL("D")},"Vacant",{RL("E")},"Offshore")')
            ws[f"I{r}"] = f"=$G{r}-$H{r}"
            ws[f"J{r}"] = f'=$E{r}-COUNTIFS({RL("F")},$B{r},{RL("E")},"Hold")'
            ws[f"K{r}"] = (f'=IF(NOT(ISNUMBER($D{r})),"Outside the archetype model",'
                           f'IF($J{r}>$D{r},"Over archetype after decisions",'
                           f'IF($J{r}<$D{r},"Under archetype","On archetype")))')
            ws[f"M{r}"] = f"=SUMIFS({RL('G')},{RL('F')},$B{r})/1000000"
            ws[f"N{r}"] = f'=IFERROR($M{r}-$L{r},"-")'
            ws[f"O{r}"] = f"=SUMIFS({RL('H')},{RL('F')},$B{r})/1000000"
            ws[f"P{r}"] = (f"=$O{r}+{ADDITIVE}*$E{r}/$E${tot}" if dt else f"=$O{r}")
            ws[f"Q{r}"] = f'=IFERROR($P{r}-$L{r},"-")'

        ws[f"B{tot}"] = "Total"
        ws[f"B{tot}"].font = BOLD
        for col in "DEFGHIJLNOPQ":
            ws[f"{col}{tot}"] = f"=SUM({col}{first}:{col}{last})"
        # E and M read the ledger directly, so a squad that fails to join shows up as a
        # gap against the sum of the rows rather than quietly shrinking the total.
        ws[f"E{tot}"] = f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},"{pf}")'
        ws[f"M{tot}"] = (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},"{pf}")/1000000')

        ws[f"B{m1}"] = "Cost to fill every vacancy at full onshore cost ($m)"
        ws[f"E{m1}"] = f'=SUMIFS({RL("G")},{RL("D")},"Vacant")/1000000'
        ws[f"B{m2}"] = "Of which currently marked Hire or Offshore ($m)"
        ws[f"E{m2}"] = (f'=(SUMIFS({RL("G")},{RL("D")},"Vacant",{RL("E")},"Hire")+Lists!$AD$5'
                        f'*SUMIFS({RL("G")},{RL("D")},"Vacant",{RL("E")},"Offshore"))/1000000')
        ws[f"B{m3}"] = ("Leadership and Head of Technology already inside this portfolio ($m) "
                        "- not re-added in column P")
        ws[f"E{m3}"] = (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},"{pf}",'
                        f'{REV}!$AQ$2:$AQ${LR},1)/1000000')
        for rr in (m1, m2, m3):
            ws[f"B{rr}"].font = ITAL

        ws.cell(rl_hdr - 1, 2).value = f"{pf} roles - one row per named role, from REVIEW"
        ws.cell(rl_hdr - 1, 2).font = ITAL
        hdr_row(ws, rl_hdr, ROLE_HDRS)
        for i, rec in enumerate(sorted(mine, key=lambda x: (order(x["squad"]),
                                                            x["vacant"], x["name"]))):
            r, src = rl0 + i, rec["row"]
            ws[f"B{r}"] = f"=INDEX({REV}!$B:$B,{src})"
            ws[f"C{r}"] = f"=INDEX({REV}!$C:$C,{src})"
            ws[f"D{r}"] = f'=IF(ISNUMBER(SEARCH("vacant",$B{r})),"Vacant","Filled")'
            ws[f"E{r}"] = rec["lever"]
            ws[f"F{r}"] = f"=INDEX({REV}!$AP:$AP,{src})"
            ws[f"G{r}"] = f"=INDEX({REV}!$AA:$AA,{src})"
            ws[f"H{r}"] = (f"=$G{r}*IFERROR(INDEX(Lists!$AD:$AD,"
                           f"MATCH($E{r},Lists!$AC:$AC,0)),1)")

        dv = DataValidation(type="list", formula1=LEVER_LIST, allow_blank=False)
        dv.errorTitle, dv.error = "Vacancy lever", "Choose Filled, Hire, Hold or Offshore."
        ws.add_data_validation(dv)
        dv.add(f"E{rl0}:E{rl1}")
        ws.cell(2, 2).value = title
        out[tab] = dict(pf=pf, first=first, last=last, tot=tot, rl0=rl0, rl1=rl1,
                        squads=squads)
    return out


def build_3x(wb, roles, geom):
    """3.2 becomes the one cost statement; 3.3 the one FTE statement; 3.1 the group view."""
    order = list(dict.fromkeys(model.TAB_PORTFOLIO.values()))
    tab_of = {v: k for k, v in model.TAB_PORTFOLIO.items()}

    # ---------------- 3.2 Total Cost ----------------
    ws = wb["3.2 Total Cost"]
    for r in range(4, ws.max_row + 1):
        for c in range(1, max(ws.max_column, 16) + 1):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).value = None
    ws["B4"] = ("One cost statement. Actual is the REVIEW ledger; archetype is the design. "
                "Overhead is compared, never added.")
    ws["B4"].font = ITAL
    H = {"B": "Portfolio", "C": "Roles", "D": "Filled", "E": "Vacant",
         "F": "Actual cost ($m)", "G": "Archetype cost ($m)",
         "H": "Variance actual vs archetype ($m)", "I": "Cost after decisions ($m)",
         "J": "Change from decisions ($m)", "K": "Roles after decisions",
         "L": "Filled cost ($m)", "M": "Vacant cost ($m)"}
    hdr_row(ws, 5, H)
    r0 = 6
    n_arch = sum(1 for p in order if tab_of[p] in DESIGN)
    for i, pf in enumerate(order):
        t = tab_of[pf]
        r = r0 + i + (1 if i >= n_arch else 0)   # leave a gap for the subtotal row
        g = geom[t]
        ws[f"B{r}"] = pf
        ws[f"C{r}"] = f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r})'
        ws[f"D{r}"] = f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r},{REV}!$AK$2:$AK${LR},"Filled")'
        ws[f"E{r}"] = f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r},{REV}!$AK$2:$AK${LR},"Vacant")'
        ws[f"F{r}"] = f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},$B{r})/1000000'
        if t in DESIGN:
            ws[f"G{r}"] = f"='{t}'!$L${g['tot']}"
            ws[f"H{r}"] = f"=$F{r}-$G{r}"
        else:
            # A COE is not built from squad archetypes. Showing 0 made its whole cost
            # read as variance; it is measured against budget on 3.1 instead.
            ws[f"G{r}"] = '="-"'
            ws[f"H{r}"] = '="-"'
            ws[f"N{r}"] = "Measured against budget, not archetype"
            ws[f"N{r}"].font = ITAL
        ws[f"I{r}"] = f"='{t}'!$O${g['tot']}"
        ws[f"J{r}"] = f"=$I{r}-$F{r}"
        ws[f"K{r}"] = f"='{t}'!$J${g['tot']}"
        ws[f"L{r}"] = (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},$B{r},'
                       f'{REV}!$AK$2:$AK${LR},"Filled")/1000000')
        ws[f"M{r}"] = (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},$B{r},'
                       f'{REV}!$AK$2:$AK${LR},"Vacant")/1000000')
    n_pf = sum(1 for pf in order if tab_of[pf] in DESIGN)
    sub = r0 + n_pf
    ws[f"B{sub}"] = "Portfolios (archetype-based)"
    ws[f"B{sub}"].font = BOLD
    for col in "CDEFGHIJKLM":
        ws[f"{col}{sub}"] = f"=SUM({col}{r0}:{col}{sub-1})"
    # shift the COE rows down one so the subtotal sits between the two groups
    tr = r0 + len(order) + 1
    ws[f"B{tr}"] = "Total"
    ws[f"B{tr}"].font = BOLD
    for col in "CDEFIJKLM":
        ws[f"{col}{tr}"] = f"=${col}{sub}+SUM({col}{sub+1}:{col}{tr-1})"
    ws[f"G{tr}"] = f"=$G{sub}"
    ws[f"H{tr}"] = f"=$H{sub}"
    ws[f"B{tr+1}"] = ("Archetype and variance cover the ten archetype-based portfolios only; "
                      "the four COEs are measured against budget.")
    ws[f"B{tr+1}"].font = ITAL
    ws[f"B{tr+2}"] = "Ledger control - roles"
    ws[f"C{tr+2}"] = f'=COUNTA({REV}!$B$2:$B${LR})-$C${tr}'
    ws[f"B{tr+3}"] = "Ledger control - cost ($m), must be 0"
    ws[f"F{tr+3}"] = (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AJ$2:$AJ${LR},"<>")/1000000'
                      f"-$F${tr}")
    for rr in (tr + 2, tr + 3):
        ws[f"B{rr}"].font = ITAL

    # shared-role class block: the recommended portrayal, allowance vs actual
    cb = tr + 5
    ws[f"B{cb}"] = "Shared roles - allowance vs actual (each role classified once in REVIEW)"
    ws[f"B{cb}"].font = BOLD
    hdr_row(ws, cb + 1, {"B": "Class", "C": "Roles", "F": "Actual ($m)",
                         "G": "Allowance ($m)", "H": "Variance ($m)"})
    for i, k in enumerate(CLASSES):
        r = cb + 2 + i
        ws[f"B{r}"] = k
        ws[f"C{r}"] = f'=COUNTIFS({REV}!$AR$2:$AR${LR},$B{r})'
        ws[f"F{r}"] = f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AR$2:$AR${LR},$B{r})/1000000'
        ws[f"G{r}"] = f"=IFERROR(INDEX(Lists!$AI:$AI,MATCH($B{r},Lists!$AF:$AF,0)),0)"
        ws[f"H{r}"] = f"=$F{r}-$G{r}"
    sr = cb + 2 + len(CLASSES)
    ws[f"B{sr}"] = "Shared roles total"
    ws[f"B{sr}"].font = BOLD
    for col in "CFGH":
        ws[f"{col}{sr}"] = f"=SUM({col}{cb+2}:{col}{sr-1})"
    ws[f"B{sr+1}"] = "Squad (delivery) roles"
    ws[f"C{sr+1}"] = f'=COUNTIFS({REV}!$AR$2:$AR${LR},"Squad")'
    ws[f"F{sr+1}"] = f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AR$2:$AR${LR},"Squad")/1000000'
    ws[f"B{sr+2}"] = "Total - ties to the ledger"
    ws[f"B{sr+2}"].font = BOLD
    ws[f"C{sr+2}"] = f"=$C{sr}+$C{sr+1}"
    ws[f"F{sr+2}"] = f"=$F{sr}+$F{sr+1}"
    g32 = dict(r0=r0, tr=tr, cb=cb, sr=sr)

    # ---------------- 3.3 FTE View ----------------
    ws = wb["3.3 FTE View"]
    for r in range(4, ws.max_row + 1):
        for c in range(1, max(ws.max_column, 16) + 1):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).value = None
    ws["B4"] = ("One FTE statement, every squad in the ledger. Counts only - no dollars in "
                "this table.")
    ws["B4"].font = ITAL
    H = {"B": "Portfolio", "C": "Squad", "D": "Archetype type", "E": "Archetype size",
         "F": "Archetype roles", "G": "Filled", "H": "Vacant", "I": "Total roles",
         "J": "Variance to archetype", "K": "Roles after decisions", "L": "Vacancy %"}
    hdr_row(ws, 5, H)
    r = 6
    starts = {}
    for pf in order:
        t = tab_of[pf]
        g = geom[t]
        starts[pf] = r
        for i, sq in enumerate(g["squads"]):
            sr_ = g["first"] + i
            ws[f"B{r}"] = pf
            ws[f"C{r}"] = f"='{t}'!$B${sr_}"
            ws[f"D{r}"] = f"='{t}'!$C${sr_}"
            ws[f"E{r}"] = f"='{t}'!$R${sr_}"
            ws[f"F{r}"] = f"='{t}'!$D${sr_}"
            ws[f"G{r}"] = (f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r},{REV}!$AP$2:$AP${LR},$C{r},'
                           f'{REV}!$AK$2:$AK${LR},"Filled")')
            ws[f"H{r}"] = (f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r},{REV}!$AP$2:$AP${LR},$C{r},'
                           f'{REV}!$AK$2:$AK${LR},"Vacant")')
            ws[f"I{r}"] = f"=$G{r}+$H{r}"
            ws[f"J{r}"] = f'=IFERROR($I{r}-$F{r},"-")'
            ws[f"K{r}"] = f"='{t}'!$J${sr_}"
            ws[f"L{r}"] = f'=IFERROR($H{r}/$I{r},"-")'
            r += 1
        ws[f"B{r}"] = pf
        ws[f"C{r}"] = "Portfolio total"
        ws[f"C{r}"].font = BOLD
        for col in "GHIJK":
            ws[f"{col}{r}"] = f"=SUM({col}{starts[pf]}:{col}{r-1})"
        ws[f"F{r}"] = (f'=IF(COUNT(F{starts[pf]}:F{r-1})=0,"-",'
                       f"SUM(F{starts[pf]}:F{r-1}))")
        ws[f"J{r}"] = (f'=IF(COUNT(F{starts[pf]}:F{r-1})=0,"-",'
                       f"SUM(J{starts[pf]}:J{r-1}))")
        ws[f"I{r}"] = f'=COUNTIFS({REV}!$AJ$2:$AJ${LR},$B{r})'
        ws[f"L{r}"] = f'=IFERROR($H{r}/$I{r},"-")'
        starts[pf + "_tot"] = r
        r += 1
    gt = r
    ws[f"C{gt}"] = "Group total"
    ws[f"C{gt}"].font = BOLD
    for col in "GHIJK":
        ws[f"{col}{gt}"] = "=" + "+".join(f"${col}${starts[p+'_tot']}" for p in order)
    # F and J carry "-" on the COE total rows, so the group total sums the squad rows
    # directly (SUM ignores text) rather than adding the portfolio totals together.
    for col in ("F", "J"):
        ws[f"{col}{gt}"] = (f'=SUMIF($C$6:$C${gt-1},"<>Portfolio total",'
                            f"{col}6:{col}{gt-1})")
    ws[f"I{gt}"] = f"=COUNTA({REV}!$B$2:$B${LR})"
    ws[f"L{gt}"] = f"=IFERROR($H{gt}/$I{gt},\"-\")"
    ws[f"B{gt+2}"] = "Ledger control - roles, must be 0"
    ws[f"B{gt+2}"].font = ITAL
    ws[f"I{gt+2}"] = f"=COUNTA({REV}!$B$2:$B${LR})-$I${gt}"
    g33 = dict(gt=gt, starts=starts)

    # ---------------- 3.1 Group Summary ----------------
    ws = wb["3.1 Group Summary"]
    for r in range(4, ws.max_row + 1):
        for c in range(1, max(ws.max_column, 14) + 1):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).value = None
    ws["B4"] = "Group view. Every figure is pulled from 3.2 or 3.3 - nothing is restated here."
    ws["B4"].font = ITAL
    H = {"B": "Portfolio", "C": "Budget ($m)", "D": "Actual cost ($m)",
         "E": "Variance to budget ($m)", "F": "Archetype cost ($m)",
         "G": "Variance to archetype ($m)", "H": "Roles", "I": "Vacant",
         "J": "Cost after decisions ($m)"}
    hdr_row(ws, 5, H)
    n_arch = sum(1 for p in order if tab_of[p] in DESIGN)
    for i, pf in enumerate(order):
        r = 6 + i
        s = g32["r0"] + i + (1 if i >= n_arch else 0)
        ws[f"B{r}"] = f"='3.2 Total Cost'!$B${s}"
        # budget lines are matched through the map on Lists: Customer, Cyber, BP&T and
        # SA&D each draw on more than one line, so a direct name match returned zero
        ws[f"C{r}"] = (f"=SUMPRODUCT(('0.2 Data Config'!$B$6:$B$25<>\"\")*"
                       f"(IFERROR(INDEX(Lists!$AL$2:$AL$19,MATCH('0.2 Data Config'!$B$6:$B$25,"
                       f"Lists!$AK$2:$AK$19,0)),\"\")=$B{r})*'0.2 Data Config'!$E$6:$E$25)")
        ws[f"D{r}"] = f"='3.2 Total Cost'!$F${s}"
        ws[f"E{r}"] = f"=$C{r}-$D{r}"
        ws[f"F{r}"] = f"='3.2 Total Cost'!$G${s}"
        ws[f"G{r}"] = f"='3.2 Total Cost'!$H${s}"
        ws[f"H{r}"] = f"='3.2 Total Cost'!$C${s}"
        ws[f"I{r}"] = f"='3.2 Total Cost'!$E${s}"
        ws[f"J{r}"] = f"='3.2 Total Cost'!$I${s}"
    tr1 = 6 + len(order)
    ws[f"B{tr1}"] = "Total"
    ws[f"B{tr1}"].font = BOLD
    for col in "CDEFGHIJ":
        ws[f"{col}{tr1}"] = f"=SUM({col}6:{col}{tr1-1})"
    return g32, g33


def run(src, dst):
    roles = model.load(src)
    wb = openpyxl.load_workbook(src)
    geom = build_2x(wb, roles)
    build_3x(wb, roles, geom)
    exec_summary(wb)
    wb.save(dst)
    return geom




def exec_summary(wb):
    """Re-point the Exec Summary at the rebuilt 3.2 / 3.3 anchors.

    Every headline was reading a cell that no longer holds what its label claims, so the
    whole story block returned 0 - a silent wrong answer, not an error. Labels that
    described the old double-counted overhead treatment are restated to match the model.
    """
    ws = wb["Exec Summary"]
    G32, G33, G31 = "'3.2 Total Cost'", "'3.3 FTE View'", "'3.1 Group Summary'"
    TOT, ARCH = 21, 16          # 3.2 grand total row, archetype-based subtotal row
    GT = 95                     # 3.3 group total row
    DESIGNED = f"({G32}!$G${ARCH}+Lists!$AI$6)"   # squad archetypes + shared allowance

    txt = {
        12: ("Shared roles - leadership, Head of Technology, business partners and "
             "architects - are classified once in REVIEW and compared to the allowance "
             "on Lists. They are never added on top of a portfolio's cost."),
        13: ("Strategic Programs squads (AmPOS, CTRM, P&C, the EGI squads) have no rate "
             "in 0.3 - it is marked TBC - so they carry no archetype comparison."),
        15: ("Business partners and architects sit in the COEs, so a portfolio's column P "
             "adds only its share of those two, never its own leadership."),
    }
    for r, t in txt.items():
        ws.cell(r, 2).value = t

    f = {
        5:  f"={G33}!$F${GT}",
        6:  f"={G33}!$I${GT}",
        7:  f"={G32}!$F${TOT}",
        8:  f"={G32}!$H${ARCH}",
        9:  f"={G32}!$I${TOT}",
        20: f"={G31}!$C$20",
        21: f"='0.2 Data Config'!$E$27-{G31}!$C$20",
        23: f"={G32}!$G${ARCH}",
        24: "=Lists!$AI$6",
        25: f"={DESIGNED}",
        26: f"={G32}!$F${TOT}",
        27: f"={G32}!$F${TOT}-{DESIGNED}",
        28: f"={G31}!$C$20-{G32}!$F${TOT}",
        30: f"={G32}!$F${TOT}",
        31: f"={G32}!$L${TOT}",
        32: f"={G32}!$M${TOT}",
        33: f"={G32}!$F${TOT}-{DESIGNED}",
        36: f"={G33}!$F${GT}",
        37: f"={G32}!$C${ARCH}",
        38: f"={G32}!$C${ARCH}-{G33}!$F${GT}",
        39: f"={G32}!$C${TOT}-{G32}!$C${ARCH}",
        40: f"={G33}!$I${GT}",
        41: f"={G33}!$G${GT}",
        42: f"={G33}!$H${GT}",
        45: f"={G33}!$L${GT}",
        49: f"={G32}!$L${TOT}",
        50: f"={G32}!$L${TOT}-{G32}!$G${ARCH}",
        51: f"={G32}!$M${TOT}",
        52: (f'=SUMIFS({REV}!$AA$2:$AA${LR},{REV}!$AR$2:$AR${LR},"Squad",'
             f'{REV}!$AK$2:$AK${LR},"Vacant")/1000000'),
        54: f"={G32}!$H${ARCH}",
        57: f"=COUNTIFS({REV}!$AP$2:$AP${LR},\"Unassigned\")",
    }
    for r, v in f.items():
        ws.cell(r, 3).value = v

    lab = {
        23: "Squad archetype cost - the ten archetype-based portfolios ($m)",
        24: "Shared role allowance - leadership, Head of Tech, BP, architecture ($m)",
        25: "Total designed cost ($m)",
        26: "Actual cost of the organisation today ($m)",
        27: "Actual over the designed cost by ($m)",
        28: "Budget allocated less actual cost ($m)",
        33: "Difference vs the designed cost ($m)",
        37: "Roles actually raised in the archetype-based portfolios",
        38: "Roles raised beyond the archetypes",
        39: "Roles in the COEs and EGI - measured against budget, not archetypes",
        50: "Filled roles over/(under) the squad archetype cost ($m)",
        52: "of which squad roles - the 2.x working tab lever ($m)",
        57: "Roles with no squad in the ledger",
    }
    for r, t in lab.items():
        ws.cell(r, 2).value = t
    return ws


if __name__ == "__main__":
    g = run("core.xlsx", "flow.xlsx")
    print(f"built {len(g)} working-copy tabs + 3.1/3.2/3.3")
