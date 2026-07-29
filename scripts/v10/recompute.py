"""Recompute every figure a reader sees, from REVIEW, and compare it to the tab.

The other passes test the workbook against itself: a control row, a cross-tab agreement, a
formula error. All of those can pass while the whole model is consistently wrong, and twice
in this build they did - a subtotal that summed row variances and dropped every row carrying
a dash, and a subtotal with no archetype that reported the whole cost as a variance. Both
added up. Both were wrong.

This pass builds the numbers again from the ledger, in Python, and asserts the tab matches.
It knows nothing about the workbook's formulas.
"""
import collections
import json
import sys

import openpyxl

REVIEW = "REVIEW - Complete Role Mapping"
# the reading order the builders use, not a list of what exists: a name with no ledger
# rows and no working tab yet contributes nothing here and is skipped in silence
PF = ["Ampol Retail", "Customer", "Enterprise Data", "TDD Group Functions", "P&C",
      "Finance", "Infrastructure", "Energy Solutions & B2B", "Commercial Fuels",
      "Z Retail", "TDD Cyber"]
COE = ["COE Cyber", "COE BP&T", "COE SA&D", "EGI"]
TOL = 1e-6


def last(wv):
    R = wv[REVIEW]
    r = R.max_row
    while r > 1 and not str(R.cell(r, 2).value or '').strip():
        r -= 1
    return r


def ledger(wv):
    """Every figure 3.1, 3.2, 3.3 and 3.4 state, rebuilt from the ledger.

    line_n / line_c are the organisation's count and cost per overhead line - every role
    carrying the line, wherever the person sits - which is what 3.2's roles and cost
    columns state. line_pfn / line_pfc narrow the same two to the people who sit in a
    portfolio, which is what its "Where they sit" column and its first band state.
    """
    R = wv[REVIEW]
    cost, roles, filled, vacant, ohcost, ohroles = (collections.Counter() for _ in range(6))
    line_n, line_c, line_pfn, line_pfc = (collections.Counter() for _ in range(4))
    for i in range(2, last(wv) + 1):
        if not str(R.cell(i, 2).value or "").strip():
            continue
        pf, grp = str(R.cell(i, 36).value), str(R.cell(i, 46).value)
        c = R.cell(i, 27).value or 0
        cost[(pf, grp)] += c
        roles[(pf, grp)] += 1
        if str(R.cell(i, 37).value) == "Filled":
            filled[(pf, grp)] += 1
        else:
            vacant[(pf, grp)] += 1
        if str(R.cell(i, 44).value) != "Squad":
            # AR == AT selects the roles that sit in a portfolio; a COE role carrying an
            # overhead title has AT set to its COE squad
            line = str(R.cell(i, 44).value)
            key = "pf" if line == grp else "coe"
            ohcost[key] += c
            ohroles[key] += 1
            line_n[line] += 1
            line_c[line] += c
            if key == "pf":
                line_pfn[line] += 1
                line_pfc[line] += c
    return (cost, roles, filled, vacant, ohcost, ohroles,
            line_n, line_c, line_pfn, line_pfc)


def find(ws, label, col=2, figure=None):
    """Exact match first, then a prefix, so a label that gains a clause still resolves.

    `figure` is the column a wanted row must carry a number in. 3.x tabs put a pale
    section label and the grey subtotal under it on rows that start with the same words -
    "COEs and EGI ..." over "COEs and EGI" - and the prefix pass reads top down, so it
    finds the label, which carries no figures, and every check against that step then
    compares a real total to a blank. Passing the figure column makes the search skip
    rows that cannot be the one being asked for.
    """
    for exact in (True, False):
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, col).value
            if not isinstance(v, str):
                continue
            if not ((v.strip() == label) if exact else v.strip().startswith(label)):
                continue
            if figure is not None and not isinstance(ws.cell(r, figure).value,
                                                     (int, float)):
                continue
            return r
    return None


def col_of(ws, header, limit=12):
    for r in range(1, limit + 1):
        for c in range(2, 24):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == header:
                return c
    return None


def check(out, name, got, exp, count=False):
    if got is None or exp is None:
        out.append(f"{name}: tab {got!r}, recomputed {exp!r}")
        return False
    ok = (got == exp) if count else abs(got - exp) < TOL
    if not ok:
        out.append(f"{name}: tab {got!r}, recomputed {exp!r}")
    return ok


def run(path, anchors="anchors_final.json"):
    wv = openpyxl.load_workbook(path, data_only=True)
    a = json.load(open(anchors))
    (cost, roles, filled, vacant, ohcost, ohroles,
     line_n, line_c, line_pfn, line_pfc) = ledger(wv)
    # tabs are renamed after the anchors are written, so map by the portfolio cell
    tab_of = {str(wv[t]["C3"].value): t for t in wv.sheetnames if t.startswith("2.")}
    out = []

    # ---- every squad row on every working tab ----
    S = a[list(a)[0]]["cols"]
    for v in a.values():
        # a portfolio named on the anchors whose working tab is not in this workbook has
        # nothing to recompute against, so it is skipped rather than stopping the pass
        t = tab_of.get(v["pf"])
        if t is None:
            continue
        ws = wv[t]
        for g in v["squads"] + v["direct"] + v["nofig"] + v["overhead"]:
            r, k = v["srow"].get(g), (v["pf"], g)
            if not r:
                continue
            for key, exp, cnt in (("actual", cost[k] / 1e6, False),
                                  ("roles", roles[k], True),
                                  ("filled", filled[k], True),
                                  ("vacant", vacant[k], True)):
                check(out, f"{t} {g} {key}", ws.cell(r, S[key]).value, exp, cnt)

    # ---- the bridge on 3.1 ----
    b = next(t for t in wv.sheetnames if t.startswith("3.1 "))
    ws = wv[b]
    ca, cc, cr = (col_of(ws, "Actual cost ($m)"), col_of(ws, "Archetype cost ($m)"),
                  col_of(ws, "Total roles"))
    arch = sum(cost[(v["pf"], g)] for v in a.values() for g in v["squads"])
    narch = sum(roles[(v["pf"], g)] for v in a.values() for g in v["squads"])
    direct = sum(cost[(v["pf"], g)] for v in a.values() if v["pf"] in PF
                 for g in v["direct"])
    ndir = sum(roles[(v["pf"], g)] for v in a.values() if v["pf"] in PF
               for g in v["direct"])
    nofig = sum(cost[(v["pf"], g)] for v in a.values() if v["pf"] in PF
                for g in v["nofig"])
    nnof = sum(roles[(v["pf"], g)] for v in a.values() if v["pf"] in PF
               for g in v["nofig"])
    coe = sum(cost[(v["pf"], g)] for v in a.values() if v["pf"] in COE
              for g in v["direct"] + v["squads"] + v["nofig"] + v["overhead"])
    ncoe = sum(roles[(v["pf"], g)] for v in a.values() if v["pf"] in COE
               for g in v["direct"] + v["squads"] + v["nofig"] + v["overhead"])
    # the directly funded step is two subtotals, split by whether a funded figure is set,
    # so its actual cost is the pair added
    DIRECT = ("Directly funded, where the funded figure is set",
              "Directly funded, where no funded figure is set yet")
    for labs, c, n in ((("Squads priced by an archetype",), arch, narch),
                       (DIRECT, direct, ndir),
                       (("COEs and EGI",), coe, ncoe),
                       (("Groups with no archetype and no funded figure",), nofig, nnof),
                       (("Overhead roles in the portfolios",), ohcost["pf"],
                        ohroles["pf"])):
        rs = [find(ws, x, figure=ca) for x in labs]
        if any(x is None for x in rs):
            out.append(f"3.1 has no row {labs!r}")
            continue
        name = labs[0]
        check(out, f"3.1 {name} cost",
              round(sum(ws.cell(x, ca).value or 0 for x in rs), 9), round(c / 1e6, 9))
        check(out, f"3.1 {name} roles", sum(ws.cell(x, cr).value or 0 for x in rs), n, True)
    total = sum(cost.values())
    r = find(ws, "Cost of the")
    check(out, "3.1 ledger cost", ws.cell(r, ca).value, total / 1e6)
    check(out, "3.1 ledger roles", ws.cell(r, cr).value, sum(roles.values()), True)
    # The archetype side is a design figure, so it is checked for internal consistency:
    # the steps that carry one must add to the comparable subtotal. It used to be checked
    # against the ledger line, which only worked while every step had an archetype - and
    # three of them only had one because the column was restating the actual.
    steps = [find(ws, x) for x in
             ("Squads priced by an archetype",
              "Directly funded, where the funded figure is set",
              "Overhead roles in the portfolios - the allowance is on 3.2")]
    cmp_row = find(ws, "Everything with a figure to compare")
    if None in steps or cmp_row is None:
        out.append("3.1 is missing one of the comparable steps")
    else:
        for col, what in ((cc, "archetype"), (ca, "actual")):
            check(out, f"3.1 comparable steps add up - {what}",
                  round(sum(v for v in (ws.cell(x, col).value for x in steps)
                            if isinstance(v, (int, float))), 6),
                  round(ws.cell(cmp_row, col).value, 6))
    # the ledger and grand rows carry a DASH in the archetype and variance columns since
    # wave K (D115): their actual side covers all 531 roles and the archetype prices only
    # the comparable set, so a figure there mixed bases and read as "TDD is $40.7m over".
    # What is checked now is that the dash is a dash - a number reappearing here is the
    # defect coming back.
    k = find(ws, "Cost of the")
    check(out, "3.1 ledger row archetype is a dash, not a mixed-basis figure",
          str(ws.cell(k, cc).value), "-", count=True)
    check(out, "3.1 ledger row variance is a dash",
          str(ws.cell(k, 6).value), "-", count=True)
    g2 = find(ws, "Total cost of TDD including the GM layer")
    check(out, "3.1 grand row archetype is a dash", str(ws.cell(g2, cc).value), "-",
          count=True)
    check(out, "3.1 grand row variance is a dash", str(ws.cell(g2, 6).value), "-",
          count=True)
    gm = wv["Lists"]["AG12"].value
    g = find(ws, "Total cost of TDD including the GM layer")
    check(out, "3.1 grand total", ws.cell(g, ca).value, total / 1e6 + gm)

    # ---- 3.2 overhead: applied to the portfolios against the organisation ----
    # One row per line. The roles and the cost are the organisation's, so they are rebuilt
    # from every ledger role carrying the line and not from the portfolio half of it, and
    # the two gaps are rebuilt as the difference against what the allowance applies.
    o = next(t for t in wv.sheetnames if t.startswith("3.2 "))
    ws = wv[o]
    # 3.2 was rebuilt to the owner's own layout and every one of these headings changed
    # with it. The six names below were the old ones, so col_of returned None for all six,
    # so this whole block short-circuited to one line of output and stopped checking
    # anything - every 3.2 figure has been unverified since that rebuild, and the line it
    # printed read like a note about the tab rather than a dead check. His headings now,
    # with what each one is in the old vocabulary beside it.
    cE = col_of(ws, "Roles priced for in archetype")            # applied to portfolios
    cF = col_of(ws, "Total Archetype cost ($m)")                 # applied, in money
    cG = col_of(ws, "Actual number of leadership roles")         # the organisation's roles
    cH = col_of(ws, "Actual cost of leadership roles")           # the organisation's cost
    cI = col_of(ws, "# of roles not applied in archetype")       # the roles gap
    cJ = col_of(ws, "Variance between archetype and actuals")    # the cost gap
    if None in (cE, cF, cG, cH, cI, cJ):
        missing = [n for n, c in (("roles priced for", cE), ("archetype cost", cF),
                                  ("actual roles", cG), ("actual cost", cH),
                                  ("roles gap", cI), ("cost gap", cJ)) if c is None]
        out.append(f"3.2 is missing columns: {', '.join(missing)} - nothing on the tab "
                   f"was checked")
    else:
        gm_n, gm_c = wv["Lists"]["AG11"].value, wv["Lists"]["AG12"].value
        for line in ("Head of Technology", "Business Partner", "Domain Architect",
                     "Delivery Manager", "Technology Manager", "Leadership - 8 GMs"):
            r = find(ws, line)
            if r is None:
                out.append(f"3.2 has no row {line!r}")
                continue
            n = gm_n if line.startswith("Leadership") else line_n[line]
            c = gm_c if line.startswith("Leadership") else line_c[line] / 1e6
            check(out, f"3.2 {line} roles in the organisation",
                  ws.cell(r, cG).value, n, True)
            check(out, f"3.2 {line} cost in the organisation", ws.cell(r, cH).value, c)
            check(out, f"3.2 {line} roles gap", round(ws.cell(r, cI).value or 0, 6),
                  round(n - (ws.cell(r, cE).value or 0), 6))
            check(out, f"3.2 {line} cost gap", round(ws.cell(r, cJ).value or 0, 6),
                  round(c - (ws.cell(r, cF).value or 0), 6))
        t = find(ws, "Overheads incl. GMs")
        check(out, "3.2 overhead roles in the organisation", ws.cell(t, cG).value,
              sum(line_n.values()) + gm_n, True)
        check(out, "3.2 overhead cost in the organisation", ws.cell(t, cH).value,
              sum(line_c.values()) / 1e6 + gm_c)
        r = find(ws, "Of which sits in the portfolios")
        check(out, "3.2 overhead roles sitting in the portfolios",
              ws.cell(r, cG).value, ohroles["pf"], True)
        check(out, "3.2 overhead cost sitting in the portfolios",
              ws.cell(r, cH).value, ohcost["pf"] / 1e6)
        b = find(ws, "Of which sits in the COEs and EGI")
        check(out, "3.2 overhead roles sitting outside the portfolios",
              ws.cell(b, cG).value, ohroles["coe"] + gm_n, True)
        check(out, "3.2 overhead cost sitting outside the portfolios",
              ws.cell(b, cH).value, ohcost["coe"] / 1e6 + gm_c)
        # the one row that states the ledger itself, and the sentence that reads it out
        allrow = find(ws, "Roles in the organisation, all lines and squads")
        if allrow is None:
            out.append("3.2 has no all-roles row")
        else:
            n_pf = sum(v for (p, _), v in roles.items() if p in PF)
            n_coe = sum(v for (p, _), v in roles.items() if p in COE)
            check(out, "3.2 all roles", ws.cell(allrow, cG).value,
                  sum(roles.values()), True)
            want = (f"Roles in the organisation, all lines and squads: "
                    f"{sum(roles.values())} - portfolios {n_pf}, "
                    f"COEs and EGI {n_coe}, each counted once")
            got = str(ws.cell(allrow, 2).value or "")
            if got != want:
                out.append(f"3.2 all-roles line: tab {got!r}, recomputed {want!r}")

    # ---- 3.3 group total and 3.4 COE total ----
    d = next(t for t in wv.sheetnames if t.startswith("3.3 "))
    ws = wv[d]
    r = find(ws, "Group total")
    check(out, "3.3 group cost", ws.cell(r, col_of(ws, "Actual cost ($m)")).value,
          total / 1e6)
    check(out, "3.3 group roles", ws.cell(r, col_of(ws, "Total roles")).value,
          sum(roles.values()), True)
    c4 = next(t for t in wv.sheetnames if t.startswith("3.4 "))
    ws = wv[c4]
    r = find(ws, "COEs and EGI total")
    check(out, "3.4 COE cost", ws.cell(r, col_of(ws, "Cost ($m)")).value, coe / 1e6)

    # ---- Exec ----
    ws = wv["Exec Summary"]
    filled_cost = 0.0
    R = wv[REVIEW]
    for i in range(2, last(wv) + 1):
        if str(R.cell(i, 2).value or "").strip() and str(R.cell(i, 37).value) == "Filled":
            filled_cost += R.cell(i, 27).value or 0
    for lab, exp, cnt in (("Roles in the role mapping", sum(roles.values()), True),
                          ("Cost of the", total / 1e6,
                           False),
                          ("Of which filled roles ($m)", filled_cost / 1e6, False),
                          ("Cost today including the GM layer ($m)", total / 1e6 + gm,
                           False)):
        r = find(ws, lab)
        if r is None:
            out.append(f"Exec has no row {lab!r}")
            continue
        check(out, f"Exec {lab}", ws.cell(r, 3).value, exp, cnt)
    return out


if __name__ == "__main__":
    bad = run(sys.argv[1] if len(sys.argv) > 1 else "cand.xlsx")
    print("=" * 74)
    print(f"FIGURES DISAGREEING WITH A RECOMPUTATION FROM THE LEDGER: {len(bad)}")
    for x in bad[:40]:
        print("   ", x)
