"""The new chain input: the owner's review workbook, plus his 2707-only typed inputs.

rev.xlsx descends from the same generation as base_ship.xlsx - its only extra sheet is a
Claude Log - so his 0.2 / 0.3 / 1.x / Lists / REVIEW changes transplant by using the whole
file as the base rather than by cell surgery. The chain rebuilds 2.x / 3.x / 4.0 / Exec
from REVIEW and the 1.x tabs, so the old-generation content on those tabs is overwritten
the same way it always is.

His 2707 copy carries two typed inputs on tabs this base now takes from rev. They are
re-applied here unless rev changed the same cell itself - in that conflict the review
workbook wins, because applying its changes to 2707 is the instruction.

Everything else he changed on 2707 - labels, the deleted 3.1 bar, the footer shape, the
Actuals box, the two levers - is adopted inside the builders or re-applied after the
build, and is documented in docs/ORCHESTRATION_2707.md.
"""
import openpyxl

# (tab, cell, value_2707, value_in_the_common_ancestor)
TYPED_2707 = [
    ("1.2 Customer", "I54", 2.21, '=IFERROR($H54*$G54,"")'),
    ("1.8 Energy Solutions & B2B", "E12", 7.2, "=SUM(I13:I20)"),
]


def run(src, dst, ancestor="base_ship.xlsx"):
    wb = openpyxl.load_workbook(src)
    anc = openpyxl.load_workbook(ancestor)
    out = []
    if "Claude Log" in wb.sheetnames:
        del wb["Claude Log"]
        out.append("Claude Log dropped - a working log, not a model tab")
    for tab, cell, val, anc_val in TYPED_2707:
        ws = wb[tab]
        cur = ws[cell].value
        base = anc[tab][cell].value if tab in anc.sheetnames else None
        if cur != base and cur != anc_val:
            out.append(f"{tab}!{cell}: rev changed it to {cur!r}, keeping rev over the "
                       f"2707 value {val!r}")
            continue
        ws[cell] = val
        out.append(f"{tab}!{cell} = {val!r} (his 2707 input carried onto the new base)")
    wb.save(dst)
    return out


if __name__ == "__main__":
    import sys
    for x in run(sys.argv[1], sys.argv[2]):
        print("  ", x)
