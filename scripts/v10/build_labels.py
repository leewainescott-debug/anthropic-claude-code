"""Stage 14: readability. Headers that say what the number is.

"Total to fund" was the header over a block containing two variances and their sum. The
total IS the amount to fund; the header was naming the block after one of its own rows.
The block is a funding position, so it now reads:

    Funding position ($m)
      Over/(under) TDD budget          actual less TDD budget, + = over
      Still to fund outside TDD        the part no TDD budget covers
      Total to fund                    the two added

Same three rows, same formulas, same order on all ten portfolio tabs. The block still
sits at a different row on some tabs because 1.4's platform table starts at row 19 and
moving it would collide; the labels and formulas are identical, which is what a reader
follows.
"""
import re
import openpyxl
from openpyxl.styles import Font
import build_1x

BOLD = Font(bold=True)

def run(src, dst):
    wb = openpyxl.load_workbook(src); out = []
    for tab, (tv, ov, tot, ltf) in build_1x.VAR.items():
        ws = wb[tab]
        # the block header sits directly above the first variance row
        hdr = tv - 1
        if str(ws.cell(hdr, 2).value or "").strip() == "":
            hdr = tv - 2
        ws.cell(hdr, 2).value = "Funding position ($m)"
        ws.cell(hdr, 2).font = BOLD
        ws[f"B{tv}"] = "Over/(under) TDD budget"
        ws[f"B{ov}"] = "Still to fund outside TDD"
        ws[f"B{tot}"] = "Total to fund"
        ws[f"B{tot}"].font = BOLD
        out.append(f"{tab}: funding position block relabelled at rows {hdr}-{tot}")

    # the same three quantities, named the same way, on the summaries
    s1 = wb["3.1 Group Summary"]
    s1["E5"] = "Over/(under) budget ($m)"
    s1["H5"] = "Over/(under) archetype ($m)"
    g = wb["3.2 Total Cost"]
    g["K5"] = "Over/(under) archetype ($m)"
    for r in range(27, 45):
        if str(g.cell(r, 2).value or "").startswith("Overhead line"):
            g.cell(r, 8).value = "Over/(under) allowance ($m)"
    s4 = wb["3.4 COE Summary"]
    s4["H5"] = "Over/(under) budget ($m)"
    dc = wb["0.2 Data Config"]
    dc["G5"] = "Over/(under) budget ($m)"
    out.append("3.1 / 3.2 / 3.4 / 0.2: every variance column now reads "
               "'Over/(under) X', one direction, one wording")

    # the empty overhead block on the two tabs that have no overhead roles
    for tab in ("2.11 TDD Cyber", "2.14 EGI"):
        ws = wb[tab]
        for r in range(6, 20):
            if str(ws.cell(r, 2).value or "") == "Overhead roles":
                for c in (5, 7, 8, 9, 10, 13, 15, 16, 17):
                    if isinstance(ws.cell(r, c).value, (int, float)):
                        ws.cell(r, c).value = "=0"
                ws.cell(r, 3).value = "No overhead roles sit in this portfolio"
    out.append("2.11 / 2.14: the empty overhead subtotal reads as a formula, not a typed 0")
    wb.save(dst); return out

if __name__ == "__main__":
    for x in run("f1.xlsx", "f2.xlsx"): print("  ", x)
