#!/usr/bin/env python3
"""Pass 2: finish the old-ledger retirement on clean_v8.xlsx.
- 1.11 / 1.12 / 1.13 rosters re-pointed from Sheet2 to REVIEW rows (same column
  layout: Name / Title / Department / Country / Status / Cost if hired / On-Off / T cost).
- 2.12 / 2.13 rosters re-pointed to the same REVIEW rows (statuses via AK).
- 1.x strategic-program notes and 3.2's restatement row read REVIEW.
- 3.3 leftover Squads/Sheet2 cells cleared (recon block I/J, old EGI cost col).
- 4.0 Data QA: every remaining cell referencing Squads / Added data / Sheet2 is
  cleared; the tab keeps the REVIEW-based headline and coverage checks.
3.5 Source Reconciliation deliberately keeps its Squads references - it exists to
compare REVIEW against the old sheet."""
import openpyxl
from openpyxl.utils import column_index_from_string as ci, get_column_letter as gl
from openpyxl.styles import Font, Alignment
from collections import Counter

P="clean_v8.xlsx"
wb=openpyxl.load_workbook(P)
REVN="REVIEW - Complete Role Mapping"; Q="'"+REVN+"'"
rev=wb[REVN]
def rc(r,cc): return rev.cell(r,ci(cc)).value
RLO,RHI=2,530
IT=Font(italic=True,size=10)
CEN=Alignment(horizontal="center",vertical="center"); RT=Alignment(horizontal="right")
DOL='#,##0'

def review_rows(portfolio):
    return [r for r in range(RLO,RHI)
            if str(rc(r,'I') or "").strip()==portfolio]

BPT=review_rows("COE - Partnering & Transformation")        # 24
SAD=review_rows("COE - Strategy, Architecture, Data")       # 23
CYB=review_rows("COE - Cyber, Risk & Operations")           # 46
print("REVIEW rows: BP&T",len(BPT)," SA&D",len(SAD)," Cyber",len(CYB))

def write_role_row(ws,r,dr,oncol=None,tpause=False):
    """one roster row on a 1.1x tab pointing at REVIEW row dr"""
    ws.cell(r,2).value=f"={Q}!$B${dr}"
    ws.cell(r,3).value=f"={Q}!$C${dr}"
    ws.cell(r,4).value=f"={Q}!$G${dr}"
    ws.cell(r,5).value=f"={Q}!$M${dr}"
    ws.cell(r,6).value=f"={Q}!$AK${dr}"
    vac=str(rc(dr,'B') or "").strip().lower().find("vacant")>=0
    ws.cell(r,7).value=(f"={Q}!$AA${dr}" if vac else None)
    if ws.cell(r,7).value: ws.cell(r,7).number_format=DOL
    base=f"{Q}!$AA${dr}/1000000"
    if tpause: base=f'IF($F{r}="Paused",0,{base})'
    if oncol:
        ws.cell(r,20).value=f'=({base})*IF(${oncol}{r}="Offshore",0.4,1)'
        if ws.cell(r,ci(oncol)).value in (None,""): ws.cell(r,ci(oncol)).value="Onshore"
    else:
        ws.cell(r,20).value=f"=({base})"

def clear_role_row(ws,r,cols):
    for col in cols: ws.cell(r,col).value=None

# ---------- 1.11 BP&T: rows 21-44 = 24 REVIEW rows ----------
w=wb["1.11 BP&T"]
for i,dr in enumerate(BPT):
    write_role_row(w,21+i,dr,oncol="H")
# ---------- 1.12 SA&D: rows 22-50 (29 slots) -> 23 rows ----------
w=wb["1.12 SA&D"]
for i,dr in enumerate(SAD):
    write_role_row(w,22+i,dr,oncol="H",tpause=True)
for r in range(22+len(SAD),51):
    clear_role_row(w,r,[2,3,4,5,6,7,8,20])
# ---------- 1.13 Cyber Roles: rows 19-70 (52 slots) -> 46 rows ----------
w=wb["1.13 Cyber Roles"]
for i,dr in enumerate(CYB):
    write_role_row(w,19+i,dr)
for r in range(19+len(CYB),71):
    clear_role_row(w,r,[2,3,4,5,6,7,20])

# ---------- 2.12 / 2.13: same REVIEW rows, working-tab shape ----------
def redo_2x(tab,rows,r1,slots):
    ws=wb[tab]
    for i,dr in enumerate(rows):
        r=r1+i
        ws.cell(r,2).value=f"={Q}!$B${dr}"
        ws.cell(r,3).value=f"={Q}!$C${dr}"
        ws.cell(r,4).value=f"={Q}!$G${dr}"
        ws.cell(r,5).value=f"={Q}!$AK${dr}"
        vac=str(rc(dr,'B') or "").strip().lower().find("vacant")>=0
        if vac:
            if ws.cell(r,6).value in (None,""): ws.cell(r,6).value="Hold"
            ws.cell(r,7).value=f"={Q}!$AA${dr}"; ws.cell(r,7).number_format=DOL
        else:
            if ws.cell(r,6).value in (None,""): ws.cell(r,6).value="Filled"
            ws.cell(r,7).value=None
    for r in range(r1+len(rows),r1+slots):
        for col in (2,3,4,5,6,7): ws.cell(r,col).value=None
redo_2x("2.12 BP&T",BPT,14,24)
# find 2.13's roster start: header row with B='Name'
w13=wb["2.13 SA&D"]
hdr=None
for r in range(1,40):
    if w13.cell(r,2).value=="Name": hdr=r; break
# count existing roster slots (rows with a value in B below header)
slots=0
r=hdr+1
while w13.cell(r,2).value is not None: slots+=1; r+=1
redo_2x("2.13 SA&D",SAD,hdr+1,max(slots,len(SAD)))
print("2.13 roster header row:",hdr,"old slots:",slots)

# ---------- 1.x strategic-program notes -> REVIEW ----------
for t,cellr in [("1.1 Ampol Retail","N47"),("1.1 Ampol Retail","N65"),("1.2 Customer","N53"),
                ("1.4 TDD Group Functions","N33"),("1.5 P&C","N31"),("1.6 Finance","N32")]:
    c=wb[t][cellr]; v=c.value
    if isinstance(v,str) and "Added data" in v:
        c.value=(v.replace("'Added data'!$AA:$AA",f"{Q}!$AA${RLO}:$AA${RHI}")
                  .replace("'Added data'!$AE:$AE",f"{Q}!$AP${RLO}:$AP${RHI}"))

# ---------- 3.2 restatement row -> REVIEW ----------
tc=wb["3.2 Total Cost"]
tc["B26"].value="Restatement vs the REVIEW ledger ($m) - must be 0"
tc["D26"].value=f"=ROUND(D24-SUM({Q}!$AA${RLO}:$AA${RHI})/1000000,3)"
tc["B27"].value="Every count and dollar on this tab reads the REVIEW - Complete Role Mapping sheet."

# ---------- 3.3 leftovers ----------
f3=wb["3.3 FTE View"]
for r in range(154,166):
    for col in (9,10): f3.cell(r,col).value=None       # old Squads I/J counts
f3["B167"].value=None
for r in range(169,185): f3.cell(r,6).value=None       # old Sheet2 cost col F

# ---------- 4.0 Data QA: clear every remaining old-ledger cell ----------
qa=wb["4.0 Data QA"]
cleared=0
for row in qa.iter_rows():
    for c in row:
        v=c.value
        if isinstance(v,str) and (("Squads!" in v) or ("'Added data'" in v) or ("Sheet2!" in v) or v.startswith("=Sheet2")):
            c.value=None; cleared+=1
qa["B105"].value="Old-ledger checks retired: the model reads REVIEW only. Comparison to the old Squads sheet lives on 3.5 Source Reconciliation."
qa["B105"].font=IT
print("4.0 old-ledger cells cleared:",cleared)

wb.save(P)
print("saved",P)
# final census
n=Counter()
wb2=openpyxl.load_workbook(P)
for ws in wb2.worksheets:
    if ws.title in ("Squads","Added data","Sheet2","Claude Log","3.5 Source Reconciliation"): continue
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value.startswith("="):
                if "Squads!" in c.value: n[("Squads",ws.title)]+=1
                if "Added data" in c.value: n[("Added data",ws.title)]+=1
                if "Sheet2" in c.value: n[("Sheet2",ws.title)]+=1
print("remaining old-ledger refs (excl 3.5 by design):", dict(n) if n else "NONE")
