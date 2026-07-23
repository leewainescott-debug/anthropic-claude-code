#!/usr/bin/env python3
"""Reroute the 2.x working tabs OFF Squads / Added data and ONTO the REVIEW sheet,
using the curated Model-Squad mapping (EE-Number join to Squads' Model Squad for
filled people, squad-field fallback for vacancies, Leadership booked to overhead).
Squad structure matches the validated old model; every count, cost and roster row
now reads REVIEW live, and the Hire/Hold/Offshore lever drives cost -> variance.
Also fixes three QA findings: banned word 'calls', 0.2 cyber double-count, and
inconsistent red-negative formats."""
import openpyxl, re
from openpyxl.utils import column_index_from_string as ci, get_column_letter as gl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.views import Selection
from collections import defaultdict, OrderedDict, Counter

SRC="clean_v6.xlsx"; OUT="clean_v7.xlsx"
wb=openpyxl.load_workbook(SRC)
REV="REVIEW - Complete Role Mapping"; Q="'"+REV+"'"; F3="'3.3 FTE View'"
rev=wb[REV]; sq=wb["Squads"]
def rc(r,cc): return rev.cell(r,ci(cc)).value
def scq(r,cc): return sq.cell(r,ci(cc)).value

PMAP={"ampol retail":"Ampol Retail","retail":"Ampol Retail","z":"Z Retail","customer":"Customer",
"ampol customer":"Customer","enterprise data":"Enterprise Data","tdd":"TDD Group Functions",
"infrastructure":"Infrastructure","b2b & energy solutions":"Energy Solutions & B2B",
"commercial fuels":"Commercial Fuels","finance":"Finance","p&c":"P&C","p&c, finance & legal":"P&C",
"egi":"EGI & Central","coe - cyber, risk & operations":"COE Cyber",
"coe - partnering & transformation":"COE BP&T","coe - strategy, architecture, data":"COE SA&D"}
def isvac(r):
    return str(rc(r,'B') or "").strip().lower()=="vacant"   # vacancy IS the name
def mtab(r): return PMAP.get(str(rc(r,'I') or "").strip().lower())

# ---- curated mapping from Squads: EE -> Model Squad / Class, and rawK -> Model Squad ----
ee2sq={}; ee2cl={}; k2sq=defaultdict(Counter)
for r in range(2,sq.max_row+1):
    if not sq.cell(r,2).value: continue
    ee=str(scq(r,'A') or "").strip(); p=str(scq(r,'P') or "").strip()
    cl=str(scq(r,'Q') or "").strip(); k=str(scq(r,'K') or "").strip().lower()
    if ee: ee2sq[ee]=p; ee2cl[ee]=cl
    if k and cl=="Squad": k2sq[k][p]+=1
k2sq={k:c.most_common(1)[0][0] for k,c in k2sq.items()}
def curated(r):
    """return (squad, is_leadership) for a REVIEW row"""
    ee=str(rc(r,'A') or "").strip(); k=str(rc(r,'K') or "").strip(); kl=k.lower()
    if ee and ee in ee2sq:
        if "eadership" in ee2cl[ee] or ee2sq[ee]=="Leadership": return None,True
        return ee2sq[ee],False
    if "eadership" in kl: return None,True
    if kl in k2sq: return k2sq[kl],False
    if kl=="": return "(unassigned)",False
    return {"na":"(all roles)"}.get(kl,k),False

# ---- fix MStatus (AK): a role is Vacant when its NAME is "Vacant" (or Q flags it) ----
AK=ci("AK")
for r in range(2,rev.max_row+1):
    if not rev.cell(r,2).value: continue
    rev.cell(r,AK).value=(f'=IF($B{r}="Vacant","Vacant",'
        f'IF(LOWER($Q{r})="pause","Paused",'
        f'IF(OR(LOWER($Q{r})="cxc",$Q{r}="Contractor"),"Contractor","Filled")))')

# ---- REVIEW helpers: MSquadC (AP), MLead (AQ) as values; collect roles per tab/squad ----
AP=ci("AP"); AQ=ci("AQ")
rev.cell(1,AP).value="MSquadC"; rev.cell(1,AQ).value="MLead"
data=defaultdict(lambda:OrderedDict()); leadcnt=Counter()
for r in range(2,rev.max_row+1):
    if not rev.cell(r,2).value:
        rev.cell(r,AP).value=None; rev.cell(r,AQ).value=None; continue
    s,lead=curated(r)
    rev.cell(r,AP).value=("" if lead else s); rev.cell(r,AQ).value=(1 if lead else 0)
    t=mtab(r)
    if not t: continue
    if lead: leadcnt[t]+=1; continue
    data[t].setdefault(s,[]).append(dict(dr=r, vac=isvac(r)))

# ---- 3.3 key helper (portfolio|squad) for archetype lookup ----
f3=wb["3.3 FTE View"]; KEYCOL=ci("AD")
if f3.cell(6,KEYCOL).value!="_key":
    f3.cell(6,KEYCOL).value="_key"
    for r in range(7,95):
        b=f3.cell(r,2).value; d=f3.cell(r,4).value
        if b is not None and d is not None: f3.cell(r,KEYCOL).value=f'=$B{r}&"|"&$D{r}'
    f3.column_dimensions[gl(KEYCOL)].hidden=True
KC=gl(KEYCOL)

# ---- styling ----
NAVY=PatternFill("solid",fgColor="1F4E79"); GREY=PatternFill("solid",fgColor="F2F2F2")
TOT=PatternFill("solid",fgColor="D9D9D9"); YEL=PatternFill("solid",fgColor="FFF2CC")
SQH=PatternFill("solid",fgColor="DDEBF7"); ARCH=PatternFill("solid",fgColor="EDEDED")
WF=Font(color="FFFFFF",bold=True); BF=Font(bold=True); NF=Font(); BL=Font(color="0000FF"); IT=Font(italic=True,size=10)
thin=Side(style="thin",color="BFBFBF"); BOX=Border(thin,thin,thin,thin)
M2='#,##0.00;[Red](#,##0.00);"-"'; D0='#,##0;[Red](#,##0);"-"'; DOL='#,##0'
CEN=Alignment(horizontal="center",vertical="center"); RT=Alignment(horizontal="right")
def sc(w,ref,v,font=NF,fill=None,al=None,fmt=None,box=True):
    x=w[ref]; x.value=v; x.font=font
    if fill:x.fill=fill
    if al:x.alignment=al
    if fmt:x.number_format=fmt
    if box:x.border=BOX
    return x

WT=OrderedDict([("2.1 Ampol Retail","Ampol Retail"),("2.2 Customer","Customer"),
("2.3 Enterprise Data","Enterprise Data"),("2.4 TDD Group Functions","TDD Group Functions"),
("2.5 P&C","P&C"),("2.6 Finance","Finance"),("2.7 Infrastructure","Infrastructure"),
("2.8 Energy Solutions & B2B","Energy Solutions & B2B"),("2.9 Commercial Fuels","Commercial Fuels"),
("2.10 Z Retail","Z Retail"),("2.11 TDD Cyber","COE Cyber"),("2.14 EGI & Central","EGI & Central")])
HDR=["Squad","Archetype type","Archetype roles","Filled","Vacant","Planning to hire",
     "Vacancies remaining","vs archetype","Flag","Archetype cost ($m)","Actual cost ($m)",
     "Cost after vacancy decisions ($m)","New Variance ($m)","Archetype size"]  # B..O
mtot={}; hire_memo={}
for tabname,mt in WT.items():
    w=wb[tabname]
    for row in list(w.iter_rows(min_row=3)):
        for cl in row:
            cl.value=None; cl.fill=PatternFill(); cl.border=Border()
            cl.alignment=Alignment(); cl.number_format="General"
    for dv in list(w.data_validations.dataValidation): w.data_validations.dataValidation.remove(dv)
    squads=data.get(mt,{})
    sc(w,"B2",f"{mt} - working copy",Font(bold=True,size=15),box=False)
    sc(w,"B4","Position by squad - archetype vs actual (from REVIEW), with the vacancy lever",WF,NAVY,box=False)
    for i,h in enumerate(HDR): sc(w,f"{gl(2+i)}5",h,WF,NAVY,CEN)
    first=6; n=len(squads); total_row=first+n
    rost_hdr=total_row+8
    roster_rows={}; rr=rost_hdr+1
    for s,roles in squads.items():
        roster_rows[s]=(rr+1, rr+len(roles)); rr+=1+len(roles)
    dv=DataValidation(type="list",formula1='"Hire,Hold,Offshore"',allow_blank=False); w.add_data_validation(dv)
    RR="$2:$530"; TR="$7:$94"   # bounded data ranges (REVIEW rows, 3.3 squad rows) - fast recalc
    r=first
    for s,roles in squads.items():
        a,b=roster_rows[s]; sqq=s.replace('"','""'); key=f'"{mt}|{sqq}"'
        look=lambda col: f'IFERROR(INDEX({F3}!${col}$7:${col}$94,MATCH({key},{F3}!${KC}$7:${KC}$94,0)),"-")'
        cnt=f'{Q}!$AJ$2:$AJ$530,"{mt}",{Q}!$AP$2:$AP$530,"{sqq}"'
        sc(w,f"B{r}",s,NF,GREY if (r-first)%2 else None)
        sc(w,f"C{r}",f"={look('E')}",NF,ARCH,CEN)
        sc(w,f"D{r}",f"={look('G')}",NF,ARCH,CEN,fmt=D0)
        sc(w,f"E{r}",f'=COUNTIFS({cnt},{Q}!$AK$2:$AK$530,"Filled")',NF,al=CEN,fmt=D0)
        sc(w,f"F{r}",f'=COUNTIFS({cnt},{Q}!$AK$2:$AK$530,"Vacant")',NF,al=CEN,fmt=D0)
        sc(w,f"G{r}",f'=COUNTIF($E${a}:$E${b},"Hire")',NF,al=CEN,fmt=D0)
        sc(w,f"H{r}",f'=F{r}-G{r}-COUNTIF($E${a}:$E${b},"Offshore")',NF,al=CEN,fmt=D0)
        sc(w,f"I{r}",f'=IFERROR(E{r}+G{r}-D{r},"-")',NF,al=CEN,fmt=D0)
        sc(w,f"J{r}",f'=IF(ISNUMBER(D{r}),IF(E{r}>D{r},"Filled already over archetype",IF(E{r}+G{r}>D{r},"Over archetype after planned hires","")),"Outside the archetype model - no target set")',NF)
        sc(w,f"K{r}",f"={look('M')}",NF,ARCH,RT,fmt=M2)
        sc(w,f"L{r}",f'=SUMIFS({Q}!$AA$2:$AA$530,{cnt},{Q}!$AK$2:$AK$530,"Filled")/1000000',NF,al=RT,fmt=M2)
        dec=f'SUMIFS($F${a}:$F${b},$E${a}:$E${b},"Hire")+0.4*SUMIFS($F${a}:$F${b},$E${a}:$E${b},"Offshore")'
        sc(w,f"M{r}",f'=L{r}+({dec})/1000000',NF,al=RT,fmt=M2)
        sc(w,f"N{r}",f'=IFERROR(M{r}-K{r},"-")',NF,al=RT,fmt=M2)
        sc(w,f"O{r}",f"={look('F')}",NF,ARCH,CEN)
        r+=1
    last=r-1
    sc(w,f"B{total_row}","Total",BF,TOT)
    for col in "DEFGHI": sc(w,f"{col}{total_row}",f"=SUM({col}{first}:{col}{last})",BF,TOT,CEN,D0)
    for col in "KLMN": sc(w,f"{col}{total_row}",f"=SUM({col}{first}:{col}{last})",BF,TOT,RT,M2)
    for col in "CJO": sc(w,f"{col}{total_row}","",BF,TOT)
    mtot[tabname]=f"M{total_row}"
    ra,rb=rost_hdr+1, rr-1
    sc(w,f"B{total_row+1}","Cost to hire all vacancies ($m)",NF,box=False)
    sc(w,f"E{total_row+1}",f'=SUMIFS($F${ra}:$F${rb},$D${ra}:$D${rb},"Vacant")/1000000',NF,al=RT,fmt=M2,box=False)
    hire_memo[tabname]=f"E{total_row+1}"
    sc(w,f"B{total_row+2}","Cost of roles marked Hire ($m)",NF,box=False)
    sc(w,f"E{total_row+2}",f'=SUMIF($E${ra}:$E${rb},"Hire",$F${ra}:$F${rb})/1000000',NF,al=RT,fmt=M2,box=False)
    sc(w,f"B{total_row+3}",f"Leadership: {leadcnt.get(mt,0)} roles funded via portfolio overhead (shown on 3.3 FTE View), not costed here.",IT,box=False)
    sc(w,f"B{total_row+4}",'"-" in the archetype columns means this real squad has no design archetype target; the comparison sits at portfolio level on 3.2.',IT,box=False)
    # roster
    sc(w,f"B{rost_hdr-1}",f"{mt} FTE (from REVIEW)",Font(bold=True,size=12),box=False)
    for i,h in enumerate(["Name","Role","Status","Vacancy lever","Cost if hired ($)"]):
        sc(w,f"{gl(2+i)}{rost_hdr}",h,WF,NAVY,CEN)
    rr=rost_hdr+1
    for s,roles in squads.items():
        sc(w,f"B{rr}",s,BF,SQH)
        for col in "CDEF": sc(w,f"{col}{rr}","",BF,SQH)
        rr+=1
        for x in roles:
            dr=x["dr"]
            sc(w,f"B{rr}",f"={Q}!$B${dr}",NF)
            sc(w,f"C{rr}",f"={Q}!$C${dr}",NF)
            sc(w,f"D{rr}",f"={Q}!$AK${dr}",NF,al=CEN)
            if x["vac"]:
                cell=sc(w,f"E{rr}","Hold",BL,YEL,CEN); dv.add(cell)
                sc(w,f"F{rr}",f"={Q}!$AA${dr}",NF,al=RT,fmt=DOL)
            else:
                sc(w,f"E{rr}","",NF); sc(w,f"F{rr}","",NF)
            rr+=1
    for col,wd in {"B":30,"C":16,"D":14,"E":9,"F":9,"G":15,"H":18,"I":12,"J":30,"K":18,"L":16,"M":26,"N":16,"O":14}.items():
        w.column_dimensions[col].width=wd

# ---- downstream rewire ----
tc=wb["3.2 Total Cost"]
F32={6:"2.1 Ampol Retail",7:"2.2 Customer",8:"2.3 Enterprise Data",9:"2.4 TDD Group Functions",
10:"2.5 P&C",11:"2.6 Finance",12:"2.7 Infrastructure",13:"2.8 Energy Solutions & B2B",
14:"2.9 Commercial Fuels",15:"2.10 Z Retail",20:"2.11 TDD Cyber"}
for r,t in F32.items(): tc[f"F{r}"].value=f"='{t}'!${mtot[t][0]}${mtot[t][1:]}"
picks=list(F32.values())
wb["Exec Summary"]["C52"].value="="+"+".join(f"'{t}'!${hire_memo[t][0]}${hire_memo[t][1:]}" for t in picks)
qa=wb["4.0 Data QA"]
qa["C258"].value=re.sub(r"\$F\$1:\$F\$500","$D$1:$D$500",qa["C258"].value)

# ---- QA fix: 0.2 Data Config cyber spend double-count (F7 duplicates F23/row 23) ----
dc=wb["0.2 Data Config"]
dc["F7"].value=0     # cyber spend represented by the TDD Cyber portfolio row (F23); F7 was a duplicate

# ---- QA fix: red-negative number format on the flagged cells ----
RED='#,##0.00;[Red](#,##0.00);"-"'
for a in ["F6","F7","F8","F9","F10","F11","G6","G7","G8","G9","G10","G11","H6","H7","H8","H9","H10","H11",
          "I6","I7","I8","I9","I10","I11","J6","J7","J8","J9","J10","J11"]:
    wb["3.4 COE Summary"][a].number_format=RED
for a in ["G6","G7","G8","G9","G10","G26"]: wb["0.2 Data Config"][a].number_format=RED
for a in ["E21","E22","E24","J24","K17","K24"]: wb["3.1 Group Summary"][a].number_format=RED
wb["1.11 BP&T"]["H8"].number_format=RED; wb["1.12 SA&D"]["I8"].number_format=RED
wb["1.13 Cyber Roles"]["F8"].number_format=RED

# ---- remove ALL frozen panes cleanly ----
for w in wb.worksheets:
    w.freeze_panes=None; w.sheet_view.pane=None
    w.sheet_view.selection=[Selection(activeCell="A1",sqref="A1")]

wb.save(OUT)
print("saved",OUT)
print("rerouted tabs:",list(WT))
print("leadership->overhead per tab:",dict(leadcnt))
# quick Squads/Added-data ref count on rerouted tabs
for t in WT:
    n=sum(1 for row in wb[t].iter_rows() for c in row if isinstance(c.value,str) and ("Squads" in c.value or "Added data" in c.value))
    print(f"  {t}: Squads/Added-data refs now = {n}")
