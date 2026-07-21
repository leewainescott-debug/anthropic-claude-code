#!/usr/bin/env python3
"""v6 on the user's Updated_view:
 1. Total-to-fund block (positive variances) aligned on every 1.x tab
 2. Full sweep: strip all comments, AI-ish grey notes, random orange fills on
    raw data, delete Sheet1 (user's own J/K/L notes are kept)
 3. 1.11 Cyber tidy: structured table, Cyber CapEx 0.5 in the funding block
 4. 2.2 BP&T / 2.3 SA&D / 2.4 Cyber Roles detail tabs from Added data
    (per-role On/Off + vacancy + live cost; budgets to draw down; totals)
 5. 2.1 COE hub rewired to the detail tabs
 6. 3.0 FTE View rebuilt: KPIs, clean table with subtotals, named leadership,
    vacancy analysis, red/green CF
 7. 4.0 Insights: CTO dashboard with portfolio selector + key callouts
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

U = "/root/.claude/uploads/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/"
SRC = U + "7fc99890-TDD_Cost_Calc_Updated_view.xlsx"
OUT = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/TDD_Cost_Calc_v6.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=False)
wbv = openpyxl.load_workbook(SRC, data_only=True)

NAVY = "FF1F4E79"; DKNAVY = "FF002F6C"; GREY = "FFD9D9D9"; YEL = "FFFFF2CC"
YELLOW = PatternFill("solid", fgColor=YEL)
NAVY_F = PatternFill("solid", fgColor=NAVY); DK_F = PatternFill("solid", fgColor=DKNAVY)
GREY_F = PatternFill("solid", fgColor=GREY)
W = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
B = Font(name="Calibri", size=10, bold=True)
N = Font(name="Calibri", size=10)
BLUE = Font(name="Calibri", size=10, color="FF0000FF")
TITLE = Font(name="Calibri", size=16, bold=True)
thin = Side(style="thin", color="FFB8C9CC")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '#,##0.00;\\(#,##0.00\\);\\-'
DOLLAR = '$#,##0;($#,##0);\\-'
PCT = '0%'
GREEN_F = PatternFill("solid", fgColor="FFE2EFDA"); GREEN_T = Font(color="FF006100", bold=True)
RED_F = PatternFill("solid", fgColor="FFFBE4D5"); RED_T = Font(color="FF9C0006", bold=True)

def pos_bad(ws, rng):
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=RED_F, font=RED_T))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=GREEN_F, font=GREEN_T))
def pos_good(ws, rng):
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_F, font=GREEN_T))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_F, font=RED_T))

def frow(ws, col, text, exact=True, start=1):
    ci = openpyxl.utils.column_index_from_string(col)
    for r in range(start, ws.max_row + 2):
        v = ws.cell(r, ci).value
        if v is None: continue
        s = str(v).strip()
        if (exact and s == text) or (not exact and text in s):
            return r
    return None

def set_cell(ws, coord, value, font=None, fill=None, fmt=None, border=True, align=None):
    c = ws[coord]; c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if border: c.border = BOX
    if align: c.alignment = Alignment(horizontal=align, vertical="center")
    return c

# =====================================================================
# 1. sweep: comments off, Sheet1 gone, orange fills off raw data,
#    my grey-italic notes gone
# =====================================================================
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if c.comment is not None:
                c.comment = None
if "Sheet1" in wb.sheetnames:
    del wb["Sheet1"]
rd = wb["raw data"]
NOFILL = PatternFill()
for row in rd.iter_rows():
    for c in row:
        if c.fill and c.fill.patternType == "solid":
            rgb = getattr(c.fill.fgColor, "rgb", None)
            if isinstance(rgb, str) and rgb[2:] in ("FFA500","ED7D31","F4B084","FCE4D6","FFC000","FFD966"):
                c.fill = NOFILL
fin04 = wb["0.4 Budget Table (Fin)"]
if fin04["B35"].value and "printed Finance pack" in str(fin04["B35"].value):
    fin04["B35"] = None
coe = wb["2.1 COE"]
for r in range(22, 28):
    v = coe[f"B{r}"].value
    if isinstance(v, str) and ("depreciation" in v.lower() or "note:" in v.lower()):
        coe[f"B{r}"] = None
# Added data repairs: pasted #N/A artifact in Unit col; SUBTOTAL cells frozen to
# their current displayed values (kept verbatim from the source's cached results)
ad1 = wb["Added data"]; adv1 = wbv["Added data"]
if ad1["R468"].data_type == "e" or ad1["R468"].value == "#N/A":
    ad1["R468"] = "Daily"
for cc in ("AA551", "AB551"):
    v = ad1[cc].value
    if isinstance(v, str) and v.startswith("=SUBTOTAL"):
        ad1[cc] = adv1[cc].value

# =====================================================================
# 2. Total-to-fund block, aligned, positive, on every 1.x tab
#    (placed in B/C/D beside the funding block, under 'Reconciled to Finance')
# =====================================================================
TABS = ["1.1 Ampol Retail","1.2 Customer","1.3 Enterprise Data","1.4 TDD Group Functions",
        "1.5 P&C","1.6 Finance","1.7 Infrastructure","1.8 Energy Solutions & B2B",
        "1.9 Commercial Fuels","1.10 Z Retail","1.11 TDD Cyber"]
anchors = {}
for t in TABS:
    ws = wb[t]
    vr = frow(ws, "G", "Variance (budget less cost)")
    hdr = frow(ws, "G", "Budget line") or frow(ws, "G", "Funding position") or 10
    lr = frow(ws, "G", "Left to fund", start=hdr)
    ar = frow(ws, "G", "Total applied", start=hdr) or frow(ws, "G", "Funded from TDD Corporate pool", exact=False, start=hdr)
    anchors[t] = dict(var=vr, left=lr, applied=ar)
    # clear any prior variants of the block in B/C/D (incl 1.2's x's and old Total to fund)
    for r in range(hdr + 1, (lr or hdr + 9) + 4):
        for col in "BCD":
            v = ws[f"{col}{r}"].value
            if isinstance(v, str) and v.strip() in ("Total to fund","TDD Variance","Other Variance","Total","x"):
                ws[f"{col}{r}"] = None
            elif col == "C" and isinstance(v, str) and v.startswith("=SUM(-H"):
                ws[f"{col}{r}"] = None
    base = (frow(ws, "B", "Reconciled to Finance", exact=False) or hdr) + 2
    rows = [("Total to fund", None), ("TDD Variance", f"=MAX(0,-H{vr})"),
            ("Other Variance", f"=MAX(0,I{lr})"), ("Total", f"=C{base+1}+C{base+2}")]
    for i, (label, f) in enumerate(rows):
        rr = base + i
        set_cell(ws, f"B{rr}", label, B if i in (0, 3) else N, NAVY_F if i == 0 else (GREY_F if i == 3 else None),
                 align="left")
        if i == 0:
            ws[f"B{rr}"].font = W
            set_cell(ws, f"C{rr}", None, W, NAVY_F)
        else:
            set_cell(ws, f"C{rr}", f, B if i == 3 else N, GREY_F if i == 3 else None, MONEY, align="right")
    pos_bad(ws, f"C{base+3}")

# =====================================================================
# 3. 1.11 Cyber tidy + CapEx 0.5
# =====================================================================
cy = wb["1.11 TDD Cyber"]
cy["H16"] = 0.5
cy["H16"].fill = YELLOW; cy["H16"].font = BLUE; cy["H16"].number_format = MONEY
cy["I16"] = 0
cy["I16"].fill = YELLOW; cy["I16"].font = BLUE; cy["I16"].number_format = MONEY
set_cell(cy, "G15", "Cyber funding buckets", W, NAVY_F, align="left")
set_cell(cy, "H15", "Budget ($m)", W, NAVY_F, align="center")
set_cell(cy, "I15", "Applied ($m)", W, NAVY_F, align="center")
set_cell(cy, "J15", "Remaining ($m)", W, NAVY_F, align="center")
set_cell(cy, "J16", "=H16-I16", N, None, MONEY, align="right")
cy["G16"] = "Cyber CapEx (Monitoring)"
cy["G16"].font = N; cy["G16"].border = BOX
# tidy the stray planning notes into one column block (keep user's words verbatim, values only moved if scattered)
# (user notes at B16:B19/D16:D17 stay untouched)

# =====================================================================
# 4. detail tabs from Added data
# =====================================================================
ad = wb["Added data"]; adv = wbv["Added data"]
BPT_D = {"tdd business partner": "Business Partnering", "commercial": "Business Partnering",
         "transformation": "Transformation"}
SAD_D = {"architecture", "technology strategy & ai capability", "delivery, sada", "group data"}
CYB_D = {"cyber strat & tech", "cyber sec ops", "cyber risk", "cyber grc", "service op & assurance"}
SQ13 = {"data science","coe, data science","reporting & analytics","data platform","data platforms",
        "enterprise data delivery","coe, data operations"}
groups = {"BPT": [], "SAD": [], "CYB": []}
for r in range(2, ad.max_row + 1):
    g = lambda cc: (str(adv.cell(r, cc).value).strip() if adv.cell(r, cc).value is not None else "")
    name, title, dept, sq, ctry = g(2), g(3), (g(7) or g(6)).lower(), g(11).lower(), g(13)
    if not name and not title: continue
    if dept in BPT_D:
        groups["BPT"].append((r, name, title, dept, ctry, BPT_D[dept]))
    elif dept in SAD_D:
        if dept == "group data":
            cat = "Data - Portfolio squad (1.3)" if sq in SQ13 else ("Strategy & Architecture" if "architect" in title.lower() else "Data - COE")
        elif "architect" in title.lower() or dept != "group data":
            cat = "Strategy & Architecture"
        groups["SAD"].append((r, name, title, dept, ctry, cat))
    elif dept in CYB_D:
        groups["CYB"].append((r, name, title, dept, ctry, dept.title()))

ls = wb["Lists"]
ls["E1"] = "BPT Category"; ls["E2"] = "Business Partnering"; ls["E3"] = "Transformation"; ls["E4"] = "Exclude"
ls["F1"] = "SAD Category"; ls["F2"] = "Strategy & Architecture"; ls["F3"] = "Data - COE"
ls["F4"] = "Data - Portfolio squad (1.3)"; ls["F5"] = "Exclude"
wb.defined_names["BPTCat"] = openpyxl.workbook.defined_name.DefinedName("BPTCat", attr_text="Lists!$E$2:$E$4")
wb.defined_names["SADCat"] = openpyxl.workbook.defined_name.DefinedName("SADCat", attr_text="Lists!$F$2:$F$5")

OFF = "'0.1 Squads'!$J$5"   # 0.4 offshore factor - same logic as squads
def role_tab(title_txt, sheetname, rows, cat_named, cat_options_note, budgets, group_field=6):
    ws = wb.create_sheet(sheetname)
    ws.sheet_view.showGridLines = False
    for col, wdt in {"A":3,"B":26,"C":42,"D":24,"E":10,"F":9,"G":24,"H":11,"I":13,"J":13,"K":3}.items():
        ws.column_dimensions[col].width = wdt
    ws.row_dimensions[2].height = 21
    set_cell(ws, "B2", title_txt, TITLE, border=False)
    hdrs = ["Name","Position Title","Department","Country","Status","Category","On/Off",
            "Full Cost AUD ($)","Model cost ($m)"]
    for j, h in enumerate(hdrs):
        set_cell(ws, f"{'BCDEFGHIJ'[j]}5", h, W, NAVY_F, align="center")
    dvo = DataValidation(type="list", formula1="OnOff", allow_blank=True)
    dvc = DataValidation(type="list", formula1=cat_named, allow_blank=True) if cat_named else None
    ws.add_data_validation(dvo)
    if dvc: ws.add_data_validation(dvc)
    r = 6
    first = r
    for (srcrow, name, title, dept, ctry, cat) in sorted(rows, key=lambda x: (x[5], x[3], x[1])):
        set_cell(ws, f"B{r}", name if name.lower() != "vacant" else "Vacant", N, align="left")
        set_cell(ws, f"C{r}", title, N, align="left")
        set_cell(ws, f"D{r}", dept.title() if dept.islower() else dept, N, align="left")
        set_cell(ws, f"E{r}", ctry, N, align="left")
        set_cell(ws, f"F{r}", f"=IF('Added data'!$B${srcrow}=\"Vacant\",\"Vacant\",\"Filled\")", N, align="center")
        cc = set_cell(ws, f"G{r}", cat, N, YELLOW if cat_named else None, align="left")
        oo = set_cell(ws, f"H{r}", "Onshore", N, YELLOW, align="center")
        set_cell(ws, f"I{r}", f"='Added data'!$AA${srcrow}", N, None, DOLLAR, align="right")
        set_cell(ws, f"J{r}", f"=IF(H{r}=\"Offshore\",I{r}*{OFF},I{r})/1000000", N, None, MONEY, align="right")
        dvo.add(f"H{r}")
        if dvc: dvc.add(f"G{r}")
        r += 1
    last = r - 1
    ws.conditional_formatting.add(f"F{first}:F{last}",
        CellIsRule(operator="equal", formula=['"Vacant"'], fill=RED_F, font=RED_T))
    r += 1
    # summary block
    set_cell(ws, f"B{r}", "Summary", W, DK_F, align="left")
    for col in "CDEFGHIJ": set_cell(ws, f"{col}{r}", None, W, DK_F)
    r += 1
    hdr2 = ["Category","Roles","Filled","Vacant","Planned spend ($m)","Budget to draw down ($m)","To fund ($m)"]
    for j, h in enumerate(hdr2):
        set_cell(ws, f"{'BCDEFGH'[j]}{r}", h, W, NAVY_F, align="center")
    r += 1
    sum_first = r
    for (cat, budget_formula, budget_label) in budgets:
        set_cell(ws, f"B{r}", cat, N, align="left")
        set_cell(ws, f"C{r}", f"=COUNTIF(G{first}:G{last},B{r})", N, align="center")
        set_cell(ws, f"D{r}", f"=COUNTIFS(G{first}:G{last},B{r},F{first}:F{last},\"Filled\")", N, align="center")
        set_cell(ws, f"E{r}", f"=COUNTIFS(G{first}:G{last},B{r},F{first}:F{last},\"Vacant\")", N, align="center")
        set_cell(ws, f"F{r}", f"=SUMIF(G{first}:G{last},B{r},J{first}:J{last})", N, None, MONEY, align="right")
        set_cell(ws, f"G{r}", budget_formula, N, None, MONEY, align="right")
        set_cell(ws, f"H{r}", f"=MAX(0,F{r}-G{r})", B, None, MONEY, align="right")
        ws[f"I{r}"] = budget_label; ws[f"I{r}"].font = N
        pos_bad(ws, f"H{r}")
        r += 1
    set_cell(ws, f"B{r}", "Total", B, GREY_F, align="left")
    for col, cl in [("C","C"),("D","D"),("E","E"),("F","F"),("G","G"),("H","H")]:
        set_cell(ws, f"{col}{r}", f"=SUM({cl}{sum_first}:{cl}{r-1})", B, GREY_F,
                 MONEY if col in "FGH" else None, align="right" if col in "FGH" else "center")
    return ws, first, last, sum_first, r

DC = "'0.0 Data Config'"
bpt_ws, bpt_f, bpt_l, bpt_sf, bpt_tot = role_tab(
    "Business Partnering & Transformation - roles and funding", "2.2 BP&T",
    groups["BPT"], "BPTCat", None,
    budgets=[("Business Partnering", f"=11*{DC}!$L$7", "Business Partner overhead paid by all 11 portfolios"),
             ("Transformation", f"={DC}!$E$8", "COE - Transformation allocation")])
sad_ws, sad_f, sad_l, sad_sf, sad_tot = role_tab(
    "Strategy, Architecture & Data - roles and funding", "2.3 SA&D",
    groups["SAD"], "SADCat", None,
    budgets=[("Strategy & Architecture", f"={DC}!$E$6+11*{DC}!$L$8",
              "COE - Strategy Architecture allocation + Domain Architect overhead paid by 11 portfolios"),
             ("Data - COE", f"={DC}!$E$10", "COE - Data allocation"),
             ("Data - Portfolio squad (1.3)", "=0", "Funded inside the Enterprise Data portfolio (1.3)")])
sad_ws["H" + str(sad_sf + 2)] = 0
cyb_ws, cyb_f, cyb_l, cyb_sf, cyb_tot = role_tab(
    "Cyber, Risk & Service Operations - roles and funding", "2.4 Cyber Roles",
    groups["CYB"], None, None,
    budgets=[("Cyber Strat & Tech", "=0", ""), ("Cyber Sec Ops", "=0", ""), ("Cyber Risk", "=0", ""),
             ("Cyber Grc", "=0", ""), ("Service Op & Assurance", "=0", "")])
# cyber: single budget = bucket + capex; per-team budget/to-fund columns are noise
for rr in range(cyb_sf, cyb_tot + 1):
    cyb_ws[f"G{rr}"] = None
    cyb_ws[f"H{rr}"] = None
r = cyb_tot + 2
set_cell(cyb_ws, f"B{r}", "Cyber funding", W, DK_F, align="left")
for col in "CDEFGH": set_cell(cyb_ws, f"{col}{r}", None, W, DK_F)
r += 1
for label, f in [("TDD Cyber bucket (Data Config)", f"={DC}!$E$23"),
                 ("Cyber CapEx - Monitoring", "='1.11 TDD Cyber'!$H$16"),
                 ("Total budget", f"=C{r}+C{r+1}"),
                 ("Planned spend (all cyber roles)", f"=SUM(J{cyb_f}:J{cyb_l})"),
                 ("To fund", f"=MAX(0,C{r+3}-C{r+2})")]:
    set_cell(cyb_ws, f"B{r}", label, B if "Total" in label or "To fund" in label else N, align="left")
    set_cell(cyb_ws, f"C{r}", f, B if "Total" in label or "To fund" in label else N,
             GREY_F if "To fund" in label else None, MONEY, align="right")
    r += 1
pos_bad(cyb_ws, f"C{r-1}")

# =====================================================================
# 5. 2.1 COE hub rewired to detail tabs (user's K/L notes untouched)
# =====================================================================
coe["D8"] = f"='2.3 SA&D'!$F${sad_sf}"          # Strategy & Architecture planned
coe["D10"] = f"='2.2 BP&T'!$F${bpt_sf+1}"       # Transformation planned
coe["D11"] = f"='2.2 BP&T'!$F${bpt_sf}"         # Business Partnering planned
coe["D12"] = f"='2.3 SA&D'!$F${sad_sf+1}"       # Data - COE planned
for c in ("D8","D10","D11","D12"):
    coe[c].fill = PatternFill(); coe[c].font = N

# =====================================================================
# 6. 3.0 FTE View rebuild
# =====================================================================
old = wb["3.0 FTE View"]
idx = wb.sheetnames.index("3.0 FTE View")
wb.remove(old)
ft = wb.create_sheet("3.0 FTE View", idx)
ft.sheet_view.showGridLines = False
for col, wdt in {"A":3,"B":24,"C":26,"D":30,"E":26,"F":7,"G":13,"H":8,"I":8,"J":9,"K":15,"L":11,"M":16}.items():
    ft.column_dimensions[col].width = wdt
ft.row_dimensions[2].height = 21
set_cell(ft, "B2", "FTE View - archetypes vs actual organisation", TITLE, border=False)
RD = "'raw data'"
# KPI strip
kpis = [("Org roles", f"=COUNTA({RD}!$R$2:$R$1000)"),
        ("Filled", f"=COUNTIF({RD}!$R$2:$R$1000,\"Filled\")"),
        ("Vacant", f"=COUNTIF({RD}!$R$2:$R$1000,\"Vacant\")"),
        ("Vacancy rate", "=E4/C4"),
        ("Archetype FTE", None),          # filled after table build
        ("Seats over archetype", None)]
kcols = "CDEFGH"
for j, (label, f) in enumerate(kpis):
    set_cell(ft, f"{kcols[j]}3", label, W, NAVY_F, align="center")
    set_cell(ft, f"{kcols[j]}4", f, B, GREY_F, PCT if label == "Vacancy rate" else None, align="center")

HDR = ["Portfolio","Platform","Squad","Squad Type","Size","Archetype FTE","Filled","Vacant",
       "Seats","Seats vs archetype","Vacancy %","Archetype cost ($m)"]
r = 6
for j, h in enumerate(HDR):
    set_cell(ft, f"{'BCDEFGHIJKLM'[j]}{r}", h, W, NAVY_F, align="center")
r += 1
def blocks(w):
    out = []
    rows = [rr for rr in range(1, w.max_row + 1)
            if isinstance(w.cell(rr, 2).value, str) and w.cell(rr, 2).value.startswith("Platform: ")]
    for tr in rows:
        hr = tr + 1; ohr = None
        for rr in range(hr + 1, w.max_row + 2):
            bv = w.cell(rr, 2).value
            if bv == "Platform Overhead" or (isinstance(bv, str) and (bv.endswith(" Total") or bv.startswith("(combined"))):
                ohr = rr; break
        out.append((tr, hr, hr + 1, (ohr or hr + 2) - 1))
    return out
PORT_LABEL = {t: t.split(" ", 1)[1] for t in TABS}
first_data = r
port_rows = {}
for t in TABS:
    w = wb[t]
    pstart = r
    for (tr, hrr, s0, s1) in blocks(w):
        pname = str(w.cell(tr, 2).value).replace("Platform: ", "")
        if "combined into" in pname: continue
        for sr in range(s0, s1 + 1):
            sqn = w.cell(sr, 2).value
            if not sqn: continue
            pl = PORT_LABEL[t]
            set_cell(ft, f"B{r}", pl, N, align="left")
            set_cell(ft, f"C{r}", pname, N, align="left")
            set_cell(ft, f"D{r}", sqn, N, align="left")
            set_cell(ft, f"E{r}", f"='{t}'!$C${sr}", N, align="left")
            set_cell(ft, f"F{r}", f"='{t}'!$D${sr}", N, align="center")
            set_cell(ft, f"G{r}", (f"=IFERROR(INDEX('0.1 Squads'!$F$5:$F$23,MATCH('{t}'!$C${sr}&\"|\"&'{t}'!$D${sr},"
                                   f"'0.1 Squads'!$A$5:$A$23,0)),\"-\")"), N, align="center")
            set_cell(ft, f"H{r}", f"=COUNTIFS({RD}!$N:$N,\"{pl}\",{RD}!$P:$P,\"{sqn}\",{RD}!$R:$R,\"Filled\")", N, align="center")
            set_cell(ft, f"I{r}", f"=COUNTIFS({RD}!$N:$N,\"{pl}\",{RD}!$P:$P,\"{sqn}\",{RD}!$R:$R,\"Vacant\")", N, align="center")
            set_cell(ft, f"J{r}", f"=H{r}+I{r}", N, align="center")
            set_cell(ft, f"K{r}", f'=IFERROR(J{r}-G{r},"-")', N, align="center")
            set_cell(ft, f"L{r}", f'=IFERROR(I{r}/J{r},"-")', N, None, PCT, align="center")
            set_cell(ft, f"M{r}", f"='{t}'!$G${sr}", N, None, MONEY, align="right")
            r += 1
    # portfolio subtotal
    set_cell(ft, f"B{r}", f"{PORT_LABEL[t]} total", B, GREY_F, align="left")
    for colL in "CDEF": set_cell(ft, f"{colL}{r}", None, B, GREY_F)
    for colL in "GHIJ":
        set_cell(ft, f"{colL}{r}", f"=SUM({colL}{pstart}:{colL}{r-1})", B, GREY_F, align="center")
    set_cell(ft, f"K{r}", f"=J{r}-G{r}", B, GREY_F, align="center")
    set_cell(ft, f"L{r}", f'=IFERROR(I{r}/J{r},"-")', B, GREY_F, PCT, align="center")
    set_cell(ft, f"M{r}", f"=SUM(M{pstart}:M{r-1})", B, GREY_F, MONEY, align="right")
    port_rows[t] = r
    r += 1
last_all = r - 1
set_cell(ft, f"B{r}", "TOTAL - all modelled squads", W, DK_F, align="left")
for colL in "CDEF": set_cell(ft, f"{colL}{r}", None, W, DK_F)
sub = ",".join(f"{{c}}{rr}".format(c="{c}", rr=rr) for rr in port_rows.values())
for colL in "GHIJ":
    set_cell(ft, f"{colL}{r}", "=" + "+".join(f"{colL}{rr}" for rr in port_rows.values()), W, DK_F, align="center")
set_cell(ft, f"K{r}", f"=J{r}-G{r}", W, DK_F, align="center")
set_cell(ft, f"L{r}", f"=I{r}/J{r}", W, DK_F, PCT, align="center")
set_cell(ft, f"M{r}", "=" + "+".join(f"M{rr}" for rr in port_rows.values()), W, DK_F, MONEY, align="right")
grand = r
ft["G4"] = f"=G{grand}"
ft["H4"] = f"=K{grand}"
pos_bad(ft, f"K{first_data}:K{grand}")
ft.conditional_formatting.add(f"L{first_data}:L{last_all}",
    ColorScaleRule(start_type="num", start_value=0, start_color="FFE2EFDA",
                   end_type="num", end_value=0.6, end_color="FFF4B084"))
r += 2

# Leadership: named list
set_cell(ft, f"B{r}", "Leadership roles (funded by platform & portfolio overheads, not squads)", W, DK_F, align="left")
for col in "CDEFGHIJKLM": set_cell(ft, f"{col}{r}", None, W, DK_F)
r += 1
for j, h in enumerate(["Portfolio","Name","Position Title","Org platform field","","","","","","","",""]):
    if h: set_cell(ft, f"{'BCDEFGHIJKLM'[j]}{r}", h, W, NAVY_F, align="center")
r += 1
lead_first = r
rdv2 = wbv["raw data"]
leads = []
for rr in range(2, rdv2.max_row + 1):
    if rdv2.cell(rr, 17).value == "Leadership":
        leads.append((str(rdv2.cell(rr, 14).value or ""), str(rdv2.cell(rr, 2).value or ""),
                      str(rdv2.cell(rr, 3).value or ""), str(rdv2.cell(rr, 10).value or "")))
for (pl, nm, ti, plat) in sorted(leads):
    set_cell(ft, f"B{r}", pl, N, align="left")
    set_cell(ft, f"C{r}", nm, N, align="left")
    if nm.lower() == "vacant":
        ft[f"C{r}"].fill = RED_F; ft[f"C{r}"].font = RED_T
    set_cell(ft, f"D{r}", ti, N, align="left")
    set_cell(ft, f"E{r}", plat, N, align="left")
    r += 1
set_cell(ft, f"B{r}", "Total leadership roles", B, GREY_F, align="left")
set_cell(ft, f"C{r}", len(leads), B, GREY_F, align="center")
set_cell(ft, f"D{r}", f'{sum(1 for x in leads if x[1].lower()!="vacant")} filled / '
                      f'{sum(1 for x in leads if x[1].lower()=="vacant")} vacant', B, GREY_F, align="left")
r += 2

# not-in-archetype section
set_cell(ft, f"B{r}", "In op model / org data but NOT in the archetype model", W, DK_F, align="left")
for col in "CDEFGHIJKLM": set_cell(ft, f"{col}{r}", None, W, DK_F)
r += 1
for j, h in enumerate(["Portfolio","Group","","","","","Filled","Vacant","Seats"]):
    if h: set_cell(ft, f"{'BCDEFGHIJ'[j]}{r}", h, W, NAVY_F, align="center")
r += 1
un_first = r
for (p, s) in [("Customer","Customer AI"),("Z Retail","Data NZ"),("EGI","EGI (unassigned)")]:
    set_cell(ft, f"B{r}", p, N, RED_F, align="left")
    set_cell(ft, f"C{r}", s, N, RED_F, align="left")
    set_cell(ft, f"H{r}", f"=COUNTIFS({RD}!$N:$N,\"{p}\",{RD}!$P:$P,\"{s}\",{RD}!$R:$R,\"Filled\")", N, RED_F, align="center")
    set_cell(ft, f"I{r}", f"=COUNTIFS({RD}!$N:$N,\"{p}\",{RD}!$P:$P,\"{s}\",{RD}!$R:$R,\"Vacant\")", N, RED_F, align="center")
    set_cell(ft, f"J{r}", f"=H{r}+I{r}", N, RED_F, align="center")
    r += 1
set_cell(ft, f"B{r}", "Other unmapped", N, RED_F, align="left")
set_cell(ft, f"H{r}", f"=COUNTIFS({RD}!$Q:$Q,\"Unmapped\",{RD}!$R:$R,\"Filled\")-SUM(H{un_first}:H{r-1})", N, RED_F, align="center")
set_cell(ft, f"I{r}", f"=COUNTIFS({RD}!$Q:$Q,\"Unmapped\",{RD}!$R:$R,\"Vacant\")-SUM(I{un_first}:I{r-1})", N, RED_F, align="center")
set_cell(ft, f"J{r}", f"=H{r}+I{r}", N, RED_F, align="center")
un_last = r
r += 2
# COE seats (live from raw data classes) + cross-check
set_cell(ft, f"B{r}", "COE seats (detail on 2.2 / 2.3 / 2.4)", W, DK_F, align="left")
for col in "CDEFGHIJKLM": set_cell(ft, f"{col}{r}", None, W, DK_F)
r += 1
set_cell(ft, f"B{r}", "COE roles (raw data)", N, align="left")
set_cell(ft, f"H{r}", f"=COUNTIFS({RD}!$Q:$Q,\"COE\",{RD}!$R:$R,\"Filled\")", N, align="center")
set_cell(ft, f"I{r}", f"=COUNTIFS({RD}!$Q:$Q,\"COE\",{RD}!$R:$R,\"Vacant\")", N, align="center")
set_cell(ft, f"J{r}", f"=H{r}+I{r}", N, align="center")
coe_row = r
r += 2
set_cell(ft, f"B{r}", "Cross-check: org records", B, align="left")
set_cell(ft, f"C{r}", f"=C4", B, align="center")
set_cell(ft, f"B{r+1}", "Accounted for in the sections above", B, align="left")
set_cell(ft, f"C{r+1}", f"=J{grand}+C{lead_first-1+len(leads)+1}+SUM(J{un_first}:J{un_last})+J{coe_row}", B, align="center")
ft[f"C{r+1}"] = f"=J{grand}+{len(leads)}+SUM(J{un_first}:J{un_last})+J{coe_row}"
set_cell(ft, f"B{r+2}", "Difference (must be 0)", B, GREY_F, align="left")
set_cell(ft, f"C{r+2}", f"=C{r}-C{r+1}", B, GREY_F, align="center")
pos_bad(ft, f"C{r+2}")

# =====================================================================
# 7. 4.0 Insights
# =====================================================================
ins = wb.create_sheet("4.0 Insights", wb.sheetnames.index("3.0 FTE View") + 1)
ins.sheet_view.showGridLines = False
for col, wdt in {"A":3,"B":30,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16}.items():
    ins.column_dimensions[col].width = wdt
ins.row_dimensions[2].height = 21
set_cell(ins, "B2", "Insights - the position in one page", TITLE, border=False)
GS = "'2.0 Group Summary'"
AD = "'Added data'"
r = 4
set_cell(ins, f"B{r}", "The money", W, DK_F, align="left")
for col in "CDEFGH": set_cell(ins, f"{col}{r}", None, W, DK_F)
r += 1
money = [
 ("Total TDD people budget ($m)", f"={DC}!$E$27"),
 ("Allocated to portfolios + COEs ($m)", f"={GS}!$C$30"),
 ("Unallocated headroom ($m)", f"={DC}!$E$27-{GS}!$C$30"),
 ("Model TDD cost - all portfolios ($m)", f"={GS}!$D$24"),
 ("Model cost funded outside TDD ($m)", f"={GS}!$G$24"),
 ("Total model cost ($m)", f"={GS}!$J$24"),
 ("Full org people cost - Added data ($m)", f"=SUM({AD}!$AA$2:$AA$550)/1000000"),
 ("Archetype cost - all squads ($m)", f"='3.0 FTE View'!$M${grand}"),
]
for label, f in money:
    set_cell(ins, f"B{r}", label, N, align="left")
    set_cell(ins, f"C{r}", f, B, None, MONEY, align="right")
    r += 1
r += 1
set_cell(ins, f"B{r}", "The people", W, DK_F, align="left")
for col in "CDEFGH": set_cell(ins, f"{col}{r}", None, W, DK_F)
r += 1
people = [
 ("Org roles (raw data)", f"=COUNTA({RD}!$R$2:$R$1000)"),
 ("Filled", f"=COUNTIF({RD}!$R$2:$R$1000,\"Filled\")"),
 ("Vacant", f"=COUNTIF({RD}!$R$2:$R$1000,\"Vacant\")"),
 ("Vacancy rate", "SENTINEL_VACRATE"),
 ("Archetype FTE (model)", f"='3.0 FTE View'!$G${grand}"),
 ("Seats over archetype", f"='3.0 FTE View'!$K${grand}"),
 ("Leadership roles (overhead-funded)", len(leads)),
 ("AU-based roles", f"=COUNTIF({RD}!$M$2:$M$1000,\"Australia\")"),
 ("NZ-based roles", f"=COUNTIF({RD}!$M$2:$M$1000,\"NZ\")"),
]
for label, f in people:
    if f == "SENTINEL_VACRATE":
        f = f"=C{r-1}/C{r-3}"
    set_cell(ins, f"B{r}", label, N, align="left")
    set_cell(ins, f"C{r}", f, B, None, PCT if label == "Vacancy rate" else None, align="right")
    r += 1
r += 1
# portfolio selector
set_cell(ins, f"B{r}", "Portfolio drill-down - pick a portfolio", W, DK_F, align="left")
for col in "CDEFGH": set_cell(ins, f"{col}{r}", None, W, DK_F)
r += 1
sel = r
set_cell(ins, f"B{r}", "Portfolio", N, align="left")
set_cell(ins, f"C{r}", "Ampol Retail", B, YELLOW, align="center")
ins[f"C{r}"].font = BLUE
dvp = DataValidation(type="list",
    formula1='"' + ",".join(PORT_LABEL[t] for t in TABS) + '"', allow_blank=True)
ins.add_data_validation(dvp); dvp.add(f"C{r}")
r += 1
drill = [
 ("TDD Lights On budget ($m)", f"=INDEX({GS}!$C$6:$C$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
 ("TDD cost ($m)", f"=INDEX({GS}!$D$6:$D$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
 ("Variance ($m)", f"=INDEX({GS}!$E$6:$E$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
 ("Funded outside TDD ($m)", f"=INDEX({GS}!$G$6:$G$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
 ("Left to fund ($m)", f"=INDEX({GS}!$I$6:$I$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
 ("Total cost ($m)", f"=INDEX({GS}!$J$6:$J$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
 ("Org seats", f"=COUNTIFS({RD}!$N:$N,$C${sel})-COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$Q:$Q,\"Leadership\")", None),
 ("Filled", f"=COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$R:$R,\"Filled\")-COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$Q:$Q,\"Leadership\",{RD}!$R:$R,\"Filled\")", None),
 ("Vacant", f"=COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$R:$R,\"Vacant\")-COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$Q:$Q,\"Leadership\",{RD}!$R:$R,\"Vacant\")", None),
 ("Leadership roles", f"=COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$Q:$Q,\"Leadership\")", None),
]
for label, f, fm in drill:
    set_cell(ins, f"B{r}", label, N, align="left")
    set_cell(ins, f"C{r}", f, B, None, fm, align="right")
    r += 1
vr_row = r - 8
pos_good(ins, f"C{vr_row}")
r += 1
set_cell(ins, f"B{r}", "Where the pressure is", W, DK_F, align="left")
for col in "CDEFGH": set_cell(ins, f"{col}{r}", None, W, DK_F)
r += 1
press = [
 ("Largest portfolio overrun ($m)", f"=MAX(0,-MIN({GS}!$E$6:$E$16))"),
 ("Portfolio", f"=IFERROR(INDEX({GS}!$B$6:$B$16,MATCH(MIN({GS}!$E$6:$E$16),{GS}!$E$6:$E$16,0)),\"-\")"),
 ("Largest left-to-fund ($m)", f"=MAX({GS}!$I$6:$I$16)"),
 ("Portfolio", f"=IFERROR(INDEX({GS}!$B$6:$B$16,MATCH(MAX({GS}!$I$6:$I$16),{GS}!$I$6:$I$16,0)),\"-\")"),
 ("COE planned spend vs allocation ($m)", f"='2.1 COE'!$D$13-'2.1 COE'!$C$13"),
 ("Cyber roles cost vs bucket ($m)", f"='2.4 Cyber Roles'!$C${cyb_tot+6}-'2.4 Cyber Roles'!$C${cyb_tot+5}"),
]
for label, f in press:
    set_cell(ins, f"B{r}", label, N, align="left")
    set_cell(ins, f"C{r}", f, B, None, MONEY if "$m" in label else None, align="right")
    r += 1

# order: ... 2.0, 2.1, 2.2, 2.3, 2.4, 3.0, 4.0, squad mapping, raw data, Added data, Lists
ORDER = ["0.0 Data Config","0.1 Squads","0.2 FY26 Budget","0.3 For Presentation Pack (2)",
         "0.4 Budget Table (Fin)"] + TABS + ["2.0 Group Summary","2.1 COE","2.2 BP&T","2.3 SA&D",
         "2.4 Cyber Roles","3.0 FTE View","4.0 Insights","squad mapping","raw data","Added data","Lists"]
wb._sheets = [wb[s] for s in ORDER if s in wb.sheetnames]

wb.save(OUT)

# ---------------------------------------------------------------------
# restore cached results for 'Added data' formula cells (openpyxl drops
# them on save; the source file carried them, so put them back verbatim)
# ---------------------------------------------------------------------
import re as _re, zipfile, os, shutil
import xml.etree.ElementTree as ET
_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", _NS)
with zipfile.ZipFile(OUT) as zin:
    wbxml = zin.read("xl/workbook.xml").decode()
    relxml = zin.read("xl/_rels/workbook.xml.rels").decode()
    m = _re.search(r'<sheet[^>]*name="Added data"[^>]*r:id="(rId\d+)"', wbxml) or \
        _re.search(r'<sheet[^>]*r:id="(rId\d+)"[^>]*name="Added data"', wbxml)
    rid = m.group(1)
    m2 = _re.search(r'<Relationship[^>]*Id="%s"[^>]*Target="([^"]+)"' % rid, relxml) or \
         _re.search(r'<Relationship[^>]*Target="([^"]+)"[^>]*Id="%s"' % rid, relxml)
    target = m2.group(1).lstrip("/")
    if not target.startswith("xl/"):
        target = "xl/" + target
    tree = ET.fromstring(zin.read(target))
    injected = 0
    for cell in tree.iter("{%s}c" % _NS):
        fe = cell.find("{%s}f" % _NS)
        ve = cell.find("{%s}v" % _NS)
        if fe is not None and (ve is None or not (ve.text or "").strip()):
            cached = adv1[cell.attrib["r"]].value
            if isinstance(cached, (int, float)) and not isinstance(cached, bool):
                if ve is None:
                    ve = ET.SubElement(cell, "{%s}v" % _NS)
                ve.text = repr(cached)
                injected += 1
    newxml = ET.tostring(tree, xml_declaration=True, encoding="UTF-8")
    tmp = OUT + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            zout.writestr(item, newxml if item.filename == target else zin.read(item.filename))
shutil.move(tmp, OUT)
print("cached values injected:", injected)
print("saved", OUT)
print("groups:", {k: len(v) for k, v in groups.items()}, "leads:", len(leads))
print("anchors:", {t: a for t, a in anchors.items()})
print("3.0 grand row:", grand)
