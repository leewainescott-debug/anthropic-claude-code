#!/usr/bin/env python3
"""Merge best-of-both: base = WB-A (user's latest) + fixes/uplift.
 1. purge junk defined names
 2. 1.1: re-align spare-row SUM coverage (H already spans rows 35/42)
 3. funding blocks: IFERROR guards; re-point 1.3/1.4/1.11 to their own
    Finance sub-rows (people element restored); 1.10 sig-items ref restored
 4. 0.4: complete the live conversion the user started (table 2 + links)
 5. add 2.0 Group Summary (incl. COE) rebuilt on WB-A styles
 6. sheet order, conditional formatting (position colours), legends
"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from copy import copy

U = "/root/.claude/uploads/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/"
SRC = U + "086bb1b3-TDD_Cost_Calculator_.xlsx"
OUT = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/TDD_Cost_Calc_final.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=False)
tpl = wb["1.1 Ampol Retail"]

# ---------- 1. purge junk names ----------
KEEP = {"SquadTypes", "SquadSizes", "OnOff", "SupportPct"}
removed = 0
for n in list(wb.defined_names.keys()):
    if n not in KEEP:
        del wb.defined_names[n]; removed += 1
for ws_ in wb.worksheets:
    for n in list(getattr(ws_, "defined_names", {}).keys()):
        del ws_.defined_names[n]; removed += 1
print("names removed:", removed)

# ---------- 2. 1.1 spare-row SUM alignment ----------
tpl["C10"] = "=SUM(H25,H26,H27,H28,H34,H35,H41,H42,H48,H54,H60)"
tpl["D10"] = "=SUM(I25,I26,I27,I28,I34,I35,I41,I42,I48,I54,I60)"
tpl["G37"] = "=SUM(G34:G35)"; tpl["I37"] = "=SUM(I34:I35)"
tpl["G44"] = "=SUM(G41:G42)"; tpl["I44"] = "=SUM(I41:I42)"
tpl["B38"] = None  # stray note "Squad type is an exception" no longer applies
tpl["C10"].comment = Comment(
    "Includes the spare squad rows 35 and 42 so anything typed there is counted "
    "everywhere (the H-column platform totals already spanned them).", "QA")

# ---------- 3. funding blocks ----------
FIN = "'0.4 Budget Table (Fin)'"
# IFERROR guards on all J14:J18 across 1.x tabs
tabs_1x = [s for s in wb.sheetnames if s.startswith("1.")]
for t in tabs_1x:
    ws = wb[t]
    top = 14
    for r in range(14, 19):
        jc = ws[f"J{r}"]
        if isinstance(jc.value, str) and jc.value.startswith("=") and "IFERROR" not in jc.value:
            jc.value = f"=IFERROR({jc.value[1:]},0)"
# 1.1 funding rows are 14-17 (no dep line) - guard those too
for r in range(14, 18):
    jc = wb["1.1 Ampol Retail"][f"J{r}"]
    if isinstance(jc.value, str) and jc.value.startswith("=") and "IFERROR" not in jc.value:
        jc.value = f"=IFERROR({jc.value[1:]},0)"

# 1.10: restore sig-items ref (displays Finance's '-'; J17 guarded)
zr = wb["1.10 Z Retail"]
zr["H17"] = f"={FIN}!$O$6"

# Re-point the three TDD-internal tabs to their own Finance sub-rows
REPOINT = {
 "1.3 Enterprise Data": dict(row=27, label="Enterprise Data",
   note="Finance sub-row 'Strategy, Architecture and Data' (Matt Ashley) within TDD Corporate - "
        "the closest line for Enterprise Data. The whole TDD Corporate segment (112.8) was "
        "previously shown here, identically on 1.3/1.4/1.11, which triple-displays one budget "
        "and loses the people element."),
 "1.4 TDD Group Functions": dict(row=25, label="TDD Group Functions",
   note="Finance sub-row 'TDD' (Emily Mogic) within TDD Corporate - enterprise tooling/SW-heavy, "
        "matching Workplace & Enterprise Tooling, Network & Infrastructure, DevOps and Integration."),
 "1.11 TDD Cyber": dict(row=28, label="TDD Cyber",
   note="Finance sub-row 'Cyber Risk and Operations' (Frances Bouzo) within TDD Corporate. "
        "Wider than the TDD Cyber squad alone."),
}
for t, cfg in REPOINT.items():
    ws = wb[t]; r = cfg["row"]
    ws["B13"] = f"Total {cfg['label']} budget"
    ws["G14"] = f"{cfg['label']} Lights On"
    ws["E14"] = f"={FIN}!$S${r}"
    ws["H14"] = f"={FIN}!$N${r}"
    ws["I14"] = f"={FIN}!$I${r}"
    ws["I14"]._style = copy(ws["H14"]._style)   # green link style like other refs
    ws["H15"] = f"={FIN}!$P${r}"
    ws["H16"] = f"={FIN}!$R${r}"
    ws["H17"] = f"={FIN}!$Q${r}"
    ws["H18"] = f"={FIN}!$O${r}"
    ws["E14"].comment = Comment(cfg["note"], "QA")

# ---------- 4. 0.4 complete live conversion ----------
fin = wb["0.4 Budget Table (Fin)"]
# table 2 within-row totals (rows 19-29): N = SUM(I:M), S = SUM(N:R)
for r in range(19, 30):
    if fin[f"N{r}"].value is not None:
        fin[f"N{r}"] = f"=SUM(I{r}:M{r})"
    if fin[f"S{r}"].value is not None:
        fin[f"S{r}"] = f"=SUM(N{r}:R{r})"
# COE rows: W (Aligned COE's) = S ; X (Total Budget) = W
for r in range(25, 30):
    if fin[f"W{r}"].value is not None:
        fin[f"W{r}"] = f"=S{r}"
    if fin[f"X{r}"].value is not None:
        fin[f"X{r}"] = f"=W{r}"
# table 1 right side: Existing = FY26 total; Aligned Delivery = table-2 rows; U5 = COE sum; V = S+T+U
for r in range(6, 14):
    if fin[f"S{r}"].value is not None:
        fin[f"S{r}"] = f"=Q{r}"
T_MAP = {8: 19, 9: 20, 10: 21, 11: 22, 12: 23, 13: 24}   # table1 row -> table2 aligned-delivery row
for r1, r2 in T_MAP.items():
    if fin[f"T{r1}"].value is not None:
        fin[f"T{r1}"] = f"=S{r2}"
fin["U5"] = "=SUM(W25:W29)"
for r in range(5, 14):
    if fin[f"V{r}"].value is not None:
        fin[f"V{r}"] = f"=SUM(S{r}:U{r})"
fin["S14"] = "=SUM(S6:S13)"; fin["T14"] = "=SUM(T8:T13)"
fin["U14"] = "=U5"; fin["V14"] = "=SUM(V5:V13)"
# echo rows 30 (TDD Corporate) and 33 (Z-Energy) mirror table-1 rows 5 / 6
ECHO = dict(I="G", J="H", K="I", L="J", M="K", N="L", O="M", P="N", Q="O", R="P", S="Q")
for dst, src in ECHO.items():
    if fin[f"{dst}30"].value is not None:
        fin[f"{dst}30"] = f"={src}5"
    if fin[f"{dst}33"].value is not None:
        fin[f"{dst}33"] = f"={src}6"
fin["W30"] = "=U5"; fin["X30"] = "=W30"
fin["U33"] = "=S6"; fin["X33"] = "=U33"
fin["B35"] = ("Note: totals are live sums of the component cells. The printed Finance pack "
              "rounds to $0.1m, so differences of up to ±$0.1m vs the pack are expected.")
fin["B35"].font = Font(name="Calibri", size=9, italic=True, color="FF808080")

# ---------- 5. rebuild 2.0 Group Summary on WB-A base ----------
def S(ws, dst, src, value=None):
    d = ws[dst]; d._style = copy(tpl[src]._style)
    if value is not None: d.value = value
    return d

gs = wb.create_sheet("2.0 Group Summary")
gs.sheet_view.showGridLines = False
for col, w in {"A": 8.6, "B": 32, "C": 20, "D": 16, "E": 16, "F": 20, "G": 16}.items():
    gs.column_dimensions[col].width = w
gs.row_dimensions[2].height = 21
S(gs, "B2", "B2", "Group Summary - all portfolios")
COLL = "BCDEFG"
for j, col in enumerate(COLL):
    gs[f"{col}4"]._style = copy(tpl["B6"]._style)
gs["B4"] = "TDD cost vs TDD lights-on budget, by portfolio"
gs.merge_cells("B4:G4")
for j, h in enumerate(["Portfolio", "TDD Lights On Budget ($m)", "TDD Cost ($m)",
                       "Variance ($m)", "Funded outside TDD ($m)", "Total Cost ($m)"]):
    S(gs, f"{COLL[j]}5", "B7", h)
PORTS = [("1.1 Ampol Retail","Ampol Retail"),("1.2 Customer","Customer"),
         ("1.3 Enterprise Data","Enterprise Data"),("1.4 TDD Group Functions","TDD Group Functions"),
         ("1.5 P&C","P&C"),("1.6 Finance","Finance"),("1.7 Infrastructure","Infrastructure"),
         ("1.8 Energy Solutions & B2B","Energy Solutions & B2B"),("1.9 Commercial Fuels","Commercial Fuels"),
         ("1.10 Z Retail","Z Retail"),("1.11 TDD Cyber","TDD Cyber")]
r = 6
for tab, name in PORTS:
    S(gs, f"B{r}", "B8", name)
    S(gs, f"C{r}", "H25", f"='{tab}'!$H$7")
    S(gs, f"D{r}", "H25", f"='{tab}'!$C$11")
    S(gs, f"E{r}", "H25", f"='{tab}'!$H$9")
    S(gs, f"F{r}", "H25", f"='{tab}'!$D$11")
    S(gs, f"G{r}", "H25", f"='{tab}'!$E$11")
    r += 1
last = r - 1
S(gs, f"B{r}", "B11", "Total - all portfolios")
for col in "CDEFG":
    S(gs, f"{col}{r}", "C11", f"=SUM({col}6:{col}{last})")
tot = r
r += 3
for col in COLL:
    gs[f"{col}{r}"]._style = copy(tpl["B6"]._style)
gs[f"B{r}"] = "Reconciliation to Data Config TDD budget allocation"
gs.merge_cells(f"B{r}:G{r}")
r += 1
S(gs, f"B{r}", "B8", "Portfolio budgets (above)"); S(gs, f"C{r}", "H25", f"=C{tot}")
r += 1
S(gs, f"B{r}", "B8", "COE allocations (2.1 COE)"); S(gs, f"C{r}", "C8", "='2.1 COE'!$C$13")
r += 1
S(gs, f"B{r}", "B11", "Total allocation accounted for"); S(gs, f"C{r}", "C11", f"=C{r-2}+C{r-1}")
r += 1
S(gs, f"B{r}", "B8", "Data Config total allocation"); S(gs, f"C{r}", "C8", "='0.0 Data Config'!$E$26")
r += 1
S(gs, f"B{r}", "B11", "Check (should be 0)"); S(gs, f"C{r}", "C11", f"=C{r-2}-C{r-1}")
check_row = r

# ---------- 6. conditional formatting + legends ----------
GREEN_F = PatternFill("solid", fgColor="FFE2EFDA"); GREEN_T = Font(color="FF006100", name="Calibri", size=10, bold=True)
RED_F   = PatternFill("solid", fgColor="FFFBE4D5"); RED_T   = Font(color="FF9C0006", name="Calibri", size=10, bold=True)
def pos_good(ws, rng):   # >=0 green, <0 red   (variance)
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_F, font=GREEN_T))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_F, font=RED_T))
def pos_bad(ws, rng):    # >0 red, <=0 green   (left to fund)
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=RED_F, font=RED_T))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=GREEN_F, font=GREEN_T))

LEGEND = "Yellow = your selection   |   Green text = pulled from another tab   |   Grey = calculated   |   Red / green = position"
for t in tabs_1x:
    ws = wb[t]
    pos_good(ws, "H9")
    ltf = "I20" if t == "1.1 Ampol Retail" else "I21"
    pos_bad(ws, ltf)
    ws["B4"] = LEGEND
    ws["B4"].font = Font(name="Calibri", size=8, italic=True, color="FF808080")
pos_good(gs, f"E6:E{tot}")
gs[f"C{check_row}"].comment = Comment("0 means every dollar of the Data Config allocation is represented in a tab.", "QA")
pos_bad(gs, f"C{check_row}")

# ---------- 7. sheet order ----------
ORDER = ["0.0 Data Config","0.1 Squads","0.2 FY26 Budget","0.3 For Presentation Pack (2)",
         "0.4 Budget Table (Fin)","1.1 Ampol Retail","1.2 Customer","1.3 Enterprise Data",
         "1.4 TDD Group Functions","1.5 P&C","1.6 Finance","1.7 Infrastructure",
         "1.8 Energy Solutions & B2B","1.9 Commercial Fuels","1.10 Z Retail","1.11 TDD Cyber",
         "2.0 Group Summary","2.1 COE","squad mapping","Lists"]
wb._sheets = [wb[s] for s in ORDER]

wb.save(OUT)
print("saved", OUT)
