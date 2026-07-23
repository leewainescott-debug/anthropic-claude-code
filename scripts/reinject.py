#!/usr/bin/env python3
"""Clear the stray REVIEW R445 #N/A, then inject the already-computed values
(tie_out_values.json) into a populated copy. No engine re-run."""
import json, zipfile, shutil, os
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils import column_index_from_string as ci

SRC="clean_v6.xlsx"; MID="clean_v6b.xlsx"; POP="clean_v6_pop.xlsx"
vals=json.load(open("tie_out_values.json"))          # {SHEET_UPPER: {COORD: value}}

wb=openpyxl.load_workbook(SRC)
r445=wb["REVIEW - Complete Role Mapping"].cell(445, ci("R"))
if isinstance(r445.value,str) and (r445.value.startswith("=") or r445.value.strip()=="#N/A"):
    r445.value=None   # stray #N/A artifact in the otherwise-blank Unit column; feeds nothing
wb.save(MID)

# formula cells per sheet -> computed value
fcells={}
for ws in wb.worksheets:
    st=ws.title.upper(); d={}
    sv=vals.get(st,{})
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value.startswith("=") and c.coordinate in sv:
                d[c.coordinate]=sv[c.coordinate]
    if d: fcells[st]=d

# recon tab changed to Occupied = roles - vacant; recompute D/G/J from the (unchanged)
# roles/vacant counts so the injected cached values match the new formulas.
RC="3.5 SOURCE RECONCILIATION"
if RC in fcells:
    d=fcells[RC]
    def n(x): return x if isinstance(x,(int,float)) and not isinstance(x,bool) else None
    for r in range(6,20):   # data rows 6-19
        C,E,F,H=n(d.get(f"C{r}")),n(d.get(f"E{r}")),n(d.get(f"F{r}")),n(d.get(f"H{r}"))
        if C is not None and E is not None: d[f"D{r}"]=C-E
        if F is not None and H is not None: d[f"G{r}"]=F-H
        if d.get(f"D{r}") is not None and d.get(f"G{r}") is not None: d[f"J{r}"]=d[f"G{r}"]-d[f"D{r}"]
    for col in "DGJ":       # total row 20 re-summed
        d[f"{col}20"]=sum(d[f"{col}{r}"] for r in range(6,20) if isinstance(d.get(f"{col}{r}"),(int,float)))
print("sheets with values:",len(fcells),"| total formula cells:",sum(len(d) for d in fcells.values()))

NS="http://schemas.openxmlformats.org/spreadsheetml/2006/main"; NR="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("",NS); ET.register_namespace("r",NR)
ERR={"#REF!","#DIV/0!","#VALUE!","#N/A","#NAME?","#NULL!","#NUM!","#CYCLE!"}
tmp="_re"
if os.path.exists(tmp): shutil.rmtree(tmp)
os.makedirs(tmp)
with zipfile.ZipFile(MID) as z: z.extractall(tmp)
wbt=ET.parse(f"{tmp}/xl/workbook.xml"); wbr=wbt.getroot()
n2r={sh.get("name"):sh.get(f"{{{NR}}}id") for sh in wbr.iter(f"{{{NS}}}sheet")}
rels=ET.parse(f"{tmp}/xl/_rels/workbook.xml.rels").getroot()
r2t={rel.get("Id"):rel.get("Target") for rel in rels}
def resolve(tgt):
    t=tgt.lstrip("/")
    for p in (f"{tmp}/xl/{t}", f"{tmp}/{t}", f"{tmp}/xl/{t.split('xl/')[-1]}"):
        if os.path.exists(p): return p
    return None
def numstr(x): return "1" if x is True else "0" if x is False else repr(float(x))
inj=0
for name,rid in n2r.items():
    d=fcells.get((name or "").upper(),{})
    if not d: continue
    tgt=r2t.get(rid)
    path=resolve(tgt) if tgt else None
    if not path: print("  no path for",name,tgt); continue
    tr=ET.parse(path); rt=tr.getroot()
    for c in rt.iter(f"{{{NS}}}c"):
        co=c.get("r")
        if co not in d: continue
        fe=c.find(f"{{{NS}}}f")
        if fe is None: continue
        x=d[co]
        for old in c.findall(f"{{{NS}}}v"): c.remove(old)
        for old in c.findall(f"{{{NS}}}is"): c.remove(old)
        ve=ET.SubElement(c,f"{{{NS}}}v")
        if isinstance(x,str) and x in ERR: c.set("t","e"); ve.text=x
        elif isinstance(x,bool): c.set("t","b"); ve.text=numstr(x)
        elif isinstance(x,str): c.set("t","str"); ve.text=x
        else:
            if c.get("t") in ("str","e","b","s"): del c.attrib["t"]
            try: ve.text=numstr(x)
            except: c.set("t","str"); ve.text=str(x)
        c.remove(fe); c.insert(0,fe); inj+=1
    tr.write(path, xml_declaration=True, encoding="UTF-8")
cp=wbr.find(f"{{{NS}}}calcPr")
if cp is None: cp=ET.SubElement(wbr,f"{{{NS}}}calcPr")
cp.set("fullCalcOnLoad","1"); wbt.write(f"{tmp}/xl/workbook.xml", xml_declaration=True, encoding="UTF-8")
if os.path.exists(POP): os.remove(POP)
with zipfile.ZipFile(POP,"w",zipfile.ZIP_DEFLATED) as z:
    for base,_,files in os.walk(tmp):
        for fn in files: z.write(os.path.join(base,fn), os.path.relpath(os.path.join(base,fn),tmp))
shutil.rmtree(tmp)
print("injected",inj,"cached values ->",POP)