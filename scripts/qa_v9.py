#!/usr/bin/env python3
"""Quadruple-check QA for v9.
PASS 1: independent python recompute straight from Sheet2 / Squads / Added data.
PASS 2: full formula-engine evaluation - zero formula errors on model tabs.
PASS 3: assertion battery + second engine run with the offshore factor flipped.
"""
import openpyxl, re, json, gc, sys, logging
from openpyxl.utils import get_column_letter
logging.getLogger().setLevel(logging.ERROR)
SCR = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/"
SRC = SCR + "TDD_Cost_Calc_v9.xlsx"
A = json.load(open(SCR + "anchors_v9.json"))
fails = []
def chk(label, cond, detail=""):
    if not cond:
        fails.append(f"{label} :: {detail}")
        print("FAIL", label, detail)

wb = openpyxl.load_workbook(SRC, data_only=False)
def low(v): return str(v).strip().lower() if v is not None else ""

# ---------------- PASS 1: independent recompute ----------------
s2 = wb["Sheet2"]
from openpyxl.utils import column_index_from_string
REF = re.compile(r"\$?([A-Z]{1,2})\$?(\d+)")
def eval_cell(ws, col, row, depth=0):
    """evaluate a plain-arithmetic cell (numbers, same-sheet refs, + - * / parens)"""
    v = ws.cell(row, col).value
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    if not (isinstance(v, str) and v.startswith("=")): return 0.0
    if depth > 4 or ":" in v: return 0.0
    expr = REF.sub(lambda m: repr(eval_cell(ws, column_index_from_string(m.group(1)),
                                            int(m.group(2)), depth + 1)), v[1:])
    if not re.fullmatch(r"[0-9eE+\-*/(). ]+", expr): return 0.0
    return float(eval(expr, {"__builtins__": {}}, {}))
def cost_row(ws, r, ca=27):
    """recompute Full Cost AUD from the row's own inputs by evaluating the
    row's own formula - no shape assumptions"""
    return eval_cell(ws, ca, r)
rows2 = []
for r in range(2, s2.max_row + 1):
    if s2.cell(r, 6).value is None: continue
    typ = low(s2.cell(r, 17).value)
    status = ("Vacant" if typ == "v" else "Paused" if typ == "pause"
              else "Contractor" if typ == "cxc" else "Filled")
    rows2.append(dict(r=r, div=str(s2.cell(r, 6).value).strip(),
                      dept=str(s2.cell(r, 7).value or "").strip(),
                      port=low(s2.cell(r, 9).value), squad=low(s2.cell(r, 11).value),
                      status=status, cost=cost_row(s2, r)))
CYB = [x for x in rows2 if x["div"] == "Cyber, Risk & Operations"]
EGI = [x for x in rows2 if x["div"] == "EGI"]
PT  = [x for x in rows2 if x["div"] == "Partnering & Transformation"]
SAD = [x for x in rows2 if x["div"] == "Strategy, Architecture & Data"]
chk("p1.divisions", (len(CYB), len(EGI), len(PT), len(SAD)) == (52, 16, 24, 60),
    f"{len(CYB)},{len(EGI)},{len(PT)},{len(SAD)}")
# expected rosters (must mirror patch_v9 rules)
sq = wb["Squads"]
sq_rows = []
for r in range(2, sq.max_row + 1):
    if sq.cell(r, 2).value is None: continue
    sq_rows.append(dict(r=r, name=low(sq.cell(r, 2).value), title=low(sq.cell(r, 3).value),
                        div=low(sq.cell(r, 6).value), N=str(sq.cell(r, 14).value or "").strip(),
                        Q=str(sq.cell(r, 17).value or "").strip(),
                        R=str(sq.cell(r, 18).value or "").strip()))
def sq_match(x):
    nm, tt = low(s2.cell(x["r"], 2).value), low(s2.cell(x["r"], 3).value)
    c = [q for q in sq_rows if q["name"] == nm and q["title"] == tt]
    d = [q for q in c if q["div"] == low(x["div"])]
    return (d or c or [None])[0]
def blankish(v): return v in ("", "na")
def s2team(r): return str(s2.cell(r, 8).value or "").strip()
def s2fte(r):
    v = s2.cell(r, 15).value
    return float(v) if isinstance(v, (int, float)) else 1.0
sad_coe = []
for x in SAD:
    d = x["dept"]
    coe = (d in ("Architecture", "Technology Strategy & AI Capability", "na")
           or (d == "Delivery, SADA" and (x["squad"] in ("", "na") or s2fte(x["r"]) >= 1.0))
           or (d == "Group Data" and (s2team(x["r"]) == "Data Capability"
                                      or str(s2.cell(x["r"], 3).value or "").startswith("Head of Technology"))))
    if coe: sad_coe.append(x)
chk("p1.sad_coe_n", len(sad_coe) == A["N_SAD_COE"], f"{len(sad_coe)} vs {A['N_SAD_COE']}")
chk("p1.sad_covers_techstrategy", any(x["dept"] == "Technology Strategy & AI Capability" for x in sad_coe))
chk("p1.sad_covers_architecture", sum(1 for x in sad_coe if x["dept"] == "Architecture") >= 7)
def bucket_pt(x): return "Transformation" if x["dept"] == "Transformation" else "Business Partnering"
def bucket_sad(x): return "Data" if x["dept"] == "Group Data" else "Strategy & Architecture"
def bucket_cy(x): return "Service Operations" if x["dept"] == "Service Op & Assurance" else "Cyber & Risk"
def agg(rows, bucket, paused_zero):
    out = {}
    for x in rows:
        b = bucket(x)
        d = out.setdefault(b, dict(n=0, f=0, v=0, p=0, spend=0.0))
        d["n"] += 1
        if x["status"] == "Filled": d["f"] += 1
        elif x["status"] == "Vacant": d["v"] += 1
        elif x["status"] == "Paused": d["p"] += 1
        if not (paused_zero and x["status"] == "Paused"):
            d["spend"] += x["cost"] / 1e6
    return out
p1_pt = agg(PT, bucket_pt, False)
p1_sad = agg(sad_coe, bucket_sad, True)
p1_cy = agg(CYB, bucket_cy, False)
p1_egi = sum(x["cost"] for x in EGI) / 1e6
# config-derived budgets
cfg = wb["0.2 Data Config"]
def num(ws, addr):
    v = ws[addr].value
    return float(v) if isinstance(v, (int, float)) else None
L7 = num(cfg, "J7") * num(cfg, "K7")
L8 = num(cfg, "J8") * num(cfg, "K8")
E6c = num(cfg, "C6") + num(cfg, "D6"); E8c = num(cfg, "C8") + num(cfg, "D8")
E9c = num(cfg, "C9") + num(cfg, "D9"); E10c = num(cfg, "C10") + num(cfg, "D10")
g20 = wb["3.1 Group Summary"]
nport = sum(1 for r in range(6, 17) if g20.cell(r, 2).value not in (None, ""))
chk("p1.nport", nport == 10, nport)
BP_BUDGET = E9c + nport * L7
TR_BUDGET = E8c
SA_BUDGET = E6c + nport * L8
DA_DRAW = nport * L8
DATA_BUDGET = E10c
# Added data ledger recompute (portfolio actuals)
ad = wb["Added data"]
led = {}
ledger_total = 0.0
for r in range(2, ad.max_row + 1):
    if ad.cell(r, 2).value is None and ad.cell(r, 3).value is None: continue
    c = cost_row(ad, r, 27)
    ledger_total += c
    port = str(ad.cell(r, 29).value or "").strip()   # AC
    cls = str(ad.cell(r, 32).value or "").strip()    # AF
    st = str(ad.cell(r, 33).value or "").strip()     # AG
    led.setdefault((port, cls, st), [0, 0.0])
    led[(port, cls, st)][0] += 1
    led[(port, cls, st)][1] += c
def led_sum(port, sts):
    tot = 0.0
    for cls in ("Squad", "Strategic Program", "Leadership"):
        tot += led.get((port, cls, sts), [0, 0])[1]
    return tot / 1e6
# Squads counts per portfolio
def sq_cnt(port, sts):
    return sum(1 for q in sq_rows if q["N"] == port and q["R"] == sts
               and q["Q"] in ("Squad", "Strategic Program", "Leadership"))
print(f"PASS1: pt={ {k: round(v['spend'],3) for k,v in p1_pt.items()} } "
      f"sad={ {k: round(v['spend'],3) for k,v in p1_sad.items()} } "
      f"cy={ {k: round(v['spend'],3) for k,v in p1_cy.items()} } egi={p1_egi:.3f} "
      f"BP_budget={BP_BUDGET:.3f} SA_budget={SA_BUDGET:.3f} ledger={ledger_total/1e6:.3f}")

# ---------------- PASS 2 + engine values ----------------
import formulas
def engine(path):
    xl = formulas.ExcelModel().loads(path).finish()
    sol = xl.calculate()
    vals = {}
    for k, v in sol.items():
        m = re.match(r"^'?\[[^\]]*\]([^!]*?)'?!([A-Z]+\d+)$", k)
        if not m: continue
        val = v.value
        try: val = val[0, 0]
        except Exception: pass
        vals[(m.group(1).strip().upper(), m.group(2))] = val
    del xl, sol
    gc.collect()
    return vals
vals = engine(SRC)
ERRTOK = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!")
errs = [(k, v) for k, v in vals.items() if isinstance(v, str) and v in ERRTOK
        and k[0] not in ("SQUADS", "ADDED DATA", "SHEET2")]
chk("p2.zero_engine_errors", len(errs) == 0, str(errs[:8]))
def g(sheet, cell):
    v = vals.get((sheet.upper(), cell))
    return v
def gn(sheet, cell):
    v = g(sheet, cell)
    return float(v) if isinstance(v, (int, float)) else float("nan")
def close(a, b, tol=0.005): return abs(a - b) <= tol

# ---------------- PASS 3: assertions ----------------
# 1.11 BP&T
chk("bpt.nport", gn("1.11 BP&T", "C10") == nport)
chk("bpt.fte_funded", close(gn("1.11 BP&T", "C12"), nport * num(cfg, "K7"), 1e-9))
chk("bpt.draw", close(gn("1.11 BP&T", "C13"), nport * L7))
chk("bpt.budget_both", close(gn("1.11 BP&T", "C15"), BP_BUDGET))
chk("bpt.g6_budget", close(gn("1.11 BP&T", "G6"), BP_BUDGET))
chk("bpt.g7_budget", close(gn("1.11 BP&T", "G7"), TR_BUDGET))
for cat, col in (("Business Partnering", 6), ("Transformation", 7)):
    d = p1_pt[cat]
    chk(f"bpt.{col}.roles", gn("1.11 BP&T", f"C{col}") == d["n"], f"{g('1.11 BP&T', f'C{col}')} vs {d['n']}")
    chk(f"bpt.{col}.filled", gn("1.11 BP&T", f"D{col}") == d["f"])
    chk(f"bpt.{col}.vacant", gn("1.11 BP&T", f"E{col}") == d["v"])
    chk(f"bpt.{col}.spend", close(gn("1.11 BP&T", f"F{col}"), d["spend"]), f'{g("1.11 BP&T", f"F{col}")} vs {d["spend"]}')
    chk(f"bpt.{col}.aunz_foot", close(gn("1.11 BP&T", f"I{col}") + gn("1.11 BP&T", f"J{col}"),
        gn("1.11 BP&T", f"F{col}")))
    chk(f"bpt.{col}.left", close(gn("1.11 BP&T", f"H{col}"),
        max(0.0, d["spend"] - (BP_BUDGET if col == 6 else TR_BUDGET))))
chk("bpt.check0", gn("1.11 BP&T", f"C{A['PT_CHECK']}") == 0)
# 1.12 SA&D
chk("sad.nport", gn("1.12 SA&D", "C10") == nport)
chk("sad.fte_funded", close(gn("1.12 SA&D", "C12"), nport * num(cfg, "K8"), 1e-9))
chk("sad.draw", close(gn("1.12 SA&D", "C13"), DA_DRAW))
chk("sad.budget_both", close(gn("1.12 SA&D", "C15"), SA_BUDGET))
chk("sad.h6_budget", close(gn("1.12 SA&D", "H6"), SA_BUDGET))
chk("sad.h7_budget", close(gn("1.12 SA&D", "H7"), DATA_BUDGET))
for cat, col in (("Strategy & Architecture", 6), ("Data", 7)):
    d = p1_sad.get(cat, dict(n=0, f=0, v=0, p=0, spend=0.0))
    chk(f"sad.{col}.roles", gn("1.12 SA&D", f"C{col}") == d["n"], f"{g('1.12 SA&D', f'C{col}')} vs {d['n']}")
    chk(f"sad.{col}.filled", gn("1.12 SA&D", f"D{col}") == d["f"])
    chk(f"sad.{col}.vacant", gn("1.12 SA&D", f"E{col}") == d["v"])
    chk(f"sad.{col}.paused", gn("1.12 SA&D", f"F{col}") == d["p"])
    chk(f"sad.{col}.spend", close(gn("1.12 SA&D", f"G{col}"), d["spend"]), f'{g("1.12 SA&D", f"G{col}")} vs {d["spend"]}')
    chk(f"sad.{col}.aunz_foot", close(gn("1.12 SA&D", f"J{col}") + gn("1.12 SA&D", f"K{col}"),
        gn("1.12 SA&D", f"G{col}")))
paused_cost = sum(x["cost"] for x in sad_coe if x["status"] == "Paused") / 1e6
chk("sad.paused_memo", close(gn("1.12 SA&D", "C18"), paused_cost), f'{g("1.12 SA&D", "C18")} vs {paused_cost}')
chk("sad.check0", gn("1.12 SA&D", f"C{A['SAD_CHECK']}") == 0)
chk("sad.roles_total", gn("1.12 SA&D", "C8") == len(sad_coe))
# 2.5 Cyber
for cat, col in (("Cyber & Risk", 6), ("Service Operations", 7)):
    d = p1_cy[cat]
    chk(f"cy.{col}.roles", gn("1.13 Cyber Roles", f"C{col}") == d["n"])
    chk(f"cy.{col}.filled", gn("1.13 Cyber Roles", f"D{col}") == d["f"])
    chk(f"cy.{col}.vacant", gn("1.13 Cyber Roles", f"E{col}") == d["v"])
    chk(f"cy.{col}.spend", close(gn("1.13 Cyber Roles", f"F{col}"), d["spend"]), f'{g("1.13 Cyber Roles", f"F{col}")} vs {d["spend"]}')
chk("cy.total52", gn("1.13 Cyber Roles", "C8") == 52)
chk("cy.check0", gn("1.13 Cyber Roles", f"C{A['CY_CHECK']}") == 0)
CYROW = A["COE_FIRST"] + 4
chk("cy.coe_row_26", close(gn("3.2 Total Cost", f"C{CYROW}"), gn("1.13 Cyber Roles", "F8")),
    f"{g('3.2 Total Cost', 'C' + str(CYROW))} vs {g('1.13 Cyber Roles', 'F8')}")
chk("cy.buckets", close(gn("1.13 Cyber Roles", "C14"),
    gn("0.2 Data Config", "E7") + gn("0.2 Data Config", "E23") + 0.5))
chk("cy.grid_24", close(gn("3.4 COE Summary", "F10"), gn("1.13 Cyber Roles", "F8")))
chk("cy.group_row19", close(gn("3.1 Group Summary", "D19"), gn("1.13 Cyber Roles", "F8")))
# 3.4 COE Summary
chk("coe.grid_bp", close(gn("3.4 COE Summary", "F6"), gn("1.11 BP&T", "F6")))
chk("coe.grid_tr", close(gn("3.4 COE Summary", "F7"), gn("1.11 BP&T", "F7")))
chk("coe.grid_sa", close(gn("3.4 COE Summary", "F8"), gn("1.12 SA&D", "G6")))
chk("coe.grid_data", close(gn("3.4 COE Summary", "F9"), gn("1.12 SA&D", "G7")))
chk("coe.grid_budget_bp", close(gn("3.4 COE Summary", "G6"), BP_BUDGET))
chk("coe.grid_budget_sa", close(gn("3.4 COE Summary", "G8"), SA_BUDGET))
# 2.1 consolidated
DED, TOT, UNM, LEAD, RESTATE = A["DEDUP"], A["TOT"], A["UNM"], A["LEAD"], A["RESTATE"]
ports_order = ["1.1 Ampol Retail", "1.2 Customer", "1.3 Enterprise Data", "1.4 TDD Group Functions",
               "1.5 P&C", "1.6 Finance", "1.7 Infrastructure", "1.8 Energy Solutions & B2B",
               "1.9 Commercial Fuels", "1.10 Z Retail"]
ecell = {"1.7 Infrastructure": "F10"}
arch_sum = 0.0
for i, tab in enumerate(ports_order):
    r = 6 + i
    pv = gn(tab, ecell.get(tab, "F9"))
    arch_sum += pv
    chk(f"t21.{r}.arch", close(gn("3.2 Total Cost", f"C{r}"), pv), f"{g('3.2 Total Cost', f'C{r}')} vs {pv}")
    port = g("3.2 Total Cost", f"B{r}")
    chk(f"t21.{r}.filled_$", close(gn("3.2 Total Cost", f"K{r}"), led_sum(port, "Filled")),
        f"{g('3.2 Total Cost', f'K{r}')} vs {led_sum(port, 'Filled')}")
    chk(f"t21.{r}.vac_$", close(gn("3.2 Total Cost", f"L{r}"), led_sum(port, "Vacant")))
    chk(f"t21.{r}.filled_fte", gn("3.2 Total Cost", f"I{r}") == sq_cnt(port, "Filled"))
    chk(f"t21.{r}.vac_fte", gn("3.2 Total Cost", f"M{r}") == sq_cnt(port, "Vacant"))
    chk(f"t21.{r}.i", close(gn("3.2 Total Cost", f"D{r}"),
                            gn("3.2 Total Cost", f"K{r}") + gn("3.2 Total Cost", f"L{r}")))
    chk(f"t21.{r}.k", close(gn("3.2 Total Cost", f"E{r}"),
                            gn("3.2 Total Cost", f"D{r}") - gn("3.2 Total Cost", f"C{r}")))
chk("t21.dedup", close(gn("3.2 Total Cost", f"C{DED}"), -(nport * L7 + nport * L8)))
csum = sum(gn("3.2 Total Cost", f"C{r}") for r in range(6, TOT))
chk("t21.total_c", close(gn("3.2 Total Cost", f"C{TOT}"), csum))
chk("t21.total_i", close(gn("3.2 Total Cost", f"D{TOT}"),
                         sum(gn("3.2 Total Cost", f"D{r}") for r in range(6, TOT))))
chk("t21.restate", close(gn("3.2 Total Cost", f"D{RESTATE}"),
                         gn("3.2 Total Cost", f"D{TOT}") - ledger_total / 1e6, 0.01),
    f"{g('3.2 Total Cost', 'D' + str(RESTATE))} vs {gn('3.2 Total Cost', 'D' + str(TOT)) - ledger_total / 1e6}")
chk("t21.egi_memo", close(gn("3.2 Total Cost", f"I{A['EGI_MEMO']}"), p1_egi), f"{g('3.2 Total Cost', 'I' + str(A['EGI_MEMO']))} vs {p1_egi}")
chk("t30.egi_total", close(gn("3.3 FTE View", f"F{A['EGI_TOT']}"), p1_egi), f"{g('3.3 FTE View', 'F' + str(A['EGI_TOT']))} vs {p1_egi}")
chk("t30.egi_n", gn("3.3 FTE View", f"C{A['EGI_TOT']}") == 16)
chk("t30.xcheck0", gn("3.3 FTE View", "C164") == 0, g("3.3 FTE View", "C164"))
# exec ties
chk("exec.model", close(gn("Exec Summary", "C26"), gn("3.2 Total Cost", f"C{TOT}")))
chk("exec.actual", close(gn("Exec Summary", "C30"), gn("3.2 Total Cost", f"D{TOT}")))
chk("exec.filled", close(gn("Exec Summary", "C31"), gn("3.2 Total Cost", f"K{TOT}")))
chk("exec.vacant", close(gn("Exec Summary", "C32"), gn("3.2 Total Cost", f"L{TOT}")))
chk("exec.dedup", close(gn("Exec Summary", "C25"), gn("3.2 Total Cost", f"C{DED}")))
chk("exec.unmapped", close(gn("Exec Summary", "C57"), gn("3.2 Total Cost", f"D{UNM}")))
# GM tabs
wb2 = wb
lever = 0.0
for tab, anch in A["GM"].items():
    ws = wb2[tab]
    hdr, tot, rh = anch["hdr"], anch["tot"], anch["rost_hdr"]
    t = tab.upper()
    chk(f"gm.{tab}.title", isinstance(g(t, "B2"), str) and g(t, "B2").endswith(" working copy"), g(t, "B2"))
    chk(f"gm.{tab}.h_is_f_minus_g", close(gn(t, f"H{tot}"), gn(t, f"F{tot}") - gn(t, f"G{tot}")))
    for r in range(hdr + 1, tot):
        gcell = ws.cell(r, 7).value
        chk(f"gm.{tab}.g_formula_r{r}", isinstance(gcell, str) and gcell.startswith("="), repr(gcell))
    chk(f"gm.{tab}.hdrE", ws.cell(rh, 5).value == "Vacancy lever")
    chk(f"gm.{tab}.hdrH", ws.cell(hdr, 8).value == "Vacancies remaining")
    chk(f"gm.{tab}.hdrD", ws.cell(hdr, 4).value == "Archetype roles")
    chk(f"gm.{tab}.hdrC", ws.cell(hdr, 3).value == "Archetype type")
    chk(f"gm.{tab}.hdrO", ws.cell(hdr, 15).value == "Archetype size")
    chk(f"gm.{tab}.hdrK", ws.cell(hdr, 11).value == "Archetype cost ($m)")
    chk(f"gm.{tab}.hdrL", ws.cell(hdr, 12).value == "Actual cost ($m)")
    chk(f"gm.{tab}.hdrM", ws.cell(hdr, 13).value == "Cost after vacancy decisions ($m)")
    chk(f"gm.{tab}.hdrN", ws.cell(hdr, 14).value == "New Variance ($m)")
    for r_ in range(hdr + 1, tot):
        kv = str(ws.cell(r_, 11).value or "")
        lv = str(ws.cell(r_, 12).value or "")
        chk(f"gm.{tab}.kl_live_r{r_}", kv.startswith("=") and lv.startswith("="), f"{kv!r}/{lv!r}")
    for r in range(hdr + 1, tot):
        cv = ws.cell(r, 3).value
        chk(f"gm.{tab}.c_live_r{r}", isinstance(cv, str) and cv.startswith("="), repr(cv))
    # find the 'cost to hire all vacancies' row and add to lever
    for r in range(tot, rh):
        if ws.cell(r, 2).value == "Cost to hire all vacancies ($m)":
            lever += gn(t, f"C{r}")
            break
for rd in (64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75):
    vv = g("Exec Summary", f"C{rd}")
    chk(f"exec.drill.C{rd}", isinstance(vv, (int, float)) or vv is None, repr(vv))
chk("exec.lever", close(gn("Exec Summary", "C52"), lever), f"{g('Exec Summary','C52')} vs {lever}")
for t4, refs in (("2.12 BP&T", ("1.11 BP&T", "F8", "G8", "H8")),
                 ("2.13 SA&D", ("1.12 SA&D", "G8", "H8", "I8"))):
    for i, cell_ in enumerate(refs[1:]):
        chk(f"wt.{t4}.money{i}", close(gn(t4, f"F{5+i}"), gn(refs[0], cell_)),
            f"{g(t4, f'F{5+i}')} vs {g(refs[0], cell_)}")
# language / formatting guards
BAN = ["seat", "roster", "your squads, your people", "decide hire or hold on every"]
DATA_TABS = {"Squads", "Added data", "Sheet2", "0.4 Presentation Pack",
             "0.1 Budget Table (Fin)", "0.3 Squad Archetypes", "0.2 Data Config", "Sheet1",
             "squad mapping", "Lists", "FY26 Budget (ref)"}
for ws in wb2.worksheets:
    if ws.title in DATA_TABS: continue
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and not v.startswith("="):
                lv = v.lower()
                for b in BAN:
                    chk(f"lang.{ws.title}.{c.coordinate}", b not in lv, v[:60])
                chk(f"dash.{ws.title}.{c.coordinate}", "–" not in v and "—" not in v, v[:60])
            if c.font and c.font.size and c.font.size < 10 and v not in (None, ""):
                chk(f"font.{ws.title}.{c.coordinate}", False, f"size {c.font.size}")
            if c.font and c.font.italic and v not in (None, ""):
                chk(f"italic.{ws.title}.{c.coordinate}", False, str(v)[:40])
# instruction-gate regression checks
for wsx in wb2.worksheets:
    if wsx.title in ("Squads", "Added data", "Sheet2"): continue
    for rowx in wsx.iter_rows():
        for cx in rowx:
            if isinstance(cx.value, str):
                chk(f"custai.{wsx.title}.{cx.coordinate}", "AI Enablement" not in cx.value, "AI Enablement survives")
chk("recon.c32", gn("3.1 Group Summary", "C32") == 0 or abs(gn("3.1 Group Summary", "C32")) < 0.005,
    g("3.1 Group Summary", "C32"))
chk("cy.grouping_hdr", wb2["1.13 Cyber Roles"]["B5"].value == "Grouping")
chk("cy.no_stale_hdrs", all(wb2["1.13 Cyber Roles"].cell(18, c).value is None for c in (8, 9, 10)),
    "old On/Off / Full Cost headers must be cleared")
for t4 in list(A["GM"].keys()):
    dvs = wb2[t4].data_validations.dataValidation
    lever_ok = any("Offshore" in str(d.formula1 or "") for d in dvs)
    chk(f"lever.dv_offshore.{t4}", lever_ok)
for t4 in ("2.12 BP&T", "2.13 SA&D", "2.14 EGI & Central"):
    bad_call = []
    for rowx in wb2[t4].iter_rows():
        for cx in rowx:
            if cx.value == "Call": bad_call.append(cx.coordinate)
    chk(f"lever.no_call_hdr.{t4}", not bad_call, str(bad_call))
# no hard-coded 11* in funding formulas
for tab, addr in [("3.4 COE Summary", "E8"), ("3.4 COE Summary", "E11"), ("3.2 Total Cost", f"C{DED}"),
                  ("1.11 BP&T", "C13"), ("1.12 SA&D", "C13")]:
    v = wb2[tab][addr].value
    chk(f"nohard11.{tab}.{addr}", isinstance(v, str) and "11*" not in v, repr(v))
for tab in ("1.11 BP&T", "1.12 SA&D"):
    v = wb2[tab]["C10"].value
    chk(f"nohard.portcount.{tab}", isinstance(v, str) and v.startswith("=COUNTA"), repr(v))
# hidden sheets stay hidden
for t in ("FY26 Budget (superseded)", "squad mapping (superseded)", "Lists"):
    chk(f"hidden.{t}", wb2[t].sheet_state == "hidden")
chk("gone.Sheet1", "Sheet1" not in wb2.sheetnames)
# sheet order
names = wb2.sheetnames
chk("order.49_410", names.index("2.9 Commercial Fuels") < names.index("2.10 Z Retail"))
# FTE rows are live refs; NO cost against filled people anywhere
for tab, first, last_ in [("1.11 BP&T", A["PT_R1"], A["PT_CHECK"] - 1),
                          ("1.12 SA&D", A["SAD_R1"], A["SAD_CHECK"] - 1),
                          ("1.13 Cyber Roles", A["CY_R1"], A["CY_CHECK"] - 1)]:
    ws = wb2[tab]
    for r in range(first, last_ + 1):
        for col in (2, 3, 4, 5, 6):
            v = ws.cell(r, col).value
            chk(f"liveref.{tab}.r{r}c{col}", isinstance(v, str) and v.startswith("="), repr(v)[:40])
        tv = ws.cell(r, 20).value
        chk(f"helper.{tab}.r{r}", isinstance(tv, str) and tv.startswith("="), repr(tv)[:30])
        st = g(tab.upper(), f"F{r}")
        gv = ws.cell(r, 7).value
        if st == "Filled" or st == "Contractor":
            chk(f"nocost.{tab}.r{r}", gv in (None, ""), f"filled row shows cost: {gv!r}")
        else:
            chk(f"vaccost.{tab}.r{r}", isinstance(gv, str) and gv.startswith("="), repr(gv)[:30])
    chk(f"thidden.{tab}", ws.column_dimensions["T"].hidden)
# 4.x: no cost against filled people on the working copies either
for tab in list(A["GM"].keys()) + ["2.12 BP&T", "2.13 SA&D", "2.14 EGI & Central"]:
    ws = wb2[tab]
    costcol = 6 if tab in A["GM"] else 7
    statcol = 4 if tab in A["GM"] else 5
    for r in range(1, ws.max_row + 1):
        sv = g(tab.upper(), f"{get_column_letter(statcol)}{r}") if False else None
    # spot rule: every cost cell must sit on a row whose status resolves vacant-ish
    for r in range(1, ws.max_row + 1):
        cv = ws.cell(r, costcol).value
        if isinstance(cv, str) and cv.startswith("=") and ("AA$" in cv or "Added data" in cv):
            stv = g(tab.upper(), f"{get_column_letter(statcol)}{r}")
            chk(f"nocost4.{tab}.r{r}", stv in ("Vacant", "Paused", None),
                f"cost on status {stv!r}")
# ---------------- raw data integrity vs the owner's upload ----------------
UPLOAD = "/root/.claude/uploads/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/4beb5516-Cost_Calc_Lee_edits22.xlsx"
wu = openpyxl.load_workbook(UPLOAD, data_only=False)
def real_diffs(tab_v9, tab_up):
    a, b = wb[tab_v9], wu[tab_up]
    diffs = []
    for r in range(1, max(a.max_row, b.max_row) + 1):
        for c in range(1, max(a.max_column, b.max_column) + 1):
            va, vb = a.cell(r, c).value, b.cell(r, c).value
            va = getattr(va, "text", va); vb = getattr(vb, "text", vb)   # ArrayFormula
            if va == vb: continue
            if (va in (None, "") and vb in (None, "")): continue
            if isinstance(va, float) and isinstance(vb, (int, float)) and abs(va - float(vb)) < 1e-6: continue
            if str(va) == str(vb): continue
            diffs.append((r, c, va, vb))
    return diffs
logged = {(rw, cl) for rw, cl in ([tuple(x) for x in A.get("CHANGELOG", [])])}
from openpyxl.utils import get_column_letter as _gcl
for tab in ("Squads", "Sheet2", "Added data"):
    d = real_diffs(tab, tab)
    if tab == "Squads":
        bad = [x for x in d if (x[0], _gcl(x[1])) not in logged]
        chk("raw.Squads.only_logged_changes", len(bad) == 0, str(bad[:5]))
    elif tab == "Added data":
        bad = [x for x in d if x[1] < 29]   # owner data A..AB untouched; AC+ are our helper cols
        chk("raw.Added.owner_cols_untouched", len(bad) == 0, str(bad[:5]))
    else:
        chk(f"raw.{tab}.untouched", len(d) == 0, str(d[:5]))
# owner's Hire/Hold calls preserved exactly on every GM tab
for tab in A["GM"]:
    a, b = wb[tab], wu["4." + tab.split(".", 1)[1]]
    calls = [(r, a.cell(r, 5).value, b.cell(r, 5).value)
             for r in range(1, max(a.max_row, b.max_row) + 1)
             if a.cell(r, 5).value != b.cell(r, 5).value
             and (a.cell(r, 5).value in ("Hire", "Hold") or b.cell(r, 5).value in ("Hire", "Hold"))]
    chk(f"calls.{tab}.preserved", len(calls) == 0, str(calls[:4]))
# per-row GM impact formulas + G block ranges point inside the right squad block
for tab, anch in A["GM"].items():
    ws = wb[tab]
    hdr, tot, rh = anch["hdr"], anch["tot"], anch["rost_hdr"]
    blocks, cur = {}, None
    for r in range(rh + 1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if isinstance(b, str) and b.startswith("=Squads!"):
            if cur: blocks[cur][1] = r
        elif isinstance(b, str) and b.strip() and not b.startswith("="):
            if b.startswith(("Check", "Cost", "Vacant", "Leadership", "Cyber", "vs archetype")): break
            cur = b.strip(); blocks[cur] = [r + 1, r]
    for r in range(hdr + 1, tot):
        hcell, icell = ws.cell(r, 8).value, ws.cell(r, 9).value
        chk(f"gm.{tab}.h_row{r}", isinstance(hcell, str) and hcell.startswith(f"=F{r}-G{r}"), repr(hcell))
        chk(f"gm.{tab}.i_row{r}", isinstance(icell, str) and f"E{r}+G{r}-D{r}" in icell, repr(icell))
        gcell = str(ws.cell(r, 7).value or "")
        mm = re.match(r'^=COUNTIF\(E(\d+):E(\d+),"Hire"\)$', gcell)
        if mm:
            name = str(ws.cell(r, 2).value or "").strip()
            blk = blocks.get(name)
            chk(f"gm.{tab}.g_range_row{r}", blk is not None and
                int(mm.group(1)) >= blk[0] - 1 and int(mm.group(2)) <= blk[1],
                f"{gcell} vs block {blk}")
# no typed constants in the rebuilt 2.1 table body or Exec value cells
w21 = wb["3.2 Total Cost"]
for r in range(6, A["TOT"] + 1):
    for c in range(3, 13):
        v = w21.cell(r, c).value
        chk(f"nohard.21.r{r}c{c}", v is None or (isinstance(v, str) and v.startswith("=")), repr(v))
wex = wb["Exec Summary"]
for r in range(19, 60):
    v = wex.cell(r, 3).value
    if v is not None and r != 63:
        chk(f"nohard.exec.C{r}", isinstance(v, str) and v.startswith("="), repr(v)[:40])
# SA&D partition completeness, read from the ARTIFACT (2.4 + 3.1 refs), not the rule
ws24 = wb["1.12 SA&D"]
refs24 = set()
for r in range(A["SAD_R1"], A["SAD_CHECK"]):
    m = re.match(r"^=Sheet2!\$B\$(\d+)$", str(ws24.cell(r, 2).value or ""))
    if m: refs24.add(int(m.group(1)))
ws31 = wb["4.0 Data QA"]
refs31 = set()
for row in ws31.iter_rows():
    for cell in row:
        m = re.match(r"^=Sheet2!\$B\$(\d+)$", str(cell.value or ""))
        if m: refs31.add(int(m.group(1)))
unaccounted = []
for x in SAD:
    r = x["r"]
    nmx = str(s2.cell(r, 2).value)
    in24 = r in refs24
    insquad = not blankish(x["squad"])
    mq = sq_match(x)
    lead = bool(mq and mq["Q"] == "Leadership")
    ed = bool(mq and mq["N"] in ("Enterprise Data", "Group Data"))
    on31 = r in refs31
    if not (in24 or insquad or lead or ed or on31):
        unaccounted.append((r, nmx))
    model_port = bool(mq and mq["Q"] in ("Squad", "Strategic Program", "Leadership")
                      and mq["N"] not in ("COE", "Unmapped"))
    chk(f"sadpart.no_overlap.r{r}", not (in24 and model_port),
        f"on the COE tab AND in a portfolio: {nmx}")
chk("sadpart.all_accounted", len(unaccounted) == 0, str(unaccounted[:5]))
# every NAMED new joiner (no exact raw match) must be visible on 3.1
for x in rows2:
    nm = low(str(x.get("r") and s2.cell(x["r"], 2).value))
    if "vacant" in nm or "ring fenced" in nm or x["div"] == "EGI": continue
    if sq_match(x) is None:
        chk(f"named_disclosed.{nm[:20]}", x["r"] in refs31, f"Sheet2 r{x['r']} not on 3.1")
# ---------------- every single role on a 2.x working tab ----------------
refset4 = set()
for t4 in [t for t in wb.sheetnames if t.startswith("2.")]:
    for row_ in wb[t4].iter_rows(min_col=2, max_col=2):
        m4 = re.match(r"^=(?:IF\(Squads!\$R\$(\d+)=.*|Squads!\$B\$(\d+))$", str(row_[0].value or ""))
        if m4: refset4.add(int(m4.group(1) or m4.group(2)))
s2ref4 = set()
for t4 in ("2.12 BP&T", "2.13 SA&D", "2.14 EGI & Central"):
    for row_ in wb[t4].iter_rows(min_col=2, max_col=2):
        m4 = re.match(r"^=Sheet2!\$B\$(\d+)$", str(row_[0].value or ""))
        if m4: s2ref4.add(int(m4.group(1)))
def s2_twin(q):
    for x in rows2:
        if low(s2.cell(x["r"], 2).value) == q["name"] and low(s2.cell(x["r"], 3).value) == q["title"]:
            return x["r"]
    return None
missing4 = []
for q in sq_rows:
    if q["r"] in refset4: continue
    tw = s2_twin(q)
    if tw and tw in s2ref4: continue
    missing4.append((q["r"], q["name"], q["Q"]))
chk("cover.squads_all_on_4x", len(missing4) == 0, str(missing4[:6]))
missing_s2 = []
for x in rows2:
    r = x["r"]
    if r in s2ref4: continue
    mq2 = sq_match(x)
    if mq2 and mq2["r"] in refset4: continue
    missing_s2.append((r, str(s2.cell(r, 2).value)))
chk("cover.sheet2_all_on_4x", len(missing_s2) == 0, str(missing_s2[:6]))
# AU/NZ wiring ties
au_terms = [gn(t, a) for t, (a, b) in A["AUNZ"].items()]
nz_terms = [gn(t, b) for t, (a, b) in A["AUNZ"].items()]
R0 = A["AUNZ_ROW"]
chk("aunz.au_total", close(gn("3.1 Group Summary", f"C{R0+3}"), sum(au_terms)))
chk("aunz.nz_total", close(gn("3.1 Group Summary", f"C{R0+7}"), sum(nz_terms)))
dsum = sum(gn("3.1 Group Summary", f"D{r}") for r in range(6, 16))
oh_gap = dsum - (sum(au_terms) + sum(nz_terms))
chk("aunz.covers_squads_sanity", 0 < sum(au_terms) + sum(nz_terms) < dsum and 4 < oh_gap < 22,
    f"au+nz={sum(au_terms)+sum(nz_terms):.2f} dsum={dsum:.2f} overhead_gap={oh_gap:.2f}")
# 2.0 K column = total cost, ties to old semantics
chk("t20.k24_net", close(gn("3.1 Group Summary", "K26"), gn("3.2 Total Cost", f"C{A['TOT']}")))
del wu; gc.collect()

# ---------------- cost after vacancy decisions can never land below zero ----------------
for gt_, anch_ in sorted(A["GM"].items()):
    negs = []
    for rr in range(anch_["hdr"] + 1, anch_["tot"]):
        mv = gn(gt_, f"M{rr}")
        if isinstance(mv, (int, float)) and mv < -1e-9:
            negs.append((rr, round(mv, 4)))
    chk(f"gm.m_floor.{gt_}", not negs, f"negative M rows: {negs}")

# ---------------- the Offshore lever: prove the 0.4 x maths ----------------
import shutil
FLIP = SCR + "flip_v9.xlsx"
shutil.copy(SRC, FLIP)
wf = openpyxl.load_workbook(FLIP)
gt = "2.1 Ampol Retail"
anch = A["GM"][gt]
wsf = wf[gt]
target = None
for r in range(anch["rost_hdr"] + 1, wsf.max_row + 1):
    if wsf.cell(r, 5).value == "Hold" and isinstance(wsf.cell(r, 6).value, str):
        target = r; break
sqrow = None
if target:
    wsf.cell(target, 5).value = "Offshore"
    wf.save(FLIP)
del wf; gc.collect()
vals2 = engine(FLIP)
if target:
    ws4 = wb2[gt]
    for rr in range(anch["hdr"] + 1, anch["tot"]):
        gcell = str(ws4.cell(rr, 7).value or "")
        mm = re.match(r'^=COUNTIF\(E(\d+):E(\d+),"Hire"\)$', gcell)
        if mm and int(mm.group(1)) <= target <= int(mm.group(2)):
            sqrow = rr; break
if target and sqrow:
    base_m = gn(gt, f"M{sqrow}")
    cost = gn(gt, f"F{target}")
    new_m = vals2.get((gt.upper(), f"M{sqrow}"))
    chk("lever.offshore_04", isinstance(new_m, (int, float)) and close(new_m, base_m + 0.4 * cost / 1e6),
        f"before {base_m} after {new_m} cost {cost}")
else:
    chk("lever.offshore_04", False, "no Hold row with cost found to flip")
del vals2; gc.collect()

print(f"KEY: model {g('3.2 Total Cost', 'C' + str(TOT))} actual {g('3.2 Total Cost', 'D' + str(TOT))} "
      f"BPbudget {g('1.11 BP&T', 'C15')} SAbudget {g('1.12 SA&D', 'C15')} cyber {g('1.13 Cyber Roles', 'F8')} "
      f"egi {g('3.3 FTE View', 'F' + str(A['EGI_TOT']))} restate {g('3.2 Total Cost', 'D' + str(RESTATE))}")
print("FAILS:", len(fails))
for f in fails[:60]: print("  -", f)
sys.exit(0 if not fails else 1)
