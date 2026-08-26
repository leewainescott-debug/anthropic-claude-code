#!/usr/bin/env python3
"""Build TDD_NonLabour_Mapping.xlsx: every AU HW/SW line item placed in the
new op model, with Lee's six open decisions as live toggle cells.

Foundations (docs/NONLABOUR_MAPPING_ANALYSIS.md, all verified to the cent):
  leaf totals SW 51,288,134.47 + HW 25,468,998.32 = 76,757,132.79
  parents excluded: SW {HCORP, BULKFUEL, MARKETING, SIG_ITEMS},
                    HW {HCORP, MARKETING}
Raw tabs ride along verbatim (raw data rule). Ledger rows trace to raw rows.
"""
import json, warnings, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from collections import defaultdict
warnings.filterwarnings("ignore")

SRC = "/tmp/claude-0/-home-user-anthropic-claude-code/e550b440-3996-5abb-87e5-bafafe598f82/scratchpad/nonlabour/budget.xlsx"
ROLES = json.load(open("/tmp/claude-0/-home-user-anthropic-claude-code/e550b440-3996-5abb-87e5-bafafe598f82/scratchpad/roles.json"))
OUT = "deliverables/TDD_NonLabour_Mapping.xlsx"
PWD = "Tdd123"

NAVY="0F2E52"; INK="1D2939"; MUTE="667085"; BANDR="F6F8FB"; EMPH="E8EEF6"
INPUT="FFF4CC"; INPUTB="E3B505"; WHITE="FFFFFF"; TOTAL="E7E7E7"; FONT="Calibri"
MONEY='#,##0.00;(#,##0.00);""'
MONEY3='#,##0.000;(#,##0.000);""'

def f(size=11,bold=False,color=INK): return Font(name=FONT,size=size,bold=bold,color=color)
def fill(h): return PatternFill("solid",fgColor=h)
def hdr(c,t,align="center"):
    c.value=t; c.font=f(11,True,WHITE); c.fill=fill(NAVY)
    c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)

SW_PARENTS={"APPLHCORP","APPLBULKFUEL","APPLMARKETING","APPLSIG_ITEMS"}
HW_PARENTS={"APPLHCORP","APPLMARKETING"}
SIG_CC={"APOLLO":"Apollo","CTRMSIG":"CTRM","INTEMERSIG":"Emerald integration (EGI)",
        "CRPOSSIG":"Conv Retail POS","TDDEINTSIG":"TDD enterprise integration",
        "CUSTLOYSIG":"Customer loyalty","HEMERALD":"Emerald (EGI)",
        "PCTECHKCM":"P&C tech (KCM)","PCTECHMYHR":"P&C tech (MyHR)",
        "APPLSIG_ITEMS":"Device fleet (program to confirm)"}
DIRECT_CC={  # cc -> (portfolio, platform)
 "APPLB2B":("B2B & Energy Solutions","B2B"),
 "APPLAETOTAL":("B2B & Energy Solutions","Ampol Energy"),
 "INTEGM":("B2B & Energy Solutions","International & New Business"),
 "APPLFIREFINING":("Infrastructure","Refining"),
 "APPLLUBRICANTS":("Commercial Fuels","Lubricants"),
 "APPLSUPPLYOPS":("Commercial Fuels","Supply"),
 "APPLSUPFPO":("Commercial Fuels","Fuel pricing & optimisation"),
 "APPLFIFINANCE":("P&C, Finance & Legal","Finance"),
 "APPLHFIN":("P&C, Finance & Legal","Finance"),
 "APPLHHR":("P&C, Finance & Legal","People & Culture"),
 "APPLHLSA":("P&C, Finance & Legal","Legal & Secretariat"),
 "APPLHALT":("TDD Corporate","Executive (ALT & CEO)"),
 "APPLRCALSTORES":("Retail","Company stores"),
}
CC_HOME={  # fallback portfolio of each shared cc
 "APPLH_INFOTECH":"Infrastructure","APPLBULKFUEL":"Commercial Fuels",
 "APPLMTOTAL":"Retail","APPLBRNDCOMS":"Ampol Customer","APPLCCO":"Commercial Fuels",
 "APPLTDDIST":"Infrastructure","APPLGENERAL":"P&C, Finance & Legal",
}
EUC_WORDS=("LAPTOP","DESKTOP","MONITOR","DOCKING"," DOCK","SURFACE","IPAD","IPHONE",
           "PRINTER","TONER","MICROSOFT","MSFT","M365","OFFICE 365","BACKPACK","HEADSET","KEYBOARD")
SEC_WORDS=("TRELLIX","CROWDSTRIKE","ZSCALER","PALO ALTO","SPLUNK","NETSKOPE","PROOFPOINT",
           "OKTA","TENABLE","QUALYS","SENTINEL","DEFENDER","MIMECAST","CYBERARK","FORTINET")
CLOUD_WORDS=("AZURE","AWS","AMAZON WEB")
NET_WORDS=("TELSTRA","CISCO","MERAKI","SD-WAN","SDWAN","FIREWALL","SWITCH","ROUTER",
           "WIRELESS","WIFI","WI-FI","NETWORK")
DC_WORDS=("NTT","VMWARE","SERVER","STORAGE","BACKUP","VEEAM","DATA CENTRE","DATACENTRE","DATACENTER")
DATA_WORDS=("SNOWFLAKE","DATABRICKS","POWER BI","POWERBI","TABLEAU","INFORMATICA","COLLIBRA")
SAP_WORDS=("SAP ","SAP-","S/4","ARIBA","SUCCESSFACTOR"," BW ")
POS_WORDS=("AMPOS","POS ","EFTPOS","INGENICO","PINPAD","PIN PAD","VERIFONE")
CC_CENTRE_WORDS=("GENESYS","NICE LTD","TWILIO","CONTACT CENTRE","CONTACT CENTER")
SITE_WORDS=("GILBARCO","WAYNE","DISPENSER","TANK GAUGE","FORECOURT","CCTV","CAMERA","FUEL")

def classify(src,cc,ce_code,ce_name,text,vendor):
    """returns (class, basis, group, default_portfolio, default_platform)"""
    t=(text or "").upper(); v=(vendor or "").upper()
    if cc in SIG_CC:
        return ("Programs","Sig item cost centre","", "Strategic programs (sig items)", SIG_CC[cc])
    if cc in DIRECT_CC:
        p,pl=DIRECT_CC[cc]
        return ("Direct","Cost centre names the portfolio","",p,pl)
    if cc=="APPLTDDIST":
        return ("Decision","Old model Fuels, squad sits in Infrastructure","DIST","Infrastructure","Distribution, Sales & Services")
    if cc=="APPLCCO":
        return ("Decision","CCO function spans Fuels and B2B","CCO","Commercial Fuels","Portfolio-wide")
    if cc=="APPLBRNDCOMS":
        return ("Decision","Brand versus corporate comms","BRAND","Ampol Customer","Brand & communications")
    if cc=="APPLGENERAL":
        return ("Decision","HighRadius / OneStream, Finance systems","GENERAL","P&C, Finance & Legal","Finance systems")
    home=CC_HOME.get(cc,"TDD Corporate")
    # shared centres: product rules, first hit wins
    if str(ce_code) in ("801450","801440") or any(w in t for w in EUC_WORDS):
        return ("Decision","End-user device or per-user software","EUC","Infrastructure","End user compute")
    if any(w in t for w in SEC_WORDS) or any(w in v for w in SEC_WORDS):
        return ("Rule","Security product","", "TDD Corporate","Cyber COE")
    if "SERVICENOW" in t or "SERVICENOW" in v or "SERVICE NOW" in t:
        return ("Rule","ServiceNow","", "TDD Corporate","Service Ops COE")
    if any(w in t for w in CLOUD_WORDS) or any(w in v for w in CLOUD_WORDS):
        return ("Rule","Cloud consumption","", "Infrastructure","Cloud")
    if any(w in t for w in DC_WORDS) or "NTT" in v:
        return ("Rule","Data centre / compute","", "Infrastructure","Data centre & compute")
    if any(w in t for w in NET_WORDS) or "TELSTRA" in v:
        return ("Rule","Network / carriage","", "Infrastructure","Network")
    if any(w in t for w in DATA_WORDS):
        return ("Rule","Data platform product","", "Enterprise Data","Data platforms")
    if any(w in t for w in SAP_WORDS) or v.startswith("SAP"):
        return ("Rule","SAP family","", "TDD Corporate","Enterprise applications (ERP)")
    if "SALESFORCE" in t or "SALESFORCE" in v:
        if cc=="APPLMTOTAL":
            return ("Decision","Marketing Salesforce: loyalty/martech","MKT","Ampol Customer","CRM & loyalty")
        return ("Rule","Salesforce platform","", "Ampol Customer","CRM (Salesforce)")
    if "ADOBE" in t or "ADOBE" in v:
        if cc in ("APPLMTOTAL",):
            return ("Decision","Marketing Adobe stack","MKT","Ampol Customer","Digital experience")
        return ("Rule","Adobe","", "Ampol Customer","Digital experience")
    if any(w in t for w in POS_WORDS):
        return ("Rule","Store POS / payments","", "Retail","Store systems (POS & payments)")
    if "WAY4" in t:
        return ("Rule","Way4 cards platform","", "Commercial Fuels","Cards (AmpolCard)")
    if any(w in t for w in CC_CENTRE_WORDS):
        return ("Rule","Contact centre product","", "Ampol Customer","Contact centre")
    if cc in ("APPLMTOTAL",) and any(w in t for w in SITE_WORDS):
        return ("Rule","Site / forecourt equipment","", "Retail","Site & forecourt")
    if "LOYALTY" in t and cc=="APPLMTOTAL":
        return ("Decision","Loyalty spend","MKT","Ampol Customer","CRM & loyalty")
    return ("Open","No product tell yet, needs digging","", home,"Portfolio-wide")

# ---------------- read the raw workbook ----------------
src=openpyxl.load_workbook(SRC, data_only=True, read_only=True)
def leaf_rows(sheet, parents):
    ws=src[sheet]; out=[]
    for i,r in enumerate(ws.iter_rows(min_row=15, values_only=True), start=15):
        cc=r[2]; amt=r[26]
        if cc in (None,"","Overall Result") or amt is None: continue
        try: a=float(amt)
        except: continue
        if cc in parents: continue
        out.append((i,cc,r[3],r[4],r[5],r[6],r[10],r[12],r[13],a))
    return out
sw=leaf_rows("SW Line Items",SW_PARENTS)
hw=leaf_rows("HW Line Items",HW_PARENTS)
sw_tot=sum(x[-1] for x in sw); hw_tot=sum(x[-1] for x in hw)
assert abs(sw_tot-51288134.47)<0.5, sw_tot
assert abs(hw_tot-25468998.32)<0.5, hw_tot
print(f"leaves: SW {len(sw)} rows {sw_tot:,.2f} | HW {len(hw)} rows {hw_tot:,.2f}")

records=[]
stats=defaultdict(float)
for tag,rows in (("S",sw),("H",hw)):
    tab="SW Line Items" if tag=="S" else "HW Line Items"
    for (rowno,cc,ccname,ce,cename,doc,vend,post,text,a) in rows:
        cls,basis,grp,dp,dpl = classify(tag,cc,ce,cename,text,vend)
        records.append([f"{tag}{rowno}",tab,rowno,cc,ccname,ce,cename,doc,
                        (vend if vend not in (None,"","Not assigned","#") else ""),
                        post,text,a,cls,basis,grp,dp,dpl])
        stats[cls]+=a
print("class $m:", {k:round(v/1e6,3) for k,v in sorted(stats.items())})

wb=openpyxl.Workbook()

# ---------------- 2 Decisions ----------------
dec=wb.active; dec.title="2 Decisions"
dec.sheet_view.showGridLines=False
dec.column_dimensions["A"].width=2
for col,w in (("B",42),("C",34),("D",16),("E",70)):
    dec.column_dimensions[col].width=w
dec["B2"]="Decisions that move money"; dec["B2"].font=f(16,True,NAVY)
dec["B3"]="Six calls only you can make. Each yellow cell is a dropdown. The whole workbook, including the profile, moves the moment you change one. The values loaded now are my proposals, marked so."
dec["B3"].font=f(11,color=MUTE); dec["B3"].alignment=Alignment(wrap_text=True)
dec.row_dimensions[3].height=28
DEC_ROWS={}
decisions=[
 ("EUC","End-user devices and per-user software (laptops, monitors, Microsoft per-user)",
  ["Central: Infrastructure, End user compute","Follow the cost centre portfolio"],
  "About 10.6m across the shared centres. Central puts every device under Infrastructure. Follow leaves each device with the portfolio whose cost centre bought it. Sig-funded devices stay with programs either way."),
 ("MKT","Marketing digital spend (Salesforce, Adobe, loyalty in the marketing centre)",
  ["Ampol Customer","Retail"],
  "The marketing cost centre mixes store operations with loyalty and martech. This call sends the digital products to Ampol Customer or keeps everything Retail."),
 ("CCO","Chief Commercial Officer function",
  ["Commercial Fuels","B2B & Energy Solutions"],
  "0.69m of function software. The CCO spans fuels and B2B."),
 ("BRAND","Brand, Foundation & Communications",
  ["Ampol Customer","TDD Corporate"],
  "1.03m. Brand sits with Customer in the new model unless you want it held corporately."),
 ("DIST","Distribution, Sales & Services",
  ["Infrastructure","Commercial Fuels"],
  "1.02m. The old model held it under Commercial Fuels. The labour model's squad sits in Infrastructure. Proposal follows the squad."),
 ("GENERAL","Non-functional general centre (HighRadius, OneStream)",
  ["P&C, Finance & Legal","TDD Corporate"],
  "0.92m, almost all Finance systems by vendor and text."),
]
r=5
hdr(dec.cell(4,2),"Decision",align="left"); hdr(dec.cell(4,3),"Your call (dropdown)")
hdr(dec.cell(4,4),"Status"); hdr(dec.cell(4,5),"What it moves",align="left")
for key,label,options,note in decisions:
    dec.cell(r,2,label).font=f(11); dec.cell(r,2).alignment=Alignment(wrap_text=True,vertical="center")
    c=dec.cell(r,3,options[0]); c.font=f(11,True); c.fill=fill(INPUT)
    c.border=Border(*[Side("thin",color=INPUTB)]*4); c.protection=Protection(locked=False)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    dv=DataValidation(type="list", formula1='"'+",".join(options)+'"', allow_blank=False)
    dec.add_data_validation(dv); dv.add(f"C{r}")
    dec.cell(r,4,"PROPOSED").font=f(10,True,"B45309"); dec.cell(r,4).alignment=Alignment(horizontal="center",vertical="center")
    dec.cell(r,5,note).font=f(10,color=MUTE); dec.cell(r,5).alignment=Alignment(wrap_text=True,vertical="center")
    dec.row_dimensions[r].height=42
    DEC_ROWS[key]=r
    r+=1
dec.cell(r+1,2,"Sig items always show as their own Strategic programs grouping, matching the labour model and the dashboard. Tell me if you want them spread into portfolios instead.").font=f(10,color=MUTE)
DEC_SHEET="'2 Decisions'"
dec.protection=openpyxl.worksheet.protection.SheetProtection(sheet=True,password=PWD,
    selectLockedCells=False,selectUnlockedCells=False)

# ---------------- 3 Ledger ----------------
lg=wb.create_sheet("3 Ledger")
lg.sheet_view.showGridLines=False
COLS=["Line ID","Raw tab","Raw row","Cost centre","Cost centre name","Cost element",
      "Cost element name","CO document","Vendor","Posting date","Document text",
      "Total $ (12 months)","Class","Basis","Decision","Default portfolio",
      "Default platform","Final portfolio","Final platform"]
widths=[9,13,8,17,20,9,19,12,20,11,42,13,10,30,9,20,24,20,24]
lg["B1"]="Every line, placed"; lg["B1"].font=f(16,True,NAVY)
lg["B2"]="One row per leaf line item, Jul-2025 to Jun-2026 actuals. Line ID = raw tab row, so S15 is row 15 of the SW Line Items tab, untouched further right. Final columns show the landing under the proposed decisions; the Profile tab moves live when you change a decision, and this ledger is refreshed once your rulings land."
lg["B2"].font=f(10,color=MUTE)
HR=4
for j,(t,w) in enumerate(zip(COLS,widths)):
    col=2+j
    hdr(lg.cell(HR,col),t, align="left" if t in ("Cost centre name","Cost element name","Document text","Basis","Default portfolio","Default platform","Final portfolio","Final platform","Vendor") else "center")
    lg.column_dimensions[get_column_letter(col)].width=w
lg.row_dimensions[HR].height=30
euc_row=DEC_ROWS["EUC"]
row=HR+1
for rec in records:
    (lid,tab,rawrow,cc,ccname,ce,cename,doc,vend,post,text,amt,cls,basis,grp,dp,dpl)=rec
    vals=[lid,tab,rawrow,cc,ccname,ce,cename,doc,vend,
          str(post)[:10] if post else "",(text or "")[:80],amt,cls,basis,grp,dp,dpl]
    for j,v in enumerate(vals):
        c=lg.cell(row,2+j,v); c.font=f(10)
        if j==10: c.alignment=Alignment(horizontal="left")
        if j==11: c.number_format='#,##0.00;(#,##0.00)'
    lg.cell(row,19,dp).font=f(10,True,NAVY)
    lg.cell(row,20,dpl).font=f(10)
    row+=1
LG_LAST=row-1
lg.auto_filter.ref=f"B{HR}:T{LG_LAST}"
lg.freeze_panes=f"E{HR+1}"
lg.protection=openpyxl.worksheet.protection.SheetProtection(sheet=True,password=PWD,
    autoFilter=False,sort=False,selectLockedCells=False,selectUnlockedCells=False)
print("ledger rows:",LG_LAST-HR)

AMT=f"'3 Ledger'!$M${HR+1}:$M${LG_LAST}"
FPORT=f"'3 Ledger'!$S${HR+1}:$S${LG_LAST}"
FPLAT=f"'3 Ledger'!$T${HR+1}:$T${LG_LAST}"
CLS=f"'3 Ledger'!$N${HR+1}:$N${LG_LAST}"

# ---------------- 4 Profile (block-table design: fast, same maths) ----------------
# Decision groups move as whole blocks, so the profile = static base per
# (portfolio, platform) + block terms driven by the six decision cells.
from collections import defaultdict as _dd
base=_dd(float)          # non-decision lines by (portfolio, platform)
euc_total=0.0; euc_by_home=_dd(float)
mkt_crm=0.0; mkt_dx=0.0; mkt_total=0.0
cco_t=0.0; brand_t=0.0; dist_t=0.0; general_t=0.0
for rec in records:
    (lid,tab,rawrow,cc,ccname,ce,cename,doc,vend,post,text,amt,cls,basis,grp,dp,dpl)=rec
    if grp=="EUC":
        euc_total+=amt
        home={"APPLMTOTAL":"Retail","APPLBULKFUEL":"Commercial Fuels"}.get(cc,dp)
        euc_by_home[home]+=amt
    elif grp=="MKT":
        mkt_total+=amt
        if dpl=="CRM & loyalty": mkt_crm+=amt
        else: mkt_dx+=amt
    elif grp=="CCO": cco_t+=amt
    elif grp=="BRAND": brand_t+=amt
    elif grp=="DIST": dist_t+=amt
    elif grp=="GENERAL": general_t+=amt
    else:
        base[(dp,dpl)]+=amt

DEC=lambda k: f"{DEC_SHEET}!$C${DEC_ROWS[k]}"
terms=_dd(list)   # (portfolio, platform) -> list of formula terms
terms[("Infrastructure","End user compute")].append(
    f'IF(LEFT({DEC("EUC")},7)="Central",{euc_total!r},{euc_by_home.get("Infrastructure",0.0)!r})')
for home,v in euc_by_home.items():
    if home=="Infrastructure": continue
    terms[(home,"End user compute")].append(f'IF(LEFT({DEC("EUC")},7)="Central",0,{v!r})')
terms[("Ampol Customer","CRM & loyalty")].append(f'IF({DEC("MKT")}="Ampol Customer",{mkt_crm!r},0)')
terms[("Ampol Customer","Digital experience")].append(f'IF({DEC("MKT")}="Ampol Customer",{mkt_dx!r},0)')
terms[("Retail","Store systems (digital)")].append(f'IF({DEC("MKT")}="Retail",{mkt_total!r},0)')
terms[("Commercial Fuels","Portfolio-wide")].append(f'IF({DEC("CCO")}="Commercial Fuels",{cco_t!r},0)')
terms[("B2B & Energy Solutions","Portfolio-wide")].append(f'IF({DEC("CCO")}="B2B & Energy Solutions",{cco_t!r},0)')
terms[("Ampol Customer","Brand & communications")].append(f'IF({DEC("BRAND")}="Ampol Customer",{brand_t!r},0)')
terms[("TDD Corporate","Corporate communications")].append(f'IF({DEC("BRAND")}="TDD Corporate",{brand_t!r},0)')
terms[("Infrastructure","Distribution, Sales & Services")].append(f'IF({DEC("DIST")}="Infrastructure",{dist_t!r},0)')
terms[("Commercial Fuels","Distribution")].append(f'IF({DEC("DIST")}="Commercial Fuels",{dist_t!r},0)')
terms[("P&C, Finance & Legal","Finance systems")].append(f'IF({DEC("GENERAL")}="P&C, Finance & Legal",{general_t!r},0)')
terms[("TDD Corporate","Finance systems")].append(f'IF({DEC("GENERAL")}="TDD Corporate",{general_t!r},0)')

pf=wb.create_sheet("4 Profile")
pf.sheet_view.showGridLines=False
pf.column_dimensions["A"].width=2
for col,w in (("B",26),("C",34),("D",15),("E",15),("F",15)):
    pf.column_dimensions[col].width=w
pf["B1"]="Cost profile in the new operating model"; pf["B1"].font=f(16,True,NAVY)
pf["B2"]="A$, 12 months of AU actuals Jul-2025 to Jun-2026 for hardware and software, labour at the FY27 model price for AU roles. Moves live with the Decisions tab. NZ and the other blocks are named on the Reconciliation tab, never guessed."
pf["B2"].font=f(10,color=MUTE)
allpairs=set(base)|set(terms)
bypf=defaultdict(list)
for p_,pl_ in sorted(allpairs): bypf[p_].append(pl_)
PORT_ORDER=["Retail","Commercial Fuels","B2B & Energy Solutions","Infrastructure",
            "Ampol Customer","Enterprise Data","P&C, Finance & Legal","TDD Corporate",
            "Strategic programs (sig items)"]
lab=defaultdict(float)
FOLD={"RETAIL":"Retail","Retail":"Retail","Ampol Customer":"Ampol Customer",
 "Commercial Fuels":"Commercial Fuels","B2B & Energy Solutions":"B2B & Energy Solutions",
 "Infrastructure":"Infrastructure","Enterprise Data":"Enterprise Data",
 "P&C, Finance & Legal":"P&C, Finance & Legal","P&C":"P&C, Finance & Legal",
 "Finance":"P&C, Finance & Legal","TDD":"TDD Corporate"}
lab_other=0.0
for rec_r in ROLES:
    if rec_r.get("country")!="AU": continue
    p_=FOLD.get(rec_r.get("portfolio") or "")
    amt_=float(rec_r.get("today") or 0)*1e6
    if p_: lab[p_]+=amt_
    else: lab_other+=amt_
HRP=4
hdr(pf.cell(HRP,2),"Portfolio",align="left"); hdr(pf.cell(HRP,3),"Platform",align="left")
hdr(pf.cell(HRP,4),"Non-labour $"); hdr(pf.cell(HRP,5),"Labour AU $"); hdr(pf.cell(HRP,6),"Total $")
pf.row_dimensions[HRP].height=22
r=HRP+1
port_total_rows=[]
for p_ in PORT_ORDER:
    plats=bypf.get(p_,[])
    first=r
    for pl_ in plats:
        pf.cell(r,3,pl_).font=f(10)
        parts=[]
        b=base.get((p_,pl_),0.0)
        if b: parts.append(repr(b))
        parts+=terms.get((p_,pl_),[])
        form="="+("+".join(parts) if parts else "0")
        c=pf.cell(r,4,form); c.number_format=MONEY; c.font=f(10); c.alignment=Alignment(horizontal="right")
        r+=1
    pf.cell(r,2,p_+" total").font=f(11,True)
    c=pf.cell(r,4,f"=SUM(D{first}:D{r-1})" if r>first else "=0")
    c.number_format=MONEY; c.font=f(11,True); c.alignment=Alignment(horizontal="right")
    lv=lab.get(p_,0.0)
    c2=pf.cell(r,5, lv if p_!="Strategic programs (sig items)" else "TBC")
    c2.number_format=MONEY; c2.font=f(11,True); c2.alignment=Alignment(horizontal="right")
    if p_=="Strategic programs (sig items)":
        c2.font=f(10,color=MUTE); c2.alignment=Alignment(horizontal="center")
        c3=pf.cell(r,6,f"=D{r}")
    else:
        c3=pf.cell(r,6,f"=D{r}+E{r}")
    c3.number_format=MONEY; c3.font=f(11,True); c3.alignment=Alignment(horizontal="right")
    for cc_ in range(2,7): pf.cell(r,cc_).fill=fill(EMPH)
    pf.cell(r,2).alignment=Alignment(horizontal="left")
    port_total_rows.append(r)
    pf.cell(first,2,p_).font=f(11,True,NAVY)
    r+=2
gr=r
pf.cell(gr,2,"Total").font=f(12,True)
sumrefs="+".join(f"D{x}" for x in port_total_rows)
c=pf.cell(gr,4,f"={sumrefs}"); c.number_format=MONEY; c.font=f(12,True); c.alignment=Alignment(horizontal="right")
sl="+".join(f"E{x}" for x in port_total_rows if pf.cell(x,5).value!="TBC")
c=pf.cell(gr,5,f"={sl}"); c.number_format=MONEY; c.font=f(12,True); c.alignment=Alignment(horizontal="right")
tl="+".join(f"F{x}" for x in port_total_rows)
c=pf.cell(gr,6,f"={tl}"); c.number_format=MONEY; c.font=f(12,True); c.alignment=Alignment(horizontal="right")
for cc_ in range(2,7): pf.cell(gr,cc_).fill=fill(TOTAL)
chk=gr+1
pf.cell(chk,2,"Check against the leaf ledger total 76,757,132.79 (0 = ties)").font=f(9,color=MUTE)
c=pf.cell(chk,4,f"=D{gr}-76757132.79"); c.number_format='0.00;(0.00)'; c.font=f(9,color=MUTE); c.alignment=Alignment(horizontal="right")
chk2=chk+1
pf.cell(chk2,2,"Lines still Open (no product tell yet, named for digging)").font=f(9,color=MUTE)
open_t=sum(rec[11] for rec in records if rec[12]=="Open")
c=pf.cell(chk2,4,open_t); c.number_format=MONEY; c.font=f(9,color=MUTE); c.alignment=Alignment(horizontal="right")
if lab_other>0:
    ln=chk2+1
    pf.cell(ln,2,"Labour AU not folded to these portfolios (EGI, Z, other)").font=f(9,color=MUTE)
    c=pf.cell(ln,5,lab_other); c.number_format=MONEY; c.font=f(9,color=MUTE); c.alignment=Alignment(horizontal="right")
pf.protection=openpyxl.worksheet.protection.SheetProtection(sheet=True,password=PWD,
    selectLockedCells=False,selectUnlockedCells=False)

# ---------------- 5 Reconciliation ----------------
rc=wb.create_sheet("5 Reconciliation")
rc.sheet_view.showGridLines=False
rc.column_dimensions["A"].width=2
for col,w in (("B",52),("C",16),("D",60)):
    rc.column_dimensions[col].width=w
rc["B1"]="The bridge to the roughly 302m"; rc["B1"].font=f(16,True,NAVY)
rc["B2"]="Total enterprise technology spend in the FY26 dashboard = Ampol opex 189.1 + Ampol capex 48.1 + Z Energy opex 64.8 = 302.1m budget. What this workbook covers, and every named gap. TBC means the dataset is not in this file, nothing is guessed."
rc["B2"].font=f(10,color=MUTE); rc["B2"].alignment=Alignment(wrap_text=True)
rc.row_dimensions[2].height=28
hdr(rc.cell(4,2),"Block",align="left"); hdr(rc.cell(4,3),"$"); hdr(rc.cell(4,4),"Where it stands",align="left")
rows=[
 ("AU software, 12 months of actuals (this file)", 51288134.47,"Mapped line by line on the Ledger tab."),
 ("AU hardware, 12 months of actuals (this file)", 25468998.32,"Mapped line by line on the Ledger tab."),
 ("Labour, AU roles at the FY27 model price", sum(lab.values())+lab_other,"From the labour model, joined on the Profile tab."),
 ("Network (Communications-Data, Mobile)", "TBC","Not in this extract. 3.1m at B2026 enterprise level per the dashboard."),
 ("Outside services", "TBC","Not in this extract. 31.1m at B2026 enterprise level."),
 ("Depreciation", "TBC","Not in this extract. 5.8m per the budget detail."),
 ("Ampol capex", "TBC","48.1m budget FY26 per the dashboard, separate dataset."),
 ("Z Energy, all of NZ (opex 64.8m budget)", "TBC","Entire NZ side absent from this AU file. Second extract needed."),
 ("NZ labour", "TBC","Sits in the labour model at live FX, joins when the Z extract lands."),
]
r=5
for label,val,note in rows:
    rc.cell(r,2,label).font=f(11)
    c=rc.cell(r,3,val)
    if isinstance(val,str) and val=="TBC":
        c.font=f(10,color=MUTE); c.alignment=Alignment(horizontal="center")
    else:
        c.number_format=MONEY; c.font=f(11); c.alignment=Alignment(horizontal="right")
    rc.cell(r,4,note).font=f(10,color=MUTE); rc.cell(r,4).alignment=Alignment(wrap_text=True)
    if r%2==0:
        for cc_ in range(2,5): rc.cell(r,cc_).fill=fill(BANDR)
    r+=1
rc.cell(r,2,"Dashboard benchmark: total enterprise technology spend").font=f(11,True)
c=rc.cell(r,3,302100000); c.number_format=MONEY; c.font=f(11,True); c.alignment=Alignment(horizontal="right")
for cc_ in range(2,5): rc.cell(r,cc_).fill=fill(EMPH)
rc.cell(r+1,2,"The blocks above must fill in before the two columns can be compared. The AU hardware and software slice is complete and reconciled today.").font=f(9,color=MUTE)
rc.protection=openpyxl.worksheet.protection.SheetProtection(sheet=True,password=PWD,
    selectLockedCells=False,selectUnlockedCells=False)

# ---------------- 1 Start here ----------------
st=wb.create_sheet("1 Start here")
st.sheet_view.showGridLines=False
st.column_dimensions["A"].width=2
for col in "BCDEFGHIJ": st.column_dimensions[col].width=13
st.merge_cells("B2:J3")
tb=st["B2"]; tb.value="Non-labour cost in the new operating model"
tb.font=f(18,True,WHITE); tb.alignment=Alignment(horizontal="left",vertical="center",indent=1)
for rr in (2,3):
    for cc_ in range(2,11): st.cell(rr,cc_).fill=fill(NAVY)
st["B5"]="Every AU hardware and software line item, placed by portfolio and platform, with your six open calls as live dropdowns."
st["B5"].font=f(11,color=MUTE)
st["B7"]="How to use it"; st["B7"].font=f(12,True,NAVY)
steps=[
 "1.  Open 2 Decisions and set the six yellow dropdowns. Everything downstream moves with them.",
 "2.  3 Ledger is every line: 49,900 rows, each traced to its raw tab row, with the class, the basis and the landing spot.",
 "3.  4 Profile is the answer: cost by portfolio and platform, labour beside non-labour, with a zero check against the ledger.",
 "4.  5 Reconciliation is the honest bridge to the roughly 302m, every missing block named, TBC where the data is not in this file.",
 "5.  The last three tabs are your raw data, copied in verbatim and untouched.",
]
for i,s in enumerate(steps): st.cell(8+i,2,s).font=f(11)
st["B15"]="How the placing works"; st["B15"].font=f(12,True,NAVY)
notes=[
 "Direct: the cost centre names one new portfolio, no judgement involved. 18.6m.",
 "Programs: sig-item centres recharge to strategic programs, the same treatment as labour. 14.9m, including the 12.0m device fleet billed monthly per laptop.",
 "Rule: a product or vendor tell places the line (Azure to Cloud, security products to the Cyber COE, SAP to enterprise applications, POS to store systems). 39.7m before your decisions.",
 "Decision: the six dropdowns on tab 2. Open: no tell yet, named for digging, totalled on the Profile tab.",
 "Every number is A$ actuals, Jul-2025 to Jun-2026. The ledger re-adds to 76,757,132.79 exactly, and the Profile checks itself against it.",
 "Platform names for direct lines follow the cost centre wording. GMs can rename them without moving a dollar.",
 "Locked so the sums cannot break. Only the yellow cells take typing. Password Tdd123.",
]
for i,s in enumerate(notes):
    st.cell(16+i,2,s).font=f(11)
st.protection=openpyxl.worksheet.protection.SheetProtection(sheet=True,password=PWD,
    selectLockedCells=False,selectUnlockedCells=False)

# ---------------- raw tabs verbatim ----------------
for name in ("2026 Budget","SW Line Items","HW Line Items"):
    ws_src=src[name]
    tgt=wb.create_sheet(name)
    n=0
    for r_ in ws_src.iter_rows():
        for c_ in r_:
            if c_.value is not None:
                tgt.cell(c_.row,c_.column,c_.value); n+=1
    print("verbatim",name,n,"cells")

order=["1 Start here","2 Decisions","3 Ledger","4 Profile","5 Reconciliation",
       "2026 Budget","SW Line Items","HW Line Items"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active=0
wb.save(OUT)
print("saved",OUT)
