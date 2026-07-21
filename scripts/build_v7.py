#!/usr/bin/env python3
"""v7 on top of v6:
 1. Map Added data (cost source) to the model with helper cols AC-AG
    (same rules as raw data; Commercial dept -> COE Business Partnering,
    also applied to raw data for consistency)
 2. 3.0 FTE View: Actual cost ($m) + Cost variance columns per squad
 3. New '5.0 Total Cost': Model vs Actual by layer x portfolio,
    filled/vacant split, grand totals tying to $120.04m
 4. New '5.1 Data QA': raw data vs Added data discrepancy outline
 5. Leadership: actual vs allowance on 5.0
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from copy import copy

SCR = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/"
SRC = SCR + "TDD_Cost_Calc_v6.xlsx"
OUT = SCR + "TDD_Cost_Calc_v7.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=False)
wbv = openpyxl.load_workbook(SRC, data_only=True)

NAVY_F = PatternFill("solid", fgColor="FF1F4E79"); DK_F = PatternFill("solid", fgColor="FF002F6C")
GREY_F = PatternFill("solid", fgColor="FFD9D9D9")
W = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
B = Font(name="Calibri", size=10, bold=True)
N = Font(name="Calibri", size=10)
TITLE = Font(name="Calibri", size=16, bold=True)
thin = Side(style="thin", color="FFB8C9CC")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '#,##0.00;\\(#,##0.00\\);\\-'
PCT = '0%'
GREEN_F = PatternFill("solid", fgColor="FFE2EFDA"); GREEN_T = Font(color="FF006100", bold=True)
RED_F = PatternFill("solid", fgColor="FFFBE4D5"); RED_T = Font(color="FF9C0006", bold=True)
def pos_bad(ws, rng):
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=RED_F, font=RED_T))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=GREEN_F, font=GREEN_T))
def set_cell(ws, coord, value, font=None, fill=None, fmt=None, border=True, align=None):
    c = ws[coord]; c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if border: c.border = BOX
    if align: c.alignment = Alignment(horizontal=align, vertical="center")
    return c

# =====================================================================
# 1. shared mapping rules (identical to raw data build)
# =====================================================================
CI = "Configuration / Integration"; EDI = "Enterprise Data and Insights"
SQUAD_MAP = {
 "pos": ("Ampol Retail","Store Operations","POS","Squad"),
 "payments": ("Ampol Retail","Store Operations","Payments","Squad"),
 "store operations": ("Ampol Retail","Store Operations","Store Operations","Squad"),
 "deployment": ("Ampol Retail","Store Operations","Deployment","Squad"),
 "above store": ("Ampol Retail","Above Store","Above Store","Squad"),
 "merchandising & supply chain": ("Ampol Retail","Above Store","Above Store","Squad"),
 "pricing & wfm": ("Ampol Retail","Above Store","Above Store","Squad"),
 "ampos": ("Ampol Retail","AmPOS","AmPOS","Strategic Program"),
 "network & qsr": ("Ampol Retail","Network / QSR","Network / QSR","Squad"),
 "data - au": ("Ampol Retail","Data AU","Data AU","Squad"),
 "egi retail": ("Ampol Retail","EGI Retail","EGI Retail","Strategic Program"),
 "ampol app": ("Customer","Ampol Digital","Ampol App","Squad"),
 "ampol web": ("Customer","Ampol Digital","Ampol Web","Squad"),
 "digital ops support": ("Customer","Ampol Digital","Digital Operations","Squad"),
 "z energy apps": ("Customer","Customer Z","Z Energy Apps","Squad"),
 "z energy martech": ("Customer","Customer Z","Z Energy Martech","Squad"),
 "au crm & martech": ("Customer","Group Customer Platforms","AU CRM & Martech","Squad"),
 "loyalty & martech": ("Customer","Group Customer Platforms","AU CRM & Martech","Squad"),
 "customer, ai": ("Customer","-","Customer AI","Unmapped"),
 "egi customer": ("Customer","EGI Customer","EGI Customer","Strategic Program"),
 "data science": ("Enterprise Data","Group Data","Data Science","Squad"),
 "coe, data science": ("Enterprise Data","Group Data","Data Science","Squad"),
 "coe, data operations": ("Enterprise Data","Group Data","Operations","Squad"),
 "reporting & analytics": ("Enterprise Data","Group Data","Reporting & Analytics","Squad"),
 "data platform": ("Enterprise Data","Group Data","Data Platforms","Squad"),
 "enterprise data delivery": ("Enterprise Data","Group Data","Enterprise Data Delivery","Squad"),
 "coe, data capability": ("COE","COE - Data","COE - Data","COE"),
 "workplace & enterprise tooling": ("TDD Group Functions","TDD Group Functions","Workplace & Enterprise Tooling","Squad"),
 "cloud network & infra ops": ("TDD Group Functions","TDD Group Functions","Network & Infrastructure","Squad"),
 "cloud, network & infra ops": ("TDD Group Functions","TDD Group Functions","Network & Infrastructure","Squad"),
 "devops & qe": ("TDD Group Functions","TDD Group Functions","DevOps & Engineering","Squad"),
 "integration & process automation": ("TDD Group Functions","TDD Group Functions","Integration","Squad"),
 "egi tdd": ("TDD Group Functions","EGI TDD","EGI TDD","Strategic Program"),
 "p&c": ("P&C","P&C","P&C","Squad"),
 "p&c rta": ("P&C","P&C","P&C - RTA","Squad"),
 "egi p&c": ("P&C","EGI P&C","EGI P&C","Strategic Program"),
 "sap erp": ("Finance","AU Finance","AU Finance","Squad"),
 "nz finance": ("Finance","NZ Finance","NZ Finance","Squad"),
 "egi finance": ("Finance","EGI Finance","EGI Finance","Strategic Program"),
 "distribution, sales & services": ("Infrastructure","Distribution","Distribution, Sales & Services","Squad"),
 "manufacturing group projects": ("Infrastructure","Manufacturing","Manufacturing & Group Projects","Squad"),
 "manuacturing group projects": ("Infrastructure","Manufacturing","Manufacturing & Group Projects","Squad"),
 "technology suport": ("Infrastructure","Manufacturing","Technology Support","Squad"),
 "technology support": ("Infrastructure","Manufacturing","Technology Support","Squad"),
 "data & insights": ("Infrastructure","Data & Insights","Data & Insights","Squad"),
 "energy": ("Energy Solutions & B2B","Energy Solutions","Energy","Squad"),
 "evci": ("Energy Solutions & B2B","Energy Solutions","EVCI","Squad"),
 "b2b": ("Energy Solutions & B2B","B2B","B2B","Squad"),
 "trading & shipping": ("Commercial Fuels","Trading & Shipping","Trading & Shipping","Squad"),
 "trading & shipping data": ("Commercial Fuels","Trading & Shipping","Trading & Shipping Data","Squad"),
 "supply": ("Commercial Fuels","Supply","Supply","Squad"),
 "ctrm": ("Commercial Fuels","CTRM","CTRM","Strategic Program"),
 "z supply": ("Z Retail","Z Supply","Z Supply","Squad"),
 "site systems": ("Z Retail","Z Customer","Site Systems","Squad"),
 "z retail backend": ("Z Retail","Z Customer","Z Retail Backend","Squad"),
 "z retail backend & supply": ("Z Retail","Z Customer","Z Retail Backend","Squad"),
 "data - nz": ("Z Retail","-","Data NZ","Unmapped"),
 "egi": ("EGI","EGI","EGI (unassigned)","Unmapped"),
 "coe, architecture": ("COE","COE - Strategy Architecture","COE - Strategy Architecture","COE"),
 "coe": ("COE","COE (unspecified)","COE (unspecified)","COE"),
 "coe, leadership": ("COE","Leadership","Leadership","Leadership"),
}
PORT_NORM = {"z":"Z Retail","retail":"Ampol Retail","customer":"Customer",
             "b2b & energy solutions":"Energy Solutions & B2B","egi integration":"EGI","egi":"EGI",
             "tdd":"TDD Group Functions","enterprise data":"Enterprise Data",
             "infrastructure":"Infrastructure","finance":"Finance","p&c":"P&C",
             "p&c, finance & legal":"P&C","commercial fuels":"Commercial Fuels","na":"","":""}
CY_DEPTS = {"cyber strat & tech","cyber sec ops","cyber risk","cyber grc","service op & assurance"}
COE_MAP = {"transformation": "COE - Transformation", "architecture": "COE - Strategy Architecture",
           "technology strategy & ai capability": "COE - Strategy Architecture",
           "delivery, sada": "COE - Strategy Architecture",
           "tdd business partner": "COE - Business Partnering",
           "commercial": "COE - Business Partnering",
           "group data": "COE - Data"}
def classify(name, dept, port, plat, squad, title=""):
    status = "Vacant" if name.lower().strip() == "vacant" else "Filled"
    sq = squad.lower().strip()
    pn = PORT_NORM.get(port.lower().strip(), port)
    dp = dept.lower().strip()
    if sq == "leadership" or plat.lower().strip() == "leadership":
        return (pn or "Unmapped"), "Leadership", "Leadership", "Leadership", status
    if sq in SQUAD_MAP:
        mp, mpl, ms, cls = SQUAD_MAP[sq]
        return mp, mpl, ms, cls, status
    if dp in CY_DEPTS:
        return "TDD Cyber", "TDD Cyber, Risk & Service Ops", "TDD Cyber", "Squad", status
    if dp in COE_MAP:
        c = COE_MAP[dp]
        if dp == "group data":
            return "Enterprise Data", "Group Data", "Data Platforms", "Squad", status
        return "COE", c, c, "COE", status
    if dp == "loyalty & martech":
        return "Customer", "Group Customer Platforms", "AU CRM & Martech", "Squad", status
    return (pn or "Unmapped"), (plat or "-"), (squad or dept or "-"), "Unmapped", status

# apply to Added data -> cols AC(29)-AG(33)
ad = wb["Added data"]; adv = wbv["Added data"]
hdr_style = copy(ad.cell(1, 2)._style)
for j, h in enumerate(["Model Portfolio","Model Platform","Model Squad","Class","Status"]):
    c = ad.cell(1, 29 + j); c.value = h; c._style = hdr_style
ad_recs = []
for r in range(2, ad.max_row + 1):
    g = lambda cc: (str(adv.cell(r, cc).value).strip() if adv.cell(r, cc).value is not None else "")
    name, title, dept, port, plat, squad = g(2), g(3), (g(7) or g(6)), g(9), g(10), g(11)
    if not name and not title: continue
    cost = adv.cell(r, 27).value
    if squad.lower().strip() == "leadership" or plat.lower().strip() == "leadership":
        mp, mpl, ms, cls, st = classify(name, dept, port, plat, squad)
    elif dept.lower().strip() == "group data":
        # group data: squad col decides portfolio squad vs COE Data
        sq = squad.lower().strip()
        if sq in SQUAD_MAP:
            mp, mpl, ms, cls = SQUAD_MAP[sq]; st = "Vacant" if name.lower().strip()=="vacant" else "Filled"
        elif "architect" in title.lower():
            mp, mpl, ms, cls, st = "COE","COE - Strategy Architecture","COE - Strategy Architecture","COE", ("Vacant" if name.lower().strip()=="vacant" else "Filled")
        else:
            mp, mpl, ms, cls, st = "COE","COE - Data","COE - Data","COE", ("Vacant" if name.lower().strip()=="vacant" else "Filled")
    else:
        mp, mpl, ms, cls, st = classify(name, dept, port, plat, squad, title)
    for j, v in enumerate([mp, mpl, ms, cls, st]):
        ad.cell(r, 29 + j).value = v
    ad_recs.append(dict(row=r, port=mp, plat=mpl, squad=ms, cls=cls, status=st,
                        cost=float(cost) if isinstance(cost, (int, float)) else 0.0, name=name, title=title))
# raw data: remap Commercial dept rows (previously Unmapped) to COE - BP for consistency
rd = wb["raw data"]; rdv = wbv["raw data"]
for r in range(2, rd.max_row + 1):
    dept = str(rdv.cell(r, 7).value or rdv.cell(r, 6).value or "").strip().lower()
    if dept == "commercial" and rd.cell(r, 17).value == "Unmapped":
        rd.cell(r, 14).value = "COE"; rd.cell(r, 15).value = "COE - Business Partnering"
        rd.cell(r, 16).value = "COE - Business Partnering"; rd.cell(r, 17).value = "COE"

AD = "'Added data'"
def sumifs(port=None, cls=None, squad=None, status=None):
    conds = []
    if port: conds.append(f"{AD}!$AC:$AC,\"{port}\"")
    if squad: conds.append(f"{AD}!$AE:$AE,\"{squad}\"")
    if cls: conds.append(f"{AD}!$AF:$AF,\"{cls}\"")
    if status: conds.append(f"{AD}!$AG:$AG,\"{status}\"")
    return f"SUMIFS({AD}!$AA:$AA," + ",".join(conds) + ")"

# =====================================================================
# 2. 3.0 FTE View: actual cost + variance columns
# =====================================================================
ft = wb["3.0 FTE View"]
ft.column_dimensions["N"].width = 15; ft.column_dimensions["O"].width = 15
hdr_r = None; grand = None
for r in range(1, ft.max_row + 1):
    if ft.cell(r, 2).value == "Portfolio" and ft.cell(r, 13).value == "Archetype cost ($m)":
        hdr_r = r
    if ft.cell(r, 2).value == "TOTAL - all modelled squads":
        grand = r
set_cell(ft, f"N{hdr_r}", "Actual cost ($m)", W, NAVY_F, align="center")
set_cell(ft, f"O{hdr_r}", "Cost var ($m)", W, NAVY_F, align="center")
sub_rows = []
for r in range(hdr_r + 1, grand):
    b = ft.cell(r, 2).value; d = ft.cell(r, 4).value
    if b and str(b).endswith(" total"):
        sub_rows.append(r); continue
    if not d: continue
    port = b
    set_cell(ft, f"N{r}", f"=({sumifs(port=port, squad=str(d))})/1000000", N, None, MONEY, align="right")
    set_cell(ft, f"O{r}", f"=N{r}-IFERROR(M{r},0)", N, None, MONEY, align="right")
prev = hdr_r
for sr in sub_rows:
    set_cell(ft, f"N{sr}", f"=SUM(N{prev+1}:N{sr-1})", B, GREY_F, MONEY, align="right")
    set_cell(ft, f"O{sr}", f"=N{sr}-M{sr}", B, GREY_F, MONEY, align="right")
    prev = sr
set_cell(ft, f"N{grand}", "=" + "+".join(f"N{sr}" for sr in sub_rows), W, DK_F, MONEY, align="right")
set_cell(ft, f"O{grand}", f"=N{grand}-M{grand}", W, DK_F, MONEY, align="right")
pos_bad(ft, f"O{hdr_r+1}:O{grand}")

# =====================================================================
# 3. 5.0 Total Cost - Model vs Actual by layer
# =====================================================================
TABS = ["1.1 Ampol Retail","1.2 Customer","1.3 Enterprise Data","1.4 TDD Group Functions",
        "1.5 P&C","1.6 Finance","1.7 Infrastructure","1.8 Energy Solutions & B2B",
        "1.9 Commercial Fuels","1.10 Z Retail","1.11 TDD Cyber"]
PORT = {t: t.split(" ", 1)[1] for t in TABS}
# summary anchor rows per tab (Total Cost row = C col SUM)
tc_row = {}
for t in TABS:
    w = wb[t]
    for r in range(1, 15):
        if w.cell(r, 2).value == "Total Cost": tc_row[t] = r
oh_rows = {}
for t in TABS:
    w = wb[t]
    for r in range(1, 15):
        if w.cell(r, 2).value == "Portfolio Overhead": oh_rows[t] = r

tc = wb.create_sheet("5.0 Total Cost", wb.sheetnames.index("4.0 Insights") + 1)
tc.sheet_view.showGridLines = False
for col, wd in {"A":3,"B":30,"C":14,"D":14,"E":14,"F":14,"G":14}.items():
    tc.column_dimensions[col].width = wd
tc.row_dimensions[2].height = 21
set_cell(tc, "B2", "Total Cost - Model vs Actual (every layer of the operating model)", TITLE, border=False)
r = 4
HD = ["","Model ($m)","Actual Filled ($m)","Actual Vacant ($m)","Actual Total ($m)","Variance ($m)"]
def section(title):
    global r
    set_cell(tc, f"B{r}", title, W, DK_F, align="left")
    for col in "CDEFG": set_cell(tc, f"{col}{r}", None, W, DK_F)
    r += 1
    for j, h in enumerate(HD):
        if h: set_cell(tc, f"{'BCDEFG'[j]}{r}", h, W, NAVY_F, align="center")
    set_cell(tc, f"B{r}", "Portfolio", W, NAVY_F, align="center")
    r += 1
def row(label, model_f, port=None, cls_list=("Squad", "Strategic Program"), squad=None, cls_single=None, bold=False, fillc=None):
    global r
    set_cell(tc, f"B{r}", label, B if bold else N, fillc, align="left")
    set_cell(tc, f"C{r}", model_f, B if bold else N, fillc, MONEY, align="right")
    if cls_single:
        f_fill = f"=({sumifs(cls=cls_single, port=port, squad=squad, status='Filled')})/1000000"
        f_vac = f"=({sumifs(cls=cls_single, port=port, squad=squad, status='Vacant')})/1000000"
    else:
        f_fill = "=(" + "+".join(sumifs(cls=c, port=port, squad=squad, status="Filled") for c in cls_list) + ")/1000000"
        f_vac = "=(" + "+".join(sumifs(cls=c, port=port, squad=squad, status="Vacant") for c in cls_list) + ")/1000000"
    set_cell(tc, f"D{r}", f_fill, B if bold else N, fillc, MONEY, align="right")
    set_cell(tc, f"E{r}", f_vac, B if bold else N, fillc, MONEY, align="right")
    set_cell(tc, f"F{r}", f"=D{r}+E{r}", B if bold else N, fillc, MONEY, align="right")
    set_cell(tc, f"G{r}", f"=F{r}-C{r}", B if bold else N, fillc, MONEY, align="right")
    r += 1
def subtotal(label, first, fillc=GREY_F):
    global r
    set_cell(tc, f"B{r}", label, B, fillc, align="left")
    for col in "CDEFG":
        set_cell(tc, f"{col}{r}", f"=SUM({col}{first}:{col}{r-1})", B, fillc, MONEY, align="right")
    out = r; r += 1
    return out

ss_rows = {}
for t in TABS:
    w = wb[t]
    for rr in range(1, 15):
        if w.cell(rr, 2).value == "Squad Support Costs": ss_rows[t] = rr
section("Squads & strategic programmes (support costs)")
a_first = r
for t in TABS:
    sr_ = ss_rows[t]
    row(PORT[t], f"='{t}'!$C${sr_}+'{t}'!$D${sr_}", port=PORT[t])
subA = subtotal("Subtotal - squads & programmes", a_first)
r += 1
section("Leadership & overheads (model allowance vs actual leadership cost)")
b_first = r
for t in TABS:
    ohr = oh_rows[t]
    row(PORT[t], f"='{t}'!$C${ohr}+'{t}'!$C${ohr+1}", port=PORT[t], cls_single="Leadership")
row("COE / central leadership", "=0", port="COE", cls_single="Leadership")
row("Unmapped leadership", "=0", port="Unmapped", cls_single="Leadership")
subB = subtotal("Subtotal - leadership & overheads", b_first)
r += 1
section("Centres of Excellence")
c_first = r
row("COE - Business Partnering", "='2.1 COE'!$D$11", squad="COE - Business Partnering", cls_single="COE")
row("COE - Transformation", "='2.1 COE'!$D$10", squad="COE - Transformation", cls_single="COE")
row("COE - Strategy Architecture", "='2.1 COE'!$D$8", squad="COE - Strategy Architecture", cls_single="COE")
row("COE - Data", "='2.1 COE'!$D$12", squad="COE - Data", cls_single="COE")
row("COE (unspecified)", "=0", squad="COE (unspecified)", cls_single="COE")
subC = subtotal("Subtotal - Centres of Excellence", c_first)
r += 1
section("Not in the model")
d_first = r
row("Unmapped roles", "=0", cls_single="Unmapped")
subD = subtotal("Subtotal - not in the model", d_first)
r += 2
set_cell(tc, f"B{r}", "TOTAL OPERATING MODEL", W, DK_F, align="left")
for col, lbl in [("C", None), ("D", None), ("E", None), ("F", None), ("G", None)]:
    set_cell(tc, f"{col}{r}", f"={col}{subA}+{col}{subB}+{col}{subC}+{col}{subD}", W, DK_F, MONEY, align="right")
GRAND = r
r += 1
set_cell(tc, f"B{r}", "Check: Actual Total = Added data full cost", B, align="left")
last_ad = max(rec["row"] for rec in ad_recs)
set_cell(tc, f"C{r}", f"=F{GRAND}-SUM({AD}!$AA$2:$AA${last_ad})/1000000", B, GREY_F, MONEY, align="right")
pos_bad(tc, f"C{r}")
CHECK = r
r += 2
set_cell(tc, f"B{r}", "Model TDD-funded (portfolios, from 2.0)", N, align="left")
set_cell(tc, f"C{r}", "='2.0 Group Summary'!$D$17", N, None, MONEY, align="right")
r += 1
set_cell(tc, f"B{r}", "Model funded outside TDD (portfolios, from 2.0)", N, align="left")
set_cell(tc, f"C{r}", "='2.0 Group Summary'!$G$17", N, None, MONEY, align="right")
pos_bad(tc, f"G5:G{GRAND}")

# =====================================================================
# 4. 5.1 Data QA - discrepancy outline (build-time snapshot)
# =====================================================================
qa = wb.create_sheet("5.1 Data QA", wb.sheetnames.index("5.0 Total Cost") + 1)
qa.sheet_view.showGridLines = False
for col, wd in {"A":3,"B":34,"C":16,"D":16,"E":16,"F":40}.items():
    qa.column_dimensions[col].width = wd
qa.row_dimensions[2].height = 21
set_cell(qa, "B2", "Data QA - raw data (roles) vs Added data (costs)", TITLE, border=False)
# recompute census
def census(ws, name_c=2, title_c=3):
    recs = {}
    vac = 0; n = 0
    for r0 in range(2, ws.max_row + 1):
        nm = str(ws.cell(r0, name_c).value or "").strip()
        ti = str(ws.cell(r0, title_c).value or "").strip()
        if not nm and not ti: continue
        n += 1
        if nm.lower() == "vacant": vac += 1
        else: recs[(nm.lower(), )] = recs.get((nm.lower(),), 0) + 1
    return n, vac, recs
rn, rvac, rnames = census(rdv)
an, avac, anames = census(adv)
r = 4
set_cell(qa, f"B{r}", "Headline", W, DK_F, align="left")
for col in "CDEF": set_cell(qa, f"{col}{r}", None, W, DK_F)
r += 1
for j, h in enumerate(["", "raw data", "Added data", "Difference"]):
    if h: set_cell(qa, f"{'BCDE'[j]}{r}", h, W, NAVY_F, align="center")
r += 1
for label, a, b_ in [("Records", rn, an), ("Vacant", rvac, avac),
                     ("Filled", rn - rvac, an - avac),
                     ("Distinct named people", len(rnames), len(anames))]:
    set_cell(qa, f"B{r}", label, N, align="left")
    set_cell(qa, f"C{r}", a, N, align="center")
    set_cell(qa, f"D{r}", b_, N, align="center")
    set_cell(qa, f"E{r}", b_ - a, B, RED_F if b_ != a else GREEN_F, align="center")
    r += 1
r += 1
only_ad = sorted(set(anames) - set(rnames)); only_rd = sorted(set(rnames) - set(anames))
set_cell(qa, f"B{r}", f"In Added data only ({len(only_ad)}) - joined the org after the raw data cut, or naming mismatch", W, DK_F, align="left")
for col in "CDEF": set_cell(qa, f"{col}{r}", None, W, DK_F)
r += 1
adv_names = {}
for r0 in range(2, adv.max_row + 1):
    nm = str(adv.cell(r0, 2).value or "").strip()
    if nm and nm.lower() != "vacant":
        adv_names.setdefault(nm.lower(), (nm, str(adv.cell(r0, 3).value or ""), str(adv.cell(r0, 7).value or adv.cell(r0, 6).value or "")))
for k in only_ad:
    nm, ti, dp = adv_names[k[0] if isinstance(k, tuple) else k]
    set_cell(qa, f"B{r}", nm, N, align="left"); set_cell(qa, f"C{r}", None, N)
    set_cell(qa, f"D{r}", None, N); set_cell(qa, f"F{r}", f"{ti} | {dp}", N, align="left")
    r += 1
r += 1
set_cell(qa, f"B{r}", f"In raw data only ({len(only_rd)}) - left the org, or naming mismatch", W, DK_F, align="left")
for col in "CDEF": set_cell(qa, f"{col}{r}", None, W, DK_F)
r += 1
rdv_names = {}
for r0 in range(2, rdv.max_row + 1):
    nm = str(rdv.cell(r0, 2).value or "").strip()
    if nm and nm.lower() != "vacant":
        rdv_names.setdefault(nm.lower(), (nm, str(rdv.cell(r0, 3).value or ""), str(rdv.cell(r0, 16).value or "")))
for k in only_rd:
    nm, ti, sq = rdv_names[k[0] if isinstance(k, tuple) else k]
    set_cell(qa, f"B{r}", nm, N, align="left")
    set_cell(qa, f"F{r}", f"{ti} | model squad: {sq}", N, align="left")
    r += 1
r += 1
# per-squad seat comparison
set_cell(qa, f"B{r}", "Seats by model squad - raw data vs Added data (differences only)", W, DK_F, align="left")
for col in "CDEF": set_cell(qa, f"{col}{r}", None, W, DK_F)
r += 1
for j, h in enumerate(["Portfolio / squad", "raw seats", "Added seats", "Difference"]):
    set_cell(qa, f"{'BCDE'[j]}{r}", h, W, NAVY_F, align="center")
r += 1
from collections import Counter
rc = Counter(); ac_ = Counter()
for r0 in range(2, rdv.max_row + 1):
    p = rd.cell(r0, 14).value; s = rd.cell(r0, 16).value
    if p and s: rc[(p, s)] += 1
for rec in ad_recs:
    ac_[(rec["port"], rec["squad"])] += 1
diffs = 0
for k in sorted(set(rc) | set(ac_)):
    d = ac_.get(k, 0) - rc.get(k, 0)
    if d == 0: continue
    diffs += 1
    set_cell(qa, f"B{r}", f"{k[0]} / {k[1]}", N, align="left")
    set_cell(qa, f"C{r}", rc.get(k, 0), N, align="center")
    set_cell(qa, f"D{r}", ac_.get(k, 0), N, align="center")
    set_cell(qa, f"E{r}", d, B, RED_F if abs(d) > 2 else None, align="center")
    r += 1

wb.save(OUT)
print("saved", OUT)
print("ad_recs:", len(ad_recs), "| raw:", rn, "added:", an, "| squad diffs:", diffs)
print("FTE hdr:", hdr_r, "grand:", grand, "| 5.0 grand row:", GRAND, "check row:", CHECK)

# restore Added data cached values (openpyxl strips them; SUMIFS reads need them at preview time)
import zipfile, shutil, os
import xml.etree.ElementTree as ET
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NSR = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", NS); ET.register_namespace("r", NSR)
tmp = SCR + "_v7tmp"
if os.path.exists(tmp): shutil.rmtree(tmp)
os.makedirs(tmp)
with zipfile.ZipFile(OUT) as z: z.extractall(tmp)
wbt = ET.parse(f"{tmp}/xl/workbook.xml"); root = wbt.getroot()
rid = None
for sh in root.iter(f"{{{NS}}}sheet"):
    if sh.get("name") == "Added data": rid = sh.get(f"{{{NSR}}}id")
rels = ET.parse(f"{tmp}/xl/_rels/workbook.xml.rels")
target = None
for rel in rels.getroot():
    if rel.get("Id") == rid: target = rel.get("Target")
path = f"{tmp}/xl/{target}" if not target.startswith("/") else f"{tmp}{target}"
tree = ET.parse(path); troot = tree.getroot()
fixed = 0
for c in troot.iter(f"{{{NS}}}c"):
    coord = c.get("r")
    f_el = c.find(f"{{{NS}}}f")
    if f_el is None: continue
    col = "".join(ch for ch in coord if ch.isalpha()); rown = int("".join(ch for ch in coord if ch.isdigit()))
    cached = adv[coord].value
    if cached is None: continue
    v_el = c.find(f"{{{NS}}}v")
    if v_el is None:
        v_el = ET.SubElement(c, f"{{{NS}}}v")
    if v_el.text in (None, ""):
        if isinstance(cached, (int, float)):
            for a in ("t",):
                if c.get(a) in ("str", "e"): del c.attrib[a]
            v_el.text = repr(float(cached)); fixed += 1
        else:
            c.set("t", "str"); v_el.text = str(cached); fixed += 1
tree.write(path, xml_declaration=True, encoding="UTF-8")
os.remove(OUT)
with zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED) as z:
    for base, _, files in os.walk(tmp):
        for fn in files:
            full = os.path.join(base, fn)
            z.write(full, os.path.relpath(full, tmp))
shutil.rmtree(tmp)
print("Added data caches restored:", fixed)
