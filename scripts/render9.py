#!/usr/bin/env python3
"""Render workbook sheets to a faithful HTML preview using actual cell values,
fills, fonts, merges & number formats (openpyxl)."""
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
import html, datetime

SRC="TDD_Cost_Calc_v9.xlsx"
wbv=openpyxl.load_workbook(SRC, data_only=True)
wbf=openpyxl.load_workbook(SRC, data_only=False)

def argb(color):
    if color is None: return None
    rgb=getattr(color,"rgb",None)
    if not isinstance(rgb,str): return None
    if len(rgb)==8: rgb=rgb[2:]
    if rgb in ("000000",) and getattr(color,"type",None)!="rgb": return None
    return "#"+rgb

def fmt_num(v, nf):
    if v is None or v=="": return ""
    if isinstance(v,str): return html.escape(v)
    if isinstance(v,(int,float)):
        if nf and ("%" in nf):
            return f"{v*100:.0f}%"
        if nf and "$" in nf:
            neg=v<0; a=abs(v)
            s=f"${a:,.2f}"
            return f"({s})" if neg else s
        if nf and "0.0" in nf:
            return f"{v:,.1f}"
        if float(v)==int(v): return f"{int(v):,}"
        return f"{v:,.2f}"
    return html.escape(str(v))

def render_sheet(title, maxr, maxc, minc=2):
    wsv=wbv[title]; wsf=wbf[title]
    # merged map: anchor -> (rowspan,colspan); covered set
    covered=set(); spans={}
    for mr in wsf.merged_cells.ranges:
        c1,r1,c2,r2=range_boundaries(str(mr))
        spans[(r1,c1)]=(r2-r1+1,c2-c1+1)
        for rr in range(r1,r2+1):
            for cc in range(c1,c2+1):
                if (rr,cc)!=(r1,c1): covered.add((rr,cc))
    out=[f'<table class="sheet"><caption>{html.escape(title)}</caption>']
    for r in range(1,maxr+1):
        h=wsf.row_dimensions[r].height
        rowstyle=f"height:{h}px" if h else ""
        out.append(f'<tr style="{rowstyle}">')
        for c in range(minc,maxc+1):
            if (r,c) in covered: continue
            cell=wsf.cell(r,c); cellv=wsv.cell(r,c)
            fillc=argb(cell.fill.fgColor) if cell.fill and cell.fill.patternType else None
            fontc=argb(cell.font.color) if cell.font else None
            styles=[]
            if fillc: styles.append(f"background:{fillc}")
            if fontc: styles.append(f"color:{fontc}")
            if cell.font and cell.font.bold: styles.append("font-weight:700")
            if cell.font and cell.font.italic: styles.append("font-style:italic")
            if cell.font and cell.font.size: styles.append(f"font-size:{float(cell.font.size)+1}px")
            al=cell.alignment.horizontal if cell.alignment else None
            if al: styles.append(f"text-align:{al}")
            else: styles.append("text-align:left")
            val=fmt_num(cellv.value, cell.number_format)
            rs,cs=spans.get((r,c),(1,1))
            attr=""
            if rs>1: attr+=f' rowspan="{rs}"'
            if cs>1: attr+=f' colspan="{cs}"'
            out.append(f'<td{attr} style="{";".join(styles)}">{val}</td>')
        out.append("</tr>")
    out.append("</table>")
    return "\n".join(out)

# column widths (approx px) for Ampol Retail
doc=["""<!doctype html><meta charset=utf-8><style>
body{font-family:Arial,Helvetica,sans-serif;background:#EEF2F3;margin:0;padding:24px;color:#1A1A1A}
h1{font-size:20px;color:#0B2E3C;margin:0 0 4px}
.meta{color:#5A6B6F;font-size:12px;margin-bottom:18px}
table.sheet{border-collapse:collapse;background:#fff;margin:0 0 34px;box-shadow:0 2px 10px rgba(0,0,0,.08);table-layout:fixed}
caption{caption-side:top;text-align:left;font-size:13px;font-weight:700;color:#1C5A66;padding:8px 2px}
td{border:1px solid #D4DEE0;padding:3px 7px;font-size:12px;white-space:nowrap;overflow:hidden;vertical-align:middle;max-width:230px}
</style>"""]
doc.append("<h1>TDD Cost Calc v9 (preview)</h1>")
doc.append('<div class="meta">Faithful render of the workbook cells &amp; formatting • generated '+datetime.date.today().isoformat()+'</div>')
import sys
PAGES = {
 "exec":   [("Exec Summary", 76, 7)],
 "port":   [("1.1 Ampol Retail", 62, 10), ("1.11 TDD Cyber", 42, 10)],
 "group":  [("2.0 Group Summary", 36, 10), ("2.1 Total Cost", 30, 12)],
 "coe":    [("2.2 COE", 24, 8), ("2.3 BP&T", 48, 10)],
 "sadcyb": [("2.4 SA&D", 45, 10), ("2.5 Cyber Roles", 74, 10)],
 "fte":    [("3.0 FTE View", 187, 15)],
 "gm":     [("4.1 Ampol Retail", 106, 10), ("4.11 TDD Cyber", 64, 10)],
 "qa":     [("3.1 Data QA", 190, 6)],
}
page = sys.argv[1] if len(sys.argv) > 1 else "exec"
for (nm, mr, mc) in PAGES[page]:
    doc.append(render_sheet(nm, mr, mc))
open(f"preview_{page}.html","w").write("\n".join(doc))
print(f"wrote preview_{page}.html")
