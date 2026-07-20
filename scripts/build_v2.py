#!/usr/bin/env python3
"""Extend the user's TDD_Cost_Calc.xlsx:
 - QA fixes on 1.1 Ampol Retail + Data Config
 - clean broken legacy defined names
 - build tabs 1.2-1.11 from squad mapping, cloning 1.1 styles exactly
 - add 2.0 Group Summary reconciliation tab
"""
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from copy import copy

SRC = "/root/.claude/uploads/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/08c7b762-TDD_Cost_Calc.xlsx"
OUT = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/TDD_Cost_Calc.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=False)
wbv = openpyxl.load_workbook(SRC, data_only=True)
tpl = wb["1.1 Ampol Retail"]

# =====================================================================
# 0. Freeze external-workbook references to their cached values.
#    0.3 links to '[1]TDD Pack (2)' (the original Cost_calculator file);
#    openpyxl strips cached values on save, which would leave permanent
#    #REF!. Replace those formulas with their cached literals.
# =====================================================================
import re as _re
EXT = _re.compile(r"\[\d+\]")
frozen = 0
for _ws in wb.worksheets:
    _wsv = wbv[_ws.title]
    for _row in _ws.iter_rows():
        for _c in _row:
            if isinstance(_c.value, str) and _c.value.startswith("=") and EXT.search(_c.value):
                _c.value = _wsv[_c.coordinate].value
                frozen += 1
wb._external_links = []
print("external-ref formulas frozen to cached values:", frozen)

# =====================================================================
# 1. Clean broken legacy defined names (keep the 4 list names + valid)
# =====================================================================
KEEP = {"SquadTypes", "SquadSizes", "OnOff", "SupportPct"}
removed = 0
for name in list(wb.defined_names.keys()):
    if name in KEEP:
        continue
    dn = wb.defined_names[name]
    v = str(dn.value) if dn.value is not None else ""
    if ("#REF!" in v) or ("#N/A" in v) or v == "" or name.startswith("\\") or name.startswith("_"):
        del wb.defined_names[name]
        removed += 1
    else:
        # drop names pointing at sheets that don't exist / external books
        target_ok = any(f"'{s.title}'" in v or v.startswith(s.title + "!") for s in wb.worksheets)
        if not target_ok:
            del wb.defined_names[name]
            removed += 1
for ws_ in wb.worksheets:
    try:
        for name in list(ws_.defined_names.keys()):
            del ws_.defined_names[name]
            removed += 1
    except Exception:
        pass
print("defined names removed:", removed, "kept:", list(wb.defined_names.keys()))

# =====================================================================
# 2. Surgical QA fixes on 1.1
# =====================================================================
tpl["B62"] = "Data AU Total"
for c in ["I33", "I40", "I47", "I53", "I59"]:
    tpl[c] = "Funded outside TDD ($m)"
# extend platform totals to cover the spare DV'd rows 35 / 42
tpl["G37"] = "=SUM(G34:G35)"
tpl["I37"] = "=SUM(I34:I35)"
tpl["G44"] = "=SUM(G41:G42)"
tpl["I44"] = "=SUM(I41:I42)"
# summary sums must include the spare rows too
tpl["C10"] = "=SUM(H25,H26,H27,H28,H34,H35,H41,H42,H48,H54,H60)"
tpl["D10"] = "=SUM(I25,I26,I27,I28,I34,I35,I41,I42,I48,I54,I60)"
tpl["E14"].comment = Comment(
    "Finance table Q9 = 33.9 but its own components (13.9+1.3+6.8+11.8) = 33.8. "
    "0.1 rounding sits in the Finance source table, not this model.", "QA")
dc = wb["0.0 Data Config"]
dc["C27"].comment = Comment(
    "= 0.4 Fin People AU (35.2) + 1 manual adjustment. Confirm the +1 is intended.", "QA")
sm = wb["squad mapping"]
sm["D7"].comment = Comment(
    "Shorthand types normalised in the calculator tabs: Config -> Configuration / Integration; "
    "Data -> Enterprise Data and Insights; Ops -> Operations; Support & Maintain -> Build and Run; "
    "Strat (CTRM) -> Configuration / Integration. Missing types/sizes given sensible defaults - "
    "all remain editable dropdowns on each tab.", "QA")

# =====================================================================
# 3. Style cloning helpers
# =====================================================================
def S(dst_ws, dst_coord, src_coord, value=None):
    src = tpl[src_coord]
    d = dst_ws[dst_coord]
    d._style = copy(src._style)
    if value is not None:
        d.value = value
    return d

def band(ws, row, c1, c2, src, text):
    for col in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=col)
        cell._style = copy(tpl[src]._style)
    ws.cell(row=row, column=c1).value = text
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)

COLL = "BCDEFGHIJ"

# =====================================================================
# 4. Tab specs from squad mapping (normalised defaults, all editable)
# =====================================================================
CI = "Configuration / Integration"
EDI = "Enterprise Data and Insights"
SPECS = [
 dict(tab="1.2 Customer", name="Customer",
      budget="='0.0 Data Config'!$E$13+'0.0 Data Config'!$E$14",
      budget_note="Ampol Customer (2.5) + Z Customer (2.5) from Data Config - mapping treats Customer as one portfolio spanning AU & NZ.",
      fin=dict(table=1, row=13), fin_note=None,
      platforms=[
        ("Ampol Digital", [("Ampol App","Engineering","L","Onshore",.2,None),
                           ("Ampol Web","Engineering","L","Onshore",.2,None),
                           ("Digital Operations","Operations","M","Onshore",.2,None)]),
        ("Customer Z", [("Z Energy Apps","Product","M","Onshore",.2,None),
                        ("Z Energy Martech","Product","L","Onshore",.2,None)]),
        ("Group Customer Platforms", [("Loyalty & Martech","Product","L","Onshore",.2,None)]),
      ]),
 dict(tab="1.3 Enterprise Data", name="Enterprise Data",
      budget="='0.0 Data Config'!$E$22",
      budget_note="TDD Data allocation (3.5) from Data Config.",
      fin=dict(table=2, row=27),
      fin_note="Closest Finance line is 'Strategy, Architecture and Data' (Matt Ashley) - includes Strategy & Architecture, not only data.",
      platforms=[
        ("Group Data", [("Data Science",EDI,"M","Onshore",.2,None),
                        ("Reporting & Analytics","Build and Run","L","Onshore",.2,"Mapping said 'Support & Maintain' - normalised to Build and Run."),
                        ("Data Platforms","Build and Run","L","Onshore",.2,"Mapping said 'Support & Maintain' - normalised to Build and Run."),
                        ("Enterprise Data Delivery",EDI,"M","Offshore",.2,"Size not set in mapping - defaulted M. Mapping flags this squad Offshore.")]),
      ]),
 dict(tab="1.4 TDD Group Functions", name="TDD Group Functions",
      budget="='0.0 Data Config'!$E$21",
      budget_note="TDD allocation (5.5) from Data Config.",
      fin=dict(table=2, row=25), fin_note=None,
      platforms=[
        ("TDD Group Functions", [("Workplace & Enterprise Tooling",CI,"S","Onshore",.2,None),
                                 ("Network & Infrastructure","Operations","M","Onshore",.2,"No type/size in mapping - defaulted Operations M."),
                                 ("DevOps & Engineering","Engineering","S","Onshore",.2,None),
                                 ("Integration",CI,"M","Onshore",.2,"No type/size in mapping - defaulted Configuration / Integration M.")]),
      ]),
 dict(tab="1.5 P&C", name="P&C",
      budget="='0.0 Data Config'!$E$18", budget_note=None,
      fin=dict(table=1, row=12), fin_note=None,
      platforms=[("P&C", [("P&C",CI,"M","Onshore",.2,None),
                          ("P&C - RTA",CI,"S","Onshore",.2,None)])]),
 dict(tab="1.6 Finance", name="Finance",
      budget="='0.0 Data Config'!$E$19", budget_note=None,
      fin=dict(table=1, row=11), fin_note=None,
      platforms=[("Finance", [("AU Finance",CI,"L","Onshore",.2,None),
                              ("NZ Finance",CI,"L","Onshore",.2,None)])]),
 dict(tab="1.7 Infrastructure", name="Infrastructure",
      budget="='0.0 Data Config'!$E$17", budget_note=None,
      fin=dict(table=1, row=10), fin_note=None,
      platforms=[
        ("Distribution", [("Distribution, Sales & Services",CI,"M","Onshore",.2,None)]),
        ("Manufacturing", [("Manufacturing & Group Projects",CI,"M","Onshore",.2,None),
                           ("Technology Support","Operations","S","Onshore",.2,None)]),
        ("Data & Insights", [("Data & Insights",CI,"S","Onshore",.2,None)]),
      ]),
 dict(tab="1.8 Energy Solutions & B2B", name="Energy Solutions & B2B",
      budget="='0.0 Data Config'!$E$16", budget_note=None,
      fin=dict(table=1, row=7),
      fin_note="Finance line 'EnergySolutions' - B2B has no separate Finance row. Components (4+1+2.3=7.3) vs Finance total 7.2: 0.1 rounding in the Finance table.",
      platforms=[
        ("Energy Solutions", [("Energy","Engineering","L","Onshore",.2,None),
                              ("EVCI","Engineering","M","Onshore",.2,None)]),
        ("B2B", [("B2B","Engineering","L","Onshore",.2,None)]),
      ]),
 dict(tab="1.9 Commercial Fuels", name="Commercial Fuels",
      budget="='0.0 Data Config'!$E$15", budget_note=None,
      fin=dict(table=1, row=8),
      fin_note="Components (13.3+6.1+14.2+6.8=40.4) vs Finance total 40.5: 0.1 rounding in the Finance table.",
      platforms=[
        ("Trading & Shipping", [("Trading & Shipping",CI,"M","Onshore",.2,None),
                                ("Trading & Shipping Data",EDI,"M","Onshore",.2,None)]),
        ("Supply", [("Supply",CI,"M","Onshore",.2,None)]),
        ("CTRM", [("CTRM",CI,"M","Onshore",.2,"Mapping said 'Strat' - normalised to Configuration / Integration (vendor-managed platform).")]),
      ]),
 dict(tab="1.10 Z Retail", name="Z Retail",
      budget="='0.0 Data Config'!$E$12", budget_note=None,
      fin=dict(table=1, row=6),
      fin_note="Finance line is the whole Z-Energy segment (incl. Z Customer) - no Z Retail-only Finance row exists.",
      platforms=[
        ("Z Supply", [("Z Supply",CI,"M","Onshore",.2,"No type/size in mapping - defaulted Configuration / Integration M.")]),
        ("Z Customer", [("Site Systems",CI,"M","Onshore",.2,"No type/size in mapping - defaulted Configuration / Integration M."),
                        ("Z Retail Backend","Product","M","Onshore",.2,"No type/size in mapping - defaulted Product M (was 'Balanced Product Squad' in the original op-model sheet).")]),
      ]),
 dict(tab="1.11 TDD Cyber", name="TDD Cyber",
      budget="='0.0 Data Config'!$E$23", budget_note=None,
      fin=dict(table=2, row=28),
      fin_note="Finance line 'Cyber Risk and Operations' is the full COE budget, wider than the TDD Cyber squad alone.",
      platforms=[("TDD Cyber", [("TDD Cyber","Operations","M","Onshore",.2,"No type/size in mapping - defaulted Operations M.")])]),
]

FIN = "'0.4 Budget Table (Fin)'"
def fin_refs(f):
    r = f["row"]
    if f["table"] == 1:   # cols: G people, L lights-on, M dep, N init, O sig, P capex, Q total
        return dict(people=f"={FIN}!$G${r}", lightson=f"={FIN}!$L${r}", init=f"={FIN}!$N${r}",
                    capex=f"={FIN}!$P${r}", sig=f"={FIN}!$O${r}", dep=f"={FIN}!$M${r}", total=f"={FIN}!$Q${r}")
    else:                  # cols: I people, N lights-on, O dep, P init, Q sig, R capex, S total
        return dict(people=f"={FIN}!$I${r}", lightson=f"={FIN}!$N${r}", init=f"={FIN}!$P${r}",
                    capex=f"={FIN}!$R${r}", sig=f"={FIN}!$Q${r}", dep=f"={FIN}!$O${r}", total=f"={FIN}!$S${r}")

LOOKUP = ('=IFERROR(IF($E{r}="Onshore",INDEX(\'0.1 Squads\'!$G$5:$G$23,MATCH($C{r}&"|"&$D{r},'
          "'0.1 Squads'!$A$5:$A$23,0)),INDEX('0.1 Squads'!$H$5:$H$23,MATCH($C{r}&\"|\"&$D{r},"
          "'0.1 Squads'!$A$5:$A$23,0))),\"check size\")")

# =====================================================================
# 5. Build each portfolio tab
# =====================================================================
tab_registry = []
base_index = wb.sheetnames.index("1.1 Ampol Retail")

for k, spec in enumerate(SPECS):
    ws = wb.create_sheet(spec["tab"], base_index + 1 + k)
    ws.sheet_view.showGridLines = False
    for col, dim in tpl.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[col].width = dim.width
    ws.row_dimensions[2].height = 21

    S(ws, "B2", "B2", spec["name"])

    # ---- build platform blocks first (start row 24) to collect refs ----
    pr = 24
    oh_cells, sq_h, sq_i, input_rows, plat_totals = [], [], [], [], []
    for pname, squads in spec["platforms"]:
        band(ws, pr, 2, 9, "B23", f"Platform: {pname}")
        pr += 1
        hdr = ["Squad", "Squad Type", "Size", "On/Off", "Support %",
               "Total Squad Cost ($m)", "TDD Cost ($m)", "Funded outside TDD ($m)"]
        for j, h in enumerate(hdr):
            S(ws, f"{COLL[j]}{pr}", f"{COLL[j]}24", h)
        pr += 1
        first = pr
        for (sqname, typ, size, shore, sup, note) in squads:
            S(ws, f"B{pr}", "B25", sqname)
            S(ws, f"C{pr}", "C25", typ)
            S(ws, f"D{pr}", "D25", size)
            S(ws, f"E{pr}", "E25", shore)
            S(ws, f"F{pr}", "F25", sup)
            S(ws, f"G{pr}", "G25", LOOKUP.format(r=pr))
            S(ws, f"H{pr}", "H25", f'=IFERROR($G{pr}*$F{pr},"")')
            S(ws, f"I{pr}", "I25", f'=IFERROR($G{pr}*(1-$F{pr}),"")')
            if note:
                ws[f"C{pr}"].comment = Comment(note, "QA")
            sq_h.append(f"H{pr}"); sq_i.append(f"I{pr}"); input_rows.append(pr)
            pr += 1
        last = pr - 1
        S(ws, f"B{pr}", "B29", "Platform Overhead")
        S(ws, f"H{pr}", "H29", "='0.0 Data Config'!$J$16")
        oh_cells.append(f"H{pr}")
        oh_row = pr
        pr += 1
        S(ws, f"B{pr}", "B30", f"{pname} Total")
        S(ws, f"G{pr}", "G30", f"=SUM(G{first}:G{last})")
        S(ws, f"H{pr}", "H30", f"=SUM(H{first}:H{oh_row})")
        S(ws, f"I{pr}", "I30", f"=SUM(I{first}:I{last})")
        for col in "CDEF":
            S(ws, f"{col}{pr}", f"{col}30")
        plat_totals.append(pr)
        pr += 2

    # ---- summary block ----
    band(ws, 6, 2, 5, "B6", "Portfolio Summary")
    S(ws, "G6", "G6", "Budget vs TDD cost")
    for j, h in enumerate(["Cost", "TDD ($m)", "Other ($m)", f"Total {spec['name']} Cost  ($m)"]):
        S(ws, f"{COLL[j]}7", f"{COLL[j]}7", h)
    S(ws, "B8", "B8", "Portfolio Overhead")
    S(ws, "C8", "C8", "='0.0 Data Config'!$J$10")
    S(ws, "E8", "E8", "=C8")
    S(ws, "B9", "B8", "Platform Overheads")
    S(ws, "C9", "C9" if tpl["C9"].value else "C8", f"=SUM({','.join(oh_cells)})")
    ws["C9"]._style = copy(tpl["C9"]._style)
    S(ws, "E9", "E8", "=C9")
    S(ws, "B10", "B8", "Squad Support Costs")
    S(ws, "C10", "C10", f"=SUM({','.join(sq_h)})")
    S(ws, "D10", "D10", f"=SUM({','.join(sq_i)})")
    S(ws, "E10", "E8", "=C10+D10")
    S(ws, "B11", "B11", "Total Cost")
    S(ws, "C11", "C11", "=SUM(C8:C10)")
    S(ws, "D11", "C11", "=SUM(D8:D10)")
    S(ws, "E11", "C11", "=C11+D11")
    # budget vs cost (right)
    S(ws, "G7", "G7", "TDD Lights On Budget")
    S(ws, "H7", "H7", spec["budget"])
    if spec.get("budget_note"):
        ws["H7"].comment = Comment(spec["budget_note"], "QA")
    S(ws, "G8", "G7", "Overheads & Support (this model)")
    S(ws, "H8", "H7", "=C11")
    ws["H8"].font = copy(tpl["H8"].font) if tpl["H8"].font else ws["H8"].font
    ws["H8"]._style = copy(tpl["H8"]._style)
    S(ws, "G9", "G9" if tpl["G9"].value else "G7", "Variance (budget less cost)")
    ws["G9"]._style = copy(tpl["G9"]._style)
    S(ws, "H9", "H9", "=H7-H8")

    # ---- funding block ----
    fr = fin_refs(spec["fin"])
    for col in "BCD":
        ws[f"{col}13"]._style = copy(tpl[f"{col}13"]._style)
        ws[f"{col}14"]._style = copy(tpl[f"{col}14"]._style)
    ws["B13"] = f"Total {spec['name']} budget"
    ws.merge_cells("B13:D13")
    S(ws, "E13", "E13", "=SUM(H14:H18)")
    ws["B14"] = "Reconciled to Finance"
    ws.merge_cells("B14:D14")
    S(ws, "E14", "E14", fr["total"])
    if spec.get("fin_note"):
        ws["E14"].comment = Comment(spec["fin_note"], "QA")
    band(ws, 12, 7, 10, "G12", "Other funding")
    for j, h in enumerate(["Budget line", "Budget ($m)", "Amount that can be allocated to people",
                           "Remaining for non-people ($m)"]):
        S(ws, f"{'GHIJ'[j]}13", f"{'GHIJ'[j]}13", h)
    fund = [("%s Lights On" % spec["name"], fr["lightson"], fr["people"], True),
            ("OpEx Initiatives", fr["init"], 0, False),
            ("%s CapEx" % spec["name"], fr["capex"], 0, False),
            ("Significant Items", fr["sig"], 0, False),
            ("Depreciation", fr["dep"], 0, None)]
    r = 14
    for (label, href, ival, ilink) in fund:
        S(ws, f"G{r}", "G14", label)
        S(ws, f"H{r}", "H14", href)
        if ilink is True:
            S(ws, f"I{r}", "H14", ival)          # green link style (formula ref)
        elif ilink is False:
            S(ws, f"I{r}", "I16", ival)          # yellow editable input
        else:
            S(ws, f"I{r}", "H25", ival)          # plain (dep cannot fund people)
            ws[f"I{r}"].comment = Comment("Depreciation cannot fund people - shown so the total reconciles to Finance.", "QA")
        S(ws, f"J{r}", "J14" if tpl["J14"].value else "H25", f"=IFERROR(H{r}-I{r},0)")
        ws[f"J{r}"]._style = copy(tpl["J14"]._style)
        r += 1
    S(ws, "G19", "G18", "Total applied")
    S(ws, "I19", "I18" if tpl["I18"].value else "H25", "=SUM(I14:I18)")
    ws["I19"]._style = copy(tpl["I18"]._style)
    S(ws, "G20", "G19", "Other cost (this model)")
    S(ws, "I20", "I19" if tpl["I19"].value else "H25", "=D11")
    ws["I20"]._style = copy(tpl["I19"]._style)
    S(ws, "G21", "G20", "Left to fund")
    S(ws, "I21", "I20" if tpl["I20"].value else "H25", "=I20-I19")
    ws["I21"]._style = copy(tpl["I20"]._style)

    # ---- data validations ----
    dvt = DataValidation(type="list", formula1="SquadTypes", allow_blank=True)
    dvs = DataValidation(type="list", formula1="SquadSizes", allow_blank=True)
    dvo = DataValidation(type="list", formula1="OnOff", allow_blank=True)
    dvp = DataValidation(type="list", formula1="SupportPct", allow_blank=True)
    for dv in (dvt, dvs, dvo, dvp):
        ws.add_data_validation(dv)
    for rr in input_rows:
        dvt.add(f"C{rr}"); dvs.add(f"D{rr}"); dvo.add(f"E{rr}"); dvp.add(f"F{rr}")

    tab_registry.append(dict(tab=spec["tab"], name=spec["name"]))

# =====================================================================
# 6. 2.0 Group Summary
# =====================================================================
gs = wb.create_sheet("2.0 Group Summary", base_index + 1 + len(SPECS))
gs.sheet_view.showGridLines = False
for col, w in {"A": 8.6, "B": 30, "C": 18, "D": 16, "E": 16, "F": 20, "G": 16, "H": 3}.items():
    gs.column_dimensions[col].width = w
gs.row_dimensions[2].height = 21
S(gs, "B2", "B2", "Group Summary - all portfolios")
band(gs, 4, 2, 7, "B6", "TDD cost vs TDD lights-on budget, by portfolio")
for j, h in enumerate(["Portfolio", "TDD Lights On Budget ($m)", "TDD Cost ($m)",
                       "Variance ($m)", "Funded outside TDD ($m)", "Total Cost ($m)"]):
    S(gs, f"{COLL[j]}5", f"{COLL[j]}7" if j < 4 else "E7", h)
    gs[f"{COLL[j]}5"]._style = copy(tpl["B7"]._style)
all_tabs = [("1.1 Ampol Retail", "Ampol Retail")] + [(t["tab"], t["name"]) for t in tab_registry]
r = 6
first = r
for (tab, name) in all_tabs:
    S(gs, f"B{r}", "B8", name)
    S(gs, f"C{r}", "H25", f"='{tab}'!$H$7")
    S(gs, f"D{r}", "H25", f"='{tab}'!$C$11")
    S(gs, f"E{r}", "H25", f"='{tab}'!$H$9")
    S(gs, f"F{r}", "H25", f"='{tab}'!$D$11")
    S(gs, f"G{r}", "H25", f"='{tab}'!$E$11")
    r += 1
last = r - 1
S(gs, f"B{r}", "B11", "Total - all portfolios")
for col in "CDEFG":
    S(gs, f"{col}{r}", "C11", f"=SUM({col}{first}:{col}{last})")
tot_row = r
r += 3
band(gs, r, 2, 7, "B6", "Reconciliation to Data Config TDD budget allocation")
r += 1
S(gs, f"B{r}", "B8", "Portfolio budgets used above")
S(gs, f"C{r}", "H25", f"=C{tot_row}")
r += 1
S(gs, f"B{r}", "B8", "COE allocations (Data Config rows 6-10)")
S(gs, f"C{r}", "C8", "=SUM('0.0 Data Config'!$E$6:$E$10)")
r += 1
S(gs, f"B{r}", "B11", "Total allocation accounted for")
S(gs, f"C{r}", "C11", f"=C{r-2}+C{r-1}")
r += 1
S(gs, f"B{r}", "B8", "Data Config total allocation (E26)")
S(gs, f"C{r}", "C8", "='0.0 Data Config'!$E$26")
r += 1
S(gs, f"B{r}", "B11", "Check (should be 0)")
S(gs, f"C{r}", "C11", f"=C{r-2}-C{r-1}")

wb.save(OUT)
print("Saved", OUT)
print("Sheet order:", wb.sheetnames)
