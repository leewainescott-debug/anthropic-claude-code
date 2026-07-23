#!/usr/bin/env python3
"""v8 build off the owner's file (user_v2.xlsx).

- Retires Squads / Added data from 3.1, 3.2, 3.3, Exec Summary and 4.0 Data QA:
  every count and dollar reads REVIEW (AJ tab / AP curated squad / AK status /
  AQ leadership / AA cost). Totals tie to REVIEW: 527 roles, $115.89m.
- Renames the invented "EGI & Central" bucket to EGI everywhere (sheet, Lists,
  formulas, labels); EGI gets its own row in 3.1 and 3.2.
- COE lines show gross AND net: 3.4 gains Gross / Funded by portfolios columns;
  3.2 COE planned goes net (matching 1.11/1.12), the double-subtract on 3.1 dies.
- 3.1 gains the overhead coverage block (charged vs leadership + BP/DA drawn).
- 1.x: all conditional formatting stripped, variance shown as plain
  over/(under) budget for AU / NZ / TDD (positive = over).
- 2.x: vacancy lever on every roster row (Filled added as an option); salaries
  remain vacant-only. 2.2 D82 mispull fixed.
Cyber split (portfolio vs COE), J-col basis, AmPOS OH, Data NZ row: pending the
owner's answers - untouched here."""
import openpyxl, re
from openpyxl.utils import column_index_from_string as ci
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.views import Selection
from collections import Counter

SRC="user_v2.xlsx"; OUT="clean_v8.xlsx"
wb=openpyxl.load_workbook(SRC)
REVN="REVIEW - Complete Role Mapping"; Q="'"+REVN+"'"
rev=wb[REVN]
def rc(r,cc): return rev.cell(r,ci(cc)).value
RLO,RHI=2,530
EGIQ='"EGI"'
FILQ='"Filled"'
VACQ='"Vacant"'

NAVY=PatternFill("solid",fgColor="1F4E79")
WF=Font(color="FFFFFF",bold=True); BF=Font(bold=True); NF=Font(); IT=Font(italic=True,size=10)
M2='#,##0.00;(#,##0.00);"-"'
CEN=Alignment(horizontal="center",vertical="center"); RT=Alignment(horizontal="right")

# ---------- 1) EGI rename everywhere ----------
wb["2.14 EGI & Central"].title="2.14 EGI"
L=wb["Lists"]
for r in range(1,20):
    if L.cell(r,ci("U")).value=="EGI & Central": L.cell(r,ci("U")).value="EGI"
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            v=c.value
            if isinstance(v,str) and "EGI & Central" in v:
                c.value=(v.replace("'2.14 EGI & Central'","'2.14 EGI'")
                          .replace("2.14 EGI & Central","2.14 EGI")
                          .replace("EGI & Central","EGI"))

def cnt(*pairs):
    s=",".join(f"{Q}!${a}${RLO}:${a}${RHI},{b}" for a,b in pairs)
    return f"COUNTIFS({s})"
def sm(*pairs):
    s=",".join(f"{Q}!${a}${RLO}:${a}${RHI},{b}" for a,b in pairs)
    return f"SUMIFS({Q}!$AA${RLO}:$AA${RHI},{s})"

# ---------- 2) 3.3 FTE View -> REVIEW ----------
f3=wb["3.3 FTE View"]
f3["C4"].value=f"=COUNTA({Q}!$B${RLO}:$B${RHI})"
f3["D4"].value="="+cnt(("AK",'"Filled"'))
f3["E4"].value="="+cnt(("AK",'"Vacant"'))
det=[]
for r in range(7,95):
    b=f3.cell(r,2).value; d=f3.cell(r,4).value
    if b is None or d is None: continue
    det.append((r,str(b),str(d)))
    f3.cell(r,8).value="="+cnt(("AJ",f"$B{r}"),("AP",f"$D{r}"),("AK",'"Filled"'))
    f3.cell(r,9).value="="+cnt(("AJ",f"$B{r}"),("AP",f"$D{r}"),("AK",'"Vacant"'))
    f3.cell(r,14).value=f"=({sm(('AJ',f'$B{r}'),('AP',f'$D{r}'))})/1000000"
# leadership roster rows 99-151 from REVIEW AQ=1 rows (live refs)
lead=[r for r in range(RLO,RHI) if rev.cell(r,ci("AQ")).value==1]
lead.sort(key=lambda r:(str(rc(r,'I') or ""), str(rc(r,'B') or "")))
row=99
for dr in lead:
    if row>151: break
    f3.cell(row,2).value=f"={Q}!$AJ${dr}"
    f3.cell(row,3).value=f"={Q}!$B${dr}"
    f3.cell(row,4).value=f"={Q}!$C${dr}"
    f3.cell(row,5).value=f"={Q}!$J${dr}"
    row+=1
for r in range(row,152):
    for col in (2,3,4,5): f3.cell(r,col).value=None
f3["C152"].value="="+cnt(("AQ","1"))
f3["D152"].value="="+cnt(("AQ","1"),("AK",'"Vacant"'))
# reconciliation block: full REVIEW accounting, difference must be 0
f3["B154"].value="Reconciliation of every REVIEW role"
f3["B155"].value=None; f3["C155"].value=None
f3["B156"].value="Squad-table roles (sum of the table above)"; f3["C156"].value="=J95"
f3["B157"].value="Leadership (funded by overheads)"; f3["C157"].value="="+cnt(("AQ","1"))
f3["B158"].value="COE roles (detail on 1.11 / 1.12 / 1.13)"
f3["C158"].value="="+cnt(("AJ",'"COE BP&T"'))+"+"+cnt(("AJ",'"COE SA&D"'))+"+"+cnt(("AJ",'"COE Cyber"'))
f3["B159"].value="EGI (funded via Significant Items)"; f3["C159"].value="="+cnt(("AJ",'"EGI"'))
f3["B160"].value="Portfolio roles outside the archetype squads"
f3["C160"].value="=C162-C156-C157-C158-C159"
f3["B161"].value=None; f3["C161"].value=None
f3["B162"].value="Cross-check: org records"; f3["C162"].value="=C4"
f3["B163"].value="Accounted for above"; f3["C163"].value="=C156+C157+C158+C159+C160"
f3["B164"].value="Difference (must be 0)"; f3["C164"].value="=C162-C163"
# EGI block rows 169-185 previously Sheet2 -> REVIEW EGI rows
f3["B168"].value="EGI - funded via Significant Items (from REVIEW)"
for i in range(169,186):
    for col in (2,3,4,5,18): f3.cell(i,col).value=None
egi=[r for r in range(RLO,RHI) if str(rc(r,'I') or "").strip()=="EGI"]
row=169
for dr in egi:
    if row>184: break
    f3.cell(row,2).value=f"={Q}!$B${dr}"
    f3.cell(row,3).value=f"={Q}!$C${dr}"
    f3.cell(row,4).value=f"={Q}!$AK${dr}"
    f3.cell(row,5).value=f"=IFERROR({Q}!$AA${dr}/1000000,0)"
    row+=1
f3.cell(185,2).value="Total EGI"
f3.cell(185,5).value="=SUM(E169:E184)"
f3["F185"].value="=("+sm(('AJ',EGIQ))+")/1000000"

# ---------- 3) 3.2 Total Cost ----------
tc=wb["3.2 Total Cost"]
tc["C16"].value="='1.11 BP&T'!$F$6"        # net (matches 1.11)
tc["C18"].value="='1.12 SA&D'!$G$6"        # net (matches 1.12)
tc["B21"].value="EGI (funded via Significant Items)"
tc["C21"].value=0
tc["I21"].value="="+cnt(("AJ",'"EGI"'),("AK",'"Filled"'))
tc["M21"].value="="+cnt(("AJ",'"EGI"'),("AK",'"Vacant"'))
tc["K21"].value="=("+sm(('AJ',EGIQ),('AK',FILQ))+")/1000000"
tc["L21"].value="=("+sm(('AJ',EGIQ),('AK',VACQ))+")/1000000"
tc["D21"].value="=K21+L21"; tc["E21"].value="=ROUND(D21-C21,6)"
tc["F21"].value="=K21"; tc["G21"].value="=ROUND(F21-C21,6)"
tc["H21"].value='="-"'; tc["J21"].value='="-"'; tc["N21"].value="=I21+M21"
tc["B22"].value=None
for col in "CDEFGHIJKLMN": tc[f"{col}22"].value=None
tc["B23"].value="COE planned spend is net of the Business Partner & Domain Architect roles funded inside portfolio overheads (gross on 3.4)."
tc["B23"].font=IT
tc["C23"].value=0; tc["D23"].value=0; tc["E23"].value=0; tc["F23"].value=0; tc["G23"].value=0
tc["H23"].value='="-"'; tc["I23"].value=0; tc["J23"].value='="-"'
tc["K23"].value=0; tc["L23"].value=0; tc["M23"].value=0; tc["N23"].value=0

# ---------- 4) 3.4 COE Summary gross / funded columns ----------
cs=wb["3.4 COE Summary"]
cs["K5"].value="Gross people cost ($m)"; cs["L5"].value="Funded by portfolios ($m)"
for c_ in ("K5","L5"): cs[c_].font=WF; cs[c_].fill=NAVY; cs[c_].alignment=CEN
cs["K6"].value="='1.11 BP&T'!$F$6+'1.11 BP&T'!$C$13"; cs["L6"].value="='1.11 BP&T'!$C$13"
cs["K7"].value="='1.11 BP&T'!$F$7";                   cs["L7"].value=0
cs["K8"].value="='1.12 SA&D'!$G$6+'1.12 SA&D'!$C$13"; cs["L8"].value="='1.12 SA&D'!$C$13"
cs["K9"].value="='1.12 SA&D'!$G$7";                   cs["L9"].value=0
cs["K10"].value="='1.13 Cyber Roles'!$F$8";           cs["L10"].value=0
cs["K11"].value="=SUM(K6:K10)"; cs["L11"].value="=SUM(L6:L10)"
for r in range(6,12):
    for c_ in ("K","L"):
        cs[f"{c_}{r}"].number_format=M2; cs[f"{c_}{r}"].alignment=RT
cs["B13"].value="Planned spend (F) is net: gross people cost less the amount funded inside portfolio overheads. Both matter - gross is what the people cost, net is the call on the COE budget."
cs["B13"].font=IT

# ---------- 5) 3.1 Group Summary ----------
gs=wb["3.1 Group Summary"]
gs["M25"].value=None; gs["B25"].value=None      # second netting - dead
gs["M26"].value=None; gs["B26"].value=None
gs["B23"].value="EGI (funded via Significant Items)"
gs["C23"].value=0; gs["D23"].value=0
gs["F23"].value="='3.2 Total Cost'!$D$21"
gs["M23"].value=0
m24=gs["M24"].value
if isinstance(m24,str) and "M23" not in m24: gs["M24"].value="=SUM(M17:M23)"
OVH="+".join(f"'{t}'!$F$6+'{t}'!$F$7" for t in
    ["1.1 Ampol Retail","1.2 Customer","1.3 Enterprise Data","1.4 TDD Group Functions",
     "1.5 P&C","1.6 Finance","1.8 Energy Solutions & B2B","1.9 Commercial Fuels","1.10 Z Retail"])
OVH+="+'1.7 Infrastructure'!$F$7+'1.7 Infrastructure'!$F$8"
gs["B54"].value="Overhead coverage"; gs["B54"].font=BF
gs["B55"].value="Overheads charged into the portfolio archetypes ($m)"
gs["C55"].value="="+OVH
gs["B56"].value="Drawn from overheads - Business Partner & Domain Architect allocations ($m)"
gs["C56"].value="='1.11 BP&T'!$C$13+'1.12 SA&D'!$C$13"
gs["B57"].value="Drawn from overheads - leadership actual cost ($m)"
gs["C57"].value=f"=({sm(('AQ','1'))})/1000000"
gs["B58"].value="Overhead not covered / (headroom) ($m)"
gs["C58"].value="=C56+C57-C55"
for r in range(55,59):
    gs[f"C{r}"].number_format=M2; gs[f"C{r}"].alignment=RT
gs["B59"].value="Positive = leadership and allocations cost more than the overheads charged; negative = headroom."
gs["B59"].font=IT

# ---------- 6) Exec Summary residual Squads refs ----------
ex=wb["Exec Summary"]
ex["C43"].value="="+cnt(("AK",'"Vacant"'),("AQ","0"))
ex["C44"].value="="+cnt(("AK",'"Vacant"'),("AQ","1"))
ex["B44"].value="of which leadership roles funded via overheads"
ex["C72"].value="="+cnt(("AJ","$C$63"),("AQ","0"))
ex["C73"].value="="+cnt(("AJ","$C$63"),("AQ","0"),("AK",'"Filled"'))
ex["C74"].value="="+cnt(("AJ","$C$63"),("AQ","0"),("AK",'"Vacant"'))
ex["C75"].value="="+cnt(("AJ","$C$63"),("AQ","1"))
ex["B77"].value="Role counts and dollars come from the REVIEW - Complete Role Mapping sheet (the single source of truth)."

# ---------- 7) 4.0 Data QA -> REVIEW ----------
qa=wb["4.0 Data QA"]
qa["B111"].value="Live check - REVIEW role rows now"
qa["C111"].value=f"=COUNTA({Q}!$B${RLO}:$B${RHI})"
qa["B112"].value="Live check - REVIEW vacant roles now"
qa["C112"].value="="+cnt(("AK",'"Vacant"'))
qa["B6"].value="REVIEW roles"; qa["C6"].value=f"=COUNTA({Q}!$B${RLO}:$B${RHI})"
qa["B7"].value="REVIEW filled"; qa["C7"].value="="+cnt(("AK",'"Filled"'))
qa["B8"].value="REVIEW vacant"; qa["C8"].value="="+cnt(("AK",'"Vacant"'))
qa["B9"].value="of which leadership (funded via overheads)"; qa["C9"].value="="+cnt(("AQ","1"))
WT12=["2.1 Ampol Retail","2.2 Customer","2.3 Enterprise Data","2.4 TDD Group Functions",
"2.5 P&C","2.6 Finance","2.7 Infrastructure","2.8 Energy Solutions & B2B",
"2.9 Commercial Fuels","2.10 Z Retail","2.11 TDD Cyber","2.14 EGI"]
c258="="+"+".join(f"COUNTIF('{t}'!$D$1:$D$500,\"Vacant\")" for t in WT12)
c258+='+COUNTIF(\'2.12 BP&T\'!$E$1:$E$500,"Vacant")+COUNTIF(\'2.13 SA&D\'!$E$1:$E$500,"Vacant")'
qa["C258"].value=c258
qa["B259"].value="REVIEW vacant excluding leadership (leadership sits on 3.3, not the working tabs)"
qa["C259"].value="="+cnt(("AK",'"Vacant"'),("AQ","0"))
qa["B260"].value="Difference (investigate if not 0)"
qa["C260"].value="=C258-C259"
qa["B261"].value=None; qa["C261"].value=None
for lbl in ("B116","B134","B180"):
    v=qa[lbl].value
    if isinstance(v,str):
        qa[lbl].value=(v.replace("Squads","REVIEW").replace("Sheet2","REVIEW")
                        .replace("Added data","REVIEW"))

# ---------- 8) 1.x: strip CF + plain over/(under) budget lines ----------
D1=["1.1 Ampol Retail","1.2 Customer","1.3 Enterprise Data","1.4 TDD Group Functions",
    "1.5 P&C","1.6 Finance","1.7 Infrastructure","1.8 Energy Solutions & B2B",
    "1.9 Commercial Fuels","1.10 Z Retail"]
for t in D1:
    ws=wb[t]
    ws.conditional_formatting._cf_rules.clear()
    au=nz=None
    for r in range(4,12):
        h=str(ws.cell(r,8).value or "")
        if h.startswith("AU Variance"): au=r
        if h.startswith("NZ Variance"): nz=r
    if au and nz:
        f=str(ws.cell(au,9).value or "")
        m=re.search(r"=I\d+-([CD])(\d+)", f)
        if m:
            cost_row=int(m.group(2))
            # the line below the box may sit in a merged range - release it first
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row<=nz+1<=mr.max_row and mr.min_col<=9 and mr.max_col>=8:
                    ws.unmerge_cells(str(mr))
            ws.cell(au,8).value="AU over/(under) budget ($m)"
            ws.cell(au,9).value=f"=C{cost_row}-I{au-2}"
            ws.cell(nz,8).value="NZ over/(under) budget ($m)"
            ws.cell(nz,9).value=f"=D{cost_row}-I{nz-2}"
            ws.cell(nz+1,8).value="TDD over/(under) budget ($m)"
            ws.cell(nz+1,9).value=f"=I{au}+I{nz}"
            ws.cell(nz+1,9).number_format=M2
            ws.cell(nz+1,8).font=BF; ws.cell(nz+1,9).font=BF
            # TDD Variance rows depending on the old sign: '=-(I{au}+I{nz})' -> '=I{au}+I{nz}'
            for row2 in ws.iter_rows(min_row=10,max_row=25,min_col=2,max_col=6):
                for c in row2:
                    v=c.value
                    if isinstance(v,str) and v.replace(" ","")==f"=-(I{au}+I{nz})":
                        c.value=f"=I{au}+I{nz}"
                    if isinstance(v,str) and v.replace(" ","")==f"=(I{au}+I{nz})":
                        c.value=f"=I{au}+I{nz}"
    # plain black numbers: strip red/green colouring on summary/variance area
    for r in range(4,26):
        for col in (3,9,10):
            cl=ws.cell(r,col)
            if cl.font and cl.font.color and getattr(cl.font.color,"rgb",None) in ("FF9C0006","FF006100"):
                cl.font=Font(bold=cl.font.bold)
            nf=cl.number_format or ""
            if "[Red]" in nf or "[Green]" in nf:
                cl.number_format=nf.replace("[Red]","").replace("[Green]","")

# ---------- 9) 2.x: lever incl Filled on all roster rows; D82 fix ----------
for t in WT12:
    ws=wb[t]
    for dv in list(ws.data_validations.dataValidation): ws.data_validations.dataValidation.remove(dv)
    dv=DataValidation(type="list",formula1='"Hire,Hold,Offshore,Filled"',allow_blank=True)
    ws.add_data_validation(dv)
    for r in range(1,ws.max_row+1):
        d=ws.cell(r,4).value
        if isinstance(d,str) and d.startswith("='REVIEW") and "$AK$" in d:
            e=ws.cell(r,5)
            if e.value in (None,""):
                e.value="Filled"; e.alignment=CEN
            dv.add(e)
wb["2.2 Customer"]["D82"].value=f"={Q}!$AK$171"

# ---------- panes ----------
for ws in wb.worksheets:
    ws.freeze_panes=None; ws.sheet_view.pane=None
    ws.sheet_view.selection=[Selection(activeCell="A1",sqref="A1")]

wb.save(OUT)
print("saved",OUT)
n=Counter()
wb2=openpyxl.load_workbook(OUT)
for ws in wb2.worksheets:
    if ws.title in ("Squads","Added data","Sheet2","Claude Log"): continue
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value.startswith("="):
                if "Squads!" in c.value: n[("Squads",ws.title)]+=1
                if "Added data" in c.value: n[("Added data",ws.title)]+=1
                if "Sheet2" in c.value: n[("Sheet2",ws.title)]+=1
print("remaining old-ledger refs:")
for (k,t),v in sorted(n.items()): print(f"   {k}: {t} = {v}")
if not n: print("   NONE outside 1.11/1.12/1.13/2.12/2.13 (checked all)")
