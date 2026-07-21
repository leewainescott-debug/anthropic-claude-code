#!/usr/bin/env python3
"""Implement user's round-4 changes on their FINAL file (surgical, style-preserving):
 1. Strategic Programs treatment: AmPOS (1.1), CTRM (1.9), EGI Retail (1.1, finish
    half-built block), + new EGI platforms on 1.2/1.4/1.5/1.6. Hard-coded yellow
    squad cost, no size, Onshore, support editable; 'other' flows into the tab's
    Significant Items funded cell (left-to-fund unaffected).
 2. Double-counting fix on 1.3/1.4/1.11: lights-on is not a drawdown envelope.
 3. 2.0: Funded / Left to fund columns; COE row in main table; ladder to Data
    Config allocation (43.5) and total TDD budget (53.8) with headroom.
 4. 2.1 COE: project funding envelopes (TDDEGM capex 4.9 / sig 20.2 / init 3.0)
    with drawdowns incl. EGI TDD.
 5. Data Config: complete the Spend/Variance wiring the user started.
 6. Replicate user's 'Total to fund' line to every tab.
"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

U = "/root/.claude/uploads/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/"
SRC = U + "4519c91f-TDD_Cost_Calc_FINAL.xlsx"
OUT = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/TDD_Cost_Calc_v4.xlsx"
FIN = "'0.4 Budget Table (Fin)'"

wb = openpyxl.load_workbook(SRC, data_only=False)

# ---------------- helpers ----------------
def frow(ws, col, text, exact=True, start=1):
    for r in range(start, ws.max_row + 2):
        v = ws.cell(r, openpyxl.utils.column_index_from_string(col)).value
        if v is None: continue
        s = str(v).strip()
        if (exact and s == text) or (not exact and text in s):
            return r
    return None

def anchors(ws):
    a = {}
    a["totalcost"]  = frow(ws, "B", "Total Cost")
    a["platoh_sum"] = frow(ws, "B", "Platform Overheads")
    a["squadsum"]   = frow(ws, "B", "Squad Support Costs")
    a["variance"]   = frow(ws, "G", "Variance (budget less cost)")
    hdr = frow(ws, "G", "Budget line") or 1
    a["sig"]        = frow(ws, "G", "Significant Items", start=hdr)
    a["lightson"]   = frow(ws, "G", "Lights On", exact=False, start=hdr)
    a["applied"]    = frow(ws, "G", "Total applied")
    a["leftfund"]   = frow(ws, "G", "Left to fund")
    return a

def blocks(ws):
    """[(title_row, header_row, first_squad, last_squad, oh_row, total_row)]"""
    out = []
    rows = [r for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 2).value, str) and ws.cell(r, 2).value.startswith("Platform: ")]
    for tr in rows:
        hr = tr + 1
        ohr = None
        for r in range(hr + 1, ws.max_row + 2):
            if ws.cell(r, 2).value == "Platform Overhead":
                ohr = r; break
        out.append((tr, hr, hr + 1, ohr - 1, ohr, ohr + 1))
    return out

def rebuild_summary_sums(ws, a):
    bl = blocks(ws)
    oh = [f"H{b[4]}" for b in bl]
    sq = []
    for b in bl:
        sq += [r for r in range(b[2], b[3] + 1)]
    ws.cell(a["platoh_sum"], 3).value = f"=SUM({','.join(oh)})"
    ws.cell(a["squadsum"], 3).value = "=SUM(" + ",".join(f"H{r}" for r in sq) + ")"
    ws.cell(a["squadsum"], 4).value = "=SUM(" + ",".join(f"I{r}" for r in sq) + ")"

LOOKUP_TAIL = "'0.1 Squads'"
def strategic_convert(ws, r, seed, note):
    """Turn squad row r into a Strategic Programs row with hard-coded cost."""
    c = ws.cell(r, 3); c.value = "Strategic Programs"
    ws.cell(r, 4).value = None                      # no size
    ws.cell(r, 5).value = "Onshore"                 # always on
    if ws.cell(r, 6).value in (None, ""):
        ws.cell(r, 6).value = 0                     # support editable, default 0
    g = ws.cell(r, 7)
    # yellow hard-coded money input: clone style from the C (yellow) cell, money fmt from H
    g._style = copy(ws.cell(r, 3)._style)
    g.number_format = ws.cell(r, 8).number_format
    g.font = Font(name="Calibri", size=10, color="FF0000FF")
    g.value = seed
    g.comment = Comment(note, "QA")
    # H / I formulas stay (G*F and G*(1-F))
    ws.cell(r, 8).value = f'=IFERROR($G{r}*$F{r},"")'
    ws.cell(r, 9).value = f'=IFERROR($G{r}*(1-$F{r}),"")'

def clone_row_style(ws, src_r, dst_r, cols=range(2, 11)):
    for c in cols:
        ws.cell(dst_r, c)._style = copy(ws.cell(src_r, c)._style)

def add_platform_block(ws, pname, sqname, seed, note):
    """Append a strategic-program platform block, cloning styles from the last block."""
    tpl_t, tpl_h, tpl_s0, tpl_s1, tpl_oh, tpl_tot = blocks(ws)[-1]
    r0 = ws.max_row + 2
    # title band (merged B:I like existing)
    clone_row_style(ws, tpl_t, r0)
    ws.cell(r0, 2).value = f"Platform: {pname}"
    for m in list(ws.merged_cells.ranges):
        pass
    ws.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=9)
    # header
    clone_row_style(ws, tpl_h, r0 + 1)
    for j, h in enumerate(["Squad", "Squad Type", "Size", "On/Off", "Support %",
                           "Total Squad Cost ($m)", "TDD Cost ($m)", "Funded outside TDD ($m)"]):
        ws.cell(r0 + 1, 2 + j).value = h
    # squad row
    clone_row_style(ws, tpl_s0, r0 + 2)
    ws.cell(r0 + 2, 2).value = sqname
    ws.cell(r0 + 2, 6).value = 0
    strategic_convert(ws, r0 + 2, seed, note)
    # overhead row
    clone_row_style(ws, tpl_oh, r0 + 3)
    ws.cell(r0 + 3, 2).value = "Platform Overhead"
    ws.cell(r0 + 3, 8).value = "='0.0 Data Config'!$L$16"
    ws.cell(r0 + 3, 9).value = None
    # total row
    clone_row_style(ws, tpl_tot, r0 + 4)
    ws.cell(r0 + 4, 2).value = f"{pname} Total"
    ws.cell(r0 + 4, 7).value = f"=SUM(G{r0+2}:G{r0+2})"
    ws.cell(r0 + 4, 8).value = f"=SUM(H{r0+2}:H{r0+3})"
    ws.cell(r0 + 4, 9).value = f"=SUM(I{r0+2}:I{r0+2})"
    return r0 + 2   # squad row

def wire_sig(ws, a, extra_cells, keep_base=True):
    """Sig-items funded cell = existing base + strategic-program other cells."""
    cell = ws.cell(a["sig"], 9)
    base = cell.value
    if not keep_base or base in (None, "") or isinstance(base, str):
        base_txt = "0"
    else:
        base_txt = repr(base)
    cell.value = "=" + base_txt + "".join(f"+{c}" for c in extra_cells)
    cell.comment = Comment(
        "Strategic Programs draw down directly against Significant Items: "
        "base amount + " + " + ".join(extra_cells) + ". Their cost raises total squad cost "
        "and is funded here, so Left to fund is unchanged by them.", "QA")

def add_dv_cells(ws, row):
    for dv in ws.data_validations.dataValidation:
        f = str(dv.formula1)
        if f == "SquadTypes": dv.add(f"C{row}")
        elif f == "SquadSizes": dv.add(f"D{row}")
        elif f == "OnOff": dv.add(f"E{row}")
        elif f == "SupportPct": dv.add(f"F{row}")

# ---------------- 0. Lists: add Strategic Programs type ----------------
ls = wb["Lists"]
ls["A8"] = "Strategic Programs"
wb.defined_names["SquadTypes"].value = "Lists!$A$2:$A$8"

SEED_NOTE = ("Hard-coded strategic-program cost (seeded from the actual cost in the original "
             "cost calculator - overwrite with the agreed programme cost).")

# ---------------- 1.1 Ampol Retail ----------------
ws = wb["1.1 Ampol Retail"]; a = anchors(ws)
# finish the half-built EGI block
ws["B64"] = "EGI Retail"; ws["B66"] = "EGI Total"
ws["F64"] = 0            # leftover 0.2 from the copied row; strategic default = 0
strategic_convert(ws, 64, 0.97, SEED_NOTE)
add_dv_cells(ws, 64)
# AmPOS -> strategic (user note J46)
strategic_convert(ws, 46, 2.66, SEED_NOTE)
rebuild_summary_sums(ws, a)
wire_sig(ws, a, ["I46", "I64"])          # base 2.3 preserved

# ---------------- 1.2 Customer: convert the user's own EGI block ----------------
ws = wb["1.2 Customer"]; a = anchors(ws)
ws["B49"] = "EGI Customer"; ws["B51"] = "EGI Total"   # was copy-paste 'Group Customer Platforms Total'
strategic_convert(ws, 49, 2.26, SEED_NOTE)
add_dv_cells(ws, 49)
rebuild_summary_sums(ws, a)
wire_sig(ws, a, ["I49"])

# ---------------- new EGI platforms (tabs with no existing EGI block) ----------------
EGI_ADDS = [
 ("1.4 TDD Group Functions", "EGI TDD",      "EGI TDD",      1.06),
 ("1.5 P&C",                 "EGI P&C",      "EGI P&C",      0.24),
 ("1.6 Finance",             "EGI Finance",  "EGI Finance",  0.24),
]
egi_rows = {"1.2 Customer": 49}
for tab, pname, sqname, seed in EGI_ADDS:
    ws = wb[tab]; a = anchors(ws)
    sr = add_platform_block(ws, pname, sqname, seed, SEED_NOTE)
    add_dv_cells(ws, sr)
    rebuild_summary_sums(ws, a)
    wire_sig(ws, a, [f"I{sr}"])
    egi_rows[tab] = sr

# ---------------- 1.9 CTRM -> strategic ----------------
ws = wb["1.9 Commercial Fuels"]; a = anchors(ws)
ctrm_row = None
for b in blocks(ws):
    if ws.cell(b[0], 2).value == "Platform: CTRM":
        ctrm_row = b[2]
strategic_convert(ws, ctrm_row, 3.22, SEED_NOTE)
rebuild_summary_sums(ws, a)
wire_sig(ws, a, [f"I{ctrm_row}"])

# ---------------- 2. double-counting fix on TDD-internal tabs ----------------
DC_NOTE = ("Lights-on shown for reference only - it is the funding source of the Data Config "
           "TDD allocations (squad config), so drawing people from it here would double count. "
           "Real project envelopes (TDD Corporate CapEx 4.9 / Sig Items 20.2 / Initiatives 3.0) "
           "are on 2.1 COE.")
for tab in ["1.3 Enterprise Data", "1.4 TDD Group Functions", "1.11 TDD Cyber"]:
    ws = wb[tab]; a = anchors(ws)
    c = ws.cell(a["lightson"], 9)
    c.value = 0
    c.comment = Comment(DC_NOTE, "QA")

# ---------------- 3. 2.1 COE uplift (envelopes) ----------------
coe = wb["2.1 COE"]
tpl11 = wb["1.1 Ampol Retail"]
def S21(dst, src_cell, value=None):
    coe[dst]._style = copy(src_cell._style)
    if value is not None: coe[dst].value = value
egit = egi_rows["1.4 TDD Group Functions"]
# section band
for cc in "BCDE":
    coe[f"{cc}15"]._style = copy(coe["B6"]._style)
coe["B15"] = "COE project funding envelopes (TDD Corporate - 0.4 Fin, TDDEGM row)"
coe.merge_cells("B15:E15")
for j, h in enumerate(["Envelope", "Budget ($m)", "Drawn down ($m)", "Remaining ($m)"]):
    S21(f"{'BCDE'[j]}16", coe["B7"], h)
ENV = [
 ("OpEx Initiatives",  f"={FIN}!$P$26", "=0"),
 ("Significant Items", f"={FIN}!$Q$26", f"=0+'1.4 TDD Group Functions'!$I${egit}"),
 ("CapEx",             f"={FIN}!$R$26", "=0"),
]
r = 17
for name, bud, drawn in ENV:
    S21(f"B{r}", coe["B8"], name)
    S21(f"C{r}", coe["C8"], bud)
    S21(f"D{r}", coe["D8"], drawn if drawn != "=0" else 0)
    if drawn == "=0":
        coe[f"D{r}"].comment = Comment("Editable - record COE drawdowns against this envelope here.", "QA")
    else:
        coe[f"D{r}"].comment = Comment(
            "0 (editable base) + EGI TDD strategic programme cost from 1.4 - strategic programmes "
            "in TDD portfolios draw against the central Significant Items envelope.", "QA")
    S21(f"E{r}", coe["E8"], f"=C{r}-D{r}")
    r += 1
S21(f"B{r}", coe["B13"], "Total envelopes")
S21(f"C{r}", coe["C13"], f"=SUM(C17:C{r-1})")
S21(f"D{r}", coe["D13"], f"=SUM(D17:D{r-1})")
S21(f"E{r}", coe["E13"], f"=SUM(E17:E{r-1})")
env_tot = r
r += 2
S21(f"B{r}", coe["B8"], "COE project spend (planned)")
S21(f"C{r}", coe["D8"], 0)
coe[f"C{r}"].comment = Comment("Editable - planned COE project spend to be funded from the envelopes above.", "QA")
plan_row = r
r += 1
S21(f"B{r}", coe["B13"], "Left to fund (project spend - drawn)")
S21(f"C{r}", coe["C13"], f"=C{plan_row}-D{env_tot}")
coe_left_row = r
# note re TDD depreciation
r += 2
coe[f"B{r}"] = ("Note: TDD (Emily Mogic) also carries $4.9m depreciation; lights-on lines of TDD portfolios "
                "are not drawdown envelopes - they already fund the Data Config allocations.")
coe[f"B{r}"].font = Font(name="Calibri", size=9, italic=True, color="FF808080")

# ---------------- 4. 2.0 Group Summary rework ----------------
gs = wb["2.0 Group Summary"]
TABS = ["1.1 Ampol Retail","1.2 Customer","1.3 Enterprise Data","1.4 TDD Group Functions",
        "1.5 P&C","1.6 Finance","1.7 Infrastructure","1.8 Energy Solutions & B2B",
        "1.9 Commercial Fuels","1.10 Z Retail","1.11 TDD Cyber"]
# per-tab anchor cells
tabinfo = {}
for t in TABS:
    w = wb[t]; aa = anchors(w)
    tabinfo[t] = dict(applied=aa["applied"], left=aa["leftfund"], total=aa["totalcost"])
# F -> Funded, G -> Left to fund, H -> Total Cost
gs["F5"] = "Funded ($m)"
gs["G5"] = "Left to fund ($m)"
gs["H5"]._style = copy(gs["G5"]._style); gs["H5"] = "Total Cost ($m)"
gs.column_dimensions["H"].width = gs.column_dimensions["G"].width
for i, t in enumerate(TABS):
    r = 6 + i
    ti = tabinfo[t]
    gs[f"F{r}"] = f"='{t}'!$I${ti['applied']}"
    gs[f"G{r}"] = f"='{t}'!$I${ti['left']}"
    gs[f"H{r}"]._style = copy(gs[f"G{r}"]._style)
    gs[f"H{r}"] = f"='{t}'!$E${ti['total']}"
# COE row at 17, totals to 18
gs.insert_rows(17)
for cc in "BCDEFGH":
    gs[f"{cc}17"]._style = copy(gs[f"{cc}16"]._style)
gs["B17"] = "COE (2.1)"
gs["C17"] = "='2.1 COE'!$C$13"
gs["D17"] = "='2.1 COE'!$D$13"
gs["E17"] = "='2.1 COE'!$E$13"
gs["F17"] = f"='2.1 COE'!$D${env_tot}"
gs["G17"] = f"='2.1 COE'!$C${coe_left_row}"
gs["H17"] = "='2.1 COE'!$D$13"
gs["B18"] = "Total - all portfolios + COE"
for cc in "CDEFGH":
    gs[f"{cc}18"] = f"=SUM({cc}6:{cc}17)"
    gs[f"{cc}18"]._style = copy(gs[f"{cc}17"]._style)
# restyle total row from old total style (was row 17 pre-insert -> now 18 already has values/styles set above)
# ladder + reconciliation block (rewrite rows 21-27); unmerge first
for mr in list(gs.merged_cells.ranges):
    if mr.min_row >= 19:
        gs.unmerge_cells(str(mr))
# purge stale MergedCell placeholders left behind by insert_rows
from openpyxl.cell.cell import MergedCell, Cell
for (rr, cc), cell in list(gs._cells.items()):
    if rr >= 19 and isinstance(cell, MergedCell):
        del gs._cells[(rr, cc)]
for r in range(21, 28):
    for cc in "BCDEFGH":
        gs[f"{cc}{r}"].value = None
band_style_src = gs["B21"]
gs["B21"] = "Reconciliation - does everything ladder up?"
lbl = gs["B22"]; val = gs["C22"]
rows = [
 ("Total allocations above (portfolios + COE)", "=C18"),
 ("Data Config total allocation (E26)", "='0.0 Data Config'!$E$26"),
 ("Check (should be 0)", "=C22-C23"),
 ("Total TDD budget - people, AU + NZ (Data Config E27)", "='0.0 Data Config'!$E$27"),
 ("Headroom (total budget less allocations)", "=C25-C23"),
]
r = 22
for text, f in rows:
    gs[f"B{r}"] = text
    gs[f"C{r}"] = f
    r += 1
gs["C24"].comment = Comment("0 = the summary table covers every dollar of the Data Config allocation.", "QA")
gs["C26"].comment = Comment("The $53.8m total TDD people budget less the $43.5m allocated - unallocated capacity.", "QA")

# ---------------- 5. Data Config Spend / Variance wiring ----------------
dc = wb["0.0 Data Config"]
def tc_ref(t):  # total-cost C-cell ref for a tab
    return f"='{t}'!$C${tabinfo[t]['total']}"
WIRE = {
 6:  "='2.1 COE'!$D$8", 7: "='2.1 COE'!$D$9", 8: "='2.1 COE'!$D$10",
 9:  "='2.1 COE'!$D$11", 10: "='2.1 COE'!$D$12",
 11: tc_ref("1.1 Ampol Retail"), 12: tc_ref("1.10 Z Retail"),
 13: tc_ref("1.2 Customer"),
 16: tc_ref("1.8 Energy Solutions & B2B"), 17: tc_ref("1.7 Infrastructure"),
 18: tc_ref("1.5 P&C"), 19: tc_ref("1.6 Finance"),
 21: tc_ref("1.4 TDD Group Functions"), 22: tc_ref("1.3 Enterprise Data"),
 23: tc_ref("1.11 TDD Cyber"),
}
f_style = dc["F15"]._style; g_style = dc["G15"]._style
for r, f in WIRE.items():
    if dc[f"F{r}"].value in (None, ""):
        dc[f"F{r}"].value = f
        dc[f"F{r}"]._style = copy(f_style)
    if dc[f"G{r}"].value in (None, ""):
        dc[f"G{r}"].value = f"=SUM(E{r}-F{r})" if r != 13 else "=SUM(E13+E14-F13)"
        dc[f"G{r}"]._style = copy(g_style)
dc["G13"].comment = Comment("Customer tab spans Ampol Customer + Z Customer, so variance is vs E13+E14.", "QA")
dc["F14"].comment = Comment("Included in the Ampol Customer row above (1.2 Customer covers both).", "QA")

# ---------------- 6. replicate user's 'Total to fund' line ----------------
src_b = wb["1.1 Ampol Retail"]["B16"]; src_c = wb["1.1 Ampol Retail"]["C16"]
for t in TABS[1:]:
    w = wb[t]; aa = anchors(w)
    r = aa["leftfund"]
    if w.cell(r, 2).value in (None, ""):
        w.cell(r, 2)._style = copy(src_b._style)
        w.cell(r, 2).value = "Total to fund"
        w.cell(r, 3)._style = copy(src_c._style)
        w.cell(r, 3).value = f"=SUM(-H{aa['variance']},I{aa['leftfund']})"
        w.cell(r, 3).comment = Comment(
            "Over-budget TDD cost (negative variance) plus Left to fund - the total still to be found.", "QA")

# ---------------- 7. squad mapping addendum ----------------
sm = wb["squad mapping"]
r0 = 56
sm[f"C{r0}"] = "Strategic Programs (project platforms - blue on the op model)"
sm[f"C{r0}"].font = Font(name="Calibri", size=10, bold=True)
SP = [("Ampol Retail","AmPOS","AmPOS"),("Ampol Retail","EGI","EGI Retail"),
      ("Customer","EGI Customer","EGI Customer"),("TDD Group Functions","EGI TDD","EGI TDD"),
      ("P&C","EGI P&C","EGI P&C"),("Finance","EGI Finance","EGI Finance"),
      ("Commercial Fuels","CTRM","CTRM")]
for i,(p,pl,sq) in enumerate(SP):
    rr = r0 + 1 + i
    sm[f"D{rr}"] = p; sm[f"E{rr}"] = pl; sm[f"F{rr}"] = sq
    sm[f"G{rr}"] = "Strategic Programs"; sm[f"I{rr}"] = "Onshore"

wb.save(OUT)
print("saved", OUT)
print("EGI squad rows:", egi_rows, "CTRM row:", ctrm_row)
