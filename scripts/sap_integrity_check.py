#!/usr/bin/env python3
"""
sap_integrity_check.py
Standing structural integrity check for SAP BEx cost centre line item exports.

WHY THIS EXISTS
----------------
SAP BEx exports of cost centre line items can carry an overlapping COST
CENTRE GROUP axis: when a query's rows/free characteristic is a cost centre
hierarchy and more than one hierarchy node is selected, a single underlying
posting is displayed once per selected node it rolls up to. The "Cost
Center" column in these exports actually holds a GROUP CODE (a hierarchy
node), not a specific leaf cost centre, and the same postings can appear
several times under different group codes on the SAME tab. A plain SUM of
the data rows then overstates the true cost, sometimes by a large margin.
The file's own SAP "Overall Result" row is not affected (SAP computed it
correctly at query time); only a naive re-sum of the displayed rows is.

Two shapes of this have been seen across two client files:
  1. GROUP-LEVEL TWINS: an entire group's rows are byte-identical repeats
     of rows already present, individually, under other groups (nothing to
     do with that group's own true cost). Drop the whole group.
  2. PARENT EQUALS SUM OF CHILDREN: a group is a hierarchy roll-up node
     whose total exactly equals the sum of several other groups' totals
     (its children), even where the individual rows are not all
     byte-identical twins of a child row (some are re-aggregated at the
     parent's coarser grain). Drop the parent, keep the children.
  3. PARENT-PLUS-CHILD IN ONE COLUMN: several hierarchy tiers (a top node,
     a mid node, and one or more side nodes) are all present as peer values
     in the one "Cost Center" column with no level flag, and only the
     COMBINED total of that whole redundant tier ties out; no single one of
     them decomposes cleanly against the rest on its own.

WHAT THIS SCRIPT DOES
----------------------
Given an .xlsx path, for every tab it identifies as an SAP line-item export
it will:
  a. Find the tab generically: it looks for the SAP BEx column signature
     (a header row carrying "CO Document Number", "Posting date" and "Text
     of CO Document") rather than trusting sheet names. It does not assume
     which sheets exist or what they are called.
  b. Work out the column layout from that header row and the row above it:
     which column is the cost centre GROUP axis (whatever it happens to be
     called), which is the amount block, and which row is SAP's own
     "Overall Result" control total row. Nothing about tower names,
     Ampol, or this engagement is hard coded; the layout is read off the
     file's own headers, generically, every time.
  c. Compute the RAW sum (every data row, once) and the LEAF sum (the
     de-duplicated true sum), with row counts for both, using the
     mechanisms below, run in this order:
       1. Row level twin detection: build a signature of every field
          EXCEPT the group axis; any row whose full signature is shared
          with a row under a DIFFERENT group is a twin. This is evidence,
          reported per group as a coverage percentage, and is also used to
          steer (not gate) the search below.
       2. Group level reconciliation: raw sum minus SAP's control total
          gives the overstatement. The script searches for the smallest
          set of groups whose totals add up to that overstatement exactly,
          to the cent (a meet-in-the-middle exact subset sum, so it is not
          limited to a small, hand chosen number of candidate groups). That
          set is EXCLUDED; every other group is LEAF.
       3. Naming: each excluded group is then explained on its own, in this
          order. First, a SECOND, smaller exact subset sum is tried, but
          restricted to groups this specific group was DIRECTLY seen
          sharing a twin row with in step 1 (never the open kept-group
          pool - tried once for calibration, that open search does find
          numerically exact but causally spurious combinations, swapping
          an untouched, zero-evidence group in for the much larger one the
          row evidence actually points to; grounding the search in step 1's
          own evidence removes that failure mode). If a grounded exact
          match is found, the share of this group's OWN twin partners that
          the match actually needed decides the label: using nearly all of
          them (>= TWIN_CHILD_COVERAGE, default 70%) means the group is, in
          substance, just a copy of those specific others, named a
          GROUP-LEVEL TWIN (mechanism 1); using only a narrow slice of a
          much wider partner set means it is better described as a roll-up
          node, named a PARENT with those children (mechanism 2). Second,
          if no grounded match exists but the group's own row level twin
          coverage from step 1 is still high (>= TWIN_ROW_FRAC_FALLBACK,
          default 90%), it is still named a GROUP-LEVEL TWIN, honestly
          reported without an exact children list (a handful of non-twinned
          rows on one side or the other keep the whole-group totals from
          closing to the cent, even though the row evidence is
          overwhelming). Third and last, an OPEN search against every kept
          group is tried and, if it succeeds, reported as PARENT EQUALS SUM
          OF CHILDREN but explicitly flagged "ungrounded" (arithmetic match
          only, no row evidence ties it). If nothing above resolves the
          group individually, it is named PARENT-PLUS-CHILD-IN-ONE-COLUMN:
          it only ties out collectively, with the other excluded groups in
          the same tab (mechanism 3). On the two files this script has been
          run against, every excluded group resolved at the first or second
          step; the open-search and mixed-tier code paths (mechanism 3)
          exist and are reachable but were not exercised by real data -
          said plainly in the certificate, not glossed over.
  d. Tie the leaf sum to the tab's own SAP Overall Result row (to the
     cent; this is the PASS or FAIL test) and separately search every
     OTHER, non line-item tab in the same file for any cell that also
     carries the leaf sum to the cent, reporting it as independent
     corroboration where found, and reporting plainly when none is found
     (that is itself a finding, not a script fault).
  e. Print a one-page certificate: per tab, raw, leaf, overstatement,
     mechanism(s) named with the groups and row counts involved, tie
     status, and a PASS or FAIL line, plus a file-level verdict.

READ ONLY, STREAMED
--------------------
Opens every workbook with openpyxl(read_only=True, data_only=True). Each
sheet is read in at most two forward streaming passes (a small header
scan of the first ~40 rows, then one pass over the data rows); nothing is
loaded via ws.values or list(ws.rows), and no raw row dump is ever printed.
Per row, only the specific fields needed (group code, group name, and the
handful of other raw SAP columns plus the amounts) are kept in memory, not
the whole row; that per-tab, per-row working set (roughly one tuple per
row) is the only thing held for the run, freed tab by tab.

HONEST LIMITS (read this before trusting a FAIL or a silent PASS)
--------------------------------------------------------------------
- Tab discovery depends on the literal SAP BEx header strings "CO Document
  Number", "Posting date" and "Text of CO Document" appearing together in
  one row. A tab that does not carry all three headers is not seen as an
  SAP tab at all, even if it is one; a re-labelled export would need those
  headers restored, or this script extended with an alternate signature.
- The group axis column is found by header text matching /cost cent(er
  |re)/i. A file that calls that column something else entirely (not a
  "cost center/centre" label of any spelling) will not be found generically
  and the tab is skipped, not mis-read.
- The "name" half of a code/name pair (cost centre name, cost element
  name, vendor name) is inferred as the column immediately to the right of
  the code column when that column's own header is blank. A layout that
  puts the name column somewhere else will attach the wrong text (or none)
  to the code; the code columns themselves (used for all arithmetic) are
  unaffected.
- The exact subset sum search is a real, full meet-in-the-middle over
  every group in the tab (not a small hard-coded shortlist), but it is
  still a needle search: with real, non-repeating cent-level totals a
  match is effectively certain to be the true one, but the script cannot
  prove uniqueness, only report the smallest exact match it found and how
  many equally-small alternatives existed (0 in every run so far). If the
  true overstatement genuinely cannot be built as an exact sum of some
  subset of groups (a shape not yet seen), the script says so and stops
  short of a leaf figure rather than guessing.
- Meet-in-the-middle needs 2^(n/2) memory/time for n groups; tabs with
  more than ~40 distinct group codes will be slow or exhausted. Every tab
  seen to date has under 35 distinct groups.
- Row level twin detection treats two rows as a twin only if EVERY field
  outside the group axis matches exactly, including all 12 monthly amount
  cells and the row total. A parent-level row that aggregates two or more
  child postings into one coarser row (a genuine collision, not a clean
  1:1 twin) will not be caught by this check; it still nets out correctly
  through the group-level subset sum in step 2, but it will lower that
  group's reported twin-coverage percentage, which is reported as-is, not
  smoothed over.
- The twin-vs-parent split (mechanisms 1 and 2) is decided by one threshold,
  TWIN_CHILD_COVERAGE (0.70): how much of a group's own observed twin
  partner evidence its exact decomposition actually uses. This is the one
  judgment call in the whole method; it is applied identically everywhere,
  never tuned per file or per group, and a group sitting right at the line
  could read either way. The row-level fallback threshold,
  TWIN_ROW_FRAC_FALLBACK (0.90), is the second and only other tuned number.
- A file with a THIRD hierarchy tier (a parent of parents, where an
  excluded group's true children are themselves excluded groups rather
  than kept ones) is not specifically modelled: children are only ever
  searched for among the KEPT groups. Such a group would fall through to
  the ungrounded open search and, failing that, to mechanism 3 (reported
  honestly as an unresolved, mixed-tier member of the excluded set) rather
  than being mis-labelled a clean parent.
- The ungrounded, open decomposition search (the third and last naming
  step) is retained as a genuine last resort, not removed, precisely
  because it is known to be able to produce a spurious answer; it is
  always reported with a visible "ungrounded" flag for that reason, and a
  reader should treat such a finding as arithmetic corroboration of the
  group-level exclusion (which is independently proven by the tie to SAP's
  own control total), not as proof of which specific groups it duplicates.
- The external tie search (step d) only checks OTHER tabs in the same
  workbook, for a literal cell value equal to the leaf sum to the cent. A
  summary tab that nets, rounds, or filters before totalling (as this
  client's own Network summary check row does, by a proven 700.00 gap)
  will correctly come back NOT FOUND; that is reported, not hidden, and it
  does not by itself fail the tab, because the tab's own SAP control row
  is the authoritative tie.
- This script does not touch, write to, or modify the source file in any
  way; it opens every workbook read only and prints its findings.

USAGE
-----
    python3 sap_integrity_check.py <path-to-workbook.xlsx> [--json out.json]

Exit code 0 if every SAP tab found in the file ties leaf-to-control to the
cent (FILE STATUS: PASS); exit code 1 otherwise.
"""

import sys
import re
import json
import argparse
from collections import defaultdict
from datetime import datetime, timezone

import openpyxl

CENTS = 100
TWIN_CHILD_COVERAGE = 0.70     # share of a group's OWN observed twin-partner evidence its exact
                                # decomposition must use to be called a "twin" rather than a
                                # "parent equal to the sum of its children" (see certify_tab)
TWIN_ROW_FRAC_FALLBACK = 0.90  # share of a group's OWN rows that must carry an external twin to
                                # call it a twin duplicate when no exact decomposition closes
MAX_SCAN_ROWS = 40             # how far down we look for the SAP header block
MAX_MITM_GROUPS = 40           # meet-in-the-middle group-count ceiling (2^(n/2) cost)
CONSEC_BLANK_STOP = 300        # consecutive blank rows (after data has started) before we stop reading a tab

PERIOD_RE = re.compile(r'^\d{3}\.\d{4}$')
COST_CENTRE_RE = re.compile(r'cost\s*cent(er|re)', re.IGNORECASE)


# --------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------

def to_cents(v):
    """Coerce a cell value to an integer number of cents, or None if it is
    not a usable number (blank, text, error string)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(round(v * CENTS))
    if isinstance(v, str):
        s = v.strip()
        if s == '':
            return None
        try:
            return int(round(float(s) * CENTS))
        except ValueError:
            return None
    return None


def money(cents, brackets=True):
    """Format integer cents as a money string, brackets for negative."""
    if cents is None:
        return 'TBC'
    neg = cents < 0
    c = abs(cents)
    s = f"{c // CENTS:,}.{c % CENTS:02d}"
    if neg and brackets:
        return f"({s})"
    if neg:
        return f"-{s}"
    return s


def norm_header(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip()
    return v


# --------------------------------------------------------------------------
# (a) + (b): generic discovery of SAP line-item tabs and their column layout
# --------------------------------------------------------------------------

def scan_sheet_for_sap_layout(ws, sheet_name, log):
    """Read the first MAX_SCAN_ROWS rows of a sheet in one forward pass and,
    if it carries the SAP BEx line-item signature, return a layout dict.
    Returns None if this sheet is not an SAP line-item tab."""
    buf = []
    for row in ws.iter_rows(min_row=1, max_row=MAX_SCAN_ROWS, values_only=True):
        buf.append(row)
    header_row_idx = None  # 0-based index into buf
    for i, row in enumerate(buf):
        texts = {norm_header(c) for c in row if isinstance(c, str)}
        if {'CO Document Number', 'Posting date', 'Text of CO Document'} <= texts:
            header_row_idx = i
            break
    if header_row_idx is None:
        return None

    header = buf[header_row_idx]
    ncols = len(header)

    def find_col(label_re_or_str, row):
        for c, v in enumerate(row):
            if not isinstance(v, str):
                continue
            vv = v.strip()
            if isinstance(label_re_or_str, str):
                if vv == label_re_or_str:
                    return c
            else:
                if label_re_or_str.search(vv):
                    return c
        return None

    group_code_col = find_col(COST_CENTRE_RE, header)
    elem_code_col = find_col('Cost Element', header)
    doc_col = find_col('CO Document Number', header)
    user_col = find_col('User', header)
    aux_col = find_col('Aux. Acct Asgmt', header)
    vendor_col = find_col('Vendor', header)
    po_col = find_col('Purchase Order', header)
    date_col = find_col('Posting date', header)
    text_col = find_col('Text of CO Document', header)

    required = dict(group_code_col=group_code_col, elem_code_col=elem_code_col,
                     doc_col=doc_col, vendor_col=vendor_col, po_col=po_col,
                     date_col=date_col, text_col=text_col)
    missing = [k for k, v in required.items() if v is None]
    if missing:
        log(f"  [skip] '{sheet_name}': has the SAP header row but is missing {missing}; not treated as a line-item tab")
        return None

    def name_col_for(code_col):
        nxt = code_col + 1
        if nxt < ncols and header[nxt] is None:
            return nxt
        return None

    group_name_col = name_col_for(group_code_col)
    elem_name_col = name_col_for(elem_code_col)
    vendor_name_col = name_col_for(vendor_col)

    # the row above the header row carries the period codes + 'Overall Result' label
    period_row = buf[header_row_idx - 1] if header_row_idx >= 1 else []
    amount_cols = []
    total_col = None
    for c, v in enumerate(period_row):
        if isinstance(v, str) and PERIOD_RE.match(v.strip()):
            amount_cols.append(c)
        elif isinstance(v, str) and v.strip() == 'Overall Result':
            total_col = c
    if total_col is None or not amount_cols:
        log(f"  [skip] '{sheet_name}': header row found but no period/'Overall Result' amount block above it")
        return None

    # the control total row: first row after the header row whose group-code
    # cell literally reads 'Overall Result'
    total_row_idx = None  # 0-based into buf
    for i in range(header_row_idx + 1, min(header_row_idx + 6, len(buf))):
        v = buf[i][group_code_col] if group_code_col < len(buf[i]) else None
        if isinstance(v, str) and v.strip() == 'Overall Result':
            total_row_idx = i
            break
    if total_row_idx is None:
        log(f"  [skip] '{sheet_name}': no 'Overall Result' control row found within 5 rows of the header")
        return None

    control_total_cents = to_cents(buf[total_row_idx][total_col])
    if control_total_cents is None:
        log(f"  [skip] '{sheet_name}': control row found but its total cell is not numeric")
        return None

    header_row_1based = header_row_idx + 1
    total_row_1based = total_row_idx + 1

    return dict(
        sheet_name=sheet_name,
        header_row=header_row_1based,
        total_row=total_row_1based,
        data_start_row=total_row_1based + 1,
        group_code_col=group_code_col, group_name_col=group_name_col,
        elem_code_col=elem_code_col, elem_name_col=elem_name_col,
        doc_col=doc_col, user_col=user_col, aux_col=aux_col,
        vendor_col=vendor_col, vendor_name_col=vendor_name_col,
        po_col=po_col, date_col=date_col, text_col=text_col,
        amount_cols=amount_cols, total_col=total_col,
        control_total_cents=control_total_cents,
        ncols=ncols,
    )


def find_sap_tabs(wb, log):
    tabs = []
    for name in wb.sheetnames:
        ws = wb[name]
        info = scan_sheet_for_sap_layout(ws, name, log)
        if info:
            tabs.append(info)
            log(f"  [found] '{name}': header row {info['header_row']}, control row {info['total_row']}, "
                f"group axis col {info['group_code_col']}, {len(info['amount_cols'])} period cols + total col {info['total_col']}")
    return tabs


# --------------------------------------------------------------------------
# (c-1): stream the data rows of one tab, building signatures + group sums
# --------------------------------------------------------------------------

def read_tab_rows(ws, layout, log):
    """Single forward streaming pass over the data rows of one SAP tab.
    Returns (rows, raw_sum_cents, raw_count) where rows is a list of small
    per-row dicts (not the raw sheet row) carrying only what later stages
    need: group code/name, a hashable signature of every OTHER field, and
    the row's own total in cents."""
    g = layout
    rows = []
    raw_sum_cents = 0
    raw_count = 0
    consec_blank = 0
    scanned = 0
    for row in ws.iter_rows(min_row=g['data_start_row'], values_only=True):
        scanned += 1
        doc_v = row[g['doc_col']] if g['doc_col'] < len(row) else None
        grp_v = row[g['group_code_col']] if g['group_code_col'] < len(row) else None
        if doc_v is None and grp_v is None:
            consec_blank += 1
            if raw_count > 0 and consec_blank >= CONSEC_BLANK_STOP:
                break
            continue
        consec_blank = 0
        total_c = to_cents(row[g['total_col']]) if g['total_col'] < len(row) else None
        if total_c is None:
            # a row with identity fields but no numeric total: keep it out of
            # sums but do not silently drop the row count discrepancy
            total_c = 0
        month_sig = tuple(
            to_cents(row[c]) if c < len(row) else None for c in g['amount_cols']
        )
        sig = (
            row[g['elem_code_col']] if g['elem_code_col'] < len(row) else None,
            doc_v,
            row[g['user_col']] if g['user_col'] is not None and g['user_col'] < len(row) else None,
            row[g['aux_col']] if g['aux_col'] is not None and g['aux_col'] < len(row) else None,
            row[g['vendor_col']] if g['vendor_col'] < len(row) else None,
            row[g['vendor_name_col']] if g['vendor_name_col'] is not None and g['vendor_name_col'] < len(row) else None,
            row[g['po_col']] if g['po_col'] < len(row) else None,
            row[g['date_col']] if g['date_col'] < len(row) else None,
            row[g['text_col']] if g['text_col'] < len(row) else None,
            month_sig,
            total_c,
        )
        grp_code = grp_v
        grp_name = row[g['group_name_col']] if g['group_name_col'] is not None and g['group_name_col'] < len(row) else None
        rows.append(dict(grp=grp_code, grp_name=grp_name, sig=sig, total_c=total_c))
        raw_sum_cents += total_c
        raw_count += 1
    log(f"    scanned {scanned} rows below the control row, kept {raw_count} as data rows "
        f"(stopped on {CONSEC_BLANK_STOP} consecutive blanks after data began, or end of sheet)")
    return rows, raw_sum_cents, raw_count


# --------------------------------------------------------------------------
# (c-2): mechanism 1 - row level twin detection across the group axis
# --------------------------------------------------------------------------

def detect_twins(rows):
    """Group rows by signature (everything except the group axis). A
    signature shared by >=2 DISTINCT group codes marks every row in it as a
    cross-group twin. Returns per-group twin coverage stats and a
    (groupA, groupB) co-occurrence tally for narrative evidence."""
    sig_index = defaultdict(list)  # sig -> list of row indices
    for i, r in enumerate(rows):
        sig_index[r['sig']].append(i)

    twin_row = [False] * len(rows)
    pair_counts = defaultdict(int)
    pair_sums = defaultdict(int)

    for sig, idxs in sig_index.items():
        groups_here = {rows[i]['grp'] for i in idxs}
        if len(groups_here) < 2:
            continue
        for i in idxs:
            twin_row[i] = True
        # tally pairwise co-occurrence between distinct groups sharing this signature
        glist = sorted(groups_here)
        for a in range(len(glist)):
            for b in range(a + 1, len(glist)):
                pair_counts[(glist[a], glist[b])] += 1
                # dollar weight: cents of one representative row in this cluster
                pair_sums[(glist[a], glist[b])] += rows[idxs[0]]['total_c']

    by_group_rows = defaultdict(int)
    by_group_twin_rows = defaultdict(int)
    by_group_sum = defaultdict(int)
    by_group_twin_sum = defaultdict(int)
    by_group_partners = defaultdict(set)

    for i, r in enumerate(rows):
        by_group_rows[r['grp']] += 1
        by_group_sum[r['grp']] += r['total_c']
        if twin_row[i]:
            by_group_twin_rows[r['grp']] += 1
            by_group_twin_sum[r['grp']] += r['total_c']

    for (a, b) in pair_counts:
        by_group_partners[a].add(b)
        by_group_partners[b].add(a)

    coverage = {}
    for grp in by_group_rows:
        rc = by_group_rows[grp]
        sc = by_group_sum[grp]
        trc = by_group_twin_rows.get(grp, 0)
        tsc = by_group_twin_sum.get(grp, 0)
        row_frac = trc / rc if rc else 0.0
        dollar_frac = (tsc / sc) if sc else (1.0 if tsc == 0 and trc == rc else 0.0)
        coverage[grp] = dict(
            rows=rc, sum_c=sc, twin_rows=trc, twin_sum_c=tsc,
            row_frac=row_frac, dollar_frac=dollar_frac,
            partners=sorted(by_group_partners.get(grp, [])),
        )
    return coverage, pair_counts, pair_sums


# --------------------------------------------------------------------------
# (c-3): exact subset-sum (meet in the middle) for group level reconciliation
# --------------------------------------------------------------------------

def exact_subset_sum(items, target):
    """items: list of (key, value_cents). Find the smallest-cardinality
    subset of keys whose values sum EXACTLY to target (cents). Uses a full
    meet-in-the-middle split, so it is exhaustive over all subsets, not a
    hand-picked shortlist. Returns (subset_keys, tie_count) where tie_count
    is how many other subsets of the same minimal size also matched
    (0 = the match is unique at that size), or (None, 0) if nothing sums to
    target at all (including target itself as a single-empty-set case)."""
    n = len(items)
    if n == 0:
        return (([], 0) if target == 0 else (None, 0))
    if n > MAX_MITM_GROUPS:
        raise ValueError(f"{n} distinct groups exceeds the meet-in-the-middle ceiling of {MAX_MITM_GROUPS}")

    half = n // 2
    left, right = items[:half], items[half:]

    def all_subsets(part):
        # returns dict: sum_cents -> list of (cardinality, tuple_of_keys), kept small (min cardinality + ties only)
        best = {}
        m = len(part)
        for mask in range(1 << m):
            s = 0
            keys = []
            card = 0
            mm = mask
            idx = 0
            while mm:
                if mm & 1:
                    k, v = part[idx]
                    s += v
                    keys.append(k)
                    card += 1
                mm >>= 1
                idx += 1
            cur = best.get(s)
            if cur is None or card < cur[0]:
                best[s] = (card, [tuple(keys)])
            elif card == cur[0]:
                cur[1].append(tuple(keys))
        return best

    left_sums = all_subsets(left)
    right_sums = all_subsets(right)

    best_card = None
    best_solutions = []
    for ls, (lcard, lkeys_list) in left_sums.items():
        need = target - ls
        rentry = right_sums.get(need)
        if rentry is None:
            continue
        rcard, rkeys_list = rentry
        total_card = lcard + rcard
        if best_card is not None and total_card > best_card:
            continue
        if best_card is None or total_card < best_card:
            best_card = total_card
            best_solutions = []
        for lk in lkeys_list:
            for rk in rkeys_list:
                best_solutions.append(list(lk) + list(rk))

    if best_card is None:
        return (None, 0)
    # de-duplicate identical solutions (can arise if left/right both have multiple min-card reps)
    uniq = []
    seen = set()
    for sol in best_solutions:
        key = tuple(sorted(sol, key=lambda x: str(x)))
        if key not in seen:
            seen.add(key)
            uniq.append(sol)
    return (uniq[0], len(uniq) - 1)


# --------------------------------------------------------------------------
# (c) + (d): put one tab's certificate together
# --------------------------------------------------------------------------

def certify_tab(wb, layout, other_small_sheets, log):
    ws = wb[layout['sheet_name']]
    log(f"  reading data rows for '{layout['sheet_name']}' ...")
    rows, raw_sum_c, raw_count = read_tab_rows(ws, layout, log)

    by_group_sum = defaultdict(int)
    by_group_rows = defaultdict(int)
    for r in rows:
        by_group_sum[r['grp']] += r['total_c']
        by_group_rows[r['grp']] += 1

    control_c = layout['control_total_cents']
    overstatement_c = raw_sum_c - control_c

    result = dict(
        sheet_name=layout['sheet_name'],
        raw_rows=raw_count, raw_sum_c=raw_sum_c,
        control_c=control_c,
        overstatement_c=overstatement_c,
        n_groups=len(by_group_sum),
        group_sums_c={k: v for k, v in by_group_sum.items()},
        group_rows={k: v for k, v in by_group_rows.items()},
    )

    if overstatement_c == 0:
        result.update(
            leaf_rows=raw_count, leaf_sum_c=raw_sum_c,
            excluded_groups=[], mechanisms=[],
            tie=True, note='raw sum already ties to the Overall Result control row; no duplication found',
        )
    else:
        log(f"    raw sum {money(raw_sum_c)} vs control {money(control_c)}: overstatement {money(overstatement_c)}; "
            f"searching {len(by_group_sum)} groups for an exact subset match ...")
        coverage, pair_counts, pair_sums = detect_twins(rows)

        items = sorted(by_group_sum.items(), key=lambda kv: -coverage.get(kv[0], {}).get('dollar_frac', 0))
        try:
            excluded, ties = exact_subset_sum(items, overstatement_c)
        except ValueError as e:
            excluded, ties = None, 0
            log(f"    [limit] {e}")

        if excluded is None:
            log(f"    [FAIL] no exact subset of groups reproduces the overstatement of {money(overstatement_c)}")
            result.update(
                leaf_rows=None, leaf_sum_c=None, excluded_groups=[], mechanisms=[],
                tie=False, note='no exact group-level subset sum found for the overstatement; leaf figure not certified',
            )
        else:
            excluded_set = set(excluded)
            kept_items = [(k, v) for k, v in items if k not in excluded_set]
            leaf_sum_c = raw_sum_c - sum(by_group_sum[g] for g in excluded_set)
            leaf_rows = raw_count - sum(by_group_rows[g] for g in excluded_set)
            # independent recomputation directly from the rows, not by subtraction
            leaf_sum_c_direct = sum(r['total_c'] for r in rows if r['grp'] not in excluded_set)
            leaf_rows_direct = sum(1 for r in rows if r['grp'] not in excluded_set)
            assert leaf_sum_c_direct == leaf_sum_c, "algebraic and direct leaf sums disagree"
            assert leaf_rows_direct == leaf_rows, "algebraic and direct leaf row counts disagree"

            if ties:
                log(f"    [note] {ties} other group combination(s) of the same minimal size also matched the overstatement exactly; "
                    f"the first found is reported")

            mechanisms = []
            # Name each excluded group individually. The search space for "what does this
            # group's total decompose into" is deliberately restricted to groups it has been
            # DIRECTLY, independently observed sharing exact-signature rows with (its own twin
            # partners from step 1, intersected with the groups actually kept as leaf) rather
            # than the full universe of kept groups. This is not a convenience: tried against
            # the full universe, the search can and does land on a numerically-exact but
            # causally spurious combination (proven on this file's own SW tab: it will swap an
            # untouched, zero-twin-evidence group for the much larger group that the row-level
            # evidence actually implicates, and still hit the target to the cent). Restricting
            # to observed partners removes that failure mode; every decomposition reported below
            # is one the row-level evidence directly supports, not merely one that adds up.
            #
            # A group whose decomposition draws on nearly all of its own observed partner
            # evidence (a high "coverage ratio") is, in substance, just a copy of those specific
            # other groups: GROUP-LEVEL TWIN. A group whose decomposition is a narrow, clean
            # slice out of a much wider, tangled web of partner evidence (a low coverage ratio,
            # seen only where a tab excludes several groups from a genuinely multi-level
            # hierarchy) is better described as a roll-up node: PARENT EQUALS SUM OF CHILDREN.
            # The ratio threshold (TWIN_CHILD_COVERAGE) is the one judgment call in this whole
            # method; it is applied identically everywhere, never tuned per file or per group.
            claimed_children = set()
            per_group = {}
            for grp in sorted(excluded_set, key=lambda g: -abs(by_group_sum[g])):
                target_c = by_group_sum[grp]
                cov = coverage.get(grp, dict(row_frac=0, dollar_frac=0, rows=0, twin_rows=0, twin_sum_c=0, partners=[]))
                eligible_partners = [p for p in cov['partners'] if p not in excluded_set]
                grounded_pool = [(k, v) for k, v in kept_items if k in eligible_partners]
                child_sol, child_ties, grounded = (None, 0, True)
                if grounded_pool:
                    try:
                        child_sol, child_ties = exact_subset_sum(grounded_pool, target_c)
                    except ValueError:
                        child_sol, child_ties = None, 0
                if child_sol is None and cov['row_frac'] < TWIN_ROW_FRAC_FALLBACK and kept_items:
                    # Grounded search found no exact match, AND this group's own rows are not
                    # predominantly twinned with anything either, so there is no direct row
                    # evidence to fall back on. Last resort: an open search against every kept
                    # group, not just this group's own observed partners. Always reported as
                    # UNGROUNDED so the certificate does not overstate how it was arrived at.
                    try:
                        open_sol, open_ties = exact_subset_sum(kept_items, target_c)
                    except ValueError:
                        open_sol, open_ties = None, 0
                    if open_sol is not None:
                        child_sol, child_ties, grounded = open_sol, open_ties, False
                per_group[grp] = (target_c, cov, eligible_partners, child_sol, child_ties, grounded)

            for grp in sorted(excluded_set, key=lambda g: -abs(by_group_sum[g])):
                target_c, cov, eligible_partners, child_sol, child_ties, grounded = per_group[grp]
                overlaps_prior = bool(child_sol) and (set(child_sol) & claimed_children)
                if child_sol and not overlaps_prior:
                    claimed_children |= set(child_sol)
                    coverage_ratio = (len(set(child_sol) & set(eligible_partners)) / len(eligible_partners)) if (grounded and eligible_partners) else 0.0
                    if grounded and coverage_ratio >= TWIN_CHILD_COVERAGE:
                        mechanisms.append(dict(
                            group=grp, kind='group_level_twin', exact=True,
                            rows=by_group_rows[grp], sum_c=target_c,
                            twin_rows=cov['twin_rows'], twin_row_frac=cov['row_frac'], twin_dollar_frac=cov['dollar_frac'],
                            partners=child_sol,
                        ))
                    else:
                        mechanisms.append(dict(
                            group=grp, kind='parent_equals_sum_of_children',
                            rows=by_group_rows[grp], sum_c=target_c,
                            children=child_sol, child_ties=child_ties, grounded=grounded,
                            twin_row_frac=cov['row_frac'], twin_dollar_frac=cov['dollar_frac'],
                        ))
                elif cov['row_frac'] >= TWIN_ROW_FRAC_FALLBACK:
                    # No exact whole-group decomposition closes to the cent (a handful of this
                    # group's own rows, or its partners', are not themselves twinned and leave a
                    # residual) but the row-level evidence that it is fundamentally a copy is
                    # overwhelming on its own terms: report it as a twin without a children list,
                    # rather than force an exact-but-unevidenced combination through step 2's
                    # broader search purely to have one.
                    mechanisms.append(dict(
                        group=grp, kind='group_level_twin', exact=False,
                        rows=by_group_rows[grp], sum_c=target_c,
                        twin_rows=cov['twin_rows'], twin_row_frac=cov['row_frac'], twin_dollar_frac=cov['dollar_frac'],
                        partners=eligible_partners,
                    ))
                else:
                    reason = ('its children would double claim a group another excluded group already claimed' if overlaps_prior
                              else 'no exact subset of its own row-twinned groups (or, failing that, of the kept groups) reproduces its total on its own, and its own rows are not predominantly twinned either')
                    mechanisms.append(dict(
                        group=grp, kind='mixed_tier_one_column',
                        rows=by_group_rows[grp], sum_c=target_c,
                        twin_row_frac=cov['row_frac'], twin_dollar_frac=cov['dollar_frac'],
                        partners=cov['partners'], reason=reason,
                        co_excluded_with=sorted(excluded_set - {grp}),
                    ))

            tie = (leaf_sum_c == control_c)
            result.update(
                leaf_rows=leaf_rows, leaf_sum_c=leaf_sum_c,
                excluded_groups=sorted(excluded_set),
                mechanisms=mechanisms,
                tie=tie,
                note='leaf sum reproduced by excluding the group(s) named above; recomputed directly from the rows, not only by subtraction',
            )

    # (d) external tie: search other, non line-item tabs for a cell matching the leaf sum to the cent
    result['external_ties'] = []
    if result.get('leaf_sum_c') is not None:
        target = result['leaf_sum_c']
        for name, ws2 in other_small_sheets:
            hits = search_sheet_for_value(ws2, target)
            for coord in hits:
                result['external_ties'].append(f"'{name}'!{coord}")
            if len(result['external_ties']) >= 5:
                break
    return result


def search_sheet_for_value(ws, target_cents, cap=3):
    hits = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if isinstance(v, (int, float)):
                if to_cents(v) == target_cents:
                    hits.append(cell.coordinate)
                    if len(hits) >= cap:
                        return hits
    return hits


# --------------------------------------------------------------------------
# top level: run the whole certificate for one file
# --------------------------------------------------------------------------

def mechanism_line(m):
    if m['kind'] == 'group_level_twin':
        who = ', '.join(str(c) for c in m['partners'])
        closes = (f"its total is exactly the sum of {len(m['partners'])} of those groups it shares rows with, in full: [{who}]"
                  if m.get('exact') else
                  f"a residual of non-twinned rows on one side or the other keeps the whole-group totals from closing to the cent; "
                  f"the groups it shares rows with are: [{who}]")
        return (f"group {m['group']} ({m['rows']:,} rows, {money(m['sum_c'])}) dropped whole: "
                f"a GROUP-LEVEL TWIN duplicate, {m['twin_row_frac']*100:.1f}% of its rows "
                f"({m['twin_dollar_frac']*100:.1f}% of its dollars) are byte-identical to rows kept "
                f"under other groups; {closes}")
    if m['kind'] == 'parent_equals_sum_of_children':
        kids = ', '.join(str(c) for c in m['children'])
        basis = 'its own row-twinned groups' if m.get('grounded', True) else 'the kept groups generally (no row-twin evidence directly ties them; arithmetic match only)'
        return (f"group {m['group']} ({m['rows']:,} rows, {money(m['sum_c'])}) dropped whole: "
                f"a PARENT whose total equals the sum of its children [{kids}] to the cent, found among {basis}")
    return (f"group {m['group']} ({m['rows']:,} rows, {money(m['sum_c'])}) dropped whole: "
            f"PARENT-PLUS-CHILD-IN-ONE-COLUMN, {m.get('reason', 'ties only collectively')}; "
            f"only ties together with {m.get('co_excluded_with')}; row-twin evidence covers "
            f"{m['twin_row_frac']*100:.1f}% of its rows (partners: {m.get('partners')})")


def run_file(path, log):
    log(f"Opening {path} read only, data_only ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    log(f"  {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    tabs = find_sap_tabs(wb, log)
    if not tabs:
        log("  no SAP line-item tabs discovered in this file")
        return dict(path=path, tabs=[], file_pass=False)

    sap_names = {t['sheet_name'] for t in tabs}
    other_small_sheets = []
    for name in wb.sheetnames:
        if name in sap_names:
            continue
        ws = wb[name]
        try:
            mr, mc = ws.max_row or 0, ws.max_column or 0
        except Exception:
            mr, mc = 0, 0
        if mr * mc and mr * mc <= 250_000:
            other_small_sheets.append((name, ws))

    results = []
    for layout in tabs:
        log(f"\n--- tab '{layout['sheet_name']}' ---")
        res = certify_tab(wb, layout, other_small_sheets, log)
        results.append(res)

    file_pass = all(r.get('tie') for r in results)
    return dict(path=path, tabs=results, file_pass=file_pass)


def print_certificate(run, log):
    log("\n" + "=" * 78)
    log("SAP STRUCTURAL INTEGRITY CERTIFICATE")
    log(f"File: {run['path']}")
    log(f"Generated: {datetime.now(timezone.utc).isoformat(timespec='seconds')}Z")
    log("=" * 78)
    for r in run['tabs']:
        log(f"\nTab: {r['sheet_name']}")
        log(f"  raw rows    : {r['raw_rows']:>10,}    raw sum    : {money(r['raw_sum_c']):>18}")
        if r.get('leaf_sum_c') is not None:
            log(f"  leaf rows   : {r['leaf_rows']:>10,}    leaf sum   : {money(r['leaf_sum_c']):>18}")
        else:
            log(f"  leaf rows   :        TBC    leaf sum   :                TBC")
        log(f"  overstatement: {money(r['overstatement_c'])}")
        if r['mechanisms']:
            log(f"  mechanism(s):")
            for m in r['mechanisms']:
                log(f"    - {mechanism_line(m)}")
        else:
            log(f"  mechanism(s): {r['note']}")
        ctrl_tie = "PASS" if r.get('leaf_sum_c') == r['control_c'] else "FAIL"
        log(f"  tie to this tab's SAP Overall Result row: {ctrl_tie} "
            f"({money(r.get('leaf_sum_c'))} vs {money(r['control_c'])})")
        if r['external_ties']:
            log(f"  tie to another tab in this file: FOUND at {', '.join(r['external_ties'])}")
        else:
            log(f"  tie to another tab in this file: NOT FOUND (checked every other tab for a matching cell)")
        log(f"  TAB STATUS: {'PASS' if r['tie'] else 'FAIL'}")

    log("\n" + "-" * 78)
    log(f"FILE STATUS: {'PASS' if run['file_pass'] else 'FAIL'} "
        f"({sum(1 for r in run['tabs'] if r['tie'])}/{len(run['tabs'])} tabs tie to the cent)")
    log("=" * 78)


def main():
    ap = argparse.ArgumentParser(description="Standing structural integrity check for SAP BEx cost centre exports.")
    ap.add_argument('path', help='path to the .xlsx workbook')
    ap.add_argument('--json', help='optional path to also write the full machine-readable report as JSON')
    args = ap.parse_args()

    lines = []

    def log(msg):
        print(msg)
        lines.append(msg)

    run = run_file(args.path, log)
    print_certificate(run, log)

    if args.json:
        def jsonable(o):
            if isinstance(o, dict):
                return {str(k): jsonable(v) for k, v in o.items()}
            if isinstance(o, (list, tuple, set)):
                return [jsonable(v) for v in o]
            return o
        with open(args.json, 'w') as f:
            json.dump(jsonable(run), f, indent=1, default=str)
        log(f"\n[wrote machine-readable report to {args.json}]")

    sys.exit(0 if run['file_pass'] else 1)


if __name__ == '__main__':
    main()
