#!/usr/bin/env python3
"""v8 - full restructure per user direction:
 A. Offshore toggle FIX: real offshore rate cell on 0.1 Squads (K5=40%),
    every role model-cost formula uses it
 B. Role tabs rebuilt with SUMMARY AT TOP; SA&D drops portfolio-squad roles;
    BP/DA "how many times is the overhead applied" lines (11x, FTE-equivalents)
 C. Cyber single-source: 1.11 squads are live refs to 2.5 groupings -> no
    double count, no drift
 D. 2.1 Total Cost: de-duplicates overhead-funded COE money (BP+DA netting row)
 E. 3.0 FTE View: platform subtotals (incl platform OH) and portfolio totals
    (incl portfolio OH) -> platform & portfolio costs visible; summary on top
 F. Exec Summary as the first tab (story + drill-down); 4.0 Insights removed
 G. Tab renumbering: 2.0 Group Summary, 2.1 Total Cost, 2.2 COE, 2.3 BP&T,
    2.4 SA&D, 2.5 Cyber Roles, 3.0 FTE View, 3.1 Data QA
 H. Emits anchors_v8.json for the QA harness
"""
import openpyxl, re, json
from collections import Counter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

SCR = "/tmp/claude-0/-home-user-anthropic-claude-code/6161aafe-2dad-5bc5-ad55-d8a92ce554cc/scratchpad/"
SRC = SCR + "TDD_Cost_Calc_v7.xlsx"
OUT = SCR + "TDD_Cost_Calc_v8.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=False)
wbv = openpyxl.load_workbook(SRC, data_only=True)

NAVY_F = PatternFill("solid", fgColor="FF1F4E79"); DK_F = PatternFill("solid", fgColor="FF002F6C")
GREY_F = PatternFill("solid", fgColor="FFD9D9D9"); YELLOW = PatternFill("solid", fgColor="FFFFF2CC")
LT_F = PatternFill("solid", fgColor="FFEDF2F7")
W = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
B = Font(name="Calibri", size=10, bold=True)
N = Font(name="Calibri", size=10)
BLUE = Font(name="Calibri", size=10, color="FF0000FF")
GREEN = Font(name="Calibri", size=10, color="FF008000")
TITLE = Font(name="Calibri", size=16, bold=True)
BIG = Font(name="Calibri", size=14, bold=True)
thin = Side(style="thin", color="FFB8C9CC")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '#,##0.00;\\(#,##0.00\\);\\-'
DOLLAR = '$#,##0;($#,##0);\\-'
PCT = '0%'
GREEN_F = PatternFill("solid", fgColor="FFE2EFDA"); GREEN_T = Font(color="FF006100", bold=True)
RED_F = PatternFill("solid", fgColor="FFFBE4D5"); RED_T = Font(color="FF9C0006", bold=True)
def pos_bad(ws, rng):
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=RED_F, font=RED_T))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=GREEN_F, font=GREEN_T))
def sc(ws, coord, value, font=None, fill=None, fmt=None, border=True, align=None):
    c = ws[coord]; c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if border: c.border = BOX
    if align: c.alignment = Alignment(horizontal=align, vertical="center")
    return c
def frow(ws, col, text, exact=True, start=1):
    ci = openpyxl.utils.column_index_from_string(col)
    for r in range(start, ws.max_row + 2):
        v = ws.cell(r, ci).value
        if v is None: continue
        s = str(v).strip()
        if (exact and s == text) or (not exact and text in s):
            return r
    return None

anchors = {}
TABS = ["1.1 Ampol Retail","1.2 Customer","1.3 Enterprise Data","1.4 TDD Group Functions",
        "1.5 P&C","1.6 Finance","1.7 Infrastructure","1.8 Energy Solutions & B2B",
        "1.9 Commercial Fuels","1.10 Z Retail","1.11 TDD Cyber"]
PORT = {t: t.split(" ", 1)[1] for t in TABS}

# =====================================================================
# A. offshore rate cell + rename 2.1 COE -> 2.2 COE + drop old tabs
# =====================================================================
sq = wb["0.1 Squads"]
sc(sq, "K4", "Offshore rate", W, NAVY_F, align="center")
sc(sq, "K5", 0.4, B, GREY_F, PCT, align="center")
OFFRATE = "'0.1 Squads'!$K$5"

for old in ["2.2 BP&T", "2.3 SA&D", "2.4 Cyber Roles", "5.0 Total Cost", "5.1 Data QA", "4.0 Insights"]:
    if old in wb.sheetnames: del wb[old]
wb["2.1 COE"].title = "2.2 COE"
# rewrite every formula + label referencing the old COE name
pat = re.compile(r"'2\.1 COE'")
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                if "'2.1 COE'" in c.value:
                    c.value = c.value.replace("'2.1 COE'", "'2.2 COE'")
                elif "(2.1 COE)" in c.value:
                    c.value = c.value.replace("(2.1 COE)", "(2.2 COE)")
coe = wb["2.2 COE"]

# =====================================================================
# B. role tabs rebuilt - summary on top, working offshore toggle
# =====================================================================
# RAW DATA is the source of truth for every roster, mapping and vacancy (the new model);
# Added data is used ONLY to map a cost onto each role.
ad = wb["Added data"]; adv = wbv["Added data"]
rdv0 = wbv["raw data"]
_adn = Counter(str(adv.cell(r, 2).value or "").strip().lower() for r in range(2, 550))
BPT_P = {"COE - Business Partnering": "Business Partnering", "COE - Transformation": "Transformation"}
SAD_P = {"COE - Strategy Architecture": "Strategy & Architecture", "COE - Data": "Data - COE"}
CYB_D = {"cyber strat & tech": "TDD COE", "cyber risk": "TDD COE", "cyber grc": "TDD COE",
         "cyber sec ops": "TDD Cyber", "service op & assurance": "TDD Cyber"}
groups = {"BPT": [], "SAD": [], "CYB": []}
unspec = []
def cost_kind(name, status):
    # unique filled names match the cost ledger by name; vacant seats and duplicate or
    # placeholder names ('Vacant', 'ring fenced') take the ledger's rate for the title
    if status == "Vacant" or _adn.get(name.strip().lower(), 0) != 1:
        return "title"
    return "name"
for r in range(2, rdv0.max_row + 1):
    msq = str(rdv0.cell(r, 16).value or "").strip()
    port = str(rdv0.cell(r, 14).value or "").strip()
    name = str(rdv0.cell(r, 2).value or "").strip()
    title = str(rdv0.cell(r, 3).value or "").strip()
    dept = str(rdv0.cell(r, 7).value or "").strip()
    ctry = str(rdv0.cell(r, 13).value or "").strip()
    st = str(rdv0.cell(r, 18).value or "").strip()
    if not name and not title: continue
    kind = cost_kind(name, st)
    if msq in BPT_P:
        groups["BPT"].append((r, name, title, dept, ctry, BPT_P[msq], kind))
    elif msq in SAD_P:
        groups["SAD"].append((r, name, title, dept, ctry, SAD_P[msq], kind))
    elif msq == "COE (unspecified)":
        unspec.append((r, name, title, dept, ctry, "COE (unspecified)", kind))
    elif port == "TDD Cyber" and dept.lower() in CYB_D:
        groups["CYB"].append((r, name, title, dept, ctry, CYB_D[dept.lower()], kind))
# cost criteria reference the RAW source cells directly, so displayed titles can be
# dash-sanitised without breaking the name/title matching into the cost ledger
AD_COST = ("=IFERROR(INDEX('Added data'!$AA$2:$AA$549,MATCH('raw data'!$B${src},'Added data'!$B$2:$B$549,0)),"
           "SUMIF('Added data'!$C$2:$C$549,'raw data'!$C${src},'Added data'!$AA$2:$AA$549)"
           "/COUNTIF('Added data'!$C$2:$C$549,'raw data'!$C${src}))")
AD_COST_T = ("=SUMIF('Added data'!$C$2:$C$549,'raw data'!$C${src},'Added data'!$AA$2:$AA$549)"
             "/COUNTIF('Added data'!$C$2:$C$549,'raw data'!$C${src})")

ls = wb["Lists"]
ls["F1"] = "SAD Category"; ls["F2"] = "Strategy & Architecture"; ls["F3"] = "Data - COE"; ls["F4"] = "Exclude"
if ls["F5"].value: ls["F5"] = None
wb.defined_names["SADCat"].value = "Lists!$F$2:$F$4"
if "CYBCat" not in wb.defined_names:
    ls["G1"] = "CYB Category"; ls["G2"] = "TDD COE"; ls["G3"] = "TDD Cyber"; ls["G4"] = "Exclude"
    wb.defined_names["CYBCat"] = openpyxl.workbook.defined_name.DefinedName("CYBCat", attr_text="Lists!$G$2:$G$4")

DC = "'0.0 Data Config'"
def role_tab(sheetname, title_txt, rows, cat_named, cats, budget_map, extra_info):
    ws = wb.create_sheet(sheetname)
    ws.sheet_view.showGridLines = False
    for col, wd in {"A":3,"B":30,"C":42,"D":26,"E":10,"F":9,"G":20,"H":11,"I":13,"J":13,"K":46}.items():
        ws.column_dimensions[col].width = wd
    ws.row_dimensions[2].height = 21
    sc(ws, "B2", title_txt, TITLE, border=False)
    # ---- SUMMARY ON TOP ----
    r = 4
    sc(ws, f"B{r}", "Summary", W, DK_F, align="left")
    for col in "CDEFGHIJ": sc(ws, f"{col}{r}", None, W, DK_F)
    r += 1
    hdr = ["Category","Roles","Filled","Vacant","Planned spend ($m)","Budget to draw down ($m)","Left to fund ($m)"]
    for j, h in enumerate(hdr):
        sc(ws, f"{'BCDEFGH'[j]}{r}", h, W, NAVY_F, align="center")
    r += 1
    sum_first = r
    cat_cells = {}
    n_roles = len(rows)
    tbl_first = None   # filled later; formulas reference planned range below
    for cat in cats:
        cat_cells[cat] = r
        sc(ws, f"B{r}", cat, N, align="left")
        r += 1
    tot_r = r
    sc(ws, f"B{tot_r}", "Total", B, GREY_F, align="left")
    r += 2
    info_first = r
    for label, formula, fmt in extra_info:
        sc(ws, f"B{r}", label, N, LT_F, align="left")
        sc(ws, f"C{r}", formula, B, LT_F, fmt, align="right")
        r += 1
    r += 1
    # ---- ROLES TABLE ----
    sc(ws, f"B{r}", "Roles", W, DK_F, align="left")
    for col in "CDEFGHIJ": sc(ws, f"{col}{r}", None, W, DK_F)
    r += 1
    rhdr = ["Name","Position Title","Department","Country","Status","Category","On/Off",
            "Full Cost AUD ($)","Model cost ($m)"]
    for j, h in enumerate(rhdr):
        sc(ws, f"{'BCDEFGHIJ'[j]}{r}", h, W, NAVY_F, align="center")
    r += 1
    tbl_first = r
    dvo = DataValidation(type="list", formula1="OnOff", allow_blank=True)
    dvc = DataValidation(type="list", formula1=cat_named, allow_blank=True) if cat_named else None
    ws.add_data_validation(dvo)
    if dvc: ws.add_data_validation(dvc)
    for (srcrow, name, title, dept, ctry, cat, kind) in sorted(rows, key=lambda x: (x[5], x[3], x[1])):
        sc(ws, f"B{r}", f"=IF('raw data'!$R${srcrow}=\"Vacant\",\"Vacant\",'raw data'!$B${srcrow})", N, align="left")
        sc(ws, f"C{r}", f"=SUBSTITUTE(SUBSTITUTE('raw data'!$C${srcrow},\"–\",\"-\"),\"—\",\"-\")", N, align="left")
        sc(ws, f"D{r}", dept, N, align="left")
        sc(ws, f"E{r}", ctry.title() if ctry.islower() else ctry, N, align="left")
        sc(ws, f"F{r}", f"='raw data'!$R${srcrow}", N, align="center")
        sc(ws, f"G{r}", cat, N, YELLOW if cat_named else None, align="left")
        sc(ws, f"H{r}", "Onshore", N, YELLOW, align="center")
        cost_f = (AD_COST if kind == "name" else AD_COST_T).format(src=srcrow)
        sc(ws, f"I{r}", cost_f, N, None, DOLLAR, align="right")
        sc(ws, f"J{r}", f"=IF(H{r}=\"Offshore\",I{r}*{OFFRATE},I{r})/1000000", N, None, MONEY, align="right")
        dvo.add(f"H{r}")
        if dvc: dvc.add(f"G{r}")
        r += 1
    tbl_last = r - 1
    ws.conditional_formatting.add(f"F{tbl_first}:F{tbl_last}",
        CellIsRule(operator="equal", formula=['"Vacant"'], fill=RED_F, font=RED_T))
    # wire the summary block
    for cat, cr in cat_cells.items():
        sc(ws, f"C{cr}", f"=COUNTIF(G{tbl_first}:G{tbl_last},B{cr})", N, align="center")
        sc(ws, f"D{cr}", f"=COUNTIFS(G{tbl_first}:G{tbl_last},B{cr},F{tbl_first}:F{tbl_last},\"Filled\")", N, align="center")
        sc(ws, f"E{cr}", f"=COUNTIFS(G{tbl_first}:G{tbl_last},B{cr},F{tbl_first}:F{tbl_last},\"Vacant\")", N, align="center")
        sc(ws, f"F{cr}", f"=SUMIF(G{tbl_first}:G{tbl_last},B{cr},J{tbl_first}:J{tbl_last})", N, None, MONEY, align="right")
        bud = budget_map.get(cat, "=0")
        sc(ws, f"G{cr}", bud, N, None, MONEY, align="right")
        sc(ws, f"H{cr}", f"=MAX(0,F{cr}-G{cr})", B, None, MONEY, align="right")
        pos_bad(ws, f"H{cr}")
    for col, cl in [("C","C"),("D","D"),("E","E"),("F","F"),("G","G"),("H","H")]:
        sc(ws, f"{col}{tot_r}", f"=SUM({cl}{sum_first}:{cl}{tot_r-1})", B, GREY_F,
           MONEY if col in "FGH" else None, align="right" if col in "FGH" else "center")
    ckr = tbl_last + 1
    sc(ws, f"B{ckr}", "Check - roles listed vs counted (must be 0)", N, LT_F, align="left")
    sc(ws, f"C{ckr}", f"=COUNTA(B{tbl_first}:B{tbl_last})-C{tot_r}", B, LT_F, align="center")
    pos_bad(ws, f"C{ckr}")
    return ws, dict(sum_first=sum_first, tot=tot_r, cats=cat_cells,
                    tbl_first=tbl_first, tbl_last=tbl_last, info_first=info_first, ck=ckr)

bp_ws, bp_a = role_tab("2.3 BP&T", "Business Partnering & Transformation - roles and funding",
    groups["BPT"], "BPTCat", ["Business Partnering", "Transformation"],
    {"Business Partnering": f"=11*{DC}!$L$7", "Transformation": f"={DC}!$E$8"},
    [("Business Partner overhead applied - portfolios", 11, None),
     ("Funding per portfolio ($m) - Data Config", f"={DC}!$L$7", MONEY),
     ("Total BP funding from portfolio overheads ($m)", f"=11*{DC}!$L$7", MONEY),
     ("BP FTEs funded (11 x 40% allocation)", f"=11*{DC}!$K$7", None),
     ("Transformation funding - COE allocation ($m)", f"={DC}!$E$8", MONEY)])
sad_ws, sad_a = role_tab("2.4 SA&D", "Strategy, Architecture & Data - roles and funding",
    groups["SAD"], "SADCat", ["Strategy & Architecture", "Data - COE"],
    {"Strategy & Architecture": f"={DC}!$E$6+11*{DC}!$L$8", "Data - COE": f"={DC}!$E$10"},
    [("Domain Architect overhead applied - portfolios", 11, None),
     ("Funding per portfolio ($m) - Data Config", f"={DC}!$L$8", MONEY),
     ("Total DA funding from portfolio overheads ($m)", f"=11*{DC}!$L$8", MONEY),
     ("DA FTEs funded (11 x 50% allocation)", f"=11*{DC}!$K$8", None),
     ("Strategy & Architecture COE allocation ($m)", f"={DC}!$E$6", MONEY),
     ("Data COE allocation ($m)", f"={DC}!$E$10", MONEY)])
cy_ws, cy_a = role_tab("2.5 Cyber Roles", "Cyber, Risk & Service Operations - roles and funding",
    groups["CYB"], "CYBCat", ["TDD COE", "TDD Cyber"],
    {"TDD COE": "=0", "TDD Cyber": "=0"},
    [("TDD Cyber people budget - Data Config ($m)", f"={DC}!$E$23", MONEY),
     ("Cyber CapEx - Monitoring, tracked separately ($m)", "='1.11 TDD Cyber'!$H$16", MONEY),
     ("Planned spend - all cyber roles ($m)", None, MONEY),
     ("Left to fund - roles vs people budget ($m)", None, MONEY),
     ("Portfolio + platform overheads charged on 1.11 ($m)", None, MONEY),
     ("Total to fund - ties to 1.11 TDD Cyber ($m)", None, MONEY)])
# finish cyber info rows (need anchor refs; overheads line is wired in the post-pass)
cyt = cy_a["tot"]
inf = cy_a["info_first"]
cy_ws[f"C{inf+2}"] = f"=F{cyt}"
cy_ws[f"C{inf+3}"] = f"=H{cyt}"
cy_ws[f"C{inf+5}"] = f"=C{inf+3}+C{inf+4}"
cy_tie = inf + 5
# category budget cells stay blank; the people budget applies to the tab total only
for cat, cr in cy_a["cats"].items():
    cy_ws[f"G{cr}"] = None; cy_ws[f"H{cr}"] = None
cy_ws[f"G{cyt}"] = f"={DC}!$E$23"
cy_ws[f"H{cyt}"] = f"=MAX(0,F{cyt}-G{cyt})"
pos_bad(cy_ws, f"H{cyt}")

# =====================================================================
# C. 1.11 cyber single-source: squads = live refs to 2.5 groupings
# =====================================================================
cy11 = wb["1.11 TDD Cyber"]
coe_r = cy_a["cats"]["TDD COE"]; cyb_r = cy_a["cats"]["TDD Cyber"]
cy11["G24"] = f"='2.5 Cyber Roles'!$F${coe_r}"
cy11["G24"].font = GREEN; cy11["G24"].fill = PatternFill(); cy11["G24"].number_format = MONEY
cy11["G25"] = f"='2.5 Cyber Roles'!$F${cyb_r}"
cy11["G25"].font = GREEN; cy11["G25"].fill = PatternFill(); cy11["G25"].number_format = MONEY
cy11["C24"] = "Strategic Programs"; cy11["D24"] = None
cy11["C25"] = "Strategic Programs"; cy11["D25"] = None

# =====================================================================
# 2.2 COE hub: planned refs to new tabs
# =====================================================================
coe["D8"] = f"='2.4 SA&D'!$F${sad_a['cats']['Strategy & Architecture']}"
coe["D10"] = f"='2.3 BP&T'!$F${bp_a['cats']['Transformation']}"
coe["D11"] = f"='2.3 BP&T'!$F${bp_a['cats']['Business Partnering']}"
coe["D12"] = f"='2.4 SA&D'!$F${sad_a['cats']['Data - COE']}"
# ONE definition of "Left to fund" everywhere: 2.2 pulls the role tabs' netted figures,
# and the hub foots on its face: Planned (D) - Budget to draw down (E) = Left to fund (F)
coe["E7"] = "Budget to draw down ($m)"
coe["E8"] = f"={DC}!$E$6+11*{DC}!$L$8"
coe["E9"] = 0
coe["E10"] = f"={DC}!$E$8"
coe["E11"] = f"=11*{DC}!$L$7"
coe["E12"] = f"={DC}!$E$10"
coe["E13"] = "=SUM(E8:E12)"
coe["F8"] = f"='2.4 SA&D'!$H${sad_a['cats']['Strategy & Architecture']}"
coe["F10"] = f"='2.3 BP&T'!$H${bp_a['cats']['Transformation']}"
coe["F11"] = f"='2.3 BP&T'!$H${bp_a['cats']['Business Partnering']}"
coe["F12"] = f"='2.4 SA&D'!$H${sad_a['cats']['Data - COE']}"
# COE (unspecified): raw maps these roles to no COE bucket - surface them for review
ur = 23
sc(coe, f"B{ur}", "COE (unspecified) - mapped to no COE bucket in the raw model, review these", W, DK_F, align="left")
for col in "CDEF": sc(coe, f"{col}{ur}", None, W, DK_F)
ur += 1
for j, h in enumerate(["Name", "Position Title", "Department", "Status", "Mapped cost ($)"]):
    sc(coe, f"{'BCDEF'[j]}{ur}", h, W, NAVY_F, align="center")
ur += 1
un_first2 = ur
for (srcrow, name, title, dept, ctry, cat, kind) in unspec:
    isvac = name.lower() == "vacant"
    sc(coe, f"B{ur}", f"=IF('raw data'!$R${srcrow}=\"Vacant\",\"Vacant\",'raw data'!$B${srcrow})", N,
       RED_F if isvac else None, align="left")
    sc(coe, f"C{ur}", f"=SUBSTITUTE(SUBSTITUTE('raw data'!$C${srcrow},\"–\",\"-\"),\"—\",\"-\")", N, align="left")
    sc(coe, f"D{ur}", dept, N, align="left")
    sc(coe, f"E{ur}", f"='raw data'!$R${srcrow}", N, align="center")
    sc(coe, f"F{ur}", (AD_COST if kind == "name" else AD_COST_T).format(src=srcrow), N, None, DOLLAR, align="right")
    ur += 1
sc(coe, f"B{ur}", "Total", B, GREY_F, align="left")
sc(coe, f"F{ur}", f"=SUM(F{un_first2}:F{ur-1})", B, GREY_F, DOLLAR, align="right")
unspec_tot_row = ur

# =====================================================================
# D. 2.1 Total Cost (rebuilt, with BP/DA de-duplication row)
# =====================================================================
AD = "'Added data'"
def sumifs(port=None, cls=None, squad=None, status=None):
    conds = []
    if port: conds.append(f"{AD}!$AC:$AC,\"{port}\"")
    if squad: conds.append(f"{AD}!$AE:$AE,\"{squad}\"")
    if cls: conds.append(f"{AD}!$AF:$AF,\"{cls}\"")
    if status: conds.append(f"{AD}!$AG:$AG,\"{status}\"")
    return f"SUMIFS({AD}!$AA:$AA," + ",".join(conds) + ")"
ss_rows = {}; oh_rows = {}
for t in TABS:
    w = wb[t]
    for rr in range(1, 15):
        if w.cell(rr, 2).value == "Squad Support Costs": ss_rows[t] = rr
        if w.cell(rr, 2).value == "Portfolio Overhead": oh_rows[t] = rr
tc = wb.create_sheet("2.1 Total Cost", wb.sheetnames.index("2.0 Group Summary") + 1)
tc.sheet_view.showGridLines = False
for col, wd in {"A":3,"B":34,"C":14,"D":14,"E":14,"F":14,"G":14}.items():
    tc.column_dimensions[col].width = wd
tc.row_dimensions[2].height = 21
sc(tc, "B2", "Total Cost - archetype model vs actual org, every layer of the operating model", TITLE, border=False)
HD = ["","Archetype model ($m)","Actual Filled ($m)","Actual Vacant ($m)","Actual Total ($m)","Actual over/(under) archetype ($m)"]
r = 4
def section(title):
    global r
    sc(tc, f"B{r}", title, W, DK_F, align="left")
    for col in "CDEFG": sc(tc, f"{col}{r}", None, W, DK_F)
    r += 1
    for j, h in enumerate(HD):
        if h: sc(tc, f"{'BCDEFG'[j]}{r}", h, W, NAVY_F, align="center")
    sc(tc, f"B{r}", "Portfolio", W, NAVY_F, align="center")
    r += 1
def rowx(label, model_f, actual_pair, bold=False, fillc=None):
    global r
    sc(tc, f"B{r}", label, B if bold else N, fillc, align="left")
    sc(tc, f"C{r}", model_f, B if bold else N, fillc, MONEY, align="right")
    sc(tc, f"D{r}", actual_pair[0], B if bold else N, fillc, MONEY, align="right")
    sc(tc, f"E{r}", actual_pair[1], B if bold else N, fillc, MONEY, align="right")
    sc(tc, f"F{r}", f"=D{r}+E{r}", B if bold else N, fillc, MONEY, align="right")
    sc(tc, f"G{r}", f"=ROUND(F{r}-C{r},6)", B if bold else N, fillc, MONEY, align="right")
    r += 1
def subtotal(label):
    global r
    first = subtotal.first
    sc(tc, f"B{r}", label, B, GREY_F, align="left")
    for col in "CDEFG":
        sc(tc, f"{col}{r}", f"=SUM({col}{first}:{col}{r-1})", B, GREY_F, MONEY, align="right")
    out = r; r += 1
    return out
def actual_pair(port=None, cls_list=("Squad","Strategic Program"), squad=None, cls_single=None):
    if cls_single:
        return (f"=({sumifs(cls=cls_single, port=port, squad=squad, status='Filled')})/1000000",
                f"=({sumifs(cls=cls_single, port=port, squad=squad, status='Vacant')})/1000000")
    return ("=(" + "+".join(sumifs(cls=c, port=port, squad=squad, status="Filled") for c in cls_list) + ")/1000000",
            "=(" + "+".join(sumifs(cls=c, port=port, squad=squad, status="Vacant") for c in cls_list) + ")/1000000")

section("Squads & strategic programs (support costs)")
subtotal.first = r
for t in TABS:
    sr_ = ss_rows[t]
    rowx(PORT[t], f"='{t}'!$C${sr_}+'{t}'!$D${sr_}", actual_pair(port=PORT[t]))
subA = subtotal("Subtotal - squads & programs")
r += 1
section("Leadership & overheads (allowance vs actual leadership cost)")
subtotal.first = r
for t in TABS:
    ohr = oh_rows[t]
    rowx(PORT[t], f"='{t}'!$C${ohr}+'{t}'!$C${ohr+1}", actual_pair(port=PORT[t], cls_single="Leadership"))
rowx("COE / central leadership", "=0", actual_pair(port="COE", cls_single="Leadership"))
rowx("Unmapped leadership", "=0", actual_pair(port="Unmapped", cls_single="Leadership"))
subB = subtotal("Subtotal - leadership & overheads")
r += 1
section("Centres of Excellence")
subtotal.first = r
rowx("COE - Business Partnering", "='2.2 COE'!$D$11", actual_pair(squad="COE - Business Partnering", cls_single="COE"))
rowx("COE - Transformation", "='2.2 COE'!$D$10", actual_pair(squad="COE - Transformation", cls_single="COE"))
rowx("COE - Strategy Architecture", "='2.2 COE'!$D$8", actual_pair(squad="COE - Strategy Architecture", cls_single="COE"))
rowx("COE - Data", "='2.2 COE'!$D$12", actual_pair(squad="COE - Data", cls_single="COE"))
rowx("COE (unspecified)", "=0", actual_pair(squad="COE (unspecified)", cls_single="COE"))
rowx("Less: BP & Domain Architect already funded in portfolio overheads",
     f"=-(11*{DC}!$L$7+11*{DC}!$L$8)", ("=0", "=0"))
dedup_row = r - 1
subC = subtotal("Subtotal - Centres of Excellence (net of overhead funding)")
r += 1
section("Not in the model")
subtotal.first = r
rowx("Unmapped roles", "=0", actual_pair(cls_single="Unmapped"))
subD = subtotal("Subtotal - not in the model")
r += 2
sc(tc, f"B{r}", "TOTAL OPERATING MODEL", W, DK_F, align="left")
for col in "CDEFG":
    sc(tc, f"{col}{r}", f"={col}{subA}+{col}{subB}+{col}{subC}+{col}{subD}", W, DK_F, MONEY, align="right")
GRAND = r
r += 1
sc(tc, f"B{r}", "Check: Actual Total = Added data full cost", B, align="left")
sc(tc, f"C{r}", f"=F{GRAND}-SUM({AD}!$AA$2:$AA$549)/1000000", B, GREY_F, MONEY, align="right")
pos_bad(tc, f"C{r}")
CHECK = r
pos_bad(tc, f"G6:G{GRAND}")

# =====================================================================
# E. 3.0 FTE View rebuilt with platform + portfolio cost rollups
# =====================================================================
old = wb["3.0 FTE View"]; idx = wb.sheetnames.index("3.0 FTE View")
wb.remove(old)
ft = wb.create_sheet("3.0 FTE View", idx)
ft.sheet_view.showGridLines = False
for col, wd in {"A":3,"B":26,"C":28,"D":30,"E":26,"F":7,"G":13,"H":8,"I":8,"J":9,"K":15,"L":11,"M":16,"N":15,"O":15}.items():
    ft.column_dimensions[col].width = wd
ft.row_dimensions[2].height = 21
sc(ft, "B2", "FTE View - archetype model vs actual organisation", TITLE, border=False)
RD = "'raw data'"
kpis = [("Org roles", f"=COUNTA({RD}!$R$2:$R$1000)"),
        ("Filled", f"=COUNTIF({RD}!$R$2:$R$1000,\"Filled\")"),
        ("Vacant", f"=COUNTIF({RD}!$R$2:$R$1000,\"Vacant\")"),
        ("Vacancy rate", "=E4/C4"),
        ("Archetype roles (squads)", None), ("Seats above archetypes", None),
        ("Archetype cost incl overheads ($m)", None), ("Actual squad cost ($m)", None)]
kcols = "CDEFGHIJ"
for j, (label, f) in enumerate(kpis):
    sc(ft, f"{kcols[j]}3", label, W, NAVY_F, align="center")
    sc(ft, f"{kcols[j]}4", f, B, GREY_F, PCT if label == "Vacancy rate" else (MONEY if "$m" in label else None), align="center")
HDR = ["Portfolio","Platform","Squad","Squad Type","Size","Archetype FTE","Filled","Vacant",
       "Seats","Seats vs archetype","Vacancy %","Archetype cost ($m)","Actual cost ($m)","Cost var - squads ($m)"]
r = 6
for j, h in enumerate(HDR):
    if h: sc(ft, f"{'BCDEFGHIJKLMNO'[j]}{r}", h, W, NAVY_F, align="center")
r += 1
def blocks(w):
    out = []
    rows = [rr for rr in range(1, w.max_row + 1)
            if isinstance(w.cell(rr, 2).value, str) and w.cell(rr, 2).value.startswith("Platform: ")]
    for tr in rows:
        hr = tr + 1; ohr = None
        for rr in range(hr + 1, w.max_row + 2):
            bv = w.cell(rr, 2).value
            if bv == "Platform Overhead" or (isinstance(bv, str) and (bv.endswith(" Total") or bv.startswith("(combined"))):
                ohr = rr; break
        out.append((tr, hr, hr + 1, (ohr or hr + 2) - 1))
    return out
first_data = r
port_sub = {}
ft_rows = {}
for t in TABS:
    w = wb[t]
    plat_rows = []
    for (tr, hrr, s0, s1) in blocks(w):
        pname = str(w.cell(tr, 2).value).replace("Platform: ", "")
        if "combined into" in pname: continue
        p_first = r
        if t == "1.11 TDD Cyber" and s0 <= 24 <= s1:
            # single consolidated cyber row: model and actual on the same population basis
            pl = PORT[t]
            sc(ft, f"B{r}", pl, N, align="left")
            sc(ft, f"C{r}", pname, N, align="left")
            sc(ft, f"D{r}", "All cyber squads (detail on 2.5 Cyber Roles)", N, align="left")
            sc(ft, f"E{r}", "Strategic Programs", N, align="left")
            sc(ft, f"F{r}", "-", N, align="center")
            sc(ft, f"G{r}", "-", N, align="center")
            sc(ft, f"H{r}", (f"=COUNTIFS({RD}!$N:$N,\"{pl}\",{RD}!$Q:$Q,\"Squad\",{RD}!$R:$R,\"Filled\")"
                             f"+COUNTIFS({RD}!$N:$N,\"{pl}\",{RD}!$Q:$Q,\"Strategic Program\",{RD}!$R:$R,\"Filled\")"), N, align="center")
            sc(ft, f"I{r}", (f"=COUNTIFS({RD}!$N:$N,\"{pl}\",{RD}!$Q:$Q,\"Squad\",{RD}!$R:$R,\"Vacant\")"
                             f"+COUNTIFS({RD}!$N:$N,\"{pl}\",{RD}!$Q:$Q,\"Strategic Program\",{RD}!$R:$R,\"Vacant\")"), N, align="center")
            sc(ft, f"J{r}", f"=H{r}+I{r}", N, align="center")
            sc(ft, f"K{r}", "-", N, align="center")
            sc(ft, f"L{r}", f'=IFERROR(I{r}/J{r},"-")', N, None, PCT, align="center")
            sc(ft, f"M{r}", "='1.11 TDD Cyber'!$G$24+'1.11 TDD Cyber'!$G$25", N, None, MONEY, align="right")
            sc(ft, f"N{r}", f"=({sumifs(port=pl)})/1000000", N, None, MONEY, align="right")
            sc(ft, f"O{r}", f"=N{r}-M{r}", N, None, MONEY, align="right")
            ft_rows.setdefault(t, []).append(("__CYBER__", r))
            r += 1
            s0, s1 = 1, 0   # suppress the per-squad loop below
        for srr in range(s0, s1 + 1):
            sqn = w.cell(srr, 2).value
            if not sqn: continue
            pl = PORT[t]
            sc(ft, f"B{r}", pl, N, align="left")
            sc(ft, f"C{r}", pname, N, align="left")
            sc(ft, f"D{r}", sqn, N, align="left")
            sc(ft, f"E{r}", f"='{t}'!$C${srr}", N, align="left")
            sc(ft, f"F{r}", f"='{t}'!$D${srr}", N, align="center")
            sc(ft, f"G{r}", (f"=IFERROR(INDEX('0.1 Squads'!$F$5:$F$23,MATCH('{t}'!$C${srr}&\"|\"&'{t}'!$D${srr},"
                             f"'0.1 Squads'!$A$5:$A$23,0)),\"-\")"), N, align="center")
            sc(ft, f"H{r}", f"=COUNTIFS({RD}!$N:$N,\"{pl}\",{RD}!$P:$P,\"{sqn}\",{RD}!$R:$R,\"Filled\")", N, align="center")
            sc(ft, f"I{r}", f"=COUNTIFS({RD}!$N:$N,\"{pl}\",{RD}!$P:$P,\"{sqn}\",{RD}!$R:$R,\"Vacant\")", N, align="center")
            sc(ft, f"J{r}", f"=H{r}+I{r}", N, align="center")
            sc(ft, f"K{r}", f'=IFERROR(J{r}-G{r},"-")', N, align="center")
            sc(ft, f"L{r}", f'=IFERROR(I{r}/J{r},"-")', N, None, PCT, align="center")
            sc(ft, f"M{r}", f"='{t}'!$G${srr}", N, None, MONEY, align="right")
            sc(ft, f"N{r}", f"=({sumifs(port=pl, squad=str(sqn))})/1000000", N, None, MONEY, align="right")
            sc(ft, f"O{r}", f"=N{r}-IFERROR(M{r},0)", N, None, MONEY, align="right")
            ft_rows.setdefault(t, []).append((str(sqn), r))
            r += 1
        # platform subtotal incl platform overhead (EGI programs carry NO platform overhead -
        # they are funded via Significant Items, matching the portfolio tabs)
        egi = pname.startswith("EGI")
        pl_label = (f"{pname} - platform total (no overhead - funded via Significant Items)" if egi
                    else f"{pname} - platform total (incl overhead)")
        ohtxt = "" if egi else f"+{DC}!$L$16"
        sc(ft, f"B{r}", None, B, LT_F); sc(ft, f"C{r}", pl_label, B, LT_F, align="left")
        sc(ft, f"D{r}", None, B, LT_F); sc(ft, f"E{r}", None, B, LT_F); sc(ft, f"F{r}", None, B, LT_F)
        for colL in "GHIJ":
            sc(ft, f"{colL}{r}", f"=SUM({colL}{p_first}:{colL}{r-1})", B, LT_F, align="center")
        sc(ft, f"K{r}", f"=SUM(K{p_first}:K{r-1})", B, LT_F, align="center")
        sc(ft, f"L{r}", f'=IFERROR(I{r}/J{r},"-")', B, LT_F, PCT, align="center")
        sc(ft, f"M{r}", f"=SUM(M{p_first}:M{r-1}){ohtxt}", B, LT_F, MONEY, align="right")
        sc(ft, f"N{r}", f"=SUM(N{p_first}:N{r-1})", B, LT_F, MONEY, align="right")
        sc(ft, f"O{r}", f"=SUM(O{p_first}:O{r-1})", B, LT_F, MONEY, align="right")
        plat_rows.append(r)
        r += 1
    # portfolio total incl portfolio overhead
    sc(ft, f"B{r}", f"{PORT[t]} - portfolio total (incl OHs)", B, GREY_F, align="left")
    for colL in "CDEF": sc(ft, f"{colL}{r}", None, B, GREY_F)
    for colL in "GHIJ":
        sc(ft, f"{colL}{r}", "=" + "+".join(f"{colL}{pr}" for pr in plat_rows), B, GREY_F, align="center")
    sc(ft, f"K{r}", "=" + "+".join(f"K{pr}" for pr in plat_rows), B, GREY_F, align="center")
    sc(ft, f"L{r}", f'=IFERROR(I{r}/J{r},"-")', B, GREY_F, PCT, align="center")
    sc(ft, f"M{r}", "=" + "+".join(f"M{pr}" for pr in plat_rows) + f"+{DC}!$L$10", B, GREY_F, MONEY, align="right")
    sc(ft, f"N{r}", "=" + "+".join(f"N{pr}" for pr in plat_rows), B, GREY_F, MONEY, align="right")
    sc(ft, f"O{r}", "=" + "+".join(f"O{pr}" for pr in plat_rows), B, GREY_F, MONEY, align="right")
    port_sub[t] = r
    r += 1
sc(ft, f"B{r}", "TOTAL - delivery organisation (archetype cost incl overheads; actual = squad roles only)", W, DK_F, align="left")
for colL in "CDEF": sc(ft, f"{colL}{r}", None, W, DK_F)
for colL in "GHIJ":
    sc(ft, f"{colL}{r}", "=" + "+".join(f"{colL}{pr}" for pr in port_sub.values()), W, DK_F, align="center")
sc(ft, f"K{r}", "=" + "+".join(f"K{pr}" for pr in port_sub.values()), W, DK_F, align="center")
sc(ft, f"L{r}", f"=I{r}/J{r}", W, DK_F, PCT, align="center")
sc(ft, f"M{r}", "=" + "+".join(f"M{pr}" for pr in port_sub.values()), W, DK_F, MONEY, align="right")
sc(ft, f"N{r}", "=" + "+".join(f"N{pr}" for pr in port_sub.values()), W, DK_F, MONEY, align="right")
sc(ft, f"O{r}", "=" + "+".join(f"O{pr}" for pr in port_sub.values()), W, DK_F, MONEY, align="right")
grand_ft = r
ft["G4"] = f"=G{grand_ft}"; ft["H4"] = f"=K{grand_ft}"
ft["I4"] = f"=M{grand_ft}"; ft["J4"] = f"=N{grand_ft}"
for colcf in ("K", "O"):
    ft.conditional_formatting.add(f"{colcf}{first_data}:{colcf}{grand_ft}",
        FormulaRule(formula=[f"AND(ISNUMBER({colcf}{first_data}),{colcf}{first_data}>0)"], fill=RED_F, font=RED_T))
    ft.conditional_formatting.add(f"{colcf}{first_data}:{colcf}{grand_ft}",
        FormulaRule(formula=[f"AND(ISNUMBER({colcf}{first_data}),{colcf}{first_data}<=0)"], fill=GREEN_F, font=GREEN_T))
ft.conditional_formatting.add(f"L{first_data}:L{grand_ft}",
    FormulaRule(formula=[f"AND(ISNUMBER(L{first_data}),L{first_data}>=0.5)"], fill=RED_F, font=RED_T))
r += 2
# leadership (names) - kept, compact
sc(ft, f"B{r}", "Leadership roles - funded by the platform & portfolio overheads above", W, DK_F, align="left")
for col in "CDEFGHIJKLMNO": sc(ft, f"{col}{r}", None, W, DK_F)
r += 1
for j, h in enumerate(["Portfolio","Name","Position Title","Org platform field"]):
    sc(ft, f"{'BCDE'[j]}{r}", h, W, NAVY_F, align="center")
r += 1
rdv2 = wbv["raw data"]
leads = []
for rr in range(2, rdv2.max_row + 1):
    if rdv2.cell(rr, 17).value == "Leadership":
        leads.append((str(rdv2.cell(rr, 14).value or ""), str(rdv2.cell(rr, 2).value or ""),
                      str(rdv2.cell(rr, 3).value or ""), str(rdv2.cell(rr, 10).value or ""), rr))
lead_first = r
for (pl, nm, ti, plat, rr) in sorted(leads):
    sc(ft, f"B{r}", pl, N, align="left")
    sc(ft, f"C{r}", f"=IF({RD}!$R${rr}=\"Vacant\",\"Vacant\",{RD}!$B${rr})", N, align="left")
    if nm.strip().lower() == "vacant": ft[f"C{r}"].fill = RED_F; ft[f"C{r}"].font = RED_T
    sc(ft, f"D{r}", f"=SUBSTITUTE(SUBSTITUTE({RD}!$C${rr},\"–\",\"-\"),\"—\",\"-\")", N, align="left")
    sc(ft, f"E{r}", plat, N, align="left")
    r += 1
sc(ft, f"B{r}", "Total leadership roles", B, GREY_F, align="left")
sc(ft, f"C{r}", f"=COUNTIF({RD}!$Q:$Q,\"Leadership\")", B, GREY_F, align="center")
sc(ft, f"D{r}", (f"=COUNTIFS({RD}!$Q:$Q,\"Leadership\",{RD}!$R:$R,\"Filled\")&\" filled / \"&"
                 f"COUNTIFS({RD}!$Q:$Q,\"Leadership\",{RD}!$R:$R,\"Vacant\")&\" vacant\""), B, GREY_F, align="left")
lead_tot = r
r += 2
sc(ft, f"B{r}", "In op model / org data but NOT in the archetype model", W, DK_F, align="left")
for col in "CDEFGHIJKLMNO": sc(ft, f"{col}{r}", None, W, DK_F)
r += 1
for j, h in enumerate(["Portfolio","Group","","","","","","Filled","Vacant","Seats"]):
    if h: sc(ft, f"{'BCDEFGHIJK'[j]}{r}", h, W, NAVY_F, align="center")
r += 1
un_first = r
for (p, s) in [("Customer","Customer AI"),("Z Retail","Data NZ"),("EGI","EGI (unassigned)")]:
    sc(ft, f"B{r}", p, N, RED_F, align="left")
    sc(ft, f"C{r}", s, N, RED_F, align="left")
    sc(ft, f"I{r}", f"=COUNTIFS({RD}!$N:$N,\"{p}\",{RD}!$P:$P,\"{s}\",{RD}!$R:$R,\"Filled\")", N, RED_F, align="center")
    sc(ft, f"J{r}", f"=COUNTIFS({RD}!$N:$N,\"{p}\",{RD}!$P:$P,\"{s}\",{RD}!$R:$R,\"Vacant\")", N, RED_F, align="center")
    sc(ft, f"K{r}", f"=I{r}+J{r}", N, RED_F, align="center")
    r += 1
sc(ft, f"B{r}", "Other unmapped", N, RED_F, align="left")
sc(ft, f"I{r}", f"=COUNTIFS({RD}!$Q:$Q,\"Unmapped\",{RD}!$R:$R,\"Filled\")-SUM(I{un_first}:I{r-1})", N, RED_F, align="center")
sc(ft, f"J{r}", f"=COUNTIFS({RD}!$Q:$Q,\"Unmapped\",{RD}!$R:$R,\"Vacant\")-SUM(J{un_first}:J{r-1})", N, RED_F, align="center")
sc(ft, f"K{r}", f"=I{r}+J{r}", N, RED_F, align="center")
un_last = r
r += 2
sc(ft, f"B{r}", "COE seats (detail on 2.3 / 2.4 / 2.5)", N, align="left")
sc(ft, f"I{r}", f"=COUNTIFS({RD}!$Q:$Q,\"COE\",{RD}!$R:$R,\"Filled\")", N, align="center")
sc(ft, f"J{r}", f"=COUNTIFS({RD}!$Q:$Q,\"COE\",{RD}!$R:$R,\"Vacant\")", N, align="center")
sc(ft, f"K{r}", f"=I{r}+J{r}", N, align="center")
coe_row = r
r += 1
sc(ft, f"B{r}", "Cross-check: org records", B, align="left")
sc(ft, f"C{r}", "=C4", B, align="center")
sc(ft, f"B{r+1}", "Accounted for above", B, align="left")
ft[f"C{r+1}"] = f"=J{grand_ft}+COUNTIF({RD}!$Q:$Q,\"Leadership\")+SUM(K{un_first}:K{un_last})+K{coe_row}"
ft[f"C{r+1}"].font = B; ft[f"C{r+1}"].border = BOX
sc(ft, f"B{r+2}", "Difference (must be 0)", B, GREY_F, align="left")
sc(ft, f"C{r+2}", f"=C{r}-C{r+1}", B, GREY_F, align="center")
pos_bad(ft, f"C{r+2}")
xcheck_row = r + 2

# =====================================================================
# F. Exec Summary (first tab) - the story + drill-down
# =====================================================================
ex = wb.create_sheet("Exec Summary", 0)
ex.sheet_view.showGridLines = False
for col, wd in {"A":3,"B":56,"C":15,"D":15,"E":15,"F":15,"G":15}.items():
    ex.column_dimensions[col].width = wd
ex.row_dimensions[2].height = 24
sc(ex, "B2", "TDD Operating Model - Executive Summary", Font(name="Calibri", size=18, bold=True), border=False)
GS = "'2.0 Group Summary'"
TC = "'2.1 Total Cost'"
r = 4
sc(ex, f"B{r}", "Why this workbook exists", W, DK_F, align="left")
for col in "CDEFG": sc(ex, f"{col}{r}", None, W, DK_F)
r += 1
for txt in [
    "Purpose: each GM designed a new operating model against set squad archetypes (type x size = people and cost). This workbook prices the archetypes and checks the actual design against them.",
    "The test: can you fund your archetype cost (TDD + business)? If yes, live within it. If not, pull the levers - first the vacancies (your 4.x GM tab), then archetype size itself (the 1.x dropdowns).",
    "Your portfolio: your 1.x tab has your squads, sizes, support %, budget draw-downs and what is left to fund.",
    "Next step: agree funding for what is left to fund, and decide which vacant seats to hire or hold."]:
    sc(ex, f"B{r}", txt, N, align="left")
    for col in "CDEFG": sc(ex, f"{col}{r}", None, N)
    r += 1
r += 1
sc(ex, f"B{r}", "How the model is built - key decisions", W, DK_F, align="left")
for col in "CDEFG": sc(ex, f"{col}{r}", None, W, DK_F)
r += 1
for txt in [
    "Squads are priced from the archetype library on 0.1 Squads (type x size). Offshore costs 40% of onshore.",
    "Each portfolio pays one overhead: Head of Tech + 0.4 Business Partner + Domain Architect + leadership (0.0 Data Config). Each platform pays a platform overhead.",
    "Strategic programs (AmPOS, CTRM, EGI) are priced at their entered cost. EGI is funded from Significant Items, not overheads.",
    "TDD Cyber is priced from its actual roles on 2.5 Cyber Roles, not archetypes.",
    "Business Partner & Domain Architect money sits inside the portfolio overheads, so it is taken out of the COE total - never counted twice."]:
    sc(ex, f"B{r}", txt, N, align="left")
    for col in "CDEFG": sc(ex, f"{col}{r}", None, N)
    r += 1
r += 1
sc(ex, f"B{r}", "The money", W, DK_F, align="left")
for col in "CDEFG": sc(ex, f"{col}{r}", None, W, DK_F)
r += 1
for label, f in [
    ("The TDD budget", "HDR"),
    ("Total TDD people budget ($m)", f"={DC}!$E$27"),
    ("Allocated to portfolios + COEs ($m)", f"={GS}!$C$30"),
    ("Not yet allocated ($m)", f"={DC}!$E$27-{GS}!$C$30"),
    ("What the archetype model costs", "HDR"),
    ("TDD Cost - funded by TDD ($m)", f"={GS}!$D$24"),
    ("Funded outside TDD ($m)", f"={GS}!$G$24"),
    ("Less: BP & Domain Architect already funded in portfolio overheads ($m)", f"={TC}!$C${dedup_row}"),
    ("Total archetype model cost ($m)", f"={TC}!$C${GRAND}"),
    ("of which TDD pays - after the double-count ($m)", f"={GS}!$D$24+{TC}!$C${dedup_row}"),
    ("Left to fund - portfolios + COEs ($m)", f"={GS}!$I$24"),
    ("What the org actually costs today", "HDR"),
    ("All roles today - filled + vacant ($m)", f"={TC}!$F${GRAND}"),
    ("of which filled ($m)", f"={TC}!$D${GRAND}"),
    ("of which vacant ($m)", f"={TC}!$E${GRAND}"),
    ("Difference vs archetype model ($m)", f"={TC}!$G${GRAND}")]:
    if f == "HDR":
        sc(ex, f"B{r}", label, B, LT_F, align="left")
        sc(ex, f"C{r}", None, B, LT_F)
    else:
        sc(ex, f"B{r}", label, N, align="left")
        sc(ex, f"C{r}", f, B, None, MONEY, align="right")
    r += 1
r += 1
sc(ex, f"B{r}", "The people - archetypes vs the org", W, DK_F, align="left")
for col in "CDEFG": sc(ex, f"{col}{r}", None, W, DK_F)
r += 1
FTS = "'3.0 FTE View'"
for label, f, fm in [
    ("Seats the archetypes allow - your squads at their set sizes", f"={FTS}!$G$4", None),
    ("Seats actually raised in those squads - filled + vacant", f"={FTS}!$G${grand_ft}+{FTS}!$K${grand_ft}", None),
    ("Seats raised beyond the archetypes", f"={FTS}!$H$4", None),
    ("Seats in squads priced outside archetypes (AmPOS, EGI, cyber)", f"={FTS}!$J${grand_ft}-{FTS}!$G${grand_ft}-{FTS}!$K${grand_ft}", None),
    ("All org roles (squads + leadership + COEs + unmapped)", f"={FTS}!$C$4", None),
    ("Filled - people in seats today", f"={FTS}!$D$4", None),
    ("Vacant - raised, not yet hired", f"={FTS}!$E$4", None),
    ("of which squad seats - the GM hire or hold lever", (f"=COUNTIFS({RD}!$Q:$Q,\"Squad\",{RD}!$R:$R,\"Vacant\")"
        f"+COUNTIFS({RD}!$Q:$Q,\"Strategic Program\",{RD}!$R:$R,\"Vacant\")"), None),
    ("of which leadership, COE and unmapped seats", (f"={FTS}!$E$4-COUNTIFS({RD}!$Q:$Q,\"Squad\",{RD}!$R:$R,\"Vacant\")"
        f"-COUNTIFS({RD}!$Q:$Q,\"Strategic Program\",{RD}!$R:$R,\"Vacant\")"), None),
    ("Vacancy rate", f"={FTS}!$F$4", PCT)]:
    sc(ex, f"B{r}", label, N, align="left")
    sc(ex, f"C{r}", f, B, None, fm, align="right")
    r += 1
sc(ex, f"B{r}", "Vacant = open seats in the raw data. Ring-fenced roles count as filled. Costs come from the cost data (548 records / 156 vacant) - lined up on 3.1.", N, align="left")
r += 2
sc(ex, f"B{r}", "What it means", W, DK_F, align="left")
for col in "CDEFG": sc(ex, f"{col}{r}", None, W, DK_F)
r += 1
wim_first = r
for label, f, fm in [
    ("Today's filled seats cost ($m)", f"={TC}!$D${GRAND}", MONEY),
    ("Filled seats over/(under) the archetype cost ($m)", f"=ROUND({TC}!$D${GRAND}-{TC}!$C${GRAND},6)", MONEY),
    ("Hiring every vacant seat would add ($m)", f"={TC}!$E${GRAND}", MONEY),
    ("of which squad seats - your 4.x GM lever ($m)", "=0", MONEY),
    ("Vacant seats are priced at standard title rates - indicative until an offer is made.", "TXT", None),
    ("Taking the design OVER the archetypes by ($m)", f"={TC}!$G${GRAND}", MONEY),
    ("The main lever is the vacancies: they are raised but not hired, so holding them impacts nobody. Make the call seat by seat on your 4.x GM tab.", "TXT", None),
    ("Resizing the archetypes themselves (squad size, support %, onshore or offshore) is done with the dropdowns on the 1.x tabs - cost, variance and left to fund update live.", "TXT", None),
    ("Roles not mapped to any squad or COE - inside the 120.0 ($m)", f"={TC}!$F${subD}", MONEY),
    ("TDD Cyber - needs more than its budget, see 1.11 ($m)", f"='2.5 Cyber Roles'!$C${cy_tie}", MONEY),
    ("COEs - left to fund after budgets, see 2.2 ($m)", "='2.2 COE'!$F$13", MONEY)]:
    if f == "TXT":
        sc(ex, f"B{r}", label, N, align="left")
        for col in "CDEFG": sc(ex, f"{col}{r}", None, N)
    else:
        sc(ex, f"B{r}", label, N, align="left")
        sc(ex, f"C{r}", f, B, None, fm, align="right")
    r += 1
r += 1
sc(ex, f"B{r}", "Portfolio drill-down", W, DK_F, align="left")
for col in "CDEFG": sc(ex, f"{col}{r}", None, W, DK_F)
r += 1
sc(ex, f"B{r}", "Yellow cell = dropdown input - pick a portfolio", N, align="left")
r += 1
sel = r
sc(ex, f"B{r}", "Portfolio", N, align="left")
sc(ex, f"C{r}", "Ampol Retail", BLUE, YELLOW, None, align="center")
dvp = DataValidation(type="list", formula1='"' + ",".join(PORT[t] for t in TABS) + '"', allow_blank=True)
ex.add_data_validation(dvp); dvp.add(f"C{r}")
r += 1
for label, f, fm in [
    ("TDD Lights On budget ($m)", f"=INDEX({GS}!$C$6:$C$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
    ("TDD Cost ($m)", f"=INDEX({GS}!$D$6:$D$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
    ("TDD Variance - to fund ($m)", f"=MAX(0,-INDEX({GS}!$E$6:$E$16,MATCH($C${sel},{GS}!$B$6:$B$16,0)))", MONEY),
    ("Funded outside TDD ($m)", f"=INDEX({GS}!$G$6:$G$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
    ("Allocated outside TDD ($m)", f"=INDEX({GS}!$H$6:$H$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
    ("Left to fund ($m)", f"=INDEX({GS}!$I$6:$I$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
    ("Total cost ($m)", f"=INDEX({GS}!$J$6:$J$16,MATCH($C${sel},{GS}!$B$6:$B$16,0))", MONEY),
    ("Archetype squad seats allowed", f"=INDEX(Lists!$K$2:$K$12,MATCH($C${sel},Lists!$J$2:$J$12,0))", "0.0"),
    ("Org seats (excl leadership)", f"=COUNTIFS({RD}!$N:$N,$C${sel})-COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$Q:$Q,\"Leadership\")", None),
    ("Filled", f"=COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$R:$R,\"Filled\")-COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$Q:$Q,\"Leadership\",{RD}!$R:$R,\"Filled\")", None),
    ("Vacant", f"=COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$R:$R,\"Vacant\")-COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$Q:$Q,\"Leadership\",{RD}!$R:$R,\"Vacant\")", None),
    ("Leadership roles", f"=COUNTIFS({RD}!$N:$N,$C${sel},{RD}!$Q:$Q,\"Leadership\")", None)]:
    sc(ex, f"B{r}", label, N, align="left")
    sc(ex, f"C{r}", f, B, None, fm, align="right")
    r += 1
sc(ex, f"B{r}", "Org seats include roles outside the archetype model; squad-only counts are on 3.0 FTE View.", N, align="left")
var_row = sel + 3
pos_bad(ex, f"C{var_row}")

# =====================================================================
# F2. panel-fix post-pass (single-sourcing, labels, formats, hygiene)
# =====================================================================
# stale bare "2.1 COE" prose pointers left by the renumber
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            try:
                if isinstance(c.value, str) and not c.value.startswith("=") and "2.1 COE" in c.value:
                    c.value = c.value.replace("2.1 COE", "2.2 COE")
            except AttributeError:
                pass
# offshore rate single-source: squad offshore column derives from K5
for rr in range(5, 24):
    if sq.cell(rr, 7).value is not None:
        sq.cell(rr, 8).value = f"=G{rr}*$K$5"
sq["K5"].fill = YELLOW
# 0.1 Squads font -> Calibri (was Aptos, inconsistent with the model)
for row in sq.iter_rows():
    for c in row:
        try:
            if c.font is not None and c.font.name != "Calibri":
                f0 = c.font
                c.font = Font(name="Calibri", size=f0.size, bold=f0.bold, italic=f0.italic, color=f0.color)
        except AttributeError:
            pass
# SupportPct dropdown must cover every support % actually in use
sup_vals = [0, 0.2, 0.3, 0.4, 0.5, 0.7, 0.8, 0.9, 1]
for i, v in enumerate(sup_vals):
    ls.cell(2 + i, 4).value = v
wb.defined_names["SupportPct"].value = f"Lists!$D$2:$D${1 + len(sup_vals)}"
# 2.0 Group Summary: units on money headers, % convention, net tie to 2.1
gs2 = wb["2.0 Group Summary"]
for cc in ("G5", "H5", "I5"):
    v = gs2[cc].value
    if isinstance(v, str) and "($m)" not in v:
        gs2[cc].value = v.rstrip() + " ($m)"
gs2["C35"].number_format = "0%"
from openpyxl.utils import range_boundaries
for mr in list(gs2.merged_cells.ranges):
    c1_, r1_, c2_, r2_ = range_boundaries(str(mr))
    if not (r2_ < 25 or r1_ > 26):
        gs2.unmerge_cells(str(mr))
sc(gs2, "B25", "Less: BP & Domain Architect already funded in portfolio overheads (see 2.1)", N, LT_F, align="left")
sc(gs2, "J25", f"='2.1 Total Cost'!$C${dedup_row}", N, LT_F, MONEY, align="right")
sc(gs2, "B26", "Net operating model cost - ties to 2.1 Total Cost ($m)", B, GREY_F, align="left")
sc(gs2, "J26", f"=J24+'2.1 Total Cost'!$C${dedup_row}", B, GREY_F, MONEY, align="right")
# number-format drift: Data Config budget rows + portfolio Reconciled-to-Finance cells
for cc in ("C27", "D27", "E27", "C28", "D28", "E28"):
    wb["0.0 Data Config"][cc].number_format = MONEY
for t in TABS:
    w = wb[t]
    for rr in range(10, 15):
        if isinstance(w.cell(rr, 2).value, str) and "Reconciled to Finance" in str(w.cell(rr, 2).value):
            w.cell(rr, 5).number_format = MONEY
# header-fill bleed on the TDD Variance rows (1.2 / 1.3)
for t, cells in [("1.2 Customer", ("B15", "C15")), ("1.3 Enterprise Data", ("C14",))]:
    for cc in cells:
        wb[t][cc].fill = PatternFill()
        wb[t][cc].font = N
# float-residue negative zero on 1.10 (flows into 2.0 I15); it is a left-to-fund cell so MAX(0,) fits
w110 = wb["1.10 Z Retail"]
if isinstance(w110["I19"].value, str) and w110["I19"].value.startswith("="):
    w110["I19"].value = "=MAX(0,ROUND(" + w110["I19"].value[1:] + ",6))"
# cyber tie line: overheads charged on 1.11 (portfolio + platform rows)
ohr11 = oh_rows["1.11 TDD Cyber"]
cy_ws[f"C{cy_a['info_first']+4}"] = f"='1.11 TDD Cyber'!$C${ohr11}+'1.11 TDD Cyber'!$C${ohr11+1}"
# 2.0 variance header states its own math
gs2["E5"] = "Variance ($m) = budget - cost"
# disambiguate the two "Lights On" labels on the portfolio tabs (Finance block labels are the owner's)
for t in TABS:
    w = wb[t]
    for rr in (4, 5, 6):
        for cc in range(2, 9):
            v = w.cell(rr, cc).value
            if isinstance(v, str) and "TDD Lights On" in v and "Data Config" not in v:
                w.cell(rr, cc).value = v.rstrip() + " (people - 0.0 Data Config)"
# portfolio overhead label points at its breakdown
for t, ohr_ in oh_rows.items():
    w = wb[t]
    if w.cell(ohr_, 2).value == "Portfolio Overhead":
        w.cell(ohr_, 2).value = "Portfolio Overhead (see 0.0 Data Config)"
# one phrasing for the rolled-in cyber COE note
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            try:
                if isinstance(c.value, str) and "rolled into" in c.value and "1.11 TDD Cyber" not in c.value:
                    c.value = c.value.replace("rolled into 1.11", "rolled into 1.11 TDD Cyber") \
                        if "rolled into 1.11" in c.value else \
                        c.value.replace("rolled into TDD Cyber", "rolled into 1.11 TDD Cyber")
            except AttributeError:
                pass
# error alerts ON for every list dropdown - off-list typing gets rejected
for ws in wb.worksheets:
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list":
            dv.showErrorMessage = True
            dv.errorStyle = "stop"
            dv.errorTitle = "Invalid entry"
            dv.error = "Pick a value from the list"
# plumbing and legacy sheets are hidden, not shown to GMs (nothing references them in formulas
# except the dropdown lists, which keep working from a hidden sheet)
for s in ("Lists", "squad mapping", "0.2 FY26 Budget"):
    if s in wb.sheetnames:
        wb[s].sheet_state = "hidden"
# gridlines off on the four stragglers
for s in ("0.3 For Presentation Pack (2)", "0.4 Budget Table (Fin)", "squad mapping", "Lists"):
    wb[s].sheet_view.showGridLines = False
# 2.2: G is simply the total cost (F is now genuinely net, so no duplicate columns)
coe["G7"] = "Total cost ($m)"
# Left-to-fund conditional format on role-tab total rows (2.5 already has it)
pos_bad(bp_ws, f"H{bp_a['tot']}")
pos_bad(sad_ws, f"H{sad_a['tot']}")
# centre the blue column-header layer on tabs still left-aligned (labels in B stay left)
NAVY_RGB = "FF1F4E79"
for t in TABS + ["2.0 Group Summary", "2.2 COE"]:
    w = wb[t]
    for row in w.iter_rows():
        for c in row:
            try:
                if (c.value is not None and c.fill is not None and c.fill.patternType == "solid"
                        and getattr(c.fill.fgColor, "rgb", None) == NAVY_RGB
                        and c.font is not None and c.font.bold and c.column > 2):
                    c.alignment = Alignment(horizontal="center", vertical="center")
            except AttributeError:
                pass

# =====================================================================
# F3. GM working-copy tabs (4.x): names, vacancies, hire/hold decisions
# =====================================================================
def esc(s):
    return str(s).replace('"', '""')
rd_by_port = {}
for rr in range(2, rdv2.max_row + 1):
    port = rdv2.cell(rr, 14).value
    if port is None: continue
    cls_ = str(rdv2.cell(rr, 17).value or "")
    if cls_ not in ("Squad", "Strategic Program"): continue
    sqn_ = str(rdv2.cell(rr, 16).value or "").strip()
    if not sqn_: sqn_ = "(no squad recorded)"
    rd_by_port.setdefault(str(port).strip(), {}).setdefault(sqn_, []).append(rr)
gm_anchors = {}
GM_TABS = []
for i, t in enumerate(TABS, 1):
    pl = PORT[t]
    gname = f"4.{i} GM {pl}"[:31]
    GM_TABS.append(gname)
    gm = wb.create_sheet(gname)
    gm.sheet_view.showGridLines = False
    for col, wd in {"A":3,"B":44,"C":46,"D":12,"E":13,"F":14,"G":15,"H":17,"I":15,"J":30}.items():
        gm.column_dimensions[col].width = wd
    gm.row_dimensions[2].height = 21
    sc(gm, "B2", f"{pl} - GM working copy: your squads, your people, your hire or hold calls", TITLE, border=False)
    arch_squads = {nm: fr for (nm, fr) in ft_rows.get(t, [])}
    port_squads = rd_by_port.get(pl, {})
    ordered = [nm for (nm, _) in ft_rows.get(t, []) if nm != "__CYBER__" and nm in port_squads]
    ordered += [nm for nm in port_squads if nm not in ordered]
    dv_dec = DataValidation(type="list", formula1='"Hire,Hold"', allow_blank=False,
                            showErrorMessage=True, errorStyle="stop",
                            errorTitle="Invalid entry", error="Pick Hire or Hold")
    gm.add_data_validation(dv_dec)
    # summary skeleton on TOP; formulas filled after the roster is written
    r = 4
    sc(gm, f"B{r}", "Your position by squad", W, DK_F, align="left")
    for col in "CDEFGHIJ": sc(gm, f"{col}{r}", None, W, DK_F)
    r += 1
    hdrs = ["Squad", "", "Archetype seats", "Filled", "Vacant", "Planning to hire",
            "Seats after your calls", "vs archetype", "Flag"]
    for j, h in enumerate(hdrs):
        if h: sc(gm, f"{'BCDEFGHIJ'[j]}{r}", h, W, NAVY_F, align="center")
    r += 1
    sum_first = r
    sum_rows = {}
    for nm in ordered:
        sum_rows[nm] = r
        r += 1
    tot_row = r
    r += 1
    vac_cost_row = r
    r += 1
    plan_cost_row = r
    r += 1
    caveat_row = r
    r += 1
    note_row = r
    r += 1
    cyb_note_row = r if t == "1.11 TDD Cyber" else None
    if cyb_note_row: r += 1
    r += 1
    # roster blocks
    sc(gm, f"B{r}", "Your people - decide Hire or Hold on every vacant seat (yellow cells)", W, DK_F, align="left")
    for col in "CDEFGHIJ": sc(gm, f"{col}{r}", None, W, DK_F)
    r += 1
    for j, h in enumerate(["Name", "Role", "Status", "Your call", "Cost if hired ($)"]):
        sc(gm, f"{'BCDEF'[j]}{r}", h, W, NAVY_F, align="center")
    r += 1
    roster_first = r
    dec_ranges = {}
    for nm in ordered:
        in_arch = nm in arch_squads or "__CYBER__" in arch_squads
        label = nm if in_arch else f"{nm} (no archetype)"
        sc(gm, f"B{r}", label, B, LT_F if in_arch else RED_F, align="left")
        for col in "CDE": sc(gm, f"{col}{r}", None, B, LT_F if in_arch else RED_F)
        r += 1
        d_first = None
        for rr in port_squads[nm]:
            st = str(rdv2.cell(rr, 18).value or "")
            sc(gm, f"B{r}", f"='raw data'!$B${rr}", N, align="left")
            sc(gm, f"C{r}", f"=SUBSTITUTE(SUBSTITUTE('raw data'!$C${rr},\"–\",\"-\"),\"—\",\"-\")", N, align="left")
            sc(gm, f"D{r}", f"='raw data'!$R${rr}", N, align="center")
            if st == "Vacant":
                sc(gm, f"E{r}", "Hold", B, YELLOW, align="center")
                sc(gm, f"F{r}", AD_COST_T.format(src=rr), N, None, DOLLAR, align="right")
                dv_dec.add(f"E{r}")
                if d_first is None: d_first = r
                d_last = r
            r += 1
        if d_first is not None:
            dec_ranges[nm] = (d_first, d_last)
    tbl_last = r - 1
    gm.conditional_formatting.add(f"D{roster_first}:D{tbl_last}",
        CellIsRule(operator="equal", formula=['"Vacant"'], fill=RED_F, font=RED_T))
    # fill the summary
    for nm in ordered:
        sr_ = sum_rows[nm]
        in_arch = nm in arch_squads
        cyber = "__CYBER__" in arch_squads
        sc(gm, f"B{sr_}", nm if (in_arch or cyber) else f"{nm} (no archetype)", N,
           None if (in_arch or cyber) else RED_F, align="left")
        if in_arch:
            sc(gm, f"D{sr_}", f"='3.0 FTE View'!$G${arch_squads[nm]}", N, align="center")
        elif cyber:
            sc(gm, f"D{sr_}", "priced from 2.5", N, align="center")
        else:
            sc(gm, f"D{sr_}", "-", N, align="center")
        sc(gm, f"E{sr_}", f"=COUNTIFS({RD}!$N:$N,\"{esc(pl)}\",{RD}!$P:$P,\"{esc(nm)}\",{RD}!$R:$R,\"Filled\")", N, align="center")
        sc(gm, f"F{sr_}", f"=COUNTIFS({RD}!$N:$N,\"{esc(pl)}\",{RD}!$P:$P,\"{esc(nm)}\",{RD}!$R:$R,\"Vacant\")", N, align="center")
        if nm in dec_ranges:
            df, dl = dec_ranges[nm]
            sc(gm, f"G{sr_}", f"=COUNTIF(E{df}:E{dl},\"Hire\")", N, align="center")
        else:
            sc(gm, f"G{sr_}", 0, N, align="center")
        sc(gm, f"H{sr_}", f"=E{sr_}+G{sr_}", B, align="center")
        sc(gm, f"I{sr_}", f'=IFERROR(H{sr_}-D{sr_},"-")', N, align="center")
        sc(gm, f"J{sr_}", (f"=IF(ISNUMBER(D{sr_}),IF(E{sr_}>D{sr_},\"Filled already over archetype\","
                           f"IF(H{sr_}>D{sr_},\"Over archetype after your calls\",\"\")),"
                           f"\"No archetype - review these roles\")"), N, align="left")
    gm.conditional_formatting.add(f"I{sum_first}:I{tot_row}",
        FormulaRule(formula=[f"AND(ISNUMBER(I{sum_first}),I{sum_first}>0)"], fill=RED_F, font=RED_T))
    gm.conditional_formatting.add(f"I{sum_first}:I{tot_row}",
        FormulaRule(formula=[f"AND(ISNUMBER(I{sum_first}),I{sum_first}<=0)"], fill=GREEN_F, font=GREEN_T))
    sc(gm, f"B{tot_row}", "Total", B, GREY_F, align="left")
    for colL in "DEFGH":
        sc(gm, f"{colL}{tot_row}", f"=SUM({colL}{sum_first}:{colL}{tot_row-1})", B, GREY_F, align="center")
    sc(gm, f"I{tot_row}", f'=IF(COUNT(I{sum_first}:I{tot_row-1})=0,"-",SUM(I{sum_first}:I{tot_row-1}))', B, GREY_F, align="center")
    sc(gm, f"B{vac_cost_row}", "Cost to hire all vacant squad seats ($m)", N, LT_F, align="left")
    sc(gm, f"C{vac_cost_row}", f"=SUM(F{roster_first}:F{tbl_last})/1000000", B, LT_F, MONEY, align="right")
    sc(gm, f"B{plan_cost_row}", "Cost of the seats you chose to Hire ($m)", N, LT_F, align="left")
    sc(gm, f"C{plan_cost_row}", f"=SUMIF(E{roster_first}:E{tbl_last},\"Hire\",F{roster_first}:F{tbl_last})/1000000",
       B, LT_F, MONEY, align="right")
    sc(gm, f"B{caveat_row}", "Vacant seats are priced at standard title rates - indicative until an offer is made.", N, align="left")
    sc(gm, f"B{note_row}", "Leadership roles are funded via the portfolio overheads and sit on 3.0 FTE View, not here.", N, align="left")
    if cyb_note_row:
        sc(gm, f"B{cyb_note_row}", "Cyber vacancies include risk and security roles - confirm with cyber risk before holding.", N, align="left")
    gm_anchors[gname] = dict(sum_first=sum_first, tot=tot_row, vac_cost=vac_cost_row,
                             plan_cost=plan_cost_row, roster_first=roster_first,
                             tbl_last=tbl_last, n_squads=len(ordered))

# =====================================================================
# G. order + anchors json
# =====================================================================
ORDER = ["Exec Summary","0.0 Data Config","0.1 Squads","0.2 FY26 Budget",
         "0.3 For Presentation Pack (2)","0.4 Budget Table (Fin)"] + TABS + \
        ["2.0 Group Summary","2.1 Total Cost","2.2 COE","2.3 BP&T","2.4 SA&D","2.5 Cyber Roles",
         "3.0 FTE View","3.1 Data QA"] + GM_TABS + ["squad mapping","raw data","Added data","Lists"]
# rebuild 3.1 Data QA name (old 5.1 was deleted; recreate from v7's content copy in wbv)
qa_src = wbv["5.1 Data QA"]
qa = wb.create_sheet("3.1 Data QA")
qa.sheet_view.showGridLines = False
for col, wd in {"A":3,"B":34,"C":16,"D":16,"E":16,"F":44}.items():
    qa.column_dimensions[col].width = wd
for row in qa_src.iter_rows():
    for c in row:
        if c.value is not None:
            qa.cell(c.row, c.column).value = c.value
wf7 = openpyxl.load_workbook(SRC)
qs = wf7["5.1 Data QA"]
for row in qs.iter_rows():
    for c in row:
        if c.value is not None or (c.fill and c.fill.patternType):
            qa.cell(c.row, c.column)._style = copy(c._style)
qa["B2"] = str(qa["B2"].value) + " - point-in-time snapshot"
for row in qa.iter_rows():
    for c in row:
        if c.value == "Distinct named people": c.value = "Distinct named records"
mrq = qa.max_row
sc(qa, f"B{mrq+2}", "Live check - raw data rows now", B, LT_F, align="left")
sc(qa, f"C{mrq+2}", "=COUNTA('raw data'!$B$2:$B$1000)", B, LT_F, align="center")
sc(qa, f"B{mrq+3}", "Live check - cost data rows now", B, LT_F, align="left")
sc(qa, f"C{mrq+3}", "=COUNTA('Added data'!$B$2:$B$549)", B, LT_F, align="center")
sc(qa, f"B{mrq+4}", "If these differ from the snapshot above, refresh this tab.", N, align="left")
# drill-down archetype-seat lookup (hidden Lists helper), exec lever rewired to the GM tabs
ls["J1"] = "Portfolio"; ls["K1"] = "Archetype seats"
for i, t in enumerate(TABS, 1):
    ls.cell(1 + i, 10).value = PORT[t]
    ls.cell(1 + i, 11).value = f"='3.0 FTE View'!$G${port_sub[t]}"
lever_row = None
for rr in range(1, ex.max_row + 1):
    if ex.cell(rr, 2).value == "of which squad seats - your 4.x GM lever ($m)": lever_row = rr
ex[f"C{lever_row}"] = "=" + "+".join(
    f"'{gname}'!$C${gm_anchors[gname]['vac_cost']}" for gname in GM_TABS)
ex[f"C{lever_row}"].number_format = MONEY
for gname in GM_TABS:
    ga = gm_anchors[gname]
    w = wb[gname]
    for rr in range(ga["sum_first"], ga["tot"] + 1):
        for cc in (4, 9):
            w.cell(rr, cc).number_format = "0.0"
# no en or em dashes, no italics, no tiny fonts on the model tabs we author
MODEL_TABS = ["Exec Summary", "2.0 Group Summary", "2.1 Total Cost", "2.2 COE", "2.3 BP&T",
              "2.4 SA&D", "2.5 Cyber Roles", "3.0 FTE View", "3.1 Data QA"] + GM_TABS + TABS
GREYS = {"FF808080", "FF7F7F7F", "FFA6A6A6", "FF999999", "FFBFBFBF"}
for tname in MODEL_TABS:
    ws = wb[tname]
    col_cap = 9 if tname in TABS else 10000   # owner scratch on 1.x lives in J+
    for row in ws.iter_rows():
        for c in row:
            if c.column > col_cap: continue
            try:
                if isinstance(c.value, str) and not c.value.startswith("="):
                    if "–" in c.value or "—" in c.value:
                        c.value = c.value.replace("–", "-").replace("—", "-")
                    if "TDD cost" in c.value: c.value = c.value.replace("TDD cost", "TDD Cost")
                    if "lights-on" in c.value.lower():
                        c.value = c.value.replace("Lights-on", "Lights On").replace("lights-on", "Lights On")
                if c.value is not None and c.font is not None and (c.font.italic or (c.font.size or 10) < 10
                        or (getattr(c.font.color, "rgb", None) in GREYS)):
                    f0 = c.font
                    c.font = Font(name=f0.name or "Calibri", size=max(10, f0.size or 10),
                                  bold=f0.bold, italic=False)
            except AttributeError:
                pass
# stray General-format money cells on 1.x + leftover yellow on empty squad slots
for t in TABS:
    w = wb[t]
    i12 = w.cell(12, 9)
    if i12.number_format == "General" and i12.value is not None:
        i12.number_format = MONEY
    for rr in range(4, 71):
        for cc in (7, 9):
            c = w.cell(rr, cc)
            if c.number_format == "General" and (isinstance(c.value, (int, float))
                    or (isinstance(c.value, str) and c.value.startswith("=SUM"))):
                c.number_format = MONEY
        if w.cell(rr, 2).value in (None, ""):
            for cc in range(3, 7):
                cf = w.cell(rr, cc)
                if cf.fill is not None and getattr(cf.fill.fgColor, "rgb", None) == "FFFFF2CC" and cf.value in (None, ""):
                    cf.fill = PatternFill()
wb._sheets = [wb[s] for s in ORDER if s in wb.sheetnames]

anchors = dict(grand=GRAND, check=CHECK, subA=subA, subB=subB, subC=subC, subD=subD,
               dedup=dedup_row, ft_grand=grand_ft, ft_xcheck=xcheck_row,
               gs_net=26, cy_tie=cy_tie, unspec_first=un_first2, unspec_tot=unspec_tot_row,
               bp=bp_a, sad=sad_a, cyb=cy_a, cyt=cyt, gm=gm_anchors,
               n_bpt=len(groups["BPT"]), n_sad=len(groups["SAD"]), n_cyb=len(groups["CYB"]),
               n_unspec=len(unspec))
def ser(o):
    if isinstance(o, dict): return {k: ser(v) for k, v in o.items()}
    return o
with open(SCR + "anchors_v8.json", "w") as fjson:
    json.dump(ser(anchors), fjson, indent=1)

# strip any comments once more
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if c.comment is not None: c.comment = None

wb.save(OUT)
print("saved", OUT)
print("groups:", {k: len(v) for k, v in groups.items()})
print("anchors:", {k: v for k, v in anchors.items() if not isinstance(v, dict)})

# restore Added data cached values (same machinery as v7)
import zipfile, shutil, os
import xml.etree.ElementTree as ET
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NSR = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", NS); ET.register_namespace("r", NSR)
tmp = SCR + "_v8tmp"
if os.path.exists(tmp): shutil.rmtree(tmp)
os.makedirs(tmp)
with zipfile.ZipFile(OUT) as z: z.extractall(tmp)
wbt = ET.parse(f"{tmp}/xl/workbook.xml"); root = wbt.getroot()
rid = None
for sh in root.iter(f"{{{NS}}}sheet"):
    if sh.get("name") == "Added data": rid = sh.get(f"{{{NSR}}}id")
rels = ET.parse(f"{tmp}/xl/_rels/workbook.xml.rels")
target = None
for rel in rels.getroot():
    if rel.get("Id") == rid: target = rel.get("Target")
path = f"{tmp}/xl/{target}" if not target.startswith("/") else f"{tmp}{target}"
tree = ET.parse(path); troot = tree.getroot()
fixed = 0
for c in troot.iter(f"{{{NS}}}c"):
    coord = c.get("r")
    f_el = c.find(f"{{{NS}}}f")
    if f_el is None: continue
    cached = adv[coord].value
    if cached is None: continue
    v_el = c.find(f"{{{NS}}}v")
    if v_el is None: v_el = ET.SubElement(c, f"{{{NS}}}v")
    if v_el.text in (None, ""):
        if isinstance(cached, (int, float)):
            if c.get("t") in ("str", "e"): del c.attrib["t"]
            v_el.text = repr(float(cached)); fixed += 1
        else:
            c.set("t", "str"); v_el.text = str(cached); fixed += 1
tree.write(path, xml_declaration=True, encoding="UTF-8")
os.remove(OUT)
with zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED) as z:
    for base, _, files in os.walk(tmp):
        for fn in files:
            full = os.path.join(base, fn)
            z.write(full, os.path.relpath(full, tmp))
shutil.rmtree(tmp)
print("Added data caches restored:", fixed)
